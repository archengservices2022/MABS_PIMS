# project_number_generator.py
import sys
import os
import json
import re
import tempfile
import shutil
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import secrets
import hashlib
import threading
from payment_tracker import get_payment_tracker
from payment_dialog import PaymentDialog, PaymentHistoryDialog
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import pyqtSignal
from main import ItemRowWidget  # Add this import
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal


from app_logger import get_logger
from app_theme import configure_filter_button
_log = get_logger(__name__)


class FrozenProjectTable(QtWidgets.QTableWidget):
    """QTableWidget with a fixed Project # column synced to the main grid."""

    def __init__(self, *args, frozen_column=1, **kwargs):
        super().__init__(*args, **kwargs)
        self._frozen_column = frozen_column
        self._frozen_width = 118
        self._frozen_view = QtWidgets.QTableView(self)
        self._frozen_view.setModel(self.model())
        self._frozen_view.setSelectionModel(self.selectionModel())
        self._frozen_view.setFocusPolicy(QtCore.Qt.NoFocus)
        self._frozen_view.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self._frozen_view.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._frozen_view.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self._frozen_view.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self._frozen_view.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self._frozen_view.setFrameShape(QtWidgets.QFrame.NoFrame)
        self._frozen_view.verticalHeader().hide()
        self._frozen_view.horizontalHeader().setHighlightSections(False)
        self._frozen_view.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        self._frozen_view.clicked.connect(self._handle_frozen_click)
        self._frozen_view.setStyleSheet("""
            QTableView {
                background: #ffffff;
                border-right: 2px solid #cbd5e1;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 12px;
                gridline-color: #e5e7eb;
            }
            QTableView::item {
                background: #ffffff;
                padding: 8px 10px;
                color: #0f172a;
                border-bottom: 1px solid #f3f4f6;
            }
            QTableView::item:alternate {
                background: #ffffff;
            }
            QTableView::item:selected {
                background: #dbeafe;
                color: #1e40af;
            }
            QHeaderView::section {
                background: #f9fafb;
                color: #374151;
                font-weight: 800;
                font-size: 12px;
                padding: 11px 8px;
                border: none;
                border-bottom: 2px solid #e5e7eb;
                border-right: 2px solid #cbd5e1;
            }
        """)

        self.verticalHeader().sectionResized.connect(self._sync_frozen_row_height)
        self.horizontalHeader().sectionResized.connect(self._handle_section_resized)
        self.verticalScrollBar().valueChanged.connect(
            self._frozen_view.verticalScrollBar().setValue
        )
        self._frozen_view.verticalScrollBar().valueChanged.connect(
            self.verticalScrollBar().setValue
        )
        self.viewport().stackUnder(self._frozen_view)

    def refresh_frozen_project_column(self, width=None):
        if width is not None:
            self._frozen_width = width
        for col in range(self.model().columnCount()):
            self._frozen_view.setColumnHidden(col, col != self._frozen_column)
        self._frozen_view.setColumnWidth(self._frozen_column, self._frozen_width)
        self.setColumnHidden(self._frozen_column, False)
        self.setViewportMargins(self._frozen_width, 0, 0, 0)
        self._frozen_view.horizontalHeader().setFixedHeight(self.horizontalHeader().height())
        self._frozen_view.verticalHeader().setDefaultSectionSize(
            self.verticalHeader().defaultSectionSize()
        )
        for row in range(self.rowCount()):
            self._frozen_view.setRowHeight(row, self.rowHeight(row))
        self._update_frozen_geometry()

    def _handle_section_resized(self, logical_index, old_size, new_size):
        if logical_index == self._frozen_column:
            self._frozen_width = new_size
            self._frozen_view.setColumnWidth(self._frozen_column, new_size)
            self.setViewportMargins(new_size, 0, 0, 0)
        self._update_frozen_geometry()

    def _sync_frozen_row_height(self, logical_index, old_size, new_size):
        self._frozen_view.setRowHeight(logical_index, new_size)

    def _handle_frozen_click(self, index):
        self.setCurrentCell(index.row(), self._frozen_column)
        self.selectRow(index.row())
        self.cellClicked.emit(index.row(), self._frozen_column)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_frozen_geometry()

    def _update_frozen_geometry(self):
        self._frozen_view.setGeometry(
            self.frameWidth(),
            self.frameWidth(),
            self._frozen_width,
            self.viewport().height() + self.horizontalHeader().height(),
        )
        self._frozen_view.raise_()


_PAYMENT_STAGES = ["1st Installment", "2nd Installment", "3rd Installment", "Final Payment"]


def _detect_payment_stage(project: dict, existing_invoices: list) -> dict:
    """Auto-detect next payment stage and amount for a project.
    Returns: {stage, amount, blocked, reason}
    """
    pn = project.get("project_number", "").strip().upper()
    total = float(project.get("project_amount", 0) or 0)

    # Build stage → {status, amount} from existing invoices
    invoiced = {}
    for inv in existing_invoices:
        inv_status = (inv.get("meta") or {}).get("status", inv.get("status", "Pending"))
        for item in inv.get("items", []):
            if item.get("project_number", "").strip().upper() != pn:
                continue
            cat = item.get("payment_category", "").strip().lower()
            if "down" in cat or "deposit" in cat or "50%" in cat or "1st" in cat or "first" in cat:
                key = "1st Installment"
            elif "2nd" in cat or "term 2" in cat or "second" in cat:
                key = "2nd Installment"
            elif "3rd" in cat or "term 3" in cat or "third" in cat:
                key = "3rd Installment"
            elif "4th" in cat or "term 4" in cat or "fourth" in cat:
                key = "4th Installment"
            elif "final" in cat:
                key = "Final Payment"
            elif "full" in cat or "due" in cat or "balance" in cat:
                key = "Final Payment"
            else:
                continue
            invoiced[key] = {"status": inv_status, "amount": float(item.get("unit_price", 0) or 0)}

    # Determine deposit amount
    stored_deposit = float(project.get("deposit_amount", 0) or 0)
    deposit_rule = project.get("deposit_rule", "")
    if stored_deposit > 0:
        down_amt = stored_deposit
    elif "50%" in deposit_rule:
        down_amt = total * 0.50
    elif "25%" in deposit_rule:
        down_amt = total * 0.25
    else:
        down_amt = total  # full amount if no deposit rule

    for i, stage in enumerate(_PAYMENT_STAGES):
        if stage in invoiced:
            continue  # already invoiced for this stage
        # Calculate amount
        if stage == "1st Installment":
            amt = down_amt
        elif stage == "Final Payment":
            prev = invoiced.get("1st Installment", {}).get("amount", down_amt)
            amt = max(total - prev, 0)
        else:
            prev_total = sum(v["amount"] for v in invoiced.values())
            amt = max(total - prev_total, 0)
        # Check if previous stage paid (check both invoice status AND payment tracker)
        blocked, reason = False, ""
        if i > 0:
            prev_stage = _PAYMENT_STAGES[i - 1]
            if prev_stage in invoiced:
                ps = invoiced[prev_stage]["status"].lower()
                # Also check payment tracker — payment may have been recorded manually
                try:
                    tracker_paid = float(
                        get_payment_tracker().get_payment_summary(pn, total).get("total_paid", 0)
                    )
                except Exception:
                    tracker_paid = 0.0
                prev_amt = invoiced[prev_stage].get("amount", 0)
                if "paid" not in ps and tracker_paid < prev_amt:
                    blocked = True
                    reason = f"{prev_stage} invoice not yet paid (status: {invoiced[prev_stage]['status']})"
        if not reason:
            reason = "Ready" if i == 0 else f"{_PAYMENT_STAGES[i-1]} paid — ready for {stage}"
        return {"stage": stage, "amount": amt, "blocked": blocked, "reason": reason}

    return {"stage": None, "amount": 0, "blocked": True, "reason": "All payment stages already invoiced"}


def _split_percentages(raw) -> list:
    if isinstance(raw, (list, tuple)):
        parts = raw
    else:
        parts = re.findall(r"\d+(?:\.\d+)?", str(raw or ""))

    values = []
    for part in parts:
        try:
            value = float(part)
            if value > 0:
                values.append(value)
        except (TypeError, ValueError):
            continue
    return values[:4]


def _project_invoice_stage_plan(project: dict) -> list:
    """Return ordered list of {stage, amount} dicts for a project's payment plan."""
    total = float(project.get("project_amount", 0) or 0)

    # ── NEW model (payment_type / payment_stages) ───────────────────────
    ptype  = project.get("payment_type", "")
    stages = project.get("payment_stages") or []

    if ptype and stages:
        # Stages are ordered label strings; distribute amounts
        n = len(stages)
        down_pct = float(project.get("down_payment_percent", 0) or 0)
        result = []
        remaining = total
        for i, stage_label in enumerate(stages):
            if i == 0 and down_pct > 0:
                amt = total * down_pct / 100.0
            elif i == n - 1:
                amt = remaining  # last stage gets whatever is left
            else:
                amt = remaining / (n - i)   # split equally across remaining stages
            amt = max(round(amt, 2), 0.0)
            remaining = max(remaining - amt, 0.0)
            result.append({"stage": stage_label, "amount": amt})
        return result

    # ── Legacy / backward-compat fallback ──────────────────────────────
    split = _split_percentages(project.get("payment_split_percentages", ""))
    plan  = str(project.get("payment_terms") or project.get("payment_category") or "Single Payment")

    if split:
        current_total = sum(split)
        if current_total and abs(current_total - 100) > 0.01:
            split = [(v / current_total) * 100 for v in split]
        labels_by_count = {
            1: [ItemRowWidget.FULL_AMOUNT_LABEL],
            2: [ItemRowWidget.DEPOSIT_LABEL, ItemRowWidget.FINAL_PAYMENT_LABEL],
            3: [ItemRowWidget.DEPOSIT_LABEL, ItemRowWidget.TERM_2_LABEL, ItemRowWidget.FINAL_PAYMENT_LABEL],
            4: [ItemRowWidget.DEPOSIT_LABEL, ItemRowWidget.TERM_2_LABEL,
                ItemRowWidget.TERM_3_LABEL, ItemRowWidget.FINAL_PAYMENT_LABEL],
        }
        labels = labels_by_count.get(len(split), labels_by_count[4])
        return [{"stage": labels[i], "amount": total * pct / 100}
                for i, pct in enumerate(split)]

    if any(x in plan.lower() for x in ("50%", "deposit", "down payment", "25%")):
        default_pct = 0.25 if "25%" in plan else 0.50
        deposit_amt = float(project.get("deposit_amount", 0) or 0) or total * default_pct
        return [
            {"stage": ItemRowWidget.DEPOSIT_LABEL, "amount": deposit_amt},
            {"stage": ItemRowWidget.FINAL_PAYMENT_LABEL, "amount": max(total - deposit_amt, 0)},
        ]

    if "installment" in plan.lower() or "custom" in plan.lower():
        inst = int(project.get("installment_count", 2) or 2)
        amt  = round(total / inst, 2)
        ordinals = ["1st", "2nd", "3rd", "4th", "5th", "6th"]
        return [{"stage": f"{ordinals[i] if i < len(ordinals) else str(i+1)+'th'} Installment",
                 "amount": amt if i < inst - 1 else max(total - amt * (inst - 1), 0)}
                for i in range(inst)]

    return [{"stage": ItemRowWidget.FULL_AMOUNT_LABEL, "amount": total}]


def _detect_payment_stage(project: dict, existing_invoices: list) -> dict:
    """Auto-detect next invoice stage and amount for a project."""
    pn = project.get("project_number", "").strip().upper()
    total = float(project.get("project_amount", 0) or 0)
    planned_stages = _project_invoice_stage_plan(project)

    invoiced = {}
    for inv in existing_invoices:
        inv_status = (inv.get("meta") or {}).get("status", inv.get("status", "Pending"))
        for item in inv.get("items", []):
            if item.get("project_number", "").strip().upper() != pn:
                continue
            key = ItemRowWidget.normalize_payment_label(item.get("payment_category", ""))
            if key:
                invoiced[key] = {"status": inv_status, "amount": float(item.get("unit_price", 0) or 0)}

    for index, planned in enumerate(planned_stages):
        stage = planned["stage"]
        if stage in invoiced:
            continue

        blocked, reason = False, ""
        if index > 0:
            prev_stage = planned_stages[index - 1]["stage"]
            if prev_stage in invoiced:
                ps = invoiced[prev_stage]["status"].lower()
                try:
                    tracker_paid = float(
                        get_payment_tracker().get_payment_summary(pn, total).get("total_paid", 0)
                    )
                except Exception:
                    tracker_paid = 0.0
                prev_amt = invoiced[prev_stage].get("amount", 0)
                if "paid" not in ps and tracker_paid < prev_amt:
                    blocked = True
                    reason = f"{prev_stage} invoice not yet paid (status: {invoiced[prev_stage]['status']})"

        if not reason:
            reason = "Ready" if index == 0 else f"{planned_stages[index - 1]['stage']} paid - ready for {stage}"
        return {"stage": stage, "amount": planned["amount"], "blocked": blocked, "reason": reason}

    return {"stage": None, "amount": 0, "blocked": True, "reason": "All payment stages already invoiced"}


class PaymentStageConfirmDialog(QtWidgets.QDialog):
    """Confirmation dialog showing auto-detected payment stages before loading to invoice."""

    def __init__(self, project_stages: list, parent=None):
        super().__init__(parent)
        self.project_stages = project_stages
        self.setWindowTitle("Auto-Detected Payment Stages")
        self.setModal(True)
        self.resize(760, 400)
        self.setStyleSheet("QDialog { background: #f8fafc; }")
        self._build_ui()

    def _build_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        title = QtWidgets.QLabel("Payment Stage Detection")
        title.setStyleSheet("font-size:16px;font-weight:700;color:#0f172a;font-family:'Inter','Segoe UI';")
        sub = QtWidgets.QLabel("System auto-detected the next invoice stage for each project. Review then confirm.")
        sub.setStyleSheet("font-size:12px;color:#64748b;font-family:'Inter','Segoe UI';")
        layout.addWidget(title)
        layout.addWidget(sub)

        tbl = QtWidgets.QTableWidget(0, 5)
        tbl.setHorizontalHeaderLabels(["Project #", "Name", "Next Stage", "Amount", "Status"])
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        tbl.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        tbl.setAlternatingRowColors(True)
        tbl.setStyleSheet("""
            QTableWidget { background:white; border:1px solid #e2e8f0; font-size:12px; font-family:'Inter','Segoe UI'; }
            QTableWidget::item { padding:8px; color:#1e293b; }
            QHeaderView::section { background:#f1f5f9; color:#334155; font-weight:600; padding:8px; border:none; border-bottom:1px solid #e2e8f0; }
            QTableWidget::item:alternate { background:#f8fafc; }
        """)
        hdr = tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        hdr.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        hdr.setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        hdr.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        tbl.setColumnWidth(0, 130)
        tbl.setColumnWidth(2, 130)
        tbl.setColumnWidth(3, 110)

        blocked_count = 0
        for ps in self.project_stages:
            p, si = ps["project"], ps["stage_info"]
            row = tbl.rowCount()
            tbl.insertRow(row)
            tbl.setRowHeight(row, 40)

            tbl.setItem(row, 0, QtWidgets.QTableWidgetItem(p.get("project_number", "")))
            tbl.setItem(row, 1, QtWidgets.QTableWidgetItem(p.get("project_name", "")))

            stage_lbl = si["stage"] or "Fully Invoiced"
            si_item = QtWidgets.QTableWidgetItem(stage_lbl)
            color = QtGui.QColor("#dc2626" if si["blocked"] else "#16a34a")
            si_item.setForeground(color)
            tbl.setItem(row, 2, si_item)

            amt_item = QtWidgets.QTableWidgetItem(f"${si['amount']:,.2f}" if si["amount"] > 0 else "—")
            amt_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            tbl.setItem(row, 3, amt_item)

            reason_item = QtWidgets.QTableWidgetItem(si["reason"])
            reason_item.setForeground(QtGui.QColor("#b45309" if si["blocked"] else "#0f766e"))
            tbl.setItem(row, 4, reason_item)

            if si["blocked"]:
                blocked_count += 1

        layout.addWidget(tbl)

        if blocked_count:
            warn = QtWidgets.QLabel(
                f"⚠  {blocked_count} project(s) have a prior stage not yet paid. "
                "They will still be loaded — confirm with client before sending."
            )
            warn.setStyleSheet(
                "color:#92400e;font-size:12px;background:#fef3c7;"
                "padding:8px 12px;border-radius:6px;font-family:'Inter','Segoe UI';"
            )
            warn.setWordWrap(True)
            layout.addWidget(warn)

        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch()
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setFixedSize(100, 36)
        cancel_btn.setStyleSheet(
            "QPushButton{background:#f1f5f9;color:#334155;border:1px solid #cbd5e1;"
            "border-radius:7px;font-size:13px;font-weight:600;font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:#e2e8f0;}"
        )
        cancel_btn.clicked.connect(self.reject)

        go_btn = QtWidgets.QPushButton("Load to Invoice →")
        go_btn.setFixedSize(150, 36)
        go_btn.setStyleSheet(
            "QPushButton{background:#00756f;color:white;border:none;"
            "border-radius:7px;font-size:13px;font-weight:600;font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:#00645f;}"
        )
        go_btn.clicked.connect(self.accept)

        btn_row.addWidget(cancel_btn)
        btn_row.addSpacing(8)
        btn_row.addWidget(go_btn)
        layout.addLayout(btn_row)


def _load_local_sales_people() -> list:
    try:
        path = Path(__file__).resolve().parent / "data" / "sales_persons.json"
        if not path.exists():
            return []
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [person for person in data if isinstance(person, dict)]
    except Exception as exc:
        _log.warning("Could not load local sales people: %s", exc)
    return []


def _load_local_job_forms() -> list:
    try:
        path = Path(__file__).resolve().parent / "data" / "job_forms.json"
        if not path.exists():
            return []
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        jobs = data if isinstance(data, list) else list(data.values())
        return [job for job in jobs if isinstance(job, dict)]
    except Exception as exc:
        _log.warning("Could not load local quote forms: %s", exc)
    return []


class InvoiceItem:
    """Simple Invoice Item class for loading projects"""
    def __init__(self, project_number="", description="", plant="", quantity=1, unit_price=0.0, down_payment=0.0):
        self.project_number = project_number
        self.description = description
        self.plant = plant
        self.quantity = quantity
        self.unit_price = Decimal(str(unit_price))  # Convert to Decimal
        self.down_payment = Decimal(str(down_payment))  # Convert to Decimal
        self.total = self.quantity * self.unit_price
        self.payment_due = self.total - self.down_payment
    
    def to_dict(self) -> Dict:
        """Convert InvoiceItem to dictionary"""
        return {
            "project_number": self.project_number,
            "description": self.description,
            "plant": self.plant,
            "quantity": self.quantity,
            "unit_price": float(self.unit_price),
            "down_payment": float(self.down_payment),
            "total": float(self.total),
            "payment_due": float(self.payment_due)
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'InvoiceItem':
        """Create InvoiceItem from dictionary"""
        return cls(
            project_number=data.get("project_number", ""),
            description=data.get("description", ""),
            plant=data.get("plant", ""),
            quantity=data.get("quantity", 1),
            unit_price=data.get("unit_price", 0.0),
            down_payment=data.get("down_payment", 0.0)
        )
# Import necessary classes from main
try:
    from main import Currency, FirebaseManager, FIREBASE_AVAILABLE, Config, db
except ImportError as e:
    _log.warning("Error importing from main: %s", e)
    
    # Fallback classes
    class Currency:
        @staticmethod
        def format(value, symbol: str = "$") -> str:
            try:
                if isinstance(value, (int, float)):
                    return f"{symbol}{value:,.2f}"
                return f"{symbol}{value}"
            except:
                return f"{symbol}0.00"
    
    class FirebaseManager:
        @staticmethod
        def load_clients():
            _log.warning("⚠️ FirebaseManager not available - using empty clients list")
            return {}
        
        @staticmethod
        def load_projects():
            _log.warning("⚠️ FirebaseManager not available - using empty projects list")
            return []
        
        @staticmethod
        def save_project(project_data):
            _log.warning("⚠️ FirebaseManager not available - project not saved")
            return True
    
    FIREBASE_AVAILABLE = False
    Config = None

class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year
        self.end_year = end_year
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Navigation buttons
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton()
        self.prev_block_btn.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_ArrowBack))
        self.prev_block_btn.setIconSize(QtCore.QSize(18, 18))
        self.prev_block_btn.setFixedSize(44, 34)
        self.prev_block_btn.setToolTip("Previous years")
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                border: none;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)

        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)

        self.next_block_btn = QtWidgets.QPushButton()
        self.next_block_btn.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_ArrowForward))
        self.next_block_btn.setIconSize(QtCore.QSize(18, 18))
        self.next_block_btn.setFixedSize(44, 34)
        self.next_block_btn.setToolTip("Next years")
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                border: none;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        
        layout.addLayout(nav_layout)
        
        # Year grid container
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create 3x3 grid of year buttons
        self.year_buttons = []
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        # Current year display
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        
        layout.addLayout(current_layout)
        
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        """Calculate which 9-year block a year belongs to"""
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        """Update the 3x3 grid with years from current 9-year block"""
        years = []
        for i in range(9):
            year = self.current_block_start + i
            years.append(year)
        
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        
        current_year = datetime.now().year
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #2980b9;
                    }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        """Select a year"""
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        """Go to previous 9-year block"""
        self.current_block_start -= 9
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        """Go to next 9-year block"""
        self.current_block_start += 9
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        """Set the selected year"""
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        
    def get_selected_year(self):
        """Get the selected year"""
        return self.selected_year
    
    year_selected = QtCore.pyqtSignal(int)

class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection with unlimited years"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📅 Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        # Create YearCalendarGrid
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        # Selected year display
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        
        self.selected_label = QtWidgets.QLabel(f"")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        
        layout.addLayout(selected_layout)
        
        # Action buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Connect signals
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        """Update selected year display when year is changed in calendar"""
        self.current_year = year
    
    def on_select_clicked(self):
        """Emit signal with selected year and close popup"""
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        """Get the selected year"""
        return self.current_year

class _WheelBlocker(QtCore.QObject):
    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Wheel:
            return True
        return False

class _NoScrollComboBox(QtWidgets.QComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._wb = _WheelBlocker(self)
        self.installEventFilter(self._wb)

class _NoScrollDateEdit(QtWidgets.QDateEdit):
    def stepBy(self, steps):
        pass  # block scroll wheel and arrow-key stepping; use calendar popup to change date


class ProjectsExportDialog(QtWidgets.QDialog):
    """Professional PDF/Excel Export Dialog for Projects with Tabs"""
    
    def __init__(self, parent=None, available_dates=None):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowContextHelpButtonHint)
        self.available_dates = available_dates or []
        self.export_range = "all"
        self.selected_dates = []
        self.export_type = "pdf"
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("📊 Export Projects")
        self.setFixedSize(700, 750)
        self.setStyleSheet("""
            ProjectsExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 4, 10, 10)

        # Header
        header = QtWidgets.QLabel("📤 Export Manager - Projects")
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 8px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Export Type Tabs
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #d5dbdb;
            }
        """)
        
        # PDF Export Tab
        self.pdf_tab = QtWidgets.QWidget()
        self.setup_pdf_tab()
        self.tab_widget.addTab(self.pdf_tab, "📄 PDF Export")
        
        # Excel Export Tab
        self.excel_tab = QtWidgets.QWidget()
        self.setup_excel_tab()
        self.tab_widget.addTab(self.excel_tab, "📊 Excel Export")
        
        layout.addWidget(self.tab_widget)
        
        # Connect tab change signal
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
        # Progress Bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
    
    def setup_pdf_tab(self):
        """Setup the PDF export tab"""
        layout = QtWidgets.QVBoxLayout(self.pdf_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Projects")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)

        # From date section
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(200, 40)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)

        # To date section
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(200, 40)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)

        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month Selector
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)

        # Month and Year in two-column form layout
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(24)

        # Month column (label + combo stacked)
        month_col = QtWidgets.QVBoxLayout()
        month_col.setSpacing(6)
        month_label = QtWidgets.QLabel("Month")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_col.addWidget(month_label)
        self.month_combo = _NoScrollComboBox()
        self.month_combo.setFixedHeight(42)
        self.month_combo.setMinimumWidth(160)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: 600;
            }
            QComboBox:hover { border-color: #3498db; }
            QComboBox::drop-down { border: none; width: 24px; }
        """)
        self.populate_months()
        month_col.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_col)

        # Year column (label + [input + button] stacked)
        year_container_col = QtWidgets.QVBoxLayout()
        year_container_col.setSpacing(6)
        year_label_month = QtWidgets.QLabel("Year")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container_col.addWidget(year_label_month)
        year_container = QtWidgets.QHBoxLayout()
        year_container.setSpacing(8)

        # Year field
        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn_month.setFixedHeight(45)
        self.year_calendar_btn_month.setMinimumWidth(72)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_container.addWidget(self.year_edit_month)
        year_container.addWidget(self.year_calendar_btn_month)
        year_container_col.addLayout(year_container)
        month_year_row_layout.addLayout(year_container_col)

        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        # Year Selector
        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        # Year selection row
        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Year")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)
        
        # Year field
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        
        # Calendar button
        self.year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn.setFixedHeight(45)
        self.year_calendar_btn.setMinimumWidth(72)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)

        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        # Initially hide all date selection components
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        # Preview Section
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel("Ready to export all projects as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)

        # Connect signals for live preview updates
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.month_combo.currentTextChanged.connect(self.update_preview)
    
    def setup_excel_tab(self):
        """Setup the Excel export tab"""
        layout = QtWidgets.QVBoxLayout(self.excel_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 Excel Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.excel_all_radio = QtWidgets.QRadioButton("📋 Export All Projects")
        self.excel_all_radio.setChecked(True)
        self.excel_all_radio.toggled.connect(lambda: self.on_excel_range_changed("all"))
        
        self.excel_date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.excel_date_range_radio.toggled.connect(lambda: self.on_excel_range_changed("date_range"))
        
        self.excel_month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.excel_month_radio.toggled.connect(lambda: self.on_excel_range_changed("month"))
        
        self.excel_year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.excel_year_radio.toggled.connect(lambda: self.on_excel_range_changed("year"))
        
        options_layout.addWidget(self.excel_all_radio)
        options_layout.addWidget(self.excel_date_range_radio)
        options_layout.addWidget(self.excel_month_radio)
        options_layout.addWidget(self.excel_year_radio)
        
        range_group.addButton(self.excel_all_radio)
        range_group.addButton(self.excel_date_range_radio)
        range_group.addButton(self.excel_month_radio)
        range_group.addButton(self.excel_year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container for Excel
        self.excel_date_selection_container = QtWidgets.QWidget()
        self.excel_date_selection_layout = QtWidgets.QVBoxLayout(self.excel_date_selection_container)
        self.excel_date_selection_layout.setSpacing(15)
        self.excel_date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector for Excel
        self.excel_date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.excel_date_range_group.setMinimumHeight(120)
        self.excel_date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_date_range_layout = QtWidgets.QHBoxLayout(self.excel_date_range_group)
        excel_date_range_layout.setSpacing(20)

        # From date section
        excel_from_layout = QtWidgets.QVBoxLayout()
        excel_from_label = QtWidgets.QLabel("From Date:")
        excel_from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_from_layout.addWidget(excel_from_label)
        self.excel_from_date = _NoScrollDateEdit()
        self.excel_from_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.excel_from_date.setCalendarPopup(True)
        self.excel_from_date.setFixedSize(200, 40)
        self.excel_from_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_from_layout.addWidget(self.excel_from_date)
        excel_date_range_layout.addLayout(excel_from_layout)

        # To date section
        excel_to_layout = QtWidgets.QVBoxLayout()
        excel_to_label = QtWidgets.QLabel("To Date:")
        excel_to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_to_layout.addWidget(excel_to_label)
        self.excel_to_date = _NoScrollDateEdit()
        self.excel_to_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_to_date.setDate(QtCore.QDate.currentDate())
        self.excel_to_date.setCalendarPopup(True)
        self.excel_to_date.setFixedSize(200, 40)
        self.excel_to_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_to_layout.addWidget(self.excel_to_date)
        excel_date_range_layout.addLayout(excel_to_layout)

        excel_date_range_layout.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_date_range_group)

        # Month Selector for Excel
        self.excel_month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.excel_month_group.setMinimumHeight(150)
        self.excel_month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_month_layout = QtWidgets.QVBoxLayout(self.excel_month_group)
        excel_month_layout.setSpacing(15)

        # Month and Year in two-column form layout (matches PDF tab)
        excel_month_year_row_layout = QtWidgets.QHBoxLayout()
        excel_month_year_row_layout.setSpacing(24)

        # Month column (label + combo stacked)
        excel_month_col = QtWidgets.QVBoxLayout()
        excel_month_col.setSpacing(6)
        excel_month_label = QtWidgets.QLabel("Month")
        excel_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_month_col.addWidget(excel_month_label)
        self.excel_month_combo = _NoScrollComboBox()
        self.excel_month_combo.setFixedHeight(42)
        self.excel_month_combo.setMinimumWidth(160)
        self.excel_month_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: 600;
            }
            QComboBox:hover { border-color: #3498db; }
            QComboBox::drop-down { border: none; width: 24px; }
        """)
        self.populate_months_excel()
        excel_month_col.addWidget(self.excel_month_combo)
        excel_month_year_row_layout.addLayout(excel_month_col)

        # Year column (label + [input + button] stacked)
        excel_year_container_col = QtWidgets.QVBoxLayout()
        excel_year_container_col.setSpacing(6)
        excel_year_month_label = QtWidgets.QLabel("Year")
        excel_year_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_container_col.addWidget(excel_year_month_label)
        excel_year_inner = QtWidgets.QHBoxLayout()
        excel_year_inner.setSpacing(8)

        self.excel_year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit_month.setFixedSize(150, 45)
        self.excel_year_edit_month.setReadOnly(True)
        self.excel_year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.excel_year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn_month.setFixedHeight(45)
        self.excel_year_calendar_btn_month.setMinimumWidth(72)
        self.excel_year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month_excel)

        excel_year_inner.addWidget(self.excel_year_edit_month)
        excel_year_inner.addWidget(self.excel_year_calendar_btn_month)
        excel_year_container_col.addLayout(excel_year_inner)
        excel_month_year_row_layout.addLayout(excel_year_container_col)

        excel_month_year_row_layout.addStretch()
        excel_month_layout.addLayout(excel_month_year_row_layout)
        self.excel_date_selection_layout.addWidget(self.excel_month_group)

        # Year Selector for Excel
        self.excel_year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.excel_year_group.setMinimumHeight(120)
        self.excel_year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_year_layout = QtWidgets.QVBoxLayout(self.excel_year_group)
        excel_year_layout.setSpacing(15)

        # Year selection row (matches PDF tab layout)
        excel_year_row_layout = QtWidgets.QHBoxLayout()
        excel_year_label = QtWidgets.QLabel("Year")
        excel_year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_row_layout.addWidget(excel_year_label)

        self.excel_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit.setFixedSize(150, 45)
        self.excel_year_edit.setReadOnly(True)
        self.excel_year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.excel_year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn.setFixedHeight(45)
        self.excel_year_calendar_btn.setMinimumWidth(72)
        self.excel_year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn.clicked.connect(self.show_year_popup_excel)

        excel_year_row_layout.addWidget(self.excel_year_edit)
        excel_year_row_layout.addWidget(self.excel_year_calendar_btn)
        excel_year_row_layout.addStretch()
        excel_year_layout.addLayout(excel_year_row_layout)

        self.excel_date_selection_layout.addWidget(self.excel_year_group)

        layout.addWidget(self.excel_date_selection_container)

        # Initially hide all date selection components for Excel
        self.excel_date_selection_container.setVisible(False)
        self.excel_date_range_group.setVisible(False)
        self.excel_month_group.setVisible(False)
        self.excel_year_group.setVisible(False)
        
        # Preview Section for Excel
        excel_preview_card = QtWidgets.QGroupBox("👁️ Excel Export Preview")
        excel_preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        excel_preview_layout = QtWidgets.QVBoxLayout(excel_preview_card)
        
        self.excel_preview_label = QtWidgets.QLabel("Ready to export all projects as Excel")
        self.excel_preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.excel_preview_label.setWordWrap(True)
        excel_preview_layout.addWidget(self.excel_preview_label)
        
        layout.addWidget(excel_preview_card)

        # Connect signals for live preview updates for Excel
        self.excel_from_date.dateChanged.connect(self.update_excel_preview)
        self.excel_to_date.dateChanged.connect(self.update_excel_preview)
        self.excel_month_combo.currentTextChanged.connect(self.update_excel_preview)
    
    def show_year_popup(self):
        """Show year calendar popup for PDF year selection"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show year calendar popup for PDF month+year selection"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_excel(self):
        """Show year calendar popup for Excel year selection"""
        try:
            current_year = int(self.excel_year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month_excel(self):
        """Show year calendar popup for Excel month+year selection"""
        try:
            current_year = int(self.excel_year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for PDF year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for PDF month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_excel(self, year):
        """Handle year selection from popup for Excel year export"""
        self.excel_year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def on_year_selected_for_month_excel(self, year):
        """Handle year selection from popup for Excel month+year export"""
        self.excel_year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def populate_months(self):
        """Populate months combo box for PDF"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)

    def populate_months_excel(self):
        """Populate months combo box for Excel"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.excel_month_combo.addItems(months)
        self.excel_month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_tab_changed(self, index):
        """Handle tab changes"""
        if index == 0:  # PDF tab
            self.export_type = "pdf"
            self.export_btn.setText("🚀 Export PDF")
            self.update_preview()
        elif index == 1:  # Excel tab
            self.export_type = "excel"
            self.export_btn.setText("🚀 Export Excel")
            self.update_excel_preview()
    
    def on_range_changed(self, range_type):
        """Handle export range changes for PDF"""
        self.export_range = range_type
        
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        
        self.date_selection_container.setVisible(range_type != "all")
        self.update_preview()

    def on_excel_range_changed(self, range_type):
        """Handle export range changes for Excel"""
        self.excel_export_range = range_type
        
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        self.excel_date_range_group.setVisible(date_range_visible)
        self.excel_month_group.setVisible(month_visible)
        self.excel_year_group.setVisible(year_visible)
        
        self.excel_date_selection_container.setVisible(range_type != "all")
        self.update_excel_preview()
    
    def update_preview(self):
        """Update the PDF preview text"""
        if self.export_range == "all":
            self.preview_label.setText("📋 Will export ALL projects as PDF")
        
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toString("MM/dd/yyyy")
            to_date = self.to_date.date().toString("MM/dd/yyyy")
            self.preview_label.setText(f"📅 Will export projects from {from_date} to {to_date} as PDF")
        
        elif self.export_range == "month":
            month = self.month_combo.currentText()
            year = self.year_edit_month.text()
            self.preview_label.setText(f"🗓️ Will export projects for {month} {year} as PDF")
        
        elif self.export_range == "year":
            year = self.year_edit.text()
            self.preview_label.setText(f"📊 Will export projects for the year {year} as PDF")

    def update_excel_preview(self):
        """Update the Excel preview text"""
        if hasattr(self, 'excel_export_range'):
            range_type = self.excel_export_range
        else:
            range_type = "all"
        
        if range_type == "all":
            self.excel_preview_label.setText("📋 Will export ALL projects as Excel")
        
        elif range_type == "date_range":
            from_date = self.excel_from_date.date().toString("MM/dd/yyyy")
            to_date = self.excel_to_date.date().toString("MM/dd/yyyy")
            self.excel_preview_label.setText(f"📅 Will export projects from {from_date} to {to_date} as Excel")
        
        elif range_type == "month":
            month = self.excel_month_combo.currentText()
            year = self.excel_year_edit_month.text()
            self.excel_preview_label.setText(f"🗓️ Will export projects for {month} {year} as Excel")
        
        elif range_type == "year":
            year = self.excel_year_edit.text()
            self.excel_preview_label.setText(f"📊 Will export projects for the year {year} as Excel")
        
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_type == "pdf":
            if self.export_range == "all":
                return {"range": "all", "type": "pdf"}
            
            elif self.export_range == "date_range":
                from_date = self.from_date.date().toPyDate()
                to_date = self.to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "pdf"}
            
            elif self.export_range == "month":
                month = self.month_combo.currentIndex() + 1
                year = int(self.year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "pdf"}
            
            elif self.export_range == "year":
                year = int(self.year_edit.text())
                return {"range": "year", "year": year, "type": "pdf"}
        
        elif self.export_type == "excel":
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            
            if range_type == "all":
                return {"range": "all", "type": "excel"}
            
            elif range_type == "date_range":
                from_date = self.excel_from_date.date().toPyDate()
                to_date = self.excel_to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "excel"}
            
            elif range_type == "month":
                month = self.excel_month_combo.currentIndex() + 1
                year = int(self.excel_year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "excel"}
            
            elif range_type == "year":
                year = int(self.excel_year_edit.text())
                return {"range": "year", "year": year, "type": "excel"}
    
    def start_export(self):
        """Start the export process based on selected type"""
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
            
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            self._export_params = export_params
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False

from PyQt5 import QtWidgets, QtCore, QtGui
from pathlib import Path
from datetime import datetime
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Import necessary classes from main
try:
    from main import Currency, FirebaseManager, FIREBASE_AVAILABLE, db
except ImportError:
    # Fallback classes if main is not available
    class Currency:
        @staticmethod
        def format(value, symbol: str = "$") -> str:
            try:
                if isinstance(value, (int, float)):
                    return f"{symbol}{value:,.2f}"
                return f"{symbol}{value}"
            except:
                return f"{symbol}0.00"
    
    class FirebaseManager:
        @staticmethod
        def load_clients():
            _log.warning("⚠️ FirebaseManager not available - using empty clients list")
            return {}
        
        @staticmethod
        def load_projects():
            _log.warning("⚠️ FirebaseManager not available - using empty projects list")
            return []
        
        @staticmethod
        def save_project(project_data):
            _log.warning("⚠️ FirebaseManager not available - project not saved")
            return True
    
    FIREBASE_AVAILABLE = False
    db = None

def _norm_stage_simple(s: str) -> str:
    """Lightweight stage normaliser for module-level use (no self required)."""
    lo = (s or "").strip().lower()
    if any(x in lo for x in ("1st", "first", "deposit", "down", "term 1")):
        return "first"
    if any(x in lo for x in ("2nd", "second", "term 2")):
        return "second"
    if any(x in lo for x in ("3rd", "third", "term 3")):
        return "third"
    if any(x in lo for x in ("4th", "fourth", "final", "term 4")):
        return "final"
    if any(x in lo for x in ("full amount", "full payment", "due payment")):
        return "full_amount"
    return lo


def is_project_fully_paid(project: dict) -> bool:
    """Return True only when:
    1. Total amount paid >= project total (100 %), AND
    2. Every planned stage (with amount > 0) has at least one matching payment.
    Both conditions must be met to prevent a single over-payment on one stage
    from falsely marking the project as Paid before all stages are settled.
    """
    try:
        total_amount = float(project.get("project_amount", 0) or 0)
        if total_amount == 0:
            return False
        from payment_tracker import get_payment_tracker
        tracker = get_payment_tracker()
        if not tracker:
            return False

        pn = project.get("project_number", "")
        summary = tracker.get_payment_summary(pn, total_amount)

        # ── Condition 1: 100 % amount reached ────────────────────────────
        if summary.get("payment_percentage", 0) < 100.0:
            return False

        # ── Condition 2: every stage has at least one payment ─────────────
        planned_stages = _project_invoice_stage_plan(project)
        # Single-stage projects only need the amount check
        if len(planned_stages) <= 1:
            return True

        payments = [
            p for p in summary.get("payments", [])
            if (p.payment_stage or "").strip().lower() != "tax"
        ]
        paid_stage_keys = {_norm_stage_simple(p.payment_stage) for p in payments}

        for stage_info in planned_stages:
            stage_name   = stage_info.get("stage", "")
            stage_amount = float(stage_info.get("amount") or 0)
            if stage_amount <= 0:
                continue  # skip zero-value stages
            if _norm_stage_simple(stage_name) not in paid_stage_keys:
                return False  # this stage has no payment yet

        return True

    except Exception as e:
        _log.error("Error checking if project is fully paid: %s", e)
        return False


def update_project_status_on_full_payment(project_number: str, project_data: dict) -> bool:
    """Set project status to 'Paid' when 100% of the project amount has been received.
    Does NOT override a terminal status the user has already set manually."""
    try:
        from main import FirebaseManager
        if is_project_fully_paid(project_data):
            current = (project_data.get('status') or '').strip()
            # Respect any status the user set manually — never auto-override it.
            if project_data.get('status_manual'):
                return False
            # Also protect terminal statuses regardless of manual flag.
            if current in ('Paid', 'Cancelled', 'On Hold', 'Completed Not Invoiced', 'Completed & Invoiced'):
                return False
            project_data['status'] = 'Paid'
            project_data['completion_date'] = datetime.now().isoformat()
            if FIREBASE_AVAILABLE:
                try:
                    from firebase_admin import db as _db
                    ref = _db.reference('/projects')
                    data = ref.order_by_child('project_number').equal_to(project_number).get()
                    if data:
                        pid = list(data.keys())[0]
                        _db.reference(f'/projects/{pid}').update({
                            'status': 'Paid',
                            'completion_date': project_data['completion_date'],
                            'updated_at': datetime.now().isoformat(),
                        })
                except Exception as fe:
                    _log.warning("Firebase status update failed: %s", fe)
            from balance_sheet_tab import BalanceSheetFirebaseManager
            BalanceSheetFirebaseManager.update_balance_sheet_on_project_completion(
                project_number, project_data
            )
            _log.info("Project %s marked as Paid — 100%% payment received", project_number)
            return True
        return False
    except Exception as e:
        _log.error("Error updating project status on full payment: %s", e)
        return False


def check_and_update_project_completion(project_number: str, project_data: dict) -> bool:
    """Check project payment completion and update status if fully paid"""
    if update_project_status_on_full_payment(project_number, project_data):
        _log.info(f"Project {project_number} status updated to Completed")
        return True
    return False


class ProjectNumberGeneratorTab(QtWidgets.QWidget):
    """Project Number Generator Tab - UI similar to JobFormTab with enhanced features"""
    
    PROJECT_STATUSES = ["Not Started", "In Progress", "On Hold", "Completed Not Invoiced", "Completed & Invoiced", "Paid", "Cancelled"]
    COL_SNO = 0
    COL_PROJECT_NUMBER = 1
    COL_PROJECT_NAME = 2
    COL_CLIENT = 3
    COL_PO_WO = 4
    COL_RECEIVED = 5
    COL_PLANT = 6
    COL_SALES = 7
    COL_PRICE = 8
    COL_PAID = 9
    COL_REMAINING = 10
    COL_PAYMENT = 11
    COL_START_DATE = 12
    COL_DUE_DATE = 13
    COL_STATUS = 14
    COL_ACTIONS = 15

    # Emitted from background thread after Firebase invoice statuses are updated;
    # Qt queued-connection dispatches the handler safely on the main thread.
    _invoice_sync_done = pyqtSignal()

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self._invoice_sync_done.connect(self._on_invoice_sync_done)
        self.generated_projects = []
        self._last_clicked_row = None
        self.cached_projects = []
        self.selected_projects = set()
        self.project_being_edited = None
        self.selected_client_filter = "All Clients"
        self.selected_plant_filter = "All Plants"
        self.revenue_label = None
        self._view_mode = "all"          # "active" | "completed" | "all"
        self._period_mode = "all"       # "month" | "year" | "all"

        # Initialize UI
        self.init_ui()
        
        # Load data after UI is initialized
        QtCore.QTimer.singleShot(100, self.initial_data_load)
    
    def initial_data_load(self):
        """Load initial data after UI is fully initialized"""
        self.load_projects()
        self.refresh_clients_list()
        self.update_stats()
    
    def init_ui(self):
        """Initialize the tab UI with scroll area matching JobFormTab"""
        # OUTER layout (nothing is added directly into the tab)
        outer_layout = QtWidgets.QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        # SCROLL AREA covering the entire tab
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        scroll.viewport().setStyleSheet("background: transparent;")

        # SCROLL CONTAINER (this will hold all your UI)
        container = QtWidgets.QWidget()
        container.setStyleSheet("background: transparent;")
        container.setMinimumWidth(1100)
        scroll.setWidget(container)

        # MAIN LAYOUT inside scroll area
        main_layout = QtWidgets.QVBoxLayout(container)
        main_layout.setContentsMargins(10, 14, 10, 18)
        main_layout.setSpacing(10)
        self.main_layout = main_layout
        self.inline_project_editor = None

        # Add scroll area to the tab
        outer_layout.addWidget(scroll)

        # 1. Header Section
        self.create_header_section(main_layout)
        
        # 2. Stats Section with Generate Button
        self.create_stats_section(main_layout)
        
        # 3. Projects Table Section (Below search/filter)
        self.create_projects_table_section(main_layout)
        self.create_project_workspace_section(main_layout)

        # 4. Recent Invoices Section
        self.create_recent_invoices_section(main_layout)

    def create_header_section(self, layout):
        """Create compact left-aligned professional header matching JobFormTab"""
        header_frame = QtWidgets.QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-radius: 8px;
            }
        """)
        
        # Use QHBoxLayout to have title on left and button on right
        header_layout = QtWidgets.QHBoxLayout(header_frame)
        header_layout.setContentsMargins(20, 16, 20, 16)
        
        # Left side: Title and subtitle
        text_container = QtWidgets.QWidget()
        text_layout = QtWidgets.QVBoxLayout(text_container)
        text_layout.setContentsMargins(0, 0, 0, 0)
        text_layout.setSpacing(2)
        
        title = QtWidgets.QLabel("Project Workspace")
        title.setStyleSheet("""
            QLabel {
                color: #0f172a;
                font-size: 20px;
                font-weight: 800;
                margin-bottom: 0px;
                background: transparent;
            }
        """)

        subtitle = QtWidgets.QLabel(
            "Create, track, load to invoices, and export project records."
        )
        subtitle.setStyleSheet("""
            QLabel {
                color: #53657d;
                font-size: 13px;
                margin-top: 0px;
                font-weight: 600;
                background: transparent;
            }
        """)

        text_layout.addWidget(title)
        text_layout.addWidget(subtitle)
        
        # Add text container to left
        header_layout.addWidget(text_container)
        
        # Add stretch to push button to right
        header_layout.addStretch()
        
        # Right side: Load to Invoice Button
        self.load_selected_btn = QtWidgets.QPushButton("📄 Load to Invoice")
        self.load_selected_btn.setMinimumHeight(40)
        self.load_selected_btn.setFixedSize(160, 44)
        self.load_selected_btn.setEnabled(False)  # Disabled by default
        
        # Updated style with red gradient
        self.load_selected_btn.setStyleSheet("""
            QToolTip {
                background: #ffffff;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: 600;
            }
            QPushButton {
                background: #334155;
                color: white;
                font-weight: 800;
                border-radius: 8px;
                border: none;
                padding: 8px 12px;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #1f2937;
            }
            QPushButton:pressed {
                background: #111827;
            }
            QPushButton:disabled {
                background: #d7dee8;
                color: #7d8ba1;
            }
        """)
        
        self.load_selected_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.load_selected_btn.clicked.connect(self.load_multiple_projects)
        self.load_selected_btn.setToolTip("Load selected projects to Invoice tab")
        
        header_layout.addWidget(self.load_selected_btn)
        
        self.header_frame = header_frame
        # header_frame is kept off-layout; stats_section already has the title
        header_frame.setVisible(False)
    
    
    def create_stats_section(self, layout):
        """Create the main project workspace header bar with stat cards."""
        stats_frame = QtWidgets.QFrame()
        stats_frame.setObjectName("ProjectDashboardHero")
        stats_frame.setStyleSheet("""
            QFrame#ProjectDashboardHero {
                background: transparent;
                border: none;
                border-radius: 12px;
            }
        """)

        main_layout = QtWidgets.QHBoxLayout(stats_frame)
        main_layout.setContentsMargins(20, 16, 20, 16)
        main_layout.setSpacing(16)
        main_layout.setAlignment(QtCore.Qt.AlignVCenter)

        # Left: Title stack
        section_title = QtWidgets.QLabel("Project Workspace")
        section_title.setStyleSheet("""
            QLabel {
                color: #0f172a;
                font-size: 17px;
                font-weight: 800;
                background: transparent;
                border: none;
            }
        """)
        section_hint = QtWidgets.QLabel("Create, track, and manage project records.")
        section_hint.setStyleSheet("""
            QLabel {
                color: #94a3b8;
                font-size: 12px;
                font-weight: 600;
                background: transparent;
                border: none;
            }
        """)
        title_stack = QtWidgets.QVBoxLayout()
        title_stack.setContentsMargins(0, 0, 0, 0)
        title_stack.setSpacing(2)
        title_stack.addWidget(section_title)
        title_stack.addWidget(section_hint)
        main_layout.addLayout(title_stack)

        # Separator
        sep = QtWidgets.QFrame()
        sep.setFrameShape(QtWidgets.QFrame.VLine)
        sep.setFixedWidth(1)
        sep.setFixedHeight(48)
        sep.setStyleSheet("background: #e2e8f0; border: none;")
        main_layout.addWidget(sep)

        # Center: Stat cards
        cards_layout = QtWidgets.QHBoxLayout()
        cards_layout.setSpacing(10)
        cards_layout.setAlignment(QtCore.Qt.AlignCenter)

        total_card = self.create_stat_card("Total Projects", "0", "#2563eb", "fa5s.briefcase")
        self.total_projects_label = total_card.findChild(QtWidgets.QLabel, "StatValue")
        self.total_projects_trend  = total_card.findChild(QtWidgets.QLabel, "StatTrend")
        cards_layout.addWidget(total_card)

        active_card = self.create_stat_card("Active Projects", "0", "#0ea5e9", "fa5s.bolt")
        self.active_projects_label = active_card.findChild(QtWidgets.QLabel, "StatValue")
        self.active_projects_trend  = active_card.findChild(QtWidgets.QLabel, "StatTrend")
        cards_layout.addWidget(active_card)

        completed_card = self.create_stat_card("Completed", "0", "#00756f", "fa5s.check-circle")
        self.completed_label = completed_card.findChild(QtWidgets.QLabel, "StatValue")
        self.completed_trend  = completed_card.findChild(QtWidgets.QLabel, "StatTrend")
        cards_layout.addWidget(completed_card)

        on_hold_card = self.create_stat_card("On Hold", "0", "#d97706", "fa5s.pause-circle")
        self.on_hold_projects_label = on_hold_card.findChild(QtWidgets.QLabel, "StatValue")
        self.on_hold_trend           = on_hold_card.findChild(QtWidgets.QLabel, "StatTrend")
        cards_layout.addWidget(on_hold_card)

        overdue_card = self.create_stat_card("Overdue", "0", "#ef4444", "fa5s.clock")
        self.overdue_projects_label = overdue_card.findChild(QtWidgets.QLabel, "StatValue")
        self.overdue_trend           = overdue_card.findChild(QtWidgets.QLabel, "StatTrend")
        cards_layout.addWidget(overdue_card)

        # Hidden cancelled card kept for update_stats compatibility
        cancelled_card = self.create_stat_card("Cancelled", "0", "#b45309", "fa5s.ban")
        self.cancelled_label = cancelled_card.findChild(QtWidgets.QLabel, "StatValue")
        cards_layout.addWidget(cancelled_card)
        cancelled_card.setVisible(False)

        main_layout.addLayout(cards_layout)

        # Right: Action buttons
        self.generate_project_btn = QtWidgets.QPushButton("+ New Project")
        self.generate_project_btn.setFixedSize(132, 38)
        self.generate_project_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.generate_project_btn.setStyleSheet("""
            QPushButton {
                background: #00756f;
                color: white;
                font-weight: 800;
                font-size: 13px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover { background: #00645f; }
            QPushButton:pressed { background: #00514d; }
        """)
        self.generate_project_btn.clicked.connect(self.show_project_dialog)

        self.export_btn = QtWidgets.QPushButton("Export")
        self.export_btn.setFixedSize(96, 38)
        self.export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: #ffffff;
                color: #334155;
                border: 1.5px solid #cbd5e1;
                border-radius: 8px;
                font-weight: 800;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #f1f5f9;
                border-color: #94a3b8;
            }
            QPushButton:pressed { background: #e2e8f0; }
        """)
        self.export_btn.clicked.connect(self.open_projects_pdf_export_dialog)

        self.load_selected_btn.setVisible(True)
        self.load_selected_btn.setText("Load to Invoice")
        self.load_selected_btn.setFixedSize(138, 38)

        btn_container = QtWidgets.QWidget()
        btn_container.setStyleSheet("background: transparent; border: none;")
        btn_layout = QtWidgets.QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.setSpacing(8)
        btn_layout.addWidget(self.generate_project_btn)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.load_selected_btn)

        main_layout.addWidget(btn_container, alignment=QtCore.Qt.AlignRight)
        self.stats_frame = stats_frame
        layout.addWidget(stats_frame)
    
    def refresh_clients_immediately(self):
        """Refresh the company dropdown with latest clients from Firebase"""
        try:
            _log.info("Refreshing company dropdown with latest clients...")
            
            # Get current selection
            current_selection = self.company_combo.currentText()
            
            # Load clients from Firebase
            from main import FirebaseManager
            clients = FirebaseManager.load_clients()
            
            # Clear and repopulate dropdown
            self.company_combo.clear()
            self.company_combo.addItem("-- Select Company --")
            
            # Sort and add clients
            for client_name in sorted(clients.keys()):
                self.company_combo.addItem(client_name)
            
            # Try to restore previous selection
            if current_selection and current_selection != "-- Select Company --":
                index = self.company_combo.findText(current_selection)
                if index >= 0:
                    self.company_combo.setCurrentIndex(index)
            
            _log.info("Company dropdown refreshed with %s clients", len(clients))
            
        except Exception as e:
            _log.warning("Error refreshing company dropdown: %s", e)
        
    def create_stat_card(self, title, value, color, icon=""):
        """Create a modern stat card: icon box left, value + label + trend right."""
        card = QtWidgets.QWidget()
        card.setObjectName("StatCard")
        card.setFixedSize(154, 72)
        card.setStyleSheet("""
            QWidget#StatCard {
                background: #ffffff;
                border: 1px solid #edf2f7;
                border-radius: 10px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        # Subtle drop shadow
        shadow = QtWidgets.QGraphicsDropShadowEffect(card)
        shadow.setBlurRadius(10)
        shadow.setOffset(0, 1)
        shadow.setColor(QtGui.QColor(15, 23, 42, 12))
        card.setGraphicsEffect(shadow)
        row = QtWidgets.QHBoxLayout(card)
        row.setContentsMargins(12, 9, 12, 9)
        row.setSpacing(10)

        # Icon — soft circle background, white icon inside (matches reference)
        tile_bg, tile_fg = {
            "#2563eb": ("#eff6ff", "#2563eb"),
            "#0ea5e9": ("#ecfeff", "#0891b2"),
            "#00756f": ("#ecfdf5", "#16a34a"),
            "#d97706": ("#fff7ed", "#d97706"),
            "#ef4444": ("#fff1f2", "#ef4444"),
            "#b45309": ("#fff7ed", "#b45309"),
        }.get(color, ("#f8fafc", "#64748b"))
        icon_box = QtWidgets.QLabel()
        icon_box.setObjectName("StatIcon")
        icon_box.setFixedSize(34, 34)
        icon_box.setAlignment(QtCore.Qt.AlignCenter)
        icon_box.setStyleSheet(
            f"QLabel {{ background: {tile_bg}; border-radius: 9px; border: none; }}"
        )
        _icon_emoji = {
            "fa5s.briefcase": "💼", "fa5s.bolt": "⚡", "fa5s.check-circle": "✔",
            "fa5s.pause-circle": "⏸", "fa5s.clock": "⏰", "fa5s.ban": "🚫",
        }
        if icon and "." in icon:
            try:
                import qtawesome as qta
                icon_color = "#ffffff" if color == "#ef4444" else tile_fg
                _ic = qta.icon(icon, color=icon_color)
                icon_box.setPixmap(_ic.pixmap(18, 18))
            except Exception:
                fallback = _icon_emoji.get(icon, "•")
                icon_box.setText(fallback)
                icon_box.setStyleSheet(
                    f"QLabel {{ background: {tile_bg}; border-radius: 9px; border: none; "
                    f"font-size: 15px; color: {tile_fg}; font-family: 'Segoe UI Emoji', sans-serif; }}"
                )
        else:
            icon_box.setText(icon)
            icon_box.setStyleSheet(
                f"QLabel {{ background: {tile_bg}; border-radius: 9px; border: none; font-size: 18px; color: {tile_fg}; font-family: 'Segoe UI Emoji', sans-serif; }}"
            )
        if color == "#ef4444":
            icon_box.setStyleSheet("""
                QLabel {
                    background: #fff1f2;
                    border-radius: 9px;
                    border: none;
                }
            """)
            overdue_dot = QtWidgets.QLabel(icon_box)
            overdue_dot.setFixedSize(18, 18)
            overdue_dot.move(8, 8)
            overdue_dot.setAlignment(QtCore.Qt.AlignCenter)
            overdue_dot.setStyleSheet("""
                QLabel {
                    background: #fb7185;
                    border-radius: 9px;
                    border: none;
                }
            """)
            try:
                import qtawesome as qta
                overdue_dot.setPixmap(qta.icon(icon, color="#ffffff").pixmap(11, 11))
            except Exception:
                overdue_dot.setText("!")
                overdue_dot.setStyleSheet("""
                    QLabel {
                        background: #fb7185;
                        color: #ffffff;
                        border-radius: 9px;
                        border: none;
                        font-size: 11px;
                        font-weight: 900;
                    }
                """)
        row.addWidget(icon_box, 0, QtCore.Qt.AlignVCenter)

        # Right side: value + label + trend
        right = QtWidgets.QVBoxLayout()
        right.setSpacing(0)
        right.setContentsMargins(0, 0, 0, 0)

        value_label = QtWidgets.QLabel(value)
        value_label.setObjectName("StatValue")
        value_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 900;
                color: #0f172a;
                background: transparent;
                border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        value_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

        desc_label = QtWidgets.QLabel(title)
        desc_label.setObjectName("StatDesc")
        desc_label.setStyleSheet("""
            QLabel {
                font-size: 9px;
                color: #64748b;
                font-weight: 700;
                background: transparent;
                border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        desc_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

        trend_label = QtWidgets.QLabel("")
        trend_label.setObjectName("StatTrend")
        trend_label.setStyleSheet("""
            QLabel {
                font-size: 8px;
                font-weight: 800;
                color: #22c55e;
                background: transparent;
                border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        trend_label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

        right.addWidget(value_label)
        right.addWidget(desc_label)
        right.addWidget(trend_label)
        row.addLayout(right)
        return card

    # ── Summary cards ────────────────────────────────────────────────────────
    def _build_summary_cards(self, layout):
        """Three KPI cards: This Month / This Year / Total Active."""
        row = QtWidgets.QHBoxLayout()
        row.setSpacing(12)

        def _card(title, val_attr, sub_attr, accent):
            frame = QtWidgets.QFrame()
            frame.setStyleSheet(f"""
                QFrame {{
                    background: #ffffff;
                    border: 1.5px solid #e2eaf3;
                    border-left: 4px solid {accent};
                    border-radius: 10px;
                }}
            """)
            vb = QtWidgets.QVBoxLayout(frame)
            vb.setContentsMargins(14, 10, 14, 10)
            vb.setSpacing(3)

            hdr = QtWidgets.QLabel(title.upper())
            hdr.setStyleSheet(
                "font-size:9px;font-weight:800;color:#94a3b8;"
                "letter-spacing:0.5px;background:transparent;border:none;"
            )
            val = QtWidgets.QLabel("—")
            val.setStyleSheet(
                f"font-size:22px;font-weight:900;color:{accent};"
                "background:transparent;border:none;"
            )
            sub = QtWidgets.QLabel("")
            sub.setWordWrap(True)
            sub.setStyleSheet(
                "font-size:11px;font-weight:600;color:#64748b;"
                "background:transparent;border:none;"
            )
            vb.addWidget(hdr)
            vb.addWidget(val)
            vb.addWidget(sub)
            setattr(self, val_attr, val)
            setattr(self, sub_attr, sub)
            return frame

        row.addWidget(_card("This Month",  "_sm_val",  "_sm_sub",  "#2563eb"))
        row.addWidget(_card("This Year",   "_sy_val",  "_sy_sub",  "#0f766e"))
        row.addWidget(_card("Active Jobs", "_sa_val",  "_sa_sub",  "#d97706"))
        layout.addLayout(row)

    def _refresh_summary_cards(self):
        """Recompute and redraw the three KPI summary cards."""
        from datetime import datetime as _dt
        now = _dt.now()
        cur_month, cur_year = now.month, now.year

        month_projects, month_done, month_rev = 0, 0, 0.0
        year_projects,  year_done,  year_rev  = 0, 0, 0.0
        active_count = 0

        def _is_done(status: str) -> bool:
            s = (status or "").lower()
            return any(x in s for x in ("completed", "paid"))

        def _is_active(status: str) -> bool:
            s = (status or "").lower()
            return not any(x in s for x in ("& invoiced", "cancelled", "canceled")) \
                   and "paid" not in s.split()

        for p in self.generated_projects:
            amt = self._project_total_amount(p)
            rd  = p.get("date_received", "") or p.get("created_at", "") or ""
            date_obj = None
            for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    from datetime import datetime as _dt2
                    date_obj = _dt2.strptime(rd[:10], fmt)
                    break
                except Exception:
                    pass
            if date_obj is None:
                try:
                    from datetime import datetime as _dt2
                    date_obj = _dt2.fromisoformat(rd[:10])
                except Exception:
                    pass

            status = p.get("status", "")
            is_done   = _is_done(status)
            is_active = _is_active(status)
            if is_active:
                active_count += 1

            if date_obj:
                if date_obj.month == cur_month and date_obj.year == cur_year:
                    month_projects += 1
                    if is_done:
                        month_done += 1
                    month_rev += amt
                if date_obj.year == cur_year:
                    year_projects += 1
                    if is_done:
                        year_done += 1
                    year_rev += amt

        if hasattr(self, "_sm_val"):
            self._sm_val.setText(str(month_projects))
            self._sm_sub.setText(f"{month_done} done  •  ${month_rev:,.0f}")
        if hasattr(self, "_sy_val"):
            self._sy_val.setText(str(year_projects))
            self._sy_sub.setText(f"{year_done} done  •  ${year_rev:,.0f}")
        if hasattr(self, "_sa_val"):
            self._sa_val.setText(str(active_count))
            self._sa_sub.setText("In Progress / On Hold")

    # ── View tabs + quick period filters ────────────────────────────────────
    def _build_view_tabs(self, layout):
        """Active | Completed | All tabs  +  This Month | This Year | All Time quick filters."""
        bar = QtWidgets.QHBoxLayout()
        bar.setSpacing(0)

        # --- View tabs (left) ---
        tab_frame = QtWidgets.QFrame()
        tab_frame.setStyleSheet("""
            QFrame {
                background: #f1f5f9;
                border-radius: 8px;
                border: 1px solid #e2e8f0;
            }
        """)
        tab_row = QtWidgets.QHBoxLayout(tab_frame)
        tab_row.setContentsMargins(4, 4, 4, 4)
        tab_row.setSpacing(4)

        self._view_btns = {}
        for label, key in [("Active", "active"), ("Completed", "completed"), ("All", "all")]:
            btn = QtWidgets.QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(32)
            btn.setMinimumWidth(110)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            btn.setStyleSheet("""
                QPushButton {
                    background: transparent;
                    color: #64748b;
                    border: none;
                    border-radius: 6px;
                    font-size: 12px;
                    font-weight: 800;
                    padding: 0 12px;
                }
                QPushButton:checked {
                    background: #ffffff;
                    color: #0f172a;
                    border: 1px solid #d1d5db;
                }
                QPushButton:hover:!checked { color: #0f172a; }
            """)
            btn.clicked.connect(lambda _, k=key: self._set_view_mode(k))
            tab_row.addWidget(btn)
            self._view_btns[key] = btn

        self._view_btns["active"].setChecked(True)
        bar.addWidget(tab_frame)
        bar.addStretch()

        # --- Quick period filters (right) ---
        period_frame = QtWidgets.QFrame()
        period_frame.setStyleSheet(tab_frame.styleSheet())
        period_row = QtWidgets.QHBoxLayout(period_frame)
        period_row.setContentsMargins(4, 4, 4, 4)
        period_row.setSpacing(4)

        self._period_btns = {}
        now = datetime.now()
        for label, key in [
            (f"{now.strftime('%B')}", "month"),
            (f"{now.year}", "year"),
            ("All Time", "all"),
        ]:
            btn = QtWidgets.QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(32)
            btn.setMinimumWidth(82)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            btn.setStyleSheet("""
                QPushButton {
                    background: transparent;
                    color: #64748b;
                    border: none;
                    border-radius: 6px;
                    font-size: 12px;
                    font-weight: 800;
                    padding: 0 12px;
                }
                QPushButton:checked {
                    background: #ffffff;
                    color: #0f766e;
                    border: 1px solid #d1d5db;
                }
                QPushButton:hover:!checked { color: #0f172a; }
            """)
            btn.clicked.connect(lambda _, k=key: self._set_period_mode(k))
            period_row.addWidget(btn)
            self._period_btns[key] = btn

        self._period_btns["all"].setChecked(True)
        bar.addWidget(period_frame)
        layout.addLayout(bar)

    def _set_view_mode(self, mode: str):
        self._view_mode = mode
        self._page_num = 1
        for k, btn in self._view_btns.items():
            btn.setChecked(k == mode)
        self.filter_projects()

    def _set_period_mode(self, mode: str):
        self._period_mode = mode
        for k, btn in self._period_btns.items():
            btn.setChecked(k == mode)
        self.filter_projects()

    def create_projects_table_section(self, layout):
        """Create projects table section with modern filter bar and table."""
        table_group = QtWidgets.QGroupBox()
        self.table_group = table_group
        table_group.setTitle("")
        table_group.setStyleSheet(self.get_group_box_style())
        table_layout = QtWidgets.QVBoxLayout(table_group)
        table_layout.setContentsMargins(10, 8, 10, 8)
        table_layout.setSpacing(10)

        # Initialize view/period mode attributes (no separate tab row)
        self._view_btns = {}
        self._period_btns = {}
        if not hasattr(self, "_view_mode"):
            self._view_mode = "all"
        if not hasattr(self, "_period_mode"):
            self._period_mode = "all"

        # ── Filter bar ────────────────────────────────────────────────────────
        search_filter_frame = QtWidgets.QFrame()
        search_filter_frame.setObjectName("ProjectFilterBar")
        search_filter_frame.setStyleSheet("""
            QFrame#ProjectFilterBar {
                background: transparent;
                border: none;
                border-radius: 9px;
            }
        """)
        search_filter_layout = QtWidgets.QHBoxLayout(search_filter_frame)
        search_filter_layout.setSpacing(10)
        search_filter_layout.setContentsMargins(12, 8, 12, 8)

        # Search input
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("Search by project #, name, client, amount...")
        self.search_edit.setMinimumHeight(36)
        self.search_edit.setMinimumWidth(260)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 6px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                font-size: 13px;
                background: #f8fafc;
            }
            QLineEdit:focus { border-color: #00756f; background: #ffffff; }
        """)
        self.search_edit.textChanged.connect(self.search_projects)
        search_filter_layout.addWidget(self.search_edit)

        # Status label + combo
        from app_theme import CHEVRON_URL
        _combo_style = f"""
            QComboBox {{
                padding: 5px 22px 5px 10px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                color: #0f172a;
                font-size: 12px;
                font-weight: 700;
                min-width: 140px;
            }}
            QComboBox:focus {{ border-color: #00756f; }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px;
                border: none;
                background: transparent;
            }}
            QComboBox::down-arrow {{
                image: url("{CHEVRON_URL}");
                width: 13px; height: 13px; margin-right: 3px;
            }}
        """

        status_label = QtWidgets.QLabel("Status:")
        status_label.setStyleSheet("font-weight: 700; color: #64748b; font-size: 12px; background: transparent; border: none;")
        search_filter_layout.addWidget(status_label)

        self.status_filter_combo = QtWidgets.QComboBox()
        self.status_filter_combo.setMinimumHeight(36)
        self.status_filter_combo.setStyleSheet(_combo_style)
        self.status_filter_combo.currentTextChanged.connect(self.on_status_filter_changed)
        self.status_filter_combo.wheelEvent = lambda e: e.ignore()
        self.status_filter_combo.keyPressEvent = lambda e, c=self.status_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.status_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.status_filter_combo.clearFocus))
        search_filter_layout.addWidget(self.status_filter_combo)

        # Client filter button (dropdown menu)
        self.client_filter_button = QtWidgets.QPushButton("Client")
        self.client_filter_button.setFixedHeight(36)
        self.client_filter_button.setMinimumWidth(86)
        self.client_filter_button.setMaximumWidth(300)
        self.client_filter_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.client_filter_button.setStyleSheet("""
            QPushButton {
                background: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                font-size: 12px;
                font-weight: 700;
                color: #334155;
                padding: 0 12px;
            }
            QPushButton:hover { color: #00756f; border-color: #00756f; }
        """)
        self.client_filter_menu = QtWidgets.QMenu(self)
        self.client_filter_menu.setStyleSheet("""
            QMenu {
                background: white; border: 1px solid #cfd4da;
                border-radius: 8px; padding: 6px; margin-top: 4px;
            }
            QMenu::item { padding: 7px 14px; border-radius: 6px; font-size: 13px; }
            QMenu::item:selected { background: #e6f6f4; color: #0f172a; }
        """)
        self.client_filter_button.clicked.connect(
            lambda: self.client_filter_menu.exec_(
                self.client_filter_button.mapToGlobal(
                    QtCore.QPoint(0, self.client_filter_button.height() + 4)
                )
            )
        )
        search_filter_layout.addWidget(self.client_filter_button)

        # Plant filter — plain QComboBox, same style as Status filter
        plant_label = QtWidgets.QLabel("Plant:")
        plant_label.setStyleSheet(
            "font-weight: 700; color: #64748b; font-size: 12px;"
            " background: transparent; border: none;"
        )
        search_filter_layout.addWidget(plant_label)

        self.plant_filter_combo = QtWidgets.QComboBox()
        self.plant_filter_combo.addItem("All Plants")
        self.plant_filter_combo.setMinimumHeight(36)
        self.plant_filter_combo.setMinimumWidth(160)
        self.plant_filter_combo.setStyleSheet(_combo_style)
        self.plant_filter_combo.currentTextChanged.connect(self.on_plant_filter_changed)
        self.plant_filter_combo.wheelEvent = lambda e: e.ignore()
        self.plant_filter_combo.keyPressEvent = lambda e, c=self.plant_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.plant_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.plant_filter_combo.clearFocus))
        search_filter_layout.addWidget(self.plant_filter_combo)

        # Date range button
        self.date_range_button = configure_filter_button(QtWidgets.QPushButton(), height=36)
        self.date_range_button.clicked.connect(self.show_date_range_dialog)
        search_filter_layout.addWidget(self.date_range_button)

        search_filter_layout.addStretch()

        # Compact view toggle: All | Active | Completed
        view_frame = QtWidgets.QFrame()
        view_frame.setStyleSheet("""
            QFrame { background: #f1f5f9; border-radius: 7px; border: 1px solid #e2e8f0; }
        """)
        view_row = QtWidgets.QHBoxLayout(view_frame)
        view_row.setContentsMargins(3, 3, 3, 3)
        view_row.setSpacing(2)
        for label, key in [("All", "all"), ("Active", "active"), ("Done", "completed")]:
            btn = QtWidgets.QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(28)
            btn.setMinimumWidth(52)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            btn.setStyleSheet("""
                QPushButton {
                    background: transparent; color: #64748b; border: none;
                    border-radius: 5px; font-size: 11px; font-weight: 700; padding: 0 8px;
                }
                QPushButton:checked { background: #ffffff; color: #0f172a; border: 1px solid #d1d5db; }
                QPushButton:hover:!checked { color: #0f172a; }
            """)
            btn.clicked.connect(lambda _, k=key: self._set_view_mode(k))
            view_row.addWidget(btn)
            self._view_btns[key] = btn
        self._view_btns["all"].setChecked(True)
        search_filter_layout.addWidget(view_frame)

        table_layout.addWidget(search_filter_frame)

        # Results info label
        self.results_label = QtWidgets.QLabel("")
        self.results_label.setMinimumWidth(180)
        self.results_label.setStyleSheet(
            "color:#94a3b8; font-size:11px; font-weight:700;"
            " background:transparent; border:none; padding:0 2px;")
        table_layout.addWidget(self.results_label)

        # Table shell
        table_shell = QtWidgets.QFrame()
        table_shell.setObjectName("ProjectGridFrame")
        table_shell.setStyleSheet("QFrame#ProjectGridFrame { background: transparent; border: none; }")
        table_shell_layout = QtWidgets.QVBoxLayout(table_shell)
        table_shell_layout.setContentsMargins(0, 0, 0, 0)
        table_shell_layout.setSpacing(0)
        table_shell_layout.setAlignment(QtCore.Qt.AlignTop)

        # Projects table
        self.projects_table = QtWidgets.QTableWidget()
        self.projects_table.setColumnCount(16)
        self.projects_table.setHorizontalHeaderLabels([
            "#", "Project #", "Project Name", "Client", "PO/WO",
            "Received", "Plant", "Sales", "Price", "Paid",
            "Remaining", "Progress", "Start Date", "Due Date", "Status", "Actions"
        ])
        self.projects_table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                font-size: 12px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                gridline-color: #f1f5f9;
                outline: none;
            }
            QTableWidget::item {
                padding: 6px 10px;
                border-bottom: 1px solid #f1f5f9;
                color: #111827;
            }
            QTableWidget::item:selected {
                background: #e6f6f4;
                color: #0f172a;
            }
            QTableWidget::item:hover { background: #f8fafc; }
            QHeaderView::section {
                background: #f8fafc;
                color: #64748b;
                font-weight: 800;
                font-size: 11px;
                padding: 10px 8px;
                border: none;
                border-right: 1px solid #e2e8f0;
                border-bottom: 1.5px solid #e2e8f0;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                letter-spacing: 0.3px;
                text-transform: uppercase;
            }
            QTableWidget::item:alternate { background: #fafbfc; }
            QToolTip {
                background: #ffffff;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 11px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        self.projects_table.setMinimumHeight(132)
        self.projects_table.setMaximumHeight(600)
        self.projects_table.setShowGrid(False)
        self.projects_table.horizontalHeader().setStretchLastSection(False)
        self.projects_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.projects_table.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.projects_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.projects_table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.projects_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.projects_table.setWordWrap(False)
        self.projects_table.setAlternatingRowColors(False)
        self.projects_table.verticalHeader().setVisible(False)
        self.projects_table.cellClicked.connect(self.on_project_row_clicked)
        self.projects_table.itemSelectionChanged.connect(self.on_project_selected)
        self.projects_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.projects_table.customContextMenuRequested.connect(self._on_projects_table_context_menu)

        header = self.projects_table.horizontalHeader()
        header.setDefaultAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)
        header.setHighlightSections(False)
        header.setFixedHeight(40)
        header.setMinimumSectionSize(80)
        for col in range(self.projects_table.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Interactive)

        # Hidden columns
        self.projects_table.setColumnHidden(self.COL_SNO,        True)
        self.projects_table.setColumnHidden(self.COL_PO_WO,      True)
        self.projects_table.setColumnHidden(self.COL_RECEIVED,   True)
        self.projects_table.setColumnHidden(self.COL_PRICE,      True)
        self.projects_table.setColumnHidden(self.COL_PAID,       True)
        self.projects_table.setColumnHidden(self.COL_REMAINING,  True)
        self.projects_table.setColumnHidden(self.COL_START_DATE, True)
        self.projects_table.setColumnHidden(self.COL_DUE_DATE,   False)
        self.projects_table.setColumnHidden(self.COL_PLANT,      True)

        # Fixed columns; PROJECT_NAME stretch fills remainder
        self.projects_table.setColumnWidth(self.COL_PROJECT_NUMBER, 195)
        self.projects_table.setColumnWidth(self.COL_PROJECT_NAME,   140)
        self.projects_table.setColumnWidth(self.COL_CLIENT,         230)
        self.projects_table.setColumnWidth(self.COL_SALES,          215)
        self.projects_table.setColumnWidth(self.COL_DUE_DATE,       120)
        self.projects_table.setColumnWidth(self.COL_PAYMENT,        225)
        self.projects_table.setColumnWidth(self.COL_STATUS,         210)
        self.projects_table.setColumnWidth(self.COL_ACTIONS,        120)

        header.setSectionResizeMode(self.COL_PROJECT_NUMBER, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_PROJECT_NAME,   QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(self.COL_CLIENT,         QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_SALES,          QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_DUE_DATE,       QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_PAYMENT,        QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_STATUS,         QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(self.COL_ACTIONS,        QtWidgets.QHeaderView.Fixed)

        table_shell_layout.addWidget(self.projects_table)
        table_layout.addWidget(table_shell)

        # ── Pagination bar ────────────────────────────────────────────────────
        self._page_num = 1
        self._page_size = 10

        pagination_frame = QtWidgets.QFrame()
        pagination_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        pagination_hbox = QtWidgets.QHBoxLayout(pagination_frame)
        pagination_hbox.setContentsMargins(4, 6, 4, 2)
        pagination_hbox.setSpacing(6)

        self._pagination_info_label = QtWidgets.QLabel("")
        self._pagination_info_label.setStyleSheet(
            "color: #94a3b8; font-size: 11px; font-weight: 600;"
            " background: transparent; border: none;")
        pagination_hbox.addWidget(self._pagination_info_label)
        pagination_hbox.addStretch()

        _page_btn_style = """
            QPushButton {
                background: #ffffff; color: #334155;
                border: 1px solid #e2e8f0; border-radius: 6px;
                font-size: 12px; font-weight: 700;
                min-width: 32px; min-height: 28px; padding: 0 8px;
            }
            QPushButton:hover { background: #f1f5f9; border-color: #cbd5e1; }
            QPushButton:disabled { color: #cbd5e1; }
            QPushButton[active="true"] {
                background: #00756f; color: white; border-color: #00756f;
            }
        """
        self._prev_page_btn = QtWidgets.QPushButton("‹")
        self._prev_page_btn.setStyleSheet(_page_btn_style)
        self._prev_page_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._prev_page_btn.clicked.connect(self._go_prev_page)
        pagination_hbox.addWidget(self._prev_page_btn)

        self._page_btns_container = QtWidgets.QHBoxLayout()
        self._page_btns_container.setSpacing(4)
        pagination_hbox.addLayout(self._page_btns_container)

        self._next_page_btn = QtWidgets.QPushButton("›")
        self._next_page_btn.setStyleSheet(_page_btn_style)
        self._next_page_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._next_page_btn.clicked.connect(self._go_next_page)
        pagination_hbox.addWidget(self._next_page_btn)

        self._page_btn_style_ref = _page_btn_style
        table_layout.addWidget(pagination_frame)

        layout.addWidget(table_group)

        # Connect selection changed signal
        self.projects_table.selectionModel().selectionChanged.connect(self.on_project_selected)

    def adjust_projects_table_height(self):
        """Keep the grid compact and anchored to the top as row counts change."""
        if not hasattr(self, "projects_table"):
            return
        row_count = max(1, self.projects_table.rowCount())
        header_height = self.projects_table.horizontalHeader().height()
        row_height = self.projects_table.verticalHeader().defaultSectionSize()
        target_height = header_height + (row_count * row_height) + 16
        target_height = max(132, min(target_height, 760))
        self.projects_table.setFixedHeight(target_height)

    # ── Pagination helpers ───────────────────────────────────────────────────
    def _go_prev_page(self):
        if getattr(self, "_page_num", 1) > 1:
            self._page_num -= 1
            self.filter_projects()

    def _go_next_page(self):
        total = getattr(self, "_paginated_total", 0)
        page_size = getattr(self, "_page_size", 10)
        import math
        max_page = max(1, math.ceil(total / page_size))
        if getattr(self, "_page_num", 1) < max_page:
            self._page_num += 1
            self.filter_projects()

    def _rebuild_pagination(self, total_count):
        """Rebuild the page number buttons and update info label."""
        import math
        if not hasattr(self, "_page_btns_container"):
            return
        page_size = getattr(self, "_page_size", 10)
        page_num  = getattr(self, "_page_num", 1)
        max_page  = max(1, math.ceil(total_count / page_size))
        self._paginated_total = total_count

        # Info text
        start = (page_num - 1) * page_size + 1 if total_count else 0
        end   = min(page_num * page_size, total_count)
        if hasattr(self, "_pagination_info_label"):
            self._pagination_info_label.setText(
                f"Showing {start}–{end} of {total_count} projects"
            )

        # Rebuild page buttons
        while self._page_btns_container.count():
            item = self._page_btns_container.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        _s = getattr(self, "_page_btn_style_ref", "")
        _win_start = max(1, min(page_num, max_page - 2))
        for p in range(_win_start, min(_win_start + 3, max_page + 1)):
            btn = QtWidgets.QPushButton(str(p))
            btn.setFixedSize(32, 28)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            if p == page_num:
                btn.setStyleSheet("""QPushButton {
                    background-color: #00756f; color: #ffffff;
                    border: 1px solid #00756f; border-radius: 6px;
                    font-size: 12px; font-weight: 700;
                    min-width: 32px; min-height: 28px; padding: 0 8px;
                }
                QPushButton:hover { background-color: #005f5a; color: #ffffff; }""")
            else:
                btn.setStyleSheet(_s)
                btn.clicked.connect(lambda _, pg=p: self._go_to_page(pg))
            self._page_btns_container.addWidget(btn)

        if hasattr(self, "_prev_page_btn"):
            self._prev_page_btn.setEnabled(page_num > 1)
        if hasattr(self, "_next_page_btn"):
            self._next_page_btn.setEnabled(page_num < max_page)

    def _go_to_page(self, page):
        self._page_num = page
        self.filter_projects()

    def create_project_workspace_section(self, layout):
        self.project_workspace_frame = QtWidgets.QFrame()
        self.project_workspace_frame.setObjectName("ProjectWorkspaceFrame")
        self.project_workspace_frame.setStyleSheet("""
            QFrame#ProjectWorkspaceFrame {
                background: transparent;
                border: none;
            }
        """)
        self.project_workspace_layout = QtWidgets.QVBoxLayout(self.project_workspace_frame)
        self.project_workspace_layout.setContentsMargins(18, 16, 18, 18)
        self.project_workspace_layout.setSpacing(14)
        self.project_workspace_frame.setVisible(False)
        layout.addWidget(self.project_workspace_frame)

    def create_recent_invoices_section(self, layout):
        """Bottom section: latest 4 invoices as cards."""
        self._recent_invoices_frame = QtWidgets.QFrame()
        self._recent_invoices_frame.setObjectName("RecentInvoicesFrame")
        self._recent_invoices_frame.setStyleSheet("""
            QFrame#RecentInvoicesFrame {
                background: transparent;
                border: none;
            }
        """)
        outer = QtWidgets.QVBoxLayout(self._recent_invoices_frame)
        outer.setContentsMargins(0, 8, 0, 0)
        outer.setSpacing(10)

        # Header row
        header_row = QtWidgets.QHBoxLayout()
        title_lbl = QtWidgets.QLabel("Recent Invoices")
        title_lbl.setStyleSheet("""
            font-size: 15px; font-weight: 800; color: #0f172a;
            background: transparent; border: none;
            font-family: 'Inter', 'Segoe UI', sans-serif;
        """)
        sub_lbl = QtWidgets.QLabel("Latest invoices across all projects.")
        sub_lbl.setStyleSheet("""
            font-size: 11px; font-weight: 600; color: #94a3b8;
            background: transparent; border: none;
            font-family: 'Inter', 'Segoe UI', sans-serif;
        """)
        title_stack = QtWidgets.QVBoxLayout()
        title_stack.setSpacing(1)
        title_stack.setContentsMargins(0, 0, 0, 0)
        title_stack.addWidget(title_lbl)
        title_stack.addWidget(sub_lbl)
        header_row.addLayout(title_stack)
        header_row.addStretch()

        view_all_btn = QtWidgets.QPushButton("View All Invoices")
        view_all_btn.setFixedHeight(34)
        view_all_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        view_all_btn.setStyleSheet("""
            QPushButton {
                background: #ffffff; color: #334155;
                border: 1.5px solid #e2e8f0; border-radius: 8px;
                font-size: 12px; font-weight: 700; padding: 0 16px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background: #f8fafc; border-color: #cbd5e1; }
        """)
        view_all_btn.clicked.connect(self._on_view_all_invoices)
        header_row.addWidget(view_all_btn)
        outer.addLayout(header_row)

        # Cards row
        self._recent_invoice_cards_row = QtWidgets.QHBoxLayout()
        self._recent_invoice_cards_row.setSpacing(12)
        outer.addLayout(self._recent_invoice_cards_row)

        layout.addWidget(self._recent_invoices_frame)

        # Populate on first render (deferred so Firebase is ready)
        QtCore.QTimer.singleShot(600, self.refresh_recent_invoices)

    def refresh_recent_invoices(self):
        """Load latest 4 invoices and rebuild the cards row."""
        if not hasattr(self, "_recent_invoice_cards_row"):
            return
        # Clear existing cards
        while self._recent_invoice_cards_row.count():
            item = self._recent_invoice_cards_row.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        try:
            from main import FirebaseManager
            invoices = FirebaseManager.load_invoices() or []
        except Exception:
            invoices = []

        # Load live status overrides from Invoice History's status cache
        try:
            from main import Config, FileManager
            _cache_file = Config.INVOICES_DIR / "status_cache.json"
            _status_cache = FileManager.load_json(_cache_file, {}) if _cache_file.exists() else {}
        except Exception:
            _status_cache = {}

        # Sort by invoice date descending, take 4 (exclude deleted invoices)
        def _inv_date(inv):
            meta = inv.get("meta", {}) if isinstance(inv, dict) else {}
            raw = meta.get("date", "") or meta.get("invoice_date", "") or ""
            for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y"):
                try:
                    from datetime import datetime as _dt
                    return _dt.strptime(raw[:10], fmt)
                except Exception:
                    pass
            try:
                from datetime import datetime as _dt
                return _dt.fromisoformat(raw[:19])
            except Exception:
                pass
            return None

        def _is_deleted(inv):
            meta = inv.get("meta", {}) if isinstance(inv, dict) else {}
            status = (meta.get("status", "") or "").strip().lower()
            return bool(meta.get("deleted")) or status == "deleted"

        dated = [(inv, _inv_date(inv)) for inv in invoices if not _is_deleted(inv)]
        dated = [(inv, d) for inv, d in dated if d is not None]
        dated.sort(key=lambda x: x[1], reverse=True)
        recent = [inv for inv, _ in dated[:4]]

        if not recent:
            placeholder = QtWidgets.QLabel("No invoices found.")
            placeholder.setStyleSheet("color:#94a3b8; font-size:12px; background:transparent; border:none;")
            self._recent_invoice_cards_row.addWidget(placeholder)
            return

        _status_styles = {
            "paid":          ("#dcfce7", "#15803d"),
            "unpaid":        ("#fee2e2", "#dc2626"),
            "partially paid":("#fef9c3", "#b45309"),
            "pending":       ("#f1f5f9", "#64748b"),
        }

        for inv in recent:
            meta = inv.get("meta", {}) if isinstance(inv, dict) else {}
            inv_num  = meta.get("invoice_number", "INV-????")
            # Prefer live status from Invoice History's status cache over Firebase meta
            status   = (_status_cache.get(inv_num) or meta.get("status", "Pending") or "Pending").strip()
            due_date = meta.get("due_date", meta.get("date", "—")) or "—"

            # Project name from first item
            items = inv.get("items", [])
            proj_name = ""
            total_amt = 0.0
            for it in items:
                if not proj_name:
                    proj_name = it.get("project_name", "") or it.get("description", "")
                try:
                    total_amt += float(it.get("payment_due") or it.get("total") or it.get("unit_price") or 0)
                except Exception:
                    pass

            status_key = status.lower()
            bg_s, fg_s = _status_styles.get(status_key, ("#f1f5f9", "#64748b"))

            # Build card
            card = QtWidgets.QFrame()
            card.setStyleSheet("""
                QFrame {
                    background: #ffffff;
                    border: 1px solid #e2e8f0;
                    border-radius: 10px;
                }
                QFrame:hover {
                    border-color: #00756f;
                    background: #f0fdf4;
                }
            """)
            card.setMinimumWidth(200)
            card.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            # Click → navigate to Invoice History and filter by this invoice number
            _inv_num_cap = inv_num
            _client_name_cap = (meta.get("client_name", "") or "").strip()
            def _open_invoice_in_history(_checked=False, _num=_inv_num_cap, _client=_client_name_cap, _mw=self.main_window):
                try:
                    if hasattr(_mw, "_nav_to"):
                        _mw._nav_to(2)
                    if hasattr(_mw, "_switch_inner_tab"):
                        _mw._switch_inner_tab(3)
                    elif hasattr(_mw, "project_invoice_inner_tabs"):
                        _mw.project_invoice_inner_tabs.setCurrentIndex(3)
                    hist = getattr(_mw, "history_tab", None)
                    if hist is None:
                        return
                    if _client and hasattr(hist, "show_invoice_history"):
                        hist.show_invoice_history(_client)
                    def _set_search():
                        try:
                            cw = hist.stacked_widget.currentWidget()
                            if hasattr(cw, "date_range_widget"):
                                cw.date_range_widget.search_bar.setText(_num)
                        except Exception:
                            pass
                    QtCore.QTimer.singleShot(150, _set_search)
                except Exception:
                    pass
            card.mousePressEvent = lambda event, fn=_open_invoice_in_history: fn()
            cvbox = QtWidgets.QVBoxLayout(card)
            cvbox.setContentsMargins(14, 12, 14, 12)
            cvbox.setSpacing(6)

            # Top: invoice # + status badge
            top_row = QtWidgets.QHBoxLayout()
            inv_lbl = QtWidgets.QLabel(inv_num)
            inv_lbl.setStyleSheet("""
                font-size: 13px; font-weight: 800; color: #0f172a;
                background: transparent; border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            """)
            inv_lbl.setToolTip("Right-click to copy invoice number")
            inv_lbl.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            def _make_inv_copy_menu(label=inv_lbl, num=inv_num):
                def _show(pos):
                    menu = QtWidgets.QMenu(label)
                    menu.setStyleSheet("""
                        QMenu { background: #ffffff; border: 1px solid #e2e8f0;
                                border-radius: 7px; padding: 4px; }
                        QMenu::item { padding: 7px 20px; font-size: 13px; color: #0f172a;
                                      font-family: 'Inter','Segoe UI'; border-radius: 5px; }
                        QMenu::item:selected { background: #e6f6f4; color: #00756f; }
                    """)
                    copy_action = menu.addAction("Copy Invoice Number")
                    action = menu.exec_(label.mapToGlobal(pos))
                    if action == copy_action:
                        QtWidgets.QApplication.clipboard().setText(num)
                return _show
            inv_lbl.customContextMenuRequested.connect(_make_inv_copy_menu())
            status_badge = QtWidgets.QLabel(status)
            status_badge.setStyleSheet(f"""
                background: {bg_s}; color: {fg_s};
                border-radius: 5px; padding: 2px 8px;
                font-size: 10px; font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                border: none;
            """)
            top_row.addWidget(inv_lbl)
            top_row.addStretch()
            top_row.addWidget(status_badge)
            cvbox.addLayout(top_row)

            # Project name subtitle
            proj_lbl = QtWidgets.QLabel(proj_name or "—")
            proj_lbl.setStyleSheet("""
                font-size: 11px; font-weight: 600; color: #64748b;
                background: transparent; border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            """)
            proj_lbl.setWordWrap(True)
            cvbox.addWidget(proj_lbl)

            # Divider
            div = QtWidgets.QFrame()
            div.setFrameShape(QtWidgets.QFrame.HLine)
            div.setStyleSheet("background: #f1f5f9; border: none; max-height: 1px;")
            cvbox.addWidget(div)

            # Bottom: Total + Due date
            bottom_row = QtWidgets.QHBoxLayout()
            total_col = QtWidgets.QVBoxLayout()
            total_col.setSpacing(1)
            QtWidgets.QLabel("Total").setStyleSheet("")
            t_hdr = QtWidgets.QLabel("Total")
            t_hdr.setStyleSheet("font-size:10px; color:#94a3b8; font-weight:600; background:transparent; border:none;")
            t_val = QtWidgets.QLabel(f"${total_amt:,.2f}")
            t_val.setStyleSheet("font-size:13px; font-weight:800; color:#0f172a; background:transparent; border:none;")
            total_col.addWidget(t_hdr)
            total_col.addWidget(t_val)

            due_col = QtWidgets.QVBoxLayout()
            due_col.setSpacing(1)
            d_hdr = QtWidgets.QLabel("Due Date")
            d_hdr.setStyleSheet("font-size:10px; color:#94a3b8; font-weight:600; background:transparent; border:none;")
            d_val = QtWidgets.QLabel(due_date)
            d_val.setStyleSheet("font-size:12px; font-weight:700; color:#334155; background:transparent; border:none;")
            due_col.addWidget(d_hdr)
            due_col.addWidget(d_val)

            bottom_row.addLayout(total_col)
            bottom_row.addStretch()
            bottom_row.addLayout(due_col)
            cvbox.addLayout(bottom_row)

            self._recent_invoice_cards_row.addWidget(card, 1)

    def _on_view_all_invoices(self):
        """Switch to Invoice History inner tab and show the client list view."""
        try:
            mw = self.main_window
            if hasattr(mw, "_nav_to"):
                mw._nav_to(2)           # outer tab: Projects & Invoice
            if hasattr(mw, "_switch_inner_tab"):
                mw._switch_inner_tab(3) # inner tab index 3: Invoice History
            elif hasattr(mw, "project_invoice_inner_tabs"):
                mw.project_invoice_inner_tabs.setCurrentIndex(3)
            # Always show the client list (first view), not a previously-open client history
            hist = getattr(mw, "history_tab", None)
            if hist and hasattr(hist, "show_client_view"):
                hist.show_client_view()
        except Exception:
            pass

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            child_layout = item.layout()
            if widget is not None:
                widget.deleteLater()
            elif child_layout is not None:
                self._clear_layout(child_layout)

    def _workspace_metric_card(self, title, value, color="#0f172a", accent="#e2e8f0"):
        card = QtWidgets.QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: #f8fafc;
                border: 1px solid {accent};
                border-radius: 9px;
            }}
        """)
        layout = QtWidgets.QVBoxLayout(card)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(4)

        title_label = QtWidgets.QLabel(title)
        title_label.setStyleSheet("font-size: 12px; font-weight: 800; color: #64748b;")
        value_label = QtWidgets.QLabel(str(value))
        value_label.setStyleSheet(f"font-size: 18px; font-weight: 900; color: {color};")
        value_label.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        return card

    def _workspace_table_item(self, value, *, align=QtCore.Qt.AlignCenter, bold=False, color="#0f172a"):
        item = QtWidgets.QTableWidgetItem(str(value))
        item.setTextAlignment(align | QtCore.Qt.AlignVCenter)
        item.setToolTip(str(value))
        item.setForeground(QtGui.QColor(color))
        item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold if bold else QtGui.QFont.Normal))
        return item

    def _prepare_workspace_table(self, table):
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                gridline-color: #edf2f7;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 8px 10px;
                color: #0f172a;
            }
            QHeaderView::section {
                background: #f8fafc;
                color: #334155;
                font-weight: 900;
                padding: 9px 10px;
                border: none;
                border-right: 1px solid #e2e8f0;
                border-bottom: 1px solid #d8e2ec;
            }
            QToolTip {
                background: #ffffff;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 11px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)

    def _parse_payment_split_percentages(self, project_data):
        raw = (
            project_data.get("payment_split_percentages")
            or project_data.get("custom_payment_percentages")
            or project_data.get("payment_split")
            or ""
        )
        if isinstance(raw, (list, tuple)):
            values = raw
        else:
            values = re.findall(r"\d+(?:\.\d+)?", str(raw))

        percentages = []
        for value in values:
            try:
                number = float(value)
                if number > 0:
                    percentages.append(number)
            except (TypeError, ValueError):
                continue

        percentages = percentages[:4]
        current_total = sum(percentages)
        if 0 < current_total < 100 and len(percentages) < 4:
            percentages.append(100 - current_total)
        if current_total > 100:
            percentages = [(p / current_total) * 100 for p in percentages]
        return percentages

    def _planned_payment_rows(self, project_data):
        total_amount = self._project_total_amount(project_data)
        planned = _project_invoice_stage_plan(project_data)
        rows = []
        for row in planned:
            amount = float(row.get("amount", 0) or 0)
            rows.append({
                "stage": row.get("stage", ItemRowWidget.FULL_AMOUNT_LABEL),
                "percent": (amount / total_amount * 100) if total_amount else 0,
                "amount": amount,
            })
        return rows

    def _stage_matches(self, left, right):
        left = str(left or "").lower()
        right = str(right or "").lower()
        if not left or not right:
            return False
        normalized = {
            # "payment 1" / "1st installment" do NOT include "deposit"/"down payment" —
            # those are a separate stage in multi-stage payment plans.
            "payment 1": ("payment 1", "term 1", "1st", "first"),
            "1st installment": ("payment 1", "term 1", "1st installment", "1st", "first"),
            "payment 2": ("payment 2", "term 2", "2nd", "second"),
            "2nd installment": ("payment 2", "term 2", "2nd installment", "2nd", "second"),
            "payment 3": ("payment 3", "term 3", "3rd", "third"),
            "3rd installment": ("payment 3", "term 3", "3rd installment", "3rd", "third"),
            "final payment": ("final", "payment 4", "term 4", "4th", "balance"),
            "full amount": ("full", "due", "full amount"),
            # "deposit" and "down payment" only match each other, not numbered installments
            "deposit": ("deposit", "down payment"),
            "down payment": ("deposit", "down payment"),
        }
        left_terms = normalized.get(left, (left,))
        return any(term in right for term in left_terms) or left in right or right in left

    def _payment_plan_rows_for_workspace(self, project_data):
        project_number = project_data.get("project_number", "")
        invoice_rows = self._get_project_invoice_rows(project_number)
        payments = get_payment_tracker().get_project_payments(project_number)
        planned_rows = self._planned_payment_rows(project_data)
        rows = []
        matched_payment_ids = set()

        # Tax payments are shown as a separate row after all project stages —
        # exclude them from stage matching and the unassigned bucket entirely.
        tax_payment_ids = {
            p.payment_id for p in payments
            if (p.payment_stage or "").strip().lower() == "tax"
        }
        tax_payments_list = [p for p in payments if p.payment_id in tax_payment_ids]

        for planned in planned_rows:
            stage = planned["stage"]
            matched_invoices = [
                row for row in invoice_rows
                if self._stage_matches(stage, row.get("stage", ""))
            ]
            if len(planned_rows) == 1:
                matched_payments = [p for p in payments if p.payment_id not in tax_payment_ids]
            else:
                # Invoice-number fallback: payments auto-recorded from an invoice that
                # belongs to THIS stage should be counted even when the stored
                # payment_stage string doesn't match (e.g. "Installment 3" vs "3rd Installment").
                stage_invoice_numbers = {
                    row.get("invoice_number", "")
                    for row in matched_invoices
                    if row.get("invoice_number")
                }
                other_stage_names = [r["stage"] for r in planned_rows if r["stage"] != stage]
                def _payment_belongs_here(p, _stage=stage,
                                          _inv_nums=stage_invoice_numbers,
                                          _others=other_stage_names):
                    # Primary: stage name matches
                    if self._stage_matches(_stage, p.payment_stage):
                        return True
                    # Fallback: payment linked to an invoice for this stage AND its
                    # payment_stage does NOT better match a different planned stage
                    p_inv = (p.invoice_number or "").strip()
                    if p_inv and p_inv in _inv_nums:
                        for other in _others:
                            if self._stage_matches(other, p.payment_stage):
                                return False  # belongs to another stage
                        return True
                    return False
                matched_payments = [
                    p for p in payments
                    if p.payment_id not in tax_payment_ids
                    and _payment_belongs_here(p)
                ]
            for payment in matched_payments:
                matched_payment_ids.add(payment.payment_id)

            paid = sum(float(payment.amount) for payment in matched_payments)
            planned_amount = planned["amount"]
            # Derive status from actual payments first, fall back to invoice record
            if paid >= planned_amount > 0:
                invoice_status = "Paid"
            elif paid > 0:
                invoice_status = "Partially Paid"
            elif matched_invoices:
                raw_status = matched_invoices[0].get("status", "Invoice Created")
                # Don't inherit invoice-level payment statuses when this project has
                # no payments — the invoice status may reflect payments for other
                # projects on the same multi-project invoice.
                if raw_status in ("Partially Paid", "Paid"):
                    invoice_status = "Invoice Created"
                else:
                    invoice_status = raw_status
            else:
                invoice_status = "Not Invoiced"
            rows.append({
                "stage": stage,
                "percent": planned["percent"],
                "amount": planned_amount,
                "invoice_status": invoice_status,
                "paid": paid,
                "remaining": max(planned_amount - paid, 0),
            })

        # Only truly unmatched, non-Tax payments appear as "Unassigned Payments"
        unassigned = [
            p for p in payments
            if p.payment_id not in matched_payment_ids
            and p.payment_id not in tax_payment_ids
        ]
        if unassigned and len(planned_rows) > 1:
            paid = sum(float(payment.amount) for payment in unassigned)
            rows.append({
                "stage": "Unassigned Payments",
                "percent": 0,
                "amount": 0,
                "invoice_status": "Recorded",
                "paid": paid,
                "remaining": 0,
            })

        return rows

    def _set_list_view_visible(self, visible):
        for attr in ("stats_frame", "table_group"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setVisible(visible)
        if hasattr(self, "project_workspace_frame"):
            self.project_workspace_frame.setVisible(not visible)

    def hide_project_workspace(self):
        self._set_list_view_visible(True)
        self.filter_projects()

    def show_project_workspace(self, project_data):
        if not project_data:
            return
        self._ws_project_data = project_data  # store for live refresh
        self._clear_layout(self.project_workspace_layout)
        self._set_list_view_visible(False)

        project_number = project_data.get("project_number", "")
        project_name = project_data.get("project_name", "")
        client_name = project_data.get("company", "")
        status = project_data.get("status", "Not Started")
        summary = self._project_payment_summary(project_data)
        total_amount = float(summary.get("total_amount", 0) or 0)
        paid_amount = float(summary.get("total_paid", 0) or 0)
        remaining_amount = max(float(summary.get("remaining", 0) or 0), 0.0)
        pct = int((paid_amount / total_amount * 100) if total_amount > 0 else 0)

        # ── HEADER ───────────────────────────────────────────────────────────
        header = QtWidgets.QFrame()
        header.setStyleSheet(
            "QFrame { background:transparent; border:none; border-radius:10px; }")
        h_lay = QtWidgets.QHBoxLayout(header)
        h_lay.setContentsMargins(16, 12, 16, 12)
        h_lay.setSpacing(12)

        back_btn = QtWidgets.QPushButton("← Back")
        back_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        back_btn.setFixedSize(90, 36)
        back_btn.setStyleSheet("""
            QPushButton { background:#1e293b; color:#ffffff;
                border:none; border-radius:7px;
                font-weight:700; font-size:13px; }
            QPushButton:hover { background:#0f172a; }
        """)
        back_btn.clicked.connect(self.hide_project_workspace)
        h_lay.addWidget(back_btn)

        vdiv = QtWidgets.QFrame()
        vdiv.setFrameShape(QtWidgets.QFrame.VLine)
        vdiv.setFixedWidth(1)
        vdiv.setStyleSheet("background:#e2e8f0; border:none;")
        h_lay.addWidget(vdiv)

        pn_badge = QtWidgets.QLabel(project_number)
        pn_badge.setStyleSheet("""
            background:#0f766e; color:#ffffff;
            font-family:'Consolas','Courier New',monospace;
            font-size:13px; font-weight:800;
            border:none; border-radius:6px;
            padding:5px 12px;
        """)
        h_lay.addWidget(pn_badge)

        name_col = QtWidgets.QVBoxLayout()
        name_col.setSpacing(2)
        title_lbl = QtWidgets.QLabel(project_name)
        title_lbl.setStyleSheet(
            "font-size:18px; font-weight:800; color:#0f172a; background:transparent;")
        sub_parts = [p for p in [
            client_name,
            f"PO/WO: {project_data.get('po_wo_number','')}" if project_data.get('po_wo_number') else None,
            project_data.get('plant',''),
            project_data.get('sales',''),
        ] if p]
        sub_lbl = QtWidgets.QLabel("  ·  ".join(sub_parts))
        sub_lbl.setStyleSheet(
            "font-size:12px; color:#64748b; font-weight:500; background:transparent;")
        name_col.addWidget(title_lbl)
        name_col.addWidget(sub_lbl)
        h_lay.addLayout(name_col, 1)

        # Use live status from generated_projects so workspace always reflects current state
        live_project = next(
            (p for p in self.generated_projects if p.get('project_number') == project_number),
            project_data
        )
        status = live_project.get("status", "Not Started")

        status_combo = QtWidgets.QComboBox()
        status_combo.setObjectName("projectStatusCombo")
        status_combo.addItems(self.PROJECT_STATUSES)
        idx = status_combo.findText(status)
        if idx >= 0:
            status_combo.setCurrentIndex(idx)
        self.style_project_status_combo(status_combo, status)
        status_combo.setFixedHeight(36)
        status_combo.setMinimumWidth(160)
        status_combo.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        status_combo.currentTextChanged.connect(
            lambda new_st, pd=live_project, c=status_combo:
                self.on_project_status_changed(pd, new_st, c)
        )
        # Store reference so project list status changes can sync back
        self._workspace_status_combo = status_combo
        self._workspace_project_data = live_project
        h_lay.addWidget(status_combo)

        for label, bg, hover, fg in [
            ("+ Invoice", "#0f766e", "#0d625c", "#fff"),
            ("+ Payment", "#2563eb", "#1d4ed8", "#fff"),
            ("✎ Edit",    "#f1f5f9", "#e2e8f0",  "#334155"),
        ]:
            b = QtWidgets.QPushButton(label)
            b.setFixedHeight(34)
            b.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            b.setStyleSheet(f"""
                QPushButton {{ background:{bg}; color:{fg}; border:none;
                    border-radius:7px; font-weight:700; font-size:12px; padding:0 14px; }}
                QPushButton:hover {{ background:{hover}; }}
            """)
            h_lay.addWidget(b)

        # wire button slots after loop
        for btn, slot in zip(
            [h_lay.itemAt(h_lay.count()-3).widget(),
             h_lay.itemAt(h_lay.count()-2).widget(),
             h_lay.itemAt(h_lay.count()-1).widget()],
            [lambda c=False: self.load_projects_to_invoice_direct([project_data]),
             lambda c=False: self._open_workspace_payment_history(project_data),
             lambda c=False: self.edit_single_project(project_data)]
        ):
            btn.clicked.connect(slot)

        self.project_workspace_layout.addWidget(header)

        # ── SUMMARY CARDS ────────────────────────────────────────────────────
        cards_row = QtWidgets.QHBoxLayout()
        cards_row.setSpacing(10)

        def _card(label, value, value_color="#0f172a"):
            f = QtWidgets.QFrame()
            f.setStyleSheet(
                "QFrame { background:#ffffff; border:1px solid #e2e8f0; border-radius:10px; }")
            lay = QtWidgets.QVBoxLayout(f)
            lay.setContentsMargins(18, 14, 18, 14)
            lay.setSpacing(4)
            lbl = QtWidgets.QLabel(label)
            lbl.setStyleSheet("font-size:11px; font-weight:600; color:#94a3b8; border:none; background:transparent;")
            val = QtWidgets.QLabel(str(value))
            val.setStyleSheet(f"font-size:22px; font-weight:800; color:{value_color}; border:none; background:transparent;")
            lay.addWidget(lbl)
            lay.addWidget(val)
            return f

        # Progress card
        prog_f = QtWidgets.QFrame()
        prog_f.setStyleSheet(
            "QFrame { background:#ffffff; border:1px solid #e2e8f0; border-radius:10px; }")
        pf_lay = QtWidgets.QVBoxLayout(prog_f)
        pf_lay.setContentsMargins(18, 14, 18, 14)
        pf_lay.setSpacing(6)
        pf_lbl = QtWidgets.QLabel("Payment Progress")
        pf_lbl.setStyleSheet("font-size:11px; font-weight:600; color:#94a3b8; border:none; background:transparent;")
        pf_val = QtWidgets.QLabel(f"{pct}% Paid")
        pf_val.setStyleSheet("font-size:22px; font-weight:800; color:#0f766e; border:none; background:transparent;")
        prog_bar = QtWidgets.QProgressBar()
        prog_bar.setValue(pct)
        prog_bar.setFixedHeight(8)
        prog_bar.setTextVisible(False)
        prog_bar.setStyleSheet("""
            QProgressBar { background:#e2e8f0; border-radius:4px; border:none; }
            QProgressBar::chunk { background:#0f766e; border-radius:4px; }
        """)
        pf_lay.addWidget(pf_lbl)
        pf_lay.addWidget(pf_val)
        pf_lay.addWidget(prog_bar)

        cards_row.addWidget(_card("Project Total",
            self._format_project_money(total_amount)))

        cards_row.addWidget(_card("Amount Paid",
            self._format_project_money(paid_amount), "#047857"))
        cards_row.addWidget(_card("Remaining",
            self._format_project_money(remaining_amount),
            "#dc2626" if remaining_amount else "#047857"))
        cards_row.addWidget(prog_f, 1)
        self.project_workspace_layout.addLayout(cards_row)

        # ── TABS ─────────────────────────────────────────────────────────────
        tabs = QtWidgets.QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: transparent;
                top: -1px;
            }
            QTabBar { background: transparent; }
            QTabBar::tab {
                background: transparent; color: #64748b;
                padding: 11px 24px; border: none;
                border-bottom: 2px solid transparent;
                font-weight: 600; font-size: 13px; min-width: 90px;
            }
            QTabBar::tab:selected { color: #0f766e; border-bottom: 2px solid #0f766e; font-weight: 800; }
            QTabBar::tab:hover:!selected { color: #334155; }
        """)
        tabs.addTab(self._create_workspace_overview_tab(project_data), "  Overview  ")
        tabs.addTab(self._create_workspace_invoices_tab(project_data), "  Invoices  ")
        tabs.addTab(self._create_workspace_payments_tab(project_data), "  Payments  ")
        self.project_workspace_layout.addWidget(tabs)

    def _open_workspace_payment_history(self, project_data):
        self.show_payment_history(project_data)
        # NOTE: _refresh_finance_tabs() removed — same race condition as open_payment_dialog.
        # Annual summary is event-driven via _trigger_annual_summary_refresh in _bg_update.
        # Refresh workspace after short delay so Firebase write completes before re-render.
        # Reload payments from disk so any tax/stage entries written by background threads
        # are included in the plan before the workspace is re-rendered.
        def _ws_refresh_after_payment(_pd=project_data):
            try:
                get_payment_tracker()._load_payments()
                self.show_project_workspace(_pd)
            except Exception:
                pass
        QtCore.QTimer.singleShot(600, _ws_refresh_after_payment)

    def _create_workspace_overview_tab(self, project_data):
        tab = QtWidgets.QWidget()
        tab.setStyleSheet("background: transparent;")
        layout = QtWidgets.QVBoxLayout(tab)
        layout.setContentsMargins(0, 12, 0, 12)
        layout.setSpacing(14)

        # ── Payment Plan section ─────────────────────────────────────────────
        plan_card = QtWidgets.QFrame()
        plan_card.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-radius: 12px;
            }
        """)
        plan_lay = QtWidgets.QVBoxLayout(plan_card)
        plan_lay.setContentsMargins(16, 14, 16, 14)
        plan_lay.setSpacing(10)

        plan_hdr = QtWidgets.QHBoxLayout()
        plan_title = QtWidgets.QLabel("Payment Plan")
        plan_title.setStyleSheet("font-size:14px; font-weight:800; color:#0f172a;")
        plan_hdr.addWidget(plan_title)
        plan_hdr.addStretch()
        plan_lay.addLayout(plan_hdr)

        table = QtWidgets.QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Stage", "%", "Planned Amount", "Invoice Status", "Paid", "Remaining"])
        self._prepare_workspace_table(table)
        rows = self._payment_plan_rows_for_workspace(project_data)
        table.setRowCount(max(len(rows), 1))
        table.setRowHeight(0, 42)

        INV_STATUS_STYLE = {
            "Paid":           ("#065f46", "#d1fae5", "#6ee7b7"),
            "Partially Paid": ("#78350f", "#fef3c7", "#fcd34d"),
            "Unpaid":         ("#92400e", "#fef3c7", "#fcd34d"),
            "Not Invoiced":   ("#374151", "#f3f4f6", "#d1d5db"),
            "Not Created":    ("#374151", "#f3f4f6", "#d1d5db"),
            "Invoice Created":("#1e40af", "#dbeafe", "#93c5fd"),
            "Pending":        ("#78350f", "#fef3c7", "#fcd34d"),
            "Overdue":        ("#7f1d1d", "#fee2e2", "#fca5a5"),
        }

        for row_index, row in enumerate(rows):
            table.setRowHeight(row_index, 42)
            values = [
                row["stage"],
                f"{row['percent']:.0f}%" if row["percent"] else "—",
                self._format_project_money(row["amount"]),
                row["invoice_status"],
                self._format_project_money(row["paid"]),
                self._format_project_money(row["remaining"]),
            ]
            for col, value in enumerate(values):
                if col == 3:
                    # Colored badge widget for invoice status
                    sc = INV_STATUS_STYLE.get(value, ("#374151", "#f3f4f6", "#d1d5db"))
                    cell_w = QtWidgets.QWidget()
                    cell_w.setStyleSheet("background: transparent;")
                    cell_l = QtWidgets.QHBoxLayout(cell_w)
                    cell_l.setContentsMargins(6, 0, 6, 0)
                    cell_l.setAlignment(QtCore.Qt.AlignCenter)
                    badge = QtWidgets.QLabel(f"  {value}  ")
                    badge.setAlignment(QtCore.Qt.AlignCenter)
                    badge.setStyleSheet(f"""
                        color: {sc[0]}; background: {sc[1]};
                        border: 1px solid {sc[2]};
                        border-radius: 8px; font-size: 11px; font-weight: 800;
                        padding: 3px 6px;
                    """)
                    cell_l.addWidget(badge)
                    table.setCellWidget(row_index, col, cell_w)
                else:
                    color = "#dc2626" if col == 5 and row["remaining"] > 0 else "#0f172a"
                    if col == 4 and row["paid"] > 0:
                        color = "#047857"
                    item = self._workspace_table_item(value, bold=col in (0, 2, 4, 5), color=color)
                    table.setItem(row_index, col, item)

        hdr = table.horizontalHeader()
        hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        for col in range(1, 6):
            hdr.setSectionResizeMode(col, QtWidgets.QHeaderView.Fixed)
        table.setColumnWidth(1, 95)
        table.setColumnWidth(2, 140)
        table.setColumnWidth(3, 155)
        table.setColumnWidth(4, 120)
        table.setColumnWidth(5, 130)
        row_count = max(len(rows), 1)
        table.setFixedHeight(hdr.sizeHint().height() + row_count * 42 + 4)
        plan_lay.addWidget(table)
        layout.addWidget(plan_card)

        # ── Project Information card ─────────────────────────────────────────
        info_card = QtWidgets.QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """)
        info_lay = QtWidgets.QVBoxLayout(info_card)
        info_lay.setContentsMargins(16, 14, 16, 16)
        info_lay.setSpacing(12)

        info_hdr = QtWidgets.QHBoxLayout()
        info_title = QtWidgets.QLabel("Project Information")
        info_title.setStyleSheet("font-size:14px; font-weight:800; color:#0f172a;")
        info_hdr.addWidget(info_title)
        info_hdr.addStretch()
        info_lay.addLayout(info_hdr)

        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(10)
        info_fields = [
            ("Project #",        project_data.get("project_number", "—")),
            ("Project Name",     project_data.get("project_name", "—")),
            ("Client",           project_data.get("company", "—")),
            ("PO / WO",          project_data.get("po_wo_number", "—")),
            ("Date Received",    project_data.get("date_received", "—")),
            ("Plant",            project_data.get("plant", "—")),
            ("Sales",            project_data.get("sales", "—")),
            ("Payment Category", project_data.get("payment_category", "—")),
        ]
        for idx, (lbl, val) in enumerate(info_fields):
            row_i = idx // 2
            col_i = idx % 2

            field_frame = QtWidgets.QFrame()
            field_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
            f_lay = QtWidgets.QVBoxLayout(field_frame)
            f_lay.setContentsMargins(10, 6, 10, 6)
            f_lay.setSpacing(2)

            lbl_w = QtWidgets.QLabel(lbl.upper())
            lbl_w.setStyleSheet(
                "font-size:10px; font-weight:700; color:#94a3b8; letter-spacing:0.8px; background:transparent; border:none;")
            val_w = QtWidgets.QLabel(str(val or "—"))
            val_w.setStyleSheet("font-size:13px; font-weight:700; color:#0f172a; background:transparent; border:none;")
            val_w.setWordWrap(True)

            f_lay.addWidget(lbl_w)
            f_lay.addWidget(val_w)
            grid.addWidget(field_frame, row_i, col_i)

        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        info_lay.addLayout(grid)
        layout.addWidget(info_card)
        layout.addStretch()
        return tab

    def _create_workspace_invoices_tab(self, project_data):
        tab = QtWidgets.QWidget()
        tab.setStyleSheet("background: transparent;")
        layout = QtWidgets.QVBoxLayout(tab)
        layout.setContentsMargins(0, 12, 0, 12)
        layout.setSpacing(12)

        card = QtWidgets.QFrame()
        card.setStyleSheet("QFrame { background:transparent; border:none; border-radius:10px; }")
        card_lay = QtWidgets.QVBoxLayout(card)
        card_lay.setContentsMargins(0, 0, 0, 0)
        card_lay.setSpacing(10)

        hdr_lbl = QtWidgets.QLabel("Invoices")
        hdr_lbl.setStyleSheet("font-size:15px; font-weight:900; color:#0f172a;")
        card_lay.addWidget(hdr_lbl)

        table = QtWidgets.QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Invoice #", "Date", "Stage", "Amount", "Status", "Received Date"])
        self._prepare_workspace_table(table)

        INV_STATUS_STYLE = {
            "Paid":           ("#065f46", "#d1fae5", "#6ee7b7"),
            "Partially Paid": ("#78350f", "#fef3c7", "#fcd34d"),
            "Unpaid":         ("#7f1d1d", "#fee2e2", "#fca5a5"),
            "Pending":        ("#78350f", "#fef3c7", "#fcd34d"),
            "Overdue":        ("#7f1d1d", "#fee2e2", "#fca5a5"),
            "Sent":           ("#1e40af", "#dbeafe", "#93c5fd"),
            "Draft":          ("#374151", "#f3f4f6", "#d1d5db"),
        }

        rows = self._get_project_invoice_rows(project_data.get("project_number", ""))
        table.setRowCount(max(len(rows), 1))
        if rows:
            for row_index, row in enumerate(rows):
                table.setRowHeight(row_index, 42)
                values = [
                    row.get("invoice_number", "N/A"),
                    row.get("date", "N/A"),
                    row.get("stage", "N/A"),
                    row.get("amount", "$0.00"),
                    row.get("status", "Unpaid"),
                    row.get("received_date", "N/A"),
                ]
                for col, value in enumerate(values):
                    if col == 4:
                        sc = INV_STATUS_STYLE.get(value, ("#374151", "#f3f4f6", "#d1d5db"))
                        cell_w = QtWidgets.QWidget()
                        cell_w.setStyleSheet("background: transparent;")
                        cell_l = QtWidgets.QHBoxLayout(cell_w)
                        cell_l.setContentsMargins(6, 0, 6, 0)
                        cell_l.setAlignment(QtCore.Qt.AlignCenter)
                        badge = QtWidgets.QLabel(f"  {value}  ")
                        badge.setAlignment(QtCore.Qt.AlignCenter)
                        badge.setStyleSheet(f"""
                            color: {sc[0]}; background: {sc[1]};
                            border: 1px solid {sc[2]};
                            border-radius: 8px; font-size: 11px; font-weight: 800;
                            padding: 3px 6px;
                        """)
                        cell_l.addWidget(badge)
                        table.setCellWidget(row_index, col, cell_w)
                    else:
                        table.setItem(row_index, col, self._workspace_table_item(value, bold=col in (0, 3)))
        else:
            table.setRowHeight(0, 48)
            item = self._workspace_table_item("No invoices created yet.", align=QtCore.Qt.AlignLeft, color="#94a3b8")
            table.setItem(0, 0, item)
            table.setSpan(0, 0, 1, 6)

        # Column widths: Invoice # wider (stretch), Stage narrower (fixed)
        hdr = table.horizontalHeader()
        for col in range(6):
            hdr.setSectionResizeMode(col, QtWidgets.QHeaderView.Fixed)
        hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)  # Invoice # stretches
        table.setColumnWidth(1, 115)   # Date
        table.setColumnWidth(2, 180)   # Stage
        table.setColumnWidth(3, 135)   # Amount
        table.setColumnWidth(4, 125)   # Status
        table.setColumnWidth(5, 135)   # Received Date

        # Right-click on any row → copy invoice number
        table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        def _ws_inv_ctx_menu(pos, t=table):
            idx = t.indexAt(pos)
            if not idx.isValid():
                return
            inv_item = t.item(idx.row(), 0)
            if inv_item is None:
                return
            num = inv_item.text().strip()
            if not num:
                return
            menu = QtWidgets.QMenu(t)
            menu.setStyleSheet("""
                QMenu { background:#ffffff; border:1px solid #e2e8f0; border-radius:7px; padding:4px; }
                QMenu::item { padding:7px 20px; font-size:13px; color:#0f172a;
                              font-family:'Inter','Segoe UI'; border-radius:5px; }
                QMenu::item:selected { background:#e6f6f4; color:#00756f; }
            """)
            act = menu.addAction("Copy Invoice Number")
            if menu.exec_(t.viewport().mapToGlobal(pos)) == act:
                QtWidgets.QApplication.clipboard().setText(num)
        table.customContextMenuRequested.connect(_ws_inv_ctx_menu)

        table.setMinimumHeight(300)
        card_lay.addWidget(table)
        layout.addWidget(card)
        return tab

    def _create_workspace_payments_tab(self, project_data):
        tab = QtWidgets.QWidget()
        tab.setStyleSheet("background: transparent;")
        layout = QtWidgets.QVBoxLayout(tab)
        layout.setContentsMargins(0, 12, 0, 12)
        layout.setSpacing(12)

        card = QtWidgets.QFrame()
        card.setStyleSheet("QFrame { background:transparent; border:none; border-radius:10px; }")
        card_lay = QtWidgets.QVBoxLayout(card)
        card_lay.setContentsMargins(0, 0, 0, 0)
        card_lay.setSpacing(10)

        hdr_lbl = QtWidgets.QLabel("Payment History")
        hdr_lbl.setStyleSheet("font-size:15px; font-weight:900; color:#0f172a;")
        card_lay.addWidget(hdr_lbl)

        table = QtWidgets.QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Term", "Invoice #", "Paid Date", "Amount Paid", "Method", "Notes"])
        self._prepare_workspace_table(table)
        _all_pays = get_payment_tracker().get_project_payments(project_data.get("project_number", ""))
        def _pay_date_key(p):
            d = p.payment_date or ""
            for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(d, fmt)
                except ValueError:
                    pass
            return datetime.min
        payments = sorted(
            [p for p in _all_pays if (p.payment_stage or "").strip().lower() != "tax"],
            key=_pay_date_key
        )
        table.setRowCount(max(len(payments), 1))
        if payments:
            for row_index, payment in enumerate(payments):
                table.setRowHeight(row_index, 42)
                values = [
                    payment.payment_stage or "General Payment",
                    payment.invoice_number or "—",
                    payment.payment_date or "—",
                    self._format_project_money(payment.amount),
                    payment.payment_method or "—",
                    payment.notes or "",
                ]
                for col, value in enumerate(values):
                    if col == 3:
                        item = self._workspace_table_item(value, bold=True, color="#047857")
                    else:
                        align = QtCore.Qt.AlignLeft if col == 5 else QtCore.Qt.AlignCenter
                        item = self._workspace_table_item(value, align=align)
                    table.setItem(row_index, col, item)
        else:
            table.setRowHeight(0, 48)
            item = self._workspace_table_item("No payments recorded yet.", align=QtCore.Qt.AlignLeft, color="#94a3b8")
            table.setItem(0, 0, item)
            table.setSpan(0, 0, 1, 6)

        table.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.Stretch)
        for col in range(5):
            table.horizontalHeader().setSectionResizeMode(col, QtWidgets.QHeaderView.Fixed)
        table.setColumnWidth(0, 200)
        table.setColumnWidth(1, 175)
        table.setColumnWidth(2, 155)
        table.setColumnWidth(3, 160)
        table.setColumnWidth(4, 130)
        table.setMinimumHeight(300)
        card_lay.addWidget(table)
        layout.addWidget(card)
        return tab

    def _set_project_item_readability(self, item, text="", *, emphasis=False, muted=False, left=False):
        """Apply readable table text defaults and keep full values available on hover."""
        item.setToolTip(str(text or item.text()))
        item.setTextAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)
        font = QtGui.QFont("Inter", 9)
        font.setWeight(QtGui.QFont.DemiBold if emphasis else QtGui.QFont.Normal)
        font.setUnderline(False)
        item.setFont(font)
        item.setForeground(QtGui.QColor("#64748b" if muted else "#0f172a"))

    def _style_project_number_cell(self, item, project):
        project_number = project.get("project_number", item.text())
        item.setToolTip(f"Click to view details for {project_number}")
        font = QtGui.QFont("Inter", 9)
        font.setWeight(QtGui.QFont.DemiBold)
        font.setUnderline(False)
        item.setFont(font)
        item.setForeground(QtGui.QColor("#2563eb"))
        item.setTextAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)

    def _create_project_number_widget(self, project_number: str, project: dict):
        """Cell widget: blue dot + project number in clean Inter font."""
        widget = QtWidgets.QWidget()
        widget.setStyleSheet("background: transparent; border: none;")
        hbox = QtWidgets.QHBoxLayout(widget)
        hbox.setContentsMargins(8, 0, 6, 0)
        hbox.setSpacing(7)

        # Status-colored dot
        status = self._project_effective_status(project)
        dot_color = {
            "In Progress":             "#2563eb",
            "Not Started":             "#2563eb",
            "Completed Not Invoiced":  "#22c55e",
            "Completed & Invoiced":    "#22c55e",
            "On Hold":                 "#f59e0b",
            "Cancelled":               "#94a3b8",
        }.get(status, "#2563eb")

        dot = QtWidgets.QLabel("●")
        dot.setFixedSize(14, 14)
        dot.setAlignment(QtCore.Qt.AlignCenter)
        dot.setStyleSheet(f"""
            QLabel {{
                color: {dot_color};
                font-size: 10px;
                background: transparent;
                border: none;
            }}
        """)

        num_lbl = QtWidgets.QLabel(project_number)
        num_lbl.setToolTip(f"Click to view details for {project_number}")
        num_lbl.setStyleSheet("""
            QLabel {
                color: #2563eb;
                font-size: 13px;
                font-weight: 700;
                background: transparent;
                border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QToolTip {
                background: #ffffff;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 4px 9px;
                font-size: 11px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        num_lbl.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        num_lbl.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)

        hbox.addWidget(dot)
        hbox.addWidget(num_lbl, 1)
        return widget

    def _payment_category_colors(self, category: str):
        category = (category or "N/A").lower()
        if "single" in category or "full" in category:
            return "#e0f2fe", "#075985", "#7dd3fc", "#bae6fd"
        if "custom" in category or "installment" in category:
            return "#ede9fe", "#5b21b6", "#c4b5fd", "#ddd6fe"
        if "25%" in category or "1st" in category:
            return "#dcfce7", "#166534", "#86efac", "#bbf7d0"
        if "down" in category or "deposit" in category or "50%" in category:
            return "#dcfce7", "#166534", "#86efac", "#bbf7d0"   # green  — stage 1
        if "2nd" in category or "3rd" in category or "4th" in category:
            return "#dbeafe", "#1e40af", "#93c5fd", "#bfdbfe"   # blue   — progress
        if "balance" in category or "due" in category:
            return "#fef3c7", "#92400e", "#fcd34d", "#fde68a"   # amber  — balance due
        if "final" in category:
            return "#ede9fe", "#5b21b6", "#c4b5fd", "#ddd6fe"
        return "#f1f5f9", "#475569", "#cbd5e1", "#e2e8f0"

    def _style_payment_category_button(self, button: QtWidgets.QPushButton, category: str):
        display_names = {
            "Single Payment": "Plan: Single",
            "25% Deposit + Balance": "Plan: 25% + Balance",
            "50% Deposit + Final": "Plan: 50% + Final",
            "Custom Installments": "Plan: Custom",
        }
        button.setText(display_names.get(category, category))
        button.setToolTip(f"Payment plan: {category}")
        button.setStyleSheet("""
            QPushButton {
                background: #f8fafc;
                color: #475569;
                border: 1px solid #cbd5e1;
                border-radius: 7px;
                padding: 0 8px;
                font-size: 10px;
                font-weight: 800;
                text-align: center;
            }
            QPushButton:hover {
                background: #eef6f8;
                border-color: #0f766e;
                color: #0f766e;
            }
        """)
        
    def get_group_box_style(self):
        """Get group box style matching JobFormTab"""
        return """
            QGroupBox {
                font-weight: 800;
                font-size: 14px;
                color: #0f172a;
                border: none;
                border-radius: 10px;
                margin-top: 0px;
                padding-top: 0px;
                background: transparent;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 12px 0 12px;
                color: #0f172a;
                font-weight: 800;
                background: transparent;
            }
        """
    
    def _on_projects_table_context_menu(self, pos):
        """Show Copy context menu when right-clicking the project number column."""
        index = self.projects_table.indexAt(pos)
        if not index.isValid():
            return
        row = index.row()
        col = index.column()
        # Allow copy from project number column or any column (get project number from row data)
        pn_item = self.projects_table.item(row, self.COL_PROJECT_NUMBER)
        if pn_item is None:
            return
        project_data = pn_item.data(QtCore.Qt.UserRole)
        project_number = (project_data or {}).get("project_number", "") or pn_item.toolTip()
        if not project_number:
            return
        menu = QtWidgets.QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 7px; padding: 4px; }
            QMenu::item { padding: 7px 20px; font-size: 13px; color: #0f172a;
                          font-family: 'Inter','Segoe UI'; border-radius: 5px; }
            QMenu::item:selected { background: #e6f6f4; color: #00756f; }
        """)
        copy_action = menu.addAction(f"Copy Project Number")
        action = menu.exec_(self.projects_table.viewport().mapToGlobal(pos))
        if action == copy_action:
            QtWidgets.QApplication.clipboard().setText(project_number)

    def on_project_row_clicked(self, row, column):
        self._last_clicked_row = row
        self.on_project_selected()
        if column == self.COL_PROJECT_NUMBER:
            project_item = self.projects_table.item(row, self.COL_PROJECT_NUMBER)
            project_data = project_item.data(QtCore.Qt.UserRole) if project_item else None
            if project_data:
                self.show_project_workspace(project_data)
        return
        selection_model = self.projects_table.selectionModel()

        # If user clicks the same already-selected row → unselect it
        if self._last_clicked_row == row and selection_model.isRowSelected(row, QtCore.QModelIndex()):
            self.projects_table.clearSelection()
            self.selected_projects.clear()
            self.load_selected_btn.setEnabled(False)
            self._last_clicked_row = None
            return

        # Otherwise, normal selection
        self._last_clicked_row = row

        # Update selected projects list
        self.selected_projects.clear()
        project_item = self.projects_table.item(row, 1)  # Project Number column
        if project_item:
            project_data = project_item.data(QtCore.Qt.UserRole)
            if project_data:
                self.selected_projects.add(project_data["project_number"])

        self.load_selected_btn.setEnabled(True)

    def load_multiple_projects(self):
        """Load multiple selected projects to invoice tab with client details"""
        # Get selected projects from the table selection
        selected_rows = set()
        for item in self.projects_table.selectedItems():
            selected_rows.add(item.row())
        
        if not selected_rows:
            QtWidgets.QMessageBox.warning(self, "Load Projects", "Please select projects first.")
            return
        
        _log.info("📥 Loading %s selected projects...", len(selected_rows))
        
        selected_project_objects = []
        for row in selected_rows:
            project_item = self.projects_table.item(row, 1)  # Project Number column
            if project_item:
                project_data = project_item.data(QtCore.Qt.UserRole)
                if project_data:
                    selected_project_objects.append(project_data)
        
        if not selected_project_objects:
            QtWidgets.QMessageBox.warning(self, "Load Projects", "No valid projects selected.")
            return

        return self.load_projects_to_invoice_direct(selected_project_objects)
        
        # Navigate to the correct tab
        main_window = self.main_window
        main_window.tabs.setCurrentIndex(1)  # Projects & Invoice tab (index 1)
        
        lockable_tab = main_window.tabs.widget(1)
        
        # Access the inner content properly
        if hasattr(lockable_tab, 'content_container'):
            # Find the inner tab widget that contains the invoice tab
            inner_tab_widget = None
            
            # Look for QTabWidget inside the content container
            for child in lockable_tab.content_container.findChildren(QtWidgets.QTabWidget):
                if child:
                    inner_tab_widget = child
                    break
            
            if inner_tab_widget:
                # Switch to "Create Invoice" tab (index 1)
                inner_tab_widget.setCurrentIndex(1)
                create_tab_widget = inner_tab_widget.widget(1)
                
                # Get references to the invoice tab's components
                # These are the actual attribute names used in MainWindow's setup_create_tab method
                client_combo = None
                client_email_edit = None
                client_address_edit = None
                clear_all_items_method = None
                update_totals_method = None
                add_item_row_method = None
                item_rows = None
                items_layout = None
                
                # Try to get attributes directly from the create_tab_widget
                if hasattr(create_tab_widget, 'client_combo'):
                    client_combo = create_tab_widget.client_combo
                if hasattr(create_tab_widget, 'client_email_edit'):
                    client_email_edit = create_tab_widget.client_email_edit
                if hasattr(create_tab_widget, 'client_address_edit'):
                    client_address_edit = create_tab_widget.client_address_edit
                if hasattr(create_tab_widget, 'clear_all_items'):
                    clear_all_items_method = create_tab_widget.clear_all_items
                if hasattr(create_tab_widget, 'update_totals'):
                    update_totals_method = create_tab_widget.update_totals
                if hasattr(create_tab_widget, 'add_item_row'):
                    add_item_row_method = create_tab_widget.add_item_row
                if hasattr(create_tab_widget, 'item_rows'):
                    item_rows = create_tab_widget.item_rows
                if hasattr(create_tab_widget, 'items_layout'):
                    items_layout = create_tab_widget.items_layout
                
                # If not found, search through all child widgets
                if not client_combo:
                    for widget in create_tab_widget.findChildren(QtWidgets.QComboBox):
                        if hasattr(widget, 'objectName') and widget.objectName() == 'client_combo':
                            client_combo = widget
                            break
                    # If still not found, try to find by placeholder text
                    if not client_combo:
                        for widget in create_tab_widget.findChildren(QtWidgets.QComboBox):
                            if widget.lineEdit() and widget.lineEdit().placeholderText() == "Select or type client name...":
                                client_combo = widget
                                break
                
                if not client_email_edit:
                    for widget in create_tab_widget.findChildren(QtWidgets.QLineEdit):
                        if hasattr(widget, 'objectName') and widget.objectName() == 'client_email_edit':
                            client_email_edit = widget
                            break
                    if not client_email_edit:
                        for widget in create_tab_widget.findChildren(QtWidgets.QLineEdit):
                            if widget.placeholderText() == "company@example.com":
                                client_email_edit = widget
                                break
                
                if not client_address_edit:
                    for widget in create_tab_widget.findChildren(QtWidgets.QTextEdit):
                        if hasattr(widget, 'objectName') and widget.objectName() == 'client_address_edit':
                            client_address_edit = widget
                            break
                    if not client_address_edit:
                        for widget in create_tab_widget.findChildren(QtWidgets.QTextEdit):
                            if widget.placeholderText() == "Street Address, City, State ZIP":
                                client_address_edit = widget
                                break
                
                # If clear_all_items method doesn't exist, create a fallback
                if not clear_all_items_method and hasattr(create_tab_widget, 'item_rows') and hasattr(create_tab_widget, 'items_layout'):
                    def fallback_clear():
                        if hasattr(create_tab_widget, 'item_rows'):
                            for row in create_tab_widget.item_rows[:]:
                                if hasattr(row, 'setParent'):
                                    row.setParent(None)
                            create_tab_widget.item_rows.clear()
                        if hasattr(create_tab_widget, 'items_layout'):
                            for i in reversed(range(create_tab_widget.items_layout.count())):
                                widget = create_tab_widget.items_layout.itemAt(i).widget()
                                if widget:
                                    widget.deleteLater()
                    clear_all_items_method = fallback_clear
                
                # If add_item_row method doesn't exist, create a fallback
                if not add_item_row_method and hasattr(create_tab_widget, 'items_layout') and hasattr(create_tab_widget, 'item_rows'):
                    def fallback_add_item(item):
                        from project_number_generator import InvoiceItem
                        row = ItemRowWidget(item)
                        if hasattr(row, 'removed'):
                            row.removed.connect(lambda: create_tab_widget.item_rows.remove(row) if row in create_tab_widget.item_rows else None)
                        create_tab_widget.items_layout.addWidget(row)
                        create_tab_widget.item_rows.append(row)
                        if hasattr(row, 'update_total'):
                            row.update_total()
                        if hasattr(create_tab_widget, 'update_totals'):
                            create_tab_widget.update_totals()
                    add_item_row_method = fallback_add_item
                
                # If update_totals method doesn't exist, create a fallback
                if not update_totals_method and hasattr(main_window, 'update_totals'):
                    update_totals_method = main_window.update_totals
                
                if not client_combo:
                    QtWidgets.QMessageBox.warning(self, "Error", "Cannot find invoice tab components.")
                    return
                
                # Check if all projects are from the same company
                companies = set(project["company"] for project in selected_project_objects)
                
                if len(companies) == 1:
                    # Single company - set client info
                    first_project = selected_project_objects[0]
                    company_name = first_project["company"]
                    
                    client_combo.blockSignals(True)
                    index = client_combo.findText(company_name)
                    if index >= 0:
                        client_combo.setCurrentIndex(index)
                    else:
                        client_combo.setEditText(company_name)
                    client_combo.blockSignals(False)
                    
                    # Load client details if available
                    if hasattr(main_window, 'clients') and company_name in main_window.clients:
                        client_data = main_window.clients[company_name]
                        if client_email_edit:
                            client_email_edit.setText(client_data.get("company_email", ""))
                        if client_address_edit:
                            client_address_edit.setPlainText(client_data.get("address", ""))
                else:
                    # Multiple companies - show warning but still load projects
                    QtWidgets.QMessageBox.warning(
                        self, 
                        "Multiple Companies", 
                        f"You have selected projects from {len(companies)} different companies.\n\n"
                        f"Companies: {', '.join(companies)}\n\n"
                        f"Projects will be loaded but client information won't be auto-filled.",
                        QtWidgets.QMessageBox.Ok
                    )
                
                # Clear existing items
                if clear_all_items_method:
                    clear_all_items_method()
                else:
                    # Manual clear as last resort
                    if hasattr(main_window, 'clear_all_items'):
                        main_window.clear_all_items()
                    elif hasattr(create_tab_widget, 'item_rows') and hasattr(create_tab_widget, 'items_layout'):
                        for row in create_tab_widget.item_rows[:]:
                            if hasattr(row, 'setParent'):
                                row.setParent(None)
                        create_tab_widget.item_rows.clear()
                        for i in reversed(range(create_tab_widget.items_layout.count())):
                            widget = create_tab_widget.items_layout.itemAt(i).widget()
                            if widget:
                                widget.deleteLater()
                
                # Load projects
                loaded_count = 0
                down_payment_projects = 0
                final_payment_projects = 0
                due_payment_projects = 0
                na_payment_projects = 0
                
                for project in selected_project_objects:
                    project_amount = project.get("project_amount", 0.0)
                    plant = project.get("plant", "")
                    
                    payment_category = project.get("payment_category", "N/A")
                    
                    if payment_category == "Down Payment":
                        down_payment_projects += 1
                    elif payment_category == "Final Payment":
                        final_payment_projects += 1
                    elif payment_category == "Due Payment":
                        due_payment_projects += 1
                    elif not payment_category or payment_category == "N/A":
                        na_payment_projects += 1
                    
                    # Create invoice item
                    from project_number_generator import InvoiceItem
                    item = InvoiceItem(
                        project_number=project["project_number"],
                        description=project["project_name"],
                        plant=plant,
                        quantity=1,
                        unit_price=project_amount,
                        down_payment=0.0
                    )
                    
                    # Add item to invoice tab
                    if add_item_row_method:
                        add_item_row_method(item)
                    elif hasattr(main_window, 'add_item_row'):
                        main_window.add_item_row(item)
                    else:
                        # Manual add as last resort
                        row = ItemRowWidget(item)
                        if hasattr(row, 'removed'):
                            row.removed.connect(lambda r=row: create_tab_widget.item_rows.remove(r) if r in create_tab_widget.item_rows else None)
                        create_tab_widget.items_layout.addWidget(row)
                        create_tab_widget.item_rows.append(row)
                        if hasattr(row, 'update_total'):
                            row.update_total()
                    
                    # Get the last added row to set payment category
                    last_row = None
                    if add_item_row_method and hasattr(create_tab_widget, 'item_rows') and create_tab_widget.item_rows:
                        last_row = create_tab_widget.item_rows[-1]
                    elif hasattr(main_window, 'item_rows') and main_window.item_rows:
                        last_row = main_window.item_rows[-1]
                    elif hasattr(create_tab_widget, 'item_rows') and create_tab_widget.item_rows:
                        last_row = create_tab_widget.item_rows[-1]
                    
                    if last_row:
                        if payment_category == "Down Payment":
                            last_row.down_payment_combo.setCurrentText("Down Payment (50%)")
                        elif payment_category == "Due Payment":
                            last_row.down_payment_combo.setCurrentText("Due Payment")
                        elif payment_category == "Final Payment":
                            last_row.down_payment_combo.setCurrentText("Final Payment")
                        else:
                            last_row.down_payment_combo.setCurrentIndex(0)
                        
                        if hasattr(last_row, 'update_total'):
                            last_row.update_total()
                    
                    loaded_count += 1
                
                # Update totals
                if update_totals_method:
                    update_totals_method()
                elif hasattr(main_window, 'update_totals'):
                    main_window.update_totals()
                elif hasattr(create_tab_widget, 'update_totals'):
                    create_tab_widget.update_totals()

                updated_status_count = self.mark_loaded_projects_invoiced(selected_project_objects)
                
                # Clear selection in projects table
                self.projects_table.clearSelection()
                self.selected_projects.clear()
                self.load_selected_btn.setEnabled(False)
                
                # Show success message
                if len(companies) == 1:
                    company_name = list(companies)[0]
                    message = f"<h3>{loaded_count} project(s) loaded to invoice!</h3>" \
                            f"<b>Company:</b> {company_name}<br>"
                else:
                    message = f"<h3>{loaded_count} project(s) loaded to invoice!</h3>" \
                            f"<b>Multiple companies detected</b><br>"
                
                if down_payment_projects > 0:
                    message += f"<b>Down Payment (50%) projects:</b> {down_payment_projects}<br>"
                if final_payment_projects > 0:
                    message += f"<b>Final Payment projects:</b> {final_payment_projects}<br>"
                if due_payment_projects > 0:
                    message += f"<b>Due Payment projects:</b> {due_payment_projects}<br>"
                if na_payment_projects > 0:
                    message += f"<b>N/A (No payment category set):</b> {na_payment_projects}<br>"
                
                if updated_status_count > 0:
                    message += f"<b>Status updated:</b> {updated_status_count} project(s) marked Completed & Invoiced<br>"
                
                message += "<br><i>Down payments have been auto-set based on payment category where specified.</i>"
                
                QtWidgets.QMessageBox.information(
                    self, "Projects Loaded", 
                    message,
                    QtWidgets.QMessageBox.Ok
                )

                if updated_status_count > 0:
                    QtCore.QTimer.singleShot(0, self.load_projects)
            else:
                QtWidgets.QMessageBox.warning(self, "Error", "Cannot find invoice tab inside the locked tab.")
        else:
            QtWidgets.QMessageBox.warning(self, "Error", "Cannot access the Projects & Invoice tab.")

    def load_projects_to_invoice_direct(self, selected_project_objects):
        """Load selected projects using the main invoice widgets directly."""
        main_window = self.main_window
        if not main_window:
            QtWidgets.QMessageBox.warning(self, "Error", "Cannot access the main window.")
            return

        try:
            # ── Auto-detect payment stage for each project ───────────────
            try:
                from main import FirebaseManager
                existing_invoices = FirebaseManager.load_invoices() or []
            except Exception:
                existing_invoices = []

            project_stages = [
                {"project": p, "stage_info": _detect_payment_stage(p, existing_invoices)}
                for p in selected_project_objects
            ]

            # Block projects that are fully invoiced
            fully_done = [
                ps["project"].get("project_number", "")
                for ps in project_stages
                if ps["stage_info"]["stage"] is None
            ]
            if fully_done and len(fully_done) == len(project_stages):
                QtWidgets.QMessageBox.information(
                    self, "All Invoiced",
                    "All selected projects have already been fully invoiced."
                )
                return

            confirm = PaymentStageConfirmDialog(project_stages, self)
            if confirm.exec_() != QtWidgets.QDialog.Accepted:
                return

            # Build stage map: project_number → stage_info
            stage_map = {
                ps["project"].get("project_number", ""): ps["stage_info"]
                for ps in project_stages
                if ps["stage_info"]["stage"] is not None
            }
            # Filter to only projects with a valid next stage
            selected_project_objects = [
                ps["project"] for ps in project_stages
                if ps["stage_info"]["stage"] is not None
            ]
            # ─────────────────────────────────────────────────────────────

            if hasattr(main_window, "_nav_to"):
                if not main_window._nav_to(2):
                    return
            if hasattr(main_window, "_switch_inner_tab"):
                main_window._switch_inner_tab(1)   # index 1 = Invoice Management
            elif hasattr(main_window, "project_invoice_inner_tabs"):
                main_window.project_invoice_inner_tabs.setCurrentIndex(1)

            client_combo = getattr(main_window, "client_combo", None)
            line_items_client_combo = getattr(main_window, "line_items_client_combo", None)
            client_email_edit = getattr(main_window, "client_email_edit", None)
            client_address_edit = getattr(main_window, "client_address_edit", None)

            if client_combo is None or not hasattr(main_window, "add_item_row"):
                QtWidgets.QMessageBox.warning(self, "Error", "Cannot find invoice tab components.")
                return

            companies = {project.get("company", "") for project in selected_project_objects if project.get("company", "")}
            if len(companies) == 1:
                company_name = next(iter(companies)).strip()
                saved_client_name = company_name
                client_data = {}
                if hasattr(main_window, "_find_client_record"):
                    saved_client_name, client_data = main_window._find_client_record(company_name)
                    saved_client_name = saved_client_name or company_name
                else:
                    clients = getattr(main_window, "clients", {}) or {}
                    client_data = clients.get(company_name, {})

                if not client_data:
                    try:
                        from main import FirebaseManager
                        refreshed_clients = FirebaseManager.load_clients() or {}
                        if refreshed_clients:
                            main_window.clients = refreshed_clients
                            if hasattr(main_window, "_find_client_record"):
                                saved_client_name, client_data = main_window._find_client_record(company_name)
                                saved_client_name = saved_client_name or company_name
                            else:
                                client_data = refreshed_clients.get(company_name, {})
                    except Exception:
                        client_data = {}

                for combo in (client_combo, line_items_client_combo):
                    if combo is not None:
                        combo.blockSignals(True)
                        index = combo.findText(saved_client_name)
                        if index >= 0:
                            combo.setCurrentIndex(index)
                        else:
                            combo.addItem(saved_client_name)
                            combo.setCurrentIndex(combo.findText(saved_client_name))
                        combo.blockSignals(False)

                if client_data:
                    if hasattr(main_window, "load_client_details"):
                        main_window.load_client_details(saved_client_name)
                    else:
                        if client_email_edit:
                            client_email_edit.setText(client_data.get("company_email", client_data.get("email", "")))
                        if client_address_edit:
                            client_address_edit.setPlainText(client_data.get("address", ""))
                    if hasattr(main_window, "update_invoice_client_summary"):
                        main_window.update_invoice_client_summary(
                            saved_client_name,
                            client_data.get("company_email", client_data.get("email", "")),
                            client_data.get("address", ""),
                        )
                elif hasattr(main_window, "update_invoice_client_summary"):
                    main_window.update_invoice_client_summary(company_name, "", "")
            elif len(companies) > 1:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Multiple Companies",
                    f"You selected projects from {len(companies)} different companies.\n\n"
                    "Projects will be loaded, but client information will not be auto-filled.",
                    QtWidgets.QMessageBox.Ok,
                )

            if hasattr(main_window, "clear_all_items"):
                main_window.clear_all_items()

            loaded_count = 0
            invoice_item_cls = getattr(main_window, "InvoiceItem", None)
            if invoice_item_cls is None:
                from main import InvoiceItem as invoice_item_cls

            for project in selected_project_objects:
                pn = project.get("project_number", "")
                si = stage_map.get(pn, {})
                stage = si.get("stage", "Down Payment")
                stage_amount = si.get("amount", float(project.get("project_amount", 0) or 0))

                invoice_payment_category = ItemRowWidget.normalize_payment_label(stage) or ItemRowWidget.FULL_AMOUNT_LABEL

                item = invoice_item_cls(
                    project_number=pn,
                    description=project.get("project_name", ""),
                    plant=project.get("plant", ""),
                    quantity=1,
                    unit_price=stage_amount,
                    down_payment=0.0,
                    payment_category=invoice_payment_category,
                )
                main_window.add_item_row(item)

                last_row = main_window.item_rows[-1] if getattr(main_window, "item_rows", None) else None
                if last_row:
                    # Lock the row to this exact payment stage — hides all other options
                    if hasattr(last_row, "lock_to_stage") and invoice_payment_category:
                        last_row.lock_to_stage(invoice_payment_category)
                    if hasattr(last_row, "update_total"):
                        last_row.update_total()
                    # Qty and plant are auto-filled from the project — prevent accidental edits
                    if hasattr(last_row, "qty_spin"):
                        last_row.qty_spin.setEnabled(False)
                    if hasattr(last_row, "plant_edit"):
                        last_row.plant_edit.setReadOnly(True)
                        last_row.plant_edit.setStyleSheet(
                            last_row.plant_edit.styleSheet() +
                            "QLineEdit { background: #f1f5f9; color: #64748b; }"
                        )

                loaded_count += 1

            if hasattr(main_window, "update_totals"):
                main_window.update_totals()

            # ── Fix 6: Set invoice date / due date from project dates ────────
            try:
                if len(selected_project_objects) == 1:
                    _proj = selected_project_objects[0]
                    _start = str(_proj.get("start_date", "") or "").strip()
                    _due   = str(_proj.get("due_date", "")   or "").strip()
                elif len(selected_project_objects) > 1:
                    _starts = {str(p.get("start_date", "") or "").strip() for p in selected_project_objects}
                    _dues   = {str(p.get("due_date", "")   or "").strip() for p in selected_project_objects}
                    _start = next(iter(_starts)) if len(_starts) == 1 else ""
                    _due   = next(iter(_dues))   if len(_dues)   == 1 else ""
                else:
                    _start = _due = ""

                _date_fmts = ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y")
                def _parse_to_qdate(s):
                    for _fmt in _date_fmts:
                        try:
                            _dt = datetime.strptime(s, _fmt)
                            return QtCore.QDate(_dt.year, _dt.month, _dt.day)
                        except Exception:
                            pass
                    return None

                if _start and hasattr(main_window, "date_edit"):
                    _qd = _parse_to_qdate(_start)
                    if _qd and _qd.isValid():
                        main_window.date_edit.setDate(_qd)
                if _due and hasattr(main_window, "due_date_edit"):
                    _qd = _parse_to_qdate(_due)
                    if _qd and _qd.isValid():
                        main_window.due_date_edit.setDate(_qd)
            except Exception as _date_err:
                _log.warning("Could not set invoice dates from project: %s", _date_err)

            # ── Update Payment Status label from actual stage-matched payments ──
            try:
                from payment_tracker import get_payment_tracker as _get_tracker
                _tracker = _get_tracker()
                _tracker._load_payments()

                _inv_total   = 0.0
                _stage_paid  = 0.0
                _date_fmts2  = ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y")

                for _p in selected_project_objects:
                    _pn        = _p.get("project_number", "")
                    _si        = stage_map.get(_pn, {})
                    _stage_lbl = _si.get("stage", "")
                    _stage_amt = float(_si.get("amount", float(_p.get("project_amount", 0) or 0)))
                    _inv_total += _stage_amt
                    _norm_lbl  = self._norm_stage(_stage_lbl) if _stage_lbl else ""

                    for _pay in _tracker.payments:
                        if _pay.project_number != _pn:
                            continue
                        if (_pay.payment_stage or "").strip().lower() == "tax":
                            continue
                        # If we know the stage, only count payments for that stage
                        if _norm_lbl and self._norm_stage(_pay.payment_stage) != _norm_lbl:
                            continue
                        _stage_paid += float(_pay.amount)

                # Compute base status without tax, then apply tax override for display
                _base_status = "Unpaid"
                if _stage_paid >= _inv_total - 0.005 and _inv_total > 0:
                    _base_status = "Paid"
                elif _stage_paid > 0:
                    _base_status = "Partially Paid"
                else:
                    _today = datetime.now().date()
                    _overdue = False
                    for _p in selected_project_objects:
                        _dd = str(_p.get("due_date", "") or "").strip()
                        if _dd:
                            for _fmt in _date_fmts2:
                                try:
                                    if datetime.strptime(_dd, _fmt).date() < _today:
                                        _overdue = True
                                    break
                                except Exception:
                                    pass
                    if _overdue:
                        _base_status = "Overdue"

                # Store base status so update_totals can re-apply tax override dynamically
                main_window._ps_base_status = _base_status

                _computed_status = _base_status

                _status_styles = {
                    "Paid":           "color:#00756f;background-color:#e9fbf7;border-color:#8edbd2;",
                    "Partially Paid": "color:#1e40af;background-color:#dbeafe;border-color:#93c5fd;",
                    "Overdue":        "color:#9d174d;background-color:#fce7f3;border-color:#f9a8d4;",
                    "Unpaid":         "color:#7a1f1f;background-color:#fff1f0;border-color:#f0b4b4;",
                }
                _ss = _status_styles.get(_computed_status, _status_styles["Unpaid"])
                if hasattr(main_window, "payment_status_label") and main_window.payment_status_label:
                    main_window.payment_status_label.setText(_computed_status)
                    main_window.payment_status_label.setStyleSheet(f"""
                        QLabel {{
                            font-size: 13px; font-weight: 800;
                            border: 1px solid; border-radius: 7px;
                            padding: 8px 12px; {_ss}
                        }}
                    """)
            except Exception as _st_err:
                _log.warning("Could not compute payment status label: %s", _st_err)

            if hasattr(main_window, "_switch_inner_tab"):
                main_window._switch_inner_tab(1)  # Invoice Management
            elif hasattr(main_window, "project_invoice_inner_tabs"):
                main_window.project_invoice_inner_tabs.setCurrentIndex(1)
            if line_items_client_combo is not None and line_items_client_combo.lineEdit():
                line_items_client_combo.lineEdit().setFocus()

            updated_status_count = self.mark_loaded_projects_invoiced(selected_project_objects)

            if len(companies) == 1:
                message = f"{loaded_count} project(s) loaded to invoice for {next(iter(companies))}."
            else:
                message = f"{loaded_count} project(s) loaded to invoice."
            if updated_status_count:
                message += f" {updated_status_count} project status updated."

            self.selected_projects.clear()
            self.load_selected_btn.setEnabled(False)
            if hasattr(main_window, "statusBar"):
                main_window.statusBar().showMessage(message, 6000)
            _log.info(message)

            # Ensure Generate PDF button is enabled after all rows are locked
            if hasattr(main_window, "_update_pdf_btn_state"):
                QtCore.QTimer.singleShot(150, main_window._update_pdf_btn_state)

        except Exception as exc:
            _log.exception("Load to Invoice failed")
            QtWidgets.QMessageBox.critical(self, "Load to Invoice", f"Could not load project to invoice:\n{exc}")

    def mark_loaded_projects_invoiced(self, projects):
        """Loading a project to invoice only fills the invoice screen.

        The invoice is saved/uploaded only after the user reviews it and clicks the
        invoice save/generate action, so project status is not changed here.
        """
        return 0
            
    def on_project_selected(self):
        """When a project is selected in the table - updated for multi-selection"""
        selected_items = self.projects_table.selectedItems()
        
        self.selected_projects.clear()
        
        if selected_items:
            selected_rows = set()
            for item in selected_items:
                row = item.row()
                selected_rows.add(row)
                
                if item.column() == 1:  # Project Number column
                    project_item = self.projects_table.item(row, 1)
                    if project_item:
                        project_data = project_item.data(QtCore.Qt.UserRole)
                        if project_data:
                            self.selected_projects.add(project_data.get("project_number", ""))
            
            for row in range(self.projects_table.rowCount()):
                if row in selected_rows:
                    for col in range(self.projects_table.columnCount()):
                        item = self.projects_table.item(row, col)
                        if item:
                            item.setBackground(QtGui.QColor("#e3f2fd"))
                            item.setForeground(QtGui.QColor("#2c3e50"))
                else:
                    for col in range(self.projects_table.columnCount()):
                        item = self.projects_table.item(row, col)
                        if item:
                            item.setBackground(QtGui.QColor("#ffffff"))
                            item.setForeground(QtGui.QColor("#2c3e50"))

            _log.info("%s project(s) selected", len(self.selected_projects))
        else:
            for row in range(self.projects_table.rowCount()):
                for col in range(self.projects_table.columnCount()):
                    item = self.projects_table.item(row, col)
                    if item:
                        item.setBackground(QtGui.QColor("#ffffff"))
                        item.setForeground(QtGui.QColor("#2c3e50"))
            
            _log.info("No projects selected")
        
        self.load_selected_btn.setEnabled(len(self.selected_projects) > 0)
        
    def update_stats(self):
        """Update statistics cards based on ALL active filters"""
        search_text = self.search_edit.text().lower()
        
        status_filter = self._clean_project_status_filter(self.status_filter_combo.currentText())
        
        client_filter = self.selected_client_filter
        plant_filter  = getattr(self, "selected_plant_filter", "All Plants")
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')

        filtered_projects = []
        for project in self.generated_projects:
            matches_client = (
                client_filter == "All Clients" or
                client_filter == "📂 All Clients" or
                project.get('company', '') == client_filter
            )

            matches_plant = (
                plant_filter in ("All Plants", "") or
                project.get("plant", "").strip().lower() == plant_filter.strip().lower()
            )
            
            matches_search = False
            if not search_text:
                matches_search = True
            else:
                if search_text in project.get('project_number', '').lower():
                    matches_search = True
                elif search_text in project.get('project_name', '').lower():
                    matches_search = True
                elif search_text in project.get('company', '').lower():
                    matches_search = True
                elif search_text in project.get('sales', '').lower():
                    matches_search = True
                elif search_text in project.get('plant', '').lower():
                    matches_search = True
                elif search_text in project.get('payment_category', '').lower():
                    matches_search = True
                else:
                    amount = project.get('project_amount', 0.0)
                    if amount:
                        try:
                            amount_str = str(amount).replace(',', '').replace('$', '')
                            if search_text in amount_str.lower():
                                matches_search = True
                            elif search_text in Currency.format(amount).lower():
                                matches_search = True
                        except:
                            pass
                    if not matches_search:
                        payment_lower = search_text.replace(' ', '')
                        if payment_lower == 'downpayment' and project.get('payment_category', '').lower() == 'down payment':
                            matches_search = True
                        elif payment_lower == 'duepayment' and project.get('payment_category', '').lower() == 'due payment':
                            matches_search = True
                        elif payment_lower == 'finalpayment' and project.get('payment_category', '').lower() == 'final payment':
                            matches_search = True
            
            matches_status = self._project_status_matches(project, status_filter)
            
            matches_date = True
            if date_range_active:
                matches_date = self.is_project_in_date_range_by_received(
                    project, self.current_from_date, self.current_to_date
                )
            
            if matches_client and matches_plant and matches_search and matches_status and matches_date:
                filtered_projects.append(project)

        total_projects = len(filtered_projects)
        completed_projects = len([p for p in filtered_projects if p.get('status') in ['Completed Not Invoiced', 'Completed & Invoiced']])
        cancelled_projects = len([p for p in filtered_projects if p.get('status') == 'Cancelled'])
        active_projects = len([
            p for p in filtered_projects
            if self._project_effective_status(p) in ("Not Started", "In Progress")
        ])
        on_hold_projects = len([
            p for p in filtered_projects
            if self._project_effective_status(p) == "On Hold"
        ])
        today = datetime.now().date()
        overdue_projects = 0
        for project in filtered_projects:
            due_raw = str(project.get("due_date", "") or "").strip()
            due_date = None
            for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    due_date = datetime.strptime(due_raw[:10], fmt).date()
                    break
                except Exception:
                    pass
            if due_date and due_date < today and self._project_effective_status(project) not in (
                "Completed Not Invoiced", "Completed & Invoiced", "Paid", "Cancelled"
            ):
                overdue_projects += 1

        if hasattr(self, 'total_projects_label'):
            self.total_projects_label.setText(str(total_projects))
        if hasattr(self, 'completed_label'):
            self.completed_label.setText(str(completed_projects))
        if hasattr(self, 'cancelled_label'):
            self.cancelled_label.setText(str(cancelled_projects))
        if hasattr(self, 'active_projects_label'):
            self.active_projects_label.setText(str(active_projects))
        if hasattr(self, 'on_hold_projects_label'):
            self.on_hold_projects_label.setText(str(on_hold_projects))
        if hasattr(self, 'overdue_projects_label'):
            self.overdue_projects_label.setText(str(overdue_projects))

        # ── Trend labels: percentage change vs last month ─────────────────
        now = datetime.now()
        def _month_count(projects, year, month, status_fn=None):
            count = 0
            for p in projects:
                rd = p.get("created_at") or p.get("date_received") or ""
                try:
                    dt = datetime.fromisoformat(str(rd)[:19].replace("Z", ""))
                except Exception:
                    dt = None
                if dt and dt.year == year and dt.month == month:
                    if status_fn is None or status_fn(p):
                        count += 1
            return count

        # Last month date
        first_of_this = now.replace(day=1)
        last_month = (first_of_this - timedelta(days=1))
        ly, lm = last_month.year, last_month.month

        def _pct_trend(this, last):
            """Return (text, is_up) for percentage trend."""
            if last == 0 and this == 0:
                return "— no change", True
            if last == 0:
                return f"↑ {this} this month", True
            pct = round((this - last) / last * 100)
            if pct > 0:
                return f"↑ {pct}% this month", True
            elif pct < 0:
                return f"↓ {abs(pct)}% this month", False
            else:
                return "— no change", True

        def _set_trend(attr, text, up=True):
            lbl = getattr(self, attr, None)
            if lbl:
                clr = "#22c55e" if up else "#ef4444"
                lbl.setStyleSheet(
                    f"font-size:10px; font-weight:600; color:{clr};"
                    "background:transparent; border:none;"
                    "font-family:'Inter','Segoe UI',sans-serif;"
                )
                lbl.setText(text)

        all_ps = filtered_projects
        active_fn  = lambda p: self._project_effective_status(p) in ("Not Started", "In Progress")
        done_fn    = lambda p: self._project_effective_status(p) in ("Completed Not Invoiced", "Completed & Invoiced")
        hold_fn    = lambda p: self._project_effective_status(p) == "On Hold"
        overdue_fn = lambda p: self._project_effective_status(p) not in ("Paid", "Cancelled", "Completed Not Invoiced", "Completed & Invoiced", "On Hold") and p.get("due_date", "")

        for attr, fn in [
            ("total_projects_trend",  None),
            ("active_projects_trend", active_fn),
            ("completed_trend",       done_fn),
            ("on_hold_trend",         hold_fn),
            ("overdue_trend",         overdue_fn),
        ]:
            this = _month_count(all_ps, now.year, now.month, fn)
            last = _month_count(all_ps, ly, lm, fn)
            text, up = _pct_trend(this, last)
            _set_trend(attr, text, up)

        self.update_status_filter_counts()
    
    def show_project_dialog(self):
        """Show project creation in a focused dialog so the dashboard layout stays stable."""
        dialog = ProjectDialog(
            self.main_window,
            self,
            project_data=None,
            firebase_available=FIREBASE_AVAILABLE
        )
        result = dialog.exec_()
        if result == QtWidgets.QDialog.Accepted:
            self.load_projects()
            self.update_stats()

    def show_inline_project_editor(self, project_data=None):
        """Open an existing project in a focused dialog instead of expanding the dashboard."""
        dialog = ProjectDialog(self.main_window, self, project_data=project_data, firebase_available=FIREBASE_AVAILABLE)
        result = dialog.exec_()
        if result == QtWidgets.QDialog.Accepted:
            self.load_projects()
            self.update_stats()
    
    def open_projects_pdf_export_dialog(self):
        """Open PDF/Excel export dialog for projects"""
        try:
            available_dates = []
            for project in self.generated_projects:
                try:
                    project_date = datetime.strptime(project.get('date_received', ''), "%m-%d-%Y")
                    available_dates.append(project_date)
                except (ValueError, TypeError):
                    continue
            
            dialog = ProjectsExportDialog(self, available_dates)
            result = dialog.exec_()
            
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                export_params = dialog._export_params
                if export_params["type"] == "pdf":
                    self.perform_projects_pdf_export(export_params)
                elif export_params["type"] == "excel":
                    self.perform_projects_excel_export(export_params)
                        
        except Exception as e:
            _log.warning("Error opening export dialog: %s", e)
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error opening export dialog: {str(e)}")
    
    def clear_selection_colors(self):
        """Clear selection colors from all rows"""
        for row in range(self.projects_table.rowCount()):
            for col in range(self.projects_table.columnCount()):
                item = self.projects_table.item(row, col)
                if item:
                    item.setBackground(QtGui.QColor("#ffffff"))
                    item.setForeground(QtGui.QColor("#2c3e50"))
    
    def sort_by_created_desc(self, projects):
        def get_created_date(project):
            created_date = project.get('created_at', '')
            if not created_date:
                return datetime.min

            try:
                # ✅ KEEP FULL DATE + TIME
                return datetime.fromisoformat(created_date.replace("Z", ""))
            except:
                return datetime.min

        return sorted(projects, key=get_created_date, reverse=True)

    def update_projects_table(self, projects_list=None):
        # ✅ SORT BY CREATED DATE (NEWEST FIRST)
        def get_created_date(project):
            created_date = project.get('created_at', '')
            if not created_date:
                return datetime.min
            try:
                return datetime.fromisoformat(created_date.replace("Z", ""))
            except:
                return datetime.min
    

        """Update the projects table with generated projects"""
        if projects_list is None:
            projects_list = self.generated_projects
        projects_list = self.sort_by_created_desc(projects_list)
        
        valid_projects = []
        for project in projects_list:
            project_number = project.get("project_number", "")
            if project_number and project_number != "-0-0003":
                valid_projects.append(project)
        
        self.projects_table.clearContents()
        self.projects_table.setRowCount(len(valid_projects))

        for row, project in enumerate(valid_projects):
            bg_color = QtGui.QColor("#ffffff")

            # S.No.
            sno_item = QtWidgets.QTableWidgetItem(str(row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            sno_item.setForeground(QtGui.QColor('#2c3e50'))
            sno_item.setFont(QtGui.QFont("Inter", 9))
            sno_item.setBackground(bg_color)
            self.projects_table.setItem(row, 0, sno_item)

            # Project Number — dot bullet + clean Inter font widget
            pn_widget = self._create_project_number_widget(project["project_number"], project)
            # Store project data on a hidden item so row-click handlers can retrieve it
            _pn_item = QtWidgets.QTableWidgetItem("")
            _pn_item.setData(QtCore.Qt.UserRole, project)
            _pn_item.setToolTip(project["project_number"])
            self.projects_table.setItem(row, self.COL_PROJECT_NUMBER, _pn_item)
            self.projects_table.setCellWidget(row, self.COL_PROJECT_NUMBER, pn_widget)
            
            # Project Name
            name_item = QtWidgets.QTableWidgetItem(project.get("project_name", ""))
            name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            name_item.setForeground(QtGui.QColor('#0f172a'))
            name_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            name_item.setBackground(bg_color)
            self._set_project_item_readability(name_item, project.get("project_name", ""), emphasis=True, left=True)
            self.projects_table.setItem(row, 2, name_item)
            
            # Client
            company_item = QtWidgets.QTableWidgetItem(project.get("company", ""))
            company_item.setTextAlignment(QtCore.Qt.AlignCenter)
            company_item.setForeground(QtGui.QColor('#0f172a'))
            company_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            company_item.setBackground(bg_color)
            self._set_project_item_readability(company_item, project.get("company", ""), emphasis=True, left=True)
            self.projects_table.setItem(row, 3, company_item)
            
            # PO/WO Number
            po_wo_item = QtWidgets.QTableWidgetItem(project.get("po_wo_number", ""))
            po_wo_item.setTextAlignment(QtCore.Qt.AlignCenter)
            po_wo_item.setForeground(QtGui.QColor('#7f8c8d'))
            po_wo_item.setFont(QtGui.QFont("Inter", 9))
            po_wo_item.setBackground(bg_color)
            self._set_project_item_readability(po_wo_item, project.get("po_wo_number", ""), muted=True, left=True)
            self.projects_table.setItem(row, 4, po_wo_item)
            
            # Received Date
            date_received_item = QtWidgets.QTableWidgetItem(project.get("date_received", ""))
            date_received_item.setTextAlignment(QtCore.Qt.AlignCenter)
            date_received_item.setForeground(QtGui.QColor('#0f172a'))
            date_received_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            date_received_item.setBackground(bg_color)
            self._set_project_item_readability(date_received_item, project.get("date_received", ""), emphasis=True)
            self.projects_table.setItem(row, 5, date_received_item)
            
            # Plant
            plant_item = QtWidgets.QTableWidgetItem(project.get("plant", ""))
            plant_item.setTextAlignment(QtCore.Qt.AlignCenter)
            plant_item.setForeground(QtGui.QColor('#7f8c8d'))
            plant_item.setFont(QtGui.QFont("Inter", 9))
            plant_item.setBackground(bg_color)
            self._set_project_item_readability(plant_item, project.get("plant", ""), muted=True, left=True)
            self.projects_table.setItem(row, 6, plant_item)
            
            # Sales Person — avatar + name widget
            self.projects_table.setCellWidget(
                row, self.COL_SALES,
                self._create_sales_person_widget(project.get("sales", ""))
            )

            # Price (Project Amount)
            amount = project.get("project_amount", 0.0)
            amount_item = QtWidgets.QTableWidgetItem(Currency.format(amount))
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            if amount != 0.0:
                amount_item.setForeground(QtGui.QColor('#27ae60'))
                amount_item.setFont(QtGui.QFont("Inter", 11))
            else:
                amount_item.setForeground(QtGui.QColor('#95a5a6'))
            amount_item.setBackground(bg_color)
            amount_item.setToolTip(Currency.format(amount))
            amount_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            self.projects_table.setItem(row, self.COL_PRICE, amount_item)
            self.projects_table.setItem(row, self.COL_PAID, self._create_paid_amount_item(project, bg_color))
            self.projects_table.setItem(row, self.COL_REMAINING, self._create_remaining_due_item(project, bg_color))
            self.projects_table.setCellWidget(row, self.COL_PAYMENT, self._create_payment_cell_widget(project))
            
            # Start Date
            start_date_item = QtWidgets.QTableWidgetItem(project.get("start_date", ""))
            start_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            start_date_item.setForeground(QtGui.QColor('#0f172a'))
            start_date_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            start_date_item.setBackground(bg_color)
            self._set_project_item_readability(start_date_item, project.get("start_date", ""), emphasis=True)
            self.projects_table.setItem(row, self.COL_START_DATE, start_date_item)
            
            # Due Date
            due_date_item = QtWidgets.QTableWidgetItem(project.get("due_date", ""))
            due_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            due_date_item.setForeground(QtGui.QColor('#0f172a'))
            due_date_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            due_date_item.setBackground(bg_color)
            self._set_project_item_readability(due_date_item, project.get("due_date", ""), emphasis=True)
            self.projects_table.setItem(row, self.COL_DUE_DATE, due_date_item)
            
            # Status pill badge
            self.projects_table.setCellWidget(row, self.COL_STATUS, self._build_project_status_badge(project))

            # Enhanced action buttons
            self.add_enhanced_action_buttons(row, project)
        
        for row in range(self.projects_table.rowCount()):
            self.projects_table.setRowHeight(row, 65)
        self.projects_table.verticalHeader().setDefaultSectionSize(65)
        if hasattr(self.projects_table, "refresh_frozen_project_column"):
            self.projects_table.refresh_frozen_project_column(195)
        self.adjust_projects_table_height()
        
        self.update_stats()
        self.update_results_label()
        
        self.projects_table.clearSelection()
        self.selected_projects.clear()
        self.load_selected_btn.setEnabled(False)
        
    def project_number_sort_key(self, project):
        """
        Sort project numbers with suffix support
        Format: YYMMSEQ[_suffix]
        Examples: 2401001, 2401001_1000, 2401001_a, 2401001_3a, 2401001_3a1000
        """
        import re
        job_num = project.get('project_number', '').upper()
        
        if not job_num:
            return (0, 0, 0, 0, 0, 0, 0, 0)
        
        pattern = r'^(\d{2})(\d{2})(\d{3})(?:_?([a-zA-Z]+)?(\d+)?)?(?:_?(\d+)([a-zA-Z]+)?(\d+)?)?$'
        match = re.match(pattern, job_num, re.IGNORECASE)
        
        if not match:
            return (0, 0, 0, 0, 0, 0, 0, 0)
        
        year = match.group(1)
        month = match.group(2)
        main_seq = match.group(3)
        alpha1 = (match.group(4) or '').lower()
        num1 = match.group(5) or ''
        num2 = match.group(6) or ''
        alpha2 = (match.group(7) or '').lower()
        num3 = match.group(8) or ''
        
        year_value = -int(year)
        month_value = -int(month)
        
        try:
            seq_value = -int(main_seq.lstrip('0')) if main_seq.lstrip('0') else 0
        except:
            seq_value = 0
        
        variant_priority = 0
        
        if num2 and alpha2 and num3:
            variant_priority = -6
        elif num2 and alpha2 and not num3:
            variant_priority = -5
        elif num2 and not alpha2:
            variant_priority = -4
        elif alpha1 and num1 and not (num2 or alpha2):
            variant_priority = -3
        elif alpha1 and not num1:
            variant_priority = -2
        elif not (alpha1 or num1 or num2 or alpha2 or num3):
            variant_priority = -1
        
        alpha1_value = 0
        if alpha1:
            for i, char in enumerate(reversed(alpha1)):
                char_val = ord(char) - 96
                alpha1_value += char_val * (26 ** i)
        alpha1_value = -alpha1_value
        
        num1_value = 0
        try:
            num1_value = -int(num1) if num1 else 0
        except:
            num1_value = 0
        
        num2_value = 0
        try:
            num2_value = -int(num2) if num2 else 0
        except:
            num2_value = 0
        
        alpha2_value = 0
        if alpha2:
            for i, char in enumerate(reversed(alpha2)):
                char_val = ord(char) - 96
                alpha2_value += char_val * (26 ** i)
        alpha2_value = -alpha2_value
        
        num3_value = 0
        try:
            num3_value = -int(num3) if num3 else 0
        except:
            num3_value = 0
        
        return (
            year_value,
            month_value,
            seq_value,
            variant_priority,
            num2_value,
            alpha2_value,
            num3_value,
            alpha1_value,
            num1_value
        )
    
    # Short readable labels for the status badge (full name in tooltip)
    _STATUS_ABBR = {
        "Not Started":            "Not Started",
        "In Progress":            "In Progress",
        "On Hold":                "On Hold",
        "Completed Not Invoiced": "Completed",
        "Completed & Invoiced":   "Invoiced ✓",
        "Paid":                   "Paid ✓",
        "Cancelled":              "Cancelled",
    }

    def style_project_status_combo(self, combo: QtWidgets.QComboBox, status: str):
        """Style the status combo box as a single unified colored badge (no visible arrow)."""
        colors = {
            "Not Started":            ("#f1f5f9", "#475569", "#cbd5e1"),
            "In Progress":            ("#fff7ed", "#c2410c", "#fed7aa"),
            "On Hold":                ("#faf5ff", "#7e22ce", "#d8b4fe"),
            "Completed Not Invoiced": ("#fef2f2", "#b91c1c", "#fca5a5"),
            "Completed & Invoiced":   ("#f0fdf4", "#166534", "#86efac"),
            "Paid":                   ("#eff6ff", "#1e40af", "#93c5fd"),
            "Cancelled":              ("#f9fafb", "#6b7280", "#d1d5db"),
        }
        bg, fg, border = colors.get(status, colors["Not Started"])
        # Use object name selector to beat the global app stylesheet
        combo.setStyleSheet(f"""
            QComboBox#projectStatusCombo {{
                background: {bg};
                color: {fg};
                border: 1.5px solid {border};
                border-radius: 8px;
                padding: 3px 8px 3px 10px;
                font-size: 12px;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }}
            QComboBox#projectStatusCombo::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: right center;
                width: 0px;
                border: none;
                background: transparent;
                image: none;
            }}
            QComboBox#projectStatusCombo::down-arrow {{
                width: 0; height: 0; image: none; border: none;
            }}
            QComboBox#projectStatusCombo QAbstractItemView {{
                background: #ffffff;
                border: 1px solid #d1d5db;
                border-radius: 8px;
                selection-background-color: #e6f6f4;
                selection-color: #0f172a;
                font-size: 12px;
                padding: 4px;
            }}
        """)
        # Show abbreviated label, full status in tooltip
        abbr = self._STATUS_ABBR.get(status, status)
        idx = combo.findText(status)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        combo.setObjectName("projectStatusCombo")
        combo.setToolTip("")
        combo.setEditable(False)
    
    def _center_cell_widget(self, widget: QtWidgets.QWidget) -> QtWidgets.QWidget:
        """Wrap a widget in a transparent container that centers it vertically in a table cell."""
        container = QtWidgets.QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        lay = QtWidgets.QHBoxLayout(container)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.setAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
        lay.addWidget(widget)
        return container

    def _build_status_combo(self, project: dict) -> QtWidgets.QComboBox:
        """Build a styled status combo for a project row."""
        current_status = project.get("status", "Not Started")
        combo = QtWidgets.QComboBox()
        combo.addItems(self.PROJECT_STATUSES)
        idx = combo.findText(current_status)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        self.style_project_status_combo(combo, current_status)
        combo.setFixedWidth(120)
        combo.setFixedHeight(26)

        combo.currentTextChanged.connect(
            lambda status, project_data=project, c=combo:
            self.on_project_status_changed(project_data, status, c)
        )
        return combo

    def _build_status_cell_widget(self, project: dict) -> QtWidgets.QWidget:
        """Return a centered scroll-proof dropdown combo for the status column."""
        current_status = project.get("status", "Not Started")
        if current_status == "Completed":
            if is_project_fully_paid(project):
                current_status = "Paid"
            else:
                from payment_tracker import get_payment_tracker as _gpt
                _pn = project.get("project_number", "")
                _has = any(
                    float(_p.amount) > 0 and (_p.payment_stage or "").strip().lower() != "tax"
                    for _p in _gpt().get_project_payments(_pn)
                )
                current_status = "In Progress" if _has else "Not Started"

        def _style(st):
            bg, fg, bdr = self._PROJECT_STATUS_PALETTE.get(
                st, self._PROJECT_STATUS_PALETTE["Not Started"])
            return f"""
                QComboBox {{
                    background: {bg}; color: {fg};
                    border: 1.5px solid {bdr};
                    border-radius: 8px;
                    padding: 4px 8px 4px 12px;
                    font-size: 12px; font-weight: 700;
                    font-family: 'Inter', 'Segoe UI', sans-serif;
                }}
                QComboBox::drop-down {{
                    subcontrol-origin: padding;
                    subcontrol-position: right center;
                    width: 20px;
                    border-left: 1px solid {bdr};
                    border-top-right-radius: 7px;
                    border-bottom-right-radius: 7px;
                }}
                QComboBox QAbstractItemView {{
                    background: white;
                    border: 1px solid #d1d5db;
                    selection-background-color: #e6f6f4;
                    selection-color: #0f172a;
                    font-size: 12px; padding: 4px;
                }}
            """

        combo = _NoScrollComboBox()
        combo.addItems(self.PROJECT_STATUSES)
        idx = combo.findText(current_status)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        combo.setFixedHeight(34)
        combo.setMinimumWidth(175)
        combo.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        combo.setStyleSheet(_style(current_status))

        def _on_changed(new_status):
            combo.setStyleSheet(_style(new_status))
            project['status'] = new_status
            project['status_manual'] = True
            if FIREBASE_AVAILABLE:
                FirebaseManager.save_project(project)
            self.update_stats()
            ws_combo = getattr(self, '_workspace_status_combo', None)
            ws_data  = getattr(self, '_workspace_project_data', None)
            if ws_combo and ws_data and ws_data.get('project_number') == project.get('project_number'):
                ws_combo.blockSignals(True)
                i = ws_combo.findText(new_status)
                if i >= 0:
                    ws_combo.setCurrentIndex(i)
                self.style_project_status_combo(ws_combo, new_status)
                ws_combo.blockSignals(False)
                ws_data['status'] = new_status

        combo.currentTextChanged.connect(_on_changed)
        return self._center_cell_widget(combo)

    # Colours shared by badge + existing combo styler
    _PROJECT_STATUS_PALETTE = {
        "Not Started":            ("#f1f5f9", "#475569", "#cbd5e1"),
        "In Progress":            ("#fff7ed", "#c2410c", "#fed7aa"),
        "On Hold":                ("#faf5ff", "#7e22ce", "#d8b4fe"),
        "Completed Not Invoiced": ("#fef2f2", "#b91c1c", "#fca5a5"),
        "Completed & Invoiced":   ("#f0fdf4", "#166534", "#86efac"),
        "Paid":                   ("#eff6ff", "#1e40af", "#93c5fd"),
        "Cancelled":              ("#f9fafb", "#6b7280", "#d1d5db"),
    }

    def _style_project_badge(self, btn: QtWidgets.QPushButton, status: str):
        """Apply pill styling identical to the quotes table status badge."""
        from app_theme import CHEVRON_URL
        bg, fg, border = self._PROJECT_STATUS_PALETTE.get(
            status, self._PROJECT_STATUS_PALETTE["Not Started"])
        btn.setText(status)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg};
                color: {fg};
                border: 1px solid {border};
                border-radius: 7px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 30px 0 14px;
                text-align: left;
            }}
            QPushButton:hover {{
                border-width: 1.5px;
                opacity: 0.9;
            }}
            QPushButton::menu-indicator {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                subcontrol-origin: padding;
                subcontrol-position: center right;
                right: 12px;
            }}
        """)

    def _build_project_status_badge(self, project: dict) -> QtWidgets.QWidget:
        """Return a centred QPushButton pill with a status-change menu."""
        current_status = project.get("status", "Not Started")
        # Legacy "Completed" that wasn't yet migrated: derive display status from
        # payment state so the badge never shows "Completed" to the user.
        if current_status == "Completed":
            if is_project_fully_paid(project):
                current_status = "Paid"
            else:
                from payment_tracker import get_payment_tracker as _gpt_b
                _pn_b = project.get("project_number", "")
                _has_b = any(
                    float(_pb.amount) > 0
                    and (_pb.payment_stage or "").strip().lower() != "tax"
                    for _pb in _gpt_b().get_project_payments(_pn_b)
                )
                current_status = "In Progress" if _has_b else "Not Started"

        container = QtWidgets.QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        lay = QtWidgets.QHBoxLayout(container)
        lay.setContentsMargins(3, 2, 3, 2)
        lay.setSpacing(0)
        lay.addStretch(1)

        btn = QtWidgets.QPushButton(current_status)
        btn.setFixedSize(155, 30)
        btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._style_project_badge(btn, current_status)

        from PyQt5.QtWidgets import QMenu, QAction
        menu = QMenu(btn)
        menu.setStyleSheet("""
            QMenu { background: white; border: 1px solid #d0d7de; border-radius: 8px;
                    padding: 4px 0; font-family: 'Inter','Segoe UI',sans-serif; font-size: 12px; }
            QMenu::item { padding: 7px 18px; color: #24292f; }
            QMenu::item:selected { background: #f6f8fa; color: #0969da; }
            QMenu::separator { height: 1px; background: #e5e7eb; margin: 3px 0; }
        """)
        groups = [
            ["Not Started", "In Progress"],
            ["On Hold", "Completed Not Invoiced", "Completed & Invoiced"],
            ["Paid"],
            ["Cancelled"],
        ]
        first = True
        for grp in groups:
            if not first:
                menu.addSeparator()
            first = False
            for s in grp:
                act = QAction(s, menu)
                act.triggered.connect(
                    lambda _, st=s, b=btn, pd=project:
                        self._apply_project_badge_status(st, b, pd))
                menu.addAction(act)
        btn.setMenu(menu)

        lay.addWidget(btn)
        lay.addStretch(1)
        return container

    def _apply_project_badge_status(self, new_status: str, btn: QtWidgets.QPushButton, project_data: dict):
        """Re-style badge and persist the status change."""
        self._style_project_badge(btn, new_status)
        project_data['status'] = new_status
        if FIREBASE_AVAILABLE:
            FirebaseManager.save_project(project_data)
        self.update_stats()
        # Keep workspace status combo in sync if the same project is open
        ws_combo = getattr(self, '_workspace_status_combo', None)
        ws_data  = getattr(self, '_workspace_project_data', None)
        if ws_combo and ws_data and ws_data.get('project_number') == project_data.get('project_number'):
            ws_combo.blockSignals(True)
            idx = ws_combo.findText(new_status)
            if idx >= 0:
                ws_combo.setCurrentIndex(idx)
            self.style_project_status_combo(ws_combo, new_status)
            ws_combo.blockSignals(False)
            ws_data['status'] = new_status

    def on_project_status_changed(self, project_data: dict, new_status: str, combo: QtWidgets.QComboBox):
        """Handle project status changes and update styling"""
        try:
            project_data['status'] = new_status
            project_data['status_manual'] = True  # Mark as user-set — suppress auto-overrides
            self.style_project_status_combo(combo, new_status)

            if FIREBASE_AVAILABLE:
                FirebaseManager.save_project(project_data)

            _log.info("(converted from print, see git history)")
            self.update_stats()

            # Sync status combo in open workspace if it shows the same project
            ws_combo = getattr(self, '_workspace_status_combo', None)
            ws_data = getattr(self, '_workspace_project_data', None)
            if ws_combo and ws_data and ws_data.get('project_number') == project_data.get('project_number'):
                ws_combo.blockSignals(True)
                idx = ws_combo.findText(new_status)
                if idx >= 0:
                    ws_combo.setCurrentIndex(idx)
                self.style_project_status_combo(ws_combo, new_status)
                ws_combo.blockSignals(False)
                ws_data['status'] = new_status
        except Exception as e:
            _log.error(f"Error updating project status: {e}")
    
    def add_enhanced_action_buttons(self, row, project):
        """Three icon buttons: view (eye), edit (pencil), more (three-dot menu)."""
        actions_widget = QtWidgets.QWidget()
        actions_widget.setStyleSheet("background: transparent; border: none;")
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(4, 0, 4, 0)
        actions_layout.setSpacing(6)
        actions_layout.setAlignment(QtCore.Qt.AlignCenter)

        _btn_base = """
            QToolTip {{
                background: #ffffff;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 11px;
                font-weight: 600;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }}
            QPushButton {{
                background: {bg};
                color: {fg};
                border: 1px solid {border};
                border-radius: 7px;
                font-size: {fs}px;
                font-weight: 700;
                font-family: 'Segoe UI Symbol', 'Segoe UI Emoji', 'Segoe UI', sans-serif;
                min-width: 28px; max-width: 28px;
                min-height: 30px; max-height: 30px;
                padding: 0;
            }}
            QPushButton:hover {{ background: {hover}; border-color: {hover_border}; }}
            QPushButton:pressed {{ background: {hover}; }}
        """

        def _make_icon_btn(text, tip, bg, fg, border, hover, hover_border, fs=14):
            b = QtWidgets.QPushButton(text)
            b.setFixedSize(28, 30)
            b.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            b.setToolTip(tip)
            b.setStyleSheet(_btn_base.format(
                bg=bg, fg=fg, border=border, hover=hover,
                hover_border=hover_border, fs=fs
            ))
            return b

        # View — eye shape using Unicode
        view_btn = _make_icon_btn(
            "◉", "View Project",
            "#eff6ff", "#2563eb", "#bfdbfe", "#dbeafe", "#93c5fd", 13
        )
        view_btn.clicked.connect(lambda checked=False, p=project: self.show_project_workspace(p))
        actions_layout.addWidget(view_btn)

        # Edit — pencil shape
        edit_btn = _make_icon_btn(
            "✎", "Edit Project",
            "#f0fdf4", "#16a34a", "#bbf7d0", "#dcfce7", "#86efac", 15
        )
        edit_btn.clicked.connect(lambda checked=False, p=project: self.edit_single_project(p))
        actions_layout.addWidget(edit_btn)

        # More — vertical dots
        more_btn = _make_icon_btn(
            "⋮", "More actions",
            "#f8fafc", "#475569", "#e2e8f0", "#f1f5f9", "#cbd5e1", 16
        )

        menu = QtWidgets.QMenu(more_btn)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 6px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QMenu::item {
                color: #334155;
                padding: 7px 16px;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 700;
                min-width: 160px;
            }
            QMenu::item:selected { background: #f1f5f9; color: #0f172a; }
            QMenu::separator { height: 1px; background: #f1f5f9; margin: 4px 8px; }
        """)
        invoice_action = menu.addAction("Generate Stage Invoice")
        invoice_action.triggered.connect(
            lambda checked=False, p=project: self.load_projects_to_invoice_direct([p]))
        history_action = menu.addAction("Payment History")
        history_action.triggered.connect(
            lambda checked=False, p=project: self.show_payment_history(p))
        menu.addSeparator()
        delete_action = menu.addAction("Delete Project")
        delete_action.triggered.connect(lambda checked=False, p=project: self.delete_single_project(p))

        more_btn.clicked.connect(
            lambda: menu.exec_(more_btn.mapToGlobal(QtCore.QPoint(0, more_btn.height() + 2)))
        )
        actions_layout.addWidget(more_btn)

        self.projects_table.setCellWidget(row, self.COL_ACTIONS, actions_widget)
    
    def show_payment_history(self, project):
        """Show payment history dialog for a project"""
        project_number = project.get("project_number", "")
        project_name = project.get("project_name", "")
        total_amount = float(project.get("project_amount", 0))
        
        if not project_number:
            QtWidgets.QMessageBox.warning(self, "Error", "Project number is required.")
            return
        
        dialog = PaymentHistoryDialog(
            self,
            project_number,
            project_name,
            total_amount,
            invoice_rows=self._get_project_payment_schedule_rows(project),
        )
        dialog.payments_changed.connect(self._on_payment_changed)
        dialog.exec_()
        # Defer so the UI returns to the caller before the heavier project-list
        # rebuild runs; avoids the "can't click other tabs" lag after close.
        QtCore.QTimer.singleShot(100, self.filter_projects)
        # NOTE: Do NOT call _refresh_finance_tabs() here — it starts a Firebase
        # read immediately before _bg_update has written, creating a race that
        # overwrites the fresh annual_revenue_data with stale data.
        # Annual summary is refreshed event-driven from _trigger_annual_summary_refresh
        # in payment_tracker._bg_update (fires after write completes).
        # Paid revenues + stats cards are refreshed by the 1500ms timer in _on_payment_changed.

    def _on_payment_changed(self, project_number: str):
        """Live-refresh the Payment cell and all finance tabs when a payment is added/edited/deleted."""
        # Immediately update annual summary from in-memory payments — zero latency,
        # no Firebase read. The signal-based refresh confirms later via Firebase.
        try:
            bs_tab = getattr(self.main_window, "balance_sheet_tab", None)
            if bs_tab and hasattr(bs_tab, "update_annual_summary_from_payments"):
                bs_tab.update_annual_summary_from_payments()
        except Exception:
            pass

        # Update project table cell immediately
        _found_row = -1
        _found_pd = None
        for row in range(self.projects_table.rowCount()):
            item = self.projects_table.item(row, 1)
            if item is None:
                continue
            project_data = item.data(QtCore.Qt.UserRole)
            if project_data and project_data.get("project_number") == project_number:
                bg_color = QtGui.QColor("#ffffff")
                self.projects_table.setItem(row, self.COL_PAID, self._create_paid_amount_item(project_data, bg_color))
                self.projects_table.setItem(row, self.COL_REMAINING, self._create_remaining_due_item(project_data, bg_color))
                self.projects_table.setCellWidget(
                    row, self.COL_PAYMENT, self._create_payment_cell_widget(project_data)
                )
                _found_row = row
                _found_pd = project_data
                break

        # Auto-update project status based on payment state.
        #
        # Rules (upward-only, no auto-demotion):
        #   • User-stable statuses are NEVER auto-changed:
        #       "On Hold", "Cancelled", "Completed Not Invoiced", "Completed & Invoiced"
        #   • All stages fully paid  → "Paid"   (if current is Not Started / In Progress)
        #   • Partial payments exist → "In Progress"  (only if current is "Not Started")
        #   • Deleting a payment never demotes "Paid" or "In Progress" automatically.
        #
        _USER_STABLE = {
            "On Hold", "Cancelled",
            "Completed Not Invoiced", "Completed & Invoiced",
        }
        if _found_pd is not None:
            _cur = (_found_pd.get("status") or "Not Started").strip()
            _is_manual = bool(_found_pd.get("status_manual"))
            if _cur not in _USER_STABLE:
                _new_status = None

                if is_project_fully_paid(_found_pd):
                    # Fully paid always auto-promotes to "Paid", even if user set status manually
                    if _cur != "Paid":
                        _new_status = "Paid"
                elif not _is_manual and _cur == "Not Started":
                    # Only auto-promote to "In Progress" if user hasn't manually set a status
                    from payment_tracker import get_payment_tracker as _gpt2
                    _has_pay = any(
                        float(p.amount) > 0
                        and (p.payment_stage or "").strip().lower() != "tax"
                        for p in _gpt2().get_project_payments(project_number)
                    )
                    if _has_pay:
                        _new_status = "In Progress"

                if _new_status:
                    _found_pd["status"] = _new_status
                    # Auto-"Paid" clears the manual flag; other auto-changes keep it
                    if _new_status == "Paid":
                        _found_pd.pop("status_manual", None)
                    self.projects_table.setCellWidget(
                        _found_row, self.COL_STATUS,
                        self._build_project_status_badge(_found_pd)
                    )
                    _pn_cap, _st_cap = project_number, _new_status
                    _clear_manual = (_new_status == "Paid")
                    def _set_status_bg(_p=_pn_cap, _st=_st_cap, _clr=_clear_manual):
                        try:
                            from firebase_admin import db as _fdb
                            _d = _fdb.reference('/projects') \
                                     .order_by_child('project_number').equal_to(_p).get()
                            if _d:
                                _pid = list(_d.keys())[0]
                                _upd = {'status': _st, 'updated_at': datetime.now().isoformat()}
                                if _clr:
                                    _upd['status_manual'] = None
                                _fdb.reference(f'/projects/{_pid}').update(_upd)
                        except Exception:
                            pass
                    threading.Thread(target=_set_status_bg, daemon=True).start()

        # Rebuild workspace cards immediately if this project is currently shown.
        # Use in-memory tracker.payments directly — add/delete/update all mutate it
        # synchronously, so it is always up-to-date.  Calling _load_payments() here
        # would race with in-flight Firebase writes and wipe payments from memory.
        ws_data = getattr(self, "_ws_project_data", None)
        if ws_data and ws_data.get("project_number") == project_number:
            def _rebuild_from_payment(_wd=ws_data):
                try:
                    self.show_project_workspace(_wd)
                except Exception:
                    pass
            QtCore.QTimer.singleShot(50, _rebuild_from_payment)

        def _refresh_all_revenue():
            try:
                bs_tab = getattr(self.main_window, "balance_sheet_tab", None)
                if bs_tab and hasattr(bs_tab, "_refresh_all_revenue_background"):
                    bs_tab._refresh_all_revenue_background()
            except Exception:
                pass

        # Annual summary is refreshed event-driven from _bg_update in payment_tracker.py
        # (fires immediately after Firebase sync + orphan cleanup complete).
        # Paid revenues need _recompute_invoice_status to finish first (~1s), so keep timer.
        QtCore.QTimer.singleShot(1500, _refresh_all_revenue)

        # Auto-sync invoice statuses in background thread after a short delay
        # so Firebase reads don't block the main Qt thread.
        QtCore.QTimer.singleShot(
            400,
            lambda pn=project_number: threading.Thread(
                target=lambda: self._auto_sync_invoice_statuses(pn),
                daemon=True,
            ).start(),
        )
        # NOTE: _refresh_finance_tabs() intentionally NOT called here.
        # It would race with _bg_update (Firebase write takes 1-2s, 800ms debounce
        # is not enough) and overwrite annual_revenue_data with stale data.
        # Annual summary is updated event-driven by _trigger_annual_summary_refresh
        # after the write completes. Paid revenues + stats are covered by the
        # 1500ms _refresh_all_revenue timer above.

    def _refresh_finance_tabs(self):
        """Debounced refresh — coalesces rapid calls into one 800ms deferred refresh."""
        if not hasattr(self, "_finance_refresh_timer"):
            self._finance_refresh_timer = QtCore.QTimer(self)
            self._finance_refresh_timer.setSingleShot(True)
            self._finance_refresh_timer.timeout.connect(self._do_refresh_finance_tabs)
        self._finance_refresh_timer.start(800)  # restart timer on each call

    def _on_invoice_sync_done(self):
        """Called on the main thread after _auto_sync_invoice_statuses finishes.
        Finance data is reloaded in a background thread so the UI stays responsive."""
        main_win = self.main_window
        if not main_win:
            return

        tabs_to_refresh = []
        for attr in ("balance_sheet_tab", "finance_overview_tab"):
            tab = getattr(main_win, attr, None)
            if tab and hasattr(tab, "_fetch_data_background") and hasattr(tab, "_apply_fetched_data_ui"):
                tabs_to_refresh.append(tab)

        def _bg():
            # Pure Firebase reads — no Qt calls allowed here
            for tab in tabs_to_refresh:
                try:
                    tab._fetch_data_background()
                except Exception:
                    pass
            QtCore.QTimer.singleShot(0, _ui)

        def _ui():
            # Qt widget updates — main thread only
            for tab in tabs_to_refresh:
                try:
                    tab._apply_fetched_data_ui()
                except Exception:
                    pass
            hist = getattr(main_win, "history_tab", None)
            if hist and hasattr(hist, "refresh_invoices_immediately"):
                try:
                    if hist.isVisible():
                        hist.refresh_invoices_immediately()
                except Exception:
                    pass
            # Workspace refresh — use in-memory tracker.payments directly.
            # Do NOT call _load_payments() here: tracker.payments is already
            # up-to-date (add/delete/update all mutate it synchronously before
            # starting any Firebase write thread).  Calling _load_payments() at
            # this point could race with any still-in-flight Firebase writes and
            # overwrite correct memory with a partial snapshot, making payments
            # for P2/P3 disappear from the workspace display.
            ws_data = getattr(self, "_ws_project_data", None)
            if ws_data:
                try:
                    self.show_project_workspace(ws_data)
                except Exception as _e:
                    _log.warning("Could not refresh workspace from sync done: %s", _e)

        threading.Thread(target=_bg, daemon=True).start()

    def _do_refresh_finance_tabs(self):
        """Refresh finance tabs — Firebase data loads in a background thread so the
        main thread (and all UI interactions) stay fully responsive."""
        main_win = self.main_window
        if not main_win:
            return

        # Collect tab references now (main thread is safe for attribute lookup).
        # Only include tabs that have the split _fetch_data_background / _apply_fetched_data_ui
        # API so we guarantee no Qt calls happen on the background thread.
        tabs_to_refresh = []
        for attr in ("balance_sheet_tab", "expenses_tab", "finance_overview_tab"):
            tab = getattr(main_win, attr, None)
            if tab and hasattr(tab, "_fetch_data_background") and hasattr(tab, "_apply_fetched_data_ui"):
                tabs_to_refresh.append(tab)

        def _bg_load():
            # Pure Firebase reads — no Qt calls allowed in this thread
            for tab in tabs_to_refresh:
                try:
                    tab._fetch_data_background()
                except Exception as _e:
                    _log.warning("Background finance data load failed: %s", _e)
            # Marshal UI rebuild back to the main thread
            QtCore.QTimer.singleShot(0, _ui_update)

        def _ui_update():
            # Qt widget updates — must run on the main thread
            for tab in tabs_to_refresh:
                try:
                    tab._apply_fetched_data_ui()
                except Exception as _e:
                    _log.warning("Finance tab UI update failed: %s", _e)
            # Invoice history: only refresh if visible (its load_clients does Firebase reads)
            hist = getattr(main_win, "history_tab", None)
            if hist and hasattr(hist, "refresh_invoices_immediately"):
                try:
                    if hist.isVisible():
                        hist.refresh_invoices_immediately()
                except Exception as _e:
                    _log.warning("Could not refresh history_tab: %s", _e)

        threading.Thread(target=_bg_load, daemon=True).start()

    # kept for backward-compat callers in this file
    def _refresh_balance_sheet_tab(self):
        self._refresh_finance_tabs()

    # ------------------------------------------------------------------ #
    # Auto-sync invoice status from payments                               #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _norm_stage(stage: str) -> str:
        """Collapse any stage label to a canonical keyword for payment matching."""
        lo = (stage or "").strip().lower()
        if any(x in lo for x in ("down payment", "deposit", "50%")):
            return "down_payment"
        if any(x in lo for x in ("remaining balance", "remaining", "balance due")):
            return "remaining_balance"
        if any(x in lo for x in ("1st installment", "first installment", "payment 1", "term 1")):
            return "installment_1"
        if any(x in lo for x in ("2nd installment", "second installment", "payment 2", "term 2")):
            return "installment_2"
        if any(x in lo for x in ("3rd installment", "third installment", "payment 3", "term 3")):
            return "installment_3"
        if any(x in lo for x in ("final payment", "4th", "fourth", "term 4")):
            return "final_payment"
        if any(x in lo for x in ("full amount", "full payment", "due payment")):
            return "full_amount"
        return lo

    def _auto_sync_invoice_statuses(self, project_number: str,
                                      target_invoice: str = ""):
        """Recalculate invoice status based on actual payments.

        Rules (strict — prevents cross-invoice contamination):
        1. Only count payments EXPLICITLY linked to THIS invoice (invoice_number matches).
        2. Also count UNLINKED payments (invoice_number="") whose stage matches the
           invoice item's stage — and immediately link them to this invoice.
        3. NEVER count payments linked to a DIFFERENT invoice.
        4. Skip invoices whose current Firebase status is already "Paid" unless
           target_invoice is provided (i.e. leave completed invoices alone).
        5. If target_invoice is provided, only process that specific invoice
           (used when a new invoice is first saved so old invoices aren't touched).
        """
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()
            tracker._load_payments()
            invoices = FirebaseManager.load_invoices() or []

            for invoice_data in invoices:
                if not isinstance(invoice_data, dict):
                    continue
                meta = invoice_data.get("meta", {}) or {}
                invoice_number = meta.get("invoice_number", "")
                if not invoice_number:
                    continue

                # ── Rule 5: target filter ────────────────────────────────────
                if target_invoice and invoice_number != target_invoice:
                    continue

                # ── Rule 4: skip completed invoices unless explicitly targeted ─
                current_status = (meta.get("status") or "Unpaid").strip()
                if current_status == "Paid" and not target_invoice:
                    continue

                items = invoice_data.get("items", []) or []
                pns_in_invoice = [
                    str(it.get("project_number", "")).strip()
                    for it in items if it.get("project_number")
                ]
                if project_number not in pns_in_invoice:
                    continue

                invoice_total = 0.0
                total_paid_for_invoice = 0.0
                all_payment_dates: list = []
                # seen is global across all items so a payment is never counted
                # twice (once per item) even when it matches multiple items.
                seen: set = set()

                for item in items:
                    pn = str(item.get("project_number", "")).strip()
                    raw_amt = (
                        item.get("payment_due")
                        if item.get("payment_due") not in (None, "", 0, 0.0)
                        else item.get("total", item.get("unit_price", 0))
                    )
                    item_amount = float(raw_amt or 0)
                    invoice_total += item_amount

                    if not pn or item_amount <= 0:
                        continue

                    item_stage = self._norm_stage(item.get("payment_category", ""))
                    project_payments = tracker.get_project_payments(pn)

                    # ── Rule 2: link truly unlinked payments for this stage ───
                    # Only link if the payment has NO invoice_number yet.
                    # Never re-link payments already bound to another invoice.
                    for _p in project_payments:
                        p_inv = (_p.invoice_number or "").strip()
                        if (
                            not p_inv                                         # unlinked
                            and self._norm_stage(_p.payment_stage) == item_stage
                        ):
                            tracker.update_payment(_p.payment_id,
                                                   invoice_number=invoice_number)
                    project_payments = tracker.get_project_payments(pn)

                    # ── Rules 1 & 3: count ONLY this-invoice OR newly-unlinked ─
                    # CRITICAL: stage_match is ALWAYS required so that a payment
                    # for Stage 1 cannot be counted toward Stage 2 or any other
                    # item.  This prevents a single payment from inflating
                    # total_paid_for_invoice across multiple items and falsely
                    # marking the invoice as "Paid" before all stages are paid.
                    stage_paid = 0.0
                    for _p in project_payments:
                        if _p.payment_id in seen:
                            continue
                        p_inv = (_p.invoice_number or "").strip()
                        inv_match   = p_inv == invoice_number.strip()
                        is_unlinked = not p_inv
                        stage_match = self._norm_stage(_p.payment_stage) == item_stage

                        # Count only if stage matches AND payment belongs here
                        # (linked to this invoice, or still unlinked after rule 2)
                        if stage_match and (inv_match or is_unlinked):
                            stage_paid += float(_p.amount)
                            seen.add(_p.payment_id)
                            if _p.payment_date:
                                all_payment_dates.append(_p.payment_date)

                    total_paid_for_invoice += min(stage_paid, item_amount)

                # Include tax amount in the invoice total so that paying only
                # project item amounts never flips the invoice to "Paid" when
                # there is outstanding tax.  Tax payments (stage="Tax") are
                # counted separately here.
                try:
                    tax_amount = float(meta.get("tax_amount") or 0)
                    if tax_amount > 0.005:
                        invoice_total += tax_amount
                        for _pn in set(pns_in_invoice):
                            for _tp in tracker.get_project_payments(_pn):
                                if ((_tp.payment_stage or "").strip().lower() == "tax"
                                        and (_tp.invoice_number or "").strip()
                                        == invoice_number.strip()):
                                    total_paid_for_invoice += float(_tp.amount)
                                    if _tp.payment_date:
                                        all_payment_dates.append(_tp.payment_date)
                except Exception:
                    pass

                if invoice_total <= 0 or total_paid_for_invoice <= 0:
                    continue

                if total_paid_for_invoice >= invoice_total - 0.005:
                    new_status = "Paid"
                else:
                    new_status = "Partially Paid"

                latest_date = ""
                if all_payment_dates:
                    raw_latest = max(all_payment_dates)
                    for fmt in ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y"):
                        try:
                            latest_date = datetime.strptime(
                                raw_latest, fmt).strftime("%m-%d-%Y")
                            break
                        except (ValueError, TypeError):
                            pass
                    if not latest_date:
                        latest_date = raw_latest

                self._update_invoice_status_firebase(invoice_number, new_status,
                                                     latest_date)
                self._sync_invoice_to_balance_sheet(
                    invoice_number, new_status, latest_date,
                    total_paid_for_invoice, invoice_total,
                )

        except Exception as e:
            _log.warning("Error in _auto_sync_invoice_statuses for %s: %s",
                         project_number, e)

        # Signal the main thread to refresh UI (thread-safe cross-thread signal)
        try:
            self._invoice_sync_done.emit()
        except Exception:
            pass

    def _update_invoice_status_firebase(self, invoice_number: str, status: str, received_date: str):
        """Update invoice meta/status and meta/received_date in Firebase."""
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db as _db
            ref = _db.reference('/invoices')
            matches = ref.order_by_child('meta/invoice_number').equal_to(invoice_number).get()
            if not matches:
                return
            inv_id = list(matches.keys())[0]
            from datetime import timezone
            ref.child(inv_id).child('meta').update({
                'status': status,
                'received_date': received_date,
                'updated_at': datetime.now(timezone.utc).isoformat(),
            })
            _log.info("Auto-updated invoice %s → %s (received %s)", invoice_number, status, received_date)
        except Exception as e:
            _log.warning("Error updating invoice status in Firebase: %s", e)

    def _sync_invoice_to_balance_sheet(
        self, invoice_number: str, status: str,
        received_date: str, total_paid: float, invoice_total: float
    ):
        """Update the balance-sheet revenue entry for this invoice.

        Only called when total_paid > 0 (caller must ensure this).
        Sets has_payment_entries=True so update_annual_summary skips the
        invoice-level row (payment-tracker entries handle per-month breakdown).
        """
        if total_paid <= 0:
            # Nothing actually paid — do not change balance-sheet status.
            return
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db as _db
            from datetime import timezone
            rev_ref = _db.reference('revenue')
            all_rev = rev_ref.get() or {}
            for rev_id, rev in all_rev.items():
                if not isinstance(rev, dict):
                    continue
                if rev.get('is_invoice') and rev.get('invoice_number') == invoice_number:
                    unpaid = max(invoice_total - total_paid, 0.0)
                    update_data = {
                        'status': status,
                        'paid_amount': str(round(total_paid, 2)),
                        'unpaid_amount': str(round(unpaid, 2)),
                        'has_payment_entries': True,
                        'updated_at': datetime.now(timezone.utc).isoformat(),
                    }
                    # Always store latest payment date as received_date so balance
                    # sheet matches invoice history exactly for all paid statuses.
                    update_data['received_date'] = received_date or "N/A"
                    update_data['down_payment_received_date'] = received_date or "N/A"
                    rev_ref.child(rev_id).update(update_data)
                    _log.info(
                        "Synced invoice %s → %s (paid %.2f / %.2f)",
                        invoice_number, status, total_paid, invoice_total,
                    )
                    break
        except Exception as e:
            _log.warning("Error syncing invoice %s to balance sheet: %s",
                         invoice_number, e)

    def _build_tax_section_for_workspace(self, project_data):
        """Return a QFrame showing tax info for this project's invoices, or None."""
        project_number = project_data.get("project_number", "")
        try:
            invoices = FirebaseManager.load_invoices() or []
            tracker = get_payment_tracker()
            tax_entries = []
            for invoice_data in invoices:
                if not isinstance(invoice_data, dict):
                    continue
                meta = invoice_data.get("meta") or {}
                inv_no = meta.get("invoice_number", "")
                if not inv_no:
                    continue
                items = invoice_data.get("items") or []
                pns = [str(it.get("project_number", "")).strip() for it in items]
                if project_number not in pns:
                    continue
                try:
                    tax_amount = float(meta.get("tax_amount") or 0)
                except (TypeError, ValueError):
                    tax_amount = 0.0
                if tax_amount <= 0.005:
                    continue
                tax_paid = sum(
                    float(p.amount) for p in tracker.payments
                    if (p.payment_stage or "").strip().lower() == "tax"
                    and (p.invoice_number or "").strip() == inv_no.strip()
                )
                tax_entries.append({
                    "invoice_number": inv_no,
                    "tax_amount": tax_amount,
                    "tax_paid": tax_paid,
                    "is_paid": tax_paid >= tax_amount - 0.005,
                })
            if not tax_entries:
                return None

            card = QtWidgets.QFrame()
            card.setStyleSheet("""
                QFrame { background:#ffffff; border:1px solid #e2e8f0; border-radius:12px; }
            """)
            card_lay = QtWidgets.QVBoxLayout(card)
            card_lay.setContentsMargins(16, 14, 16, 16)
            card_lay.setSpacing(10)

            hdr_row = QtWidgets.QHBoxLayout()
            tax_badge = QtWidgets.QLabel("TAX")
            tax_badge.setStyleSheet("""
                background:#0f766e; color:#ffffff;
                font-family:'Consolas','Courier New',monospace;
                font-size:13px; font-weight:800;
                border:none; border-radius:6px;
                padding:5px 12px;
            """)
            note = QtWidgets.QLabel("Recorded when invoice is marked Paid")
            note.setStyleSheet("font-size:11px; color:#64748b;")
            hdr_row.addWidget(tax_badge)
            hdr_row.addStretch()
            hdr_row.addWidget(note)
            card_lay.addLayout(hdr_row)

            tbl = QtWidgets.QTableWidget(len(tax_entries), 4)
            tbl.setHorizontalHeaderLabels(["Invoice", "Tax Amount", "Paid", "Status"])
            tbl.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
            tbl.setSelectionBehavior(QtWidgets.QTableWidget.SelectRows)
            tbl.setAlternatingRowColors(True)
            tbl.verticalHeader().setVisible(False)
            tbl.setStyleSheet("""
                QTableWidget { border:none; font-size:12px; font-family:'Segoe UI'; }
                QHeaderView::section {
                    background:#f8fafc; font-weight:700; font-size:11px;
                    padding:6px; border:none; border-bottom:1px solid #e2e8f0;
                }
            """)
            STATUS_SC = {
                "Paid":   ("#065f46", "#d1fae5", "#6ee7b7"),
                "Unpaid": ("#92400e", "#fef3c7", "#fcd34d"),
            }
            for i, e in enumerate(tax_entries):
                tbl.setRowHeight(i, 38)
                for col, (val, align, color) in enumerate([
                    (e["invoice_number"],                       QtCore.Qt.AlignCenter, "#0f172a"),
                    (f"${e['tax_amount']:,.2f}",                QtCore.Qt.AlignCenter, "#0f172a"),
                    (f"${e['tax_paid']:,.2f}" if e["tax_paid"] > 0 else "$0.00",
                                                                QtCore.Qt.AlignCenter,
                                                                "#047857" if e["tax_paid"] > 0 else "#94a3b8"),
                ]):
                    item = QtWidgets.QTableWidgetItem(val)
                    item.setTextAlignment(align)
                    item.setForeground(QtGui.QColor(color))
                    tbl.setItem(i, col, item)
                # Status badge
                status_text = "Paid" if e["is_paid"] else "Unpaid"
                sc = STATUS_SC[status_text]
                cell_w = QtWidgets.QWidget()
                cell_w.setStyleSheet("background:transparent;")
                cell_l = QtWidgets.QHBoxLayout(cell_w)
                cell_l.setContentsMargins(4, 0, 4, 0)
                cell_l.setAlignment(QtCore.Qt.AlignCenter)
                badge = QtWidgets.QLabel(f"  {status_text}  ")
                badge.setAlignment(QtCore.Qt.AlignCenter)
                badge.setStyleSheet(
                    f"color:{sc[0]};background:{sc[1]};border:1px solid {sc[2]};"
                    "border-radius:8px;font-size:11px;font-weight:800;padding:3px 6px;"
                )
                cell_l.addWidget(badge)
                tbl.setCellWidget(i, 3, cell_w)

            hdr = tbl.horizontalHeader()
            hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
            for col in range(1, 4):
                hdr.setSectionResizeMode(col, QtWidgets.QHeaderView.Fixed)
            tbl.setColumnWidth(1, 145)
            tbl.setColumnWidth(2, 120)
            tbl.setColumnWidth(3, 145)
            tbl.setFixedHeight(hdr.sizeHint().height() + len(tax_entries) * 38 + 4)
            card_lay.addWidget(tbl)
            return card
        except Exception as exc:
            _log.warning("Could not build tax section for %s: %s", project_number, exc)
            return None

    def _get_project_invoice_rows(self, project_number: str):
        rows = []
        try:
            invoices = FirebaseManager.load_invoices() or []
        except Exception as exc:
            _log.warning("Could not load invoices for project %s: %s", project_number, exc)
            invoices = []

        # Load actual payments to compute per-stage payment status
        try:
            from payment_tracker import get_payment_tracker as _gpt_inv
            _proj_payments = _gpt_inv().get_project_payments(project_number)
        except Exception:
            _proj_payments = []

        for invoice_data in invoices:
            meta = invoice_data.get("meta", {}) if isinstance(invoice_data, dict) else {}
            invoice_number = meta.get("invoice_number", "")
            invoice_date = meta.get("date", "")
            received_date = meta.get("received_date", "N/A")

            for item in invoice_data.get("items", []):
                if str(item.get("project_number", "")).strip() != str(project_number).strip():
                    continue

                amount_raw = (
                    item.get("payment_due")
                    if item.get("payment_due") not in (None, "", 0, 0.0)
                    else item.get("total", item.get("unit_price", 0))
                )
                try:
                    stage_amount = float(amount_raw or 0)
                except Exception:
                    stage_amount = 0.0

                stage = item.get("payment_category", "N/A") or "N/A"

                # Sum payments linked to this invoice OR matching this stage
                stage_paid = sum(
                    float(p.amount) for p in _proj_payments
                    if (p.payment_stage or "").strip().lower() != "tax"
                    and (
                        (p.invoice_number or "") == invoice_number
                        or self._stage_matches(stage, p.payment_stage or "")
                    )
                )
                if stage_amount > 0 and stage_paid >= stage_amount - 0.01:
                    pay_status = "Paid"
                elif stage_paid > 0:
                    pay_status = "Partially Paid"
                else:
                    pay_status = "Unpaid"

                rows.append({
                    "invoice_number": invoice_number,
                    "date": invoice_date,
                    "stage": stage,
                    "amount": self._format_project_money(amount_raw),
                    "status": pay_status,
                    "received_date": received_date or "N/A",
                })

        return rows

    def _get_project_payment_schedule_rows(self, project_data: dict):
        project_number = project_data.get("project_number", "")
        invoice_rows = self._get_project_invoice_rows(project_number)
        payments = get_payment_tracker().get_project_payments(project_number)
        planned_rows = self._planned_payment_rows(project_data)
        rows = []

        for planned in planned_rows:
            stage = planned["stage"]
            matched_invoice = next(
                (
                    invoice for invoice in invoice_rows
                    if self._stage_matches(stage, invoice.get("stage", ""))
                ),
                None,
            )
            matched_payments = [
                payment for payment in payments
                if self._stage_matches(stage, payment.payment_stage)
                or self._stage_matches(stage, payment.invoice_number)
            ]
            paid = sum(float(payment.amount) for payment in matched_payments)
            planned_amount = float(planned.get("amount", 0) or 0)
            remaining = max(planned_amount - paid, 0)

            if remaining <= 0 and planned_amount > 0:
                status = "Paid"
            elif paid > 0:
                status = "Partially Paid"
            elif matched_invoice:
                raw_status = matched_invoice.get("status", "Invoice Created") or "Invoice Created"
                # Don't inherit invoice-level payment statuses when this project has
                # no payments — the invoice status may reflect payments for other
                # projects on the same multi-project invoice.
                if raw_status in ("Partially Paid", "Paid"):
                    status = "Invoice Created"
                else:
                    status = raw_status
            else:
                status = "Not Invoiced"

            rows.append({
                "invoice_number": matched_invoice.get("invoice_number", "") if matched_invoice else "",
                "date": matched_invoice.get("date", "") if matched_invoice else "",
                "stage": stage,
                "amount": self._format_project_money(planned_amount),
                "amount_value": planned_amount,
                "paid_value": paid,
                "remaining_value": remaining,
                "status": status,
                "received_date": matched_invoice.get("received_date", "") if matched_invoice else "",
            })

        return rows

    def _format_project_money(self, value):
        try:
            return Currency.format(float(value or 0))
        except Exception:
            return "$0.00"

    def _project_total_amount(self, project_data):
        raw_amount = project_data.get("project_amount", 0) if isinstance(project_data, dict) else 0
        try:
            if isinstance(raw_amount, str):
                raw_amount = raw_amount.replace("$", "").replace(",", "").strip()
            return float(raw_amount or 0)
        except (TypeError, ValueError):
            return 0.0

    def _project_payment_summary(self, project_data):
        project_number = project_data.get("project_number", "")
        total_amount = self._project_total_amount(project_data)
        return get_payment_tracker().get_payment_summary(project_number, total_amount)

    def _next_unpaid_payment_stage(self, project_data):
        project_number = project_data.get("project_number", "")
        payments = get_payment_tracker().get_project_payments(project_number)
        for planned in self._planned_payment_rows(project_data):
            stage = planned.get("stage", "")
            planned_amount = float(planned.get("amount", 0) or 0)
            paid_amount = sum(
                float(payment.amount)
                for payment in payments
                if self._stage_matches(stage, payment.payment_stage)
                or self._stage_matches(stage, payment.invoice_number)
            )
            remaining_amount = max(planned_amount - paid_amount, 0.0)
            if remaining_amount > 0.009:
                return {
                    "stage": stage,
                    "planned_value": planned_amount,
                    "paid_value": paid_amount,
                    "remaining_value": remaining_amount,
                }
        return None

    def _create_sales_person_widget(self, sales_name: str):
        """Cell widget: colored initial avatar + name text."""
        widget = QtWidgets.QWidget()
        widget.setStyleSheet("background: transparent; border: none;")
        hbox = QtWidgets.QHBoxLayout(widget)
        hbox.setContentsMargins(6, 0, 6, 0)
        hbox.setSpacing(7)

        name = (sales_name or "").strip()
        if not name:
            lbl = QtWidgets.QLabel("—")
            lbl.setStyleSheet("color:#94a3b8; font-size:12pt; background:transparent; border:none;")
            hbox.addWidget(lbl)
            return widget

        initials = "".join(part[:1] for part in name.split()[:2]).upper() or "?"
        avatar_bg, avatar_fg = {
            0: ("#eff6ff", "#2563eb"),
            1: ("#ecfeff", "#0891b2"),
            2: ("#ecfdf5", "#059669"),
            3: ("#f5f3ff", "#7c3aed"),
        }[sum(ord(ch) for ch in name) % 4]

        avatar = QtWidgets.QLabel(initials)
        avatar.setFixedSize(30, 30)
        avatar.setAlignment(QtCore.Qt.AlignCenter)
        avatar.setStyleSheet(f"""
            QLabel {{
                background: {avatar_bg};
                color: {avatar_fg};
                border: none;
                border-radius: 15px;
                font-size: 11px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }}
        """)

        name_lbl = QtWidgets.QLabel(name)
        name_lbl.setStyleSheet(f"""
            QLabel {{
                color: #0f172a;
                font-size: 13px;
                font-weight: 800;
                background: transparent;
                border: none;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }}
        """)
        name_lbl.setAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)
        name_lbl.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)

        hbox.addWidget(avatar)
        hbox.addWidget(name_lbl, 1)
        return widget

    def _create_paid_amount_item(self, project_data, bg_color):
        summary = self._project_payment_summary(project_data)
        total_amount = float(summary.get("total_amount", 0) or 0)
        paid_amount = float(summary.get("total_paid", 0) or 0)
        remaining_amount = max(float(summary.get("remaining", 0) or 0), 0.0)

        paid_item = QtWidgets.QTableWidgetItem(self._format_project_money(paid_amount))
        paid_item.setTextAlignment(QtCore.Qt.AlignCenter)
        paid_item.setBackground(bg_color)
        paid_item.setToolTip(
            f"Project total: {self._format_project_money(total_amount)}\n"
            f"Paid: {self._format_project_money(paid_amount)}\n"
            f"Remaining: {self._format_project_money(remaining_amount)}"
        )
        paid_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
        paid_item.setForeground(QtGui.QColor("#047857" if paid_amount > 0 else "#64748b"))
        return paid_item

    def _create_remaining_due_item(self, project_data, bg_color):
        summary = self._project_payment_summary(project_data)
        total_amount = float(summary.get("total_amount", 0) or 0)
        paid_amount = float(summary.get("total_paid", 0) or 0)
        remaining_amount = max(float(summary.get("remaining", 0) or 0), 0.0)

        remaining_item = QtWidgets.QTableWidgetItem(self._format_project_money(remaining_amount))
        remaining_item.setTextAlignment(QtCore.Qt.AlignCenter)
        remaining_item.setBackground(bg_color)
        remaining_item.setToolTip(
            f"Project total: {self._format_project_money(total_amount)}\n"
            f"Paid: {self._format_project_money(paid_amount)}\n"
            f"Remaining: {self._format_project_money(remaining_amount)}"
        )
        remaining_font = QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold)
        remaining_item.setFont(remaining_font)

        if total_amount <= 0:
            remaining_item.setForeground(QtGui.QColor("#64748b"))
        elif remaining_amount <= 0:
            remaining_item.setForeground(QtGui.QColor("#047857"))
        else:
            remaining_item.setForeground(QtGui.QColor("#dc2626"))

        return remaining_item

    def _add_project_payment_section(self, content_layout, project_data):
        project_number = project_data.get("project_number", "")
        total_amount = float(project_data.get("project_amount", 0) or 0)
        tracker = get_payment_tracker()
        payments = sorted(
            tracker.get_project_payments(project_number),
            key=lambda payment: payment.payment_date or ""
        )
        summary = tracker.get_payment_summary(project_number, total_amount)

        title = QtWidgets.QLabel("Payment Terms & Paid History")
        title.setStyleSheet("""
            font-size: 15px;
            font-weight: 800;
            color: #0f172a;
            padding-top: 12px;
        """)
        content_layout.addWidget(title)

        summary_frame = QtWidgets.QFrame()
        summary_frame.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
            }
            QLabel {
                color: #334155;
                font-size: 12px;
            }
        """)
        summary_grid = QtWidgets.QGridLayout(summary_frame)
        summary_grid.setContentsMargins(14, 10, 14, 10)
        summary_grid.setHorizontalSpacing(18)
        summary_grid.setVerticalSpacing(6)

        payment_terms = project_data.get("payment_terms", "N/A") or "N/A"
        deposit_amount = project_data.get("deposit_amount", 0) or 0
        deposit_date = project_data.get("deposit_received_date", "") or "N/A"
        deposit_received = project_data.get("deposit_received", "No") or "No"
        summary_items = [
            ("Payment Terms", payment_terms),
            ("Deposit Rule", project_data.get("deposit_rule", "N/A") or "N/A"),
            ("Deposit Received", str(deposit_received)),
            ("Deposit Amount", self._format_project_money(deposit_amount)),
            ("Deposit Date", deposit_date),
            ("Project Total", self._format_project_money(total_amount)),
            ("Total Paid", self._format_project_money(summary["total_paid"])),
            ("Balance", self._format_project_money(summary["remaining"])),
        ]

        for index, (label, value) in enumerate(summary_items):
            label_widget = QtWidgets.QLabel(f"{label}:")
            label_widget.setStyleSheet("font-weight: 800; color: #475569;")
            value_widget = QtWidgets.QLabel(str(value))
            value_widget.setStyleSheet("font-weight: 700; color: #0f172a;")
            row = index // 2
            col = (index % 2) * 2
            summary_grid.addWidget(label_widget, row, col)
            summary_grid.addWidget(value_widget, row, col + 1)

        content_layout.addWidget(summary_frame)

        payments_table = QtWidgets.QTableWidget()
        payments_table.setColumnCount(5)
        payments_table.setHorizontalHeaderLabels([
            "Term", "Paid Date", "Amount Paid", "Method", "Notes"
        ])
        payments_table.verticalHeader().setVisible(False)
        payments_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        payments_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        payments_table.setAlternatingRowColors(True)
        payments_table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                gridline-color: #edf2f7;
                font-size: 12px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QTableWidget::item {
                padding: 7px 10px;
                color: #0f172a;
            }
            QHeaderView::section {
                background: #f8fafc;
                color: #334155;
                font-weight: 800;
                padding: 8px 10px;
                border: none;
                border-bottom: 1px solid #e2e8f0;
            }
        """)

        header = payments_table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        payments_table.setColumnWidth(0, 70)
        payments_table.setColumnWidth(1, 115)
        payments_table.setColumnWidth(2, 120)
        payments_table.setColumnWidth(3, 120)

        if payments:
            payments_table.setRowCount(len(payments))
            for row, payment in enumerate(payments):
                values = [
                    f"Term {row + 1}",
                    payment.payment_date or "N/A",
                    self._format_project_money(payment.amount),
                    payment.payment_method or "N/A",
                    payment.notes or "",
                ]
                for col, value in enumerate(values):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    align = QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter if col == 2 else QtCore.Qt.AlignCenter
                    if col == 4:
                        align = QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter
                    item.setTextAlignment(align)
                    item.setToolTip(str(value))
                    payments_table.setItem(row, col, item)
                payments_table.setRowHeight(row, 36)
        else:
            payments_table.setRowCount(1)
            item = QtWidgets.QTableWidgetItem("No payment records added yet.")
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            payments_table.setItem(0, 0, item)
            payments_table.setSpan(0, 0, 1, 5)
            payments_table.setRowHeight(0, 38)

        table_height = payments_table.horizontalHeader().height() + (payments_table.rowCount() * 38) + 12
        payments_table.setMinimumHeight(max(88, min(table_height, 230)))
        payments_table.setMaximumHeight(max(88, min(table_height, 230)))
        content_layout.addWidget(payments_table)
     
    def view_project_details(self, project_data):
        """View project details in a dialog"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"Project Details - {project_data['project_number']}")
        dialog.setModal(True)
        dialog.resize(760, 640)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        header = QtWidgets.QLabel(f"Project: {project_data['project_number']}")
        header.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 20px;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
        """)
        layout.addWidget(header)
        
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        content = QtWidgets.QWidget()
        scroll.setWidget(content)
        
        content_layout = QtWidgets.QVBoxLayout(content)
        
        details = [
            ("Project Number:", project_data.get("project_number", "")),
            ("Project Name:", project_data.get("project_name", "")),
            ("Client:", project_data.get("company", "")),
            ("PO/WO Number:", project_data.get("po_wo_number", "")),
            ("Date Received:", project_data.get("date_received", "")),
            ("Plant:", project_data.get("plant", "")),
            ("Sales:", project_data.get("sales", "")),
            ("Project Amount:", Currency.format(project_data.get("project_amount", 0))),
            ("Payment Category:", project_data.get("payment_category", "")),
            ("Payment Terms:", project_data.get("payment_terms", "")),
            ("Deposit Rule:", project_data.get("deposit_rule", "")),
            ("Deposit Received:", project_data.get("deposit_received", "")),
            ("Deposit Amount:", Currency.format(project_data.get("deposit_amount", 0))),
            ("Deposit Date:", project_data.get("deposit_received_date", "")),
            ("Start Date:", project_data.get("start_date", "")),
            ("Due Date:", project_data.get("due_date", "")),
            ("Status:", project_data.get("status", "")),
            ("Notes:", project_data.get("notes", "")),
        ]
        
        for label, value in details:
            row_layout = QtWidgets.QHBoxLayout()
            
            label_widget = QtWidgets.QLabel(label)
            label_widget.setStyleSheet("font-weight: bold; color: #555; min-width: 120px;")
            row_layout.addWidget(label_widget)
            
            value_widget = QtWidgets.QLabel(str(value))
            value_widget.setStyleSheet("color: #333;")
            value_widget.setWordWrap(True)
            row_layout.addWidget(value_widget, 1)
            
            content_layout.addLayout(row_layout)
            content_layout.addSpacing(8)
        
        self._add_project_payment_section(content_layout, project_data)
        content_layout.addStretch()
        layout.addWidget(scroll)
        
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()
    
    def apply_current_filters_after_reload(self):
        """Apply existing filters after reloading projects - PRESERVE FILTER STATE"""
        self._refresh_summary_cards()
        current_search = self.search_edit.text()
        current_status = self.status_filter_combo.currentText()
        
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        if date_range_active:
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date
        
        current_client_filter = self.selected_client_filter
        
        self.update_projects_table(self.generated_projects)
        
        if current_search:
            self.search_edit.setText(current_search)
        
        if current_status:
            for i in range(self.status_filter_combo.count()):
                if self.status_filter_combo.itemText(i).startswith(current_status.split(" (")[0]):
                    self.status_filter_combo.setCurrentIndex(i)
                    break
        
        if date_range_active:
            from_date_formatted = current_from_date.toString("MM/dd/yy")
            to_date_formatted = current_to_date.toString("MM/dd/yy")
            configure_filter_button(
                self.date_range_button,
                f"{from_date_formatted} to {to_date_formatted}",
                active=True,
                height=38,
            )
            self.current_from_date = current_from_date
            self.current_to_date = current_to_date
        
        self.selected_client_filter = current_client_filter
        if current_client_filter != "All Clients" and current_client_filter != "📂 All Clients":
            display_name = current_client_filter[:15] + "..." if len(current_client_filter) > 15 else current_client_filter
            self.client_filter_button.setText(f"🏢 {display_name}")
            self.client_filter_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: bold;
                    padding: 6px 10px;
                    min-width: 50px;
                    max-width: 150px;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
                QPushButton:pressed {
                    background-color: #21618c;
                }
            """)
            self.client_filter_button.setToolTip(f"Client: {current_client_filter}\nClick to change")
        else:
            self.client_filter_button.setText("Client")
            self.client_filter_button.setStyleSheet("""
                QPushButton {
                    background: #ffffff;
                    border: 1.5px solid #d8e2ec;
                    border-radius: 8px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #334155;
                    padding: 0 12px;
                }
                QPushButton:hover { color: #00756f; border-color: #00756f; }
            """)
            self.client_filter_button.setToolTip("Filter by Client")
        
        self.filter_projects()
        self.update_stats()
        self.update_client_filter_menu()
        self.update_status_filter_counts()
        
        _log.info("Filters preserved after project reload")
    
    def load_projects(self):
        """Load saved projects from Firebase - SORT BY CREATED DATE (newest first)"""
        try:
            if FIREBASE_AVAILABLE:
                from main import db
                ref = db.reference('/projects')
                data = ref.get() or {}

                loaded = []
                seen_project_numbers = set()
                
                for pid, p in data.items():
                    p["firebase_id"] = pid
                    project_number = p.get("project_number", "")
                    
                    # Skip invalid project numbers
                    if project_number and project_number != "-0-0003":
                        if project_number not in seen_project_numbers:
                            seen_project_numbers.add(project_number)
                            
                            # Ensure payment category has valid value
                            if p.get("payment_category") in ["Select", "Select Payment Category", ""]:
                                p["payment_category"] = "N/A"

                            # Migrate legacy plain "Completed" status to the correct
                            # current status based on actual payment state.
                            # Old code blindly set "Completed" regardless of payment
                            # progress, so we must inspect payments rather than just
                            # mapping every "Completed" → "Paid".
                            if p.get("status") == "Completed":
                                if is_project_fully_paid(p):
                                    correct = "Paid"
                                else:
                                    from payment_tracker import get_payment_tracker as _gpt_m
                                    _pn_m = p.get("project_number", "")
                                    _has_m = any(
                                        float(_pm.amount) > 0
                                        and (_pm.payment_stage or "").strip().lower() != "tax"
                                        for _pm in _gpt_m().get_project_payments(_pn_m)
                                    )
                                    correct = "In Progress" if _has_m else "Not Started"
                                p["status"] = correct
                                try:
                                    from firebase_admin import db as _fdb
                                    _fdb.reference(f'/projects/{pid}').update({
                                        'status': correct,
                                        'updated_at': datetime.now().isoformat(),
                                    })
                                except Exception:
                                    pass

                            # Ensure created_at exists (for older projects, use current date as fallback)
                            if 'created_at' not in p or not p['created_at']:
                                # For older projects without created_at, use current date
                                p['created_at'] = datetime.now().isoformat()

                            loaded.append(p)
                
                # Sort by CREATED DATE (newest first)
                def get_created_date(project):
                    """Extract and parse created date for sorting"""
                    created_date = project.get('created_at', '')
                    if not created_date:
                        return datetime.min  # Put projects without date at the end
                    
                    # Handle ISO format with time (YYYY-MM-DDTHH:MM:SS)
                    try:
                        if 'T' in created_date:
                            # Parse ISO format
                            return datetime.fromisoformat(created_date.replace("Z", ""))
                        else:
                            return datetime.strptime(created_date, "%Y-%m-%d")
                    except (ValueError, TypeError):
                        pass
                    
                    # Try other date formats as fallback
                    for date_format in ["%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"]:
                        try:
                            return datetime.strptime(created_date, date_format)
                        except (ValueError, TypeError):
                            continue
                    
                    # If parsing fails, return min date
                    return datetime.min
                
                # Sort in descending order (newest first)
                loaded.sort(key=get_created_date, reverse=True)

                self.generated_projects = loaded
                self.cached_projects = loaded.copy()
                self.update_sales_filter_options()
                
                # Apply current filters after reload
                self.apply_current_filters_after_reload()
                
                _log.info("📊 Total valid projects loaded: %s", len(self.cached_projects))
                _log.info("   Sorted by created date (newest first)")
                
                # Debug: Show first 5 projects with their dates
                if self.cached_projects:
                    _log.info("   First 5 projects (newest first):")
                    for i, proj in enumerate(self.cached_projects[:5], 1):
                        proj_num = proj.get('project_number', 'N/A')
                        created_date = proj.get('created_at', 'No date')
                        received_date = proj.get('date_received', 'No date')
                        _log.info("     %s. %s - Created: %s | Received: %s", i, proj_num, created_date, received_date)

        except Exception as e:
            _log.error("Error loading projects: %s", e)
            traceback.print_exc()
                
    def refresh_clients_list(self):
        """Refresh the company dropdown"""
        _log.info("Refreshing clients list in project generator tab...")
        self.update_client_filter_menu()

    def update_plant_filter_options(self):
        """Rebuild the plant filter dropdown from current project data."""
        if not hasattr(self, "plant_filter_combo"):
            return
        current = getattr(self, "selected_plant_filter", "All Plants") or "All Plants"
        plants = sorted({
            str(p.get("plant", "")).strip()
            for p in self.generated_projects
            if str(p.get("plant", "")).strip()
        })
        self.plant_filter_combo.blockSignals(True)
        self.plant_filter_combo.clear()
        self.plant_filter_combo.addItem("All Plants")
        self.plant_filter_combo.addItems(plants)
        index = self.plant_filter_combo.findText(current)
        self.plant_filter_combo.setCurrentIndex(index if index >= 0 else 0)
        self.plant_filter_combo.blockSignals(False)

    def update_sales_filter_options(self):
        """Alias kept for backward compatibility."""
        self.update_plant_filter_options()

    def on_plant_filter_changed(self, filter_text: str):
        self.selected_plant_filter = filter_text.strip() or "All Plants"
        self._page_num = 1
        self.filter_projects()
        self.update_stats()

    def on_sales_filter_changed(self, filter_text):
        pass   # superseded by on_plant_filter_changed
    
    def update_client_filter_menu(self):
        """Update the client filter menu with client counts"""
        self.client_filter_menu.clear()
        
        search_text = self.search_edit.text().lower()
        status_filter = self._clean_project_status_filter(self.status_filter_combo.currentText())
        
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        
        all_action = QtWidgets.QAction("📂 All Clients", self)
        all_action.triggered.connect(lambda: self.apply_client_menu_filter("All Clients"))
        self.client_filter_menu.addAction(all_action)
        
        self.client_filter_menu.addSeparator()
        
        client_counts = {}
        total_filtered_projects = 0
        
        for project in self.generated_projects:
            matches_search = False
            if not search_text:
                matches_search = True
            else:
                if search_text in project.get('project_number', '').lower():
                    matches_search = True
                elif search_text in project.get('project_name', '').lower():
                    matches_search = True
                elif search_text in project.get('company', '').lower():
                    matches_search = True
                elif search_text in project.get('sales', '').lower():
                    matches_search = True
                elif search_text in project.get('plant', '').lower():
                    matches_search = True
                elif search_text in project.get('payment_category', '').lower():
                    matches_search = True
                else:
                    amount = project.get('project_amount', 0.0)
                    if amount:
                        try:
                            amount_str = str(amount).replace(',', '').replace('$', '')
                            if search_text in amount_str.lower():
                                matches_search = True
                            elif search_text in Currency.format(amount).lower():
                                matches_search = True
                        except:
                            pass
                    if not matches_search:
                        payment_lower = search_text.replace(' ', '')
                        if payment_lower == 'downpayment' and project.get('payment_category', '').lower() == 'down payment':
                            matches_search = True
                        elif payment_lower == 'finalpayment' and project.get('payment_category', '').lower() == 'final payment':
                            matches_search = True
            
            matches_status = self._project_status_matches(project, status_filter)
            
            matches_date = True
            if date_range_active:
                matches_date = self.is_project_in_date_range_by_received(
                    project, self.current_from_date, self.current_to_date
                )
            
            if matches_search and matches_status and matches_date:
                client = project.get("company", "")
                if client:
                    client_counts[client] = client_counts.get(client, 0) + 1
                    total_filtered_projects += 1
        
        self.client_filter_menu.actions()[0].setText(f"📂 All Clients ({total_filtered_projects})")
        
        for client, count in sorted(client_counts.items()):
            action_text = f"🏢 {client} ({count})"
            action = QtWidgets.QAction(action_text, self)
            action.triggered.connect(lambda checked, c=client: self.apply_client_menu_filter(c))
            self.client_filter_menu.addAction(action)

    def _clean_project_status_filter(self, status_text: str) -> str:
        """Strip count suffixes from status dropdown labels."""
        status = (status_text or "").strip()
        if " (" in status:
            status = status.split(" (", 1)[0].strip()
        return status

    def _canonical_project_status(self, status: str) -> str:
        """Normalize legacy/project status spellings to current dropdown values."""
        raw = (status or "").strip()
        key = raw.lower().replace("&", "and")
        key = " ".join(key.replace("-", " ").split())
        aliases = {
            "": "Not Started",
            "select": "Not Started",
            "active": "In Progress",
            "started": "In Progress",
            "progress": "In Progress",
            "hold": "On Hold",
            "onhold": "On Hold",
            "completed": "Completed & Invoiced",
            "complete": "Completed & Invoiced",
            "completed invoiced": "Completed & Invoiced",
            "completed and invoiced": "Completed & Invoiced",
            "completed not invoiced": "Completed Not Invoiced",
            "not invoiced": "Completed Not Invoiced",
            "invoice pending": "Completed Not Invoiced",
            "fully paid": "Paid",
            "paid": "Paid",
            "cancel": "Cancelled",
            "canceled": "Cancelled",
            "cancelled": "Cancelled",
        }
        return aliases.get(key, raw if raw in self.PROJECT_STATUSES else raw.title())

    def _project_effective_status(self, project: dict) -> str:
        """Return the status the UI should filter/count, including fully-paid projects."""
        status = self._canonical_project_status(project.get("status", "Not Started"))
        if status != "Cancelled":
            try:
                total = self._project_total_amount(project)
                pn = project.get("project_number", "")
                summary = get_payment_tracker().get_payment_summary(pn, total)
                if total > 0 and float(summary.get("payment_percentage", 0)) >= 100.0:
                    return "Paid"
            except Exception:
                pass
        return status

    def _project_status_matches(self, project: dict, status_filter: str) -> bool:
        status_filter = self._clean_project_status_filter(status_filter)
        if status_filter in ("", "All Status"):
            return True
        return self._project_effective_status(project) == self._canonical_project_status(status_filter)
            
    def update_status_filter_counts(self):
        """Update status filter dropdown with counts"""
        if not hasattr(self, 'generated_projects'):
            return
        
        search_text = self.search_edit.text().lower()
        client_filter = self.selected_client_filter
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        
        status_counts = {}
        for project in self.generated_projects:
            matches_client = (
                client_filter == "All Clients" or
                client_filter == "📂 All Clients" or
                project.get('company', '') == client_filter
            )
            
            matches_search = False
            if not search_text:
                matches_search = True
            else:
                if search_text in project.get('project_number', '').lower():
                    matches_search = True
                elif search_text in project.get('project_name', '').lower():
                    matches_search = True
                elif search_text in project.get('company', '').lower():
                    matches_search = True
                elif search_text in project.get('sales', '').lower():
                    matches_search = True
                elif search_text in project.get('plant', '').lower():
                    matches_search = True
                elif search_text in project.get('payment_category', '').lower():
                    matches_search = True
                else:
                    amount = project.get('project_amount', 0.0)
                    if amount:
                        try:
                            amount_str = str(amount).replace(',', '').replace('$', '')
                            if search_text in amount_str.lower():
                                matches_search = True
                            elif search_text in Currency.format(amount).lower():
                                matches_search = True
                        except:
                            pass
                    if not matches_search:
                        payment_lower = search_text.replace(' ', '')
                        if payment_lower == 'downpayment' and project.get('payment_category', '').lower() == 'down payment':
                            matches_search = True
                        elif payment_lower == 'finalpayment' and project.get('payment_category', '').lower() == 'final payment':
                            matches_search = True
            
            matches_date = True
            if date_range_active:
                matches_date = self.is_project_in_date_range_by_received(
                    project, self.current_from_date, self.current_to_date
                )
            
            if matches_client and matches_search and matches_date:
                status = self._project_effective_status(project)
                status_counts[status] = status_counts.get(status, 0) + 1
        
        all_statuses = ["All Status"] + self.PROJECT_STATUSES
        
        current_selection = self.status_filter_combo.currentText()
        
        self.status_filter_combo.blockSignals(True)
        self.status_filter_combo.clear()
        
        total_filtered = sum(status_counts.values())
        
        for status in all_statuses:
            if status == "All Status":
                display_text = f"All Status ({total_filtered})"
            else:
                count = status_counts.get(status, 0)
                display_text = f"{status} ({count})"
            
            self.status_filter_combo.addItem(display_text)
        
        if current_selection:
            status_name = current_selection.split(" (")[0]
            for i in range(self.status_filter_combo.count()):
                item_text = self.status_filter_combo.itemText(i)
                if item_text.startswith(status_name):
                    self.status_filter_combo.setCurrentIndex(i)
                    break
        
        self.status_filter_combo.blockSignals(False)
        
    def update_results_label(self):
        """Update the results counter label"""
        if hasattr(self, 'results_label'):
            filtered_count = self.projects_table.rowCount()
            total_count = len(self.generated_projects)
            self.results_label.setText(f"{filtered_count} of {total_count} projects shown")
            return
            self.results_label.setText(f"📋 Showing {filtered_count} of {total_count} projects")
    
    def on_status_filter_changed(self, filter_text):
        """Handle status filter change"""
        self._page_num = 1
        self.update_client_filter_menu()
        self.filter_projects()
    
    def apply_client_menu_filter(self, company_name):
        """Apply filter from client menu"""
        self.selected_client_filter = company_name
        
        if company_name == "All Clients" or company_name == "📂 All Clients":
            self.client_filter_button.setText("Client")
            self.client_filter_button.setStyleSheet("""
                QPushButton {
                    background: #ffffff;
                    border: 1.5px solid #d8e2ec;
                    border-radius: 8px;
                    font-size: 12px;
                    font-weight: 700;
                    color: #334155;
                    padding: 0 12px;
                }
                QPushButton:hover { color: #00756f; border-color: #00756f; }
            """)
            self.client_filter_button.setToolTip("Filter by Client")
        else:
            display_name = company_name[:15] + "..." if len(company_name) > 15 else company_name
            self.client_filter_button.setText(f"🏢 {display_name}")
            self.client_filter_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: bold;
                    padding: 6px 10px;
                    min-width: 50px;
                    max-width: 150px;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
                QPushButton:pressed {
                    background-color: #21618c;
                }
            """)
            self.client_filter_button.setToolTip(f"Client: {company_name}\nClick to change")
        
        self.update_status_filter_counts()
        self.filter_projects()
    
    def search_projects(self, search_text):
        """Search projects by text"""
        self._page_num = 1
        self.filter_projects()
    
    def is_project_in_date_range_by_received(self, project, from_date, to_date):
        """Check if project is within date range based on RECEIVED DATE"""
        try:
            received_date_str = project.get('date_received', '')
            
            if not received_date_str:
                return False
            
            project_date = None
            
            # Try parsing in MM-DD-YYYY format
            try:
                from datetime import datetime
                dt_obj = datetime.strptime(received_date_str, "%m-%d-%Y")
                project_date = QtCore.QDate(dt_obj.year, dt_obj.month, dt_obj.day)
            except:
                # Try YYYY-MM-DD format
                try:
                    dt_obj = datetime.strptime(received_date_str, "%Y-%m-%d")
                    project_date = QtCore.QDate(dt_obj.year, dt_obj.month, dt_obj.day)
                except:
                    # Try MM/DD/YYYY format
                    try:
                        dt_obj = datetime.strptime(received_date_str, "%m/%d/%Y")
                        project_date = QtCore.QDate(dt_obj.year, dt_obj.month, dt_obj.day)
                    except:
                        pass
            
            if project_date and project_date.isValid():
                return from_date <= project_date <= to_date
            
            return False
        except:
            return False
    
    def show_date_range_dialog(self):
        """Show date range selection dialog for projects based on RECEIVED DATE"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📅 Select Date Range (by Received Date)")
        dialog.setModal(True)
        dialog.resize(400, 200)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        title = QtWidgets.QLabel("Select Date Range (Filter by Received Date)")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title)
        
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(20, 10, 20, 10)
        
        current_from_date = QtCore.QDate.currentDate().addMonths(-1)
        current_to_date = QtCore.QDate.currentDate()
        
        if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date
        
        self.from_date_edit = QtWidgets.QDateEdit()
        self.from_date_edit.setDate(current_from_date)
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("MM/dd/yyyy")
        self.from_date_edit.wheelEvent = lambda e: e.ignore()
        self.from_date_edit.stepBy = lambda x: None

        self.to_date_edit = QtWidgets.QDateEdit()
        self.to_date_edit.setDate(current_to_date)
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("MM/dd/yyyy")
        self.to_date_edit.wheelEvent = lambda e: e.ignore()
        self.to_date_edit.stepBy = lambda x: None

        form_layout.addRow("From Date:", self.from_date_edit)
        form_layout.addRow("To Date:", self.to_date_edit)
        
        layout.addLayout(form_layout)
        
        button_layout = QtWidgets.QHBoxLayout()
        
        clear_btn = QtWidgets.QPushButton("Clear Filter")
        clear_btn.setMinimumHeight(40)
        clear_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        
        apply_btn = QtWidgets.QPushButton("Apply Filter")
        apply_btn.setMinimumHeight(40)
        apply_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #2ecc71;
            }
        """)
        
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        button_layout.addWidget(apply_btn)
        
        layout.addLayout(button_layout)
        
        def apply_filter():
            from_date_qdate = self.from_date_edit.date()
            to_date_qdate = self.to_date_edit.date()
            
            self.current_from_date = from_date_qdate
            self.current_to_date = to_date_qdate
            
            from_date_formatted = from_date_qdate.toString("MM/dd/yy")
            to_date_formatted = to_date_qdate.toString("MM/dd/yy")
            
            configure_filter_button(
                self.date_range_button,
                f"{from_date_formatted} to {to_date_formatted}",
                active=True,
                height=38,
            )
            self.update_status_filter_counts()
            self.update_client_filter_menu()
            self.filter_projects()
            
            dialog.accept()

        def clear_filter():
            configure_filter_button(self.date_range_button, height=38)
            
            if hasattr(self, 'current_from_date'):
                del self.current_from_date
            if hasattr(self, 'current_to_date'):
                del self.current_to_date
            
            self.update_client_filter_menu()
            self.update_status_filter_counts()
            self.filter_projects()
            self.update_stats()
            
            dialog.accept()
        
        apply_btn.clicked.connect(apply_filter)
        clear_btn.clicked.connect(clear_filter)
        
        dialog.exec_()
            
    def filter_projects(self):
        """Filter projects based on all criteria - using RECEIVED DATE for date range"""
        self._refresh_summary_cards()

        search_text = self.search_edit.text().lower()
        status_filter = self.status_filter_combo.currentText()
        if "(" in status_filter:
            status_filter = status_filter.split(" (")[0]

        client_filter = self.selected_client_filter
        plant_filter = getattr(self, "selected_plant_filter", "All Plants")
        date_range_active = hasattr(self, 'date_range_button') and "to" in self.date_range_button.text()

        view_mode   = getattr(self, "_view_mode",  "all")
        period_mode = getattr(self, "_period_mode", "all")
        now = datetime.now()

        def _view_matches(status: str) -> bool:
            s = (status or "").lower().strip()
            if view_mode == "active":
                # Active = work in progress — not cancelled and not fully closed
                closed = ("completed", "cancelled", "canceled", "paid")
                return not any(s == c or s.startswith(c) for c in closed)
            if view_mode == "completed":
                # Completed = any terminal/done state
                done = ("completed", "paid")
                return any(s == d or s.startswith(d) for d in done)
            return True  # "all"

        filtered_projects = []

        for project in self.generated_projects:
            # ── View-mode filter ────────────────────────────────────────
            if status_filter == "All Status" and not _view_matches(self._project_effective_status(project)):
                continue

            # ── Period filter ────────────────────────────────────────────
            if period_mode in ("month", "year"):
                rd = (project.get("date_received") or project.get("created_at") or "")
                p_date = None
                for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        p_date = datetime.strptime(rd[:10], fmt)
                        break
                    except Exception:
                        pass
                if p_date is None:
                    try:
                        p_date = datetime.fromisoformat(rd[:19])
                    except Exception:
                        pass
                if p_date is None:
                    continue
                if period_mode == "month" and (p_date.year != now.year or p_date.month != now.month):
                    continue
                if period_mode == "year" and p_date.year != now.year:
                    continue
            _cf = client_filter.strip()
            matches_client = (
                _cf in ("All Clients", "📂 All Clients") or
                project.get('company', '').strip().lower() == _cf.lower() or
                project.get('client', '').strip().lower() == _cf.lower()
            )
            
            matches_plant = (
                plant_filter in ("All Plants", "") or
                project.get("plant", "").strip().lower() == plant_filter.strip().lower()
            )

            matches_search = False
            if not search_text:
                matches_search = True
            else:
                if search_text in project.get('project_number', '').lower():
                    matches_search = True
                elif search_text in project.get('project_name', '').lower():
                    matches_search = True
                elif search_text in project.get('company', '').lower():
                    matches_search = True
                elif search_text in project.get('plant', '').lower():
                    matches_search = True
                elif search_text in project.get('payment_category', '').lower():
                    matches_search = True
                else:
                    amount = project.get('project_amount', 0.0)
                    if amount:
                        try:
                            amount_str = str(amount).replace(',', '').replace('$', '')
                            if search_text in amount_str.lower():
                                matches_search = True
                            elif search_text in Currency.format(amount).lower():
                                matches_search = True
                        except:
                            pass
                    if not matches_search:
                        payment_lower = search_text.replace(' ', '')
                        if payment_lower == 'downpayment' and project.get('payment_category', '').lower() == 'down payment':
                            matches_search = True
                        elif payment_lower == 'finalpayment' and project.get('payment_category', '').lower() == 'final payment':
                            matches_search = True
            
            matches_status = self._project_status_matches(project, status_filter)
            
            matches_date = True
            if date_range_active and hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
                matches_date = self.is_project_in_date_range_by_received(
                    project, self.current_from_date, self.current_to_date
                )
            
            if matches_client and matches_plant and matches_search and matches_status and matches_date:
                filtered_projects.append(project)
        def get_created_date(project):
            created_date = project.get('created_at', '')
            if not created_date:
                return datetime.min
            try:
                return datetime.fromisoformat(created_date.replace("Z", ""))
            except:
                return datetime.min

        filtered_projects = self.sort_by_created_desc(filtered_projects)

        total_count = len(filtered_projects)
        self._rebuild_pagination(total_count)
        page_size = getattr(self, "_page_size", 10)
        page_num  = getattr(self, "_page_num", 1)
        start_idx = (page_num - 1) * page_size
        filtered_projects = filtered_projects[start_idx: start_idx + page_size]

        self.projects_table.clearContents()
        self.projects_table.setRowCount(len(filtered_projects))

        for row, project in enumerate(filtered_projects):
            bg_color = QtGui.QColor("#ffffff")

            # S.No.
            sno_item = QtWidgets.QTableWidgetItem(str(row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            sno_item.setForeground(QtGui.QColor('#2c3e50'))
            sno_item.setFont(QtGui.QFont("Inter", 9))
            sno_item.setBackground(bg_color)
            self.projects_table.setItem(row, 0, sno_item)
            
            # Project Number — dot bullet + clean Inter font widget
            pn_widget = self._create_project_number_widget(project["project_number"], project)
            # Store project data on a hidden item so row-click handlers can retrieve it
            _pn_item = QtWidgets.QTableWidgetItem("")
            _pn_item.setData(QtCore.Qt.UserRole, project)
            _pn_item.setToolTip(project["project_number"])
            self.projects_table.setItem(row, self.COL_PROJECT_NUMBER, _pn_item)
            self.projects_table.setCellWidget(row, self.COL_PROJECT_NUMBER, pn_widget)
            
            # Project Name
            name_item = QtWidgets.QTableWidgetItem(project.get("project_name", ""))
            name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            name_item.setForeground(QtGui.QColor('#0f172a'))
            name_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            name_item.setBackground(bg_color)
            self._set_project_item_readability(name_item, project.get("project_name", ""), emphasis=True, left=True)
            self.projects_table.setItem(row, 2, name_item)
            
            # Client
            company_item = QtWidgets.QTableWidgetItem(project.get("company", ""))
            company_item.setTextAlignment(QtCore.Qt.AlignCenter)
            company_item.setForeground(QtGui.QColor('#0f172a'))
            company_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            company_item.setBackground(bg_color)
            self._set_project_item_readability(company_item, project.get("company", ""), emphasis=True, left=True)
            self.projects_table.setItem(row, 3, company_item)
            
            # PO/WO
            po_wo_item = QtWidgets.QTableWidgetItem(project.get("po_wo_number", ""))
            po_wo_item.setTextAlignment(QtCore.Qt.AlignCenter)
            po_wo_item.setForeground(QtGui.QColor('#7f8c8d'))
            po_wo_item.setFont(QtGui.QFont("Inter", 9))
            po_wo_item.setBackground(bg_color)
            self._set_project_item_readability(po_wo_item, project.get("po_wo_number", ""), muted=True, left=True)
            self.projects_table.setItem(row, 4, po_wo_item)
            
            # Received Date
            date_received_item = QtWidgets.QTableWidgetItem(project.get("date_received", ""))
            date_received_item.setTextAlignment(QtCore.Qt.AlignCenter)
            date_received_item.setForeground(QtGui.QColor('#0f172a'))
            date_received_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            date_received_item.setBackground(bg_color)
            self._set_project_item_readability(date_received_item, project.get("date_received", ""), emphasis=True)
            self.projects_table.setItem(row, 5, date_received_item)
            
            # Plant
            plant_item = QtWidgets.QTableWidgetItem(project.get("plant", ""))
            plant_item.setTextAlignment(QtCore.Qt.AlignCenter)
            plant_item.setForeground(QtGui.QColor('#7f8c8d'))
            plant_item.setFont(QtGui.QFont("Inter", 9))
            plant_item.setBackground(bg_color)
            self._set_project_item_readability(plant_item, project.get("plant", ""), muted=True, left=True)
            self.projects_table.setItem(row, 6, plant_item)
            
            # Sales Person — avatar + name widget
            self.projects_table.setCellWidget(
                row, self.COL_SALES,
                self._create_sales_person_widget(project.get("sales", ""))
            )

            # Price
            amount = project.get("project_amount", 0.0)
            amount_item = QtWidgets.QTableWidgetItem(Currency.format(amount))
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            if amount != 0.0:
                amount_item.setForeground(QtGui.QColor('#27ae60'))
                amount_item.setFont(QtGui.QFont("Inter", 9))
            else:
                amount_item.setForeground(QtGui.QColor('#95a5a6'))
            amount_item.setBackground(bg_color)
            amount_item.setToolTip(Currency.format(amount))
            amount_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            self.projects_table.setItem(row, self.COL_PRICE, amount_item)
            self.projects_table.setItem(row, self.COL_PAID, self._create_paid_amount_item(project, bg_color))
            self.projects_table.setItem(row, self.COL_REMAINING, self._create_remaining_due_item(project, bg_color))
            
            # Payment Category Combo Box
            self.projects_table.setCellWidget(row, self.COL_PAYMENT, self._create_payment_cell_widget(project))
            
            # Start Date
            start_date_item = QtWidgets.QTableWidgetItem(project.get("start_date", ""))
            start_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            start_date_item.setForeground(QtGui.QColor('#0f172a'))
            start_date_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            start_date_item.setBackground(bg_color)
            self._set_project_item_readability(start_date_item, project.get("start_date", ""), emphasis=True)
            self.projects_table.setItem(row, self.COL_START_DATE, start_date_item)
            
            # Due Date
            due_date_item = QtWidgets.QTableWidgetItem(project.get("due_date", ""))
            due_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            due_date_item.setForeground(QtGui.QColor('#0f172a'))
            due_date_item.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
            due_date_item.setBackground(bg_color)
            self._set_project_item_readability(due_date_item, project.get("due_date", ""), emphasis=True)
            self.projects_table.setItem(row, self.COL_DUE_DATE, due_date_item)
            
            # Status pill badge
            self.projects_table.setCellWidget(row, self.COL_STATUS, self._build_project_status_badge(project))

            # Action buttons
            self.add_enhanced_action_buttons(row, project)
        
        for row in range(self.projects_table.rowCount()):
            self.projects_table.setRowHeight(row, 65)
        self.projects_table.verticalHeader().setDefaultSectionSize(65)
        if hasattr(self.projects_table, "refresh_frozen_project_column"):
            self.projects_table.refresh_frozen_project_column(195)
        self.adjust_projects_table_height()
        
        self.update_stats()
        self.update_results_label()
        
        self.projects_table.clearSelection()
        self.selected_projects.clear()
        self.load_selected_btn.setEnabled(False)
        
    def copy_single_project(self, project):
        """Copy single project number from action button"""
        clipboard = QtWidgets.QApplication.clipboard()
        clipboard.setText(project["project_number"])
        
        QtWidgets.QMessageBox.information(
            self, "Copied", 
            f"Project number '{project['project_number']}' copied to clipboard!"
        )

    def create_payment_category_button(self, project_data):
        """Create a compact menu button for payment category inside the table."""
        valid_categories = [
            "Single Payment",
            "25% Deposit + Balance",
            "50% Deposit + Final",
            "Custom Installments",
        ]
        raw = project_data.get("payment_category", "N/A") or "N/A"
        # Migrate old names to new unified names
        _migrate = {
            "Down Payment (50%)":  "50% Deposit + Final",
            "Due Payment":         "Single Payment",
            "Final Payment Due":   "50% Deposit + Final",
            "Full Amount Due":     "Single Payment",
            "25% Deposit + Balance": "25% Deposit + Balance",
            "Deposit Received (50%)": "50% Deposit + Final",
            "Term 2 Payment":      "Custom Installments",
            "Term 3 Payment":      "Custom Installments",
            "Term 4 Payment":      "Custom Installments",
            "Custom 4 Payments":   "Custom Installments",
            "Full Amount":         "Single Payment",
            "Down Payment":        "50% Deposit + Final",
            "Final Payment":       "50% Deposit + Final",
        }
        current = _migrate.get(raw, raw)
        if current not in valid_categories:
            current = "Single Payment"

        button = QtWidgets.QPushButton(current)
        button.setFixedSize(174, 24)
        button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._style_payment_category_button(button, current)

        menu = QtWidgets.QMenu(button)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 6px;
            }
            QMenu::item {
                color: #334155;
                padding: 8px 18px;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 700;
                min-width: 128px;
            }
            QMenu::item:selected {
                background: #e6f6f4;
                color: #0f172a;
            }
        """)

        def apply_category(category):
            self._style_payment_category_button(button, category)
            self.update_payment_category(project_data, category)

        for category in valid_categories:
            action = menu.addAction(category)
            action.triggered.connect(
                lambda checked=False, cat=category: apply_category(cat)
            )

        button.clicked.connect(lambda: menu.exec_(button.mapToGlobal(QtCore.QPoint(0, button.height() + 2))))
        return button

    def _create_payment_cell_widget(self, project_data):
        """Progress column: stage badge + paid/remaining + next stage, fits 72px rows."""
        container = QtWidgets.QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        vbox = QtWidgets.QVBoxLayout(container)
        vbox.setContentsMargins(6, 3, 6, 3)
        vbox.setSpacing(2)
        vbox.setAlignment(QtCore.Qt.AlignCenter)

        total   = self._project_total_amount(project_data)
        pn      = project_data.get("project_number", "")
        summary = get_payment_tracker().get_payment_summary(pn, total)
        paid    = float(summary.get("total_paid", 0))
        rem     = max(float(summary.get("remaining", 0)), 0.0)
        pct     = float(summary.get("payment_percentage", 0))

        next_stage = self._next_unpaid_payment_stage(project_data)

        # Find the stage AFTER the current one
        after_stage_name = ""
        planned_rows  = self._planned_payment_rows(project_data)
        total_stages  = len(planned_rows) if planned_rows else 1
        if next_stage and planned_rows:
            cur_stage_name = next_stage.get("stage", "")
            for i, pr in enumerate(planned_rows):
                if pr.get("stage") == cur_stage_name:
                    after_stage_name = (
                        planned_rows[i + 1].get("stage", "")
                        if i + 1 < len(planned_rows)
                        else "— Final Stage"
                    )
                    break

        if pct >= 100.0 or rem <= 0:
            stage_text = "Paid in Full ✓"
            bg, fg, border = "#dcfce7", "#047857", "#86efac"
            paid_text = f"Total Paid: {self._format_project_money(total)}"
            after_stage_name = ""
        elif next_stage:
            stage = next_stage.get("stage", "")
            display_stage = stage if stage else "Payment Due"
            if display_stage in ("Full Amount", "Full Amt"):
                display_stage = "Full Amount Due"
            bg, fg, border, _ = self._payment_category_colors(stage)
            stage_text = display_stage
            paid_text = (
                f"Paid: {self._format_project_money(paid)}  ·  Rem: {self._format_project_money(rem)}"
                if paid > 0 else
                f"Amt Due: {self._format_project_money(rem)}"
            )
        else:
            stage_text = project_data.get("payment_category", "—") or "—"
            if stage_text in ("Full Amount", "Full Amt"):
                stage_text = "Full Amount Due"
            bg, fg, border, _ = self._payment_category_colors(stage_text)
            paid_text = f"Amt Due: {self._format_project_money(rem)}"

        # ── Stage badge ───────────────────────────────────────────────────
        stage_lbl = QtWidgets.QLabel(stage_text)
        stage_lbl.setAlignment(QtCore.Qt.AlignCenter)
        stage_lbl.setFixedHeight(22)
        stage_lbl.setStyleSheet(
            f"background:{bg}; color:{fg}; border:1px solid {border};"
            "border-radius:7px; padding:0 8px;"
            "font-size:11px; font-weight:900;"
            "font-family:'Inter','Segoe UI',sans-serif;"
        )
        stage_lbl.setMinimumWidth(130)
        stage_lbl.setMaximumWidth(211)
        stage_lbl.setToolTip(
            f"Current: {stage_text}\n"
            f"Paid: {self._format_project_money(paid)}\n"
            f"Remaining: {self._format_project_money(rem)}"
        )
        vbox.addWidget(stage_lbl, 0, QtCore.Qt.AlignCenter)

        # ── Paid / remaining line ─────────────────────────────────────────
        info_lbl = QtWidgets.QLabel(paid_text)
        info_lbl.setAlignment(QtCore.Qt.AlignCenter)
        info_lbl.setWordWrap(False)
        info_lbl.setMaximumWidth(211)
        info_lbl.setStyleSheet(
            f"font-size:10px; font-weight:700; color:{fg};"
            "font-family:'Inter','Segoe UI',sans-serif;"
            "background:transparent; border:none;"
        )
        vbox.addWidget(info_lbl, 0, QtCore.Qt.AlignCenter)

        # ── Next stage hint ───────────────────────────────────────────────
        if after_stage_name and after_stage_name != "— Final Stage":
            next_text, next_color = f"Next → {after_stage_name}", "#64748b"
        elif after_stage_name == "— Final Stage":
            next_text, next_color = "Final payment stage", "#b45309"
        elif total_stages == 1 and pct < 100:
            next_text, next_color = "Single payment plan", "#94a3b8"
        else:
            next_text = ""
            next_color = "#94a3b8"

        if next_text:
            next_lbl = QtWidgets.QLabel(next_text)
            next_lbl.setAlignment(QtCore.Qt.AlignCenter)
            next_lbl.setWordWrap(False)
            next_lbl.setMaximumWidth(211)
            next_lbl.setStyleSheet(
                f"font-size:9px; font-weight:700; color:{next_color};"
                "font-family:'Inter','Segoe UI',sans-serif;"
                "background:transparent; border:none;"
            )
            vbox.addWidget(next_lbl, 0, QtCore.Qt.AlignCenter)

        return container

    def edit_single_project(self, project):
        """Edit single project details"""
        try:
            self.show_inline_project_editor(project)
                
        except Exception as e:
            _log.warning("Error editing project: %s", e)
            QtWidgets.QMessageBox.critical(
                self, "Error",
                f"Error loading project for editing: {str(e)}"
            )

    def edit_project(self, project):
        """Backward-compatible action handler."""
        return self.edit_single_project(project)
    
    def delete_single_project(self, project):
        """Delete single project from action button.

        Only the project record itself is removed.  All invoices, balance-sheet
        entries, and payment histories are intentionally preserved so financial
        history stays intact.
        """
        project_number = project["project_number"]

        reply = QtWidgets.QMessageBox.question(
            self, "Delete Project",
            f"Delete this project?\n\n"
            f"{project['project_name']}\n\n"
            f"Note: Invoices, balance sheet entries, and payment histories\n"
            f"for this project will be preserved.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )

        if reply == QtWidgets.QMessageBox.Yes:
            for i, proj in enumerate(self.generated_projects):
                if proj["project_number"] == project_number:
                    self.generated_projects.pop(i)
                    break

            if FIREBASE_AVAILABLE and "firebase_id" in project:
                try:
                    db.reference(f'/projects/{project["firebase_id"]}').delete()
                    _log.info("Project deleted from Firebase: %s", project_number)
                except Exception as e:
                    _log.warning("Error deleting project from Firebase: %s", e)

            # Do NOT cascade-delete invoices, balance-sheet entries, or payments.
            # Financial history must be preserved after project deletion.

            self.update_projects_table()
            self.update_client_filter_menu()
            self.update_stats()

            QtWidgets.QMessageBox.information(
                self, "Deleted",
                f"Project '{project_number}' deleted.\n"
                f"Invoices, balance sheet entries, and payment history preserved."
            )

    def delete_project(self, project):
        """Backward-compatible action handler."""
        return self.delete_single_project(project)

    def _cascade_delete_project_data(self, project_number: str):
        """Intentionally preserved: invoices, balance sheet entries, and payment histories
        must survive a project deletion so financial history stays intact.
        This method is kept as a no-op stub for backward compatibility."""
        _log.info(
            "Project %s removed. Invoices, balance sheet, and payments preserved.",
            project_number,
        )

    def perform_projects_pdf_export(self, export_params):
        """Perform PDF export for projects based on RECEIVED DATE"""
        try:
            projects_to_export = []
            
            for project in self.generated_projects:
                try:
                    project_datetime = None
                    
                    # Use RECEIVED DATE for filtering
                    received_date = project.get('date_received', '')
                    if received_date:
                        date_formats = ["%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"]
                        
                        for date_format in date_formats:
                            try:
                                project_datetime = datetime.strptime(received_date, date_format)
                                break
                            except ValueError:
                                continue
                    
                    if project_datetime is None:
                        _log.warning("Warning: Could not parse received date for project - EXCLUDING from export")
                        continue
                    
                    include_project = False
                    
                    if export_params["range"] == "all":
                        include_project = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        project_date_only = project_datetime.date()
                        
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        if from_date_only <= project_date_only <= to_date_only:
                            include_project = True
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if project_datetime.month == month and project_datetime.year == year:
                            include_project = True
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if project_datetime.year == year:
                            include_project = True
                    
                    if include_project:
                        projects_to_export.append(project)
                            
                except Exception as e:
                    _log.warning("Error processing project: %s", e)
                    continue
            
            if not projects_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No projects found matching the selected criteria.")
                return
            
            _log.info("PDF Export: Found %s projects to export (filtered by received date)", len(projects_to_export))
            
            # Sort by RECEIVED DATE (oldest first for export)
            def sort_by_received_date_asc(project):
                received_date = project.get('date_received', '')
                try:
                    return datetime.strptime(received_date, "%m-%d-%Y")
                except:
                    try:
                        return datetime.strptime(received_date, "%Y-%m-%d")
                    except:
                        return datetime.now()
            projects_to_export.sort(
                key=lambda p: (
                    sort_by_received_date_asc(p),  # ✅ keep your old received date logic
                    datetime.fromisoformat(p.get('created_at', '').replace("Z", "")) 
                    if p.get('created_at') else datetime.min
                )
            )
            
            self.generate_projects_combined_pdf(projects_to_export, export_params)

        except Exception as e:
            _log.warning("Error performing PDF export: %s", e)
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during PDF export: {str(e)}")
            
    def update_payment_category(self, project_data: dict, new_category: str):
        """Update payment category and sync it everywhere"""
        try:
            project_number = project_data.get("project_number")
            
            if new_category == "Select Payment Category" or not new_category:
                stored_category = "N/A"
            else:
                stored_category = new_category

            project_data["payment_category"] = stored_category
            project_data["payment_terms"] = stored_category

            for row in range(self.projects_table.rowCount()):
                table_item = self.projects_table.item(row, 1)
                if table_item:
                    stored_data = table_item.data(QtCore.Qt.UserRole)
                    if stored_data and stored_data.get("project_number") == project_number:
                        stored_data["payment_category"] = stored_category
                        stored_data["payment_terms"] = stored_category
                        table_item.setData(QtCore.Qt.UserRole, stored_data)
                        break

            for p in self.generated_projects:
                if p.get("project_number") == project_number:
                    p["payment_category"] = stored_category
                    p["payment_terms"] = stored_category
                    break

            for p in self.cached_projects:
                if p.get("project_number") == project_number:
                    p["payment_category"] = stored_category
                    p["payment_terms"] = stored_category
                    break

            if FIREBASE_AVAILABLE and "firebase_id" in project_data:
                project_data_copy = project_data.copy()
                project_data_copy["payment_category"] = stored_category
                project_data_copy["payment_terms"] = stored_category
                FirebaseManager.save_project(project_data_copy)

            _log.info("💳 Payment category synced: %s → %s", project_number, stored_category)

        except Exception as e:
            _log.warning("❌ Error updating payment category: %s", e)

    def generate_projects_combined_pdf(self, projects, export_params):
        """Generate a professional combined PDF report for projects - LANDSCAPE MODE"""
        try:
            export_dir = Path.home() / "Downloads" / "Project_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Projects_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Projects_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Projects_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Projects_{year}.pdf"

            pdf_path = export_dir / filename

            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # Use LANDSCAPE mode
            doc = SimpleDocTemplate(str(pdf_path), pagesize=landscape(A4), 
                                topMargin=0.2*inch, bottomMargin=0.2*inch,
                                leftMargin=0.2*inch, rightMargin=0.2*inch)
            elements = []

            styles = getSampleStyleSheet()
            
            mabs_header_style = ParagraphStyle(
                'MABSHeader',
                parent=styles['Normal'],
                fontSize=16,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,  # Center alignment
                fontName='Helvetica-Bold',
                spaceAfter=2,
                spaceBefore=5
            )
            
            report_title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=12,
                spaceAfter=6,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold'
            )

            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=6,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold'
            )

            period_style = ParagraphStyle(
                'PeriodStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#7f8c8d'),
                alignment=0,  # Right alignment
                fontName='Helvetica',
                spaceAfter=3
            )

            # Header with company name centered
            try:
                from main import Config as _Cfg
                _co_hdr = _Cfg.COMPANY.get('name', 'MABS Engineering LLC')
            except Exception:
                _co_hdr = 'MABS Engineering LLC'
            page_width = landscape(A4)[0]
            header_data = [[Paragraph(_co_hdr, mabs_header_style)]]
            header_table = Table(header_data, colWidths=[page_width])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.15*inch))
            
            report_title = Paragraph("Project Overview Report", report_title_style)
            elements.append(report_title)
            
            total_projects = len(projects)
            completed_projects = len([p for p in projects if p.get('status') in ['Completed Not Invoiced', 'Completed & Invoiced']])
            cancelled_projects = len([p for p in projects if p.get('status') == 'Cancelled'])
            
            elements.append(Spacer(1, 0.08*inch))
            
            stats_text = f"Total Projects: {total_projects}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Completed: {completed_projects}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Cancelled: {cancelled_projects}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            # Period information - right aligned
            if export_params["range"] == "all":
                export_range_text = "All Projects"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"Period: {from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"Period: {month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Period: Year {year}"
            
            period_data = [[Paragraph(export_range_text, period_style)]]
            period_table = Table(period_data, colWidths=[page_width])
            period_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (0, 0), 20),
                ('RIGHTPADDING', (0, 0), (0, 0), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(period_table)
            elements.append(Spacer(1, 0.15*inch))

            if projects:
                # Column widths for landscape mode
                col_widths = [
                    0.35 * inch,   # S.No.
                    1.25 * inch,   # Project Number
                    1.1 * inch,    # Project Name
                    1.3 * inch,    # Client (increased)
                    0.9 * inch,    # PO/WO
                    0.8 * inch,    # Received Date
                    0.9 * inch,    # Plant
                    0.9 * inch,    # Sales
                    0.8 * inch,    # Price
                    0.8 * inch,    # Due Date
                    1.2 * inch,    # Status (increased)
                ]

                _cs = ParagraphStyle('_cs', fontName='Helvetica', fontSize=6.5,
                    leading=9, alignment=1, textColor=colors.HexColor('#2c3e50'))
                _ch = ParagraphStyle('_ch', fontName='Helvetica-Bold', fontSize=7,
                    leading=9, alignment=1, textColor=colors.whitesmoke)

                table_data = [[
                    Paragraph("S.No.", _ch), Paragraph("Project Number", _ch),
                    Paragraph("Project Name", _ch), Paragraph("Client", _ch),
                    Paragraph("PO/WO", _ch), Paragraph("Received Date", _ch),
                    Paragraph("Plant", _ch), Paragraph("Sales", _ch),
                    Paragraph("Price", _ch), Paragraph("Due Date", _ch),
                    Paragraph("Status", _ch),
                ]]
                
                for idx, project in enumerate(projects, 1):
                    start_date = project.get('start_date', '')
                    due_date = project.get('due_date', '')
                    date_received = project.get('date_received', '')
                    
                    # Format dates
                    try:
                        start_date_dt = datetime.strptime(start_date, "%m-%d-%Y")
                        start_date_us = start_date_dt.strftime("%m/%d/%Y")
                    except:
                        try:
                            start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
                            start_date_us = start_date_dt.strftime("%m/%d/%Y")
                        except:
                            start_date_us = start_date
                    
                    try:
                        due_date_dt = datetime.strptime(due_date, "%m-%d-%Y")
                        due_date_us = due_date_dt.strftime("%m/%d/%Y")
                    except:
                        try:
                            due_date_dt = datetime.strptime(due_date, "%Y-%m-%d")
                            due_date_us = due_date_dt.strftime("%m/%d/%Y")
                        except:
                            due_date_us = due_date
                    
                    try:
                        date_received_dt = datetime.strptime(date_received, "%m-%d-%Y")
                        date_received_us = date_received_dt.strftime("%m/%d/%Y")
                    except:
                        try:
                            date_received_dt = datetime.strptime(date_received, "%Y-%m-%d")
                            date_received_us = date_received_dt.strftime("%m/%d/%Y")
                        except:
                            date_received_us = date_received
                    
                    amount = project.get('project_amount', 0.0)
                    if isinstance(amount, (int, float)):
                        amount_formatted = f"${amount:,.2f}"
                    else:
                        amount_formatted = str(amount)
                    
                    def truncate(text, max_chars):
                        if len(str(text)) > max_chars:
                            return str(text)[:max_chars-3] + "..."
                        return str(text)
                    
                    payment_category = project.get('payment_category', 'N/A')
                    if payment_category in ["Select", "Select Payment Category", ""]:
                        payment_category = "N/A"
                    
                    status_value = project.get('status', 'Not Started')
                    
                    table_data.append([
                        Paragraph(str(idx), _cs),
                        Paragraph(project.get('project_number', ''), _cs),
                        Paragraph(project.get('project_name', ''), _cs),
                        Paragraph(project.get('company', ''), _cs),
                        Paragraph(project.get('po_wo_number', ''), _cs),
                        Paragraph(date_received_us, _cs),
                        Paragraph(project.get('plant', ''), _cs),
                        Paragraph(project.get('sales', ''), _cs),
                        Paragraph(amount_formatted, _cs),
                        Paragraph(due_date_us, _cs),
                        Paragraph(status_value, _cs),
                    ])
                
                projects_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                # Apply styles
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                    ('TOPPADDING', (0, 0), (-1, 0), 5),
                    
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffffff')),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                    
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8f9fa'), colors.white]),
                    
                    ('TOPPADDING', (0, 1), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                    
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ]
                
                projects_table.setStyle(TableStyle(table_style))
                elements.append(projects_table)
            else:
                no_data_style = ParagraphStyle(
                    'NoData',
                    parent=styles['Normal'],
                    fontSize=12,
                    textColor=colors.HexColor('#7f8c8d'),
                    alignment=1
                )
                elements.append(Paragraph("No projects found for the selected criteria.", no_data_style))

            doc.build(elements)

            if self.open_project_pdf_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"The PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            _log.warning("Error generating combined PDF: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", 
                                        f"Error generating PDF: {str(e)}")
            
    def open_project_pdf_file(self, file_path):
        """Open file with default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == "Darwin":
                subprocess.call(("open", file_path))
            elif platform.system() == "Windows":
                os.startfile(file_path)
            else:
                subprocess.call(("xdg-open", file_path))
            return True
        except Exception as e:
            _log.warning("Error opening file: %s", e)
            return False

    def perform_projects_excel_export(self, export_params):
        """Perform Excel export for projects based on RECEIVED DATE - includes all fields"""
        try:
            projects_to_export = []
            
            for project in self.generated_projects:
                try:
                    project_datetime = None
                    date_formats = ["%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"]
                    
                    # Use RECEIVED DATE for filtering
                    for date_format in date_formats:
                        try:
                            project_datetime = datetime.strptime(project.get('date_received', ''), date_format)
                            break
                        except ValueError:
                            continue
                    
                    if project_datetime is None:
                        _log.warning("Warning: Could not parse received date for project - EXCLUDING from export")
                        continue
                    
                    include_project = False
                    
                    if export_params["range"] == "all":
                        include_project = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        project_date_only = project_datetime.date()
                        
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        if from_date_only <= project_date_only <= to_date_only:
                            include_project = True
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if project_datetime.month == month and project_datetime.year == year:
                            include_project = True
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if project_datetime.year == year:
                            include_project = True
                    
                    if include_project:
                        projects_to_export.append(project)
                            
                except Exception as e:
                    _log.warning("Error processing project: %s", e)
                    continue
            
            if not projects_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No projects found matching the selected criteria.")
                return
            
            _log.info("Excel Export: Found %s projects to export (filtered by received date)", len(projects_to_export))
            
            # Sort by RECEIVED DATE (oldest first for export)
            def sort_by_received_date_asc(project):
                received_date = project.get('date_received', '')
                try:
                    return datetime.strptime(received_date, "%m-%d-%Y")
                except:
                    try:
                        return datetime.strptime(received_date, "%Y-%m-%d")
                    except:
                        return datetime.now()
            projects_to_export.sort(
                key=lambda p: (
                    sort_by_received_date_asc(p),  # ✅ keep your old received date logic
                    datetime.fromisoformat(p.get('created_at', '').replace("Z", "")) 
                    if p.get('created_at') else datetime.min
                )
            )
            
            self.generate_projects_combined_excel(projects_to_export, export_params)
 
        except Exception as e:
            _log.warning("Error performing Excel export: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during Excel export: {str(e)}")

    def generate_projects_combined_excel(self, projects, export_params):
        """Generate a professional combined Excel report for projects - ALL FIELDS from dialog"""
        try:
            export_dir = Path.home() / "Downloads" / "Project_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Projects_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Projects_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Projects_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Projects_{year}.xlsx"

            excel_path = export_dir / filename

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Projects"

            # Merge cells for title
            ws.merge_cells('A1:N1')
            try:
                from main import Config as _Cfg
                _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co = 'MABS ENGINEERING LLC'
            ws['A1'] = f"{_co} - PROJECTS REPORT"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            generated_date = datetime.now().strftime("%m-%d-%Y")
            ws['A2'] = f"Generated: {generated_date}"
            
            if export_params["range"] == "all":
                export_range_text = "All Projects"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m-%d-%Y")
                to_date = export_params["to_date"].strftime("%m-%d-%Y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws['A3'] = f"Period: {export_range_text}"

            # ALL FIELDS from dialog in same order as form
            headers = [
                "S.No.", "Project Number", "Project Name", "Client", "PO/WO Number",
                "Site Address", "Mail Address", "Received Date", "Plant", "Price",
                "Sales", "Due Date", "Status", "Notes"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
        
            for row_idx, project in enumerate(projects, 6):
                start_date = project.get('start_date', '')
                due_date = project.get('due_date', '')
                date_received = project.get('date_received', '')
                
                # Format dates
                try:
                    start_date_dt = datetime.strptime(start_date, "%m-%d-%Y")
                    start_date_formatted = start_date_dt.strftime("%m-%d-%Y")
                except:
                    try:
                        start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
                        start_date_formatted = start_date_dt.strftime("%m-%d-%Y")
                    except:
                        start_date_formatted = start_date
                
                try:
                    due_date_dt = datetime.strptime(due_date, "%m-%d-%Y")
                    due_date_formatted = due_date_dt.strftime("%m-%d-%Y")
                except:
                    try:
                        due_date_dt = datetime.strptime(due_date, "%Y-%m-%d")
                        due_date_formatted = due_date_dt.strftime("%m-%d-%Y")
                    except:
                        due_date_formatted = due_date
                
                try:
                    date_received_dt = datetime.strptime(date_received, "%m-%d-%Y")
                    date_received_formatted = date_received_dt.strftime("%m-%d-%Y")
                except:
                    try:
                        date_received_dt = datetime.strptime(date_received, "%Y-%m-%d")
                        date_received_formatted = date_received_dt.strftime("%m-%d-%Y")
                    except:
                        date_received_formatted = date_received
                
                amount = project.get('project_amount', 0.0)
                if isinstance(amount, (int, float)):
                    amount_formatted = f"${amount:,.2f}"
                else:
                    amount_formatted = str(amount)
                
                payment_category = project.get('payment_category', 'N/A')
                if payment_category in ["Select", "Select Payment Category", ""]:
                    payment_category = "N/A"
                
                data = [
                    row_idx - 5,
                    project.get('project_number', ''),
                    project.get('project_name', ''),
                    project.get('company', ''),
                    project.get('po_wo_number', ''),
                    project.get('site_address', ''),
                    project.get('mail_address', ''),
                    date_received_formatted,
                    project.get('plant', ''),
                    amount_formatted,
                    project.get('sales', ''),
                    due_date_formatted,
                    project.get('status', 'Not Started'),
                    project.get('notes', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)

                    if col == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

                    if col == 13:  # Status column
                        status = str(value).lower()
                        if 'completed & invoiced' in status:
                            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        elif 'completed not invoiced' in status:
                            cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                        elif 'in progress' in status:
                            cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E9", fill_type="solid")
                        elif 'paid' in status:
                            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                        elif 'cancelled' in status:
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            # Set column widths
            column_widths = {
                1: 6,    # S.No.
                2: 20,   # Project Number
                3: 30,   # Project Name
                4: 32,   # Client (increased)
                5: 15,   # PO/WO Number
                6: 35,   # Site Address
                7: 35,   # Mail Address
                8: 14,   # Received Date
                9: 15,   # Plant
                10: 14,  # Price
                11: 15,  # Sales
                12: 14,  # Due Date
                13: 28,  # Status (increased)
                14: 40   # Notes
            }
            
            for col_idx in range(1, len(headers) + 1):
                if col_idx in column_widths:
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = column_widths[col_idx]
                else:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[column_letter]:
                        if cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Alternate row colors
            for row in range(6, ws.max_row + 1):
                if row % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.fill.start_color.index == '00000000':
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

            wb.save(str(excel_path))

            if self.open_project_excel_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"The Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            _log.warning("Error generating combined Excel: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error", 
                                        f"Error generating Excel: {str(e)}")
        
    def open_project_excel_file(self, file_path):
        """Open Excel file with default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == "Darwin":
                subprocess.call(("open", file_path))
            elif platform.system() == "Windows":
                os.startfile(file_path)
            else:
                subprocess.call(("xdg-open", file_path))
            return True
        except Exception as e:
            _log.warning("Error opening file: %s", e)
            return False
        

class ProjectDialog(QtWidgets.QDialog):
    """Project Creation/Editing Dialog - Enhanced with all required fields"""
    
    def __init__(self, main_window, parent=None, project_data=None, firebase_available=False):
        super().__init__(parent)

        self.main_window = main_window
        self.project_data = project_data
        self.FIREBASE_AVAILABLE = firebase_available

        # MUST be defined BEFORE init_ui
        self.is_editing = bool(project_data)
        self.loading_job_details = False
        self.initializing_form = True
        self.was_in_balance_sheet = False  # Track if project was in balance sheet
        self._enter_on_last_field = False  # Track if Enter was pressed on last field

        # Dialog properties FIRST
        self.setWindowTitle("Edit Project" if self.is_editing else "Add Project")
        self.setModal(True)
        self.resize(1120, 720)
        self.setMinimumSize(980, 640)
        self.setStyleSheet("""
            QDialog {
                background: #f5f8fb;
            }
        """)

        # Build UI ONCE
        self.init_ui()

        # Check if project was in balance sheet BEFORE populating
        if self.is_editing:
            self.was_in_balance_sheet = self.check_if_in_balance_sheet()
            self.populate_form_data()
        else:
            self.auto_generate_project_number()  # Generate based on current date

        # Enable autofill logic AFTER everything is ready
        self.initializing_form = False
        self.setup_enter_key_navigation()
        self.setup_date_edit_key_handling()
        
        # Set up Ctrl+S shortcut
        save_shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+S"), self)
        save_shortcut.activated.connect(self.create_project)
        
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # ===== Header =====
        header = QtWidgets.QFrame()
        header.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border-bottom: 1px solid #d8e2ec;
                padding: 8px 18px;
            }
        """)
        header_layout = QtWidgets.QVBoxLayout(header)
        header_layout.setContentsMargins(18, 10, 18, 10)
        header_layout.setSpacing(3)

        title = QtWidgets.QLabel("Edit Project" if self.is_editing else "New Project")
        title.setStyleSheet("font-size: 22px; font-weight: 900; color: #0f172a; font-family:'Inter','Segoe UI';")
        subtitle = QtWidgets.QLabel("Update project details." if self.is_editing else "Fill the project details here. The dashboard stays visible after saving.")
        subtitle.setStyleSheet("font-size: 13px; font-weight: 700; color: #53657d; font-family:'Inter','Segoe UI';")

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)

        # ===== Scrollable Content =====
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: #f5f8fb;
            }
        """)

        # Create container for full scroll
        full_container = QtWidgets.QWidget()
        scroll.setWidget(full_container)

        full_layout = QtWidgets.QVBoxLayout(full_container)
        full_layout.setContentsMargins(0, 0, 0, 0)
        full_layout.setSpacing(0)
    
        # Move header INTO scroll
        full_layout.addWidget(header)

        # Create scroll_widget for the form
        scroll_widget = QtWidgets.QWidget()
        form_layout = QtWidgets.QVBoxLayout(scroll_widget)
        form_layout.setContentsMargins(24, 16, 24, 18)
        form_layout.setSpacing(10)

        # Add form into full container
        full_layout.addWidget(scroll_widget)
        layout.addWidget(scroll)

        # ===== BASIC INFORMATION SECTION =====
        # Create header layout with Quote Number on the right (like old code)
        basic_header_layout = QtWidgets.QHBoxLayout()
        
        basic_title = QtWidgets.QLabel("📝 Basic Information")
        basic_title.setStyleSheet("""
            QLabel {
                font-weight: 900;
                font-size: 15px;
                color: #0f172a;
                border-bottom: 1px solid #d8e2ec;
                padding-bottom: 5px;
                font-family: 'Inter', 'Segoe UI';
            }
        """)
        basic_header_layout.addWidget(basic_title)
        basic_title.setText("Basic Information")
        
        # Push Quote Number to the right
        basic_header_layout.addStretch(1)
        
        # Quote Number label + field
        job_label = QtWidgets.QLabel("Quote Number:")
        job_label.setStyleSheet("""
            font-size: 13px;
            font-weight: 800;
            color: #0f172a;
            font-family: 'Inter', 'Segoe UI';
        """)
        
        self.job_number_edit = self.create_styled_line_edit("Quote Number")
        self.job_number_edit.setFixedWidth(230)
        self.job_number_edit.textChanged.connect(self.on_job_number_changed)
        self.job_number_edit.editingFinished.connect(self.load_job_details)
        self.job_number_edit.returnPressed.connect(self.load_job_details)
        
        basic_header_layout.addWidget(job_label)
        basic_header_layout.addWidget(self.job_number_edit)
        
        form_layout.addLayout(basic_header_layout)
        basic_grid = QtWidgets.QGridLayout()
        basic_grid.setContentsMargins(0, 4, 0, 0)
        basic_grid.setHorizontalSpacing(18)
        basic_grid.setVerticalSpacing(10)
        form_layout.addLayout(basic_grid)

        # Project Number
        self.project_number_edit = self.create_styled_line_edit("Auto-generated (MABS-YYYYMM###)")
        self.add_field(basic_grid, "Project Number:", self.project_number_edit)
        if self.is_editing:
            self.job_number_edit.setReadOnly(True)
            self.project_number_edit.setReadOnly(True)
            self._apply_locked_identifier_style(self.job_number_edit)
            self._apply_locked_identifier_style(self.project_number_edit)

        # Add validation for duplicate project numbers
        def check_duplicate_project_number():
            if not self.is_editing:  # Only check for new projects
                project_number = self.project_number_edit.text().strip()
                if project_number and project_number != "Auto-generated (MABS-YYYYMM###)":
                    parent = self.parent()
                    if hasattr(parent, 'generated_projects') and parent.generated_projects:
                        for project in parent.generated_projects:
                            if project.get('project_number', '').upper() == project_number.upper():
                                QtWidgets.QMessageBox.warning(
                                    self,
                                    "Duplicate Project Number",
                                    f"Project number '{project_number}' already exists!"
                                )
                                self.project_number_edit.clear()
                                self.project_number_edit.setPlaceholderText("Auto-generated (MABS-YYYYMM###)")
                                return
        
        self.project_number_edit.editingFinished.connect(check_duplicate_project_number)

        # Project Name
        self.project_name_edit = self.create_styled_line_edit("Enter descriptive project name")
        self.add_field(basic_grid, "Project Name:", self.project_name_edit)

        # Client (with dropdown) - EXACTLY LIKE OLD CODE (NO AUTO-FILL)
        self.company_combo = self.create_styled_combo_box([])
        self.company_combo.setEditable(True)
        line = self.company_combo.lineEdit()

        # Remove placeholder when typing
        line.textEdited.connect(lambda t: line.setPlaceholderText(""))

        # Remove placeholder when clicking ANYWHERE
        old_mouse = line.mousePressEvent
        def new_mouse(event):
            line.setPlaceholderText("")
            old_mouse(event)

        line.mousePressEvent = new_mouse
        line.setPlaceholderText("Enter or select Company/Client name")
        
        # Connect company selection - NO AUTO-FILL, just set the value
        self.company_combo.activated[str].connect(self.on_company_selected)
        # When typing finishes OR Enter pressed
        self.company_combo.lineEdit().editingFinished.connect(
            lambda: self.on_company_selected(self.company_combo.currentText())
        )
        self.add_field(basic_grid, "Client:", self.company_combo)

        # PO/WO Number
        self.po_wo_edit = self.create_styled_line_edit("Enter PO or WO number")
        self.add_field(basic_grid, "PO/WO Number:", self.po_wo_edit)

        # Site Address
        self.site_address_edit = self.create_styled_line_edit("Enter site address")
        self.add_field(basic_grid, "Site Address:", self.site_address_edit)

        # Mail Address
        self.mail_address_edit = self.create_styled_line_edit("Enter mailing address")
        self.add_field(basic_grid, "Mail Address:", self.mail_address_edit)

        # Date of Receive
        self.date_received_edit = self.create_styled_date_edit_no_scroll(QtCore.QDate.currentDate())
        self.add_field(basic_grid, "Date Received:", self.date_received_edit)

        # Plant — editable combo with US state abbreviations
        _US_STATES = [
            "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA",
            "HI","ID","IL","IN","IA","KS","KY","LA","ME","MD",
            "MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
            "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC",
            "SD","TN","TX","UT","VT","VA","WA","WV","WI","WY",
        ]
        _combo_style = """
            QComboBox {
                padding: 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
            }
            QComboBox:focus { border-color: #00756f; background: #ffffff; }
            QComboBox QAbstractItemView { selection-background-color: #00756f; }
        """
        self.plant_edit = QtWidgets.QComboBox()
        self.plant_edit.setEditable(True)
        self.plant_edit.addItem("")
        self.plant_edit.addItems(_US_STATES)
        self.plant_edit.setCurrentIndex(0)
        self.plant_edit.lineEdit().setPlaceholderText("State / facility (e.g. MO, TX)")
        self.plant_edit.setStyleSheet(_combo_style)
        # Prevent scroll wheel from changing selection
        self.plant_edit.wheelEvent = lambda event: event.ignore()
        # Smart popup: show all items when no text typed; when text typed let completer handle it
        _plant_orig_popup = self.plant_edit.showPopup
        def _plant_smart_popup(_combo=self.plant_edit, _orig=_plant_orig_popup):
            text = _combo.lineEdit().text().strip()
            if not text:
                _orig()
        self.plant_edit.showPopup = _plant_smart_popup
        # Type-to-filter: show only states starting with typed letters
        _plant_completer = QtWidgets.QCompleter(_US_STATES, self.plant_edit)
        _plant_completer.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
        _plant_completer.setFilterMode(QtCore.Qt.MatchStartsWith)
        _plant_completer.setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
        _plant_completer.popup().setStyleSheet("""
            QListView {
                background: #ffffff; border: 1px solid #d8e2ec;
                border-radius: 7px; outline: none;
            }
            QListView::item {
                padding: 3px 10px; min-height: 20px;
                font-size: 13px; font-family: 'Inter', 'Segoe UI';
                color: #0f172a; border-radius: 4px;
            }
            QListView::item:selected { background: #e6f6f4; color: #00756f; }
        """)
        self.plant_edit.setCompleter(_plant_completer)
        self.add_field(basic_grid, "Plant (State):", self.plant_edit)

        # Price (same as cost)
        self.project_amount_edit = self.create_styled_line_edit("$0.00")
        self.project_amount_edit.textChanged.connect(self.validate_amount_input)
        self.project_amount_edit.textChanged.connect(self.update_deposit_fields)
        self.add_field(basic_grid, "Price:", self.project_amount_edit)

        # ===== SALES DROPDOWN (NEW - Like Client dropdown) =====
        self.sales_combo = self.create_styled_combo_box([])
        self.sales_combo.setEditable(True)
        sales_line = self.sales_combo.lineEdit()
        
        # Remove placeholder when typing
        sales_line.textEdited.connect(lambda t: sales_line.setPlaceholderText(""))
        
        # Remove placeholder when clicking ANYWHERE
        old_mouse_sales = sales_line.mousePressEvent
        def new_mouse_sales(event):
            sales_line.setPlaceholderText("")
            old_mouse_sales(event)
        
        sales_line.mousePressEvent = new_mouse_sales
        sales_line.setPlaceholderText("Enter or select Sales Person")
        # Prevent scroll wheel from changing selection
        self.sales_combo.wheelEvent = lambda event: event.ignore()
        # Force popup to open below the combo (not above)
        _sales_orig_show_popup = self.sales_combo.showPopup
        def _sales_show_popup_below(_combo=self.sales_combo, _orig=_sales_orig_show_popup):
            _orig()
            _popup = _combo.view().window()
            _global_pos = _combo.mapToGlobal(QtCore.QPoint(0, _combo.height()))
            _popup.move(_global_pos)
        self.sales_combo.showPopup = _sales_show_popup_below

        self.add_field(basic_grid, "Sales:", self.sales_combo)

        # Payment Category
        self.payment_category_combo = QtWidgets.QComboBox()
        self.payment_category_combo.addItems(["Single Payment", "25% Deposit + Balance", "50% Deposit + Final", "Custom Installments"])
        self.payment_category_combo.setEditable(False)
        self.payment_category_combo.currentTextChanged.connect(self.update_deposit_fields)
        self.payment_category_combo.wheelEvent = lambda event: event.ignore()

        self.payment_category_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
            }
            QComboBox:focus { 
                border-color: #00756f; 
                background: #ffffff; 
            }
            QComboBox QAbstractItemView {
                selection-background-color: #00756f;
            }
        """)

        self.payment_category_frame = self.add_field(basic_grid, "Payment Category:", self.payment_category_combo)
        self.payment_category_frame.setVisible(False)  # hidden — derived from payment plan

        # ── Payment Plan Section ──────────────────────────────────────────
        pay_section_label = QtWidgets.QLabel("💳 Payment Plan")
        pay_section_label.setStyleSheet(
            "font-weight:900;font-size:14px;color:#0f172a;"
            "font-family:'Inter','Segoe UI';margin-top:6px;"
        )
        basic_grid.addWidget(pay_section_label, basic_grid.rowCount(), 0, 1, 2)

        pay_panel = QtWidgets.QFrame()
        pay_panel.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: 1.5px solid #d8e2ec;
                border-radius: 10px;
            }
        """)
        pay_vbox = QtWidgets.QVBoxLayout(pay_panel)
        pay_vbox.setContentsMargins(16, 14, 16, 14)
        pay_vbox.setSpacing(10)
        basic_grid.addWidget(pay_panel, basic_grid.rowCount(), 0, 1, 2)

        # Row 1 — Down payment type (3 pill buttons)
        row1 = QtWidgets.QHBoxLayout()
        row1.setSpacing(0)
        type_lbl = QtWidgets.QLabel("Down Payment:")
        type_lbl.setStyleSheet("font-size:13px;font-weight:700;color:#334155;min-width:130px;")
        row1.addWidget(type_lbl)

        self._ptype_btns: dict = {}
        btn_frame = QtWidgets.QFrame()
        btn_frame.setStyleSheet("QFrame{background:#e2e8f0;border-radius:8px;border:none;}")
        btn_row = QtWidgets.QHBoxLayout(btn_frame)
        btn_row.setContentsMargins(3, 3, 3, 3)
        btn_row.setSpacing(2)
        _pill_ss = lambda active: (
            "QPushButton{background:#0f766e;color:white;border:none;border-radius:6px;"
            "font-size:12px;font-weight:800;padding:6px 14px;font-family:'Inter','Segoe UI';}"
            if active else
            "QPushButton{background:transparent;color:#64748b;border:none;border-radius:6px;"
            "font-size:12px;font-weight:700;padding:6px 14px;font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:white;color:#0f172a;}"
        )
        for label, key in [("No Down Payment", "no_down"),
                            ("50% Down Payment", "50_down"),
                            ("Custom %", "custom_down")]:
            btn = QtWidgets.QPushButton(label)
            btn.setCheckable(True)
            btn.clicked.connect(lambda _, k=key: self._set_payment_type(k))
            btn_row.addWidget(btn)
            self._ptype_btns[key] = btn
        row1.addWidget(btn_frame)
        row1.addStretch()
        pay_vbox.addLayout(row1)

        # Custom down payment input — toggle between % and $
        self._custom_pct_widget = QtWidgets.QWidget()
        _cpct_vbox = QtWidgets.QVBoxLayout(self._custom_pct_widget)
        _cpct_vbox.setContentsMargins(0, 0, 0, 0)
        _cpct_vbox.setSpacing(6)

        # Toggle row: % | $
        _toggle_row = QtWidgets.QHBoxLayout()
        _toggle_lbl = QtWidgets.QLabel("Down Payment:")
        _toggle_lbl.setStyleSheet("font-size:13px;font-weight:700;color:#334155;min-width:130px;")
        _toggle_row.addWidget(_toggle_lbl)
        self._dp_mode_btns: dict = {}
        _mode_frame = QtWidgets.QFrame()
        _mode_frame.setStyleSheet(
            "QFrame{background:#f1f5f9;border:1px solid #e2e8f0;border-radius:8px;}"
        )
        _mode_btn_row = QtWidgets.QHBoxLayout(_mode_frame)
        _mode_btn_row.setContentsMargins(3, 3, 3, 3)
        _mode_btn_row.setSpacing(2)
        _dp_btn_style = (
            "QPushButton{border:none;border-radius:6px;padding:4px 14px;"
            "font-size:12px;font-weight:700;background:transparent;color:#64748b;}"
            "QPushButton:checked{background:#0f766e;color:white;}"
        )
        for _lbl, _key in [("Percentage (%)", "pct"), ("Amount ($)", "amt")]:
            _btn = QtWidgets.QPushButton(_lbl)
            _btn.setCheckable(True)
            _btn.setFixedHeight(30)
            _btn.setStyleSheet(_dp_btn_style)
            _btn.clicked.connect(lambda _, k=_key: self._set_dp_mode(k))
            _mode_btn_row.addWidget(_btn)
            self._dp_mode_btns[_key] = _btn
        _toggle_row.addWidget(_mode_frame)
        _toggle_row.addStretch()
        _cpct_vbox.addLayout(_toggle_row)

        # Input row (% spin or $ spin, swapped by toggle)
        _input_row = QtWidgets.QHBoxLayout()
        _input_row.addSpacing(135)
        self.custom_pct_spin = QtWidgets.QSpinBox()
        self.custom_pct_spin.setRange(1, 99)
        self.custom_pct_spin.setValue(25)
        self.custom_pct_spin.setSuffix(" %")
        self.custom_pct_spin.setFixedWidth(110)
        self.custom_pct_spin.setFixedHeight(36)
        _spin_style = (
            "QSpinBox,QDoubleSpinBox{padding:6px 10px;border:1.5px solid #d8e2ec;border-radius:7px;"
            "font-size:13px;font-weight:600;background:white;color:#0f172a;"
            "font-family:'Inter','Segoe UI';}"
            "QSpinBox:focus,QDoubleSpinBox:focus{border-color:#0f766e;}"
        )
        self.custom_pct_spin.setStyleSheet(_spin_style)
        self.custom_amt_spin = QtWidgets.QDoubleSpinBox()
        self.custom_amt_spin.setRange(0.01, 9999999.99)
        self.custom_amt_spin.setDecimals(2)
        self.custom_amt_spin.setPrefix("$ ")
        self.custom_amt_spin.setValue(0.00)
        self.custom_amt_spin.setFixedWidth(140)
        self.custom_amt_spin.setFixedHeight(36)
        self.custom_amt_spin.setStyleSheet(_spin_style)
        self.custom_amt_spin.setVisible(False)
        _input_row.addWidget(self.custom_pct_spin)
        _input_row.addWidget(self.custom_amt_spin)
        _input_row.addStretch()
        _cpct_vbox.addLayout(_input_row)

        pay_vbox.addWidget(self._custom_pct_widget)
        self._dp_mode = "pct"
        self._dp_mode_btns["pct"].setChecked(True)

        # Row 2 — Remaining balance type (hidden for 50% down)
        self._remaining_widget = QtWidgets.QWidget()
        rem_vbox = QtWidgets.QVBoxLayout(self._remaining_widget)
        rem_vbox.setContentsMargins(0, 0, 0, 0)
        rem_vbox.setSpacing(6)

        rem_type_row = QtWidgets.QHBoxLayout()
        rem_lbl = QtWidgets.QLabel("Remaining Balance:")
        rem_lbl.setStyleSheet("font-size:13px;font-weight:700;color:#334155;min-width:130px;")
        rem_type_row.addWidget(rem_lbl)

        self._rem_btns: dict = {}
        rem_frame = QtWidgets.QFrame()
        rem_frame.setStyleSheet(btn_frame.styleSheet())
        rem_btn_row = QtWidgets.QHBoxLayout(rem_frame)
        rem_btn_row.setContentsMargins(3, 3, 3, 3)
        rem_btn_row.setSpacing(2)
        for label, key in [("Full Payment", "full"), ("Installments", "installments")]:
            btn = QtWidgets.QPushButton(label)
            btn.setCheckable(True)
            btn.clicked.connect(lambda _, k=key: self._set_remaining_type(k))
            rem_btn_row.addWidget(btn)
            self._rem_btns[key] = btn
        rem_type_row.addWidget(rem_frame)
        rem_type_row.addStretch()
        rem_vbox.addLayout(rem_type_row)

        # Installment count (shown only when installments chosen)
        self._inst_count_widget = QtWidgets.QWidget()
        inst_row = QtWidgets.QHBoxLayout(self._inst_count_widget)
        inst_row.setContentsMargins(0, 0, 0, 0)
        inst_lbl = QtWidgets.QLabel("Number of Installments:")
        inst_lbl.setStyleSheet("font-size:13px;font-weight:700;color:#334155;min-width:130px;")
        inst_row.addWidget(inst_lbl)
        self.installment_count_spin = QtWidgets.QSpinBox()
        self.installment_count_spin.setRange(2, 6)
        self.installment_count_spin.setValue(2)
        self.installment_count_spin.setFixedWidth(80)
        self.installment_count_spin.setFixedHeight(36)
        self.installment_count_spin.setStyleSheet(self.custom_pct_spin.styleSheet())
        inst_row.addWidget(self.installment_count_spin)
        inst_row.addWidget(QtWidgets.QLabel("payments"))
        inst_row.addStretch()
        rem_vbox.addWidget(self._inst_count_widget)

        pay_vbox.addWidget(self._remaining_widget)

        # Payment summary label (live preview) + View Plan button on same row
        self._pay_summary_lbl = QtWidgets.QLabel("")
        self._pay_summary_lbl.setStyleSheet(
            "font-size:12px;font-weight:700;color:#0f766e;"
            "background:#f0fdf9;border:1px solid #99f6e4;"
            "border-radius:7px;padding:6px 12px;"
        )

        self._view_plan_btn = QtWidgets.QPushButton("View Plan")
        self._view_plan_btn.setFixedHeight(34)
        self._view_plan_btn.setVisible(False)
        self._view_plan_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._view_plan_btn.setStyleSheet(
            "QPushButton{background:#0f766e;color:white;border:none;border-radius:7px;"
            "font-size:12px;font-weight:800;padding:0 18px;"
            "font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:#0d9488;}"
            "QPushButton:pressed{background:#0a6960;}"
        )
        self._view_plan_btn.clicked.connect(self._show_payment_plan_dialog)

        _plan_row = QtWidgets.QHBoxLayout()
        _plan_row.setSpacing(10)
        _plan_row.addWidget(self._pay_summary_lbl, 1)
        _plan_row.addWidget(self._view_plan_btn)
        pay_vbox.addLayout(_plan_row)

        # Due amount (project total - any upfront payment)
        self.due_amount_label = QtWidgets.QLabel("$0.00")
        self.due_amount_label.setStyleSheet(
            "font-size: 13px; font-weight: 700; color: #0f766e;"
            " background: #f0fdf9; border: 1px solid #99f6e4;"
            " border-radius: 6px; padding: 6px 12px;"
        )
        self.project_amount_edit.textChanged.connect(self._update_due_amount_label)

        # Legacy stubs so existing code that references them doesn't crash
        self.payment_terms_combo = QtWidgets.QComboBox()
        self.payment_terms_combo.addItems(["Single Payment", "50% Deposit + Final",
                                           "Custom Installments", "25% Deposit + Balance"])
        self.payment_terms_combo.setVisible(False)
        self.payment_split_edit = QtWidgets.QLineEdit()
        self.payment_split_edit.setVisible(False)
        self.payment_split_frame = QtWidgets.QWidget()
        self.deposit_rule_combo = QtWidgets.QComboBox()
        self.deposit_rule_combo.setVisible(False)
        self.deposit_rule_frame = QtWidgets.QWidget()
        self.deposit_received_combo = QtWidgets.QComboBox()
        self.deposit_received_combo.addItems(["No", "Yes"])
        self.deposit_received_combo.setVisible(False)
        self.deposit_received_frame = QtWidgets.QWidget()
        self.deposit_amount_edit = QtWidgets.QLineEdit()
        self.deposit_amount_edit.setVisible(False)
        self.deposit_amount_frame = QtWidgets.QWidget()
        self.due_amount_frame = QtWidgets.QWidget()
        self.deposit_received_date_edit = self.create_styled_date_edit_no_scroll(QtCore.QDate.currentDate())
        self.deposit_received_date_edit.setVisible(False)
        self.deposit_date_frame = QtWidgets.QWidget()

        # Connect live signals
        self.custom_pct_spin.valueChanged.connect(self._update_payment_summary)
        self.custom_amt_spin.valueChanged.connect(self._update_payment_summary)
        self.installment_count_spin.valueChanged.connect(self._update_payment_summary)
        self.project_amount_edit.textChanged.connect(self._update_payment_summary)
        self.project_amount_edit.textChanged.connect(self._update_view_plan_btn)

        # Default state: No Down Payment + Full Payment
        self._set_payment_type("no_down")
        self._set_remaining_type("full")

        # Notes (without scrolling)
        self.notes_edit = self.create_styled_line_edit("Enter any additional notes or comments")
        self.add_field(basic_grid, "Notes:", self.notes_edit)

        # ===== TIMELINE SECTION =====
        timeline_title = QtWidgets.QLabel("📅 Timeline")
        timeline_title.setStyleSheet("""
            QLabel {
                font-weight: 900;
                font-size: 15px;
                color: #0f172a;
                border-bottom: 1px solid #d8e2ec;
                padding-bottom: 5px;
                margin-top: 4px;
                font-family: 'Inter', 'Segoe UI';
            }
        """)
        form_layout.addWidget(timeline_title)
        timeline_grid = QtWidgets.QGridLayout()
        timeline_grid.setContentsMargins(0, 4, 0, 0)
        timeline_grid.setHorizontalSpacing(18)
        timeline_grid.setVerticalSpacing(10)
        form_layout.addLayout(timeline_grid)

        # Start Date, Duration, and Due Date on the same row
        self.start_date_edit = self.create_styled_date_edit_no_scroll(QtCore.QDate.currentDate())
        self.due_date_edit = self.create_styled_date_edit_no_scroll(QtCore.QDate.currentDate().addDays(30))

        self.duration_spin = QtWidgets.QSpinBox()
        self.duration_spin.setRange(1, 365)
        self.duration_spin.setValue(30)
        self.duration_spin.setSuffix(" days")
        self.duration_spin.setFixedHeight(36)
        self.duration_spin.setStyleSheet("""
            QSpinBox {
                padding: 6px 10px; border: 1.5px solid #d8e2ec;
                border-radius: 7px; font-size: 13px; font-weight: 600;
                background: white; color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
            }
            QSpinBox:focus { border-color: #00756f; }
        """)
        # Prevent scroll wheel and arrow keys from changing duration
        self.duration_spin.wheelEvent = lambda event: event.ignore()
        _dur_orig_key = self.duration_spin.keyPressEvent
        def _dur_key(event, _orig=_dur_orig_key):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            _orig(event)
        self.duration_spin.keyPressEvent = _dur_key
        self.duration_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        date_row = QtWidgets.QHBoxLayout()
        date_row.setSpacing(10)

        _lbl = lambda t: (lambda l: (l.setStyleSheet("font-weight:600;color:#334155;font-size:13px;"), l)[1])(QtWidgets.QLabel(t))
        date_row.addWidget(_lbl("Start:"))
        date_row.addWidget(self.start_date_edit, 2)
        date_row.addSpacing(6)
        date_row.addWidget(_lbl("Duration:"))
        date_row.addWidget(self.duration_spin, 1)
        date_row.addSpacing(6)
        date_row.addWidget(_lbl("Due Date:"))
        date_row.addWidget(self.due_date_edit, 2)

        form_layout.addLayout(date_row)

        # Timeline Info
        self.timeline_info = QtWidgets.QLabel()
        self.timeline_info.setStyleSheet(
            "font-size: 12px; color: #53657d; background: #eef6f5; padding: 8px 10px; border-radius: 7px; margin-top: 2px; font-weight:700;"
        )
        form_layout.addWidget(self.timeline_info)
        self.start_date_edit.dateChanged.connect(self._auto_calc_due_date)
        self.duration_spin.valueChanged.connect(self._auto_calc_due_date)
        self.due_date_edit.dateChanged.connect(self._on_due_date_manually_changed)
        self.update_timeline_info()

        # Sequence Field (hidden)
        self.sequence_edit = self.create_styled_line_edit("Auto-generated sequence")
        self.sequence_edit.setReadOnly(True)
        self.sequence_edit.setVisible(False)

        # ===== Action Buttons =====
        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.setSpacing(20)

        self.create_btn = QtWidgets.QPushButton("Update Project" if self.is_editing else "🚀 Add Project")
        self.create_btn.setText("Update Project" if self.is_editing else "Save Project")
        self.create_btn.setMinimumHeight(46)
        self.create_btn.setMinimumWidth(180)
        self.create_btn.setStyleSheet("""
            QPushButton {
                background-color: #00756f;
                color: white;
                font-weight: 800;
                font-size: 15px;
                border-radius: 8px;
                padding: 10px 20px;
                font-family: 'Inter', 'Segoe UI';
            }
            QPushButton:hover { background-color: #00645f; }
        """)
        self.create_btn.clicked.connect(self.create_project)

        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setMinimumHeight(46)
        self.cancel_btn.setMinimumWidth(140)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #ffffff;
                color: #334155;
                border: 1.5px solid #cbd5e1;
                font-weight: 800;
                border-radius: 8px;
                padding: 10px 20px;
                font-family: 'Inter', 'Segoe UI';
            }
            QPushButton:hover { background: #f8fafc; border-color: #00756f; color: #00756f; }
        """)
        self.cancel_btn.clicked.connect(self.reject)

        btn_layout.addWidget(self.create_btn)
        btn_layout.addWidget(self.cancel_btn)

        form_layout.addLayout(btn_layout)
        self.update_deposit_fields()
        self._update_payment_split_visibility()
                
        # Load saved companies and populate form if editing
        self.load_saved_companies()
        self.load_sales_persons()  # NEW: Load sales persons
        
        if self.is_editing:
            self.populate_form_data()

    def load_sales_persons(self):
        """Load sales persons into dropdown."""
        try:
            self.sales_combo.clear()
            sales_list = []
            
            if self.FIREBASE_AVAILABLE:
                from main import FirebaseManager
                sales_list = [
                    person.get("name", "")
                    for person in FirebaseManager.load_sales_people()
                    if person.get("name")
                ]
                _log.info("Loaded %s sales persons from Firebase", len(sales_list))
            else:
                sales_list = [
                    person.get("name", "")
                    for person in _load_local_sales_people()
                    if person.get("name")
                ]
                _log.info("Loaded %s sales persons from local data", len(sales_list))

            self.sales_combo.addItems(sorted(set(sales_list)))
            
            self.sales_combo.setCurrentIndex(-1)
            self.sales_combo.lineEdit().clear()
            
        except Exception as e:
            _log.error("Error loading sales persons: %s", e)
        
    def on_company_selected(self, company_name):
        """Handle company selection - JUST SET THE VALUE, NO AUTO-FILL"""
        # Handle index → text safely
        if isinstance(company_name, int):
            company_name = self.company_combo.itemText(company_name)

        company_name = str(company_name).strip()

        if getattr(self, "initializing_form", False) or getattr(self, "loading_job_details", False):
            return

        if not company_name:
            return

        # Just set the text, no auto-fill of email or address
        # The user will manually enter email and address
        pass

    # ── New payment plan helpers ─────────────────────────────────────────
    _PILL_ON  = ("QPushButton{background:#0f766e;color:white;border:none;border-radius:6px;"
                 "font-size:12px;font-weight:800;padding:6px 14px;font-family:'Inter','Segoe UI';}")
    _PILL_OFF = ("QPushButton{background:transparent;color:#64748b;border:none;border-radius:6px;"
                 "font-size:12px;font-weight:700;padding:6px 14px;font-family:'Inter','Segoe UI';}"
                 "QPushButton:hover{background:white;color:#0f172a;}")

    def _set_payment_type(self, key: str):
        self._payment_type = key
        for k, btn in self._ptype_btns.items():
            btn.setStyleSheet(self._PILL_ON if k == key else self._PILL_OFF)
            btn.setChecked(k == key)
        self._custom_pct_widget.setVisible(key == "custom_down")
        # 50% down: remaining is always full (no installments choice)
        self._remaining_widget.setVisible(key != "50_down")
        self._update_payment_summary()

    def _set_remaining_type(self, key: str):
        self._remaining_type = key
        for k, btn in self._rem_btns.items():
            btn.setStyleSheet(self._PILL_ON if k == key else self._PILL_OFF)
            btn.setChecked(k == key)
        self._inst_count_widget.setVisible(key == "installments")
        self._update_payment_summary()

    def _set_dp_mode(self, mode: str):
        """Toggle custom down payment input between % and $."""
        self._dp_mode = mode
        for k, btn in self._dp_mode_btns.items():
            btn.setChecked(k == mode)
        pct_visible = (mode == "pct")
        self.custom_pct_spin.setVisible(pct_visible)
        self.custom_amt_spin.setVisible(not pct_visible)
        self._update_payment_summary()

    def _update_payment_summary(self):
        """Show a one-line preview of the payment plan."""
        if not hasattr(self, "_payment_type"):
            return
        try:
            total_text = self.project_amount_edit.text().replace("$", "").replace(",", "").strip()
            total = float(total_text) if total_text else 0.0
        except ValueError:
            total = 0.0

        ptype = getattr(self, "_payment_type", "no_down")
        rtype = getattr(self, "_remaining_type", "full")
        dp_mode = getattr(self, "_dp_mode", "pct")
        stages = self._build_stage_labels(ptype, rtype,
                                          getattr(self, "custom_pct_spin", None),
                                          getattr(self, "installment_count_spin", None),
                                          total,
                                          dp_mode=dp_mode,
                                          amt_spin=getattr(self, "custom_amt_spin", None))
        if hasattr(self, "_pay_summary_lbl"):
            self._pay_summary_lbl.setText("Plan:  " + "  →  ".join(stages))
        self._update_view_plan_btn()

        # Also update the legacy payment_terms_combo so collect_project_data still works
        if hasattr(self, "payment_terms_combo"):
            label = self._derive_payment_terms_label(ptype, rtype)
            idx = self.payment_terms_combo.findText(label)
            if idx >= 0:
                self.payment_terms_combo.blockSignals(True)
                self.payment_terms_combo.setCurrentIndex(idx)
                self.payment_terms_combo.blockSignals(False)

    @staticmethod
    def _build_stage_labels(ptype, rtype, pct_spin, inst_spin, total=0.0,
                            dp_mode="pct", amt_spin=None):
        """Return the ordered list of stage names for this payment plan."""
        inst = inst_spin.value() if inst_spin else 2
        if ptype == "no_down":
            if rtype == "full":
                return ["Full Payment"]
            return [f"{i}{'st' if i==1 else 'nd' if i==2 else 'rd' if i==3 else 'th'} Installment"
                    for i in range(1, inst + 1)]
        if ptype == "50_down":
            return ["Down Payment (50%)", "Remaining Balance"]
        # custom_down — label depends on whether user entered % or $
        if dp_mode == "amt" and amt_spin is not None:
            amt = amt_spin.value()
            pct = round((amt / total * 100), 1) if total else 0
            dp_label = f"Down Payment ({int(pct) if pct == int(pct) else pct}%)"
        else:
            pct = pct_spin.value() if pct_spin else 25
            dp_label = f"Down Payment ({pct}%)"
        stages = [dp_label]
        if rtype == "full":
            stages.append("Remaining Balance")
        else:
            stages += [f"{i}{'st' if i==1 else 'nd' if i==2 else 'rd' if i==3 else 'th'} Installment"
                       for i in range(1, inst + 1)]
        return stages

    @staticmethod
    def _derive_payment_terms_label(ptype, rtype):
        if ptype == "no_down":
            return "Single Payment" if rtype == "full" else "Custom Installments"
        if ptype == "50_down":
            return "50% Deposit + Final"
        return "25% Deposit + Balance" if rtype == "full" else "Custom Installments"

    def _update_view_plan_btn(self):
        """Show View Plan button only when a cost > 0 is entered."""
        if not hasattr(self, "_view_plan_btn"):
            return
        try:
            cost_text = self.project_amount_edit.text().replace("$", "").replace(",", "").strip()
            total = float(cost_text) if cost_text else 0.0
        except ValueError:
            total = 0.0
        self._view_plan_btn.setVisible(total > 0)

    def _compute_stage_amounts(self, ptype, rtype, total):
        """Return a dollar amount for each payment stage in order."""
        dp_mode = getattr(self, "_dp_mode", "pct")
        if ptype == "no_down":
            if rtype == "full":
                return [total]
            inst = getattr(self.installment_count_spin, "value", lambda: 2)()
            base = total / inst if inst else 0
            amounts = [round(base, 2)] * inst
            # Adjust last to cover any rounding diff
            amounts[-1] = round(total - sum(amounts[:-1]), 2)
            return amounts
        if ptype == "50_down":
            dp = round(total * 0.5, 2)
            return [dp, round(total - dp, 2)]
        # custom_down
        if dp_mode == "amt" and hasattr(self, "custom_amt_spin"):
            dp = round(self.custom_amt_spin.value(), 2)
        else:
            pct = getattr(self.custom_pct_spin, "value", lambda: 25)()
            dp = round(total * pct / 100, 2)
        remaining = round(total - dp, 2)
        if rtype == "full":
            return [dp, remaining]
        inst = getattr(self.installment_count_spin, "value", lambda: 2)()
        base = remaining / inst if inst else 0
        inst_amounts = [round(base, 2)] * inst
        inst_amounts[-1] = round(remaining - sum(inst_amounts[:-1]), 2)
        return [dp] + inst_amounts

    def _show_payment_plan_dialog(self):
        """Open a read-only payment plan overview window."""
        try:
            cost_text = self.project_amount_edit.text().replace("$", "").replace(",", "").strip()
            total = float(cost_text) if cost_text else 0.0
        except ValueError:
            total = 0.0

        ptype = getattr(self, "_payment_type", "no_down")
        rtype = getattr(self, "_remaining_type", "full")
        dp_mode = getattr(self, "_dp_mode", "pct")
        stages = self._build_stage_labels(ptype, rtype,
                                          getattr(self, "custom_pct_spin", None),
                                          getattr(self, "installment_count_spin", None),
                                          total, dp_mode=dp_mode,
                                          amt_spin=getattr(self, "custom_amt_spin", None))
        amounts = self._compute_stage_amounts(ptype, rtype, total)

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Payment Plan Overview")
        dlg.setModal(True)
        dlg.setMinimumWidth(460)
        dlg.setStyleSheet("QDialog{background:#f8fafc;}")

        lay = QtWidgets.QVBoxLayout(dlg)
        lay.setContentsMargins(24, 22, 24, 20)
        lay.setSpacing(14)

        # ── Title ──────────────────────────────────────────────────
        title_lbl = QtWidgets.QLabel("Payment Plan Overview")
        title_lbl.setStyleSheet(
            "font-size:17px;font-weight:900;color:#0f172a;"
            "font-family:'Inter','Segoe UI';"
        )
        lay.addWidget(title_lbl)

        # ── Total cost banner ───────────────────────────────────────
        cost_frame = QtWidgets.QFrame()
        cost_frame.setStyleSheet(
            "background:#f0fdf9;border:1.5px solid #6ee7b7;border-radius:9px;"
        )
        cost_lay = QtWidgets.QHBoxLayout(cost_frame)
        cost_lay.setContentsMargins(16, 10, 16, 10)
        cost_title = QtWidgets.QLabel("Total Project Cost")
        cost_title.setStyleSheet(
            "font-size:13px;font-weight:700;color:#334155;"
            "font-family:'Inter','Segoe UI';background:transparent;border:none;"
        )
        cost_val = QtWidgets.QLabel(f"${total:,.2f}")
        cost_val.setStyleSheet(
            "font-size:16px;font-weight:900;color:#065f46;"
            "font-family:'Inter','Segoe UI';background:transparent;border:none;"
        )
        cost_lay.addWidget(cost_title)
        cost_lay.addStretch()
        cost_lay.addWidget(cost_val)
        lay.addWidget(cost_frame)

        # ── Stages table ────────────────────────────────────────────
        table_frame = QtWidgets.QFrame()
        table_frame.setStyleSheet(
            "background:white;border:1.5px solid #e2e8f0;border-radius:10px;"
        )
        table_lay = QtWidgets.QVBoxLayout(table_frame)
        table_lay.setContentsMargins(0, 0, 0, 0)
        table_lay.setSpacing(0)

        # Header
        hdr_frame = QtWidgets.QFrame()
        hdr_frame.setStyleSheet(
            "background:#0f766e;border-radius:8px 8px 0 0;"
        )
        hdr_lay = QtWidgets.QHBoxLayout(hdr_frame)
        hdr_lay.setContentsMargins(16, 9, 16, 9)
        hdr_lay.setSpacing(0)
        for text, stretch in [("Payment Stage", 3), ("Amount", 2), ("% of Total", 2)]:
            h = QtWidgets.QLabel(text)
            h.setStyleSheet(
                "color:white;font-size:12px;font-weight:800;"
                "font-family:'Inter','Segoe UI';background:transparent;border:none;"
            )
            hdr_lay.addWidget(h, stretch)
        table_lay.addWidget(hdr_frame)

        for idx, (stage, amount) in enumerate(zip(stages, amounts)):
            is_last = (idx == len(stages) - 1)
            bg = "#f8fafc" if idx % 2 == 0 else "white"
            br = "border-radius:0 0 8px 8px;" if is_last else ""
            row_frame = QtWidgets.QFrame()
            row_frame.setStyleSheet(f"background:{bg};{br}")
            row_lay = QtWidgets.QHBoxLayout(row_frame)
            row_lay.setContentsMargins(16, 10, 16, 10)
            row_lay.setSpacing(0)

            stage_lbl = QtWidgets.QLabel(stage)
            stage_lbl.setStyleSheet(
                "font-size:13px;font-weight:600;color:#0f172a;"
                "font-family:'Inter','Segoe UI';background:transparent;border:none;"
            )
            amt_lbl = QtWidgets.QLabel(f"${amount:,.2f}")
            amt_lbl.setStyleSheet(
                "font-size:13px;font-weight:800;color:#0f766e;"
                "font-family:'Inter','Segoe UI';background:transparent;border:none;"
            )
            pct_str = f"{(amount / total * 100):.1f}%" if total > 0 else "—"
            pct_lbl = QtWidgets.QLabel(pct_str)
            pct_lbl.setStyleSheet(
                "font-size:13px;color:#64748b;"
                "font-family:'Inter','Segoe UI';background:transparent;border:none;"
            )
            row_lay.addWidget(stage_lbl, 3)
            row_lay.addWidget(amt_lbl, 2)
            row_lay.addWidget(pct_lbl, 2)
            table_lay.addWidget(row_frame)

        lay.addWidget(table_frame)

        # ── Close button ────────────────────────────────────────────
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setFixedHeight(38)
        close_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        close_btn.setStyleSheet(
            "QPushButton{background:#e2e8f0;color:#334155;border:none;"
            "border-radius:8px;font-size:13px;font-weight:700;"
            "padding:0 28px;font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:#cbd5e1;}"
        )
        close_btn.clicked.connect(dlg.accept)
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        lay.addLayout(btn_row)

        dlg.exec_()

    def update_deposit_fields(self):
        """No-op — deposit fields replaced by the new payment plan UI."""
        if getattr(self, "initializing_form", False):
            return
        # Kept as stub so any old signal connections don't crash.

    def _update_deposit_fields_legacy(self):
        """Legacy deposit field updater (only used internally when old fields exist)."""
        try:
            if not hasattr(self, "deposit_received_combo") or getattr(self, "initializing_form", False):
                return

            deposit_received = self.deposit_received_combo.currentText() == "Yes"
            plan = self.payment_terms_combo.currentText() if hasattr(self, "payment_terms_combo") else ""
            is_deposit_plan = "Deposit" in plan
            enabled = deposit_received and is_deposit_plan

            self.deposit_amount_edit.setEnabled(enabled)
            self.deposit_received_date_edit.setEnabled(enabled)

            # Lock payment category once deposit is marked received — prevents accidental changes
            if deposit_received and is_deposit_plan:
                self.payment_category_combo.setEnabled(False)
                self.payment_category_combo.setToolTip("Payment category is locked after deposit is received.")
            else:
                self.payment_category_combo.setEnabled(True)
                self.payment_category_combo.setToolTip("")

            if enabled and self.deposit_rule_combo.currentText() != "Custom":
                amount_text = self.project_amount_edit.text().replace("$", "").replace(",", "").strip()
                try:
                    amount = float(amount_text) if amount_text else 0.0
                except (ValueError, TypeError):
                    amount = 0.0
                percent = self.deposit_rule_percent()
                suggested = amount * percent
                # Auto-fill due payment as remainder
                self.deposit_amount_edit.setText(f"${suggested:,.2f}")
            elif not enabled:
                self.deposit_amount_edit.setText("$0.00")
        except Exception as exc:
            _log.warning("Could not update deposit fields: %s", exc)

    def _update_due_amount_label(self):
        """Recalculate and display Due Amount = Project Amount - Deposit Amount."""
        try:
            def _parse(text):
                return float(text.replace("$", "").replace(",", "").strip() or "0")
            total = _parse(self.project_amount_edit.text())
            deposit = _parse(self.deposit_amount_edit.text())
            due = max(total - deposit, 0.0)
            self.due_amount_label.setText(f"${due:,.2f}")
            color = "#0f766e" if due > 0 else "#64748b"
            self.due_amount_label.setStyleSheet(
                f"font-size: 13px; font-weight: 700; color: {color};"
                " background: #f0fdf9; border: 1px solid #99f6e4;"
                " border-radius: 6px; padding: 6px 12px;"
            )
        except Exception:
            pass

    def deposit_rule_percent(self):
        rule = self.deposit_rule_combo.currentText() if hasattr(self, "deposit_rule_combo") else ""
        if "50%" in rule:
            return 0.50
        if "25%" in rule:
            return 0.25
        return 0.0

    def _update_payment_split_visibility(self):
        """No-op — new payment plan UI replaced the old split/deposit fields."""
        pass

    def _payment_split_values(self):
        raw = self.payment_split_edit.text().strip() if hasattr(self, "payment_split_edit") else ""
        values = []
        for part in re.split(r"[,/ ]+", raw):
            if not part:
                continue
            try:
                values.append(float(part))
            except (TypeError, ValueError):
                continue
        return values

    def populate_form_data(self):
        """Populate form with existing project data for editing"""
        if not self.project_data:
            return
        
        self.initializing_form = True
        
        # Quote Number
        self.job_number_edit.setText(self._quote_number_from_project_data(self.project_data))
        
        # Project Number
        self.project_number_edit.setText(self.project_data.get('project_number', ''))
        
        # Project Name
        self.project_name_edit.setText(self.project_data.get('project_name', ''))
        
        # Client
        company = self.project_data.get('company', '')
        index = self.company_combo.findText(company)
        if index >= 0:
            self.company_combo.setCurrentIndex(index)
        else:
            self.company_combo.setEditText(company)
        
        # PO/WO Number
        self.po_wo_edit.setText(self.project_data.get('po_wo_number', ''))
        
        # Site Address
        self.site_address_edit.setText(self.project_data.get('site_address', ''))
        
        # Mail Address
        self.mail_address_edit.setText(self.project_data.get('mail_address', ''))
        
        # Date of Receive
        date_received_str = self.project_data.get('date_received', '')
        if date_received_str:
            try:
                for date_format in ["MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy"]:
                    date_received = QtCore.QDate.fromString(date_received_str, date_format)
                    if date_received.isValid():
                        self.date_received_edit.setDate(date_received)
                        break
            except:
                self.date_received_edit.setDate(QtCore.QDate.currentDate())
        else:
            self.date_received_edit.setDate(QtCore.QDate.currentDate())
        
        # Plant
        plant_val = self.project_data.get('plant', '')
        idx = self.plant_edit.findText(plant_val)
        if idx >= 0:
            self.plant_edit.setCurrentIndex(idx)
        else:
            self.plant_edit.setEditText(plant_val)
        
        # Price
        project_amount = self.project_data.get('project_amount', 0.0)
        try:
            amount_value = float(project_amount)
            self.project_amount_edit.setText(f"${amount_value:,.2f}")
        except (ValueError, TypeError):
            self.project_amount_edit.setText("$0.00")
        
        # ===== SALES (Updated to work with combo box) =====
        sales_value = self.project_data.get('sales', '')
        index = self.sales_combo.findText(sales_value)
        if index >= 0:
            self.sales_combo.setCurrentIndex(index)
        else:
            self.sales_combo.setEditText(sales_value)
        
        # Payment Category
        payment_category = self.project_data.get('payment_category', 'N/A')
        if payment_category == "N/A" or not payment_category:
            self.payment_category_combo.setCurrentIndex(0)
        else:
            index = self.payment_category_combo.findText(payment_category)
            if index >= 0:
                self.payment_category_combo.setCurrentIndex(index)
            else:
                self.payment_category_combo.setCurrentIndex(0)

        # ── Restore new payment plan fields ─────────────────────────────
        ptype = self.project_data.get("payment_type", "")
        if not ptype:
            # Derive from legacy fields
            pct = float(self.project_data.get("down_payment_percent", 0) or 0)
            terms = str(self.project_data.get("payment_terms", "") or "")
            if pct == 50 or "50%" in terms:
                ptype = "50_down"
            elif pct > 0:
                ptype = "custom_down"
            else:
                ptype = "no_down"

        rtype = self.project_data.get("remaining_type", "")
        if not rtype:
            terms = str(self.project_data.get("payment_terms", "") or "")
            rtype = "installments" if "installment" in terms.lower() or "custom" in terms.lower() else "full"

        inst     = int(self.project_data.get("installment_count", 2) or 2)
        pct      = int(self.project_data.get("down_payment_percent", 25) or 25)
        dp_mode  = self.project_data.get("down_payment_mode", "pct") or "pct"
        dep_amt  = float(self.project_data.get("deposit_amount", 0) or 0)
        if hasattr(self, "custom_pct_spin"):
            self.custom_pct_spin.blockSignals(True)
            self.custom_pct_spin.setValue(pct)
            self.custom_pct_spin.blockSignals(False)
        if hasattr(self, "custom_amt_spin"):
            self.custom_amt_spin.blockSignals(True)
            self.custom_amt_spin.setValue(dep_amt if dep_amt > 0 else 0.0)
            self.custom_amt_spin.blockSignals(False)
        if hasattr(self, "installment_count_spin"):
            self.installment_count_spin.blockSignals(True)
            self.installment_count_spin.setValue(inst)
            self.installment_count_spin.blockSignals(False)

        self._set_payment_type(ptype)
        self._set_remaining_type(rtype)
        if ptype == "custom_down" and hasattr(self, "_set_dp_mode"):
            self._set_dp_mode(dp_mode)
        
        # Notes
        self.notes_edit.setText(self.project_data.get('notes', ''))
        
        # Start Date
        start_date_str = self.project_data.get('start_date', '')
        if start_date_str:
            try:
                for date_format in ["MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy"]:
                    start_date = QtCore.QDate.fromString(start_date_str, date_format)
                    if start_date.isValid():
                        self.start_date_edit.setDate(start_date)
                        break
            except:
                self.start_date_edit.setDate(QtCore.QDate.currentDate())
        else:
            self.start_date_edit.setDate(QtCore.QDate.currentDate())
        
        # Due Date
        due_date_str = self.project_data.get('due_date', '')
        if due_date_str:
            try:
                for date_format in ["MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy"]:
                    due_date = QtCore.QDate.fromString(due_date_str, date_format)
                    if due_date.isValid():
                        self.due_date_edit.setDate(due_date)
                        break
            except:
                self.due_date_edit.setDate(QtCore.QDate.currentDate().addDays(30))
        else:
            self.due_date_edit.setDate(QtCore.QDate.currentDate().addDays(30))
        
        # Sequence
        sequence = self.project_data.get('sequence', '')
        self.sequence_edit.setText(str(sequence) if sequence else "")
        
        # Update timeline info
        self.update_timeline_info()
        
        # Check if in balance sheet
        self.was_in_balance_sheet = self.check_if_in_balance_sheet()
        self.initializing_form = False

    def _quote_number_from_project_data(self, project_data):
        """Return the linked quote number, including older records with missing job_number."""
        if not isinstance(project_data, dict):
            return ""

        for key in ("job_number", "quote_number", "quote_no", "quoteNumber"):
            value = str(project_data.get(key, "") or "").strip()
            if value:
                return value

        match = self._find_quote_for_project(project_data)
        if match:
            quote_number = str(match.get("job_number", "") or "").strip()
            if quote_number:
                project_data["job_number"] = quote_number
                return quote_number
        return ""

    def _known_quote_records(self):
        """Collect quote records from the loaded quote tab, Firebase, and local backup."""
        seen = set()
        quotes = []

        def add_records(records):
            for record in records or []:
                if not isinstance(record, dict):
                    continue
                quote_number = str(record.get("job_number", "") or "").strip().upper()
                marker = quote_number or id(record)
                if marker in seen:
                    continue
                seen.add(marker)
                quotes.append(record)

        if hasattr(self.main_window, "job_form_tab"):
            add_records(getattr(self.main_window.job_form_tab, "job_forms", []))
        try:
            from main import FirebaseManager
            add_records(FirebaseManager.load_job_forms())
        except Exception as exc:
            _log.warning("Could not load quote records for project link lookup: %s", exc)
        add_records(_load_local_job_forms())
        return quotes

    def _same_project_number(self, left, right):
        left_norm = re.sub(r"[^A-Z0-9]", "", str(left or "").upper())
        right_norm = re.sub(r"[^A-Z0-9]", "", str(right or "").upper())
        return bool(left_norm and right_norm and left_norm == right_norm)

    def _find_quote_for_project(self, project_data):
        """Find a quote linked to this project by project number or unique client/name match."""
        project_number = str(project_data.get("project_number", "") or "").strip()
        project_name = str(project_data.get("project_name", "") or "").strip().upper()
        client = str(project_data.get("company", "") or "").strip().upper()
        quotes = self._known_quote_records()

        if project_number:
            project_keys = ("project_number", "linked_project_number", "converted_project_number")
            for quote in quotes:
                if any(self._same_project_number(quote.get(key, ""), project_number) for key in project_keys):
                    return quote

        candidates = []
        for quote in quotes:
            quote_name = str(quote.get("project_name", "") or quote.get("job_title", "") or "").strip().upper()
            quote_client = str(quote.get("client", "") or quote.get("company", "") or "").strip().upper()
            if project_name and client and quote_name == project_name and quote_client == client:
                candidates.append(quote)

        return candidates[0] if len(candidates) == 1 else None

    def collect_project_data(self):
        """Collect all project data into a dictionary"""
        amount_text = self.project_amount_edit.text().replace('$', '').replace(',', '').strip()
        try:
            project_amount = float(amount_text) if amount_text else 0.0
        except (ValueError, TypeError):
            project_amount = 0.0
        
        ptype   = getattr(self, "_payment_type",  "no_down")
        rtype   = getattr(self, "_remaining_type", "full")
        dp_mode = getattr(self, "_dp_mode", "pct")
        pct     = self.custom_pct_spin.value() if hasattr(self, "custom_pct_spin") else 0
        inst    = self.installment_count_spin.value() if hasattr(self, "installment_count_spin") else 2
        amt_spin = getattr(self, "custom_amt_spin", None)

        if ptype == "no_down":
            down_pct = 0.0
            deposit_amount = 0.0
        elif ptype == "50_down":
            down_pct = 50.0
            deposit_amount = project_amount * 0.5
        elif dp_mode == "amt" and amt_spin is not None:
            deposit_amount = min(amt_spin.value(), project_amount)
            down_pct = round((deposit_amount / project_amount * 100), 4) if project_amount else 0.0
        else:
            down_pct = float(pct)
            deposit_amount = project_amount * down_pct / 100.0

        stages = ProjectDialog._build_stage_labels(ptype, rtype,
                                                   self.custom_pct_spin if hasattr(self, "custom_pct_spin") else None,
                                                   self.installment_count_spin if hasattr(self, "installment_count_spin") else None,
                                                   project_amount,
                                                   dp_mode=dp_mode,
                                                   amt_spin=amt_spin)
        payment_category = stages[0] if stages else "Full Payment"
        payment_terms    = ProjectDialog._derive_payment_terms_label(ptype, rtype)

        project_data = {
            'job_number':     self.job_number_edit.text().strip(),
            'quote_number':   self.job_number_edit.text().strip(),
            'project_number': self.project_number_edit.text().strip(),
            'project_name':   self.project_name_edit.text().strip(),
            'company':        self.company_combo.currentText().strip(),
            'po_wo_number':   self.po_wo_edit.text().strip(),
            'site_address':   self.site_address_edit.text().strip(),
            'mail_address':   self.mail_address_edit.text().strip(),
            'date_received':  self.date_received_edit.date().toString("MM-dd-yyyy"),
            'plant':          self.plant_edit.currentText().strip(),
            'project_amount': project_amount,
            'sales':          self.sales_combo.currentText().strip(),
            # Payment plan (new model)
            'payment_type':          ptype,
            'down_payment_percent':  down_pct,
            'down_payment_mode':     dp_mode,
            'remaining_type':        rtype,
            'installment_count':     inst if rtype == "installments" else 1,
            'payment_stages':        stages,
            # Derived / legacy fields
            'payment_category': payment_category,
            'payment_terms':    payment_terms,
            'deposit_amount':   deposit_amount,
            'deposit_rule':     f"{int(down_pct)}% Down Payment" if down_pct else "No Down Payment",
            'deposit_received': False,
            'deposit_received_date': "",
            'payment_split_percentages': "",
            'notes':      self.notes_edit.text().strip(),
            'start_date': self.start_date_edit.date().toString("MM-dd-yyyy"),
            'due_date':   self.due_date_edit.date().toString("MM-dd-yyyy"),
            'status': (self.project_data.get('status', 'Not Started')
                       if self.is_editing and self.project_data else 'Not Started'),
            'sequence': int(self.sequence_edit.text()) if self.sequence_edit.text() else 1,
            'updated_at': datetime.now().isoformat(),
        }
        
        if not self.is_editing:
            project_data['created_at'] = datetime.now().isoformat()
        
        return project_data

    def setup_enter_key_navigation(self):
        """Set up Enter key to navigate between fields instead of submitting form - EXACTLY LIKE JobFormDialog"""
        # List of all input widgets in tab order
        self.input_widgets = [
            self.job_number_edit,
            self.project_number_edit,
            self.project_name_edit,
            self.company_combo,
            self.po_wo_edit,
            self.site_address_edit,
            self.mail_address_edit,
            self.date_received_edit,
            self.plant_edit,
            self.project_amount_edit,
            self.sales_combo,
            self.payment_category_combo,
            self.payment_terms_combo,
            self.payment_split_edit,
            self.deposit_rule_combo,
            self.deposit_received_combo,
            self.deposit_amount_edit,
            self.deposit_received_date_edit,
            self.notes_edit,
            self.start_date_edit,
            self.due_date_edit,
            self.create_btn,
            self.cancel_btn
        ]
        
        # Install event filters for all widgets
        for widget in self.input_widgets:
            if widget:
                widget.installEventFilter(self)
                
                # Special handling for date edit widgets
                if isinstance(widget, QtWidgets.QDateEdit):
                    line_edit = widget.lineEdit()
                    if line_edit:
                        line_edit.installEventFilter(self)
        
        # Set tab order
        for i in range(len(self.input_widgets) - 1):
            QtWidgets.QWidget.setTabOrder(self.input_widgets[i], self.input_widgets[i + 1])
        
        # Disable default button behavior
        self.create_btn.setAutoDefault(False)
        self.create_btn.setDefault(False)
        self.cancel_btn.setAutoDefault(False)
        self.cancel_btn.setDefault(False)
        
        # Setup date field UX
        self.setup_date_field_ux()

    def setup_date_field_ux(self):
        """Setup better UX for date fields for manual entry"""
        for date_edit in [self.start_date_edit, self.due_date_edit, self.date_received_edit]:
            if date_edit:
                date_edit.setReadOnly(False)
                line_edit = date_edit.lineEdit()
                if line_edit:
                    line_edit.setPlaceholderText("MM-DD-YY")
                    line_edit.installEventFilter(self)
                    line_edit.textEdited.connect(
                        lambda text, de=date_edit: self.validate_date_input(de, text)
                    )

    def validate_date_input(self, date_edit, text):
        """Validate date input as user types"""
        if not text:
            return
        
        cleaned = ''.join(c for c in text if c.isdigit() or c in ['-', '/'])
        
        if len(cleaned) >= 2 and '-' not in cleaned and '/' not in cleaned:
            if len(cleaned) == 2:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))
        
        elif len(cleaned) >= 5 and cleaned[2] in ['-', '/'] and cleaned[5:] == '':
            if len(cleaned) == 5:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))

    def ensureWidgetVisible(self, widget):
        """Ensure the widget is visible in the scroll area"""
        try:
            scroll_area = None
            current_widget = widget
            
            while current_widget:
                if isinstance(current_widget, QtWidgets.QScrollArea):
                    scroll_area = current_widget
                    break
                current_widget = current_widget.parent()
            
            if scroll_area and widget:
                scroll_area.ensureWidgetVisible(widget)
                
        except Exception as e:
            _log.warning("Scroll error: %s", e)

    def eventFilter(self, source, event):
        """Handle Enter key navigation with proper focus handling - EXACTLY LIKE JobFormDialog"""
        
        # =====================================================
        # 1. Handle backspace/delete in date fields
        # =====================================================
        if isinstance(source, QtWidgets.QLineEdit):
            parent = source.parent()
            if parent and isinstance(parent, QtWidgets.QDateEdit) and parent in [self.start_date_edit, self.due_date_edit, self.date_received_edit]:
                
                if event.type() == QtCore.QEvent.FocusIn:
                    QtCore.QTimer.singleShot(10, lambda s=source: s.selectAll())
                    return False
                
                if event.type() == QtCore.QEvent.KeyPress:
                    if event.key() in [QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter]:
                        return False
                    
                    elif event.key() in [QtCore.Qt.Key_Backspace, QtCore.Qt.Key_Delete]:
                        if source.hasSelectedText():
                            source.clear()
                            return True
                    
                    elif event.key() in [
                        QtCore.Qt.Key_0, QtCore.Qt.Key_1, QtCore.Qt.Key_2, QtCore.Qt.Key_3,
                        QtCore.Qt.Key_4, QtCore.Qt.Key_5, QtCore.Qt.Key_6, QtCore.Qt.Key_7,
                        QtCore.Qt.Key_8, QtCore.Qt.Key_9, QtCore.Qt.Key_Minus, QtCore.Qt.Key_Slash
                    ]:
                        if source.hasSelectedText():
                            source.clear()
                        return False
        
        # =====================================================
        # 2. Handle Enter key navigation
        # =====================================================
        if event.type() == QtCore.QEvent.KeyPress:
            key = event.key()
            is_enter = key in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter)
            is_down = key == QtCore.Qt.Key_Down
            is_up = key == QtCore.Qt.Key_Up

            if not (is_enter or is_down or is_up):
                return False

            # Find current widget index
            current_index = -1
            for i, widget in enumerate(self.input_widgets):
                if widget == source:
                    current_index = i
                    break
                
                # Handle QComboBox lineEdit
                if isinstance(widget, QtWidgets.QComboBox) and source == widget.lineEdit():
                    current_index = i
                    break

            if current_index == -1:
                return False
            
            # Direction
            step = 1 if (is_enter or is_down) else -1
            next_index = current_index + step

            # ============================
            # END → ACTION BUTTONS
            # ============================
            if next_index >= len(self.input_widgets):
                # First Enter after last field → show action buttons
                if not self._enter_on_last_field:
                    self._enter_on_last_field = True
                    self.create_btn.setFocus()
                    self.ensureWidgetVisible(self.create_btn)
                    return True

                # Second Enter → loop to first field
                self._enter_on_last_field = False
                self.job_number_edit.setFocus()
                self.job_number_edit.selectAll()
                self.ensureWidgetVisible(self.job_number_edit)
                return True

            # Reset flag if user navigates normally
            self._enter_on_last_field = False
            
            # =====================================================
            # ACTION BUTTONS → LOOP BACK TO FIRST FIELD
            # =====================================================
            if source in [self.create_btn, self.cancel_btn] and is_enter:
                self._enter_on_last_field = False
                QtCore.QTimer.singleShot(20, lambda: self.job_number_edit.setFocus())
                QtCore.QTimer.singleShot(30, lambda: self.job_number_edit.selectAll())
                QtCore.QTimer.singleShot(40, lambda: self.ensureWidgetVisible(self.job_number_edit))
                return True

            # ============================
            # NORMAL NAVIGATION
            # ============================
            attempts = 0
            while 0 <= next_index < len(self.input_widgets) and attempts < len(self.input_widgets):
                widget = self.input_widgets[next_index]

                if widget and widget.isEnabled() and widget.isVisible():

                    widget.setFocus()

                    # Select text smartly
                    if isinstance(widget, QtWidgets.QLineEdit):
                        QtCore.QTimer.singleShot(10, widget.selectAll)

                    elif isinstance(widget, QtWidgets.QComboBox):
                        le = widget.lineEdit()
                        if le:
                            QtCore.QTimer.singleShot(10, le.selectAll)

                    elif isinstance(widget, QtWidgets.QDateEdit):
                        le = widget.lineEdit()
                        if le:
                            QtCore.QTimer.singleShot(10, le.selectAll)

                    self.ensureWidgetVisible(widget)
                    return True

                next_index += step
                attempts += 1

            return True
                
        return super().eventFilter(source, event)

    def setup_date_edit_key_handling(self):
        """Setup special key handling for date edit fields"""
        for widget in [self.start_date_edit, self.due_date_edit, self.date_received_edit]:
            if widget:
                widget.installEventFilter(self)
                widget.lineEdit().installEventFilter(self)

    def on_job_number_changed(self, text):
        """Handle Quote Number changes in real-time"""
        if getattr(self, "initializing_form", False):
            return

        if hasattr(self, '_job_timer'):
            self._job_timer.stop()
        
        self._job_timer = QtCore.QTimer()
        self._job_timer.setSingleShot(True)
        self._job_timer.timeout.connect(lambda: self.load_job_details())
        self._job_timer.start(1000)

    def load_job_details(self):
        """Load job details when a Quote Number is entered or changed"""
        if self.loading_job_details:
            return
        
        job_number = self.job_number_edit.text().strip()
        if not job_number or job_number == "Enter Quote Number (Optional)":
            return
        
        try:
            self.loading_job_details = True
            
            # Temporarily disable date change signal to prevent auto-generation
            self.start_date_edit.blockSignals(True)
            
            job_data = self.find_quote_by_number(job_number)
            
            if job_data:
                # Populate fields from job details
                project_name = job_data.get('project_name', '')
                if not project_name:
                    project_name = job_data.get('job_title', '')
                self.project_name_edit.setText(project_name)
                
                # Company from Client
                self.company_combo.setEditText(job_data.get('client', ''))
                
                # Mailing and site addresses from the quote
                self.mail_address_edit.setText(job_data.get('client_address', ''))
                
                site_address = job_data.get('project_site_address', '') or job_data.get('client_address', '')
                self.site_address_edit.setText(site_address)
                
                # Plant
                plant = job_data.get('plant', '')
                if not plant:
                    plant = job_data.get('project_site_address', '')
                self.plant_edit.setText(plant)
                
                # Sales - Now using combo box
                sales_value = job_data.get('sales', '')
                index = self.sales_combo.findText(sales_value)
                if index >= 0:
                    self.sales_combo.setCurrentIndex(index)
                else:
                    self.sales_combo.setEditText(sales_value)
                
                # Price from Engineering Costs (include expedite premium if applicable)
                cost_text = job_data.get('engineering_costs', '0')
                if cost_text:
                    try:
                        cost_clean = cost_text.replace('$', '').replace(',', '').strip()
                        cost_value = float(cost_clean) if cost_clean else 0.0
                        if job_data.get('expedite') is True:
                            exp_str = str(job_data.get('expedite_amount', '50%')).strip()
                            try:
                                if '%' in exp_str:
                                    pct = float(exp_str.replace('%', '').strip())
                                    cost_value = cost_value * (1 + pct / 100)
                                elif '$' in exp_str:
                                    extra = float(exp_str.replace('$', '').replace(',', '').strip())
                                    cost_value += extra
                                else:
                                    cost_value = cost_value * 1.5
                            except Exception:
                                cost_value = cost_value * 1.5
                        self.project_amount_edit.setText(f"${cost_value:,.2f}")
                    except Exception:
                        self.project_amount_edit.setText("$0.00")
                
                # Start Date
                start_date_str = job_data.get('start_date', '')
                if start_date_str:
                    try:
                        start_date = QtCore.QDate.fromString(start_date_str, "MM-dd-yyyy")
                        if start_date.isValid():
                            self.start_date_edit.setDate(start_date)
                    except:
                        pass
                
                # Due Date
                due_date_str = job_data.get('due_date', '')
                if due_date_str:
                    try:
                        due_date = QtCore.QDate.fromString(due_date_str, "MM-dd-yyyy")
                        if due_date.isValid():
                            self.due_date_edit.setDate(due_date)
                    except:
                        pass
                
                if not self.is_editing:
                    self.auto_generate_project_number()
                
                _log.info("New Project auto-filled from quote %s", job_number)
                            
        except Exception as e:
            _log.warning("Error loading job details: %s", e)
        finally:
            self.start_date_edit.blockSignals(False)
            self.loading_job_details = False

    def find_quote_by_number(self, job_number):
        """Find a quote by number from the open tab, Firebase, or local backup."""
        target = str(job_number or "").strip().upper()
        if not target:
            return None

        def match_from_iterable(jobs):
            for job in jobs or []:
                if isinstance(job, dict) and str(job.get('job_number', '')).strip().upper() == target:
                    return job
            return None

        if hasattr(self.main_window, 'job_form_tab'):
            job_form_tab = self.main_window.job_form_tab
            job_data = match_from_iterable(getattr(job_form_tab, 'job_forms', []))
            if job_data:
                return job_data

        if self.FIREBASE_AVAILABLE:
            try:
                from main import db
                ref = db.reference('/job_forms')
                quote_data = ref.order_by_child('job_number').equal_to(job_number).get() or {}
                for _, job in quote_data.items():
                    if isinstance(job, dict):
                        return job
            except Exception as exc:
                _log.warning("Could not load quote %s from Firebase: %s", job_number, exc)

        return match_from_iterable(_load_local_job_forms())

    def create_styled_date_edit_no_scroll(self, date):
        """Create styled date edit WITHOUT scroll, arrows, or auto increment"""
        d = QtWidgets.QDateEdit(date)
        d.setCalendarPopup(True)
        d.setDisplayFormat("MM-dd-yy")

        # Disable mouse wheel
        d.wheelEvent = lambda event: None

        # Disable arrow keys
        def keyPressEvent(event, original=d.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        d.keyPressEvent = keyPressEvent

        # Disable internal stepping
        d.stepBy = lambda x: None

        # Remove spin buttons completely
        d.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        d.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
            }
            QDateEdit:focus { border-color: #00756f; background: #ffffff; }
        """)

        return d

    def create_styled_line_edit(self, placeholder="", read_only=False):
        """Create styled line edit"""
        edit = QtWidgets.QLineEdit()
        edit.setPlaceholderText(placeholder)
        edit.setReadOnly(read_only)
        edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
            }
            QLineEdit:focus { border-color: #00756f; background: #ffffff; }
            QLineEdit[readOnly="true"] { background: #f1f5f9; color: #64748b; }
        """)
        return edit

    def _apply_locked_identifier_style(self, edit):
        edit.setFocusPolicy(QtCore.Qt.NoFocus)
        edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: #f1f5f9;
                color: #334155;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
            }
            QLineEdit[readOnly="true"] {
                background: #f1f5f9;
                color: #334155;
            }
        """)

    def create_styled_combo_box(self, items):
        """Create styled combo box - EXACTLY LIKE OLD CODE"""
        from app_theme import CHEVRON_URL

        combo = QtWidgets.QComboBox()
        combo.addItems(items)
        combo.setEditable(True)

        line_edit = combo.lineEdit()

        # When user types → remove placeholder IMMEDIATELY
        def on_user_type(text, le=line_edit, cb=combo):
            if le.text().startswith("-- Select"):
                clean = le.text().replace("-- Select Company --", "").strip()
                le.blockSignals(True)
                le.setText(clean)
                le.blockSignals(False)

        line_edit.textEdited.connect(on_user_type)

        # When user focuses → clear placeholder
        def on_focus_in(event, le=line_edit):
            if le.text().startswith("-- Select"):
                le.blockSignals(True)
                le.clear()
                le.blockSignals(False)
            return QtWidgets.QLineEdit.focusInEvent(le, event)

        line_edit.focusInEvent = on_focus_in

        # Styling
        combo.setStyleSheet(f"""
            QComboBox {{
                padding: 8px 22px 8px 12px;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                background: white;
                color: #1e293b;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
            }}
            QComboBox:focus {{ border-color: #7C3AED; background: #ffffff; }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 18px;
                border: none;
                background: transparent;
            }}
            QComboBox::drop-down:hover {{
                background: transparent;
                border: none;
            }}
            QComboBox::down-arrow {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                margin-right: 2px;
            }}
            QComboBox QAbstractItemView {{
                background: white;
                color: #1e293b;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                selection-background-color: #F3EEFF;
                selection-color: #0f172a;
                padding: 8px;
            }}
        """)
        return combo

    def add_field(self, layout, label_text, widget):
        """Add field with label"""
        if isinstance(layout, QtWidgets.QGridLayout):
            index = getattr(layout, "_next_field_index", 0)
            row = index // 2
            col = index % 2
            setattr(layout, "_next_field_index", index + 1)

            field_frame = QtWidgets.QFrame()
            field_frame.setStyleSheet("""
                QFrame {
                    background: transparent;
                    border: none;
                }
            """)
            field_layout = QtWidgets.QVBoxLayout(field_frame)
            field_layout.setContentsMargins(0, 0, 0, 0)
            field_layout.setSpacing(5)

            label = QtWidgets.QLabel(label_text)
            label.setStyleSheet("""
                QLabel {
                    font-weight: 800;
                    color: #334155;
                    font-size: 12px;
                    font-family: 'Inter', 'Segoe UI';
                    background: transparent;
                    border: none;
                }
            """)
            widget.setMinimumHeight(38)
            field_layout.addWidget(label)
            field_layout.addWidget(widget)
            layout.addWidget(field_frame, row, col)
            layout.setColumnStretch(col, 1)
            return field_frame

        field_layout = QtWidgets.QHBoxLayout()
        label = QtWidgets.QLabel(label_text)
        label.setStyleSheet("font-weight: 800; color: #334155; min-width: 150px; font-family:'Inter','Segoe UI';")
        field_layout.addWidget(label)
        field_layout.addWidget(widget, 1)
        layout.addLayout(field_layout)
        return field_layout

    def validate_amount_input(self):
        """Validate cost input to accept only numbers and automatically add $ prefix"""
        self.project_amount_edit.blockSignals(True)
        
        try:
            text = self.project_amount_edit.text().strip()
            cursor_pos = self.project_amount_edit.cursorPosition()
            
            if not text:
                self.project_amount_edit.blockSignals(False)
                return
            
            # Remove any non-numeric characters except decimal point
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            cleaned = cleaned.replace('$', '')
            
            # Ensure only one decimal point
            if cleaned.count('.') > 1:
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])
            
            if cleaned:
                if not text.startswith('$'):
                    final_text = f"${cleaned}"
                else:
                    final_text = f"${cleaned}"
            else:
                final_text = ""
            
            if final_text != text:
                self.project_amount_edit.setText(final_text)
                
                if not text.startswith('$') and final_text.startswith('$'):
                    new_pos = min(cursor_pos + 1, len(final_text))
                else:
                    new_pos = min(cursor_pos, len(final_text))
                
                self.project_amount_edit.setCursorPosition(new_pos)
                
        finally:
            self.project_amount_edit.blockSignals(False)

    def _auto_calc_due_date(self):
        """Set due date = start date + N working days (Mon–Fri), skip Sat/Sun."""
        from datetime import timedelta
        start = self.start_date_edit.date().toPyDate()
        target = self.duration_spin.value()
        counted = 0
        current = start
        while counted < target:
            current += timedelta(days=1)
            if current.weekday() < 5:  # Mon–Fri only
                counted += 1
        self.due_date_edit.blockSignals(True)
        self.due_date_edit.setDate(QtCore.QDate(current.year, current.month, current.day))
        self.due_date_edit.blockSignals(False)
        self.update_timeline_info()

    def _on_due_date_manually_changed(self):
        """When due date changes manually, recalculate duration (working days) and update timeline."""
        from datetime import timedelta
        start = self.start_date_edit.date().toPyDate()
        due = self.due_date_edit.date().toPyDate()
        if due <= start:
            self.update_timeline_info()
            return
        working_days = sum(
            1 for i in range((due - start).days)
            if (start + timedelta(days=i + 1)).weekday() < 5
        )
        self.duration_spin.blockSignals(True)
        self.duration_spin.setValue(max(1, working_days))
        self.duration_spin.blockSignals(False)
        self.update_timeline_info()

    def update_timeline_info(self):
        """Update timeline info — weekday count only (Mon–Fri)."""
        from datetime import timedelta
        start = self.start_date_edit.date().toPyDate()
        due = self.due_date_edit.date().toPyDate()
        total_days = (due - start).days
        # Count Mon–Fri days only
        weekdays = sum(
            1 for i in range(total_days)
            if (start + timedelta(days=i)).weekday() < 5
        )
        color = "#dc2626" if total_days < 0 else "#53657d"
        self.timeline_info.setStyleSheet(
            f"font-size: 12px; color: {color}; background: #eef6f5; "
            "padding: 8px 10px; border-radius: 7px; margin-top: 2px; font-weight:700;"
        )
        self.timeline_info.setText(
            f"Start: {start.strftime('%b %d, %Y')}  |  Due: {due.strftime('%b %d, %Y')}  "
            f"|  {weekdays} business days  ({total_days} calendar days)"
        )

    def load_saved_companies(self):
        """Load saved companies from Firebase"""
        try:
            self.company_combo.clear()

            if self.FIREBASE_AVAILABLE:
                from main import db
                ref = db.reference('/clients')
                clients_data = ref.get()

                if clients_data:
                    clean_clients = [
                        c for c in sorted(clients_data.keys())
                        if not c.startswith("-- Select")
                    ]
                    self.company_combo.addItems(clean_clients)

            # Set placeholder
            self.company_combo.setCurrentIndex(-1)
            self.company_combo.lineEdit().setPlaceholderText("Enter or select company name")
            self.company_combo.lineEdit().clear()

        except Exception as e:
            _log.error("Error loading companies: %s", e)

    def auto_generate_project_number(self):
        """Auto-generate project number as COMPANY-YYYYMM### using the current date."""
        if self.is_editing:
            return
        
        current_date = datetime.now()
        year = str(current_date.year)
        month = f"{current_date.month:02d}"
        year_month = f"{year}{month}"
        prefix = self.get_project_number_prefix()
        
        sequence = self.get_next_sequence(year_month, prefix)
        project_number = f"{prefix}-{year_month}{sequence:03d}"
        
        self.project_number_edit.setText(project_number)
        self.sequence_edit.setText(str(sequence))
        
        _log.info("Generated project number based on current date: %s", project_number)

    def get_project_number_prefix(self):
        """Return the company prefix used for project numbers."""
        try:
            company_name = ""
            if Config and getattr(Config, "COMPANY", None):
                company_name = str(Config.COMPANY.get("name", ""))

            words = [word.strip(".,/&-") for word in company_name.split() if word.strip(".,/&-")]
            if words:
                first_word = words[0]
                if first_word:
                    return first_word.upper()
        except Exception as exc:
            _log.warning("Could not determine project number prefix: %s", exc)
        return "MABS"

    def get_next_sequence(self, year_month, prefix=None):
        """Get next sequence number for the given YYYYMM, including legacy formats."""
        try:
            parent = self.parent()
            if not hasattr(parent, 'generated_projects') or not parent.generated_projects:
                return 1
            
            max_sequence = 0
            import re
            prefix = (prefix or self.get_project_number_prefix()).upper()
            year = year_month[:4]
            month = year_month[4:6]
            
            for project in parent.generated_projects:
                project_number = str(project.get('project_number', '') or '').strip().upper()
                if not project_number:
                    continue

                base_part = project_number.split('_')[0]
                short_year = year[-2:]
                patterns = [
                    rf"^{re.escape(prefix)}-{year_month}(\d{{3,}})$",
                    rf"^{year_month}(\d{{3,}})$",
                    rf"^{short_year}{month}(\d{{3,}})$" if year.startswith("20") else r"$^",
                    rf"^{re.escape(prefix)}-{year}-{month}-(?:\d{{2}}-)?(\d{{3,}})$",
                    rf"^{re.escape(prefix)}-{year}-{month}-(\d{{3,}})$",
                ]

                for pattern in patterns:
                    match = re.match(pattern, base_part)
                    if not match:
                        continue
                    try:
                        sequence = int(match.group(1))
                    except (ValueError, TypeError):
                        continue
                    max_sequence = max(max_sequence, sequence)
                    break
            
            return max_sequence + 1
            
        except Exception as e:
            _log.warning("Error getting next sequence: %s", e)
            return 1

    def check_if_in_balance_sheet(self):
        """Check if this project exists in balance sheet revenue data"""
        if not self.project_data:
            return False
        
        balance_tab = None
        if hasattr(self.main_window, 'balance_sheet_tab'):
            balance_tab = self.main_window.balance_sheet_tab
        
        if not balance_tab or not hasattr(balance_tab, 'revenue_data'):
            return False
        
        for revenue in balance_tab.revenue_data:
            if (revenue.get('source') == self.project_data.get('company', '') and
                revenue.get('description') == self.project_data.get('project_name', '') and
                revenue.get('date') == self.project_data.get('start_date', '') and
                revenue.get('amount') == f"{self.project_data.get('project_amount', 0):.2f}"):
                return True
        
        return False

    def validate_project_number_format(self, project_number):
        """Validate project number format - allow any characters"""
        if not project_number or project_number.strip() == "":
            QtWidgets.QMessageBox.warning(
                self,
                "Invalid Project Number",
                "Please enter a project number."
            )
            return False
        return True

    def validate_form(self):
        """Validate form fields"""
        required_fields = [
            (self.project_number_edit, "Project Number"),
            (self.project_name_edit, "Project Name"),
            (self.company_combo.currentText(), "Client"),
        ]
        
        for field_value, field_name in required_fields:
            if isinstance(field_value, QtWidgets.QLineEdit):
                value = field_value.text().strip()
            else:
                value = str(field_value).strip()
            
            if not value or value == "Auto-generated (MABS-YYYYMM###)":
                QtWidgets.QMessageBox.warning(
                    self, "Validation Error", 
                    f"Please enter a {field_name.lower()}."
                )
                return False
        
        project_number = self.project_number_edit.text().strip()
        if not self.validate_project_number_format(project_number):
            return False

        # Validate new payment plan fields
        ptype = getattr(self, "_payment_type", "no_down")
        rtype = getattr(self, "_remaining_type", "full")
        if ptype == "custom_down":
            pct = self.custom_pct_spin.value() if hasattr(self, "custom_pct_spin") else 0
            if pct <= 0 or pct >= 100:
                QtWidgets.QMessageBox.warning(
                    self, "Payment Plan Error",
                    "Custom down payment percentage must be between 1% and 99%."
                )
                return False
        if rtype == "installments":
            inst = self.installment_count_spin.value() if hasattr(self, "installment_count_spin") else 0
            if inst < 2:
                QtWidgets.QMessageBox.warning(
                    self, "Payment Plan Error",
                    "Installments require at least 2 payments."
                )
                return False

        return True

    def create_project(self):
        """Create or update project"""
        if not self.validate_form():
            return
            
        try:
            project_data = self.collect_project_data()
            project_number = project_data['project_number'].strip()
            
            if not self.validate_project_number_format(project_number):
                return
            
            if not self.is_editing:
                parent = self.parent()
                if hasattr(parent, 'generated_projects') and parent.generated_projects:
                    for project in parent.generated_projects:
                        if project.get('project_number', '') == project_number:
                            QtWidgets.QMessageBox.warning(
                                self,
                                "Duplicate Project Number",
                                f"Project number '{project_number}' already exists!\n\n"
                                f"Please use a different project number."
                            )
                            return
            
            if self.is_editing and self.project_data and 'firebase_id' in self.project_data:
                project_data['firebase_id'] = self.project_data['firebase_id']
            
            success = False
            if self.FIREBASE_AVAILABLE:
                from main import db
                ref = db.reference('/projects')
                
                if 'firebase_id' in project_data and project_data['firebase_id']:
                    job_id = project_data['firebase_id']
                    project_data['updated_at'] = datetime.now().isoformat()
                    ref.child(job_id).update(project_data)
                    _log.info("(converted from print, see git history)")
                    success = True
                else:
                    existing_projects = ref.order_by_child('project_number').equal_to(project_data['project_number']).get()
                    
                    if existing_projects:
                        job_id = list(existing_projects.keys())[0]
                        project_data['updated_at'] = datetime.now().isoformat()
                        ref.child(job_id).update(project_data)
                        _log.info("(converted from print, see git history)")
                        success = True
                    else:
                        new_project_ref = ref.push()
                        firebase_id = new_project_ref.key
                        project_data['firebase_id'] = firebase_id
                        project_data['created_at'] = datetime.now().isoformat()
                        project_data['updated_at'] = datetime.now().isoformat()
                        new_project_ref.set(project_data)
                        _log.info("(converted from print, see git history)")
                        success = True
            
            if success:
                action = "updated" if self.is_editing else "created"
                
                QtWidgets.QMessageBox.information(
                    self,
                    "Success",
                    f"""
                    <div style='text-align:center; font-size:14px;'>
                        <b>Project {action.capitalize()} Successfully!</b><br><br>
                        <table style='margin-left:auto; margin-right:auto; text-align:center;'>
                            <tr>
                                <td><b>Project Number:</b></td>
                                <td>{project_data.get('project_number', 'N/A')}</td>
                            </tr>
                            <tr>
                                <td><b>Client:</b></td>
                                <td>{project_data.get('company') or 'N/A'}</td>
                            </tr>
                        </table>
                    </div>
                    """
                )
                
                parent = self.parent()
                if hasattr(parent, 'load_projects'):
                    parent.load_projects()
                    if hasattr(parent, 'filter_projects'):
                        parent.filter_projects()
                
                self.accept()
                
            else:
                QtWidgets.QMessageBox.critical(
                    self, "Error", 
                    "Failed to save project. Please try again.\n\n"
                    "Check the console for detailed error messages."
                )
                    
        except Exception as e:
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(
                self, "Error", 
                f"An error occurred while {'updating' if self.is_editing else 'creating'} the project:\n{str(e)}"
            )
