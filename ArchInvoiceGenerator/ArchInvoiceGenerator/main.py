import sys
import os
import json
import subprocess
import platform
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import traceback

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import pyqtSignal
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Image
from reportlab.lib import colors
from PIL import Image as PILImage
from reportlab.platypus import Spacer, PageBreak

import openpyxl
from openpyxl.styles import Font, Alignment
import shutil

# ---------- Configuration ----------
class Config:
    COMPANY = {
        "name": "MABS Engineering LLC",
        "address": "15555 Manchester Rd., PO Box 1144\nManchester, MO 63011",
        "email": "admin@habbengineering.com",
        "phone": "314-303-0004",
        "website": "www.mabs-engineeringg.com"
    }
    
    DATA_DIR = Path("data")
    INVOICES_DIR = DATA_DIR / "invoices"
    CLIENTS_FILE = DATA_DIR / "clients.json"
    COUNTER_FILE = DATA_DIR / "invoice_counter.json"
    LOGO_FILE = DATA_DIR / "logo.png"
    SETTINGS_FILE = DATA_DIR / "settings.json"
    
    DEFAULT_TERMS = """Thank you for your business!
Best regards,

MABS Engineering LLC"""

    @classmethod
    def setup_directories(cls):
        """Create necessary directories"""
        cls.INVOICES_DIR.mkdir(parents=True, exist_ok=True)
        cls.DATA_DIR.mkdir(exist_ok=True)

# ---------- Utility Classes ----------
class DecimalEncoder(json.JSONEncoder):
    """Custom JSON encoder for Decimal objects"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

class FileManager:
    """Handles file operations with proper error handling"""
    
    @staticmethod
    def load_json(path: Path, default: Any = None) -> Any:
        """Load JSON file with error handling"""
        try:
            if path.exists():
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading {path}: {e}")
        return default

    @staticmethod
    def save_json(path: Path, data: Any) -> bool:
        """Save JSON file with error handling"""
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False, cls=DecimalEncoder)
            return True
        except Exception as e:
            print(f"Error saving {path}: {e}")
            return False

    @staticmethod
    def open_file(path: Path) -> bool:
        """Open file with default application"""
        try:
            if platform.system() == "Windows":
                os.startfile(str(path))
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
            return True
        except Exception as e:
            print(f"Error opening file {path}: {e}")
            return False

class Currency:
    """Currency formatting utilities"""
    
    @staticmethod
    def format(value, symbol: str = "$") -> str:
        """Format decimal as currency"""
        return f"{symbol}{Currency.quantize(value)}"
    
    @staticmethod
    def quantize(value) -> Decimal:
        """Quantize decimal to 2 decimal places"""
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

class InvoiceNumberGenerator:
    """Handles invoice numbering system"""
    
    @staticmethod
    def get_next_number(date_str: str) -> str:
        """Generate next invoice number for given date"""
        yyyymm = date_str[:7].replace("-", "")
        counters = FileManager.load_json(Config.COUNTER_FILE, {})
        next_num = counters.get(yyyymm, 1)
        counters[yyyymm] = next_num + 1
        FileManager.save_json(Config.COUNTER_FILE, counters)
        return f"INV-{yyyymm}-{next_num:03d}"
    
    @staticmethod
    def get_preview_number(date_str: str) -> str:
        """Preview next invoice number without persisting"""
        yyyymm = date_str[:7].replace("-", "")
        counters = FileManager.load_json(Config.COUNTER_FILE, {})
        preview_next = counters.get(yyyymm, 0)
        
        # Count existing invoices for this month
        month_dir = Config.INVOICES_DIR / date_str[:7]
        if month_dir.exists():
            count = len(list(month_dir.glob("*.json")))
            preview_next = max(preview_next, count)
        
        return f"INV-{yyyymm}-{(preview_next + 1):03d}"

class PDFGenerator:
    """Generates simple, clean PDF invoices using ReportLab's SimpleDocTemplate and Flowables"""
    
    @staticmethod
    def generate(invoice: "Invoice", output_path: Path, logo_path: Optional[Path] = None):
        """Generate PDF invoice"""
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4, 
                                  topMargin=5*mm, bottomMargin=5*mm,
                                  leftMargin=5*mm, rightMargin=5*mm)
            story = []
            styles = getSampleStyleSheet()

            # Custom styles
            styles.add(ParagraphStyle(name='CenteredBold20', alignment=1, fontName='Helvetica-Bold', fontSize=20, leading=24))
            styles.add(ParagraphStyle(name='Centered10', alignment=1, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='LeftBold16', alignment=0, fontName='Helvetica-Bold', fontSize=16, leading=18))
            styles.add(ParagraphStyle(name='LeftBold12', alignment=0, fontName='Helvetica-Bold', fontSize=12, leading=14))
            styles.add(ParagraphStyle(name='Left10', alignment=0, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='LeftBold10', alignment=0, fontName='Helvetica-Bold', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Right10', alignment=2, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='RightBold10', alignment=2, fontName='Helvetica-Bold', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Left9', alignment=0, fontName='Helvetica', fontSize=9, leading=11))
            styles.add(ParagraphStyle(name='Left8', alignment=0, fontName='Helvetica', fontSize=8, leading=10))
            styles.add(ParagraphStyle(name='CenterBold12', alignment=1, fontName='Helvetica-Bold', fontSize=12, leading=14))
            styles.add(ParagraphStyle(name='CenterBold10', alignment=1, fontName='Helvetica-Bold', fontSize=10, leading=12))

            # --- Company Header ---
            story.append(Paragraph(Config.COMPANY["name"], styles['CenteredBold20']))
            for line in Config.COMPANY["address"].split('\n'):
                story.append(Paragraph(line, styles['Centered10']))
            story.append(Paragraph(f"Phone: {Config.COMPANY['phone']} • Email: {Config.COMPANY['email']}", styles['Centered10']))
            story.append(Paragraph(Config.COMPANY["website"], styles['Centered10']))
            story.append(Spacer(1, 2*mm))

            # Separator line (using a Table for a simple line)
            story.append(Table([['']], colWidths=[doc.width], style=[
                ('LINEBELOW', (0,0), (-1,-1), 1, colors.black),
            ]))
            story.append(Spacer(1, 2*mm))

            # --- Invoice Title and Details ---
            story.append(Paragraph("INVOICE", styles['LeftBold16']))
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(f"Invoice Number: {invoice.invoice_number}", styles['Left10']))
            story.append(Paragraph(f"Date: {invoice.date}", styles['Left10']))
            story.append(Paragraph(f"Due Date: {invoice.due_date}", styles['Left10']))
            story.append(Spacer(1, 2*mm))

            # --- Client Information ---
            story.append(Paragraph("Bill To:", styles['LeftBold12']))
            story.append(Paragraph(invoice.client_name, styles['Left10']))
            for line in invoice.client_address.split('\n'):
                if line.strip():
                    story.append(Paragraph(line.strip(), styles['Left9']))
            if invoice.client_email:
                story.append(Paragraph(invoice.client_email, styles['Left9']))
            story.append(Spacer(1, 2*mm))

            # --- Line Items ---
            story.append(Paragraph("Items:", styles['LeftBold12']))
            story.append(Spacer(1, 2*mm))

            # Updated headers to match the image
            item_data = [[
                Paragraph("Project Number", styles['LeftBold10']),
                Paragraph("Description (PO & Address)", styles['LeftBold10']), 
                Paragraph("Plant", styles['LeftBold10']),
                Paragraph("Quantity", styles['LeftBold10']),
                Paragraph("Unit Price", styles['LeftBold10']),
                Paragraph("Down Payment", styles['LeftBold10']),
                Paragraph("Payment Due", styles['LeftBold10'])
            ]]
            
            for item in invoice.items:
                item_data.append([
                    Paragraph(item.project_number or "", styles['Left10']),
                    Paragraph(item.description, styles['Left10']),
                    Paragraph(item.plant or "", styles['Left10']),
                    Paragraph(str(item.quantity), styles['Right10']),
                    Paragraph(Currency.format(item.unit_price), styles['Right10']),
                    Paragraph(Currency.format(item.down_payment), styles['Right10']),
                    Paragraph(Currency.format(item.payment_due), styles['Right10'])
                ])
            
            # Calculate column widths based on available space
            available_width = doc.width
            item_table = Table(item_data, colWidths=[
                available_width * 0.12,  # Project Number
                available_width * 0.22,  # Description
                available_width * 0.07,  # Plant
                available_width * 0.08,  # Quantity
                available_width * 0.11,  # Unit Price
                available_width * 0.13,  # Down Payment
                available_width * 0.13   # Payment Due
            ])
            item_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
                ('ALIGN', (0,0), (2,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 1),
                ('RIGHTPADDING', (0,0), (-1,-1), 1),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                ('TOPPADDING', (0,0), (-1,-1), 1),
                ('FONTSIZE', (0,0), (-1,-1), 7),
            ]))
            story.append(item_table)
            story.append(Spacer(1, 2*mm))

            # --- Totals ---
            total_data = []
            total_data.append([Paragraph("Subtotal:", styles['RightBold10']), Paragraph(Currency.format(invoice.subtotal), styles['RightBold10'])])
            if invoice.tax_rate > 0:
                total_data.append([Paragraph(f"Tax ({invoice.tax_rate}%):", styles['Right10']), Paragraph(Currency.format(invoice.tax_amount), styles['Right10'])])
            
            total_data.append([Paragraph("Total Amount Due:", styles['RightBold10']), Paragraph(Currency.format(invoice.total), styles['RightBold10'])])

            total_table = Table(total_data, colWidths=[available_width * 0.7, available_width * 0.3])
            total_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 1),
                ('RIGHTPADDING', (0,0), (-1,-1), 1),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1),
                ('TOPPADDING', (0,0), (-1,-1), 1),
            ]))
            story.append(total_table)
            story.append(Spacer(1, 3*mm))

            # --- Thank You Message ---
            story.append(Paragraph("Thank you for your business!", styles['CenterBold12']))
            story.append(Paragraph("Best regards,", styles['CenterBold10']))
            story.append(Paragraph(Config.COMPANY["name"], styles['CenterBold12']))
            story.append(Spacer(1, 2*mm))

            # --- Payment Options Section ---
            story.append(Paragraph("PAYMENT OPTIONS", styles['CenterBold12']))
            story.append(Spacer(1, 2*mm))
            
            # Modified payment options table: Option 1 & 3 on left, Option 2 on right with image
            payment_data = [
                [
                    Paragraph("<b>Option 1: Check</b>", styles['Left10']),
                    Paragraph("<b>Option 2: Venmo</b>", styles['Left10'])
                ],
                [
                    Paragraph("Payable to: " + Config.COMPANY["name"] + "<br/>" + "Mailing Address:<br/>" + Config.COMPANY["address"].replace('\n', '<br/>'), styles['Left9']),
                    [Paragraph("Please scan the QR code to pay via Venmo.", styles['Left9']),
                     Image("C:\\Users\\potin\\Documents\\ArchInvoiceGenerator\\ArchInvoiceGenerator\\assets\\payment.jpeg", width=25*mm, height=25*mm)]
                ],
                [
                    Paragraph("<b>Option 3: Bank ACH Transfer</b>", styles['Left10']),
                    Paragraph("", styles['Left9'])
                ],
                [
                    Paragraph("Please contact " + Config.COMPANY["name"].split()[0] + " Admin to get<br/>our bank information for ACH transfers.", styles['Left9']),
                    Paragraph("", styles['Left9'])
                ]
            ]
            
            payment_table = Table(payment_data, colWidths=[available_width * 0.5, available_width * 0.5])
            payment_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING', (0,0), (-1,-1), 2),
                ('RIGHTPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BACKGROUND', (0,0), (0,0), colors.lightblue),  # Option 1 title
                ('BACKGROUND', (1,0), (1,0), colors.lightgreen),  # Option 2 title
                ('BACKGROUND', (0,2), (0,2), colors.lightcoral),  # Option 3 title
                ('BACKGROUND', (0,1), (0,1), colors.whitesmoke),
                ('BACKGROUND', (1,1), (1,1), colors.whitesmoke),
                ('BACKGROUND', (0,3), (0,3), colors.whitesmoke),
                ('BACKGROUND', (1,3), (1,3), colors.whitesmoke),
                ('BOX', (0,0), (-1,-1), 0.5, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
            ]))
            story.append(payment_table)
            story.append(Spacer(1, 2*mm))

            # --- Footer (Notes/Terms) ---
            if invoice.notes and invoice.notes.strip() and invoice.notes != Config.DEFAULT_TERMS:
                story.append(Paragraph("Notes/Terms:", styles['LeftBold12']))
                for line in invoice.notes.split('\n'):
                    story.append(Paragraph(line, styles['Left9']))
                story.append(Spacer(1, 2*mm))

            doc.build(story)
            return True
        except Exception as e:
            print(f"PDF generation error: {e}")
            traceback.print_exc()
            return False

# ---------- Data Models ----------
class InvoiceItem:
    """Represents an invoice line item"""
    
    def __init__(self, project_number: str = "", description: str = "", plant: str = "", 
                 quantity: int = 1, unit_price: float = 0.0, down_payment: float = 0.0, 
                 payment_due: float = 0.0):
        self.project_number = project_number
        self.description = description
        self.plant = plant
        self.quantity = quantity
        self.unit_price = Currency.quantize(unit_price)
        
        # Convert all values to Decimal for proper calculation
        unit_price_decimal = Decimal(str(unit_price))
        down_payment_decimal = Decimal(str(down_payment))
        quantity_decimal = Decimal(str(quantity))
        
        # Auto-calculate 50% down payment
        total_amount = quantity_decimal * unit_price_decimal
        if down_payment_decimal > 0:
            self.down_payment = Currency.quantize(down_payment_decimal)
        else:
            self.down_payment = Currency.quantize(total_amount * Decimal('0.5'))
        
        # Auto-calculate payment due
        payment_due_decimal = Decimal(str(payment_due))
        if payment_due_decimal > 0:
            self.payment_due = Currency.quantize(payment_due_decimal)
        else:
            self.payment_due = Currency.quantize(total_amount - self.down_payment)
    
    @property
    def total(self) -> Decimal:
        return Currency.quantize(self.quantity * self.unit_price)
    
    def to_dict(self) -> Dict:
        return {
            "project_number": self.project_number,
            "description": self.description,
            "plant": self.plant,
            "quantity": self.quantity,
            "unit_price": float(self.unit_price),
            "down_payment": float(self.down_payment),
            "payment_due": float(self.payment_due),
            "total": float(self.total)
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'InvoiceItem':
        return cls(
            project_number=data.get("project_number", ""),
            description=data.get("description", ""),
            plant=data.get("plant", ""),
            quantity=data.get("quantity", 1),
            unit_price=data.get("unit_price", 0.0),
            down_payment=data.get("down_payment", 0.0),
            payment_due=data.get("payment_due", 0.0)
        )

class Invoice:
    """Represents a complete invoice with downpayment support"""
    
    def __init__(self):
        self.invoice_number = ""
        self.date = datetime.now().strftime("%Y-%m-%d")
        self.due_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        self.client_name = ""
        self.client_email = ""
        self.client_address = ""
        self.items: List[InvoiceItem] = []
        self.tax_rate = Decimal("0.0")
        self.discount_rate = Decimal("0.0")
        self.notes = Config.DEFAULT_TERMS
        self.logo_path = None
    
    @property
    def subtotal(self) -> Decimal:
        return Currency.quantize(sum(item.total for item in self.items))
    
    @property
    def tax_amount(self) -> Decimal:
        return Currency.quantize(self.subtotal * self.tax_rate / Decimal("100"))
    
    @property
    def total(self) -> Decimal:
        return Currency.quantize(self.subtotal + self.tax_amount)
    
    def add_item(self, item: InvoiceItem):
        self.items.append(item)
    
    def remove_item(self, index: int):
        if 0 <= index < len(self.items):
            self.items.pop(index)
    
    def to_dict(self) -> Dict:
        return {
            "meta": {
                "invoice_number": self.invoice_number,
                "date": self.date,
                "due_date": self.due_date,
                "client_name": self.client_name,
                "client_email": self.client_email,
                "client_address": self.client_address,
                "tax_rate": float(self.tax_rate),
                "subtotal": float(self.subtotal),
                "tax_amount": float(self.tax_amount),
                "total": float(self.total),
                "notes": self.notes
            },
            "items": [item.to_dict() for item in self.items]
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'Invoice':
        invoice = cls()
        meta = data.get("meta", {})
        
        invoice.invoice_number = meta.get("invoice_number", "")
        invoice.date = meta.get("date", invoice.date)
        invoice.due_date = meta.get("due_date", invoice.due_date)
        invoice.client_name = meta.get("client_name", "")
        invoice.client_email = meta.get("client_email", "")
        invoice.client_address = meta.get("client_address", "")
        invoice.tax_rate = Decimal(str(meta.get("tax_rate", 0.0)))
        invoice.notes = meta.get("notes", Config.DEFAULT_TERMS)
        
        invoice.items = [InvoiceItem.from_dict(item_data) 
                        for item_data in data.get("items", [])]
        
        return invoice

# ---------- Excel Exporter ----------
class ExcelExporter:
    """Handles Excel export functionality"""
    
    @staticmethod
    def export_invoice(invoice: Invoice, output_path: Path):
        """Export single invoice to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header
        ws.merge_cells('A1:G1')
        ws['A1'] = "INVOICE"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Invoice info
        ws['A3'] = "Invoice No:"
        ws['B3'] = invoice.invoice_number
        ws['A4'] = "Date:"
        ws['B4'] = invoice.date
        ws['A5'] = "Due Date:"
        ws['B5'] = invoice.due_date
        
        # Client info
        ws['A7'] = "Bill To:"
        ws['A8'] = invoice.client_name
        for i, line in enumerate(invoice.client_address.splitlines()):
            ws.cell(row=9 + i, column=1, value=line)
        
        # Items table with new columns
        headers = ["Project Number", "Description (PO & Address)", "Plant", "Quantity", "Unit Price", "Down Payment", "Payment Due"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, item in enumerate(invoice.items, 13):
            ws.cell(row=row_idx, column=1, value=item.project_number)
            ws.cell(row=row_idx, column=2, value=item.description)
            ws.cell(row=row_idx, column=3, value=item.plant)
            ws.cell(row=row_idx, column=4, value=item.quantity)
            ws.cell(row=row_idx, column=5, value=float(item.unit_price))
            ws.cell(row=row_idx, column=6, value=float(item.down_payment))
            ws.cell(row=row_idx, column=7, value=float(item.payment_due))
        
        # Totals
        last_row = 13 + len(invoice.items)
        totals = [
            ("Subtotal:", invoice.subtotal),
        ]
        
        if invoice.tax_rate > 0:
            totals.append((f"Tax ({invoice.tax_rate}%):", invoice.tax_amount))
        
        totals.append(("Total Amount Due:", invoice.total))
        
        for i, (label, amount) in enumerate(totals):
            row = last_row + 2 + i
            ws.cell(row=row, column=6, value=label)
            ws.cell(row=row, column=7, value=float(amount))
            if "Total Amount Due" in label:
                ws.cell(row=row, column=6).font = Font(bold=True)
                ws.cell(row=row, column=7).font = Font(bold=True)
        
        wb.save(str(output_path))
    
    @staticmethod
    def export_history(invoices: List[Invoice], output_path: Path):
        """Export invoice history to Excel"""
        data = []
        for inv in invoices:
            data.append({
                "Invoice No": inv.invoice_number,
                "Date": inv.date,
                "Client": inv.client_name,
                "Subtotal": float(inv.subtotal),
                "Tax": float(inv.tax_amount),
                "Total": float(inv.total)
            })
        
        df = pd.DataFrame(data)
        df.to_excel(str(output_path), index=False, engine='openpyxl')

# ---------- Qt Widgets ----------
class ItemRowWidget(QtWidgets.QWidget):
    """Widget for a single invoice item row"""
    removed = pyqtSignal()
    
    def __init__(self, item: InvoiceItem = None):
        super().__init__()
        self.item = item or InvoiceItem()
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(8)
        
        # Project Number
        self.project_number_edit = QtWidgets.QLineEdit(self.item.project_number)
        self.project_number_edit.setPlaceholderText("Project #")
        self.project_number_edit.setMinimumHeight(35)
        self.project_number_edit.setMaximumWidth(120)
        
        # Description
        self.desc_edit = QtWidgets.QLineEdit(self.item.description)
        self.desc_edit.setPlaceholderText("Description (PO & Address)")
        self.desc_edit.setMinimumHeight(35)
        
        # Plant
        self.plant_edit = QtWidgets.QLineEdit(self.item.plant)
        self.plant_edit.setPlaceholderText("Plant")
        self.plant_edit.setMinimumHeight(35)
        self.plant_edit.setMaximumWidth(80)
        
        # Quantity
        self.qty_spin = QtWidgets.QSpinBox()
        self.qty_spin.setRange(1, 1000000)
        self.qty_spin.setValue(self.item.quantity)
        self.qty_spin.setMinimumHeight(35)
        self.qty_spin.setMinimumWidth(60)
        
        # Unit Price
        self.price_spin = QtWidgets.QDoubleSpinBox()
        self.price_spin.setRange(0, 1000000)
        self.price_spin.setDecimals(2)
        self.price_spin.setValue(float(self.item.unit_price))
        self.price_spin.setPrefix("$ ")
        self.price_spin.setMinimumHeight(35)
        self.price_spin.setMinimumWidth(100)
        
        # Down Payment (read-only, auto-calculated)
        self.down_payment_spin = QtWidgets.QDoubleSpinBox()
        self.down_payment_spin.setRange(0, 1000000)
        self.down_payment_spin.setDecimals(2)
        self.down_payment_spin.setValue(float(self.item.down_payment))
        self.down_payment_spin.setPrefix("$ ")
        self.down_payment_spin.setMinimumHeight(35)
        self.down_payment_spin.setMinimumWidth(100)
        self.down_payment_spin.setReadOnly(True)
        self.down_payment_spin.setStyleSheet("background-color: #f0f0f0; color: #666;")
        
        # Payment Due (read-only, auto-calculated)
        self.payment_due_spin = QtWidgets.QDoubleSpinBox()
        self.payment_due_spin.setRange(0, 1000000)
        self.payment_due_spin.setDecimals(2)
        self.payment_due_spin.setValue(float(self.item.payment_due))
        self.payment_due_spin.setPrefix("$ ")
        self.payment_due_spin.setMinimumHeight(35)
        self.payment_due_spin.setMinimumWidth(100)
        self.payment_due_spin.setReadOnly(True)
        self.payment_due_spin.setStyleSheet("background-color: #f0f0f0; color: #666;")
        
        # Total (read-only)
        self.total_label = QtWidgets.QLabel(Currency.format(self.item.total))
        self.total_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.total_label.setMinimumWidth(80)
        self.total_label.setMinimumHeight(35)
        self.total_label.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 8px;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        
        # Remove button
        self.remove_btn = QtWidgets.QPushButton("✕")
        self.remove_btn.setMaximumWidth(40)
        self.remove_btn.setMinimumHeight(35)
        self.remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.remove_btn.clicked.connect(self.removed)
        
        # Add widgets to layout
        layout.addWidget(self.project_number_edit, 1)
        layout.addWidget(self.desc_edit, 3)
        layout.addWidget(self.plant_edit, 1)
        layout.addWidget(self.qty_spin, 1)
        layout.addWidget(self.price_spin, 1)
        layout.addWidget(self.down_payment_spin, 1)
        layout.addWidget(self.payment_due_spin, 1)
        layout.addWidget(self.total_label, 1)
        layout.addWidget(self.remove_btn, 0)
        
        # Connect signals
        self.project_number_edit.textChanged.connect(self.update_item)
        self.desc_edit.textChanged.connect(self.update_item)
        self.plant_edit.textChanged.connect(self.update_item)
        self.qty_spin.valueChanged.connect(self.update_total)
        self.price_spin.valueChanged.connect(self.update_total)
    
    def update_total(self):
        self.item.quantity = self.qty_spin.value()
        self.item.unit_price = Currency.quantize(self.price_spin.value())
        
        # Auto-calculate 50% down payment
        total_amount = Decimal(str(self.item.quantity)) * self.item.unit_price
        self.item.down_payment = Currency.quantize(total_amount * Decimal('0.5'))
        self.down_payment_spin.setValue(float(self.item.down_payment))
        
        # Auto-calculate payment due
        self.item.payment_due = Currency.quantize(total_amount - self.item.down_payment)
        self.payment_due_spin.setValue(float(self.item.payment_due))
        
        self.total_label.setText(Currency.format(self.item.total))
        self.update_item()
    
    def update_item(self):
        self.item.project_number = self.project_number_edit.text()
        self.item.description = self.desc_edit.text()
        self.item.plant = self.plant_edit.text()
    
    def get_item(self) -> InvoiceItem:
        return self.item

class ScrollArea(QtWidgets.QScrollArea):
    """Custom scroll area with better styling"""
    def __init__(self):
        super().__init__()
        self.setWidgetResizable(True)
        self.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.setStyleSheet("""
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
            }
            QScrollBar:vertical {
                border: none;
                background: #f8f9fa;
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #ced4da;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #adb5bd;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)

class MainWindow(QtWidgets.QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.invoice = Invoice()
        self.clients = {}
        self.settings = {}
        self.item_rows = []
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        self.setWindowTitle("HABB Invoice Generator - Professional Billing System")
        self.setMinimumSize(1400, 700)  # Increased minimum width to accommodate new columns
        
        # Set window icon and style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 1em;
                padding-top: 10px;
                background-color: white;
                font-size: 11px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
                font-size: 12px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: bold;
                min-height: 20px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QPushButton.success {
                background-color: #28a745;
            }
            QPushButton.success:hover {
                background-color: #218838;
            }
            QPushButton.warning {
                background-color: #ffc107;
                color: #212529;
            }
            QPushButton.warning:hover {
                background-color: #e0a800;
            }
            QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox {
                padding: 6px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
                min-height: 20px;
            }
            QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border-color: #3498db;
            }
            QLabel {
                color: #2c3e50;
                font-size: 11px;
            }
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e9ecef;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-size: 11px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #3498db;
            }
            QFormLayout > QLabel, QFormLayout > QLineEdit, QFormLayout > QDoubleSpinBox {
                margin: 2px;
                padding: 4px;
            }
            QGroupBox {
                margin: 5px;
            }
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
                padding: 2px;
            }
            QListWidget {
                border: 1px solid #dee2e6;
                background-color: white;
                font-size: 11px;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QComboBox {
                padding: 6px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
                min-height: 20px;
                color: #2c3e50;
            }
            QComboBox:focus {
                border-color: #3498db;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #2c3e50;
                border: 1px solid #ced4da;
                selection-background-color: #3498db;
                selection-color: white;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #e9ecef;
                color: #2c3e50;
            }
        """)
        
        # Create central widget with scroll area for main content
        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QtWidgets.QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)
        
        # Create tabs
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setDocumentMode(True)
        self.create_tab = QtWidgets.QWidget()
        self.history_tab = QtWidgets.QWidget()
        
        self.tabs.addTab(self.create_tab, "📄 Create Invoice")
        self.tabs.addTab(self.history_tab, "📊 Invoice History")
        
        main_layout.addWidget(self.tabs)
        
        self.setup_create_tab()
        self.setup_history_tab()
        
        # Connect signals
        self.tax_spin.valueChanged.connect(self.update_totals)
        self.date_edit.dateChanged.connect(self.update_invoice_preview)
        
        self.update_totals()
        self.update_invoice_preview()
    
    def setup_create_tab(self):
        # Create scroll area for the create tab
        scroll_area = ScrollArea()
        scroll_content = QtWidgets.QWidget()
        scroll_area.setWidget(scroll_content)
        
        # Main layout for scroll content
        main_layout = QtWidgets.QVBoxLayout(scroll_content)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Company header with better styling
        company_frame = QtWidgets.QFrame()
        company_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                border-radius: 8px;
                padding: 12px;
            }
        """)
        company_layout = QtWidgets.QHBoxLayout(company_frame)
        
        company_info = QtWidgets.QLabel(
            f"<h2 style='color: white; margin: 0; font-size: 16px;'>{Config.COMPANY['name']}</h2>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['address'].replace(chr(10), '<br>')}</p>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['email']} | {Config.COMPANY['phone']}</p>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['website']}</p>"
        )
        company_info.setTextFormat(QtCore.Qt.RichText)
        company_layout.addWidget(company_info)
        company_layout.addStretch()
        
        main_layout.addWidget(company_frame)
        
        # Create a grid layout for the main content
        content_layout = QtWidgets.QGridLayout()
        content_layout.setSpacing(12)
        content_layout.setColumnStretch(0, 1)
        content_layout.setColumnStretch(1, 1)
        
        # Invoice Details - Column 0
        details_group = QtWidgets.QGroupBox("Invoice Details")
        details_layout = QtWidgets.QFormLayout(details_group)
        details_layout.setVerticalSpacing(8)
        details_layout.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)
        
        self.invoice_no_edit = QtWidgets.QLineEdit()
        self.invoice_no_edit.setReadOnly(True)
        self.invoice_no_edit.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 11px;")
        
        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("MMM d, yyyy")
        
        self.due_date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate().addDays(30))
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("MMM d, yyyy")
        
        details_layout.addRow("Invoice Number:", self.invoice_no_edit)
        details_layout.addRow("Invoice Date:", self.date_edit)
        details_layout.addRow("Due Date:", self.due_date_edit)
        
        content_layout.addWidget(details_group, 0, 0)
        
        # Client Information - Column 1
        client_group = QtWidgets.QGroupBox("Client Information")
        client_layout = QtWidgets.QVBoxLayout(client_group)
        
        # Client form
        client_form = QtWidgets.QFormLayout()
        client_form.setVerticalSpacing(8)
        client_form.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)
        
        self.client_name_edit = QtWidgets.QLineEdit()
        self.client_name_edit.setPlaceholderText("Client or Company Name")
        
        self.client_email_edit = QtWidgets.QLineEdit()
        self.client_email_edit.setPlaceholderText("email@example.com")
        
        self.client_address_edit = QtWidgets.QTextEdit()
        self.client_address_edit.setMaximumHeight(70)
        self.client_address_edit.setPlaceholderText("Street Address\nCity, State ZIP\nCountry")
        
        client_form.addRow("Name:", self.client_name_edit)
        client_form.addRow("Email:", self.client_email_edit)
        client_form.addRow("Address:", self.client_address_edit)
        
        client_layout.addLayout(client_form)
        
        # Client management buttons
        client_buttons_layout = QtWidgets.QHBoxLayout()
        
        self.client_combo = QtWidgets.QComboBox()
        self.client_combo.addItem("-- Select Saved Client --")
        
        load_client_btn = QtWidgets.QPushButton("📂 Load Client")
        save_client_btn = QtWidgets.QPushButton("💾 Save Client")
        
        load_client_btn.clicked.connect(self.load_client)
        save_client_btn.clicked.connect(self.save_client)
        
        client_buttons_layout.addWidget(QtWidgets.QLabel("Saved Clients:"))
        client_buttons_layout.addWidget(self.client_combo, 2)
        client_buttons_layout.addWidget(load_client_btn)
        client_buttons_layout.addWidget(save_client_btn)
        
        client_layout.addLayout(client_buttons_layout)
        
        content_layout.addWidget(client_group, 0, 1)
        
        # Items Section - Span both columns
        items_group = QtWidgets.QGroupBox("Line Items")
        items_layout = QtWidgets.QVBoxLayout(items_group)
        
        # Items header - Updated to match the image
        header_frame = QtWidgets.QFrame()
        header_layout = QtWidgets.QHBoxLayout(header_frame)
        header_layout.setContentsMargins(8, 4, 8, 4)
        
        headers = ["Project Number", "Description (PO & Address)", "Plant", "Quantity", "Unit Price", "Down Payment", "Payment Due", "Total", ""]
        widths = [1, 3, 1, 1, 1, 1, 1, 1, 0]  # Adjusted widths for new columns
        
        for header, width in zip(headers, widths):
            label = QtWidgets.QLabel(header)
            label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 11px;")
            header_layout.addWidget(label, width)
        
        items_layout.addWidget(header_frame)
        
        # Items container with scroll area
        items_scroll = ScrollArea()
        items_scroll.setMinimumHeight(180)
        
        self.items_widget = QtWidgets.QWidget()
        self.items_layout = QtWidgets.QVBoxLayout(self.items_widget)
        self.items_layout.setSpacing(4)
        items_scroll.setWidget(self.items_widget)
        
        items_layout.addWidget(items_scroll)
        
        # Add item button
        add_item_btn = QtWidgets.QPushButton("➕ Add New Item")
        add_item_btn.setStyleSheet("QPushButton { font-size: 11px; }")
        add_item_btn.clicked.connect(self.add_item_row)
        items_layout.addWidget(add_item_btn)
        
        content_layout.addWidget(items_group, 1, 0, 1, 2)
        
        # Totals and Actions - Bottom section
        bottom_layout = QtWidgets.QHBoxLayout()
        bottom_layout.setSpacing(12)
        
        # Left - Settings
        settings_group = QtWidgets.QGroupBox("Settings & Logo")
        settings_layout = QtWidgets.QVBoxLayout(settings_group)
        
        # Logo section
        logo_layout = QtWidgets.QHBoxLayout()
        self.logo_btn = QtWidgets.QPushButton("🖼️ Set Company Logo")
        self.logo_label = QtWidgets.QLabel("No logo set")
        self.logo_label.setStyleSheet("color: #6c757d; font-style: italic; font-size: 10px;")
        logo_layout.addWidget(self.logo_btn)
        logo_layout.addWidget(self.logo_label)
        logo_layout.addStretch()
        
        self.logo_btn.clicked.connect(self.set_logo)
        
        # Tax settings
        calc_layout = QtWidgets.QFormLayout()
        calc_layout.setVerticalSpacing(6)
        
        self.tax_spin = QtWidgets.QDoubleSpinBox()
        self.tax_spin.setRange(0, 100)
        self.tax_spin.setDecimals(2)
        self.tax_spin.setSuffix(" %")
        self.tax_spin.setValue(0.0)
        
        calc_layout.addRow("Tax Rate:", self.tax_spin)
        
        settings_layout.addLayout(logo_layout)
        settings_layout.addLayout(calc_layout)
        
        bottom_layout.addWidget(settings_group, 1)
        
        # Center - Totals
        totals_group = QtWidgets.QGroupBox("Invoice Totals")
        totals_layout = QtWidgets.QFormLayout(totals_group)
        totals_layout.setVerticalSpacing(8)
        
        self.subtotal_label = QtWidgets.QLabel("$0.00")
        self.tax_label = QtWidgets.QLabel("$0.00")
        
        self.total_label = QtWidgets.QLabel("$0.00")
        self.total_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                background-color: #e8f4fd;
                border: 2px solid #3498db;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        
        totals_layout.addRow("Subtotal:", self.subtotal_label)
        totals_layout.addRow("Tax Amount:", self.tax_label)
        totals_layout.addRow("Total Amount Due:", self.total_label)
        
        bottom_layout.addWidget(totals_group, 1)
        
        # Right - Actions
        actions_group = QtWidgets.QGroupBox("Actions")
        actions_layout = QtWidgets.QVBoxLayout(actions_group)
        
        self.generate_pdf_btn = QtWidgets.QPushButton("📄 Generate PDF Invoice")
        self.export_excel_btn = QtWidgets.QPushButton("📊 Export to Excel")
        self.save_json_btn = QtWidgets.QPushButton("💾 Save as JSON")
        
        # Style action buttons
        self.generate_pdf_btn.setStyleSheet("QPushButton { font-size: 11px; min-height: 35px; }")
        self.generate_pdf_btn.setProperty("class", "success")
        self.export_excel_btn.setStyleSheet("font-size: 11px; min-height: 35px;")
        self.save_json_btn.setStyleSheet("font-size: 11px; min-height: 35px;")
        
        self.generate_pdf_btn.clicked.connect(self.generate_pdf)
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.save_json_btn.clicked.connect(self.save_json)
        
        actions_layout.addWidget(self.generate_pdf_btn)
        actions_layout.addWidget(self.export_excel_btn)
        actions_layout.addWidget(self.save_json_btn)
        actions_layout.addStretch()
        
        bottom_layout.addWidget(actions_group, 1)
        
        content_layout.addLayout(bottom_layout, 2, 0, 1, 2)
        
        # Notes section
        notes_group = QtWidgets.QGroupBox("Notes & Terms")
        notes_layout = QtWidgets.QVBoxLayout(notes_group)
        self.notes_edit = QtWidgets.QTextEdit(Config.DEFAULT_TERMS)
        self.notes_edit.setMaximumHeight(80)
        notes_layout.addWidget(self.notes_edit)
        
        content_layout.addWidget(notes_group, 3, 0, 1, 2)
        
        main_layout.addLayout(content_layout)
        
        # Set the scroll area as the main widget for create tab
        self.create_tab_layout = QtWidgets.QVBoxLayout(self.create_tab)
        self.create_tab_layout.setContentsMargins(0, 0, 0, 0)
        self.create_tab_layout.addWidget(scroll_area)
        
        # Add initial item row
        self.add_item_row()

    def setup_history_tab(self):
        # Create splitter for clients and invoices
        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        
        # Left: Clients list
        clients_group = QtWidgets.QGroupBox("Clients")
        clients_layout = QtWidgets.QVBoxLayout(clients_group)
        self.clients_list = QtWidgets.QListWidget()
        self.clients_list.setMinimumWidth(200)
        clients_layout.addWidget(self.clients_list)
        
        # Right: Invoices table
        invoices_group = QtWidgets.QGroupBox("Invoices")
        invoices_layout = QtWidgets.QVBoxLayout(invoices_group)
        
        # Controls
        controls_frame = QtWidgets.QFrame()
        controls_layout = QtWidgets.QHBoxLayout(controls_frame)
        
        self.refresh_btn = QtWidgets.QPushButton("🔄 Refresh")
        self.open_pdf_btn = QtWidgets.QPushButton("📂 Open PDF")
        self.export_history_btn = QtWidgets.QPushButton("📈 Export History to Excel")
        
        controls_layout.addWidget(self.refresh_btn)
        controls_layout.addWidget(self.open_pdf_btn)
        controls_layout.addWidget(self.export_history_btn)
        controls_layout.addStretch()
        
        invoices_layout.addWidget(controls_frame)
        
        # Table
        self.history_table = QtWidgets.QTableWidget()
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels([
            "Invoice No", "Date", "Client", "Subtotal", "Tax", "Total Amount Due"
        ])
        self.history_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.history_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        
        # Set column widths
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        
        invoices_layout.addWidget(self.history_table)
        
        # Add to splitter
        splitter.addWidget(clients_group)
        splitter.addWidget(invoices_group)
        splitter.setStretchFactor(1, 3)
        
        # Main layout for history tab
        history_layout = QtWidgets.QVBoxLayout(self.history_tab)
        history_layout.setContentsMargins(10, 10, 10, 10)
        history_layout.addWidget(splitter)
        
        # Connect signals
        self.refresh_btn.clicked.connect(self.load_history)
        self.open_pdf_btn.clicked.connect(self.open_selected_pdf)
        self.export_history_btn.clicked.connect(self.export_history)
        self.clients_list.itemSelectionChanged.connect(self.load_invoices_for_client)
    
    def load_data(self):
        """Load clients and settings"""
        self.clients = FileManager.load_json(Config.CLIENTS_FILE, {})
        self.settings = FileManager.load_json(Config.SETTINGS_FILE, {})
        
        # Update client combo
        self.client_combo.clear()
        self.client_combo.addItem("-- Select Saved Client --")
        for client_name in sorted(self.clients.keys()):
            self.client_combo.addItem(client_name)
        
        # Load logo setting
        if Config.LOGO_FILE.exists():
            self.logo_label.setText(f"✓ {Config.LOGO_FILE.name}")
    
    def save_data(self):
        """Save clients and settings"""
        FileManager.save_json(Config.CLIENTS_FILE, self.clients)
        FileManager.save_json(Config.SETTINGS_FILE, self.settings)
    
    def add_item_row(self, item: InvoiceItem = None):
        """Add a new item row to the form"""
        row = ItemRowWidget(item)
        row.removed.connect(lambda: self.remove_item_row(row))
        self.items_layout.addWidget(row)
        self.item_rows.append(row)
        
        # Connect signals for auto-update
        row.qty_spin.valueChanged.connect(self.update_totals)
        row.price_spin.valueChanged.connect(self.update_totals)
        
        self.update_totals()
    
    def remove_item_row(self, row):
        """Remove an item row from the form"""
        if row in self.item_rows:
            self.item_rows.remove(row)
            row.setParent(None)
            self.update_totals()
    
    def update_totals(self):
        """Update the totals display"""
        # Update invoice object from form
        self.update_invoice_from_form()
        
        # Update UI labels
        if self.subtotal_label:
            self.subtotal_label.setText(Currency.format(self.invoice.subtotal))
        if self.tax_label:
            self.tax_label.setText(Currency.format(self.invoice.tax_amount))
        
        # Update total label
        if self.total_label:
            total_text = f"TOTAL: {Currency.format(self.invoice.total)}"
            self.total_label.setText(total_text)
    
    def update_invoice_from_form(self):
        """Update the invoice object from form data"""
        # Basic info
        self.invoice.date = self.date_edit.date().toString("yyyy-MM-dd")
        self.invoice.due_date = self.due_date_edit.date().toString("yyyy-MM-dd")
        
        # Client info
        self.invoice.client_name = self.client_name_edit.text()
        self.invoice.client_email = self.client_email_edit.text()
        self.invoice.client_address = self.client_address_edit.toPlainText()
        
        # Items
        self.invoice.items = []
        for row in self.item_rows:
            self.invoice.items.append(row.get_item())
        
        # Tax
        if self.tax_spin:
            self.invoice.tax_rate = Decimal(str(self.tax_spin.value()))
        
        # Notes
        self.invoice.notes = self.notes_edit.toPlainText()
    
    def update_invoice_preview(self):
        """Update the invoice number preview"""
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        preview_number = InvoiceNumberGenerator.get_preview_number(date_str)
        self.invoice_no_edit.setText(preview_number)
    
    def load_client(self):
        """Load selected client data"""
        client_name = self.client_combo.currentText()
        if client_name == "-- Select Saved Client --" or client_name not in self.clients:
            return
        client_data = self.clients[client_name]
        self.client_name_edit.setText(client_name)
        self.client_email_edit.setText(client_data.get("email", ""))
        self.client_address_edit.setPlainText(client_data.get("address", ""))
    
    def save_client(self):
        """Save current client data"""
        client_name = self.client_name_edit.text().strip()
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Save Client", "Client name is required.")
            return
        
        self.clients[client_name] = {
            "email": self.client_email_edit.text(),
            "address": self.client_address_edit.toPlainText()
        }
        
        if FileManager.save_json(Config.CLIENTS_FILE, self.clients):
            # Update combo if new client
            if self.client_combo.findText(client_name) == -1:
                self.client_combo.addItem(client_name)
                self.client_combo.setCurrentText(client_name)
            
            QtWidgets.QMessageBox.information(self, "Save Client", f"Client '{client_name}' saved successfully.")
        else:
            QtWidgets.QMessageBox.critical(self, "Save Client", "Failed to save client.")
    
    def set_logo(self):
        """Set company logo"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select Logo", "", "Image Files (*.png *.jpg *.jpeg *.bmp)"
        )
        
        if file_path:
            try:
                # Copy logo to data directory
                shutil.copy2(file_path, Config.LOGO_FILE)
                self.logo_label.setText(f"✓ {Path(file_path).name}")
                QtWidgets.QMessageBox.information(self, "Logo", "Logo set successfully.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Logo Error", f"Failed to set logo: {str(e)}")
    
    def generate_pdf(self):
        """Generate PDF invoice"""
        self.update_invoice_from_form()
        
        # Validate required fields
        client_name = self.invoice.client_name.strip()
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "Client name is required.")
            return
        
        if not self.invoice.items:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "At least one invoice item is required.")
            return
        
        # Set final invoice number
        self.invoice.invoice_number = InvoiceNumberGenerator.get_next_number(self.invoice.date)
        
        # Create client directory
        client_dir = Config.INVOICES_DIR / client_name
        client_dir.mkdir(parents=True, exist_ok=True)
        
        # Create output directory under client/month
        month_dir = client_dir / self.invoice.date[:7]
        month_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate PDF
        pdf_path = month_dir / f"{self.invoice.invoice_number}.pdf"
        json_path = month_dir / f"{self.invoice.invoice_number}.json"
        
        try:
            logo_path = Config.LOGO_FILE if Config.LOGO_FILE.exists() else None
            success = PDFGenerator.generate(self.invoice, pdf_path, logo_path)
            
            if success:
                # Save JSON metadata
                FileManager.save_json(json_path, self.invoice.to_dict())
                
                QtWidgets.QMessageBox.information(
                    self, "PDF Generated", 
                    f"Invoice PDF generated successfully:\n{pdf_path}"
                )
                
                # Refresh history and preview
                self.load_history()
                self.update_invoice_preview()
            else:
                QtWidgets.QMessageBox.critical(
                    self, "PDF Generation Error", 
                    "Failed to generate PDF. Check console for details."
                )
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "PDF Generation Error", 
                f"Failed to generate PDF: {str(e)}"
            )
    
    def export_excel(self):
        """Export invoice to Excel"""
        self.update_invoice_from_form()
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export to Excel", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                ExcelExporter.export_invoice(self.invoice, Path(file_path))
                QtWidgets.QMessageBox.information(
                    self, "Export Successful", 
                    f"Invoice exported to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Export Error", 
                    f"Failed to export invoice: {str(e)}"
                )
    
    def save_json(self):
        """Save invoice as JSON"""
        self.update_invoice_from_form()
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save as JSON", "", "JSON Files (*.json)"
        )
        
        if file_path:
            try:
                FileManager.save_json(Path(file_path), self.invoice.to_dict())
                QtWidgets.QMessageBox.information(
                    self, "Save Successful", 
                    f"Invoice saved to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Save Error", 
                    f"Failed to save invoice: {str(e)}"
                )
    
    def load_history(self):
        """Load clients from invoice directories"""
        self.clients_list.clear()
        
        for client_dir in Config.INVOICES_DIR.iterdir():
            if client_dir.is_dir():
                self.clients_list.addItem(client_dir.name)
        
        self.clients_list.sortItems()
        if self.clients_list.count() > 0:
            self.clients_list.setCurrentRow(0)
    
    def load_invoices_for_client(self):
        """Load invoices for the selected client"""
        current_item = self.clients_list.currentItem()
        if not current_item:
            self.history_table.setRowCount(0)
            return
        
        client_name = current_item.text()
        client_dir = Config.INVOICES_DIR / client_name
        
        invoices = []
        for month_dir in client_dir.glob("*"):
            if month_dir.is_dir():
                for json_file in month_dir.glob("*.json"):
                    try:
                        invoice_data = FileManager.load_json(json_file)
                        if invoice_data:
                            invoice = Invoice.from_dict(invoice_data)
                            invoices.append(invoice)
                    except Exception as e:
                        print(f"Error loading {json_file}: {e}")
        
        # Sort by date descending
        invoices.sort(key=lambda x: x.date, reverse=True)
        
        self.history_table.setRowCount(len(invoices))
        for row, inv in enumerate(invoices):
            self.history_table.setItem(row, 0, QtWidgets.QTableWidgetItem(inv.invoice_number))
            self.history_table.setItem(row, 1, QtWidgets.QTableWidgetItem(inv.date))
            self.history_table.setItem(row, 2, QtWidgets.QTableWidgetItem(inv.client_name))
            self.history_table.setItem(row, 3, QtWidgets.QTableWidgetItem(Currency.format(inv.subtotal)))
            self.history_table.setItem(row, 4, QtWidgets.QTableWidgetItem(Currency.format(inv.tax_amount)))
            self.history_table.setItem(row, 5, QtWidgets.QTableWidgetItem(Currency.format(inv.total)))
    
    def open_selected_pdf(self):
        """Open selected invoice PDF"""
        current_item = self.clients_list.currentItem()
        if not current_item:
            QtWidgets.QMessageBox.warning(self, "Open PDF", "Please select a client first.")
            return
        
        client_name = current_item.text()
        current_row = self.history_table.currentRow()
        if current_row < 0:
            QtWidgets.QMessageBox.warning(self, "Open PDF", "Please select an invoice first.")
            return
        
        invoice_number = self.history_table.item(current_row, 0).text()
        date = self.history_table.item(current_row, 1).text()
        
        pdf_path = Config.INVOICES_DIR / client_name / date[:7] / f"{invoice_number}.pdf"
        
        if pdf_path.exists():
            if FileManager.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Open PDF", "PDF opened successfully.")
            else:
                QtWidgets.QMessageBox.critical(self, "Open PDF", "Failed to open PDF.")
        else:
            QtWidgets.QMessageBox.warning(self, "Open PDF", f"PDF not found:\n{pdf_path}")
    
    def export_history(self):
        """Export invoice history to Excel"""
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export History", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # Collect all invoices
                invoices = []
                for client_dir in Config.INVOICES_DIR.iterdir():
                    if client_dir.is_dir():
                        for month_dir in client_dir.glob("*"):
                            if month_dir.is_dir():
                                for json_file in month_dir.glob("*.json"):
                                    try:
                                        invoice_data = FileManager.load_json(json_file)
                                        if invoice_data:
                                            invoice = Invoice.from_dict(invoice_data)
                                            invoices.append(invoice)
                                    except Exception:
                                        continue
                
                ExcelExporter.export_history(invoices, Path(file_path))
                QtWidgets.QMessageBox.information(
                    self, "Export Successful", 
                    f"History exported to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Export Error", 
                    f"Failed to export history: {str(e)}"
                )
    
    def closeEvent(self, event):
        """Handle application close"""
        self.save_data()
        event.accept()

def main():
    """Main application entry point"""
    Config.setup_directories()
    
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName("MABS Invoice Generator")
    app.setApplicationVersion("2.0")
    
    # Set application-wide font
    font = QtGui.QFont("Segoe UI", 9)
    app.setFont(font)
    
    # Set style
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.showMaximized()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()