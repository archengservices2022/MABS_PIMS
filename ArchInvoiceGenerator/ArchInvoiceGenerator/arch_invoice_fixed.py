# arch_invoice_generator.py
import sys
import os
import json
import subprocess
import platform
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import traceback

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import pyqtSignal
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment

# ---------- Configuration ----------
class Config:
    COMPANY = {
        "name": "Arch Engineering Services",
        "address": "315 Lemay Ferry Road\nSaint Louis, MO 63125",
        "email": "admin@archengineeringservices.com",
        "phone": "314-303-0004",
        "website": "www.archengineeringservices.com"
    }
    
    DATA_DIR = Path("data")
    INVOICES_DIR = DATA_DIR / "invoices"
    CLIENTS_FILE = DATA_DIR / "clients.json"
    COUNTER_FILE = DATA_DIR / "invoice_counter.json"
    LOGO_FILE = DATA_DIR / "logo.png"
    SETTINGS_FILE = DATA_DIR / "settings.json"
    
    DEFAULT_TERMS = """Payment due within 30 days.
Late payments subject to 1.5% monthly interest.
Make checks payable to Arch Engineering Services."""

    @classmethod
    def setup_directories(cls):
        """Create necessary directories"""
        cls.INVOICES_DIR.mkdir(parents=True, exist_ok=True)
        cls.DATA_DIR.mkdir(exist_ok=True)

# ---------- Utility Classes ----------
class DecimalEncoder(json.JSONEncoder):
    """Custom JSON encoder for Decimal objects"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

class FileManager:
    """Handles file operations with proper error handling"""
    
    @staticmethod
    def load_json(path: Path, default: Any = None) -> Any:
        """Load JSON file with error handling"""
        try:
            if path.exists():
                with open(path, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            print(f"Error loading {path}: {e}")
        return default

    @staticmethod
    def save_json(path: Path, data: Any) -> bool:
        """Save JSON file with error handling"""
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False, cls=DecimalEncoder)
            return True
        except Exception as e:
            print(f"Error saving {path}: {e}")
            return False

    @staticmethod
    def open_file(path: Path) -> bool:
        """Open file with default application"""
        try:
            if platform.system() == "Windows":
                os.startfile(str(path))
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
            return True
        except Exception as e:
            print(f"Error opening file {path}: {e}")
            return False

class Currency:
    """Currency formatting utilities"""
    
    @staticmethod
    def format(value, symbol: str = "$") -> str:
        """Format decimal as currency"""
        return f"{symbol}{Currency.quantize(value)}"
    
    @staticmethod
    def quantize(value) -> Decimal:
        """Quantize decimal to 2 decimal places"""
        return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

class InvoiceNumberGenerator:
    """Handles invoice numbering system"""
    
    @staticmethod
    def get_next_number(date_str: str) -> str:
        """Generate next invoice number for given date"""
        yyyymm = date_str[:7].replace("-", "")
        counters = FileManager.load_json(Config.COUNTER_FILE, {})
        next_num = counters.get(yyyymm, 1)
        counters[yyyymm] = next_num + 1
        FileManager.save_json(Config.COUNTER_FILE, counters)
        return f"INV-{yyyymm}-{next_num:03d}"
    
    @staticmethod
    def get_preview_number(date_str: str) -> str:
        """Preview next invoice number without persisting"""
        yyyymm = date_str[:7].replace("-", "")
        counters = FileManager.load_json(Config.COUNTER_FILE, {})
        preview_next = counters.get(yyyymm, 0)
        
        # Count existing invoices for this month
        month_dir = Config.INVOICES_DIR / date_str[:7]
        if month_dir.exists():
            count = len(list(month_dir.glob("*.json")))
            preview_next = max(preview_next, count)
        
        return f"INV-{yyyymm}-{(preview_next + 1):03d}"

# ---------- Data Models ----------

class PDFGenerator:
    """Generates simple, clean PDF invoices using ReportLab's SimpleDocTemplate and Flowables"""
    
    @staticmethod
    def generate(invoice: Invoice, output_path: Path, logo_path: Optional[Path] = None):
        doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                                leftMargin=25*mm, rightMargin=25*mm,
                                topMargin=25*mm, bottomMargin=25*mm)
        styles = getSampleStyleSheet()
        story = []

        # Custom styles
        styles.add(ParagraphStyle(name='CenteredBold20', alignment=1, fontName='Helvetica-Bold', fontSize=20, leading=24))
        styles.add(ParagraphStyle(name='Centered10', alignment=1, fontName='Helvetica', fontSize=10, leading=12))
        styles.add(ParagraphStyle(name='LeftBold16', alignment=0, fontName='Helvetica-Bold', fontSize=16, leading=18))
        styles.add(ParagraphStyle(name='LeftBold12', alignment=0, fontName='Helvetica-Bold', fontSize=12, leading=14))
        styles.add(ParagraphStyle(name='Left10', alignment=0, fontName='Helvetica', fontSize=10, leading=12))
        styles.add(ParagraphStyle(name='LeftBold10', alignment=0, fontName='Helvetica-Bold', fontSize=10, leading=12))
        styles.add(ParagraphStyle(name='Right10', alignment=2, fontName='Helvetica', fontSize=10, leading=12))
        styles.add(ParagraphStyle(name='RightBold10', alignment=2, fontName='Helvetica-Bold', fontSize=10, leading=12))
        styles.add(ParagraphStyle(name='Left9', alignment=0, fontName='Helvetica', fontSize=9, leading=11))

        # --- Company Header ---
        story.append(Paragraph(Config.COMPANY["name"], styles['CenteredBold20']))
        for line in Config.COMPANY["address"].split('\n'):
            story.append(Paragraph(line, styles['Centered10']))
        story.append(Paragraph(f"Phone: {Config.COMPANY['phone']} • Email: {Config.COMPANY['email']}", styles['Centered10']))
        story.append(Paragraph(Config.COMPANY["website"], styles['Centered10']))
        story.append(Spacer(1, 10*mm))

        # Separator line (using a Table for a simple line)
        story.append(Table([['']], colWidths=[doc.width], style=[
            ('LINEBELOW', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(Spacer(1, 10*mm))

        # --- Invoice Title and Details ---
        story.append(Paragraph("INVOICE", styles['LeftBold16']))
        story.append(Spacer(1, 5*mm))
        story.append(Paragraph(f"Invoice Number: {invoice.invoice_number}", styles['Left10']))
        story.append(Paragraph(f"Date: {invoice.date}", styles['Left10']))
        story.append(Paragraph(f"Due Date: {invoice.due_date}", styles['Left10']))
        story.append(Spacer(1, 10*mm))

        # --- Client Information ---
        story.append(Paragraph("Bill To:", styles['LeftBold12']))
        story.append(Paragraph(invoice.client_name, styles['Left10']))
        for line in invoice.client_address.split('\n'):
            if line.strip():
                story.append(Paragraph(line.strip(), styles['Left10']))
        if invoice.client_email:
            story.append(Paragraph(invoice.client_email, styles['Left10']))
        story.append(Spacer(1, 10*mm))

        # --- Line Items ---
        story.append(Paragraph("Items:", styles['LeftBold12']))
        story.append(Spacer(1, 5*mm))

        item_data = [[Paragraph(header, styles['LeftBold10']) for header in ["Description", "Qty", "Unit Price", "Total"]]]
        for item in invoice.items:
            item_data.append([
                Paragraph(item.description, styles['Left10']),
                Paragraph(str(item.quantity), styles['Right10']),
                Paragraph(Currency.format(item.unit_price), styles['Right10']),
                Paragraph(Currency.format(item.total), styles['Right10'])
            ])
        
        item_table = Table(item_data, colWidths=[doc.width * 0.5, doc.width * 0.15, doc.width * 0.175, doc.width * 0.175])
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(item_table)
        story.append(Spacer(1, 10*mm))

        # --- Totals ---
        total_data = []
        total_data.append([Paragraph("Subtotal:", styles['Right10']), Paragraph(Currency.format(invoice.subtotal), styles['Right10'])])
        if invoice.discount_rate > 0:
            total_data.append([Paragraph(f"Discount ({invoice.discount_rate}%):", styles['Right10']), Paragraph(f"-{Currency.format(invoice.discount_amount)}", styles['Right10'])])
        if invoice.tax_rate > 0:
            total_data.append([Paragraph(f"Tax ({invoice.tax_rate}%):", styles['Right10']), Paragraph(Currency.format(invoice.tax_amount), styles['Right10'])])
        
        if invoice.downpayment_rate > 0:
            total_data.append([Paragraph("Total Before Downpayment:", styles['Right10']), Paragraph(Currency.format(invoice.total_before_downpayment), styles['Right10'])])
            total_data.append([Paragraph(f"Downpayment ({invoice.downpayment_rate}%):", styles['Right10']), Paragraph(f"-{Currency.format(invoice.downpayment_amount)}", styles['Right10'])])
            total_data.append([Paragraph("Balance Due:", styles['RightBold10']), Paragraph(Currency.format(invoice.balance_due), styles['RightBold10'])])
        else:
            total_data.append([Paragraph("Total Due:", styles['RightBold10']), Paragraph(Currency.format(invoice.total), styles['RightBold10'])])

        total_table = Table(total_data, colWidths=[doc.width * 0.7, doc.width * 0.3])
        total_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 15*mm))

        # --- Footer (Notes/Terms) ---
        story.append(Paragraph("Notes/Terms:", styles['LeftBold12']))
        for line in invoice.notes.split('\n'):
            story.append(Paragraph(line, styles['Left9']))
        story.append(Spacer(1, 10*mm))

        try:
            doc.build(story)
            return True
        except Exception as e:
            print(f"PDF generation error: {e}")
            traceback.print_exc()
            return False




class InvoiceItem:
    """Represents an invoice line item"""
    
    def __init__(self, description: str = "", quantity: int = 1, unit_price: float = 0.0):
        self.description = description
        self.quantity = quantity
        self.unit_price = Currency.quantize(unit_price)
    
    @property
    def total(self) -> Decimal:
        return Currency.quantize(self.quantity * self.unit_price)
    
    def to_dict(self) -> Dict:
        return {
            "description": self.description,
            "quantity": self.quantity,
            "unit_price": float(self.unit_price),
            "total": float(self.total)
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'InvoiceItem':
        return cls(
            description=data.get("description", ""),
            quantity=data.get("quantity", 1),
            unit_price=data.get("unit_price", 0.0)
        )

class Invoice:
    """Represents a complete invoice with downpayment support"""
    
    def __init__(self):
        self.invoice_number = ""
        self.date = datetime.now().strftime("%Y-%m-%d")
        self.due_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        self.client_name = ""
        self.client_email = ""
        self.client_address = ""
        self.items: List[InvoiceItem] = []
        self.tax_rate = Decimal("0.0")
        self.discount_rate = Decimal("0.0")
        self.downpayment_rate = Decimal("0.0")
        self.notes = Config.DEFAULT_TERMS
        self.logo_path = None
    
    @property
    def subtotal(self) -> Decimal:
        return Currency.quantize(sum(item.total for item in self.items))
    
    @property
    def tax_amount(self) -> Decimal:
        return Currency.quantize(self.subtotal * self.tax_rate / Decimal("100"))
    
    @property
    def discount_amount(self) -> Decimal:
        return Currency.quantize(self.subtotal * self.discount_rate / Decimal("100"))
    
    @property
    def downpayment_amount(self) -> Decimal:
        return Currency.quantize(self.subtotal * self.downpayment_rate / Decimal("100"))
    
    @property
    def total_before_downpayment(self) -> Decimal:
        return Currency.quantize(self.subtotal + self.tax_amount - self.discount_amount)
    
    @property
    def balance_due(self) -> Decimal:
        return Currency.quantize(self.total_before_downpayment - self.downpayment_amount)
    
    @property
    def total(self) -> Decimal:
        if self.downpayment_amount > 0:
            return self.balance_due
        return self.total_before_downpayment
    
    def add_item(self, item: InvoiceItem):
        self.items.append(item)
    
    def remove_item(self, index: int):
        if 0 <= index < len(self.items):
            self.items.pop(index)
    
    def to_dict(self) -> Dict:
        return {
            "meta": {
                "invoice_number": self.invoice_number,
                "date": self.date,
                "due_date": self.due_date,
                "client_name": self.client_name,
                "client_email": self.client_email,
                "client_address": self.client_address,
                "tax_rate": float(self.tax_rate),
                "discount_rate": float(self.discount_rate),
                "downpayment_rate": float(self.downpayment_rate),
                "subtotal": float(self.subtotal),
                "tax_amount": float(self.tax_amount),
                "discount_amount": float(self.discount_amount),
                "downpayment_amount": float(self.downpayment_amount),
                "total_before_downpayment": float(self.total_before_downpayment),
                "balance_due": float(self.balance_due),
                "total": float(self.total),
                "notes": self.notes
            },
            "items": [item.to_dict() for item in self.items]
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'Invoice':
        invoice = cls()
        meta = data.get("meta", {})
        
        invoice.invoice_number = meta.get("invoice_number", "")
        invoice.date = meta.get("date", invoice.date)
        invoice.due_date = meta.get("due_date", invoice.due_date)
        invoice.client_name = meta.get("client_name", "")
        invoice.client_email = meta.get("client_email", "")
        invoice.client_address = meta.get("client_address", "")
        invoice.tax_rate = Decimal(str(meta.get("tax_rate", 0.0)))
        invoice.discount_rate = Decimal(str(meta.get("discount_rate", 0.0)))
        invoice.downpayment_rate = Decimal(str(meta.get("downpayment_rate", 0.0)))
        invoice.notes = meta.get("notes", Config.DEFAULT_TERMS)
        
        invoice.items = [InvoiceItem.from_dict(item_data) 
                        for item_data in data.get("items", [])]
        
        return invoice

# ---------- PDF Generator (Fixed Overlapping Issues) ----------
# ---------- Excel Exporter ----------
class ExcelExporter:
    """Handles Excel export functionality"""
    
    @staticmethod
    def export_invoice(invoice: Invoice, output_path: Path):
        """Export single invoice to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = "INVOICE"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Invoice info
        ws['A3'] = "Invoice No:"
        ws['B3'] = invoice.invoice_number
        ws['A4'] = "Date:"
        ws['B4'] = invoice.date
        ws['A5'] = "Due Date:"
        ws['B5'] = invoice.due_date
        
        # Client info
        ws['A7'] = "Bill To:"
        ws['A8'] = invoice.client_name
        for i, line in enumerate(invoice.client_address.splitlines()):
            ws.cell(row=9 + i, column=1, value=line)
        
        # Items table
        headers = ["Description", "Quantity", "Unit Price", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, item in enumerate(invoice.items, 13):
            ws.cell(row=row, column=1, value=item.description)
            ws.cell(row=row, column=2, value=item.quantity)
            ws.cell(row=row, column=3, value=float(item.unit_price))
            ws.cell(row=row, column=4, value=float(item.total))
        
        # Totals
        last_row = 13 + len(invoice.items)
        totals = [
            ("Subtotal:", invoice.subtotal),
            (f"Tax ({invoice.tax_rate}%):", invoice.tax_amount),
            (f"Discount ({invoice.discount_rate}%):", invoice.discount_amount),
        ]
        
        # Add downpayment if applicable
        if invoice.downpayment_amount > 0:
            totals.append((f"Downpayment ({invoice.downpayment_rate}%):", -invoice.downpayment_amount))
            totals.append(("BALANCE DUE:", invoice.balance_due))
        else:
            totals.append(("TOTAL:", invoice.total))
        
        for i, (label, amount) in enumerate(totals):
            row = last_row + 2 + i
            ws.cell(row=row, column=3, value=label)
            ws.cell(row=row, column=4, value=float(amount))
            if i == len(totals) - 1:  # Last row (TOTAL or BALANCE DUE)
                ws.cell(row=row, column=3).font = Font(bold=True)
                ws.cell(row=row, column=4).font = Font(bold=True)
        
        wb.save(str(output_path))
    
    @staticmethod
    def export_history(invoices: List[Invoice], output_path: Path):
        """Export invoice history to Excel"""
        data = []
        for invoice in invoices:
            data.append({
                "Invoice No": invoice.invoice_number,
                "Date": invoice.date,
                "Client": invoice.client_name,
                "Subtotal": float(invoice.subtotal),
                "Tax": float(invoice.tax_amount),
                "Discount": float(invoice.discount_amount),
                "Downpayment": float(invoice.downpayment_amount),
                "Total": float(invoice.total)
            })
        
        df = pd.DataFrame(data)
        df.to_excel(str(output_path), index=False, engine='openpyxl')

# ---------- Qt Widgets ----------
class ItemRowWidget(QtWidgets.QWidget):
    """Widget for a single invoice item row"""
    removed = pyqtSignal()
    
    def __init__(self, item: InvoiceItem = None):
        super().__init__()
        self.item = item or InvoiceItem()
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(10)
        
        self.desc_edit = QtWidgets.QLineEdit(self.item.description)
        self.desc_edit.setPlaceholderText("Item description")
        self.desc_edit.setMinimumHeight(35)
        
        self.qty_spin = QtWidgets.QSpinBox()
        self.qty_spin.setRange(1, 1000000)
        self.qty_spin.setValue(self.item.quantity)
        self.qty_spin.setMinimumHeight(35)
        self.qty_spin.setMinimumWidth(80)
        
        self.price_spin = QtWidgets.QDoubleSpinBox()
        self.price_spin.setRange(0, 1000000)
        self.price_spin.setDecimals(2)
        self.price_spin.setValue(float(self.item.unit_price))
        self.price_spin.setPrefix("$ ")
        self.price_spin.setMinimumHeight(35)
        self.price_spin.setMinimumWidth(100)
        
        self.total_label = QtWidgets.QLabel(Currency.format(self.item.total))
        self.total_label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.total_label.setMinimumWidth(80)
        self.total_label.setMinimumHeight(35)
        self.total_label.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 8px;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        
        self.remove_btn = QtWidgets.QPushButton("✕")
        self.remove_btn.setMaximumWidth(40)
        self.remove_btn.setMinimumHeight(35)
        self.remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.remove_btn.clicked.connect(self.removed)
        
        layout.addWidget(self.desc_edit, 5)
        layout.addWidget(self.qty_spin, 1)
        layout.addWidget(self.price_spin, 1)
        layout.addWidget(self.total_label, 1)
        layout.addWidget(self.remove_btn, 0)
        
        # Connect signals
        self.qty_spin.valueChanged.connect(self.update_total)
        self.price_spin.valueChanged.connect(self.update_total)
        self.desc_edit.textChanged.connect(self.update_item)
    
    def update_total(self):
        self.item.quantity = self.qty_spin.value()
        self.item.unit_price = Currency.quantize(self.price_spin.value())
        self.total_label.setText(Currency.format(self.item.total))
        self.update_item()
    
    def update_item(self):
        self.item.description = self.desc_edit.text()
    
    def get_item(self) -> InvoiceItem:
        return self.item

class ScrollArea(QtWidgets.QScrollArea):
    """Custom scroll area with better styling"""
    def __init__(self):
        super().__init__()
        self.setWidgetResizable(True)
        self.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.setStyleSheet("""
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
            }
            QScrollBar:vertical {
                border: none;
                background: #f8f9fa;
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #ced4da;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #adb5bd;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)

class MainWindow(QtWidgets.QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.invoice = Invoice()
        self.clients = {}
        self.settings = {}
        self.item_rows = []
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        self.setWindowTitle("Arch Invoice Generator - Professional Billing System")
        self.setMinimumSize(1000, 700)
        
        # Set window icon and style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 1em;
                padding-top: 10px;
                background-color: white;
                font-size: 11px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
                font-size: 12px;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: bold;
                min-height: 20px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
            QPushButton.success {
                background-color: #28a745;
            }
            QPushButton.success:hover {
                background-color: #218838;
            }
            QPushButton.warning {
                background-color: #ffc107;
                color: #212529;
            }
            QPushButton.warning:hover {
                background-color: #e0a800;
            }
            QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QComboBox {
                padding: 6px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background-color: white;
                font-size: 11px;
                min-height: 20px;
            }
            QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {
                border-color: #3498db;
            }
            QLabel {
                color: #2c3e50;
                font-size: 11px;
            }
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #e9ecef;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-size: 11px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #3498db;
            }
            QFormLayout > QLabel, QFormLayout > QLineEdit, QFormLayout > QDoubleSpinBox {
                margin: 2px;
                padding: 4px;
            }
            QGroupBox {
                margin: 5px;
            }
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
                padding: 2px;
            }
        """)
        
        # Create central widget with scroll area for main content
        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QtWidgets.QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)
        
        # Create tabs
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setDocumentMode(True)
        self.create_tab = QtWidgets.QWidget()
        self.history_tab = QtWidgets.QWidget()
        
        self.tabs.addTab(self.create_tab, "📄 Create Invoice")
        self.tabs.addTab(self.history_tab, "📊 Invoice History")
        
        main_layout.addWidget(self.tabs)
        
        self.setup_create_tab()
        self.setup_history_tab()
        
        # Connect signals
        self.tax_spin.valueChanged.connect(self.update_totals)
        self.discount_spin.valueChanged.connect(self.update_totals)
        self.downpayment_spin.valueChanged.connect(self.update_totals)
        self.date_edit.dateChanged.connect(self.update_invoice_preview)
        
        self.update_totals()
        self.update_invoice_preview()
    
    def setup_create_tab(self):
        # Create scroll area for the create tab
        scroll_area = ScrollArea()
        scroll_content = QtWidgets.QWidget()
        scroll_area.setWidget(scroll_content)
        
        # Main layout for scroll content
        main_layout = QtWidgets.QVBoxLayout(scroll_content)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Company header with better styling
        company_frame = QtWidgets.QFrame()
        company_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                border-radius: 8px;
                padding: 12px;
            }
        """)
        company_layout = QtWidgets.QHBoxLayout(company_frame)
        
        company_info = QtWidgets.QLabel(
            f"<h2 style='color: white; margin: 0; font-size: 16px;'>{Config.COMPANY['name']}</h2>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['address'].replace(chr(10), '<br>')}</p>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['email']} | {Config.COMPANY['phone']}</p>"
            f"<p style='color: #ecf0f1; margin: 3px 0; font-size: 11px;'>{Config.COMPANY['website']}</p>"
        )
        company_info.setTextFormat(QtCore.Qt.RichText)
        company_layout.addWidget(company_info)
        company_layout.addStretch()
        
        main_layout.addWidget(company_frame)
        
        # Create a grid layout for the main content
        content_layout = QtWidgets.QGridLayout()
        content_layout.setSpacing(12)
        content_layout.setColumnStretch(0, 1)
        content_layout.setColumnStretch(1, 1)
        
        # Invoice Details - Column 0
        details_group = QtWidgets.QGroupBox("Invoice Details")
        details_layout = QtWidgets.QFormLayout(details_group)
        details_layout.setVerticalSpacing(8)
        details_layout.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)
        
        self.invoice_no_edit = QtWidgets.QLineEdit()
        self.invoice_no_edit.setReadOnly(True)
        self.invoice_no_edit.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 11px;")
        
        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("MMM d, yyyy")
        
        self.due_date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate().addDays(30))
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("MMM d, yyyy")
        
        details_layout.addRow("Invoice Number:", self.invoice_no_edit)
        details_layout.addRow("Invoice Date:", self.date_edit)
        details_layout.addRow("Due Date:", self.due_date_edit)
        
        content_layout.addWidget(details_group, 0, 0)
        
        # Client Information - Column 1
        client_group = QtWidgets.QGroupBox("Client Information")
        client_layout = QtWidgets.QVBoxLayout(client_group)
        
        # Client form
        client_form = QtWidgets.QFormLayout()
        client_form.setVerticalSpacing(8)
        client_form.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)
        
        self.client_name_edit = QtWidgets.QLineEdit()
        self.client_name_edit.setPlaceholderText("Client or Company Name")
        
        self.client_email_edit = QtWidgets.QLineEdit()
        self.client_email_edit.setPlaceholderText("email@example.com")
        
        self.client_address_edit = QtWidgets.QTextEdit()
        self.client_address_edit.setMaximumHeight(70)
        self.client_address_edit.setPlaceholderText("Street Address\nCity, State ZIP\nCountry")
        
        client_form.addRow("Name:", self.client_name_edit)
        client_form.addRow("Email:", self.client_email_edit)
        client_form.addRow("Address:", self.client_address_edit)
        
        client_layout.addLayout(client_form)
        
        # Client management buttons
        client_buttons_layout = QtWidgets.QHBoxLayout()
        
        self.client_combo = QtWidgets.QComboBox()
        self.client_combo.addItem("-- Select Saved Client --")
        
        load_client_btn = QtWidgets.QPushButton("📂 Load Client")
        save_client_btn = QtWidgets.QPushButton("💾 Save Client")
        
        load_client_btn.clicked.connect(self.load_client)
        save_client_btn.clicked.connect(self.save_client)
        
        client_buttons_layout.addWidget(QtWidgets.QLabel("Saved Clients:"))
        client_buttons_layout.addWidget(self.client_combo, 2)
        client_buttons_layout.addWidget(load_client_btn)
        client_buttons_layout.addWidget(save_client_btn)
        
        client_layout.addLayout(client_buttons_layout)
        
        content_layout.addWidget(client_group, 0, 1)
        
        # Items Section - Span both columns
        items_group = QtWidgets.QGroupBox("Line Items")
        items_layout = QtWidgets.QVBoxLayout(items_group)
        
        # Items header
        header_frame = QtWidgets.QFrame()
        header_layout = QtWidgets.QHBoxLayout(header_frame)
        header_layout.setContentsMargins(8, 4, 8, 4)
        
        headers = ["Description", "Quantity", "Unit Price", "Total", ""]
        widths = [5, 1, 1, 1, 0]
        
        for header, width in zip(headers, widths):
            label = QtWidgets.QLabel(header)
            label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 11px;")
            header_layout.addWidget(label, width)
        
        items_layout.addWidget(header_frame)
        
        # Items container with scroll area
        items_scroll = ScrollArea()
        items_scroll.setMinimumHeight(180)
        
        self.items_widget = QtWidgets.QWidget()
        self.items_layout = QtWidgets.QVBoxLayout(self.items_widget)
        self.items_layout.setSpacing(4)
        items_scroll.setWidget(self.items_widget)
        
        items_layout.addWidget(items_scroll)
        
        # Add item button
        add_item_btn = QtWidgets.QPushButton("➕ Add New Item")
        add_item_btn.setStyleSheet("QPushButton { font-size: 11px; }")
        add_item_btn.clicked.connect(self.add_item_row)
        items_layout.addWidget(add_item_btn)
        
        content_layout.addWidget(items_group, 1, 0, 1, 2)
        
        # Totals and Actions - Bottom section
        bottom_layout = QtWidgets.QHBoxLayout()
        bottom_layout.setSpacing(12)
        
        # Left - Settings
        settings_group = QtWidgets.QGroupBox("Settings & Logo")
        settings_layout = QtWidgets.QVBoxLayout(settings_group)
        
        # Logo section
        logo_layout = QtWidgets.QHBoxLayout()
        self.logo_btn = QtWidgets.QPushButton("🖼️ Set Company Logo")
        self.logo_label = QtWidgets.QLabel("No logo set")
        self.logo_label.setStyleSheet("color: #6c757d; font-style: italic; font-size: 10px;")
        logo_layout.addWidget(self.logo_btn)
        logo_layout.addWidget(self.logo_label)
        logo_layout.addStretch()
        
        self.logo_btn.clicked.connect(self.set_logo)
        
        # Tax, discount, and downpayment
        calc_layout = QtWidgets.QFormLayout()
        calc_layout.setVerticalSpacing(6)
        
        self.tax_spin = QtWidgets.QDoubleSpinBox()
        self.tax_spin.setRange(0, 100)
        self.tax_spin.setDecimals(2)
        self.tax_spin.setSuffix(" %")
        self.tax_spin.setValue(0.0)
        
        self.discount_spin = QtWidgets.QDoubleSpinBox()
        self.discount_spin.setRange(0, 100)
        self.discount_spin.setDecimals(2)
        self.discount_spin.setSuffix(" %")
        self.discount_spin.setValue(0.0)
        
        self.downpayment_spin = QtWidgets.QDoubleSpinBox()
        self.downpayment_spin.setRange(0, 100)
        self.downpayment_spin.setDecimals(2)
        self.downpayment_spin.setSuffix(" %")
        self.downpayment_spin.setValue(0.0)
        
        calc_layout.addRow("Tax Rate:", self.tax_spin)
        calc_layout.addRow("Discount Rate:", self.discount_spin)
        calc_layout.addRow("Downpayment Rate:", self.downpayment_spin)
        
        settings_layout.addLayout(logo_layout)
        settings_layout.addLayout(calc_layout)
        
        bottom_layout.addWidget(settings_group, 1)
        
        # Center - Totals
        totals_group = QtWidgets.QGroupBox("Invoice Totals")
        totals_layout = QtWidgets.QFormLayout(totals_group)
        totals_layout.setVerticalSpacing(8)
        
        self.subtotal_label = QtWidgets.QLabel("$0.00")
        self.tax_label = QtWidgets.QLabel("$0.00")
        self.discount_label = QtWidgets.QLabel("$0.00")
        
        self.downpayment_label = QtWidgets.QLabel("")
        self.downpayment_label.setStyleSheet("""
            QLabel {
                font-size: 11px;
                font-weight: bold;
                color: #e74c3c;
                background-color: #fdf2f2;
                border: 1px solid #e74c3c;
                border-radius: 4px;
                padding: 6px;
                margin: 2px;
            }
        """)
        self.downpayment_label.setVisible(False)
        
        self.total_label = QtWidgets.QLabel("$0.00")
        self.total_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                background-color: #e8f4fd;
                border: 2px solid #3498db;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        
        totals_layout.addRow("Subtotal:", self.subtotal_label)
        totals_layout.addRow("Tax Amount:", self.tax_label)
        totals_layout.addRow("Discount Amount:", self.discount_label)
        totals_layout.addRow("", self.downpayment_label)
        totals_layout.addRow("TOTAL:", self.total_label)
        
        bottom_layout.addWidget(totals_group, 1)
        
        # Right - Actions
        actions_group = QtWidgets.QGroupBox("Actions")
        actions_layout = QtWidgets.QVBoxLayout(actions_group)
        
        self.generate_pdf_btn = QtWidgets.QPushButton("📄 Generate PDF Invoice")
        self.export_excel_btn = QtWidgets.QPushButton("📊 Export to Excel")
        self.save_json_btn = QtWidgets.QPushButton("💾 Save as JSON")
        
        # Style action buttons
        self.generate_pdf_btn.setStyleSheet("QPushButton.success { font-size: 11px; min-height: 35px; }")
        self.generate_pdf_btn.setProperty("class", "success")
        self.export_excel_btn.setStyleSheet("font-size: 11px; min-height: 35px;")
        self.save_json_btn.setStyleSheet("font-size: 11px; min-height: 35px;")
        
        self.generate_pdf_btn.clicked.connect(self.generate_pdf)
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.save_json_btn.clicked.connect(self.save_json)
        
        actions_layout.addWidget(self.generate_pdf_btn)
        actions_layout.addWidget(self.export_excel_btn)
        actions_layout.addWidget(self.save_json_btn)
        actions_layout.addStretch()
        
        bottom_layout.addWidget(actions_group, 1)
        
        content_layout.addLayout(bottom_layout, 2, 0, 1, 2)
        
        # Notes section
        notes_group = QtWidgets.QGroupBox("Notes & Terms")
        notes_layout = QtWidgets.QVBoxLayout(notes_group)
        self.notes_edit = QtWidgets.QTextEdit(Config.DEFAULT_TERMS)
        self.notes_edit.setMaximumHeight(80)
        notes_layout.addWidget(self.notes_edit)
        
        content_layout.addWidget(notes_group, 3, 0, 1, 2)
        
        main_layout.addLayout(content_layout)
        
        # Set the scroll area as the main widget for create tab
        self.create_tab_layout = QtWidgets.QVBoxLayout(self.create_tab)
        self.create_tab_layout.setContentsMargins(0, 0, 0, 0)
        self.create_tab_layout.addWidget(scroll_area)
        
        # Add initial item row
        self.add_item_row()

    def setup_history_tab(self):
        # Create scroll area for history tab
        scroll_area = ScrollArea()
        scroll_content = QtWidgets.QWidget()
        scroll_area.setWidget(scroll_content)
        
        layout = QtWidgets.QVBoxLayout(scroll_content)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(12)
        
        # Controls
        controls_frame = QtWidgets.QFrame()
        controls_layout = QtWidgets.QHBoxLayout(controls_frame)
        
        self.refresh_btn = QtWidgets.QPushButton("🔄 Refresh")
        self.open_pdf_btn = QtWidgets.QPushButton("📂 Open PDF")
        self.export_history_btn = QtWidgets.QPushButton("📈 Export History to Excel")
        
        controls_layout.addWidget(self.refresh_btn)
        controls_layout.addWidget(self.open_pdf_btn)
        controls_layout.addWidget(self.export_history_btn)
        controls_layout.addStretch()
        
        layout.addWidget(controls_frame)
        
        # Table
        self.history_table = QtWidgets.QTableWidget()
        self.history_table.setColumnCount(7)
        self.history_table.setHorizontalHeaderLabels([
            "Invoice No", "Date", "Client", "Subtotal", "Tax", "Downpayment", "Total"
        ])
        self.history_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.history_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        
        # Set column widths
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)
        
        layout.addWidget(self.history_table)
        
        # Set the scroll area as the main widget for history tab
        self.history_tab_layout = QtWidgets.QVBoxLayout(self.history_tab)
        self.history_tab_layout.setContentsMargins(0, 0, 0, 0)
        self.history_tab_layout.addWidget(scroll_area)
        
        # Connect signals
        self.refresh_btn.clicked.connect(self.load_history)
        self.open_pdf_btn.clicked.connect(self.open_selected_pdf)
        self.export_history_btn.clicked.connect(self.export_history)
    
    def load_data(self):
        """Load clients and settings"""
        self.clients = FileManager.load_json(Config.CLIENTS_FILE, {})
        self.settings = FileManager.load_json(Config.SETTINGS_FILE, {})
        
        # Update client combo
        self.client_combo.clear()
        self.client_combo.addItem("-- Select Saved Client --")
        for client_name in sorted(self.clients.keys()):
            self.client_combo.addItem(client_name)
        
        # Load logo setting
        if Config.LOGO_FILE.exists():
            self.logo_label.setText(f"✓ {Config.LOGO_FILE.name}")
    
    def save_data(self):
        """Save clients and settings"""
        FileManager.save_json(Config.CLIENTS_FILE, self.clients)
        FileManager.save_json(Config.SETTINGS_FILE, self.settings)
    
    def add_item_row(self, item: InvoiceItem = None):
        """Add a new item row to the form"""
        row = ItemRowWidget(item)
        row.removed.connect(lambda: self.remove_item_row(row))
        self.items_layout.addWidget(row)
        self.item_rows.append(row)
        
        # Connect signals for auto-update
        row.qty_spin.valueChanged.connect(self.update_totals)
        row.price_spin.valueChanged.connect(self.update_totals)
        
        self.update_totals()
    
    def remove_item_row(self, row):
        """Remove an item row from the form"""
        if row in self.item_rows:
            self.item_rows.remove(row)
            row.setParent(None)
            self.update_totals()
    
    def update_totals(self):
        """Update the totals display with downpayment support"""
        # Update invoice object from form
        self.update_invoice_from_form()
        
        # Update UI labels
        if self.subtotal_label:
            self.subtotal_label.setText(Currency.format(self.invoice.subtotal))
        if self.tax_label:
            self.tax_label.setText(Currency.format(self.invoice.tax_amount))
        if self.discount_label:
            self.discount_label.setText(Currency.format(self.invoice.discount_amount))
        
        # Update downpayment display if applicable
        if hasattr(self, 'downpayment_label'):
            if self.invoice.downpayment_amount > 0:
                downpayment_text = f"Downpayment ({self.invoice.downpayment_rate}%): -{Currency.format(self.invoice.downpayment_amount)}"
                self.downpayment_label.setText(downpayment_text)
                self.downpayment_label.setStyleSheet("""
                    QLabel {
                        font-size: 11px;
                        font-weight: bold;
                        color: #e74c3c;
                        background-color: #fdf2f2;
                        border: 1px solid #e74c3c;
                        border-radius: 4px;
                        padding: 6px;
                        margin: 2px;
                    }
                """)
                self.downpayment_label.setVisible(True)
            else:
                self.downpayment_label.setVisible(False)
        
        # Update total label with clear distinction
        if self.total_label:
            if self.invoice.downpayment_amount > 0:
                total_text = f"BALANCE DUE: {Currency.format(self.invoice.balance_due)}"
                self.total_label.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        font-weight: bold;
                        color: #27ae60;
                        background-color: #f0f9f4;
                        border: 2px solid #27ae60;
                        border-radius: 6px;
                        padding: 10px;
                    }
                """)
            else:
                total_text = f"TOTAL: {Currency.format(self.invoice.total)}"
                self.total_label.setStyleSheet("""
                    QLabel {
                        font-size: 14px;
                        font-weight: bold;
                        color: #2c3e50;
                        background-color: #e8f4fd;
                        border: 2px solid #3498db;
                        border-radius: 6px;
                        padding: 10px;
                    }
                """)
            
            self.total_label.setText(total_text)
    
    def update_invoice_from_form(self):
        """Update the invoice object from form data"""
        # Basic info
        self.invoice.date = self.date_edit.date().toString("yyyy-MM-dd")
        self.invoice.due_date = self.due_date_edit.date().toString("yyyy-MM-dd")
        
        # Client info
        self.invoice.client_name = self.client_name_edit.text()
        self.invoice.client_email = self.client_email_edit.text()
        self.invoice.client_address = self.client_address_edit.toPlainText()
        
        # Items
        self.invoice.items = []
        for row in self.item_rows:
            self.invoice.items.append(row.get_item())
        
        # Tax, discount, and downpayment
        if self.tax_spin:
            self.invoice.tax_rate = Decimal(str(self.tax_spin.value()))
        if self.discount_spin:
            self.invoice.discount_rate = Decimal(str(self.discount_spin.value()))
        if hasattr(self, 'downpayment_spin') and self.downpayment_spin:
            self.invoice.downpayment_rate = Decimal(str(self.downpayment_spin.value()))
        
        # Notes
        self.invoice.notes = self.notes_edit.toPlainText()
    
    def update_invoice_preview(self):
        """Update the invoice number preview"""
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        preview_number = InvoiceNumberGenerator.get_preview_number(date_str)
        self.invoice_no_edit.setText(preview_number)
    
    def load_client(self):
        """Load selected client data"""
        client_name = self.client_combo.currentText()
        if client_name in self.clients:
            client_data = self.clients[client_name]
            self.client_name_edit.setText(client_name)
            self.client_email_edit.setText(client_data.get("email", ""))
            self.client_address_edit.setPlainText(client_data.get("address", ""))
    
    def save_client(self):
        """Save current client data"""
        client_name = self.client_name_edit.text().strip()
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Save Client", "Client name is required.")
            return
        
        self.clients[client_name] = {
            "email": self.client_email_edit.text(),
            "address": self.client_address_edit.toPlainText()
        }
        
        if FileManager.save_json(Config.CLIENTS_FILE, self.clients):
            # Update combo if new client
            if self.client_combo.findText(client_name) == -1:
                self.client_combo.addItem(client_name)
                self.client_combo.setCurrentText(client_name)
            
            QtWidgets.QMessageBox.information(self, "Save Client", f"Client '{client_name}' saved successfully.")
        else:
            QtWidgets.QMessageBox.critical(self, "Save Client", "Failed to save client.")
    
    def set_logo(self):
        """Set company logo"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Select Logo", "", "Image Files (*.png *.jpg *.jpeg *.bmp)"
        )
        
        if file_path:
            try:
                # Copy logo to data directory
                import shutil
                shutil.copy2(file_path, Config.LOGO_FILE)
                self.logo_label.setText(f"✓ {Path(file_path).name}")
                QtWidgets.QMessageBox.information(self, "Logo", "Logo set successfully.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Logo Error", f"Failed to set logo: {str(e)}")
    
    def generate_pdf(self):
        """Generate PDF invoice"""
        self.update_invoice_from_form()
        
        # Validate required fields
        if not self.client_name_edit.text().strip():
            QtWidgets.QMessageBox.warning(self, "Validation Error", "Client name is required.")
            return
        
        if not self.invoice.items:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "At least one invoice item is required.")
            return
        
        # Set final invoice number
        self.invoice.invoice_number = InvoiceNumberGenerator.get_next_number(self.invoice.date)
        
        # Create output directory
        month_dir = Config.INVOICES_DIR / self.invoice.date[:7]
        month_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate PDF
        pdf_path = month_dir / f"{self.invoice.invoice_number}.pdf"
        json_path = month_dir / f"{self.invoice.invoice_number}.json"
        
        try:
            logo_path = Config.LOGO_FILE if Config.LOGO_FILE.exists() else None
            success = PDFGenerator.generate(self.invoice, pdf_path, logo_path)
            
            if success:
                # Save JSON metadata
                FileManager.save_json(json_path, self.invoice.to_dict())
                
                QtWidgets.QMessageBox.information(
                    self, "PDF Generated", 
                    f"Invoice PDF generated successfully:\n{pdf_path}"
                )
                
                # Refresh history and preview
                self.load_history()
                self.update_invoice_preview()
            else:
                QtWidgets.QMessageBox.critical(
                    self, "PDF Generation Error", 
                    "Failed to generate PDF. Check console for details."
                )
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "PDF Generation Error", 
                f"Failed to generate PDF: {str(e)}"
            )
    
    def export_excel(self):
        """Export invoice to Excel"""
        self.update_invoice_from_form()
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export to Excel", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                ExcelExporter.export_invoice(self.invoice, Path(file_path))
                QtWidgets.QMessageBox.information(
                    self, "Export Successful", 
                    f"Invoice exported to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Export Error", 
                    f"Failed to export invoice: {str(e)}"
                )
    
    def save_json(self):
        """Save invoice as JSON"""
        self.update_invoice_from_form()
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save as JSON", "", "JSON Files (*.json)"
        )
        
        if file_path:
            try:
                FileManager.save_json(Path(file_path), self.invoice.to_dict())
                QtWidgets.QMessageBox.information(
                    self, "Save Successful", 
                    f"Invoice saved to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Save Error", 
                    f"Failed to save invoice: {str(e)}"
                )
    
    def load_history(self):
        """Load invoice history"""
        self.history_table.setRowCount(0)
        
        invoices = []
        for month_dir in Config.INVOICES_DIR.glob("*"):
            if month_dir.is_dir():
                for json_file in month_dir.glob("*.json"):
                    try:
                        invoice_data = FileManager.load_json(json_file)
                        if invoice_data:
                            invoice = Invoice.from_dict(invoice_data)
                            invoices.append(invoice)
                    except Exception as e:
                        print(f"Error loading {json_file}: {e}")
        
        # Sort by date descending
        invoices.sort(key=lambda x: x.date, reverse=True)
        
        self.history_table.setRowCount(len(invoices))
        for row, invoice in enumerate(invoices):
            self.history_table.setItem(row, 0, QtWidgets.QTableWidgetItem(invoice.invoice_number))
            self.history_table.setItem(row, 1, QtWidgets.QTableWidgetItem(invoice.date))
            self.history_table.setItem(row, 2, QtWidgets.QTableWidgetItem(invoice.client_name))
            self.history_table.setItem(row, 3, QtWidgets.QTableWidgetItem(Currency.format(invoice.subtotal)))
            self.history_table.setItem(row, 4, QtWidgets.QTableWidgetItem(Currency.format(invoice.tax_amount)))
            self.history_table.setItem(row, 5, QtWidgets.QTableWidgetItem(Currency.format(invoice.downpayment_amount)))
            self.history_table.setItem(row, 6, QtWidgets.QTableWidgetItem(Currency.format(invoice.total)))
    
    def open_selected_pdf(self):
        """Open selected invoice PDF"""
        current_row = self.history_table.currentRow()
        if current_row < 0:
            QtWidgets.QMessageBox.warning(self, "Open PDF", "Please select an invoice first.")
            return
        
        invoice_number = self.history_table.item(current_row, 0).text()
        date = self.history_table.item(current_row, 1).text()
        
        pdf_path = Config.INVOICES_DIR / date[:7] / f"{invoice_number}.pdf"
        
        if pdf_path.exists():
            if FileManager.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Open PDF", "PDF opened successfully.")
            else:
                QtWidgets.QMessageBox.critical(self, "Open PDF", "Failed to open PDF.")
        else:
            QtWidgets.QMessageBox.warning(self, "Open PDF", f"PDF not found:\n{pdf_path}")
    
    def export_history(self):
        """Export invoice history to Excel"""
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export History", "", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # Collect all invoices
                invoices = []
                for month_dir in Config.INVOICES_DIR.glob("*"):
                    if month_dir.is_dir():
                        for json_file in month_dir.glob("*.json"):
                            try:
                                invoice_data = FileManager.load_json(json_file)
                                if invoice_data:
                                    invoice = Invoice.from_dict(invoice_data)
                                    invoices.append(invoice)
                            except Exception:
                                continue
                
                ExcelExporter.export_history(invoices, Path(file_path))
                QtWidgets.QMessageBox.information(
                    self, "Export Successful", 
                    f"History exported to:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Export Error", 
                    f"Failed to export history: {str(e)}"
                )
    
    def closeEvent(self, event):
        """Handle application close"""
        self.save_data()
        event.accept()

def main():
    """Main application entry point"""
    Config.setup_directories()
    
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName("Arch Invoice Generator")
    app.setApplicationVersion("2.0")
    
    # Set application-wide font
    font = QtGui.QFont("Segoe UI", 9)
    app.setFont(font)
    
    # Set style
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.showMaximized()
    
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()