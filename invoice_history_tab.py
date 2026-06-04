# invoice_history_tab.py
import os
import threading
from decimal import Decimal
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional
import base64
import tempfile
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from main import PDFGenerator
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtCore import QTimer
import pandas as pd
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from datetime import datetime


from app_logger import get_logger
from app_theme import (
    INDIGO,
    INDIGO_D,
    INDIGO_L,
    AMBER,
    AMBER_D,
    AMBER_L,
    EMERALD,
    EMERALD_L,
    VIOLET,
    VIOLET_L,
    SLATE_50,
    SLATE_100,
    SLATE_200,
    SLATE_300,
    SLATE_500,
    SLATE_600,
    SLATE_700,
    SLATE_800,
    SLATE_900,
    WHITE,
)
_log = get_logger(__name__)
_invoice_data_changed = QtCore.pyqtSignal(str)


class _InvoiceStatusRefreshSignaler(QtCore.QObject):
    """Thread-safe bridge: payment_tracker background thread emits invoice_status_changed
    to trigger invoice history UI refresh on the GUI thread after a payment deletion
    causes _recompute_invoice_status to update the invoice status."""
    invoice_status_changed = QtCore.pyqtSignal(str, str, str)  # invoice_number, new_status, received_date


# Module-level singleton — set when InvoiceHistoryTab is created, used by payment_tracker.py
_invoice_status_signaler: '_InvoiceStatusRefreshSignaler | None' = None

PROJECT_FONT = '"Inter", "Segoe UI", "Arial", sans-serif'

# Import from main application
try:
    from arch_invoice_generator import Config, FileManager, Invoice, Currency, FirebaseManager, FIREBASE_AVAILABLE, InvoiceItem
except ImportError:
    try:
        from main import Config, FileManager, Invoice, Currency, FirebaseManager, FIREBASE_AVAILABLE, InvoiceItem
    except ImportError:
        class Config: pass
        class FileManager: pass  
        class Invoice: pass
        class Currency: 
            @staticmethod
            def format(value): return f"${value}"
            @staticmethod
            def format_whole(value): return f"${int(float(value))}"
        class InvoiceItem: pass
        class FirebaseManager: 
            @staticmethod
            def load_invoices(): return []
            @staticmethod
            def sync_status_to_firebase(client_name, invoice_number, status): pass
        FIREBASE_AVAILABLE = False

def _normalize_payment_stage(stage: str) -> str:
    """Canonicalise any payment stage label for duplicate-checking. Never adds percentages."""
    s = (stage or "").strip()
    lo = s.lower()
    if any(x in lo for x in ("down payment", "deposit")):
        return "Down Payment"
    if any(x in lo for x in ("1st installment", "1st payment")):
        return "1st Installment"
    if any(x in lo for x in ("2nd installment", "2nd payment", "term 2", "second")):
        return "2nd Installment"
    if any(x in lo for x in ("3rd installment", "3rd payment", "term 3", "third")):
        return "3rd Installment"
    if any(x in lo for x in ("4th installment", "4th payment", "term 4", "fourth")):
        return "4th Installment"
    if any(x in lo for x in ("balance", "due payment", "full amount due")):
        return "Balance Payment"
    if "final" in lo:
        return "Final Payment"
    if any(x in lo for x in ("full amount", "full payment")):
        return "Full Payment"
    return s  # keep as-is if unknown


def _normalize_date(date_str: str) -> str:
    """Return date as MM-dd-YYYY. Falls back to today if unparseable."""
    from datetime import datetime as _dt
    for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return _dt.strptime((date_str or "").strip()[:10], fmt).strftime("%m-%d-%Y")
        except (ValueError, TypeError):
            pass
    return _dt.now().strftime("%m-%d-%Y")


def _normalize_search_text(value) -> str:
    """Lowercase text with punctuation removed so invoice numbers match reliably."""
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


import io
try:
    from PyPDF2 import PdfReader, PdfWriter
    PDF_AVAILABLE = True
except ImportError:
    try:
        from pypdf import PdfReader, PdfWriter
        PDF_AVAILABLE = True
    except ImportError:
        PDF_AVAILABLE = False

class TextWrapDelegate(QtWidgets.QStyledItemDelegate):
    """Custom delegate that forces text wrapping in table cells."""

    def paint(self, painter, option, index):
        """Paint cell with full text wrapping, no truncation."""
        painter.save()

        # Get the text
        text = index.data(QtCore.Qt.DisplayRole) or ""

        # If no text, use default rendering
        if not text:
            super().paint(painter, option, index)
            painter.restore()
            return

        # Set up the painter and text
        painter.setClipRect(option.rect)

        # Determine background color based on selection state
        if option.state & QtWidgets.QStyle.State_Selected:
            # Use the same blue highlight as other columns
            bg_color = QtGui.QColor("#3498db")  # Match table stylesheet
            text_color = QtGui.QColor("#ffffff")  # White text on blue
        else:
            # Use default background
            bg_color = option.palette.base().color()
            text_color = option.palette.text().color()

        # Fill the background
        painter.fillRect(option.rect, bg_color)

        # Draw the text with wrapping and center alignment
        text_rect = option.rect.adjusted(8, 4, -8, -4)  # Padding
        painter.setFont(option.font)
        painter.setPen(text_color)
        painter.drawText(text_rect, QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter | QtCore.Qt.TextWordWrap, text)

        painter.restore()

    def sizeHint(self, option, index):
        """Calculate size needed to display full wrapped text."""
        text = index.data(QtCore.Qt.DisplayRole) or ""
        if not text:
            return super().sizeHint(option, index)

        # Calculate text height with wrapping
        fm = QtGui.QFontMetrics(option.font)
        text_rect = QtCore.QRect(0, 0, 280, 10000)  # Fixed width for calculation
        text_rect = fm.boundingRect(text_rect, QtCore.Qt.AlignLeft | QtCore.Qt.AlignTop | QtCore.Qt.TextWordWrap, text)

        return QtCore.QSize(280, text_rect.height() + 8)


class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""

    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year
        self.end_year = end_year
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        self.setStyleSheet('* { font-family: "Inter", "Segoe UI", Arial, sans-serif; }')
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton("◀◀")
        self.prev_block_btn.setFixedSize(40, 30)
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #2980b9; }
            QPushButton:pressed { background: #21618c; }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)
        
        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)
        
        self.next_block_btn = QtWidgets.QPushButton("▶▶")
        self.next_block_btn.setFixedSize(40, 30)
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #2980b9; }
            QPushButton:pressed { background: #21618c; }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        layout.addLayout(nav_layout)
        
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        self.year_buttons = []
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        layout.addLayout(current_layout)
        
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        years = [self.current_block_start + i for i in range(9)]
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        current_year = datetime.now().year
        
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover { background: #2980b9; }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        self.current_block_start -= 9
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        self.current_block_start += 9
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
    
    def get_selected_year(self):
        return self.selected_year


class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        self.setStyleSheet('* { font-family: "Inter", "Segoe UI", Arial, sans-serif; }')
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)
        
        header = QtWidgets.QLabel("📅 Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        self.selected_label = QtWidgets.QLabel("")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        layout.addLayout(selected_layout)
        
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        self.current_year = year
    
    def on_select_clicked(self):
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        return self.current_year


class _WheelBlocker(QtCore.QObject):
    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Wheel:
            return True
        return False

class _NoScrollComboBox(QtWidgets.QComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._wb = _WheelBlocker(self)
        self.installEventFilter(self._wb)

class _NoScrollDateEdit(QtWidgets.QDateEdit):
    def stepBy(self, steps):
        pass


class PDFExportDialog(QtWidgets.QDialog):
    """Professional PDF/Excel Export Dialog with Tabs"""
    
    def __init__(self, parent=None, client_name="", available_dates=None):
        super().__init__(parent)
        self.client_name = client_name
        self.available_dates = available_dates or []
        self.export_range = "all"
        self.selected_dates = []
        self.export_type = "pdf"
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle(f"📊 Export Manager - {self.client_name}")
        self.setFixedSize(700, 750)
        self.setStyleSheet("""
            PDFExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        header = QtWidgets.QLabel("📤 Export Manager")
        header.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 10px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        client_card = QtWidgets.QWidget()
        client_card.setStyleSheet("""
            QWidget {
                background: white;
                border-radius: 10px;
                border: 2px solid #3498db;
                padding: 10px;
            }
        """)
        client_layout = QtWidgets.QVBoxLayout(client_card)
        client_label = QtWidgets.QLabel(f"🏢 Client: {self.client_name}")
        client_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px;
            }
        """)
        client_layout.addWidget(client_label)
        layout.addWidget(client_card)
        
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #d5dbdb;
            }
        """)
        
        self.pdf_tab = QtWidgets.QWidget()
        self.setup_pdf_tab()
        self.tab_widget.addTab(self.pdf_tab, "📄 PDF Export")
        
        self.excel_tab = QtWidgets.QWidget()
        self.setup_excel_tab()
        self.tab_widget.addTab(self.excel_tab, "📊 Excel Export")
        
        layout.addWidget(self.tab_widget)
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        layout.addLayout(button_layout)
    
    def setup_pdf_tab(self):
        layout = QtWidgets.QVBoxLayout(self.pdf_tab)
        layout.setSpacing(15)
        
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Invoices")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)
        
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(160, 45)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover { border-color: #3498db; }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)
        
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(160, 45)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover { border-color: #3498db; }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)
        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)
        
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)
        
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(24)

        month_col = QtWidgets.QVBoxLayout()
        month_col.setSpacing(6)
        month_label = QtWidgets.QLabel("Month")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_col.addWidget(month_label)
        self.month_combo = _NoScrollComboBox()
        self.month_combo.setFixedHeight(42)
        self.month_combo.setMinimumWidth(160)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover { border-color: #3498db; }
        """)
        self.populate_months()
        month_col.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_col)

        year_container_col = QtWidgets.QVBoxLayout()
        year_container_col.setSpacing(6)
        year_label_month = QtWidgets.QLabel("Year")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container_col.addWidget(year_label_month)

        year_field_row = QtWidgets.QHBoxLayout()
        year_field_row.setSpacing(6)

        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn_month.setFixedHeight(45)
        self.year_calendar_btn_month.setMinimumWidth(72)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_field_row.addWidget(self.year_edit_month)
        year_field_row.addWidget(self.year_calendar_btn_month)
        year_container_col.addLayout(year_field_row)
        month_year_row_layout.addLayout(year_container_col)
        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Year")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)

        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn.setFixedHeight(45)
        self.year_calendar_btn.setMinimumWidth(72)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)
        self.date_selection_layout.addWidget(self.year_group)
        
        layout.addWidget(self.date_selection_container)
        
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel("Ready to export all invoices as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)
        
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.month_combo.currentTextChanged.connect(self.update_preview)
    
    def setup_excel_tab(self):
        layout = QtWidgets.QVBoxLayout(self.excel_tab)
        layout.setSpacing(15)
        
        options_card = QtWidgets.QGroupBox("🎯 Excel Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        self.excel_all_radio = QtWidgets.QRadioButton("📋 Export All Invoices")
        self.excel_all_radio.setChecked(True)
        self.excel_all_radio.toggled.connect(lambda: self.on_excel_range_changed("all"))
        
        self.excel_date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.excel_date_range_radio.toggled.connect(lambda: self.on_excel_range_changed("date_range"))
        
        self.excel_month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.excel_month_radio.toggled.connect(lambda: self.on_excel_range_changed("month"))
        
        self.excel_year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.excel_year_radio.toggled.connect(lambda: self.on_excel_range_changed("year"))
        
        options_layout.addWidget(self.excel_all_radio)
        options_layout.addWidget(self.excel_date_range_radio)
        options_layout.addWidget(self.excel_month_radio)
        options_layout.addWidget(self.excel_year_radio)
        
        range_group = QtWidgets.QButtonGroup(self)
        range_group.addButton(self.excel_all_radio)
        range_group.addButton(self.excel_date_range_radio)
        range_group.addButton(self.excel_month_radio)
        range_group.addButton(self.excel_year_radio)
        
        layout.addWidget(options_card)
        
        self.excel_date_selection_container = QtWidgets.QWidget()
        self.excel_date_selection_layout = QtWidgets.QVBoxLayout(self.excel_date_selection_container)
        self.excel_date_selection_layout.setSpacing(15)
        self.excel_date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        self.excel_date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.excel_date_range_group.setMinimumHeight(120)
        self.excel_date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_date_range_layout = QtWidgets.QHBoxLayout(self.excel_date_range_group)
        excel_date_range_layout.setSpacing(20)
        
        excel_from_layout = QtWidgets.QVBoxLayout()
        excel_from_label = QtWidgets.QLabel("From Date:")
        excel_from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_from_layout.addWidget(excel_from_label)
        self.excel_from_date = _NoScrollDateEdit()
        self.excel_from_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.excel_from_date.setCalendarPopup(True)
        self.excel_from_date.setFixedSize(160, 45)
        self.excel_from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover { border-color: #3498db; }
        """)
        excel_from_layout.addWidget(self.excel_from_date)
        excel_date_range_layout.addLayout(excel_from_layout)
        
        excel_to_layout = QtWidgets.QVBoxLayout()
        excel_to_label = QtWidgets.QLabel("To Date:")
        excel_to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_to_layout.addWidget(excel_to_label)
        self.excel_to_date = _NoScrollDateEdit()
        self.excel_to_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_to_date.setDate(QtCore.QDate.currentDate())
        self.excel_to_date.setCalendarPopup(True)
        self.excel_to_date.setFixedSize(160, 45)
        self.excel_to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover { border-color: #3498db; }
        """)
        excel_to_layout.addWidget(self.excel_to_date)
        excel_date_range_layout.addLayout(excel_to_layout)
        excel_date_range_layout.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_date_range_group)
        
        self.excel_month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.excel_month_group.setMinimumHeight(150)
        self.excel_month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_month_layout = QtWidgets.QVBoxLayout(self.excel_month_group)
        excel_month_layout.setSpacing(15)
        
        excel_month_year_row_layout = QtWidgets.QHBoxLayout()
        excel_month_year_row_layout.setSpacing(24)

        excel_month_col = QtWidgets.QVBoxLayout()
        excel_month_col.setSpacing(6)
        excel_month_label = QtWidgets.QLabel("Month")
        excel_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_month_col.addWidget(excel_month_label)
        self.excel_month_combo = _NoScrollComboBox()
        self.excel_month_combo.setFixedHeight(42)
        self.excel_month_combo.setMinimumWidth(160)
        self.excel_month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover { border-color: #3498db; }
        """)
        self.populate_months_excel()
        excel_month_col.addWidget(self.excel_month_combo)
        excel_month_year_row_layout.addLayout(excel_month_col)

        excel_year_container_col = QtWidgets.QVBoxLayout()
        excel_year_container_col.setSpacing(6)
        excel_year_month_label = QtWidgets.QLabel("Year")
        excel_year_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_container_col.addWidget(excel_year_month_label)

        excel_year_field_row = QtWidgets.QHBoxLayout()
        excel_year_field_row.setSpacing(6)

        self.excel_year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit_month.setFixedSize(150, 45)
        self.excel_year_edit_month.setReadOnly(True)
        self.excel_year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.excel_year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn_month.setFixedHeight(45)
        self.excel_year_calendar_btn_month.setMinimumWidth(72)
        self.excel_year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month_excel)

        excel_year_field_row.addWidget(self.excel_year_edit_month)
        excel_year_field_row.addWidget(self.excel_year_calendar_btn_month)
        excel_year_container_col.addLayout(excel_year_field_row)
        excel_month_year_row_layout.addLayout(excel_year_container_col)
        excel_month_year_row_layout.addStretch()
        excel_month_layout.addLayout(excel_month_year_row_layout)
        self.excel_date_selection_layout.addWidget(self.excel_month_group)

        self.excel_year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.excel_year_group.setMinimumHeight(120)
        self.excel_year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_year_layout = QtWidgets.QVBoxLayout(self.excel_year_group)
        excel_year_layout.setSpacing(15)

        excel_year_row_layout = QtWidgets.QHBoxLayout()
        excel_year_label = QtWidgets.QLabel("Year")
        excel_year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_row_layout.addWidget(excel_year_label)

        self.excel_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit.setFixedSize(150, 45)
        self.excel_year_edit.setReadOnly(True)
        self.excel_year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.excel_year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn.setFixedHeight(45)
        self.excel_year_calendar_btn.setMinimumWidth(72)
        self.excel_year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn.clicked.connect(self.show_year_popup_excel)
        
        excel_year_row_layout.addWidget(self.excel_year_edit)
        excel_year_row_layout.addWidget(self.excel_year_calendar_btn)
        excel_year_row_layout.addStretch()
        excel_year_layout.addLayout(excel_year_row_layout)
        self.excel_date_selection_layout.addWidget(self.excel_year_group)
        
        layout.addWidget(self.excel_date_selection_container)
        
        self.excel_date_selection_container.setVisible(False)
        self.excel_date_range_group.setVisible(False)
        self.excel_month_group.setVisible(False)
        self.excel_year_group.setVisible(False)
        
        excel_preview_card = QtWidgets.QGroupBox("👁️ Excel Export Preview")
        excel_preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        excel_preview_layout = QtWidgets.QVBoxLayout(excel_preview_card)
        
        self.excel_preview_label = QtWidgets.QLabel("Ready to export all invoices as Excel")
        self.excel_preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.excel_preview_label.setWordWrap(True)
        excel_preview_layout.addWidget(self.excel_preview_label)
        
        layout.addWidget(excel_preview_card)
        
        self.excel_from_date.dateChanged.connect(self.update_excel_preview)
        self.excel_to_date.dateChanged.connect(self.update_excel_preview)
        self.excel_month_combo.currentTextChanged.connect(self.update_excel_preview)
    
    def show_year_popup(self):
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        self.year_calendar_popup.exec_()
    
    def show_year_popup_excel(self):
        try:
            current_year = int(self.excel_year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_excel)
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month_excel(self):
        try:
            current_year = int(self.excel_year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month_excel)
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_excel(self, year):
        self.excel_year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def on_year_selected_for_month_excel(self, year):
        self.excel_year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def populate_months(self):
        months = ["January", "February", "March", "April", "May", "June",
                 "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def populate_months_excel(self):
        months = ["January", "February", "March", "April", "May", "June",
                 "July", "August", "September", "October", "November", "December"]
        self.excel_month_combo.addItems(months)
        self.excel_month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_tab_changed(self, index):
        if index == 0:
            self.export_type = "pdf"
            self.export_btn.setText("🚀 Export PDF")
            self.update_preview()
        elif index == 1:
            self.export_type = "excel"
            self.export_btn.setText("🚀 Export Excel")
            self.update_excel_preview()
    
    def on_range_changed(self, range_type):
        self.export_range = range_type
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        self.date_selection_container.setVisible(range_type != "all")
        self.update_preview()
    
    def on_excel_range_changed(self, range_type):
        self.excel_export_range = range_type
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        self.excel_date_range_group.setVisible(date_range_visible)
        self.excel_month_group.setVisible(month_visible)
        self.excel_year_group.setVisible(year_visible)
        self.excel_date_selection_container.setVisible(range_type != "all")
        self.update_excel_preview()
    
    def update_preview(self):
        if self.export_range == "all":
            self.preview_label.setText("📋 Will export ALL invoices for this client as PDF")
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toString("MM-dd-yyyy")
            to_date = self.to_date.date().toString("MM-dd-yyyy")
            self.preview_label.setText(f"📅 Will export invoices from {from_date} to {to_date} as PDF")
        elif self.export_range == "month":
            month = self.month_combo.currentText()
            year = self.year_edit_month.text()
            self.preview_label.setText(f"🗓️ Will export invoices for {month} {year} as PDF")
        elif self.export_range == "year":
            year = self.year_edit.text()
            self.preview_label.setText(f"📊 Will export invoices for the year {year} as PDF")
    
    def update_excel_preview(self):
        if hasattr(self, 'excel_export_range'):
            range_type = self.excel_export_range
        else:
            range_type = "all"
        
        if range_type == "all":
            self.excel_preview_label.setText("📋 Will export ALL invoices for this client as Excel")
        elif range_type == "date_range":
            from_date = self.excel_from_date.date().toString("MM-dd-yyyy")
            to_date = self.excel_to_date.date().toString("MM-dd-yyyy")
            self.excel_preview_label.setText(f"📅 Will export invoices from {from_date} to {to_date} as Excel")
        elif range_type == "month":
            month = self.excel_month_combo.currentText()
            year = self.excel_year_edit_month.text()
            self.excel_preview_label.setText(f"🗓️ Will export invoices for {month} {year} as Excel")
        elif range_type == "year":
            year = self.excel_year_edit.text()
            self.excel_preview_label.setText(f"📊 Will export invoices for the year {year} as Excel")
    
    def get_export_parameters(self):
        if self.export_type == "pdf":
            if self.export_range == "all":
                return {"range": "all", "type": "pdf"}
            elif self.export_range == "date_range":
                from_date = self.from_date.date().toPyDate()
                to_date = self.to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "pdf"}
            elif self.export_range == "month":
                month = self.month_combo.currentIndex() + 1
                year = int(self.year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "pdf"}
            elif self.export_range == "year":
                year = int(self.year_edit.text())
                return {"range": "year", "year": year, "type": "pdf"}
        elif self.export_type == "excel":
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            if range_type == "all":
                return {"range": "all", "type": "excel"}
            elif range_type == "date_range":
                from_date = self.excel_from_date.date().toPyDate()
                to_date = self.excel_to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "excel"}
            elif range_type == "month":
                month = self.excel_month_combo.currentIndex() + 1
                year = int(self.excel_year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "excel"}
            elif range_type == "year":
                year = int(self.excel_year_edit.text())
                return {"range": "year", "year": year, "type": "excel"}
    
    def start_export(self):
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
        
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            self._export_params = export_params
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False


class DateRangeWidget(QtWidgets.QWidget):
    """Widget for selecting date range"""
    date_range_changed = pyqtSignal(datetime, datetime)
    date_range_cleared = pyqtSignal()
    search_changed = pyqtSignal(str)  # emitted whenever the search bar text changes
    
    def __init__(self):
        super().__init__()
        self.is_date_range_applied = False
        self.init_ui()
    
    def init_ui(self):
        self.setStyleSheet('* { font-family: "Inter", "Segoe UI", Arial, sans-serif; }')
        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(0, 3, 0, 3)
        layout.setSpacing(10)

        self.calendar_btn = QtWidgets.QPushButton("Date Range")
        self.calendar_btn.setFixedSize(106, 36)
        self.calendar_btn.setStyleSheet("""
            QPushButton {
                background-color: #00756f;
                color: #ffffff;
                border: 2px solid #00756f;
                border-radius: 8px;
                font-size: 12px;
                font-weight: 600;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background-color: #005f5a; border-color: #005f5a; }
            QPushButton:pressed { background-color: #004d49; border-color: #004d49; }
        """)
        self.calendar_btn.clicked.connect(self.toggle_date_range_visibility)
        layout.addWidget(self.calendar_btn)
        
        self.search_bar = QtWidgets.QLineEdit()
        self.search_bar.setPlaceholderText("Search invoices...")
        self.search_bar.setFixedHeight(40)
        self.search_bar.setFixedWidth(260)
        self.search_bar.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #d8e2ec;
                border-radius: 7px;
                font-size: 14px;
                background-color: white;
                color: #0f172a;
            }
            QLineEdit:focus { border-color: #00756f; }
        """)
        self.search_bar.textChanged.connect(self.on_search_changed)
        layout.addWidget(self.search_bar)
        
        self.date_range_container = QtWidgets.QWidget()
        self.date_range_container.setVisible(False)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_container)
        date_range_layout.setContentsMargins(10, 5, 10, 5)
        date_range_layout.setSpacing(10)
        
        date_range_layout.addWidget(QtWidgets.QLabel("From:"))
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setMinimumWidth(148)
        self.from_date.setFixedHeight(36)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 28px 6px 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:focus { border-color: #00756f; }
        """)
        self.from_date.dateChanged.connect(self.on_date_changed)
        date_range_layout.addWidget(self.from_date)

        date_range_layout.addWidget(QtWidgets.QLabel("To:"))
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setMinimumWidth(148)
        self.to_date.setFixedHeight(36)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 28px 6px 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:focus { border-color: #00756f; }
        """)
        self.to_date.dateChanged.connect(self.on_date_changed)
        date_range_layout.addWidget(self.to_date)
        
        self.apply_clear_btn = QtWidgets.QPushButton("Apply")
        self.apply_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        self.apply_clear_btn.clicked.connect(self.toggle_apply_clear)
        date_range_layout.addWidget(self.apply_clear_btn)
        
        layout.addWidget(self.date_range_container)
        layout.addStretch()
    
    def toggle_date_range_visibility(self):
        visible = not self.date_range_container.isVisible()
        self.date_range_container.setVisible(visible)
        _base = """
                font-size: 12px; font-weight: 600;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                border-radius: 8px;
            """
        if visible:
            self.calendar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: #005f5a;
                    color: #ffffff;
                    border: 2px solid #005f5a;
                    {_base}
                }}
                QPushButton:hover {{ background-color: #004d49; border-color: #004d49; }}
            """)
        else:
            self.calendar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: #00756f;
                    color: #ffffff;
                    border: 2px solid #00756f;
                    {_base}
                }}
                QPushButton:hover {{ background-color: #005f5a; border-color: #005f5a; }}
                QPushButton:pressed {{ background-color: #004d49; border-color: #004d49; }}
            """)
    
    def on_date_changed(self):
        if self.is_date_range_applied:
            self.is_date_range_applied = False
            self.apply_clear_btn.setText("Apply")
            self.apply_clear_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover { background-color: #229954; }
            """)
    
    def toggle_apply_clear(self):
        if self.is_date_range_applied:
            self.clear_date_range()
        else:
            self.apply_date_range()
    
    def apply_date_range(self):
        from_date_qdate = self.from_date.date()
        to_date_qdate = self.to_date.date()
        from_date_str = from_date_qdate.toString("MM-dd-yyyy")
        to_date_str = to_date_qdate.toString("MM-dd-yyyy")
        from_date = datetime.strptime(from_date_str, "%m-%d-%Y")
        to_date = datetime.strptime(to_date_str, "%m-%d-%Y")
        
        self.is_date_range_applied = True
        self.apply_clear_btn.setText("Clear")
        self.apply_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #c0392b; }
        """)
        self.date_range_changed.emit(from_date, to_date)
    
    def clear_date_range(self):
        self.is_date_range_applied = False
        self.apply_clear_btn.setText("Apply")
        self.apply_clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #229954; }
        """)
        
        if hasattr(self.parent(), 'parent') and hasattr(self.parent().parent(), 'clear_quick_filter_highlighting'):
            self.parent().parent().clear_quick_filter_highlighting()
        
        self.date_range_cleared.emit()
    
    def set_date_range(self, from_date: datetime, to_date: datetime):
        from_date_str = from_date.strftime("%m-%d-%Y")
        to_date_str = to_date.strftime("%m-%d-%Y")
        self.from_date.setDate(QtCore.QDate.fromString(from_date_str, "MM-dd-yyyy"))
        self.to_date.setDate(QtCore.QDate.fromString(to_date_str, "MM-dd-yyyy"))
    
    def hide_date_range(self):
        self.date_range_container.setVisible(False)
    
    def on_search_changed(self, text):
        self.search_changed.emit(text)


class ClientListWidget(QtWidgets.QWidget):
    """Widget for displaying client list with professional styling"""
    client_selected = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.init_ui()
    
    def init_ui(self):
        _outer = QtWidgets.QVBoxLayout(self)
        _outer.setContentsMargins(0, 0, 0, 0)
        _outer.setSpacing(0)

        self._page_scroll = QtWidgets.QScrollArea()
        self._page_scroll.setWidgetResizable(True)
        self._page_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self._page_scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self._page_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        self._page_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")

        _page_widget = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(_page_widget)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(14)

        header_widget = QtWidgets.QWidget()
        header_widget.setFixedHeight(76)
        header_widget.setStyleSheet("""
            QWidget {
                background: #0f3b57;
                border-radius: 10px;
            }
        """)
        header_layout = QtWidgets.QVBoxLayout(header_widget)
        header_layout.setContentsMargins(22, 10, 22, 10)
        
        title = QtWidgets.QLabel("📊 INVOICE HISTORY")
        title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: white;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
        """)
        title.setAlignment(QtCore.Qt.AlignCenter)
        title.setText("Invoice History")
        
        subtitle = QtWidgets.QLabel("Client & Invoice Management")
        subtitle.setStyleSheet("""
            QLabel {
                font-size: 13px;
                color: rgba(255,255,255,0.9);
                font-weight: normal;
            }
        """)
        subtitle.setAlignment(QtCore.Qt.AlignCenter)
        subtitle.setText("Client directory, invoice records, status tracking, and exports")
        
        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        layout.addWidget(header_widget)
        
        content_card = QtWidgets.QWidget()
        content_card.setStyleSheet("QWidget { background: transparent; border: none; }")
        content_layout = QtWidgets.QVBoxLayout(content_card)
        content_layout.setContentsMargins(18, 18, 18, 18)
        content_layout.setSpacing(14)
        
        header_section = QtWidgets.QWidget()
        header_layout_inner = QtWidgets.QHBoxLayout(header_section)
        header_layout_inner.setContentsMargins(0, 0, 0, 0)
        
        section_title = QtWidgets.QLabel("🏢 Client Directory")
        section_title.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                padding: 8px 0;
                border: none;
            }
        """)
        
        section_title.setText("Client Directory")

        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search clients...")
        self.search_edit.setPlaceholderText("Search clients...")
        self.search_edit.setFixedHeight(40)
        self.search_edit.setMinimumWidth(320)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                color: #2c3e50;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background-color: #f8fafc;
            }
        """)
        self.search_edit.textChanged.connect(self.filter_clients)
        
        self.export_all_btn = QtWidgets.QPushButton("📤 Export All")
        self.export_all_btn.setText("Export All")
        self.export_all_btn.setFixedHeight(40)
        self.export_all_btn.setMinimumWidth(150)
        self.export_all_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e67e22, stop:1 #f39c12);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 13px;
                padding: 8px 16px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d35400, stop:1 #e67e22);
                border: 2px solid #f39c12;
            }
            QPushButton:pressed { background: #e67e22; }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_all_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.export_all_btn.clicked.connect(self.open_export_all_dialog)
        
        header_layout_inner.addWidget(section_title)
        header_layout_inner.addStretch()
        header_layout_inner.addWidget(self.search_edit)
        header_layout_inner.addWidget(self.export_all_btn)
        content_layout.addWidget(header_section)

        stats_row = QtWidgets.QWidget()
        stats_row.setStyleSheet("QWidget { background: transparent; border: none; }")
        stats_layout = QtWidgets.QHBoxLayout(stats_row)
        stats_layout.setContentsMargins(0, 0, 0, 0)
        stats_layout.setSpacing(10)
        self.client_count_label = self.create_metric_card("Clients", "0")
        self.invoice_count_label = self.create_metric_card("Invoices", "0")
        self.revenue_total_label = self.create_metric_card("Revenue", "$0.00")
        stats_layout.addWidget(self.client_count_label)
        stats_layout.addWidget(self.invoice_count_label)
        stats_layout.addWidget(self.revenue_total_label)
        content_layout.addWidget(stats_row)
        
        list_container = QtWidgets.QWidget()
        list_container.setStyleSheet("QWidget { background: transparent; border: none; }")
        list_layout = QtWidgets.QVBoxLayout(list_container)
        list_layout.setContentsMargins(0, 0, 0, 0)
        
        self.client_list = QtWidgets.QListWidget()
        self.client_list.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.client_list.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.client_list.setStyleSheet("""
            QListWidget {
                background-color: transparent;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                padding: 5px;
                outline: none;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
            QListWidget::item {
                padding: 0px;
                margin: 4px;
                border: none;
                background: transparent;
            }
            QListWidget::item:hover {
                background: transparent;
            }
            QListWidget::item:selected {
                background: transparent;
            }
            QListWidget::item:selected:hover {
                background: transparent;
            }
            QScrollBar:vertical {
                background: #f1f5f9;
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #3498db;
                min-height: 30px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover { background: #2980b9; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: none; }
        """)
        self.client_list.setAlternatingRowColors(False)
        self.client_list.itemClicked.connect(self.on_client_selected)
        self.client_list.itemDoubleClicked.connect(self.on_client_selected)
        self.client_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        list_layout.addWidget(self.client_list)
        content_layout.addWidget(list_container)
        
        self.empty_state_label = QtWidgets.QLabel("📁 No clients found. Create your first invoice to get started!")
        self.empty_state_label.setStyleSheet("""
            QLabel {
                color: #718096;
                font-size: 14px;
                font-style: italic;
                padding: 20px;
                text-align: center;
                background-color: #f7fafc;
                border-radius: 8px;
                border: 2px dashed #cbd5e0;
            }
        """)
        self.empty_state_label.setAlignment(QtCore.Qt.AlignCenter)
        self.empty_state_label.setVisible(False)
        content_layout.addWidget(self.empty_state_label)
        
        layout.addWidget(content_card)

        actions_widget = QtWidgets.QWidget()
        self.actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        self.actions_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(actions_widget)

        footer = QtWidgets.QLabel("💡 Click on any client to view detailed invoice history and analytics")
        footer.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #718096;
                font-weight: normal;
                padding: 10px;
                text-align: center;
                background: rgba(255,255,255,0.7);
                border-radius: 8px;
                border: 1px solid #e2e8f0;
            }
        """)
        footer.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(footer)

        self._page_scroll.setWidget(_page_widget)
        _outer.addWidget(self._page_scroll)

    def _update_client_list_height(self):
        n = self.client_list.count()
        h = max(n * 112, 10)  # 104px item + ~8px margin each
        self.client_list.setFixedHeight(h)

    def create_metric_card(self, label: str, value: str):
        card = QtWidgets.QLabel(
            f"<div style='font-size:18px; font-weight:800; color:#0f172a;'>{value}</div>"
            f"<div style='font-size:12px; font-weight:700; color:#64748b; margin-top:3px;'>{label}</div>"
        )
        card.setTextFormat(QtCore.Qt.RichText)
        card.setMinimumHeight(64)
        card.setAlignment(QtCore.Qt.AlignCenter)
        card.setStyleSheet("""
            QLabel {
                background: #f8fbfd;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                padding: 10px;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
        """)
        return card

    def update_metric_card(self, card, label: str, value: str):
        if card:
            card.setText(
                f"<div style='font-size:18px; font-weight:800; color:#0f172a;'>{value}</div>"
                f"<div style='font-size:12px; font-weight:700; color:#64748b; margin-top:3px;'>{label}</div>"
            )

    def create_client_card(self, client_name: str, invoices: List, item=None):
        invoice_count = len(invoices)
        revenue = sum((getattr(invoice, "total", Decimal("0")) for invoice in invoices), Decimal("0"))
        last_date = "No invoices"
        parsed_dates = []
        for invoice in invoices:
            raw_date = getattr(invoice, "date", "")
            try:
                parsed_dates.append(datetime.strptime(raw_date, "%m-%d-%Y"))
            except Exception:
                continue
        if parsed_dates:
            last_date = max(parsed_dates).strftime("%b %d, %Y")

        accents = [
            ("#00756f", "#eefaf8", "#9ddbd4"),
            ("#0f8bd6", "#eff8ff", "#b9e0fb"),
            ("#d97706", "#fff7ed", "#fed7aa"),
            ("#7c3aed", "#f5f3ff", "#ddd6fe"),
        ]
        accent, accent_bg, accent_border = accents[sum(ord(ch) for ch in client_name) % len(accents)]

        card = QtWidgets.QFrame()
        card.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {accent_bg}, stop:0.18 #ffffff, stop:1 #ffffff);
                border: 1px solid {accent_border};
                border-radius: 9px;
            }}
            QFrame:hover {{
                border-color: {accent};
            }}
            QLabel {{
                border: none;
                background: transparent;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }}
        """)

        layout = QtWidgets.QHBoxLayout(card)
        layout.setContentsMargins(18, 14, 18, 14)
        layout.setSpacing(14)

        initials = "".join(part[:1] for part in client_name.split()[:2]).upper() or "C"
        badge = QtWidgets.QLabel(initials)
        badge.setFixedSize(44, 44)
        badge.setAlignment(QtCore.Qt.AlignCenter)
        badge.setStyleSheet("""
            QLabel {
                background: %s;
                color: %s;
                border: 1px solid %s;
                border-radius: 22px;
                font-size: 15px;
                font-weight: 900;
            }
        """ % (accent_bg, accent, accent_border))

        name_block = QtWidgets.QWidget()
        name_layout = QtWidgets.QVBoxLayout(name_block)
        name_layout.setContentsMargins(0, 0, 0, 0)
        name_layout.setSpacing(3)
        name_label = QtWidgets.QLabel(client_name)
        name_label.setStyleSheet("font-size: 16px; font-weight: 900; color: #0f172a;")
        hint_label = QtWidgets.QLabel("Click to view invoice details")
        hint_label.setStyleSheet("font-size: 13px; font-weight: 800; color: #475569;")
        name_layout.addWidget(name_label)
        name_layout.addWidget(hint_label)

        def small_metric(title, value, width=92):
            widget = QtWidgets.QWidget()
            widget.setFixedWidth(width)
            widget.setMinimumHeight(54)
            widget.setStyleSheet(f"""
                QWidget {{
                    background: {accent_bg};
                    border: 1px solid {accent_border};
                    border-radius: 8px;
                }}
                QLabel {{
                    background: transparent;
                    border: none;
                }}
            """)
            metric_layout = QtWidgets.QVBoxLayout(widget)
            metric_layout.setContentsMargins(8, 7, 8, 7)
            metric_layout.setSpacing(2)
            value_label = QtWidgets.QLabel(value)
            value_label.setAlignment(QtCore.Qt.AlignCenter)
            value_label.setWordWrap(False)
            value_label.setStyleSheet("font-size: 14px; font-weight: 900; color: #0f172a;")
            title_label = QtWidgets.QLabel(title)
            title_label.setAlignment(QtCore.Qt.AlignCenter)
            title_label.setStyleSheet("font-size: 12px; font-weight: 800; color: #475569;")
            metric_layout.addWidget(value_label)
            metric_layout.addWidget(title_label)
            return widget

        layout.addWidget(badge)
        layout.addWidget(name_block, 1)
        layout.addWidget(small_metric("Invoices", str(invoice_count), 82))
        layout.addWidget(small_metric("Revenue", Currency.format(revenue), 96))
        layout.addWidget(small_metric("Last Invoice", last_date, 120))

        def open_client(_event):
            if item is not None:
                self.client_list.setCurrentItem(item)
            self.client_selected.emit(client_name)

        card.mousePressEvent = open_client
        for child in card.findChildren(QtWidgets.QWidget):
            child.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            child.mousePressEvent = open_client
        return card
    
    def open_export_all_dialog(self):
        try:
            all_clients = self.get_all_clients()
            if not all_clients:
                QtWidgets.QMessageBox.warning(self, "No Clients", "No clients found to export.")
                return
            
            dialog = AllClientsExportDialog(self, all_clients)
            result = dialog.exec_()
            
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                export_params = dialog._export_params
                selected_client = export_params.get("client", "All Clients")
                clients_to_export = (
                    [selected_client]
                    if selected_client and selected_client != "All Clients"
                    else all_clients
                )
                if export_params["type"] == "pdf":
                    self.perform_all_clients_pdf_export(export_params, clients_to_export)
                elif export_params["type"] == "excel":
                    self.perform_all_clients_excel_export(export_params, clients_to_export)
        except Exception as e:
            _log.warning("Error opening export all dialog: %s", e)
    
    def get_all_clients(self):
        try:
            all_clients = set()
            invoices_data = FirebaseManager.load_invoices()
            if invoices_data:
                for invoice_data in invoices_data:
                    if invoice_data and 'meta' in invoice_data:
                        client_name = invoice_data['meta'].get('client_name', '')
                        if client_name:
                            all_clients.add(client_name)
            
            saved_clients = FirebaseManager.load_clients()
            if saved_clients:
                all_clients.update(saved_clients.keys())
            
            return sorted(list(all_clients))
        except Exception as e:
            _log.warning("Error getting all clients: %s", e)
            return []
    
    def perform_all_clients_pdf_export(self, export_params, clients):
        try:
            all_invoices_to_export = []
            for client_name in clients:
                client_invoices = self.load_client_invoices(client_name)
                filtered_invoices = self.filter_invoices_by_date_range(client_invoices, export_params)
                for invoice in filtered_invoices:
                    all_invoices_to_export.append((invoice, client_name))

            def _ac_pdf_key(entry):
                date_str = getattr(entry[0], 'date', '') or ''
                for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(date_str, _fmt)
                    except ValueError:
                        continue
                return datetime.min
            all_invoices_to_export.sort(key=_ac_pdf_key)

            if not all_invoices_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", "No invoices found matching the selected criteria.")
                return

            self.generate_all_clients_pdf(all_invoices_to_export, export_params, clients)
        except Exception as e:
            _log.warning("Error performing all clients PDF export: %s", e)
    
    def perform_all_clients_excel_export(self, export_params, clients):
        try:
            all_invoices_to_export = []
            for client_name in clients:
                client_invoices = self.load_client_invoices(client_name)
                filtered_invoices = self.filter_invoices_by_date_range(client_invoices, export_params)
                for invoice in filtered_invoices:
                    all_invoices_to_export.append((invoice, client_name))

            def _ac_excel_key(entry):
                date_str = getattr(entry[0], 'date', '') or ''
                for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(date_str, _fmt)
                    except ValueError:
                        continue
                return datetime.min
            all_invoices_to_export.sort(key=_ac_excel_key)

            if not all_invoices_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", "No invoices found matching the selected criteria.")
                return

            self.generate_all_clients_excel(all_invoices_to_export, export_params, clients)
        except Exception as e:
            _log.warning("Error performing all clients Excel export: %s", e)
    
    def load_client_invoices(self, client_name):
        try:
            invoices = []
            all_invoices_data = FirebaseManager.load_invoices()
            if not all_invoices_data:
                return invoices
            
            for invoice_data in all_invoices_data:
                try:
                    meta = invoice_data.get('meta', {})
                    if meta.get('client_name', '') == client_name:
                        invoice = Invoice.from_dict(invoice_data)
                        if 'meta' in invoice_data:
                            meta_data = invoice_data['meta']
                            invoice.date = meta_data.get('date', invoice.date)
                            invoice.due_date = meta_data.get('due_date', invoice.due_date)
                            invoice.invoice_number = meta_data.get('invoice_number', invoice.invoice_number)
                            invoice.status = meta_data.get('status', 'Pending')
                            invoice.client_name = meta_data.get('client_name', invoice.client_name)
                            _rd = meta_data.get('received_date', 'N/A') or 'N/A'
                            invoice.received_date = _normalize_date(_rd) if _rd not in ('N/A', '') else 'N/A'
                        invoices.append(invoice)
                except Exception as e:
                    _log.warning("Error processing invoice for %s: %s", client_name, e)
                    continue
            return invoices
        except Exception as e:
            _log.warning("Error loading invoices for %s: %s", client_name, e)
            return []
    
    def filter_invoices_by_date_range(self, invoices, export_params):
        filtered_invoices = []
        for invoice in invoices:
            try:
                invoice_date = None
                if hasattr(invoice, 'firebase_timestamp') and invoice.firebase_timestamp:
                    try:
                        if isinstance(invoice.firebase_timestamp, (int, float)):
                            invoice_date = datetime.fromtimestamp(invoice.firebase_timestamp)
                        elif isinstance(invoice.firebase_timestamp, str):
                            ts_str = invoice.firebase_timestamp.replace('Z', '+00:00')
                            invoice_date = datetime.fromisoformat(ts_str)
                    except:
                        pass
                
                if invoice_date is None:
                    invoice_date = self.parse_invoice_date(invoice.date)
                
                if invoice_date is None:
                    continue
                
                include_invoice = False
                if export_params["range"] == "all":
                    include_invoice = True
                elif export_params["range"] == "date_range":
                    from_date = export_params["from_date"]
                    to_date = export_params["to_date"]
                    invoice_date_only = invoice_date.date()
                    from_date_only = from_date.date() if isinstance(from_date, datetime) else from_date
                    to_date_only = to_date.date() if isinstance(to_date, datetime) else to_date
                    if from_date_only <= invoice_date_only <= to_date_only:
                        include_invoice = True
                elif export_params["range"] == "month":
                    month = export_params["month"]
                    year = export_params["year"]
                    if invoice_date.month == month and invoice_date.year == year:
                        include_invoice = True
                elif export_params["range"] == "year":
                    year = export_params["year"]
                    if invoice_date.year == year:
                        include_invoice = True
                
                if include_invoice:
                    filtered_invoices.append(invoice)
            except Exception as e:
                _log.warning("Error filtering invoice %s: %s", invoice.invoice_number, e)
                continue
        
        return filtered_invoices
    
    def parse_invoice_date(self, date_str):
        if not date_str:
            return None
        try:
            try:
                return datetime.strptime(date_str, "%m-%d-%Y")
            except ValueError:
                pass
            date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"]
            for date_format in date_formats:
                try:
                    return datetime.strptime(date_str, date_format)
                except ValueError:
                    continue
            return None
        except Exception as e:
            _log.info("(converted from print, see git history)")
            return None
    
    def generate_all_clients_pdf(self, all_invoices, export_params, clients):
        try:
            export_dir = Path.home() / "Downloads" / "Invoice_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if export_params["range"] == "all":
                filename = f"All_Clients_All_Invoices_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"All_Clients_Invoices_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"All_Clients_Invoices_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"All_Clients_Invoices_{year}.pdf"
            
            pdf_path = export_dir / filename
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=landscape(A4),
                leftMargin=0.32*inch,
                rightMargin=0.32*inch,
                topMargin=0.25*inch,
                bottomMargin=0.3*inch,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#7f8c8d'), fontName='Helvetica', alignment=2)
            info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#2c3e50'), fontName='Helvetica-Bold')
            main_title_style = ParagraphStyle('MainTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, textColor=colors.HexColor('#2c3e50'), alignment=1, fontName='Helvetica-Bold')
            stats_style = ParagraphStyle('StatsStyle', parent=styles['Normal'], fontSize=12, spaceAfter=20, textColor=colors.HexColor('#2c3e50'), alignment=1, fontName='Helvetica-Bold')
            
            generated_date = datetime.now().strftime("%m-%d-%Y")
            header_data = [[Paragraph(f"{generated_date}", header_style)]]
            header_table = Table(header_data, colWidths=[10.8*inch])
            header_table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BOTTOMPADDING', (0,0), (-1,-1), 10), ('TOPPADDING', (0,0), (-1,-1), 10)]))
            elements.append(header_table)
            
            main_title = Paragraph(f"{Config.COMPANY.get('name', 'MABS Engineering LLC').upper()} - INVOICE REPORT", main_title_style)
            elements.append(main_title)
            
            total_invoices = len(all_invoices)
            total_amount = sum(invoice.total for invoice, _ in all_invoices)
            total_clients = len(clients)
            stats_text = f"Total Clients: {total_clients}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Total Invoices: {total_invoices}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Total Revenue: {Currency.format(total_amount)}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            if export_params["range"] == "all":
                export_range_text = "All Invoices"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            period_para = Paragraph(f"<b>Period:</b> {export_range_text}", ParagraphStyle('PeriodLeft', parent=info_style, alignment=0, leftIndent=0, firstLineIndent=0, spaceBefore=4, spaceAfter=6))
            period_table = Table([[period_para]], colWidths=[10.8*inch])
            period_table.setStyle(TableStyle([('LEFTPADDING',(0,0),(-1,-1),0), ('RIGHTPADDING',(0,0),(-1,-1),0), ('TOPPADDING',(0,0),(-1,-1),0), ('BOTTOMPADDING',(0,0),(-1,-1),0), ('ALIGN',(0,0),(-1,-1),'LEFT')]))
            elements.append(period_table)
            elements.append(Spacer(1, 12))
            
            if all_invoices:
                invoices_by_client = {}
                for invoice, client_name in all_invoices:
                    if client_name not in invoices_by_client:
                        invoices_by_client[client_name] = []
                    invoices_by_client[client_name].append(invoice)
                
                sorted_clients = sorted(invoices_by_client.keys())
                for client_name in sorted_clients:
                    client_header = Paragraph(f"<b>Client: {client_name}</b>", ParagraphStyle('ClientHeaderLeft', parent=info_style, alignment=0, leftIndent=0, firstLineIndent=0, spaceBefore=0, spaceAfter=0))
                    client_header_table = Table([[client_header]], colWidths=[10.8 * inch])
                    client_header_table.setStyle(TableStyle([('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0), ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 0), ('ALIGN', (0,0), (-1,-1), 'LEFT')]))
                    elements.append(client_header_table)
                    elements.append(Spacer(1, 6))
                    
                    header_style_center = ParagraphStyle(
                        'header_center',
                        alignment=1,
                        fontName='Helvetica-Bold',
                        fontSize=8,
                        leading=8,          # match font size
                        spaceBefore=-2,     # 👈 THIS FIX (push slightly down visually)
                        spaceAfter=0,
                        textColor=colors.whitesmoke
                    )

                    table_data = [[
                        Paragraph("Date", header_style_center),
                        Paragraph("Invoice No", header_style_center),
                        Paragraph("Project", header_style_center),
                        Paragraph("Total Price", header_style_center),
                        Paragraph("Tax", header_style_center),
                        Paragraph("Down Payment", header_style_center),
                        Paragraph("Total Due", header_style_center),
                        Paragraph("Status", header_style_center),
                        Paragraph("Received Date", header_style_center),
                    ]]
                    cell_style = ParagraphStyle(
                        'InvoiceExportCell',
                        parent=styles['Normal'],
                        fontName='Helvetica',
                        fontSize=7.5,
                        leading=9,
                        alignment=1,
                        textColor=colors.HexColor('#2c3e50'),
                    )
                    for invoice in invoices_by_client[client_name]:
                        project_name = self.get_project_name(invoice)
                        total_down_payment = sum(item.down_payment for item in invoice.items)
                        status = invoice.status if hasattr(invoice, 'status') else "Pending"
                        received_date = getattr(invoice, 'received_date', 'N/A')
                        if not received_date or received_date == '':
                            received_date = 'N/A'
                        table_data.append([
                            Paragraph(str(invoice.date), cell_style),
                            Paragraph(str(invoice.invoice_number), cell_style),
                            Paragraph(str(project_name), cell_style),
                            Paragraph(Currency.format(invoice.subtotal), cell_style),
                            Paragraph(Currency.format(invoice.tax_amount), cell_style),
                            Paragraph(Currency.format(total_down_payment), cell_style),
                            Paragraph(Currency.format(invoice.total), cell_style),
                            Paragraph(str(status), cell_style),
                            Paragraph(str(received_date), cell_style),
                        ])
                    
                    invoice_table = Table(table_data, colWidths=[0.9*inch, 1.25*inch, 2.35*inch, 0.9*inch, 0.65*inch, 1.2*inch, 0.95*inch, 1.05*inch, 1.05*inch], repeatRows=1)
                    invoice_table.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3498db')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,0), 9), ('BOTTOMPADDING', (0,0), (-1,0), 10),
                        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#ffffff')), ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#2c3e50')),
                        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#bdc3c7')),
                        ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8f9fa'), colors.white]),
                    ]))
                    elements.append(invoice_table)
                    elements.append(Spacer(1, 20))
            else:
                no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#7f8c8d'), alignment=1)
                elements.append(Paragraph("No invoices found for the selected criteria.", no_data_style))

            doc.build(elements)

            if FileManager.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ PDF exported successfully!\n\nFile saved to: {pdf_path}\nThe PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ PDF exported successfully!\n\nFile saved to: {pdf_path}\nCould not open automatically. Please open manually.")
        except Exception as e:
            _log.warning("Error generating all clients PDF: %s", e)
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", f"Error generating PDF: {str(e)}")
    
    def generate_all_clients_excel(self, all_invoices, export_params, clients):
        try:
            export_dir = Path.home() / "Downloads" / "Invoice_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if export_params["range"] == "all":
                filename = f"All_Clients_All_Invoices_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"All_Clients_Invoices_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"All_Clients_Invoices_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"All_Clients_Invoices_{year}.xlsx"
            
            excel_path = export_dir / filename
            wb = openpyxl.Workbook()
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            ws_summary.merge_cells('A1:I1')
            ws_summary['A1'] = f"{Config.COMPANY.get('name', 'MABS Engineering LLC').upper()} - ALL CLIENTS INVOICE REPORT"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A1'].alignment = Alignment(horizontal='center')
            
            if export_params["range"] == "all":
                export_range_text = "All Invoices"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws_summary['A2'] = f"Reporting Period: {export_range_text}"
            ws_summary['A2'].font = Font(bold=True)
            
            for client_name in sorted(clients):
                client_invoices = [(inv, cl) for inv, cl in all_invoices if cl == client_name]
                if not client_invoices:
                    continue
                
                ws_client = wb.create_sheet(title=client_name[:31])
                ws_client.merge_cells('A1:I1')
                ws_client['A1'] = f"Client: {client_name}"
                ws_client['A1'].font = Font(size=14, bold=True)
                ws_client['A1'].alignment = Alignment(horizontal='center')
                
                headers = ["Date", "Invoice Number", "Project Name", "Total Price", "Tax", "Down Payment", "Total Due", "Status", "Received Date"]
                for col, header in enumerate(headers, 1):
                    cell = ws_client.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                for row, (invoice, _) in enumerate(client_invoices, 4):
                    project_name = self.get_project_name(invoice)
                    total_down_payment = sum(item.down_payment for item in invoice.items)
                    status = invoice.status if hasattr(invoice, 'status') else "Pending"
                    received_date = getattr(invoice, 'received_date', 'N/A')
                    if not received_date or received_date == '':
                        received_date = 'N/A'
                    
                    data = [invoice.date, invoice.invoice_number, project_name, float(invoice.subtotal), float(invoice.tax_amount), float(total_down_payment), float(invoice.total), status, received_date]
                    for col, value in enumerate(data, 1):
                        cell = ws_client.cell(row=row, column=col, value=value)
                        if col in [4,5,6,7]:
                            cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                column_widths = {1:14, 2:22, 3:34, 4:15, 5:12, 6:18, 7:16, 8:20, 9:18}
                for col_idx, width in column_widths.items():
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws_client.column_dimensions[column_letter].width = width
                ws_client.freeze_panes = "A4"
            
            if len(wb.sheetnames) > 1 and ws_summary.max_row <= 10:
                default_sheet = wb["Summary"]
                if default_sheet.max_row <= 10:
                    wb.remove(default_sheet)
            
            wb.save(str(excel_path))
            
            if FileManager.open_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ Excel exported successfully!\n\nFile saved to: {excel_path}\nThe Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ Excel exported successfully!\n\nFile saved to: {excel_path}\nCould not open automatically. Please open manually.")
        except Exception as e:
            _log.warning("Error generating all clients Excel: %s", e)
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error", f"Error generating Excel: {str(e)}")
    
    def get_project_name(self, invoice):
        project_names = []
        for item in invoice.items:
            if hasattr(item, 'project_number') and item.project_number:
                pn = str(item.project_number).strip()
                if pn and pn not in project_names:
                    project_names.append(pn)
        if not project_names:
            return "No Project Name"
        return ", ".join(project_names)
    
    def load_clients(self):
        self.client_list.clear()
        self.client_search_index = {}
        try:
            all_clients = self.get_all_clients()
            clients_with_invoices = []
            total_invoice_count = 0
            total_revenue = Decimal("0")
            for client_name in all_clients:
                client_invoices = self.load_client_invoices(client_name)
                if client_invoices:
                    clients_with_invoices.append(client_name)
                    total_invoice_count += len(client_invoices)
                    for invoice in client_invoices:
                        total_revenue += getattr(invoice, "total", Decimal("0"))

            clients_with_invoices.sort()
            for client_name in clients_with_invoices:
                client_invoices = self.load_client_invoices(client_name)
                searchable_parts = [client_name]
                for invoice in client_invoices:
                    searchable_parts.extend([
                        getattr(invoice, "invoice_number", ""),
                        getattr(invoice, "date", ""),
                        getattr(invoice, "status", ""),
                        str(getattr(invoice, "total", "")),
                    ])
                search_blob = " ".join(str(part or "") for part in searchable_parts).lower()
                self.client_search_index[client_name] = (
                    search_blob,
                    _normalize_search_text(search_blob),
                )
                item = QtWidgets.QListWidgetItem(client_name)
                item.setData(QtCore.Qt.UserRole, client_name)
                item.setSizeHint(QtCore.QSize(0, 104))
                self.client_list.addItem(item)
                self.client_list.setItemWidget(item, self.create_client_card(client_name, client_invoices, item))

            has_clients = len(clients_with_invoices) > 0
            self.empty_state_label.setVisible(not has_clients)
            self.client_list.setVisible(has_clients)
            self.export_all_btn.setEnabled(has_clients)
            self.update_metric_card(self.client_count_label, "Clients", str(len(clients_with_invoices)))
            self.update_metric_card(self.invoice_count_label, "Invoices", str(total_invoice_count))
            self.update_metric_card(self.revenue_total_label, "Revenue", Currency.format(total_revenue))
            self._update_client_list_height()
        except Exception as e:
            _log.warning("Error loading clients from Firebase: %s", e)
            self.empty_state_label.setVisible(True)
            self.client_list.setVisible(False)
            self.export_all_btn.setEnabled(False)
    
    def filter_clients(self, search_text: str):
        search_text = search_text.lower().strip()
        normalized_search = _normalize_search_text(search_text)
        for i in range(self.client_list.count()):
            item = self.client_list.item(i)
            client_name = item.data(QtCore.Qt.UserRole) or item.text()
            item_text = item.text().lower()
            search_blob, normalized_blob = getattr(self, "client_search_index", {}).get(
                client_name,
                (item_text, _normalize_search_text(item_text)),
            )
            matches = (
                not search_text
                or search_text in search_blob
                or (normalized_search and normalized_search in normalized_blob)
            )
            item.setHidden(not matches)
        visible = sum(1 for i in range(self.client_list.count())
                      if not self.client_list.item(i).isHidden())
        self.client_list.setFixedHeight(max(visible * 112, 10))

    def on_client_selected(self, item):
        if item and not item.isHidden():
            self.client_list.setCurrentItem(item)
            self.client_selected.emit(item.data(QtCore.Qt.UserRole) or item.text())


class AllClientsExportDialog(PDFExportDialog):
    """Dialog for exporting all clients' invoices"""
    
    def __init__(self, parent=None, all_clients=None):
        super().__init__(parent, "All Clients", [])
        self.all_clients = all_clients or []
        self.setWindowTitle("📊 Export All Clients")
    
    def init_ui(self):
        super().init_ui()
        self.setWindowTitle("📊 Export All Clients")
        client_card = self.findChild(QtWidgets.QWidget)
        if client_card:
            client_layout = client_card.layout()
            if client_layout and client_layout.itemAt(0):
                client_label = client_layout.itemAt(0).widget()
                if isinstance(client_label, QtWidgets.QLabel):
                    client_label.setText(f"🏢 Clients: {len(self.all_clients)} clients")
    
    def setup_pdf_tab(self):
        super().setup_pdf_tab()
        if hasattr(self, 'preview_label'):
            if self.export_range == "all":
                self.preview_label.setText("📋 Will export ALL invoices for ALL clients as PDF")
    
    def setup_excel_tab(self):
        super().setup_excel_tab()
        if hasattr(self, 'excel_preview_label'):
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            if range_type == "all":
                self.excel_preview_label.setText("📋 Will export ALL invoices for ALL clients as Excel")


    def __init__(self, parent=None, all_clients=None):
        self.all_clients = all_clients or []
        super().__init__(parent, "All Clients", [])
        self.setWindowTitle("Export All Clients")

    def _modern_combo_style(self):
        return (
            "QComboBox { padding:5px 30px 5px 10px; border:1.5px solid #CBD5E1; "
            "border-radius:7px; font-size:13px; font-weight:700; color:#1E293B; background:white; }"
            "QComboBox:hover { border-color:#94A3B8; } QComboBox:focus { border-color:#0F766E; }"
        )

    def _modern_input_style(self):
        return (
            "QDateEdit, QLineEdit { padding:5px 10px; border:2px solid #CBD5E1; "
            "border-radius:7px; font-size:13px; font-weight:800; color:#0F172A; background:white; }"
            "QDateEdit:hover, QLineEdit:hover { border-color:#94A3B8; }"
        )

    def _modern_button_style(self, active=False):
        border = "#0F766E" if active else "#DCE4EC"
        color = "#0F766E" if active else "#334155"
        bg = "#F0FDFA" if active else "#FFFFFF"
        return (
            f"QPushButton {{ background:{bg}; color:{color}; border:1px solid {border}; "
            "border-radius:8px; font-size:12px; font-weight:800; padding:0 12px; }}"
            "QPushButton:hover { border-color:#0F766E; color:#0F766E; }"
        )

    def init_ui(self):
        self.setFixedSize(980, 640)
        self.setStyleSheet("QDialog { background:#F8FAFC; }")
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(18, 18, 18, 18)
        layout.setSpacing(12)

        header = QtWidgets.QFrame()
        header.setStyleSheet("QFrame { background:white; border:1px solid #E2E8F0; border-radius:10px; }")
        header_layout = QtWidgets.QVBoxLayout(header)
        header_layout.setContentsMargins(18, 14, 18, 14)
        title = QtWidgets.QLabel("Export All Invoices")
        title.setStyleSheet("font-size:20px; font-weight:900; color:#0F172A; border:none;")
        subtitle = QtWidgets.QLabel(f"{len(self.all_clients)} clients available for export")
        subtitle.setStyleSheet("font-size:12px; font-weight:700; color:#64748B; border:none;")
        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        layout.addWidget(header)

        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { border:none; background:transparent; }
            QTabBar::tab { background:#E2E8F0; color:#334155; padding:9px 20px; margin-right:4px;
                border-top-left-radius:8px; border-top-right-radius:8px; font-size:13px; font-weight:900; }
            QTabBar::tab:selected { background:#0F766E; color:white; }
        """)
        self.pdf_tab = QtWidgets.QWidget()
        self.excel_tab = QtWidgets.QWidget()
        self._build_modern_export_tab(self.pdf_tab, "pdf")
        self._build_modern_export_tab(self.excel_tab, "excel")
        self.tab_widget.addTab(self.pdf_tab, "")
        self.tab_widget.addTab(self.excel_tab, "")
        self.tab_widget.tabBar().setTabButton(0, QtWidgets.QTabBar.LeftSide, self._tab_label("PDF", "PDF Export", "#DC2626"))
        self.tab_widget.tabBar().setTabButton(1, QtWidgets.QTabBar.LeftSide, self._tab_label("XLS", "Excel Export", "#059669"))
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        layout.addWidget(self.tab_widget, 1)

        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        buttons = QtWidgets.QHBoxLayout()
        buttons.addStretch()
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(110, 40)
        self.cancel_btn.setStyleSheet(self._modern_button_style(False))
        self.cancel_btn.clicked.connect(self.reject)
        self.export_btn = QtWidgets.QPushButton("Export PDF")
        self.export_btn.setFixedSize(140, 40)
        self.export_btn.setStyleSheet(
            "QPushButton { background:#0F766E; color:white; border:none; border-radius:8px; "
            "font-size:13px; font-weight:900; } QPushButton:hover { background:#115E59; }"
        )
        self.export_btn.clicked.connect(self.start_export)
        buttons.addWidget(self.cancel_btn)
        buttons.addWidget(self.export_btn)
        layout.addLayout(buttons)

        self.export_type = "pdf"
        self.export_range = "all"
        self.excel_export_range = "all"
        self.update_preview()
        self.update_excel_preview()

    def _tab_label(self, badge_text, label_text, color):
        widget = QtWidgets.QWidget()
        widget.setAttribute(QtCore.Qt.WA_TransparentForMouseEvents, True)
        layout = QtWidgets.QHBoxLayout(widget)
        layout.setContentsMargins(12, 0, 12, 0)
        layout.setSpacing(8)

        badge = QtWidgets.QLabel(badge_text)
        badge.setFixedSize(34, 22)
        badge.setAlignment(QtCore.Qt.AlignCenter)
        badge.setStyleSheet(f"""
            QLabel {{
                background: #FFFFFF;
                color: {color};
                border: 1px solid {color};
                border-radius: 5px;
                font-size: 10px;
                font-weight: 900;
            }}
        """)

        text = QtWidgets.QLabel(label_text)
        text.setStyleSheet("QLabel { background: transparent; border: none; font-size: 13px; font-weight: 900; color: inherit; }")
        layout.addWidget(badge)
        layout.addWidget(text)
        return widget

    def _build_modern_export_tab(self, host, export_type):
        is_excel = export_type == "excel"
        prefix = "excel_" if is_excel else ""
        layout = QtWidgets.QVBoxLayout(host)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        card = QtWidgets.QFrame()
        card.setStyleSheet("QFrame { background:white; border:1px solid #E2E8F0; border-radius:10px; }")
        card_layout = QtWidgets.QVBoxLayout(card)
        card_layout.setContentsMargins(18, 14, 18, 16)
        card_layout.setSpacing(12)
        heading = QtWidgets.QLabel("1   Choose Filters")
        heading.setStyleSheet("font-size:13px; font-weight:900; color:#0F172A; border:none;")
        card_layout.addWidget(heading)

        def field(label_text, widget):
            wrap = QtWidgets.QVBoxLayout()
            label = QtWidgets.QLabel(label_text)
            label.setStyleSheet("font-size:11px; font-weight:800; color:#334155; border:none;")
            wrap.addWidget(label)
            wrap.addWidget(widget)
            return wrap

        top = QtWidgets.QHBoxLayout()
        top.setSpacing(14)
        from_date = QtWidgets.QDateEdit()
        from_date.setCalendarPopup(True)
        from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        from_date.setDisplayFormat("MM/dd/yyyy")
        from_date.setFixedSize(148, 36)
        from_date.setStyleSheet(self._modern_input_style())
        from_date.wheelEvent = lambda e: e.ignore()
        from_date.stepBy = lambda x: None
        to_date = QtWidgets.QDateEdit()
        to_date.setCalendarPopup(True)
        to_date.setDate(QtCore.QDate.currentDate())
        to_date.setDisplayFormat("MM/dd/yyyy")
        to_date.setFixedSize(148, 36)
        to_date.setStyleSheet(self._modern_input_style())
        to_date.wheelEvent = lambda e: e.ignore()
        to_date.stepBy = lambda x: None
        dates = QtWidgets.QWidget()
        dates_layout = QtWidgets.QHBoxLayout(dates)
        dates_layout.setContentsMargins(0, 3, 0, 3)
        dates_layout.setSpacing(8)
        dates_layout.addWidget(from_date)
        dates_layout.addWidget(QtWidgets.QLabel("->"))
        dates_layout.addWidget(to_date)
        date_range_wrap = QtWidgets.QWidget()
        drw_lay = QtWidgets.QVBoxLayout(date_range_wrap)
        drw_lay.setContentsMargins(0, 0, 0, 0)
        drw_lay.setSpacing(4)
        _dr_lbl = QtWidgets.QLabel("Date Range")
        _dr_lbl.setStyleSheet("font-size:11px; font-weight:800; color:#334155; border:none;")
        drw_lay.addWidget(_dr_lbl)
        drw_lay.addWidget(dates)
        date_range_wrap.setVisible(False)

        quick = QtWidgets.QComboBox()
        for label, value in [
            ("All Invoices", "all"), ("Today", "today"), ("Last 7 Days", "last_7"),
            ("Last 30 Days", "last_30"), ("This Month", "this_month"), ("This Year", "this_year"),
            ("Custom Date Range", "date_range"), ("Select Month", "month"), ("Select Year", "year"),
        ]:
            quick.addItem(label, value)
        quick.setFixedSize(220, 36)
        quick.setStyleSheet(self._modern_combo_style())
        top.addLayout(field("Quick Range", quick))

        client_combo = QtWidgets.QComboBox()
        client_combo.addItem("All Clients")
        client_combo.addItems(self.all_clients)
        client_combo.setFixedSize(240, 36)
        client_combo.setStyleSheet(self._modern_combo_style())
        top.addLayout(field("Client", client_combo))
        top.addStretch()
        card_layout.addLayout(top)
        clear_btn = QtWidgets.QPushButton("↻")
        clear_btn.setText("Clear All")
        clear_btn.setFixedSize(96, 36)
        clear_btn.setStyleSheet(self._modern_button_style(False))
        action_row = QtWidgets.QHBoxLayout()
        action_row.setContentsMargins(0, 4, 0, 0)
        action_row.setSpacing(8)
        action_row.addStretch()
        action_row.addWidget(clear_btn)
        card_layout.addLayout(action_row)

        hidden = QtWidgets.QWidget()
        hidden.setStyleSheet("QWidget { background: #ffffff; border: none; }")
        hidden_layout = QtWidgets.QGridLayout(hidden)
        hidden_layout.setContentsMargins(0, 12, 0, 0)
        hidden_layout.setHorizontalSpacing(10)
        hidden_layout.setVerticalSpacing(8)

        def small_label(text):
            label = QtWidgets.QLabel(text)
            label.setFixedWidth(54)
            label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            label.setStyleSheet("border:none; color:#334155; font-size:12px; font-weight:800;")
            return label

        month_group = QtWidgets.QWidget()
        month_grid = QtWidgets.QGridLayout(month_group)
        month_grid.setContentsMargins(0, 0, 0, 0)
        month_grid.setHorizontalSpacing(10)
        month_combo = QtWidgets.QComboBox()
        month_combo.addItems(["January", "February", "March", "April", "May", "June",
                              "July", "August", "September", "October", "November", "December"])
        month_combo.setCurrentIndex(datetime.now().month - 1)
        month_combo.setFixedSize(180, 36)
        month_combo.setStyleSheet(self._modern_combo_style())
        year_month_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        year_month_edit.setFixedSize(92, 36)
        year_month_edit.setReadOnly(True)
        year_month_edit.setStyleSheet(self._modern_input_style())
        select_month = QtWidgets.QPushButton("Select")
        select_month.setFixedSize(76, 36)
        select_month.setStyleSheet(self._modern_button_style(False))
        select_month.clicked.connect(lambda: self.show_year_popup_for_month_excel() if is_excel else self.show_year_popup_for_month())
        month_grid.addWidget(small_label("Month"), 0, 0)
        month_grid.addWidget(month_combo, 0, 1)
        month_grid.addWidget(small_label("Year"), 1, 0)
        month_grid.addWidget(year_month_edit, 1, 1)
        month_grid.addWidget(select_month, 1, 2)

        year_group = QtWidgets.QWidget()
        year_grid = QtWidgets.QGridLayout(year_group)
        year_grid.setContentsMargins(0, 0, 0, 0)
        year_grid.setHorizontalSpacing(10)
        year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        year_edit.setFixedSize(92, 36)
        year_edit.setReadOnly(True)
        year_edit.setStyleSheet(self._modern_input_style())
        select_year = QtWidgets.QPushButton("Select")
        select_year.setFixedSize(76, 36)
        select_year.setStyleSheet(self._modern_button_style(False))
        select_year.clicked.connect(lambda: self.show_year_popup_excel() if is_excel else self.show_year_popup())
        year_grid.addWidget(small_label("Year"), 0, 0)
        year_grid.addWidget(year_edit, 0, 1)
        year_grid.addWidget(select_year, 0, 2)

        hidden_layout.addWidget(date_range_wrap, 0, 0)
        hidden_layout.addWidget(month_group, 0, 0)
        hidden_layout.addWidget(year_group, 0, 0)
        hidden_layout.setColumnStretch(1, 1)
        hidden.setVisible(False)
        date_range_wrap.setVisible(False)
        month_group.setVisible(False)
        year_group.setVisible(False)
        card_layout.addWidget(hidden)
        layout.addWidget(card)

        summary = QtWidgets.QFrame()
        summary.setStyleSheet("QFrame { background:white; border:1px solid #E2E8F0; border-radius:10px; }")
        summary_layout = QtWidgets.QVBoxLayout(summary)
        summary_layout.setContentsMargins(18, 14, 18, 16)
        summary_title = QtWidgets.QLabel("2   Export Summary")
        summary_title.setStyleSheet("font-size:13px; font-weight:900; color:#0F172A; border:none;")
        preview = QtWidgets.QLabel()
        preview.setWordWrap(True)
        preview.setStyleSheet(
            "QLabel { background:#EFF6FF; color:#2563EB; border:1px solid #DBEAFE; "
            "border-radius:8px; font-size:12px; font-weight:700; padding:12px 14px; }"
        )
        summary_layout.addWidget(summary_title)
        summary_layout.addWidget(preview)
        layout.addWidget(summary)
        layout.addStretch()

        setattr(self, f"{prefix}from_date", from_date)
        setattr(self, f"{prefix}to_date", to_date)
        setattr(self, f"{prefix}month_combo", month_combo)
        setattr(self, f"{prefix}year_edit_month", year_month_edit)
        setattr(self, f"{prefix}year_edit", year_edit)
        setattr(self, f"{prefix}client_combo", client_combo)
        setattr(self, f"{prefix}date_selection_container", hidden)
        setattr(self, f"{prefix}month_group", month_group)
        setattr(self, f"{prefix}year_group", year_group)
        if is_excel:
            self.excel_preview_label = preview
        else:
            self.preview_label = preview

        def _lock_combo(cb):
            cb.wheelEvent = lambda e: e.ignore()
            def _kp(event, _cb=cb):
                if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) \
                        and not _cb.view().isVisible():
                    event.ignore()
                    return
                QtWidgets.QComboBox.keyPressEvent(_cb, event)
            cb.keyPressEvent = _kp
            cb.currentIndexChanged.connect(
                lambda: QtCore.QTimer.singleShot(0, cb.clearFocus))

        _lock_combo(quick)
        _lock_combo(client_combo)
        _lock_combo(month_combo)

        def sync_hidden(selected=None, force_visible=None):
            selected = selected or quick.currentData()
            if selected == "date_range":
                hidden.setVisible(True)
                date_range_wrap.setVisible(True)
                month_group.setVisible(False)
                year_group.setVisible(False)
            elif selected in ("month", "this_month"):
                hidden.setVisible(True)
                date_range_wrap.setVisible(False)
                month_group.setVisible(True)
                year_group.setVisible(False)
            elif selected in ("year", "this_year"):
                hidden.setVisible(True)
                date_range_wrap.setVisible(False)
                month_group.setVisible(False)
                year_group.setVisible(True)
            else:
                hidden.setVisible(False)
                date_range_wrap.setVisible(False)
                month_group.setVisible(False)
                year_group.setVisible(False)
            any_open = hidden.isVisible()

        def apply_quick():
            selected = quick.currentData()
            today = QtCore.QDate.currentDate()
            if selected == "today":
                from_date.setDate(today); to_date.setDate(today); selected = "date_range"
            elif selected == "last_7":
                from_date.setDate(today.addDays(-6)); to_date.setDate(today); selected = "date_range"
            elif selected == "last_30":
                from_date.setDate(today.addDays(-29)); to_date.setDate(today); selected = "date_range"
            elif selected == "this_month":
                month_combo.setCurrentIndex(today.month() - 1); year_month_edit.setText(str(today.year()))
            elif selected == "this_year":
                year_edit.setText(str(today.year()))
            sync_hidden(selected)
            self.on_excel_range_changed(selected) if is_excel else self.on_range_changed(selected)

        def toggle_more():
            cur = quick.currentData()
            if cur == "date_range":
                sync_hidden(cur) if not date_range_wrap.isVisible() else sync_hidden("all")
            else:
                sync_hidden(cur) if not hidden.isVisible() else sync_hidden("all")

        def clear_filters():
            quick.setCurrentIndex(0)
            from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
            to_date.setDate(QtCore.QDate.currentDate())
            client_combo.setCurrentText("All Clients")
            month_combo.setCurrentIndex(datetime.now().month - 1)
            year_month_edit.setText(str(datetime.now().year))
            year_edit.setText(str(datetime.now().year))
            sync_hidden("all")
            self.on_excel_range_changed("all") if is_excel else self.on_range_changed("all")

        quick.currentIndexChanged.connect(apply_quick)
        from_date.dateChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        to_date.dateChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        month_combo.currentTextChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        year_month_edit.textChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        year_edit.textChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        client_combo.currentTextChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        clear_btn.clicked.connect(clear_filters)

    def on_range_changed(self, range_type):
        self.export_range = range_type
        self.update_preview()

    def on_excel_range_changed(self, range_type):
        self.excel_export_range = range_type
        self.update_excel_preview()

    def _selected_client(self, export_type):
        combo = getattr(self, "excel_client_combo" if export_type == "excel" else "client_combo", None)
        return combo.currentText() if combo else "All Clients"

    def update_preview(self):
        client = self._selected_client("pdf")
        target = "all clients" if client == "All Clients" else client
        if self.export_range == "all":
            text = f"Will export all invoices for {target} as PDF."
        elif self.export_range == "date_range":
            text = f"Will export invoices from {self.from_date.date().toString('MM/dd/yyyy')} to {self.to_date.date().toString('MM/dd/yyyy')} for {target} as PDF."
        elif self.export_range == "month":
            text = f"Will export invoices for {self.month_combo.currentText()} {self.year_edit_month.text()} for {target} as PDF."
        else:
            text = f"Will export invoices for {self.year_edit.text()} for {target} as PDF."
        self.preview_label.setText(text)

    def update_excel_preview(self):
        range_type = getattr(self, "excel_export_range", "all")
        client = self._selected_client("excel")
        target = "all clients" if client == "All Clients" else client
        if range_type == "all":
            text = f"Will export all invoices for {target} as Excel."
        elif range_type == "date_range":
            text = f"Will export invoices from {self.excel_from_date.date().toString('MM/dd/yyyy')} to {self.excel_to_date.date().toString('MM/dd/yyyy')} for {target} as Excel."
        elif range_type == "month":
            text = f"Will export invoices for {self.excel_month_combo.currentText()} {self.excel_year_edit_month.text()} for {target} as Excel."
        else:
            text = f"Will export invoices for {self.excel_year_edit.text()} for {target} as Excel."
        self.excel_preview_label.setText(text)

    def get_export_parameters(self):
        params = super().get_export_parameters()
        if params:
            params["client"] = self._selected_client(params["type"])
        return params


class ReceivedDateDialog(QtWidgets.QDialog):
    """Dialog to capture received date when marking invoice as paid - OPTIMIZED"""
    
    def __init__(self, invoice, parent=None):
        super().__init__(parent)
        self.invoice = invoice
        self.setWindowTitle(f"Payment Received - {invoice.invoice_number}")
        self.setModal(True)
        self.setFixedSize(400, 250)
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        header = QtWidgets.QLabel(f"💰 Payment Received")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #27ae60;
                padding: 10px;
                background: #e8f5e9;
                border-radius: 8px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        info_text = f"""
        <b>Invoice:</b> {self.invoice.invoice_number}<br>
        <b>Client:</b> {self.invoice.client_name}<br>
        <b>Amount:</b> {Currency.format(self.invoice.total)}<br>
        <b>Due Date:</b> {self.invoice.due_date}
        """
        info_label = QtWidgets.QLabel(info_text)
        info_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                padding: 10px;
                background: #f8f9fa;
                border-radius: 5px;
            }
        """)
        info_label.setTextFormat(QtCore.Qt.RichText)
        layout.addWidget(info_label)
        
        date_layout = QtWidgets.QHBoxLayout()
        date_label = QtWidgets.QLabel("Payment Received Date:")
        date_label.setStyleSheet("font-weight: bold; font-size: 12px;")
        
        self.date_edit = QtWidgets.QDateEdit()
        self.date_edit.setDate(QtCore.QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("MM-dd-yyyy")
        # Cap at today — received date must not be in the future
        self.date_edit.setMaximumDate(QtCore.QDate.currentDate())
        # Style calendar so future (disabled) dates appear grayed out
        _cal = self.date_edit.calendarWidget()
        if _cal:
            _cal.setStyleSheet("""
                QCalendarWidget QAbstractItemView {
                    selection-background-color: #00756f;
                    selection-color: white;
                }
                QCalendarWidget QAbstractItemView:disabled {
                    color: #c8c8c8;
                }
            """)
            _gray_fmt = QtGui.QTextCharFormat()
            _gray_fmt.setForeground(QtGui.QBrush(QtGui.QColor("#c8c8c8")))
            _today = QtCore.QDate.currentDate()
            _d = _today.addDays(1)
            _end = _today.addYears(1)
            while _d <= _end:
                _cal.setDateTextFormat(_d, _gray_fmt)
                _d = _d.addDays(1)
        # Disable scroll wheel
        self.date_edit.wheelEvent = lambda event: None
        # Disable arrow key changes
        def keyPressEvent(event, original=self.date_edit.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        self.date_edit.keyPressEvent = keyPressEvent
        self.date_edit.stepBy = lambda x: None
        self.date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        self.date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                font-size: 12px;
            }
        """)
        
        date_layout.addWidget(date_label)
        date_layout.addWidget(self.date_edit)
        layout.addLayout(date_layout)
        
        note_label = QtWidgets.QLabel("Note: This date will be used to track when the payment was received.")
        note_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                font-style: italic;
            }
        """)
        note_label.setWordWrap(True)
        layout.addWidget(note_label)
        
        btn_layout = QtWidgets.QHBoxLayout()
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                border: none;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #7f8c8d; }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.save_btn = QtWidgets.QPushButton("✅ Mark as Paid")
        self.save_btn.setDefault(True)  # Allow Enter key to trigger
        self.save_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                border: none;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #2ecc71; }
        """)
        self.save_btn.clicked.connect(self.accept)
        
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addWidget(self.save_btn)
        layout.addLayout(btn_layout)
        
        # Set focus to save button for quick Enter key press
        self.save_btn.setFocus()
    
    def get_received_date(self):
        """Always return received date in canonical MM-dd-YYYY format."""
        date = self.date_edit.date()
        return _normalize_date(date.toString("MM-dd-yyyy"))
class InvoiceHistoryViewWidget(QtWidgets.QWidget):
    """Main invoice history view with date range filter"""
    back_clicked = pyqtSignal()
    
    def __init__(self, client_name: str, compact: bool = False):
        super().__init__()
        self.client_name = client_name
        self._compact = compact
        self.invoices = []
        self.filtered_invoices = []
        self.current_date_filter = None
        self.current_search_text = ""
        self.projects_data = {}
        self.status_cache = {}
        self._filtering = False
        self._sort_ascending = False  # True when a specific date filter is active
        self._ih_page = 1
        self._ih_per_page = 10
        self._ih_all_displayed = []
        self.init_ui()
        self.load_projects_data()
        self.load_status_cache()
    
    def init_ui(self):
        outer = QtWidgets.QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self._scroll_area = QtWidgets.QScrollArea()
        self._scroll_area.setWidgetResizable(True)
        self._scroll_area.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self._scroll_area.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self._scroll_area.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        outer.addWidget(self._scroll_area)

        _content = QtWidgets.QWidget()
        _content.setStyleSheet("background: transparent;")
        _content.setMinimumWidth(1100)
        self._scroll_area.setWidget(_content)

        layout = QtWidgets.QVBoxLayout(_content)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        header_widget = QtWidgets.QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: #0f3b57;
                border-radius: 10px;
            }
        """)
        header_layout = QtWidgets.QHBoxLayout(header_widget)
        header_layout.setContentsMargins(15, 10, 15, 10)
        
        self.back_btn = QtWidgets.QPushButton("🔙")
        self.back_btn.setStyleSheet("""
            QPushButton {
                background-color: rgba(255,255,255,0.12);
                color: white;
                border: 1px solid rgba(255,255,255,0.25);
                border-radius: 7px;
                font-size: 13px;
                font-weight: bold;
                padding: 8px 14px;
            }
            QPushButton:hover {
                color: #f0f0f0;
                background-color: rgba(255,255,255,0.1);
                border-radius: 5px;
            }
        """)
        self.back_btn.setText("Back")
        self.back_btn.clicked.connect(self.back_clicked)
        self.back_btn.setFixedSize(92, 42)
        
        title = QtWidgets.QLabel(f"{self.client_name.upper()} - Invoice History")
        title_font = QtGui.QFont("Inter", 18, QtGui.QFont.Bold)
        title.setFont(title_font)
        title.setStyleSheet("""
            QLabel {
                font-size: 22px;
                font-weight: 800;
                color: white;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                padding: 8px;
            }
        """)
        title.setAlignment(QtCore.Qt.AlignCenter)
        
        self.pdf_export_btn = QtWidgets.QPushButton("📤 Export")
        self.pdf_export_btn.setFixedSize(120, 40)
        self.pdf_export_btn.setText("Export")
        self.pdf_export_btn.setFixedSize(130, 42)
        self.pdf_export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e67e22, stop:1 #f39c12);
                color: white;
                border: none;
                border-radius: 8px;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                font-weight: 800;
                font-size: 13px;
                padding: 8px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d35400, stop:1 #e67e22);
                border: 2px solid #f39c12;
            }
        """)
        self.pdf_export_btn.clicked.connect(self.open_pdf_export_dialog)
        
        self.refresh_btn = QtWidgets.QPushButton("⟳ Refresh")
        self.refresh_btn.setFixedSize(110, 42)
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background: rgba(255,255,255,0.15);
                color: white;
                border: 1px solid rgba(255,255,255,0.35);
                border-radius: 8px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-weight: 800;
                font-size: 13px;
                padding: 8px;
            }
            QPushButton:hover {
                background: rgba(255,255,255,0.25);
                border-color: white;
            }
        """)
        self.refresh_btn.clicked.connect(self._reload_invoices)

        header_layout.addWidget(self.back_btn)
        header_layout.addStretch()
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(self.refresh_btn)
        header_layout.addSpacing(8)
        header_layout.addWidget(self.pdf_export_btn)
        layout.addWidget(header_widget)
        
        controls_card = QtWidgets.QFrame()
        controls_card.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #d8e2ec;
                border-radius: 9px;
            }
        """)
        controls_layout = QtWidgets.QHBoxLayout(controls_card)
        controls_layout.setContentsMargins(14, 20, 14, 20)
        controls_layout.setSpacing(12)
        
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.date_range_changed.connect(self.apply_date_range_filter)
        self.date_range_widget.date_range_cleared.connect(self.clear_date_range_filter)
        # Use the proper signal so on_search_changed → search_changed → apply_search_filter
        self.date_range_widget.search_changed.connect(self.apply_search_filter)
        # Belt-and-suspenders: also connect textChanged directly in case signal chain breaks
        self.date_range_widget.search_bar.textChanged.connect(self.apply_search_filter)
        controls_layout.addWidget(self.date_range_widget)
        controls_layout.addStretch()
        
        quick_filter_layout = QtWidgets.QHBoxLayout()
        quick_filters_label = QtWidgets.QLabel("Quick Filters:")
        quick_filters_label.setStyleSheet('font-family: "Inter", "Segoe UI", Arial, sans-serif; font-size: 13px; font-weight: 800; color: #0f172a; padding: 5px;')
        quick_filter_layout.addWidget(quick_filters_label)
        
        self.last_7_days_btn = QtWidgets.QPushButton("Last 7 Days")
        self.last_30_days_btn = QtWidgets.QPushButton("Last 30 Days")
        self.this_month_btn = QtWidgets.QPushButton("This Month")
        self.this_year_btn = QtWidgets.QPushButton("This Year")
        self.all_time_btn = QtWidgets.QPushButton("All Time")
        
        quick_btn_style = """
            QPushButton {
                background-color: #f8fbfd;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                padding: 0 14px;
                border-radius: 7px;
                font-size: 12px;
                font-weight: 800;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
            QPushButton:hover { border-color: #00756f; color: #00756f; }
            QPushButton:pressed { background-color: #eefaf8; }
        """
        
        for btn in [self.last_7_days_btn, self.last_30_days_btn, self.this_month_btn, self.this_year_btn, self.all_time_btn]:
            btn.setStyleSheet(quick_btn_style)
            btn.setFixedHeight(38)
        
        self.last_7_days_btn.clicked.connect(lambda: self.apply_quick_filter(7))
        self.last_30_days_btn.clicked.connect(lambda: self.apply_quick_filter(30))
        self.this_month_btn.clicked.connect(self.apply_this_month_filter)
        self.this_year_btn.clicked.connect(self.apply_this_year_filter)
        self.all_time_btn.clicked.connect(self.apply_all_time_filter)
        
        quick_filter_layout.addWidget(self.last_7_days_btn)
        quick_filter_layout.addWidget(self.last_30_days_btn)
        quick_filter_layout.addWidget(self.this_month_btn)
        quick_filter_layout.addWidget(self.this_year_btn)
        quick_filter_layout.addWidget(self.all_time_btn)
        controls_layout.addLayout(quick_filter_layout)
        layout.addWidget(controls_card)
        
        self.stats_widget = QtWidgets.QFrame()
        self.stats_widget.setStyleSheet("QFrame { background: transparent; border: none; }")
        self.stats_layout = QtWidgets.QHBoxLayout(self.stats_widget)
        self.stats_layout.setContentsMargins(16, 14, 16, 4 if self._compact else 6)
        self.stats_layout.setSpacing(12)
        self.update_stats()
        layout.addWidget(self.stats_widget)
        layout.setSpacing(8)

        table_card = QtWidgets.QFrame()
        table_card.setStyleSheet("QFrame { background: transparent; border: none; }")
        table_layout = QtWidgets.QVBoxLayout(table_card)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(0)

        table_title = QtWidgets.QLabel("Invoice Records")
        table_title.setStyleSheet("""
            QLabel {
                color: #0f172a;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                font-size: 15px;
                font-weight: 900;
                padding: 14px 16px;
                border-bottom: 1px solid #e5edf5;
            }
        """)
        table_layout.addWidget(table_title)
        
        self.invoice_table = QtWidgets.QTableWidget()
        self.invoice_table.setColumnCount(11)
        self.invoice_table.setHorizontalHeaderLabels([
            "Date", "Invoice Number", "Project Name", "Total price", "Tax", "Down Payment", "Total Due", "Due Date", "Status", "Received Date", "Actions"
        ])
        self.invoice_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.invoice_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.invoice_table.setAlternatingRowColors(True)
        self.invoice_table.setSortingEnabled(False)
        self.invoice_table.setFont(QtGui.QFont("Inter", 9))
        self.invoice_table.setWordWrap(True)
        # Use custom delegate for proper text wrapping
        self.invoice_table.setItemDelegateForColumn(2, TextWrapDelegate(self))
        self.invoice_table.verticalHeader().setVisible(False)
        self.invoice_table.verticalHeader().setDefaultSectionSize(54)
        self.invoice_table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.invoice_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        
        self.invoice_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 8px 10px;
                border-bottom: 1px solid #ecf0f1;
                white-space: pre-wrap;
                word-wrap: break-word;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 12px 8px;
                border: none;
                font-weight: bold;
            }
            QToolTip {
                background-color: #ffffff;
                color: #1f2937;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 8px;
                font-size: 12px;
            }
        """)
        
        header = self.invoice_table.horizontalHeader()
        header.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.Bold))
        header.setMinimumSectionSize(110)
        header.setSectionResizeMode(0,  QtWidgets.QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(1,  QtWidgets.QHeaderView.ResizeToContents)  # Invoice Number
        header.setSectionResizeMode(2,  QtWidgets.QHeaderView.Fixed)             # Project Name — FIXED width to force wrapping
        header.setSectionResizeMode(3,  QtWidgets.QHeaderView.ResizeToContents)  # Total Price
        header.setSectionResizeMode(4,  QtWidgets.QHeaderView.ResizeToContents)  # Tax
        header.setSectionResizeMode(5,  QtWidgets.QHeaderView.ResizeToContents)  # Down Payment
        header.setSectionResizeMode(6,  QtWidgets.QHeaderView.ResizeToContents)  # Total Due
        header.setSectionResizeMode(7,  QtWidgets.QHeaderView.ResizeToContents)  # Due Date
        header.setSectionResizeMode(8,  QtWidgets.QHeaderView.Fixed)    # Status
        header.setSectionResizeMode(9,  QtWidgets.QHeaderView.ResizeToContents)  # Received Date
        header.setSectionResizeMode(10, QtWidgets.QHeaderView.Fixed)    # Actions
        self.invoice_table.setColumnWidth(2,  300)  # Project Name — 45 chars per line for wrapping
        self.invoice_table.setColumnWidth(8,  150)
        self.invoice_table.setColumnWidth(10, 150)
        # Keep setWordWrap(True) — don't override it to False
        self.invoice_table.verticalHeader().setDefaultSectionSize(120)  # Increased for 3+ wrapped lines
        self.invoice_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.invoice_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.invoice_table.customContextMenuRequested.connect(self._on_invoice_table_context_menu)

        table_layout.addWidget(self.invoice_table)

        # ── Pagination bar (matches project tab style) ──────────────
        _pg_s = """
            QPushButton {
                background: #ffffff; color: #334155;
                border: 1px solid #e2e8f0; border-radius: 6px;
                font-size: 12px; font-weight: 700;
                min-width: 32px; min-height: 28px; padding: 0 8px;
            }
            QPushButton:hover { background: #f1f5f9; border-color: #cbd5e1; }
            QPushButton:disabled { color: #cbd5e1; }
        """
        pg_frame = QtWidgets.QFrame()
        pg_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        pg_hbox = QtWidgets.QHBoxLayout(pg_frame)
        pg_hbox.setContentsMargins(4, 6, 4, 6)
        pg_hbox.setSpacing(6)

        self._ih_info_lbl = QtWidgets.QLabel("")
        self._ih_info_lbl.setStyleSheet(
            "color: #94a3b8; font-size: 11px; font-weight: 600;"
            " background: transparent; border: none;")
        pg_hbox.addWidget(self._ih_info_lbl)
        pg_hbox.addStretch()

        self._ih_prev_btn = QtWidgets.QPushButton("‹")
        self._ih_prev_btn.setStyleSheet(_pg_s)
        self._ih_prev_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._ih_prev_btn.clicked.connect(self._ih_go_prev)
        pg_hbox.addWidget(self._ih_prev_btn)

        self._ih_page_btns = QtWidgets.QHBoxLayout()
        self._ih_page_btns.setSpacing(4)
        pg_hbox.addLayout(self._ih_page_btns)

        self._ih_next_btn = QtWidgets.QPushButton("›")
        self._ih_next_btn.setStyleSheet(_pg_s)
        self._ih_next_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._ih_next_btn.clicked.connect(self._ih_go_next)
        pg_hbox.addWidget(self._ih_next_btn)

        self._ih_pg_style = _pg_s
        table_layout.addWidget(pg_frame)

        layout.addWidget(table_card)
        layout.addStretch()
        self.load_all_invoices()
    
    def sync_invoice_to_revenue(self, invoice: Invoice):
        try:
            if not FIREBASE_AVAILABLE:
                return
            
            from firebase_admin import db
            revenue_ref = db.reference('revenue')
            all_revenue = revenue_ref.get()
            
            if not all_revenue:
                return
            
            revenue_to_update = None
            revenue_id = None
            
            for rev_id, revenue in all_revenue.items():
                if revenue and revenue.get('is_invoice') and revenue.get('invoice_number') == invoice.invoice_number:
                    revenue_to_update = revenue
                    revenue_id = rev_id
                    break
            
            if not revenue_to_update:
                return
            
            changed = False
            updates = {}
            
            new_due_date = getattr(invoice, 'due_date', 'N/A')
            if new_due_date != revenue_to_update.get('due_date', 'N/A'):
                updates['due_date'] = new_due_date
                changed = True
            
            new_status = invoice.status if hasattr(invoice, 'status') else 'Pending'
            if new_status != revenue_to_update.get('status', 'Pending'):
                updates['status'] = new_status
                changed = True
            
            new_received_date = getattr(invoice, 'received_date', 'N/A')
            if not new_received_date or new_received_date == '':
                new_received_date = 'N/A'
            
            if new_received_date != revenue_to_update.get('received_date', 'N/A'):
                updates['received_date'] = new_received_date
                changed = True
                if new_status == "Paid" and new_received_date not in ('N/A', 'N\\A', ''):
                    try:
                        updates['year'] = datetime.strptime(
                            _normalize_date(new_received_date), "%m-%d-%Y"
                        ).year
                    except Exception:
                        pass
            
            if changed:
                updates['updated_at'] = datetime.now(timezone.utc).isoformat()
                revenue_ref.child(revenue_id).update(updates)
                self.refresh_balance_sheet()
        except Exception as e:
            _log.warning("Error syncing invoice to revenue: %s", e)
    
    def refresh_balance_sheet(self):
        try:
            main_window = self.window()
            if hasattr(main_window, 'balance_sheet_tab'):
                balance_tab = main_window.balance_sheet_tab
                if hasattr(balance_tab, 'refresh_invoice_revenues'):
                    balance_tab.refresh_invoice_revenues()
                else:
                    balance_tab.load_all_financial_data()
                    balance_tab.update_annual_summary()
                    balance_tab.on_category_changed(balance_tab.current_category)
                    balance_tab.update_stats_cards()
        except Exception as e:
            _log.warning("Error refreshing balance sheet: %s", e)
    
    def highlight_selected_quick_filter(self, selected_button):
        default_style = """
            QPushButton {
                background-color: #f8fbfd;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                padding: 0 14px;
                border-radius: 7px;
                font-size: 12px;
                font-weight: 800;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
            QPushButton:hover { border-color: #00756f; color: #00756f; }
        """
        selected_style = """
            QPushButton {
                background-color: #00756f;
                color: white;
                border: 1px solid #00756f;
                padding: 0 14px;
                border-radius: 7px;
                font-size: 12px;
                font-weight: 800;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
            QPushButton:hover { background-color: #00645f; }
        """
        
        for btn in [self.last_7_days_btn, self.last_30_days_btn, self.this_month_btn, self.this_year_btn, self.all_time_btn]:
            btn.setStyleSheet(default_style)
        selected_button.setStyleSheet(selected_style)
    
    def get_invoice_parsed_date(self, invoice):
        if hasattr(invoice, "_parsed_date"):
            return invoice._parsed_date
        parsed = self.parse_invoice_date(invoice)
        invoice._parsed_date = parsed
        return parsed
    
    def perform_excel_export(self, export_params):
        try:
            invoices_to_export = []
            for invoice, json_file in self.invoices:
                try:
                    date_str = getattr(invoice, 'date', '') or ''
                    invoice_date = None
                    for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                        try:
                            invoice_date = datetime.strptime(date_str, _fmt)
                            break
                        except ValueError:
                            continue
                    if invoice_date is None:
                        continue
                    
                    include_invoice = False
                    if export_params["range"] == "all":
                        include_invoice = True
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        invoice_date_only = invoice_date.date()
                        from_date_only = from_date.date() if isinstance(from_date, datetime) else from_date
                        to_date_only = to_date.date() if isinstance(to_date, datetime) else to_date
                        if from_date_only <= invoice_date_only <= to_date_only:
                            include_invoice = True
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if invoice_date.month == month and invoice_date.year == year:
                            include_invoice = True
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if invoice_date.year == year:
                            include_invoice = True
                    
                    if include_invoice:
                        invoices_to_export.append((invoice, json_file))
                except Exception as e:
                    continue

            def _excel_date_key(entry):
                date_str = getattr(entry[0], 'date', '') or ''
                for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(date_str, _fmt)
                    except ValueError:
                        continue
                return datetime.min
            invoices_to_export.sort(key=_excel_date_key)

            if not invoices_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", "No invoices found matching the selected criteria.")
                return

            self.generate_combined_excel(invoices_to_export, export_params)
        except Exception as e:
            _log.warning("Error performing Excel export: %s", e)
    
    def generate_combined_excel(self, invoices, export_params):
        try:
            export_dir = Path.home() / "Downloads" / "Invoice_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"{self.client_name}_All_Invoices_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"{self.client_name}_Invoices_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"{self.client_name}_Invoices_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"{self.client_name}_Invoices_{year}.xlsx"
            
            excel_path = export_dir / filename
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoice History"
            
            ws.merge_cells('A1:J1')
            ws['A1'] = f"{Config.COMPANY.get('name', 'MABS Engineering LLC').upper()} - {self.client_name.upper()} INVOICES"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            if export_params["range"] == "all":
                export_range_text = "All Invoices"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws['A2'] = f"Period: {export_range_text}"
            ws['A2'].font = Font(bold=True)
            
            headers = ["Date", "Invoice Number", "Project Name", "Total Price", "Tax", "Down Payment", "Total Due", "Due Date", "Status", "Received Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for row, (invoice, _) in enumerate(invoices, 5):
                project_name = self.get_project_name(invoice)
                total_down_payment = sum(item.down_payment for item in invoice.items)
                status = invoice.status if hasattr(invoice, 'status') else self.get_invoice_status(invoice)
                due_date = getattr(invoice, 'due_date', 'N/A')
                if not due_date or due_date == '':
                    due_date = 'N/A'
                received_date = getattr(invoice, 'received_date', 'N/A')
                if not received_date or received_date == '':
                    received_date = 'N/A'
                
                data = [invoice.date, invoice.invoice_number, project_name, float(invoice.subtotal), float(invoice.tax_amount), float(total_down_payment), float(invoice.total), due_date, status, received_date]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    if col in [4,5,6,7]:
                        cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            column_widths = {1:14, 2:30, 3:34, 4:15, 5:12, 6:18, 7:16, 8:16, 9:20, 10:18}
            for col_idx, width in column_widths.items():
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = width
            ws.freeze_panes = "A5"
            
            wb.save(str(excel_path))
            
            if FileManager.open_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ Excel exported successfully!\n\nFile saved to: {excel_path}\nThe Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ Excel exported successfully!\n\nFile saved to: {excel_path}\nCould not open automatically. Please open manually.")
        except Exception as e:
            _log.warning("Error generating combined Excel: %s", e)
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error", f"Error generating Excel: {str(e)}")
    
    def load_status_cache(self):
        try:
            cache_file = Config.INVOICES_DIR / "status_cache.json"
            if cache_file.exists():
                self.status_cache = FileManager.load_json(cache_file, {})
            else:
                self.status_cache = {}
        except Exception as e:
            _log.warning("Error loading status cache: %s", e)
            self.status_cache = {}
    
    def save_status_cache(self):
        try:
            cache_file = Config.INVOICES_DIR / "status_cache.json"
            FileManager.save_json(cache_file, self.status_cache)
        except Exception as e:
            _log.warning("Error saving status cache: %s", e)
    
    def _on_invoice_table_context_menu(self, pos):
        """Show Copy context menu when right-clicking any row in the invoice table."""
        index = self.invoice_table.indexAt(pos)
        if not index.isValid():
            return
        row = index.row()
        inv_item = self.invoice_table.item(row, 1)  # column 1 = Invoice Number
        if inv_item is None:
            return
        invoice_number = inv_item.text().strip()
        if not invoice_number:
            return
        menu = QtWidgets.QMenu(self)
        menu.setStyleSheet("""
            QMenu { background: #ffffff; border: 1px solid #e2e8f0;
                    border-radius: 7px; padding: 4px; }
            QMenu::item { padding: 7px 20px; font-size: 13px; color: #0f172a;
                          font-family: 'Inter','Segoe UI'; border-radius: 5px; }
            QMenu::item:selected { background: #e6f6f4; color: #00756f; }
        """)
        copy_action = menu.addAction("Copy Invoice Number")
        action = menu.exec_(self.invoice_table.viewport().mapToGlobal(pos))
        if action == copy_action:
            QtWidgets.QApplication.clipboard().setText(invoice_number)

    def get_cached_status(self, invoice_number: str) -> str:
        return self.status_cache.get(invoice_number, "")
    
    def set_cached_status(self, invoice_number: str, status: str):
        self.status_cache[invoice_number] = status
        self.save_status_cache()
    
    def update_invoice_status_in_file(self, invoice_number: str, new_status: str):
        try:
            for invoice, json_file in self.invoices:
                if invoice.invoice_number == invoice_number:
                    invoice.status = new_status
                    if FIREBASE_AVAILABLE:
                        success = FirebaseManager.update_invoice_status(invoice_number, new_status)
                        if success:
                            self.sync_invoice_to_revenue(invoice)
                    break
        except Exception as e:
            _log.warning("Error updating invoice file status: %s", e)
    
    def open_pdf_export_dialog(self):
        try:
            available_dates = []
            for invoice, _ in self.invoices:
                try:
                    invoice_date = self.parse_invoice_date(invoice.date)
                    if invoice_date:
                        available_dates.append(invoice_date)
                except:
                    continue
            
            dialog = PDFExportDialog(self, self.client_name, available_dates)
            result = dialog.exec_()
            
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                export_params = dialog._export_params
                if export_params["type"] == "pdf":
                    self.perform_pdf_export(export_params)
                elif export_params["type"] == "excel":
                    self.perform_excel_export(export_params)
        except Exception as e:
            _log.warning("Error opening PDF export dialog: %s", e)
    
    def perform_pdf_export(self, export_params):
        try:
            invoices_to_export = []
            for invoice, json_file in self.invoices:
                try:
                    date_str = getattr(invoice, 'date', '') or ''
                    invoice_date = None
                    for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                        try:
                            invoice_date = datetime.strptime(date_str, _fmt)
                            break
                        except ValueError:
                            continue
                    if invoice_date is None:
                        continue
                    
                    include_invoice = False
                    if export_params["range"] == "all":
                        include_invoice = True
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        invoice_date_only = invoice_date.date()
                        from_date_only = from_date.date() if isinstance(from_date, datetime) else from_date
                        to_date_only = to_date.date() if isinstance(to_date, datetime) else to_date
                        if from_date_only <= invoice_date_only <= to_date_only:
                            include_invoice = True
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if invoice_date.month == month and invoice_date.year == year:
                            include_invoice = True
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if invoice_date.year == year:
                            include_invoice = True
                    
                    if include_invoice:
                        invoices_to_export.append((invoice, json_file))

                except Exception as e:
                    continue

            def _pdf_date_key(entry):
                date_str = getattr(entry[0], 'date', '') or ''
                for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
                    try:
                        return datetime.strptime(date_str, _fmt)
                    except ValueError:
                        continue
                return datetime.min
            invoices_to_export.sort(key=_pdf_date_key)

            if not invoices_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", "No invoices found matching the selected criteria.")
                return

            self.generate_combined_pdf(invoices_to_export, export_params)
        except Exception as e:
            _log.warning("Error performing PDF export: %s", e)
    
    def generate_combined_pdf(self, invoices, export_params):
        try:
            export_dir = Path.home() / "Downloads" / "Invoice_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"{self.client_name}_All_Invoices_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"{self.client_name}_Invoices_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"{self.client_name}_Invoices_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"{self.client_name}_Invoices_{year}.pdf"
            
            pdf_path = export_dir / filename
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=landscape(A4),
                leftMargin=0.32*inch,
                rightMargin=0.32*inch,
                topMargin=0.25*inch,
                bottomMargin=0.3*inch,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#7f8c8d'), fontName='Helvetica', alignment=2)
            info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#2c3e50'), fontName='Helvetica-Bold')
            main_title_style = ParagraphStyle('MainTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, textColor=colors.HexColor('#2c3e50'), alignment=1, fontName='Helvetica-Bold')
            stats_style = ParagraphStyle('StatsStyle', parent=styles['Normal'], fontSize=12, spaceAfter=20, textColor=colors.HexColor('#2c3e50'), alignment=1, fontName='Helvetica-Bold')
            
            generated_date = datetime.now().strftime("%m-%d-%Y")
            header_data = [['', Paragraph(f"{generated_date}", header_style)]]
            header_table = Table(header_data, colWidths=[8.8*inch, 2*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,0), 'LEFT'), ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 10)
            ]))
            elements.append(header_table)
            
            main_title = Paragraph(f"{Config.COMPANY.get('name', 'MABS Engineering LLC').upper()} INVOICE HISTORY", main_title_style)
            elements.append(main_title)
            
            total_invoices = len(invoices)
            total_amount = sum(invoice.total for invoice, _ in invoices)
            stats_text = f"Total Invoices: {total_invoices}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Total Revenue: {Currency.format(total_amount)}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            if export_params["range"] == "all":
                export_range_text = "All Invoices"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            client_export_para = Paragraph(f"{self.client_name} - {export_range_text}", ParagraphStyle('ClientExportLeft', parent=info_style, alignment=0, leftIndent=0, firstLineIndent=0, spaceBefore=4, spaceAfter=12))
            client_export_table = Table([[client_export_para]], colWidths=[10.8 * inch])
            client_export_table.setStyle(TableStyle([
                ('LEFTPADDING',(0,0),(-1,-1),0), ('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0), ('BOTTOMPADDING',(0,0),(-1,-1),0),
                ('ALIGN',(0,0),(-1,-1),'LEFT')
            ]))
            elements.append(client_export_table)
            elements.append(Spacer(1, 10))
            
            if invoices:
                header_style_center = ParagraphStyle(
                    'header_center',
                    alignment=1,
                    fontName='Helvetica-Bold',
                    fontSize=8,
                    leading=8,          # ✅ important
                    spaceBefore=-2,     # ✅ visual vertical centering
                    spaceAfter=0,
                    textColor=colors.whitesmoke
                )

                table_data = [[
                    Paragraph("Date", header_style_center),
                    Paragraph("Invoice No", header_style_center),
                    Paragraph("Project", header_style_center),
                    Paragraph("Total Price", header_style_center),
                    Paragraph("Tax", header_style_center),
                    Paragraph("Down Payment", header_style_center),
                    Paragraph("Total Due", header_style_center),
                    Paragraph("Due Date", header_style_center),
                    Paragraph("Status", header_style_center),
                    Paragraph("Received Date", header_style_center),
                ]]
                cell_style = ParagraphStyle(
                    'InvoiceClientExportCell',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=7.5,
                    leading=9,
                    alignment=1,
                    textColor=colors.HexColor('#2c3e50'),
                )
                for invoice, _ in invoices:
                    project_name = self.get_project_name(invoice)
                    total_down_payment = sum(item.down_payment for item in invoice.items)
                    status = invoice.status if hasattr(invoice, 'status') else self.get_invoice_status(invoice)
                    received_date = getattr(invoice, 'received_date', 'N/A')
                    if not received_date or received_date == '':
                        received_date = 'N/A'
                    due_date = getattr(invoice, 'due_date', 'N/A')
                    if not due_date or due_date == '':
                        due_date = 'N/A'
                    table_data.append([
                        Paragraph(str(invoice.date), cell_style),
                        Paragraph(str(invoice.invoice_number), cell_style),
                        Paragraph(str(project_name), cell_style),
                        Paragraph(Currency.format(invoice.subtotal), cell_style),
                        Paragraph(Currency.format(invoice.tax_amount), cell_style),
                        Paragraph(Currency.format(total_down_payment), cell_style),
                        Paragraph(Currency.format(invoice.total), cell_style),
                        Paragraph(str(due_date), cell_style),
                        Paragraph(str(status), cell_style),
                        Paragraph(str(received_date), cell_style),
                    ])

                invoice_table = Table(table_data, colWidths=[0.85*inch, 1.4*inch, 1.8*inch, 0.85*inch, 0.6*inch, 1.1*inch, 0.85*inch, 0.8*inch, 1.0*inch, 1.0*inch], repeatRows=1)
                invoice_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 9), ('BOTTOMPADDING', (0,0), (-1,0), 10),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#ffffff')), ('TEXTCOLOR', (0,1), (-1,-1), colors.HexColor('#2c3e50')),
                    ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,1), (-1,-1), 8),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8f9fa'), colors.white]),
                ]))
                elements.append(invoice_table)
            else:
                no_data_style = ParagraphStyle('NoData', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#7f8c8d'), alignment=1)
                elements.append(Paragraph("No invoices found for the selected criteria.", no_data_style))
            
            doc.build(elements)

            if FileManager.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ PDF exported successfully!\n\nFile saved to: {pdf_path}\nThe PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", f"✅ PDF exported successfully!\n\nFile saved to: {pdf_path}\nCould not open automatically. Please open manually.")
        except Exception as e:
            _log.warning("Error generating combined PDF: %s", e)
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", f"Error generating PDF: {str(e)}")
    
    def _parse_due_date(self, invoice):
        """Parse invoice due_date into a date object; returns None on failure."""
        raw = getattr(invoice, 'due_date', '') or ''
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass
        return None

    def _schedule_overdue_update(self, invoice_number: str):
        """Queue a non-blocking Firebase update to mark invoice as Overdue."""
        if not hasattr(self, '_overdue_queue'):
            self._overdue_queue: set = set()
            self._overdue_timer = QtCore.QTimer(self)
            self._overdue_timer.setSingleShot(True)
            self._overdue_timer.timeout.connect(self._flush_overdue_updates)
        self._overdue_queue.add(invoice_number)
        self._overdue_timer.start(1200)  # debounce — fires once after all rows painted

    def _flush_overdue_updates(self):
        """Write auto-overdue status to Firebase for all queued invoices."""
        queue = getattr(self, '_overdue_queue', set()).copy()
        if hasattr(self, '_overdue_queue'):
            self._overdue_queue.clear()
        for inv_no in queue:
            try:
                from main import FirebaseManager
                FirebaseManager.update_invoice_status(inv_no, "Overdue")
                _log.info("Auto-applied Overdue for invoice %s", inv_no)
            except Exception as exc:
                _log.warning("Could not auto-overdue %s: %s", inv_no, exc)

    def get_invoice_status(self, invoice: Invoice) -> str:
        cached_status = self.get_cached_status(invoice.invoice_number)
        firebase_status = (
            invoice.status if hasattr(invoice, 'status') and invoice.status else None)

        # Firebase is the source of truth after payment deletions/edits.
        # If the on-disk cache differs from the loaded Firebase status, treat the
        # cache as stale and update it so future calls are consistent.
        if cached_status and firebase_status and cached_status != firebase_status:
            self.set_cached_status(invoice.invoice_number, firebase_status)
            cached_status = firebase_status

        raw_status = cached_status or firebase_status

        if raw_status:
            if not cached_status:
                self.set_cached_status(invoice.invoice_number, raw_status)
            # Auto-escalate Unpaid/Pending → Overdue when due date has passed
            if raw_status in ("Unpaid", "Pending"):
                due = self._parse_due_date(invoice)
                if due and due < datetime.now().date():
                    self.set_cached_status(invoice.invoice_number, "Overdue")
                    self._schedule_overdue_update(invoice.invoice_number)
                    return "Overdue"
            return raw_status

        # No status stored — derive from due date
        try:
            due = self._parse_due_date(invoice)
            if due is None:
                return "Pending"
            return "Overdue" if due < datetime.now().date() else "Pending"
        except Exception as exc:
            _log.warning("Error determining invoice status: %s", exc)
            return "Pending"
    
    def _reload_invoices(self):
        """Refresh button: reload invoices from Firebase with visual feedback."""
        _base = """
            QPushButton {
                color: white;
                border-radius: 8px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-weight: 800;
                font-size: 13px;
                padding: 8px;
            }"""
        _style_loading = _base + "QPushButton { background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.20); }"
        _style_success = _base + "QPushButton { background: rgba(52,211,153,0.45); border: 1px solid rgba(52,211,153,0.70); }"
        _style_error   = _base + "QPushButton { background: rgba(239,68,68,0.45);  border: 1px solid rgba(239,68,68,0.70);  }"
        _style_normal  = _base + "QPushButton { background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.35); }"

        def _restore():
            self.refresh_btn.setText("⟳ Refresh")
            self.refresh_btn.setStyleSheet(_style_normal)
            self.refresh_btn.setEnabled(True)

        try:
            self.refresh_btn.setEnabled(False)
            self.refresh_btn.setText("⟳ Loading…")
            self.refresh_btn.setStyleSheet(_style_loading)
            QtWidgets.QApplication.processEvents()

            self.load_all_invoices()

            self.refresh_btn.setText("✓ Refreshed")
            self.refresh_btn.setStyleSheet(_style_success)
            QtWidgets.QApplication.processEvents()
            QtCore.QTimer.singleShot(1800, _restore)
        except Exception as e:
            _log.warning("Error reloading invoices: %s", e)
            self.refresh_btn.setText("✗ Error")
            self.refresh_btn.setStyleSheet(_style_error)
            QtCore.QTimer.singleShot(1800, _restore)

    def load_projects_data(self):
        try:
            projects = FirebaseManager.load_projects()
            self.projects_data = {}
            for project in projects:
                if isinstance(project, dict):
                    project_number = project.get("project_number", "")
                    project_name = project.get("project_name", "")
                    if project_number and project_name:
                        self.projects_data[project_number] = project_name
        except Exception as e:
            _log.warning("Error loading projects data: %s", e)
            self.projects_data = {}
    
    def get_project_name(self, invoice: Invoice) -> str:
        project_names = []
        for item in invoice.items:
            project_number = item.project_number
            if project_number and project_number in self.projects_data:
                pn = self.projects_data[project_number]
            else:
                pn = (item.description or "").strip()
            if pn and pn not in project_names:
                project_names.append(pn)
        return "\n".join(project_names) if project_names else "No Project Name"
    
    def parse_invoice_date(self, invoice_or_date):
        if hasattr(invoice_or_date, 'firebase_timestamp') and invoice_or_date.firebase_timestamp:
            try:
                if isinstance(invoice_or_date.firebase_timestamp, (int, float)):
                    return datetime.fromtimestamp(invoice_or_date.firebase_timestamp)
                elif isinstance(invoice_or_date.firebase_timestamp, str):
                    ts_str = invoice_or_date.firebase_timestamp.replace('Z', '+00:00')
                    return datetime.fromisoformat(ts_str)
            except:
                pass
        
        if hasattr(invoice_or_date, 'date'):
            date_str = invoice_or_date.date
        else:
            date_str = str(invoice_or_date)
        
        if not date_str:
            return datetime.min
        
        try:
            try:
                return datetime.strptime(date_str, "%m-%d-%Y")
            except ValueError:
                pass
            date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"]
            for date_format in date_formats:
                try:
                    return datetime.strptime(date_str, date_format)
                except ValueError:
                    continue
            return datetime.min
        except Exception as e:
            _log.info("(converted from print, see git history)")
            return datetime.min
    
    def load_all_invoices(self):
        self.invoices = []
        self.verify_firebase_connection()
        
        try:
            all_invoices_data = FirebaseManager.load_invoices()
            if not all_invoices_data:
                return
            
            invoices_with_timestamp = []
            for invoice_data in all_invoices_data:
                try:
                    meta = invoice_data.get('meta', {})
                    client_name = meta.get('client_name', '')
                    if client_name == self.client_name:
                        invoice = Invoice.from_dict(invoice_data)
                        if 'meta' in invoice_data:
                            meta_data = invoice_data['meta']
                            invoice.date = meta_data.get('date', invoice.date)
                            invoice.due_date = meta_data.get('due_date', invoice.due_date)
                            invoice.invoice_number = meta_data.get('invoice_number', invoice.invoice_number)
                            invoice.status = meta_data.get('status', 'Pending')
                            invoice.client_name = meta_data.get('client_name', invoice.client_name)
                            invoice.tax_rate = Decimal(str(meta_data.get('tax_rate', 0.0)))
                            invoice.notes = meta_data.get('notes', Config.DEFAULT_TERMS)
                            _rd = meta_data.get('received_date', 'N/A') or 'N/A'
                            invoice.received_date = _normalize_date(_rd) if _rd not in ('N/A', '') else 'N/A'
                        
                        timestamp = None
                        if 'meta' in invoice_data and 'created_at' in invoice_data['meta']:
                            created_at = invoice_data['meta']['created_at']
                            try:
                                if isinstance(created_at, (int, float)):
                                    timestamp = created_at
                                elif isinstance(created_at, str):
                                    dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                                    timestamp = dt.timestamp()
                            except:
                                pass
                        
                        if timestamp is None and hasattr(invoice, 'firebase_timestamp') and invoice.firebase_timestamp:
                            try:
                                if isinstance(invoice.firebase_timestamp, (int, float)):
                                    timestamp = invoice.firebase_timestamp
                                elif isinstance(invoice.firebase_timestamp, str):
                                    dt = datetime.fromisoformat(invoice.firebase_timestamp.replace('Z', '+00:00'))
                                    timestamp = dt.timestamp()
                            except:
                                pass
                        
                        if timestamp is None:
                            parsed_date = self.parse_invoice_date(invoice)
                            if parsed_date:
                                timestamp = parsed_date.timestamp()
                            else:
                                timestamp = 0
                        
                        invoice.firebase_timestamp = timestamp
                        invoices_with_timestamp.append((timestamp, invoice, None))
                except Exception as e:
                    _log.warning("Error processing invoice data: %s", e)
                    continue
            
            invoices_with_timestamp.sort(key=lambda x: x[0] if x[0] else 0, reverse=True)
            self.invoices = [(inv, json_file) for timestamp, inv, json_file in invoices_with_timestamp]
            self.apply_all_time_filter()
        except Exception as e:
            _log.warning("Error loading invoices from Firebase: %s", e)
            QtWidgets.QMessageBox.warning(self, "Load Error", f"Error loading invoices from Firebase: {str(e)}")
    
    def clear_quick_filter_highlighting(self):
        default_style = """
            QPushButton {
                background-color: #f8fbfd;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                padding: 0 14px;
                border-radius: 7px;
                font-size: 12px;
                font-weight: 800;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }
            QPushButton:hover { border-color: #00756f; color: #00756f; }
        """
        for btn in [self.last_7_days_btn, self.last_30_days_btn, self.this_month_btn, self.this_year_btn, self.all_time_btn]:
            btn.setStyleSheet(default_style)
    
    def apply_date_range_filter(self, from_date: datetime, to_date: datetime):
        try:
            self._sort_ascending = True
            self.current_date_filter = (from_date, to_date)
            self.clear_quick_filter_highlighting()
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying date range filter: %s", e)

    def clear_date_range_filter(self):
        try:
            self._sort_ascending = False
            self.current_date_filter = None
            self.apply_filters()
        except Exception as e:
            _log.warning("Error clearing date range filter: %s", e)
    
    def apply_search_filter(self, search_text: str):
        try:
            self.current_search_text = search_text.lower().strip()
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying search filter: %s", e)
    
    def apply_filters(self):
        if getattr(self, "_filtering", False):
            return
        self._filtering = True
        try:
            q = self.current_search_text.strip()
            q_clean = _normalize_search_text(q)
            filtered_invoices = []

            for invoice, json_file in self.invoices:
                # ── Date filter — always uses invoice.date field, not created_at ──
                if not q and self.current_date_filter:
                    from_dt, to_dt = self.current_date_filter
                    try:
                        _d = getattr(invoice, 'date', '') or ''
                        inv_dt = None
                        for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                            try:
                                inv_dt = datetime.strptime(_d, _fmt)
                                break
                            except ValueError:
                                continue
                        if inv_dt and inv_dt != datetime.min:
                            if not (from_dt.date() <= inv_dt.date() <= to_dt.date()):
                                continue
                    except Exception:
                        pass  # unparseable date → include

                # ── Search filter ──────────────────────────────────────────────
                if q:
                    try:
                        parts = [
                            str(invoice.invoice_number or ""),
                            str(invoice.client_name or ""),
                            str(invoice.date or ""),
                            str(invoice.total or ""),
                            str(invoice.subtotal or ""),
                        ]
                        for item in (invoice.items or []):
                            parts.append(str(getattr(item, "description", "") or ""))
                            parts.append(str(getattr(item, "project_number", "") or ""))
                        haystack = " ".join(parts).lower()
                        haystack_clean = _normalize_search_text(haystack)
                        if q not in haystack and q_clean not in haystack_clean:
                            continue
                    except Exception as ex:
                        _log.warning("Search error for invoice %s: %s",
                                     getattr(invoice, "invoice_number", "?"), ex)
                        # include on error rather than hide

                filtered_invoices.append((invoice, json_file))

            # Sort: specific filter → ascending by date; All Time → keep created_at order
            if getattr(self, '_sort_ascending', False):
                def _date_sort_key(entry):
                    _inv, _ = entry
                    _d = getattr(_inv, 'date', '') or ''
                    for _fmt in ("%m-%d-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                        try:
                            return datetime.strptime(_d, _fmt)
                        except ValueError:
                            continue
                    return datetime.min
                filtered_invoices.sort(key=_date_sort_key)

            self.filtered_invoices = filtered_invoices
            self.display_invoices(self.filtered_invoices)
            self.update_stats(self.filtered_invoices)
        except Exception as e:
            _log.warning("Error in apply_filters: %s", e)
        finally:
            self._filtering = False
    
    def apply_quick_filter(self, days: int):
        try:
            self._sort_ascending = True
            if days == 7:
                self.highlight_selected_quick_filter(self.last_7_days_btn)
            elif days == 30:
                self.highlight_selected_quick_filter(self.last_30_days_btn)

            self.current_date_filter = None
            self.date_range_widget.is_date_range_applied = False
            self.date_range_widget.apply_clear_btn.setText("Apply")
            self.date_range_widget.apply_clear_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover { background-color: #229954; }
            """)
            self.date_range_widget.hide_date_range()
            to_date = datetime.now()
            from_date = to_date - timedelta(days=days)
            self.current_date_filter = (from_date, to_date)
            self.date_range_widget.set_date_range(from_date, to_date)
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying quick filter: %s", e)
    
    def apply_this_month_filter(self):
        try:
            self._sort_ascending = True
            self.highlight_selected_quick_filter(self.this_month_btn)
            self.date_range_widget.hide_date_range()
            self.current_date_filter = None
            self.date_range_widget.is_date_range_applied = False
            self.date_range_widget.apply_clear_btn.setText("Apply")
            self.date_range_widget.apply_clear_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover { background-color: #229954; }
            """)
            today = datetime.now()
            from_date = datetime(today.year, today.month, 1)
            to_date = today
            self.current_date_filter = (from_date, to_date)
            self.date_range_widget.set_date_range(from_date, to_date)
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying this month filter: %s", e)
    
    def apply_this_year_filter(self):
        try:
            self._sort_ascending = True
            self.highlight_selected_quick_filter(self.this_year_btn)
            self.date_range_widget.hide_date_range()
            self.current_date_filter = None
            self.date_range_widget.is_date_range_applied = False
            self.date_range_widget.apply_clear_btn.setText("Apply")
            self.date_range_widget.apply_clear_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover { background-color: #229954; }
            """)
            today = datetime.now()
            from_date = datetime(today.year, 1, 1)
            to_date = today
            self.current_date_filter = (from_date, to_date)
            self.date_range_widget.set_date_range(from_date, to_date)
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying this year filter: %s", e)
    
    def apply_all_time_filter(self):
        try:
            self._sort_ascending = False
            self.highlight_selected_quick_filter(self.all_time_btn)
            self.date_range_widget.hide_date_range()
            # No date restriction for All Time — show every invoice regardless of date
            self.current_date_filter = None
            self.date_range_widget.is_date_range_applied = False
            self.date_range_widget.apply_clear_btn.setText("Apply")
            self.date_range_widget.apply_clear_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    border: none;
                    padding: 8px 15px;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover { background-color: #229954; }
            """)
            self.apply_filters()
        except Exception as e:
            _log.warning("Error applying all time filter: %s", e)
        
    def display_invoices(self, invoices: List):
        """Store the invoice list and render page 1."""
        self._ih_all_displayed = list(invoices) if invoices else []
        self._ih_page = 1
        self._ih_render_page()

    def _ih_render_page(self):
        """Render the current page of invoices into the table."""
        import math
        all_inv   = self._ih_all_displayed
        per_page  = self._ih_per_page
        total     = len(all_inv)
        max_page  = max(1, math.ceil(total / per_page))
        self._ih_page = max(1, min(self._ih_page, max_page))
        page_num  = self._ih_page
        start_i   = (page_num - 1) * per_page
        end_i     = min(start_i + per_page, total)
        invoices  = all_inv[start_i:end_i]

        self.invoice_table.setSortingEnabled(False)
        try:
            self.invoice_table.clearContents()
            self.invoice_table.clearSpans()

            if not invoices:
                # Show "No invoices found" message instead of empty row
                self.invoice_table.setRowCount(1)
                self.invoice_table.setColumnCount(11)
                self.invoice_table.setHorizontalHeaderLabels([
                    "Date", "Invoice Number", "Project Name", "Total Price", "Tax",
                    "Down Payment", "Total Due", "Due Date", "Status", "Received Date", "Actions"
                ])

                # Create a merged cell or just put a message in the first column
                no_data_item = QtWidgets.QTableWidgetItem("📭 No invoices found")
                no_data_item.setTextAlignment(QtCore.Qt.AlignCenter)

                # Span across all columns (optional - creates a nicer look)
                self.invoice_table.setSpan(0, 0, 1, 11)
                self.invoice_table.setItem(0, 0, no_data_item)

                # Style the empty state
                no_data_item.setForeground(QtGui.QColor(120, 120, 120))
                no_data_item.setFont(QtGui.QFont("Inter", 12))
                self._ih_rebuild_pagination(total, max_page)
                return

            # Normal case - we have invoices to display
            sorted_invoices = invoices
            self.invoice_table.setRowCount(len(sorted_invoices))
            self.invoice_table.setColumnCount(11)
            self.invoice_table.setHorizontalHeaderLabels([
                "Date", "Invoice Number", "Project Name", "Total Price", "Tax",
                "Down Payment", "Total Due", "Due Date", "Status", "Received Date", "Actions"
            ])

            header = self.invoice_table.horizontalHeader()
            header.setMinimumSectionSize(110)
            header.setSectionResizeMode(0,  QtWidgets.QHeaderView.ResizeToContents)  # Date
            header.setSectionResizeMode(1,  QtWidgets.QHeaderView.ResizeToContents)  # Invoice Number
            header.setSectionResizeMode(2,  QtWidgets.QHeaderView.Fixed)             # Project Name — FIXED to force wrapping
            header.setSectionResizeMode(3,  QtWidgets.QHeaderView.Stretch)  # Total Price
            header.setSectionResizeMode(4,  QtWidgets.QHeaderView.Stretch)  # Tax
            header.setSectionResizeMode(5,  QtWidgets.QHeaderView.Stretch)  # Down Payment
            header.setSectionResizeMode(6,  QtWidgets.QHeaderView.Stretch)  # Total Due
            header.setSectionResizeMode(7,  QtWidgets.QHeaderView.Stretch)  # Due Date
            header.setSectionResizeMode(8,  QtWidgets.QHeaderView.Fixed)    # Status
            header.setSectionResizeMode(9,  QtWidgets.QHeaderView.ResizeToContents)  # Received Date
            header.setSectionResizeMode(10, QtWidgets.QHeaderView.Fixed)    # Actions
            self.invoice_table.setColumnWidth(2,  300)  # Project Name — 45 chars per line for wrapping
            self.invoice_table.setColumnWidth(8,  150)
            self.invoice_table.setColumnWidth(10, 150)
            self.invoice_table.setWordWrap(True)  # Enable word wrapping
            # Use custom delegate for proper text wrapping
            self.invoice_table.setItemDelegateForColumn(2, TextWrapDelegate(self))
            self.invoice_table.verticalHeader().setDefaultSectionSize(120)  # Increased for 3+ wrapped lines
            self.invoice_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
            
            # Populate rows with invoice data
            item_font = QtGui.QFont("Inter", 9)
            # Load tax tracker once for tax-paid status checks in each row
            try:
                from tax_payment_tracker import get_tax_payment_tracker as _get_tt
                _ih_tax_tracker = _get_tt()
                _ih_tax_tracker._load_tax_payments()
            except Exception:
                _ih_tax_tracker = None
            for row, (invoice, json_file) in enumerate(sorted_invoices):
                self.invoice_table.setRowHeight(row, 56)

                # Date
                date_item = QtWidgets.QTableWidgetItem(invoice.date)
                date_item.setFont(item_font)
                date_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 0, date_item)
                
                # Invoice Number
                invoice_item = QtWidgets.QTableWidgetItem(invoice.invoice_number)
                invoice_item.setFont(item_font)
                invoice_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 1, invoice_item)
                
                # Project Name — all project names, no client prefix (wrapping enabled)
                project_display = self.get_project_name(invoice)
                project_item = QtWidgets.QTableWidgetItem(project_display)
                project_item.setFont(item_font)
                project_item.setTextAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)  # Center-aligned wrapped text
                project_item.setToolTip(project_display)  # Full text on hover
                self.invoice_table.setItem(row, 2, project_item)
                
                # Total price (subtotal)
                subtotal_item = QtWidgets.QTableWidgetItem(Currency.format(invoice.subtotal))
                subtotal_item.setFont(item_font)
                subtotal_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 3, subtotal_item)
                
                # Tax — show green + date tooltip when a tax entry exists for this invoice
                _tax_paid_date = ""
                _tax_is_paid = False
                if _ih_tax_tracker and float(invoice.tax_amount or 0) > 0.005:
                    _tax_recs = _ih_tax_tracker.get_invoice_taxes(invoice.invoice_number or "")
                    _tax_paid_amt = sum(float(t.amount) for t in _tax_recs)
                    if _tax_recs and _tax_paid_amt >= float(invoice.tax_amount) - 0.005:
                        _tax_is_paid = True
                        _tax_paid_date = _tax_recs[0].payment_date or ""
                tax_item = QtWidgets.QTableWidgetItem(Currency.format(invoice.tax_amount))
                tax_item.setFont(item_font)
                tax_item.setTextAlignment(QtCore.Qt.AlignCenter)
                # Always show full tax amount in tooltip (visible on truncation)
                tax_item.setToolTip(f"Tax: {Currency.format(invoice.tax_amount)}")
                if _tax_is_paid:
                    tax_item.setForeground(QtGui.QColor("#047857"))
                    _tip = f"Tax paid: {Currency.format(invoice.tax_amount)}"
                    if _tax_paid_date:
                        _tip += f"  |  Date: {_tax_paid_date}"
                    tax_item.setToolTip(_tip)
                self.invoice_table.setItem(row, 4, tax_item)
                
                # Down Payment
                total_down_payment = sum(item.down_payment for item in invoice.items)
                down_payment_item = QtWidgets.QTableWidgetItem(Currency.format(total_down_payment))
                down_payment_item.setFont(item_font)
                down_payment_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 5, down_payment_item)
                
                # Total Due
                total_item = QtWidgets.QTableWidgetItem(Currency.format(invoice.total))
                total_item.setFont(item_font)
                total_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 6, total_item)
                
                # Due Date
                due_date = getattr(invoice, 'due_date', 'N/A')
                if not due_date or due_date == '':
                    due_date = 'N/A'
                due_date_item = QtWidgets.QTableWidgetItem(due_date)
                due_date_item.setFont(item_font)
                due_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 7, due_date_item)
                
                # Status pill badge
                initial_status = self.get_invoice_status(invoice)
                inv_badge = self.create_invoice_status_badge(initial_status, invoice)
                self.invoice_table.setCellWidget(row, 8, inv_badge)
                
                # Received Date
                received_date = getattr(invoice, 'received_date', '')
                if not received_date or received_date == 'N/A':
                    received_date = "N/A"
                received_date_item = QtWidgets.QTableWidgetItem(received_date)
                received_date_item.setFont(item_font)
                received_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setItem(row, 9, received_date_item)
                
                # Actions
                actions_widget = QtWidgets.QWidget()
                actions_layout = QtWidgets.QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(6, 4, 6, 4)
                actions_layout.setSpacing(8)

                open_pdf_btn = QtWidgets.QPushButton("PDF")
                open_pdf_btn.setFixedSize(52, 32)
                open_pdf_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #0f8bd6;
                        color: white;
                        border: none;
                        border-radius: 6px;
                        padding: 0px;
                        font-size: 12px;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #0b75b6; }
                """)
                open_pdf_btn.clicked.connect(lambda checked=False, inv=invoice: self.open_pdf(inv))

                more_btn = QtWidgets.QPushButton("More")
                more_btn.setFixedSize(58, 32)
                more_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #334155;
                        color: white;
                        border: none;
                        border-radius: 6px;
                        padding: 0px;
                        font-size: 12px;
                        font-weight: bold;
                    }
                    QPushButton:hover { background-color: #1f2937; }
                    QPushButton::menu-indicator { image: none; width: 0px; }
                """)

                more_menu = QtWidgets.QMenu(self)
                view_action    = QtWidgets.QAction("👁 View Details", self)
                email_action   = QtWidgets.QAction("📧 Send to Client", self)
                payment_action = QtWidgets.QAction("💳 Payment History", self)
                edit_action    = QtWidgets.QAction("✏️ Edit Invoice", self)
                delete_action  = QtWidgets.QAction("🗑️ Delete Invoice", self)
                more_menu.addAction(view_action)
                more_menu.addAction(email_action)
                more_menu.addAction(payment_action)
                more_menu.addSeparator()
                more_menu.addAction(edit_action)
                more_menu.addSeparator()
                more_menu.addAction(delete_action)

                view_action.triggered.connect(lambda checked=False, inv=invoice: self.view_invoice_details(inv))
                email_action.triggered.connect(lambda checked=False, inv=invoice: self.send_invoice_email(inv))
                payment_action.triggered.connect(lambda checked=False, inv=invoice: self.show_invoice_payments_dialog(inv))
                edit_action.triggered.connect(lambda checked=False, inv=invoice: self.edit_invoice(inv, json_file))
                delete_action.triggered.connect(lambda checked=False, inv=invoice, jf=json_file: self.delete_invoice(inv, jf))
                more_btn.setMenu(more_menu)

                actions_layout.addWidget(open_pdf_btn)
                actions_layout.addWidget(more_btn)
                actions_layout.setAlignment(QtCore.Qt.AlignCenter)
                self.invoice_table.setCellWidget(row, 10, actions_widget)

            # Resize rows to fit wrapped text, capped between 56px and 160px (allows 3+ wrapped lines)
            self.invoice_table.resizeRowsToContents()
            for _r in range(self.invoice_table.rowCount()):
                _h = self.invoice_table.rowHeight(_r)
                self.invoice_table.setRowHeight(_r, max(56, min(_h, 160)))

        except Exception as e:
            _log.warning("Error displaying invoices: %s", e)
            import traceback
            traceback.print_exc()

        self._ih_rebuild_pagination(total, max_page)

    def _ih_rebuild_pagination(self, total: int, max_page: int):
        """Rebuild page buttons and info label — mirrors project tab style."""
        if not hasattr(self, '_ih_page_btns'):
            return
        import math
        page_num = self._ih_page
        per_page = self._ih_per_page
        start = (page_num - 1) * per_page + 1 if total else 0
        end   = min(page_num * per_page, total)

        if hasattr(self, '_ih_info_lbl'):
            self._ih_info_lbl.setText(f"Showing {start}–{end} of {total} invoices")

        while self._ih_page_btns.count():
            item = self._ih_page_btns.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        _s = getattr(self, '_ih_pg_style', '')
        _win_start = max(1, min(page_num, max_page - 2))
        for p in range(_win_start, min(_win_start + 3, max_page + 1)):
            btn = QtWidgets.QPushButton(str(p))
            btn.setFixedSize(32, 28)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            if p == page_num:
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #00756f; color: #ffffff;
                        border: 1px solid #00756f; border-radius: 6px;
                        font-size: 12px; font-weight: 700;
                        min-width: 32px; min-height: 28px; padding: 0 8px;
                    }
                    QPushButton:hover { background-color: #005f5a; color: #ffffff; }
                """)
            else:
                btn.setStyleSheet(_s)
                btn.clicked.connect(lambda _, pg=p: self._ih_go_to(pg))
            self._ih_page_btns.addWidget(btn)

        if hasattr(self, '_ih_prev_btn'):
            self._ih_prev_btn.setEnabled(page_num > 1)
        if hasattr(self, '_ih_next_btn'):
            self._ih_next_btn.setEnabled(page_num < max_page)

        self._ih_resize_table()

    def _ih_resize_table(self):
        """Expand invoice_table to fit all visible rows so no internal scroll is needed."""
        t = self.invoice_table
        h = t.horizontalHeader().height()
        for i in range(t.rowCount()):
            h += t.rowHeight(i)
        t.setFixedHeight(h + 2)

    def _ih_go_prev(self):
        if self._ih_page > 1:
            self._ih_page -= 1
            self._ih_render_page()

    def _ih_go_next(self):
        self._ih_page += 1
        self._ih_render_page()

    def _ih_go_to(self, page: int):
        self._ih_page = page
        self._ih_render_page()

    def scroll_to_invoice(self, invoice_number: str):
        """Navigate to the page containing invoice_number and select its row."""
        for idx, (inv, _) in enumerate(self._ih_all_displayed):
            if getattr(inv, 'invoice_number', '') == invoice_number:
                page = idx // self._ih_per_page + 1
                if page != self._ih_page:
                    self._ih_page = page
                    self._ih_render_page()
                row_on_page = idx % self._ih_per_page
                self.invoice_table.selectRow(row_on_page)
                self.invoice_table.scrollTo(
                    self.invoice_table.model().index(row_on_page, 0),
                    QtWidgets.QAbstractItemView.PositionAtCenter,
                )
                return

    def emit_balance_sheet_refresh(self, invoice_number=None):
        """Refresh balance sheet after invoice status changes — uses background thread."""
        try:
            main_window = self.window()
            while main_window and not hasattr(main_window, 'balance_sheet_tab'):
                main_window = main_window.parent()
            if main_window and hasattr(main_window, 'balance_sheet_tab'):
                balance_tab = main_window.balance_sheet_tab
                # Immediate: annual summary (is_payment entries already written)
                if hasattr(balance_tab, '_refresh_annual_revenue_background'):
                    balance_tab._refresh_annual_revenue_background()
                # Delayed: paid revenues table needs _recompute_invoice_status to finish first
                if hasattr(balance_tab, '_refresh_all_revenue_background'):
                    QtCore.QTimer.singleShot(
                        1500,
                        lambda bt=balance_tab: bt._refresh_all_revenue_background()
                    )
        except Exception as e:
            _log.warning("Error emitting balance sheet refresh: %s", e)
    
    def show_status_update_message(self, invoice_number: str, new_status: str):
        try:
            notification = QtWidgets.QLabel(self)
            notification.setText(f"✓ Invoice {invoice_number} marked as {new_status}")
            notification.setStyleSheet("""
                QLabel {
                    background-color: #27ae60;
                    color: white;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                    font-size: 12px;
                }
            """)
            notification.adjustSize()
            notification.move(self.width() - notification.width() - 20, 10)
            notification.show()
            QtCore.QTimer.singleShot(2000, notification.deleteLater)
        except:
            pass
    
    def _get_latest_payment_date_for_invoice(self, invoice_number: str) -> str:
        """Return the most recent payment_date string for the given invoice, or '' if none."""
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()
            inv = (invoice_number or "").strip()
            pmnts = [p for p in tracker.payments
                     if (p.invoice_number or "").strip() == inv
                     and (p.payment_stage or "").strip().lower() != "tax"]
            if not pmnts:
                return ""
            latest = max(pmnts, key=lambda p: p.payment_date or "")
            return latest.payment_date or ""
        except Exception:
            return ""

    def on_status_changed_with_date(self, invoice: Invoice, new_status: str, combo: QtWidgets.QComboBox):
        """Handle status changes with received date handling - SUPPORTS ALL STATUS TYPES"""
        try:
            # Store old status for comparison
            old_status = getattr(invoice, 'status', 'Pending')

            # If no change, return immediately
            if new_status == old_status:
                return

            # Inform user when moving away from Paid that invoice-history payments will be removed
            if old_status == "Paid" and new_status != "Paid":
                msg = QtWidgets.QMessageBox(self)
                msg.setWindowTitle("Invoice Status Change")
                msg.setIcon(QtWidgets.QMessageBox.Information)
                msg.setText(
                    f"<b>Changing Status from Paid to {new_status}</b>"
                )
                msg.setInformativeText(
                    "Any payments that were automatically recorded when this invoice was "
                    "marked as <b>Paid</b> (via Invoice History) will be removed from the "
                    "payment records and the Balance Sheet.<br><br>"
                    "Payments added manually through the Project tab will <b>not</b> be affected."
                )
                msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                msg.setDefaultButton(QtWidgets.QMessageBox.Ok)
                msg.exec_()

            # Handle received date based on status
            if new_status == "Paid":
                # Only prompt for received date if changing TO Paid
                if old_status != "Paid":
                    dialog = ReceivedDateDialog(invoice, self)
                    if dialog.exec_() == QtWidgets.QDialog.Accepted:
                        received_date = _normalize_date(dialog.get_received_date())
                        invoice.received_date = received_date
                        self.update_received_date_in_table(invoice.invoice_number, received_date)
                    else:
                        # User cancelled, revert status
                        if combo is not None:
                            combo.blockSignals(True)
                            combo.setCurrentText(old_status)
                            combo.blockSignals(False)
                            self.style_status_combo(combo, old_status)
                        return
            else:
                # Partially Paid → latest payment date; all other non-Paid statuses → N/A
                if new_status == "Partially Paid":
                    _rd = self._get_latest_payment_date_for_invoice(invoice.invoice_number) or "N/A"
                else:
                    _rd = "N/A"
                invoice.received_date = _rd
                self.update_received_date_in_table(invoice.invoice_number, _rd)

            # Update local cache immediately for instant UI response
            self.set_cached_status(invoice.invoice_number, new_status)

            # Style the combo box if provided (legacy path)
            if combo is not None:
                self.style_status_combo(combo, new_status)
            
            # Update the invoice object immediately
            invoice.status = new_status
            
            # Update local stats immediately
            self.update_stats_async()
            
            # Show success message instantly
            self.show_status_update_message(invoice.invoice_number, new_status)

            # Immediately sync status/received_date to balance sheet transaction table
            # (in-memory, zero latency — Firebase round-trip arrives via real-time listener)
            try:
                main_win = self.window()
                while main_win and not hasattr(main_win, "balance_sheet_tab"):
                    main_win = main_win.parent()
                bs = getattr(main_win, "balance_sheet_tab", None)
                if bs and hasattr(bs, "update_revenue_entry_status"):
                    rd = getattr(invoice, 'received_date', 'N/A') or 'N/A'
                    bs.update_revenue_entry_status(invoice.invoice_number, new_status, rd)
            except Exception:
                pass

            # ASYNC Firebase update - this will sync ALL statuses to balance sheet
            QtCore.QTimer.singleShot(10, lambda: self._async_update_firebase(invoice))
            
        except Exception as e:
            _log.warning("Error updating status: %s", e)
            import traceback
            traceback.print_exc()

    def _async_update_firebase(self, invoice: Invoice):
        """Background Firebase update - syncs ALL statuses to balance sheet"""
        try:
            # Update status in Firebase
            success = FirebaseManager.update_invoice_status(invoice.invoice_number, invoice.status)

            if success:
                # Update received date in Firebase.
                # Paid / Partially Paid → use the date stored on the invoice object (latest payment date).
                # All other statuses → N/A.
                _inv_rd = getattr(invoice, 'received_date', 'N/A') or 'N/A'
                if invoice.status in ("Paid", "Partially Paid") and _inv_rd != 'N/A':
                    self._update_received_date_async(invoice.invoice_number, _inv_rd)
                else:
                    self._update_received_date_async(invoice.invoice_number, "N/A")

                # Sync to balance sheet for ANY status change (not just Paid)
                self._sync_to_balance_sheet_async(invoice)

                # Remove auto-recorded payments only when fully clearing the invoice
                # (going to Unpaid/Pending). Keep payments intact for Partially Paid /
                # Overdue so individual project-history deletions don't cascade.
                if invoice.status in ("Unpaid", "Pending"):
                    self._remove_auto_payments_for_invoice(invoice.invoice_number)

                # Auto-advance project stage THEN record payment, THEN refresh balance sheet
                if invoice.status == "Paid":
                    def _paid_chain(inv=invoice):
                        self._advance_then_record(inv)
                        self.emit_balance_sheet_refresh(inv.invoice_number)
                        self._refresh_active_project_workspace()
                    QtCore.QTimer.singleShot(150, _paid_chain)
                else:
                    # Non-paid status change: refresh balance sheet and workspace
                    QtCore.QTimer.singleShot(150, lambda: self.emit_balance_sheet_refresh(invoice.invoice_number))
                    QtCore.QTimer.singleShot(200, self._refresh_active_project_workspace)

        except Exception as e:
            _log.warning("Background Firebase update error: %s", e)

    def _advance_then_record(self, invoice: Invoice):
        """Advance project stages FIRST, then record payments — guaranteed order, no race condition."""
        self._advance_project_stages(invoice)
        self._record_invoice_payments(invoice)

    def _advance_project_stages(self, invoice: Invoice):
        """Advance each project in the invoice to the next payment stage after Paid."""
        try:
            raw = FirebaseManager.load_invoices() or []
            target = next(
                (inv for inv in raw
                 if (inv.get("meta") or {}).get("invoice_number") == invoice.invoice_number),
                None
            )
            if not target:
                return
            for item in target.get("items", []):
                pn = item.get("project_number", "").strip()
                if pn:
                    paid_stage = item.get("payment_category", "")
                    FirebaseManager.advance_project_payment_stage(pn, paid_stage)
                    _log.info("Auto-advanced payment stage for project %s (paid: %s)", pn, paid_stage)
        except Exception as e:
            _log.warning("Error advancing project stages: %s", e)

    def _record_invoice_payments(self, invoice: Invoice):
        """Auto-record a payment in the tracker for every project item when invoice is Paid."""
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()

            received_date = getattr(invoice, "received_date", "") or ""
            invoice_number = invoice.invoice_number

            raw = FirebaseManager.load_invoices() or []
            target = next(
                (inv for inv in raw
                 if (inv.get("meta") or {}).get("invoice_number") == invoice_number),
                None
            )
            if not target:
                return

            projects_in_invoice = []
            for item in target.get("items", []):
                pn = item.get("project_number", "").strip()
                if not pn:
                    continue
                projects_in_invoice.append(pn)

                # Determine the amount paid for this item/stage
                raw_amt = (
                    item.get("payment_due")
                    or item.get("total")
                    or item.get("unit_price")
                    or 0
                )
                try:
                    amount = float(str(raw_amt).replace("$", "").replace(",", "") or 0)
                except (ValueError, TypeError):
                    amount = 0.0
                if amount <= 0:
                    continue

                raw_stage = item.get("payment_category", "") or ""
                # Normalize stage to canonical name before comparing / storing
                payment_stage = _normalize_payment_stage(raw_stage)

                existing = tracker.get_project_payments(pn)

                # ── Link truly unlinked payments for this stage to this invoice ──
                # Only link payments with NO invoice_number (added from project tab
                # before the invoice existed).  Never re-link payments that are already
                # bound to a different invoice.
                for _p in existing:
                    p_inv = (_p.invoice_number or "").strip()
                    if (
                        not p_inv                                               # unlinked
                        and _normalize_payment_stage(_p.payment_stage) == payment_stage
                    ):
                        tracker.update_payment(_p.payment_id, invoice_number=invoice_number)
                        _log.info(
                            "Linked pre-existing payment %s ($%.2f) to invoice %s",
                            _p.payment_id, float(_p.amount), invoice_number,
                        )
                existing = tracker.get_project_payments(pn)

                # ── Count payments for this invoice + stage ──────────────────
                # Rules:
                #   1. stage_match is ALWAYS required — a payment for Stage 1 must
                #      not count toward Stage 2's remaining balance.
                #   2. Payment must be linked to THIS invoice OR still unlinked.
                #   3. Payments linked to a DIFFERENT invoice → NEVER count.
                seen_ids: set = set()
                total_existing = 0.0
                for _p in existing:
                    if _p.payment_id in seen_ids:
                        continue
                    # Tax is always tracked separately — never counts towards a stage amount
                    if (_p.payment_stage or "").strip().lower() == "tax":
                        continue
                    p_inv = (_p.invoice_number or "").strip()
                    inv_match   = p_inv == invoice_number.strip()
                    is_unlinked = not p_inv
                    stage_match = _normalize_payment_stage(_p.payment_stage) == payment_stage
                    # Stage match is mandatory — prevents a payment for the wrong stage
                    # from reducing remaining_to_record and blocking the new payment.
                    if stage_match and (inv_match or is_unlinked):
                        total_existing += float(_p.amount)
                        seen_ids.add(_p.payment_id)

                remaining_to_record = round(amount - total_existing, 2)

                if remaining_to_record <= 0.005:
                    _log.info(
                        "Payment fully covered for %s / %s (existing=%.2f, total=%.2f) — skipping",
                        pn, invoice_number, total_existing, amount,
                    )
                    continue

                # Normalise date → always "MM-dd-YYYY"
                pay_date = _normalize_date(received_date)

                if total_existing > 0:
                    note_text = (
                        f"Balance recorded from invoice {invoice_number} "
                        f"(previously paid: {Currency.format(total_existing)})"
                    )
                else:
                    note_text = f"Auto-recorded from invoice {invoice_number}"

                success = tracker.add_payment(
                    project_number=pn,
                    amount=remaining_to_record,
                    payment_date=pay_date,
                    payment_method="Invoice",
                    notes=note_text,
                    invoice_number=invoice_number,
                    payment_stage=payment_stage,
                    sync_balance_sheet=True,
                )
                if success:
                    _log.info(
                        "Auto-recorded payment $%.2f (of $%.2f total) for project %s (invoice %s, stage: %s)",
                        remaining_to_record, amount, pn, invoice_number, payment_stage,
                    )

            # Record tax in the dedicated tax_payments store when invoice is first marked Paid.
            # Tax is never stored in payments — it lives in Firebase /tax_payments/.
            # If tax was already recorded for this invoice, skip.
            try:
                meta = target.get("meta") or {}
                tax_amount = 0.0
                try:
                    tax_amount = float(str(meta.get("tax_amount") or 0).replace("$", "").replace(",", "") or 0)
                except (ValueError, TypeError):
                    tax_amount = 0.0
                if tax_amount > 0.005 and projects_in_invoice:
                    from tax_payment_tracker import get_tax_payment_tracker as _get_tt
                    _tax_tracker = _get_tt()
                    _tax_tracker._load_tax_payments()
                    existing_tax = _tax_tracker.get_invoice_taxes(invoice_number)
                    if not existing_tax:
                        pay_date = _normalize_date(received_date)
                        first_pn = projects_in_invoice[0]
                        success = _tax_tracker.add_tax_payment(
                            invoice_number=invoice_number,
                            project_number=first_pn,
                            amount=tax_amount,
                            payment_date=pay_date,
                            payment_method="Invoice",
                            notes=f"Tax recorded from invoice {invoice_number}",
                        )
                        if success:
                            _log.info(
                                "Auto-recorded tax payment $%.2f for invoice %s (project %s) "
                                "→ Firebase /tax_payments/",
                                tax_amount, invoice_number, first_pn,
                            )
            except Exception as _tex:
                _log.warning("Error recording tax payment for invoice %s: %s", invoice_number, _tex)

            # Mark the is_invoice revenue entry as having payment-tracker entries so
            # _extract_paid_entries reliably skips it (prevents double-counting in the
            # Paid Revenue dialog even before Firebase is_payment writes complete).
            if projects_in_invoice and FIREBASE_AVAILABLE:
                _inv_num_flag = invoice_number

                def _flag_has_payments(inv=_inv_num_flag):
                    try:
                        from firebase_admin import db as _db
                        from datetime import timezone as _tz
                        _ref = _db.reference('revenue')
                        _snap = _ref.get() or {}
                        for _rid, _rev in _snap.items():
                            if (isinstance(_rev, dict)
                                    and _rev.get('is_invoice')
                                    and _rev.get('invoice_number') == inv):
                                _ref.child(_rid).update({
                                    'has_payment_entries': True,
                                    'updated_at': datetime.now(_tz.utc).isoformat(),
                                })
                                break
                    except Exception as _fe:
                        _log.warning("Could not set has_payment_entries for %s: %s", inv, _fe)

                threading.Thread(target=_flag_has_payments, daemon=True).start()

            # Check if any project is now fully paid → auto-set status to Paid
            self._auto_mark_fully_paid_projects(projects_in_invoice)

            # Trigger invoice-status + balance-sheet sync for every affected project so
            # has_payment_entries is set and is_payment entries are counted correctly.
            # IMPORTANT: delay must be > the longest possible Firebase write (~300ms).
            # At 200ms some writes may still be in-flight; _auto_sync_invoice_statuses
            # calls tracker._load_payments() which would overwrite the correct in-memory
            # state with a partial Firebase snapshot, making P2/P3 payments disappear.
            # 900ms is safe: all writes finish in ≤300ms, and _late_sync_ws (700ms)
            # has already rebuilt the workspace with the correct in-memory data.
            try:
                main_win = self.window()
                project_tab = getattr(main_win, "project_tab", None)
                if project_tab and hasattr(project_tab, "_auto_sync_invoice_statuses"):
                    for _pn in set(projects_in_invoice):
                        QtCore.QTimer.singleShot(
                            900,
                            lambda p=_pn: threading.Thread(
                                target=lambda _p=p: project_tab._auto_sync_invoice_statuses(_p),
                                daemon=True,
                            ).start(),
                        )
            except Exception:
                pass

            # --- Direct workspace rebuild ----------------------------------------
            # Rebuild the currently-open workspace immediately from in-memory data.
            # add_payment() appends synchronously so self.payments already contains
            # ALL project payments before any background Firebase writes complete.
            # Do NOT call _load_payments() here — it would overwrite the correct
            # in-memory state with a stale Firebase snapshot (writes are async and
            # may not have landed yet), wiping out payments for P002, P003, etc.
            try:
                _main_win = self.window()
                _pt = getattr(_main_win, "project_tab", None)
                if _pt:
                    _wd = getattr(_pt, "_ws_project_data", None)
                    if _wd:
                        def _rebuild_ws(_ptab=_pt, _wdata=_wd):
                            try:
                                _ptab.show_project_workspace(_wdata)
                            except Exception as _e:
                                _log.warning("Direct workspace rebuild failed: %s", _e)
                        QtCore.QTimer.singleShot(80, _rebuild_ws)

                    # Late sync: after Firebase writes complete (~500ms), reload
                    # payments from Firebase and rebuild the workspace that is
                    # currently open (read _ws_project_data at fire time, not now,
                    # so we don't force the user back to a stale project).
                    def _late_sync_ws(_ptab=_pt):
                        try:
                            import threading as _t2
                            def _bg2():
                                from payment_tracker import get_payment_tracker as _gpt
                                _gpt()._load_payments()
                                cur = getattr(_ptab, "_ws_project_data", None)
                                if cur:
                                    QtCore.QTimer.singleShot(
                                        0,
                                        lambda d=cur: _ptab.show_project_workspace(d),
                                    )
                            _t2.Thread(target=_bg2, daemon=True).start()
                        except Exception as _e2:
                            _log.warning("Late workspace sync failed: %s", _e2)
                    QtCore.QTimer.singleShot(700, _late_sync_ws)
            except Exception:
                pass

            # Refresh project list cell and all finance tabs
            self._refresh_project_payment_cells()

            # Re-render invoice table so Tax column reflects paid status immediately
            try:
                QtCore.QTimer.singleShot(55, self._ih_render_page)
            except Exception:
                pass

        except Exception as e:
            _log.warning("Error recording invoice payments: %s", e)

    def _auto_mark_fully_paid_projects(self, project_numbers: list):
        """Set status='Paid' for any project where all stages AND total amount are paid."""
        try:
            from project_number_generator import (
                update_project_status_on_full_payment,
                is_project_fully_paid,
            )
            raw_projects = FirebaseManager.load_projects() or []
            for pn in set(project_numbers):
                project = next((p for p in raw_projects
                                if p.get("project_number") == pn), None)
                if project and is_project_fully_paid(project):
                    update_project_status_on_full_payment(pn, project)
        except Exception as e:
            _log.warning("Error auto-marking fully paid projects: %s", e)

    def _remove_auto_payments_for_invoice(self, invoice_number: str):
        """Remove ONLY the payments auto-recorded when the invoice was marked Paid.
        Manually-added project payments are preserved and re-synced to Firebase so
        they continue to appear in the annual financial summary."""
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()

            # Auto-recorded payments are identified by their notes prefix.
            # All three patterns produced by _record_invoice_payments must be covered.
            _auto_patterns = (
                "auto-recorded from invoice",
                "balance recorded from invoice",
                "tax recorded from invoice",
                "recorded from invoice",      # legacy / fallback
            )
            to_remove = [
                p for p in tracker.payments
                if (p.invoice_number or "").strip() == invoice_number.strip()
                and any(
                    (p.notes or "").lower().strip().startswith(pat)
                    for pat in _auto_patterns
                )
            ]

            # Collect payment_ids before deletion so the orphan purge can use them
            removed_pids = {p.payment_id for p in to_remove}

            for payment in to_remove:
                tracker.delete_payment(payment.payment_id)
                _log.info("Removed auto-payment %s for invoice %s",
                          payment.payment_id, invoice_number)

            # Orphan purge: delete any Firebase is_payment entries for removed payments
            # that weren't reached via balance_sheet_id.
            # Also purge legacy entries (no payment_id) that belong to this invoice —
            # they would otherwise cause double-counting when the invoice is re-paid.
            if FIREBASE_AVAILABLE:
                try:
                    from firebase_admin import db as _db
                    rev_ref = _db.reference('revenue')
                    snap = rev_ref.get() or {}
                    for rev_id, rev in list(snap.items()):
                        if not isinstance(rev, dict) or not rev.get('is_payment'):
                            continue
                        rev_pid = rev.get('payment_id') or ''
                        rev_inv = (rev.get('invoice_number') or '').strip()
                        # Delete if payment_id matches removed set
                        if rev_pid in removed_pids:
                            rev_ref.child(rev_id).delete()
                            _log.info("Purged orphan is_payment %s (pid match)", rev_id)
                        # Also delete legacy entries for this invoice that have no
                        # payment_id — they would cause double-counting on re-pay.
                        elif (rev_inv == invoice_number.strip()
                              and not rev_pid):
                            rev_ref.child(rev_id).delete()
                            _log.info("Purged legacy is_payment %s (no pid, inv match)", rev_id)
                except Exception as _pe:
                    _log.warning("Orphan purge failed: %s", _pe)

            # Reload tracker so remaining counts are accurate
            tracker._load_payments()
            remaining = [
                p for p in tracker.payments
                if (p.invoice_number or "").strip() == invoice_number.strip()
            ]
            manual_paid = sum(float(p.amount) for p in remaining)

            # Re-sync remaining project payments to Firebase.
            # Previous versions wiped ALL is_payment entries for the invoice; any
            # project payment whose Firebase entry was lost must be restored so the
            # annual financial summary shows the correct paid revenue.
            if FIREBASE_AVAILABLE and remaining:
                try:
                    from firebase_admin import db as _db
                    rev_ref = _db.reference('revenue')
                    snap2 = rev_ref.get() or {}
                    existing_pids = {
                        v.get('payment_id') for v in snap2.values()
                        if isinstance(v, dict) and v.get('is_payment')
                    }
                    for p in remaining:
                        if p.payment_id not in existing_pids:
                            # Entry missing from Firebase — re-create it
                            p.balance_sheet_id = ""  # force fresh write
                            bs_id = tracker._sync_payment_to_balance_sheet(p)
                            if bs_id:
                                p.balance_sheet_id = bs_id
                                tracker._update_payment_field_in_firebase(
                                    p.payment_id, {'balance_sheet_id': bs_id}
                                )
                except Exception as _re:
                    _log.warning("Re-sync of project payments failed: %s", _re)

            # Update the invoice's balance-sheet revenue node to reflect new state
            try:
                if FIREBASE_AVAILABLE:
                    from firebase_admin import db as _db
                    from datetime import timezone
                    ref = _db.reference('revenue')
                    snap3 = ref.get() or {}
                    for rev_id, rev in snap3.items():
                        if (isinstance(rev, dict) and rev.get('is_invoice')
                                and rev.get('invoice_number') == invoice_number):
                            if manual_paid <= 0:
                                ref.child(rev_id).update({
                                    'status': 'Unpaid',
                                    'paid_amount': '0.00',
                                    'unpaid_amount': str(rev.get('amount', '0')),
                                    'has_payment_entries': False,
                                    'received_date': 'N/A',
                                    'down_payment_received_date': 'N/A',
                                    'updated_at': datetime.now(timezone.utc).isoformat(),
                                })
                            else:
                                ref.child(rev_id).update({
                                    'status': 'Partially Paid',
                                    'paid_amount': f'{manual_paid:.2f}',
                                    'has_payment_entries': True,
                                    'received_date': 'N/A',
                                    'updated_at': datetime.now(timezone.utc).isoformat(),
                                })
                            break
            except Exception as _e:
                _log.warning("Could not update balance-sheet node: %s", _e)

            self._refresh_project_payment_cells()

        except Exception as e:
            _log.warning("Error removing auto payments for invoice %s: %s",
                         invoice_number, e)

    def _refresh_project_payment_cells(self):
        """Refresh project cells and finance tabs — all deferred so the UI stays responsive."""
        try:
            main_win = self.window()
            if not main_win:
                return
            project_tab = getattr(main_win, "project_tab", None)

            # Defer filter_projects so the invoice-history UI returns to the user first
            if project_tab and hasattr(project_tab, "filter_projects"):
                QtCore.QTimer.singleShot(120, project_tab.filter_projects)

            # Targeted annual-summary refresh: fetch only /revenue/ data in background
            # and redraw the summary table immediately — faster than full finance reload.
            def _refresh_annual(_mw=main_win):
                bs = getattr(_mw, "balance_sheet_tab", None)
                if bs and hasattr(bs, "_refresh_annual_revenue_background"):
                    bs._refresh_annual_revenue_background()
            QtCore.QTimer.singleShot(200, _refresh_annual)

            # Paid revenues + annual summary after _recompute_invoice_status completes
            def _refresh_all(_mw=main_win):
                bs = getattr(_mw, "balance_sheet_tab", None)
                if bs and hasattr(bs, "_refresh_all_revenue_background"):
                    bs._refresh_all_revenue_background()
            QtCore.QTimer.singleShot(1500, _refresh_all)

            # Full finance-tab refresh (expenses, salary, overview) in background.
            def _refresh_finance(_mw=main_win):
                _pt = getattr(_mw, "project_tab", None)
                if _pt and hasattr(_pt, "_refresh_finance_tabs"):
                    _pt._refresh_finance_tabs()
            QtCore.QTimer.singleShot(350, _refresh_finance)

            # Workspace panel refresh
            self._refresh_active_project_workspace(project_tab)
        except Exception as e:
            _log.warning("Could not refresh project payment cells: %s", e)

    def _refresh_active_project_workspace(self, project_tab=None):
        """Rebuild the project workspace panel from the in-memory payment tracker.
        Payment plan amounts come from the tracker (local JSON), not from Firebase,
        so no network call is needed here."""
        try:
            if project_tab is None:
                main_win = self.window()
                # Walk up the widget hierarchy in case the tab is nested
                while main_win and not hasattr(main_win, "project_tab"):
                    main_win = main_win.parent()
                project_tab = getattr(main_win, "project_tab", None)
            if not project_tab:
                return
            ws_data = getattr(project_tab, "_ws_project_data", None)
            if not ws_data:
                return

            def _do_refresh():
                try:
                    # Do NOT call _load_payments() here — it would overwrite the
                    # correct in-memory state with a partial Firebase snapshot while
                    # background write threads for multi-project invoices are still
                    # in-flight.  tracker.payments is always up-to-date because
                    # add_payment / delete_payment / update_payment all mutate it
                    # synchronously before starting any Firebase write.
                    if hasattr(project_tab, "show_project_workspace"):
                        project_tab.show_project_workspace(ws_data)
                except Exception as _e:
                    _log.warning("Could not refresh workspace: %s", _e)

            QtCore.QTimer.singleShot(100, _do_refresh)
        except Exception as e:
            _log.warning("Could not refresh active project workspace: %s", e)

    def _update_received_date_async(self, invoice_number: str, received_date: str):
        """Update received date in Firebase in background"""
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db
            invoices_ref = db.reference('/invoices')
            invoices_data = invoices_ref.get()
            if invoices_data:
                for invoice_id, invoice_data in invoices_data.items():
                    if invoice_data and 'meta' in invoice_data:
                        if invoice_data['meta'].get('invoice_number') == invoice_number:
                            invoice_ref = db.reference(f'/invoices/{invoice_id}')
                            invoice_ref.update({
                                'meta/received_date': received_date,
                                'meta/updated_at': datetime.now(timezone.utc).isoformat()
                            })
                            return
        except Exception as e:
            _log.warning("Error updating received date: %s", e)

    def _sync_to_balance_sheet_async(self, invoice: Invoice):
        """Sync to balance sheet in background - NOW HANDLES ALL STATUS TYPES"""
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db
            revenue_ref = db.reference('revenue')
            all_revenue = revenue_ref.get()
            
            if not all_revenue:
                return
            
            for rev_id, revenue in all_revenue.items():
                if revenue and revenue.get('is_invoice') and revenue.get('invoice_number') == invoice.invoice_number:
                    updates = {}
                    
                    # Always sync status regardless of what it is
                    if invoice.status != revenue.get('status', 'Pending'):
                        updates['status'] = invoice.status
                    
                    # Always sync due date
                    if invoice.due_date != revenue.get('due_date', 'N/A'):
                        updates['due_date'] = invoice.due_date
                    
                    # Handle received_date based on status.
                    # Paid / Partially Paid → use the invoice's received_date (latest payment date).
                    # All other statuses → N/A.
                    if invoice.status in ("Paid", "Partially Paid"):
                        received_date = getattr(invoice, 'received_date', 'N/A') or 'N/A'
                        updates['received_date'] = received_date
                        if received_date != 'N/A':
                            try:
                                received_date_obj = datetime.strptime(received_date, "%m-%d-%Y")
                                updates['year'] = received_date_obj.year
                            except:
                                pass
                    else:
                        updates['received_date'] = 'N/A'
                        # Keep year based on invoice date
                        try:
                            date_obj = datetime.strptime(invoice.date, "%m-%d-%Y")
                            updates['year'] = date_obj.year
                        except:
                            pass
                    
                    # Also sync the description/notes
                    project_names = []
                    for item in invoice.items:
                        if hasattr(item, 'project_name') and item.project_name:
                            project_names.append(item.project_name)
                        elif item.description:
                            project_names.append(item.description)
                    new_description = f"{invoice.client_name} - {', '.join(project_names[:2])}"
                    if len(project_names) > 2:
                        new_description += f" +{len(project_names)-2} more"
                    
                    if new_description != revenue.get('description', ''):
                        updates['description'] = new_description
                    
                    # Sync amount
                    new_amount = float(invoice.total)
                    old_amount = float(revenue.get('amount', 0))
                    if abs(new_amount - old_amount) > 0.01:
                        updates['amount'] = str(new_amount)
                    
                    # Apply all updates if any
                    if updates:
                        updates['updated_at'] = datetime.now(timezone.utc).isoformat()
                        revenue_ref.child(rev_id).update(updates)
                        _log.info("Synced invoice %s to balance sheet with status: %s", invoice.invoice_number, invoice.status)
                    break
        except Exception as e:
            _log.warning("Error syncing to balance sheet: %s", e)
            

    def update_stats_async(self):
        """Update stats without blocking UI"""
        QtCore.QTimer.singleShot(10, lambda: self.update_stats(self.filtered_invoices))
        
    
    def update_received_date_in_table(self, invoice_number: str, received_date: str):
        for row in range(self.invoice_table.rowCount()):
            invoice_item = self.invoice_table.item(row, 1)
            if invoice_item and invoice_item.text() == invoice_number:
                new_item = QtWidgets.QTableWidgetItem(received_date)
                new_item.setTextAlignment(QtCore.Qt.AlignCenter)
                new_item.setFont(QtGui.QFont("Inter", 9))
                self.invoice_table.setItem(row, 9, new_item)
                break

    def update_invoice_row_immediately(self, invoice_number: str, new_status: str, received_date: str):
        """Instant in-memory update of status badge (col 8) and received date (col 9)
        for a specific invoice row. No Firebase read — mirrors balance sheet's
        update_revenue_entry_status() pattern for zero-latency feedback."""
        for row in range(self.invoice_table.rowCount()):
            inv_item = self.invoice_table.item(row, 1)
            if not (inv_item and inv_item.text() == invoice_number):
                continue
            # Find the invoice object from self.invoices so the new badge has a
            # working click handler referencing the correct (now-updated) object.
            invoice_obj = None
            for inv, _ in (self.invoices or []):
                if getattr(inv, 'invoice_number', '') == invoice_number:
                    invoice_obj = inv
                    break
            if invoice_obj is not None:
                invoice_obj.status = new_status
                invoice_obj.received_date = received_date
                self.set_cached_status(invoice_number, new_status)
                # Replace the status badge widget with one styled for the new status
                new_badge = self.create_invoice_status_badge(new_status, invoice_obj)
                self.invoice_table.setCellWidget(row, 8, new_badge)
            # Update received-date cell regardless of whether we found the object
            rd_item = QtWidgets.QTableWidgetItem(received_date)
            rd_item.setTextAlignment(QtCore.Qt.AlignCenter)
            rd_item.setFont(QtGui.QFont("Inter", 9))
            self.invoice_table.setItem(row, 9, rd_item)
            break

    def update_invoice_received_date_in_firebase(self, invoice_number: str, received_date: str):
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db
            invoices_ref = db.reference('/invoices')
            invoices_data = invoices_ref.get()
            if invoices_data:
                for invoice_id, invoice_data in invoices_data.items():
                    if invoice_data and 'meta' in invoice_data:
                        if invoice_data['meta'].get('invoice_number') == invoice_number:
                            invoice_ref = db.reference(f'/invoices/{invoice_id}')
                            invoice_ref.update({
                                'meta/received_date': received_date,
                                'meta/updated_at': datetime.now(timezone.utc).isoformat()
                            })
                            revenue_ref = db.reference('revenue')
                            all_revenue = revenue_ref.get()
                            if all_revenue:
                                for rev_id, revenue in all_revenue.items():
                                    if revenue and revenue.get('is_invoice') and revenue.get('invoice_number') == invoice_number:
                                        try:
                                            received_date_obj = datetime.strptime(received_date, "%m-%d-%Y")
                                            revenue_ref.child(rev_id).update({
                                                'received_date': received_date,
                                                'year': received_date_obj.year,
                                                'updated_at': datetime.now(timezone.utc).isoformat()
                                            })
                                        except:
                                            pass
                                        break
                            return
        except Exception as e:
            _log.warning("Error updating received date in Firebase: %s", e)
    
    def on_status_changed(self, invoice: Invoice, new_status: str, combo: QtWidgets.QComboBox):
        try:
            self.set_cached_status(invoice.invoice_number, new_status)
            self.style_status_combo(combo, new_status)
            invoice.status = new_status
            self.update_invoice_status_in_file(invoice.invoice_number, new_status)
            self.update_stats(self.filtered_invoices)
            if FIREBASE_AVAILABLE:
                success = FirebaseManager.update_invoice_status(invoice.invoice_number, new_status)
        except Exception as e:
            _log.warning("Error updating status: %s", e)
    
    INVOICE_STATUS_PALETTE = {
        "Paid":           ("#d1fae5", "#065f46", "#6ee7b7"),
        "Unpaid":         ("#fee2e2", "#991b1b", "#fca5a5"),
        "Pending":        ("#fef3c7", "#92400e", "#fcd34d"),
        "Overdue":        ("#fce7f3", "#9d174d", "#f9a8d4"),
        "Partially Paid": ("#dbeafe", "#1e40af", "#93c5fd"),
    }

    # Pill badge color palette (bg, text, border)
    _PILL_COLORS = {
        "Paid":           ("#d1fae5", "#065f46", "#6ee7b7"),
        "Unpaid":         ("#f8d7da", "#721c24", "#f5c6cb"),
        "Pending":        ("#fff3cd", "#856404", "#ffeaa7"),
        "Overdue":        ("#ffe5d9", "#a13700", "#ffb599"),
        "Partially Paid": ("#ede7f6", "#4a148c", "#d1c4e9"),
    }

    def _pill_style(self, status: str) -> str:
        bg, fg, border = self._PILL_COLORS.get(status, ("#e2e8f0", "#64748b", "#cbd5e1"))
        return (
            f"QPushButton {{ background-color:{bg}; color:{fg}; "
            f"border:1px solid {border}; border-radius:12px; "
            f"padding:4px 14px; font-size:12px; font-weight:bold; "
            f"font-family:'Inter','Segoe UI',sans-serif; }}"
            f"QPushButton:hover {{ background-color:{border}; }}"
        )

    def create_invoice_status_badge(self, status: str, invoice) -> QtWidgets.QWidget:
        """Centered pill badge. Paid = static label; others = clickable dropdown."""
        container = QtWidgets.QWidget()
        container.setStyleSheet("background:transparent; border:none;")
        lay = QtWidgets.QHBoxLayout(container)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.addStretch()

        if status == "Paid":
            # Static non-clickable label — no dropdown for paid invoices
            lbl = QtWidgets.QLabel("Paid  ✓")
            lbl.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
            bg, fg, border = self._PILL_COLORS["Paid"]
            lbl.setStyleSheet(
                f"QLabel {{ background-color:{bg}; color:{fg}; "
                f"border:1px solid {border}; border-radius:12px; "
                f"padding:4px 14px; font-size:12px; font-weight:bold; "
                f"font-family:'Inter','Segoe UI',sans-serif; }}"
            )
            lay.addWidget(lbl)
        else:
            btn = QtWidgets.QPushButton(f"{status}  ▾")
            btn.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            btn.setStyleSheet(self._pill_style(status))

            def show_menu(checked=False, b=btn, inv=invoice):
                from PyQt5.QtWidgets import QMenu, QAction
                menu = QMenu(b)
                menu.setStyleSheet("""
                    QMenu { background:white; border:1px solid #d0d7de;
                            border-radius:8px; padding:4px 0;
                            font-family:'Inter','Segoe UI'; font-size:12px; }
                    QMenu::item { padding:7px 20px; color:#24292f; }
                    QMenu::item:selected { background:#f6f8fa; color:#0969da; }
                """)
                for s in ["Paid", "Unpaid", "Pending", "Overdue", "Partially Paid"]:
                    a = QAction(s, menu)
                    a.triggered.connect(
                        lambda _, st=s, bref=b, iref=inv:
                            self._apply_invoice_badge(st, bref, iref))
                    menu.addAction(a)
                menu.exec_(b.mapToGlobal(QtCore.QPoint(0, b.height())))

            btn.clicked.connect(show_menu)
            lay.addWidget(btn)

        lay.addStretch()
        return container

    def _apply_invoice_badge(self, new_status: str,
                              badge_btn: QtWidgets.QPushButton, invoice):
        old_status = getattr(invoice, 'status', 'Pending')
        old_icon = "✓" if old_status == "Paid" else "▾"
        new_icon = "✓" if new_status == "Paid" else "▾"

        badge_btn.setText(f"{new_status}  {new_icon}")
        badge_btn.setStyleSheet(self._pill_style(new_status))
        self.on_status_changed_with_date(invoice, new_status, None)

        # Revert badge if user cancelled (e.g. dismissed the Paid date dialog)
        if getattr(invoice, 'status', old_status) != new_status:
            badge_btn.setText(f"{old_status}  {old_icon}")
            badge_btn.setStyleSheet(self._pill_style(old_status))
            return

        # If status changed to Paid, replace the clickable button with the static label
        if new_status == "Paid":
            for row in range(self.invoice_table.rowCount()):
                inv_item = self.invoice_table.item(row, 1)
                if inv_item and inv_item.text() == invoice.invoice_number:
                    new_badge = self.create_invoice_status_badge("Paid", invoice)
                    self.invoice_table.setCellWidget(row, 8, new_badge)
                    break

    def style_status_combo(self, combo: QtWidgets.QComboBox, status: str):
        status_styles = {
            "Paid": """
                QComboBox {
                    background-color: #d1fae5;
                    color: #065f46;
                    border: 1px solid #6ee7b7;
                    border-radius: 4px;
                    padding: 4px;
                    font-weight: bold;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border: none;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 1px solid #bdc3c7;
                    selection-background-color: #3498db;
                }
            """,
            "Unpaid": """
                QComboBox {
                    background-color: #f8d7da;
                    color: #721c24;
                    border: 1px solid #f5c6cb;
                    border-radius: 4px;
                    padding: 4px;
                    font-weight: bold;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border: none;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 1px solid #bdc3c7;
                    selection-background-color: #3498db;
                }
            """,
            "Pending": """
                QComboBox {
                    background-color: #fff3cd;
                    color: #856404;
                    border: 1px solid #ffeaa7;
                    border-radius: 4px;
                    padding: 4px;
                    font-weight: bold;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border: none;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 1px solid #bdc3c7;
                    selection-background-color: #3498db;
                }
            """,
            "Overdue": """
                QComboBox {
                    background-color: #ffe5d9;  /* Soft orange */
                    color: #a13700;             /* Burnt orange text */
                    border: 1px solid #ffb599;
                    border-radius: 4px;
                    padding: 4px;
                    font-weight: bold;
                }
                QComboBox::drop-down { border: none; width: 20px; }
                QComboBox::down-arrow { image: none; border: none; }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 1px solid #bdc3c7;
                    selection-background-color: #ff7043; /* Highlight */
                }
            """,
            "Partially Paid": """
                QComboBox {
                    background-color: #ede7f6;  /* Light lavender */
                    color: #4a148c;             /* Deep purple text */
                    border: 1px solid #d1c4e9;
                    border-radius: 4px;
                    padding: 4px;
                    font-weight: bold;
                }
                QComboBox::drop-down { border: none; width: 20px; }
                QComboBox::down-arrow { image: none; border: none; }
                QComboBox QAbstractItemView {
                    background-color: white;
                    border: 1px solid #bdc3c7;
                    selection-background-color: #9575cd;
                }
                QComboBox::drop-down { border: none; width: 20px; }
                QComboBox::down-arrow { image: none; border: none; }
            """
        }
        combo.setStyleSheet(status_styles.get(status, status_styles["Pending"]))
    
    def update_stats(self, invoices=None):
        """Update statistics - OPTIMIZED for speed"""
        try:
            # Only update if there are actual changes
            if invoices is None:
                invoices = self.invoices
            
            # Calculate quickly without UI blocking
            total_invoices = len(invoices)
            total_amount = 0
            total_paid_amount = 0
            
            for invoice, _ in invoices:
                total_amount += invoice.total
                if self.get_invoice_status(invoice) == "Paid":
                    total_paid_amount += invoice.total
            
            # Update UI labels directly without recreating widgets if possible
            if hasattr(self, 'stats_labels'):
                # Update existing labels
                self.stats_labels['total_invoices'].setText(str(total_invoices))
                self.stats_labels['total_paid'].setText(Currency.format(total_paid_amount))
                self.stats_labels['total_revenue'].setText(Currency.format(total_amount))
            else:
                # First time - create stats widgets
                self.stats_labels = {}
                def clear_layout(layout):
                    while layout.count():
                        item = layout.takeAt(0)
                        if item.widget():
                            item.widget().deleteLater()
                        elif item.layout():
                            clear_layout(item.layout())
                clear_layout(self.stats_layout)
                
                stats_container = QtWidgets.QWidget()
                stats_container_layout = QtWidgets.QHBoxLayout(stats_container)
                stats_container_layout.setContentsMargins(0, 0, 0, 0)
                stats_container_layout.setSpacing(15)
                
                stats_data = [
                    ("Total Invoices", str(total_invoices), "#0f8bd6", 'total_invoices'),
                    ("Total Amount Paid", Currency.format(total_paid_amount), "#d97706", 'total_paid'),
                    ("Total Revenue", Currency.format(total_amount), "#16a34a", 'total_revenue'),
                ]
                
                for label, value, color, key in stats_data:
                    box, value_label = self.create_stat_box(label, value, color)
                    stats_container_layout.addWidget(box)
                    self.stats_labels[key] = value_label  # 👈 correct reference
    
                stats_container_layout.insertStretch(0)
                stats_container_layout.addStretch()
                self.stats_layout.addWidget(stats_container)
                
        except Exception as e:
            _log.warning("Error updating stats: %s", e)
                
    def verify_firebase_connection(self):
        if not FIREBASE_AVAILABLE:
            return
        try:
            from firebase_admin import db
            test_ref = db.reference('/test_connection')
            test_ref.set({'test_time': datetime.now(timezone.utc).isoformat()})
        except Exception as e:
            _log.warning("Firebase verification failed: %s", e)
    
    def create_stat_box(self, label: str, value: str, color: str):
        widget = QtWidgets.QFrame()
        widget.setMinimumSize(170, 78)
        widget.setStyleSheet(f"""
            QFrame {{
                background: #f8fbfd;
                border: 1px solid #d8e2ec;
                border-left: 4px solid {color};
                border-radius: 8px;
            }}
            QLabel {{
                background: transparent;
                border: none;
                font-family: "Inter", "Segoe UI", Arial, sans-serif;
            }}
        """)
        layout = QtWidgets.QVBoxLayout(widget)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(4)
        
        value_label = QtWidgets.QLabel(value)
        value_label.setStyleSheet(f"""
            QLabel {{
                font-size: 19px;
                font-weight: 900;
                color: {color};
            }}
        """)
        value_label.setAlignment(QtCore.Qt.AlignCenter)

        desc_label = QtWidgets.QLabel(label)
        desc_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #64748b;
                font-weight: 800;
            }
        """)
        desc_label.setAlignment(QtCore.Qt.AlignCenter)

        layout.addWidget(value_label)
        layout.addWidget(desc_label)

        return widget, value_label   # 👈 return value_label directly

    def send_invoice_email(self, invoice: Invoice):
        """Email the invoice PDF to the client using configured SMTP settings."""
        try:
            from email_manager import EmailManager
        except ImportError:
            QtWidgets.QMessageBox.critical(self, "Error", "email_manager module not found.")
            return

        if not EmailManager.is_configured():
            QtWidgets.QMessageBox.warning(
                self, "Email Not Configured",
                "SMTP settings are not set up.\n\n"
                "Go to Settings → Email and fill in your SMTP host, username, and password."
            )
            return

        client_email = getattr(invoice, "client_email", "") or ""
        if not client_email.strip():
            QtWidgets.QMessageBox.warning(
                self, "No Client Email",
                f"Invoice {invoice.invoice_number} has no client email address.\n"
                "Edit the invoice to add one before sending."
            )
            return

        confirm = QtWidgets.QMessageBox.question(
            self, "Send Invoice",
            f"Send invoice {invoice.invoice_number} to {client_email}?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        if confirm != QtWidgets.QMessageBox.Yes:
            return

        # Locate the PDF — prefer local file, fall back to regenerating
        pdf_path = Config.INVOICES_DIR / f"{invoice.invoice_number}.pdf"
        if not pdf_path.exists():
            QtWidgets.QMessageBox.warning(
                self, "PDF Not Found",
                f"Could not find PDF for {invoice.invoice_number}.\n"
                "Please open the invoice and regenerate the PDF first."
            )
            return

        ok = EmailManager.send_invoice(invoice, pdf_path)
        if ok:
            QtWidgets.QMessageBox.information(
                self, "Sent",
                f"Invoice {invoice.invoice_number} sent to {client_email}."
            )
        else:
            QtWidgets.QMessageBox.critical(
                self, "Send Failed",
                "Email could not be delivered. Check your SMTP settings and try again.\n"
                "See logs/pims.log for details."
            )

    def open_pdf(self, invoice: Invoice):
        try:
            current_status = self.get_invoice_status(invoice)
            if FIREBASE_AVAILABLE and hasattr(FirebaseManager, 'load_pdf_from_firebase'):
                loading_msg = QtWidgets.QMessageBox(self)
                loading_msg.setWindowTitle("Loading PDF")
                loading_msg.setText("📥 Downloading PDF from cloud...")
                loading_msg.setStandardButtons(QtWidgets.QMessageBox.Cancel)
                loading_msg.show()
                QtCore.QTimer.singleShot(100, lambda: self.download_and_open_pdf_with_watermark(invoice, current_status, loading_msg))
            else:
                QtWidgets.QMessageBox.warning(self, "PDF Open", "Firebase not available. Cannot open PDF.")
        except Exception as e:
            _log.warning("Error opening PDF: %s", e)
    
    def download_and_open_pdf_with_watermark(self, invoice: Invoice, status: str, loading_msg: QtWidgets.QMessageBox):
        try:
            temp_dir = Path(tempfile.gettempdir()) / "mabs_invoices_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            original_pdf_path = temp_dir / f"{invoice.invoice_number}_original.pdf"
            pdf_path = FirebaseManager.load_pdf_from_firebase(invoice.invoice_number, original_pdf_path)
            loading_msg.close()
            if pdf_path and pdf_path.exists():
                if FileManager.open_file(pdf_path):
                    QtWidgets.QMessageBox.information(self, "PDF Open", f"✅ PDF opened successfully!\n\nInvoice: {invoice.invoice_number}\nStatus: {status}")
                    QtCore.QTimer.singleShot(10000, lambda: self.cleanup_temp_files([pdf_path]))
                else:
                    QtWidgets.QMessageBox.critical(self, "PDF Open", "Failed to open PDF file.")
            else:
                QtWidgets.QMessageBox.warning(
                    self, "PDF Not Generated",
                    f"No PDF found for invoice {invoice.invoice_number}.\n\n"
                    "The PDF has not been generated yet.\n"
                    "Please go to Invoice Management, load this invoice, and click 'Generate PDF'."
                )
        except Exception as e:
            loading_msg.close()
            QtWidgets.QMessageBox.critical(self, "PDF Open Error", f"Error opening PDF: {str(e)}")
    
    def cleanup_temp_files(self, file_paths: List[Path]):
        for file_path in file_paths:
            self.cleanup_temp_file(file_path)
    
    def cleanup_temp_file(self, file_path: Path):
        try:
            if file_path.exists():
                file_path.unlink()
        except Exception as e:
            _log.info("Could not clean up temporary PDF: %s", e)
    
    def view_invoice_details(self, invoice: Invoice):
        try:
            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle(f"Invoice Details - {invoice.invoice_number}")
            dialog.setMinimumSize(900, 640)
            dialog.resize(1000, 700)
            layout = QtWidgets.QVBoxLayout(dialog)
            details_text = QtWidgets.QTextEdit()
            details_text.setReadOnly(True)
            total_amount = sum(item.total for item in invoice.items)
            total_down_payment = sum(item.down_payment for item in invoice.items)
            payment_due_before_tax = sum(item.payment_due for item in invoice.items)
            project_names = []
            for item in invoice.items:
                project_number = item.project_number
                if project_number and project_number in self.projects_data:
                    project_name = self.projects_data[project_number]
                    if project_name and project_name not in project_names:
                        project_names.append(project_name)
            project_info = ", ".join(project_names) if project_names else "No Project Info"
            current_status = self.get_invoice_status(invoice)
            
            details_html = f"""
            <h2>Invoice Details</h2>
            <table border="0" cellspacing="5" cellpadding="5">
            <tr><td><b>Invoice Number:</b></td><td>{invoice.invoice_number}</td></tr>
            <tr><td><b>Date:</b></td><td>{invoice.date}</td></tr>
            <tr><td><b>Due Date:</b></td><td>{invoice.due_date}</td></tr>
            <tr><td><b>Client:</b></td><td>{invoice.client_name}</td></tr>
            <tr><td><b>Project(s):</b></td><td>{project_info}</td></tr>
            <tr><td><b>Status:</b></td><td><span style='color: {'#155724' if current_status == 'Paid' else '#721c24' if current_status == 'Overdue' else '#856404'}; font-weight: bold;'>{current_status}</span></td></tr>
            <tr><td><b>Subtotal:</b></td><td>{Currency.format(total_amount)}</td></tr>
            <tr><td><b>Payment Due (before tax):</b></td><td>{Currency.format(payment_due_before_tax)}</td></tr>
            <tr><td><b>Tax ({invoice.tax_rate}% on total):</b></td><td>{Currency.format(invoice.tax_amount)}</td></tr>
            <tr><td><b>Total Amount Due:</b></td><td>{Currency.format(invoice.total)}</td></tr>
            </table>
                    
            <h3>Items</h3>
            <table border="1" cellspacing="0" cellpadding="5" style="border-collapse: collapse; width: 100%;">
            <tr style="background-color: #3498db; color: white;">
                <th>Project #</th>
                <th>Project Name</th>
                <th>Plant</th>
                <th>Qty</th>
                <th>Unit Price</th>
                <th>Payment Stage</th>
                <th>Payment Due</th>
                <th>Total</th>
            </tr>
            """
            for item in invoice.items:
                stage_label = (item.payment_category or "").strip() or "—"
                details_html += f"""
                <tr>
                    <td>{item.project_number}</td>
                    <td>{item.description}</td>
                    <td>{item.plant}</td>
                    <td>{item.quantity}</td>
                    <td>{Currency.format(item.unit_price)}</td>
                    <td>{stage_label}</td>
                    <td>{Currency.format(item.payment_due)}</td>
                    <td>{Currency.format(item.total)}</td>
                </tr>
                """
            details_html += "</table>"
            if invoice.notes and invoice.notes.strip():
                details_html += f"<h3>Notes</h3><p>{invoice.notes}</p>"
            details_text.setHtml(details_html)
            layout.addWidget(details_text)
            close_btn = QtWidgets.QPushButton("Close")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)
            dialog.exec_()
        except Exception as e:
            _log.warning("Error viewing invoice details: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Error viewing invoice details: {e}")

    def show_invoice_payments_dialog(self, invoice: Invoice):
        """Show payment history for this specific invoice with totals and remaining balance."""
        try:
            from payment_tracker import get_payment_tracker
            from tax_payment_tracker import get_tax_payment_tracker as _get_tt
            tracker = get_payment_tracker()
            # Do NOT call _load_payments() here — both trackers append synchronously
            # before starting Firebase background writes, so tracker.payments is always
            # up-to-date.  Calling _load_payments() while writes are still in-flight
            # (within ~200ms of marking an invoice Paid) overwrites correct in-memory
            # state with a partial Firebase snapshot, making payments disappear.
            _tax_tracker = _get_tt()
            inv_number = invoice.invoice_number

            # ── Collect project payments for this invoice (tax is stored separately) ──
            invoice_payments = [
                p for p in tracker.payments
                if (p.invoice_number or "").strip() == inv_number
                and (p.payment_stage or "").strip().lower() != "tax"
            ]
            # Tax payments come from the dedicated tax store
            _inv_tax_pays = _tax_tracker.get_invoice_taxes(inv_number)

            # Invoice total
            invoice_total = float(invoice.total)
            total_paid = (
                sum(float(p.amount) for p in invoice_payments)
                + sum(float(t.amount) for t in _inv_tax_pays)
            )
            remaining = max(invoice_total - total_paid, 0.0)
            current_status = self.get_invoice_status(invoice)

            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle(f"Payment History — Invoice {inv_number}")
            dialog.setWindowFlags(
                dialog.windowFlags()
                | QtCore.Qt.WindowMaximizeButtonHint
                | QtCore.Qt.WindowMinimizeButtonHint
            )
            dialog.setMinimumWidth(780)
            dialog.setMinimumHeight(500)
            layout = QtWidgets.QVBoxLayout(dialog)
            layout.setSpacing(12)
            layout.setContentsMargins(16, 16, 16, 16)

            # ── Header info card ──────────────────────────────────────────
            info_frame = QtWidgets.QFrame()
            info_frame.setStyleSheet("""
                QFrame {
                    background: #f8fafc;
                    border: 1px solid #e2e8f0;
                    border-radius: 8px;
                }
            """)
            info_layout = QtWidgets.QGridLayout(info_frame)
            info_layout.setContentsMargins(14, 10, 14, 10)
            info_layout.setHorizontalSpacing(24)
            info_layout.setVerticalSpacing(4)

            def _lbl(text, bold=False, color="#374151", wrap=False):
                l = QtWidgets.QLabel(str(text))
                l.setStyleSheet(
                    f"font-weight:{'700' if bold else '400'};"
                    f"color:{color};border:none;font-size:12px;"
                )
                if wrap:
                    l.setWordWrap(True)
                return l

            # Collect all project names from invoice items
            _all_proj_names = []
            for _it in invoice.items:
                _nm = (
                    getattr(_it, "description", "") or
                    getattr(_it, "project_number", "") or ""
                ).strip()
                if _nm and _nm not in _all_proj_names:
                    _all_proj_names.append(_nm)

            if len(_all_proj_names) == 0:
                _proj_display = self.get_project_name(invoice)
            elif len(_all_proj_names) <= 2:
                _proj_display = ', '.join(_all_proj_names)
            else:
                _proj_display = ', '.join(_all_proj_names[:2]) + f', +{len(_all_proj_names)-2} more…'

            info_layout.setVerticalSpacing(6)
            info_layout.setHorizontalSpacing(24)

            info_layout.addWidget(_lbl("Invoice #:",    True), 0, 0)
            info_layout.addWidget(_lbl(inv_number),             0, 1)
            info_layout.addWidget(_lbl("Client:",       True), 0, 2)
            info_layout.addWidget(_lbl(invoice.client_name or '—'), 0, 3)

            info_layout.addWidget(_lbl("Project:",      True), 1, 0)
            _pl = _lbl(_proj_display, wrap=True)
            _pl.setMaximumWidth(340)
            info_layout.addWidget(_pl,                          1, 1)
            info_layout.addWidget(_lbl("Invoice Date:", True), 1, 2)
            info_layout.addWidget(_lbl(invoice.date or '—'),   1, 3)

            info_layout.addWidget(_lbl("Total Due:",    True), 2, 0)
            info_layout.addWidget(_lbl(Currency.format(invoice_total)), 2, 1)
            info_layout.addWidget(_lbl("Status:",       True), 2, 2)
            _st_color = {"Paid": "#15803d", "Partially Paid": "#1e40af",
                         "Overdue": "#b91c1c"}.get(current_status, "#78350f")
            status_lbl = QtWidgets.QLabel(current_status)
            status_lbl.setStyleSheet(
                f"font-weight:800;color:{_st_color};border:none;font-size:12px;"
            )
            info_layout.addWidget(status_lbl, 2, 3)
            info_layout.setColumnStretch(1, 1)
            info_layout.setColumnStretch(3, 1)
            layout.addWidget(info_frame)

            # ── Payment History table — grouped by Project # ──────────────
            _th = QtWidgets.QLabel("Payment History by Project")
            _th.setStyleSheet(
                "font-weight:700;font-size:13px;color:#0f172a;"
                "border:none;padding:4px 0 2px 0;"
            )
            layout.addWidget(_th)
            tbl = QtWidgets.QTableWidget()
            COL_COUNT = 6
            tbl.setColumnCount(COL_COUNT)
            tbl.setHorizontalHeaderLabels(
                ["Project #", "Date", "Amount", "Method", "Stage", "Notes"]
            )
            tbl.horizontalHeader().setVisible(True)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.horizontalHeader().setDefaultSectionSize(38)
            tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            tbl.setAlternatingRowColors(False)
            tbl.verticalHeader().setVisible(False)
            tbl.setStyleSheet("""
                QTableWidget { background: white; border: 1px solid #e2e8f0;
                               border-radius: 6px; gridline-color: #f1f5f9; }
                QTableWidget::item { padding: 5px 8px; color: #1e293b; }
                QHeaderView::section { background: #f8fafc; font-weight: 700; padding: 8px;
                    border: none; border-bottom: 2px solid #e2e8f0;
                    min-height: 34px; color: #374151; }
                QTableWidget::item:selected { background: #dbeafe; color: #1e40af; }
            """)

            def _fmt_date(raw):
                for fmt in ("%Y-%m-%d", "%m-%d-%Y"):
                    try:
                        return datetime.strptime(raw, fmt).strftime("%b %d, %Y")
                    except Exception:
                        pass
                return raw or "N/A"

            def _cell(text, align=QtCore.Qt.AlignCenter):
                it = QtWidgets.QTableWidgetItem(str(text))
                it.setTextAlignment(align)
                return it

            # ── Build project order: invoice items first, then any extra from tracker ──
            pn_order = []
            seen_pns: set = set()
            for _it in invoice.items:
                _pn = (getattr(_it, "project_number", "") or "").strip()
                if _pn and _pn not in seen_pns:
                    pn_order.append(_pn)
                    seen_pns.add(_pn)
            # Payments for projects not listed as items (edge case) go at the end
            for _p in invoice_payments:
                _pn = (_p.project_number or "").strip()
                if _pn and _pn not in seen_pns:
                    pn_order.append(_pn)
                    seen_pns.add(_pn)

            def _make_bg(row, col, bg_color):
                it = QtWidgets.QTableWidgetItem("")
                it.setBackground(QtGui.QBrush(bg_color))
                return it

            # ── Planned amount per project from invoice items ─────────────
            planned_per_pn: dict = {}
            for _it in invoice.items:
                _pn = (getattr(_it, "project_number", "") or "").strip()
                if _pn:
                    _amt = float(
                        getattr(_it, "payment_due", None) or
                        getattr(_it, "unit_price", 0) or 0
                    )
                    planned_per_pn[_pn] = planned_per_pn.get(_pn, 0.0) + _amt

            # Build rows: one section header per project + payment rows + subtotal row
            _HDR_BG = QtGui.QColor("#1e3a5f")
            _HDR_FG = QtGui.QColor("#ffffff")
            _SUB_BG = QtGui.QColor("#f0f9ff")
            _SUB_FG = QtGui.QColor("#0369a1")
            _REM_GRN = QtGui.QColor("#15803d")
            _REM_RED = QtGui.QColor("#b91c1c")
            _TAX_HDR_BG = QtGui.QColor("#0f5a52")

            # "sub" data = (paid, planned)
            rows_spec = []
            for pn in pn_order:
                pn_pays = sorted(
                    [p for p in invoice_payments
                     if (p.project_number or "").strip() == pn
                     and (p.payment_stage or "").strip().lower() != "tax"],
                    key=lambda p: p.payment_date or "",
                )
                pn_paid    = sum(float(p.amount) for p in pn_pays)
                pn_planned = planned_per_pn.get(pn, 0.0)
                rows_spec.append(("header", pn))
                for pay in pn_pays:
                    rows_spec.append(("pay", pay))
                rows_spec.append(("sub", (pn_paid, pn_planned)))

            if not invoice_payments and not pn_order:
                rows_spec = [("empty", None)]

            # ── TAX section (only when invoice has tax) ───────────────────
            # Reads from tax_payment_tracker (tax_payments.json / Firebase /tax_payments/)
            try:
                _tax_amount = float(getattr(invoice, "tax_amount", 0) or 0)
            except (TypeError, ValueError):
                _tax_amount = 0.0
            if _tax_amount > 0.005:
                tax_pays = sorted(_inv_tax_pays, key=lambda t: t.payment_date or "")
                tax_paid_total = sum(float(t.amount) for t in tax_pays)
                rows_spec.append(("tax_header", _tax_amount))
                if tax_pays:
                    for tp in tax_pays:
                        rows_spec.append(("tax_pay", tp))
                else:
                    rows_spec.append(("tax_pending", (inv_number, _tax_amount)))
                rows_spec.append(("tax_sub", (tax_paid_total, _tax_amount)))

            import re as _re
            tbl.setRowCount(len(rows_spec))
            for r, (kind, data) in enumerate(rows_spec):
                if kind == "header":
                    pn_label = data
                    hdr_item = QtWidgets.QTableWidgetItem(f"  Project: {pn_label}")
                    hdr_item.setBackground(QtGui.QBrush(_HDR_BG))
                    hdr_item.setForeground(QtGui.QBrush(_HDR_FG))
                    hdr_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                    hdr_item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, hdr_item)
                    for c in range(1, COL_COUNT):
                        bg = QtWidgets.QTableWidgetItem("")
                        bg.setBackground(QtGui.QBrush(_HDR_BG))
                        tbl.setItem(r, c, bg)
                    tbl.setSpan(r, 0, 1, COL_COUNT)
                    tbl.setRowHeight(r, 30)

                elif kind == "pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.project_number or "—").strip()))
                    tbl.setItem(r, 1, _cell(_fmt_date(pay.payment_date or "")))
                    amt_it = _cell(Currency.format(float(pay.amount)))
                    amt_it.setForeground(QtGui.QColor("#15803d"))
                    tbl.setItem(r, 2, amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or "—"))
                    stage_clean = _re.sub(r'\s*\(\d+%\)', '', pay.payment_stage or '').strip() or '—'
                    tbl.setItem(r, 4, _cell(stage_clean))
                    tbl.setItem(r, 5, _cell(pay.notes or "—",
                                            QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "sub":
                    pn_paid, pn_planned = data
                    pn_remaining = max(pn_planned - pn_paid, 0.0)
                    _sub_font = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)

                    # Left cell: Paid (spans 3 columns)
                    paid_cell = QtWidgets.QTableWidgetItem(
                        f"  Paid: {Currency.format(pn_paid)}"
                    )
                    paid_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    paid_cell.setForeground(QtGui.QBrush(_SUB_FG))
                    paid_cell.setFont(_sub_font)
                    paid_cell.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, paid_cell)
                    tbl.setItem(r, 1, _make_bg(r, 1, _SUB_BG))
                    tbl.setItem(r, 2, _make_bg(r, 2, _SUB_BG))
                    tbl.setSpan(r, 0, 1, 3)

                    # Right cell: Remaining (spans 3 columns)
                    _rem_color = _REM_GRN if pn_remaining <= 0 else _REM_RED
                    _rem_text = (
                        "Fully Paid ✓" if pn_remaining <= 0
                        else f"Remaining: {Currency.format(pn_remaining)}"
                    )
                    rem_cell = QtWidgets.QTableWidgetItem(f"{_rem_text}  ")
                    rem_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    rem_cell.setForeground(QtGui.QBrush(_rem_color))
                    rem_cell.setFont(_sub_font)
                    rem_cell.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, rem_cell)
                    tbl.setItem(r, 4, _make_bg(r, 4, _SUB_BG))
                    tbl.setItem(r, 5, _make_bg(r, 5, _SUB_BG))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                elif kind == "tax_header":
                    tax_hdr_item = QtWidgets.QTableWidgetItem("  TAX")
                    tax_hdr_item.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    tax_hdr_item.setForeground(QtGui.QBrush(QtGui.QColor("#ffffff")))
                    tax_hdr_item.setFont(QtGui.QFont("Consolas", 9, QtGui.QFont.Bold))
                    tax_hdr_item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, tax_hdr_item)
                    _tax_note = QtWidgets.QTableWidgetItem(
                        f"Tax Amount: {Currency.format(data)}  — Recorded when invoice is marked Paid  "
                    )
                    _tax_note.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    _tax_note.setForeground(QtGui.QBrush(QtGui.QColor("#a7f3d0")))
                    _tax_note.setFont(QtGui.QFont("Segoe UI", 8))
                    _tax_note.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 1, _tax_note)
                    for c in range(2, COL_COUNT):
                        _bg = QtWidgets.QTableWidgetItem("")
                        _bg.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                        tbl.setItem(r, c, _bg)
                    tbl.setSpan(r, 1, 1, COL_COUNT - 1)
                    tbl.setRowHeight(r, 30)

                elif kind == "tax_pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.invoice_number or "—").strip()))
                    tbl.setItem(r, 1, _cell(_fmt_date(pay.payment_date or "")))
                    amt_it = _cell(Currency.format(float(pay.amount)))
                    amt_it.setForeground(QtGui.QColor("#15803d"))
                    tbl.setItem(r, 2, amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or "—"))
                    tbl.setItem(r, 4, _cell("Tax"))
                    tbl.setItem(r, 5, _cell(pay.notes or "—",
                                            QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_pending":
                    inv_no, tax_amt = data
                    tbl.setItem(r, 0, _cell(inv_no))
                    tbl.setItem(r, 1, _cell("—"))
                    amt_it = _cell(Currency.format(tax_amt))
                    amt_it.setForeground(QtGui.QColor("#b45309"))
                    tbl.setItem(r, 2, amt_it)
                    tbl.setItem(r, 3, _cell("—"))
                    _badge = QtWidgets.QTableWidgetItem("  Unpaid  ")
                    _badge.setBackground(QtGui.QBrush(QtGui.QColor("#fef3c7")))
                    _badge.setForeground(QtGui.QBrush(QtGui.QColor("#92400e")))
                    _badge.setTextAlignment(QtCore.Qt.AlignCenter)
                    tbl.setItem(r, 4, _badge)
                    tbl.setItem(r, 5, _cell("Pending — mark invoice as Paid to record",
                                            QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_sub":
                    _tax_p, _tax_pl = data
                    _tax_rem = max(_tax_pl - _tax_p, 0.0)
                    _tsub_bg = QtGui.QColor("#f0fdf4") if _tax_rem <= 0 else QtGui.QColor("#fef9c3")
                    _tsub_fg = QtGui.QColor("#15803d") if _tax_rem <= 0 else QtGui.QColor("#92400e")
                    _tsub_font = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)
                    _tp_cell = QtWidgets.QTableWidgetItem(f"  Paid: {Currency.format(_tax_p)}")
                    _tp_cell.setBackground(QtGui.QBrush(_tsub_bg))
                    _tp_cell.setForeground(QtGui.QBrush(_tsub_fg))
                    _tp_cell.setFont(_tsub_font)
                    _tp_cell.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, _tp_cell)
                    tbl.setItem(r, 1, _make_bg(r, 1, _tsub_bg))
                    tbl.setItem(r, 2, _make_bg(r, 2, _tsub_bg))
                    tbl.setSpan(r, 0, 1, 3)
                    _tr_text = "Tax Paid ✓" if _tax_rem <= 0 else f"Remaining: {Currency.format(_tax_rem)}"
                    _tr_cell = QtWidgets.QTableWidgetItem(f"{_tr_text}  ")
                    _tr_cell.setBackground(QtGui.QBrush(_tsub_bg))
                    _tr_cell.setForeground(QtGui.QBrush(_tsub_fg))
                    _tr_cell.setFont(_tsub_font)
                    _tr_cell.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, _tr_cell)
                    tbl.setItem(r, 4, _make_bg(r, 4, _tsub_bg))
                    tbl.setItem(r, 5, _make_bg(r, 5, _tsub_bg))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                else:  # empty
                    ni = QtWidgets.QTableWidgetItem("No payments recorded for this invoice")
                    ni.setTextAlignment(QtCore.Qt.AlignCenter)
                    tbl.setItem(r, 0, ni)
                    tbl.setSpan(r, 0, 1, COL_COUNT)
                    tbl.setRowHeight(r, 38)

            tbl.setColumnWidth(0, 165)   # Project #
            tbl.setColumnWidth(1, 115)   # Date
            tbl.setColumnWidth(2, 105)   # Amount
            tbl.setColumnWidth(3, 115)   # Method
            tbl.setColumnWidth(4, 130)   # Stage
            layout.addWidget(tbl)

            # ── Summary bar ───────────────────────────────────────────────
            summary_frame = QtWidgets.QFrame()
            summary_frame.setStyleSheet("""
                QFrame { background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 8px; }
            """ if remaining <= 0 else """
                QFrame { background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px; }
            """)
            sf_layout = QtWidgets.QHBoxLayout(summary_frame)
            sf_layout.setContentsMargins(16, 10, 16, 10)

            def _summary_val(label, value, color="#1e293b"):
                col = QtWidgets.QVBoxLayout()
                lbl = QtWidgets.QLabel(label)
                lbl.setStyleSheet("font-size:11px; color:#64748b; border:none;")
                val = QtWidgets.QLabel(value)
                val.setStyleSheet(f"font-size:15px; font-weight:800; color:{color}; border:none;")
                col.addWidget(lbl)
                col.addWidget(val)
                return col

            sf_layout.addLayout(_summary_val("Invoice Total", Currency.format(invoice_total)))
            sep1 = QtWidgets.QLabel("|")
            sep1.setStyleSheet("color:#cbd5e1; font-size:20px; border:none;")
            sf_layout.addWidget(sep1)
            sf_layout.addLayout(_summary_val("Total Paid", Currency.format(total_paid), "#15803d"))
            sep2 = QtWidgets.QLabel("|")
            sep2.setStyleSheet("color:#cbd5e1; font-size:20px; border:none;")
            sf_layout.addWidget(sep2)
            sf_layout.addLayout(_summary_val(
                "Remaining Balance",
                Currency.format(remaining),
                "#15803d" if remaining <= 0 else "#b45309",
            ))
            sf_layout.addStretch()
            layout.addWidget(summary_frame)

            close_btn = QtWidgets.QPushButton("Close")
            close_btn.setFixedHeight(36)
            close_btn.setStyleSheet("""
                QPushButton { background:#334155; color:white; border:none; border-radius:6px;
                              font-weight:bold; padding:0 24px; }
                QPushButton:hover { background:#1e293b; }
            """)
            close_btn.clicked.connect(dialog.accept)
            btn_row = QtWidgets.QHBoxLayout()
            btn_row.addStretch()
            btn_row.addWidget(close_btn)
            layout.addLayout(btn_row)
            dialog.exec_()

        except Exception as e:
            _log.warning("Error showing payment history dialog: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Could not load payment history:\n{e}")

    def edit_invoice(self, invoice, json_file):
        """Navigate to Invoice Management tab and load the invoice for editing."""
        try:
            # Walk up the widget tree to find MainWindow
            main_window = None
            widget = self
            while widget is not None:
                if hasattr(widget, '_nav_to') and hasattr(widget, 'add_item_row'):
                    main_window = widget
                    break
                widget = widget.parent()

            if main_window is None:
                # Fallback: find via QApplication
                for w in QtWidgets.QApplication.topLevelWidgets():
                    if hasattr(w, '_nav_to') and hasattr(w, 'add_item_row'):
                        main_window = w
                        break

            if main_window is None:
                QtWidgets.QMessageBox.warning(self, "Edit Invoice",
                    "Cannot reach Invoice Management tab. Please use the Edit dialog instead.")
                dialog = EditInvoiceDialog(invoice, self)
                dialog.exec_()
                return

            # Navigate to Projects & Invoice → Invoice Management
            main_window._nav_to(2)
            if hasattr(main_window, 'project_invoice_inner_tabs'):
                main_window.project_invoice_inner_tabs.setCurrentIndex(1)

            # Clear current form
            main_window.clear_all_items()
            main_window.clear_client_information()

            # --- Populate client ---
            for combo in ('client_combo', 'line_items_client_combo'):
                if hasattr(main_window, combo):
                    c = getattr(main_window, combo)
                    c.blockSignals(True)
                    idx = c.findText(invoice.client_name)
                    if idx >= 0:
                        c.setCurrentIndex(idx)
                    else:
                        c.setEditText(invoice.client_name)
                    c.blockSignals(False)

            if hasattr(main_window, 'client_email_edit'):
                main_window.client_email_edit.setText(invoice.client_email)
            if hasattr(main_window, 'client_address_edit'):
                main_window.client_address_edit.setPlainText(invoice.client_address)
            main_window.update_invoice_client_summary(
                invoice.client_name, invoice.client_email, invoice.client_address)

            # --- Populate dates ---
            try:
                from PyQt5.QtCore import QDate
                inv_date = QDate.fromString(invoice.date, "MM-dd-yyyy")
                due_date = QDate.fromString(invoice.due_date, "MM-dd-yyyy")
                if inv_date.isValid():
                    main_window.date_edit.setDate(inv_date)
                if due_date.isValid():
                    main_window.due_date_edit.setDate(due_date)
            except Exception:
                pass

            # --- Preserve existing invoice number so save = update ---
            main_window.invoice.invoice_number = invoice.invoice_number
            main_window.invoice_no_edit.setText(invoice.invoice_number)

            # --- Tax rate ---
            if hasattr(main_window, 'tax_spin'):
                main_window.tax_spin.setValue(float(invoice.tax_rate))

            # --- Notes ---
            if hasattr(main_window, 'notes_edit'):
                main_window.notes_edit.setPlainText(invoice.notes or "")

            # --- Line items ---
            from main import InvoiceItem as MainInvoiceItem, ItemRowWidget
            for item in invoice.items:
                inv_item = MainInvoiceItem(
                    project_number=item.project_number,
                    description=item.description,
                    plant=item.plant,
                    quantity=item.quantity,
                    unit_price=float(item.unit_price),
                    down_payment=float(item.down_payment),
                    payment_category=item.payment_category,
                )
                main_window.add_item_row(inv_item)

                # Lock the row to the saved payment stage so it displays correctly
                last_row = main_window.item_rows[-1] if getattr(main_window, "item_rows", None) else None
                if last_row and item.payment_category:
                    stage_label = ItemRowWidget.normalize_payment_label(item.payment_category) or item.payment_category
                    if hasattr(last_row, "lock_to_stage"):
                        last_row.lock_to_stage(stage_label)
                    if hasattr(last_row, "update_total"):
                        last_row.update_total()

            # Preserve invoice status so Payment Status label shows correctly
            main_window.invoice.status        = invoice.status or 'Unpaid'
            main_window.invoice.received_date = getattr(invoice, 'received_date', 'N/A') or 'N/A'
            main_window._ps_base_status       = invoice.status or 'Unpaid'
            # Store paid amount for dynamic recalculation if user changes amounts
            if (invoice.status or '').strip() == 'Paid':
                main_window._editing_invoice_paid_amount = float(invoice.total)
            else:
                main_window._editing_invoice_paid_amount = 0.0

            main_window.update_totals()

            # Mark as editing so Generate PDF is always enabled for existing invoices
            main_window._editing_existing_invoice = True
            QtCore.QTimer.singleShot(150, main_window._update_pdf_btn_state)

            # Close the history window
            history_win = self.window()
            if history_win and history_win is not main_window:
                history_win.close()

            if hasattr(main_window, 'statusBar'):
                main_window.statusBar().showMessage(
                    f"Editing invoice {invoice.invoice_number} — modify and click Generate PDF or Save Invoice.", 8000)

        except Exception as e:
            _log.warning("Error opening invoice for editing: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Error loading invoice for edit: {str(e)}")

    def update_invoice_in_firebase(self, invoice):
        try:
            if FIREBASE_AVAILABLE:
                invoice_dict = invoice.to_dict()
                invoice_dict['meta'] = {
                    'invoice_number': invoice.invoice_number,
                    'client_name': invoice.client_name,
                    'client_email': invoice.client_email,
                    'client_address': invoice.client_address,
                    'date': invoice.date,
                    'due_date': invoice.due_date,
                    'status': invoice.status if hasattr(invoice, 'status') else 'Pending',
                    'received_date': getattr(invoice, 'received_date', 'N/A'),
                    'tax_rate': float(invoice.tax_rate),
                    'notes': invoice.notes,
                    'updated_at': datetime.now(timezone.utc).isoformat()
                }
                if hasattr(invoice, 'firebase_timestamp') and invoice.firebase_timestamp:
                    invoice_dict['meta']['created_at'] = invoice.firebase_timestamp
                return FirebaseManager.update_invoice(invoice.invoice_number, invoice_dict)
            else:
                return False
        except Exception as e:
            _log.warning("Error updating invoice in Firebase: %s", e)
            return False

    def delete_invoice(self, invoice, json_file):
        try:
            reply = QtWidgets.QMessageBox.question(
                self, "Confirm Delete",
                f"Are you sure you want to delete this invoice?\n\n"
                f"Invoice Number: {invoice.invoice_number}\n\n"
                f"⚠️ This will also delete all associated revenue records.\n\n"
                f"This action cannot be undone!",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No
            )
            if reply == QtWidgets.QMessageBox.Yes:
                search_text = self.date_range_widget.search_bar.text()
                success = False
                if FIREBASE_AVAILABLE:
                    success = FirebaseManager.delete_invoice(invoice.invoice_number)
                else:
                    if json_file and json_file.exists():
                        json_file.unlink()
                        success = True
                if success:
                    # Also delete linked revenue entries from the balance sheet
                    self._delete_invoice_revenue_entries(invoice.invoice_number)

                    self.invoices = [(inv, file) for inv, file in self.invoices if inv.invoice_number != invoice.invoice_number]
                    self.filtered_invoices = [(inv, file) for inv, file in self.filtered_invoices if inv.invoice_number != invoice.invoice_number]
                    self.date_range_widget.search_bar.setText(search_text)
                    self.display_invoices(self.filtered_invoices)
                    self.update_stats(self.filtered_invoices)
                    QtWidgets.QMessageBox.information(self, "Success", f"Invoice {invoice.invoice_number} deleted successfully!")
                    # Refresh project tab's Recent Invoices so deleted invoice no longer appears
                    try:
                        w = self
                        while w and not hasattr(w, 'main_window'):
                            w = w.parent()
                        if w:
                            pt = getattr(w.main_window, 'project_tab', None)
                            if pt and hasattr(pt, 'refresh_recent_invoices'):
                                QtCore.QTimer.singleShot(300, pt.refresh_recent_invoices)
                    except Exception:
                        pass
                else:
                    QtWidgets.QMessageBox.critical(self, "Error", "Failed to delete invoice.")
        except Exception as e:
            _log.warning("Error deleting invoice: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Error deleting invoice: {str(e)}")

    def _delete_invoice_revenue_entries(self, invoice_number: str):
        """Remove all revenue entries in Firebase that belong to this invoice."""
        if not FIREBASE_AVAILABLE or not invoice_number:
            return
        try:
            from firebase_admin import db as _fdb
            revenue_ref = _fdb.reference('revenue')
            all_rev = revenue_ref.get() or {}
            deleted = 0
            for rev_id, rev_data in all_rev.items():
                if isinstance(rev_data, dict) and rev_data.get('invoice_number') == invoice_number:
                    revenue_ref.child(rev_id).delete()
                    deleted += 1
            if deleted:
                _log.info("Deleted %d revenue entries for invoice %s", deleted, invoice_number)
        except Exception as e:
            _log.warning("Error deleting revenue entries for invoice %s: %s", invoice_number, e)


class EditInvoiceDialog(QtWidgets.QDialog):
    """Dialog for editing existing invoices - FULLY EDITABLE ALL FIELDS"""
    
    def __init__(self, invoice, parent=None):
        super().__init__(parent)
        import copy
        self.invoice = invoice
        self.original_invoice = copy.deepcopy(self.invoice)
        self.item_rows = []
        self.setWindowTitle(f"Edit Invoice - {invoice.invoice_number}")
        self.setModal(True)
        self.resize(1200, 800)
        self.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        self.init_ui()
        self.populate_form()
        self.setup_shortcuts()
    
    def is_invoice_changed(self):
        return self.invoice.to_dict() != self.original_invoice.to_dict()

    def setup_shortcuts(self):
        self.save_shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+S"), self)
        self.save_shortcut.activated.connect(self.save_changes)
        self.installEventFilter(self)
    
    def eventFilter(self, source, event):
        if event.type() == QtCore.QEvent.KeyPress and event.key() == QtCore.Qt.Key_Return:
            focus_widgets = self.findChildren(QtWidgets.QWidget)
            focusable_widgets = [w for w in focus_widgets if w.focusPolicy() != QtCore.Qt.NoFocus]
            for i, widget in enumerate(focusable_widgets):
                if widget == self.focusWidget():
                    next_index = (i + 1) % len(focusable_widgets)
                    focusable_widgets[next_index].setFocus()
                    if isinstance(focusable_widgets[next_index], (QtWidgets.QSpinBox, QtWidgets.QDoubleSpinBox, QtWidgets.QComboBox)):
                        if hasattr(focusable_widgets[next_index], 'showPopup'):
                            focusable_widgets[next_index].showPopup()
                        else:
                            focusable_widgets[next_index].selectAll()
                    return True
        return super().eventFilter(source, event)
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        header = QtWidgets.QFrame()
        header.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                color: white;
                padding: 15px 24px;
            }
        """)
        header_layout = QtWidgets.QVBoxLayout(header)
        
        title = QtWidgets.QLabel(f"Edit Invoice: {self.invoice.invoice_number}")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: white;")
        
        header_layout.addWidget(title)
        layout.addWidget(header)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QtWidgets.QWidget()
        form_layout = QtWidgets.QVBoxLayout(scroll_widget)
        form_layout.setContentsMargins(30, 30, 30, 30)
        form_layout.setSpacing(20)
        
        self.add_section_title(form_layout, "📝 Invoice Information")

        self.invoice_number_edit = self.create_styled_line_edit("")
        self.invoice_number_edit.setReadOnly(True)
        self.invoice_number_edit.setStyleSheet("""
            QLineEdit {
                background-color: #e9ecef;
                color: #495057;
            }
        """)
        self.add_field(form_layout, "Invoice Number:", self.invoice_number_edit)

        self.date_edit = self.create_fixed_date_edit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("MM-dd-yyyy")
        self.date_edit.setStyleSheet(self.get_date_style())
        self.add_field(form_layout, "Date:", self.date_edit)
        
        self.due_date_edit = self.create_fixed_date_edit()       
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.due_date_edit.setStyleSheet(self.get_date_style())
        self.add_field(form_layout, "Due Date:", self.due_date_edit)
        
        self.tax_rate_edit = QtWidgets.QDoubleSpinBox()
        self.tax_rate_edit.valueChanged.connect(self.update_totals)
        self.tax_rate_edit.setRange(0, 100)
        self.tax_rate_edit.setDecimals(2)
        self.tax_rate_edit.setSuffix(" %")
        self.tax_rate_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)


        # 🔥 Disable scroll
        self.tax_rate_edit.wheelEvent = lambda event: None
        self.add_field(form_layout, "Tax Rate (%):", self.tax_rate_edit)
        
        self.status_combo = QtWidgets.QComboBox()
        self.status_combo.addItems(["Unpaid", "Paid", "Pending", "Overdue", "Partially Paid"])
        self.status_combo.setStyleSheet("""
            QComboBox {
                padding: 8px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QComboBox:focus { border-color: #3498db; }
        """)
        self.add_field(form_layout, "Status:", self.status_combo)
        
        self.received_date_edit = self.create_fixed_date_edit()
        self.received_date_edit.setCalendarPopup(True)
        self.received_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.received_date_edit.setMaximumDate(QtCore.QDate.currentDate())
        self.received_date_edit.setStyleSheet(self.get_date_style())
        # Style calendar so future (disabled) dates appear visually grayed out
        _rcal = self.received_date_edit.calendarWidget()
        if _rcal:
            _rcal.setStyleSheet("""
                QCalendarWidget QAbstractItemView {
                    selection-background-color: #00756f;
                    selection-color: white;
                }
                QCalendarWidget QAbstractItemView:disabled {
                    color: #c8c8c8;
                }
            """)
            _gray_fmt2 = QtGui.QTextCharFormat()
            _gray_fmt2.setForeground(QtGui.QBrush(QtGui.QColor("#c8c8c8")))
            _today2 = QtCore.QDate.currentDate()
            _d2 = _today2.addDays(1)
            _end2 = _today2.addYears(1)
            while _d2 <= _end2:
                _rcal.setDateTextFormat(_d2, _gray_fmt2)
                _d2 = _d2.addDays(1)
        self.received_date_edit.setEnabled(False)
        self.add_field(form_layout, "Payment Received Date:", self.received_date_edit)
        
        self.status_combo.currentTextChanged.connect(self.on_status_changed)
        
        self.add_section_title(form_layout, "📦 Invoice Items")
        
        header_frame = QtWidgets.QFrame()
        header_layout_inner = QtWidgets.QHBoxLayout(header_frame)
        header_layout_inner.setContentsMargins(5, 4, 5, 4)
        header_layout_inner.setSpacing(8)
        header_layout_inner.setAlignment(QtCore.Qt.AlignLeft)

        headers = ["Project Number", "Description", "Plant", "Quantity", "Unit Price", "Payment Category", "Payment Due", "Total", ""]
        widths = [140, 140, 180, 70, 120, 180, 100, 80, 40]

        for header, width in zip(headers, widths):
            label = QtWidgets.QLabel(header)
            label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12px;")
            label.setAlignment(QtCore.Qt.AlignCenter)
            label.setMinimumWidth(width)
            label.setMaximumWidth(width)
            label.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
            header_layout_inner.addWidget(label)
        
        items_scroll = QtWidgets.QScrollArea()
        items_scroll.setWidgetResizable(True)
        items_scroll.setMinimumHeight(250)
        items_scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background-color: white;
            }
        """)
        
        self.items_widget = QtWidgets.QWidget()
        self.items_layout = QtWidgets.QVBoxLayout(self.items_widget)
        self.items_layout.setSpacing(4)
        self.items_layout.setContentsMargins(5, 5, 5, 5)
        self.items_layout.addWidget(header_frame)
        items_scroll.setWidget(self.items_widget)
        
        form_layout.addWidget(items_scroll)
        
        totals_group = QtWidgets.QGroupBox("Invoice Totals")
        totals_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        totals_layout = QtWidgets.QFormLayout(totals_group)
        
        self.total_label = QtWidgets.QLabel("$0.00")
        self.down_payments_label = QtWidgets.QLabel("$0.00")
        self.tax_label = QtWidgets.QLabel("$0.00")
        self.total_amount_due_label = QtWidgets.QLabel("$0.00")
        
        self.total_amount_due_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                background-color: #e8f4fd;
                border: 2px solid #3498db;
                border-radius: 6px;
                padding: 8px;
            }
        """)
        
        totals_layout.addRow("Total:", self.total_label)
        totals_layout.addRow("Deposit Received:", self.down_payments_label)
        totals_layout.addRow("Tax Amount:", self.tax_label)
        totals_layout.addRow("Total Amount Due:", self.total_amount_due_label)
        
        form_layout.addWidget(totals_group)
        
        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.setSpacing(20)
        
        self.save_btn = QtWidgets.QPushButton("💾 Save Changes")
        self.save_btn.setMinimumHeight(48)
        self.save_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 30px;
            }
            QPushButton:hover { background: #2ecc71; }
        """)
        self.save_btn.clicked.connect(self.save_changes)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setMinimumHeight(48)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                padding: 10px 30px;
            }
            QPushButton:hover { background: #7f8c8d; }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.cancel_btn)
        btn_layout.addStretch()
        
        form_layout.addLayout(btn_layout)
        
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
    
    def on_status_changed(self, status):
        is_paid = (status == "Paid")
        self.received_date_edit.setEnabled(is_paid)
    
    def get_date_style(self):
        return """
            QDateEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """
    
    def create_fixed_date_edit(self, date=None):
        d = QtWidgets.QDateEdit(date if date else QtCore.QDate.currentDate())
        d.setCalendarPopup(True)
        d.setDisplayFormat("MM-dd-yyyy")
        d.wheelEvent = lambda event: None
        def keyPressEvent(event, original=d.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        d.keyPressEvent = keyPressEvent
        d.stepBy = lambda x: None
        d.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        d.setStyleSheet(self.get_date_style())
        return d

    
    def add_section_title(self, layout, text):
        label = QtWidgets.QLabel(text)
        label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                font-size: 16px;
                color: #2c3e50;
                border-bottom: 2px solid #dfe6e9;
                padding-bottom: 6px;
                margin-top: 10px;
            }
        """)
        layout.addWidget(label)
    
    def add_field(self, layout, label_text, widget):
        field_layout = QtWidgets.QHBoxLayout()
        label = QtWidgets.QLabel(label_text)
        label.setStyleSheet("font-weight: 500; color: #2c3e50; min-width: 120px;")
        field_layout.addWidget(label)
        field_layout.addWidget(widget, 1)
        layout.addLayout(field_layout)
    
    def create_styled_line_edit(self, placeholder=""):
        edit = QtWidgets.QLineEdit()
        edit.setPlaceholderText(placeholder)
        edit.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QLineEdit:focus { border-color: #3498db; }
        """)
        return edit
    
    def add_item_row(self, item=None):
        row = EditItemRowWidget(item)
        row.removed.connect(lambda: self.remove_item_row(row))
        row.valueChanged.connect(self.update_totals)
        self.items_layout.addWidget(row)
        self.item_rows.append(row)
        self.update_totals()
    
    def remove_item_row(self, row):
        if row in self.item_rows:
            self.item_rows.remove(row)
            row.setParent(None)
            self.update_totals()
    
    def clear_all_items(self):
        for row in self.item_rows[:]:
            self.remove_item_row(row)
    
    def update_totals(self):
        try:
            total_amount = Decimal('0.0')
            total_down_payments = Decimal('0.0')
            for row in self.item_rows:
                item = row.get_item()
                total_amount += item.total
                total_down_payments += item.down_payment
            total_payment_due = total_amount - total_down_payments
            tax_rate = Decimal(str(self.tax_rate_edit.value()))
            tax_amount = total_amount * (tax_rate / Decimal("100"))
            total_amount_due = total_payment_due + tax_amount
            self.total_label.setText(Currency.format(total_amount))
            self.down_payments_label.setText(Currency.format(total_down_payments))
            self.tax_label.setText(Currency.format(tax_amount))
            self.total_amount_due_label.setText(Currency.format(total_amount_due))
        except Exception as e:
            _log.warning("Error updating totals: %s", e)
    
    def populate_form(self):
        self.invoice_number_edit.setText(self.invoice.invoice_number)
        try:
            date = QtCore.QDate.fromString(self.invoice.date, "MM-dd-yyyy")
            if date.isValid():
                self.date_edit.setDate(date)
        except Exception as e:
            self.date_edit.setDate(QtCore.QDate.currentDate())
        try:
            due_date = QtCore.QDate.fromString(self.invoice.due_date, "MM-dd-yyyy")
            if due_date.isValid():
                self.due_date_edit.setDate(due_date)
        except Exception as e:
            self.due_date_edit.setDate(QtCore.QDate.currentDate().addDays(30))
        self.tax_rate_edit.setValue(float(self.invoice.tax_rate))
        status = getattr(self.invoice, 'status', 'Pending')
        index = self.status_combo.findText(status)
        if index >= 0:
            self.status_combo.setCurrentIndex(index)
        received_date = getattr(self.invoice, 'received_date', 'N/A')
        if received_date and received_date != 'N/A':
            try:
                rec_date = QtCore.QDate.fromString(received_date, "MM-dd-yyyy")
                if rec_date.isValid():
                    self.received_date_edit.setDate(rec_date)
            except:
                pass
        self.clear_all_items()
        for i, item in enumerate(self.invoice.items):
            self.add_item_row(item)
        self.update_totals()
    
    def update_invoice_in_firebase(self, invoice):
        try:
            if not FIREBASE_AVAILABLE:
                return False
            invoice_dict = invoice.to_dict()
            project_names = []
            for item in invoice.items:
                if hasattr(item, 'project_name') and item.project_name:
                    project_names.append(item.project_name)
                elif item.description:
                    project_names.append(item.description)
            description = f"{invoice.client_name} - {', '.join(project_names[:2])}"
            if len(project_names) > 2:
                description += f" +{len(project_names)-2} more"
            invoice_dict['meta'] = {
                'invoice_number': invoice.invoice_number,
                'client_name': invoice.client_name,
                'client_email': invoice.client_email,
                'client_address': invoice.client_address,
                'date': invoice.date,
                'due_date': invoice.due_date,
                'status': invoice.status if hasattr(invoice, 'status') else 'Pending',
                'received_date': getattr(invoice, 'received_date', 'N/A'),
                'tax_rate': float(invoice.tax_rate),
                'notes': invoice.notes,
                'description': description,
                'updated_at': datetime.now(timezone.utc).isoformat()
            }
            if hasattr(invoice, 'firebase_timestamp') and invoice.firebase_timestamp:
                invoice_dict['meta']['created_at'] = invoice.firebase_timestamp
            return FirebaseManager.update_invoice(invoice.invoice_number, invoice_dict)
        except Exception as e:
            _log.warning("Error updating invoice in Firebase: %s", e)
            return False
    
    def sync_to_balance_sheet(self, invoice):
        """Sync invoice changes to balance sheet - FIXED to properly update received_date and year"""
        try:
            if not FIREBASE_AVAILABLE:
                return
            from firebase_admin import db
            revenue_ref = db.reference('revenue')
            all_revenue = revenue_ref.get()
            if not all_revenue:
                return
            
            for rev_id, revenue in all_revenue.items():
                if revenue and revenue.get('is_invoice') and revenue.get('invoice_number') == invoice.invoice_number:
                    project_names = []
                    for item in invoice.items:
                        if hasattr(item, 'project_name') and item.project_name:
                            project_names.append(item.project_name)
                        elif item.description:
                            project_names.append(item.description)
                    new_description = f"{invoice.client_name} - {', '.join(project_names[:2])}"
                    if len(project_names) > 2:
                        new_description += f" +{len(project_names)-2} more"
                    
                    updates = {}
                    if new_description != revenue.get('description', ''):
                        updates['description'] = new_description
                    if invoice.due_date != revenue.get('due_date', 'N/A'):
                        updates['due_date'] = invoice.due_date
                    new_amount = float(invoice.total)
                    old_amount = float(revenue.get('amount', 0))
                    if abs(new_amount - old_amount) > 0.01:
                        updates['amount'] = str(new_amount)
                    if invoice.date != revenue.get('date', ''):
                        updates['date'] = invoice.date
                        try:
                            date_obj = datetime.strptime(invoice.date, "%m-%d-%Y")
                            updates['year'] = date_obj.year
                        except:
                            pass
                    if invoice.status != revenue.get('status', 'Pending'):
                        updates['status'] = invoice.status
                    
                    # Always update received_date based on status.
                    # Paid / Partially Paid → use the invoice's received_date (latest payment date).
                    # All other statuses → N/A.
                    if invoice.status in ("Paid", "Partially Paid"):
                        received_date = getattr(invoice, 'received_date', 'N/A') or 'N/A'
                        updates['received_date'] = received_date
                        if received_date != 'N/A':
                            try:
                                received_date_obj = datetime.strptime(received_date, "%m-%d-%Y")
                                updates['year'] = received_date_obj.year
                            except:
                                pass
                    else:
                        updates['received_date'] = 'N/A'
                        # Keep year based on invoice date
                        try:
                            date_obj = datetime.strptime(invoice.date, "%m-%d-%Y")
                            updates['year'] = date_obj.year
                        except:
                            pass
                    
                    if updates:
                        updates['updated_at'] = datetime.now(timezone.utc).isoformat()
                        revenue_ref.child(rev_id).update(updates)
                        _log.info("(converted from print, see git history)")
                    break
        except Exception as e:
            _log.warning("Error syncing to balance sheet: %s", e)
            import traceback
            traceback.print_exc()
        
    def refresh_balance_sheet(self):
        try:
            main_window = self.window()
            while main_window and not hasattr(main_window, 'balance_sheet_tab'):
                main_window = main_window.parent()
            if main_window and hasattr(main_window, 'balance_sheet_tab'):
                balance_tab = main_window.balance_sheet_tab
                if hasattr(balance_tab, 'refresh_invoice_revenues'):
                    balance_tab.refresh_invoice_revenues()
                else:
                    balance_tab.load_all_financial_data()
                    balance_tab.update_annual_summary()
                    balance_tab.on_category_changed(balance_tab.current_category)
                    balance_tab.update_stats_cards()
        except Exception as e:
            _log.warning("Error refreshing balance sheet: %s", e)
    
    def refresh_invoice_history(self):
        try:
            parent = self.parent()
            while parent:
                if hasattr(parent, 'refresh_invoices_immediately'):
                    parent.refresh_invoices_immediately()
                    break
                parent = parent.parent()
        except Exception as e:
            _log.warning("Error refreshing invoice history: %s", e)
    
    def regenerate_invoice_pdf(self, invoice):
        try:
            output_path = Config.INVOICES_DIR / f"{invoice.invoice_number}.pdf"
            logo_path = Config.get_logo_path()
            success = PDFGenerator.generate(invoice, output_path, logo_path)
            if success and FIREBASE_AVAILABLE:
                FirebaseManager.save_pdf_to_firebase(invoice.invoice_number, output_path)
        except Exception as e:
            _log.warning("Error regenerating PDF: %s", e)
    
    def save_changes(self):
        try:
            self.invoice.invoice_number = self.invoice_number_edit.text().strip()
            self.invoice.date = self.date_edit.date().toString("MM-dd-yyyy")
            self.invoice.due_date = self.due_date_edit.date().toString("MM-dd-yyyy")
            self.invoice.tax_rate = Decimal(str(self.tax_rate_edit.value()))
            self.invoice.status = self.status_combo.currentText()
            
            # FIX: Set received_date based on status
            if self.invoice.status == "Paid":
                # Check if received_date_edit has a valid date
                if self.received_date_edit.isEnabled() and self.received_date_edit.date().isValid():
                    self.invoice.received_date = self.received_date_edit.date().toString("MM-dd-yyyy")
                else:
                    # If received_date_edit is disabled or no valid date, use current date
                    self.invoice.received_date = datetime.now().strftime("%m-%d-%Y")
            else:
                self.invoice.received_date = "N/A"
            
            self.invoice.items = []
            for row in self.item_rows:
                self.invoice.items.append(row.get_item())
            self.invoice.calculate_totals()

            self.save_btn.setEnabled(False)
            self.save_btn.setText("Saving...")
            QtWidgets.QApplication.processEvents()

            progress = QtWidgets.QProgressDialog("Saving invoice and generating PDF...", None, 0, 0, self)
            progress.setWindowTitle("Please Wait")
            progress.setWindowModality(QtCore.Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            QtWidgets.QApplication.processEvents()

            self.regenerate_invoice_pdf(self.invoice)
            QtWidgets.QApplication.processEvents()
            if self.update_invoice_in_firebase(self.invoice):
                self.sync_to_balance_sheet(self.invoice)
            self.refresh_balance_sheet()
            self.refresh_invoice_history()

            progress.close()
            QtWidgets.QApplication.processEvents()

            QtWidgets.QMessageBox.information(self, "Success", f"Invoice {self.invoice.invoice_number} updated successfully!")
            self.accept()
        except Exception as e:
            _log.warning("Error saving changes: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Error saving changes: {str(e)}")
            

class EditItemRowWidget(QtWidgets.QWidget):
    """Widget for a single invoice item row in edit dialog"""
    removed = pyqtSignal()
    valueChanged = pyqtSignal()
    
    def __init__(self, item=None):
        super().__init__()
        self.item = item or InvoiceItem()
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(8)
        
        self.project_number_edit = QtWidgets.QLineEdit(self.item.project_number)
        self.project_number_edit.setPlaceholderText("Project #")
        self.project_number_edit.setMinimumHeight(35)
        self.project_number_edit.setMaximumWidth(180)
        self.project_number_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.project_number_edit.textChanged.connect(self.valueChanged.emit)
        
        self.desc_edit = QtWidgets.QLineEdit(self.item.description)
        self.desc_edit.setPlaceholderText("Description")
        self.desc_edit.setMinimumHeight(35)
        self.desc_edit.setMaximumWidth(180)
        self.desc_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.desc_edit.textChanged.connect(self.valueChanged.emit)
        
        self.plant_edit = QtWidgets.QLineEdit(self.item.plant)
        self.plant_edit.setPlaceholderText("Plant")
        self.plant_edit.setMinimumHeight(35)
        self.plant_edit.setMaximumWidth(180)
        self.plant_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.plant_edit.textChanged.connect(self.valueChanged.emit)
        
        self.qty_spin = QtWidgets.QSpinBox()
        self.qty_spin.setRange(1, 1000000)
        self.qty_spin.setValue(int(self.item.quantity))
        self.qty_spin.setMinimumHeight(35)
        self.qty_spin.setMinimumWidth(60)
        self.qty_spin.setAlignment(QtCore.Qt.AlignCenter)
        self.qty_spin.valueChanged.connect(self.update_total)
        self.qty_spin.valueChanged.connect(self.valueChanged.emit)
        
        initial_price = float(self.item.unit_price) if self.item.unit_price != Decimal('0') else 0.0
        self.price_edit = QtWidgets.QLineEdit(f"${initial_price:.2f}" if initial_price > 0 else "")
        self.price_edit.setPlaceholderText("$0.00")
        self.price_edit.setMinimumHeight(35)
        self.price_edit.setMinimumWidth(80)
        self.price_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.price_edit.textChanged.connect(self.update_total)
        self.price_edit.textChanged.connect(self.valueChanged.emit)
        
        self.payment_combo = QtWidgets.QComboBox()
        self.payment_combo.addItems(["Deposit Received (50%)", "Final Payment Due", "Full Amount Due"])
        self.payment_combo.setMinimumHeight(35)
        self.payment_combo.setMinimumWidth(120)
        
        category_found = False
        if hasattr(self.item, 'payment_category') and self.item.payment_category:
            stored_category = self.item.payment_category
            if "Deposit" in stored_category or "Down Payment" in stored_category or stored_category == "50%" or "50%" in stored_category:
                self.payment_combo.setCurrentText("Deposit Received (50%)")
                category_found = True
            elif "Final Payment" in stored_category:
                self.payment_combo.setCurrentText("Final Payment Due")
                category_found = True
            elif "Full Amount" in stored_category or "Due Payment" in stored_category:
                self.payment_combo.setCurrentText("Full Amount Due")
                category_found = True
        
        if not category_found:
            self.detect_payment_category_from_amounts()
        
        self.payment_combo.currentTextChanged.connect(self.update_total)
        self.payment_combo.currentTextChanged.connect(self.valueChanged.emit)
        
        self.payment_due_label = QtWidgets.QLabel(f"${float(self.item.payment_due):.2f}")
        self.payment_due_label.setAlignment(QtCore.Qt.AlignCenter)
        self.payment_due_label.setMinimumWidth(100)
        self.payment_due_label.setMinimumHeight(35)
        self.payment_due_label.setStyleSheet("""
            QLabel {
                background-color: #f0f0f0;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 8px;
                font-size: 11px;
            }
        """)
        
        self.total_label = QtWidgets.QLabel(Currency.format(self.item.total))
        self.total_label.setAlignment(QtCore.Qt.AlignCenter)
        self.total_label.setMinimumWidth(80)
        self.total_label.setMinimumHeight(35)
        self.total_label.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 8px;
                font-weight: bold;
                font-size: 11px;
            }
        """)
        
        self.remove_btn = QtWidgets.QPushButton("✕")
        self.remove_btn.setMaximumWidth(40)
        self.remove_btn.setMinimumHeight(35)
        self.remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #c82333; }
        """)
        self.remove_btn.clicked.connect(self.removed)
        
        layout.addWidget(self.project_number_edit)
        layout.addWidget(self.desc_edit)
        layout.addWidget(self.plant_edit)
        layout.addWidget(self.qty_spin)
        layout.addWidget(self.price_edit)
        layout.addWidget(self.payment_combo)
        layout.addWidget(self.payment_due_label)
        layout.addWidget(self.total_label)
        layout.addWidget(self.remove_btn)
        
        self.update_total()
    
    def detect_payment_category_from_amounts(self):
        total = Decimal(str(self.item.quantity)) * self.item.unit_price
        if total > 0:
            down = Decimal(str(self.item.down_payment))
            due = Decimal(str(self.item.payment_due))
            down_percentage = down / total if total > 0 else Decimal('0')
            if abs(down_percentage - Decimal('0.5')) < Decimal('0.01') and due > 0:
                self.payment_combo.setCurrentText("Deposit Received (50%)")
            elif due == 0:
                self.payment_combo.setCurrentText("Final Payment Due")
            elif down == 0:
                self.payment_combo.setCurrentText("Full Amount Due")
            else:
                self.payment_combo.setCurrentText("Full Amount Due")
        else:
            self.payment_combo.setCurrentText("Full Amount Due")
    
    def update_total(self):
        try:
            quantity = self.qty_spin.value()
            price_text = self.price_edit.text().replace("$", "").replace(",", "").strip()
            if price_text:
                unit_price = Decimal(str(float(price_text)))
            else:
                unit_price = Decimal('0.0')
            total = Decimal(str(quantity)) * unit_price
            payment_type = self.payment_combo.currentText()
            self.item.payment_category = payment_type
            if "Deposit" in payment_type or "Down Payment" in payment_type:
                down_payment = total * Decimal("0.5")
                payment_due = total - down_payment
            elif "Final Payment" in payment_type:
                down_payment = Decimal("0.0")
                payment_due = Decimal("0.0")
            else:
                down_payment = Decimal("0.0")
                payment_due = total
            self.item.quantity = quantity
            self.item.unit_price = unit_price
            self.item.down_payment = down_payment
            self.item.payment_due = payment_due
            self.payment_due_label.setText(f"${float(payment_due):.2f}")
            self.total_label.setText(Currency.format(total))
        except Exception as e:
            _log.warning("Error updating total: %s", e)
    
    def get_item(self):
        try:
            price_text = self.price_edit.text().replace("$", "").strip()
            unit_price = Decimal(str(float(price_text))) if price_text else Decimal("0.0")
            quantity = self.qty_spin.value()
            total = Decimal(str(quantity)) * unit_price
            payment_category = self.payment_combo.currentText()
            if "Deposit" in payment_category or "Down Payment" in payment_category:
                down_payment = total * Decimal("0.5")
                payment_due = total - down_payment
            elif "Final Payment" in payment_category:
                down_payment = Decimal("0.0")
                payment_due = Decimal("0.0")
            else:
                down_payment = Decimal("0.0")
                payment_due = total
            return InvoiceItem(
                project_number=self.project_number_edit.text(),
                description=self.desc_edit.text(),
                plant=self.plant_edit.text(),
                quantity=quantity,
                unit_price=float(unit_price),
                down_payment=float(down_payment),
                payment_due=float(payment_due),
                payment_category=payment_category
            )
        except Exception as e:
            _log.warning("Error in get_item: %s", e)
            return InvoiceItem()


class InvoiceHistoryTab(QtWidgets.QWidget):
    """Enhanced Invoice History Tab with direct invoice history view and Firebase sync"""
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.current_client = None
        self.init_ui()
    
    def init_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        self.stacked_widget = QtWidgets.QStackedWidget()
        
        self.client_view = ClientListWidget()
        self.client_view.client_selected.connect(self.show_invoice_history)
        self.stacked_widget.addWidget(self.client_view)
        
        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.addWidget(self.history_view_placeholder)
        
        main_layout.addWidget(self.stacked_widget)
        self.show_client_view()
        self._auto_refresh_timer = QtCore.QTimer(self)
        self._auto_refresh_timer.setInterval(5000)
        self._auto_refresh_timer.timeout.connect(self.refresh_invoices_immediately)
        self._auto_refresh_timer.start()
        self._auto_refresh_timer = QtCore.QTimer(self)
        self._auto_refresh_timer.setInterval(5000)
        self._auto_refresh_timer.timeout.connect(self.refresh_invoices_immediately)
        self._auto_refresh_timer.start()
    
    def add_firebase_sync_button(self):
        if not FIREBASE_AVAILABLE:
            return
    
    def sync_to_firebase(self):
        if not FIREBASE_AVAILABLE:
            QtWidgets.QMessageBox.information(self, "Cloud Sync", "Firebase integration is not available.")
            return
        try:
            progress_dialog = QtWidgets.QProgressDialog("Syncing invoices to cloud...", "Cancel", 0, 100, self)
            progress_dialog.setWindowTitle("Cloud Sync")
            progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
            progress_dialog.show()
            QtWidgets.QMessageBox.information(self, "Cloud Sync", "✅ Invoices are automatically synced to Firebase!\n\nYour data is already backed up and accessible from anywhere.")
            progress_dialog.close()
        except Exception as e:
            progress_dialog.close()
            QtWidgets.QMessageBox.critical(self, "Sync Error", f"Error during cloud sync: {e}")
    
    def show_client_view(self):
        if self.stacked_widget.indexOf(self.history_view_placeholder) != -1:
            self.stacked_widget.removeWidget(self.history_view_placeholder)
        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.insertWidget(1, self.history_view_placeholder)
        self.stacked_widget.setCurrentIndex(0)
        self.client_view.load_clients()
    
    def show_invoice_history(self, client_name: str):
        self.current_client = client_name
        history_view = InvoiceHistoryViewWidget(client_name, compact=True)
        history_view.back_clicked.connect(self.show_client_view)
        index = self.stacked_widget.indexOf(self.history_view_placeholder)
        self.stacked_widget.removeWidget(self.history_view_placeholder)
        self.stacked_widget.insertWidget(index, history_view)
        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.insertWidget(index + 1, self.history_view_placeholder)
        self.stacked_widget.setCurrentIndex(index)

    def refresh_data(self):
        current_widget = self.stacked_widget.currentWidget()
        if isinstance(current_widget, ClientListWidget):
            current_widget.load_clients()
        elif isinstance(current_widget, InvoiceHistoryViewWidget):
            self.show_invoice_history(self.current_client)

    def load_history(self):
        self.show_client_view()

    def refresh_invoices_immediately(self):
        current_widget = self.stacked_widget.currentWidget()
        if isinstance(current_widget, ClientListWidget):
            current_widget.load_clients()
        elif self.current_client and isinstance(current_widget, InvoiceHistoryViewWidget):
            search_text = current_widget.date_range_widget.search_bar.text() if hasattr(current_widget, "date_range_widget") else ""
            self.show_invoice_history(self.current_client)
            refreshed_widget = self.stacked_widget.currentWidget()
            if search_text and isinstance(refreshed_widget, InvoiceHistoryViewWidget):
                refreshed_widget.date_range_widget.search_bar.setText(search_text)


class InvoiceHistoryTab(QtWidgets.QWidget):
    """Enhanced Invoice History Tab with direct invoice history view and Firebase sync"""

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.current_client = None
        self.init_ui()

        # Wire up the thread-safe invoice-status refresh signal so payment_tracker's
        # _recompute_invoice_status (background thread) can trigger a UI refresh here.
        global _invoice_status_signaler
        if _invoice_status_signaler is None:
            _invoice_status_signaler = _InvoiceStatusRefreshSignaler()
        _invoice_status_signaler.invoice_status_changed.connect(
            self._on_invoice_status_changed_externally,
            QtCore.Qt.QueuedConnection
        )

    def _on_invoice_status_changed_externally(self, invoice_number: str, new_status: str, received_date: str):
        """Called on the GUI thread when payment_tracker recomputes an invoice status
        after a payment add/edit/delete. Updates the invoice row instantly in-memory
        (same pattern as balance sheet) then schedules a background Firebase reload
        to confirm the data."""
        # 1. Instant in-memory update — status badge + received date, no Firebase round-trip
        current = self.stacked_widget.currentWidget()
        if isinstance(current, InvoiceHistoryViewWidget):
            current.update_invoice_row_immediately(invoice_number, new_status, received_date)
        # 2. Sync balance sheet transaction table in-memory
        bs = None
        try:
            bs = getattr(self.main_window, "balance_sheet_tab", None)
            if bs and hasattr(bs, "update_revenue_entry_status"):
                bs.update_revenue_entry_status(invoice_number, new_status, received_date)
        except Exception:
            pass
        # 3. Update FirebaseManager in-memory cache so dashboard reads fresh status
        try:
            from main import FirebaseManager
            cache = FirebaseManager._invoices_cache
            if cache is not None:
                for _inv in cache:
                    _meta = _inv.get("meta", _inv)
                    if _meta.get("invoice_number") == invoice_number:
                        _meta["status"] = new_status
                        if received_date:
                            _meta["received_date"] = received_date
                        break
        except Exception:
            pass
        # 4. Refresh dashboard KPIs immediately from updated cache
        try:
            dt = getattr(self.main_window, "dashboard_tab", None)
            if dt and hasattr(dt, "refresh"):
                dt.refresh(force_firebase=False)
        except Exception:
            pass
        # 5. Delayed Firebase reload — confirms correct data for both invoice history and
        #    balance sheet transaction table (runs after _update_received_date_to_latest
        #    has had time to complete its Firebase write)
        QtCore.QTimer.singleShot(1500, self.refresh_invoices_immediately)
        if bs and hasattr(bs, "_refresh_all_revenue_background"):
            QtCore.QTimer.singleShot(1500, bs._refresh_all_revenue_background)
        # 6. Refresh project tab's Recent Invoices to show updated status
        try:
            pt = getattr(self.main_window, 'project_tab', None)
            if pt and hasattr(pt, 'refresh_recent_invoices'):
                QtCore.QTimer.singleShot(1600, pt.refresh_recent_invoices)
        except Exception:
            pass

    def init_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.stacked_widget = QtWidgets.QStackedWidget()

        self.client_view = ClientListWidget()
        self.client_view.client_selected.connect(self.show_invoice_history)
        self.stacked_widget.addWidget(self.client_view)

        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.addWidget(self.history_view_placeholder)

        main_layout.addWidget(self.stacked_widget)
        self.show_client_view()

    def add_firebase_sync_button(self):
        if not FIREBASE_AVAILABLE:
            return
    
    def sync_to_firebase(self):
        if not FIREBASE_AVAILABLE:
            QtWidgets.QMessageBox.information(self, "Cloud Sync", "Firebase integration is not available.")
            return
        try:
            progress_dialog = QtWidgets.QProgressDialog("Syncing invoices to cloud...", "Cancel", 0, 100, self)
            progress_dialog.setWindowTitle("Cloud Sync")
            progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
            progress_dialog.show()
            QtWidgets.QMessageBox.information(self, "Cloud Sync", "✅ Invoices are automatically synced to Firebase!\n\nYour data is already backed up and accessible from anywhere.")
            progress_dialog.close()
        except Exception as e:
            progress_dialog.close()
            QtWidgets.QMessageBox.critical(self, "Sync Error", f"Error during cloud sync: {e}")
    
    def show_client_view(self):
        if self.stacked_widget.indexOf(self.history_view_placeholder) != -1:
            self.stacked_widget.removeWidget(self.history_view_placeholder)
        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.insertWidget(1, self.history_view_placeholder)
        self.stacked_widget.setCurrentIndex(0)
        self.client_view.load_clients()
    
    def show_invoice_history(self, client_name: str):
        self.current_client = client_name
        history_view = InvoiceHistoryViewWidget(client_name, compact=False)
        history_view.back_clicked.connect(self.show_client_view)
        index = self.stacked_widget.indexOf(self.history_view_placeholder)
        self.stacked_widget.removeWidget(self.history_view_placeholder)
        self.stacked_widget.insertWidget(index, history_view)
        self.history_view_placeholder = QtWidgets.QWidget()
        self.stacked_widget.insertWidget(index + 1, self.history_view_placeholder)
        self.stacked_widget.setCurrentIndex(index)
    
    def refresh_data(self):
        current_widget = self.stacked_widget.currentWidget()
        if isinstance(current_widget, ClientListWidget):
            current_widget.load_clients()
        elif isinstance(current_widget, InvoiceHistoryViewWidget):
            self.show_invoice_history(self.current_client)

    def load_history(self):
        self.show_client_view()

    def navigate_to_invoice(self, client_name: str, invoice_number: str):
        """Navigate to client's invoice history and scroll to invoice_number."""
        self.show_invoice_history(client_name)
        view = self.stacked_widget.currentWidget()
        if isinstance(view, InvoiceHistoryViewWidget) and invoice_number:
            view.scroll_to_invoice(invoice_number)

    def refresh_invoices_immediately(self):
        current_widget = self.stacked_widget.currentWidget()
        if isinstance(current_widget, ClientListWidget):
            current_widget.load_clients()
        elif self.current_client and isinstance(current_widget, InvoiceHistoryViewWidget):
            search_text = current_widget.date_range_widget.search_bar.text() if hasattr(current_widget, "date_range_widget") else ""
            self.show_invoice_history(self.current_client)
            refreshed_widget = self.stacked_widget.currentWidget()
            if search_text and isinstance(refreshed_widget, InvoiceHistoryViewWidget):
                refreshed_widget.date_range_widget.search_bar.setText(search_text)
