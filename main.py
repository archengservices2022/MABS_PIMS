# arch_invoice_generator.py
import sys
import os
import csv
import json
import subprocess
import platform
import shutil
import traceback
import base64
import tempfile
from datetime import datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import secrets
import hashlib
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate, Image, Spacer, PageBreak
from reportlab.lib import colors
from PIL import Image as PILImage

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import pyqtSignal

from balance_sheet_tab import BalanceSheetTab, BalanceSheetFirebaseManager
from job_form_tab import JobFormTab
from expenses_tab import ExpensesTab
from finance_overview_tab import FinanceOverviewTab
from update_checker import UpdateChecker
from update_indicator import UpdateIndicator
from app_logger import setup_logging, get_logger
from auth_utils import hash_password, password_needs_rehash, verify_password
from access_control import (
    ACTION_CONVERT_QUOTE_TO_INVOICE,
    ACTION_CONVERT_QUOTE_TO_PROJECT,
    PAGE_FINANCIAL,
    PAGE_PROJECTS,
    PAGE_SETTINGS,
    allowed_pages_for_role,
    allowed_stack_pages_for_role,
    can_access_page,
    can_perform_action,
    first_allowed_stack_page,
    normalize_role,
    profile_is_active,
)

logger = setup_logging()
log = get_logger("main")

# Ensure stdout/stderr use UTF-8 so emoji in legacy print() calls don't crash on Windows
import io as _io
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

INVOICE_CACHE = None

def resource_path(relative_path: str) -> Path:
    """
    Resolve paths correctly for:
    - normal Python run
    - PyInstaller EXE
    """
    try:
        base_path = sys._MEIPASS  # PyInstaller temp dir
    except Exception:
        base_path = Path(__file__).resolve().parent
    return Path(base_path) / relative_path

def app_base_path() -> Path:
    """Return the editable app folder, even when running from PyInstaller."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

def cleanup_temp_backups():
    temp_dir = Path(os.environ.get("TEMP", "."))
    for f in temp_dir.glob("*_old.tmp"):
        try:
            f.unlink()
        except:
            pass

def cleanup_old_backup():
    exe = Path(sys.executable)
    bak = exe.with_suffix(".bak")
    if bak.exists():
        try:
            bak.unlink()
        except:
            pass

# ---------- UPDATE CLEANUP (VERY IMPORTANT) ----------
def cleanup_update_leftovers():
    """
    Deletes updater and backup files after update.
    This runs inside the NEW app, so Windows allows deletion.
    """
    try:
        import time
        time.sleep(2)  # allow updater process to fully exit

        app_dir = Path(sys.executable).parent

        exe = Path(sys.executable)
        targets = [
            app_dir / "invoice_updater.exe",
            app_dir / "main.bak",
            exe.with_suffix(".bak"),   # e.g. MABS_Invoice.bak
        ]

        for file in targets:
            if file.exists():
                try:
                    file.unlink()
                except Exception:
                    pass
    except Exception:
        pass

# Add this after the imports and before the Config class

def send_welcome_email(email: str, username: str, role: str, password: str = "") -> bool:
    """Send welcome email to new user with login credentials"""
    try:
        # Get settings
        settings = {}
        if Config.SETTINGS_FILE.exists():
            with open(Config.SETTINGS_FILE, encoding="utf-8") as f:
                settings = json.load(f)
        
        email_config = settings.get("email", {})
        smtp_host = email_config.get("smtp_host", "smtp.gmail.com")
        smtp_port = int(email_config.get("smtp_port", 587))
        smtp_user = email_config.get("smtp_user", "")
        smtp_pass = email_config.get("smtp_pass", "")
        app_name = email_config.get("app_name", "MABS Engineering PIMS")

        if not smtp_user or not smtp_pass:
            log.warning("SMTP not configured  -  cannot send welcome email")
            return False

        # Role display names
        role_display = {
            'admin': 'Administrator (Full Access)',
            'projects': 'Project Manager (Projects & Invoices)',
            'finance': 'Finance Manager (Financial Management)',
            'sales': 'Sales Representative (Quote Forms)'
        }
        role_name = role_display.get(role, role.capitalize())

        # Build email content
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f7fb; padding: 30px; }}
                .container {{ max-width: 500px; margin: auto; background: white; border-radius: 16px;
                            padding: 32px; border: 1px solid #d8e2ec; }}
                .header {{ text-align: center; margin-bottom: 24px; }}
                h2 {{ color: #0f172a; margin-top: 0; text-align: center; }}
                p {{ color: #475569; font-size: 15px; line-height: 1.5; }}
                .credentials {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
                               padding: 16px; margin: 20px 0; font-family: monospace; }}
                .badge {{ display: inline-block; background: #eefaf8; color: #00756f;
                         padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: bold; }}
                hr {{ border: none; border-top: 1px solid #e2e8f0; margin: 20px 0; }}
                .footer {{ text-align: center; color: #94a3b8; font-size: 11px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">    </div>
                <h2>Welcome to {app_name}!</h2>
                <p>Hello <strong>{username}</strong>,</p>
                <p>Your account has been created with <span class="badge">{role_name}</span> role permissions.</p>
                <div class="credentials">
                    <strong>Your Login Credentials:</strong><br><br>
                      "  <strong>Email:</strong> {email}<br>
                      "' <strong>Password:</strong> {password}<br><br>
                    <span style="font-size: 12px; color: #64748b;">
                        a      Please change your password after first login for security.
                    </span>
                </div>
                <p>You can now log in to the PIMS (Project & Invoice Management System)
                and start working according to your role permissions.</p>
                <hr>
                <div class="footer">
                    {app_name} a  Project & Invoice Management System<br>
                    a  2025 MABS Engineering LLC
                </div>
            </div>
        </body>
        </html>
        """
        
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import smtplib
        
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Welcome to {app_name}"
        msg["From"] = smtp_user
        msg["To"] = email
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, email, msg.as_string())
        
        log.info(f"Welcome email sent to {email}")
        return True
        
    except Exception as exc:
        log.warning(f"Welcome email failed: {exc}")
        return False

# ---------- Configuration ----------
class Config:
    BASE_DIR = app_base_path()
    RESOURCE_DIR = resource_path(".")
    DATA_DIR = BASE_DIR / "data"
    ASSETS_DIR = RESOURCE_DIR / "assets"
    INVOICES_DIR = DATA_DIR / "invoices"

    CLIENTS_FILE = DATA_DIR / "clients.json"
    COUNTER_FILE = DATA_DIR / "invoice_counter.json"
    SETTINGS_FILE = DATA_DIR / "settings.json"
    PROJECTS_FILE = DATA_DIR / "projects.json"
    LOGO_FILE = ASSETS_DIR / "logo.png"
    ZELLE_QR_FILE = ASSETS_DIR / "payment.jpeg"
    PRIVATE_SERVICE_ACCOUNT_FILE = Path.home() / ".mabs" / "servicekey.json"
    # When running as a PyInstaller .exe, bundled files land in sys._MEIPASS
    _MEIPASS_KEY = Path(getattr(sys, "_MEIPASS", "")) / "data" / "servicekey.json"
    SERVICE_ACCOUNT_FILE = Path(
        os.environ.get(
            "MABS_FIREBASE_SERVICE_ACCOUNT",
            str(PRIVATE_SERVICE_ACCOUNT_FILE) if PRIVATE_SERVICE_ACCOUNT_FILE.exists()
            else str(_MEIPASS_KEY) if _MEIPASS_KEY.exists()
            else str(DATA_DIR / "servicekey.json"),
        )
    ).expanduser()

    # Loaded from settings.json; these are the fallback defaults
    COMPANY = {
        "name": "MABS Engineering LLC",
        "address": "15455 Manchester Rd, PO Box 1144\nManchester, MO 63011",
        "email": "admin@habbengineering.com",
        "phone": "314-303-0004",
        "website": "www.mabs-engineeringg.com",
    }

    DEFAULT_TERMS = "Thank you for your business!\nBest regards,\n\nMABS Engineering LLC"

    # User accounts loaded from settings.json.
    USERS: Dict[str, Dict] = {}

    @classmethod
    def load(cls) -> None:
        """Overlay defaults with values from data/settings.json."""
        try:
            if cls.SETTINGS_FILE.exists():
                with open(cls.SETTINGS_FILE, encoding="utf-8") as f:
                    data = json.load(f)
                if "company" in data:
                    cls.COMPANY.update(data["company"])
                    # Keep DEFAULT_TERMS in sync with the saved preference
                    dt = (data["company"].get("default_terms") or "").strip()
                    if dt:
                        cls.DEFAULT_TERMS = dt
                if "users" in data:
                    cls.USERS = data["users"]
                log.info("Settings loaded from %s", cls.SETTINGS_FILE)
            else:
                log.warning("settings.json not found  -  using built-in defaults")
        except Exception as exc:
            log.error("Failed to load settings.json: %s", exc)

    @classmethod
    def get_logo_path(cls) -> Optional[Path]:
        """Return the uploaded company logo, falling back to the bundled asset."""
        saved_logo = str(cls.COMPANY.get("logo_path", "") or "").strip()
        if saved_logo:
            saved_path = Path(saved_logo)
            if saved_path.exists():
                return saved_path
        if cls.LOGO_FILE.exists():
            return cls.LOGO_FILE
        return None

    @classmethod
    def overlay_from_firebase(cls) -> bool:
        """Overlay Config with live Firebase settings (source of truth).
        Called after Firebase is connected, so FirebaseManager is available."""
        try:
            fb = FirebaseManager.load_settings_from_firebase()
            if not fb:
                return False
            if fb.get("company"):
                cls.COMPANY.update(fb["company"])
                dt = (fb["company"].get("default_terms") or "").strip()
                if dt:
                    cls.DEFAULT_TERMS = dt
            log.info("Config overlaid from Firebase /settings")
            return True
        except Exception as exc:
            log.warning("overlay_from_firebase failed: %s", exc)
            return False

    @classmethod
    def setup_directories(cls) -> None:
        cls.DATA_DIR.mkdir(parents=True, exist_ok=True)
        cls.INVOICES_DIR.mkdir(parents=True, exist_ok=True)

# Load settings as soon as Config is defined
Config.load()

# ---------- Firebase Integration ----------
# Initialize Firebase availability flag first
FIREBASE_AVAILABLE = False
FIREBASE_CONFIG = {}

try:
    import firebase_admin
    from firebase_admin import credentials, db
    from firebase_admin.exceptions import FirebaseError
    
    # Firebase Configuration
    FIREBASE_CONFIG = {
        "apiKey": "AIzaSyD6F6T_KIZ90TkCOL03-jSXTeuPM5WVwJY",
        "authDomain": "invoice-7fe93.firebaseapp.com",
        "databaseURL": "https://invoice-7fe93-default-rtdb.firebaseio.com",
        "projectId": "invoice-7fe93",
        "storageBucket": "invoice-7fe93.firebasestorage.app",
        "messagingSenderId": "600598976875",
        "appId": "1:600598976875:web:4f974b01afda7f69ddd42e",
        "measurementId": "G-EKEE52BBKH"
    }
    
    # Initialize Firebase Admin SDK
    if Config.SERVICE_ACCOUNT_FILE.exists():
        try:
            cred = credentials.Certificate(str(Config.SERVICE_ACCOUNT_FILE))
            if not firebase_admin._apps:
                firebase_admin.initialize_app(cred, {
                    'databaseURL': FIREBASE_CONFIG['databaseURL']
                })
            FIREBASE_AVAILABLE = True
            log.info("Firebase initialized successfully")
        except Exception as e:
            log.error("Firebase initialization failed: %s", e)
            FIREBASE_AVAILABLE = False
    else:
        log.warning("Service account key not found  -  Firebase disabled (expected: %s)", Config.SERVICE_ACCOUNT_FILE)
        FIREBASE_AVAILABLE = False

except ImportError:
    log.warning("firebase-admin not installed  -  Firebase disabled")
    FIREBASE_AVAILABLE = False
except Exception as e:
    log.error("Firebase setup error: %s", e)
    FIREBASE_AVAILABLE = False

# Update the FirebaseManager class in main.py

class FirebaseManager:
    """Handles Firebase Authentication and Realtime Database operations"""
    last_auth_error = ""
    last_auth_message = ""
    _invoices_cache = None  # None = never loaded; list = populated by load_invoices()

    @staticmethod
    def _set_auth_error(code: str = "", message: str = "") -> None:
        FirebaseManager.last_auth_error = code
        FirebaseManager.last_auth_message = message
    
    @staticmethod
    def initialize_default_admin():
        """Initialize default admin user if no users exist in Auth"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available, cannot init admin")
            return False
        
        try:
            from firebase_admin import auth
            
            # Check if any users exist in Firebase Auth
            users = auth.list_users().iterate_all()
            
            if not any(True for _ in users):
                log.info("No users found in Auth, creating default admin...")
                
                # Create default admin user in Firebase Auth
                admin_email = "ashajyothi.gadhi@gmail.com"
                admin_password = "admin123"
                
                try:
                    user = auth.create_user(
                        email=admin_email,
                        password=admin_password,
                        display_name="admin",
                        email_verified=True
                    )
                    
                    # Also save to Realtime Database
                    user_data = {
                        'username': 'admin',
                        'email': admin_email,
                        'role': 'admin',
                        'active': True,
                        'created_at': datetime.now(timezone.utc).isoformat(),
                        'updated_at': datetime.now(timezone.utc).isoformat(),
                        'firebase_uid': user.uid
                    }
                    
                    ref = db.reference('/users')
                    ref.child(user.uid).set(user_data)
                    
                    log.info(f"Default admin user created with email: {admin_email}")
                    return True
                    
                except auth.EmailAlreadyExistsError:
                    log.info("Admin user already exists in Auth")
                    return True
                    
            return True
            
        except Exception as e:
            log.error(f"Error initializing default admin: {e}")
            return False
    
    @staticmethod
    def send_password_reset_email(email: str) -> bool:
        """Send password reset email using Firebase REST API"""
        try:
            import requests

            API_KEY = FIREBASE_CONFIG.get("apiKey")

            if not API_KEY:
                log.error("Firebase API key missing")
                return False

            # a ... STEP 1: Check if email exists in Firebase Auth
            check_url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={API_KEY}"

            check_payload = {
                "email": email,
                "password": "dummy_password",  # intentionally wrong
                "returnSecureToken": False
            }

            check_response = requests.post(check_url, json=check_payload)

            #   "  IMPORTANT LOGIC
            if check_response.status_code != 200:
                error_msg = check_response.json().get("error", {}).get("message", "")

                # If email NOT FOUND a ' stop here
                if "EMAIL_NOT_FOUND" in error_msg:
                    log.warning(f"Email not found in Firebase Auth: {email}")
                    return False

            # a ... STEP 2: Send reset email (your existing logic)
            url = f"https://identitytoolkit.googleapis.com/v1/accounts:sendOobCode?key={API_KEY}"

            payload = {
                "requestType": "PASSWORD_RESET",
                "email": email
            }

            response = requests.post(url, json=payload)

            print("RESET RESPONSE:", response.status_code, response.text)

            if response.status_code == 200:
                log.info(f"Password reset email sent to {email}")
                return True
            else:
                log.error(f"Reset failed: {response.text}")
                return False

        except Exception as e:
            log.error(f"Error sending reset email: {e}")
            return False
    
    @staticmethod
    def save_user_to_firebase(username: str, email: str, password: str, role: str) -> bool:
        """Save user credentials to Firebase Authentication and profile to Database"""
        if not FIREBASE_AVAILABLE:
            return False
        
        try:
            from firebase_admin import auth

            username = username.strip()
            email = email.strip().lower()
            normalized_role = normalize_role(role)
            users_ref = db.reference('/users')

            existing_username = users_ref.order_by_child('username').equal_to(username).get()
            if existing_username:
                log.warning(f"Username {username} already exists in user profiles")
                return False
            
            # Check if email already exists in Auth
            try:
                existing_user = auth.get_user_by_email(email)
                log.warning(f"Email {email} already exists in Auth")
                return False
            except auth.UserNotFoundError:
                pass
            
            # Create user in Firebase Authentication
            user = auth.create_user(
                email=email,
                password=password,
                display_name=username,
                email_verified=False  # User will verify via email
            )
            
            # Save user profile to Realtime Database
            user_data = {
                'username': username,
                'email': email,
                'role': normalized_role,
                'active': True,
                'firebase_uid': user.uid,
                'created_at': datetime.now(timezone.utc).isoformat(),
                'updated_at': datetime.now(timezone.utc).isoformat(),
                'reset_token': None,
                'reset_token_expiry': None
            }
            
            try:
                users_ref.child(user.uid).set(user_data)
            except Exception:
                try:
                    auth.delete_user(user.uid)
                except Exception as cleanup_exc:
                    log.warning(f"Could not roll back Auth user after profile save failure: {cleanup_exc}")
                raise
            
            log.info(f"User {username} created in Auth with UID {user.uid}")
            return True
            
        except Exception as e:
            log.error(f"Error saving user to Firebase: {e}")
            return False
    
    @staticmethod
    def validate_user_email(email: str, password: str) -> tuple:
        """Validate user credentials using Firebase Authentication.
        Returns (is_valid, username, email, role)
        NO POPUPS - returns False if validation fails
        """
        FirebaseManager._set_auth_error()
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available")
            FirebaseManager._set_auth_error(
                "firebase_unavailable",
                "Authentication service is unavailable. Please restart the app or contact an administrator.",
            )
            return (False, "", "", "")
        
        try:
            import requests
            
            # Step 1: Verify password via Firebase REST API (THIS IS THE ONLY WAY)
            api_key = FIREBASE_CONFIG.get("apiKey")
            if not api_key:
                log.error("Firebase API key missing")
                FirebaseManager._set_auth_error(
                    "config_error",
                    "Firebase configuration is missing. Please contact an administrator.",
                )
                return (False, "", "", "")

            url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={api_key}"
            
            payload = {
                "email": email,
                "password": password,
                "returnSecureToken": True
            }
            
            response = requests.post(url, json=payload, timeout=10)
            
            # If authentication fails, return False immediately - NO POPUP
            if response.status_code != 200:
                log.warning(f"Password verification failed for {email}")
                FirebaseManager._set_auth_error(
                    "invalid_credentials",
                    "Invalid email or password. Please try again.",
                )
                return (False, "", "", "")

            auth_data = response.json()
            firebase_uid = auth_data.get('localId', '')
            if not firebase_uid:
                log.warning("Firebase auth response missing localId for %s", email)
                FirebaseManager._set_auth_error(
                    "auth_response_error",
                    "Firebase sign-in response was incomplete. Please try again.",
                )
                return (False, "", "", "")
            
            # Step 2: Load the app profile that defines role/access.
            # Firebase Auth proves identity; /users profile grants app permission.
            try:
                user_data = db.reference(f'/users/{firebase_uid}').get()

                if not user_data:
                    log.warning("Authenticated user %s has no app profile; login denied", email)
                    FirebaseManager._set_auth_error(
                        "missing_profile",
                        "This Firebase account has no app profile. Ask an administrator to create one in Settings.",
                    )
                    return (False, "", "", "")

                if not profile_is_active(user_data):
                    log.warning("Authenticated user %s has an inactive app profile; login denied", email)
                    FirebaseManager._set_auth_error(
                        "inactive_profile",
                        "This account is inactive. Please contact an administrator.",
                    )
                    return (False, "", "", "")

                username = user_data.get('username', '') or email.split('@')[0]
                raw_role = str(user_data.get('role', 'sales') or 'sales').strip().lower()
                role = normalize_role(raw_role)
                if raw_role != role:
                    log.warning("User %s has invalid role '%s'; login denied", email, raw_role)
                    FirebaseManager._set_auth_error(
                        "invalid_role",
                        "This account has an invalid role. Please contact an administrator.",
                    )
                    return (False, "", "", "")

                log.info(f"User {username} authenticated successfully with role: {role}")
                return (True, username, email, role)
                    
            except Exception as e:
                log.error(f"Error reading user profile: {e}")
                FirebaseManager._set_auth_error(
                    "profile_error",
                    "Could not read your app profile. Please try again or contact an administrator.",
                )
                return (False, "", "", "")
            
        except requests.exceptions.Timeout:
            log.error("Firebase auth request timed out")
            FirebaseManager._set_auth_error(
                "timeout",
                "Firebase sign-in timed out. Check your connection and try again.",
            )
            return (False, "", "", "")
        except requests.exceptions.ConnectionError:
            log.error("No internet connection for Firebase auth")
            FirebaseManager._set_auth_error(
                "connection_error",
                "Could not connect to Firebase. Check your internet connection and try again.",
            )
            return (False, "", "", "")
        except Exception as e:
            log.error(f"Unexpected error validating user: {e}")
            FirebaseManager._set_auth_error(
                "unexpected_error",
                "Unexpected login error. Please restart the app or contact an administrator.",
            )
            return (False, "", "", "")
    
    @staticmethod
    def get_user_by_email(email: str) -> Optional[Dict]:
        """Get user data by email address"""
        if not FIREBASE_AVAILABLE:
            return None
        
        try:
            from firebase_admin import auth
            
            # Get user from Auth first
            user_record = auth.get_user_by_email(email)
            
            # Get profile from Database
            ref = db.reference(f'/users/{user_record.uid}')
            user_data = ref.get()
            
            if user_data:
                user_data['firebase_uid'] = user_record.uid
                return user_data
            
            return None
            
        except Exception as e:
            log.error(f"Error getting user by email: {e}")
            return None
    
    @staticmethod
    def get_user_by_username(username: str) -> Optional[Dict]:
        """Get user data by username"""
        if not FIREBASE_AVAILABLE:
            return None
        
        try:
            ref = db.reference('/users')
            users = ref.order_by_child('username').equal_to(username).get()
            
            if users:
                uid = list(users.keys())[0]
                user_data = users[uid]
                user_data['firebase_uid'] = uid
                return user_data
            
            return None
            
        except Exception as e:
            log.error(f"Error getting user by username: {e}")
            return None
    
    @staticmethod
    def get_all_users() -> List[Dict]:
        """Get all users from Firebase"""
        if not FIREBASE_AVAILABLE:
            return []
        
        try:
            ref = db.reference('/users')
            users_data = ref.get()
            
            users = []
            if users_data:
                for uid, user_data in users_data.items():
                    clean_user = {
                        'firebase_uid': uid,
                        'username': user_data.get('username', ''),
                        'email': user_data.get('email', ''),
                        'role': user_data.get('role', 'sales'),
                        'active': user_data.get('active', True),
                        'created_at': user_data.get('created_at', ''),
                        'updated_at': user_data.get('updated_at', '')
                    }
                    users.append(clean_user)
            
            return users
            
        except Exception as e:
            log.error(f"Error getting all users: {e}")
            return []

    @staticmethod
    def load_sales_people() -> List[Dict]:
        """Load sales people from manual records and active sales user profiles."""
        if not FIREBASE_AVAILABLE:
            return []

        def person_key(person: Dict) -> str:
            email = str(person.get("email", "") or "").strip().lower()
            if email:
                return f"email:{email}"
            name = str(person.get("name", "") or "").strip().lower()
            return f"name:{name}" if name else ""

        try:
            people_by_key: Dict[str, Dict] = {}

            manual_data = db.reference('/sales_persons').get() or {}
            for sales_id, person in manual_data.items():
                if not isinstance(person, dict):
                    continue
                clean_person = dict(person)
                clean_person["firebase_id"] = sales_id
                clean_person.setdefault("source", "sales_persons")
                key = person_key(clean_person) or f"manual:{sales_id}"
                people_by_key[key] = clean_person

            users_data = db.reference('/users').get() or {}
            for uid, user_data in users_data.items():
                if not isinstance(user_data, dict):
                    continue
                if normalize_role(user_data.get("role")) != "sales":
                    continue
                if not profile_is_active(user_data):
                    continue

                email = str(user_data.get("email", "") or "").strip().lower()
                username = str(user_data.get("username", "") or "").strip()
                name = username or email.split("@")[0] or "Sales User"
                profile_person = {
                    "name": name,
                    "phone": str(
                        user_data.get("phone", "")
                        or user_data.get("phone_number", "")
                        or user_data.get("mobile", "")
                        or "-"
                    ),
                    "email": email,
                    "user_uid": uid,
                    "source": "user_profile",
                    "updated_at": user_data.get("updated_at", ""),
                    "created_at": user_data.get("created_at", ""),
                }

                key = person_key(profile_person) or f"user:{uid}"
                if key in people_by_key:
                    people_by_key[key]["user_uid"] = uid
                    people_by_key[key]["user_role"] = "sales"
                    people_by_key[key]["source"] = "sales_persons,user_profile"
                    if not people_by_key[key].get("email"):
                        people_by_key[key]["email"] = email
                    if not people_by_key[key].get("phone"):
                        people_by_key[key]["phone"] = profile_person["phone"]
                else:
                    people_by_key[key] = profile_person

            return sorted(
                people_by_key.values(),
                key=lambda person: str(person.get("name", "") or "").lower(),
            )

        except Exception as e:
            log.error(f"Error loading sales people: {e}")
            return []
    
    @staticmethod
    def create_password_reset_token(email: str) -> Optional[str]:
        """Send password reset email using Firebase Authentication"""
        if not FIREBASE_AVAILABLE:
            return None
        
        try:
            from firebase_admin import auth
            
            # Send password reset email via Firebase
            # This will send an email with a link to reset password
            reset_link = auth.generate_password_reset_link(email)
            
            # Store that a reset was requested (optional)
            user = auth.get_user_by_email(email)
            if user:
                ref = db.reference(f'/users/{user.uid}')
                ref.update({
                    'reset_requested_at': datetime.now(timezone.utc).isoformat(),
                    'updated_at': datetime.now(timezone.utc).isoformat()
                })
            
            # Return a token (for compatibility, though Firebase handles it)
            return "reset_email_sent"
            
        except Exception as e:
            log.error(f"Error creating password reset: {e}")
            return None
    
    @staticmethod
    def update_user_password_firebase(username: str, new_password: str) -> bool:
        """Update user password in Firebase Authentication"""
        if not FIREBASE_AVAILABLE:
            return False
        
        try:
            from firebase_admin import auth
            
            user = FirebaseManager.get_user_by_username(username)
            if not user:
                return False
            
            # Update password in Auth
            auth.update_user(user['firebase_uid'], password=new_password)
            
            # Update timestamp in Database
            ref = db.reference(f'/users/{user["firebase_uid"]}')
            ref.update({
                'updated_at': datetime.now(timezone.utc).isoformat()
            })
            
            log.info(f"Password updated for {username}")
            return True
            
        except Exception as e:
            log.error(f"Error updating password: {e}")
            return False
    
    @staticmethod
    def update_user_role_firebase(username: str, new_role: str) -> bool:
        """Update user role in Realtime Database"""
        if not FIREBASE_AVAILABLE:
            return False
        
        try:
            user = FirebaseManager.get_user_by_username(username)
            if not user:
                return False
            
            ref = db.reference(f'/users/{user["firebase_uid"]}')
            ref.update({
                'role': normalize_role(new_role),
                'updated_at': datetime.now(timezone.utc).isoformat()
            })
            
            log.info(f"Updated role for {username} to {new_role}")
            return True
            
        except Exception as e:
            log.error(f"Error updating user role: {e}")
            return False
    
    @staticmethod
    def delete_user_from_firebase(username: str) -> bool:
        """Delete user from Firebase Authentication and Database"""
        if not FIREBASE_AVAILABLE:
            return False
        
        try:
            from firebase_admin import auth

            user = FirebaseManager.get_user_by_username(username)
            if not user:
                return False

            if username == 'admin':
                return False

            firebase_uid = user.get('firebase_uid')

            # Try to delete from Authentication; skip if no Auth account exists
            try:
                if firebase_uid:
                    auth.delete_user(firebase_uid)
            except auth.UserNotFoundError:
                log.info(f"No Auth account for {username}, skipping Auth deletion")
            except Exception as e:
                log.warning(f"Could not delete Auth account for {username}: {e}")

            # Always delete from Database
            ref = db.reference(f'/users/{firebase_uid}')
            ref.delete()

            log.info(f"User {username} deleted from Firebase")
            return True

        except Exception as e:
            log.error(f"Error deleting user: {e}")
            return False
    @staticmethod
    def is_admin(email: str) -> bool:
        """Check if user has admin role"""
        user = FirebaseManager.get_user_by_email(email)
        return user and user.get('role') == 'admin' if user else False
    
    @staticmethod
    def get_user_role(email: str) -> str:
        """Get user role by email"""
        user = FirebaseManager.get_user_by_email(email)
        return normalize_role(user.get('role', 'sales')) if user else 'sales'
    
    @staticmethod
    def save_project(project_data: Dict) -> bool:
        """Save project data to Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - project not saved to Firebase")
            return False
            
        try:
            ref = db.reference('/projects')
            # Check if project already exists (by project_number)
            existing_projects = ref.order_by_child('project_number').equal_to(project_data['project_number']).get()
            
            if existing_projects:
                # Update existing project
                project_id = list(existing_projects.keys())[0]
                project_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(project_id).update(project_data)
                log.info("Project updated in Firebase: %s", project_data['project_number'])
                return True
            else:
                # Create new project
                new_project_ref = ref.push()
                project_data['firebase_id'] = new_project_ref.key
                project_data['created_at'] = datetime.now(timezone.utc).isoformat()
                project_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                new_project_ref.set(project_data)
                log.info("Project saved to Firebase with ID: %s", new_project_ref.key)
                return True
        except Exception as e:
            log.warning("Error saving project to Firebase: %s", e)
            return False
    @staticmethod
    def save_combined_job_forms_pdf_to_firebase(filename, pdf_path):
        """Save combined quote forms PDF to Firebase Storage"""
        try:
            from firebase_admin import storage
            
            bucket = storage.bucket()
            blob = bucket.blob(f"job_forms_exports/{filename}")
            
            with open(pdf_path, 'rb') as pdf_file:
                blob.upload_from_file(pdf_file, content_type='application/pdf')
            
            log.info("Combined quote forms PDF saved to Firebase: %s", filename)
            return True
            
        except Exception as e:
            log.warning("Error saving combined quote forms PDF to Firebase: %s", e)
            return False
    @staticmethod
    def save_invoice(invoice_data: Dict) -> bool:
        """Save invoice data to Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - invoice not saved to Firebase")
            return False
            
        try:
            ref = db.reference('/invoices')
            # Check if invoice already exists
            existing_invoices = ref.order_by_child('meta/invoice_number').equal_to(invoice_data['meta']['invoice_number']).get()
            
            if existing_invoices:
                # Update existing invoice - preserve created_at
                invoice_id = list(existing_invoices.keys())[0]
                # If created_at is not already in invoice_data, fetch from existing record
                if 'created_at' not in invoice_data.get('meta', {}):
                    existing_data = ref.child(invoice_id).get().val()
                    if existing_data and 'meta' in existing_data:
                        # Preserve the original created_at timestamp
                        if 'created_at' in existing_data['meta']:
                            invoice_data['meta']['created_at'] = existing_data['meta']['created_at']
                invoice_data['meta']['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(invoice_id).update(invoice_data)
                log.info("Invoice updated in Firebase: %s", invoice_data['meta']['invoice_number'])
                return True
            else:
                # Create new invoice
                new_invoice_ref = ref.push()
                invoice_data['firebase_id'] = new_invoice_ref.key
                invoice_data['meta']['created_at'] = datetime.now(timezone.utc).isoformat()
                invoice_data['meta']['updated_at'] = datetime.now(timezone.utc).isoformat()
                new_invoice_ref.set(invoice_data)
                log.info("Invoice saved to Firebase with ID: %s", new_invoice_ref.key)
                return True
        except Exception as e:
            log.warning("Error saving invoice to Firebase: %s", e)
            return False
    @staticmethod
    def update_invoice(invoice_number, invoice_data):
        """Update an existing invoice in Firebase"""
        try:
            if not FIREBASE_AVAILABLE:
                return False
                
            from firebase_admin import db
            ref = db.reference('/invoices')
            
            # Find invoice by invoice_number
            invoices = ref.order_by_child('meta/invoice_number').equal_to(invoice_number).get()
            
            if invoices:
                # Update the first matching invoice
                invoice_id = list(invoices.keys())[0]
                ref.child(invoice_id).update(invoice_data)
                log.info("Invoice updated in Firebase: %s", invoice_number)
                return True
            else:
                log.warning("Invoice not found in Firebase: %s", invoice_number)
                return False
                
        except Exception as e:
            log.warning("Error updating invoice in Firebase: %s", e)
            return False

    @staticmethod
    def validate_user(username: str, password: str) -> bool:
        """Validate user credentials against hashed passwords in settings.json."""
        _log = get_logger("auth")
        users = Config.USERS

        if not users:
            _log.warning("No users configured in settings.json  -  login rejected")
            return False

        user = users.get(username)
        stored_hash = user.get("password_hash", "") if user else ""
        if user and verify_password(password, stored_hash):
            if password_needs_rehash(stored_hash):
                FirebaseManager.update_user_password_hash(username, hash_password(password))
            _log.info("Login success: %s", username)
            return True

        _log.warning("Failed login attempt for username: %s", username)
        return False

    @staticmethod
    def update_user_password_hash(username: str, password_hash: str) -> None:
        """Persist an upgraded password hash when settings.json is writable."""
        try:
            Config.USERS.setdefault(username, {})["password_hash"] = password_hash
            if Config.SETTINGS_FILE.exists():
                with open(Config.SETTINGS_FILE, encoding="utf-8") as f:
                    settings = json.load(f)
                settings.setdefault("users", {}).setdefault(username, {})["password_hash"] = password_hash
                with open(Config.SETTINGS_FILE, "w", encoding="utf-8") as f:
                    json.dump(settings, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            get_logger("auth").warning("Could not upgrade password hash for %s: %s", username, exc)

    @staticmethod
    def get_user_role(username: str) -> str:
        """Return the configured role for a user, or an empty string."""
        user = Config.USERS.get(username, {})
        return normalize_role(user.get("role", "sales"))

    @staticmethod
    def delete_invoice(invoice_number):
        """Delete an invoice from Firebase"""
        try:
            if not FIREBASE_AVAILABLE:
                return False
                
            from firebase_admin import db
            ref = db.reference('/invoices')
            
            # Find invoice by invoice_number
            invoices = ref.order_by_child('meta/invoice_number').equal_to(invoice_number).get()
            
            if invoices:
                invoice_id = list(invoices.keys())[0]
                ref.child(invoice_id).delete()
                
                # Also delete PDF if it exists
                pdf_ref = db.reference('/pdfs')
                pdfs_data = pdf_ref.order_by_child('invoice_number').equal_to(invoice_number).get()
                if pdfs_data:
                    pdf_id = list(pdfs_data.keys())[0]
                    pdf_ref.child(pdf_id).delete()
                
                log.info("Invoice deleted from Firebase: %s", invoice_number)
                return True
            else:
                log.warning("Invoice not found in Firebase: %s", invoice_number)
                return False
                
        except Exception as e:
            log.warning("Error deleting invoice from Firebase: %s", e)
            return False
    @staticmethod
    def save_client(client_name: str, client_data: Dict) -> bool:
        """Save client data to Firebase, or local clients.json when offline."""
        if not FIREBASE_AVAILABLE:
            try:
                Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
                clients = {}
                if Config.CLIENTS_FILE.exists():
                    with open(Config.CLIENTS_FILE, encoding="utf-8") as f:
                        loaded = json.load(f)
                        if isinstance(loaded, dict):
                            clients = loaded
                client_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                clients[client_name] = client_data
                with open(Config.CLIENTS_FILE, "w", encoding="utf-8") as f:
                    json.dump(clients, f, indent=2, ensure_ascii=False)
                log.info("Client saved locally: %s", client_name)
                return True
            except Exception as e:
                log.warning("Error saving client locally: %s", e)
                return False
            
        try:
            ref = db.reference('/clients')
            client_data['updated_at'] = datetime.now(timezone.utc).isoformat()
            ref.child(client_name).set(client_data)
            log.info("Client saved to Firebase: %s", client_name)
            return True
        except Exception as e:
            log.warning("Error saving client to Firebase: %s", e)
            return False
    
    @staticmethod
    def load_projects() -> List[Dict]:
        """Load projects from Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot load projects")
            return []

        try:
            ref = db.reference('/projects')
            projects_data = ref.get()
            if projects_data:
                projects = []
                for project_id, project_data in projects_data.items():
                    if project_data:
                        project_data['firebase_id'] = project_id
                        projects.append(project_data)
                log.info("Loaded %s projects from Firebase", len(projects))
                return projects
            log.info("No projects found in Firebase")
            return []
        except Exception as e:
            log.warning("Error loading projects from Firebase: %s", e)
            return []

    _listeners = {}  # Store all listeners

    @staticmethod
    def add_realtime_listener(path, callback, name="listener"):
        """Generic method to add real-time listener for any Firebase path"""
        if not FIREBASE_AVAILABLE:
            return False

        try:
            ref = db.reference(path)

            def on_change(message):
                """Called whenever data changes in Firebase"""
                try:
                    data = ref.get()
                    callback(data)
                except Exception as e:
                    log.warning(f"Error in {name} listener callback: %s", e)

            ref.listen(on_change)
            if name not in FirebaseManager._listeners:
                FirebaseManager._listeners[name] = []
            FirebaseManager._listeners[name].append((ref, on_change))
            log.info(f"Real-time {name} listener added for {path}")
            return True
        except Exception as e:
            log.warning(f"Error adding {name} listener: %s", e)
            return False

    @staticmethod
    def add_projects_listener(callback):
        """Add a real-time listener for projects changes"""
        return FirebaseManager.add_realtime_listener('/projects', callback, 'projects')

    @staticmethod
    def add_invoices_listener(callback):
        """Add a real-time listener for invoices changes"""
        return FirebaseManager.add_realtime_listener('/invoices', callback, 'invoices')

    @staticmethod
    def add_quotes_listener(callback):
        """Add a real-time listener for quotes/job forms changes"""
        return FirebaseManager.add_realtime_listener('/job_forms', callback, 'quotes')

    @staticmethod
    def add_clients_listener(callback):
        """Add a real-time listener for clients changes"""
        return FirebaseManager.add_realtime_listener('/clients', callback, 'clients')

    @staticmethod
    def add_expenses_listener(callback):
        """Add a real-time listener for expenses changes"""
        return FirebaseManager.add_realtime_listener('/balance_sheet_expenses', callback, 'expenses')

    @staticmethod
    def add_balance_sheet_listener(callback):
        """Add a real-time listener for balance sheet changes"""
        return FirebaseManager.add_realtime_listener('/balance_sheet_data', callback, 'balance_sheet')

    @staticmethod
    def load_job_forms() -> List[Dict]:
        """Load quote/job forms from Firebase."""
        if not FIREBASE_AVAILABLE:
            try:
                path = Path(__file__).resolve().parent / "data" / "job_forms.json"
                if path.exists():
                    import json
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    if isinstance(data, list):
                        return data
                    if isinstance(data, dict):
                        return list(data.values())
            except Exception as e:
                log.warning("Error loading job forms from local: %s", e)
            return []
        try:
            ref = db.reference('/job_forms')
            raw = ref.get()
            if not raw:
                return []
            forms = []
            for fid, fdata in raw.items():
                if fdata:
                    fdata['firebase_id'] = fid
                    fdata.setdefault('status', 'Not Started')
                    forms.append(fdata)
            log.info("Loaded %s job forms from Firebase", len(forms))
            return forms
        except Exception as e:
            log.warning("Error loading job forms from Firebase: %s", e)
            return []

    @staticmethod
    def advance_project_payment_stage(project_number: str) -> bool:
        """Advance a project's payment_category to the next stage after payment received."""
        if not FIREBASE_AVAILABLE:
            return False
        _STAGES = ["Down Payment", "Due Payment", "Final Payment"]
        try:
            ref = db.reference('/projects')
            data = ref.order_by_child('project_number').equal_to(project_number).get()
            if not data:
                return False
            project_id = list(data.keys())[0]
            project = data[project_id]
            current = project.get('payment_category', 'Down Payment')
            if current in _STAGES:
                idx = _STAGES.index(current)
                if idx < len(_STAGES) - 1:
                    next_stage = _STAGES[idx + 1]
                    updates = {'payment_category': next_stage, 'updated_at': datetime.now(timezone.utc).isoformat()}
                    db.reference(f'/projects/{project_id}').update(updates)
                    log.info("Project %s advanced to %s", project_number, next_stage)
                    return True
            return False
        except Exception as e:
            log.warning("Error advancing project stage: %s", e)
            return False

    @staticmethod
    def update_project_status(project_id: str, new_status: str) -> bool:
        """Update only the status field of a project in Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - status not updated")
            return False
            
        try:
            ref = db.reference(f'/projects/{project_id}')
            ref.update({
                'status': new_status,
                'updated_at': datetime.now(timezone.utc).isoformat()
            })
            log.info("Project status updated in Firebase: %s -> %s", project_id, new_status)
            return True
        except Exception as e:
            log.warning("Error updating project status in Firebase: %s", e)
            return False
    
    @staticmethod
    def load_invoices() -> List[Dict]:
        """Load invoices from Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot load invoices")
            return []  # Return empty list, not None
        
        try:
            ref = db.reference('/invoices')
            invoices_data = ref.get()
            if invoices_data:
                invoices = []
                for invoice_id, invoice_data in invoices_data.items():
                    if invoice_data:
                        invoice_data['firebase_id'] = invoice_id
                        invoices.append(invoice_data)
                log.info("Loaded %s invoices from Firebase", len(invoices))
                FirebaseManager._invoices_cache = invoices
                return invoices
            log.info("No invoices found in Firebase")
            FirebaseManager._invoices_cache = []
            return []  # Return empty list, not None
        except Exception as e:
            log.warning("Error loading invoices from Firebase: %s", e)
            return []  # Return empty list, not None
        
    
    # Add to FirebaseManager class in main.py:

    @staticmethod
    def save_job_pdf_to_firebase(job_number: str, pdf_path: Path) -> bool:
        """Save quote form PDF to Firebase Realtime Database as Base64"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - job PDF not saved")
            return False
            
        try:
            # Read PDF file
            with open(pdf_path, "rb") as pdf_file:
                pdf_data = pdf_file.read()
            
            # Convert PDF to base64 for storage in Firebase Realtime Database
            pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')
            
            # Save to Firebase under /job_pdfs node
            ref = db.reference('/job_pdfs')
            pdf_record = {
                'job_number': job_number,
                'pdf_base64': pdf_base64,
                'file_name': f"{job_number}_job_form.pdf",
                'created_at': datetime.now(timezone.utc).isoformat(),
                'size_bytes': len(pdf_data)
            }
            
            # Check if PDF already exists
            existing_pdfs = ref.order_by_child('job_number').equal_to(job_number).get()
            
            if existing_pdfs:
                # Update existing PDF
                pdf_id = list(existing_pdfs.keys())[0]
                ref.child(pdf_id).update(pdf_record)
                log.info("Job PDF updated in Firebase: %s", job_number)
            else:
                # Create new PDF entry
                new_pdf_ref = ref.push()
                pdf_record['firebase_id'] = new_pdf_ref.key
                new_pdf_ref.set(pdf_record)
                log.info("Job PDF saved to Firebase with ID: %s", new_pdf_ref.key)
            
            return True
            
        except Exception as e:
            log.warning("Error saving job PDF to Firebase: %s", e)
            return False

    @staticmethod
    def load_job_pdf_from_firebase(job_number: str, output_path: Path = None) -> Optional[Path]:
        """Load quote form PDF from Firebase and decode from Base64"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot load job PDF")
            return None
            
        try:
            ref = db.reference('/job_pdfs')
            pdfs_data = ref.order_by_child('job_number').equal_to(job_number).get()
            
            if not pdfs_data:
                log.warning("Job PDF not found in Firebase: %s", job_number)
                return None
            
            # Get the first matching PDF
            pdf_id = list(pdfs_data.keys())[0]
            pdf_data = pdfs_data[pdf_id]
            
            # Decode Base64 back to PDF
            pdf_base64 = pdf_data.get('pdf_base64', '')
            if not pdf_base64:
                log.warning("No PDF data found for: %s", job_number)
                return None
            
            pdf_bytes = base64.b64decode(pdf_base64)
            
            # Save to temporary file or use provided output path
            if output_path:
                pdf_path = output_path
            else:
                temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
                temp_dir.mkdir(parents=True, exist_ok=True)
                pdf_path = temp_dir / f"{job_number}_job_form.pdf"
            
            with open(pdf_path, "wb") as pdf_file:
                pdf_file.write(pdf_bytes)
            
            log.info("Job PDF loaded from Firebase: %s", job_number)
            return pdf_path
            
        except Exception as e:
            log.warning("Error loading job PDF from Firebase: %s", e)
            return None

    @staticmethod
    def delete_job_pdf_from_firebase(job_number: str) -> bool:
        """Delete quote form PDF from Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot delete job PDF")
            return False
            
        try:
            ref = db.reference('/job_pdfs')
            pdfs_data = ref.order_by_child('job_number').equal_to(job_number).get()
            
            if not pdfs_data:
                log.warning("Job PDF not found in Firebase: %s", job_number)
                return False
            
            pdf_id = list(pdfs_data.keys())[0]
            ref.child(pdf_id).delete()
            log.info("Job PDF deleted from Firebase: %s", job_number)
            return True
            
        except Exception as e:
            log.warning("Error deleting job PDF from Firebase: %s", e)
            return False
        
    @staticmethod
    def save_job_form(job_data: Dict) -> bool:
        """Save or update a quote form in Firebase, keyed on job_number."""
        _log = get_logger("firebase")
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available  -  job form not saved")
            return False
        try:
            ref = db.reference('/job_forms')
            existing = ref.order_by_child('job_number').equal_to(job_data['job_number']).get()
            if existing:
                job_id = list(existing.keys())[0]
                job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(job_id).update(job_data)
                _log.info("Quote form updated: %s", job_data['job_number'])
            else:
                new_ref = ref.push()
                job_data['firebase_id'] = new_ref.key
                job_data['created_at'] = datetime.now(timezone.utc).isoformat()
                job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                new_ref.set(job_data)
                _log.info("Quote form saved: %s", job_data['job_number'])
            return True
        except Exception as e:
            _log.error("Error saving job form: %s", e)
            return False
        
    @staticmethod
    def load_clients() -> Dict:
        """Load clients from Firebase, or local clients.json when offline."""
        if not FIREBASE_AVAILABLE:
            try:
                if Config.CLIENTS_FILE.exists():
                    with open(Config.CLIENTS_FILE, encoding="utf-8") as f:
                        clients = json.load(f)
                    if isinstance(clients, dict):
                        log.info("Loaded %s clients locally", len(clients))
                        return clients
                return {}
            except Exception as e:
                log.warning("Error loading clients locally: %s", e)
                return {}
            
        try:
            ref = db.reference('/clients')
            clients_data = ref.get()
            if clients_data:
                log.info("Loaded %s clients from Firebase", len(clients_data))
                return clients_data
            log.info("No clients found in Firebase")
            return {}
        except Exception as e:
            log.warning("Error loading clients from Firebase: %s", e)
            return {}
    
    @staticmethod
    def update_invoice_status(invoice_number: str, new_status: str) -> bool:
        """Update invoice status in Firebase using correct path"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - status not updated")
            return False

        try:
            from firebase_admin import db
            invoices_ref = db.reference('/invoices')
            invoices_data = invoices_ref.get()

            if not invoices_data:
                log.warning("No invoices found in Firebase")
                return False

            # Find invoice based on invoice_number under meta
            for invoice_id, invoice_data in invoices_data.items():
                if not invoice_data or 'meta' not in invoice_data:
                    continue

                if invoice_data['meta'].get('invoice_number') == invoice_number:
                    log.debug("Updating invoice '%s' at Firebase ID '%s'", invoice_number, invoice_id)

                    status_ref = db.reference(f'/invoices/{invoice_id}/meta')
                    status_ref.update({
                        "status": new_status,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    })

                    log.info("Status updated successfully: %s -> %s", invoice_number, new_status)
                    return True

            log.warning("Invoice '%s' not found in Firebase", invoice_number)
            return False

        except Exception as e:
            log.warning("Error updating status in Firebase: %s", e)
            return False

    @staticmethod
    def auto_mark_overdue_invoices() -> int:
        """Flip Unpaid/Pending invoices to Overdue when their due date has passed.

        Returns the number of invoices updated.
        """
        from status_enums import InvoiceStatus
        if not FIREBASE_AVAILABLE:
            return 0

        today = datetime.now().date()
        updated = 0
        try:
            ref = db.reference('/invoices')
            all_invoices = ref.get()
            if not all_invoices:
                return 0

            for invoice_id, invoice_data in all_invoices.items():
                if not isinstance(invoice_data, dict):
                    continue
                meta = invoice_data.get('meta', {})
                status = meta.get('status', '')
                if status not in InvoiceStatus.OPEN or status == InvoiceStatus.OVERDUE:
                    continue
                if status in InvoiceStatus.CLOSED:
                    continue

                due_raw = meta.get('due_date', '')
                if not due_raw or due_raw == 'N/A':
                    continue

                try:
                    due_date = datetime.strptime(due_raw, "%m-%d-%Y").date()
                except ValueError:
                    try:
                        due_date = datetime.strptime(due_raw, "%Y-%m-%d").date()
                    except ValueError:
                        continue

                if due_date < today:
                    ref.child(invoice_id).child('meta').update({
                        'status': InvoiceStatus.OVERDUE,
                        'updated_at': datetime.now(timezone.utc).isoformat(),
                    })
                    log.info("Auto-marked overdue: %s (due %s)", meta.get('invoice_number', invoice_id), due_raw)
                    updated += 1

        except Exception as exc:
            log.warning("auto_mark_overdue_invoices failed: %s", exc)

        if updated:
            log.info("Auto-overdue sweep: %d invoice(s) updated", updated)
        return updated

    @staticmethod
    def auto_complete_projects() -> int:
        """Mark a project Completed when every invoice linked to it is Paid.

        Returns the number of projects updated.
        """
        from status_enums import InvoiceStatus, ProjectStatus
        if not FIREBASE_AVAILABLE:
            return 0

        updated = 0
        try:
            invoices = FirebaseManager.load_invoices()
            projects = FirebaseManager.load_projects()

            # Build map: project_number a ' set of invoice statuses
            project_statuses: Dict[str, List[str]] = {}
            for inv in invoices:
                meta = inv.get('meta', inv)
                status = meta.get('status', InvoiceStatus.UNPAID)
                items = inv.get('items', [])
                for item in items:
                    pnum = item.get('project_number', '').strip() if isinstance(item, dict) else ''
                    if pnum:
                        project_statuses.setdefault(pnum, []).append(status)

            for project in projects:
                pnum = project.get('project_number', '')
                if not pnum:
                    continue
                # Never auto-override a status the user set manually
                if project.get('status_manual'):
                    continue
                if project.get('status') in ProjectStatus.INACTIVE:
                    continue

                inv_statuses = project_statuses.get(pnum, [])
                if not inv_statuses:
                    continue
                if all(s == InvoiceStatus.PAID for s in inv_statuses):
                    firebase_id = project.get('firebase_id', '')
                    if firebase_id:
                        FirebaseManager.update_project_status(firebase_id, ProjectStatus.COMPLETED)
                        log.info("Auto-completed project %s (all invoices paid)", pnum)
                        updated += 1

        except Exception as exc:
            log.warning("auto_complete_projects failed: %s", exc)

        if updated:
            log.info("Auto-complete sweep: %d project(s) completed", updated)
        return updated

    @staticmethod
    def auto_expire_quotes() -> int:
        """Flip quotes to Expired when their due_date has passed and they haven't been actioned.

        Statuses that are left unchanged: Completed, Converted, Expired, Cancel, Cancelled.
        Returns the number of quotes updated.
        """
        if not FIREBASE_AVAILABLE:
            return 0

        INACTIVE = {"Completed", "Converted", "Expired", "Cancel", "Cancelled"}
        today = datetime.now().date()
        updated = 0
        try:
            ref = db.reference('/job_forms')
            all_forms = ref.get()
            if not all_forms:
                return 0

            for form_id, form_data in all_forms.items():
                if not isinstance(form_data, dict):
                    continue
                if form_data.get('status', '') in INACTIVE:
                    continue

                due_raw = form_data.get('due_date', '')
                if not due_raw or due_raw == 'N/A':
                    continue

                try:
                    due_date = datetime.strptime(due_raw, "%m-%d-%Y").date()
                except ValueError:
                    continue

                if due_date < today:
                    ref.child(form_id).update({
                        'status': 'Expired',
                        'updated_at': datetime.now(timezone.utc).isoformat(),
                    })
                    log.info("Auto-expired quote %s (due %s)", form_data.get('job_number', form_id), due_raw)
                    updated += 1

        except Exception as exc:
            log.warning("auto_expire_quotes failed: %s", exc)

        if updated:
            log.info("Quote expiry sweep: %d quote(s) expired", updated)
        return updated

    # ADD THESE PDF METHODS TO THE FIREBASEMANAGER CLASS:

    @staticmethod
    def save_pdf_to_firebase(invoice_number: str, pdf_path: Path) -> bool:
        """Save PDF to Firebase Realtime Database as Base64"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - PDF not saved")
            return False
            
        try:
            # Read PDF file
            with open(pdf_path, "rb") as pdf_file:
                pdf_data = pdf_file.read()
            
            # Convert PDF to base64 for storage in Firebase Realtime Database
            pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')
            
            # Save to Firebase under /pdfs node
            ref = db.reference('/pdfs')
            pdf_record = {
                'invoice_number': invoice_number,
                'pdf_base64': pdf_base64,
                'file_name': f"{invoice_number}.pdf",
                'created_at': datetime.now(timezone.utc).isoformat(),
                'size_bytes': len(pdf_data)
            }
            
            # Check if PDF already exists
            existing_pdfs = ref.order_by_child('invoice_number').equal_to(invoice_number).get()
            
            if existing_pdfs:
                # Update existing PDF
                pdf_id = list(existing_pdfs.keys())[0]
                ref.child(pdf_id).update(pdf_record)
                log.info("PDF updated in Firebase: %s", invoice_number)
            else:
                # Create new PDF entry
                new_pdf_ref = ref.push()
                pdf_record['firebase_id'] = new_pdf_ref.key
                new_pdf_ref.set(pdf_record)
                log.info("PDF saved to Firebase with ID: %s", new_pdf_ref.key)
            
            return True
            
        except Exception as e:
            log.warning("Error saving PDF to Firebase: %s", e)
            return False
    
    @staticmethod
    def load_pdf_from_firebase(invoice_number: str, output_path: Path = None) -> Optional[Path]:
        """Load PDF from Firebase and decode from Base64"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot load PDF")
            return None
            
        try:
            ref = db.reference('/pdfs')
            pdfs_data = ref.order_by_child('invoice_number').equal_to(invoice_number).get()
            
            if not pdfs_data:
                log.warning("PDF not found in Firebase: %s", invoice_number)
                return None
            
            # Get the first matching PDF (should be only one per invoice number)
            pdf_id = list(pdfs_data.keys())[0]
            pdf_data = pdfs_data[pdf_id]
            
            # Decode Base64 back to PDF
            pdf_base64 = pdf_data.get('pdf_base64', '')
            if not pdf_base64:
                log.warning("No PDF data found for: %s", invoice_number)
                return None
            
            pdf_bytes = base64.b64decode(pdf_base64)
            
            # Save to temporary file or use provided output path
            if output_path:
                pdf_path = output_path
            else:
                temp_dir = Path(tempfile.gettempdir()) / "mabs_invoices_temp"
                temp_dir.mkdir(parents=True, exist_ok=True)
                pdf_path = temp_dir / f"{invoice_number}.pdf"
            
            with open(pdf_path, "wb") as pdf_file:
                pdf_file.write(pdf_bytes)
            
            log.info("PDF loaded from Firebase: %s", invoice_number)
            return pdf_path
            
        except Exception as e:
            log.warning("Error loading PDF from Firebase: %s", e)
            return None
    
    @staticmethod
    def delete_pdf_from_firebase(invoice_number: str) -> bool:
        """Delete PDF from Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot delete PDF")
            return False
            
        try:
            ref = db.reference('/pdfs')
            pdfs_data = ref.order_by_child('invoice_number').equal_to(invoice_number).get()
            
            if not pdfs_data:
                log.warning("PDF not found in Firebase: %s", invoice_number)
                return False
            
            pdf_id = list(pdfs_data.keys())[0]
            ref.child(pdf_id).delete()
            log.info("PDF deleted from Firebase: %s", invoice_number)
            return True
            
        except Exception as e:
            log.warning("Error deleting PDF from Firebase: %s", e)
            return False
    
    @staticmethod
    def get_pdf_list() -> List[Dict]:
        """Get list of all PDFs in Firebase"""
        if not FIREBASE_AVAILABLE:
            log.warning("Firebase not available - cannot get PDF list")
            return []
            
        try:
            ref = db.reference('/pdfs')
            pdfs_data = ref.get()
            
            if pdfs_data:
                pdfs = []
                for pdf_id, pdf_data in pdfs_data.items():
                    if pdf_data:
                        pdf_data['firebase_id'] = pdf_id
                        pdfs.append(pdf_data)
                log.info("Loaded %s PDFs from Firebase", len(pdfs))
                return pdfs
            log.info("No PDFs found in Firebase")
            return []
            
        except Exception as e:
            log.warning("Error loading PDF list from Firebase: %s", e)
            return []

    @staticmethod
    def save_quotation(quotation_data: Dict) -> bool:
        """Save quotation data to Firebase"""
        if not FIREBASE_AVAILABLE:
            return False
            
        try:
            ref = db.reference('/quotations')
            new_quotation_ref = ref.push()
            quotation_data['firebase_id'] = new_quotation_ref.key
            new_quotation_ref.set(quotation_data)
            return True
        except Exception as e:
            log.warning("Error saving quotation to Firebase: %s", e)
            return False
    
    @staticmethod
    def save_expense(expense_data: Dict) -> bool:
        """Save expense data to Firebase"""
        if not FIREBASE_AVAILABLE:
            return False
            
        try:
            ref = db.reference('/expenses')
            new_expense_ref = ref.push()
            expense_data['firebase_id'] = new_expense_ref.key
            new_expense_ref.set(expense_data)
            return True
        except Exception as e:
            log.warning("Error saving expense to Firebase: %s", e)
            return False

    @staticmethod
    def save_settings_to_firebase(data: dict) -> bool:
        """Save company info and preferences to /settings in Firebase."""
        if not FIREBASE_AVAILABLE:
            return False
        try:
            import base64 as _b64
            company = dict(data.get("company") or {})
            payload: dict = {
                "company": {k: v for k, v in company.items() if k != "logo_path"},
                "app":     dict(data.get("app") or {}),
                "github":  dict(data.get("github") or {}),
            }
            # Store logo as base64 so it syncs across devices
            logo_str = (company.get("logo_path") or "").strip()
            if logo_str:
                lp = Path(logo_str)
                if lp.exists():
                    try:
                        payload["logo_base64"] = _b64.b64encode(lp.read_bytes()).decode("utf-8")
                        payload["logo_ext"] = lp.suffix.lower() or ".png"
                    except Exception:
                        pass
            db.reference("/settings").set(payload)
            log.info("Settings saved to Firebase")
            return True
        except Exception as exc:
            log.warning("Failed to save settings to Firebase: %s", exc)
            return False

    @staticmethod
    def load_settings_from_firebase() -> dict:
        """Load company info and preferences from /settings in Firebase. Returns {} on failure."""
        if not FIREBASE_AVAILABLE:
            return {}
        try:
            import base64 as _b64
            raw = db.reference("/settings").get()
            if not raw or not isinstance(raw, dict):
                return {}
            result: dict = {}
            if "company" in raw:
                result["company"] = dict(raw["company"])
            if "app" in raw:
                result["app"] = dict(raw["app"])
            if "github" in raw:
                result["github"] = dict(raw["github"])
            # Restore logo from base64 to a local cache file
            logo_b64 = (raw.get("logo_base64") or "").strip()
            logo_ext = raw.get("logo_ext") or ".png"
            if logo_b64:
                try:
                    logo_cache = Config.DATA_DIR / f"company_logo{logo_ext}"
                    logo_cache.write_bytes(_b64.b64decode(logo_b64))
                    result.setdefault("company", {})["logo_path"] = str(logo_cache)
                except Exception as exc:
                    log.warning("Failed to restore logo from Firebase: %s", exc)
            return result
        except Exception as exc:
            log.warning("Failed to load settings from Firebase: %s", exc)
            return {}

# ---------- Utility Classes ----------
class DecimalEncoder(json.JSONEncoder):
    """Custom JSON encoder for Decimal objects"""
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

class FileManager:
    """Handles file operations with Firebase integration"""
    
    @staticmethod
    def load_json(path: Path, default: Any = None) -> Any:
        """Load JSON file with error handling - tries Firebase first, then local."""
        try:
            if FIREBASE_AVAILABLE:
                if path.name == "clients.json":
                    clients = FirebaseManager.load_clients()
                    if clients is not None:
                        return clients
                
                elif path.name == "projects.json":
                    projects = FirebaseManager.load_projects()
                    if projects is not None:
                        return projects
                
                elif path.name == "invoice_counter.json":
                    return FileManager._load_local_json(path, default)
            
        except Exception as e:
            log.warning("Firebase load failed for %s: %s", path.name, e)
        
        return FileManager._load_local_json(path, default)

    @staticmethod
    def save_json(path: Path, data: Any) -> bool:
        """Save JSON file with error handling - saves to Firebase and local fallback."""
        firebase_saved = False
        try:
            if FIREBASE_AVAILABLE:
                if path.name == "clients.json" and isinstance(data, dict):
                    for client_name, client_data in data.items():
                        FirebaseManager.save_client(client_name, client_data)
                    log.info("Clients saved to Firebase")
                    firebase_saved = True
                
                elif path.name == "projects.json" and isinstance(data, list):
                    for project in data:
                        FirebaseManager.save_project(project)
                    log.info("Projects saved to Firebase")
                    firebase_saved = True
                
                elif path.name == "invoice_counter.json":
                    firebase_saved = True

        except Exception as e:
            log.warning("Firebase save failed for %s: %s", path.name, e)

        local_saved = FileManager._save_local_json(path, data)
        return firebase_saved or local_saved

    @staticmethod
    def _load_local_json(path: Path, default: Any = None) -> Any:
        try:
            if path.exists():
                with open(path, encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            log.warning("Local load failed for %s: %s", path, e)
        return default

    @staticmethod
    def _save_local_json(path: Path, data: Any) -> bool:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False, cls=DecimalEncoder)
            log.info("Data saved locally: %s", path)
            return True
        except Exception as e:
            log.warning("Local save failed for %s: %s", path, e)
            return False

    @staticmethod
    def open_file(path: Path) -> bool:
        """Open file with default application"""
        try:
            if platform.system() == "Windows":
                os.startfile(str(path))
            elif platform.system() == "Darwin":
                subprocess.run(["open", str(path)])
            else:
                subprocess.run(["xdg-open", str(path)])
            return True
        except Exception as e:
            log.warning("Error opening file %s: %s", path, e)
            return False

class Currency:
    """Currency formatting utilities"""
    
    @staticmethod
    def format(value, symbol: str = "$") -> str:
        """Format decimal as currency"""
        return f"{symbol}{Currency.quantize(value)}"
    
    @staticmethod
    def format_whole(value, symbol: str = "$") -> str:
        """Format decimal as currency without decimal places and without thousands separators"""
        try:
            # Convert to integer (truncate decimal part)
            whole_value = int(float(value))
            # Return without commas: 12470310 not 12,470,310
            return f"{symbol}{whole_value}"
        except (ValueError, TypeError):
            # If conversion fails, fall back to regular format
            return Currency.format(value, symbol)
    
    @staticmethod
    def quantize(value) -> Decimal:
        """Quantize decimal to 2 decimal places"""
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
class InvoiceNumberGenerator:
    """Handles invoice numbering system"""

    @staticmethod
    def _date_code(date_str: str) -> str:
        """Return YYYYMM for the date formats used by the invoice form."""
        text = str(date_str or "").strip()
        for fmt in ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text[:10], fmt).strftime("%Y%m")
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text[:19]).strftime("%Y%m")
        except ValueError:
            log.warning("Could not parse invoice date '%s'; using current month for invoice number", date_str)
            return datetime.now().strftime("%Y%m")
    
    @staticmethod
    def get_next_number(date_str: str) -> str:
        """Generate next invoice number for given date"""
        yyyymm = InvoiceNumberGenerator._date_code(date_str)

        if FIREBASE_AVAILABLE:
            try:
                next_num = InvoiceNumberGenerator._reserve_firebase_number(yyyymm)
                return InvoiceNumberGenerator._format(yyyymm, next_num)
            except Exception as e:
                log.warning("Firebase invoice counter failed: %s", e)

        counters = FileManager.load_json(Config.COUNTER_FILE, {})
        next_num = int(counters.get(yyyymm, 0)) + 1
        counters[yyyymm] = next_num
        FileManager.save_json(Config.COUNTER_FILE, counters)
        return InvoiceNumberGenerator._format(yyyymm, next_num)
    
    @staticmethod
    def get_preview_number(date_str: str) -> str:
        """Preview next invoice number without persisting"""
        yyyymm = InvoiceNumberGenerator._date_code(date_str)
        preview_next = 1
        
        if FIREBASE_AVAILABLE:
            try:
                invoices = FirebaseManager.load_invoices()
                highest_existing = InvoiceNumberGenerator._highest_existing_number(invoices, yyyymm)
                if highest_existing:
                    preview_next = highest_existing + 1
            except Exception as e:
                log.warning("Error getting preview invoice number: %s", e)
        
        return InvoiceNumberGenerator._format(yyyymm, preview_next)

    @staticmethod
    def _reserve_firebase_number(yyyymm: str) -> int:
        invoices = FirebaseManager.load_invoices()
        highest_existing = InvoiceNumberGenerator._highest_existing_number(invoices, yyyymm)
        counter_ref = db.reference(f"/counters/invoices/{yyyymm}")

        def reserve(current):
            current_int = int(current or 0)
            return max(current_int, highest_existing) + 1

        return int(counter_ref.transaction(reserve))

    @staticmethod
    def _highest_existing_number(invoices: List[Dict], yyyymm: str) -> int:
        numbers = []
        for inv in invoices or []:
            try:
                invoice_number = inv["meta"]["invoice_number"]
                if invoice_number.startswith(f"INV-{yyyymm}"):
                    numbers.append(int(invoice_number.split("-")[-1]))
            except (KeyError, TypeError, ValueError, IndexError):
                continue
        return max(numbers, default=0)

    @staticmethod
    def _format(yyyymm: str, number: int) -> str:
        return f"INV-{yyyymm}-{number:03d}"


def truncate_text(text: str, max_chars: int = 45) -> str:
    """
    Truncate text to max_chars and add ellipsis if needed
    """
    if not text:
        return ""
    text = text.strip()
    return text if len(text) <= max_chars else text[:max_chars - 3] + "..."

# ---------- PDF Generator ----------
class PDFGenerator:
    """Generates simple, clean PDF invoices using ReportLab's SimpleDocTemplate and Flowables"""
    
    @staticmethod
    def generate(invoice: "Invoice", output_path: Path, logo_path: Optional[Path] = None):
        """Generate PDF invoice with logo - LOGO LEFT, COMPANY INFO RIGHT (VERTICALLY CENTERED)"""
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4, 
                                  topMargin=5*mm, bottomMargin=5*mm,
                                  leftMargin=10*mm, rightMargin=10*mm)
            story = []
            styles = getSampleStyleSheet()

            # Custom styles
            styles.add(ParagraphStyle(name='CenteredBold20', alignment=1, fontName='Helvetica-Bold', fontSize=20, leading=24))
            styles.add(ParagraphStyle(name='Centered10', alignment=1, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Centered8',  alignment=1, fontName='Helvetica', fontSize=8,  leading=10))
            styles.add(ParagraphStyle(name='LeftBold16', alignment=0, fontName='Helvetica-Bold', fontSize=16, leading=18))
            styles.add(ParagraphStyle(name='LeftBold12', alignment=0, fontName='Helvetica-Bold', fontSize=12, leading=14))
            styles.add(ParagraphStyle(name='Left10', alignment=0, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='LeftBold10', alignment=0, fontName='Helvetica-Bold', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Right10', alignment=2, fontName='Helvetica', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='RightBold10', alignment=2, fontName='Helvetica-Bold', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Left9', alignment=0, fontName='Helvetica', fontSize=9, leading=11))
            styles.add(ParagraphStyle(name='Left8', alignment=0, fontName='Helvetica', fontSize=8, leading=10))
            styles.add(ParagraphStyle(name='CenterBold16', alignment=1, fontName='Helvetica-Bold', fontSize=16, leading=18))
            styles.add(ParagraphStyle(name='CenterBold12', alignment=1, fontName='Helvetica-Bold', fontSize=12, leading=14))
            styles.add(ParagraphStyle(name='CenterBold10', alignment=1, fontName='Helvetica-Bold', fontSize=10, leading=12))
            styles.add(ParagraphStyle(name='Center9', parent=styles['Normal'], alignment=1, fontSize=9, leading=11))
            styles.add(ParagraphStyle(name='ProjectNumberCell', alignment=1, fontName='Helvetica', fontSize=9, leading=11))
            # Add this line with your other style definitions
            styles.add(ParagraphStyle(name='Left14', alignment=0, fontName='Helvetica', fontSize=14, leading=16))
            if 'CenterBoldHeader10' not in styles:
                styles.add(ParagraphStyle(
                    name='CenterBoldHeader10',
                    alignment=1,  # CENTER
                    fontName='Helvetica-Bold',
                    fontSize=10,
                    leading=12
                ))

            # --- Company Header with Logo - LOGO LEFT, COMPANY INFO RIGHT (VERTICALLY CENTERED) ---
            if logo_path and logo_path.exists():
                try:
                    # Verify image is readable
                    try:
                        PILImage.open(str(logo_path))
                    except Exception:
                        logo_path = None


                    _logo_w = 30 * mm
                    logo_img = Image(str(logo_path), width=27*mm, height=27*mm)

                    # Company info — centered across the full page width
                    _addr_lines = Config.COMPANY.get('address', '').split('\n')
                    _addr1 = _addr_lines[0].strip() if _addr_lines else ''
                    _addr2 = _addr_lines[1].strip() if len(_addr_lines) > 1 else ''
                    _phone = Config.COMPANY.get('phone', '')
                    _email = Config.COMPANY.get('email', '')
                    _website = Config.COMPANY.get('website', '')
                    _phone_str = f"Phone:{_phone}" if _phone else ''
                    _contact = '  •  '.join(p for p in [_phone_str, _email, _website] if p)

                    _mid_w = doc.width - 2 * _logo_w
                    line1 = Paragraph(f"<b>{Config.COMPANY.get('name', '')}</b>", styles['CenteredBold20'])
                    line2 = Paragraph(_addr1, styles['Centered10'])
                    line3 = Paragraph(_addr2, styles['Centered10'])
                    line4 = Paragraph(_contact, styles['Centered8'])

                    company_info_table = Table(
                        [[line1], [line2], [line3], [line4]],
                        colWidths=[_mid_w],
                        style=TableStyle([
                            ('ALIGN',   (0,0), (-1,-1), 'CENTER'),
                            ('VALIGN',  (0,0), (-1,-1), 'MIDDLE'),
                            ('TOPPADDING',    (0,0), (-1,-1), 0),
                            ('BOTTOMPADDING', (0,0), (0,0), 4),
                            ('BOTTOMPADDING', (0,1), (0,1), 2),
                            ('BOTTOMPADDING', (0,2), (0,2), 4),
                            ('BOTTOMPADDING', (0,3), (0,3), 0),
                            ('LEFTPADDING',  (0,0), (-1,-1), 0),
                            ('RIGHTPADDING', (0,0), (-1,-1), 0),
                        ])
                    )

                    # 3-column layout: logo | centered text | equal spacer
                    # The equal spacer on the right makes the middle column
                    # truly centred over the full page width.
                    header_table = Table(
                        [[logo_img, company_info_table, '']],
                        colWidths=[_logo_w, _mid_w, _logo_w]
                    )
                    header_table.setStyle(TableStyle([
                        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                        ('LEFTPADDING',   (0,0), (-1,-1), 0),
                        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
                        ('TOPPADDING',    (0,0), (-1,-1), 0),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                    ]))

                    
                    story.append(header_table)
                    
                except Exception as e:
                    log.warning("Logo loading error: %s", e)
                    # Fallback: show company info without logo (centered)
                    story.append(Paragraph(Config.COMPANY.get("name", ""), styles['CenteredBold20']))
                    for line in Config.COMPANY.get("address", "").split('\n'):
                        story.append(Paragraph(line, styles['Centered10']))
                    _fb_ph = f"Phone:{Config.COMPANY.get('phone','')}" if Config.COMPANY.get('phone') else ''
                    _fb_ct = '  •  '.join(p for p in [_fb_ph, Config.COMPANY.get('email',''), Config.COMPANY.get('website','')] if p)
                    story.append(Paragraph(_fb_ct, styles['Centered8']))
            else:
                # No logo - just show company info centered
                story.append(Paragraph(Config.COMPANY.get("name", ""), styles['CenteredBold20']))
                for line in Config.COMPANY.get("address", "").split('\n'):
                    story.append(Paragraph(line, styles['Centered10']))
                _nl_ph = f"Phone:{Config.COMPANY.get('phone','')}" if Config.COMPANY.get('phone') else ''
                _nl_ct = '  •  '.join(p for p in [_nl_ph, Config.COMPANY.get('email',''), Config.COMPANY.get('website','')] if p)
                story.append(Paragraph(_nl_ct, styles['Centered8']))
                
            # Separator line - thinner and closer to header
            story.append(Table(
                [['']],
                colWidths=[doc.width],
                rowHeights=[1],  #   "' critical
                style=TableStyle([
                    ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.black),
                    ('TOPPADDING', (0,0), (-1,-1), 0),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                ])
            ))

            story.append(Spacer(1, 1*mm))  # very small gap

            
            # --- Invoice Details and Bill To Section (Side by side) ---
            # Build Bill-To block dynamically
            bill_to_lines = [invoice.client_name]

            if invoice.client_email.strip():
                bill_to_lines.append(invoice.client_email)

            if invoice.client_address.strip():
                bill_to_lines.append(invoice.client_address)

            bill_to_paragraph = Paragraph("<br/>".join(bill_to_lines), styles['Left9'])

            # Invoice details section
            invoice_details_data = [
                [Paragraph("Invoice : ", styles['LeftBold12']),
                Paragraph("<b>Bill To:</b>", styles['LeftBold12'])],

                [Paragraph(
                    f"<b>Invoice Number:</b> {invoice.invoice_number}<br/>"
                    f"<b>Date:</b> {invoice.date}<br/>"
                    f"<b>Due Date:</b> {invoice.due_date}",
                    styles['Left9']
                ),
                bill_to_paragraph]
            ]

            invoice_details_table = Table(invoice_details_data, colWidths=[doc.width * 0.5, doc.width * 0.5])
            invoice_details_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('ALIGN', (1,0), (1,-1), 'LEFT'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ]))
            
            story.append(invoice_details_table)
            story.append(Spacer(1, 5*mm))

            # --- Line Items ---
            story.append(Paragraph("ITEMS", styles['LeftBold12']))
            story.append(Spacer(1, 2*mm))

            # Updated headers with light gray color
            item_data = [[
                Paragraph("<b>Project</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Project Name</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Plant</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Qty</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Unit Price</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Payment Stage</b>", styles['CenterBoldHeader10']),
                Paragraph("<b>Payment Due</b>", styles['CenterBoldHeader10'])
            ]]

            for item in invoice.items:
                unit_price_whole = Currency.format_whole(item.unit_price)
                payment_due_whole = Currency.format_whole(item.payment_due)
                stage_label = (item.payment_category or "").strip() or "—"

                item_data.append([
                    Paragraph(item.project_number or "", styles['ProjectNumberCell']),
                    Paragraph(truncate_text(item.description, max_chars=45), styles['Center9']),
                    Paragraph(item.plant or "", styles['Center9']),
                    Paragraph(str(item.quantity), styles['Center9']),
                    Paragraph(unit_price_whole, styles['Center9']),
                    Paragraph(stage_label, styles['Center9']),
                    Paragraph(payment_due_whole, styles['Center9'])
                ])
            
            # Column widths — Project Number wide enough to fit on one line
            available_width = doc.width
            item_table = Table(item_data, colWidths=[
                available_width * 0.21,  # Project Number
                available_width * 0.21,  # Description
                available_width * 0.11,  # Plant
                available_width * 0.09,  # Quantity
                available_width * 0.11,  # Unit Price
                available_width * 0.12,  # Payment Stage
                available_width * 0.15   # Payment Due
            ])

            # Updated table style with light gray header background and CENTERED content
            item_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9D9D9')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),
                ('BACKGROUND', (0,1), (-1,-1), colors.white),
                ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
                ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,1), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('LEFTPADDING', (0,0), (-1,-1), 3),
                ('RIGHTPADDING', (0,0), (-1,-1), 3),
                ('LEFTPADDING', (0,1), (0,-1), 2),
                ('RIGHTPADDING', (0,1), (0,-1), 2),
                ('TOPPADDING', (0,1), (-1,-1), 4),
                ('BOTTOMPADDING', (0,1), (-1,-1), 4),
            ]))

            story.append(item_table)
            story.append(Spacer(1, 5*mm))

            # Create totals table — Deposit Received removed; unit_price is already the stage amount
            totals_data = [
                [Paragraph("<b>Total:</b>", styles['RightBold10']),
                 Paragraph(f"{Currency.format(invoice.subtotal)}", styles['RightBold10'])],
                [Paragraph(f"<b>Tax ({invoice.tax_rate}%):</b>", styles['RightBold10']),
                 Paragraph(f"{Currency.format(invoice.tax_amount)}", styles['RightBold10'])],
                [Paragraph("<b>TOTAL AMOUNT DUE:</b>", styles['RightBold10']),
                 Paragraph(f"{Currency.format(invoice.total)}", styles['RightBold10'])],
            ]

            # Calculate column widths
            label_width = available_width * 0.7
            amount_width = available_width * 0.3

            totals_table = Table(totals_data, colWidths=[label_width, amount_width])
            totals_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.white),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.black),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('ALIGN', (1,0), (1,-1), 'RIGHT'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('LINEBELOW', (0,0), (-1,0), 0, colors.white),
                ('LINEAFTER', (0,2), (0,2), 0, colors.white),
                ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#D9D9D9')),
            ]))

            story.append(totals_table)
            story.append(Spacer(1, 8*mm))

            # --- PAYMENT OPTIONS SECTION ---
            story.append(Paragraph("PAYMENT OPTIONS", ParagraphStyle(name='Left12', parent=styles['CenterBold12'], alignment=0)))
            story.append(Spacer(1, 3*mm))
            zelle_qr_path = Config.ZELLE_QR_FILE if Config.ZELLE_QR_FILE.exists() else resource_path("assets/venmo.png")

            # --- LEFT SECTION (Option 1 & 3) ---
            left_section = [
                # Option 1 Header
                Table(
                    [[Paragraph("<b>Option 1: Check</b>", styles['CenterBold10'])]],
                    colWidths=[available_width * 0.60],
                    rowHeights=[14],
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#B6DDE8")),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 1),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ])
                ),

                # Option 1 content
                Table(
                    [[Paragraph(
                        f"<para align='left' leading='14' color='black'>"
                        f"<font size='10'><b>Payable to:</b> {Config.COMPANY['name']}<br/>"
                        f"<b>Mailing Address:</b> {Config.COMPANY['address'].replace(chr(10), '<br/>')}</font>"
                        f"</para>",
                        styles['Left9']
                    )]],
                    colWidths=[available_width * 0.60],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 8),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ])
                ),

                # Option 3 Header
                Table(
                    [[Paragraph("<b>Option 3: Bank ACH Transfer</b>", styles['CenterBold10'])]],
                    colWidths=[available_width * 0.60],
                    rowHeights=[14],
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#EA9999")),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 1),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ])
                ),

                # Option 3 content
                Table(
                    [[Paragraph(
                        f"<para align='center' leading='11' color='black'>"
                        f"<font size='10'>Please contact {Config.COMPANY.get('name', 'MABS Engineering LLC').split()[0]} Admin to get our bank information for ACH transfers.</font>"
                        "</para>",
                        styles['Left9']
                    )]],
                    colWidths=[available_width * 0.60],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ])
                ),
            ]

            # --- RIGHT SECTION (Zelle) ---
            right_section = [
                # Option 2 Header
                Table(
                    [[Paragraph("<b>Option 2: Zelle QR code</b>", styles['CenterBold10'])]],
                    colWidths=[available_width * 0.40],
                    rowHeights=[14],
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#B6D7A8")),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 1),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ])
                ),

                Spacer(1, 2*mm),

                # Scanner image
                Table(
                    [[Image(str(zelle_qr_path), width=98, height=98)]],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ])
                ),
                Paragraph("<para align='center'><font color='black'>Scan to pay with Zelle</font></para>", styles['Left9'])
            ]

            # --- MAIN OUTER TABLE ---
            payment_data = [[left_section, right_section]]
            payment_table = Table(payment_data, colWidths=[available_width * 0.60, available_width * 0.40])
            payment_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))

            story.append(payment_table)
            story.append(Spacer(1, 5*mm))

            # --- Footer (Notes/Terms) ---
            _terms = Config.DEFAULT_TERMS
            if invoice.notes and invoice.notes.strip() and invoice.notes.strip() != _terms.strip():
                _terms = invoice.notes.strip()
            story.append(Spacer(1, 2*mm))
            story.append(Paragraph(f"<b>Note:</b> {_terms}", styles['Left9']))

            doc.build(story)
            return True
        except Exception as e:
            log.warning("PDF generation error: %s", e)
            traceback.print_exc()
            return False
# ---------- Data Models ----------
class InvoiceItem:
    """Represents an invoice line item"""
    
    def __init__(self, project_number: str = "", description: str = "", plant: str = "", 
                 quantity: int = 1, unit_price: float = 0.0, down_payment: float = 0.0, 
                 payment_due: float = 0.0, project_name: str = "", payment_category: str = ""):  # ADD payment_category parameter
        self.project_number = project_number
        self.description = description
        self.plant = plant
        self.quantity = quantity
        self.unit_price = Currency.quantize(unit_price)
        self.project_name = project_name
        self.payment_category = payment_category  # Store the payment category string
        
        # Convert all values to Decimal for proper calculation
        unit_price_decimal = Decimal(str(unit_price))
        down_payment_decimal = Decimal(str(down_payment))
        quantity_decimal = Decimal(str(quantity))
        
        # Store down payment as provided
        self.down_payment = Currency.quantize(down_payment_decimal)

        # Calculate payment due based on down payment
        total_amount = quantity_decimal * unit_price_decimal
        self.payment_due = Currency.quantize(total_amount - self.down_payment)
        
    @property
    def total(self) -> Decimal:
        return Currency.quantize(self.quantity * self.unit_price)
    
    def to_dict(self) -> Dict:
        return {
            "project_number": self.project_number,
            "description": self.description,
            "plant": self.plant,
            "quantity": self.quantity,
            "unit_price": float(self.unit_price),
            "down_payment": float(self.down_payment),
            "payment_due": float(self.payment_due),
            "total": float(self.total),
            "payment_category": self.payment_category  # ADD this to dictionary
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'InvoiceItem':
        return cls(
            project_number=data.get("project_number", ""),
            description=data.get("description", ""),
            plant=data.get("plant", ""),
            quantity=data.get("quantity", 1),
            unit_price=data.get("unit_price", 0.0),
            down_payment=data.get("down_payment", 0.0),
            payment_due=data.get("payment_due", 0.0),
            payment_category=data.get("payment_category", "")  # Make sure this is being loaded
        )

class Invoice:
    """Represents a complete invoice with downpayment support"""
    
    def __init__(self):
        self.invoice_number = ""
        self.date = datetime.now().strftime("%Y-%m-%d")
        self.due_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        self.client_name = ""
        self.client_email = ""
        self.client_address = ""
        self.items: List[InvoiceItem] = []
        self.tax_rate = Decimal("0.0")
        self.discount_rate = Decimal("0.0")
        self.notes = Config.DEFAULT_TERMS
        self.payment_terms = "Net 30"
        self.logo_path = None
        self.status = "Pending"
        self.received_date = "N/A"
        self.firebase_id = None
    
    @property
    def subtotal(self) -> Decimal:
        """Calculate subtotal from all items"""
        if not self.items:
            return Currency.quantize(Decimal("0.0"))
        total = sum(item.total for item in self.items)
        return Currency.quantize(total)
    
    @property
    def tax_amount(self) -> Decimal:
        """Calculate tax amount based on total amount"""
        if not self.items:
            return Currency.quantize(Decimal("0.0"))
        total_amount = sum(item.total for item in self.items)
        return Currency.quantize(total_amount * self.tax_rate / Decimal("100"))
    
    @property
    def total(self) -> Decimal:
        """Calculate total amount due (payment due + tax)"""
        if not self.items:
            return Currency.quantize(Decimal("0.0"))
        total_payment_due = sum(item.payment_due for item in self.items)
        total_amount = sum(item.total for item in self.items)
        tax_amount = total_amount * self.tax_rate / Decimal("100")
        return Currency.quantize(total_payment_due + tax_amount)
    
    def calculate_totals(self):
        """Recalculate all invoice totals after editing"""
        # This method can be empty as properties handle calculations dynamically
        pass
    
    def add_item(self, item: InvoiceItem):
        self.items.append(item)
    
    def remove_item(self, index: int):
        if 0 <= index < len(self.items):
            self.items.pop(index)
    
    def to_dict(self) -> Dict:
        return {
            "meta": {
                "invoice_number": self.invoice_number,
                "date": self.date,
                "due_date": self.due_date,
                "client_name": self.client_name,
                "client_email": self.client_email,
                "client_address": self.client_address,
                "tax_rate": float(self.tax_rate),
                "subtotal": float(self.subtotal),
                "tax_amount": float(self.tax_amount),
                "total": float(self.total),
                "notes": self.notes,
                "payment_terms": self.payment_terms,
                "status": self.status,
                "received_date": self.received_date
            },
            "items": [item.to_dict() for item in self.items]
        }

    @classmethod
    def from_dict(cls, data: Dict) -> 'Invoice':
        invoice = cls()
        meta = data.get("meta", {})
        
        invoice.invoice_number = meta.get("invoice_number", "")
        invoice.date = meta.get("date", invoice.date)
        invoice.due_date = meta.get("due_date", invoice.due_date)
        invoice.client_name = meta.get("client_name", "")
        invoice.client_email = meta.get("client_email", "")
        invoice.client_address = meta.get("client_address", "")
        invoice.tax_rate = Decimal(str(meta.get("tax_rate", 0.0)))
        invoice.notes = meta.get("notes", Config.DEFAULT_TERMS)
        invoice.status = meta.get("status", "Pending")
        invoice.received_date = meta.get("received_date", "N/A")
        
        # Store Firebase ID if available
        if 'firebase_id' in data:
            invoice.firebase_id = data['firebase_id']
        
        # Create items from data - PRESERVE payment_category
        invoice.items = []
        for item_data in data.get("items", []):
            try:
                raw_unit_price = item_data.get("unit_price", 0.0)
                raw_down_payment = item_data.get("down_payment", 0.0)
                raw_payment_category = item_data.get("payment_category", "")

                # Fix legacy invoices where "Down Payment (50%)" was mistakenly applied
                # to a stage amount (e.g. $50 stage → down_payment=$25, payment_due=$25).
                # Detect: down_payment == unit_price * 0.5 for a "Down Payment" category item.
                try:
                    _up = Decimal(str(raw_unit_price))
                    _dp = Decimal(str(raw_down_payment))
                    _cat_lo = (raw_payment_category or "").lower()
                    _is_deposit_cat = "down payment" in _cat_lo or "deposit" in _cat_lo
                    if _is_deposit_cat and _dp > 0 and abs(_dp - _up * Decimal("0.5")) < Decimal("0.01"):
                        raw_down_payment = 0.0
                except Exception:
                    pass

                item = InvoiceItem(
                    project_number=item_data.get("project_number", ""),
                    description=item_data.get("description", ""),
                    plant=item_data.get("plant", ""),
                    quantity=item_data.get("quantity", 1),
                    unit_price=raw_unit_price,
                    down_payment=raw_down_payment,
                    payment_due=item_data.get("payment_due", 0.0),
                    payment_category=raw_payment_category,
                )
                invoice.items.append(item)
            except Exception as e:
                log.warning("Error creating invoice item: %s", e)
                continue
        
        return invoice
# ---------- Excel Exporter ----------
class ExcelExporter:
    """Handles Excel export functionality"""
    
    @staticmethod
    def export_invoice(invoice: Invoice, output_path: Path):
        """Export single invoice to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"
        
        # Header
        ws.merge_cells('A1:G1')
        ws['A1'] = "INVOICE"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Invoice info
        ws['A3'] = "Invoice No:"
        ws['B3'] = invoice.invoice_number
        ws['A4'] = "Date:"
        ws['B4'] = invoice.date
        ws['A5'] = "Due Date:"
        ws['B5'] = invoice.due_date
        
        # Client info
        ws['A7'] = "Bill To:"
        ws['A8'] = invoice.client_name
        for i, line in enumerate(invoice.client_address.splitlines()):
            ws.cell(row=9 + i, column=1, value=line)
        
        # Items table with new columns
        headers = ["Project Number", "Description (PO & Address)", "Plant(Other)", "Quantity", "Unit Price", "Down Payment %", "Payment Due", "Total", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row_idx, item in enumerate(invoice.items, 13):
            ws.cell(row=row_idx, column=1, value=item.project_number)
            ws.cell(row=row_idx, column=2, value=item.description)
            ws.cell(row=row_idx, column=3, value=item.plant)
            ws.cell(row=row_idx, column=4, value=item.quantity)
            ws.cell(row=row_idx, column=5, value=float(item.unit_price))
            ws.cell(row=row_idx, column=6, value=float(item.down_payment))
            ws.cell(row=row_idx, column=7, value=float(item.payment_due))
        
        # Totals
        last_row = 13 + len(invoice.items)
        totals = [
            ("Subtotal:", invoice.subtotal),
        ]
        
        if invoice.tax_rate > 0:
            totals.append((f"Tax ({invoice.tax_rate}%):", invoice.tax_amount))
        
        totals.append(("Total Amount Due:", invoice.total))
        
        for i, (label, amount) in enumerate(totals):
            row = last_row + 2 + i
            ws.cell(row=row, column=6, value=label)
            ws.cell(row=row, column=7, value=float(amount))
            if "Total Amount Due" in label:
                ws.cell(row=row, column=6).font = Font(bold=True)
                ws.cell(row=row, column=7).font = Font(bold=True)
        
        wb.save(str(output_path))
    
    @staticmethod
    def export_history(invoices: List[Invoice], output_path: Path):
        """Export invoice history to Excel"""
        data = []
        for inv in invoices:
            data.append({
                "Invoice No": inv.invoice_number,
                "Date": inv.date,
                "Client": inv.client_name,
                "Subtotal": float(inv.subtotal),
                "Tax": float(inv.tax_amount),
                "Total": float(inv.total)
            })
        
        df = pd.DataFrame(data)
        df.to_excel(str(output_path), index=False, engine='openpyxl')

# ---------- Qt Widgets ----------

class ProjectLookupDialog(QtWidgets.QDialog):
    """Professional project browser — opens when the user clicks the lookup button on a project # field."""

    STATUS_COLORS = {
        "Paid":        ("#064e3b", "#d1fae5"),
        "In Progress": ("#1e40af", "#dbeafe"),
        "Not Started": ("#374151", "#f3f4f6"),
        "Cancelled":   ("#7f1d1d", "#fee2e2"),
        "On Hold":     ("#78350f", "#fef3c7"),
        "Completed":   ("#065f46", "#ecfdf5"),
    }

    def __init__(self, current_number="", projects=None, parent=None):
        super().__init__(parent)
        self.selected_project = None
        self._projects = sorted(projects or [], key=lambda p: p.get("project_number", ""))
        self._filtered = list(self._projects)
        self.setWindowTitle("Project Lookup")
        self.setMinimumSize(900, 600)
        self.setModal(True)
        self.setWindowFlags(self.windowFlags() | QtCore.Qt.WindowMaximizeButtonHint)
        self._build_ui(current_number)

    # ── UI Construction ────────────────────────────────────────────────────

    def _build_ui(self, current_number):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._make_header())
        root.addWidget(self._make_search_bar(current_number))
        root.addWidget(self._make_body(), 1)
        root.addWidget(self._make_footer())

        self._populate_table()
        if current_number:
            self._auto_select(current_number)

    def _make_header(self):
        header = QtWidgets.QFrame()
        header.setFixedHeight(70)
        header.setStyleSheet(
            "QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #0f172a, stop:1 #1a3a5c); }"
        )
        lay = QtWidgets.QHBoxLayout(header)
        lay.setContentsMargins(24, 0, 24, 0)
        lay.setSpacing(14)

        badge = QtWidgets.QLabel("P")
        badge.setFixedSize(38, 38)
        badge.setAlignment(QtCore.Qt.AlignCenter)
        badge.setStyleSheet(
            "background:#0f766e; color:#fff; font-size:18px; font-weight:900;"
            " border-radius:10px; font-family:'Inter','Segoe UI';"
        )

        col = QtWidgets.QVBoxLayout()
        col.setSpacing(2)
        t1 = QtWidgets.QLabel("Project Lookup")
        t1.setStyleSheet(
            "color:#fff; font-size:17px; font-weight:900; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        t2 = QtWidgets.QLabel("Search, select and link a project to this invoice line")
        t2.setStyleSheet(
            "color:rgba(255,255,255,0.5); font-size:11px; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        col.addWidget(t1)
        col.addWidget(t2)

        lay.addWidget(badge)
        lay.addLayout(col, 1)
        return header

    def _make_search_bar(self, current_number):
        bar = QtWidgets.QFrame()
        bar.setStyleSheet(
            "QFrame { background:#f8fafc; border-bottom:1px solid #e2e8f0; }"
        )
        lay = QtWidgets.QHBoxLayout(bar)
        lay.setContentsMargins(20, 10, 20, 10)
        lay.setSpacing(10)

        icon = QtWidgets.QLabel("🔍")
        icon.setStyleSheet("font-size:14px; background:transparent;")

        self._search_edit = QtWidgets.QLineEdit()
        self._search_edit.setPlaceholderText(
            "Search by project number, name, or client…"
        )
        self._search_edit.setMinimumHeight(36)
        self._search_edit.setStyleSheet(
            "QLineEdit { background:#fff; border:1.5px solid #cbd5e1; border-radius:8px;"
            " padding:0 12px; font-size:13px; font-family:'Inter','Segoe UI'; color:#0f172a; }"
            "QLineEdit:focus { border-color:#0f766e; }"
        )
        self._search_edit.textChanged.connect(self._filter)
        if current_number:
            self._search_edit.setText(current_number)

        clear = QtWidgets.QPushButton("✕")
        clear.setFixedSize(32, 32)
        clear.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        clear.setToolTip("Clear search")
        clear.setStyleSheet(
            "QPushButton { background:#e2e8f0; color:#64748b; border:none;"
            " border-radius:7px; font-size:11px; font-weight:800; }"
            "QPushButton:hover { background:#cbd5e1; }"
        )
        clear.clicked.connect(self._search_edit.clear)

        lay.addWidget(icon)
        lay.addWidget(self._search_edit, 1)
        lay.addWidget(clear)
        return bar

    def _make_body(self):
        body = QtWidgets.QFrame()
        body.setStyleSheet("QFrame { background:#f4f7fb; }")
        lay = QtWidgets.QHBoxLayout(body)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(14)
        lay.addWidget(self._make_table(), 3)
        lay.addWidget(self._make_detail_panel())
        return body

    def _make_table(self):
        self._table = QtWidgets.QTableWidget()
        self._table.setColumnCount(5)
        self._table.setHorizontalHeaderLabels(
            ["Project #", "Project Name", "Client", "Status", "Contract Value"]
        )
        self._table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        self._table.setShowGrid(False)
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        self._table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self._table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 13px;
                gridline-color: transparent;
            }
            QTableWidget::item { padding: 10px 12px; color: #334155; }
            QTableWidget::item:selected { background: #f0fdf4; color: #065f46; }
            QTableWidget::item:alternate { background: #f8fafc; }
            QHeaderView::section {
                background: #f1f5f9; color: #475569; font-weight: 800;
                font-size: 11px; padding: 8px 12px; border: none;
                border-bottom: 2px solid #e2e8f0;
                font-family: 'Inter', 'Segoe UI';
            }
        """)
        self._table.itemSelectionChanged.connect(self._on_selection_changed)
        self._table.itemDoubleClicked.connect(self._accept_selection)
        return self._table

    def _make_detail_panel(self):
        self._detail_panel = QtWidgets.QFrame()
        self._detail_panel.setFixedWidth(268)
        self._detail_panel.setStyleSheet(
            "QFrame { background:#fff; border:1px solid #e2e8f0; border-radius:10px; }"
        )
        lay = QtWidgets.QVBoxLayout(self._detail_panel)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        title = QtWidgets.QLabel("Project Details")
        title.setStyleSheet(
            "color:#0f172a; font-size:13px; font-weight:900;"
            " font-family:'Inter','Segoe UI'; border:none; border-bottom:1px solid #e2e8f0;"
            " padding-bottom:8px;"
        )
        lay.addWidget(title)

        self._detail_scroll = QtWidgets.QScrollArea()
        self._detail_scroll.setWidgetResizable(True)
        self._detail_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        self._detail_scroll.setStyleSheet("QScrollArea { background:transparent; border:none; }")

        self._detail_inner = QtWidgets.QWidget()
        self._detail_inner.setStyleSheet("background:transparent;")
        self._detail_inner_lay = QtWidgets.QVBoxLayout(self._detail_inner)
        self._detail_inner_lay.setContentsMargins(0, 0, 0, 0)
        self._detail_inner_lay.setSpacing(6)

        placeholder = QtWidgets.QLabel("Select a project\nto view its details")
        placeholder.setAlignment(QtCore.Qt.AlignCenter)
        placeholder.setStyleSheet(
            "color:#94a3b8; font-size:12px; font-family:'Inter','Segoe UI';"
        )
        self._detail_inner_lay.addWidget(placeholder)
        self._detail_inner_lay.addStretch()

        self._detail_scroll.setWidget(self._detail_inner)
        lay.addWidget(self._detail_scroll, 1)
        return self._detail_panel

    def _make_footer(self):
        footer = QtWidgets.QFrame()
        footer.setFixedHeight(58)
        footer.setStyleSheet(
            "QFrame { background:#fff; border-top:1px solid #e2e8f0; }"
        )
        lay = QtWidgets.QHBoxLayout(footer)
        lay.setContentsMargins(20, 0, 20, 0)
        lay.setSpacing(10)

        self._count_lbl = QtWidgets.QLabel()
        self._count_lbl.setStyleSheet(
            "color:#64748b; font-size:12px; font-family:'Inter','Segoe UI'; border:none;"
        )
        lay.addWidget(self._count_lbl, 1)

        cancel = QtWidgets.QPushButton("Cancel")
        cancel.setFixedSize(100, 36)
        cancel.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel.setStyleSheet(
            "QPushButton { background:#f1f5f9; color:#475569; border:none;"
            " border-radius:8px; font-weight:700; font-size:13px; }"
            "QPushButton:hover { background:#e2e8f0; }"
        )
        cancel.clicked.connect(self.reject)

        self._select_btn = QtWidgets.QPushButton("✔  Select Project")
        self._select_btn.setFixedSize(150, 36)
        self._select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._select_btn.setEnabled(False)
        self._select_btn.setStyleSheet(
            "QPushButton { background:#0f766e; color:#fff; border:none;"
            " border-radius:8px; font-weight:700; font-size:13px; }"
            "QPushButton:hover { background:#0d625c; }"
            "QPushButton:disabled { background:#cbd5e1; color:#94a3b8; }"
        )
        self._select_btn.clicked.connect(self._accept_selection)

        lay.addWidget(cancel)
        lay.addWidget(self._select_btn)
        return footer

    # ── Table population ───────────────────────────────────────────────────

    def _populate_table(self):
        self._table.setRowCount(0)
        for r, proj in enumerate(self._filtered):
            self._table.insertRow(r)

            pn = proj.get("project_number", "")
            pn_item = QtWidgets.QTableWidgetItem(pn)
            pn_item.setFont(QtGui.QFont("Consolas", 11, QtGui.QFont.Bold))
            pn_item.setForeground(QtGui.QColor("#0f766e"))
            pn_item.setTextAlignment(QtCore.Qt.AlignCenter)
            pn_item.setData(QtCore.Qt.UserRole, proj)

            name_item = QtWidgets.QTableWidgetItem(proj.get("project_name", ""))
            name_item.setFont(QtGui.QFont("Inter", 12))

            client_item = QtWidgets.QTableWidgetItem(proj.get("company", ""))
            client_item.setForeground(QtGui.QColor("#64748b"))

            status = proj.get("status", "Not Started")
            sc = self.STATUS_COLORS.get(status, ("#374151", "#f3f4f6"))
            status_item = QtWidgets.QTableWidgetItem(f"  {status}  ")
            status_item.setTextAlignment(QtCore.Qt.AlignCenter)
            status_item.setForeground(QtGui.QColor(sc[0]))
            status_item.setBackground(QtGui.QColor(sc[1]))

            raw = proj.get("contract_value", proj.get("total_amount", ""))
            try:
                contract_txt = f"${float(raw):,.2f}" if raw else "—"
            except (ValueError, TypeError):
                contract_txt = str(raw) if raw else "—"
            contract_item = QtWidgets.QTableWidgetItem(contract_txt)
            contract_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)

            for col, itm in enumerate([pn_item, name_item, client_item, status_item, contract_item]):
                self._table.setItem(r, col, itm)
            self._table.setRowHeight(r, 44)

        shown, total = len(self._filtered), len(self._projects)
        self._count_lbl.setText(f"Showing {shown} of {total} project{'s' if total != 1 else ''}")

    # ── Filtering ──────────────────────────────────────────────────────────

    def _filter(self, text):
        q = text.strip().lower()
        self._filtered = [
            p for p in self._projects
            if not q
            or q in p.get("project_number", "").lower()
            or q in p.get("project_name", "").lower()
            or q in p.get("company", "").lower()
        ]
        self._populate_table()

    # ── Selection / Detail panel ───────────────────────────────────────────

    def _on_selection_changed(self):
        row = self._table.currentRow()
        item = self._table.item(row, 0) if row >= 0 else None
        proj = item.data(QtCore.Qt.UserRole) if item else None
        self._select_btn.setEnabled(proj is not None)
        self._refresh_detail(proj)

    def _refresh_detail(self, proj):
        # Clear old content
        while self._detail_inner_lay.count():
            child = self._detail_inner_lay.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not proj:
            ph = QtWidgets.QLabel("Select a project\nto view its details")
            ph.setAlignment(QtCore.Qt.AlignCenter)
            ph.setStyleSheet("color:#94a3b8; font-size:12px; font-family:'Inter','Segoe UI';")
            self._detail_inner_lay.addWidget(ph)
            self._detail_inner_lay.addStretch()
            return

        # Status badge
        status = proj.get("status", "Not Started")
        sc = self.STATUS_COLORS.get(status, ("#374151", "#f3f4f6"))
        status_lbl = QtWidgets.QLabel(f"  {status}  ")
        status_lbl.setFixedHeight(28)
        status_lbl.setAlignment(QtCore.Qt.AlignCenter)
        status_lbl.setStyleSheet(
            f"color:{sc[0]}; background:{sc[1]}; border-radius:8px;"
            " font-size:11px; font-weight:800; font-family:'Inter','Segoe UI';"
        )
        self._detail_inner_lay.addWidget(status_lbl)

        def _row(label, value):
            if not value:
                return
            card = QtWidgets.QFrame()
            card.setStyleSheet(
                "QFrame { background:#f8fafc; border-radius:6px; border:none; }"
            )
            cl = QtWidgets.QVBoxLayout(card)
            cl.setContentsMargins(10, 7, 10, 7)
            cl.setSpacing(2)
            lbl_w = QtWidgets.QLabel(label.upper())
            lbl_w.setStyleSheet(
                "color:#94a3b8; font-size:9px; font-weight:800;"
                " letter-spacing:0.8px; font-family:'Inter','Segoe UI'; background:transparent;"
            )
            val_w = QtWidgets.QLabel(str(value))
            val_w.setWordWrap(True)
            val_w.setStyleSheet(
                "color:#0f172a; font-size:12px; font-weight:700;"
                " font-family:'Inter','Segoe UI'; background:transparent;"
            )
            cl.addWidget(lbl_w)
            cl.addWidget(val_w)
            self._detail_inner_lay.addWidget(card)

        _row("Project #", proj.get("project_number"))
        _row("Project Name", proj.get("project_name"))
        _row("Client", proj.get("company"))
        _row("PO / WO Number", proj.get("po_wo_number"))
        _row("Plant", proj.get("plant"))
        _row("Sales Person", proj.get("sales"))
        _row("Start Date", proj.get("start_date"))

        raw = proj.get("contract_value", proj.get("total_amount", ""))
        if raw:
            try:
                _row("Contract Value", f"${float(raw):,.2f}")
            except (ValueError, TypeError):
                _row("Contract Value", str(raw))

        self._detail_inner_lay.addStretch()

    def _auto_select(self, number):
        for row in range(self._table.rowCount()):
            item = self._table.item(row, 0)
            if item and item.text().strip() == number.strip():
                self._table.selectRow(row)
                self._table.scrollToItem(item)
                break

    def _accept_selection(self, *_):
        row = self._table.currentRow()
        if row < 0:
            return
        item = self._table.item(row, 0)
        if item:
            self.selected_project = item.data(QtCore.Qt.UserRole)
        self.accept()

    def keyPressEvent(self, event):
        if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            if self._select_btn.isEnabled():
                self._accept_selection()
        elif event.key() == QtCore.Qt.Key_Escape:
            self.reject()
        else:
            super().keyPressEvent(event)


class ItemRowWidget(QtWidgets.QWidget):
    """Widget for a single invoice item row"""
    removed = pyqtSignal()
    project_changed = pyqtSignal(str)

    # ── Payment stage labels ─────────────────────────────────────────────
    DEPOSIT_LABEL        = "Down Payment (50%)"
    REMAINING_LABEL      = "Remaining Balance"
    INSTALLMENT_1_LABEL  = "1st Installment"
    INSTALLMENT_2_LABEL  = "2nd Installment"
    INSTALLMENT_3_LABEL  = "3rd Installment"
    INSTALLMENT_4_LABEL  = "4th Installment"
    TERM_2_LABEL         = "2nd Installment"
    TERM_3_LABEL         = "3rd Installment"
    TERM_4_LABEL         = "4th Installment"
    FINAL_PAYMENT_LABEL  = "Final Payment"
    FULL_AMOUNT_LABEL    = "Full Amount"

    @classmethod
    def normalize_payment_label(cls, label: str) -> str:
        """Map any saved payment name to the current unified label."""
        text = (label or "").strip()
        if not text or text in ("Select payment category", "Select payment stage"):
            return ""
        lo = text.lower()
        if "down payment (50%)" in lo or "deposit received (50%)" in lo:
            return cls.DEPOSIT_LABEL
        if lo.startswith("down payment (") and "%" in lo:
            return text
        if any(x in lo for x in ("deposit", "down payment", "down_payment")):
            return cls.DEPOSIT_LABEL
        if any(x in lo for x in ("remaining balance", "remaining 50%", "balance due")):
            return cls.REMAINING_LABEL
        if "1st installment" in lo or "first installment" in lo:
            return cls.INSTALLMENT_1_LABEL
        if "2nd installment" in lo or "second installment" in lo:
            return cls.INSTALLMENT_2_LABEL
        if "3rd installment" in lo or "third installment" in lo:
            return cls.INSTALLMENT_3_LABEL
        if "4th installment" in lo or "fourth installment" in lo:
            return cls.INSTALLMENT_4_LABEL
        if any(x in lo for x in ("term 2", "2nd payment")):
            return cls.TERM_2_LABEL
        if any(x in lo for x in ("term 3", "3rd payment")):
            return cls.TERM_3_LABEL
        if any(x in lo for x in ("term 4", "4th payment")):
            return cls.TERM_4_LABEL
        if any(x in lo for x in ("final payment", "final payment due")):
            return cls.FINAL_PAYMENT_LABEL
        if any(x in lo for x in ("full amount", "full amount due", "due payment")):
            return cls.FULL_AMOUNT_LABEL
        return text

    def __init__(self, item: InvoiceItem = None):
        super().__init__()
        self.item = item or InvoiceItem()
        self.project_number_edit = None
        self.desc_edit = None
        self.plant_edit = None
        self.qty_spin = None
        self.price_edit = None
        self.down_payment_combo = None
        self.payment_due_spin = None
        self.total_label = None
        self.remove_btn = None
        self.item_number_label = None
        self._stage_locked = False  # True when loaded from a project payment stage
        self.init_ui()

    @classmethod
    def is_deposit_payment(cls, label: str) -> bool:
        return cls.normalize_payment_label(label) == cls.DEPOSIT_LABEL
    
    def init_ui(self):
        layout = QtWidgets.QGridLayout(self)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setHorizontalSpacing(10)
        layout.setVerticalSpacing(8)
        self.setStyleSheet("""
            ItemRowWidget {
                background: #ffffff;
                border: 1px solid #ffffff;
                border-radius: 10px;
            }
        """)
        
        # Project Number - Center aligned
        self.project_number_edit = QtWidgets.QLineEdit(self.item.project_number)
        self.project_number_edit.setPlaceholderText("Project #")
        self.project_number_edit.setMinimumHeight(35)
        self.project_number_edit.setMinimumWidth(110)
        self.project_number_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.project_number_edit.editingFinished.connect(
            lambda: self.on_project_number_changed(self.project_number_edit.text())
        )

        # Set Enter key navigation for Project Number field
        self.project_number_edit.installEventFilter(self)

        # Wrap project number field + lookup button in a container
        self._pn_container = QtWidgets.QFrame()
        self._pn_container.setStyleSheet("QFrame { background: transparent; border: none; }")
        self._pn_container.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed
        )
        _pn_lay = QtWidgets.QHBoxLayout(self._pn_container)
        _pn_lay.setContentsMargins(0, 0, 0, 0)
        _pn_lay.setSpacing(6)
        _pn_lay.addWidget(self.project_number_edit)

        self._lookup_btn = QtWidgets.QPushButton("⊕")
        self._lookup_btn.setFixedSize(35, 35)
        self._lookup_btn.setToolTip("Browse & select a project")
        self._lookup_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._lookup_btn.setStyleSheet("""
            QPushButton {
                background: #0f766e;
                color: white;
                border: none;
                border-radius: 7px;
                font-size: 17px;
                font-weight: 900;
            }
            QPushButton:hover { background: #0d6360; }
            QPushButton:pressed { background: #0a4f4c; }
        """)
        self._lookup_btn.clicked.connect(self.open_project_lookup)
        _pn_lay.addWidget(self._lookup_btn)

        # Description - DECREASED WIDTH and center aligned
        self.desc_edit = QtWidgets.QLineEdit(self.item.description)
        self.desc_edit.setPlaceholderText("Project Name")
        self.desc_edit.setMinimumHeight(35)
        self.desc_edit.setAlignment(QtCore.Qt.AlignCenter)
        
        # Set Enter key navigation for Description field
        self.desc_edit.installEventFilter(self)
        
        # Plant - Center aligned
        self.plant_edit = QtWidgets.QLineEdit(self.item.plant)
        self.plant_edit.setPlaceholderText("Plant(Other)")
        self.plant_edit.setMinimumHeight(35)
        self.plant_edit.setAlignment(QtCore.Qt.AlignCenter)
        
        # Set Enter key navigation for Plant field
        self.plant_edit.installEventFilter(self)
        
        # Quantity - Center aligned
        self.qty_spin = QtWidgets.QSpinBox()
        self.qty_spin.setRange(1, 1000000)
        self.qty_spin.setValue(self.item.quantity)
        self.qty_spin.setMinimumHeight(35)
        self.qty_spin.setMinimumWidth(70)
        self.qty_spin.setAlignment(QtCore.Qt.AlignCenter)
        
        # Set Enter key navigation for Quantity field
        self.qty_spin.installEventFilter(self)
        
        # Unit Price - Center aligned - CHANGED TO LINE EDIT LIKE JOB FORM
        self.price_edit = QtWidgets.QLineEdit(str(float(self.item.unit_price)) if self.item.unit_price != Decimal('0') else "")
        self.price_edit.setPlaceholderText("$0.00")
        self.price_edit.setMinimumHeight(35)
        self.price_edit.setMinimumWidth(95)
        self.price_edit.setAlignment(QtCore.Qt.AlignCenter)
        self.price_edit.textChanged.connect(self.validate_price_input)
        self.price_edit.textChanged.connect(self.update_total)
        
        # Set Enter key navigation for Price field
        self.price_edit.installEventFilter(self)

        # Payment dropdown - uses business-friendly labels while accepting old saved values.
        self.down_payment_combo = QtWidgets.QComboBox()
        
        # Add placeholder item that's not selectable
        self.down_payment_combo.addItem("Select payment stage")  # Placeholder
        self.down_payment_combo.addItems([
            self.DEPOSIT_LABEL,        # Down Payment (50%)
            self.REMAINING_LABEL,      # Remaining Balance
            self.INSTALLMENT_1_LABEL,  # 1st Installment
            self.INSTALLMENT_2_LABEL,  # 2nd Installment
            self.INSTALLMENT_3_LABEL,  # 3rd Installment
            self.INSTALLMENT_4_LABEL,  # 4th Installment
            self.FINAL_PAYMENT_LABEL,  # Final Payment
            self.FULL_AMOUNT_LABEL,    # Full Amount
        ])
        
        # Set placeholder as initial selection but make it non-selectable after selection
        self.down_payment_combo.setCurrentIndex(0)
        
        self.down_payment_combo.setMinimumHeight(35)
        self.down_payment_combo.setMinimumWidth(160)
        self.down_payment_combo.currentIndexChanged.connect(self.on_down_payment_changed)
        self.down_payment_combo.currentIndexChanged.connect(self.update_total)
        
        # Install event filter to prevent editing/backspace
        self.down_payment_combo.installEventFilter(self)
        
        # Center align the dropdown text
        self.down_payment_combo.setStyleSheet("""
            QComboBox {
                text-align: center;
                color: #666;
            }
            QComboBox:focus {
                color: #000;
            }
            QComboBox QAbstractItemView {
                text-align: center;
            }
        """)
        
        # Set initial value based on existing item
        self.down_payment_combo.blockSignals(True)
        normalized_payment = self.normalize_payment_label(self.item.payment_category)
        if normalized_payment:
            self.down_payment_combo.setCurrentText(normalized_payment)
        elif self.item.down_payment > 0:
            self.down_payment_combo.setCurrentText(self.DEPOSIT_LABEL)
        else:
            # If it's a new item, show placeholder
            if self.item.down_payment == 0 and self.item.payment_due == 0:
                self.down_payment_combo.setCurrentIndex(0)  # Show placeholder
        self.down_payment_combo.blockSignals(False)
        
        # Set Enter key navigation for Down Payment field
        self.down_payment_combo.installEventFilter(self)
        
        # Payment Due (read-only, auto-calculated) - Center aligned
        self.payment_due_spin = QtWidgets.QDoubleSpinBox()  # THIS MUST BE DEFINED
        self.payment_due_spin.setMinimum(0)  # Only set minimum to 0, no maximum
        self.payment_due_spin.setMaximum(999999999) 
        self.payment_due_spin.setDecimals(2)
        self.payment_due_spin.setValue(float(self.item.payment_due))
        self.payment_due_spin.setPrefix("$ ")
        self.payment_due_spin.setMinimumHeight(35)
        self.payment_due_spin.setMinimumWidth(105)
        self.payment_due_spin.setReadOnly(True)
        self.payment_due_spin.setStyleSheet("background-color: #f0f0f0; color: #666; text-align: center;")
        self.payment_due_spin.setAlignment(QtCore.Qt.AlignCenter)
        
        # Total (read-only) - Center aligned
        self.total_label = QtWidgets.QLabel(Currency.format(self.item.total))
        self.total_label.setAlignment(QtCore.Qt.AlignCenter)  # CHANGED: Center aligned
        self.total_label.setMinimumWidth(90)
        self.total_label.setMinimumHeight(35)
        self.total_label.setStyleSheet("""
            QLabel {
                background-color: #ffffff;
                border: 1px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px;
                font-weight: 800;
                font-size: 12px;
                text-align: center;  /* ADDED: Ensure text is centered */
            }
        """)
        
        # Remove button
        self.remove_btn = QtWidgets.QPushButton("a *")
        self.remove_btn.setMinimumWidth(90)
        self.remove_btn.setText("X")
        self.remove_btn.setText("Remove")
        self.remove_btn.setToolTip("Remove this line item")
        self.remove_btn.setMinimumHeight(35)
        self.remove_btn.setStyleSheet("""
            QPushButton {
                background-color: #fff1f2;
                color: #be123c;
                border: 1px solid #fecdd3;
                border-radius: 7px;
                font-weight: 800;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #ffe4e6;
                border-color: #fb7185;
            }
        """)
        self.remove_btn.clicked.connect(self.removed)

        self.item_number_label = QtWidgets.QLabel("Line Item")
        self.item_number_label.setStyleSheet("""
            QLabel {
                color: #0f766e;
                font-size: 14px;
                font-weight: 900;
                padding-bottom: 2px;
            }
        """)
        layout.addWidget(self.item_number_label, 0, 0, 1, 9)
        
        for widget in [
            self.project_number_edit,
            self.desc_edit,
            self.plant_edit,
            self.qty_spin,
            self.price_edit,
            self.down_payment_combo,
            self.payment_due_spin,
            self.total_label,
            self.remove_btn,
        ]:
            widget.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)

        def add_field(row, col, label_text, widget, colspan=1):
            label = QtWidgets.QLabel(label_text)
            label.setStyleSheet("font-weight: 800; color: #314155; font-size: 12px;")
            label.setAlignment(QtCore.Qt.AlignLeft)
            layout.addWidget(label, row, col, 1, colspan)
            layout.addWidget(widget, row + 1, col, 1, colspan)

        add_field(1, 0, "Project #", self._pn_container)
        add_field(1, 1, "Project Name", self.desc_edit)
        add_field(1, 2, "Plant", self.plant_edit)
        add_field(1, 3, "Qty", self.qty_spin)
        add_field(1, 4, "Unit Price", self.price_edit)
        add_field(1, 5, "Payment", self.down_payment_combo)
        add_field(1, 6, "Due", self.payment_due_spin)
        add_field(1, 7, "Total", self.total_label)
        add_field(1, 8, "Action", self.remove_btn)

        layout.setColumnStretch(0, 2)
        layout.setColumnStretch(1, 4)
        layout.setColumnStretch(2, 1)
        layout.setColumnStretch(3, 1)
        layout.setColumnStretch(4, 1)
        layout.setColumnStretch(5, 2)
        layout.setColumnStretch(6, 1)
        layout.setColumnStretch(7, 1)
        layout.setColumnStretch(8, 1)
        
        # Connect signals
        self.project_number_edit.textChanged.connect(self.update_item)
        self.desc_edit.textChanged.connect(self.update_item)
        self.plant_edit.textChanged.connect(self.update_item)
        self.qty_spin.valueChanged.connect(self.update_total)

    def set_item_number(self, number: int):
        """Update the visible line item label."""
        if self.item_number_label:
            self.item_number_label.setText(f"Line Item {number}")
        
    def eventFilter(self, source, event):
        """Handle Enter key navigation between fields"""
        if event.type() == QtCore.QEvent.KeyPress and event.key() == QtCore.Qt.Key_Return:
            # Get all widgets in tab order
            widgets = [
                self.project_number_edit,
                self.desc_edit,
                self.plant_edit,
                self.qty_spin,
                self.price_edit,
                self.down_payment_combo,
                self.payment_due_spin,
                self.remove_btn
            ]
            
            # Find current widget and move to next
            for i, widget in enumerate(widgets):
                if widget == source:
                    next_index = (i + 1) % len(widgets)
                    widgets[next_index].setFocus()
                    
                    # Special handling for spin boxes
                    if isinstance(widgets[next_index], QtWidgets.QSpinBox) or isinstance(widgets[next_index], QtWidgets.QDoubleSpinBox):
                        widgets[next_index].selectAll()
                    
                    return True
        
        # Special handling for QComboBox to prevent editing
        if source == self.down_payment_combo:
            if event.type() == QtCore.QEvent.KeyPress:
                # Prevent backspace and delete keys
                if event.key() in [QtCore.Qt.Key_Backspace, QtCore.Qt.Key_Delete]:
                    return True
                # Prevent any text input
                if event.text():
                    return True
        
        return super().eventFilter(source, event)
    
    def on_down_payment_changed(self, index):
        """Handle dropdown selection change"""
        if index == 0:  # Placeholder item selected
            # Change color to indicate placeholder
            self.down_payment_combo.setStyleSheet("""
                QComboBox {
                    text-align: center;
                    color: #999;
                    font-style: italic;
                }
                QComboBox:focus {
                    color: #999;
                }
                QComboBox QAbstractItemView {
                    text-align: center;
                }
            """)
        else:
            # Change color to indicate actual selection
            self.down_payment_combo.setStyleSheet("""
                QComboBox {
                    text-align: center;
                    color: #000;
                    font-style: normal;
                }
                QComboBox:focus {
                    color: #000;
                }
                QComboBox QAbstractItemView {
                    text-align: center;
                }
            """)
        
    def lock_to_stage(self, label: str):
        """Lock this row to a single project payment stage — replaces the full dropdown."""
        self._stage_locked = True
        self.down_payment_combo.blockSignals(True)
        self.down_payment_combo.clear()
        self.down_payment_combo.addItem(label)
        self.down_payment_combo.setCurrentIndex(0)
        self.down_payment_combo.setEnabled(False)
        self.down_payment_combo.setStyleSheet("""
            QComboBox {
                text-align: center;
                color: #0f766e;
                font-weight: 800;
                background: #f0fdf9;
                border: 1.5px solid #99f6e4;
                border-radius: 7px;
                padding: 4px 8px;
            }
            QComboBox::drop-down { width: 0px; border: none; }
            QComboBox::down-arrow { image: none; }
        """)
        self.down_payment_combo.blockSignals(False)

        # Make project name read-only so it always reflects the loaded project
        if self.desc_edit is not None:
            self.desc_edit.setReadOnly(True)
            self.desc_edit.setStyleSheet(
                self.desc_edit.styleSheet() +
                "QLineEdit { background: #f0fdf9; color: #0f766e; font-weight: 600; }"
            )
        # Also lock amount and project number
        self.lock_project_fields()

    def lock_project_fields(self):
        """Lock project number and amount so they cannot be changed after loading from a project."""
        _locked_style = "QLineEdit { background: #f0fdf9; color: #0f766e; font-weight: 600; border: 1.5px solid #99f6e4; }"
        if self.project_number_edit is not None:
            self.project_number_edit.setReadOnly(True)
            self.project_number_edit.setStyleSheet(self.project_number_edit.styleSheet() + _locked_style)
        if self.price_edit is not None:
            self.price_edit.setReadOnly(True)
            self.price_edit.setStyleSheet(self.price_edit.styleSheet() + _locked_style)
        if hasattr(self, "_lookup_btn") and self._lookup_btn is not None:
            self._lookup_btn.setEnabled(False)
            self._lookup_btn.setToolTip("Project number is locked after loading from project")

    def validate_price_input(self):
        """Validate price input to accept only numbers and automatically add $ prefix"""
        # Block signals to prevent recursive calls
        self.price_edit.blockSignals(True)
        
        try:
            text = self.price_edit.text().strip()
            cursor_pos = self.price_edit.cursorPosition()
            
            # If text is empty, just return
            if not text:
                self.price_edit.blockSignals(False)
                return
            
            # Remove any non-numeric characters except decimal point
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            
            # Remove any existing $ signs
            cleaned = cleaned.replace('$', '')
            
            # Ensure only one decimal point
            if cleaned.count('.') > 1:
                # Remove extra decimal points
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])
            
            # Add $ prefix if we have any valid number
            if cleaned:
                # Ensure it starts with $
                if not text.startswith('$'):
                    final_text = f"${cleaned}"
                else:
                    # Keep $ and clean the rest
                    final_text = f"${cleaned}"
            else:
                final_text = ""
            
            # Update the field if it changed
            if final_text != text:
                self.price_edit.setText(final_text)
                
                # Adjust cursor position
                # If we added $ at the beginning, move cursor right by 1
                if not text.startswith('$') and final_text.startswith('$'):
                    new_pos = min(cursor_pos + 1, len(final_text))
                else:
                    new_pos = min(cursor_pos, len(final_text))
                
                self.price_edit.setCursorPosition(new_pos)
                
        finally:
            self.price_edit.blockSignals(False)
            
    def on_project_number_changed(self, text):
        """When project number is changed, emit signal for auto-fill"""
        self.update_item()
        self.project_changed.emit(text.strip())

    def open_project_lookup(self):
        """Open the professional project lookup dialog from the ⊕ button."""
        try:
            projects = FirebaseManager.load_projects() if FIREBASE_AVAILABLE else []
        except Exception:
            projects = []
        current = self.project_number_edit.text().strip()
        dialog = ProjectLookupDialog(current, projects, parent=self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted and dialog.selected_project:
            p = dialog.selected_project
            pn = p.get("project_number", "")
            self.project_number_edit.setText(pn)
            if not self.desc_edit.text().strip() and p.get("project_name"):
                self.desc_edit.setText(p.get("project_name", ""))
            if not self.plant_edit.text().strip() and p.get("plant"):
                self.plant_edit.setText(p.get("plant", ""))
            self.on_project_number_changed(pn)

    def update_total(self):
        try:
            # Skip calculation if placeholder is selected (index 0 = unselected)
            # but NOT when locked to a project stage — there index 0 IS the valid label
            if self.down_payment_combo.currentIndex() == 0 and not self._stage_locked:
                self.item.down_payment = Currency.quantize(Decimal('0.0'))
                self.item.payment_due = Currency.quantize(Decimal('0.0'))
                self.item.payment_category = ""
                if hasattr(self, 'payment_due_spin') and self.payment_due_spin is not None:
                    self.payment_due_spin.setValue(0.0)
                if self.total_label is not None:
                    self.total_label.setText("$0.00")
                self.update_item()
                return
            
            self.item.quantity = self.qty_spin.value()
            
            # Get unit price from line edit
            price_text = self.price_edit.text().replace("$", "").replace(",", "").strip()
            if price_text:
                try:
                    self.item.unit_price = Currency.quantize(float(price_text))
                except ValueError:
                    self.item.unit_price = Currency.quantize(0.0)
            else:
                self.item.unit_price = Currency.quantize(0.0)
            
            # Calculate total amount
            total_amount = Decimal(str(self.item.quantity)) * self.item.unit_price

            # Get the payment category - FIXED: Store it
            payment_category = self.down_payment_combo.currentText()
            self.item.payment_category = payment_category

            # The payment category is a label only — the entered unit_price IS the full
            # amount due for this invoice stage. Never re-apply a percentage here.
            self.item.down_payment = Currency.quantize(Decimal('0.0'))
            self.item.payment_due = Currency.quantize(total_amount)
            
            # Check if payment_due_spin exists before setting value
            if hasattr(self, "payment_due_spin") and self.payment_due_spin is not None:
                self.payment_due_spin.setValue(float(self.item.payment_due))
            else:
                log.warning("a   Warning: payment_due_spin not available")
            
            if self.total_label is not None:
                self.total_label.setText(Currency.format(self.item.total))
            self.update_item()
            
            # Emit update to parent window
            if hasattr(self.parent(), 'parent') and hasattr(self.parent().parent(), 'update_totals'):
                self.parent().parent().update_totals()
                
        except Exception as e:
            log.warning("Error in update_total: %s", e)
            import traceback
            traceback.print_exc()
        
    def update_item(self):
        self.item.project_number = self.project_number_edit.text()
        self.item.description = self.desc_edit.text()
        self.item.plant = self.plant_edit.text()
    
    def get_item(self) -> InvoiceItem:
        """Get the invoice item with current values"""
        # Ensure price is updated from line edit before returning
        price_text = self.price_edit.text().replace("$", "").replace(",", "").strip()
        if price_text:
            try:
                self.item.unit_price = Currency.quantize(float(price_text))
            except ValueError:
                self.item.unit_price = Currency.quantize(0.0)
        
        # Get the payment category - FIXED: Always save the selected category
        payment_category = self.normalize_payment_label(self.down_payment_combo.currentText())
        # Don't save placeholder text - but if it's the placeholder, treat as empty
        if not payment_category:
            payment_category = ""
        
        # Update the item's payment category
        self.item.payment_category = payment_category
        
        # Calculate down payment based on selection
        total_amount = Decimal(str(self.item.quantity)) * self.item.unit_price
        
        # Parse the selection text to determine payment calculation
        # Payment category is a label only — the entered unit_price IS the full amount
        # due for this invoice stage. Never re-apply a percentage to the stage amount.
        down_payment = Currency.quantize(Decimal('0.0'))
        payment_due = Currency.quantize(total_amount)
        
        # Update the item's down payment and payment due
        self.item.down_payment = down_payment
        self.item.payment_due = payment_due
        
        return InvoiceItem(
            project_number=self.item.project_number,
            description=self.item.description,
            plant=self.item.plant,
            quantity=self.item.quantity,
            unit_price=float(self.item.unit_price),
            down_payment=float(down_payment),
            payment_due=float(payment_due),
            payment_category=payment_category  # Pass the category
        )

class ScrollArea(QtWidgets.QScrollArea):
    """Custom scroll area with better styling"""
    def __init__(self):
        super().__init__()
        self.setWidgetResizable(True)
        self.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.setStyleSheet("""
            QScrollArea {
                border: 1px solid #dee2e6;
                border-radius: 6px;
                background-color: white;
            }
            QScrollBar:vertical {
                border: none;
                background: #f8f9fa;
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #ced4da;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #adb5bd;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)

class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year  # Minimum year (1 AD)
        self.end_year = end_year      # Maximum year (9999 AD)
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Navigation buttons
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton("a--a--")
        self.prev_block_btn.setFixedSize(40, 30)
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)
        
        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)
        
        self.next_block_btn = QtWidgets.QPushButton("a- a- ")
        self.next_block_btn.setFixedSize(40, 30)
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        
        layout.addLayout(nav_layout)
        
        # Year grid container
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create 3x3 grid of year buttons
        self.year_buttons = []
        
        # Calculate current 9-year block start
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        # Current year display
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        
        layout.addLayout(current_layout)
        
        # Update the grid
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        """Calculate which 9-year block a year belongs to"""
        # Formula: ((year - 1) // 9) * 9 + 1
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        """Update the 3x3 grid with years from current 9-year block"""
        # Generate 9 consecutive years starting from current_block_start
        years = []
        
        for i in range(9):
            year = self.current_block_start + i
            years.append(year)
        
        # Update block label
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        
        # Update button texts and styles
        current_year = datetime.now().year
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            
            # Check if year is within valid range (1-9999)
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            # Style based on selection and current year
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #2980b9;
                    }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            # Connect button click
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        """Select a year"""
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        """Go to previous 9-year block (unlimited past)"""
        self.current_block_start -= 9
        
        # Unlimited past - no lower bound check
        # If we go below year 1, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        """Go to next 9-year block (unlimited future)"""
        self.current_block_start += 9
        
        # Unlimited future - no upper bound check
        # If we go above year 9999, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        """Set the selected year"""
        # Ensure year is within valid range
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        
    def get_selected_year(self):
        """Get the selected year"""
        return self.selected_year
    
    # Add signal for year selection
    year_selected = QtCore.pyqtSignal(int)

class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection with unlimited years"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        # Create YearCalendarGrid with unlimited years
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        # Selected year display
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        
        self.selected_label = QtWidgets.QLabel(f"")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        
        layout.addLayout(selected_layout)
        
        # Action buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Connect signals
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        """Update selected year display when year is changed in calendar"""
        self.current_year = year
    
    def on_select_clicked(self):
        """Emit signal with selected year and close popup"""
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        """Get the selected year"""
        return self.current_year

          
class PDFWorker(QtCore.QThread):
    finished = QtCore.pyqtSignal(bool, str)  # success flag + message

    def __init__(self, invoice, output_path):
        super().__init__()
        self.invoice = invoice
        self.output_path = output_path

    def run(self):
        try:
            success = PDFGenerator.generate(self.invoice, self.output_path)
            self.finished.emit(success, "PDF generated successfully!")
        except Exception as e:
            self.finished.emit(False, str(e))

# MAIN WINDOW CLASS - With Role-Based Access Control
# ====================================================================

class MainWindow(QtWidgets.QMainWindow):
    logout_requested = QtCore.pyqtSignal()
    window_ready = QtCore.pyqtSignal()  # emitted once after first show

    def showEvent(self, event):
        super().showEvent(event)
        if not self._initialization_complete:
            self._initialization_complete = True
            QtCore.QTimer.singleShot(0, self.window_ready.emit)
            # Pre-load invoice cache after window is shown
            QtCore.QTimer.singleShot(500, self._preload_invoice_cache)

    def __init__(self, username="", role="", parent=None):
        super().__init__(parent)
        self.current_username = username
        self.current_role = normalize_role(role)
        self._allowed_pages = set(allowed_pages_for_role(self.current_role))
        self._initialization_complete = False  # Add this flag
        self._logging_out = False
        self.invoice = Invoice()
        self.clients = {}
        self.settings = {}
        self.item_rows = []
        self.FIREBASE_AVAILABLE = FIREBASE_AVAILABLE
        self.db = db
        self.init_ui()
        self.load_data()
        self.apply_role_based_access()
        

    def apply_role_based_access(self):
        """
        Role-based access control:
        - admin: all workspace pages and Settings
        - sales: only Quote Forms
        - projects: only Projects & Invoice
        - finance: only Financial Management
        """
        visible_pages = allowed_stack_pages_for_role(self.current_role)
        self._allowed_pages = set(allowed_pages_for_role(self.current_role))

        projects_allowed = can_access_page(self.current_role, PAGE_PROJECTS)
        finance_allowed = can_access_page(self.current_role, PAGE_FINANCIAL)

        if hasattr(self, 'project_invoice_inner_tabs'):
            for i in range(self.project_invoice_inner_tabs.count()):
                self.project_invoice_inner_tabs.setTabEnabled(i, projects_allowed)
        if hasattr(self, 'finance_inner_tabs'):
            for i in range(self.finance_inner_tabs.count()):
                self.finance_inner_tabs.setTabEnabled(i, finance_allowed)

        self.sidebar.set_visible_pages(allowed_pages_for_role(self.current_role))

        if hasattr(self, 'topbar'):
            self.topbar.set_role(self.current_role)
            if hasattr(self.topbar, 'settings_btn'):
                self.topbar.settings_btn.setVisible(True)
                self.topbar.settings_btn.setEnabled(True)

        # Map 5-button access list to 4-page stack indices
        stack_visible = set()
        for pg in visible_pages:
            stack_visible.add(self._SIDEBAR_TO_STACK.get(pg, pg))
        for i in range(self.stack.count()):
            w = self.stack.widget(i)
            if w:
                w.setVisible(i in stack_visible)

        first_sidebar = first_allowed_stack_page(self.current_role)
        first_stack   = self._SIDEBAR_TO_STACK.get(first_sidebar, first_sidebar)
        self._nav_to(first_stack)

        log.info(f"{self.current_role} user '{self.current_username}' access: {visible_pages}")

    def _disable_unauthorized_tabs(self, widget, allowed_tabs):
        """
        Disable specific tabs within a QTabWidget for non-admin users.
        
        Args:
            widget: The QTabWidget to modify
            allowed_tabs: List of tab indices that should stay enabled
        """
        if not widget or not hasattr(widget, 'count'):
            return
        
        for i in range(widget.count()):
            if i not in allowed_tabs:
                # Disable the tab so it cannot be clicked
                widget.setTabEnabled(i, False)
                
                # Also set the tab style to visually indicate disabled state
                if hasattr(widget, 'tabBar'):
                    tab_bar = widget.tabBar()
                    if tab_bar:
                        tab_bar.setTabEnabled(i, False)
                    
    def on_tab_changed(self, index):
        """Called whenever the active page changes (stack index)."""
        # When navigating TO the Financial page, refresh the balance sheet if it's
        # the active inner tab — currentChanged on finance_inner_tabs won't fire
        # because the inner tab index hasn't changed.
        if index == 3 and hasattr(self, 'finance_inner_tabs') and hasattr(self, 'balance_sheet_tab'):
            if self.finance_inner_tabs.currentIndex() == 2:
                self.balance_sheet_tab.refresh_on_tab_show()

    def init_ui(self):
        self.setWindowTitle(f"{Config.COMPANY.get('name', 'MABS Engineering LLC')} - Project & Invoice Management")
        # Set application icon
        icon_path = Path(__file__).resolve().parent / "assets" / "icons" / "icon.png"
        if icon_path.exists():
            self.setWindowIcon(QtGui.QIcon(str(icon_path)))
        self.setMinimumSize(1024, 640)

        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QHBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # a"a" sidebar a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"
        from sidebar import Sidebar
        co_name = Config.COMPANY.get("name", "MABS Engineering")
        self.sidebar = Sidebar(co_name)
        self.sidebar.page_changed.connect(self._on_nav)
        root.addWidget(self.sidebar)

        # a"a" right side: topbar + content a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"
        right = QtWidgets.QWidget()
        right.setStyleSheet("background:#F6F8FB;")
        right_lay = QtWidgets.QVBoxLayout(right)
        right_lay.setContentsMargins(0, 0, 0, 0)
        right_lay.setSpacing(0)

        # Top bar
        from topbar import TopBar
        self.topbar = TopBar()
        self.topbar.set_company(co_name)
        self.topbar.set_logo(Config.get_logo_path())
        self.topbar.settings_clicked.connect(self.open_settings)
        self.topbar.logout_clicked.connect(self.request_logout)
        self.topbar.refresh_clicked.connect(self.refresh_workspace)
        self.topbar.search_submitted.connect(self.handle_global_search)
        self.topbar.search_text_changed.connect(self._queue_global_search)

        # update indicator a ' plug into topbar
        self.update_indicator = UpdateIndicator(self)
        self.update_indicator.setFixedSize(28, 28)
        self.update_indicator.update_available.connect(self.on_update_available)
        self.update_indicator.hide()

        right_lay.addWidget(self.topbar)

        # Content stack
        self.stack = QtWidgets.QStackedWidget()
        self.stack.setStyleSheet("background:#F6F8FB;")
        right_lay.addWidget(self.stack, 1)
        root.addWidget(right, 1)

        # a"a" build page widgets a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"
        from project_number_generator import ProjectNumberGeneratorTab
        from dashboard_tab import DashboardTab

        from invoice_history_tab import InvoiceHistoryTab

        self.dashboard_tab        = DashboardTab(self, FIREBASE_AVAILABLE)
        self.job_form_tab         = JobFormTab(self, FIREBASE_AVAILABLE)
        self.project_tab          = ProjectNumberGeneratorTab(self)
        self.create_tab           = QtWidgets.QWidget()
        self.client_info_tab      = QtWidgets.QWidget()
        self.client_list_tab      = QtWidgets.QWidget()
        self.history_tab          = InvoiceHistoryTab(self)
        self.finance_overview_tab = FinanceOverviewTab(self)
        self.expenses_tab         = ExpensesTab(self)
        self.balance_sheet_tab    = BalanceSheetTab(self)

        # Projects & Invoice container (no lock overlay)
        pi_widget = QtWidgets.QWidget()
        pi_lay    = QtWidgets.QVBoxLayout(pi_widget)
        pi_lay.setContentsMargins(0, 0, 0, 0)
        self.project_invoice_inner_tabs = QtWidgets.QTabWidget()
        self.project_invoice_inner_tabs.setDocumentMode(False)
        self.project_invoice_inner_tabs.setTabPosition(QtWidgets.QTabWidget.North)
        self.project_invoice_inner_tabs.setUsesScrollButtons(True)
        self.project_invoice_inner_tabs.setStyleSheet("""
            QTabWidget::pane {
                border-top: 1px solid #d8e2ec;
                background: #f4f7fb;
            }
            QTabBar::tab {
                background: #ffffff;
                color: #314155;
                border: 1px solid #d8e2ec;
                border-bottom: 2px solid #d8e2ec;
                min-width: 150px;
                min-height: 42px;
                padding: 0 16px;
                margin-right: 2px;
                font-size: 14px;
                font-weight: 700;
            }
            QTabBar::tab:selected {
                color: #00756f;
                border-bottom: 3px solid #00756f;
                background: #ffffff;
            }
            QTabBar::tab:hover {
                background: #f8fbfd;
                color: #00756f;
            }
        """)
        self.project_invoice_inner_tabs.addTab(self.project_tab,      "Projects")
        self.project_invoice_inner_tabs.addTab(self.create_tab,      "Invoice Management")
        self.project_invoice_inner_tabs.addTab(self.client_info_tab, "Client Information")
        self.project_invoice_inner_tabs.addTab(self.history_tab,     "Invoice History")
        self.project_invoice_inner_tabs.tabBar().setVisible(True)
        pi_lay.addWidget(self.project_invoice_inner_tabs)

        # Financial container (no lock overlay)
        fin_widget = QtWidgets.QWidget()
        fin_lay    = QtWidgets.QVBoxLayout(fin_widget)
        fin_lay.setContentsMargins(0, 0, 0, 0)
        self.finance_inner_tabs = QtWidgets.QTabWidget()
        self.finance_inner_tabs.setDocumentMode(True)
        self.finance_inner_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: #f4f7fb;
                top: -1px;
            }
            QTabBar::tab {
                background: #ffffff;
                color: #314155;
                border: 1px solid #d8e2ec;
                border-bottom: 2px solid #d8e2ec;
                min-width: 170px;
                min-height: 42px;
                padding: 0 18px;
                margin-right: 2px;
                font-size: 14px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
            }
            QTabBar::tab:selected {
                color: #00756f;
                border-bottom: 3px solid #00756f;
                background: #ffffff;
            }
            QTabBar::tab:hover {
                background: #f8fbfd;
                color: #00756f;
            }
        """)
        self.finance_inner_tabs.addTab(self.finance_overview_tab, "Overview")
        self.finance_inner_tabs.addTab(self.expenses_tab,     "Expenses")
        self.finance_inner_tabs.addTab(self.balance_sheet_tab,"Balance Sheet")
        fin_lay.addWidget(self.finance_inner_tabs)

        # Refresh balance sheet whenever the user switches to it
        self.finance_inner_tabs.currentChanged.connect(self._on_finance_tab_changed)

        # Stack: 4 pages. Projects & Invoicing share one page, switching via inner tab.
        self.stack.addWidget(self.dashboard_tab)   # 0 - Dashboard
        self.stack.addWidget(self.job_form_tab)    # 1 - Quote Forms
        self.stack.addWidget(pi_widget)            # 2 - Projects / Invoicing (inner tabs)
        self.stack.addWidget(fin_widget)           # 3 - Financial

        # Dashboard quick-action signals
        self.dashboard_tab.open_quotes.connect(self.open_dashboard_new_quote)
        self.dashboard_tab.open_invoices.connect(self.open_dashboard_new_invoice)
        self.dashboard_tab.open_overdue_invoices.connect(self.open_dashboard_invoice_history)
        self.dashboard_tab.open_projects.connect(self.open_dashboard_projects)
        self.dashboard_tab.open_expenses.connect(self.open_dashboard_expenses)
        self.dashboard_tab.open_project_record.connect(self.open_dashboard_project_record)
        self.dashboard_tab.open_invoice_record.connect(self.open_dashboard_invoice_record)

        # Quote → Invoice / Project conversion
        self.job_form_tab.convert_to_invoice.connect(self.prefill_invoice_from_quote)
        self.job_form_tab.convert_to_project.connect(self.create_project_from_quote)

        # Page titles shown in top bar
        self._page_titles = [
            "Dashboard", "Quote Forms",
            "Project & Invoice", "Financial Management"]

        self._nav_to(first_allowed_stack_page(self.current_role))
        self._setup_automation()
        self.setup_create_tab()

        if hasattr(self, 'tax_spin'):
            self.tax_spin.valueChanged.connect(self.update_totals)
        if hasattr(self, 'date_edit'):
            self.date_edit.dateChanged.connect(self.update_invoice_preview)
        if hasattr(self, 'due_date_edit'):
            self.due_date_edit.dateChanged.connect(self._on_due_date_changed)
        self.update_totals()
        self.update_invoice_preview()

    # a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"a"
    def _nav_to(self, index: int):
        """Switch page and sync sidebar + topbar title."""
        if not can_access_page(self.current_role, index):
            log.warning(
                "Blocked navigation for %s user '%s' to page index %s",
                self.current_role,
                self.current_username,
                index,
            )
            if hasattr(self, 'topbar'):
                self.topbar.set_status("Access restricted", "warning")
            return False

        if index == PAGE_SETTINGS:
            self.open_settings()
            return True

        if index < 0 or index >= self.stack.count():
            log.warning("Blocked navigation to unknown page index %s", index)
            return False

        self.stack.setCurrentIndex(index)
        self.sidebar.select(index)
        if hasattr(self, '_page_titles') and index < len(self._page_titles):
            self.topbar.set_title(self._page_titles[index])
        # fire legacy on_tab_changed logic
        self.on_tab_changed(index)
        return True

    # Sidebar 4 buttons → stack 4 pages (1-to-1)
    _SIDEBAR_TO_STACK = {0: 0, 1: 1, 2: 2, 3: 3}
    _SIDEBAR_TO_INNER = {}

    def _on_nav(self, sidebar_idx: int):
        stack_idx = self._SIDEBAR_TO_STACK.get(sidebar_idx, sidebar_idx)
        self._nav_to(stack_idx)

    def _switch_inner_tab(self, tab_index: int):
        """Switch the inner tab in the Projects/Invoicing page."""
        if hasattr(self, 'project_invoice_inner_tabs'):
            self.project_invoice_inner_tabs.setCurrentIndex(tab_index)

    def open_dashboard_new_quote(self):
        if self._nav_to(1) and hasattr(self.job_form_tab, "workflow_tabs"):
            self.job_form_tab.workflow_tabs.setCurrentIndex(1)

    def open_dashboard_new_invoice(self):
        if self._nav_to(2) and hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(1)

    def open_dashboard_projects(self):
        if self._nav_to(2) and hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(0)

    def open_dashboard_invoice_history(self):
        if self._nav_to(2) and hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(3)

    def _on_finance_tab_changed(self, index: int):
        """Refresh balance sheet data whenever the user switches to the Balance Sheet tab."""
        if index == 2 and hasattr(self, "balance_sheet_tab"):
            self.balance_sheet_tab.refresh_on_tab_show()

    def open_dashboard_expenses(self):
        if self._nav_to(3) and hasattr(self, "finance_inner_tabs"):
            self.finance_inner_tabs.setCurrentIndex(1)

    def open_dashboard_project_record(self, project_data):
        if not isinstance(project_data, dict):
            return
        project_number = project_data.get("project_number", "")
        if not self._nav_to(2):
            return
        if hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(0)
        project_tab = getattr(self, "project_tab", None)
        if not project_tab:
            return

        projects = (
            list(getattr(project_tab, "generated_projects", None) or [])
            or list(getattr(project_tab, "cached_projects", None) or [])
        )
        match = next(
            (p for p in projects if p.get("project_number") == project_number),
            project_data,
        )
        if hasattr(project_tab, "search_edit"):
            project_tab.search_edit.setText(project_number)
        if hasattr(project_tab, "filter_projects"):
            project_tab.filter_projects()
        if hasattr(project_tab, "show_project_workspace"):
            QtCore.QTimer.singleShot(0, lambda p=match: project_tab.show_project_workspace(p))
        if hasattr(self, "topbar") and project_number:
            self.topbar.set_status(f"Opened {project_number}", "success")

    def open_dashboard_invoice_record(self, invoice_data):
        if not isinstance(invoice_data, dict):
            return
        meta = invoice_data.get("meta", invoice_data)
        client_name = meta.get("client_name", "")
        invoice_number = meta.get("invoice_number", "")
        if not client_name or not self._nav_to(2):
            return
        if hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(3)
        if hasattr(self, "history_tab"):
            self.history_tab.navigate_to_invoice(client_name, invoice_number)
        if hasattr(self, "topbar") and invoice_number:
            self.topbar.set_status(f"Invoice {invoice_number}", "success")

    def _open_search_result(self, kind: str, tab_idx: int, popup=None):
        if not self._nav_to(tab_idx):
            return
        if tab_idx == 2 and hasattr(self, "project_invoice_inner_tabs"):
            if kind == "Invoice":
                self.project_invoice_inner_tabs.setCurrentIndex(1)
            elif kind == "Project":
                self.project_invoice_inner_tabs.setCurrentIndex(0)
        if popup is not None:
            popup.close()

    def _normalize_global_search(self, value):
        return "".join(ch for ch in str(value or "").lower() if ch.isalnum())

    def _queue_global_search(self, query):
        self._pending_global_search = (query or "").strip()
        if not hasattr(self, "_search_debounce"):
            return
        if len(self._pending_global_search) >= 2:
            self._search_debounce.start(280)
        else:
            self._search_debounce.stop()
            self._hide_global_search_menu()
            if hasattr(self, "topbar") and hasattr(self.topbar, "clear_status"):
                self.topbar.clear_status()

    def _run_pending_global_search(self):
        query = getattr(self, "_pending_global_search", "")
        if len((query or "").strip()) >= 2:
            self._show_global_search_menu(query)

    def handle_global_search(self, query):
        """Open the best result from the top-bar advanced search."""
        query = (query or "").strip()
        if not query:
            return

        try:
            results = self._collect_global_search_results(query, limit=1)
            if results:
                self._hide_global_search_menu()
                self._open_global_search_result(results[0])
                return
            if hasattr(self, "topbar"):
                self.topbar.set_status(f"No match found for {query}", "warning")
        except Exception as exc:
            log.warning("Global search failed for %s: %s", query, exc)
            if hasattr(self, "topbar"):
                self.topbar.set_status("Search failed", "warning")

    def _hide_global_search_menu(self):
        panel = getattr(self, "_global_search_menu", None)
        if panel is not None:
            try:
                panel.hide()
            except RuntimeError:
                pass
            self._global_search_menu = None
        # Disconnect focus watcher
        conn = getattr(self, '_search_focus_conn', None)
        if conn is not None:
            try:
                QtWidgets.QApplication.instance().focusChanged.disconnect(conn)
            except Exception:
                pass
            self._search_focus_conn = None

    def _result_matches_query(self, query, fields):
        q = (query or "").strip().lower()
        q_norm = self._normalize_global_search(q)
        haystack = " ".join(str(field or "") for field in fields).lower()
        haystack_norm = self._normalize_global_search(haystack)
        return q in haystack or (q_norm and q_norm in haystack_norm)

    def _collect_global_search_results(self, query, limit=8):
        results = []

        try:
            quotes = list(getattr(getattr(self, "job_form_tab", None), "job_forms", None) or [])
            if not quotes:
                quotes = FirebaseManager.load_job_forms() or []
            for quote in quotes:
                if not isinstance(quote, dict):
                    continue
                fields = [
                    quote.get("job_number", ""),
                    quote.get("client", ""),
                    quote.get("project_name", ""),
                    quote.get("job_title", ""),
                    quote.get("status", ""),
                    quote.get("sales", ""),
                ]
                if self._result_matches_query(query, fields):
                    results.append({
                        "kind": "Quote",
                        "title": quote.get("job_number", "Quote"),
                        "subtitle": " | ".join(part for part in [
                            quote.get("client", ""),
                            quote.get("project_name", "") or quote.get("job_title", ""),
                            quote.get("status", ""),
                        ] if part),
                        "data": quote,
                    })
                    if len(results) >= limit:
                        return results
        except Exception as exc:
            log.warning("Quote advanced search failed: %s", exc)

        try:
            # Use already-loaded in-memory invoices — avoids a synchronous Firebase call
            history_tab = getattr(self, 'history_tab', None)
            _raw_invoices = []
            if history_tab:
                _src = (getattr(history_tab, '_ih_all_displayed', None)
                        or getattr(history_tab, 'invoices', []) or [])
                for _item in _src:
                    if isinstance(_item, tuple) and _item:
                        inv_obj = _item[0]
                        _raw_invoices.append(inv_obj.to_dict() if hasattr(inv_obj, 'to_dict') else {})
                    elif isinstance(_item, dict):
                        _raw_invoices.append(_item)
            if not _raw_invoices:
                _raw_invoices = FirebaseManager.load_invoices() or []
            for invoice in _raw_invoices:
                if not isinstance(invoice, dict):
                    continue
                meta = invoice.get("meta", invoice)
                fields = [
                    meta.get("invoice_number", ""),
                    meta.get("client_name", ""),
                    meta.get("status", ""),
                    meta.get("date", ""),
                    str(meta.get("total", "")),
                ]
                if self._result_matches_query(query, fields):
                    results.append({
                        "kind": "Invoice",
                        "title": meta.get("invoice_number", "Invoice"),
                        "subtitle": " | ".join(part for part in [
                            meta.get("client_name", ""),
                            meta.get("status", ""),
                            str(meta.get("total", "")),
                        ] if part),
                        "data": invoice,
                    })
                    if len(results) >= limit:
                        return results
        except Exception as exc:
            log.warning("Invoice advanced search failed: %s", exc)

        try:
            projects = list(getattr(getattr(self, "project_tab", None), "generated_projects", None) or [])
            if not projects:
                projects = FirebaseManager.load_projects() or []
            for project in projects:
                if not isinstance(project, dict):
                    continue
                fields = [
                    project.get("project_number", ""),
                    project.get("project_name", ""),
                    project.get("company", ""),
                    project.get("status", ""),
                    project.get("payment_category", ""),
                ]
                if self._result_matches_query(query, fields):
                    results.append({
                        "kind": "Project",
                        "title": project.get("project_number", "Project"),
                        "subtitle": " | ".join(part for part in [
                            project.get("company", ""),
                            project.get("project_name", ""),
                            project.get("status", ""),
                        ] if part),
                        "data": project,
                    })
                    if len(results) >= limit:
                        return results
        except Exception as exc:
            log.warning("Project advanced search failed: %s", exc)

        return results

    def _show_global_search_menu(self, query):
        results = self._collect_global_search_results(query)
        search_widget = self.topbar.get_search_widget() if hasattr(self, "topbar") else None
        if not search_widget:
            return

        self._hide_global_search_menu()

        # ── frameless panel — WA_ShowWithoutActivating keeps focus on search bar ──
        panel = QtWidgets.QFrame(None)
        panel.setWindowFlags(QtCore.Qt.Tool | QtCore.Qt.FramelessWindowHint)
        panel.setAttribute(QtCore.Qt.WA_ShowWithoutActivating, True)
        panel.setAttribute(QtCore.Qt.WA_DeleteOnClose, False)
        sw_parent = search_widget.parent()
        panel.setFixedWidth(sw_parent.width() if sw_parent else 460)
        panel.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 12px;
            }
        """)
        # Drop shadow
        shadow = QtWidgets.QGraphicsDropShadowEffect(panel)
        shadow.setBlurRadius(20)
        shadow.setOffset(0, 6)
        shadow.setColor(QtGui.QColor(0, 0, 0, 40))
        panel.setGraphicsEffect(shadow)

        outer = QtWidgets.QVBoxLayout(panel)
        outer.setContentsMargins(8, 8, 8, 8)
        outer.setSpacing(1)

        _KIND_ICONS = {"Quote": "💼", "Invoice": "📄", "Project": "⚙"}
        _LABEL_CSS = ("color:#94a3b8;font-size:11px;font-weight:700;padding:6px 10px 2px;"
                      "background:transparent;border:none;font-family:'Inter','Segoe UI';")
        _SEP_CSS   = "background:#e2e8f0;border:none;"
        _BTN_CSS   = """
            QPushButton {
                text-align:left; padding:0; border:none; border-radius:8px;
                background:transparent; font-family:'Inter','Segoe UI';
            }
            QPushButton:hover { background:#ecfdf5; }
        """

        if not results:
            lbl = QtWidgets.QLabel(f"  No results for \"{query}\"")
            lbl.setStyleSheet("color:#94a3b8;font-size:13px;padding:10px 8px;"
                              "background:transparent;border:none;font-family:'Inter','Segoe UI';")
            lbl.setFocusPolicy(QtCore.Qt.NoFocus)
            outer.addWidget(lbl)
        else:
            last_kind = None
            for result in results:
                if result["kind"] != last_kind:
                    if last_kind is not None:
                        sep = QtWidgets.QFrame()
                        sep.setFrameShape(QtWidgets.QFrame.HLine)
                        sep.setFixedHeight(1)
                        sep.setStyleSheet(_SEP_CSS)
                        sep.setFocusPolicy(QtCore.Qt.NoFocus)
                        outer.addWidget(sep)
                    cat = QtWidgets.QLabel(
                        f"  {_KIND_ICONS.get(result['kind'],'')}  {result['kind']}s")
                    cat.setStyleSheet(_LABEL_CSS)
                    cat.setFocusPolicy(QtCore.Qt.NoFocus)
                    outer.addWidget(cat)
                    last_kind = result["kind"]

                # Each result: full-width button with title + subtitle rows
                btn_wrap = QtWidgets.QFrame()
                btn_wrap.setStyleSheet(
                    "QFrame{background:transparent;border:none;border-radius:8px;}"
                    "QFrame:hover{background:#ecfdf5;}")
                btn_wrap.setFocusPolicy(QtCore.Qt.NoFocus)
                btn_wrap.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                blay = QtWidgets.QVBoxLayout(btn_wrap)
                blay.setContentsMargins(12, 7, 12, 7)
                blay.setSpacing(2)

                t_lbl = QtWidgets.QLabel(result["title"])
                t_lbl.setStyleSheet("font-size:13px;font-weight:700;color:#0f172a;"
                                    "background:transparent;border:none;"
                                    "font-family:'Inter','Segoe UI';")
                t_lbl.setFocusPolicy(QtCore.Qt.NoFocus)
                blay.addWidget(t_lbl)

                if result.get("subtitle"):
                    s_lbl = QtWidgets.QLabel(result["subtitle"])
                    s_lbl.setStyleSheet("font-size:11px;color:#64748b;"
                                        "background:transparent;border:none;"
                                        "font-family:'Inter','Segoe UI';")
                    s_lbl.setFocusPolicy(QtCore.Qt.NoFocus)
                    blay.addWidget(s_lbl)

                def _click(ev, r=result, p=panel):
                    if ev.button() == QtCore.Qt.LeftButton:
                        p.hide()
                        QtCore.QTimer.singleShot(0, lambda _r=r: self._open_global_search_result(_r))
                btn_wrap.mousePressEvent = _click
                outer.addWidget(btn_wrap)

        # ── position below the search wrap ──────────────────────────────────
        pos = (sw_parent.mapToGlobal(QtCore.QPoint(0, sw_parent.height() + 4))
               if sw_parent else
               search_widget.mapToGlobal(QtCore.QPoint(0, search_widget.height() + 4)))
        panel.adjustSize()
        panel.move(pos)
        panel.show()
        self._global_search_menu = panel

        # Close panel when search bar loses focus (but not when clicking results)
        def _on_focus_change(old_w, new_w, _p=panel, _sw=search_widget):
            if new_w is None:
                return
            if new_w is _sw:
                return
            try:
                is_inside = any(new_w is child or _p.isAncestorOf(new_w)
                                for child in _p.findChildren(QtWidgets.QWidget))
            except RuntimeError:
                is_inside = False
            if not is_inside:
                self._hide_global_search_menu()

        self._search_focus_conn = (
            QtWidgets.QApplication.instance().focusChanged.connect(_on_focus_change))

    def _open_global_search_result(self, result):
        kind = result.get("kind")
        data = result.get("data") or {}
        self._hide_global_search_menu()
        if kind == "Quote":
            self._open_quote_search_result(data)
        elif kind == "Invoice":
            self._open_invoice_search_result(data)
        elif kind == "Project":
            self._open_project_search_result(data)

    def _open_invoice_search_result(self, invoice_data):
        meta = invoice_data.get("meta", invoice_data) if isinstance(invoice_data, dict) else {}
        invoice_number = meta.get("invoice_number", "")
        client_name = meta.get("client_name", "")
        if client_name and self._nav_to(2):
            if hasattr(self, "project_invoice_inner_tabs"):
                self.project_invoice_inner_tabs.setCurrentIndex(3)
            if hasattr(self, "history_tab"):
                self.history_tab.show_invoice_history(client_name)
                current = self.history_tab.stacked_widget.currentWidget()
                if hasattr(current, "date_range_widget"):
                    current.date_range_widget.search_bar.clear()
            if hasattr(self, "topbar"):
                self.topbar.set_status(f"Showing {client_name} invoices", "success")

    def _open_project_search_result(self, project):
        project_number = project.get("project_number", "")
        if not self._nav_to(2):
            return
        if hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(0)
        tab = getattr(self, "project_tab", None)
        if tab:
            if hasattr(tab, "search_edit"):
                tab.search_edit.setText(project_number)
            if hasattr(tab, "filter_projects"):
                tab.filter_projects()
        if hasattr(self, "topbar"):
            self.topbar.set_status(f"Found {project_number}", "success")

    def _find_quote_by_query(self, query, query_norm):
        try:
            quotes = []
            if hasattr(self, "job_form_tab"):
                quotes = list(getattr(self.job_form_tab, "job_forms", None) or [])
            if not quotes:
                quotes = FirebaseManager.load_job_forms() or []

            for quote in quotes:
                if not isinstance(quote, dict):
                    continue
                quote_number = quote.get("job_number", "")
                quote_norm = self._normalize_global_search(quote_number)
                if not quote_norm:
                    continue
                if query_norm == quote_norm or query.lower() in str(quote_number).lower():
                    return quote
        except Exception as exc:
            log.warning("Quote search failed for %s: %s", query, exc)
        return None

    def _open_quote_search_result(self, quote):
        quote_number = quote.get("job_number", "")
        client_name = quote.get("client", "")
        if not self._nav_to(1):
            return
        tab = getattr(self, "job_form_tab", None)
        if tab:
            if hasattr(tab, "workflow_tabs"):
                tab.workflow_tabs.setCurrentIndex(0)
            if hasattr(tab, "status_filter_combo"):
                tab.status_filter_combo.setCurrentText("All Status")
            if hasattr(tab, "year_filter_combo"):
                tab.year_filter_combo.setCurrentIndex(0)
            if hasattr(tab, "month_filter_combo"):
                tab.month_filter_combo.setCurrentIndex(0)
            if hasattr(tab, "date_filter_active_chk"):
                tab.date_filter_active_chk.setChecked(False)
            if hasattr(tab, "search_edit"):
                tab.search_edit.clear()
            if client_name and hasattr(tab, "apply_client_filter"):
                tab.apply_client_filter(client_name)
            elif hasattr(tab, "filter_job_forms"):
                if hasattr(tab, "selected_client_filter"):
                    tab.selected_client_filter = "All Clients"
                tab.filter_job_forms()
        if hasattr(self, "topbar"):
            if client_name:
                self.topbar.set_status(f"Showing {client_name} quotes", "success")
            else:
                self.topbar.set_status(f"Found {quote_number}", "success")
        
    def _setup_automation(self):
        """Wire up workspace automation: shortcuts, debounced search, and auto-refresh."""
        self._search_debounce = QtCore.QTimer(self)
        self._search_debounce.setSingleShot(True)
        self._search_debounce.timeout.connect(self._run_pending_global_search)

        # Escape key on the search bar closes the dropdown without losing focus
        class _SearchKeyFilter(QtCore.QObject):
            def __init__(self, main_win):
                super().__init__(main_win)
                self._mw = main_win
            def eventFilter(self, obj, event):
                if event.type() == QtCore.QEvent.KeyPress:
                    if event.key() == QtCore.Qt.Key_Escape:
                        self._mw._hide_global_search_menu()
                        return True
                return False
        if hasattr(self, "topbar"):
            self._search_key_filter = _SearchKeyFilter(self)
            self.topbar.search_edit.installEventFilter(self._search_key_filter)

        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+R"), self, activated=self.refresh_workspace)
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+N"), self, activated=lambda: self._nav_to(1))
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+K"), self, activated=self._focus_global_search)


    def _focus_global_search(self):
        if hasattr(self, "topbar"):
            search = self.topbar.get_search_widget()
            search.setFocus()
            search.selectAll()

    def refresh_workspace(self, auto=False):
        """Refresh visible workspace data and dashboard summaries."""
        try:
            if hasattr(self, 'topbar'):
                self.topbar.set_status("Syncing", "busy")
                QtWidgets.QApplication.processEvents()
            if hasattr(self, 'dashboard_tab'):
                self.dashboard_tab.refresh(force_firebase=True)
            if hasattr(self, 'project_tab') and hasattr(self.project_tab, 'load_projects'):
                self.project_tab.load_projects()
            if hasattr(self, 'history_tab') and hasattr(self.history_tab, 'refresh_invoices_immediately'):
                self.history_tab.refresh_invoices_immediately()
            if hasattr(self, 'expenses_tab') and hasattr(self.expenses_tab, 'refresh_data'):
                self.expenses_tab.refresh_data(auto=auto)
            if hasattr(self, 'balance_sheet_tab'):
                if hasattr(self.balance_sheet_tab, 'refresh_data'):
                    self.balance_sheet_tab.refresh_data(auto=auto)
                elif hasattr(self.balance_sheet_tab, 'update_stats_cards'):
                    self.balance_sheet_tab.update_stats_cards()
            if hasattr(self, 'finance_overview_tab') and hasattr(self.finance_overview_tab, 'refresh_data'):
                self.finance_overview_tab.refresh_data(auto=auto)
            label = "Auto-synced" if auto else "Synced"
            if hasattr(self, 'topbar'):
                self.topbar.set_status(datetime.now().strftime(f"{label} %I:%M %p"), "success")
        except Exception as exc:
            log.warning("Workspace refresh failed: %s", exc)
            if hasattr(self, 'topbar'):
                self.topbar.set_status("Refresh failed", "warning")

    def open_settings(self):
        """Admin → full Settings dialog. All other roles → Software Updates only."""
        if self.current_role == "admin":
            from settings_dialog import SettingsDialog
            dlg = SettingsDialog(self, role=self.current_role)
            dlg.settingsSaved.connect(self._on_settings_saved)
            dlg.exec_()
        else:
            from settings_dialog import SoftwareUpdatesDialog
            SoftwareUpdatesDialog(self).exec_()

    def _apply_company_to_ui(self):
        """Push current Config.COMPANY values to every UI element that shows them."""
        name = Config.COMPANY.get("name", "MABS Engineering LLC")
        logo = Config.get_logo_path()
        self.setWindowTitle(f"{name} - Project & Invoice Management")
        if hasattr(self, 'sidebar'):
            self.sidebar.update_company(name)
        if hasattr(self, 'topbar'):
            self.topbar.set_company(name)
            self.topbar.set_logo(logo)
        if hasattr(self, 'tax_spin'):
            self.tax_spin.setValue(float(Config.COMPANY.get("default_tax_rate", 0.0)))
        if hasattr(self, 'notes_edit'):
            self.notes_edit.setPlainText(Config.DEFAULT_TERMS)

    def _on_settings_saved(self):
        """Reload config after settings are saved so changes take effect immediately."""
        Config.load()
        # Do NOT call overlay_from_firebase() here — Firebase save is still in flight
        # so it still holds the old values and would overwrite the freshly saved config.
        self._apply_company_to_ui()
        if hasattr(self, 'dashboard_tab'):
            self.dashboard_tab.refresh(force_firebase=True)
        if hasattr(self, 'topbar'):
            self.topbar.set_status("Settings saved", "success")
        log.info("Settings reloaded after save")

    def request_logout(self):
        """Log out the current user and return to the login window."""
        reply = QtWidgets.QMessageBox.question(
            self,
            "Logout",
            "Log out and return to the login screen?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return

        log.info("User %s logged out", self.current_username)
        self._logging_out = True
        self.logout_requested.emit()
        self.close()

    def on_projects_loaded(self, projects):
        """Called when projects are loaded in project tab"""
        if hasattr(self, 'project_tab'):
            self.project_tab.cached_projects = projects.copy()
            log.info("  Main window received %s projects for auto-fill", len(projects))
    
    def on_client_changed(self, client_name):
        """When client is selected from dropdown, load its details"""
        if hasattr(self, "line_items_client_combo") and self.sender() is self.line_items_client_combo:
            if hasattr(self, "client_combo"):
                self.client_combo.blockSignals(True)
                self.client_combo.setEditText(client_name)
                self.client_combo.blockSignals(False)
        elif hasattr(self, "line_items_client_combo") and self.sender() is self.client_combo:
            self.line_items_client_combo.blockSignals(True)
            self.line_items_client_combo.setEditText(client_name)
            self.line_items_client_combo.blockSignals(False)
        if client_name.strip():
            self.load_client_details(client_name)
            if hasattr(self, "save_client_btn"):
                self.save_client_btn.setEnabled(False)
        else:
            self.update_invoice_client_summary("", "", "")

    def _find_client_record(self, client_name):
        """Return the saved client key and data for a user-entered client name."""
        target = str(client_name or "").strip()
        if not target:
            return "", None
        if target in self.clients:
            return target, self.clients[target]

        target_lower = target.lower()
        for saved_name, client_data in self.clients.items():
            if str(saved_name).strip().lower() == target_lower:
                return saved_name, client_data
        return "", None

    def _sync_client_combos(self, client_name):
        for combo_name in ("client_combo", "line_items_client_combo"):
            combo = getattr(self, combo_name, None)
            if combo is None or combo.currentText() == client_name:
                continue
            combo.blockSignals(True)
            combo.setEditText(client_name)
            combo.blockSignals(False)

    def update_invoice_client_summary(self, client_name="", email="", address=""):
        if not hasattr(self, "invoice_client_name_label"):
            return
        clean_address = (address or "").replace("\n", ", ").strip()
        self.invoice_client_name_label.setText(f"Client: {client_name or 'Not selected'}")
        self.invoice_client_email_label.setText(f"Email: {email or '-'}")
        self.invoice_client_address_label.setText(f"Address: {clean_address or '-'}")

    def show_client_information_tab(self):
        """Focus the inline client management tab."""
        if not self._nav_to(PAGE_PROJECTS):
            return
        if hasattr(self, "project_invoice_inner_tabs") and hasattr(self, "client_info_tab"):
            index = self.project_invoice_inner_tabs.indexOf(self.client_info_tab)
            if index >= 0:
                self.project_invoice_inner_tabs.setCurrentIndex(index)
        if hasattr(self, "client_combo") and getattr(self, "client_editor_panel", None) and self.client_editor_panel.isVisible():
            self.client_combo.setFocus()

    def show_client_editor_panel(self):
        """Reveal the inline client editor only when the user asks to manage a record."""
        if hasattr(self, "client_editor_panel"):
            self.client_editor_panel.setVisible(True)

    def prepare_new_client_inline(self):
        """Start a new client record without opening a separate window."""
        self.show_client_information_tab()
        self.show_client_editor_panel()
        self.clear_client_information()
        self.editing_client_name = None
        if hasattr(self, "save_client_btn"):
            self.save_client_btn.setEnabled(True)
        if hasattr(self, "client_combo") and self.client_combo.lineEdit():
            self.client_combo.lineEdit().setFocus()

    def open_new_client_popup_dialog(self):
        """Open popup to add a new client (no pre-populated name)."""
        self.open_client_popup_dialog(client_name=None)

    def _switch_and_open_new_client_popup(self):
        """Switch to Client tab, then open an empty New Client popup."""
        self.show_client_information_tab()
        QtCore.QTimer.singleShot(150, lambda: self.open_client_popup_dialog(client_name=None))

    def edit_current_client_popup(self):
        """Switch to Client tab, then open the edit popup for the selected client."""
        client_name = ""
        if hasattr(self, "line_items_client_combo"):
            client_name = self.line_items_client_combo.currentText().strip()
        if not client_name and hasattr(self, "client_combo"):
            client_name = self.client_combo.currentText().strip()
        self.show_client_information_tab()
        QtCore.QTimer.singleShot(
            150,
            lambda cn=client_name: self.open_client_popup_dialog(client_name=cn or None),
        )

    def open_client_popup_dialog(self, client_name=None, navigate_after=False):
        """Open a modal popup dialog to add or edit a client."""
        is_edit = client_name is not None
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Edit Client" if is_edit else "New Client")
        dialog.setModal(True)
        dialog.resize(620, 520)
        dialog.setMinimumWidth(560)
        dialog.setMinimumHeight(460)
        dialog.setStyleSheet("QDialog { background: #f5f8fb; }")

        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        header = QtWidgets.QFrame()
        header.setStyleSheet("QFrame { background: #ffffff; border-bottom: 1px solid #d8e2ec; }")
        header_lay = QtWidgets.QVBoxLayout(header)
        header_lay.setContentsMargins(24, 16, 24, 14)
        header_lay.setSpacing(3)
        title_lbl = QtWidgets.QLabel("Edit Client" if is_edit else "New Client")
        title_lbl.setStyleSheet("font-size: 20px; font-weight: 900; color: #0f172a; font-family:'Inter','Segoe UI';")
        sub_lbl = QtWidgets.QLabel(
            f"Editing details for {client_name}." if is_edit else "Fill in the client details below."
        )
        sub_lbl.setStyleSheet("font-size: 12px; font-weight: 600; color: #53657d; font-family:'Inter','Segoe UI';")
        header_lay.addWidget(title_lbl)
        header_lay.addWidget(sub_lbl)
        layout.addWidget(header)

        # Scrollable form area
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea { border: none; background: #f5f8fb; }")
        form_widget = QtWidgets.QWidget()
        form_widget.setStyleSheet("background: #f5f8fb;")
        form_lay = QtWidgets.QVBoxLayout(form_widget)
        form_lay.setContentsMargins(24, 20, 24, 20)
        form_lay.setSpacing(12)

        _lbl_style = "font-size: 12px; font-weight: 700; color: #314155; font-family:'Inter','Segoe UI';"
        _field_style = """
            QLineEdit {
                background: #ffffff; border: 1.5px solid #d8e2ec; border-radius: 7px;
                padding: 7px 10px; font-size: 13px; color: #0f172a;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QLineEdit:focus { border-color: #00756f; }
        """

        def _field(placeholder):
            w = QtWidgets.QLineEdit()
            w.setMinimumHeight(38)
            w.setPlaceholderText(placeholder)
            w.setStyleSheet(_field_style)
            return w

        def _pair(lbl1, w1, lbl2, w2):
            row = QtWidgets.QHBoxLayout()
            row.setSpacing(16)
            for lbl_txt, widget in ((lbl1, w1), (lbl2, w2)):
                col = QtWidgets.QVBoxLayout()
                col.setSpacing(4)
                lbl = QtWidgets.QLabel(lbl_txt)
                lbl.setStyleSheet(_lbl_style)
                col.addWidget(lbl)
                col.addWidget(widget)
                row.addLayout(col, 1)
            form_lay.addLayout(row)

        # Row 1 — Client Name (full width)
        lbl_name = QtWidgets.QLabel("Client / Company Name *")
        lbl_name.setStyleSheet(_lbl_style)
        d_client_name = QtWidgets.QLineEdit()
        d_client_name.setMinimumHeight(40)
        d_client_name.setPlaceholderText("Enter client / company name...")
        d_client_name.setStyleSheet(_field_style)
        if is_edit:
            d_client_name.setText(client_name)
        form_lay.addWidget(lbl_name)
        form_lay.addWidget(d_client_name)

        # Row 2 — Contact | Phone
        d_contact = _field("Primary contact person")
        d_phone   = _field("(xxx) xxx-xxxx")
        _pair("Contact Person", d_contact, "Phone", d_phone)

        # Row 3 — Primary Email | Company Email
        d_primary_email = _field("primary@example.com")
        d_company_email = _field("company@example.com")
        _pair("Primary Email", d_primary_email, "Company Email", d_company_email)

        # Row 4 — Address (full width)
        addr_lbl = QtWidgets.QLabel("Address")
        addr_lbl.setStyleSheet(_lbl_style)
        d_address = QtWidgets.QTextEdit()
        d_address.setMinimumHeight(68)
        d_address.setMaximumHeight(80)
        d_address.setPlaceholderText("Street Address, City, State ZIP")
        d_address.setStyleSheet("""
            QTextEdit {
                background: #ffffff; border: 1.5px solid #d8e2ec; border-radius: 7px;
                padding: 7px 10px; font-size: 13px; color: #0f172a;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QTextEdit:focus { border-color: #00756f; }
        """)
        form_lay.addWidget(addr_lbl)
        form_lay.addWidget(d_address)

        # Pre-populate fields when editing
        if is_edit:
            _, client_data = self._find_client_record(client_name)
            if client_data:
                d_contact.setText(client_data.get("contact_person", ""))
                d_phone.setText(client_data.get("phone", ""))
                d_primary_email.setText(client_data.get("email", ""))
                d_company_email.setText(client_data.get("company_email", ""))
                d_address.setPlainText(client_data.get("address", ""))

        scroll.setWidget(form_widget)
        layout.addWidget(scroll, 1)

        # Footer buttons
        footer = QtWidgets.QFrame()
        footer.setStyleSheet("QFrame { background: #ffffff; border-top: 1px solid #d8e2ec; }")
        footer_lay = QtWidgets.QHBoxLayout(footer)
        footer_lay.setContentsMargins(24, 14, 24, 14)
        footer_lay.setSpacing(10)

        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setFixedHeight(40)
        cancel_btn.setMinimumWidth(100)
        cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        cancel_btn.setStyleSheet("""
            QPushButton { background: #f1f5f9; color: #334155; border: 1.5px solid #d8e2ec;
                          border-radius: 8px; font-size: 13px; font-weight: 700;
                          font-family:'Inter','Segoe UI'; padding: 0 16px; }
            QPushButton:hover { background: #e2e8f0; }
        """)
        cancel_btn.clicked.connect(dialog.reject)

        save_btn = QtWidgets.QPushButton("Save Client")
        save_btn.setFixedHeight(40)
        save_btn.setMinimumWidth(130)
        save_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        save_btn.setStyleSheet("""
            QPushButton { background: #00756f; color: #ffffff; border: none;
                          border-radius: 8px; font-size: 13px; font-weight: 900;
                          font-family:'Inter','Segoe UI'; padding: 0 16px; }
            QPushButton:hover { background: #00645f; }
            QPushButton:disabled { background: #cbd5e1; color: #64748b; }
        """)

        # --- Enter / Tab key navigation between fields ---
        nav_fields = [d_client_name, d_contact, d_phone, d_primary_email, d_company_email, d_address]

        class _FieldNav(QtCore.QObject):
            def eventFilter(self_inner, source, event):
                if event.type() == QtCore.QEvent.KeyPress:
                    key = event.key()
                    is_enter = key in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter)
                    is_tab = key == QtCore.Qt.Key_Tab
                    is_back = key == QtCore.Qt.Key_Backtab or (
                        is_tab and bool(event.modifiers() & QtCore.Qt.ShiftModifier)
                    )
                    if is_enter or is_tab:
                        step = -1 if is_back else 1
                        for i, w in enumerate(nav_fields):
                            if source is w or (isinstance(w, QtWidgets.QTextEdit) and source is w):
                                next_i = i + step
                                if 0 <= next_i < len(nav_fields):
                                    nw = nav_fields[next_i]
                                    nw.setFocus()
                                    if isinstance(nw, QtWidgets.QLineEdit):
                                        QtCore.QTimer.singleShot(0, nw.selectAll)
                                    return True
                                elif is_enter and i == len(nav_fields) - 1:
                                    save_btn.click()
                                    return True
                return False

        nav = _FieldNav(dialog)
        for w in nav_fields:
            w.installEventFilter(nav)

        def _do_save():
            new_name = d_client_name.text().strip()
            if not new_name:
                QtWidgets.QMessageBox.warning(dialog, "Save Client", "Client name is required.")
                d_client_name.setFocus()
                return
            # Push values into main window's client form fields then delegate to save_client
            if hasattr(self, "client_combo"):
                if self.client_combo.findText(new_name) == -1:
                    self.client_combo.addItem(new_name)
                self.client_combo.blockSignals(True)
                self.client_combo.setEditText(new_name)
                self.client_combo.blockSignals(False)
            if hasattr(self, "client_contact_edit"):
                self.client_contact_edit.setText(d_contact.text().strip())
            if hasattr(self, "client_phone_edit"):
                self.client_phone_edit.setText(d_phone.text().strip())
            if hasattr(self, "client_primary_email_edit"):
                self.client_primary_email_edit.setText(d_primary_email.text().strip())
            if hasattr(self, "client_email_edit"):
                self.client_email_edit.setText(d_company_email.text().strip())
            if hasattr(self, "client_address_edit"):
                self.client_address_edit.setPlainText(d_address.toPlainText().strip())
            self.editing_client_name = client_name if is_edit else None
            if hasattr(self, "save_client_btn"):
                self.save_client_btn.setEnabled(True)
            self.save_client()
            dialog.accept()

        save_btn.clicked.connect(_do_save)

        footer_lay.addStretch()
        footer_lay.addWidget(cancel_btn)
        footer_lay.addWidget(save_btn)
        layout.addWidget(footer)

        QtCore.QTimer.singleShot(100, d_client_name.setFocus)
        dialog.exec_()

    def load_client_details(self, client_name):
        """Load client details from saved clients"""
        saved_client_name, client_data = self._find_client_record(client_name)
        if client_data:
            self.editing_client_name = saved_client_name
            self._sync_client_combos(saved_client_name)
            company_email = client_data.get("company_email", "")
            regular_email = client_data.get("email", "")

            if hasattr(self, "client_contact_edit"):
                self.client_contact_edit.setText(client_data.get("contact_person", ""))
            if hasattr(self, "client_phone_edit"):
                self.client_phone_edit.setText(client_data.get("phone", ""))
            if hasattr(self, "client_primary_email_edit"):
                self.client_primary_email_edit.setText(regular_email)
            
            if company_email:
                self.client_email_edit.setText(company_email)
            else:
                self.client_email_edit.setText(regular_email)
                
            self.client_address_edit.setPlainText(client_data.get("address", ""))
            self.update_invoice_client_summary(
                saved_client_name,
                self.client_email_edit.text().strip(),
                client_data.get("address", ""),
            )
            if hasattr(self, "save_client_btn"):
                self.save_client_btn.setEnabled(False)
        else:
            if getattr(self, "editing_client_name", None):
                self.update_invoice_client_summary(client_name, self.client_email_edit.text().strip(), self.client_address_edit.toPlainText().strip())
                return
            if hasattr(self, "client_contact_edit"):
                self.client_contact_edit.clear()
            if hasattr(self, "client_phone_edit"):
                self.client_phone_edit.clear()
            if hasattr(self, "client_primary_email_edit"):
                self.client_primary_email_edit.clear()
            self.client_email_edit.clear()
            self.client_address_edit.clear()
            self.update_invoice_client_summary(client_name, "", "")
            
    def check_password(self, tab_name):
        password, ok = QtWidgets.QInputDialog.getText(
            self,
            "Password Required",
            f"Enter password to open {tab_name}:",
            QtWidgets.QLineEdit.Password
        )

        if not ok:
            return False

        #   "  Set your password here
        if password == "admin123":
            return True
        else:
            QtWidgets.QMessageBox.warning(self, "Access Denied", "Incorrect Password")
            return False
           
    def setup_create_tab(self):
        # Create scroll area for the create tab
        scroll_area = ScrollArea()
        scroll_content = QtWidgets.QWidget()
        scroll_area.setWidget(scroll_content)
        
        # Main layout for scroll content
        main_layout = QtWidgets.QVBoxLayout(scroll_content)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(20, 14, 20, 18)
        
        
        # Create a grid layout for the main content
        content_layout = QtWidgets.QGridLayout()
        content_layout.setSpacing(14)
        content_layout.setColumnStretch(0, 2)
        content_layout.setColumnStretch(1, 1)

        panel_style = """
            QGroupBox {
                font-weight: 800;
                font-size: 15px;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 12px;
                background: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 14px;
                padding: 0 8px;
                color: #0f172a;
                background: #ffffff;
            }
        """
        field_style = """
            QLineEdit, QComboBox, QDateEdit, QDoubleSpinBox, QSpinBox, QTextEdit {
                background: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px 10px;
                font-size: 15px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                color: #0f172a;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QTextEdit:focus {
                border-color: #00756f;
            }
        """
        form_label_style = """
            QLabel {
                background: transparent;
                border: none;
                color: #0f172a;
                font-size: 15px;
                font-weight: 500;
                padding: 0px;
                margin: 0px;
            }
        """
        
        # Invoice Details
        details_group = QtWidgets.QGroupBox("Invoice Details")
        details_group.setMinimumSize(0, 180)
        details_group.setMaximumSize(16777215, 240)
        details_group.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        details_group.setStyleSheet(panel_style)
        details_layout = QtWidgets.QFormLayout(details_group)
        details_layout.setContentsMargins(28, 24, 28, 22)
        details_layout.setVerticalSpacing(14)
        details_layout.setHorizontalSpacing(14)
        details_layout.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)
        
        self.invoice_no_edit = QtWidgets.QLineEdit()
        self.invoice_no_edit.setReadOnly(True)
        self.invoice_no_edit.setFixedHeight(36)
        self.invoice_no_edit.setMaximumWidth(260)
        self.invoice_no_edit.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        self.invoice_no_edit.setStyleSheet(field_style + "QLineEdit { font-weight: 800; background: #f8fbfd; }")
        
        _date_style = """
            QDateEdit {
                background: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px 34px 7px 10px;
                font-size: 15px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                color: #0f172a;
            }
            QDateEdit:focus { border-color: #00756f; }
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 30px;
                border-left: 1px solid #d8e2ec;
                border-top-right-radius: 7px;
                border-bottom-right-radius: 7px;
                background: #f1f5f9;
            }
            QDateEdit::down-arrow { image: none; }
        """

        self.date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("MMM d, yyyy")
        self.date_edit.setFixedHeight(36)
        self.date_edit.setFixedWidth(200)
        self.date_edit.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        self.date_edit.setStyleSheet(_date_style)
        self.date_edit.wheelEvent = lambda event: event.ignore()
        self.date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        self.date_edit.stepBy = lambda x: None

        self.due_date_edit = QtWidgets.QDateEdit(QtCore.QDate.currentDate().addDays(30))
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("MMM d, yyyy")
        self.due_date_edit.setFixedHeight(36)
        self.due_date_edit.setFixedWidth(200)
        self.due_date_edit.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        self.due_date_edit.setStyleSheet(_date_style)
        self.due_date_edit.wheelEvent = lambda event: event.ignore()
        self.due_date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        self.due_date_edit.stepBy = lambda x: None
        # Keep due date >= invoice date at all times
        self.due_date_edit.setMinimumDate(self.date_edit.date())

        # Gray out disabled (before-invoice) dates in the calendar popup
        _due_cal = self.due_date_edit.calendarWidget()
        if _due_cal:
            _due_cal.setStyleSheet("""
                QCalendarWidget QAbstractItemView {
                    selection-background-color: #00756f;
                    selection-color: white;
                }
                QCalendarWidget QAbstractItemView:disabled {
                    color: #c8c8c8;
                }
            """)

            def _apply_due_cal_grays(cal, min_date):
                # Clear previously set formats then re-gray dates before min_date
                cal.setDateTextFormat(QtCore.QDate(), QtGui.QTextCharFormat())
                gray_fmt = QtGui.QTextCharFormat()
                gray_fmt.setForeground(QtGui.QBrush(QtGui.QColor("#c8c8c8")))
                d = min_date.addYears(-5)
                end = min_date.addDays(-1)
                while d <= end:
                    cal.setDateTextFormat(d, gray_fmt)
                    d = d.addDays(1)

            _apply_due_cal_grays(_due_cal, self.date_edit.date())

            self.date_edit.dateChanged.connect(
                lambda d, cal=_due_cal: (
                    self.due_date_edit.setMinimumDate(d),
                    _apply_due_cal_grays(cal, d)
                )
            )
        else:
            self.date_edit.dateChanged.connect(
                lambda d: self.due_date_edit.setMinimumDate(d)
            )

        invoice_number_label = QtWidgets.QLabel("Invoice Number:")
        invoice_date_label = QtWidgets.QLabel("Invoice Date:")
        due_date_label = QtWidgets.QLabel("Due Date (Payment Deadline):")
        for label in (invoice_number_label, invoice_date_label, due_date_label):
            label.setStyleSheet(form_label_style)

        details_layout.addRow(invoice_number_label, self.invoice_no_edit)
        details_layout.addRow(invoice_date_label, self.date_edit)
        details_layout.addRow(due_date_label, self.due_date_edit)
        
        content_layout.addWidget(details_group, 0, 0)
        
        # Client Information - moved to its own tab, but the same widgets are kept
        # so invoice/PDF logic still reads the selected client values.
        client_group = QtWidgets.QGroupBox("Client Information")
        client_group.setMinimumSize(0, 360)
        client_group.setMaximumSize(16777215, 440)
        client_group.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        client_group.setStyleSheet(panel_style)
        client_layout = QtWidgets.QFormLayout(client_group)
        client_layout.setContentsMargins(20, 18, 20, 16)
        client_layout.setVerticalSpacing(10)
        client_layout.setHorizontalSpacing(12)
        client_layout.setFieldGrowthPolicy(QtWidgets.QFormLayout.ExpandingFieldsGrow)

        # ====== Company Name Dropdown with Add Client Info Button ======
        company_name_label = QtWidgets.QLabel("Client:")
        company_name_label.setStyleSheet("font-size: 13px; font-weight: 700; color: #314155; background: transparent; border: none; font-family: 'Inter', 'Segoe UI', sans-serif;")
        company_name_layout = QtWidgets.QHBoxLayout()
        company_name_layout.setSpacing(10)

        self.client_combo = QtWidgets.QComboBox()
        self.client_combo.setEditable(True)
        self.client_combo.setMinimumHeight(36)
        self.client_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px 32px 7px 10px;
                font-size: 15px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                min-height: 20px;
                min-width: 350px;
                color: #0f172a;
            }
            QComboBox:hover  { border-color: #94a3b8; }
            QComboBox:focus  { border-color: #00756f; background-color: #ffffff; }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 28px;
                border-left: 1px solid #d8e2ec;
                border-top-right-radius: 7px;
                border-bottom-right-radius: 7px;
                background: #f1f5f9;
            }
            QComboBox::down-arrow { image: none; }
            QComboBox QAbstractItemView {
                background-color: white;
                border: 1px solid #d8e2ec;
                border-radius: 6px;
                selection-background-color: #f0fdf4;
                selection-color: #00756f;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 14px;
                padding: 4px;
            }
        """)

        # Placeholder
        self.client_combo.lineEdit().setPlaceholderText("Select or type client name...")
        # Prevent scroll wheel and arrow keys from changing client selection
        self.client_combo.wheelEvent = lambda event: event.ignore()
        _cc_orig_key = self.client_combo.keyPressEvent
        def _cc_key(event, _orig=_cc_orig_key):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            _orig(event)
        self.client_combo.keyPressEvent = _cc_key

        # Only this signal is needed
        self.client_combo.currentTextChanged.connect(self.on_client_changed)

        # Keep client management inside this tab instead of opening another window.
        self.add_client_btn = QtWidgets.QPushButton("a * Add Client Info")
        self.add_client_btn.setMinimumHeight(36)
        self.add_client_btn.setFixedWidth(120)
        self.add_client_btn.setText("New Client")
        self.add_client_btn.setStyleSheet("""
            QPushButton {
                background: #334155;
                color: white;
                font-weight: 800;
                padding: 5px 10px;
                border-radius: 7px;
                font-size: 12px;
                border: none;
            }
            QPushButton:hover {
                background: #1f2937;
            }
            QPushButton:pressed {
                background: #111827;
            }
        """)
        self.add_client_btn.clicked.connect(self.prepare_new_client_inline)
        self.add_client_btn.setVisible(True)

        company_name_layout.addWidget(self.client_combo)
        company_name_layout.addSpacing(12)
        company_name_layout.addWidget(self.add_client_btn)

        client_layout.addRow(company_name_label, company_name_layout)
        company_name_label.setVisible(True)
        self.client_combo.setVisible(True)
        self.add_client_btn.setVisible(True)

        _lbl_style = "font-size: 13px; font-weight: 700; color: #314155; background: transparent; border: none; font-family: 'Inter', 'Segoe UI', sans-serif;"
        _field_no_autofill = """
            QLineEdit {
                background-color: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px 10px;
                font-size: 15px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                color: #0f172a;
            }
            QLineEdit:hover  { border-color: #94a3b8; background-color: #ffffff; }
            QLineEdit:focus  { border-color: #00756f; background-color: #ffffff; }
            QLineEdit:!focus { background-color: #ffffff; }
        """

        contact_label = QtWidgets.QLabel("Contact Person:")
        contact_label.setStyleSheet(_lbl_style)
        self.client_contact_edit = QtWidgets.QLineEdit()
        self.client_contact_edit.setMinimumHeight(36)
        self.client_contact_edit.setStyleSheet(_field_no_autofill)
        self.client_contact_edit.setPlaceholderText("Primary contact person")
        client_layout.addRow(contact_label, self.client_contact_edit)

        phone_label = QtWidgets.QLabel("Phone:")
        phone_label.setStyleSheet(_lbl_style)
        self.client_phone_edit = QtWidgets.QLineEdit()
        self.client_phone_edit.setMinimumHeight(36)
        self.client_phone_edit.setStyleSheet(_field_no_autofill)
        self.client_phone_edit.setPlaceholderText("(xxx) xxx-xxxx")
        client_layout.addRow(phone_label, self.client_phone_edit)

        primary_email_label = QtWidgets.QLabel("Primary Email:")
        primary_email_label.setStyleSheet(_lbl_style)
        self.client_primary_email_edit = QtWidgets.QLineEdit()
        self.client_primary_email_edit.setMinimumHeight(36)
        self.client_primary_email_edit.setStyleSheet(_field_no_autofill)
        self.client_primary_email_edit.setPlaceholderText("primary@example.com")
        client_layout.addRow(primary_email_label, self.client_primary_email_edit)

        # ====== Client Email ======
        email_label = QtWidgets.QLabel("Company Email:")
        email_label.setStyleSheet(_lbl_style)

        self.client_email_edit = QtWidgets.QLineEdit()
        self.client_email_edit.setMinimumHeight(36)
        self.client_email_edit.setStyleSheet(_field_no_autofill)
        self.client_email_edit.setPlaceholderText("company@example.com")
        client_layout.addRow(email_label, self.client_email_edit)

        # ====== Mailing Address ======
        address_label = QtWidgets.QLabel("Address:")
        address_label.setStyleSheet(_lbl_style)

        self.client_address_edit = QtWidgets.QTextEdit()
        self.client_address_edit.setMinimumHeight(42)
        self.client_address_edit.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 7px 10px;
                font-size: 17px;
                color: #0f172a;
            }
            QTextEdit:focus {
                border-color: #00756f;
                background-color: #ffffff;
            }
        """)
        self.client_address_edit.setPlaceholderText("Street Address, City, State ZIP")
        client_layout.addRow(address_label, self.client_address_edit)

        client_page_layout = QtWidgets.QVBoxLayout(self.client_info_tab)
        client_page_layout.setContentsMargins(20, 18, 20, 20)
        client_page_layout.setSpacing(14)

        self.setup_client_information_directory(client_page_layout)

        client_actions = QtWidgets.QHBoxLayout()
        client_actions.setSpacing(10)

        self.save_client_btn = QtWidgets.QPushButton("Save Client")
        self.save_client_btn.setMinimumHeight(42)
        self.save_client_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save_client_btn.setStyleSheet("""
            QPushButton {
                background: #00756f;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: 900;
                font-size: 17px;
                padding: 0 18px;
            }
            QPushButton:hover {
                background: #00645f;
            }
            QPushButton:disabled {
                background: #cbd5e1;
                color: #64748b;
            }
        """)
        self.save_client_btn.clicked.connect(self.save_client)

        self.clear_client_btn = QtWidgets.QPushButton("Clear")
        self.clear_client_btn.setMinimumHeight(42)
        self.clear_client_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.clear_client_btn.setStyleSheet("""
            QPushButton {
                background: #334155;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: 800;
                font-size: 13px;
                padding: 0 18px;
            }
            QPushButton:hover {
                background: #1f2937;
            }
        """)
        self.clear_client_btn.clicked.connect(self.clear_client_information)

        client_actions.addWidget(self.save_client_btn)
        client_actions.addWidget(self.clear_client_btn)
        client_actions.addStretch()

        self.client_combo.lineEdit().textEdited.connect(lambda _text: self.save_client_btn.setEnabled(True))
        self.client_combo.currentTextChanged.connect(self._on_client_name_edit_changed)
        self.client_contact_edit.textEdited.connect(lambda _text: self.save_client_btn.setEnabled(True))
        self.client_phone_edit.textEdited.connect(lambda _text: self.save_client_btn.setEnabled(True))
        self.client_primary_email_edit.textEdited.connect(lambda _text: self.save_client_btn.setEnabled(True))
        self.client_email_edit.textEdited.connect(lambda _text: self.save_client_btn.setEnabled(True))
        self.client_address_edit.textChanged.connect(lambda: self.save_client_btn.setEnabled(True))

        self.client_editor_panel = QtWidgets.QWidget()
        self.client_editor_panel.setVisible(False)
        client_editor_layout = QtWidgets.QVBoxLayout(self.client_editor_panel)
        client_editor_layout.setContentsMargins(0, 0, 0, 0)
        client_editor_layout.setSpacing(12)
        client_editor_layout.addWidget(client_group)
        client_editor_layout.addLayout(client_actions)
        client_page_layout.addWidget(self.client_editor_panel)
        client_page_layout.addStretch()

        self.setup_inline_client_list_tab()

        # Items Section - Span both columns
        items_group = QtWidgets.QGroupBox("Line Items")
        items_group.setStyleSheet(panel_style)
        items_layout = QtWidgets.QVBoxLayout(items_group)
        items_layout.setContentsMargins(20, 20, 20, 16)
        items_layout.setSpacing(12)

        client_select_frame = QtWidgets.QFrame()
        client_select_frame.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-radius: 8px;
            }
            QLabel {
                background: transparent;
                border: none;
                color: #0f172a;
                font-size: 14px;
                font-weight: 900;
            }
        """)
        client_select_layout = QtWidgets.QHBoxLayout(client_select_frame)
        client_select_layout.setContentsMargins(16, 12, 16, 12)
        client_select_layout.setSpacing(12)

        select_client_label = QtWidgets.QLabel("Select Client")
        self.line_items_client_combo = QtWidgets.QComboBox()
        self.line_items_client_combo.setEditable(True)
        self.line_items_client_combo.setMinimumHeight(40)
        self.line_items_client_combo.setMinimumWidth(360)
        self.line_items_client_combo.setStyleSheet("""
            QComboBox {
                background: white;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                padding: 7px 10px;
                color: #0f172a;
                font-size: 14px;
                font-weight: 700;
            }
            QComboBox:focus {
                border-color: #00756f;
            }
        """)
        self.line_items_client_combo.lineEdit().setPlaceholderText("Select or type client name before generating invoice...")
        # Prevent scroll wheel and arrow keys from changing client selection
        self.line_items_client_combo.wheelEvent = lambda event: event.ignore()
        _lic_orig_key = self.line_items_client_combo.keyPressEvent
        def _lic_key(event, _orig=_lic_orig_key):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            _orig(event)
        self.line_items_client_combo.keyPressEvent = _lic_key
        self.line_items_client_combo.currentTextChanged.connect(self.on_client_changed)

        new_client_btn = QtWidgets.QPushButton("+ New Client")
        new_client_btn.setFixedHeight(40)
        new_client_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        new_client_btn.setStyleSheet("""
            QPushButton {
                background: #334155;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 800;
                padding: 0 16px;
            }
            QPushButton:hover { background: #1f2937; }
        """)
        new_client_btn.clicked.connect(self._switch_and_open_new_client_popup)

        manage_client_btn = QtWidgets.QPushButton("Edit Client")
        manage_client_btn.setFixedHeight(40)
        manage_client_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        manage_client_btn.setStyleSheet("""
            QPushButton {
                background: #0f766e;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 800;
                padding: 0 16px;
            }
            QPushButton:hover { background: #0d625c; }
        """)
        manage_client_btn.clicked.connect(self.edit_current_client_popup)

        client_select_layout.addWidget(select_client_label)
        client_select_layout.addWidget(self.line_items_client_combo, 1)
        client_select_layout.addWidget(new_client_btn)
        client_select_layout.addWidget(manage_client_btn)
        items_layout.addWidget(client_select_frame)

        self.invoice_client_info_frame = QtWidgets.QFrame()
        self.invoice_client_info_frame.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-radius: 8px;
            }
            QLabel {
                background: transparent;
                border: none;
                color: #0f172a;
                font-size: 13px;
            }
        """)
        invoice_client_layout = QtWidgets.QGridLayout(self.invoice_client_info_frame)
        invoice_client_layout.setContentsMargins(16, 12, 16, 12)
        invoice_client_layout.setHorizontalSpacing(18)
        invoice_client_layout.setVerticalSpacing(6)

        invoice_client_title = QtWidgets.QLabel("Client Information")
        invoice_client_title.setStyleSheet("font-size: 14px; font-weight: 900; color: #00756f;")
        self.invoice_client_name_label = QtWidgets.QLabel("Client: Not selected")
        self.invoice_client_email_label = QtWidgets.QLabel("Email: -")
        self.invoice_client_address_label = QtWidgets.QLabel("Address: -")
        self.invoice_client_address_label.setWordWrap(True)
        for info_label in (
            self.invoice_client_name_label,
            self.invoice_client_email_label,
            self.invoice_client_address_label,
        ):
            info_label.setStyleSheet("font-size: 13px; font-weight: 700; color: #334155;")

        invoice_client_layout.addWidget(invoice_client_title, 0, 0, 1, 2)
        invoice_client_layout.addWidget(self.invoice_client_name_label, 1, 0)
        invoice_client_layout.addWidget(self.invoice_client_email_label, 1, 1)
        invoice_client_layout.addWidget(self.invoice_client_address_label, 2, 0, 1, 2)
        items_layout.addWidget(self.invoice_client_info_frame)
        
        # Items container expands naturally with the page-level scroll area.
        self.items_widget = QtWidgets.QWidget()
        self.items_layout = QtWidgets.QVBoxLayout(self.items_widget)
        self.items_layout.setContentsMargins(0, 0, 0, 0)
        self.items_layout.setSpacing(10)
        
        items_layout.addWidget(self.items_widget)
        
        # Add item button
        add_item_btn = QtWidgets.QPushButton("a * Add New Item")
        add_item_btn.setText("Add Item")
        add_item_btn.setFixedHeight(36)
        add_item_btn.setMinimumWidth(150)
        add_item_btn.setMaximumWidth(190)
        add_item_btn.setStyleSheet("""
            QPushButton {
                background: #00756f;
                color: #ffffff;
                border: none;
                border-radius: 7px;
                font-size: 13px;
                font-weight: 800;
            }
            QPushButton:hover {
                background: #00645f;
            }
            QPushButton:pressed {
                background: #00514d;
            }
        """)
        add_item_btn.clicked.connect(self.add_item_row)
        add_item_row_layout = QtWidgets.QHBoxLayout()
        add_item_row_layout.setContentsMargins(0, 0, 0, 0)
        add_item_row_layout.addStretch()
        add_item_row_layout.addWidget(add_item_btn)
        add_item_row_layout.addStretch()
        items_layout.addLayout(add_item_row_layout)
        
        # Invoice summary
        summary_group = QtWidgets.QGroupBox("Invoice Summary")
        summary_group.setStyleSheet(panel_style)
        summary_layout = QtWidgets.QHBoxLayout(summary_group)
        summary_layout.setContentsMargins(20, 20, 20, 18)
        summary_layout.setSpacing(14)
        
        # Left - Settings
        settings_group = QtWidgets.QGroupBox("Tax Deatils")
        settings_group.setTitle("Tax Details")
        settings_group.setStyleSheet(panel_style)
        settings_group.setCheckable(False)
        settings_layout = QtWidgets.QVBoxLayout(settings_group)
        settings_layout.setContentsMargins(22, 24, 22, 22)
        
        # Tax settingsssss
        calc_layout = QtWidgets.QFormLayout()
        calc_layout.setVerticalSpacing(12)
        calc_layout.setHorizontalSpacing(12)
        
        self.tax_spin = QtWidgets.QDoubleSpinBox()
        self.tax_spin.setRange(0, 100)
        self.tax_spin.setDecimals(2)
        self.tax_spin.setSuffix(" %")
        self.tax_spin.setValue(float(Config.COMPANY.get("default_tax_rate", 0.0)))
        self.tax_spin.setFixedHeight(36)
        self.tax_spin.setMaximumWidth(170)
        self.tax_spin.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        self.tax_spin.setStyleSheet(field_style)
        # Prevent scroll and arrow keys from accidentally changing tax rate
        self.tax_spin.wheelEvent = lambda event: event.ignore()
        _tax_orig_key = self.tax_spin.keyPressEvent
        def _tax_key(event, _orig=_tax_orig_key):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            _orig(event)
        self.tax_spin.keyPressEvent = _tax_key
        
        tax_rate_label = QtWidgets.QLabel("Tax Rate:")
        tax_rate_label.setStyleSheet(form_label_style)
        calc_layout.addRow(tax_rate_label, self.tax_spin)
        
        settings_layout.addLayout(calc_layout)
        settings_layout.addStretch()
        self.tax_spin.setVisible(True)
        
        summary_layout.addWidget(settings_group, 1)
        
        # Center - Totals
        totals_group = QtWidgets.QGroupBox("Invoice Totals")
        totals_group.setStyleSheet(panel_style)
        totals_layout = QtWidgets.QFormLayout(totals_group)
        totals_layout.setContentsMargins(22, 24, 22, 22)
        totals_layout.setVerticalSpacing(10)
        totals_layout.setHorizontalSpacing(14)

        self.total_label = QtWidgets.QLabel("$0.00")
        self.down_payments_label = QtWidgets.QLabel("$0.00")
        self.tax_label = QtWidgets.QLabel("$0.00")
        self.total_amount_due_label = QtWidgets.QLabel("$0.00")
        self.payment_status_label = QtWidgets.QLabel("Not Started")
        total_label_title = QtWidgets.QLabel("Total:")
        down_payments_title = QtWidgets.QLabel("Deposit Received:")
        tax_amount_title = QtWidgets.QLabel("Tax Amount:")
        total_due_title = QtWidgets.QLabel("Remaining Due:")
        payment_status_title = QtWidgets.QLabel("Payment Status:")
        for label in (
            total_label_title,
            down_payments_title,
            tax_amount_title,
            total_due_title,
            payment_status_title,
            self.total_label,
            self.down_payments_label,
            self.tax_label,
        ):
            label.setMinimumHeight(28)
            label.setStyleSheet(form_label_style)

        self.total_amount_due_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: 800;
                color: #00756f;
                background-color: #eefaf8;
                border: 1.5px solid #9ddbd4;
                border-radius: 7px;
                padding: 10px;
            }
        """)
        self.total_amount_due_label.setMinimumHeight(40)
        self.payment_status_label.setMinimumHeight(34)
        self.payment_status_label.setAlignment(QtCore.Qt.AlignCenter)
        self.payment_status_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: 800;
                color: #42526e;
                background-color: #f2f6fa;
                border: 1px solid #d8e2ec;
                border-radius: 7px;
                padding: 8px 12px;
            }
        """)

        totals_layout.addRow(total_label_title, self.total_label)
        totals_layout.addRow(tax_amount_title, self.tax_label)
        totals_layout.addRow(total_due_title, self.total_amount_due_label)
        totals_layout.addRow(payment_status_title, self.payment_status_label)

        summary_layout.addWidget(totals_group, 2)
        items_layout.addWidget(summary_group)
        content_layout.addWidget(items_group, 1, 0, 1, 2)
        
        # Invoice actions - shown below the summary so the workflow stays in order.
        actions_group = QtWidgets.QGroupBox("Actions")
        actions_group.setStyleSheet(panel_style)
        actions_layout = QtWidgets.QHBoxLayout(actions_group)
        actions_layout.setContentsMargins(22, 20, 22, 20)
        actions_layout.setSpacing(14)

        # Enhanced PDF Button with icon and increased width
        self.generate_pdf_btn = QtWidgets.QPushButton("Save & Generate Invoice")
        self.generate_pdf_btn.setMinimumHeight(52)
        self.generate_pdf_btn.setMaximumHeight(52)
        self.generate_pdf_btn.setMinimumWidth(220)
        self.generate_pdf_btn.setMaximumWidth(420)
        self.generate_pdf_btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        self.generate_pdf_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        # Enhanced styling with icon layout
        self.generate_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #00756f;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 0 20px;
                font-weight: 800;
                font-size: 17px;
                min-height: 52px;
                max-height: 52px;
                margin: 0px;
            }
            QPushButton:hover {
                background-color: #00645f;
            }
            QPushButton:pressed {
                background-color: #00514d;
            }
        """)

        self.generate_pdf_btn.clicked.connect(self.generate_pdf)
        # Wire client name changes to re-evaluate button state
        self.client_combo.editTextChanged.connect(self._update_pdf_btn_state)
        # Initial state — disabled until a client is entered
        QtCore.QTimer.singleShot(0, self._update_pdf_btn_state)

        actions_layout.addStretch(1)
        self.save_invoice_btn = QtWidgets.QPushButton("Save Invoice")
        self.save_invoice_btn.setMinimumHeight(52)
        self.save_invoice_btn.setMaximumHeight(52)
        self.save_invoice_btn.setMinimumWidth(200)
        self.save_invoice_btn.setMaximumWidth(360)
        self.save_invoice_btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        self.save_invoice_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.save_invoice_btn.setStyleSheet("""
            QPushButton {
                background: #334155;
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: 800;
                font-size: 15px;
                min-height: 52px;
                max-height: 52px;
                padding: 0 16px;
            }
            QPushButton:hover {
                background: #1f2937;
            }
            QPushButton:pressed {
                background: #111827;
            }
        """)
        self.save_invoice_btn.clicked.connect(self.save_invoice)
        actions_layout.addWidget(self.save_invoice_btn)
        self.save_invoice_btn.hide()  # Hidden - only Generate PDF is used
        actions_layout.addWidget(self.generate_pdf_btn)

        self.history_btn = QtWidgets.QPushButton("Invoice History")
        self.history_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.history_btn.clicked.connect(self.open_invoice_history)
        self.history_btn.setMinimumHeight(52)
        self.history_btn.setMaximumHeight(52)
        self.history_btn.setMinimumWidth(200)
        self.history_btn.setMaximumWidth(360)
        self.history_btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        self.history_btn.setStyleSheet("""
            QPushButton {
                background: #475569;
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: 800;
                font-size: 15px;
                min-height: 52px;
                max-height: 52px;
                padding: 0 16px;
            }
            QPushButton:hover {
                background: #1f2937;
            }
            QPushButton:pressed {
                background: #111827;
            }
        """)
        actions_layout.addWidget(self.history_btn)
        actions_layout.addStretch(1)

        actions_group.setMinimumHeight(92)
        actions_group.setMaximumHeight(130)
        actions_group.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        items_layout.addWidget(actions_group)
        
        # Notes section
        notes_group = QtWidgets.QGroupBox("Notes and Terms")
        notes_group.setStyleSheet(panel_style)
        notes_group.setCheckable(False)
        notes_layout = QtWidgets.QVBoxLayout(notes_group)
        self.notes_edit = QtWidgets.QTextEdit(Config.DEFAULT_TERMS)
        self.notes_edit.setMaximumHeight(80)
        self.notes_edit.setVisible(True)
        notes_layout.addWidget(self.notes_edit)
        
        content_layout.addWidget(notes_group, 2, 0, 1, 2)
        
        main_layout.addLayout(content_layout)
        
        # Set the scroll area as the main widget for create tab
        self.create_tab_layout = QtWidgets.QVBoxLayout(self.create_tab)
        self.create_tab_layout.setContentsMargins(0, 0, 0, 0)
        self.create_tab_layout.addWidget(scroll_area)
        
        # Add initial item row
        self.add_item_row()

    def refresh_balance_sheet(self):
        """Force refresh balance sheet tab"""
        try:
            if hasattr(self, 'balance_sheet_tab'):
                balance_tab = self.balance_sheet_tab
                
                # Reload ALL financial data from Firebase
                balance_tab.load_all_financial_data()
                
                # Reload annual summary data for current year
                balance_tab.load_annual_summary_data_for_year(balance_tab.annual_summary_year)
                
                # Also reload for the transaction table year if different
                if balance_tab.current_year != balance_tab.annual_summary_year:
                    balance_tab.load_annual_summary_data_for_year(balance_tab.current_year)
                
                # Update the annual summary table
                balance_tab.update_annual_summary()
                
                # Refresh the current transaction table view
                balance_tab.on_category_changed(balance_tab.current_category)
                
                # Update stats cards
                balance_tab.update_stats_cards()
                
                log.info("Balance sheet fully refreshed")
        except Exception as e:
            log.warning("Error refreshing balance sheet: %s", e)
            import traceback
            traceback.print_exc()
        
    def on_update_available(self):
        """Called when UpdateIndicator detects an update — dialog is shown by UpdateChecker."""
        log.info("Update available signal received.")

    def show_notification(self, title, message):
        """Show a notification message"""
        QtWidgets.QMessageBox.information(self, title, message)

    def show_settings(self):
        """Show settings dialog"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Settings")
        dialog.setFixedSize(400, 300)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Update settings section
        update_group = QtWidgets.QGroupBox("Update Settings")
        update_layout = QtWidgets.QVBoxLayout(update_group)
        
        # Auto-check for updates
        self.auto_check_checkbox = QtWidgets.QCheckBox("Automatically check for updates")
        self.auto_check_checkbox.setChecked(True)
        update_layout.addWidget(self.auto_check_checkbox)
        
        # Check now button
        check_now_btn = QtWidgets.QPushButton("Check for Updates Now")
        check_now_btn.clicked.connect(lambda: self.update_indicator.on_update_clicked())
        update_layout.addWidget(check_now_btn)
        
        layout.addWidget(update_group)
        layout.addStretch()
        
        # OK button
        ok_btn = QtWidgets.QPushButton("OK")
        ok_btn.clicked.connect(dialog.accept)
        layout.addWidget(ok_btn)
        
        dialog.exec_()
    
    def _on_client_name_edit_changed(self, text):
        """Enable client saving when an existing client name is edited/renamed."""
        if not hasattr(self, "save_client_btn"):
            return
        original = getattr(self, "editing_client_name", None)
        if original and str(text).strip() != str(original).strip():
            self.save_client_btn.setEnabled(True)

    def load_data(self):
        """Load clients and settings from Firebase only"""
        
        # Load from Firebase
        self.clients = FirebaseManager.load_clients()
        self.settings = {}
            
        # Update client combo
        self.client_combo.blockSignals(True)
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.blockSignals(True)
        self.client_combo.clear()
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.clear()

        self.client_combo.setEditable(True)
        self.client_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.setEditable(True)
            self.line_items_client_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)

        # a ... Placeholder (modern way)
        self.client_combo.lineEdit().setPlaceholderText("Select or type client name...")
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.lineEdit().setPlaceholderText("Select or type client name before generating invoice...")

        # Add clients
        for client_name in sorted(self.clients.keys()):
            self.client_combo.addItem(client_name)
            if hasattr(self, "line_items_client_combo"):
                self.line_items_client_combo.addItem(client_name)

        # a ... No default selection
        self.client_combo.setCurrentIndex(-1)
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.setCurrentIndex(-1)

        self.client_combo.blockSignals(False)
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.blockSignals(False)
        if hasattr(self, "save_client_btn"):
            self.save_client_btn.setEnabled(False)
        self.refresh_inline_clients_table()
        
        # Refresh project tab
        if hasattr(self, 'project_tab'):
            self.refresh_project_tab_clients()
        
        # Overlay Firebase settings (source of truth) and refresh all UI elements
        if Config.overlay_from_firebase():
            self._apply_company_to_ui()

        # Run automation sweeps 3 s after startup so the UI is fully painted first
        QtCore.QTimer.singleShot(3000, self._run_startup_automation)

    def _run_startup_automation(self):
        """Deferred startup sweep: mark overdue, complete projects, send reminders.
        Runs silently without interfering with the UI."""
        
        # Suppress all warnings and popups during startup automation
        original_warning = QtWidgets.QMessageBox.warning
        original_information = QtWidgets.QMessageBox.information
        original_critical = QtWidgets.QMessageBox.critical
        original_question = QtWidgets.QMessageBox.question
        
        # Temporarily replace QMessageBox methods with silent versions
        def silent_warning(*args, **kwargs):
            pass
        
        def silent_information(*args, **kwargs):
            pass
        
        def silent_critical(*args, **kwargs):
            pass
        
        def silent_question(*args, **kwargs):
            return QtWidgets.QMessageBox.No
        
        QtWidgets.QMessageBox.warning = silent_warning
        QtWidgets.QMessageBox.information = silent_information
        QtWidgets.QMessageBox.critical = silent_critical
        QtWidgets.QMessageBox.question = silent_question
        
        try:
            if not FIREBASE_AVAILABLE:
                return
            overdue_count = FirebaseManager.auto_mark_overdue_invoices()
            completed_count = FirebaseManager.auto_complete_projects()
            expired_quotes = FirebaseManager.auto_expire_quotes()
            
            if overdue_count or completed_count or expired_quotes:
                if hasattr(self, 'history_tab') and hasattr(self.history_tab, 'refresh_invoices_immediately'):
                    self.history_tab.refresh_invoices_immediately()
                if expired_quotes and hasattr(self, 'job_form_tab'):
                    self.job_form_tab.load_job_forms_from_firebase()

            # Payment reminders (only if email is configured)
            reminder_count = 0
            try:
                from email_manager import EmailManager
                if EmailManager.is_configured():
                    import json
                    settings_path = Config.SETTINGS_FILE
                    days_before = 3
                    if settings_path.exists():
                        with open(settings_path, encoding="utf-8") as f:
                            days_before = json.load(f).get("email", {}).get("reminder_days_before", 3)
                    invoices = FirebaseManager.load_invoices()
                    reminder_count = EmailManager.send_payment_reminders(invoices, days_before)
            except Exception as exc:
                log.warning("Payment reminder sweep failed: %s", exc)

            # Log quietly without popups
            if overdue_count or completed_count or expired_quotes or reminder_count:
                log.info(
                    "Startup automation: %d overdue, %d projects completed, %d quotes expired, %d reminders sent",
                    overdue_count, completed_count, expired_quotes, reminder_count,
                )
                
        finally:
            # Restore original QMessageBox methods
            QtWidgets.QMessageBox.warning = original_warning
            QtWidgets.QMessageBox.information = original_information
            QtWidgets.QMessageBox.critical = original_critical
            QtWidgets.QMessageBox.question = original_question

    def _preload_invoice_cache(self):
        """Pre-load invoice cache in background after window is shown."""
        try:
            from invoice_history_tab import preload_invoices_cache
            preload_invoices_cache()
            log.info("Invoice cache pre-loading initiated")
        except Exception as e:
            log.warning("Failed to initiate invoice cache pre-loading: %s", e)

    def save_data(self):
        """Save clients and settings to Firebase only"""
        # Clients are saved individually to Firebase when created
        # This method is kept for compatibility but doesn't need to do anything
        pass

    # Add/modify this method in the MainWindow class in main.py:

    def create_project_from_quote(self, job_data: dict):
        """Open New Project from a quote and convert the quote after project save.
          "  Sales users cannot access this feature."""
        
        #   "  CHECK: Sales users cannot create projects
        if not can_perform_action(self.current_role, ACTION_CONVERT_QUOTE_TO_PROJECT):
            QtWidgets.QMessageBox.warning(
                self,
                "Access Denied",
                "You do not have permission to create projects.\n\n"
                "Please contact an administrator or a project manager to convert this quote to a project."
            )
            return
        
        if not job_data:
            return

        job_number = job_data.get('job_number', '').strip()
        if not job_number:
            QtWidgets.QMessageBox.warning(self, "Create Project", "This quote has no quote number.")
            return

        # Fast-path: quote status already marked Converted — warn before navigating away
        if job_data.get('status') == 'Converted':
            linked_proj = job_data.get('project_number', 'N/A')
            msg = QtWidgets.QMessageBox(self)
            msg.setWindowTitle("Project Already Created")
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText("<b>This quote has already been converted to a project.</b>")
            msg.setInformativeText(
                f"Quote <b>{job_number}</b> was already pushed to:<br><br>"
                f"&nbsp;&nbsp;Project #: <b>{linked_proj}</b><br><br>"
                f"To make changes, open that project directly."
            )
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        if not self._nav_to(PAGE_PROJECTS):
            return
        if hasattr(self, 'project_invoice_inner_tabs'):
            self.project_invoice_inner_tabs.setCurrentIndex(0)

        if hasattr(self, 'project_tab') and hasattr(self.project_tab, 'load_projects'):
            self.project_tab.load_projects()

        existing_project = self.find_project_by_quote_number(job_number, job_data.get('project_number', ''))
        if existing_project:
            if job_data.get('status') != 'Converted' or job_data.get('project_number') != existing_project.get('project_number'):
                self.mark_quote_converted_to_project(job_data, existing_project)
            proj_num  = existing_project.get('project_number', 'N/A')
            proj_name = existing_project.get('project_name', '') or existing_project.get('project_name', 'Untitled')
            msg = QtWidgets.QMessageBox(self)
            msg.setWindowTitle("Project Already Created")
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setText(f"<b>This quote has already been converted to a project.</b>")
            msg.setInformativeText(
                f"Quote <b>{job_number}</b> was already pushed to:<br><br>"
                f"&nbsp;&nbsp;Project #: <b>{proj_num}</b><br>"
                f"&nbsp;&nbsp;Name: <b>{proj_name}</b><br><br>"
                f"To make changes, open that project directly."
            )
            msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            msg.setDefaultButton(QtWidgets.QMessageBox.Ok)
            msg.exec_()
            return

        from project_number_generator import ProjectDialog
        dialog = ProjectDialog(
            self,
            self.project_tab if hasattr(self, 'project_tab') else self,
            project_data=None,
            firebase_available=FIREBASE_AVAILABLE
        )
        dialog.job_number_edit.setText(job_number)
        dialog.load_job_details()
        try:
            dialog.job_number_edit.setReadOnly(True)
            if hasattr(dialog, "_apply_locked_identifier_style"):
                dialog._apply_locked_identifier_style(dialog.job_number_edit)
        except Exception:
            pass

        # Auto-fill price from quote — prefer project_amount (already includes expedite premium)
        try:
            price_raw = job_data.get('project_amount', '') or job_data.get('engineering_costs', '')
            if price_raw:
                price_val = float(str(price_raw).replace('$', '').replace(',', ''))
                dialog.project_amount_edit.setText(f"${price_val:,.2f}")
        except Exception:
            pass
        try:
            sales_val = job_data.get('sales', '')
            if sales_val:
                idx = dialog.sales_combo.findText(sales_val)
                if idx >= 0:
                    dialog.sales_combo.setCurrentIndex(idx)
                else:
                    dialog.sales_combo.setEditText(sales_val)
        except Exception:
            pass

        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            project_data = dialog.collect_project_data()
            project_data['job_number'] = job_number
            project_data.setdefault('quote_number', job_number)
            self.mark_quote_converted_to_project(job_data, project_data)
            if hasattr(self, 'project_tab') and hasattr(self.project_tab, 'load_projects'):
                self.project_tab.load_projects()
                
    def find_project_by_quote_number(self, job_number: str, project_number: str = ""):
        """Return an existing project linked to a quote number, if any."""
        target = str(job_number or '').strip().upper()
        project_target = str(project_number or '').strip().upper()
        if not target or not hasattr(self, 'project_tab'):
            return None

        project_sources = []
        for attr in ('generated_projects', 'cached_projects'):
            project_sources.extend(getattr(self.project_tab, attr, []) or [])

        seen_numbers = set()
        for project in project_sources:
            if not isinstance(project, dict):
                continue
            project_number = project.get('project_number', '')
            if project_number in seen_numbers:
                continue
            seen_numbers.add(project_number)
            if str(project.get('job_number', '')).strip().upper() == target:
                return project
            if project_target and str(project_number).strip().upper() == project_target:
                return project
        return None

    def mark_quote_converted_to_project(self, job_data: dict, project_data: dict):
        """Persist the project link back to the quote and refresh quote stats."""
        if not job_data or not project_data:
            return False

        project_number = project_data.get('project_number', '').strip()
        if not project_number:
            return False

        updates = {
            'status': 'Converted',
            'project_number': project_number,
            'converted_at': datetime.now(timezone.utc).isoformat(),
            'updated_at': datetime.now(timezone.utc).isoformat(),
        }
        job_data.update(updates)

        saved = False
        if FIREBASE_AVAILABLE:
            try:
                ref = db.reference('/job_forms')
                firebase_id = job_data.get('firebase_id')
                if firebase_id:
                    ref.child(firebase_id).update(updates)
                    saved = True
                else:
                    existing = ref.order_by_child('job_number').equal_to(job_data.get('job_number', '')).get() or {}
                    for quote_id in existing.keys():
                        ref.child(quote_id).update(updates)
                        saved = True
            except Exception as exc:
                log.warning("Could not link quote %s to project %s: %s", job_data.get('job_number'), project_number, exc)

        if hasattr(self, 'job_form_tab'):
            for quote in getattr(self.job_form_tab, 'job_forms', []) or []:
                if quote.get('job_number', '').strip().upper() == job_data.get('job_number', '').strip().upper():
                    quote.update(updates)
                    break
            if not saved and hasattr(self.job_form_tab, 'save_job_form_locally'):
                saved = self.job_form_tab.save_job_form_locally(job_data)
            QtCore.QTimer.singleShot(0, self.job_form_tab.update_job_forms_table)

        log.info("Quote %s converted to project %s", job_data.get('job_number'), project_number)
        return saved
    
    # Update this method in MainWindow class:

    def prefill_invoice_from_quote(self, job_data: dict):
        """Navigate to Invoice Management and pre-fill fields from a quote/job form.
          "  Sales users cannot create invoices."""
        
        #   "  CHECK: Sales users cannot create invoices
        if not can_perform_action(self.current_role, ACTION_CONVERT_QUOTE_TO_INVOICE):
            QtWidgets.QMessageBox.warning(
                self,
                "Access Denied",
                "You do not have permission to create invoices.\n\n"
                "Please contact a project manager or finance manager to convert this quote to an invoice."
            )
            return
        
        if not self._nav_to(PAGE_PROJECTS):
            return
        if hasattr(self, 'project_invoice_inner_tabs'):
            self.project_invoice_inner_tabs.setCurrentIndex(1)

        client_name = job_data.get('client', '').strip()
        project_name = job_data.get('project_name', '').strip()
        job_number = job_data.get('job_number', '').strip()
        scope = job_data.get('scope_of_work', '').strip()
        eng_costs_raw = job_data.get('engineering_costs', '0')

        if client_name:
            for combo in ('client_combo', 'line_items_client_combo'):
                if hasattr(self, combo):
                    getattr(self, combo).setEditText(client_name)
            self.load_client_details(client_name)

        if hasattr(self, 'item_rows') and self.item_rows:
            row = self.item_rows[0]
            if hasattr(row, 'project_number_edit'):
                row.project_number_edit.setText(job_number)
            if hasattr(row, 'project_name_edit'):
                row.project_name_edit.setText(project_name)
            if hasattr(row, 'description_edit'):
                row.description_edit.setText(scope if scope else project_name)
            if hasattr(row, 'unit_price_spin'):
                try:
                    amount = float(str(eng_costs_raw).replace('$', '').replace(',', '')) if eng_costs_raw else 0.0
                    row.unit_price_spin.setValue(amount)
                except (ValueError, TypeError):
                    pass

        self.update_totals()
        log.info("Invoice pre-filled from quote %s (client: %s)", job_number, client_name)

    def auto_fill_project_details(self, project_number: str):
        """Auto-fill invoice item details when project number is entered"""
        if not project_number or not project_number.strip():
            return
        
        # Find which row triggered this
        for row in self.item_rows:
            if row.project_number_edit.text().strip() == project_number:
                self.fill_project_details_for_row(project_number, row)
                break
    
    def fill_project_details_for_row(self, project_number: str, row: ItemRowWidget):
        """Fill project details for a specific row"""
        try:
            # Look for the project in cached projects
            if hasattr(self, 'project_tab') and hasattr(self.project_tab, 'cached_projects'):
                for project in self.project_tab.cached_projects:
                    if project.get("project_number", "") == project_number:
                        # Found the project - auto-fill details
                        self.auto_populate_row_from_project(row, project)
                        return
                
                # If not found in cached projects, try to load from Firebase
                if FIREBASE_AVAILABLE:
                    self.load_project_from_firebase(project_number, row)
        
        except Exception as e:
            log.warning("Error auto-filling project details: %s", e)
    
    def open_invoice_history(self):
        """Open invoice history in a maximized window with proper window controls"""
        try:
            # Create a QMainWindow instead of QDialog for better window controls
            history_window = QtWidgets.QMainWindow(self)
            history_window.setWindowTitle("  Invoice History - MABS Engineering")
            
            # Set window flags to include minimize/maximize buttons
            history_window.setWindowFlags(
                QtCore.Qt.Window |  # Makes it a standalone window
                QtCore.Qt.WindowMinimizeButtonHint |
                QtCore.Qt.WindowMaximizeButtonHint |
                QtCore.Qt.WindowCloseButtonHint
            )
            
            # Create the invoice history widget - PASS THE MAIN WINDOW REFERENCE
            from invoice_history_tab import InvoiceHistoryTab
            history_widget = InvoiceHistoryTab(self)
            
            # Set as central widget
            history_window.setCentralWidget(history_widget)
            
            # Add a toolbar with close button
            toolbar = QtWidgets.QToolBar("History Toolbar")
            toolbar.setMovable(False)
            history_window.addToolBar(toolbar)
            
            # Close action
            close_action = QtWidgets.QAction("✕ Close", history_window)
            close_action.setShortcut("Esc")
            close_action.triggered.connect(history_window.close)
            toolbar.addAction(close_action)
            
            # Add separator
            toolbar.addSeparator()
            
            # Refresh action
            refresh_action = QtWidgets.QAction("  Refresh", history_window)
            refresh_action.triggered.connect(history_widget.refresh_data)
            toolbar.addAction(refresh_action)
            
            # Show window maximized (non-modal)
            history_window.showMaximized()
            
            # Return the window reference in case you need to track it
            return history_window
            
        except ImportError as e:
            log.warning("Import error: %s", e)
            QtWidgets.QMessageBox.warning(self, "Import Error", 
                                        "Invoice history module not available.")
            return None
        except Exception as e:
            log.warning("Error opening invoice history: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", 
                                        f"Failed to open invoice history: {str(e)}")
            return None
    
    def auto_populate_row_from_project(self, row: ItemRowWidget, project: dict):
        try:
            row.desc_edit.blockSignals(True)
            row.plant_edit.blockSignals(True)
            row.price_edit.blockSignals(True)
            row.down_payment_combo.blockSignals(True)

            # Store project name separately
            row.item.project_name = project.get("project_name", "")

            # Description can be a combination or just the project name
            row.desc_edit.setText(project.get("project_name", ""))

            # Plant
            row.plant_edit.setText(project.get("plant", ""))

            # Unit price (QLineEdit)
            project_amount = project.get("project_amount", 0.0)
            if project_amount:
                row.price_edit.setText(f"${float(project_amount):.2f}")

            payment_category = row.normalize_payment_label(project.get("payment_category", row.FINAL_PAYMENT_LABEL))
            row.down_payment_combo.setCurrentText(payment_category or row.FINAL_PAYMENT_LABEL)

        finally:
            row.desc_edit.blockSignals(False)
            row.plant_edit.blockSignals(False)
            row.price_edit.blockSignals(False)
            row.down_payment_combo.blockSignals(False)

        # Determine status badge based on project payment stage
        project_status = (project.get("status", "") or "").lower().strip()
        pay_cat = row.normalize_payment_label(project.get("payment_category", "") or "")
        _FIRST_STAGES = {ItemRowWidget.DEPOSIT_LABEL, ItemRowWidget.FULL_AMOUNT_LABEL,
                         ItemRowWidget.INSTALLMENT_1_LABEL, ""}
        _PAID_PROJECT_STATUSES = {"paid", "completed", "closed", "invoiced"}
        if project_status in _PAID_PROJECT_STATUSES:
            self._ps_base_status = "Paid"   # update_totals will flip to Partially Paid if tax > 0
        elif pay_cat not in _FIRST_STAGES:
            # A previous stage was already billed/paid → partial
            self._ps_base_status = "Partially Paid"
        else:
            # First stage or unknown → Unpaid (update_totals will flip to Overdue if past due)
            self._ps_base_status = "Unpaid"

        # Force recalculation
        row.update_total()
        self.update_totals()


    def load_project_from_firebase(self, project_number: str, row: ItemRowWidget):
        """Load project details from Firebase for auto-fill"""
        if not FIREBASE_AVAILABLE:
            return
        
        try:
            from main import db
            ref = db.reference('/projects')
            projects_data = ref.order_by_child('project_number').equal_to(project_number).get()
            
            if projects_data:
                # Get the first matching project
                project_id = list(projects_data.keys())[0]
                project = projects_data[project_id]
                
                # Auto-populate the row
                self.auto_populate_row_from_project(row, project)
                
                # Also update cached projects for future lookups
                if hasattr(self, 'project_tab'):
                    # Check if project already exists in cache
                    existing = False
                    for cached_project in self.project_tab.cached_projects:
                        if cached_project.get("project_number") == project_number:
                            existing = True
                            break
                    
                    if not existing:
                        project["firebase_id"] = project_id
                        self.project_tab.cached_projects.append(project)
                        log.info("Added project to cache: %s", project_number)
                
        except Exception as e:
            log.warning("Error loading project from Firebase: %s", e)
            
    def add_item_row(self, item: InvoiceItem = None):
        """Add a new item row to the form"""
        row = ItemRowWidget(item)
        row.removed.connect(lambda: self.remove_item_row(row))
        self.items_layout.addWidget(row)
        self.item_rows.append(row)
        self.renumber_item_rows()
        
        # Connect signals for auto-update
        row.qty_spin.valueChanged.connect(self.update_totals)
        row.price_edit.textChanged.connect(self.update_totals)
        row.down_payment_combo.currentTextChanged.connect(self.update_totals)

        # Connect required-field signals so PDF button reacts immediately
        row.project_number_edit.textChanged.connect(self._update_pdf_btn_state)
        row.desc_edit.textChanged.connect(self._update_pdf_btn_state)
        row.price_edit.textChanged.connect(self._update_pdf_btn_state)
        row.down_payment_combo.currentIndexChanged.connect(self._update_pdf_btn_state)

        # Connect project number changed signal to auto-fill method
        row.project_changed.connect(self.auto_fill_project_details)
        
        # Restore payment category -- prefer saved label, fall back to amount-based detection
        if item:
            normalized = ItemRowWidget.normalize_payment_label(item.payment_category)
            if normalized:
                row.down_payment_combo.setCurrentText(normalized)
            elif item.down_payment > 0:
                total = Decimal(str(item.quantity)) * Decimal(str(item.unit_price))
                down_payment = Decimal(str(item.down_payment))
                if total > 0 and abs((down_payment / total) - Decimal("0.5")) < Decimal("0.01"):
                    row.down_payment_combo.setCurrentText(row.DEPOSIT_LABEL)
                else:
                    row.down_payment_combo.setCurrentText(row.FINAL_PAYMENT_LABEL)
        
        self.update_totals()
        self._update_pdf_btn_state()

    def remove_item_row(self, row):
        """Remove an item row from the form"""
        if row in self.item_rows:
            self.item_rows.remove(row)
            row.setParent(None)
            self.renumber_item_rows()
            self.update_totals()
            self._update_pdf_btn_state()

    def renumber_item_rows(self):
        """Keep visible line item labels in the same order as the form."""
        for index, row in enumerate(self.item_rows, start=1):
            if hasattr(row, "set_item_number"):
                row.set_item_number(index)
    
    def update_totals(self):
        """Update the totals display with correct tax calculation on total amount"""
        # Update invoice object from form
        self.update_invoice_from_form()
        
        # Calculate totals from the actual invoice items
        total_amount = sum(item.total for item in self.invoice.items)
        total_payment_due = sum(item.payment_due for item in self.invoice.items)

        tax_amount = total_amount * (self.invoice.tax_rate / Decimal("100"))
        total_amount_due = total_payment_due + tax_amount

        # Update UI labels
        if self.total_label:
            self.total_label.setText(Currency.format(total_amount))

        if self.tax_label:
            self.tax_label.setText(Currency.format(tax_amount))

        if self.total_amount_due_label:
            self.total_amount_due_label.setText(Currency.format(total_amount_due))

        if hasattr(self, "payment_status_label") and self.payment_status_label:
            _ps_base = getattr(self, "_ps_base_status", None)
            if _ps_base is not None:
                status_text = _ps_base
                if getattr(self, '_editing_existing_invoice', False):
                    # Dynamic recalculation: compare stored paid amount against current total
                    _paid = getattr(self, '_editing_invoice_paid_amount', None)
                    if _paid is not None and _paid > 0:
                        _current_total = float(total_amount + tax_amount)
                        if _current_total > 0 and _paid >= _current_total:
                            status_text = "Paid"
                        else:
                            status_text = "Partially Paid"
                elif _ps_base == "Paid" and tax_amount > 0:
                    # Project marked Paid but tax not yet settled → partially paid.
                    status_text = "Partially Paid"
                # Recheck overdue against current due date when status is unpaid/overdue
                if status_text in ("Unpaid", "Overdue") and hasattr(self, "due_date_edit"):
                    try:
                        from PyQt5.QtCore import QDate
                        due_qdate = self.due_date_edit.date()
                        today = QDate.currentDate()
                        if status_text == "Unpaid" and due_qdate < today:
                            status_text = "Overdue"
                        elif status_text == "Overdue" and due_qdate >= today:
                            status_text = "Unpaid"
                    except Exception:
                        pass
                _status_styles = {
                    "Paid":           "color:#00756f;background-color:#e9fbf7;border-color:#8edbd2;",
                    "Partially Paid": "color:#1e40af;background-color:#dbeafe;border-color:#93c5fd;",
                    "Overdue":        "color:#9d174d;background-color:#fce7f3;border-color:#f9a8d4;",
                    "Unpaid":         "color:#7a1f1f;background-color:#fff1f0;border-color:#f0b4b4;",
                    "Not Started":    "color:#42526e;background-color:#f2f6fa;border-color:#d8e2ec;",
                }
                status_css = _status_styles.get(status_text, _status_styles["Unpaid"])
            else:
                if total_amount == 0:
                    status_text = "Not Started"
                    status_css = "color: #42526e; background-color: #f2f6fa; border-color: #d8e2ec;"
                elif total_amount_due <= 0:
                    status_text = "Paid"
                    status_css = "color: #00756f; background-color: #e9fbf7; border-color: #8edbd2;"
                else:
                    status_text = "Unpaid"
                    status_css = "color: #7a1f1f; background-color: #fff1f0; border-color: #f0b4b4;"
            self.payment_status_label.setText(status_text)
            self.payment_status_label.setStyleSheet(f"""
                QLabel {{
                    font-size: 13px;
                    font-weight: 800;
                    border: 1px solid;
                    border-radius: 7px;
                    padding: 8px 12px;
                    {status_css}
                }}
            """)
        
        # Debug output to verify calculation
        log.info("  Totals Calculation:")
        log.info("   Total Amount: %s", Currency.format(total_amount))
        log.info("   Payment Due: %s", Currency.format(total_payment_due))
        log.info("   Payment Due (before tax): %s", Currency.format(total_payment_due))
        log.info("   Tax Rate: %s%%", self.invoice.tax_rate)
        log.info("   Tax Amount (on total): %s", Currency.format(tax_amount))
        log.info("   Total Amount Due: %s", Currency.format(total_amount_due))
        
    def _on_due_date_changed(self):
        """Recompute Overdue/Unpaid status when the due date is edited by the user."""
        _ps_base = getattr(self, "_ps_base_status", None)
        if _ps_base not in ("Unpaid", "Overdue"):
            self.update_totals()
            return
        try:
            from PyQt5.QtCore import QDate
            due_qdate = self.due_date_edit.date()
            today = QDate.currentDate()
            self._ps_base_status = "Overdue" if due_qdate < today else "Unpaid"
        except Exception:
            pass
        self.update_totals()

    def clear_all_items(self):
        """Clear all invoice items without adding an empty row back"""
        self._ps_base_status = None  # reset so update_totals reverts to default logic
        self._editing_existing_invoice = False
        self._editing_invoice_paid_amount = None
        for row in self.item_rows[:]:
            self.remove_item_row(row)
    
    def update_invoice_from_form(self):
        """Update the invoice object from form data"""
        # Basic info
        self.invoice.date = self.date_edit.date().toString("MM-dd-yyyy")
        self.invoice.due_date = self.due_date_edit.date().toString("MM-dd-yyyy")
        
        # Client info - FIXED: Use client_combo instead of client_name_edit
        client_name = self.client_combo.currentText()
        if "-- Select Company --" in client_name:
            client_name = client_name.replace("-- Select Company --", "").strip()
        self.invoice.client_name = client_name
        
        self.invoice.client_email = self.client_email_edit.text()
        self.invoice.client_address = self.client_address_edit.toPlainText()
        
        # Items
        self.invoice.items = []
        for row in self.item_rows:
            self.invoice.items.append(row.get_item())
        
        # Tax
        if self.tax_spin:
            self.invoice.tax_rate = Decimal(str(self.tax_spin.value()))
        
        # Notes
        self.invoice.notes = self.notes_edit.toPlainText()
        
    def update_invoice_preview(self):
        """Update the invoice number preview using CURRENT date (not invoice date)"""
        from datetime import datetime as _dt
        today_str = _dt.now().strftime("%Y-%m-%d")
        preview_number = InvoiceNumberGenerator.get_preview_number(today_str)
        self.invoice_no_edit.setText(preview_number)
    
    def load_client(self):
        """Load selected client data WITHOUT clearing existing line items"""
        client_name = self.client_combo.currentText()
        saved_client_name, client_data = self._find_client_record(client_name)
        if not client_data:
            return
        
        # DON'T CLEAR EXISTING ITEMS - just update client information
        self._sync_client_combos(saved_client_name)
        if hasattr(self, "client_contact_edit"):
            self.client_contact_edit.setText(client_data.get("contact_person", ""))
        if hasattr(self, "client_phone_edit"):
            self.client_phone_edit.setText(client_data.get("phone", ""))
        if hasattr(self, "client_primary_email_edit"):
            self.client_primary_email_edit.setText(client_data.get("email", ""))
        self.client_email_edit.setText(client_data.get("company_email", client_data.get("email", "")))
        self.client_address_edit.setPlainText(client_data.get("address", ""))
        self.update_invoice_client_summary(
            saved_client_name,
            self.client_email_edit.text().strip(),
            client_data.get("address", ""),
        )
        
        # Ask user if they want to load completed projects (this will add to existing items)
        reply = self.ask_load_completed_projects(saved_client_name)
        
        if reply == QtWidgets.QMessageBox.Yes:
            self.load_client_projects(saved_client_name)
    
    def load_client_projects(self, client_name):
        """Load completed projects for the selected client and add them as invoice items"""
        try:
            saved_client_name, client_data = self._find_client_record(client_name)
            if client_data:
                self.load_client_details(saved_client_name)
                client_name = saved_client_name

            # Clear existing items first
            for row in self.item_rows[:]:
                self.remove_item_row(row)
            
            # Load projects data from Firebase only
            projects = FirebaseManager.load_projects()
            
            # Filter projects for this client with status "Completed"
            completed_projects = [
                project for project in projects 
                if isinstance(project, dict) and 
                project.get("company") == client_name and 
                project.get("status") == "Completed"
            ]
            
            if not completed_projects:
                QtWidgets.QMessageBox.information(
                    self, "No Completed Projects", 
                    f"No completed projects found for <b>{client_name}</b>.<br><br>",
                    QtWidgets.QMessageBox.Ok
                )
                # Add one empty row if no projects found
                self.add_item_row()
                return
            
            # Add completed projects as invoice items
            loaded_count = 0
            for project in completed_projects:
                item = InvoiceItem(
                    project_number=project.get("project_number", ""),
                    description=project.get("project_name", ""),
                    plant="",
                    quantity=1,
                    unit_price=project.get("project_amount", 0.0)
                )
                self.add_item_row(item)
                loaded_count += 1
            
            # Show success message
            QtWidgets.QMessageBox.information(
                self, "Projects Loaded", 
                f"Successfully loaded {loaded_count} completed projects for {client_name}.",
                QtWidgets.QMessageBox.Ok
            )
            
        except Exception as e:
            log.warning("Error loading client projects: %s", e)
            QtWidgets.QMessageBox.warning(
                self, "Load Error", 
                f"Error loading projects for {client_name}: {str(e)}",
                QtWidgets.QMessageBox.Ok
            )
            # Ensure at least one row exists even if there's an error
            if not self.item_rows:
                self.add_item_row()
            
    def ask_load_client_details(self, client_name):
        """Show simple dialog asking to load client details"""
        reply = QtWidgets.QMessageBox.question(
            self, "Load Client Details",
            f"Load details for client: <b>{client_name}</b>?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        return reply == QtWidgets.QMessageBox.Yes

    def save_client(self):
        """Save current client data to Firebase only with immediate refresh"""
        original_client_name = getattr(self, "editing_client_name", None)
        client_name = self.client_combo.currentText().strip()
        
        if "-- Select Company --" in client_name:
            client_name = client_name.replace("-- Select Company --", "").strip()
        
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Save Client", "Client name is required.")
            return
        
        current_contact = self.client_contact_edit.text().strip() if hasattr(self, "client_contact_edit") else ""
        current_phone = self.client_phone_edit.text().strip() if hasattr(self, "client_phone_edit") else ""
        current_primary_email = self.client_primary_email_edit.text().strip() if hasattr(self, "client_primary_email_edit") else ""
        current_company_email = self.client_email_edit.text().strip()
        current_address = self.client_address_edit.toPlainText().strip()
        client_data = {
            "email": current_primary_email,
            "address": current_address,
            "contact_person": current_contact,
            "phone": current_phone,
            "company_email": current_company_email,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        is_new_client = not original_client_name and client_name not in self.clients
        has_changes = True
        
        existing_key = original_client_name or client_name
        if not is_new_client and existing_key in self.clients:
            existing_client = self.clients[existing_key]
            has_changes = any(
                client_data.get(key, "") != existing_client.get(key, "")
                for key in ["email", "address", "contact_person", "phone", "company_email"]
            ) or client_name != existing_key
        
        if not has_changes and not is_new_client:
            stored_clients = FirebaseManager.load_clients()
            if client_name in stored_clients:
                QtWidgets.QMessageBox.information(self, "Save Client", "No changes detected to save.")
                if hasattr(self, "save_client_btn"):
                    self.save_client_btn.setEnabled(False)
                return
            has_changes = True
        
        self.clients[client_name] = client_data
        
        if FirebaseManager.save_client(client_name, self.clients[client_name]):
            if original_client_name and original_client_name != client_name:
                self._delete_client_storage(original_client_name)
                self.clients.pop(original_client_name, None)
                old_index = self.client_combo.findText(original_client_name)
                if old_index >= 0:
                    self.client_combo.removeItem(old_index)
                if hasattr(self, "line_items_client_combo"):
                    old_line_index = self.line_items_client_combo.findText(original_client_name)
                    if old_line_index >= 0:
                        self.line_items_client_combo.removeItem(old_line_index)

            if self.client_combo.findText(client_name) == -1:
                self.client_combo.addItem(client_name)
            
            self.client_combo.blockSignals(True)
            self.client_combo.setEditText(client_name)
            self.client_combo.blockSignals(False)
            if hasattr(self, "line_items_client_combo"):
                self.line_items_client_combo.blockSignals(True)
                if self.line_items_client_combo.findText(client_name) == -1:
                    self.line_items_client_combo.addItem(client_name)
                self.line_items_client_combo.setEditText(client_name)
                self.line_items_client_combo.blockSignals(False)
            self.update_invoice_client_summary(client_name, current_company_email or current_primary_email, current_address)
            
            # Refresh project tab dropdown
            self.refresh_project_tab_clients()
            self.refresh_inline_clients_table()
            
            if hasattr(self, "save_client_btn"):
                self.save_client_btn.setEnabled(False)
            self.editing_client_name = client_name
            
            message_type = "created" if is_new_client else "updated"
            QtWidgets.QMessageBox.information(
                self, "Save Client", 
                f"Client '{client_name}' {message_type} successfully!"
            )
        else:
            QtWidgets.QMessageBox.critical(self, "Save Client", "Failed to save client to Firebase.")

    def _delete_client_storage(self, client_name: str) -> bool:
        """Delete a client from Firebase or local storage."""
        try:
            if FIREBASE_AVAILABLE:
                from firebase_admin import db
                db.reference('/clients').child(client_name).delete()
                return True

            clients = FirebaseManager.load_clients()
            if client_name in clients:
                clients.pop(client_name, None)
                Config.DATA_DIR.mkdir(parents=True, exist_ok=True)
                with open(Config.CLIENTS_FILE, "w", encoding="utf-8") as f:
                    json.dump(clients, f, indent=2, ensure_ascii=False)
            return True
        except Exception as exc:
            log.warning("Failed to delete client '%s': %s", client_name, exc)
            return False
            
    def verify_client_refresh(self):
        """Verify that the client refresh worked and retry if needed"""
        if hasattr(self, 'project_tab'):
            current_clients = [self.project_tab.company_combo.itemText(i) 
                            for i in range(self.project_tab.company_combo.count())]
            # FIXED: Use client_combo instead of client_name_edit
            new_client = self.client_combo.currentText().strip()
            
            if "-- Select Company --" in new_client:
                new_client = new_client.replace("-- Select Company --", "").strip()
            
            if new_client and new_client not in current_clients:
                log.info("Client not found in dropdown, retrying refresh...")
                self.project_tab.refresh_clients_immediately()

    def setup_client_information_directory(self, layout):
        """Build the polished client directory shown at the top of Client Information."""
        directory = QtWidgets.QFrame()
        directory.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-radius: 9px;
            }
        """)
        outer = QtWidgets.QVBoxLayout(directory)
        outer.setContentsMargins(0, 0, 0, 14)
        outer.setSpacing(0)

        header = QtWidgets.QFrame()
        header.setStyleSheet("""
            QFrame {
                background: transparent;
                border: none;
                border-bottom: 1px solid #e6edf5;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }
        """)
        header_lay = QtWidgets.QHBoxLayout(header)
        header_lay.setContentsMargins(18, 16, 16, 16)
        header_lay.setSpacing(14)

        title_col = QtWidgets.QVBoxLayout()
        title_col.setContentsMargins(0, 0, 0, 0)
        title_col.setSpacing(4)
        title = QtWidgets.QLabel("Client Information")
        title.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                color: #0f172a;
                font-size: 19px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        subtitle = QtWidgets.QLabel("Manage your client details and contact information.")
        subtitle.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                color: #53657d;
                font-size: 12px;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        title_col.addWidget(title)
        title_col.addWidget(subtitle)
        header_lay.addLayout(title_col, 1)

        export_btn = QtWidgets.QPushButton("Export")
        export_btn.setFixedSize(100, 42)
        export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        export_btn.setStyleSheet("""
            QPushButton {
                background: #ffffff;
                color: #334155;
                border: 1px solid #dfe7f0;
                border-radius: 7px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover {
                background: #ffffff;
                border-color: #cbd5e1;
            }
        """)
        export_btn.clicked.connect(self.export_clients_csv)
        header_lay.addWidget(export_btn)

        new_btn = QtWidgets.QPushButton("+  New Client")
        new_btn.setFixedSize(126, 42)
        new_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        new_btn.setStyleSheet("""
            QPushButton {
                background: #00756f;
                color: #ffffff;
                border: none;
                border-radius: 7px;
                font-size: 13px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background: #00645f; }
        """)
        new_btn.clicked.connect(self.open_new_client_popup_dialog)
        header_lay.addWidget(new_btn)
        outer.addWidget(header)

        toolbar = QtWidgets.QFrame()
        toolbar.setStyleSheet("QFrame { background:transparent; border:none; }")
        toolbar_lay = QtWidgets.QHBoxLayout(toolbar)
        toolbar_lay.setContentsMargins(18, 14, 16, 12)
        toolbar_lay.setSpacing(12)

        self.client_info_search = QtWidgets.QLineEdit()
        self.client_info_search.setPlaceholderText("Search by client name, contact, phone, or email...")
        self.client_info_search.setFixedHeight(38)
        self.client_info_search.setMinimumWidth(360)
        self.client_info_search.setStyleSheet("""
            QLineEdit {
                background: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 7px;
                padding: 0 12px;
                font-size: 13px;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QLineEdit:focus {
                border-color: #00756f;
            }
        """)
        self._client_search_timer = QtCore.QTimer(self)
        self._client_search_timer.setSingleShot(True)
        self._client_search_timer.setInterval(200)
        self._client_search_timer.timeout.connect(self.refresh_client_information_directory)
        self.client_info_search.textChanged.connect(self._client_search_timer.start)
        toolbar_lay.addWidget(self.client_info_search, 1)

        self.client_count_lbl = QtWidgets.QLabel("Clients: 0")
        self.client_count_lbl.setStyleSheet("""
            QLabel {
                background: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 7px;
                color: #475569;
                font-size: 13px;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 14px;
            }
        """)
        self.client_count_lbl.setFixedHeight(38)
        self.client_count_lbl.setMinimumWidth(90)
        toolbar_lay.addWidget(self.client_count_lbl)
        outer.addWidget(toolbar)

        self.client_info_table = QtWidgets.QTableWidget()
        self.client_info_table.setColumnCount(7)
        self.client_info_table.setHorizontalHeaderLabels([
            "Client", "Contact Person", "Phone", "Email", "Company Email", "Address", "Actions"
        ])
        self.client_info_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.client_info_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.client_info_table.verticalHeader().setVisible(False)
        self.client_info_table.setAlternatingRowColors(False)
        self.client_info_table.setShowGrid(False)
        self.client_info_table.setMinimumHeight(640)
        self.client_info_table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                border: 1px solid #e6edf5;
                border-radius: 7px;
                gridline-color: #edf2f7;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QTableWidget::item {
                padding: 8px 10px;
                color: #10213a;
                border-bottom: 1px solid #edf2f7;
                border-right: 1px solid #edf2f7;
            }
            QHeaderView::section {
                background: #ffffff;
                color: #53657d;
                border: none;
                border-bottom: 1px solid #dfe7f0;
                border-right: 1px solid #dfe7f0;
                padding: 9px 10px;
                font-size: 11px;
                font-weight: 900;
                text-transform: uppercase;
            }
        """)
        header_view = self.client_info_table.horizontalHeader()
        for col in range(6):
            header_view.setSectionResizeMode(col, QtWidgets.QHeaderView.Stretch)
        header_view.setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)
        self.client_info_table.setColumnWidth(6, 156)
        outer.addWidget(self.client_info_table)

        footer = QtWidgets.QHBoxLayout()
        footer.setContentsMargins(18, 12, 16, 0)
        footer.setSpacing(10)
        self.client_info_footer = QtWidgets.QLabel("Showing 0 of 0 clients")
        self.client_info_footer.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                color: #53657d;
                font-size: 12px;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        footer.addWidget(self.client_info_footer, 1)

        for text in ("<", "1", ">"):
            pager = QtWidgets.QPushButton(text)
            pager.setFixedSize(34, 34)
            pager.setEnabled(text == "1")
            pager.setStyleSheet("""
                QPushButton {
                    background: #ffffff;
                    color: #64748b;
                    border: 1px solid #e2e8f0;
                    border-radius: 7px;
                    font-size: 13px;
                    font-weight: 900;
                }
                QPushButton:enabled {
                    color: #ffffff;
                    background: #00756f;
                    border-color: #00756f;
                }
            """)
            footer.addWidget(pager)
        outer.addLayout(footer)

        layout.addWidget(directory)
        self.refresh_client_information_directory()

    def _filtered_clients(self, search_text=""):
        self.clients = FirebaseManager.load_clients()
        query = (search_text or "").strip().lower()
        records = []
        for client_name, client_data in getattr(self, "clients", {}).items():
            searchable = " ".join([
                client_name,
                client_data.get("contact_person", ""),
                client_data.get("phone", ""),
                client_data.get("email", ""),
                client_data.get("company_email", ""),
                client_data.get("address", "")
            ]).lower()
            if query and query not in searchable:
                continue
            records.append((client_name, client_data))
        records.sort(key=lambda item: item[0].lower())
        return records, len(getattr(self, "clients", {}))

    def _client_manage_widget(self, client_name):
        actions = QtWidgets.QWidget()
        actions_layout = QtWidgets.QHBoxLayout(actions)
        actions_layout.setContentsMargins(0, 0, 0, 0)
        actions_layout.setSpacing(8)
        actions_layout.setAlignment(QtCore.Qt.AlignCenter)

        manage_btn = QtWidgets.QToolButton()
        manage_btn.setText("Manage")
        manage_btn.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        manage_btn.setFixedSize(84, 30)
        manage_btn.setStyleSheet("""
            QToolButton {
                background: #ffffff;
                color: #0f766e;
                border: 1px solid #d6e4e7;
                border-radius: 5px;
                font-size: 11px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QToolButton:hover {
                border-color: #00756f;
                background: #f0fdfa;
            }
            QToolButton::menu-indicator {
                image: none;
                width: 0px;
            }
        """)

        menu = QtWidgets.QMenu(manage_btn)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                padding: 4px;
            }
            QMenu::item {
                padding: 8px 24px;
                color: #0f172a;
                font-weight: 700;
            }
            QMenu::item:selected {
                background: #eefaf8;
                color: #00756f;
            }
        """)
        edit_action = menu.addAction("Edit")
        delete_action = menu.addAction("Delete")
        edit_action.triggered.connect(lambda checked=False, name=client_name: self.edit_inline_client(name))
        delete_action.triggered.connect(lambda checked=False, name=client_name: self.delete_inline_client(name))
        manage_btn.setMenu(menu)

        actions_layout.addWidget(manage_btn)
        return actions

    def _client_identity_widget(self, client_name):
        wrap = QtWidgets.QWidget()
        lay = QtWidgets.QHBoxLayout(wrap)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        initials = "".join(part[:1] for part in client_name.split()[:2]).upper() or "C"
        color_sets = [
            ("#d1fae5", "#0f766e"),
            ("#dbeafe", "#2563eb"),
            ("#fef3c7", "#d97706"),
            ("#ede9fe", "#7c3aed"),
            ("#fce7f3", "#db2777"),
        ]
        bg, fg = color_sets[sum(ord(ch) for ch in client_name) % len(color_sets)]

        badge = QtWidgets.QLabel(initials)
        badge.setFixedSize(34, 34)
        badge.setAlignment(QtCore.Qt.AlignCenter)
        badge.setStyleSheet(f"""
            QLabel {{
                background: {bg};
                color: {fg};
                border: none;
                border-radius: 17px;
                font-size: 11px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }}
        """)
        lay.addWidget(badge)

        name = QtWidgets.QLabel(client_name)
        name.setWordWrap(True)
        name.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                color: #10213a;
                font-size: 12px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
        """)
        lay.addWidget(name, 1)
        return wrap

    def _populate_client_table(self, table, clients_list):
        table.setUpdatesEnabled(False)
        table.blockSignals(True)
        table.setRowCount(len(clients_list))
        table.setColumnWidth(6, 156 if table is getattr(self, "client_info_table", None) else 170)
        for row, (client_name, client_data) in enumerate(clients_list):
            values = [
                client_data.get("contact_person", "") or "-",
                client_data.get("phone", "") or "-",
                client_data.get("email", "") or "-",
                client_data.get("company_email", "") or "-",
                (client_data.get("address", "") or "-").replace("\n", ", "),
            ]
            table.setCellWidget(row, 0, self._client_identity_widget(client_name))
            for col, value in enumerate(values, start=1):
                item = QtWidgets.QTableWidgetItem(value)
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                table.setItem(row, col, item)
            table.setCellWidget(row, 6, self._client_manage_widget(client_name))
            table.setRowHeight(row, 58 if table is getattr(self, "client_info_table", None) else 52)
        table.blockSignals(False)
        table.setUpdatesEnabled(True)

    def refresh_client_information_directory(self):
        """Refresh the Client Information directory panel."""
        if not hasattr(self, "client_info_table"):
            return
        search = self.client_info_search.text() if hasattr(self, "client_info_search") else ""
        clients_list, total = self._filtered_clients(search)
        shown = len(clients_list)
        self._populate_client_table(self.client_info_table, clients_list)
        if hasattr(self, "client_info_count_btn"):
            self.client_info_count_btn.setText(f"{total} Client{'s' if total != 1 else ''}")
        if hasattr(self, "client_info_footer"):
            self.client_info_footer.setText(f"Showing {shown} of {total} client{'s' if total != 1 else ''}")
        if hasattr(self, "client_count_lbl"):
            self.client_count_lbl.setText(f"Clients: {total}")

    def export_clients_csv(self):
        """Export the currently visible client directory rows to Excel."""
        search = self.client_info_search.text() if hasattr(self, "client_info_search") else ""
        clients_list, _total = self._filtered_clients(search)
        export_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Export Clients",
            "clients_export.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not export_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clients"

            ws.merge_cells('A1:G1')
            ws['A1'] = f"{Config.COMPANY.get('name', 'MABS Engineering LLC').upper()} - CLIENT DIRECTORY"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            ws['A2'] = f"Generated: {__import__('datetime').datetime.now().strftime('%m-%d-%Y')}"
            ws['A2'].font = Font(bold=True)

            headers = ["S.No.", "Client", "Contact Person", "Phone", "Email", "Company Email", "Address"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            for row_idx, (client_name, client_data) in enumerate(clients_list, 5):
                data = [
                    row_idx - 4,
                    client_name,
                    client_data.get("contact_person", ""),
                    client_data.get("phone", ""),
                    client_data.get("email", ""),
                    client_data.get("company_email", ""),
                    (client_data.get("address", "") or "").replace("\n", ", "),
                ]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    if col == 1:
                        cell.font = Font(bold=True)

            column_widths = {1: 8, 2: 35, 3: 25, 4: 16, 5: 35, 6: 35, 7: 45}
            for col_idx, width in column_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

            for row in range(5, ws.max_row + 1):
                if row % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.fill.start_color.index == '00000000':
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

            wb.save(export_path)
            QtWidgets.QMessageBox.information(self, "Export Clients", "Client export completed successfully.")
        except Exception as exc:
            log.warning("Client Excel export failed: %s", exc)
            QtWidgets.QMessageBox.critical(self, "Export Clients", f"Failed to export clients: {exc}")

    def setup_inline_client_list_tab(self):
        """Build the inline saved-client list tab."""
        layout = QtWidgets.QVBoxLayout(self.client_list_tab)
        layout.setContentsMargins(20, 18, 20, 20)
        layout.setSpacing(14)

        header = QtWidgets.QFrame()
        header.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
            }
        """)
        header_layout = QtWidgets.QHBoxLayout(header)
        header_layout.setContentsMargins(24, 22, 24, 22)

        title_wrap = QtWidgets.QVBoxLayout()
        title = QtWidgets.QLabel("Saved Clients")
        title.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                font-size: 24px;
                font-weight: 900;
                color: #0f172a;
            }
        """)
        subtitle = QtWidgets.QLabel("Search saved clients, edit details, or remove old records.")
        subtitle.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                font-size: 14px;
                font-weight: 700;
                color: #53657d;
            }
        """)
        title_wrap.addWidget(title)
        title_wrap.addWidget(subtitle)
        header_layout.addLayout(title_wrap)
        header_layout.addStretch()

        new_client_btn = QtWidgets.QPushButton("New Client")
        new_client_btn.setFixedSize(360, 180)
        new_client_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        new_client_btn.setStyleSheet("""
            QPushButton {
                background: #00756f;
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: 900;
                font-size: 28px;
                padding: 0 38px;
            }
            QPushButton:hover { background: #00645f; }
        """)
        new_client_btn.clicked.connect(self.prepare_new_client_inline)
        header_layout.addWidget(new_client_btn)

        toolbar = QtWidgets.QFrame()
        toolbar.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
            }
        """)
        toolbar_layout = QtWidgets.QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(14, 12, 14, 12)

        self.inline_client_search = QtWidgets.QLineEdit()
        self.inline_client_search.setPlaceholderText("Search by client, contact, phone, or email...")
        self.inline_client_search.setMinimumHeight(38)
        self.inline_client_search.setStyleSheet("""
            QLineEdit {
                background: #ffffff;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 13px;
                color: #0f172a;
            }
            QLineEdit:focus { border-color: #00756f; }
        """)
        self.inline_client_search.textChanged.connect(self.refresh_inline_clients_table)
        toolbar_layout.addWidget(self.inline_client_search, 1)

        self.client_count_btn = QtWidgets.QPushButton("0 Clients")
        self.client_count_btn.setFixedHeight(38)
        self.client_count_btn.setMinimumWidth(110)
        self.client_count_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.client_count_btn.setStyleSheet("""
            QPushButton {
                background: #f0fdf4;
                color: #00756f;
                border: 1.5px solid #6ee7b7;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 16px;
            }
            QPushButton:hover {
                background: #dcfce7;
                border-color: #00756f;
            }
        """)
        self.client_count_btn.setToolTip("Total saved clients")
        self.client_count_btn.clicked.connect(lambda: self.inline_client_search.clear())
        toolbar_layout.addWidget(self.client_count_btn)

        self.inline_clients_table = QtWidgets.QTableWidget()
        self.inline_clients_table.setColumnCount(7)
        self.inline_clients_table.setHorizontalHeaderLabels([
            "Client", "Contact", "Phone", "Primary Email", "Company Email", "Address", "Actions"
        ])
        self.inline_clients_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.inline_clients_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.inline_clients_table.verticalHeader().setVisible(False)
        self.inline_clients_table.setAlternatingRowColors(True)
        self.inline_clients_table.setStyleSheet("""
            QTableWidget {
                background: transparent;
                border: none;
                border-radius: 8px;
                gridline-color: #e5edf5;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 8px;
                color: #0f172a;
            }
            QHeaderView::section {
                background: #1f2937;
                color: white;
                border: none;
                padding: 10px 8px;
                font-size: 12px;
                font-weight: 900;
            }
        """)
        table_header = self.inline_clients_table.horizontalHeader()
        for col in range(6):
            table_header.setSectionResizeMode(col, QtWidgets.QHeaderView.Stretch)
        table_header.setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)
        self.inline_clients_table.setColumnWidth(6, 170)

        layout.addWidget(header)
        layout.addWidget(toolbar)
        layout.addWidget(self.inline_clients_table, 1)
        self.refresh_inline_clients_table()

    def refresh_inline_clients_table(self):
        """Refresh the inline saved-client list."""
        if not hasattr(self, "inline_clients_table"):
            return

        self.clients = FirebaseManager.load_clients()

        search = ""
        if hasattr(self, "inline_client_search"):
            search = self.inline_client_search.text().strip().lower()

        clients_list = []
        for client_name, client_data in getattr(self, "clients", {}).items():
            searchable = " ".join([
                client_name,
                client_data.get("contact_person", ""),
                client_data.get("phone", ""),
                client_data.get("email", ""),
                client_data.get("company_email", ""),
                client_data.get("address", "")
            ]).lower()
            if search and search not in searchable:
                continue
            clients_list.append((client_name, client_data))

        clients_list.sort(key=lambda item: item[0].lower())
        total = len(self.clients) if hasattr(self, 'clients') else 0
        shown = len(clients_list)
        if hasattr(self, 'client_count_btn'):
            if search and shown != total:
                self.client_count_btn.setText(f"{shown} / {total} Clients")
            else:
                self.client_count_btn.setText(f"{total} Client{'s' if total != 1 else ''}")
        self.inline_clients_table.setRowCount(shown)
        self.inline_clients_table.setColumnWidth(6, 170)

        for row, (client_name, client_data) in enumerate(clients_list):
            values = [
                client_name,
                client_data.get("contact_person", "") or "-",
                client_data.get("phone", "") or "-",
                client_data.get("email", "") or "-",
                client_data.get("company_email", "") or "-",
                (client_data.get("address", "") or "-").replace("\n", ", "),
            ]
            for col, value in enumerate(values):
                item = QtWidgets.QTableWidgetItem(value)
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.inline_clients_table.setItem(row, col, item)

            actions = QtWidgets.QWidget()
            actions_layout = QtWidgets.QHBoxLayout(actions)
            actions_layout.setContentsMargins(0, 0, 0, 0)
            actions_layout.setSpacing(0)
            actions_layout.setAlignment(QtCore.Qt.AlignCenter)

            manage_btn = QtWidgets.QToolButton()
            manage_btn.setText("Manage")
            manage_btn.setPopupMode(QtWidgets.QToolButton.InstantPopup)
            manage_btn.setFixedSize(112, 26)
            manage_btn.setStyleSheet("""
                QToolButton {
                    background: #eefaf8;
                    color: #0f172a;
                    border: 1px solid #9ddbd4;
                    border-radius: 5px;
                    font-weight: 900;
                }
                QToolButton:hover {
                    border-color: #00756f;
                    color: #00756f;
                    background: #dcf7f2;
                }
                QToolButton::menu-indicator {
                    image: none;
                    width: 0px;
                }
            """)

            menu = QtWidgets.QMenu(manage_btn)
            menu.setStyleSheet("""
                QMenu {
                    background: #ffffff;
                    border: 1px solid #d8e2ec;
                    border-radius: 8px;
                    padding: 4px;
                }
                QMenu::item {
                    padding: 8px 24px;
                    color: #0f172a;
                    font-weight: 700;
                }
                QMenu::item:selected {
                    background: #eefaf8;
                    color: #00756f;
                }
            """)
            edit_action = menu.addAction("Edit")
            delete_action = menu.addAction("Delete")
            edit_action.triggered.connect(lambda checked=False, name=client_name: self.edit_inline_client(name))
            delete_action.triggered.connect(lambda checked=False, name=client_name: self.delete_inline_client(name))
            manage_btn.setMenu(menu)

            actions_layout.addWidget(manage_btn)
            self.inline_clients_table.setCellWidget(row, 6, actions)
            self.inline_clients_table.setRowHeight(row, 52)

        self.refresh_client_information_directory()

    def edit_inline_client(self, client_name):
        """Edit a saved client via the popup dialog."""
        self.open_client_popup_dialog(client_name=client_name)

    def delete_inline_client(self, client_name):
        """Delete a saved client from the inline list."""
        reply = QtWidgets.QMessageBox.question(
            self,
            "Delete Client",
            f"Delete client '{client_name}'?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return

        if not self._delete_client_storage(client_name):
            QtWidgets.QMessageBox.critical(self, "Delete Client", f"Failed to delete client: {client_name}")
            return

        self.clients.pop(client_name, None)
        index = self.client_combo.findText(client_name)
        if index >= 0:
            self.client_combo.removeItem(index)
        self.refresh_project_tab_clients()
        self.refresh_inline_clients_table()
        QtWidgets.QMessageBox.information(self, "Delete Client", f"Client '{client_name}' deleted successfully.")
    
    def open_client_info_dialog(self):
        """Route legacy client-management actions to the inline Client Information tab."""
        self.prepare_new_client_inline()
        return

    def refresh_project_tab_clients(self):
        """Refresh the company dropdown in project tab"""
        try:
            if hasattr(self, 'project_tab') and hasattr(self.project_tab, 'company_combo'):
                # Store current selection
                current_selection = self.project_tab.company_combo.currentText()
                
                # Clear and repopulate
                self.project_tab.company_combo.blockSignals(True)
                self.project_tab.company_combo.clear()
                self.project_tab.company_combo.addItem("-- Select Company --")
                
                # Add sorted clients
                for client in sorted(self.clients.keys()):
                    self.project_tab.company_combo.addItem(client)
                
                # Restore selection if possible
                if current_selection and current_selection in self.clients:
                    index = self.project_tab.company_combo.findText(current_selection)
                    if index >= 0:
                        self.project_tab.company_combo.setCurrentIndex(index)
                
                self.project_tab.company_combo.blockSignals(False)
                
                log.info("Project tab clients refreshed with %s clients", len(self.clients))
                
        except Exception as e:
            log.warning("Error refreshing project tab clients: %s", e)
               
    def _update_pdf_btn_state(self, *_):
        """Enable Generate PDF only when every item row has project number, project name,
        a valid amount (> 0), and a payment category selected."""
        if not hasattr(self, 'generate_pdf_btn'):
            return
        # Always enable when editing an existing invoice
        if getattr(self, '_editing_existing_invoice', False):
            self.generate_pdf_btn.setEnabled(True)
            self.generate_pdf_btn.setStyleSheet("""
                QPushButton {
                    background-color: #00756f; color: white; border: none;
                    border-radius: 10px; padding: 0 20px; font-weight: 800;
                    font-size: 17px; min-height: 52px; max-height: 52px;
                }
                QPushButton:hover  { background-color: #00645f; }
                QPushButton:pressed{ background-color: #00514d; }
            """)
            return
        rows = getattr(self, 'item_rows', [])
        enabled = False
        if rows:
            enabled = True
            for row in rows:
                # Skip locked stage rows — their values are pre-filled and immutable
                if getattr(row, '_stage_locked', False):
                    continue
                project_number = row.project_number_edit.text().strip() if row.project_number_edit else ""
                project_name   = row.desc_edit.text().strip()           if row.desc_edit            else ""
                payment_cat_ok = row.down_payment_combo.currentIndex() != 0 if row.down_payment_combo else False
                amount_raw     = row.price_edit.text().replace("$", "").replace(",", "").strip() if row.price_edit else ""
                try:
                    amount_ok = float(amount_raw) > 0 if amount_raw else False
                except ValueError:
                    amount_ok = False
                if not (project_number and project_name and amount_ok and payment_cat_ok):
                    enabled = False
                    break

        self.generate_pdf_btn.setEnabled(enabled)
        if enabled:
            self.generate_pdf_btn.setStyleSheet("""
                QPushButton {
                    background-color: #00756f; color: white; border: none;
                    border-radius: 10px; padding: 0 20px; font-weight: 800;
                    font-size: 17px; min-height: 52px; max-height: 52px;
                }
                QPushButton:hover  { background-color: #00645f; }
                QPushButton:pressed{ background-color: #00514d; }
            """)
        else:
            self.generate_pdf_btn.setStyleSheet("""
                QPushButton {
                    background-color: #94a3b8; color: #e2e8f0; border: none;
                    border-radius: 10px; padding: 0 20px; font-weight: 800;
                    font-size: 17px; min-height: 52px; max-height: 52px;
                }
            """)

    def generate_pdf(self):
        """Generate PDF invoice and save to Firebase only - ADD revenue tracking"""
        self.update_invoice_from_form()
        # Recalculate status after form update (important when tax/total changes)
        if getattr(self, '_editing_existing_invoice', False):
            self.update_totals()

        # Validate required fields
        client_name = self.invoice.client_name.strip()
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "Client name is required.")
            return
        
        if not self.invoice.items:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "At least one invoice item is required.")
            return
        
        # Validate due date is not before invoice date
        if hasattr(self, 'due_date_edit') and hasattr(self, 'date_edit'):
            if self.due_date_edit.date() < self.date_edit.date():
                QtWidgets.QMessageBox.warning(
                    self, "Validation Error",
                    "Due date cannot be before the invoice date."
                )
                self.due_date_edit.setFocus()
                return

        # Validate that all items have a valid payment category selected (not placeholder)
        for i, row in enumerate(self.item_rows, 1):
            # Locked rows have a single stage label at index 0 — that IS valid
            if getattr(row, "_stage_locked", False):
                continue
            if row.down_payment_combo.currentIndex() == 0:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Validation Error",
                    "Please select a payment category for all items."
                )
                row.down_payment_combo.setFocus()
                return

        # Auto-save new client if needed
        if client_name not in self.clients:
            email = self.client_email_edit.text().strip()
            address = self.client_address_edit.toPlainText().strip()
            if self.auto_save_new_client(client_name, email, address):
                log.info("New client '%s' auto-saved before invoice generation", client_name)

        missing_client_fields = []
        if not self.client_email_edit.text().strip():
            missing_client_fields.append("email")
        if not self.client_address_edit.toPlainText().strip():
            missing_client_fields.append("address")
        if missing_client_fields:
            reply = QtWidgets.QMessageBox.question(
                self,
                "Client Information Missing",
                "Client " + " and ".join(missing_client_fields) + " is missing.\n\n"
                "Do you want to continue generating the invoice PDF anyway?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if reply != QtWidgets.QMessageBox.Yes:
                self.show_client_information_tab()
                return
        
        # Set final invoice number based on current date (not invoice date). Reuse if already set.
        if not self.invoice.invoice_number:
            from datetime import datetime as _dt
            today_str = _dt.now().strftime("%Y-%m-%d")
            self.invoice.invoice_number = InvoiceNumberGenerator.get_next_number(today_str)
            self.invoice_no_edit.setText(self.invoice.invoice_number)
        
        # Generate PDF with wait cursor
        temp_dir = Path(tempfile.gettempdir()) / "mabs_invoices_temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = temp_dir / f"{self.invoice.invoice_number}.pdf"

        try:
            logo_path = Config.get_logo_path()

            # Generate PDF locally with wait cursor (no network needed)
            QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
            QtWidgets.QApplication.processEvents()
            try:
                success = PDFGenerator.generate(self.invoice, pdf_path, logo_path)
            finally:
                QtWidgets.QApplication.restoreOverrideCursor()

            if success:
                _saved_inv_number = self.invoice.invoice_number
                _saved_inv_items  = list(self.invoice.items)

                invoice_data = self.invoice.to_dict()
                if not getattr(self, '_editing_existing_invoice', False):
                    invoice_data['meta']['status'] = 'Unpaid'
                    invoice_data['meta']['received_date'] = 'N/A'
                else:
                    # Preserve original created_at and firebase_id when editing (so invoice stays in original position)
                    if hasattr(self, '_original_created_at') and self._original_created_at:
                        invoice_data['meta']['created_at'] = self._original_created_at
                    if hasattr(self, '_original_firebase_id') and self._original_firebase_id:
                        invoice_data['firebase_id'] = self._original_firebase_id
                _inv_total_fmt = Currency.format(self.invoice.total)
                _inv_copy = Invoice.from_dict(invoice_data)

                # Clear form BEFORE showing popup
                self.tax_spin.setValue(float(Config.COMPANY.get("default_tax_rate", 0.0)))
                self.notes_edit.setPlainText(Config.DEFAULT_TERMS)
                self.clear_all_items()
                self.clear_client_information()
                self.invoice.invoice_number = ""
                self.invoice_no_edit.setText("")
                # Add one empty line item immediately
                self.add_item_row()

                log.info("Invoice %s generated", _saved_inv_number)

                # Show popup with BLOCKING dialog (user must close before other interactions)
                is_editing = getattr(self, '_editing_existing_invoice', False)
                title_text = "Invoice Saved & Updated Successfully!" if is_editing else "Invoice Saved & Generated Successfully!"
                message = (
                    f"<div style='text-align:center;'>"
                    f"<h3 style='color:#28a745;margin-bottom:15px;'>{title_text}</h3>"
                    f"<p style='font-size:14px;color:#495057;line-height:1.6;'>"
                    f"<b>Invoice Number:</b> {_saved_inv_number}<br>"
                    f"<b>Client:</b> {client_name}<br>"
                    f"<b>Amount:</b> {_inv_total_fmt}"
                    f"</p></div>"
                )

                msg_box = QtWidgets.QMessageBox(self)
                msg_box.setWindowTitle("Success")
                msg_box.setTextFormat(QtCore.Qt.RichText)
                msg_box.setText(message)
                msg_box.setIcon(QtWidgets.QMessageBox.Information)
                ok_btn = QtWidgets.QPushButton("OK")
                open_btn = QtWidgets.QPushButton("📄 Open PDF")
                msg_box.addButton(ok_btn, QtWidgets.QMessageBox.AcceptRole)
                msg_box.addButton(open_btn, QtWidgets.QMessageBox.ActionRole)
                msg_box.setDefaultButton(ok_btn)
                msg_box.setEscapeButton(ok_btn)
                msg_box.exec_()

                if msg_box.clickedButton() == open_btn:
                    FileManager.open_file(pdf_path)

                # Background: Firebase saves only (no Qt calls allowed)
                def _bg_save(inv_data=invoice_data, inv_num=_saved_inv_number,
                              inv_copy=_inv_copy, _pdf_path=pdf_path):
                    try:
                        FirebaseManager.save_invoice(inv_data)
                    except Exception as _e:
                        log.warning("bg save_invoice error: %s", _e)
                    try:
                        FirebaseManager.save_pdf_to_firebase(inv_num, _pdf_path)
                    except Exception as _e:
                        log.warning("bg save_pdf_to_firebase error: %s", _e)
                    try:
                        self.save_invoice_as_revenue(inv_copy, _skip_ui_refresh=True)
                    except Exception as _e:
                        log.warning("bg save_invoice_as_revenue error: %s", _e)

                import threading as _threading
                _threading.Thread(target=_bg_save, daemon=True).start()

                # Post-save UI refresh and syncs (main thread only)
                self._trigger_post_save_payment_sync(_saved_inv_number, _saved_inv_items)
                self.update_invoice_preview()

                # Live update: Update invoice in-memory immediately (handles both new and edited)
                try:
                    if hasattr(self, 'history_tab'):
                        current_widget = self.history_tab.stacked_widget.currentWidget() if hasattr(self.history_tab, 'stacked_widget') else None
                        if current_widget and hasattr(current_widget, 'invoices'):
                            updated_inv = Invoice.from_dict(invoice_data)
                            client_name = invoice_data.get('meta', {}).get('client_name', '')

                            # Check if invoice already exists (editing) or new
                            found = False
                            for i, (inv, json_file) in enumerate(current_widget.invoices):
                                if inv.invoice_number == _saved_inv_number:
                                    # Update existing invoice (editing) - preserve firebase_timestamp so it stays in place
                                    original_timestamp = getattr(inv, 'firebase_timestamp', None)
                                    if original_timestamp is not None:
                                        updated_inv.firebase_timestamp = original_timestamp
                                    current_widget.invoices[i] = (updated_inv, json_file)
                                    found = True
                                    break

                            if not found and client_name == current_widget.client_name:
                                # Add new invoice to list (for current client only)
                                from datetime import datetime, timezone
                                timestamp = datetime.now(timezone.utc).timestamp()
                                updated_inv.firebase_timestamp = timestamp
                                current_widget.invoices.append((updated_inv, None))  # Add to list, sort will order it

                            # Refresh table display immediately (live, like balance sheet)
                            QtCore.QTimer.singleShot(0, current_widget.apply_all_time_filter)
                except Exception as _e:
                    log.warning("Live update error: %s", _e)
                if hasattr(self, 'balance_sheet_tab'):
                    QtCore.QTimer.singleShot(1500, self.balance_sheet_tab.load_invoice_revenues)
                    QtCore.QTimer.singleShot(1500, self.balance_sheet_tab.update_annual_summary)
                QtCore.QTimer.singleShot(8000, lambda: self.cleanup_temp_file(pdf_path))

            else:
                QtWidgets.QMessageBox.critical(self, "PDF Generation Error", "Failed to generate PDF.")
        
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", f"Failed to generate PDF: {str(e)}")

    def save_invoice(self):
        """Save invoice record in background - instant form clear, responsive UI."""
        self.update_invoice_from_form()

        client_name = self.invoice.client_name.strip()
        if not client_name:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "Client name is required.")
            return

        if not self.invoice.items:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "At least one invoice item is required.")
            return

        if hasattr(self, 'due_date_edit') and hasattr(self, 'date_edit'):
            if self.due_date_edit.date() < self.date_edit.date():
                QtWidgets.QMessageBox.warning(
                    self, "Validation Error",
                    "Due date cannot be before the invoice date."
                )
                self.due_date_edit.setFocus()
                return

        for row in self.item_rows:
            if getattr(row, "_stage_locked", False):
                continue
            if row.down_payment_combo.currentIndex() == 0:
                QtWidgets.QMessageBox.warning(
                    self, "Validation Error",
                    "Please select a payment category for all items before saving the invoice."
                )
                row.down_payment_combo.setFocus()
                return

        if client_name not in self.clients:
            email = self.client_email_edit.text().strip()
            address = self.client_address_edit.toPlainText().strip()
            if self.auto_save_new_client(client_name, email, address):
                log.info("New client '%s' auto-saved before invoice save", client_name)

        if not self.invoice.invoice_number:
            from datetime import datetime as _dt
            today_str = _dt.now().strftime("%Y-%m-%d")
            self.invoice.invoice_number = InvoiceNumberGenerator.get_next_number(today_str)
            self.invoice_no_edit.setText(self.invoice.invoice_number)

        invoice_data = self.invoice.to_dict()
        if not getattr(self, '_editing_existing_invoice', False):
            invoice_data['meta']['status'] = 'Unpaid'
        invoice_data['meta']['received_date'] = self.invoice.received_date or 'N/A'

        saved_number = self.invoice.invoice_number
        saved_items  = list(self.invoice.items)
        invoice_snapshot = Invoice.from_dict(invoice_data)
        _save_total_fmt = Currency.format(self.invoice.total) if hasattr(self, 'invoice') else "$0.00"

        # Show saved notification FIRST - popup appears instantly
        # All other work happens in background so UI remains responsive
        msg = (
            f"<div style='text-align:center;'>"
            f"<h3 style='color:#15803d;margin-bottom:10px;'>✓ Invoice Saved</h3>"
            f"<p style='font-size:12px;color:#495057;line-height:1.5;margin:10px 0;'>"
            f"<b>{saved_number}</b><br>"
            f"{client_name}<br>"
            f"{_save_total_fmt}"
            f"</p></div>"
        )
        msg_box = QtWidgets.QMessageBox(self)
        msg_box.setWindowTitle("Invoice Saved")
        msg_box.setTextFormat(QtCore.Qt.RichText)
        msg_box.setText(msg)
        msg_box.setIcon(QtWidgets.QMessageBox.Information)
        msg_box.setWindowModality(QtCore.Qt.NonModal)
        msg_box.setStandardButtons(QtWidgets.QMessageBox.Ok)

        # Auto-close after 3 seconds
        _timer = QtCore.QTimer()
        def _auto_close():
            if msg_box.isVisible():
                msg_box.close()
        _timer.setSingleShot(True)
        _timer.timeout.connect(_auto_close)
        _timer.start(3000)
        msg_box.show()

        log.info("Invoice %s saved", saved_number)

        # Background: form reset + save invoice + PDF + revenue (all non-blocking)
        def _bg_save(inv_data=invoice_data, inv_num=saved_number,
                     inv_copy=invoice_snapshot, inv_items=saved_items):
            # Schedule form clearing on main thread (non-blocking, happens after popup shows)
            try:
                from datetime import datetime as _dt
                QtCore.QTimer.singleShot(100, lambda: self.tax_spin.setValue(float(Config.COMPANY.get("default_tax_rate", 0.0))))
                QtCore.QTimer.singleShot(100, lambda: self.notes_edit.setPlainText(Config.DEFAULT_TERMS))
                QtCore.QTimer.singleShot(100, self.clear_all_items)
                QtCore.QTimer.singleShot(100, self.clear_client_information)
                next_inv_number = InvoiceNumberGenerator.get_next_number(_dt.now().strftime("%Y-%m-%d"))
                QtCore.QTimer.singleShot(100, lambda: setattr(self.invoice, 'invoice_number', next_inv_number))
                QtCore.QTimer.singleShot(100, lambda: self.invoice_no_edit.setText(next_inv_number))
                QtCore.QTimer.singleShot(200, self.add_item_row)
            except Exception as _e:
                log.warning("Form reset schedule error: %s", _e)

            # Save to Firebase
            try:
                FirebaseManager.save_invoice(inv_data)
            except Exception as _e:
                log.warning("bg save_invoice error: %s", _e)

            # Generate and save PDF
            temp_dir = Path(tempfile.gettempdir()) / "mabs_invoices_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            pdf_path = temp_dir / f"{inv_num}.pdf"
            try:
                logo_path = Config.get_logo_path()
                PDFGenerator.generate(inv_copy, pdf_path, logo_path)
                try:
                    FirebaseManager.save_pdf_to_firebase(inv_num, pdf_path)
                except Exception as _e:
                    log.warning("bg save_pdf_to_firebase error: %s", _e)
            except Exception as _e:
                log.warning("bg PDF generation error: %s", _e)

            try:
                self.save_invoice_as_revenue(inv_copy, _skip_ui_refresh=True)
            except Exception as _e:
                log.warning("bg save_invoice_as_revenue error: %s", _e)

        import threading as _threading
        _threading.Thread(target=_bg_save, daemon=True).start()

        # Schedule UI refresh (main thread only)
        self._trigger_post_save_payment_sync(saved_number, saved_items)
        if hasattr(self, 'history_tab'):
            QtCore.QTimer.singleShot(1000, self.history_tab.refresh_invoices_immediately)
        if hasattr(self, 'balance_sheet_tab'):
            QtCore.QTimer.singleShot(1500, self.balance_sheet_tab.load_invoice_revenues)
            QtCore.QTimer.singleShot(1500, self.balance_sheet_tab.update_annual_summary)

    def _clear_invoice_form_after_save(self):
        """Clear invoice form and add 1 empty line item after saving"""
        try:
            # Clear all line items
            self.clear_all_items()
            # Clear client information
            self.clear_client_information()
            # Reset editing flag
            self._editing_existing_invoice = False
            # Add 1 empty line item with proper delay
            if hasattr(self, 'add_item_row'):
                QtCore.QTimer.singleShot(200, self.add_item_row)
        except Exception as e:
            log.warning("Error clearing invoice form after save: %s", e)

    def edit_invoice_by_number(self, invoice_number: str):
        """Find invoice in Firebase and load it into the Invoice Management form.
        Called from balance_sheet_tab so all Firebase/Invoice access stays in main.py."""
        if not invoice_number:
            return

        # Navigate to Invoice Management tab
        self._nav_to(2)
        if hasattr(self, "project_invoice_inner_tabs"):
            self.project_invoice_inner_tabs.setCurrentIndex(1)

        # Load invoice from Firebase
        invoice_obj = None
        original_created_at = None
        original_firebase_id = None
        try:
            raw = FirebaseManager.load_invoices() or []
            for inv_data in raw:
                meta = inv_data.get('meta', {})
                stored = str(meta.get('invoice_number', '') or '').strip()
                if stored == invoice_number.strip():
                    invoice_obj = Invoice.from_dict(inv_data)
                    original_created_at = meta.get('created_at')
                    original_firebase_id = inv_data.get('firebase_id')
                    break
        except Exception as e:
            log.warning("edit_invoice_by_number load error: %s", e)

        if invoice_obj is None:
            QtWidgets.QMessageBox.warning(
                self, "Invoice Not Found",
                f"Invoice {invoice_number} could not be loaded.\n"
                "Please edit it directly from Invoice History.",
            )
            return

        # Populate the form — same steps as InvoiceHistoryViewWidget.edit_invoice
        self.clear_all_items()
        self.clear_client_information()

        for combo_name in ('client_combo', 'line_items_client_combo'):
            if hasattr(self, combo_name):
                c = getattr(self, combo_name)
                c.blockSignals(True)
                idx = c.findText(invoice_obj.client_name)
                if idx >= 0:
                    c.setCurrentIndex(idx)
                else:
                    c.setEditText(invoice_obj.client_name)
                c.blockSignals(False)

        if hasattr(self, 'client_email_edit'):
            self.client_email_edit.setText(invoice_obj.client_email)
        if hasattr(self, 'client_address_edit'):
            self.client_address_edit.setPlainText(invoice_obj.client_address)
        self.update_invoice_client_summary(
            invoice_obj.client_name, invoice_obj.client_email, invoice_obj.client_address)

        try:
            inv_date = QtCore.QDate.fromString(invoice_obj.date, "MM-dd-yyyy")
            due_date = QtCore.QDate.fromString(invoice_obj.due_date, "MM-dd-yyyy")
            if inv_date.isValid() and hasattr(self, 'date_edit'):
                self.date_edit.setDate(inv_date)
            if due_date.isValid() and hasattr(self, 'due_date_edit'):
                self.due_date_edit.setDate(due_date)
        except Exception:
            pass

        self.invoice.invoice_number = invoice_obj.invoice_number
        self.invoice_no_edit.setText(invoice_obj.invoice_number)
        # Preserve the existing invoice status so Payment Status label shows correctly
        self.invoice.status        = invoice_obj.status or 'Unpaid'
        self.invoice.received_date = invoice_obj.received_date or 'N/A'
        self._ps_base_status       = invoice_obj.status or 'Unpaid'
        # Store paid amount so update_totals can recalculate dynamically if amounts change
        if (invoice_obj.status or '').strip() == 'Paid':
            self._editing_invoice_paid_amount = float(invoice_obj.total)
        else:
            self._editing_invoice_paid_amount = 0.0

        if hasattr(self, 'tax_spin'):
            self.tax_spin.setValue(float(invoice_obj.tax_rate))
        if hasattr(self, 'notes_edit'):
            self.notes_edit.setPlainText(invoice_obj.notes or "")

        for item in invoice_obj.items:
            inv_item = InvoiceItem(
                project_number=item.project_number,
                description=item.description,
                plant=item.plant,
                quantity=item.quantity,
                unit_price=float(item.unit_price),
                down_payment=float(item.down_payment),
                payment_category=item.payment_category,
            )
            self.add_item_row(inv_item)
            last_row = self.item_rows[-1] if getattr(self, "item_rows", None) else None
            if last_row and item.payment_category:
                stage = ItemRowWidget.normalize_payment_label(item.payment_category) or item.payment_category
                if hasattr(last_row, "lock_to_stage"):
                    last_row.lock_to_stage(stage)
                if hasattr(last_row, "update_total"):
                    last_row.update_total()

        self.update_totals()
        # Mark as editing so Generate PDF is always enabled for existing invoices
        self._editing_existing_invoice = True
        self._original_created_at = original_created_at  # Preserve original creation date
        self._original_firebase_id = original_firebase_id  # Preserve Firebase ID for proper update
        QtCore.QTimer.singleShot(150, self._update_pdf_btn_state)
        if hasattr(self, 'statusBar'):
            self.statusBar().showMessage(
                f"Editing invoice {invoice_obj.invoice_number} — modify and click Generate PDF.",
                8000,
            )

    def _sync_invoice_revenue_after_save(self, invoice_snapshot: 'Invoice'):
        """Update the balance sheet revenue entry for a saved/edited invoice."""
        try:
            self.save_invoice_as_revenue(invoice_snapshot)
        except Exception as e:
            log.warning("_sync_invoice_revenue_after_save error: %s", e)
            self.refresh_balance_sheet_after_invoice()

    def _link_unlinked_payments_to_invoice(
            self, invoice_number: str, project_numbers: list):
        """Bind every payment-tracker record that has no invoice_number yet
        (for the given projects) to this invoice, then sync each one to the
        balance sheet as an is_payment=True revenue entry.

        Called immediately when a new invoice is saved so that payments that
        were recorded before the invoice was created are correctly reflected
        in the annual financial summary without needing stage-label matching.
        """
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()
            tracker._load_payments()
            linked = 0
            for pn in project_numbers:
                for pay in tracker.get_project_payments(pn):
                    if not (pay.invoice_number or '').strip():
                        tracker.update_payment(
                            pay.payment_id, invoice_number=invoice_number)
                        linked += 1
            if linked:
                log.info(
                    "Linked %d pre-invoice payment(s) to invoice %s",
                    linked, invoice_number)
        except Exception as e:
            log.warning("_link_unlinked_payments_to_invoice error: %s", e)

    def _trigger_post_save_payment_sync(self, invoice_number: str, items: list):
        """After an invoice is saved, link any pre-existing project payments to it
        and recalculate only THAT invoice's status via _auto_sync_invoice_statuses.

        Passing target_invoice ensures old invoices for the same projects are
        never touched — only the freshly saved invoice is processed.
        """
        try:
            project_numbers = list({
                str(it.project_number).strip()
                for it in items
                if getattr(it, "project_number", "")
            })
            if not project_numbers:
                return
            # Immediately link all unlinked payments for these projects to this
            # invoice so that _auto_sync_invoice_statuses (500 ms later) finds
            # them already bound and correctly updates paid_amount / status.
            self._link_unlinked_payments_to_invoice(invoice_number, project_numbers)
            project_tab = getattr(self, "project_tab", None)
            if project_tab and hasattr(project_tab, "_auto_sync_invoice_statuses"):
                for pn in project_numbers:
                    QtCore.QTimer.singleShot(
                        500,
                        lambda p=pn, inv=invoice_number: project_tab._auto_sync_invoice_statuses(
                            p, target_invoice=inv
                        ),
                    )
            # Refresh the balance sheet AFTER _auto_sync_invoice_statuses (500 ms)
            # has written the updated paid_amount / status / has_payment_entries
            # back to Firebase so the UI reads the correct values.
            QtCore.QTimer.singleShot(750, self.refresh_balance_sheet_after_invoice)
        except Exception as e:
            log.warning("_trigger_post_save_payment_sync error: %s", e)

    def save_invoice_as_revenue(self, invoice: Invoice, _skip_ui_refresh: bool = False):
        """Save invoice as revenue entry in balance sheet with due date"""
        try:
            log.info("  Attempting to save invoice %s as revenue...", invoice.invoice_number)
            
            # Build description with client name + project names
            project_names = []
            for item in invoice.items:
                if hasattr(item, 'project_name') and item.project_name:
                    project_names.append(item.project_name)
                elif item.description and item.description.strip():
                    project_names.append(item.description.strip())
                elif item.project_number and item.project_number.strip():
                    project_names.append(item.project_number.strip())
            
            # Create a meaningful description
            # Create a meaningful description with ALL project names
            if project_names:
                # Join ALL project names with commas - NO TRUNCATION
                project_text = ", ".join(project_names)
            else:
                project_text = "No project specified"
            description = f"{invoice.client_name} - {project_text}"
            source = f"Invoice - {invoice.invoice_number}"
            
            # Get invoice date for year determination
            try:
                date_obj = datetime.strptime(invoice.date, "%m-%d-%Y")
                year = date_obj.year
            except Exception as e:
                log.warning("a      Error parsing date: %s, using current year", e)
                year = datetime.now().year
            
            # Check if revenue entry already exists for this invoice
            existing_revenue = None
            existing_revenue_id = None
            
            if FIREBASE_AVAILABLE:
                from firebase_admin import db
                revenue_ref = db.reference('revenue')
                all_revenue = revenue_ref.get()
                
                if all_revenue:
                    for rev_id, revenue in all_revenue.items():
                        if revenue and revenue.get('is_invoice') and revenue.get('invoice_number') == invoice.invoice_number:
                            existing_revenue = revenue
                            existing_revenue_id = rev_id
                            break
            
            total_down_payments = sum(item.down_payment for item in invoice.items)
            paid_amount_dp = Currency.quantize(total_down_payments)

            # Check payment tracker for pre-existing payments for this invoice's
            # projects — payments recorded before the invoice was created are
            # already in memory and give us the true paid amount + received date.
            _tracker_paid = Decimal('0')
            _tracker_received_date = "N/A"
            try:
                from payment_tracker import get_payment_tracker as _get_pt
                _pt = _get_pt()
                _date_fmts = ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y")
                _best_dt = None
                _best_ds = "N/A"
                for _item in invoice.items:
                    _pn = str(getattr(_item, 'project_number', '') or '').strip()
                    if not _pn:
                        continue
                    for _pay in _pt.get_project_payments(_pn):
                        if (_pay.payment_stage or '').strip().lower() == 'tax':
                            continue
                        # Only count payments linked to THIS invoice or unlinked payments.
                        # Payments for other invoices on the same project must not inflate
                        # this invoice's paid amount and falsely mark it as "Paid".
                        _pay_inv = (getattr(_pay, 'invoice_number', '') or '').strip()
                        if _pay_inv and _pay_inv != invoice.invoice_number:
                            continue
                        _tracker_paid += _pay.amount
                        _d = _pay.payment_date or ''
                        for _fmt in _date_fmts:
                            try:
                                _dt = datetime.strptime(_d, _fmt)
                                if _best_dt is None or _dt > _best_dt:
                                    _best_dt = _dt
                                    _best_ds = _d
                                break
                            except (ValueError, TypeError):
                                pass
                _tracker_received_date = _best_ds
            except Exception as _pe:
                log.warning("save_invoice_as_revenue: payment lookup error: %s", _pe)

            # Use the larger of invoice down-payments or tracker payments
            paid_amount = max(paid_amount_dp, Currency.quantize(float(_tracker_paid)))
            unpaid_amount = max(Currency.quantize(invoice.total) - paid_amount, Decimal('0'))
            invoice_total_d = Currency.quantize(invoice.total)

            if paid_amount >= invoice_total_d > 0:
                revenue_status = "Paid"
                received_date = _tracker_received_date if _tracker_received_date != "N/A" else invoice.date
            elif paid_amount > 0:
                revenue_status = "Partially Paid"
                received_date = _tracker_received_date if _tracker_received_date != "N/A" else invoice.date
            else:
                revenue_status = "Unpaid"
                received_date = "N/A"

            # When UPDATING an existing revenue entry that was already "Paid", preserve
            # that status only if the stored paid amount still covers the new invoice total.
            # If the user raised the total above what was paid, downgrade to "Partially Paid".
            if existing_revenue and existing_revenue.get('status') == 'Paid' and revenue_status != 'Paid':
                _ex_paid = float(existing_revenue.get('paid_amount', 0) or 0)
                _new_total = float(invoice.total)
                if _ex_paid >= _new_total > 0:
                    revenue_status = 'Paid'
                    received_date = existing_revenue.get('received_date', invoice.date) or invoice.date

            down_payment_received_date = received_date

            # CRITICAL: Create revenue entry with DUE DATE from invoice
            revenue_entry = {
                'source': source,
                'description': description,
                'amount': str(float(invoice.total)),
                'paid_amount': str(float(paid_amount)),
                'unpaid_amount': str(float(unpaid_amount)),
                'down_payment_amount': str(float(paid_amount)),
                'down_payment_received_date': down_payment_received_date,
                'date': invoice.date,
                'due_date': invoice.due_date,
                'invoice_number': invoice.invoice_number,
                'status': revenue_status,
                'received_date': received_date,
                'is_invoice': True,
                'year': year,
                'updated_at': datetime.now(timezone.utc).isoformat()
            }
            
            if existing_revenue:
                # Update existing revenue
                revenue_entry['created_at'] = existing_revenue.get('created_at', datetime.now(timezone.utc).isoformat())
                revenue_entry['firebase_id'] = existing_revenue_id
                
                from firebase_admin import db
                revenue_ref = db.reference('revenue')
                revenue_ref.child(existing_revenue_id).update(revenue_entry)
                log.info("UPDATED revenue entry for invoice %s", invoice.invoice_number)
            else:
                # Create new revenue entry
                revenue_entry['created_at'] = datetime.now(timezone.utc).isoformat()
                from balance_sheet_tab import BalanceSheetFirebaseManager
                success = BalanceSheetFirebaseManager.save_revenue(revenue_entry)
                
                if success:
                    log.info("CREATED revenue entry for invoice %s", invoice.invoice_number)
                else:
                    log.warning("a   FAILED: Could not save revenue entry for invoice %s", invoice.invoice_number)
                    return False
            
            log.info("  - Source: %s", source)
            log.info("  - Description: %s", description)
            log.info("  - Amount: $%.2f", float(invoice.total))
            log.debug("  - Status: %s", revenue_entry['status'])
            log.info("  - Down Payment Received: $%.2f", float(paid_amount))
            log.info("  - Down Payment Received Date: %s", down_payment_received_date)
            log.info("  - Remaining Invoice Due: $%.2f", float(unpaid_amount))
            log.info("  - Due Date: %s", invoice.due_date)
            log.info("  - Year: %s", year)
            
            # Refresh the balance sheet data immediately (skip when called from a
            # background thread — Qt UI must only be updated from the main thread)
            if not _skip_ui_refresh:
                self.refresh_balance_sheet_after_invoice()
            return True
            
        except Exception as e:
            log.warning("a   ERROR saving invoice as revenue: %s", e)
            import traceback
            traceback.print_exc()
            return False
            
    def refresh_balance_sheet_after_invoice(self):
        """Refresh balance sheet after invoice creation — runs Firebase fetch in background."""
        try:
            if not hasattr(self, 'balance_sheet_tab'):
                return
            balance_tab = self.balance_sheet_tab
            import threading as _threading
            def _bg():
                try:
                    balance_tab._fetch_data_background()
                except Exception as _e:
                    log.warning("bg balance sheet fetch error: %s", _e)
                QtCore.QTimer.singleShot(0, balance_tab._apply_fetched_data_ui)
            _threading.Thread(target=_bg, daemon=True).start()
        except Exception as e:
            log.warning("Error refreshing balance sheet: %s", e)

    def clear_client_information(self):
        """Clear all client information fields"""
        self.client_combo.blockSignals(True)
        self.client_combo.setCurrentIndex(-1)  # Set to -1 (no selection) instead of 0
        self.client_combo.blockSignals(False)
        if hasattr(self, "client_contact_edit"):
            self.client_contact_edit.clear()
        if hasattr(self, "client_phone_edit"):
            self.client_phone_edit.clear()
        if hasattr(self, "client_primary_email_edit"):
            self.client_primary_email_edit.clear()
        self.client_email_edit.clear()
        self.client_address_edit.clear()
        if hasattr(self, "line_items_client_combo"):
            self.line_items_client_combo.blockSignals(True)
            self.line_items_client_combo.setCurrentIndex(-1)
            self.line_items_client_combo.blockSignals(False)
        self.update_invoice_client_summary("", "", "")
        self.editing_client_name = None
        if hasattr(self, "save_client_btn"):
            self.save_client_btn.setEnabled(False)
    
    def cleanup_temp_file(self, file_path: Path):
        """Clean up temporary file"""
        try:
            if file_path.exists():
                file_path.unlink()
                log.info("Cleaned up temporary file: %s", file_path)
        except Exception as e:
            log.warning("Could not clean up temporary file: %s", e)
    
    def open_pdf_from_firebase(self, invoice_number: str):
        """Open PDF directly from Firebase"""
        if not FIREBASE_AVAILABLE:
            QtWidgets.QMessageBox.warning(self, "PDF Open", "Firebase not available. Cannot open PDF from cloud.")
            return
        
        try:
            # Create temporary file for PDF
            temp_dir = Path(tempfile.gettempdir()) / "mabs_invoices_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_pdf_path = temp_dir / f"{invoice_number}.pdf"
            
            # Load PDF from Firebase
            pdf_path = FirebaseManager.load_pdf_from_firebase(invoice_number, temp_pdf_path)
            
            if pdf_path and pdf_path.exists():
                # Open the PDF
                if FileManager.open_file(pdf_path):
                    log.info("PDF opened successfully: %s", invoice_number)
                    
                    # Clean up temporary file after 10 seconds
                    QtCore.QTimer.singleShot(10000, lambda: self.cleanup_temp_file(pdf_path))
                else:
                    QtWidgets.QMessageBox.critical(self, "PDF Open", "Failed to open PDF file.")
            else:
                QtWidgets.QMessageBox.warning(self, "PDF Open", 
                                            f"PDF not found in Firebase:\n{invoice_number}\n\n"
                                            f"The PDF may not have been generated yet or uploaded to the cloud.")
                                            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "PDF Open Error", f"Error opening PDF: {str(e)}")
    
    def ask_load_completed_projects(self, client_name):
        """Ask user if they want to load completed projects for this client"""
        reply = QtWidgets.QMessageBox.question(
            self, 
            "Load Completed Projects",
            f"Do you want to load completed projects for {client_name} as invoice items?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        return reply
    
    def auto_save_new_client(self, client_name, email, address):
        """Auto-save a new client when generating invoice"""
        try:
            if not client_name or client_name in self.clients:
                return False
            
            # Prepare client data with all fields
            client_data = {
                "email": email,
                "address": address,
                "contact_person": "",
                "phone": "",
                "company_email": "",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "auto_saved": True
            }
            
            # Save to Firebase
            if FirebaseManager.save_client(client_name, client_data):
                # Update local cache
                self.clients[client_name] = client_data
                
                # Update dropdown if not already there
                if self.client_combo.findText(client_name) == -1:
                    self.client_combo.addItem(client_name)
                
                # Refresh project tab dropdown
                self.refresh_project_tab_clients()
                self.refresh_inline_clients_table()
                
                log.info("Auto-saved new client: %s", client_name)
                return True
            else:
                log.warning("Failed to auto-save client: %s", client_name)
                return False
                
        except Exception as e:
            log.warning("Error auto-saving client: %s", e)
            return False
    
    def confirm_close(self):
        """Show a bold, unmissable exit confirmation dialog."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Exit PIMS")
        dlg.setFixedSize(440, 240)
        dlg.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowTitleHint)
        dlg.setStyleSheet("QDialog { background: #ffffff; }")

        root = QtWidgets.QVBoxLayout(dlg)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Red header bar ──────────────────────────────────────────────
        header = QtWidgets.QFrame()
        header.setFixedHeight(72)
        header.setStyleSheet(
            "QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #dc2626, stop:1 #b91c1c); }"
        )
        hlay = QtWidgets.QHBoxLayout(header)
        hlay.setContentsMargins(24, 0, 24, 0)
        hlay.setSpacing(14)

        icon_lbl = QtWidgets.QLabel("⚠")
        icon_lbl.setStyleSheet(
            "color:#fff; font-size:28px; background:transparent;"
        )
        title_col = QtWidgets.QVBoxLayout()
        title_col.setSpacing(2)
        t1 = QtWidgets.QLabel("Exit PIMS?")
        t1.setStyleSheet(
            "color:#fff; font-size:18px; font-weight:900; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        t2 = QtWidgets.QLabel("Any unsaved changes will be lost.")
        t2.setStyleSheet(
            "color:rgba(255,255,255,0.75); font-size:12px; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        title_col.addWidget(t1)
        title_col.addWidget(t2)
        hlay.addWidget(icon_lbl)
        hlay.addLayout(title_col, 1)
        root.addWidget(header)

        # ── Body message ────────────────────────────────────────────────
        body = QtWidgets.QLabel(
            "Are you sure you want to exit\n<b>PIMS — Project & Invoice Management</b>?"
        )
        body.setTextFormat(QtCore.Qt.RichText)
        body.setAlignment(QtCore.Qt.AlignCenter)
        body.setStyleSheet(
            "color:#1e293b; font-size:14px; font-family:'Inter','Segoe UI';"
            " padding:20px 28px 16px 28px; background:transparent;"
        )
        body.setWordWrap(True)
        root.addWidget(body, 1)

        # ── Buttons ─────────────────────────────────────────────────────
        btn_bar = QtWidgets.QFrame()
        btn_bar.setFixedHeight(66)
        btn_bar.setStyleSheet(
            "QFrame { background:#f8fafc; border-top:1px solid #e2e8f0; }"
        )
        blay = QtWidgets.QHBoxLayout(btn_bar)
        blay.setContentsMargins(24, 0, 24, 0)
        blay.setSpacing(12)

        stay_btn = QtWidgets.QPushButton("  No, Stay")
        stay_btn.setFixedSize(150, 42)
        stay_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        stay_btn.setStyleSheet("""
            QPushButton {
                background: #16a34a; color: #fff;
                border: none; border-radius: 9px;
                font-size: 14px; font-weight: 800;
                font-family: 'Inter','Segoe UI';
            }
            QPushButton:hover { background: #15803d; }
        """)
        stay_btn.clicked.connect(dlg.reject)

        exit_btn = QtWidgets.QPushButton("  Yes, Exit")
        exit_btn.setFixedSize(150, 42)
        exit_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        exit_btn.setStyleSheet("""
            QPushButton {
                background: #dc2626; color: #fff;
                border: none; border-radius: 9px;
                font-size: 14px; font-weight: 800;
                font-family: 'Inter','Segoe UI';
            }
            QPushButton:hover { background: #b91c1c; }
        """)
        exit_btn.clicked.connect(dlg.accept)

        blay.addStretch()
        blay.addWidget(stay_btn)
        blay.addWidget(exit_btn)
        root.addWidget(btn_bar)

        return dlg.exec_() == QtWidgets.QDialog.Accepted

    def closeEvent(self, event):
        """Show confirmation popup; if user confirms, exit immediately."""
        if getattr(self, "_logging_out", False):
            event.accept()
            return
        if self.confirm_close():
            import os as _os
            _os._exit(0)
        else:
            event.ignore()

# ====================================================================
# MAIN FUNCTION WITH LOGIN INTEGRATION
# ====================================================================

def create_main_window(username: str, role: str):
    """Factory function to create MainWindow with user credentials."""
    # Clean up leftover update files
    cleanup_old_backup()
    cleanup_temp_backups()
    
    window = MainWindow(username=username, role=role)
    return window

def _make_app_icon() -> QtGui.QIcon:
    """Build a crisp 256×256 app icon: teal rounded-rect background,
    white invoice document with fold, horizontal content lines, and a
    small teal checkmark badge — all drawn with QPainter, no files needed."""
    SIZE = 256
    px = QtGui.QPixmap(SIZE, SIZE)
    px.fill(QtCore.Qt.transparent)

    p = QtGui.QPainter(px)
    p.setRenderHint(QtGui.QPainter.Antialiasing)

    # ── background ────────────────────────────────────────────────────────
    grad = QtGui.QLinearGradient(0, 0, 0, SIZE)
    grad.setColorAt(0.0, QtGui.QColor("#0F766E"))
    grad.setColorAt(1.0, QtGui.QColor("#065F46"))
    p.setBrush(QtGui.QBrush(grad))
    p.setPen(QtCore.Qt.NoPen)
    p.drawRoundedRect(0, 0, SIZE, SIZE, 48, 48)

    # ── document body ──────────────────────────────────────────────────────
    DX, DY, DW, DH = 62, 44, 132, 162          # document rect
    FOLD = 32                                    # fold-corner size
    doc_path = QtGui.QPainterPath()
    doc_path.moveTo(DX, DY)
    doc_path.lineTo(DX + DW - FOLD, DY)         # top edge → fold start
    doc_path.lineTo(DX + DW, DY + FOLD)         # fold diagonal
    doc_path.lineTo(DX + DW, DY + DH)           # right edge down
    doc_path.lineTo(DX, DY + DH)                # bottom edge
    doc_path.closeSubpath()
    p.setBrush(QtGui.QColor(255, 255, 255, 245))
    p.setPen(QtCore.Qt.NoPen)
    p.drawPath(doc_path)

    # fold triangle (slightly darker white)
    fold_path = QtGui.QPainterPath()
    fold_path.moveTo(DX + DW - FOLD, DY)
    fold_path.lineTo(DX + DW - FOLD, DY + FOLD)
    fold_path.lineTo(DX + DW, DY + FOLD)
    fold_path.closeSubpath()
    p.setBrush(QtGui.QColor(200, 230, 228))
    p.drawPath(fold_path)

    # ── invoice lines ──────────────────────────────────────────────────────
    line_pen = QtGui.QPen(QtGui.QColor("#0F766E"), 7, QtCore.Qt.SolidLine, QtCore.Qt.RoundCap)
    p.setPen(line_pen)
    short_pen = QtGui.QPen(QtGui.QColor("#CBD5E1"), 7, QtCore.Qt.SolidLine, QtCore.Qt.RoundCap)

    line_x1, line_x2_long, line_x2_short = DX + 18, DX + DW - 20, DX + DW - 58
    for i, y in enumerate([DY + 54, DY + 76, DY + 98, DY + 120]):
        pen = line_pen if i == 0 else short_pen
        x2  = line_x2_long if i % 2 == 0 else line_x2_short
        p.setPen(pen)
        p.drawLine(line_x1, y, x2, y)

    # amount line (bold, full width)
    amt_pen = QtGui.QPen(QtGui.QColor("#0F766E"), 9, QtCore.Qt.SolidLine, QtCore.Qt.RoundCap)
    p.setPen(amt_pen)
    p.drawLine(line_x1, DY + 142, line_x2_long, DY + 142)

    # ── checkmark badge (bottom-right) ─────────────────────────────────────
    BX, BY, BR = SIZE - 68, SIZE - 68, 34       # badge centre x, y, radius
    p.setBrush(QtGui.QColor("#10B981"))
    p.setPen(QtCore.Qt.NoPen)
    p.drawEllipse(QtCore.QPoint(BX, BY), BR, BR)

    ck = QtGui.QPen(QtGui.QColor(255, 255, 255), 10, QtCore.Qt.SolidLine,
                    QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin)
    p.setPen(ck)
    p.drawPolyline(QtGui.QPolygon([
        QtCore.QPoint(BX - 16, BY),
        QtCore.QPoint(BX - 4,  BY + 13),
        QtCore.QPoint(BX + 18, BY - 14),
    ]))

    p.end()
    return QtGui.QIcon(px)


def main():
    # Single-instance check BEFORE creating QApplication — prevent multiple instances from running
    import socket
    single_instance_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        single_instance_socket.bind(('127.0.0.1', 25519))
        single_instance_socket.listen(1)
    except OSError:
        # Port is already in use — another instance is running
        print("ERROR: PIMS is already running. Aborting this instance.")
        sys.exit(1)

    Config.setup_directories()
    cleanup_old_backup()   # delete MABS_Invoice.bak left by the previous update
    cleanup_temp_backups()

    global FIREBASE_AVAILABLE

    if FIREBASE_AVAILABLE:
        log.info("Firebase integration: ENABLED")

    app = QtWidgets.QApplication(sys.argv)

    # Remove the "?" (What's This) button from every dialog globally
    _orig_qdialog_init = QtWidgets.QDialog.__init__
    def _no_help_btn_init(self, *args, **kwargs):
        _orig_qdialog_init(self, *args, **kwargs)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowContextHelpButtonHint)
    QtWidgets.QDialog.__init__ = _no_help_btn_init

    _co = Config.COMPANY.get("name", "MABS Engineering LLC")
    app.setApplicationName(f"{_co} - Project & Invoice Management")
    app.setApplicationVersion("2.1")
    app.setStyle('Fusion')
    app.setWindowIcon(_make_app_icon())

    from font_loader import apply_font
    FONT_FAMILY = apply_font(app)
    log.info("UI font: %s", FONT_FAMILY)

    from app_theme import get_stylesheet, install_clean_dropdown_style_patch, install_calendar_style_filter
    install_clean_dropdown_style_patch()
    app.setStyleSheet(get_stylesheet())
    install_calendar_style_filter(app)

    # a"a" Initialize default admin ONCE here, silently, before UI a"a"
    if FIREBASE_AVAILABLE:
        try:
            FirebaseManager.initialize_default_admin()
        except Exception as e:
            log.warning("Default admin init failed: %s", e)

    from login_window import LoginWindow

    app.setQuitOnLastWindowClosed(False)

    class AppSession(QtCore.QObject):
        def __init__(self):
            super().__init__()
            self.login_win = None
            self.main_win = None

        def show_login(self):
            login_win = LoginWindow()
            login_win.setAttribute(QtCore.Qt.WA_DeleteOnClose, True)
            self.login_win = login_win
            login_win.login_successful.connect(self.on_login_success)
            login_win.destroyed.connect(self.on_login_closed)
            login_win.show()

        def on_login_success(self, username, email, role):
            login_win = self.login_win
            if login_win is None:
                return

            try:
                login_win.login_successful.disconnect(self.on_login_success)
            except Exception:
                pass

            def _create_main():
                main_win = MainWindow(username=username, role=role)
                main_win.setAttribute(QtCore.Qt.WA_DeleteOnClose, True)
                self.main_win = main_win
                main_win.logout_requested.connect(self.on_logout_requested)
                main_win.destroyed.connect(lambda _=None, window=main_win: self.on_main_closed(window))
                # Wire update popup BEFORE show so window_ready signal is never missed
                self.update_checker = UpdateChecker(parent=main_win)
                main_win.window_ready.connect(
                    lambda: QtCore.QTimer.singleShot(
                        30000,
                        lambda: self.update_checker.check_for_updates(silent=True)
                    )
                )

                def _do_show():
                    main_win.showMaximized()
                    main_win.raise_()
                    main_win.activateWindow()
                    login_win.close()

                dashboard = getattr(main_win, 'dashboard_tab', None)
                if dashboard is not None:
                    _shown = [False]
                    # Fallback: open the window after 8 s even if Firebase is unreachable
                    _fallback = QtCore.QTimer(main_win)
                    _fallback.setSingleShot(True)

                    def _do_show_once():
                        if not _shown[0]:
                            _shown[0] = True
                            _fallback.stop()
                            _do_show()

                    _fallback.timeout.connect(_do_show_once)
                    _fallback.start(8000)
                    dashboard.data_ready.connect(_do_show_once)
                else:
                    _do_show()

            QtCore.QTimer.singleShot(0, _create_main)

        def on_logout_requested(self):
            self.main_win = None
            self.show_login()

        def on_login_closed(self, *_):
            self.login_win = None
            if self.main_win is None:
                QtWidgets.QApplication.quit()

        def on_main_closed(self, window):
            if self.main_win is window:
                self.main_win = None
            if self.login_win is None:
                QtWidgets.QApplication.quit()

    # Global exception handler — logs unhandled exceptions instead of silent crash
    def _handle_exception(exc_type, exc_value, exc_tb):
        import traceback
        msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        log.critical("UNHANDLED EXCEPTION:\n%s", msg)
        try:
            QtWidgets.QMessageBox.critical(None, "Unexpected Error",
                f"The application encountered an error:\n\n{exc_value}\n\n"
                "Please check the log file for details.")
        except Exception:
            pass
    sys.excepthook = _handle_exception

    session = AppSession()
    session.show_login()
    exit_code = app.exec_()
    os._exit(exit_code)
    
if __name__ == "__main__":
    main()
