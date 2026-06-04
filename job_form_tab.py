from PyQt5 import QtWidgets, QtCore, QtGui
from pathlib import Path
from datetime import datetime, timedelta, timezone
import json
import re
import tempfile
import os
import threading
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from app_logger import get_logger
from app_theme import CALENDAR_URL, CHEVRON_URL, configure_filter_button
from template_manager import TemplateManager, TemplateDialog
from client_intelligence import ClientIntelligence, ClientSuggestionWidget

_log = get_logger(__name__)


def _sales_people_path() -> Path:
    return Path(__file__).resolve().parent / "data" / "sales_persons.json"


def _load_local_sales_people() -> list:
    try:
        path = _sales_people_path()
        if not path.exists():
            return []
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [person for person in data if isinstance(person, dict)]
    except Exception as exc:
        _log.warning("Could not load local sales people: %s", exc)
    return []


def _save_local_sales_people(people: list) -> bool:
    try:
        path = _sales_people_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(people, f, indent=2, ensure_ascii=False)
        return True
    except Exception as exc:
        _log.warning("Could not save local sales people: %s", exc)
    return False


def _xlsx_color(hex_color: str) -> str:
    """Return an openpyxl-safe ARGB color string."""
    value = str(hex_color or "").strip().lstrip("#")
    if len(value) == 6:
        return f"FF{value.upper()}"
    if len(value) == 8:
        return value.upper()
    return "FFFFFFFF"


def _xlsx_fill(hex_color: str) -> PatternFill:
    color = _xlsx_color(hex_color)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _load_local_clients() -> dict:
    try:
        path = Path(__file__).resolve().parent / "data" / "clients.json"
        if not path.exists():
            return {}
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception as exc:
        _log.warning("Could not load local clients: %s", exc)
    return {}
def resource_path(relative_path):
    """
    Get absolute path to resource, works for dev & PyInstaller
    """
    try:
        base_path = sys._MEIPASS  # PyInstaller temp folder
    except Exception:
        base_path = Path(__file__).resolve().parent
    return Path(base_path) / relative_path

# Import Firebase configuration from main
try:
    from main import FIREBASE_AVAILABLE, FirebaseManager, db
except ImportError:
    # Fallback if running independently
    FIREBASE_AVAILABLE = False
    FirebaseManager = None
    db = None

# ── Quote Detail Window ────────────────────────────────────────────────────────

class QuoteDetailWindow(QtWidgets.QDialog):
    """Full-detail window that opens when the user clicks a Quote # in the table."""

    status_changed  = QtCore.pyqtSignal(str)   # emits new status string
    action_requested = QtCore.pyqtSignal(str)  # emits action key

    STATUS_COLORS = {
        # fg, bg, border  (used by QuoteDetailWindow status button)
        "Draft":       ("#475569", "#f1f5f9", "#cbd5e1"),
        "Sent":        ("#1e40af", "#dbeafe", "#93c5fd"),
        "In Review":   ("#92400e", "#fef3c7", "#fcd34d"),
        "Approved":    ("#065f46", "#d1fae5", "#6ee7b7"),
        "On Hold":     ("#374151", "#f3f4f6", "#d1d5db"),
        "Completed":   ("#065f46", "#d1fae5", "#6ee7b7"),
        "Converted":   ("#4c1d95", "#ede9fe", "#c4b5fd"),
        "Rejected":    ("#991b1b", "#fee2e2", "#fca5a5"),
        "Expired":     ("#7c2d12", "#fff7ed", "#fed7aa"),
        "Cancelled":   ("#9d174d", "#fce7f3", "#f9a8d4"),
        # Legacy aliases kept for existing data
        "Not Started": ("#475569", "#f1f5f9", "#cbd5e1"),
        "Cancel":      ("#9d174d", "#fce7f3", "#f9a8d4"),
        "Low":         ("#166534", "#f0fdf4", "#86efac"),
        "Medium":      ("#1e40af", "#dbeafe", "#93c5fd"),
        "High":        ("#991b1b", "#fee2e2", "#fca5a5"),
        "Urgent":      ("#92400e", "#fef3c7", "#fcd34d"),
    }

    def __init__(self, job_data: dict, parent=None):
        super().__init__(parent)
        self._job = job_data
        self.setWindowTitle(f"Quote  —  {job_data.get('job_number', '')}")
        self.setMinimumSize(780, 580)
        self.setModal(True)
        self.setWindowFlags(self.windowFlags() | QtCore.Qt.WindowMaximizeButtonHint)
        self._build()

    # ── Build UI ───────────────────────────────────────────────────────────

    def _build(self):
        # Force light background and dark text regardless of system theme
        self.setStyleSheet("background:#f4f7fb;")
        pal = self.palette()
        pal.setColor(pal.Window,      QtGui.QColor("#f4f7fb"))
        pal.setColor(pal.WindowText,  QtGui.QColor("#0f172a"))
        pal.setColor(pal.Base,        QtGui.QColor("#ffffff"))
        pal.setColor(pal.Text,        QtGui.QColor("#0f172a"))
        self.setPalette(pal)
        self.setAutoFillBackground(True)
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)
        root.addWidget(self._make_header())
        root.addWidget(self._make_body(), 1)
        root.addWidget(self._make_footer())

    def _make_header(self):
        hdr = QtWidgets.QFrame()
        hdr.setFixedHeight(78)
        hdr.setStyleSheet(
            "QFrame{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            "stop:0 #0f172a,stop:1 #1e3a5c);}"
        )
        lay = QtWidgets.QHBoxLayout(hdr)
        lay.setContentsMargins(24, 0, 24, 0)
        lay.setSpacing(14)

        # Quote # badge
        qn = self._job.get('job_number', '')
        badge = QtWidgets.QLabel(qn)
        badge.setStyleSheet(
            "background:#0f766e; color:#fff; font-size:13px; font-weight:900;"
            " border-radius:8px; padding:5px 14px;"
            " font-family:'Inter','Segoe UI';"
        )

        # Title column
        col = QtWidgets.QVBoxLayout()
        col.setSpacing(3)
        t1 = QtWidgets.QLabel(self._job.get('project_name', 'Quote Details'))
        t1.setStyleSheet(
            "color:#fff; font-size:17px; font-weight:900; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        client = self._job.get('client', '')
        t2 = QtWidgets.QLabel(client)
        t2.setStyleSheet(
            "color:rgba(255,255,255,0.55); font-size:12px; background:transparent;"
            " font-family:'Inter','Segoe UI';"
        )
        col.addWidget(t1)
        col.addWidget(t2)

        # Status pill
        status = self._job.get('status', 'Not Started')
        sc = self.STATUS_COLORS.get(status, ("#6b7280", "#f9fafb", "#e5e7eb"))
        self._status_btn = QtWidgets.QPushButton(f"  {status}  ▾")
        self._status_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._status_btn.setFixedHeight(32)
        self._status_btn.setStyleSheet(
            f"QPushButton{{background:{sc[1]};color:{sc[0]};border:1px solid {sc[2]};"
            "border-radius:8px;font-size:12px;font-weight:800;"
            "font-family:'Inter','Segoe UI';padding:0 10px;}}"
            "QPushButton:hover{opacity:0.9;}"
        )
        self._status_btn.clicked.connect(self._show_status_menu)

        lay.addWidget(badge)
        lay.addLayout(col, 1)
        lay.addWidget(self._status_btn)
        return hdr

    def _make_body(self):
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea,QWidget{background:#f4f7fb;}")

        inner = QtWidgets.QWidget()
        inner.setStyleSheet("background:#f4f7fb; color:#0f172a;")
        grid = QtWidgets.QGridLayout(inner)
        grid.setContentsMargins(24, 20, 24, 20)
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(14)

        fields = [
            ("Quote Number",    self._job.get('job_number', '—')),
            ("Project Name",    self._job.get('project_name', '—')),
            ("Client",          self._job.get('client', '—')),
            ("Sales Person",    self._job.get('sales', '—')),
            ("Engineering Cost",self._job.get('engineering_costs', '—')),
            ("Scope of Work",   self._job.get('scope_of_work', '—')),
            ("Due Date",        self._job.get('due_date', '—')),
            ("Created",         self._job.get('created_at', '—')),
            ("Notes",           self._job.get('notes', '—')),
            ("Description",     self._job.get('description', '—')),
        ]

        r = c = 0
        for label, value in fields:
            if not value or value == '—' or str(value).strip() == '':
                value = '—'
            card = self._info_card(label, str(value))
            grid.addWidget(card, r, c)
            c += 1
            if c == 2:
                c = 0
                r += 1

        grid.setRowStretch(r + 1, 1)
        scroll.setWidget(inner)
        return scroll

    def _info_card(self, label, value):
        card = QtWidgets.QFrame()
        card.setStyleSheet(
            "QFrame{background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;}"
        )
        card.setAutoFillBackground(True)
        lay = QtWidgets.QVBoxLayout(card)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(4)

        lbl = QtWidgets.QLabel(label.upper())
        lbl.setStyleSheet(
            "color:#94a3b8 !important;font-size:10px;font-weight:800;letter-spacing:0.8px;"
            "font-family:'Inter','Segoe UI';background:transparent;border:none;"
        )
        val = QtWidgets.QLabel(value)
        val.setWordWrap(True)
        val.setStyleSheet(
            "color:#0f172a !important;font-size:14px;font-weight:700;"
            "font-family:'Inter','Segoe UI';background:transparent;border:none;"
        )
        lay.addWidget(lbl)
        lay.addWidget(val)
        return card

    def _make_footer(self):
        footer = QtWidgets.QFrame()
        footer.setStyleSheet(
            "QFrame{background:#fff;border-top:1px solid #e2e8f0;}"
        )
        footer.setFixedHeight(66)
        lay = QtWidgets.QHBoxLayout(footer)
        lay.setContentsMargins(24, 0, 24, 0)
        lay.setSpacing(10)

        def _btn(label, bg, fg, hover, border=None, width=120):
            b = QtWidgets.QPushButton(label)
            b.setFixedSize(width, 36)
            b.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            bc = f"border:1px solid {border};" if border else "border:none;"
            b.setStyleSheet(
                f"QPushButton{{background:{bg};color:{fg};{bc}"
                "border-radius:8px;font-size:13px;font-weight:700;"
                "font-family:'Inter','Segoe UI';}}"
                f"QPushButton:hover{{background:{hover};}}"
            )
            return b

        view_btn    = _btn("View Details",    "#eff6ff", "#2563eb", "#dbeafe", "#bfdbfe", 120)
        pdf_btn     = _btn("View PDF",        "#f0fdf4", "#0f766e", "#dcfce7", "#bbf7d0", 100)
        project_btn = _btn("Create Project",  "#fefce8", "#92400e", "#fef9c3", "#fde68a", 130)
        edit_btn    = _btn("Edit Quote",      "#f8fafc", "#475569", "#e2e8f0", "#cbd5e1", 110)
        close_btn   = _btn("Close",           "#f1f5f9", "#64748b", "#e2e8f0", None,      80)

        view_btn.clicked.connect(lambda: (self.action_requested.emit("view"),    self.accept()))
        pdf_btn.clicked.connect( lambda: (self.action_requested.emit("pdf"),     self.accept()))
        project_btn.clicked.connect(lambda: (self.action_requested.emit("project"), self.accept()))
        edit_btn.clicked.connect(lambda: (self.action_requested.emit("edit"),    self.accept()))
        close_btn.clicked.connect(self.reject)

        # More dropdown
        more_btn = QtWidgets.QPushButton("⋯  More")
        more_btn.setFixedSize(90, 36)
        more_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        more_btn.setStyleSheet(
            "QPushButton{background:#f8fafc;color:#475569;border:1px solid #e2e8f0;"
            "border-radius:8px;font-size:13px;font-weight:700;"
            "font-family:'Inter','Segoe UI';}"
            "QPushButton:hover{background:#e2e8f0;}"
        )
        more_btn.clicked.connect(self._show_more_menu)

        lay.addWidget(view_btn)
        lay.addWidget(pdf_btn)
        lay.addWidget(project_btn)
        lay.addWidget(edit_btn)
        lay.addWidget(more_btn)
        lay.addStretch()
        lay.addWidget(close_btn)
        return footer

    # ── Status menu ────────────────────────────────────────────────────────

    def _show_status_menu(self):
        menu = QtWidgets.QMenu(self._status_btn)
        menu.setStyleSheet("""
            QMenu{background:#fff;border:1px solid #cbd5e1;border-radius:8px;
                  padding:4px 0;font-family:'Inter','Segoe UI';font-size:13px;}
            QMenu::item{padding:9px 20px;color:#1e293b;}
            QMenu::item:selected{background:#dbeafe;color:#1d4ed8;}
            QMenu::separator{height:1px;background:#e2e8f0;margin:4px 8px;}
        """)
        groups = [
            ["Draft", "Sent", "In Review"],
            ["Approved", "On Hold"],
            ["Completed", "Converted"],
            ["Rejected", "Expired", "Cancelled"],
        ]
        first = True
        for grp in groups:
            if not first:
                menu.addSeparator()
            first = False
            for s in grp:
                a = QtWidgets.QAction(s, menu)
                a.triggered.connect(lambda _, st=s: self._set_status(st))
                menu.addAction(a)
        menu.exec_(self._status_btn.mapToGlobal(
            QtCore.QPoint(0, self._status_btn.height())))

    def _set_status(self, new_status):
        sc = self.STATUS_COLORS.get(new_status, ("#6b7280", "#f9fafb", "#e5e7eb"))
        self._status_btn.setText(f"  {new_status}  ▾")
        self._status_btn.setStyleSheet(
            f"QPushButton{{background:{sc[1]};color:{sc[0]};border:1px solid {sc[2]};"
            "border-radius:8px;font-size:12px;font-weight:800;"
            "font-family:'Inter','Segoe UI';padding:0 10px;}}"
            "QPushButton:hover{opacity:0.9;}"
        )
        self._job['status'] = new_status
        self.status_changed.emit(new_status)

    # ── More menu ──────────────────────────────────────────────────────────

    def _show_more_menu(self):
        menu = QtWidgets.QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#fff;border:1px solid #cbd5e1;border-radius:8px;
                  padding:4px 0;font-family:'Inter','Segoe UI';font-size:13px;}
            QMenu::item{padding:9px 20px;color:#1e293b;}
            QMenu::item:selected{background:#dbeafe;color:#1d4ed8;}
            QMenu::separator{height:1px;background:#e2e8f0;margin:4px 8px;}
        """)
        for label, key in [("⎘  Copy Quote Number", "copy")]:
            a = QtWidgets.QAction(label, menu)
            a.triggered.connect(lambda _, k=key: (self.action_requested.emit(k), self.accept()))
            menu.addAction(a)
        menu.addSeparator()
        del_a = QtWidgets.QAction("🗑  Delete Quote", menu)
        del_a.triggered.connect(lambda: (self.action_requested.emit("delete"), self.accept()))
        menu.addAction(del_a)
        menu.exec_(QtGui.QCursor.pos())

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Escape:
            self.reject()
        else:
            super().keyPressEvent(event)


# Add this new class for Sales Person Management Dialog

class JobFormTab(QtWidgets.QWidget):
    convert_to_invoice = QtCore.pyqtSignal(dict)  # emits job_data dict
    convert_to_project = QtCore.pyqtSignal(dict)  # emits job_data dict

    def __init__(self, main_window, firebase_available=False):
        super().__init__()
        self.main_window = main_window
        
        self.client_filter_menu = QtWidgets.QMenu()
        self.selected_client_filter = "📂 All Clients"  # Changed from "All Clients"
        
        self.FIREBASE_AVAILABLE = firebase_available
        self.job_forms = []
        self._date_filter_active = False
        self._qf_page = 1
        self._qf_per_page = 10
        self._qf_all_items = []

        # Initialize template manager
        self.template_manager = TemplateManager()
        self.template_manager.initialize_default_templates()
        
        # Initialize client intelligence
        self.client_intelligence = ClientIntelligence()
        
        self.init_ui()
        self.load_job_forms_from_firebase()

        # Add real-time listener for quotes/job forms
        try:
            from main import FirebaseManager
            FirebaseManager.add_quotes_listener(self._on_quotes_updated)
        except Exception:
            pass
        
    def calculate_next_job_numbers_numeric(self):
        """Calculate next available quote numbers based ONLY on main sequence with proper rollover logic"""
        if not self.job_forms:
            _log.info("   Starting fresh: QuoteA001")
            return
        
        from collections import defaultdict
        
        # Dictionary to store highest sequence for each category
        category_sequences = defaultdict(int)
        
        # Find highest sequence for each category (ignoring suffixes)
        for job in self.job_forms:
            job_num = job.get('job_number', '').upper()
            
            if 'QUOTE' not in job_num:
                continue
            
            # Extract category and sequence
            # Pattern: QUOTE + optional category + numbers
            match = re.match(r'^QUOTE([A-Z]?)(\d+)', job_num)
            if match:
                category = match.group(1) or 'A'  # Default to 'A' if no category
                seq_str = match.group(2)
                
                try:
                    # Convert to integer (026 becomes 26)
                    seq_num = int(seq_str)
                    
                    # Update if this is higher than current max
                    if seq_num > category_sequences[category]:
                        category_sequences[category] = seq_num
                except ValueError:
                    continue
        
        _log.info("   Next Available Quote Numbers (by Category and Sequence ONLY):")
        _log.debug("---")
        
        if not category_sequences:
            _log.debug("No Quote numbers found. Starting at QuoteA001")
            return

        for category in sorted(category_sequences.keys()):
            current_max = category_sequences[category]
            if current_max >= 999:
                if category == 'Z':
                    all_max = max(category_sequences.values())
                    next_sequence = all_max + 1
                    next_category = 'A'
                    _log.debug("Z999 rollover — new cycle at: Quote%s%03d", next_category, next_sequence)
                else:
                    next_category = chr(ord(category) + 1)
                    if next_category in category_sequences:
                        next_sequence = category_sequences[next_category] + 1
                        if next_sequence >= 1000:
                            continue
                    else:
                        next_sequence = 1
                    _log.debug("Category rollover — next: Quote%s%03d", next_category, next_sequence)
            else:
                next_sequence = current_max + 1
                _log.debug("Category %s — highest: %03d, next: %03d", category, current_max, next_sequence)
                
    def update_client_filter_menu(self):
        """Update client filter menu with unique client names from CURRENTLY FILTERED quote overview"""
        if not hasattr(self, 'client_filter_menu'):
            return
        
        # First, get the currently filtered quote overview (based on all active filters)
        filtered_clients = set()
        
        # Apply all current filters to get visible quote overview
        client_filter = self.selected_client_filter
        search_text = self.search_edit.text().lower()
        status_filter = self.status_filter_combo.currentText()
        
        # Check if date range filter is active
        date_range_active = getattr(self, '_date_filter_active', False)
        from_date = None
        to_date = None
        
        if date_range_active and hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            from_date = self.current_from_date
            to_date = self.current_to_date
        
        for job in self.job_forms:
            # Client filter (for current selection)
            matches_client = (
                client_filter == "📂 All Clients" or
                client_filter == "All Clients" or
                job.get('client', '') == client_filter.replace("📂 All Clients", "All Clients")
            )
            
            # Search filter
            matches_search = (
                not search_text or  # If no search text, match all
                search_text in job.get('job_number', '').lower() or
                search_text in job.get('project_name', '').lower() or
                search_text in job.get('job_title', '').lower() or
                search_text in job.get('client', '').lower() or
                search_text in job.get('sales', '').lower() or
                search_text in job.get('scope_of_work', '').lower() or
                search_text in job.get('engineering_costs', '').lower()
            )

            # Status filter
            job_status = job.get('status', 'Not Started')
            matches_status = (
                status_filter == "All Status" or
                job_status == status_filter
            )

            # Date range filter
            matches_date = True
            if date_range_active and from_date and to_date:
                try:
                    # Try to parse job date (format: MM-dd-yyyy)
                    job_date_str = job.get('start_date', '')
                    if job_date_str:
                        job_date = QtCore.QDate.fromString(job_date_str, "MM-dd-yyyy")
                        if job_date.isValid():
                            matches_date = (from_date <= job_date <= to_date)
                        else:
                            matches_date = False
                    else:
                        matches_date = False
                except Exception as e:
                    matches_date = False
            
            # Check if job passes ALL current filters (except client filter for this purpose)
            # We want to include clients from jobs that match all OTHER filters
            include_job = matches_search and matches_status and matches_date
            
            if include_job:
                client = job.get('client', '')
                if client:  # Only add non-empty client names
                    filtered_clients.add(client)
        
        # Clear existing menu
        self.client_filter_menu.clear()
        
        # Add "All Clients" option with emoji
        all_action = self.client_filter_menu.addAction("📂 All Clients")
        all_action.triggered.connect(lambda: self.apply_client_filter("📂 All Clients"))
        
        # Add filtered client names to menu
        for client in sorted(filtered_clients):
            action_text = f"🏢 {client}"
            act = self.client_filter_menu.addAction(action_text)
            act.triggered.connect(lambda checked, c=client: self.apply_client_filter(c))
            
    def load_job_forms_from_firebase(self):
        """Load quote overview from Firebase with improved sorting and client filter update"""
        
        if not self.FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - loading quote overview from local backup")
            self.job_forms = self.load_job_forms_from_local()
            self.update_job_forms_table()
            self.update_client_filter_menu()
            return self.job_forms

        try:
            from main import db
            ref = db.reference('/job_forms')
            job_forms_data = ref.get()

            self.job_forms = []
            
            # Check for non-Quote quote numbers
            non_quote_jobs = []

            if job_forms_data:
                for job_id, job_data in job_forms_data.items():
                    if job_data:
                        job_data['firebase_id'] = job_id
                        # Ensure status field exists
                        if 'status' not in job_data:
                            job_data['status'] = 'Not Started'
                        
                        # Ensure sales field exists (handle both 'sales' and 'sales_person' keys)
                        if 'sales' not in job_data and 'sales_person' in job_data:
                            job_data['sales'] = job_data['sales_person']
                        elif 'sales' not in job_data:
                            job_data['sales'] = ''
                        
                        # Ensure client_email field exists
                        if 'client_email' not in job_data:
                            job_data['client_email'] = ''
                        
                        # Ensure client_address field exists
                        if 'client_address' not in job_data:
                            job_data['client_address'] = ''
                        
                        # Ensure project_name field exists
                        if 'project_name' not in job_data:
                            job_data['project_name'] = ''
                        
                        # Ensure plant field exists
                        if 'plant' not in job_data:
                            job_data['plant'] = ''

                        # Ensure client and job_number always exist
                        job_data.setdefault('client', '')
                        job_data.setdefault('job_number', '')

                        # Check if quote number starts with Quote
                        job_num = job_data.get('job_number', '').upper()
                        if not job_num.startswith('QUOTE'):
                            non_quote_jobs.append(job_num)
                        
                        self.job_forms.append(job_data)

                local_jobs = self.load_job_forms_from_local()
                known_numbers = {job.get('job_number', '').upper() for job in self.job_forms}
                for local_job in local_jobs:
                    if local_job.get('job_number', '').upper() not in known_numbers:
                        self.job_forms.append(local_job)
                        known_numbers.add(local_job.get('job_number', '').upper())

                # Warn about non-Quote quote numbers
                if non_quote_jobs:
                    _log.warning("⚠️ WARNING: Found %s non-Quote quote numbers:", len(non_quote_jobs))
                    for job_num in non_quote_jobs[:10]:  # Show first 10
                        _log.info("   - %s", job_num)
                    if len(non_quote_jobs) > 10:
                        _log.info("   ... and %s more", len(non_quote_jobs) - 10)
                    
                    _log.info("\n💡 Recommendation: Convert these to Quote format for proper sorting")
                
                # Define the sorting key function
                def job_number_sort_key(job):
                    """
                    Sort Quote quote numbers in the following order (DESCENDING - Highest first):
                    
                    Hierarchy:
                    1. Main sequence number (001-∞): 100 > 099 > 002 > 001
                    2. Variant type priority (higher appears first):
                    a. Numeric suffix (quoteA001_1000)
                    b. Numeric + Alphabetic suffix (quoteA001_3z)
                    c. Numeric + Alphabetic + Numeric (quoteA001_3a1000)
                    d. Alphabetic suffix (quoteA001_z)
                    e. Alphabetic + Numeric suffix (quoteA001_a1000)
                    f. Base (no suffix)
                    3. For same variant type, sort by suffix values
                    
                    Returns tuple for sorting in DESCENDING order (negative values for highest first)
                    """
                    job_num = job.get('job_number', '').strip().upper()
                    
                    if not job_num or 'QUOTE' not in job_num:
                        return (0, 0, 0, 0, 0, 0, 0, 0)  # Non-Quote goes to bottom

                    # Enhanced pattern to capture all variant types
                    pattern = r'^QUOTE([A-Z]?)(\d+)(?:_?([a-zA-Z]+)?(\d+)?)?(?:_?(\d+)([a-zA-Z]+)?(\d+)?)?$'
                    match = re.match(pattern, job_num, re.IGNORECASE)
                    
                    if not match:
                        return (0, 0, 0, 0, 0, 0, 0, 0)
                    
                    # Extract components
                    category = match.group(1) or 'A'
                    main_seq_str = match.group(2)
                    alpha1 = (match.group(3) or '').lower()  # First alphabetic suffix
                    num1 = match.group(4) or ''              # First numeric suffix
                    num2 = match.group(5) or ''              # Second numeric (for patterns like quoteA001_3a1)
                    alpha2 = (match.group(6) or '').lower()  # Second alphabetic
                    num3 = match.group(7) or ''              # Third numeric
                    
                    # 1. Category value (A=1, Z=26) - negative for descending
                    cat_value = -(ord(category) - 64) if category else -1
                    
                    # 2. Main sequence value
                    try:
                        main_seq = int(main_seq_str.lstrip('0')) if main_seq_str.lstrip('0') else 0
                    except:
                        main_seq = 0
                    main_seq = -main_seq  # Negative for descending
                    
                    # 3. Determine variant type and assign priority (higher number = appears higher)
                    variant_priority = 0
                    
                    if num2 and alpha2 and num3:
                        # Pattern: quoteA001_3a1000 (Numeric + Alphabetic + Numeric) - HIGHEST
                        variant_priority = -6
                    elif num2 and alpha2 and not num3:
                        # Pattern: quoteA001_3a (Numeric + Alphabetic)
                        variant_priority = -5
                    elif num2 and not alpha2:
                        # Pattern: quoteA001_1000 (Pure Numeric suffix)
                        variant_priority = -4
                    elif alpha1 and num1 and not (num2 or alpha2):
                        # Pattern: quoteA001_a1000 (Alphabetic + Numeric)
                        variant_priority = -3
                    elif alpha1 and not num1:
                        # Pattern: quoteA001_a (Pure Alphabetic)
                        variant_priority = -2
                    elif not (alpha1 or num1 or num2 or alpha2 or num3):
                        # Base: quoteA001 (no suffix) - LOWEST
                        variant_priority = -1
                    
                    # 4. Convert suffixes to sortable values
                    # First alphabetic suffix (a=1, z=26, aa=27, etc.)
                    alpha1_value = 0
                    if alpha1:
                        for i, char in enumerate(reversed(alpha1)):
                            char_val = ord(char) - 96  # a=1, z=26
                            alpha1_value += char_val * (26 ** i)
                    alpha1_value = -alpha1_value  # Negative for descending
                    
                    # First numeric suffix
                    num1_value = 0
                    try:
                        num1_value = -int(num1) if num1 else 0
                    except:
                        num1_value = 0
                    
                    # Second numeric (for patterns like quoteA001_3)
                    num2_value = 0
                    try:
                        num2_value = -int(num2) if num2 else 0
                    except:
                        num2_value = 0
                    
                    # Second alphabetic (for patterns like quoteA001_3a)
                    alpha2_value = 0
                    if alpha2:
                        for i, char in enumerate(reversed(alpha2)):
                            char_val = ord(char) - 96
                            alpha2_value += char_val * (26 ** i)
                    alpha2_value = -alpha2_value
                    
                    # Third numeric (for patterns like quoteA001_3a1000)
                    num3_value = 0
                    try:
                        num3_value = -int(num3) if num3 else 0
                    except:
                        num3_value = 0
                    
                    # Return tuple for sorting (all negative for descending order)
                    return (
                        cat_value,        # Category A-Z (A first, Z last in descending)
                        main_seq,         # Main sequence (001, 002, etc.)
                        variant_priority, # Variant type priority
                        num2_value,       # First numeric in suffix (for quoteA001_1000)
                        alpha2_value,     # Alphabetic after numeric (for quoteA001_3a)
                        num3_value,       # Numeric after alphabetic (for quoteA001_3a1000)
                        alpha1_value,     # Pure alphabetic suffix (for quoteA001_a)
                        num1_value        # Numeric after alphabetic (for quoteA001_a1)
                    )
                
                # Sort using the key function
                self.job_forms.sort(key=job_number_sort_key)

                _log.info("Loaded %s quote forms from Firebase", len(self.job_forms))
                _log.debug("Sorted order (first 10): %s",
                           [j.get('job_number') for j in self.job_forms[:10]])
                self.calculate_next_job_numbers_numeric()
                
                # Update the table with loaded forms
                self.update_job_forms_table()
                
                # ⭐⭐ CRITICAL: Update client filter menu after loading quote forms
                self.update_client_filter_menu()
                
                return self.job_forms

            else:
                _log.info("No quote forms found in Firebase")
                self.job_forms = self.load_job_forms_from_local()
                self.update_job_forms_table()
                self.update_client_filter_menu()
                return self.job_forms

        except Exception as e:
            _log.warning("Error loading quote forms from Firebase: %s", e)
            _log.exception("Traceback:")
            self.job_forms = self.load_job_forms_from_local()
            self.update_job_forms_table()
            self.update_client_filter_menu()
            return self.job_forms

    def _on_quotes_updated(self, quotes_data):
        """Called when quotes/job forms are updated in Firebase - updates UI automatically"""
        try:
            QtCore.QTimer.singleShot(300, self.load_job_forms_from_firebase)
        except Exception as e:
            _log.warning("Error updating quotes in real-time: %s", e)


    def save_job_form_to_firebase(self, job_data):
        """Save quote form to Firebase - UPDATE existing job instead of creating new one"""
        
        if not self.FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - saving quote form to local backup")
            return self.save_job_form_locally(job_data)
            
        try:
            from main import db
            ref = db.reference('/job_forms')
            
            # ⭐⭐ FIXED: Check if job already exists (by firebase_id if editing, otherwise by job_number)
            if 'firebase_id' in job_data and job_data['firebase_id']:
                # Update existing job using firebase_id
                job_id = job_data['firebase_id']
                job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(job_id).update(job_data)
                _log.info("Quote form UPDATED in Firebase: %s (ID: %s)", job_data['job_number'], job_id)
                return True
            else:
                # Check if job with same job_number already exists (for new jobs with manual quote numbers)
                existing_jobs = ref.order_by_child('job_number').equal_to(job_data['job_number']).get()
                
                if existing_jobs:
                    # Update existing job found by job_number
                    job_id = list(existing_jobs.keys())[0]
                    job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    ref.child(job_id).update(job_data)
                    _log.info("Quote form UPDATED in Firebase by quote number: %s", job_data['job_number'])
                    return True
                else:
                    # Create new job
                    new_job_ref = ref.push()
                    job_data['firebase_id'] = new_job_ref.key
                    job_data['created_at'] = datetime.now(timezone.utc).isoformat()
                    job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    new_job_ref.set(job_data)
                    _log.info("✓Quote form CREATED in Firebase with ID: %s", new_job_ref.key)
                    return True
                    
        except Exception as e:
            _log.warning("Error saving Quote form to Firebase: %s", e)
            _log.exception("Traceback:")
            return self.save_job_form_locally(job_data)

    def local_job_forms_path(self):
        """Return the local quote-form backup path."""
        path = Path(__file__).resolve().parent / "data" / "job_forms.json"
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def load_job_forms_from_local(self):
        """Load quote forms saved locally when Firebase is unavailable."""
        try:
            import json
            path = self.local_job_forms_path()
            if not path.exists():
                return []
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            jobs = data if isinstance(data, list) else list(data.values())
            for idx, job in enumerate(jobs):
                if isinstance(job, dict):
                    job.setdefault("local_id", f"local_{idx}")
                    job.setdefault("status", "Not Started")
                    job.setdefault("client_email", "")
                    job.setdefault("client_address", "")
                    job.setdefault("project_name", "")
                    job.setdefault("plant", "")
                    job.setdefault("sales", "")
            _log.info("Loaded %s quote forms from local backup", len(jobs))
            return [job for job in jobs if isinstance(job, dict)]
        except Exception as e:
            _log.warning("Error loading local quote form backup: %s", e)
            _log.exception("Traceback:")
            return []

    def save_job_form_locally(self, job_data):
        """Save or update a quote form in local JSON backup."""
        try:
            import json
            path = self.local_job_forms_path()
            jobs = self.load_job_forms_from_local()
            job_number = job_data.get("job_number", "").strip()
            if not job_number:
                _log.warning("Cannot save local quote backup without job_number")
                return False

            saved = False
            for idx, job in enumerate(jobs):
                if job.get("job_number", "").upper() == job_number.upper():
                    merged = dict(job)
                    merged.update(job_data)
                    merged["updated_at"] = datetime.now(timezone.utc).isoformat()
                    jobs[idx] = merged
                    saved = True
                    break

            if not saved:
                local_job = dict(job_data)
                local_job.setdefault("created_at", datetime.now(timezone.utc).isoformat())
                local_job["updated_at"] = datetime.now(timezone.utc).isoformat()
                local_job["local_id"] = f"local_{job_number}"
                jobs.append(local_job)

            with path.open("w", encoding="utf-8") as f:
                json.dump(jobs, f, indent=2)

            _log.info("Quote form saved to local backup: %s", job_number)
            return True
        except Exception as e:
            _log.warning("Error saving quote form to local backup: %s", e)
            _log.exception("Traceback:")
            return False
    
    def delete_job_form_from_firebase(self, job_data):
        """Delete Quote form from Firebase - handles both firebase_id and job_number"""
        if not self.FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot delete quote form")
            return False
            
        try:
            from main import db
            
            # Try firebase_id first
            if 'firebase_id' in job_data and job_data['firebase_id']:
                ref = db.reference(f'/job_forms/{job_data["firebase_id"]}')
                ref.delete()
                _log.info("Quote form deleted from Firebase: %s (ID: %s)", job_data['job_number'], job_data['firebase_id'])
                return True
            else:
                # Fallback: find by job_number
                ref = db.reference('/job_forms')
                existing_jobs = ref.order_by_child('job_number').equal_to(job_data['job_number']).get()
                
                if existing_jobs:
                    job_id = list(existing_jobs.keys())[0]
                    ref.child(job_id).delete()
                    _log.info("Quote form deleted from Firebase by quote number: %s", job_data['job_number'])
                    return True
                else:
                    _log.warning("No Firebase ID or matching quote number found for job: %s", job_data['job_number'])
                    return False
                    
        except Exception as e:
            _log.warning("Error deleting quote form from Firebase: %s", e)
            return False
   
    
    def show_job_form_dialog(self):
        """Show professional quote form creation dialog"""
        # Store current filter state BEFORE opening dialog
        search_text = self.search_edit.text()
        status_filter = self.status_filter_combo.currentText()
        client_filter = self.selected_client_filter
        date_range_active = getattr(self, '_date_filter_active', False)
        
        if date_range_active and hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            from_date = self.current_from_date
            to_date = self.current_to_date
        else:
            from_date = None
            to_date = None
        
        # Pass Firebase availability to the dialog
        dialog = JobFormDialog(self.main_window, self, firebase_available=self.FIREBASE_AVAILABLE)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Refresh table after creating new quote form
            self.load_job_forms_from_firebase()
            
            # ⭐⭐ RESTORE ALL FILTERS after refresh
            QtCore.QTimer.singleShot(100, lambda: self.restore_filters(
                search_text, status_filter, client_filter, from_date, to_date
            ))

    def restore_filters(self, search_text, status_filter, client_filter, from_date, to_date):
        """Restore all filters to their previous state"""
        self.search_edit.setText(search_text)
        self.status_filter_combo.setCurrentText(status_filter)
        self.selected_client_filter = client_filter
        
        # Restore date filter if it was active
        if from_date and to_date:
            self.current_from_date = from_date
            self.current_to_date = to_date
            self._date_filter_active = True
            from_date_formatted = from_date.toString("MM-dd-yy")
            to_date_formatted = to_date.toString("MM-dd-yy")
            self.date_range_button.setText(f"📅 {from_date_formatted} to {to_date_formatted}")
            self.date_range_button.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    font-size: 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
                QPushButton:pressed {
                    background-color: #21618c;
                }
            """)
        
        # Apply the restored filters
        self.filter_job_forms()
        
    def init_ui(self):
        """Initialize layout — no outer scroll; the table handles its own scrolling."""
        outer_layout = QtWidgets.QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(0)

        self.workflow_tabs = QtWidgets.QTabWidget()
        self.workflow_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background: #F8FAFC;
                top: -1px;
            }
            QTabBar::tab {
                background: #FFFFFF;
                color: #475569;
                border: 1px solid #DCE4EC;
                border-bottom: 2px solid #DCE4EC;
                padding: 10px 22px;
                margin-right: 6px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
                min-width: 120px;
            }
            QTabBar::tab:selected {
                color: #0F766E;
                border-bottom: 3px solid #0F766E;
                background: #F8FAFC;
            }
            QTabBar::tab:hover:!selected {
                color: #111827;
                background: #F8FAFC;
            }
        """)
        outer_layout.addWidget(self.workflow_tabs)

        # All Quotes tab — wrapped in a scroll area so all content is reachable
        self.all_quotes_tab = QtWidgets.QWidget()
        _aq_outer = QtWidgets.QVBoxLayout(self.all_quotes_tab)
        _aq_outer.setContentsMargins(0, 0, 0, 0)
        _aq_outer.setSpacing(0)
        _aq_scroll = QtWidgets.QScrollArea()
        _aq_scroll.setWidgetResizable(True)
        _aq_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        _aq_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        _aq_scroll.setStyleSheet(
            "QScrollArea { background: #F6F8FB; border: none; }"
            "QScrollBar:vertical { width: 8px; background: #F1F5F9; border-radius: 4px; }"
            "QScrollBar::handle:vertical { background: #CBD5E1; border-radius: 4px; min-height: 30px; }"
            "QScrollBar::handle:vertical:hover { background: #94A3B8; }"
            "QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }"
        )
        _aq_inner = QtWidgets.QWidget()
        _aq_inner.setStyleSheet("background: #F6F8FB;")
        self.all_quotes_layout = QtWidgets.QVBoxLayout(_aq_inner)
        self.all_quotes_layout.setContentsMargins(18, 14, 18, 14)
        self.all_quotes_layout.setSpacing(12)
        _aq_scroll.setWidget(_aq_inner)
        _aq_outer.addWidget(_aq_scroll)

        self.workflow_tabs.addTab(self.all_quotes_tab, "All Quotes")
        self.workflow_tabs.addTab(self._build_new_quote_tab(), "New Quote")
        self.workflow_tabs.addTab(self._build_sales_people_tab(), "Sales People")
        self.workflow_tabs.addTab(self._build_export_tab(), "Export")
        self.workflow_tabs.currentChanged.connect(self._on_workflow_tab_changed)

        self.create_stats_section(self.all_quotes_layout)
        self.create_job_forms_table_section(self.all_quotes_layout)

    def _on_workflow_tab_changed(self, index: int):
        """Refresh visible Quote Forms data when switching inner tabs."""
        label = self.workflow_tabs.tabText(index)
        if label == "All Quotes":
            self.load_job_forms_from_firebase()
        elif label == "New Quote":
            form = getattr(self, "embedded_quote_form", None)
            if form and hasattr(form, "load_sales_persons"):
                form.load_sales_persons()
        elif label == "Sales People":
            reload_sales = getattr(self, "_reload_sales_people", None)
            if callable(reload_sales):
                reload_sales()
        elif label == "Export":
            exp_dlg = getattr(self, "embedded_export", None)
            if exp_dlg:
                # Scroll back to top
                p = exp_dlg.parent()
                if isinstance(p, QtWidgets.QScrollArea):
                    p.verticalScrollBar().setValue(0)
                # Refresh record counts and recent exports
                try:
                    exp_dlg.refresh_recent_exports()
                    exp_dlg._refresh_export_filter_options("pdf")
                    exp_dlg._refresh_export_filter_options("excel")
                    if hasattr(exp_dlg, '_excel_rec_lbl'):
                        exp_dlg._excel_rec_lbl.setText(f"{len(self.job_forms or [])} Quotes")
                    if hasattr(exp_dlg, '_pdf_rec_lbl'):
                        exp_dlg._pdf_rec_lbl.setText(f"{len(self.job_forms or [])} Quotes")
                except Exception:
                    pass

    def create_header_section(self, layout):
        header_frame = QtWidgets.QFrame()
        header_frame.setMinimumHeight(86)
        header_frame.setStyleSheet("""
            QFrame {
                background: #FFFFFF;
                border: 1px solid #DCE4EC;
                border-left: 5px solid #0F766E;
                border-radius: 10px;
            }
        """)
        
        # Use QHBoxLayout to place title and export button on same row
        header_layout = QtWidgets.QHBoxLayout(header_frame)
        header_layout.setContentsMargins(22, 14, 16, 14)
        header_layout.setSpacing(16)
        
        # Left side: Title and subtitle
        left_container = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(5)
        
        title = QtWidgets.QLabel("Quote Form Management")
        title.setStyleSheet("""
            QLabel {
                color: #111827;
                font-size: 23px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                background: transparent;
                border: none;
                padding: 0;
            }
        """)

        subtitle = QtWidgets.QLabel("Manage and create professional quote forms for engineering projects")
        subtitle.setStyleSheet("""
            QLabel {
                color: #64748B;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-weight: 500;
                background: transparent;
                border: none;
                padding: 0;
            }
        """)
        
        left_layout.addWidget(title)
        left_layout.addWidget(subtitle)
        
        # Right side: Export button
        self.export_btn = QtWidgets.QPushButton("Export")
        self.export_btn.setFixedSize(110, 38)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: #FFFFFF;
                color: #334155;
                border: 1.5px solid #CBD5E1;
                border-radius: 8px;
                font-weight: 800;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 6px 14px;
            }
            QPushButton:hover {
                background: #F8FAFC;
                border-color: #0F766E;
                color: #0F766E;
            }
            QPushButton:pressed {
                background: #ECFDF5;
            }
        """)
        self.export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.export_btn.clicked.connect(self.open_job_forms_pdf_export_dialog)
        
        header_layout.addWidget(left_container, stretch=1)
        header_layout.addWidget(self.export_btn, alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        
        layout.addWidget(header_frame)
    
    def open_job_forms_pdf_export_dialog(self):
        """Open PDF/Excel export dialog for quote forms with tabs"""
        try:
            # Collect available dates from quote forms
            available_dates = []
            for job in self.job_forms:
                try:
                    job_date = datetime.strptime(job.get('start_date', ''), "%Y-%m-%d")
                    available_dates.append(job_date)
                except (ValueError, TypeError):
                    continue
            
            # Create and show the export dialog
            dialog = JobFormsExportDialog(self, available_dates)
            result = dialog.exec_()
            
            # Only perform export if dialog was accepted AND has export parameters
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                # Get export parameters and perform actual export
                export_params = dialog._export_params
                if export_params["type"] == "pdf":
                    self.perform_job_forms_pdf_export(export_params)
                elif export_params["type"] == "excel":
                    self.perform_job_forms_excel_export(export_params)
                        
        except Exception as e:
            _log.warning("Error opening export dialog: %s", e)
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error opening export dialog: {str(e)}")

    def _build_action_page(self, title, subtitle, button_text, button_color, callback):
        page = QtWidgets.QFrame()
        page.setStyleSheet("""
            QFrame {
                background: #FFFFFF;
                border: 1px solid #DCE4EC;
                border-radius: 10px;
            }
        """)
        layout = QtWidgets.QVBoxLayout(page)
        layout.setContentsMargins(28, 28, 28, 28)
        layout.setSpacing(14)

        title_label = QtWidgets.QLabel(title)
        title_label.setStyleSheet("""
            QLabel {
                color: #111827;
                font-size: 20px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI';
                background: transparent;
                border: none;
            }
        """)
        subtitle_label = QtWidgets.QLabel(subtitle)
        subtitle_label.setWordWrap(True)
        subtitle_label.setStyleSheet("""
            QLabel {
                color: #64748B;
                font-size: 13px;
                font-weight: 600;
                font-family: 'Inter', 'Segoe UI';
                background: transparent;
                border: none;
            }
        """)
        button = QtWidgets.QPushButton(button_text)
        button.setFixedSize(180, 40)
        button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        button.setStyleSheet(f"""
            QPushButton {{
                background: {button_color};
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
            }}
            QPushButton:hover {{
                background: #115E59;
            }}
        """)
        button.clicked.connect(callback)

        layout.addWidget(title_label)
        layout.addWidget(subtitle_label)
        layout.addSpacing(6)
        layout.addWidget(button)
        layout.addStretch()
        return page

    def _build_new_quote_tab(self):
        page = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.embedded_quote_form = self._create_full_quote_form()
        layout.addWidget(self.embedded_quote_form)
        return page

    def _create_full_quote_form(self):
        form = JobFormDialog(
            self.main_window,
            self,
            firebase_available=self.FIREBASE_AVAILABLE,
        )
        form.setWindowFlags(QtCore.Qt.Widget)
        form.setModal(False)
        form.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)

        def after_saved():
            self.load_job_forms_from_firebase()
            self.workflow_tabs.setCurrentIndex(0)
            QtCore.QTimer.singleShot(400, self._rebuild_full_quote_form)

        form.accept = after_saved
        try:
            form.cancel_btn.clicked.disconnect()
        except Exception:
            pass
        form.cancel_btn.setText("Back to Quotes")
        form.cancel_btn.clicked.connect(lambda: self.workflow_tabs.setCurrentIndex(0))

        # Re-generate quote number after Firebase data is loaded.
        # auto_generate_enabled is set False during __init__ (before job_forms loads),
        # so we reset it here and call generate_job_number directly.
        def _refresh_quote_number(f=form):
            f.auto_generate_enabled = True
            f.generate_job_number()

        QtCore.QTimer.singleShot(300, _refresh_quote_number)
        return form

    def _rebuild_full_quote_form(self):
        if not hasattr(self, "embedded_quote_form"):
            return
        parent = self.embedded_quote_form.parentWidget()
        if not parent or not parent.layout():
            return
        layout = parent.layout()
        old = self.embedded_quote_form
        layout.removeWidget(old)
        old.deleteLater()
        self.embedded_quote_form = self._create_full_quote_form()
        layout.addWidget(self.embedded_quote_form)

    def _build_sales_people_tab(self):
        _AVATAR_PAL = [
            ("#3B82F6", "#FFFFFF"), ("#10B981", "#FFFFFF"), ("#6366F1", "#FFFFFF"),
            ("#F59E0B", "#FFFFFF"), ("#EF4444", "#FFFFFF"), ("#8B5CF6", "#FFFFFF"),
            ("#EC4899", "#FFFFFF"), ("#14B8A6", "#FFFFFF"),
        ]

        def _av_color(name):
            return _AVATAR_PAL[hash(name or "?") % len(_AVATAR_PAL)]

        def _initials(name):
            parts = str(name or "").split()
            return (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else str(name or "?")[:2].upper()

        page = QtWidgets.QWidget()
        page.setStyleSheet("QWidget { background: #F8FAFC; }")
        outer = QtWidgets.QVBoxLayout(page)
        outer.setContentsMargins(32, 24, 32, 24)
        outer.setSpacing(18)

        # ── Header ────────────────────────────────────────────────
        hdr = QtWidgets.QHBoxLayout()
        tc = QtWidgets.QVBoxLayout()
        tc.setSpacing(3)
        t1 = QtWidgets.QLabel("Sales People")
        t1.setStyleSheet(
            "font-size:22px; font-weight:900; color:#111827;"
            " background:transparent; border:none; font-family:'Inter','Segoe UI';"
        )
        t2 = QtWidgets.QLabel("Manage your sales team and their contact information")
        t2.setStyleSheet(
            "font-size:13px; color:#6B7280; background:transparent; border:none;"
            " font-family:'Inter','Segoe UI';"
        )
        tc.addWidget(t1)
        tc.addWidget(t2)
        add_btn = QtWidgets.QPushButton("+ Add Sales Person")
        add_btn.setFixedHeight(40)
        add_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        add_btn.setStyleSheet("""
            QPushButton {
                background:#0F766E; color:white; border:none; border-radius:8px;
                font-size:13px; font-weight:800; padding:0 20px;
                font-family:'Inter','Segoe UI';
            }
            QPushButton:hover { background:#115E59; }
        """)
        hdr.addLayout(tc)
        hdr.addStretch()
        hdr.addWidget(add_btn)
        outer.addLayout(hdr)

        # ── Search bar ────────────────────────────────────────────
        scard = QtWidgets.QFrame()
        scard.setStyleSheet(
            "QFrame { background:white; border:1px solid #E2E8F0; border-radius:10px; }"
        )
        sl = QtWidgets.QHBoxLayout(scard)
        sl.setContentsMargins(16, 10, 16, 10)
        sl.setSpacing(10)

        unified_search = QtWidgets.QLineEdit()
        unified_search.setPlaceholderText("🔍  Search by name, phone, or email...")
        unified_search.setMinimumHeight(40)
        unified_search.setStyleSheet("""
            QLineEdit {
                background:white; border:1.5px solid #E2E8F0; border-radius:8px;
                padding:9px 12px; font-size:13px; color:#374151;
                font-family:'Inter','Segoe UI';
            }
            QLineEdit:focus { border-color:#0F766E; }
        """)

        sp_count_lbl = QtWidgets.QLabel("0 Sales")
        sp_count_lbl.setFixedHeight(40)
        sp_count_lbl.setMinimumWidth(90)
        sp_count_lbl.setStyleSheet("""
            QLabel {
                background:#F1F5F9; border:1px solid #E2E8F0; border-radius:7px;
                color:#475569; font-size:12px; font-weight:700;
                font-family:'Inter','Segoe UI'; padding:0 14px;
            }
        """)

        sl.addWidget(unified_search, 1)
        sl.addWidget(sp_count_lbl)
        outer.addWidget(scard)

        # ── Table ─────────────────────────────────────────────────
        table = QtWidgets.QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["NAME", "PHONE", "EMAIL", "ACTIONS"])
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setFocusPolicy(QtCore.Qt.NoFocus)
        table.setStyleSheet("""
            QTableWidget {
                background:white; border:1px solid #E2E8F0; border-radius:10px;
                gridline-color:#F1F5F9; font-size:13px; font-family:'Inter','Segoe UI';
            }
            QTableWidget::item { padding:0; color:#1E293B; border-bottom:1px solid #F1F5F9; }
            QTableWidget::item:selected { background:#F0FDF9; color:#111827; }
            QHeaderView::section {
                background:#F8FAFC; color:#6B7280; padding:10px 16px;
                border:none; border-bottom:1px solid #E2E8F0;
                font-weight:700; font-size:12px; font-family:'Inter','Segoe UI';
            }
        """)
        for c in range(3):
            table.horizontalHeader().setSectionResizeMode(c, QtWidgets.QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
        table.setColumnWidth(3, 160)
        table.verticalHeader().setDefaultSectionSize(64)
        outer.addWidget(table, 1)
        self.sales_people_table = table

        # ── Footer ────────────────────────────────────────────────
        ftr = QtWidgets.QHBoxLayout()
        count_lbl = QtWidgets.QLabel("Showing 0 entries")
        count_lbl.setStyleSheet(
            "font-size:13px; color:#6B7280; background:transparent; border:none;"
            " font-family:'Inter','Segoe UI';"
        )
        _pg_s = """
            QPushButton {
                background:white; color:#374151; border:1px solid #E2E8F0;
                border-radius:6px; font-size:13px; font-weight:700;
            }
            QPushButton:hover { background:#F8FAFC; }
            QPushButton:disabled { color:#D1D5DB; background:#F9FAFB; }
        """
        prev_btn = QtWidgets.QPushButton("<")
        prev_btn.setFixedSize(32, 32)
        prev_btn.setStyleSheet(_pg_s)
        prev_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        _sp_page_btns_w = QtWidgets.QWidget()
        _sp_page_btns_w.setStyleSheet("background:transparent;")
        self._sp_page_btns_layout = QtWidgets.QHBoxLayout(_sp_page_btns_w)
        self._sp_page_btns_layout.setContentsMargins(0, 0, 0, 0)
        self._sp_page_btns_layout.setSpacing(4)
        next_btn = QtWidgets.QPushButton(">")
        next_btn.setFixedSize(32, 32)
        next_btn.setStyleSheet(_pg_s)
        next_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        ftr.addWidget(count_lbl)
        ftr.addStretch()
        ftr.addWidget(prev_btn)
        ftr.addWidget(_sp_page_btns_w)
        ftr.addWidget(next_btn)
        outer.addLayout(ftr)

        # ── State ─────────────────────────────────────────────────
        self._sales_people = []
        self._sp_filtered = []
        self._sp_page = 0
        _PER_PAGE = 10

        # ── Render page ───────────────────────────────────────────
        def render_page(filtered, page_no=0):
            self._sp_filtered = filtered
            self._sp_page = page_no
            total = len(filtered)
            start = page_no * _PER_PAGE
            end = min(start + _PER_PAGE, total)
            visible = filtered[start:end]
            table.setRowCount(len(visible))

            for row, person in enumerate(visible):
                table.setRowHeight(row, 64)
                name = person.get("name", "") or ""
                phone = person.get("phone", "") or ""
                email = person.get("email", "") or ""

                # Col 0: Avatar + Name
                cw0 = QtWidgets.QWidget()
                cw0.setStyleSheet("background:transparent;")
                cl0 = QtWidgets.QHBoxLayout(cw0)
                cl0.setContentsMargins(12, 0, 12, 0)
                cl0.setSpacing(10)
                bg, fg = _av_color(name)
                av = QtWidgets.QLabel(_initials(name))
                av.setFixedSize(38, 38)
                av.setAlignment(QtCore.Qt.AlignCenter)
                av.setStyleSheet(
                    f"background:{bg}; color:{fg}; border-radius:19px;"
                    " font-size:13px; font-weight:900; font-family:'Inter','Segoe UI';"
                )
                nm = QtWidgets.QLabel(name)
                nm.setStyleSheet(
                    "font-size:13px; font-weight:600; color:#111827;"
                    " background:transparent; border:none; font-family:'Inter','Segoe UI';"
                )
                cl0.addWidget(av)
                cl0.addWidget(nm)
                cl0.addStretch()
                table.setCellWidget(row, 0, cw0)

                # Col 1: Phone
                cw1 = QtWidgets.QWidget()
                cw1.setStyleSheet("background:transparent;")
                cl1 = QtWidgets.QHBoxLayout(cw1)
                cl1.setContentsMargins(16, 0, 16, 0)
                cl1.setSpacing(6)
                pi = QtWidgets.QLabel("✆")
                pi.setStyleSheet("color:#10B981; font-size:14px; background:transparent; border:none;")
                pl = QtWidgets.QLabel(phone if phone and phone != "-" else "—")
                pl.setStyleSheet(
                    "font-size:13px; color:#374151; background:transparent;"
                    " border:none; font-family:'Inter','Segoe UI';"
                )
                cl1.addWidget(pi)
                cl1.addWidget(pl)
                cl1.addStretch()
                table.setCellWidget(row, 1, cw1)

                # Col 2: Email
                cw2 = QtWidgets.QWidget()
                cw2.setStyleSheet("background:transparent;")
                cl2 = QtWidgets.QHBoxLayout(cw2)
                cl2.setContentsMargins(16, 0, 16, 0)
                cl2.setSpacing(6)
                ei = QtWidgets.QLabel("✉")
                ei.setStyleSheet("color:#6B7280; font-size:14px; background:transparent; border:none;")
                el = QtWidgets.QLabel(email or "—")
                el.setStyleSheet(
                    "font-size:13px; color:#374151; background:transparent;"
                    " border:none; font-family:'Inter','Segoe UI';"
                )
                cl2.addWidget(ei)
                cl2.addWidget(el)
                cl2.addStretch()
                table.setCellWidget(row, 2, cw2)

                # Col 3: Actions button
                cw3 = QtWidgets.QWidget()
                cw3.setStyleSheet("background:transparent;")
                cl3 = QtWidgets.QHBoxLayout(cw3)
                cl3.setContentsMargins(16, 0, 16, 0)
                cl3.setAlignment(QtCore.Qt.AlignCenter)
                ab = QtWidgets.QPushButton("Actions  ▾")
                ab.setFixedSize(110, 32)
                ab.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                ab.setStyleSheet("""
                    QPushButton {
                        background:white; color:#1E293B; border:1.5px solid #E2E8F0;
                        border-radius:7px; font-size:12px; font-weight:700;
                        font-family:'Inter','Segoe UI'; padding:0 8px;
                    }
                    QPushButton:hover { background:#F1F5F9; border-color:#CBD5E1; }
                """)

                def _menu(checked=False, b=ab, p=person):
                    m = QtWidgets.QMenu(b)
                    m.setStyleSheet("""
                        QMenu { background:white; border:1px solid #E2E8F0; border-radius:8px;
                                padding:4px 0; font-family:'Inter','Segoe UI'; font-size:13px; }
                        QMenu::item { padding:9px 20px; color:#1E293B; }
                        QMenu::item:selected { background:#EFF6FF; color:#1D4ED8; }
                        QMenu::separator { height:1px; background:#EEF2F6; margin:3px 8px; }
                    """)
                    ea = QtWidgets.QAction("✏️  Edit", m)
                    ea.triggered.connect(lambda: open_dialog(p))
                    da = QtWidgets.QAction("🗑️  Delete", m)
                    da.triggered.connect(lambda: delete_sp(p))
                    m.addAction(ea)
                    m.addSeparator()
                    if p.get("user_uid") and not p.get("firebase_id"):
                        ea.setEnabled(False)
                        da.setEnabled(False)
                    m.addAction(da)
                    m.exec_(b.mapToGlobal(QtCore.QPoint(0, b.height())))

                ab.clicked.connect(_menu)
                cl3.addWidget(ab)
                table.setCellWidget(row, 3, cw3)

            if total == 0:
                count_lbl.setText("No entries found")
            else:
                count_lbl.setText(f"Showing {start + 1} to {end} of {total} entries")
            while self._sp_page_btns_layout.count():
                _it = self._sp_page_btns_layout.takeAt(0)
                if _it.widget():
                    _it.widget().deleteLater()
            _max_pg = max(1, (total + _PER_PAGE - 1) // _PER_PAGE)
            _pg_num = page_no + 1
            _win_start = max(1, min(_pg_num, _max_pg - 2))
            for _p in range(_win_start, min(_win_start + 3, _max_pg + 1)):
                _pb = QtWidgets.QPushButton(str(_p))
                _pb.setFixedSize(32, 32)
                _pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                if _p == _pg_num:
                    _pb.setStyleSheet("""QPushButton {
                        background:#0F766E; color:white; border:none; border-radius:6px;
                        font-size:13px; font-weight:800;
                    }
                    QPushButton:hover { background:#115E59; color:white; }""")
                else:
                    _pb.setStyleSheet(_pg_s)
                    _pb.clicked.connect(lambda _, pg=_p: render_page(self._sp_filtered, pg - 1))
                self._sp_page_btns_layout.addWidget(_pb)
            prev_btn.setEnabled(page_no > 0)
            next_btn.setEnabled(end < total)

        # ── Search ────────────────────────────────────────────────
        def apply_search():
            q = unified_search.text().strip().lower()
            filtered = [
                sp for sp in self._sales_people
                if (not q
                    or q in (sp.get("name", "") or "").lower()
                    or q in (sp.get("phone", "") or "").lower()
                    or q in (sp.get("email", "") or "").lower())
            ]
            total = len(filtered)
            sp_count_lbl.setText(f"{total} Sale{'s' if total != 1 else ''}")
            render_page(filtered, 0)

        unified_search.textChanged.connect(apply_search)

        # ── Load ──────────────────────────────────────────────────
        def load_sales():
            try:
                if self.FIREBASE_AVAILABLE:
                    from main import FirebaseManager
                    data = FirebaseManager.load_sales_people()
                else:
                    data = _load_local_sales_people()
            except Exception as exc:
                _log.warning("Error loading sales people: %s", exc)
                data = list(getattr(self, '_sales_people', None) or [])
            self._sales_people = sorted(
                data, key=lambda sp: (sp.get("name", "") or "").lower())
            apply_search()
            for wid in (
                getattr(self, "embedded_quote_form", None),
                getattr(getattr(self, "main_window", None), "project_tab", None),
            ):
                if wid and hasattr(wid, "load_sales_persons"):
                    wid.load_sales_persons()

        # ── Add / Edit dialog ─────────────────────────────────────
        def open_dialog(person=None):
            dlg = QtWidgets.QDialog(self)
            dlg.setWindowTitle("Edit Sales Person" if person else "Add Sales Person")
            dlg.setMinimumWidth(420)
            dlg.setModal(True)
            dlg.setStyleSheet("QDialog { background:white; }")
            vl = QtWidgets.QVBoxLayout(dlg)
            vl.setContentsMargins(24, 20, 24, 20)
            vl.setSpacing(14)
            ttl = QtWidgets.QLabel("Edit Sales Person" if person else "Add Sales Person")
            ttl.setStyleSheet(
                "font-size:17px; font-weight:900; color:#111827;"
                " font-family:'Inter','Segoe UI';"
            )
            vl.addWidget(ttl)
            _FLD = """
                QLineEdit {
                    background:white; border:1.5px solid #E2E8F0; border-radius:8px;
                    padding:9px 12px; font-size:13px; color:#374151;
                    font-family:'Inter','Segoe UI';
                }
                QLineEdit:focus { border-color:#0F766E; }
            """

            def _row(lbl_t, ph, val=""):
                rw = QtWidgets.QWidget()
                rw.setStyleSheet("background:transparent;")
                rl = QtWidgets.QVBoxLayout(rw)
                rl.setContentsMargins(0, 0, 0, 0)
                rl.setSpacing(4)
                l = QtWidgets.QLabel(lbl_t)
                l.setStyleSheet(
                    "font-size:12px; font-weight:700; color:#374151; background:transparent;"
                )
                e = QtWidgets.QLineEdit()
                e.setPlaceholderText(ph)
                e.setText(val)
                e.setMinimumHeight(40)
                e.setStyleSheet(_FLD)
                rl.addWidget(l)
                rl.addWidget(e)
                return rw, e

            nrw, ne = _row("Name *", "Enter full name", person.get("name", "") if person else "")
            prw, pe = _row("Phone", "Enter phone number", (person.get("phone", "") or "") if person else "")
            erw, ee = _row("Email", "Enter email address", person.get("email", "") if person else "")
            vl.addWidget(nrw)
            vl.addWidget(prw)
            vl.addWidget(erw)

            _dlg_fields = [ne, pe, ee]

            def _dlg_ef(obj, event):
                if event.type() == QtCore.QEvent.KeyPress and event.key() in (
                    QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter
                ):
                    if obj in _dlg_fields:
                        nxt = _dlg_fields[(_dlg_fields.index(obj) + 1) % len(_dlg_fields)]
                        nxt.setFocus()
                        nxt.selectAll()
                        return True
                return False

            for _w in _dlg_fields:
                _w.installEventFilter(dlg)
            dlg.eventFilter = _dlg_ef

            br = QtWidgets.QHBoxLayout()
            c_btn = QtWidgets.QPushButton("Cancel")
            c_btn.setFixedHeight(38)
            c_btn.setStyleSheet("""
                QPushButton { background:#F1F5F9; color:#475569; border:1.5px solid #CBD5E1;
                              border-radius:8px; font-size:13px; font-weight:700; padding:0 20px; }
                QPushButton:hover { background:#E2E8F0; }
            """)
            c_btn.clicked.connect(dlg.reject)
            s_btn = QtWidgets.QPushButton("Update" if person else "Save")
            s_btn.setFixedHeight(38)
            s_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            s_btn.setStyleSheet("""
                QPushButton { background:#0F766E; color:white; border:none; border-radius:8px;
                              font-size:13px; font-weight:800; padding:0 20px; }
                QPushButton:hover { background:#115E59; }
            """)

            def do_save():
                name = ne.text().strip()
                if not name:
                    QtWidgets.QMessageBox.warning(dlg, "Sales People", "Name is required.")
                    return
                pd_data = {
                    "name": name,
                    "phone": pe.text().strip() or "-",
                    "email": ee.text().strip(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
                if self.FIREBASE_AVAILABLE:
                    try:
                        from main import db
                        if person and person.get("firebase_id"):
                            db.reference(f'/sales_persons/{person["firebase_id"]}').update(pd_data)
                        else:
                            ref = db.reference('/sales_persons').push()
                            pd_data["firebase_id"] = ref.key
                            pd_data["created_at"] = datetime.now(timezone.utc).isoformat()
                            ref.set(pd_data)
                    except Exception as exc:
                        QtWidgets.QMessageBox.critical(dlg, "Sales People", f"Could not save: {exc}")
                        return
                else:
                    people = _load_local_sales_people()
                    if person and person.get("local_id"):
                        pd_data["local_id"] = person["local_id"]
                        pd_data["created_at"] = person.get("created_at", pd_data["updated_at"])
                        people = [
                            pd_data if p.get("local_id") == person["local_id"] else p
                            for p in people
                        ]
                    else:
                        pd_data["local_id"] = datetime.now().strftime("%Y%m%d%H%M%S%f")
                        pd_data["created_at"] = pd_data["updated_at"]
                        people.append(pd_data)
                    if not _save_local_sales_people(people):
                        QtWidgets.QMessageBox.critical(dlg, "Sales People", "Could not save local file.")
                        return
                dlg.accept()
                load_sales()

            s_btn.clicked.connect(do_save)
            br.addStretch()
            br.addWidget(c_btn)
            br.addWidget(s_btn)
            vl.addLayout(br)
            dlg.exec_()

        # ── Delete ────────────────────────────────────────────────
        def delete_sp(person):
            if self.FIREBASE_AVAILABLE and person.get("firebase_id"):
                try:
                    from main import db
                    db.reference(f'/sales_persons/{person["firebase_id"]}').delete()
                except Exception as exc:
                    QtWidgets.QMessageBox.critical(self, "Sales People", f"Could not delete: {exc}")
                    return
            elif person.get("local_id"):
                people = [
                    i for i in _load_local_sales_people()
                    if i.get("local_id") != person.get("local_id")
                ]
                if not _save_local_sales_people(people):
                    QtWidgets.QMessageBox.critical(self, "Sales People", "Could not update local file.")
                    return
            load_sales()

        # ── Wire up ───────────────────────────────────────────────
        add_btn.clicked.connect(lambda: open_dialog())
        prev_btn.clicked.connect(lambda: render_page(self._sp_filtered, self._sp_page - 1))
        next_btn.clicked.connect(lambda: render_page(self._sp_filtered, self._sp_page + 1))

        self._reload_sales_people = load_sales
        QtCore.QTimer.singleShot(0, load_sales)
        return page

    def _build_export_tab(self):
        # Use a scroll area so the export dialog always fills from the top
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                background: #F8FAFC;
                border: none;
            }
            QScrollArea > QWidget > QWidget {
                background: #F8FAFC;
            }
        """)

        self.embedded_export = JobFormsExportDialog(self, [])
        self.embedded_export.setWindowFlags(QtCore.Qt.Widget)
        self.embedded_export.setModal(False)
        self.embedded_export.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Expanding,
        )
        self.embedded_export.setMinimumHeight(820)
        scroll.setWidget(self.embedded_export)

        def run_embedded_export():
            params = getattr(self.embedded_export, "_export_params", None)
            if not params:
                return
            if params.get("type") == "pdf":
                self.perform_job_forms_pdf_export(params)
            elif params.get("type") == "excel":
                self.perform_job_forms_excel_export(params)
            self.workflow_tabs.setCurrentIndex(0)

        self.embedded_export.accept = run_embedded_export
        self.embedded_export.reject = lambda: self.workflow_tabs.setCurrentIndex(0)
        try:
            self.embedded_export.cancel_btn.clicked.disconnect()
        except Exception:
            pass
        self.embedded_export.cancel_btn.setText("Back to Quotes")
        self.embedded_export.cancel_btn.clicked.connect(lambda: self.workflow_tabs.setCurrentIndex(0))

        return scroll

    def perform_job_forms_excel_export(self, export_params):
        """Perform Excel export for quote forms - LOWEST TO HIGHEST ORDER"""
        try:
            # Filter quote forms based on export parameters
            jobs_to_export = []
            
            for job in self.job_forms:
                try:
                    # Parse job creation date (created_at) instead of start_date
                    job_datetime = None
                    created_at_str = job.get('created_at', '')
                    
                    if created_at_str:
                        # Extract date part from ISO string
                        date_part = created_at_str.split('T')[0] if 'T' in created_at_str else created_at_str
                        
                        # Try parsing the created_at date
                        try:
                            job_datetime = datetime.strptime(date_part, "%Y-%m-%d")
                        except ValueError:
                            # Try alternative formats if needed
                            date_formats = ["%m-%d-%Y", "%Y/%m/%d", "%d/%m/%Y"]
                            for date_format in date_formats:
                                try:
                                    job_datetime = datetime.strptime(date_part, date_format)
                                    break
                                except ValueError:
                                    continue
                    
                    if job_datetime is None:
                        _log.warning("Warning: Could not parse creation date for job - EXCLUDING from export")
                        continue
                    
                    include_job = False
                    
                    if export_params["range"] == "all":
                        include_job = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        job_date_only = job_datetime.date()
                        
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        if from_date_only <= job_date_only <= to_date_only:
                            include_job = True
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if job_datetime.month == month and job_datetime.year == year:
                            include_job = True
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if job_datetime.year == year:
                            include_job = True

                    if include_job:
                        status_filter = export_params.get("status", "All Status")
                        client_filter = export_params.get("client", "All Clients")
                        if status_filter != "All Status" and job.get("status", "Not Started") != status_filter:
                            include_job = False
                        if client_filter != "All Clients" and job.get("client", "") != client_filter:
                            include_job = False
                    
                    if include_job:
                        jobs_to_export.append(job)
                            
                except Exception as e:
                    _log.warning("Error processing job: %s", e)
                    continue
            
            if not jobs_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No quote forms found matching the selected criteria.")
                return
            
            _log.info("Excel Export: Found %s quote forms to export", len(jobs_to_export))
            
            # ====== CRITICAL CHANGE: SORT FROM LOWEST TO HIGHEST ======
            def job_number_sort_key_low_to_high(job):
                """
                Sort Quote quote numbers in ASCENDING order (LOWEST to HIGHEST)
                Same logic as PDF export
                """
                job_num = job.get('job_number', '').strip().upper()
                
                if not job_num or 'QUOTE' not in job_num:
                    return (999, 999, 999, 999, 999, 999, 999, 999)

                pattern = r'^QUOTE([A-Z]?)(\d+)(?:_?([a-zA-Z]+)?(\d+)?)?(?:_?(\d+)([a-zA-Z]+)?(\d+)?)?$'
                match = re.match(pattern, job_num, re.IGNORECASE)
                
                if not match:
                    return (999, 999, 999, 999, 999, 999, 999, 999)
                
                category = match.group(1) or 'A'
                main_seq_str = match.group(2)
                alpha1 = (match.group(3) or '').lower()
                num1 = match.group(4) or ''
                num2 = match.group(5) or ''
                alpha2 = (match.group(6) or '').lower()
                num3 = match.group(7) or ''
                
                # 1. Category value (A=1, Z=26) - positive for ascending
                cat_value = ord(category) - 64 if category else 1
                
                # 2. Main sequence value
                try:
                    main_seq = int(main_seq_str.lstrip('0')) if main_seq_str.lstrip('0') else 0
                except:
                    main_seq = 0
                
                # 3. Variant type priority (lower number = appears earlier)
                variant_priority = 0
                
                if num2 and alpha2 and num3:
                    # Numeric + Alphabetic + Numeric - LOWEST priority
                    variant_priority = 6
                elif num2 and alpha2 and not num3:
                    # Numeric + Alphabetic
                    variant_priority = 5
                elif num2 and not alpha2:
                    # Pure Numeric suffix
                    variant_priority = 4
                elif alpha1 and num1 and not (num2 or alpha2):
                    # Alphabetic + Numeric
                    variant_priority = 3
                elif alpha1 and not num1:
                    # Pure Alphabetic
                    variant_priority = 2
                elif not (alpha1 or num1 or num2 or alpha2 or num3):
                    # Base (no suffix) - HIGHEST priority
                    variant_priority = 1
                
                # 4. Convert suffixes to sortable values
                alpha1_value = 0
                if alpha1:
                    for i, char in enumerate(reversed(alpha1)):
                        char_val = ord(char) - 96
                        alpha1_value += char_val * (26 ** i)
                
                num1_value = 0
                try:
                    num1_value = int(num1) if num1 else 0
                except:
                    num1_value = 0
                
                num2_value = 0
                try:
                    num2_value = int(num2) if num2 else 0
                except:
                    num2_value = 0
                
                alpha2_value = 0
                if alpha2:
                    for i, char in enumerate(reversed(alpha2)):
                        char_val = ord(char) - 96
                        alpha2_value += char_val * (26 ** i)
                
                num3_value = 0
                try:
                    num3_value = int(num3) if num3 else 0
                except:
                    num3_value = 0
                
                # Return tuple for ASCENDING order
                return (
                    cat_value,        # A-Z
                    main_seq,         # 001-999
                    variant_priority, # 1-6
                    alpha1_value,     # a-z
                    num1_value,       # 0-999
                    num2_value,       # 0-999
                    alpha2_value,     # a-z
                    num3_value        # 0-999
                )
            
            def _received_date_sort_key(job):
                for field in ('start_date', 'created_at'):
                    raw = job.get(field, '') or ''
                    date_part = raw.split('T')[0] if 'T' in raw else raw
                    for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
                        try:
                            return datetime.strptime(date_part, fmt)
                        except ValueError:
                            continue
                return datetime.min

            jobs_to_export.sort(key=_received_date_sort_key)

            _log.debug("Excel export order (first 10): %s",
                       [j.get('job_number') for j in jobs_to_export[:10]])

            self.generate_job_forms_combined_excel(jobs_to_export, export_params)
                
        except Exception as e:
            _log.warning("Error performing Excel export: %s", e)
            _log.exception("Traceback:")
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during Excel export: {str(e)}")
    
    def open_sales_person_dialog(self):
        """Open the Sales Person Management dialog with Firebase storage"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QTabWidget, QWidget, QGroupBox, QFormLayout, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView
        from PyQt5.QtCore import Qt, QEvent
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Sales Person Management")
        dialog.setModal(True)
        dialog.resize(860, 660)
        
        # Set up Ctrl+S shortcut
        save_shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+S"), dialog)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(10)
        
        # Tab widget
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #d0d7de;
                border-radius: 0px 8px 8px 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #f6f8fa;
                color: #57606a;
                padding: 10px 28px;
                min-width: 170px;
                margin-right: 2px;
                border: 1px solid #d0d7de;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 600;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QTabBar::tab:selected {
                background-color: white;
                color: #0969da;
                border-bottom: 2px solid #0969da;
                font-weight: bold;
            }
            QTabBar::tab:hover:!selected {
                background-color: #eaeef2;
                color: #24292f;
            }
        """)
        
        # ========== LOAD SALES PERSONS FROM FIREBASE (NOT QSettings) ==========
        sales_persons = []  # List to store sales persons from Firebase
        
        def load_sales_persons_from_firebase():
            """Load sales persons from Firebase - like clients loading"""
            nonlocal sales_persons
            
            if not self.FIREBASE_AVAILABLE:
                _log.warning("Firebase not available - cannot load sales persons")
                QMessageBox.warning(dialog, "Firebase Error", "Firebase is not available. Cannot load sales persons.")
                return
            
            try:
                from main import db
                ref = db.reference('/sales_persons')
                sales_data = ref.get()
                
                sales_persons = []
                
                if sales_data:
                    for person_id, person_data in sales_data.items():
                        if person_data:
                            person_data['firebase_id'] = person_id
                            sales_persons.append(person_data)
                
                # Sort by name
                sales_persons.sort(key=lambda x: x.get('name', ''))
                
                _log.info("Loaded %s sales persons from Firebase", len(sales_persons))
                filter_table()
                
            except Exception as e:
                _log.warning("Error loading sales persons from Firebase: %s", e)
                QMessageBox.critical(dialog, "Error", f"Error loading sales persons: {str(e)}")
        
        def save_sales_person_to_firebase(person_data):
            """Save a sales person to Firebase - like saving a client"""
            if not self.FIREBASE_AVAILABLE:
                _log.warning("Firebase not available - cannot save sales person")
                return False
            
            try:
                from main import db
                ref = db.reference('/sales_persons')
                
                # Check if we're updating an existing person
                if 'firebase_id' in person_data and person_data['firebase_id']:
                    # Update existing
                    person_id = person_data['firebase_id']
                    person_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    ref.child(person_id).update(person_data)
                    _log.info("Sales person updated in Firebase: %s", person_data.get('name'))
                else:
                    # Create new
                    new_ref = ref.push()
                    person_data['firebase_id'] = new_ref.key
                    person_data['created_at'] = datetime.now(timezone.utc).isoformat()
                    person_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    new_ref.set(person_data)
                    _log.info("Sales person created in Firebase: %s", person_data.get('name'))
                
                return True
                
            except Exception as e:
                _log.warning("Error saving sales person to Firebase: %s", e)
                return False
        
        def delete_sales_person_from_firebase(person_id):
            """Delete a sales person from Firebase"""
            if not self.FIREBASE_AVAILABLE:
                _log.warning("Firebase not available - cannot delete sales person")
                return False
            
            try:
                from main import db
                ref = db.reference(f'/sales_persons/{person_id}')
                ref.delete()
                _log.info("Sales person deleted from Firebase (ID: %s)", person_id)
                return True
                
            except Exception as e:
                _log.warning("Error deleting sales person from Firebase: %s", e)
                return False
        
        # Tab 1: Add/Edit
        add_tab = QWidget()
        add_layout = QVBoxLayout(add_tab)
        add_layout.setSpacing(15)
        add_layout.setContentsMargins(20, 20, 20, 20)
        
        form_group = QGroupBox("Sales Person Details")
        form_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        form_layout = QFormLayout(form_group)
        form_layout.setSpacing(12)
        form_layout.setContentsMargins(20, 20, 20, 20)
        
        # Name - Required
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("Full Name (Required)")
        name_edit.setStyleSheet("padding: 10px; border: 1px solid #bdc3c7; border-radius: 6px; font-size: 13px;")
        form_layout.addRow("Name:", name_edit)
        
        # Phone - Optional
        phone_edit = QLineEdit()
        phone_edit.setPlaceholderText("(xxx) xxx-xxxx")
        phone_edit.setStyleSheet("padding: 10px; border: 1px solid #bdc3c7; border-radius: 6px; font-size: 13px;")
        form_layout.addRow("Phone:", phone_edit)
        
        # Email - Optional
        email_edit = QLineEdit()
        email_edit.setPlaceholderText("email@example.com")
        email_edit.setStyleSheet("padding: 10px; border: 1px solid #bdc3c7; border-radius: 6px; font-size: 13px;")
        form_layout.addRow("Email:", email_edit)
        
        add_layout.addWidget(form_group)

        input_widgets = [name_edit, phone_edit, email_edit]

        # Button layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)
        
        # Save Details button
        save_btn = QPushButton("Save")
        save_btn.setFixedSize(100, 36)
        save_btn.setAutoDefault(False)
        save_btn.setDefault(False)
        save_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #1a7f37;
                color: white;
                border: none;
                border-radius: 7px;
                font-weight: 600;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background-color: #166b2e; }
            QPushButton:pressed { background-color: #125826; }
        """)
        
        # Update Details button (Visible only when editing)
        update_btn = QPushButton("Update")
        update_btn.setFixedSize(100, 36)
        update_btn.setAutoDefault(False)
        update_btn.setDefault(False)
        update_btn.setVisible(False)
        update_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        update_btn.setStyleSheet("""
            QPushButton {
                background-color: #9a3800;
                color: white;
                border: none;
                border-radius: 7px;
                font-weight: 600;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background-color: #7d2d00; }
            QPushButton:pressed { background-color: #5f2200; }
        """)
        
        clear_btn = QPushButton("Clear")
        clear_btn.setFixedSize(90, 36)
        clear_btn.setAutoDefault(False)
        clear_btn.setDefault(False)
        clear_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #f6f8fa;
                color: #24292f;
                border: 1px solid #d0d7de;
                border-radius: 7px;
                font-weight: 600;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
            }
            QPushButton:hover { background-color: #eaeef2; }
            QPushButton:pressed { background-color: #d0d7de; }
        """)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(update_btn)
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        add_layout.addLayout(button_layout)
        add_layout.addStretch()
        
        # Tab 2: View
        view_tab = QWidget()
        view_layout = QVBoxLayout(view_tab)
        view_layout.setSpacing(10)
        view_layout.setContentsMargins(10, 10, 10, 10)
        
        # Search bar
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("🔍 Search by name, phone, or email...")
        search_edit.setStyleSheet("""
            QLineEdit {
                padding: 10px 15px;
                border: 2px solid #e1e8ed;
                border-radius: 10px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
        """)
        search_edit.setMinimumHeight(38)

        sales_count_lbl = QtWidgets.QLabel("0 Sales")
        sales_count_lbl.setStyleSheet("""
            QLabel {
                background: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 7px;
                color: #475569;
                font-size: 12px;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 14px;
            }
        """)
        sales_count_lbl.setFixedHeight(38)
        sales_count_lbl.setMinimumWidth(80)

        search_layout = QHBoxLayout()
        search_layout.addWidget(search_edit, 1)
        search_layout.addWidget(sales_count_lbl)
        view_layout.addLayout(search_layout)
        
        # Table with action buttons
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Name", "Phone", "Email", "Actions"])
        table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 2px solid #e1e8ed;
                border-radius: 10px;
                gridline-color: #e1e8ed;
            }
            QTableWidget::item {
                padding: 6px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background: #e3f2fd;
                color: black;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2c3e50, stop:1 #34495e);
                color: white;
                font-weight: bold;
                padding: 8px;
                border: none;
            }
        """)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setDefaultSectionSize(44)
        table.verticalHeader().setVisible(False)

        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)       # Name stretches
        header.setSectionResizeMode(1, QHeaderView.Stretch)       # Phone stretches
        header.setSectionResizeMode(2, QHeaderView.Stretch)       # Email stretches
        header.setSectionResizeMode(3, QHeaderView.Fixed)         # Actions fixed
        table.setColumnWidth(3, 140)                              # single dropdown button
        view_layout.addWidget(table)
        
        refresh_btn = QPushButton("🔄 Refresh List")
        refresh_btn.setFixedHeight(40)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #5dade2, stop:1 #3498db);
            }
        """)
        view_layout.addWidget(refresh_btn)
        
        # Prevent Enter key on buttons from saving
        def prevent_button_enter(obj, event):
            if event.type() == QEvent.KeyPress and event.key() in (Qt.Key_Return, Qt.Key_Enter):
                return True
            return False
        
        for btn in [save_btn, update_btn, clear_btn, refresh_btn]:
            btn.installEventFilter(dialog)
            btn.eventFilter = prevent_button_enter
        
        def display_table(persons):
            """Display sales persons in table"""
            persons = list(reversed(persons))
            table.setRowCount(len(persons))
            n = len(persons)
            sales_count_lbl.setText(f"{n} Sale{'s' if n != 1 else ''}")
            
            for row, person in enumerate(persons):
                name_item = QTableWidgetItem(person.get('name', '-') if person.get('name') else "-")
                name_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 0, name_item)
                
                phone_item = QTableWidgetItem(person.get('phone', '-') if person.get('phone') else "-")
                phone_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 1, phone_item)
                
                email_item = QTableWidgetItem(person.get('email', '-') if person.get('email') else "-")
                email_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, 2, email_item)
                
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(10, 6, 10, 6)
                actions_layout.setAlignment(Qt.AlignVCenter | Qt.AlignCenter)

                # Single dropdown action button — professional, no clipping
                from PyQt5.QtWidgets import QMenu, QAction
                action_btn = QPushButton("Actions  ▾")
                action_btn.setFixedSize(110, 32)
                action_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                action_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #f6f8fa;
                        color: #24292f;
                        border: 1px solid #d0d7de;
                        border-radius: 6px;
                        font-size: 12px;
                        font-weight: 600;
                        font-family: 'Inter', 'Segoe UI', sans-serif;
                        text-align: center;
                        padding: 0 8px;
                    }
                    QPushButton:hover {
                        background-color: #eaeef2;
                        border-color: #0969da;
                        color: #0969da;
                    }
                    QPushButton:pressed { background-color: #d0d7de; }
                """)

                def show_action_menu(checked=False, p=person, btn=action_btn):
                    menu = QMenu(btn)
                    menu.setStyleSheet("""
                        QMenu {
                            background-color: white;
                            border: 1px solid #d0d7de;
                            border-radius: 8px;
                            padding: 4px 0;
                            font-family: 'Inter', 'Segoe UI', sans-serif;
                            font-size: 13px;
                        }
                        QMenu::item {
                            padding: 8px 20px;
                            color: #24292f;
                        }
                        QMenu::item:selected {
                            background-color: #f6f8fa;
                            color: #0969da;
                        }
                        QMenu::separator {
                            height: 1px;
                            background: #d0d7de;
                            margin: 4px 0;
                        }
                    """)
                    edit_action = QAction("✏  Edit", menu)
                    delete_action = QAction("🗑  Delete", menu)
                    edit_action.triggered.connect(lambda: edit_sales_person(p))
                    delete_action.triggered.connect(lambda: delete_sales_person(p))
                    menu.addAction(edit_action)
                    menu.addSeparator()
                    menu.addAction(delete_action)
                    menu.exec_(btn.mapToGlobal(QtCore.QPoint(0, btn.height())))

                action_btn.clicked.connect(show_action_menu)
                action_btn.installEventFilter(dialog)
                action_btn.eventFilter = prevent_button_enter

                actions_layout.addWidget(action_btn)
                table.setCellWidget(row, 3, actions_widget)
        
        def filter_table():
            search_text = search_edit.text().lower()
            if not search_text:
                display_table(sales_persons)
            else:
                filtered = [p for p in sales_persons 
                        if search_text in p.get('name', '').lower() 
                        or search_text in p.get('phone', '').lower() 
                        or search_text in p.get('email', '').lower()]
                display_table(filtered)
        
        def add_sales_person():
            name = name_edit.text().strip()
            phone = phone_edit.text().strip()
            email = email_edit.text().strip()
            
            if not name:
                QMessageBox.warning(dialog, "Validation Error", "Name is required.")
                name_edit.setFocus()
                return
            
            # Check for duplicate name
            for person in sales_persons:
                if person.get('name', '').lower() == name.lower():
                    QMessageBox.warning(dialog, "Duplicate", f"Sales person '{name}' already exists!")
                    return
            
            new_person = {
                'name': name,
                'phone': phone if phone else "",
                'email': email if email else ""
            }
            
            if save_sales_person_to_firebase(new_person):
                load_sales_persons_from_firebase()
                clear_form()
                QMessageBox.information(dialog, "Success", f"Sales person '{name}' saved successfully!")
        
        editing_id = None
        
        def edit_sales_person(person):
            nonlocal editing_id
            editing_id = person.get('firebase_id')
            name_edit.setText(person.get('name', ''))
            phone_edit.setText(person.get('phone', ''))
            email_edit.setText(person.get('email', ''))
            
            save_btn.setVisible(False)
            update_btn.setVisible(True)
            
            tab_widget.setCurrentIndex(0)
            name_edit.setFocus()
            name_edit.selectAll()
        
        def update_sales_person():
            nonlocal editing_id
            if editing_id is None:
                return
            
            name = name_edit.text().strip()
            phone = phone_edit.text().strip()
            email = email_edit.text().strip()
            
            if not name:
                QMessageBox.warning(dialog, "Validation Error", "Name is required.")
                name_edit.setFocus()
                return
            
            # Find the person and update
            for i, person in enumerate(sales_persons):
                if person.get('firebase_id') == editing_id:
                    updated_person = {
                        'firebase_id': editing_id,
                        'name': name,
                        'phone': phone if phone else "",
                        'email': email if email else ""
                    }
                    if save_sales_person_to_firebase(updated_person):
                        load_sales_persons_from_firebase()
                        clear_form()
                        QMessageBox.information(dialog, "Success", "Sales person updated successfully!")
                    break
        
        def delete_sales_person(person):
            reply = QMessageBox.question(
                dialog, 
                "Confirm Delete", 
                f"Delete '{person.get('name', 'this entry')}'?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if 'firebase_id' in person:
                    if delete_sales_person_from_firebase(person['firebase_id']):
                        load_sales_persons_from_firebase()
                        QMessageBox.information(dialog, "Success", "Sales person deleted successfully!")
        
        def clear_form():
            name_edit.clear()
            phone_edit.clear()
            email_edit.clear()
            nonlocal editing_id
            editing_id = None
            
            save_btn.setVisible(True)
            update_btn.setVisible(False)
            
            name_edit.setFocus()
        
        # Connect signals
        save_btn.clicked.connect(add_sales_person)
        update_btn.clicked.connect(update_sales_person)
        clear_btn.clicked.connect(clear_form)
        search_edit.textChanged.connect(filter_table)
        refresh_btn.clicked.connect(load_sales_persons_from_firebase)
        
        # Connect Ctrl+S shortcut
        def handle_save_shortcut():
            if save_btn.isVisible():
                add_sales_person()
            elif update_btn.isVisible():
                update_sales_person()
        save_shortcut.activated.connect(handle_save_shortcut)
        
        tab_widget.addTab(add_tab, "Add Sales Person")
        tab_widget.addTab(view_tab, "Sales Information")
        layout.addWidget(tab_widget)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setFixedHeight(40)
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        close_btn.installEventFilter(dialog)
        close_btn.eventFilter = prevent_button_enter
        layout.addWidget(close_btn)
        
        # Initial display
        load_sales_persons_from_firebase()
        
        def eventFilter(obj, event):
            if event.type() == QEvent.KeyPress and event.key() in (Qt.Key_Return, Qt.Key_Enter):
                if obj in input_widgets:
                    index = input_widgets.index(obj)
                    next_index = (index + 1) % len(input_widgets)
                    input_widgets[next_index].setFocus()
                    input_widgets[next_index].selectAll()
                    return True
            return False

        for w in input_widgets:
            w.installEventFilter(dialog)

        dialog.eventFilter = eventFilter
        
        dialog.exec_()
            
    def generate_job_forms_combined_excel(self, jobs, export_params):
        """Generate a professional combined Excel report for quote forms in"""
        try:
            # Create export directory if it doesn't exist
            temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_exports_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Job_Forms_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Job_Forms_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Job_Forms_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Job_Forms_{year}.xlsx"

            excel_path = temp_dir / filename

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote Forms"

            # Header information - CLEAN VERSION WITHOUT STATISTICS
            ws.merge_cells('A1:I1')  # Changed to I1 for 9 columns
            try:
                from main import Config as _Cfg
                _co_name = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co_name = 'MABS ENGINEERING LLC'
            ws['A1'] = f"{_co_name} - QUOTE FORMS REPORT"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Export range info
            if export_params["range"] == "all":
                export_range_text = "All Quote Forms"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m-%d-%Y")  # Changed to MM-dd-YYYY
                to_date = export_params["to_date"].strftime("%m-%d-%Y")      # Changed to MM-dd-YYYY
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws['A2'] = f"Period: {export_range_text}"

            # Table headers - start at row 4 (one row closer after removing Generated line)
            # Added Sales column (9 columns total)
            headers = ["S.No.", "Quote Number", "Project Name", "Client", "Sales", "Start Date", "Due Date", "Cost", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = _xlsx_fill("D9D9D9")
                cell.alignment = Alignment(horizontal='center')

            # Job data
            for row_idx, job in enumerate(jobs, 5):  # Start at row 5 (after headers)
                # Parse creation date
                # Parse start date
                start_date = job.get('start_date', '')
                try:
                    # Try MM-dd-yyyy format first
                    start_date_dt = datetime.strptime(start_date, "%m-%d-%Y")
                    start_date_formatted = start_date_dt.strftime("%m-%d-%Y")
                except:
                    try:
                        # Try alternative formats
                        start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
                        start_date_formatted = start_date_dt.strftime("%m-%d-%Y")
                    except:
                        start_date_formatted = start_date
                # Parse due date
                due_date = job.get('due_date', '')
                try:
                    due_date_dt = datetime.strptime(due_date, "%m-%d-%Y")
                    due_date_formatted = due_date_dt.strftime("%m-%d-%Y")
                except:
                    try:
                        due_date_dt = datetime.strptime(due_date, "%Y-%m-%d")
                        due_date_formatted = due_date_dt.strftime("%m-%d-%Y")
                    except:
                        due_date_formatted = due_date
                
                data = [
                    row_idx - 4,  # Sequential number
                    job.get('job_number', ''),
                    job.get('project_name', ''),  # Changed from job_title to project_name
                    job.get('client', ''),
                    job.get('sales', ''),  # Added Sales column
                    start_date_formatted,  # Changed: Use Start Date instead of Created Date
                    due_date_formatted,    # Keep due date
                    job.get('engineering_costs', 'N/A'),
                    job.get('status', 'Not Started')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Style for sequential number column - SMALLER WIDTH
                    if col == 1:
                        cell.font = Font(bold=True)
                        cell.fill = _xlsx_fill("F0F8FF")  # Light blue
                    
                    # Style for status column
                    if col == 9:  # Status column (now column 9)
                        status = str(value).lower()
                        if 'completed' in status:
                            cell.fill = _xlsx_fill("E8F5E9")  # Light green
                        elif 'urgent' in status or 'high' in status:
                            cell.fill = _xlsx_fill("FFEBEE")  # Light red
                        elif 'cancel' in status:
                            cell.fill = _xlsx_fill("F5F5F5")  # Light gray

            # Auto-adjust column widths - with specific width for S.No. column
            column_widths = {
                1: 8,   # S.No. - smaller width (4 characters + 2 padding)
                2: 25,  # Quote Number
                3: 40,  # Project Name (was Job Title)
                4: 40,  # Client
                5: 25,  # Sales
                6: 15,  # Start Date (was Created Date)
                7: 15,  # Due Date
                8: 15,  # Cost
                9: 15   # Status
            }
            
            for col_idx in range(1, len(headers) + 1):
                if col_idx in column_widths:
                    # Use specified width
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = column_widths[col_idx]
                else:
                    # Auto-adjust for any additional columns
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[column_letter]:
                        if cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Add alternating row colors for better readability
            for row in range(6, ws.max_row + 1):
                if row % 2 == 0:  # Even rows
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        # Don't override existing fills
                        if cell.fill.start_color.index == '00000000':  # Default fill
                            cell.fill = _xlsx_fill("F9F9F9")

            # Save the workbook
            wb.save(str(excel_path))

            # Open the Excel file
            if self.open_job_form_pdf_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success",
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"The Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success",
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"Could not open automatically. Please open manually.")

            # Save to export history and refresh dialog
            try:
                exp_dlg = getattr(self, 'embedded_export', None)
                if exp_dlg:
                    scope_map = {"all": "All Quote Forms", "date_range": "Date Range",
                                 "month": "By Month", "year": "By Year"}
                    scope = scope_map.get(export_params.get("range", "all"), "All Quote Forms")
                    if export_params.get("range") == "date_range":
                        fd = export_params.get("from_date")
                        td = export_params.get("to_date")
                        if fd and td:
                            scope = f"Date Range ({fd.strftime('%b %d, %Y')} - {td.strftime('%b %d, %Y')})"
                    exp_dlg._save_export_entry("Excel", scope, len(jobs_to_export), str(excel_path))
                    exp_dlg.refresh_recent_exports()
            except Exception:
                pass

        except Exception as e:
            _log.warning("Error generating combined Excel: %s", e)
            _log.exception("Traceback:")
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error",
                                        f"Error generating Excel: {str(e)}")
            
    def perform_job_forms_pdf_export(self, export_params):
        """Perform PDF export for quote forms - LOWEST TO HIGHEST ORDER - SAVE LOCALLY"""
        try:
            # Filter quote forms based on export parameters
            jobs_to_export = []
            
            for job in self.job_forms:
                try:
                    # Parse job creation date (created_at) instead of start_date
                    job_datetime = None
                    created_at_str = job.get('created_at', '')
                    
                    if created_at_str:
                        # Extract date part from ISO string
                        date_part = created_at_str.split('T')[0] if 'T' in created_at_str else created_at_str
                        
                        # Try parsing the created_at date
                        try:
                            job_datetime = datetime.strptime(date_part, "%Y-%m-%d")
                        except ValueError:
                            # Try alternative formats if needed
                            date_formats = ["%m-%d-%Y", "%Y/%m/%d", "%d/%m/%Y"]
                            for date_format in date_formats:
                                try:
                                    job_datetime = datetime.strptime(date_part, date_format)
                                    break
                                except ValueError:
                                    continue
                    
                    if job_datetime is None:
                        _log.warning("Warning: Could not parse creation date for job - EXCLUDING from export")
                        continue
                    
                    include_job = False
                    
                    if export_params["range"] == "all":
                        include_job = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        job_date_only = job_datetime.date()
                        
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        if from_date_only <= job_date_only <= to_date_only:
                            include_job = True
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if job_datetime.month == month and job_datetime.year == year:
                            include_job = True
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if job_datetime.year == year:
                            include_job = True

                    if include_job:
                        status_filter = export_params.get("status", "All Status")
                        client_filter = export_params.get("client", "All Clients")
                        if status_filter != "All Status" and job.get("status", "Not Started") != status_filter:
                            include_job = False
                        if client_filter != "All Clients" and job.get("client", "") != client_filter:
                            include_job = False
                    
                    if include_job:
                        jobs_to_export.append(job)
                            
                except Exception as e:
                    _log.warning("Error processing job: %s", e)
                    continue
            
            if not jobs_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No quote forms found matching the selected criteria.")
                return
            
            _log.info("PDF Export: Found %s quote forms to export", len(jobs_to_export))
            
            # ====== CRITICAL CHANGE: SORT FROM LOWEST TO HIGHEST ======
            def job_number_sort_key_low_to_high(job):
                """
                Sort Quote quote numbers in ASCENDING order (LOWEST to HIGHEST)
                Same logic as PDF export
                """
                job_num = job.get('job_number', '').strip().upper()
                
                if not job_num or 'QUOTE' not in job_num:
                    return (999, 999, 999, 999, 999, 999, 999, 999)

                pattern = r'^QUOTE([A-Z]?)(\d+)(?:_?([a-zA-Z]+)?(\d+)?)?(?:_?(\d+)([a-zA-Z]+)?(\d+)?)?$'
                match = re.match(pattern, job_num, re.IGNORECASE)
                
                if not match:
                    return (999, 999, 999, 999, 999, 999, 999, 999)
                
                category = match.group(1) or 'A'
                main_seq_str = match.group(2)
                alpha1 = (match.group(3) or '').lower()
                num1 = match.group(4) or ''
                num2 = match.group(5) or ''
                alpha2 = (match.group(6) or '').lower()
                num3 = match.group(7) or ''
                
                # 1. Category value (A=1, Z=26) - positive for ascending
                cat_value = ord(category) - 64 if category else 1
                
                # 2. Main sequence value
                try:
                    main_seq = int(main_seq_str.lstrip('0')) if main_seq_str.lstrip('0') else 0
                except:
                    main_seq = 0
                
                # 3. Variant type priority (lower number = appears earlier)
                variant_priority = 0
                
                if num2 and alpha2 and num3:
                    variant_priority = 6
                elif num2 and alpha2 and not num3:
                    variant_priority = 5
                elif num2 and not alpha2:
                    variant_priority = 4
                elif alpha1 and num1 and not (num2 or alpha2):
                    variant_priority = 3
                elif alpha1 and not num1:
                    variant_priority = 2
                elif not (alpha1 or num1 or num2 or alpha2 or num3):
                    variant_priority = 1
                
                # 4. Convert suffixes to sortable values
                alpha1_value = 0
                if alpha1:
                    for i, char in enumerate(reversed(alpha1)):
                        char_val = ord(char) - 96
                        alpha1_value += char_val * (26 ** i)
                
                num1_value = 0
                try:
                    num1_value = int(num1) if num1 else 0
                except:
                    num1_value = 0
                
                num2_value = 0
                try:
                    num2_value = int(num2) if num2 else 0
                except:
                    num2_value = 0
                
                alpha2_value = 0
                if alpha2:
                    for i, char in enumerate(reversed(alpha2)):
                        char_val = ord(char) - 96
                        alpha2_value += char_val * (26 ** i)
                
                num3_value = 0
                try:
                    num3_value = int(num3) if num3 else 0
                except:
                    num3_value = 0
                
                return (
                    cat_value,
                    main_seq,
                    variant_priority,
                    alpha1_value,
                    num1_value,
                    num2_value,
                    alpha2_value,
                    num3_value
                )
            
            def _received_date_sort_key(job):
                for field in ('start_date', 'created_at'):
                    raw = job.get(field, '') or ''
                    date_part = raw.split('T')[0] if 'T' in raw else raw
                    for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
                        try:
                            return datetime.strptime(date_part, fmt)
                        except ValueError:
                            continue
                return datetime.min

            jobs_to_export.sort(key=_received_date_sort_key)

            _log.debug("PDF export order (first 10): %s",
                       [j.get('job_number') for j in jobs_to_export[:10]])

            self.generate_job_forms_combined_pdf(jobs_to_export, export_params)
                    
        except Exception as e:
            _log.warning("Error performing PDF export: %s", e)
            _log.exception("Traceback:")
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during PDF export: {str(e)}")

    def generate_job_forms_combined_pdf(self, jobs, export_params):
        """Generate a professional combined PDF report for quote forms - SAVE TO DOWNLOADS FOLDER"""
        try:
            # Create export directory in Downloads folder
            from pathlib import Path
            import os
            
            export_dir = Path.home() / "Downloads" / "Job_Forms_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Job_Forms_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Job_Forms_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Job_Forms_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Job_Forms_{year}.pdf"

            pdf_path = export_dir / filename

            # Create PDF document
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            doc = SimpleDocTemplate(str(pdf_path), pagesize=landscape(A4),
                                topMargin=0.2*inch, bottomMargin=0.2*inch,
                                leftMargin=0.3*inch, rightMargin=0.3*inch)
            elements = []

            # Get styles
            styles = getSampleStyleSheet()
            
            # MABS Engineering LLC Header Style - Large and centered
            mabs_header_style = ParagraphStyle(
                'MABSHeader',
                parent=styles['Normal'],
                fontSize=20,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold',
                spaceAfter=3,
                spaceBefore=2
            )
            
            # Date Style - CHANGED TO MM-dd-yyyy FORMAT
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#7f8c8d'),
                fontName='Helvetica',
                alignment=2,
                spaceAfter=3
            )
            
            # Quote form Report Title Style
            report_title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=14,
                spaceAfter=5,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold'
            )

            # Statistics Style
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=8,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold'
            )

            # 1. Company name — full-width centered so short or long names stay balanced
            try:
                from main import Config as _Cfg
                _co_display = _Cfg.COMPANY.get('name', 'MABS Engineering LLC')
            except Exception:
                _co_display = 'MABS Engineering LLC'
            avail_w = 9.5*inch + 1.59*inch  # total usable width
            header_table = Table(
                [[Paragraph(_co_display, mabs_header_style)]],
                colWidths=[avail_w]
            )
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
            ]))

            elements.append(header_table)
            
            # Add gap between Row 1 and Row 2
            elements.append(Spacer(1, 0.22*inch))
            
            # 2. Quote form Report Title (centered below MABS)
            report_title = Paragraph("Quote Forms Report", report_title_style)
            elements.append(report_title)
            
            # Calculate statistics
            total_jobs = len(jobs)
            completed_jobs = len([job for job in jobs if job.get('status') == 'Completed'])
            high_priority_jobs = len([job for job in jobs if job.get('status') in ['High', 'Urgent']])
            
            # 3. Statistics
            elements.append(Spacer(1, 0.1*inch))
            
            stats_text = f"Total Jobs: {total_jobs}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Completed: {completed_jobs}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;High Priority: {high_priority_jobs}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            # 4. Export range info - CHANGED TO MM-dd-yyyy FORMAT
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#2c3e50'),
                fontName='Helvetica-Bold'
            )
            
            # Generate export range text with MM-dd-yyyy format
            if export_params["range"] == "all":
                export_range_text = "All Quote Forms"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%Y")
                to_date = export_params["to_date"].strftime("%m/%d/%Y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            info_data = [[Paragraph(f"{export_range_text}", info_style)]]

            info_table = Table(info_data, colWidths=[11.09 * inch])
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))

            elements.append(info_table)
            elements.append(Spacer(1, 0.2*inch))

            # 5. Quote Forms Table with sequential numbers in ASCENDING order (LOWEST to HIGHEST)
            if jobs:
                table_data = [["S.No.", "Quote Number", "Project Name", "Client", "Sales", "Cost", "Status"]]

                cell_s = ParagraphStyle('CellS', parent=styles['Normal'],
                    fontName='Helvetica', fontSize=7, leading=9, alignment=1,
                    textColor=colors.HexColor('#2c3e50'))

                # Add sequential numbers for FIFO order clarity (already sorted LOWEST to HIGHEST)
                for idx, job in enumerate(jobs, 1):
                    table_data.append([
                        Paragraph(str(idx), cell_s),
                        Paragraph(str(job.get('job_number', '') or ''), cell_s),
                        Paragraph(str(job.get('project_name', '') or ''), cell_s),
                        Paragraph(str(job.get('client', '') or ''), cell_s),
                        Paragraph(str(job.get('sales', '') or ''), cell_s),
                        Paragraph(str(job.get('engineering_costs', '') or ''), cell_s),
                        Paragraph(str(job.get('status', '') or ''), cell_s),
                    ])
                
                # Column widths for landscape A4 (available ~11.09 inches)
                col_widths = [
                    0.4*inch,   # S.No.
                    1.2*inch,   # Quote Number
                    2.5*inch,   # Project Name
                    2.0*inch,   # Client
                    1.8*inch,   # Sales
                    1.0*inch,   # Cost
                    1.5*inch,   # Status
                ]
                
                job_table = Table(table_data, colWidths=col_widths)
                job_table.setStyle(TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    
                    # Data row styling
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffffff')),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8f9fa'), colors.white]),
                    
                    # Special styling for S.No. column
                    ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#e8f4f8')),
                    ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 1), (0, -1), 8),
                    ('TEXTCOLOR', (0, 1), (0, -1), colors.HexColor('#2c3e50')),
                    
                    # Row padding
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                elements.append(job_table)
            else:
                no_data_style = ParagraphStyle(
                    'NoData',
                    parent=styles['Normal'],
                    fontSize=12,
                    textColor=colors.HexColor('#7f8c8d'),
                    alignment=1
                )
                elements.append(Paragraph("No quote forms found for the selected criteria.", no_data_style))

            # Build PDF
            doc.build(elements)

            # Ensure PDF has full permissions for Adobe Reader compatibility
            try:
                from main import PDFPermissions
                PDFPermissions.ensure_full_permissions(Path(pdf_path))
            except Exception as e:
                _log.warning("Error ensuring PDF permissions: %s", e)

            # Open the PDF file
            if self.open_job_form_pdf_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success",
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"The PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success",
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"Could not open automatically. Please open manually.")

            # Save to export history and refresh dialog
            try:
                exp_dlg = getattr(self, 'embedded_export', None)
                if exp_dlg:
                    scope_map = {"all": "All Quote Forms", "date_range": "Date Range",
                                 "month": "By Month", "year": "By Year"}
                    scope = scope_map.get(export_params.get("range", "all"), "All Quote Forms")
                    if export_params.get("range") == "date_range":
                        fd = export_params.get("from_date")
                        td = export_params.get("to_date")
                        if fd and td:
                            scope = f"Date Range ({fd.strftime('%b %d, %Y')} - {td.strftime('%b %d, %Y')})"
                    exp_dlg._save_export_entry("PDF", scope, len(jobs_to_export), str(pdf_path))
                    exp_dlg.refresh_recent_exports()
            except Exception:
                pass

        except Exception as e:
            _log.warning("Error generating combined PDF: %s", e)
            _log.exception("Traceback:")
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error",
                                        f"Error generating PDF: {str(e)}")
        
    def open_job_form_pdf_file(self, file_path):
        """Open file with default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == "Darwin":  # macOS
                subprocess.call(("open", file_path))
            elif platform.system() == "Windows":  # Windows
                os.startfile(file_path)
            else:  # linux variants
                subprocess.call(("xdg-open", file_path))
            return True
        except Exception as e:
            _log.warning("Error opening file: %s", e)
            return False

    def create_stats_section(self, layout):
        """Create a compact quote actions toolbar with useful metrics."""
        stats_frame = QtWidgets.QFrame()
        stats_frame.setStyleSheet("""
            QFrame {
                background: #FFFFFF;
                border: 1px solid #DCE4EC;
                border-radius: 10px;
            }
        """)

        main_layout = QtWidgets.QHBoxLayout(stats_frame)
        main_layout.setContentsMargins(20, 16, 18, 16)
        main_layout.setSpacing(16)
        main_layout.setAlignment(QtCore.Qt.AlignVCenter)

        title_col = QtWidgets.QVBoxLayout()
        title_col.setSpacing(2)
        title = QtWidgets.QLabel("Quote Workspace")
        title.setStyleSheet("""
            QLabel {
                color: #111827;
                font-size: 18px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI';
                background: transparent;
                border: none;
            }
        """)
        subtitle = QtWidgets.QLabel("Create, filter, export, and review quote forms.")
        subtitle.setStyleSheet("""
            QLabel {
                color: #64748B;
                font-size: 12px;
                font-weight: 600;
                font-family: 'Inter', 'Segoe UI';
                background: transparent;
                border: none;
            }
        """)
        title_col.addWidget(title)
        title_col.addWidget(subtitle)
        main_layout.addLayout(title_col, 1)

        def _stat_btn(text, bg, hover):
            b = QtWidgets.QPushButton(text)
            b.setFixedHeight(36)
            b.setMinimumWidth(122)
            b.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            b.setStyleSheet(f"""
                QPushButton {{
                    background: {bg}; color: white; border: none;
                    border-radius: 8px; font-size: 12px; font-weight: 800;
                    font-family: 'Inter', 'Segoe UI'; padding: 0 14px;
                }}
                QPushButton:hover {{ background: {hover}; }}
            """)
            return b

        # ── Stat cards ────────────────────────────────────────────────────
        total_jobs          = len(self.job_forms)
        high_priority_count = len([j for j in self.job_forms if j.get('status') in ('High', 'Urgent')])
        completed_count     = len([j for j in self.job_forms if j.get('status') == 'Completed'])
        total_value         = sum(
            float(str(j.get('engineering_costs', '0')).replace('$', '').replace(',', '') or 0)
            for j in self.job_forms
        )
        win_rate = f"{int(completed_count / total_jobs * 100)}%" if total_jobs else "—"

        cards_layout = QtWidgets.QHBoxLayout()
        cards_layout.setSpacing(10)
        for title, val, color in [
            ("Total Quotes",  str(total_jobs),                           "#2563EB"),
            ("High Priority", str(high_priority_count),                  "#B45309"),
            ("Completed",     str(completed_count),                      "#0F766E"),
            ("Total Value",   f"${total_value:,.0f}",                    "#7C3AED"),
            ("Win Rate",      win_rate,                                   "#0891B2"),
        ]:
            card = self.create_invoice_style_stat_card(title, val, color)
            value_label = getattr(card, "value_label", None)
            if title == "Total Quotes":
                self.total_jobs_label = value_label
            elif title == "High Priority":
                self.high_priority_label = value_label
            elif title == "Completed":
                self.completed_label = value_label
            elif title == "Total Value":
                self.total_value_label = value_label
            elif title == "Win Rate":
                self.win_rate_label = value_label
            cards_layout.addWidget(card)
        main_layout.addLayout(cards_layout, 1)

        self.new_quote_tab_btn = _stat_btn("New Quote", "#0F766E", "#115E59")
        self.new_quote_tab_btn.clicked.connect(lambda: self.workflow_tabs.setCurrentIndex(1))
        self.export_tab_btn = _stat_btn("Export", "#475569", "#334155")
        self.export_tab_btn.clicked.connect(lambda: self.workflow_tabs.setCurrentIndex(3))
        action_stack = QtWidgets.QVBoxLayout()
        action_stack.setContentsMargins(0, 0, 0, 0)
        action_stack.setSpacing(10)
        action_stack.addWidget(self.new_quote_tab_btn)
        action_stack.addWidget(self.export_tab_btn)
        main_layout.addLayout(action_stack)

        layout.addWidget(stats_frame)
                
    def eventFilter(self, obj, event):
        """Handle table events including right-click context menu for actions"""
        if event.type() == QtCore.QEvent.ContextMenu and obj is self.job_forms_table.viewport():
            pos = event.pos()
            row = self.job_forms_table.rowAt(pos.y())
            col = self.job_forms_table.columnAt(pos.x())
            if row >= 0:
                job_data = self.get_job_data_from_row(row)
                if job_data:
                    if col == 0:
                        self._show_copy_quote_menu(pos, job_data)
                    else:
                        self.show_actions_context_menu(pos, job_data)
                    return True

        return super().eventFilter(obj, event)

    def _show_copy_quote_menu(self, pos, job_data):
        """Show copy-only context menu for the Quote # column"""
        menu = QtWidgets.QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 7px;
                padding: 4px;
            }
            QMenu::item {
                padding: 7px 20px;
                font-size: 13px;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                border-radius: 5px;
            }
            QMenu::item:selected {
                background: #e6f6f4;
                color: #00756f;
            }
        """)
        copy_action = menu.addAction("Copy Quote Number")
        copy_action.triggered.connect(lambda: self.copy_job_number(job_data))
        menu.exec_(self.job_forms_table.viewport().mapToGlobal(pos))
    
    def show_actions_context_menu(self, pos, job_data):
        """Show context menu for actions"""
        menu = QtWidgets.QMenu(self)
        
        view_action = menu.addAction("👁️ View Details")
        view_action.triggered.connect(lambda: self.view_job_details(job_data))
        
        pdf_action = menu.addAction("📄 Generate PDF")
        pdf_action.triggered.connect(lambda: self.generate_job_form_pdf(job_data))
        
        if self.current_role == 'admin':
            project_action = menu.addAction("🚀 Convert to Project")
            project_action.triggered.connect(lambda: self._on_invoice_click(job_data))
        
        menu.addSeparator()
        
        copy_action = menu.addAction("📋 Copy Quote Number")
        copy_action.triggered.connect(lambda: self.copy_job_number(job_data))
        
        
        menu.exec_(self.job_forms_table.viewport().mapToGlobal(pos))

    def get_job_data_from_row(self, row):
        """Get job data from table row"""
        if row < 0 or row >= len(self.job_forms):
            return None
        
        # Find the job data for this row
        displayed_forms = getattr(self, 'current_displayed_forms', [])
        if row < len(displayed_forms):
            return displayed_forms[row]
        
        return None

    def create_invoice_style_stat_card(self, title, value, color):
        """Stat card — white with coloured left accent bar."""
        card = QtWidgets.QFrame()
        card.setFixedSize(152, 72)
        card.setStyleSheet(f"""
            QFrame {{
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
            }}
            QFrame:hover {{
                background: #F8FAFC;
                border-color: #CBD5E1;
            }}
        """)
        layout = QtWidgets.QHBoxLayout(card)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(10)

        icon = QtWidgets.QLabel()
        icon.setFixedSize(34, 34)
        icon.setAlignment(QtCore.Qt.AlignCenter)
        icon.setStyleSheet(f"""
            QLabel {{
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 8px;
                color: {color};
                font-size: 18px;
                font-weight: 900;
            }}
        """)
        if title == "Total Quotes":
            icon.setPixmap(self._make_briefcase_icon().pixmap(22, 22))
        else:
            icon_map = {
                "High Priority": "☆",
                "Completed": "✓",
                "Total Value": "$",
                "Win Rate": "↗",
            }
            icon.setText(icon_map.get(title, "•"))

        value_label = QtWidgets.QLabel(value)
        value_label.setStyleSheet(f"""
            QLabel {{
                font-size: 18px;
                font-weight: 900;
                color: {color};
                font-family: 'Inter', 'Segoe UI', sans-serif;
                background: transparent;
                border: none;
            }}
        """)
        card.value_label = value_label

        desc_label = QtWidgets.QLabel(title)
        desc_label.setStyleSheet("""
            QLabel {
                font-size: 11px;
                color: #64748B;
                font-weight: 700;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                background: transparent;
                border: none;
            }
        """)
        text_col = QtWidgets.QVBoxLayout()
        text_col.setSpacing(2)
        text_col.addWidget(value_label)
        text_col.addWidget(desc_label)

        layout.addWidget(icon)
        layout.addLayout(text_col, 1)
        return card
    
    def show_date_range_dialog(self):
        """Show date range selection dialog for quote forms"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📅 Select Date Range")
        dialog.setModal(True)
        dialog.resize(400, 200)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        # Title
        title = QtWidgets.QLabel("Select Date Range")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title)
        
        # Date inputs
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(20, 10, 20, 10)
        
        # Check if there's an active date range filter
        current_from_date = QtCore.QDate.currentDate().addMonths(-1)
        current_to_date = QtCore.QDate.currentDate()
        
        # Parse current filter if active
        button_text = self.date_range_button.text()
        if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            # Use stored QDate objects directly
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date
        elif "to" in button_text and button_text != "📅":
            try:
                date_text = button_text.replace("📅 ", "")
                from_str, to_str = date_text.split(" to ")
                # Parse from MM-DD-YY format
                current_from_date = QtCore.QDate.fromString(from_str, "MM-dd-yy")
                current_to_date = QtCore.QDate.fromString(to_str, "MM-dd-yy")
            except:
                # If parsing fails, try MM-dd-yyyy format as fallback
                try:
                    current_from_date = QtCore.QDate.fromString(from_str, "MM-dd-yyyy")
                    current_to_date = QtCore.QDate.fromString(to_str, "MM-dd-yyyy")
                except:
                    # If all parsing fails, use default dates
                    current_from_date = QtCore.QDate.currentDate().addMonths(-1)
                    current_to_date = QtCore.QDate.currentDate()
        
        # From Date - set display format to MM-dd-yyyy for user selection
        # Use local variables to avoid overwriting the inline filter bar's self.from_date_edit
        _dlg_from_date_edit = QtWidgets.QDateEdit()
        _dlg_from_date_edit.setDate(current_from_date)
        _dlg_from_date_edit.setCalendarPopup(True)
        _dlg_from_date_edit.setDisplayFormat("MM-dd-yyyy")
        _dlg_from_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """)

        # To Date - set display format to MM-dd-yyyy for user selection
        _dlg_to_date_edit = QtWidgets.QDateEdit()
        _dlg_to_date_edit.setDate(current_to_date)
        _dlg_to_date_edit.setCalendarPopup(True)
        _dlg_to_date_edit.setDisplayFormat("MM-dd-yyyy")
        _dlg_to_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """)

        form_layout.addRow("From Date:", _dlg_from_date_edit)
        form_layout.addRow("To Date:", _dlg_to_date_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons - Clear on left, Apply on right
        button_layout = QtWidgets.QHBoxLayout()
        
        clear_btn = QtWidgets.QPushButton("Clear Filter")
        clear_btn.setMinimumHeight(40)
        clear_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        
        apply_btn = QtWidgets.QPushButton("Apply Filter")
        apply_btn.setMinimumHeight(40)
        apply_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #2ecc71;
            }
        """)
        
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        button_layout.addWidget(apply_btn)
        
        layout.addLayout(button_layout)
        
        # Connect signals
        def apply_filter():
            # Pass dates from the dialog's local widgets directly so we don't
            # need to store them on self (which would overwrite the inline filter widgets).
            self.apply_date_range_filter(
                _dlg_from_date_edit.date(), _dlg_to_date_edit.date()
            )
            dialog.accept()

        def clear_filter():
            self.clear_date_range_filter()
            dialog.accept()
        
        apply_btn.clicked.connect(apply_filter)
        clear_btn.clicked.connect(clear_filter)
        
        dialog.exec_()
            
    def apply_date_range_filter(self, from_date_qdate=None, to_date_qdate=None):
        """Apply date range filter to quote forms - displays dates in MM-DD-YY format"""
        if from_date_qdate is None:
            from_date_qdate = self.from_date_edit.date()
        if to_date_qdate is None:
            to_date_qdate = self.to_date_edit.date()
        
        # Store the actual QDate objects for later use
        self.current_from_date = from_date_qdate
        self.current_to_date = to_date_qdate
        self._date_filter_active = True
        
        # Format dates as MM-DD-YY
        from_date_formatted = from_date_qdate.toString("MM-dd-yy")
        to_date_formatted = to_date_qdate.toString("MM-dd-yy")
        
        # Update button text to show active filter in MM-DD-YY format
        self.date_range_button.setText(f"📅 {from_date_formatted} to {to_date_formatted}")
        self.date_range_button.setStyleSheet("""
            QPushButton {
                background-color: #0F766E;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #115E59;
            }
            QPushButton:pressed {
                background-color: #134E4A;
            }
        """)
        
        # DO NOT clear other filters when date filter is applied
        # Just apply the filter
        self.filter_job_forms()
        
        # ✅ UPDATE CLIENT FILTER MENU
        self.update_client_filter_menu()

    def clear_date_range_filter(self):
        """Clear date range filter for quote forms WITHOUT clearing other filters"""
        # Clear date filter
        self.date_range_button.setText("Date")
        self.date_range_button.setStyleSheet("""
            QPushButton {
                background-color: #0F766E;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #115E59;
            }
            QPushButton:pressed {
                background-color: #134E4A;
            }
        """)
        
        # Clear the stored date objects
        self._date_filter_active = False
        if hasattr(self, 'current_from_date'):
            del self.current_from_date
        if hasattr(self, 'current_to_date'):
            del self.current_to_date
        
        # ⭐⭐ CRITICAL CHANGE: DO NOT CLEAR OTHER FILTERS
        # Only clear what's being cleared - the date filter
        # DO NOT clear search, client, or status filters
        
        # Apply the cleared filters (only date filter is cleared)
        self.filter_job_forms()
        
        # ✅ UPDATE CLIENT FILTER MENU
        self.update_client_filter_menu()

    @staticmethod
    def _make_filter_icon() -> QtGui.QIcon:
        """Create a compact outlined funnel icon for the filter button."""
        pixmap = QtGui.QPixmap(20, 20)
        pixmap.fill(QtCore.Qt.transparent)

        painter = QtGui.QPainter(pixmap)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor("#64748B"))
        pen.setWidthF(2.2)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        pen.setCapStyle(QtCore.Qt.RoundCap)
        painter.setPen(pen)
        painter.setBrush(QtCore.Qt.NoBrush)

        path = QtGui.QPainterPath()
        path.moveTo(3.0, 4.0)
        path.lineTo(17.0, 4.0)
        path.lineTo(12.0, 10.0)
        path.lineTo(12.0, 15.0)
        path.lineTo(8.0, 17.0)
        path.lineTo(8.0, 10.0)
        path.closeSubpath()
        painter.drawPath(path)
        painter.end()

        return QtGui.QIcon(pixmap)

    @staticmethod
    def _make_briefcase_icon() -> QtGui.QIcon:
        """Create the outlined briefcase-style icon used for Total Quotes."""
        pixmap = QtGui.QPixmap(24, 24)
        pixmap.fill(QtCore.Qt.transparent)

        painter = QtGui.QPainter(pixmap)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor("#2563EB"))
        pen.setWidthF(2.2)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        pen.setCapStyle(QtCore.Qt.RoundCap)
        painter.setPen(pen)
        painter.setBrush(QtCore.Qt.NoBrush)

        body = QtCore.QRectF(4.0, 8.0, 16.0, 11.0)
        painter.drawRoundedRect(body, 1.8, 1.8)
        painter.drawLine(QtCore.QPointF(4.0, 11.0), QtCore.QPointF(20.0, 11.0))

        handle = QtGui.QPainterPath()
        handle.moveTo(9.0, 8.0)
        handle.lineTo(9.0, 6.0)
        handle.lineTo(15.0, 6.0)
        handle.lineTo(15.0, 8.0)
        painter.drawPath(handle)
        painter.end()

        return QtGui.QIcon(pixmap)
        
    def create_job_forms_table_section(self, layout):
        """Create quote forms table section below stats cards, with date filter, search, and status filter"""
        # Work directly with the passed layout — no extra GroupBox wrapper
        table_layout = layout
        table_layout.setSpacing(12)

        # 🔍 Integrated Search + Filters Section
        search_filter_frame = QtWidgets.QFrame()
        search_filter_frame.setStyleSheet("QFrame { background: transparent; border: none; }")

        search_filter_layout = QtWidgets.QHBoxLayout(search_filter_frame)
        search_filter_layout.setSpacing(12)
        search_filter_layout.setContentsMargins(0, 0, 0, 0)

        # Search bar
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search by client, job title, project details...")
        self.search_edit.setPlaceholderText("Search by client, quote number, project, sales, or scope...")
        self.search_edit.setMinimumHeight(38)
        self.search_edit.setMinimumWidth(380)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 9px;
                font-size: 13px;
                background: white;
                min-width: 300px;
            }
            QLineEdit:focus { border-color: #0F766E; background: #F8FAFC; }
        """)
        self.search_edit.textChanged.connect(self.filter_job_forms)
        search_filter_layout.addWidget(self.search_edit)

        # Flexible space
        search_filter_layout.addStretch(1)

        # Filters container (right side)
        filters_container = QtWidgets.QHBoxLayout()
        filters_container.setSpacing(8)
        filters_container.setContentsMargins(0, 0, 0, 0)

        # Status filter
        status_label = QtWidgets.QLabel("Status:")
        status_label.setStyleSheet("font-weight: 800; color: #1E293B; font-size: 13px; padding: 5px 0px;")
        filters_container.addWidget(status_label)

        self.status_filter_combo = QtWidgets.QComboBox()
        self.status_filter_combo.addItems([
            "All Status",
            "Draft", "Sent", "In Review",
            "Approved", "On Hold",
            "Completed", "Converted",
            "Rejected", "Expired", "Cancelled",
        ])
        self.status_filter_combo.setMinimumHeight(38)
        self.status_filter_combo.setStyleSheet(f"""
            QComboBox {{
                padding: 6px 38px 6px 14px;
                border: 1.5px solid #D9E2EC;
                border-radius: 10px;
                background: #FFFFFF;
                min-width: 176px;
                font-size: 14px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-weight: 800;
                color: #0F172A;
                margin-right: 2px;
            }}
            QComboBox:hover  {{
                border-color: #CBD5E1;
                background: #F8FAFC;
            }}
            QComboBox:focus  {{
                border-color: #0F766E;
                background: #FFFFFF;
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 34px;
                border-left: none;
                border-top-right-radius: 10px;
                border-bottom-right-radius: 10px;
                background: transparent;
            }}
            QComboBox::down-arrow {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                margin-right: 8px;
            }}
            QComboBox QAbstractItemView {{
                border: 1px solid #D9E2EC;
                border-radius: 10px;
                background: #FFFFFF;
                selection-background-color: #ECFDF5;
                selection-color: #0F766E;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 13px;
                padding: 6px;
            }}
        """)
        self.status_filter_combo.currentTextChanged.connect(self.on_status_filter_changed)
        self.status_filter_combo.wheelEvent = lambda e: e.ignore()
        self.status_filter_combo.keyPressEvent = lambda e, c=self.status_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.status_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.status_filter_combo.clearFocus))
        filters_container.addWidget(self.status_filter_combo)

        # Client filter button
        self.client_filter_button = QtWidgets.QPushButton("🗂️")
        self.client_filter_button.setText("Client")
        self.client_filter_button.setFixedHeight(38)
        self.client_filter_button.setMinimumWidth(78)
        self.client_filter_button.setMaximumWidth(380)  # Allow it to grow with text
        self.client_filter_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.client_filter_button.setStyleSheet("""
            QPushButton {
                background: #FFFFFF;
                border: 1.5px solid #E2E8F0;
                border-radius: 9px;
                font-size: 13px;
                font-weight: 800;
                color: #334155;
                padding: 0px 12px;
            }
            QPushButton:hover {
                color: #2563EB;
                border-color: #93C5FD;
            }
        """)

        # Client filter menu
        self.client_filter_menu = QtWidgets.QMenu(self)
        self.client_filter_menu.setStyleSheet("""
            QMenu {
                background: white;
                border: 1px solid #cfd4da;
                border-radius: 8px;
                padding: 6px;
                margin-top: 5px;
            }
            QMenu::item {
                padding: 8px 14px;
                border-radius: 6px;
                font-size: 13px;
            }
            QMenu::item:selected {
                background: #3498db;
                color: white;
            }
        """)

        self.client_filter_button.clicked.connect(
            lambda: self.client_filter_menu.exec_(
                self.client_filter_button.mapToGlobal(
                    QtCore.QPoint(-40, self.client_filter_button.height() + 5)
                )
            )
        )

        filters_container.addWidget(self.client_filter_button)
        search_filter_layout.addLayout(filters_container)
        table_layout.addWidget(search_filter_frame)

        # ── Year / Month filter bar ───────────────────────────────────────────
        ym_frame = QtWidgets.QFrame()
        ym_frame.setStyleSheet(
            "QFrame { background:transparent; border:none; }")
        ym_layout = QtWidgets.QHBoxLayout(ym_frame)
        ym_layout.setContentsMargins(0, 6, 0, 6)
        ym_layout.setSpacing(12)

        ym_lbl = QtWidgets.QLabel("📆  Period:")
        ym_lbl.setStyleSheet(
            "font-weight:800; color:#1E293B; font-size:13px;"
            " background:transparent; border:none;")
        ym_layout.addWidget(ym_lbl)

        _combo_qss = (
            "QComboBox { padding:4px 28px 4px 10px; border:1.5px solid #e1e8ed;"
            " border-radius:7px; background:white; font-size:13px; font-weight:600;"
            " color:#1e293b; }"
            "QComboBox:hover { border-color:#94a3b8; }"
            "QComboBox:focus { border-color:#0F766E; }"
            "QComboBox::drop-down { width:24px; border:none; }"
            "QComboBox QAbstractItemView { background:white; border:1px solid #e1e8ed;"
            " selection-background-color:#f0fdf4; selection-color:#0F766E;"
            " font-size:13px; padding:4px; }"
        )

        current_year = datetime.now().year
        self.year_filter_combo = QtWidgets.QComboBox()
        self.year_filter_combo.addItems(
            ["All Years"] + [str(y) for y in range(current_year, current_year - 7, -1)])
        self.year_filter_combo.setFixedHeight(40)
        self.year_filter_combo.setMinimumWidth(115)
        self.year_filter_combo.setStyleSheet(_combo_qss)
        self.year_filter_combo.currentTextChanged.connect(self.filter_job_forms)
        self.year_filter_combo.wheelEvent = lambda e: e.ignore()
        self.year_filter_combo.keyPressEvent = lambda e, c=self.year_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.year_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.year_filter_combo.clearFocus))
        ym_layout.addWidget(self.year_filter_combo)

        self.month_filter_combo = QtWidgets.QComboBox()
        self.month_filter_combo.addItems([
            "All Months", "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"])
        self.month_filter_combo.setFixedHeight(40)
        self.month_filter_combo.setMinimumWidth(135)
        self.month_filter_combo.setStyleSheet(_combo_qss)
        self.month_filter_combo.currentTextChanged.connect(self.filter_job_forms)
        self.month_filter_combo.wheelEvent = lambda e: e.ignore()
        self.month_filter_combo.keyPressEvent = lambda e, c=self.month_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.month_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.month_filter_combo.clearFocus))
        ym_layout.addWidget(self.month_filter_combo)

        clear_period_btn = QtWidgets.QPushButton("✕ Clear")
        clear_period_btn.setFixedSize(92, 40)
        clear_period_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        clear_period_btn.setStyleSheet(
            "QPushButton { background:transparent; border:1px solid #cbd5e1;"
            " border-radius:6px; font-size:12px; font-weight:700; color:#64748b; }"
            "QPushButton:hover { background:#fee2e2; border-color:#fca5a5; color:#dc2626; }")
        clear_period_btn.clicked.connect(self._clear_period_filter)
        ym_layout.addWidget(clear_period_btn)

        # ── Date range pickers ──────────────────────────────────────────
        _date_sep = QtWidgets.QFrame()
        _date_sep.setFixedWidth(1)
        _date_sep.setStyleSheet("background:#e2e8f0; border:none;")
        ym_layout.addSpacing(6)
        ym_layout.addWidget(_date_sep)
        ym_layout.addSpacing(6)

        date_lbl = QtWidgets.QLabel("📅  Date Range:")
        date_lbl.setStyleSheet(
            "font-weight:800; color:#1E293B; font-size:13px;"
            " background:transparent; border:none;")
        ym_layout.addWidget(date_lbl)

        _date_qss = (
            "QDateEdit { padding:4px 8px; border:1.5px solid #e1e8ed;"
            " border-radius:7px; background:white; font-size:12px; font-weight:600;"
            " color:#1e293b; }"
            "QDateEdit:focus { border-color:#0F766E; }"
            "QDateEdit::drop-down { width:20px; border:none; }"
        )
        self.from_date_edit = QtWidgets.QDateEdit()
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date_edit.setDisplayFormat("MM/dd/yyyy")
        self.from_date_edit.setMinimumWidth(175)
        self.from_date_edit.setFixedHeight(40)
        self.from_date_edit.setStyleSheet(_date_qss)
        ym_layout.addWidget(self.from_date_edit)

        to_lbl = QtWidgets.QLabel("→")
        to_lbl.setStyleSheet("color:#64748b; font-size:13px; background:transparent; border:none;")
        ym_layout.addWidget(to_lbl)

        self.to_date_edit = QtWidgets.QDateEdit()
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDate(QtCore.QDate.currentDate())
        self.to_date_edit.setDisplayFormat("MM/dd/yyyy")
        self.to_date_edit.setMinimumWidth(175)
        self.to_date_edit.setFixedHeight(40)
        self.to_date_edit.setStyleSheet(_date_qss)
        ym_layout.addWidget(self.to_date_edit)

        self.date_filter_active_chk = configure_filter_button(
            QtWidgets.QPushButton(),
            height=40,
        )
        self.date_filter_active_chk.setCheckable(True)

        def _toggle_inline_date_filter(checked):
            configure_filter_button(
                self.date_filter_active_chk,
                "Filter",
                active=checked,
                height=40,
            )
            self.date_filter_active_chk.setChecked(checked)
            self.filter_job_forms()

        self.date_filter_active_chk.toggled.connect(_toggle_inline_date_filter)
        self.from_date_edit.dateChanged.connect(
            lambda: self.date_filter_active_chk.isChecked() and self.filter_job_forms())
        self.to_date_edit.dateChanged.connect(
            lambda: self.date_filter_active_chk.isChecked() and self.filter_job_forms())
        self.from_date_edit.wheelEvent = lambda e: e.ignore()
        self.from_date_edit.stepBy = lambda x: None
        self.to_date_edit.wheelEvent = lambda e: e.ignore()
        self.to_date_edit.stepBy = lambda x: None
        ym_layout.addWidget(self.date_filter_active_chk)

        ym_layout.addStretch()

        self.period_summary_lbl = QtWidgets.QLabel("")
        self.period_summary_lbl.setStyleSheet(
            "color:#0f766e; font-size:13px; font-weight:700;"
            " background:transparent; border:none;")
        ym_layout.addWidget(self.period_summary_lbl)

        table_layout.addWidget(ym_frame)

        # Quote Forms Table with updated columns
        self.job_forms_table = QtWidgets.QTableWidget()
        self.job_forms_table.setColumnCount(6)
        self.job_forms_table.setHorizontalHeaderLabels([
            "QUOTE #", "SALES", "PROJECT NAME", "CLIENT", "STATUS", "COST"
        ])
        
        self.job_forms_table.setStyleSheet("""
        QTableWidget {
            background: #ffffff;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            gridline-color: #f1f5f9;
            alternate-background-color: #f8fafc;
            outline: none;
        }
        QTableWidget::item {
            padding: 10px 12px;
            font-size: 13px;
            font-family: 'Inter', 'Segoe UI', sans-serif;
            border-bottom: 1px solid #f1f5f9;
        }
        QTableWidget::item:selected {
            background: #f0fdf4;
        }
        QTableWidget::item:hover {
            background: #f8fafc;
        }
        QHeaderView::section {
            background: #f8fafc;
            color: #475569;
            font-size: 12px;
            font-weight: 800;
            font-family: 'Inter', 'Segoe UI', sans-serif;
            letter-spacing: 0.6px;
            padding: 10px 12px;
            border: none;
            border-bottom: 2px solid #0f766e;
            border-right: 1px solid #e2e8f0;
        }
        QHeaderView::section:last { border-right: none; }
    """)
        
        # Enable grid to show vertical lines
        self.job_forms_table.setShowGrid(True)  # Changed to True to show grid lines
        self.job_forms_table.setGridStyle(QtCore.Qt.SolidLine)
        
        # Table properties for compact look
        self.job_forms_table.horizontalHeader().setStretchLastSection(False)
        self.job_forms_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.job_forms_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.job_forms_table.setAlternatingRowColors(True)
        
        self.job_forms_table.verticalHeader().setVisible(False)
        self.job_forms_table.viewport().installEventFilter(self)
        
        # Header properties
        header= self.job_forms_table.horizontalHeader()
        header.setDefaultAlignment(QtCore.Qt.AlignCenter)
        header.setHighlightSections(False)
        header.setFixedHeight(48)

        # Quote# / Sales / Client / Status / Cost fixed; Project Name stretches.
        for col in range(self.job_forms_table.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)   # Project Name only

        self.job_forms_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.job_forms_table.setColumnWidth(0, 170)   # Quote Number
        self.job_forms_table.setColumnWidth(1, 220)   # Sales
        self.job_forms_table.setColumnWidth(3, 320)   # Client (fixed)
        self.job_forms_table.setColumnWidth(4, 210)   # Status
        self.job_forms_table.setColumnWidth(5, 180)   # Cost
        self.job_forms_table.setTextElideMode(QtCore.Qt.ElideNone)
        self.job_forms_table.verticalHeader().setDefaultSectionSize(58)
        self.job_forms_table.setWordWrap(True)
        self.job_forms_table.cellClicked.connect(self._on_table_cell_clicked)

        table_layout.addWidget(self.job_forms_table, 1)  # stretch=1 so table fills remaining height

        # ── Quotes pagination bar ──────────────────────────────────────────────
        self._qf_pg_s = """
            QPushButton {
                background:white; color:#374151; border:1px solid #E2E8F0;
                border-radius:6px; font-size:12px; font-weight:700;
                min-width:32px; min-height:28px; padding:0 8px;
            }
            QPushButton:hover { background:#F8FAFC; }
            QPushButton:disabled { color:#D1D5DB; background:#F9FAFB; }
        """
        qf_ftr = QtWidgets.QHBoxLayout()
        qf_ftr.setContentsMargins(0, 6, 0, 0)
        self.qf_count_lbl = QtWidgets.QLabel("Showing 0 entries")
        self.qf_count_lbl.setStyleSheet(
            "font-size:13px; color:#6B7280; background:transparent; border:none;"
        )
        self._qf_prev_btn = QtWidgets.QPushButton("<")
        self._qf_prev_btn.setFixedSize(32, 28)
        self._qf_prev_btn.setStyleSheet(self._qf_pg_s)
        self._qf_prev_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._qf_prev_btn.clicked.connect(self._qf_go_prev)
        _qf_page_btns_w = QtWidgets.QWidget()
        _qf_page_btns_w.setStyleSheet("background:transparent;")
        self._qf_page_btns_layout = QtWidgets.QHBoxLayout(_qf_page_btns_w)
        self._qf_page_btns_layout.setContentsMargins(0, 0, 0, 0)
        self._qf_page_btns_layout.setSpacing(4)
        self._qf_next_btn = QtWidgets.QPushButton(">")
        self._qf_next_btn.setFixedSize(32, 28)
        self._qf_next_btn.setStyleSheet(self._qf_pg_s)
        self._qf_next_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._qf_next_btn.clicked.connect(self._qf_go_next)
        qf_ftr.addWidget(self.qf_count_lbl)
        qf_ftr.addStretch()
        qf_ftr.addWidget(self._qf_prev_btn)
        qf_ftr.addWidget(_qf_page_btns_w)
        qf_ftr.addWidget(self._qf_next_btn)
        table_layout.addLayout(qf_ftr)

    def display_filtered_forms(self, forms):
        self._qf_all_items = forms
        self._qf_page = 1
        self._qf_render_page()

    def _qf_render_page(self):
        all_items = self._qf_all_items
        per = self._qf_per_page
        page = self._qf_page
        total = len(all_items)
        start_i = (page - 1) * per
        end_i = min(start_i + per, total)
        forms = all_items[start_i:end_i]

        self.job_forms_table.setRowCount(len(forms))
        self.current_displayed_forms = forms

        for row, job in enumerate(forms):
          try:
            job_number = job.get('job_number', '')
            job_number_item = QtWidgets.QTableWidgetItem(job_number)
            job_number_item.setForeground(QtGui.QColor('#2563EB'))
            f = QtGui.QFont("Inter", 11, QtGui.QFont.Bold)
            f.setUnderline(True)
            job_number_item.setFont(f)
            job_number_item.setTextAlignment(QtCore.Qt.AlignCenter)
            job_number_item.setBackground(QtGui.QColor('#EFF6FF'))
            job_number_item.setToolTip(f"Click {job_number} to open full quote details")
            job_number_item.setData(QtCore.Qt.UserRole, job)
            self.job_forms_table.setItem(row, 0, job_number_item)

            project_name = job.get('project_name', '')
            age_str = ""
            try:
                created_raw = job.get('created_at', '')
                if created_raw:
                    created_dt = datetime.fromisoformat(str(created_raw).replace("Z", ""))
                    days = (datetime.now() - created_dt).days
                    if days == 0:
                        age_str = "Today"
                    elif days == 1:
                        age_str = "1 day ago"
                    else:
                        age_str = f"{days}d ago"
            except Exception:
                pass

            project_name_item = QtWidgets.QTableWidgetItem(project_name)
            project_name_item.setForeground(QtGui.QColor('#0f172a'))
            project_name_item.setFont(QtGui.QFont("Inter", 11))
            project_name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            if age_str:
                project_name_item.setToolTip(f"{project_name} — created {age_str}")

            sales_val = job.get("sales", "")
            sales_item = QtWidgets.QTableWidgetItem(sales_val)
            sales_item.setFont(QtGui.QFont("Inter", 11))
            sales_item.setForeground(QtGui.QColor("#334155"))
            if sales_val:
                sales_item.setToolTip(sales_val)
            sales_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.job_forms_table.setItem(row, 1, sales_item)

            self.job_forms_table.setItem(row, 2, project_name_item)

            client_val = job.get('client', '')
            client_item = QtWidgets.QTableWidgetItem(client_val)
            client_item.setForeground(QtGui.QColor('#0f172a'))
            client_item.setFont(QtGui.QFont("Inter", 11))
            client_item.setTextAlignment(QtCore.Qt.AlignCenter)
            client_item.setToolTip(client_val)
            self.job_forms_table.setItem(row, 3, client_item)

            current_status = job.get('status', 'Not Started')
            status_badge = self.create_status_badge(current_status, job, row)
            self.job_forms_table.setCellWidget(row, 4, status_badge)

            cost = job.get('engineering_costs', 'N/A')
            display_cost = cost
            is_expedited = job.get('expedite') is True
            if is_expedited and cost and cost != 'N/A':
                try:
                    base_val = float(str(cost).replace('$', '').replace(',', '').strip())
                    display_cost = f"${base_val * 1.5:,.2f} (Exp.)"
                except Exception:
                    display_cost = cost
            cost_item = QtWidgets.QTableWidgetItem(display_cost)
            if display_cost != 'N/A' and '$' in str(display_cost):
                cost_item.setForeground(QtGui.QColor('#047857'))
                cost_item.setFont(QtGui.QFont("Inter", 11, QtGui.QFont.Bold))
            else:
                cost_item.setForeground(QtGui.QColor('#94a3b8'))
                cost_item.setFont(QtGui.QFont("Inter", 11))
            cost_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.job_forms_table.setItem(row, 5, cost_item)

          except Exception as e:
            _log.warning("_qf_render_page: skipped row %d (%s): %s",
                         row, job.get('job_number', '?'), e)

        for row in range(self.job_forms_table.rowCount()):
            self.job_forms_table.setRowHeight(row, 58)

        if hasattr(self, 'qf_count_lbl'):
            if total == 0:
                self.qf_count_lbl.setText("No entries found")
            else:
                self.qf_count_lbl.setText(f"Showing {start_i + 1} to {end_i} of {total} entries")

        self.job_forms_table.viewport().update()
        dynamic_height = 48 + (self.job_forms_table.rowCount() * 62)
        final_height = max(400, min(dynamic_height, 870))
        self.job_forms_table.setMinimumHeight(final_height)
        self.job_forms_table.setMaximumHeight(16777215)

        self._qf_rebuild_pagination()

    def _qf_rebuild_pagination(self):
        if not hasattr(self, '_qf_page_btns_layout'):
            return
        while self._qf_page_btns_layout.count():
            _it = self._qf_page_btns_layout.takeAt(0)
            if _it.widget():
                _it.widget().deleteLater()
        total = len(self._qf_all_items)
        max_page = max(1, (total + self._qf_per_page - 1) // self._qf_per_page)
        page_num = self._qf_page
        self._qf_prev_btn.setEnabled(page_num > 1)
        self._qf_next_btn.setEnabled(page_num < max_page)
        _win_start = max(1, min(page_num, max_page - 2))
        for p in range(_win_start, min(_win_start + 3, max_page + 1)):
            btn = QtWidgets.QPushButton(str(p))
            btn.setFixedSize(32, 28)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            if p == page_num:
                btn.setStyleSheet("""QPushButton {
                    background-color: #00756f; color: #ffffff;
                    border: 1px solid #00756f; border-radius: 6px;
                    font-size: 12px; font-weight: 700;
                    min-width: 32px; min-height: 28px; padding: 0 8px;
                }
                QPushButton:hover { background-color: #005f5a; color: #ffffff; }""")
            else:
                btn.setStyleSheet(self._qf_pg_s)
                btn.clicked.connect(lambda _, pg=p: self._qf_go_to(pg))
            self._qf_page_btns_layout.addWidget(btn)

    def _qf_go_prev(self):
        if self._qf_page > 1:
            self._qf_page -= 1
            self._qf_render_page()

    def _qf_go_next(self):
        total = len(self._qf_all_items)
        max_page = max(1, (total + self._qf_per_page - 1) // self._qf_per_page)
        if self._qf_page < max_page:
            self._qf_page += 1
            self._qf_render_page()

    def _qf_go_to(self, p):
        self._qf_page = p
        self._qf_render_page()

    # ── status palette: (background, text, border) ───────────────────────────
    STATUS_PALETTE = {
        # bg, fg, border  (used by table badge buttons)
        "Draft":       ("#f1f5f9", "#475569", "#cbd5e1"),
        "Sent":        ("#dbeafe", "#1e40af", "#93c5fd"),
        "In Review":   ("#fef3c7", "#92400e", "#fcd34d"),
        "Approved":    ("#d1fae5", "#065f46", "#6ee7b7"),
        "On Hold":     ("#f3f4f6", "#374151", "#d1d5db"),
        "Completed":   ("#d1fae5", "#065f46", "#6ee7b7"),
        "Converted":   ("#ede9fe", "#4c1d95", "#c4b5fd"),
        "Rejected":    ("#fee2e2", "#991b1b", "#fca5a5"),
        "Expired":     ("#fff7ed", "#7c2d12", "#fed7aa"),
        "Cancelled":   ("#fce7f3", "#9d174d", "#f9a8d4"),
        # Legacy aliases kept for existing data
        "Not Started": ("#f9fafb", "#6b7280", "#e5e7eb"),
        "Cancel":      ("#fce7f3", "#9d174d", "#f9a8d4"),
        "Low":         ("#f0fdf4", "#166534", "#86efac"),
        "Medium":      ("#dbeafe", "#1e40af", "#93c5fd"),
        "High":        ("#fee2e2", "#991b1b", "#fca5a5"),
        "Urgent":      ("#fef3c7", "#92400e", "#fcd34d"),
    }

    # ── Quote # click → detail window ─────────────────────────────────────

    def _on_table_cell_clicked(self, row, col):
        if col != 0:
            return
        item = self.job_forms_table.item(row, 0)
        if not item:
            return
        job_data = item.data(QtCore.Qt.UserRole)
        if job_data:
            dlg = QuoteDetailWindow(job_data, self)
            dlg.status_changed.connect(
                lambda new_status, jd=job_data: self._apply_status_from_detail(new_status, jd, row)
            )
            dlg.action_requested.connect(
                lambda action, jd=job_data: self._handle_detail_action(action, jd)
            )
            dlg.exec_()

    def _apply_status_from_detail(self, new_status, job_data, row):
        """Refresh the status badge in the table after a change in the detail window."""
        job_data['status'] = new_status
        for j in self.job_forms:
            if j.get('job_number') == job_data.get('job_number'):
                j['status'] = new_status
                break
        # Explicitly update the table item's stored data so reopening the popup
        # shows the updated status rather than the original one.
        item = self.job_forms_table.item(row, 0)
        if item:
            stored = item.data(QtCore.Qt.UserRole)
            if isinstance(stored, dict):
                stored['status'] = new_status
                item.setData(QtCore.Qt.UserRole, stored)
        badge = self.create_status_badge(new_status, job_data, row)
        self.job_forms_table.setCellWidget(row, 4, badge)
        if self.FIREBASE_AVAILABLE:
            self.update_job_status_in_firebase(job_data, new_status)
        self.update_filtered_stats(self.job_forms)

    def _handle_detail_action(self, action, job_data):
        if action == "view":
            self.view_job_details(job_data)
        elif action == "pdf":
            self.open_job_form_pdf(job_data)
        elif action == "project":
            try:
                enhanced = self._prepare_enhanced_project_data(job_data)
                self.convert_to_project.emit(enhanced)
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", str(e))
        elif action == "edit":
            self.edit_job_form(job_data)
        elif action == "copy":
            self.copy_job_number(job_data)
        elif action == "email":
            self.email_quote_to_client(job_data)
        elif action == "delete":
            self.delete_job_form(job_data)

    def create_status_badge(self, status: str, job_data: dict, row: int) -> QtWidgets.QWidget:
        """Return a clickable pill badge for the given status."""
        bg, fg, border = self.STATUS_PALETTE.get(
            status, ("#f9fafb", "#6b7280", "#e5e7eb"))

        container = QtWidgets.QWidget()
        lay = QtWidgets.QHBoxLayout(container)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(0)
        lay.addStretch(1)
        badge = QtWidgets.QPushButton(status)
        badge.setFixedSize(178, 34)
        badge.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        badge.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg};
                color: {fg};
                border: 1px solid {border};
                border-radius: 7px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 30px 0 14px;
                text-align: left;
            }}
            QPushButton:hover {{
                border-width: 1.5px;
                opacity: 0.9;
            }}
            QPushButton::menu-indicator {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                subcontrol-origin: padding;
                subcontrol-position: center right;
                right: 12px;
            }}
        """)

        from PyQt5.QtWidgets import QMenu, QAction
        menu = QMenu(badge)
        menu.setStyleSheet("""
            QMenu {
                background: white;
                border: 1px solid #d0d7de;
                border-radius: 8px;
                padding: 4px 0;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 12px;
            }
            QMenu::item { padding: 7px 20px; color: #24292f; }
            QMenu::item:selected { background: #f6f8fa; color: #0969da; }
            QMenu::separator { height:1px; background:#e5e7eb; margin:3px 0; }
        """)
        groups = [
            ["Draft", "Sent", "In Review"],
            ["Approved", "On Hold"],
            ["Completed", "Converted"],
            ["Rejected", "Expired", "Cancelled"],
        ]
        first = True
        for grp in groups:
            if not first:
                menu.addSeparator()
            first = False
            for s in grp:
                a = QAction(s, menu)
                a.triggered.connect(
                    lambda _, st=s, bref=badge, jref=job_data, rref=row:
                        self._apply_status_badge(st, bref, jref, rref))
                menu.addAction(a)
        badge.setMenu(menu)
        lay.addWidget(badge)
        lay.addStretch(1)
        return container

    def _apply_status_badge(self, new_status: str, badge_btn: QtWidgets.QPushButton,
                             job_data: dict, row: int):
        """Update badge colour + persist status change."""
        bg, fg, border = self.STATUS_PALETTE.get(
            new_status, ("#f9fafb", "#6b7280", "#e5e7eb"))
        badge_btn.setText(new_status)
        badge_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg};
                color: {fg};
                border: 1px solid {border};
                border-radius: 7px;
                font-size: 13px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                padding: 0 30px 0 14px;
                text-align: left;
            }}
            QPushButton:hover {{ border-width: 1.5px; }}
            QPushButton::menu-indicator {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                subcontrol-origin: padding;
                subcontrol-position: center right;
                right: 12px;
            }}
        """)
        job_data['status'] = new_status
        for job in self.job_forms:
            if job.get('job_number') == job_data.get('job_number'):
                job['status'] = new_status
                break
        # Explicitly update the item stored data so reopening the popup reflects the change.
        item = self.job_forms_table.item(row, 0)
        if item:
            stored = item.data(QtCore.Qt.UserRole)
            if isinstance(stored, dict):
                stored['status'] = new_status
                item.setData(QtCore.Qt.UserRole, stored)
        if self.FIREBASE_AVAILABLE:
            self.update_job_status_in_firebase(job_data, new_status)
        self.update_filtered_stats(self.job_forms)

    def show_job_status_notification(self, message: str, msg_type: str = "info"):
        pass   # Disable all popups completely

    def get_status_style(self, status):
        """Get professional color scheme for status badges"""
        status_styles = {
            'Not Started': ('#f8f9fa', '#6c757d'),
            'Completed': ('#e8f5e8', '#2e7d32'),
            'Urgent': ('#FEF3C7', "#92400E"),
            'High': ("#FFEDD5", "#9A3412"),
            'Medium': ('#fff3cd', '#f39c12'),
            'Low': ('#e8f5e8', "#9ed4dc")
        }
        return status_styles.get(status, ('#f8f9fa', '#6c757d'))

    def _prepare_enhanced_project_data(self, job_data):
        """Prepare enhanced project data with intelligent auto-population"""
        from datetime import datetime, timedelta
        import json
        
        enhanced_data = job_data.copy()
        
        # Auto-generate intelligent project number
        if not enhanced_data.get('project_number'):
            company_code = 'MABS'  # Get from company settings
            date_str = datetime.now().strftime("%Y%m")
            enhanced_data['project_number'] = f"{company_code}-{date_str}-001"
        
        # Auto-set project dates based on quote data
        if not enhanced_data.get('start_date'):
            start_date = datetime.now().strftime("%m-%d-%Y")
            enhanced_data['start_date'] = start_date
        
        # Auto-calculate due date (30 working days from start)
        if not enhanced_data.get('due_date'):
            start_date = datetime.strptime(enhanced_data['start_date'], "%m-%d-%Y")
            due_date = start_date + timedelta(days=30)
            enhanced_data['due_date'] = due_date.strftime("%m-%d-%Y")
        
        # Auto-detect payment category based on project amount
        amount = enhanced_data.get('engineering_costs', '0')
        try:
            amount_float = float(str(amount).replace('$', '').replace(',', ''))
            # Include expedite premium in project amount
            if enhanced_data.get('expedite') is True:
                exp_str = str(enhanced_data.get('expedite_amount', '50%')).strip()
                try:
                    if '%' in exp_str:
                        pct = float(exp_str.replace('%', '').strip())
                        amount_float = amount_float * (1 + pct / 100)
                    elif '$' in exp_str:
                        extra = float(exp_str.replace('$', '').replace(',', '').strip())
                        amount_float = amount_float + extra
                    else:
                        amount_float = amount_float * 1.5
                except Exception:
                    amount_float = amount_float * 1.5
            enhanced_data['project_amount'] = amount_float  # Store total (with expedite) for project
            
            # Auto-detect payment category
            if amount_float < 5000:
                enhanced_data['payment_category'] = 'Small Project'
                enhanced_data['payment_structure'] = 'Single Payment'
            elif amount_float < 20000:
                enhanced_data['payment_category'] = 'Medium Project'
                enhanced_data['payment_structure'] = 'Down Payment + Final'
            else:
                enhanced_data['payment_category'] = 'Large Project'
                enhanced_data['payment_structure'] = 'Milestone Payments'
                
            # Auto-calculate payment amounts
            if amount_float < 5000:
                enhanced_data['down_payment'] = 0
                enhanced_data['final_payment'] = amount_float
            elif amount_float < 20000:
                enhanced_data['down_payment'] = amount_float * 0.5
                enhanced_data['final_payment'] = amount_float * 0.5
            else:
                enhanced_data['down_payment'] = amount_float * 0.3
                enhanced_data['milestone1'] = amount_float * 0.4
                enhanced_data['final_payment'] = amount_float * 0.3
                
        except:
            enhanced_data['payment_category'] = 'Standard'
            enhanced_data['payment_structure'] = 'Single Payment'
            enhanced_data['project_amount'] = 0
        
        # Auto-set initial status
        enhanced_data['status'] = 'Not Started'
        
        # Auto-populate project name from job title if missing
        if not enhanced_data.get('project_name'):
            enhanced_data['project_name'] = enhanced_data.get('job_title', 'Untitled Project')
        
        # Auto-transfer sales person from quote
        if enhanced_data.get('sales'):
            enhanced_data['sales_person'] = enhanced_data['sales']
            enhanced_data['assigned_sales'] = enhanced_data['sales']
        
        # Auto-transfer client information
        if enhanced_data.get('client'):
            enhanced_data['client_name'] = enhanced_data['client']
            enhanced_data['company'] = enhanced_data['client']
        
        # Auto-transfer client contact details
        if enhanced_data.get('client_email'):
            enhanced_data['client_contact_email'] = enhanced_data['client_email']
            enhanced_data['contact_email'] = enhanced_data['client_email']
        
        if enhanced_data.get('client_address'):
            enhanced_data['client_address'] = enhanced_data['client_address']
            enhanced_data['mailing_address'] = enhanced_data['client_address']
        
        # Auto-set created date
        enhanced_data['created_at'] = datetime.now(timezone.utc).isoformat()
        
        # Auto-detect services from scope
        scope = enhanced_data.get('scope_of_work', '').lower()
        services = []
        if 'structural' in scope:
            services.append('Structural')
        if 'civil' in scope:
            services.append('Civil')
        if 'electrical' in scope:
            services.append('Electrical')
        if 'mechanical' in scope:
            services.append('Mechanical')
        if 'plumbing' in scope:
            services.append('Plumbing Design')
        
        enhanced_data['services'] = services
        
        # Auto-calculate project duration
        try:
            start_dt = datetime.strptime(enhanced_data['start_date'], "%m-%d-%Y")
            due_dt = datetime.strptime(enhanced_data['due_date'], "%m-%d-%Y")
            duration_days = (due_dt - start_dt).days
            enhanced_data['estimated_duration'] = f"{duration_days} days"
        except:
            enhanced_data['estimated_duration'] = "30 days"
        
        return enhanced_data

    def show_template_dialog(self):
        """Show template selection dialog"""
        dialog = TemplateDialog(self.template_manager, 'quotes', self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            template = dialog.get_selected_template()
            if template:
                self.apply_template(template)

    def apply_template(self, template):
        """Apply selected template to current quote form"""
        # Apply scope of work
        if 'scope_of_work' in template:
            self.scope_of_work_edit.setText(template['scope_of_work'])
        
        # Apply services
        if 'services' in template:
            # Clear all service checkboxes first
            self.structural_checkbox.setChecked(False)
            self.civil_checkbox.setChecked(False)
            self.electrical_checkbox.setChecked(False)
            self.mechanical_checkbox.setChecked(False)
            self.plumbing_checkbox.setChecked(False)
            self.anchor_calc_checkbox.setChecked(False)
            self.solidworks_checkbox.setChecked(False)
            self.foundation_checkbox.setChecked(False)
            self.other_checkbox.setChecked(False)
            self.others_input.clear()
            self.others_input.setVisible(False)
            
            # Check services from template
            for service in template['services']:
                if service == 'Structural':
                    self.structural_checkbox.setChecked(True)
                elif service == 'Civil':
                    self.civil_checkbox.setChecked(True)
                elif service == 'Electrical':
                    self.electrical_checkbox.setChecked(True)
                elif service == 'Mechanical':
                    self.mechanical_checkbox.setChecked(True)
                elif service == 'Plumbing Design':
                    self.plumbing_checkbox.setChecked(True)
                elif service == 'Anchor Calculations':
                    self.anchor_calc_checkbox.setChecked(True)
                elif service == 'Solid Works':
                    self.solidworks_checkbox.setChecked(True)
                elif service == 'Foundation':
                    self.foundation_checkbox.setChecked(True)
                else:
                    # Handle other services
                    self.other_checkbox.setChecked(True)
                    self.others_input.setVisible(True)
                    current_others = self.others_input.text()
                    if current_others:
                        self.others_input.setText(f"{current_others}, {service}")
                    else:
                        self.others_input.setText(service)
        
        # Apply default price range as suggested cost
        if 'default_price_range' in template:
            # Extract the lower bound from price range
            price_range = template['default_price_range']
            if '$' in price_range and '-' in price_range:
                try:
                    lower_price = price_range.split('$')[1].split('-')[0].strip()
                    if lower_price.isdigit():
                        self.engineering_costs_edit.setText(f"${lower_price}")
                except:
                    pass
        
        # Update project name if template has a name and current name is empty
        if 'name' in template and not self.project_name_edit.text().strip():
            # Generate a project name based on template
            project_name = f"{template['name']} - {self.client_combo.currentText()}"
            self.project_name_edit.setText(project_name)
        
        # Show success message
        QtWidgets.QMessageBox.information(
            self, 
            "Template Applied", 
            f"Template '{template['name']}' has been applied successfully!"
        )

    def on_client_changed(self, client_name):
        """Handle client selection change"""
        if hasattr(self, 'client_suggestion_widget'):
            self.client_suggestion_widget.update_client(client_name)
        
        # Update client intelligence
        if client_name and client_name != "-- Select Client --":
            self.client_intelligence.update_client_activity(
                client_name, 
                'quote_started', 
                {'timestamp': datetime.now(timezone.utc).isoformat()}
            )

    def toggle_client_intelligence(self):
        """Toggle client intelligence widget visibility"""
        if hasattr(self, 'client_suggestion_widget'):
            current_client = self.client_combo.currentText()
            self.client_suggestion_widget.update_client(current_client)
            self.client_suggestion_widget.setVisible(not self.client_suggestion_widget.isVisible())

    def add_enhanced_action_buttons(self, row, job_data):
        """Add compact action buttons: View, PDF, and ⋯ dropdown (Project / Edit / Copy / Email / Delete)."""
        actions_widget = QtWidgets.QWidget()
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(6, 6, 6, 6)
        actions_layout.setSpacing(5)

        BTN_H = 32

        MENU_STYLE = """
            QMenu {
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 4px 0;
                font-family: 'Inter', 'Segoe UI'; font-size: 13px;
            }
            QMenu::item { padding: 9px 20px; color: #1e293b; background: #ffffff; }
            QMenu::item:selected { background: #dbeafe; color: #1d4ed8; }
            QMenu::item:disabled { color: #94a3b8; }
            QMenu::separator { height: 1px; background: #e2e8f0; margin: 4px 8px; }
        """

        def _mk_btn(label, bg, fg, hover_bg, border=None):
            b = QtWidgets.QPushButton(label)
            b.setFixedSize(52, BTN_H)
            b.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            border_css = f"border:1px solid {border};" if border else "border:none;"
            b.setStyleSheet(f"""
                QPushButton {{
                    background:{bg}; color:{fg};
                    {border_css}
                    border-radius:7px;
                    font-size:12px; font-weight:800;
                    font-family:'Inter','Segoe UI',sans-serif;
                }}
                QPushButton:hover {{ background:{hover_bg}; }}
            """)
            return b

        # ── View ──
        view_btn = _mk_btn("View", "#eff6ff", "#2563eb", "#dbeafe", "#bfdbfe")
        view_btn.setToolTip("View Details")
        view_btn.clicked.connect(lambda: self.view_job_details(job_data))

        # ── PDF ──
        pdf_btn = _mk_btn("PDF", "#f0fdf4", "#0f766e", "#dcfce7", "#bbf7d0")
        pdf_btn.setToolTip("Open Generated PDF")
        pdf_btn.clicked.connect(lambda: self.open_job_form_pdf(job_data))

        # ── ⋯ dropdown ──
        dots_btn = QtWidgets.QPushButton("⋯")
        dots_btn.setFixedSize(36, BTN_H)
        dots_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        dots_btn.setToolTip("More Actions")
        dots_btn.setStyleSheet("""
            QPushButton {
                background:#f1f5f9; color:#475569; border:1px solid #e2e8f0;
                border-radius:7px; font-size:16px; font-weight:900;
            }
            QPushButton:hover { background:#e2e8f0; color:#0f172a; }
        """)

        main_window = getattr(self, 'main_window', None)
        is_sales = (main_window and hasattr(main_window, 'current_role')
                    and main_window.current_role == "sales")

        def _on_invoice_click(jd=job_data):
            try:
                enhanced_job_data = self._prepare_enhanced_project_data(jd)
                self.convert_to_project.emit(enhanced_job_data)
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Conversion Error",
                    f"Failed to convert quote to project:\n{str(e)}"
                )

        def _show_dots_menu(checked=False, b=dots_btn, jd=job_data):
            menu = QtWidgets.QMenu(b)
            menu.setStyleSheet(MENU_STYLE)

            proj_action = QtWidgets.QAction("📋  Create Project", menu)
            proj_action.triggered.connect(lambda: _on_invoice_click(jd))
            if is_sales:
                proj_action.setEnabled(False)
                proj_action.setText("📋  Create Project  (admin only)")
            menu.addAction(proj_action)

            menu.addSeparator()

            edit_action = QtWidgets.QAction("✎  Edit Quote", menu)
            edit_action.triggered.connect(lambda: self.edit_job_form(jd))
            menu.addAction(edit_action)

            copy_action = QtWidgets.QAction("⎘  Copy Quote Number", menu)
            copy_action.triggered.connect(lambda: self.copy_job_number(jd))
            menu.addAction(copy_action)

            menu.addSeparator()

            del_action = QtWidgets.QAction("🗑  Delete", menu)
            del_action.triggered.connect(lambda: self.delete_job_form(jd))
            menu.addAction(del_action)

            menu.exec_(b.mapToGlobal(QtCore.QPoint(0, b.height())))

        dots_btn.clicked.connect(_show_dots_menu)

        actions_layout.addStretch()
        actions_layout.addWidget(view_btn)
        actions_layout.addWidget(pdf_btn)
        actions_layout.addWidget(dots_btn)
        self.job_forms_table.setCellWidget(row, 7, actions_widget)
        
    def copy_job_number(self, job_data):
        """Copy quote number to clipboard"""
        job_number = job_data.get('job_number', '')
        QtWidgets.QApplication.clipboard().setText(job_number)
        if hasattr(self.main_window, 'statusBar'):
            self.main_window.statusBar().showMessage(
                f"Copied: {job_number}", 2500
            )

    def email_quote_to_client(self, job_data: dict):
        """Email the quote PDF to the client using EmailManager."""
        try:
            from email_manager import EmailManager
        except ImportError:
            QtWidgets.QMessageBox.critical(self, "Error", "email_manager module not found.")
            return

        if not EmailManager.is_configured():
            QtWidgets.QMessageBox.warning(
                self, "Email Not Configured",
                "SMTP settings are not set up.\n\nGo to Settings → Email and fill in your SMTP details."
            )
            return

        client_email = job_data.get('client_email', '').strip()
        if not client_email:
            QtWidgets.QMessageBox.warning(
                self, "No Client Email",
                f"Quote {job_data.get('job_number', '')} has no client email.\n"
                "Edit the quote to add one."
            )
            return

        confirm = QtWidgets.QMessageBox.question(
            self, "Send Quote",
            f"Email quote {job_data.get('job_number', '')} to {client_email}?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        if confirm != QtWidgets.QMessageBox.Yes:
            return

        import tempfile
        job_number = job_data.get('job_number', '')
        temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_pdf_path = temp_dir / f"{job_number}_job_form.pdf"

        pdf_candidates = [
            temp_pdf_path,
            Path(__file__).resolve().parent / "data" / f"{job_number}_job_form.pdf",
        ]
        pdf_path = next((p for p in pdf_candidates if p.exists()), None)

        # Download from Firebase if not cached locally
        if pdf_path is None and self.FIREBASE_AVAILABLE:
            try:
                from main import FirebaseManager
                pdf_path = FirebaseManager.load_job_pdf_from_firebase(job_number, temp_pdf_path)
            except Exception as e:
                _log.warning("Could not download PDF for email: %s", e)

        try:
            from email_manager import _load_company_name
            import smtplib, ssl
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.base import MIMEBase
            from email import encoders
            import json

            cfg_path = Path(__file__).resolve().parent / "data" / "settings.json"
            with open(cfg_path, encoding="utf-8") as f:
                settings = json.load(f)
            cfg = settings.get("email", {})
            company = settings.get("company", {}).get("name", "MABS Engineering LLC")

            subject = f"Quote {job_number} from {company}"
            body = f"""<html><body style="font-family:Arial,sans-serif;color:#2c3e50;">
            <p>Dear {job_data.get('client', 'Valued Client')},</p>
            <p>Please find attached quote <b>{job_number}</b> for <b>{job_data.get('project_name', '')}</b>.</p>
            <p><b>Scope:</b> {job_data.get('scope_of_work', 'See attached PDF')}</p>
            <p><b>Agreed Cost:</b> ${job_data.get('engineering_costs', 'N/A')}</p>
            <p><b>Due Date:</b> {job_data.get('due_date', 'N/A')}</p>
            <p>Please don't hesitate to reach out with any questions.</p>
            <br><p>Best regards,<br><b>{company}</b></p>
            </body></html>"""

            ok = EmailManager.send_email(client_email, subject, body, pdf_path)
            if ok:
                QtWidgets.QMessageBox.information(self, "Sent", f"Quote {job_number} sent to {client_email}.")
            else:
                QtWidgets.QMessageBox.critical(self, "Send Failed",
                    "Email could not be delivered. Check SMTP settings.\nSee logs/pims.log for details.")
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Error", str(exc))

    def update_job_status_in_firebase(self, job_data, new_status):
        """Update job status in Firebase"""
        try:
            from main import db
            if 'firebase_id' in job_data:
                ref = db.reference(f'/job_forms/{job_data["firebase_id"]}')
                ref.update({
                    'status': new_status,
                    'updated_at': datetime.now(timezone.utc).isoformat()
                })
                _log.info("Status updated in Firebase: %s -> %s", job_data['job_number'], new_status)
                return True
        except Exception as e:
            _log.warning("Error updating status in Firebase: %s", e)
        return False

    def view_job_details(self, job_data):
        """View job details"""
        dialog = JobDetailsDialog(job_data, self)
        dialog.exec_()

    def update_job_forms_table(self):
        """Update the quote forms table with created forms"""
        self.display_filtered_forms(self.job_forms)
        self.update_stats()
        
        # Force UI refresh
        if hasattr(self, 'job_forms_table'):
            self.job_forms_table.repaint()
        self.update()
        QtWidgets.QApplication.processEvents()

        # Populate sales filter options
        # Populate client filter options
        client_names = sorted({job.get('client', '') for job in self.job_forms if job.get('client')})
        self.client_filter_menu.clear()
        
        # Add "All Clients" option only once
        all_action = self.client_filter_menu.addAction("All Clients")
        all_action.triggered.connect(lambda: self.apply_client_filter("All Clients"))

        # Add client names only once
        for client in client_names:
            act = self.client_filter_menu.addAction(client)
            act.triggered.connect(lambda checked, c=client: self.apply_client_filter(c))
            

    def on_status_filter_changed(self, text):
        """Handle status filter change - clear search text only, keep client and date filters"""
        self.filter_job_forms()
        
        # ✅ UPDATE CLIENT FILTER MENU
        self.update_client_filter_menu()
        
    def apply_client_filter(self, company_name):
        """Apply filter from client menu - matching project tab style"""
        self.selected_client_filter = company_name
        
        # Update the client filter button — consistent base style, teal accent when active
        _btn_base = (
            "QPushButton { background:#FFFFFF; border:1.5px solid %s;"
            " border-radius:9px; font-size:13px; font-weight:800; color:%s;"
            " padding:0 12px; }"
            "QPushButton:hover { border-color:#0F766E; color:#0F766E; }"
        )
        if company_name in ("All Clients", "📂 All Clients"):
            self.client_filter_button.setText("Client")
            self.client_filter_button.setStyleSheet(_btn_base % ("#E2E8F0", "#334155"))
            self.client_filter_button.setToolTip("Filter by client")
        else:
            clean = company_name.removeprefix("🏢 ").split(" (")[0]
            display = clean[:15] + "..." if len(clean) > 15 else clean
            self.client_filter_button.setText(f"  {display}  ×")
            self.client_filter_button.setStyleSheet(_btn_base % ("#0F766E", "#0F766E"))
            self.client_filter_button.setToolTip(f"Filtered: {clean}  (click to clear)")
            
        # ⭐⭐ CRITICAL CHANGE: DO NOT CLEAR STATUS FILTER
        # self.status_filter_combo.setCurrentText("All Status")  # REMOVE THIS LINE
        
        # DO NOT CLEAR DATE FILTER HERE - Keep date filter active if it exists
        
        # Apply filter - this will use all current filters (date, client, status)
        self.filter_job_forms()
                    
    def update_stats(self):
        """Update all statistics cards."""
        total_jobs          = len(self.job_forms)
        high_priority_count = len([j for j in self.job_forms if j.get('status') in ('High', 'Urgent')])
        completed_count     = len([j for j in self.job_forms if j.get('status') == 'Completed'])
        total_value         = sum(
            float(str(j.get('engineering_costs', '0')).replace('$', '').replace(',', '') or 0)
            for j in self.job_forms
        )
        win_rate = f"{int(completed_count / total_jobs * 100)}%" if total_jobs else "—"

        def _set(attr, text):
            lbl = getattr(self, attr, None)
            if lbl:
                lbl.setText(text)

        _set("total_jobs_label",    str(total_jobs))
        _set("high_priority_label", str(high_priority_count))
        _set("completed_label",     str(completed_count))
        _set("total_value_label",   f"${total_value:,.0f}")
        _set("win_rate_label",      win_rate)

    def get_group_box_style(self):
        return """
            QGroupBox {
                font-weight: 800;
                font-size: 16px;
                color: #1E293B;
                border: 1px solid #DCE4EC;
                border-radius: 10px;
                margin-top: 14px;
                padding-top: 18px;
                background: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 20px;
                padding: 0 10px;
                color: #1E293B;
                font-weight: 800;
            }
        """

    def filter_job_forms(self):
        """Filter quote forms based on search criteria, status filter, client filter, date range, and year/month."""
        client_filter = self.selected_client_filter
        search_text = self.search_edit.text().lower()
        status_filter = self.status_filter_combo.currentText()

        if client_filter == "📂 All Clients":
            client_filter = "All Clients"

        selected_year  = self.year_filter_combo.currentText()  if hasattr(self, 'year_filter_combo')  else "All Years"
        selected_month = self.month_filter_combo.currentText() if hasattr(self, 'month_filter_combo') else "All Months"
        _month_names = ["January","February","March","April","May","June",
                        "July","August","September","October","November","December"]

        filtered_forms = []
        for job in self.job_forms:
            matches_client = (
                client_filter == "All Clients" or
                job.get('client', '') == client_filter
            )
            matches_search = (
                not search_text or
                search_text in job.get('job_number', '').lower() or
                search_text in job.get('project_name', '').lower() or
                search_text in job.get('job_title', '').lower() or
                search_text in job.get('client', '').lower() or
                search_text in job.get('sales', '').lower() or
                search_text in job.get('scope_of_work', '').lower() or
                search_text in job.get('engineering_costs', '').lower()
            )
            job_status = job.get('status', 'Not Started')
            matches_status = (status_filter == "All Status" or job_status == status_filter)

            # Year / Month filter — based on created_at
            matches_period = True
            if selected_year != "All Years" or selected_month != "All Months":
                try:
                    created_raw = job.get('created_at', '')
                    if created_raw:
                        created_dt = datetime.fromisoformat(str(created_raw).replace("Z", ""))
                        if selected_year != "All Years" and str(created_dt.year) != selected_year:
                            matches_period = False
                        if selected_month != "All Months":
                            if created_dt.month != (_month_names.index(selected_month) + 1):
                                matches_period = False
                    else:
                        matches_period = False
                except Exception:
                    matches_period = False

            # Date range filter — based on created_at
            # Two separate mechanisms: inline checkbox filter and the 📅 Date dialog filter.
            matches_date_range = True
            _inline_active = hasattr(self, 'date_filter_active_chk') and self.date_filter_active_chk.isChecked()
            _dialog_active = (getattr(self, '_date_filter_active', False)
                              and hasattr(self, 'current_from_date')
                              and hasattr(self, 'current_to_date'))
            if _inline_active or _dialog_active:
                try:
                    created_raw = job.get('created_at', '')
                    if created_raw:
                        created_dt = datetime.fromisoformat(str(created_raw).replace("Z", ""))
                        created_qdate = QtCore.QDate(created_dt.year, created_dt.month, created_dt.day)
                        if _dialog_active:
                            matches_date_range = (
                                self.current_from_date <= created_qdate <= self.current_to_date
                            )
                        else:
                            matches_date_range = (
                                self.from_date_edit.date() <= created_qdate <= self.to_date_edit.date()
                            )
                    else:
                        matches_date_range = False
                except Exception:
                    matches_date_range = False

            if matches_client and matches_search and matches_status and matches_period and matches_date_range:
                filtered_forms.append(job)

        self.display_filtered_forms(filtered_forms)
        self.update_filtered_stats(filtered_forms)
        self.update_client_filter_menu()

        # Update period summary label
        if hasattr(self, 'period_summary_lbl'):
            if selected_year != "All Years" or selected_month != "All Months":
                total_val = sum(
                    float(str(j.get('engineering_costs', '0')).replace('$', '').replace(',', '') or 0)
                    for j in filtered_forms
                )
                period_str = " ".join(filter(lambda x: x not in ("All Years", "All Months"),
                                             [selected_month, selected_year])) or "Selected"
                self.period_summary_lbl.setText(
                    f"{len(filtered_forms)} quotes  |  ${total_val:,.0f} total  [{period_str}]")
            else:
                self.period_summary_lbl.setText("")
    
    def _clear_period_filter(self):
        """Reset year and month dropdowns to 'All'."""
        if hasattr(self, 'year_filter_combo'):
            self.year_filter_combo.setCurrentIndex(0)
        if hasattr(self, 'month_filter_combo'):
            self.month_filter_combo.setCurrentIndex(0)

    def update_filtered_stats(self, filtered_jobs):
        """Update all stats cards based on filtered quote forms."""
        total_jobs          = len(filtered_jobs)
        high_priority_count = len([j for j in filtered_jobs if j.get('status') in ('High', 'Urgent')])
        completed_count     = len([j for j in filtered_jobs if j.get('status') == 'Completed'])
        total_value         = sum(
            float(str(j.get('engineering_costs', '0')).replace('$', '').replace(',', '') or 0)
            for j in filtered_jobs
        )
        win_rate = f"{int(completed_count / total_jobs * 100)}%" if total_jobs else "—"

        def _set(attr, text):
            lbl = getattr(self, attr, None)
            if lbl:
                lbl.setText(text)

        _set("total_jobs_label",    str(total_jobs))
        _set("high_priority_label", str(high_priority_count))
        _set("completed_label",     str(completed_count))
        _set("total_value_label",   f"${total_value:,.0f}")
        _set("win_rate_label",      win_rate)
        
        
    def refresh_job_forms_immediately(self):
        """Force immediate refresh of quote forms from Firebase"""
        _log.info("Immediate quote forms refresh requested...")
        self.load_job_forms_from_firebase()
        
        # Force complete UI refresh
        QtCore.QTimer.singleShot(100, self.force_ui_refresh)

    def force_ui_refresh(self):
        """Force UI components to refresh"""
        if hasattr(self, 'job_forms_table'):
            self.job_forms_table.repaint()
        self.update()
        QtWidgets.QApplication.processEvents()

    def edit_job_form(self, job_data):
        """Edit existing quote form - PRESERVE ALL FILTERS WITHOUT SHOWING ALL JOBS"""
        # Store current filter state BEFORE opening dialog
        search_text = self.search_edit.text()
        status_filter = self.status_filter_combo.currentText()
        client_filter = self.selected_client_filter
        date_range_active = getattr(self, '_date_filter_active', False)
        
        if date_range_active and hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            from_date = self.current_from_date
            to_date = self.current_to_date
        else:
            from_date = None
            to_date = None
        
        # Store the current quote forms (already filtered)
        current_filtered_jobs = []
        for row in range(self.job_forms_table.rowCount()):
            job_number_item = self.job_forms_table.item(row, 0)
            if job_number_item:
                # Find this job in the original job_forms list
                job_number = job_number_item.text()
                for job in self.job_forms:
                    if job.get('job_number') == job_number:
                        current_filtered_jobs.append(job)
                        break
        
        dialog = JobFormDialog(self.main_window, self, job_data, firebase_available=self.FIREBASE_AVAILABLE)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Collect updated data once and reuse
            updated_data = dialog.collect_job_form_data()
            if 'firebase_id' in job_data:
                updated_data['firebase_id'] = job_data['firebase_id']

            # Update the specific job in our local list
            for i, job in enumerate(self.job_forms):
                if job.get('job_number') == job_data.get('job_number'):
                    self.job_forms[i] = updated_data
                    break

            # Update Firebase without reloading everything
            try:
                if self.FIREBASE_AVAILABLE:
                    self.save_job_form_directly(updated_data)
            except Exception as e:
                _log.warning("Error updating Firebase: %s", e)
            
            # ⭐ DIRECTLY APPLY FILTERS WITHOUT RELOADING ALL FROM FIREBASE
            # Restore ALL filter controls
            self.search_edit.setText(search_text)
            self.status_filter_combo.setCurrentText(status_filter)
            self.selected_client_filter = client_filter
            
            # Restore date filter if it was active
            if from_date and to_date:
                self.current_from_date = from_date
                self.current_to_date = to_date
                self._date_filter_active = True
                from_date_formatted = from_date.toString("MM-dd-yy")
                to_date_formatted = to_date.toString("MM-dd-yy")
                self.date_range_button.setText(f"📅 {from_date_formatted} to {to_date_formatted}")
                self.date_range_button.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        border-radius: 5px;
                        font-size: 16px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                    QPushButton:pressed {
                        background-color: #21618c;
                    }
                """)
            
            # Apply filters directly to current quote form
            self.filter_job_forms()

    def generate_pdf(self, job_data):
        """Generate PDF for quote form"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            import os
            
            # Create PDF directory if not exists
            temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            filename = temp_dir / f"{job_data['job_number']}_job_form.pdf"
            doc = SimpleDocTemplate(str(filename), pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title = Paragraph(f"<b>QUOTE FORM: {job_data['job_number']}</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Job Details Table
            data = [
                ["Quote Number:", job_data['job_number']],
                ["Job Title:", job_data['job_title']],
                ["Client:", job_data['client']],
                ["Project Site:", job_data['project_site_address']],
                ["Priority:", job_data['priority']],
                ["Status:", job_data.get('status', 'Not Started')],
                ["Start Date:", job_data['start_date']],
                ["Due Date:", job_data['due_date']],
                ["Agreed Cost:", job_data.get('engineering_costs', 'N/A')],
                ["Scope of Work:", job_data['scope_of_work']],
                ["Services:", ", ".join(job_data.get('services', []))],
                ["Deliverables:", ", ".join(job_data.get('deliverables', []))],
                ["Created:", job_data['created_at'][:10]],
            ]
            
            table = Table(data, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            QtWidgets.QMessageBox.information(
                self, "PDF Generated", 
                f"PDF has been generated successfully!\n\nLocation: {filename}"
            )
            
        except Exception as e:
            QtWidgets.QMessageBox.warning(
                self, "PDF Generation Failed", 
                f"Could not generate PDF:\n{str(e)}"
            )

    def open_job_form_pdf(self, job_data):
        """Open the generated quote form PDF - ONLY from Firebase"""
        job_number = job_data.get('job_number', '')
        
        if not job_number:
            QtWidgets.QMessageBox.warning(
                self, "PDF Not Found", 
                "No quote number found."
            )
            return
        
        try:
            # ✅ ONLY use Firebase - remove local directory check
            if self.FIREBASE_AVAILABLE:
                from main import FirebaseManager
                
                # Create temporary directory for PDF download
                temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
                temp_dir.mkdir(parents=True, exist_ok=True)
                temp_pdf_path = temp_dir / f"{job_number}_job_form.pdf"
                
                # Load PDF from Firebase
                pdf_path = FirebaseManager.load_job_pdf_from_firebase(job_number, temp_pdf_path)
                
                if pdf_path and pdf_path.exists():
                    # Open the PDF
                    if self.open_job_form_pdf_file(pdf_path):
                        _log.info("Opened quote form PDF from Firebase: %s", job_number)
                        
                        # Clean up temporary file after a delay
                        QtCore.QTimer.singleShot(10000, lambda: self.cleanup_temp_pdf_file(pdf_path))
                        return
                    else:
                        _log.warning("Failed to open PDF from Firebase: %s", job_number)
                else:
                    _log.warning("PDF not found in Firebase for job: %s", job_number)
            
            # If we get here, PDF wasn't found in Firebase
            QtWidgets.QMessageBox.warning(
                self, "PDF Not Found", 
                f"No PDF found in Firebase for job: {job_number}\n\n"
                f"Please generate the PDF first by clicking 'Generate Quote form PDF'."
            )
                    
        except Exception as e:
            _log.warning("Error opening quote form PDF from Firebase: %s", e)
            QtWidgets.QMessageBox.critical(
                self, "Error Opening PDF", 
                f"Could not open PDF from Firebase:\n{str(e)}"
            )

    def cleanup_temp_pdf_file(self, file_path: Path):
        """Clean up temporary PDF file"""
        try:
            if file_path.exists():
                file_path.unlink()
                _log.info("Cleaned up temporary PDF file: %s", file_path)
        except Exception as e:
            _log.warning("Could not clean up temporary PDF file: %s", e)
            
    def delete_job_form(self, job_data):
        """Delete quote form with confirmation - PRESERVE FILTERS WITHOUT SHOWING ALL JOBS"""
        reply = QtWidgets.QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete quote form:\n\n"
            f"Quote Number: {job_data['job_number']}\n"
            f"Project Name: {job_data.get('project_name', 'n/a')}\n\n"
            f"This action cannot be undone.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            try:
                # Store current filter state BEFORE deletion
                search_text = self.search_edit.text()
                status_filter = self.status_filter_combo.currentText()
                client_filter = self.selected_client_filter
                date_range_active = getattr(self, '_date_filter_active', False)
                
                if date_range_active and hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
                    from_date = self.current_from_date
                    to_date = self.current_to_date
                else:
                    from_date = None
                    to_date = None
                
                # Delete from Firebase first
                success = self.delete_job_form_from_firebase(job_data)
                
                if success:
                    # ⭐ CRITICAL: Remove job from local list WITHOUT reloading all from Firebase
                    # Remove the deleted job from our local job_forms list
                    self.job_forms = [job for job in self.job_forms 
                                    if job.get('job_number') != job_data.get('job_number')]
                    
                    # RESTORE FILTERS
                    self.search_edit.setText(search_text)
                    self.status_filter_combo.setCurrentText(status_filter)
                    self.selected_client_filter = client_filter
                    
                    # Restore date filter if it was active
                    if from_date and to_date:
                        self.current_from_date = from_date
                        self.current_to_date = to_date
                        from_date_formatted = from_date.toString("MM-dd-yy")
                        to_date_formatted = to_date.toString("MM-dd-yy")
                        self.date_range_button.setText(f"📅 {from_date_formatted} to {to_date_formatted}")
                        self.date_range_button.setStyleSheet("""
                            QPushButton {
                                background-color: #3498db;
                                color: white;
                                border: none;
                                border-radius: 5px;
                                font-size: 16px;
                                font-weight: bold;
                            }
                            QPushButton:hover {
                                background-color: #2980b9;
                            }
                            QPushButton:pressed {
                                background-color: #21618c;
                            }
                        """)
                    
                    # Apply filters to the updated local list
                    self.filter_job_forms()
                    
                    QtWidgets.QMessageBox.information(
                        self, "Success", 
                        f"Quote form '{job_data['job_number']}' has been deleted successfully."
                    )
                else:
                    QtWidgets.QMessageBox.critical(
                        self, "Error", 
                        f"Failed to delete quote form from Firebase."
                    )
                
            except Exception as e:
                QtWidgets.QMessageBox.critical(
                    self, "Error", 
                    f"Failed to delete quote form:\n{str(e)}"
                )

class _WheelBlocker(QtCore.QObject):
    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Wheel:
            return True
        return False

class _NoScrollComboBox(QtWidgets.QComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._wb = _WheelBlocker(self)
        self.installEventFilter(self._wb)

class _NoScrollDateEdit(QtWidgets.QDateEdit):
    def stepBy(self, steps):
        pass  # block scroll wheel and arrow-key stepping; use calendar popup to change date


class JobFormsExportDialog(QtWidgets.QDialog):
    """Professional PDF/Excel Export Dialog for Quote Forms with Tabs"""

    def __init__(self, parent=None, available_dates=None):
        super().__init__(parent)
        self._owner_tab = parent  # keep reference before any reparenting
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowContextHelpButtonHint)
        self.available_dates = available_dates or []
        self.export_range = "all"  # Default export range
        self.selected_dates = []
        self.export_type = "pdf"  # Default export type
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Export Quote Forms")
        self.setMinimumSize(1280, 700)
        self.resize(1360, 760)
        self.setStyleSheet("JobFormsExportDialog { background: white; }")

        root = QtWidgets.QVBoxLayout(self)
        root.setSpacing(0)
        root.setContentsMargins(0, 0, 0, 0)

        # ── Teal header bar ─────────────────────────────────
        # ── Body ─────────────────────────────────────────────
        header_bar = QtWidgets.QFrame()
        header_bar.setFixedHeight(44)
        header_bar.setStyleSheet(
            "QFrame { background: #0F766E; border-bottom: 1px solid #0B5F59; }"
        )
        hb = QtWidgets.QHBoxLayout(header_bar)
        hb.setContentsMargins(18, 0, 18, 0)
        hb.setSpacing(10)
        title_icon = QtWidgets.QLabel("▣")
        title_icon.setStyleSheet(
            "font-size: 13px; font-weight: 900; color: #E6FFFB;"
            " background: transparent; border: none;"
        )
        title_lbl = QtWidgets.QLabel("Export Manager - Quote Forms")
        title_lbl.setStyleSheet(
            "font-size: 13px; font-weight: 800; color: white;"
            " background: transparent; border: none;"
        )
        hb.addWidget(title_icon)
        hb.addWidget(title_lbl)
        hb.addStretch()
        header_bar.setVisible(False)
        root.addWidget(header_bar)

        body = QtWidgets.QWidget()
        body.setStyleSheet("QWidget { background: #F8FAFC; }")
        body_layout = QtWidgets.QVBoxLayout(body)
        body_layout.setContentsMargins(20, 14, 20, 14)
        body_layout.setSpacing(10)

        body_layout.addWidget(self._build_export_intro())
        body_layout.addWidget(self._build_format_selector())

        # PDF / Excel sub-window tabs
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #E2E8F0;
                border-radius: 0px;
                background: white;
                top: -1px;
            }
            QTabBar::tab {
                background: #F1F5F9;
                color: #64748B;
                border: 1.5px solid #E2E8F0;
                border-bottom: none;
                padding: 10px 26px;
                margin-right: 4px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                font-family: 'Segoe UI Emoji', 'Segoe UI', 'Inter', sans-serif;
            }
            QTabBar::tab:selected {
                background: white;
                color: #0F766E;
                border-bottom: 2px solid white;
            }
            QTabBar::tab:hover:!selected {
                background: #E2E8F0;
                color: #334155;
            }
        """)

        self.pdf_tab = QtWidgets.QWidget()
        self.pdf_tab.setStyleSheet("QWidget { background: white; }")
        self.setup_pdf_tab()
        self.tab_widget.addTab(self.pdf_tab, "📄  PDF Export")

        self.excel_tab = QtWidgets.QWidget()
        self.excel_tab.setStyleSheet("QWidget { background: white; }")
        self.setup_excel_tab()
        self.tab_widget.addTab(self.excel_tab, "📊  Excel Export")
        self.tab_widget.tabBar().hide()
        self.tab_widget.setCurrentIndex(0)
        self.export_type = "pdf"
        self._sync_format_card_state(0)

        body_layout.addWidget(self.tab_widget, 1)
        self.tab_widget.currentChanged.connect(self.on_tab_changed)

        # Progress bar (thin, hidden by default)
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setFixedHeight(5)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar { background: #E2E8F0; border: none; border-radius: 3px; }
            QProgressBar::chunk { background: #0F766E; border-radius: 3px; }
        """)
        body_layout.addWidget(self.progress_bar)

        # ── Back to Quotes button row ─────────────────────────────
        self.cancel_btn = QtWidgets.QPushButton("Back to Quotes")
        self.cancel_btn.setVisible(False)
        self.cancel_btn.setFixedSize(130, 32)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: white;
                color: #475569;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                font-size: 11px;
                font-weight: 700;
                padding: 0 14px;
            }
            QPushButton:hover { background: #F8FAFC; border-color: #94A3B8; }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        back_row = QtWidgets.QHBoxLayout()
        back_row.setContentsMargins(0, 4, 0, 0)
        back_row.addWidget(self.cancel_btn)
        back_row.addStretch()
        body_layout.addLayout(back_row)

        root.addWidget(body, 1)

    # ── shared helpers used by both tab builders ─────────────
    @staticmethod
    def _field_lbl(text):
        lbl = QtWidgets.QLabel(text)
        lbl.setStyleSheet(
            "font-size: 11px; font-weight: 800; color: #64748B;"
            " letter-spacing: 0.5px; background: transparent; border: none;")
        return lbl

    def _build_export_intro(self):
        frame = QtWidgets.QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
        """)
        row = QtWidgets.QHBoxLayout(frame)
        row.setContentsMargins(18, 14, 18, 14)
        row.setSpacing(12)

        text_col = QtWidgets.QVBoxLayout()
        text_col.setSpacing(3)
        title = QtWidgets.QLabel("Export Quote Forms")
        title.setStyleSheet(
            "font-size: 18px; font-weight: 900; color: #0F172A;"
            " background: transparent; border: none;"
        )
        subtitle = QtWidgets.QLabel(
            "Export your quote forms data in PDF or Excel format with flexible date range and filters."
        )
        subtitle.setStyleSheet(
            "font-size: 12px; font-weight: 600; color: #64748B;"
            " background: transparent; border: none;"
        )
        text_col.addWidget(title)
        text_col.addWidget(subtitle)

        help_btn = QtWidgets.QPushButton("?  How export works?")
        help_btn.setFixedSize(150, 34)
        help_btn.setEnabled(False)
        help_btn.setStyleSheet("""
            QPushButton {
                background: white;
                color: #64748B;
                border: 1px solid #DCE4EC;
                border-radius: 8px;
                font-size: 12px;
                font-weight: 800;
            }
        """)

        row.addLayout(text_col, 1)
        row.addWidget(help_btn)
        return frame

    def _build_format_selector(self):
        frame = QtWidgets.QFrame()
        frame.setMinimumHeight(166)
        frame.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Minimum,
        )
        frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
        """)
        layout = QtWidgets.QVBoxLayout(frame)
        layout.setContentsMargins(18, 14, 18, 16)
        layout.setSpacing(10)

        title = QtWidgets.QLabel("1   Select Export Format")
        title.setStyleSheet(
            "font-size: 13px; font-weight: 900; color: #0F172A;"
            " background: transparent; border: none;"
        )
        layout.addWidget(title)

        row = QtWidgets.QHBoxLayout()
        row.setSpacing(14)
        row.setContentsMargins(0, 0, 0, 0)
        self.pdf_format_btn, self.pdf_format_indicator = self._make_format_card(
            "PDF Export", "Best for printing and sharing", "pdf", 0
        )
        self.excel_format_btn, self.excel_format_indicator = self._make_format_card(
            "Excel Export", "Best for data analysis", "excel", 1
        )
        row.addWidget(self.pdf_format_btn, 1)
        row.addWidget(self.excel_format_btn, 1)
        row_holder = QtWidgets.QWidget()
        row_holder.setMinimumHeight(110)
        row_holder.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed,
        )
        row_holder.setLayout(row)
        layout.addWidget(row_holder)
        return frame

    @staticmethod
    def _format_card_qss(selected):
        border = "#0F766E" if selected else "#E2E8F0"
        bg = "#F0FDFA" if selected else "#FFFFFF"
        fg = "#0F172A"
        return f"""
            QPushButton {{
                background: {bg};
                color: {fg};
                border: 1.5px solid {border};
                border-radius: 10px;
                font-size: 13px;
                font-weight: 800;
                text-align: left;
                padding: 0px;
            }}
            QPushButton:hover {{
                border-color: #0F766E;
                background: #F8FFFD;
            }}
        """

    def _make_format_card(self, title, subtitle, kind, index):
        btn = QtWidgets.QPushButton()
        btn.setCheckable(True)
        btn.setFixedHeight(110)
        btn.setSizePolicy(
            QtWidgets.QSizePolicy.Expanding,
            QtWidgets.QSizePolicy.Fixed,
        )
        btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        btn.setStyleSheet(self._format_card_qss(False))
        btn.clicked.connect(lambda: self._select_export_format(index))

        content = QtWidgets.QHBoxLayout(btn)
        content.setContentsMargins(22, 18, 18, 18)
        content.setSpacing(18)

        icon = QtWidgets.QLabel()
        icon.setFixedSize(58, 58)
        icon.setPixmap(self._make_export_format_icon(kind).pixmap(58, 58))

        text_col = QtWidgets.QVBoxLayout()
        text_col.setSpacing(5)
        title_lbl = QtWidgets.QLabel(title)
        title_lbl.setStyleSheet(
            "font-size: 14px; font-weight: 900; color: #0F172A;"
            " border: none; background: transparent;"
        )
        subtitle_lbl = QtWidgets.QLabel(subtitle)
        subtitle_lbl.setStyleSheet(
            "font-size: 12px; font-weight: 600; color: #64748B;"
            " border: none; background: transparent;"
        )
        text_col.addWidget(title_lbl)
        text_col.addWidget(subtitle_lbl)

        indicator = QtWidgets.QLabel()
        indicator.setFixedSize(18, 18)
        indicator.setAlignment(QtCore.Qt.AlignCenter)
        indicator.setStyleSheet("""
            QLabel {
                background: white;
                border: 1.5px solid #CBD5E1;
                border-radius: 9px;
                color: transparent;
                font-size: 11px;
                font-weight: 900;
            }
        """)

        content.addWidget(icon)
        content.addLayout(text_col, 1)
        content.addWidget(indicator)
        return btn, indicator

    def _select_export_format(self, index):
        self.tab_widget.setCurrentIndex(index)
        self._sync_format_card_state(index)

    def _sync_format_card_state(self, index):
        self.pdf_format_btn.setChecked(index == 0)
        self.excel_format_btn.setChecked(index == 1)
        self.pdf_format_btn.setStyleSheet(self._format_card_qss(index == 0))
        self.excel_format_btn.setStyleSheet(self._format_card_qss(index == 1))
        self._set_format_indicator(self.pdf_format_indicator, index == 0)
        self._set_format_indicator(self.excel_format_indicator, index == 1)

    @staticmethod
    def _set_format_indicator(label, selected):
        if selected:
            label.setText("✓")
            label.setStyleSheet("""
                QLabel {
                background: #0F766E;
                border: 1.5px solid #0F766E;
                border-radius: 8px;
                    color: white;
                    font-size: 11px;
                    font-weight: 900;
                }
            """)
        else:
            label.setText("")
            label.setStyleSheet("""
                QLabel {
                background: white;
                border: 1.5px solid #CBD5E1;
                border-radius: 8px;
                    color: transparent;
                    font-size: 11px;
                    font-weight: 900;
                }
            """)

    @staticmethod
    def _make_export_format_icon(kind):
        pixmap = QtGui.QPixmap(58, 58)
        pixmap.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pixmap)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)

        if kind == "pdf":
            bg = QtGui.QColor("#FEE2E2")
            fg = QtGui.QColor("#DC2626")
            text = "PDF"
        else:
            bg = QtGui.QColor("#DCFCE7")
            fg = QtGui.QColor("#15803D")
            text = "X"

        painter.setPen(QtCore.Qt.NoPen)
        painter.setBrush(bg)
        painter.drawRoundedRect(QtCore.QRectF(1, 1, 56, 56), 12, 12)

        painter.setBrush(fg)
        painter.drawRoundedRect(QtCore.QRectF(16, 13, 26, 32), 5, 5)

        painter.setPen(QtGui.QPen(QtGui.QColor("white")))
        font = QtGui.QFont("Inter", 9 if kind == "pdf" else 14, QtGui.QFont.Bold)
        painter.setFont(font)
        painter.drawText(QtCore.QRectF(16, 13, 26, 32), QtCore.Qt.AlignCenter, text)
        painter.end()
        return QtGui.QIcon(pixmap)

    _CARD = """
        QGroupBox {
            background: #F8FAFC;
            border: 1.5px solid #E2E8F0;
            border-radius: 10px;
            margin-top: 6px;
            padding-top: 6px;
            font-size: 12px;
            font-weight: 800;
            color: #334155;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 12px;
            padding: 0 6px;
            color: #0F766E;
        }
    """
    _DATE = """
        QDateEdit {
            padding: 5px 8px;
            border: 1.5px solid #CBD5E1;
            border-radius: 7px;
            font-size: 13px;
            font-weight: 600;
            color: #1E293B;
            background: white;
        }
        QDateEdit:focus { border-color: #0F766E; }
        QDateEdit:hover { border-color: #94A3B8; }
    """
    _YEAR_BTN = """
        QPushButton {
            background: #0F766E; color: white;
            border: none; border-radius: 7px; font-size: 15px;
            font-family: 'Segoe UI Emoji', 'Segoe UI', sans-serif;
        }
        QPushButton:hover { background: #115E59; }
        QPushButton:pressed { background: #0C4A45; }
    """
    _YEAR_FIELD = """
        QLineEdit {
            padding: 5px 10px;
            border: 1.5px solid #CBD5E1;
            border-radius: 7px;
            font-size: 13px; font-weight: 700;
            color: #1E293B; background: white;
        }
    """
    _PILL = """
        QRadioButton {
            background: #F8FAFC;
            border: 1.5px solid #CBD5E1;
            border-radius: 8px;
            padding: 10px 24px;
            min-height: 20px;
            min-width: 108px;
            font-size: 13px; font-weight: 800;
            color: #334155;
            spacing: 0px;
        }
        QRadioButton:checked {
            background: #ECFDF5;
            border: 2px solid #0F766E;
            color: #0F766E;
        }
        QRadioButton:hover:!checked {
            background: white;
            border-color: #94A3B8;
            color: #1E293B;
        }
        QRadioButton::indicator {
            width: 0px;
            height: 0px;
            margin: 0px;
            padding: 0px;
            border: none;
            background: transparent;
        }
    """

    def _build_range_row(self, all_r, dr_r, month_r, year_r, rg):
        """Build a pill-button row for range selection and add to a QButtonGroup."""
        for btn in (all_r, dr_r, month_r, year_r):
            btn.setStyleSheet(self._PILL)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            rg.addButton(btn)
        row = QtWidgets.QHBoxLayout()
        row.setSpacing(10)
        for btn in (all_r, dr_r, month_r, year_r):
            row.addWidget(btn)
        row.addStretch()
        return row

    def _build_year_row(self, year_edit, cal_btn):
        """Return an HBoxLayout containing a year line-edit + calendar button."""
        row = QtWidgets.QHBoxLayout()
        row.setSpacing(6)
        year_edit.setFixedSize(110, 34)
        year_edit.setStyleSheet(self._YEAR_FIELD)
        cal_btn.setFixedSize(34, 34)
        cal_btn.setStyleSheet(self._YEAR_BTN)
        cal_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        row.addWidget(year_edit)
        row.addWidget(cal_btn)
        return row

    def _build_preview_bar(self, label_widget, accent="#0F766E"):
        bar = QtWidgets.QFrame()
        bar.setStyleSheet(
            f"QFrame {{ background: #ECFDF5; border: 1.5px solid #6EE7B7;"
            f" border-radius: 8px; }}")
        bl = QtWidgets.QHBoxLayout(bar)
        bl.setContentsMargins(12, 7, 12, 7)
        bl.setSpacing(8)
        icon = QtWidgets.QLabel("👁")
        icon.setStyleSheet("font-size: 15px; background: transparent; border: none;")
        label_widget.setStyleSheet(
            "font-size: 12px; font-weight: 700; color: #065F46;"
            " background: transparent; border: none;")
        label_widget.setWordWrap(True)
        bl.addWidget(icon)
        bl.addWidget(label_widget, 1)
        return bar

    # ── Export history helpers ────────────────────────────────────
    def _export_history_path(self):
        return Path(__file__).resolve().parent / "data" / "export_history.json"

    def _load_export_history(self):
        try:
            p = self._export_history_path()
            if p.exists():
                with open(p, encoding="utf-8") as f:
                    data = json.load(f)
                return data if isinstance(data, list) else []
        except Exception:
            pass
        return []

    def _save_export_entry(self, export_type, scope, records, file_path=""):
        try:
            history = self._load_export_history()
            try:
                mw = getattr(self._owner_tab, 'main_window', None)
                username = getattr(mw, 'current_username', '') or getattr(mw, 'current_user', '') or 'You'
                initials = ''.join(w[0].upper() for w in str(username).split() if w)[:2] or 'ME'
            except Exception:
                username, initials = 'You', 'ME'
            history.insert(0, {
                "type": export_type,
                "scope": scope,
                "records": records,
                "exported_by_initials": initials,
                "exported_by_name": "You",
                "datetime": datetime.now().isoformat(),
                "status": "Completed",
                "file_path": str(file_path),
            })
            history = history[:10]
            p = self._export_history_path()
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "w", encoding="utf-8") as f:
                json.dump(history, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            _log.warning("Could not save export history: %s", exc)

    def _get_scope_label(self, tab="excel"):
        if tab == "excel":
            r = getattr(self, 'excel_export_range', 'all')
            if r == "all":
                return "All Quote Forms"
            elif r == "date_range":
                fd = self.excel_from_date.date().toString("MMM d, yyyy")
                td = self.excel_to_date.date().toString("MMM d, yyyy")
                return f"Date Range ({fd} - {td})"
            elif r == "month":
                return f"{self.excel_month_combo.currentText()} {self.excel_year_edit_month.text()}"
            elif r == "year":
                return self.excel_year_edit.text()
        else:
            r = getattr(self, 'export_range', 'all')
            if r == "all":
                return "All Quote Forms"
            elif r == "date_range":
                fd = self.from_date.date().toString("MMM d, yyyy")
                td = self.to_date.date().toString("MMM d, yyyy")
                return f"Date Range ({fd} - {td})"
            elif r == "month":
                return f"{self.month_combo.currentText()} {self.year_edit_month.text()}"
            elif r == "year":
                return self.year_edit.text()
        return "All Quote Forms"

    def _get_total_count(self):
        return len(getattr(self._owner_tab, 'job_forms', None) or [])

    def _build_status_bar(self, label_widget):
        bar = QtWidgets.QFrame()
        bar.setStyleSheet(
            "QFrame { background: #ECFDF5; border: 1.5px solid #6EE7B7; border-radius: 8px; }")
        bl = QtWidgets.QHBoxLayout(bar)
        bl.setContentsMargins(12, 8, 12, 8)
        bl.setSpacing(8)
        icon = QtWidgets.QLabel("✓")
        icon.setStyleSheet(
            "font-size: 16px; font-weight: 900; color: #059669;"
            " background: transparent; border: none;")
        label_widget.setStyleSheet(
            "font-size: 13px; font-weight: 700; color: #065F46;"
            " background: transparent; border: none;")
        label_widget.setWordWrap(True)
        bl.addWidget(icon)
        bl.addWidget(label_widget, 1)
        return bar

    def _build_export_summary_card(self, export_type="Excel"):
        card = QtWidgets.QFrame()
        card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
        """)
        card_layout = QtWidgets.QVBoxLayout(card)
        card_layout.setContentsMargins(34, 22, 34, 24)
        card_layout.setSpacing(18)

        summary_title = QtWidgets.QLabel("Export Summary")
        summary_title.setStyleSheet(
            "font-size: 24px; font-weight: 900; color: #111827;"
            " background: transparent; border: none;")
        card_layout.addWidget(summary_title)

        row_wrap = QtWidgets.QWidget()
        row_wrap.setStyleSheet("background: transparent; border: none;")
        row = QtWidgets.QHBoxLayout(row_wrap)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(0)

        # Type icon
        icon_lbl = QtWidgets.QLabel("X\nLS" if export_type == "Excel" else "PDF")
        if export_type == "Excel":
            icon_lbl.setText("XLS")
            icon_lbl.setStyleSheet("""
                QLabel {
                    background: #1E6B3E; color: white;
                    font-size: 17px; font-weight: 900;
                    border-radius: 12px; padding: 14px 10px;
                    min-width: 72px; max-width: 72px;
                    min-height: 72px; max-height: 72px;
                }
            """)
        else:
            icon_lbl.setStyleSheet("""
                QLabel {
                    background: #DC2626; color: white;
                    font-size: 16px; font-weight: 900;
                    border-radius: 12px; padding: 14px 10px;
                    min-width: 72px; max-width: 72px;
                    min-height: 72px; max-height: 72px;
                }
            """)
        icon_lbl.setAlignment(QtCore.Qt.AlignCenter)
        row.addWidget(icon_lbl)
        row.addSpacing(28)

        def _vdiv():
            d = QtWidgets.QFrame()
            d.setFrameShape(QtWidgets.QFrame.VLine)
            d.setFixedHeight(66)
            d.setStyleSheet("color: #E2E8F0; background: #E2E8F0; min-width: 1px; max-width: 1px;")
            return d

        def _stat_col(title_text, value_text, value_color="#0F766E"):
            col = QtWidgets.QVBoxLayout()
            col.setSpacing(3)
            t = QtWidgets.QLabel(title_text)
            t.setStyleSheet("font-size: 13px; color: #94A3B8; font-weight: 700;"
                           " background: transparent; border: none;")
            v = QtWidgets.QLabel(value_text)
            v.setStyleSheet(f"font-size: 18px; color: {value_color}; font-weight: 900;"
                           " background: transparent; border: none;")
            col.addWidget(t)
            col.addWidget(v)
            return col, v

        # Export Type
        c, _ = _stat_col("Export Type", export_type)
        row.addLayout(c)
        row.addSpacing(42)
        row.addWidget(_vdiv())
        row.addSpacing(42)

        # Export Scope
        c, scope_v = _stat_col("Export Scope", "All Quote Forms")
        row.addLayout(c)
        row.addSpacing(42)
        row.addWidget(_vdiv())
        row.addSpacing(42)

        # Records to Export
        total = self._get_total_count()
        c, rec_v = _stat_col("Records to Export", f"{total} Quotes")
        row.addLayout(c)
        row.addSpacing(42)
        row.addWidget(_vdiv())
        row.addSpacing(42)

        # Last Exported
        history = self._load_export_history()
        last_time = "—"
        for entry in history:
            if entry.get("type", "").lower() == export_type.lower():
                try:
                    dt = datetime.fromisoformat(entry["datetime"])
                    now = datetime.now()
                    if dt.date() == now.date():
                        last_time = f"Today, {dt.strftime('%I:%M %p')}"
                    else:
                        last_time = dt.strftime("%b %d, %I:%M %p")
                except Exception:
                    pass
                break
        c, last_v = _stat_col("Last Exported", last_time, "#1E293B")
        row.addLayout(c)
        row.addStretch()
        card_layout.addWidget(row_wrap)

        if export_type == "Excel":
            self._excel_scope_lbl = scope_v
            self._excel_rec_lbl = rec_v
            self._excel_last_lbl = last_v
        else:
            self._pdf_scope_lbl = scope_v
            self._pdf_rec_lbl = rec_v
            self._pdf_last_lbl = last_v

        return card

    def _build_recent_exports_widget(self, export_type="Excel"):
        container = QtWidgets.QFrame()
        container.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
        """)
        v = QtWidgets.QVBoxLayout(container)
        v.setContentsMargins(16, 12, 16, 12)
        v.setSpacing(6)

        title = QtWidgets.QLabel("Recent Exports")
        title.setStyleSheet("font-size: 14px; font-weight: 800; color: #1E293B;"
                           " background: transparent; border: none;")
        v.addWidget(title)

        # Header row
        hdr = QtWidgets.QWidget()
        hdr.setStyleSheet("background: #F8FAFC; border-radius: 6px;")
        hr = QtWidgets.QHBoxLayout(hdr)
        hr.setContentsMargins(8, 5, 8, 5)
        hr.setSpacing(0)
        for txt, stretch in [("Export Type", 2), ("Scope", 4), ("Records", 2),
                              ("Exported By", 2), ("Date & Time", 3), ("Status", 2), ("", 1)]:
            lbl = QtWidgets.QLabel(txt)
            lbl.setStyleSheet("font-size: 11px; font-weight: 800; color: #94A3B8;"
                             " background: transparent; border: none;")
            hr.addWidget(lbl, stretch)
        v.addWidget(hdr)

        history = self._load_export_history()
        relevant = history[:5]

        if not relevant:
            empty = QtWidgets.QLabel("No recent exports yet.")
            empty.setStyleSheet("color: #94A3B8; font-size: 13px; padding: 12px;"
                               " background: transparent; border: none;")
            empty.setAlignment(QtCore.Qt.AlignCenter)
            v.addWidget(empty)
        else:
            for i, entry in enumerate(relevant):
                row_w = QtWidgets.QWidget()
                row_w.setStyleSheet("background: transparent;")
                rh = QtWidgets.QHBoxLayout(row_w)
                rh.setContentsMargins(8, 6, 8, 6)
                rh.setSpacing(0)

                t = entry.get("type", "Excel")
                icon_txt = "📊" if t == "Excel" else "📄"
                type_lbl = QtWidgets.QLabel(f"{icon_txt}  {t}")
                type_lbl.setStyleSheet("font-size: 13px; color: #1E293B;"
                                      " background: transparent; border: none;"
                                      " font-family: 'Segoe UI Emoji', 'Segoe UI', sans-serif;")
                rh.addWidget(type_lbl, 2)

                scope_lbl = QtWidgets.QLabel(entry.get("scope", "All Quote Forms"))
                scope_lbl.setStyleSheet("font-size: 12px; color: #475569;"
                                       " background: transparent; border: none;")
                scope_lbl.setWordWrap(True)
                rh.addWidget(scope_lbl, 4)

                rec_lbl = QtWidgets.QLabel(f"{entry.get('records', 0)} Quotes")
                rec_lbl.setStyleSheet("font-size: 12px; color: #475569;"
                                     " background: transparent; border: none;")
                rh.addWidget(rec_lbl, 2)

                initials = entry.get("exported_by_initials", "ME")
                by_w = QtWidgets.QWidget()
                by_w.setStyleSheet("background: transparent;")
                by_l = QtWidgets.QHBoxLayout(by_w)
                by_l.setContentsMargins(0, 0, 0, 0)
                by_l.setSpacing(6)
                avatar = QtWidgets.QLabel(initials)
                avatar.setFixedSize(26, 26)
                avatar.setAlignment(QtCore.Qt.AlignCenter)
                avatar.setStyleSheet("background: #0F766E; color: white; border-radius: 13px;"
                                    " font-size: 10px; font-weight: 800;")
                name_lbl = QtWidgets.QLabel(entry.get("exported_by_name", "You"))
                name_lbl.setStyleSheet("font-size: 12px; color: #475569;"
                                      " background: transparent; border: none;")
                by_l.addWidget(avatar)
                by_l.addWidget(name_lbl)
                by_l.addStretch()
                rh.addWidget(by_w, 2)

                try:
                    dt = datetime.fromisoformat(entry["datetime"])
                    now = datetime.now()
                    if dt.date() == now.date():
                        dt_str = f"Today, {dt.strftime('%I:%M %p')}"
                    else:
                        dt_str = dt.strftime("%b %d, %Y %I:%M %p")
                except Exception:
                    dt_str = entry.get("datetime", "—")
                dt_lbl = QtWidgets.QLabel(dt_str)
                dt_lbl.setStyleSheet("font-size: 12px; color: #475569;"
                                    " background: transparent; border: none;")
                rh.addWidget(dt_lbl, 3)

                status_lbl = QtWidgets.QLabel("✓  Completed")
                status_lbl.setStyleSheet("""
                    QLabel {
                        background: #DCFCE7; color: #15803D;
                        border-radius: 10px; padding: 2px 10px;
                        font-size: 11px; font-weight: 700;
                    }
                """)
                rh.addWidget(status_lbl, 2)

                dl_btn = QtWidgets.QPushButton("↓")
                dl_btn.setFixedSize(26, 26)
                dl_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                dl_btn.setStyleSheet("""
                    QPushButton {
                        background: transparent; color: #2563EB;
                        border: none; border-radius: 13px;
                        font-size: 16px; font-weight: 800;
                    }
                    QPushButton:hover { background: #EFF6FF; }
                    QPushButton:disabled { color: #CBD5E1; }
                """)
                fp = entry.get("file_path", "")
                if fp:
                    dl_btn.clicked.connect(lambda checked=False, p=fp: self._open_file(p))
                else:
                    dl_btn.setEnabled(False)
                rh.addWidget(dl_btn, 1)

                v.addWidget(row_w)
                if i < len(relevant) - 1:
                    sep = QtWidgets.QFrame()
                    sep.setFrameShape(QtWidgets.QFrame.HLine)
                    sep.setStyleSheet("color: #F1F5F9; background: #F1F5F9;"
                                     " min-height: 1px; max-height: 1px;")
                    v.addWidget(sep)

        return container

    def _open_file(self, file_path):
        try:
            import subprocess
            subprocess.Popen(f'start "" "{file_path}"', shell=True)
        except Exception:
            pass

    def refresh_recent_exports(self):
        """Rebuild and refresh both Recent Exports sections."""
        for attr, export_type in [('_excel_recent_container', 'Excel'),
                                   ('_pdf_recent_container', 'PDF')]:
            old = getattr(self, attr, None)
            if old is None:
                continue
            parent_layout = old.parentWidget().layout() if old.parentWidget() else None
            if parent_layout is None:
                continue
            idx = parent_layout.indexOf(old)
            if idx < 0:
                continue
            new_w = self._build_recent_exports_widget(export_type)
            parent_layout.insertWidget(idx, new_w)
            parent_layout.removeWidget(old)
            old.deleteLater()
            setattr(self, attr, new_w)

        # Refresh Last Exported labels in summary cards
        for export_type, last_attr in [('Excel', '_excel_last_lbl'), ('PDF', '_pdf_last_lbl')]:
            lbl = getattr(self, last_attr, None)
            if lbl is None:
                continue
            history = self._load_export_history()
            last_time = "—"
            for entry in history:
                if entry.get("type", "").lower() == export_type.lower():
                    try:
                        dt = datetime.fromisoformat(entry["datetime"])
                        now = datetime.now()
                        if dt.date() == now.date():
                            last_time = f"Today, {dt.strftime('%I:%M %p')}"
                        else:
                            last_time = dt.strftime("%b %d, %I:%M %p")
                    except Exception:
                        pass
                    break
            lbl.setText(last_time)

    def _trigger_excel_export(self):
        self._active_export_btn = getattr(self, 'excel_export_btn', None)
        self.export_type = "excel"
        self.start_export()

    def _trigger_pdf_export(self):
        self._active_export_btn = getattr(self, 'pdf_export_btn', None)
        self.export_type = "pdf"
        self.start_export()

    def _export_filter_button_style(self, active=False):
        border = "#0F766E" if active else "#DCE4EC"
        color = "#0F766E" if active else "#334155"
        background = "#F0FDFA" if active else "#FFFFFF"
        return f"""
            QPushButton {{
                background: {background};
                color: {color};
                border: 1px solid {border};
                border-radius: 8px;
                font-size: 12px;
                font-weight: 800;
                padding: 0 12px;
            }}
            QPushButton:hover {{
                background: #F8FAFC;
                border-color: #0F766E;
                color: #0F766E;
            }}
        """

    def _export_combo_style(self):
        return f"""
            QComboBox {{
                padding: 5px 30px 5px 10px;
                border: 1.5px solid #CBD5E1;
                border-radius: 7px;
                font-size: 13px;
                font-weight: 700;
                color: #1E293B;
                background: white;
            }}
            QComboBox:hover {{ border-color: #94A3B8; }}
            QComboBox:focus {{ border-color: #0F766E; }}
            QComboBox::drop-down {{
                width: 24px;
                border: none;
                background: transparent;
            }}
            QComboBox::down-arrow {{
                image: url("{CHEVRON_URL}");
                width: 14px;
                height: 14px;
                margin-right: 2px;
            }}
            QComboBox QAbstractItemView {{
                background: white;
                border: 1px solid #E2E8F0;
                selection-background-color: #F0FDF4;
                selection-color: #0F766E;
                padding: 4px;
            }}
        """

    def _parse_job_export_datetime(self, job):
        created_at_str = job.get('created_at', '')
        if not created_at_str:
            return None
        date_part = created_at_str.split('T')[0] if 'T' in created_at_str else created_at_str
        for date_format in ("%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(date_part, date_format)
            except ValueError:
                continue
        return None

    def _export_filter_values(self, export_type):
        prefix = "excel_" if export_type == "excel" else ""
        status_combo = getattr(self, f"{prefix}status_combo", None)
        client_combo = getattr(self, f"{prefix}client_combo", None)
        return {
            "status": status_combo.currentText() if status_combo else "All Status",
            "client": client_combo.currentText() if client_combo else "All Clients",
        }

    def _export_filtered_records(self, export_type):
        prefix = "excel_" if export_type == "excel" else ""
        range_type = getattr(self, "excel_export_range" if export_type == "excel" else "export_range", "all")
        filters = self._export_filter_values(export_type)
        records = []

        for job in getattr(self._owner_tab, 'job_forms', None) or []:
            job_datetime = self._parse_job_export_datetime(job)
            include_job = range_type == "all"

            if range_type == "date_range" and job_datetime:
                from_date = getattr(self, f"{prefix}from_date").date().toPyDate()
                to_date = getattr(self, f"{prefix}to_date").date().toPyDate()
                include_job = from_date <= job_datetime.date() <= to_date
            elif range_type == "month" and job_datetime:
                month_combo = getattr(self, f"{prefix}month_combo")
                year_edit = getattr(self, f"{prefix}year_edit_month")
                include_job = (
                    job_datetime.month == month_combo.currentIndex() + 1
                    and job_datetime.year == int(year_edit.text())
                )
            elif range_type == "year" and job_datetime:
                year_edit = getattr(self, f"{prefix}year_edit")
                include_job = job_datetime.year == int(year_edit.text())

            if not include_job:
                continue
            if filters["status"] != "All Status" and job.get("status", "Not Started") != filters["status"]:
                continue
            if filters["client"] != "All Clients" and job.get("client", "") != filters["client"]:
                continue
            records.append(job)

        return records

    def _refresh_export_filter_options(self, export_type):
        prefix = "excel_" if export_type == "excel" else ""
        status_combo = getattr(self, f"{prefix}status_combo", None)
        client_combo = getattr(self, f"{prefix}client_combo", None)
        if not status_combo or not client_combo:
            return

        status_current = status_combo.currentText()
        client_current = client_combo.currentText()
        jobs = getattr(self._owner_tab, 'job_forms', None) or []
        default_statuses = {
            "Draft", "Sent", "In Review", "Approved", "On Hold",
            "Completed", "Converted", "Rejected", "Expired", "Cancelled",
            "Not Started", "High", "Urgent",
        }
        statuses = ["All Status"] + sorted(default_statuses | {
            str(job.get("status", "Not Started")).strip()
            for job in jobs
            if str(job.get("status", "Not Started")).strip()
        })
        clients = ["All Clients"] + sorted({
            str(job.get("client", "")).strip()
            for job in jobs
            if str(job.get("client", "")).strip()
        }, key=str.lower)

        status_combo.blockSignals(True)
        client_combo.blockSignals(True)
        status_combo.clear()
        client_combo.clear()
        status_combo.addItems(statuses)
        client_combo.addItems(clients)
        status_combo.setCurrentText(status_current if status_current in statuses else "All Status")
        client_combo.setCurrentText(client_current if client_current in clients else "All Clients")
        status_combo.blockSignals(False)
        client_combo.blockSignals(False)

    def _setup_reference_export_task(self, host, export_type):
        is_excel = export_type == "excel"
        prefix = "excel_" if is_excel else ""
        title_text = "Excel" if is_excel else "PDF"

        layout = QtWidgets.QVBoxLayout(host)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)
        layout.setAlignment(QtCore.Qt.AlignTop)

        filters_card = QtWidgets.QFrame()
        filters_card.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }
        """)
        filters_layout = QtWidgets.QVBoxLayout(filters_card)
        filters_layout.setContentsMargins(18, 14, 18, 16)
        filters_layout.setSpacing(12)
        filters_title = QtWidgets.QLabel("2   Choose Filters")
        filters_title.setStyleSheet(
            "font-size: 13px; font-weight: 900; color: #0F172A;"
            " background: transparent; border: none;"
        )
        filters_layout.addWidget(filters_title)

        top_row = QtWidgets.QHBoxLayout()
        top_row.setSpacing(14)

        def _field(label_text):
            wrap = QtWidgets.QVBoxLayout()
            wrap.setSpacing(5)
            label = QtWidgets.QLabel(label_text)
            label.setStyleSheet(
                "font-size: 11px; font-weight: 800; color: #334155;"
                " background: transparent; border: none;"
            )
            wrap.addWidget(label)
            return wrap

        from_date = QtWidgets.QDateEdit()
        from_date.setCalendarPopup(True)
        from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        from_date.setDisplayFormat("MM/dd/yyyy")
        from_date.setFixedSize(146, 36)
        from_date.setStyleSheet(self._DATE)
        to_date = QtWidgets.QDateEdit()
        to_date.setCalendarPopup(True)
        to_date.setDate(QtCore.QDate.currentDate())
        to_date.setDisplayFormat("MM/dd/yyyy")
        to_date.setFixedSize(146, 36)
        to_date.setStyleSheet(self._DATE)
        arrow = QtWidgets.QLabel("→")
        arrow.setStyleSheet("font-size: 14px; color: #64748B; border: none; background: transparent;")
        date_range_wrap = QtWidgets.QFrame()
        date_range_wrap.setStyleSheet("""
            QFrame {
                background: #F0F9FF;
                border: 1.5px solid #7DD3FC;
                border-radius: 10px;
            }
        """)
        drw_lay = QtWidgets.QVBoxLayout(date_range_wrap)
        drw_lay.setContentsMargins(14, 10, 14, 12)
        drw_lay.setSpacing(8)
        drw_hdr = QtWidgets.QHBoxLayout()
        drw_icon = QtWidgets.QLabel("📅")
        drw_icon.setStyleSheet("background: transparent; border: none; font-size: 15px;")
        drw_title = QtWidgets.QLabel("Filter by Date Range")
        drw_title.setStyleSheet(
            "background: transparent; border: none; font-size: 12px;"
            " font-weight: 900; color: #0369A1;")
        drw_hdr.addWidget(drw_icon)
        drw_hdr.addWidget(drw_title)
        drw_hdr.addStretch()
        drw_lay.addLayout(drw_hdr)
        date_row = QtWidgets.QHBoxLayout()
        date_row.setSpacing(8)
        date_row.addWidget(from_date)
        date_row.addWidget(arrow)
        date_row.addWidget(to_date)
        date_row.addStretch()
        drw_lay.addLayout(date_row)
        date_range_wrap.setVisible(False)

        range_col = _field("Quick Range")
        quick_range = QtWidgets.QComboBox()
        quick_range.addItem("All Quote Forms", "all")
        quick_range.addItem("Today", "today")
        quick_range.addItem("Last 7 Days", "last_7")
        quick_range.addItem("Last 30 Days", "last_30")
        quick_range.addItem("This Month", "this_month")
        quick_range.addItem("This Year", "this_year")
        quick_range.addItem("Custom Date Range", "date_range")
        quick_range.addItem("Select Month", "month")
        quick_range.addItem("Select Year", "year")
        quick_range.setFixedSize(220, 36)
        quick_range.setStyleSheet(self._export_combo_style())
        range_col.addWidget(quick_range)
        top_row.addLayout(range_col)

        status_col = _field("Status")
        status_combo = QtWidgets.QComboBox()
        status_combo.addItem("All Status")
        status_combo.setFixedSize(210, 36)
        status_combo.setStyleSheet(self._export_combo_style())
        status_col.addWidget(status_combo)
        top_row.addLayout(status_col)

        client_col = _field("Client")
        client_combo = QtWidgets.QComboBox()
        client_combo.addItem("All Clients")
        client_combo.setFixedSize(210, 36)
        client_combo.setStyleSheet(self._export_combo_style())
        client_col.addWidget(client_combo)
        top_row.addLayout(client_col)
        top_row.addStretch()
        filters_layout.addLayout(top_row)

        more_row = QtWidgets.QHBoxLayout()
        clear_btn = QtWidgets.QPushButton("↻  Clear All")
        clear_btn.setFixedSize(118, 36)
        clear_btn.setStyleSheet(self._export_filter_button_style(False))
        more_row.addStretch()
        more_row.addWidget(clear_btn)
        filters_layout.addLayout(more_row)
        layout.addWidget(filters_card)

        # ── "More Filters" expandable panel ──────────────────────────
        hidden_wrap = QtWidgets.QWidget()
        hidden_wrap.setStyleSheet("background: transparent;")
        hidden_layout = QtWidgets.QHBoxLayout(hidden_wrap)
        hidden_layout.setContentsMargins(0, 8, 0, 0)
        hidden_layout.setSpacing(12)

        date_group = QtWidgets.QWidget()  # kept for compat; unused

        # ── Month card ────────────────────────────────────────────
        month_group = QtWidgets.QFrame()
        month_group.setStyleSheet("""
            QFrame {
                background: #F0FDF4;
                border: 1.5px solid #6EE7B7;
                border-radius: 10px;
            }
        """)
        mg_lay = QtWidgets.QVBoxLayout(month_group)
        mg_lay.setContentsMargins(14, 10, 14, 12)
        mg_lay.setSpacing(8)

        mg_hdr = QtWidgets.QHBoxLayout()
        mg_icon = QtWidgets.QLabel("📅")
        mg_icon.setStyleSheet("background: transparent; border: none; font-size: 15px;")
        mg_title = QtWidgets.QLabel("Filter by Month")
        mg_title.setStyleSheet(
            "background: transparent; border: none; font-size: 12px;"
            " font-weight: 900; color: #065F46;"
        )
        mg_hdr.addWidget(mg_icon)
        mg_hdr.addWidget(mg_title)
        mg_hdr.addStretch()
        mg_lay.addLayout(mg_hdr)

        mg_fields = QtWidgets.QHBoxLayout()
        mg_fields.setSpacing(8)

        month_combo = QtWidgets.QComboBox()
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        month_combo.addItems(months)
        month_combo.setCurrentIndex(datetime.now().month - 1)
        month_combo.setFixedHeight(34)
        month_combo.setMinimumWidth(140)
        month_combo.setStyleSheet("""
            QComboBox {
                padding: 4px 28px 4px 10px; border: 1.5px solid #6EE7B7;
                border-radius: 7px; font-size: 13px; font-weight: 700;
                color: #065F46; background: white;
            }
            QComboBox:hover { border-color: #34D399; }
            QComboBox::drop-down { width: 22px; border: none; background: transparent; }
            QComboBox::down-arrow { image: url("%s"); width: 12px; height: 12px; }
            QComboBox QAbstractItemView { background: white; border: 1px solid #6EE7B7;
                selection-background-color: #D1FAE5; selection-color: #065F46; padding: 4px; }
        """ % CHEVRON_URL)
        month_combo.wheelEvent = lambda e: e.ignore()

        year_month_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        year_month_edit.setFixedSize(80, 34)
        year_month_edit.setReadOnly(True)
        year_month_edit.setStyleSheet("""
            QLineEdit { padding: 4px 8px; border: 1.5px solid #6EE7B7; border-radius: 7px;
                font-size: 13px; font-weight: 700; color: #065F46; background: white; }
        """)
        month_year_btn = QtWidgets.QPushButton("Select")
        month_year_btn.setFixedSize(76, 34)
        month_year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        month_year_btn.setStyleSheet("""
            QPushButton {
                background: #f9fafb; color: #374151;
                border: 1px solid #d1d5db; border-radius: 8px;
                font-size: 12px; font-weight: 600; font-family: 'Inter','Segoe UI';
            }
            QPushButton:hover { background: #f3f4f6; border-color: #9ca3af; }
            QPushButton:pressed { background: #e5e7eb; }
        """)
        month_year_btn.clicked.connect(lambda: self.show_year_popup_for_month_excel() if is_excel else self.show_year_popup_for_month())
        mg_fields.addWidget(month_combo)
        mg_fields.addWidget(year_month_edit)
        mg_fields.addWidget(month_year_btn)
        mg_fields.addStretch()
        mg_lay.addLayout(mg_fields)

        # ── Year card ─────────────────────────────────────────────
        year_group = QtWidgets.QFrame()
        year_group.setStyleSheet("""
            QFrame {
                background: #EFF6FF;
                border: 1.5px solid #93C5FD;
                border-radius: 10px;
            }
        """)
        yg_lay = QtWidgets.QVBoxLayout(year_group)
        yg_lay.setContentsMargins(14, 10, 14, 12)
        yg_lay.setSpacing(8)

        yg_hdr = QtWidgets.QHBoxLayout()
        yg_icon = QtWidgets.QLabel("📆")
        yg_icon.setStyleSheet("background: transparent; border: none; font-size: 15px;")
        yg_title = QtWidgets.QLabel("Filter by Year")
        yg_title.setStyleSheet(
            "background: transparent; border: none; font-size: 12px;"
            " font-weight: 900; color: #1D4ED8;"
        )
        yg_hdr.addWidget(yg_icon)
        yg_hdr.addWidget(yg_title)
        yg_hdr.addStretch()
        yg_lay.addLayout(yg_hdr)

        yg_fields = QtWidgets.QHBoxLayout()
        yg_fields.setSpacing(8)

        year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        year_edit.setFixedSize(80, 34)
        year_edit.setReadOnly(True)
        year_edit.setStyleSheet("""
            QLineEdit { padding: 4px 8px; border: 1.5px solid #93C5FD; border-radius: 7px;
                font-size: 13px; font-weight: 700; color: #1D4ED8; background: white; }
        """)
        year_btn = QtWidgets.QPushButton("Select")
        year_btn.setFixedSize(76, 34)
        year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        year_btn.setStyleSheet("""
            QPushButton {
                background: #f9fafb; color: #374151;
                border: 1px solid #d1d5db; border-radius: 8px;
                font-size: 12px; font-weight: 600; font-family: 'Inter','Segoe UI';
            }
            QPushButton:hover { background: #f3f4f6; border-color: #9ca3af; }
            QPushButton:pressed { background: #e5e7eb; }
        """)
        year_btn.clicked.connect(lambda: self.show_year_popup_excel() if is_excel else self.show_year_popup())
        yg_fields.addWidget(year_edit)
        yg_fields.addWidget(year_btn)
        yg_fields.addStretch()
        yg_lay.addLayout(yg_fields)

        hidden_layout.addWidget(date_range_wrap, 1)
        hidden_layout.addWidget(month_group, 1)
        hidden_layout.addWidget(year_group, 1)
        hidden_wrap.setVisible(False)
        date_range_wrap.setVisible(False)
        month_group.setVisible(False)
        year_group.setVisible(False)
        filters_layout.addWidget(hidden_wrap)

        setattr(self, f"{prefix}from_date", from_date)
        setattr(self, f"{prefix}to_date", to_date)
        setattr(self, f"{prefix}month_combo", month_combo)
        setattr(self, f"{prefix}year_edit_month", year_month_edit)
        setattr(self, f"{prefix}year_edit", year_edit)
        setattr(self, f"{prefix}date_range_group", date_group)
        setattr(self, f"{prefix}date_range_wrap", date_range_wrap)
        setattr(self, f"{prefix}month_group", month_group)
        setattr(self, f"{prefix}year_group", year_group)
        setattr(self, f"{prefix}date_selection_container", hidden_wrap)
        setattr(self, f"{prefix}quick_range_combo", quick_range)
        setattr(self, f"{prefix}status_combo", status_combo)
        setattr(self, f"{prefix}client_combo", client_combo)
        def _lock_combo(cb):
            """Prevent wheel and arrow keys from changing combo value; clear focus on select."""
            cb.wheelEvent = lambda e: e.ignore()
            def _kp(event, _cb=cb):
                if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) \
                        and not _cb.view().isVisible():
                    event.ignore()
                    return
                QtWidgets.QComboBox.keyPressEvent(_cb, event)
            cb.keyPressEvent = _kp
            cb.currentIndexChanged.connect(
                lambda: QtCore.QTimer.singleShot(0, cb.clearFocus))

        def _lock_date(de):
            de.wheelEvent = lambda e: e.ignore()
            de.stepBy = lambda x: None

        _lock_date(from_date)
        _lock_date(to_date)
        _lock_combo(quick_range)
        _lock_combo(status_combo)
        _lock_combo(client_combo)
        _lock_combo(month_combo)
        self._refresh_export_filter_options(export_type)

        summary_card = self._build_reference_summary_card(title_text)
        setattr(self, f"{prefix}summary_reference_card", summary_card)
        layout.addWidget(summary_card)

        action_row = QtWidgets.QHBoxLayout()
        action_row.addStretch()
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setFixedSize(120, 40)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: white; color: #475569; border: 1px solid #DCE4EC;
                border-radius: 8px; font-size: 13px; font-weight: 800;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        export_btn = QtWidgets.QPushButton(f"↓  Export {title_text}")
        export_btn.setFixedSize(150, 40)
        export_btn.setStyleSheet("""
            QPushButton {
                background: #0F766E; color: white; border: none;
                border-radius: 8px; font-size: 13px; font-weight: 900;
            }
            QPushButton:hover { background: #115E59; }
        """)
        if is_excel:
            self.excel_export_btn = export_btn
            export_btn.clicked.connect(self._trigger_excel_export)
        else:
            self.pdf_export_btn = export_btn
            export_btn.clicked.connect(self._trigger_pdf_export)
        action_row.addWidget(cancel_btn)
        action_row.addSpacing(10)
        action_row.addWidget(export_btn)
        layout.addLayout(action_row)
        layout.addStretch()

        preview_label = QtWidgets.QLabel()
        if is_excel:
            self.excel_preview_label = preview_label
        else:
            self.preview_label = preview_label

        def _scroll_preserve(fn):
            """Run fn() then restore parent scroll position to avoid jarring jumps."""
            sa = self.parentWidget()
            while sa and not isinstance(sa, QtWidgets.QScrollArea):
                sa = sa.parentWidget()
            saved = sa.verticalScrollBar().value() if sa else 0
            fn()
            if sa:
                QtCore.QTimer.singleShot(0, lambda: sa.verticalScrollBar().setValue(saved))

        def _apply_quick_range():
            selected = quick_range.currentData()
            today = QtCore.QDate.currentDate()
            if selected == "today":
                from_date.setDate(today)
                to_date.setDate(today)
                selected = "date_range"
            elif selected == "last_7":
                from_date.setDate(today.addDays(-6))
                to_date.setDate(today)
                selected = "date_range"
            elif selected == "last_30":
                from_date.setDate(today.addDays(-29))
                to_date.setDate(today)
                selected = "date_range"
            elif selected == "this_month":
                month_combo.setCurrentIndex(today.month() - 1)
                year_month_edit.setText(str(today.year()))
                selected = "month"
            elif selected == "this_year":
                year_edit.setText(str(today.year()))
                selected = "year"

            def _do_visibility():
                if selected == "date_range":
                    hidden_wrap.setVisible(True)
                    date_range_wrap.setVisible(True)
                    month_group.setVisible(False)
                    year_group.setVisible(False)
                elif selected in ("month", "this_month"):
                    hidden_wrap.setVisible(True)
                    date_range_wrap.setVisible(False)
                    month_group.setVisible(True)
                    year_group.setVisible(False)
                elif selected in ("year", "this_year"):
                    hidden_wrap.setVisible(True)
                    date_range_wrap.setVisible(False)
                    month_group.setVisible(False)
                    year_group.setVisible(True)
                else:
                    hidden_wrap.setVisible(False)
                    date_range_wrap.setVisible(False)
                    month_group.setVisible(False)
                    year_group.setVisible(False)
            _scroll_preserve(_do_visibility)

            if is_excel:
                self.on_excel_range_changed(selected)
            else:
                self.on_range_changed(selected)

        def _clear_filters():
            def _do_clear():
                quick_range.setCurrentIndex(0)
                from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
                to_date.setDate(QtCore.QDate.currentDate())
                month_combo.setCurrentIndex(datetime.now().month - 1)
                year_month_edit.setText(str(datetime.now().year))
                year_edit.setText(str(datetime.now().year))
                status_combo.setCurrentText("All Status")
                client_combo.setCurrentText("All Clients")
                date_range_wrap.setVisible(False)
                hidden_wrap.setVisible(False)
                month_group.setVisible(False)
                year_group.setVisible(False)
            _scroll_preserve(_do_clear)
            if is_excel:
                self.on_excel_range_changed("all")
            else:
                self.on_range_changed("all")

        def _range_changed():
            _apply_quick_range()

        quick_range.currentIndexChanged.connect(_range_changed)
        from_date.dateChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        to_date.dateChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        month_combo.currentTextChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        year_month_edit.textChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        year_edit.textChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        status_combo.currentTextChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        client_combo.currentTextChanged.connect(self.update_excel_preview if is_excel else self.update_preview)
        clear_btn.clicked.connect(_clear_filters)
        quick_range.setCurrentIndex(0)
        _range_changed()

    def _build_reference_summary_card(self, export_label):
        card = QtWidgets.QFrame()
        card.setStyleSheet("""
            QFrame { background: white; border: 1px solid #E2E8F0; border-radius: 10px; }
        """)
        layout = QtWidgets.QVBoxLayout(card)
        layout.setContentsMargins(18, 14, 18, 16)
        layout.setSpacing(12)
        title = QtWidgets.QLabel("3   Export Summary")
        title.setStyleSheet("font-size: 13px; font-weight: 900; color: #0F172A; border: none;")
        layout.addWidget(title)
        cards = QtWidgets.QHBoxLayout()
        cards.setSpacing(12)
        total = self._get_total_count()
        total_card, total_lbl = self._mini_summary_card("Total Quotes", str(total), "#2563EB", "doc")
        filtered_card, filtered_lbl = self._mini_summary_card("Filtered Quotes", str(total), "#059669", "check")
        clients_card, clients_lbl = self._mini_summary_card("Clients Included", "—", "#7C3AED", "users")
        scope_card, scope_lbl = self._mini_summary_card("Date Range", "All Quote Forms", "#D97706", "calendar")
        for widget in (total_card, filtered_card, clients_card, scope_card):
            cards.addWidget(widget, 1)
        layout.addLayout(cards)
        info = QtWidgets.QLabel(f"ⓘ  You are about to export {total} quote form(s) in {export_label} format.")
        info.setStyleSheet("""
            QLabel {
                background: #EFF6FF; color: #2563EB; border: 1px solid #DBEAFE;
                border-radius: 8px; font-size: 12px; font-weight: 700; padding: 12px 14px;
            }
        """)
        layout.addWidget(info)
        key = export_label.lower()
        setattr(self, f"_{key}_filtered_summary_lbl", filtered_lbl)
        setattr(self, f"_{key}_clients_summary_lbl", clients_lbl)
        setattr(self, f"_{key}_scope_summary_lbl", scope_lbl)
        setattr(self, f"_{key}_info_summary_lbl", info)
        return card

    def _mini_summary_card(self, title, value, color, icon_kind):
        card = QtWidgets.QFrame()
        card.setStyleSheet("QFrame { background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 8px; }")
        layout = QtWidgets.QHBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(10)
        icon = QtWidgets.QLabel()
        icon.setFixedSize(42, 42)
        icon.setAlignment(QtCore.Qt.AlignCenter)
        icon.setPixmap(self._make_summary_icon(icon_kind, color).pixmap(38, 38))
        t = QtWidgets.QLabel(title)
        t.setStyleSheet("font-size: 11px; font-weight: 800; color: #64748B; border: none;")
        v = QtWidgets.QLabel(value)
        v.setStyleSheet(f"font-size: 18px; font-weight: 900; color: {color}; border: none;")
        text_col = QtWidgets.QVBoxLayout()
        text_col.setSpacing(4)
        text_col.addWidget(t)
        text_col.addWidget(v)
        layout.addWidget(icon)
        layout.addLayout(text_col, 1)
        return card, v

    @staticmethod
    def _make_summary_icon(kind, color):
        pixmap = QtGui.QPixmap(42, 42)
        pixmap.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pixmap)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
        accent = QtGui.QColor(color)
        bg = QtGui.QColor(accent)
        bg.setAlpha(28)
        painter.setPen(QtCore.Qt.NoPen)
        painter.setBrush(bg)
        painter.drawRoundedRect(QtCore.QRectF(1, 1, 40, 40), 9, 9)
        pen = QtGui.QPen(accent)
        pen.setWidthF(2.0)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        pen.setCapStyle(QtCore.Qt.RoundCap)
        painter.setPen(pen)
        painter.setBrush(QtCore.Qt.NoBrush)
        if kind == "doc":
            painter.drawRoundedRect(QtCore.QRectF(14, 10, 14, 20), 2, 2)
            painter.drawLine(17, 16, 25, 16)
            painter.drawLine(17, 21, 25, 21)
        elif kind == "check":
            painter.drawEllipse(QtCore.QRectF(11, 11, 20, 20))
            path = QtGui.QPainterPath()
            path.moveTo(16, 21)
            path.lineTo(20, 25)
            path.lineTo(27, 16)
            painter.drawPath(path)
        elif kind == "users":
            painter.drawEllipse(QtCore.QRectF(17, 11, 8, 8))
            painter.drawRoundedRect(QtCore.QRectF(13, 22, 16, 8), 4, 4)
        else:
            painter.drawRoundedRect(QtCore.QRectF(12, 12, 18, 18), 3, 3)
            painter.drawLine(16, 9, 16, 15)
            painter.drawLine(26, 9, 26, 15)
            painter.drawLine(12, 18, 30, 18)
        painter.end()
        return QtGui.QIcon(pixmap)

    @staticmethod
    def _make_calendar_icon(color="#0F766E", size=28):
        pixmap = QtGui.QPixmap(size, size)
        pixmap.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pixmap)
        painter.setRenderHint(QtGui.QPainter.Antialiasing, True)

        accent = QtGui.QColor(color)
        bg = QtGui.QColor(accent)
        bg.setAlpha(24)

        outer = QtCore.QRectF(1, 1, size - 2, size - 2)
        painter.setPen(QtCore.Qt.NoPen)
        painter.setBrush(bg)
        painter.drawRoundedRect(outer, 6, 6)

        scale = size / 28.0
        pen = QtGui.QPen(accent)
        pen.setWidthF(max(1.6, 1.8 * scale))
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        pen.setCapStyle(QtCore.Qt.RoundCap)
        painter.setPen(pen)
        painter.setBrush(QtCore.Qt.NoBrush)
        painter.drawRoundedRect(
            QtCore.QRectF(7 * scale, 7 * scale, 14 * scale, 14 * scale),
            2.5 * scale,
            2.5 * scale,
        )
        painter.drawLine(
            QtCore.QPointF(10 * scale, 5 * scale),
            QtCore.QPointF(10 * scale, 10 * scale),
        )
        painter.drawLine(
            QtCore.QPointF(18 * scale, 5 * scale),
            QtCore.QPointF(18 * scale, 10 * scale),
        )
        painter.drawLine(
            QtCore.QPointF(7 * scale, 12 * scale),
            QtCore.QPointF(21 * scale, 12 * scale),
        )
        painter.end()
        return QtGui.QIcon(pixmap)

    def setup_pdf_tab(self):
        """Setup the PDF export sub-window."""
        self._setup_reference_export_task(self.pdf_tab, "pdf")
        return
        layout = QtWidgets.QVBoxLayout(self.pdf_tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)
        layout.setAlignment(QtCore.Qt.AlignTop)

        # ── Range label ──────────────────────────────────────
        layout.addWidget(self._field_lbl("EXPORT RANGE"))

        # ── Pill buttons ─────────────────────────────────────
        rg = QtWidgets.QButtonGroup(self)
        self.all_radio        = QtWidgets.QRadioButton("All Quote Forms")
        self.date_range_radio = QtWidgets.QRadioButton("Date Range")
        self.month_radio      = QtWidgets.QRadioButton("By Month")
        self.year_radio       = QtWidgets.QRadioButton("By Year")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(       lambda: self.on_range_changed("all"))
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        self.month_radio.toggled.connect(     lambda: self.on_range_changed("month"))
        self.year_radio.toggled.connect(      lambda: self.on_range_changed("year"))
        layout.addLayout(self._build_range_row(
            self.all_radio, self.date_range_radio,
            self.month_radio, self.year_radio, rg))

        # ── Date selection container ─────────────────────────
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_container.setStyleSheet("QWidget { background: transparent; }")
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setContentsMargins(0, 0, 0, 0)
        self.date_selection_layout.setSpacing(8)

        # Date Range card
        self.date_range_group = QtWidgets.QGroupBox("Date Range")
        self.date_range_group.setStyleSheet(self._CARD)
        dr = QtWidgets.QHBoxLayout(self.date_range_group)
        dr.setContentsMargins(14, 8, 14, 10)
        dr.setSpacing(16)

        from_col = QtWidgets.QVBoxLayout()
        from_col.setSpacing(4)
        from_col.addWidget(self._field_lbl("FROM DATE"))
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(200, 34)
        self.from_date.setStyleSheet(self._DATE)
        from_col.addWidget(self.from_date)
        dr.addLayout(from_col)

        arrow = QtWidgets.QLabel("→")
        arrow.setStyleSheet(
            "font-size: 16px; color: #94A3B8; background: transparent;"
            " border: none; padding-top: 20px;")
        dr.addWidget(arrow)

        to_col = QtWidgets.QVBoxLayout()
        to_col.setSpacing(4)
        to_col.addWidget(self._field_lbl("TO DATE"))
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(200, 34)
        self.to_date.setStyleSheet(self._DATE)
        to_col.addWidget(self.to_date)
        dr.addLayout(to_col)
        dr.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month & Year card
        self.month_group = QtWidgets.QGroupBox("Month && Year")
        self.month_group.setStyleSheet(self._CARD)
        mg = QtWidgets.QHBoxLayout(self.month_group)
        mg.setContentsMargins(14, 8, 14, 10)
        mg.setSpacing(30)

        m_col = QtWidgets.QVBoxLayout()
        m_col.setSpacing(4)
        m_col.addWidget(self._field_lbl("MONTH"))
        self.month_combo = _NoScrollComboBox()
        self.month_combo.setFixedSize(175, 34)
        self.month_combo.setStyleSheet(
            "QComboBox { padding: 5px 10px; border: 1.5px solid #CBD5E1;"
            " border-radius: 7px; font-size: 13px; font-weight: 600;"
            " color: #1E293B; background: white; }"
            "QComboBox:focus { border-color: #0F766E; }")
        self.populate_months()
        m_col.addWidget(self.month_combo)
        mg.addLayout(m_col)

        ym_col = QtWidgets.QVBoxLayout()
        ym_col.setSpacing(4)
        ym_col.addWidget(self._field_lbl("YEAR"))
        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setReadOnly(True)
        self.year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)
        ym_col.addLayout(self._build_year_row(
            self.year_edit_month, self.year_calendar_btn_month))
        mg.addLayout(ym_col)
        mg.addStretch()
        self.date_selection_layout.addWidget(self.month_group)

        # Year card
        self.year_group = QtWidgets.QGroupBox("Year")
        self.year_group.setStyleSheet(self._CARD)
        yg = QtWidgets.QHBoxLayout(self.year_group)
        yg.setContentsMargins(14, 8, 14, 10)
        yg.setSpacing(6)

        y_col = QtWidgets.QVBoxLayout()
        y_col.setSpacing(4)
        y_col.addWidget(self._field_lbl("YEAR"))
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setReadOnly(True)
        self.year_calendar_btn = QtWidgets.QPushButton("📅")
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        y_col.addLayout(self._build_year_row(self.year_edit, self.year_calendar_btn))
        yg.addLayout(y_col)
        yg.addStretch()
        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)

        # ── Status bar ──────────────────────────────────────────
        self.preview_label = QtWidgets.QLabel("Will export ALL quote forms as PDF")
        _pdf_preview = self._build_status_bar(self.preview_label)
        layout.addWidget(_pdf_preview)

        # ── Export Summary card ──────────────────────────────────
        self.pdf_summary_card = self._build_export_summary_card("PDF")
        layout.addWidget(self.pdf_summary_card)

        # ── Large Export PDF button ──────────────────────────────
        self.pdf_export_btn = QtWidgets.QPushButton("  Export PDF")
        self.pdf_export_btn.setFixedSize(240, 40)
        self.pdf_export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.pdf_export_btn.setStyleSheet("""
            QPushButton {
                background: #0F766E;
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 14px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
            }
            QPushButton:hover { background: #115E59; }
            QPushButton:disabled { background: #CBD5E1; color: #94A3B8; }
        """)
        self.pdf_export_btn.clicked.connect(self._trigger_pdf_export)
        pdf_btn_row = QtWidgets.QHBoxLayout()
        pdf_btn_row.addStretch()
        pdf_btn_row.addWidget(self.pdf_export_btn)
        pdf_btn_row.addStretch()
        layout.addLayout(pdf_btn_row)

        # ── Security note ────────────────────────────────────────
        sec = QtWidgets.QLabel("🔒  Your data is secure and will be exported in PDF format.")
        sec.setStyleSheet("font-size: 11px; color: #94A3B8; background: transparent; border: none;")
        sec.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(sec)

        # ── Recent Exports ───────────────────────────────────────
        self._pdf_recent_container = self._build_recent_exports_widget("PDF")
        layout.addWidget(self._pdf_recent_container)
        layout.addStretch()

        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.from_date.wheelEvent = lambda e: e.ignore()
        self.from_date.stepBy = lambda x: None
        self.to_date.wheelEvent = lambda e: e.ignore()
        self.to_date.stepBy = lambda x: None
        self.month_combo.wheelEvent = lambda e: e.ignore()
        self.month_combo.currentTextChanged.connect(self.update_preview)

    def setup_excel_tab(self):
        """Setup the Excel export sub-window."""
        self._setup_reference_export_task(self.excel_tab, "excel")
        return
        layout = QtWidgets.QVBoxLayout(self.excel_tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setAlignment(QtCore.Qt.AlignTop)
        layout.setSpacing(12)

        layout.addWidget(self._field_lbl("EXPORT RANGE"))

        rg = QtWidgets.QButtonGroup(self)
        self.excel_all_radio        = QtWidgets.QRadioButton("All Quote Forms")
        self.excel_date_range_radio = QtWidgets.QRadioButton("Date Range")
        self.excel_month_radio      = QtWidgets.QRadioButton("By Month")
        self.excel_year_radio       = QtWidgets.QRadioButton("By Year")
        self.excel_all_radio.setChecked(True)
        self.excel_all_radio.toggled.connect(       lambda: self.on_excel_range_changed("all"))
        self.excel_date_range_radio.toggled.connect(lambda: self.on_excel_range_changed("date_range"))
        self.excel_month_radio.toggled.connect(     lambda: self.on_excel_range_changed("month"))
        self.excel_year_radio.toggled.connect(      lambda: self.on_excel_range_changed("year"))
        layout.addLayout(self._build_range_row(
            self.excel_all_radio, self.excel_date_range_radio,
            self.excel_month_radio, self.excel_year_radio, rg))

        self.excel_date_selection_container = QtWidgets.QWidget()
        self.excel_date_selection_container.setStyleSheet("QWidget { background: transparent; }")
        self.excel_date_selection_layout = QtWidgets.QVBoxLayout(self.excel_date_selection_container)
        self.excel_date_selection_layout.setContentsMargins(0, 0, 0, 0)
        self.excel_date_selection_layout.setSpacing(8)

        # Date Range card
        self.excel_date_range_group = QtWidgets.QGroupBox("Date Range")
        self.excel_date_range_group.setStyleSheet(self._CARD)
        edr = QtWidgets.QHBoxLayout(self.excel_date_range_group)
        edr.setContentsMargins(14, 8, 14, 10)
        edr.setSpacing(16)

        efrom_col = QtWidgets.QVBoxLayout()
        efrom_col.setSpacing(4)
        efrom_col.addWidget(self._field_lbl("FROM DATE"))
        self.excel_from_date = _NoScrollDateEdit()
        self.excel_from_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.excel_from_date.setCalendarPopup(True)
        self.excel_from_date.setFixedSize(200, 34)
        self.excel_from_date.setStyleSheet(self._DATE)
        efrom_col.addWidget(self.excel_from_date)
        edr.addLayout(efrom_col)

        earrow = QtWidgets.QLabel("→")
        earrow.setStyleSheet(
            "font-size: 16px; color: #94A3B8; background: transparent;"
            " border: none; padding-top: 20px;")
        edr.addWidget(earrow)

        eto_col = QtWidgets.QVBoxLayout()
        eto_col.setSpacing(4)
        eto_col.addWidget(self._field_lbl("TO DATE"))
        self.excel_to_date = _NoScrollDateEdit()
        self.excel_to_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_to_date.setDate(QtCore.QDate.currentDate())
        self.excel_to_date.setCalendarPopup(True)
        self.excel_to_date.setFixedSize(200, 34)
        self.excel_to_date.setStyleSheet(self._DATE)
        eto_col.addWidget(self.excel_to_date)
        edr.addLayout(eto_col)
        edr.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_date_range_group)

        # Month & Year card
        self.excel_month_group = QtWidgets.QGroupBox("Month && Year")
        self.excel_month_group.setStyleSheet(self._CARD)
        emg = QtWidgets.QHBoxLayout(self.excel_month_group)
        emg.setContentsMargins(14, 8, 14, 10)
        emg.setSpacing(30)

        em_col = QtWidgets.QVBoxLayout()
        em_col.setSpacing(4)
        em_col.addWidget(self._field_lbl("MONTH"))
        self.excel_month_combo = _NoScrollComboBox()
        self.excel_month_combo.setFixedSize(175, 34)
        self.excel_month_combo.setStyleSheet(
            "QComboBox { padding: 5px 10px; border: 1.5px solid #CBD5E1;"
            " border-radius: 7px; font-size: 13px; font-weight: 600;"
            " color: #1E293B; background: white; }"
            "QComboBox:focus { border-color: #0F766E; }")
        self.populate_months_excel()
        em_col.addWidget(self.excel_month_combo)
        emg.addLayout(em_col)

        eym_col = QtWidgets.QVBoxLayout()
        eym_col.setSpacing(4)
        eym_col.addWidget(self._field_lbl("YEAR"))
        self.excel_year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit_month.setReadOnly(True)
        self.excel_year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.excel_year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month_excel)
        eym_col.addLayout(self._build_year_row(
            self.excel_year_edit_month, self.excel_year_calendar_btn_month))
        emg.addLayout(eym_col)
        emg.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_month_group)

        # Year card
        self.excel_year_group = QtWidgets.QGroupBox("Year")
        self.excel_year_group.setStyleSheet(self._CARD)
        eyg = QtWidgets.QHBoxLayout(self.excel_year_group)
        eyg.setContentsMargins(14, 8, 14, 10)
        eyg.setSpacing(6)

        ey_col = QtWidgets.QVBoxLayout()
        ey_col.setSpacing(4)
        ey_col.addWidget(self._field_lbl("YEAR"))
        self.excel_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit.setReadOnly(True)
        self.excel_year_calendar_btn = QtWidgets.QPushButton("📅")
        self.excel_year_calendar_btn.clicked.connect(self.show_year_popup_excel)
        ey_col.addLayout(self._build_year_row(
            self.excel_year_edit, self.excel_year_calendar_btn))
        eyg.addLayout(ey_col)
        eyg.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_year_group)

        layout.addWidget(self.excel_date_selection_container)

        self.excel_date_selection_container.setVisible(False)
        self.excel_date_range_group.setVisible(False)
        self.excel_month_group.setVisible(False)
        self.excel_year_group.setVisible(False)

        # ── Status bar ──────────────────────────────────────────
        self.excel_preview_label = QtWidgets.QLabel("Will export ALL quote forms as Excel")
        _xl_preview = self._build_status_bar(self.excel_preview_label)
        layout.addWidget(_xl_preview)

        # ── Export Summary card ──────────────────────────────────
        self.excel_summary_card = self._build_export_summary_card("Excel")
        layout.addWidget(self.excel_summary_card)

        # ── Large Export Excel button ────────────────────────────
        self.excel_export_btn = QtWidgets.QPushButton("  Export Excel")
        self.excel_export_btn.setFixedSize(240, 40)
        self.excel_export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_export_btn.setStyleSheet("""
            QPushButton {
                background: #0F766E;
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 14px;
                font-weight: 800;
                font-family: 'Inter', 'Segoe UI';
            }
            QPushButton:hover { background: #115E59; }
            QPushButton:disabled { background: #CBD5E1; color: #94A3B8; }
        """)
        self.excel_export_btn.clicked.connect(self._trigger_excel_export)
        excel_btn_row = QtWidgets.QHBoxLayout()
        excel_btn_row.addStretch()
        excel_btn_row.addWidget(self.excel_export_btn)
        excel_btn_row.addStretch()
        layout.addLayout(excel_btn_row)

        # ── Security note ────────────────────────────────────────
        sec = QtWidgets.QLabel("🔒  Your data is secure and will be exported in Excel format.")
        sec.setStyleSheet("font-size: 11px; color: #94A3B8; background: transparent; border: none;")
        sec.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(sec)

        # ── Recent Exports ───────────────────────────────────────
        self._excel_recent_container = self._build_recent_exports_widget("Excel")
        layout.addWidget(self._excel_recent_container)
        layout.addStretch()

        self.excel_from_date.dateChanged.connect(self.update_excel_preview)
        self.excel_to_date.dateChanged.connect(self.update_excel_preview)
        self.excel_from_date.wheelEvent = lambda e: e.ignore()
        self.excel_from_date.stepBy = lambda x: None
        self.excel_to_date.wheelEvent = lambda e: e.ignore()
        self.excel_to_date.stepBy = lambda x: None
        self.excel_month_combo.wheelEvent = lambda e: e.ignore()
        self.excel_month_combo.currentTextChanged.connect(self.update_excel_preview)
    
    def show_year_popup(self):
        """Show year calendar popup for PDF year selection"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        # Center the popup
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show year calendar popup for PDF month+year selection"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_excel(self):
        """Show year calendar popup for Excel year selection"""
        try:
            current_year = int(self.excel_year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month_excel(self):
        """Show year calendar popup for Excel month+year selection"""
        try:
            current_year = int(self.excel_year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for PDF year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for PDF month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_excel(self, year):
        """Handle year selection from popup for Excel year export"""
        self.excel_year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def on_year_selected_for_month_excel(self, year):
        """Handle year selection from popup for Excel month+year export"""
        self.excel_year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def populate_months(self):
        """Populate months combo box for PDF"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)

    def populate_months_excel(self):
        """Populate months combo box for Excel"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.excel_month_combo.addItems(months)
        self.excel_month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_tab_changed(self, index):
        """Handle tab changes"""
        self._sync_format_card_state(index)
        if index == 0:  # PDF tab
            self.export_type = "pdf"
            self._active_export_btn = getattr(self, "pdf_export_btn", None)
            self.update_preview()
        elif index == 1:  # Excel tab
            self.export_type = "excel"
            self._active_export_btn = getattr(self, "excel_export_btn", None)
            self.update_excel_preview()
    
    def on_range_changed(self, range_type):
        """Handle export range changes for PDF"""
        self.export_range = range_type
        month_visible = (range_type == "month")
        year_visible  = (range_type == "year")
        date_visible  = (range_type == "date_range")
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        if hasattr(self, 'date_range_wrap'):
            self.date_range_wrap.setVisible(date_visible)
        self.date_selection_container.setVisible(month_visible or year_visible or date_visible)
        self.update_preview()

    def on_excel_range_changed(self, range_type):
        """Handle export range changes for Excel"""
        self.excel_export_range = range_type
        month_visible = (range_type == "month")
        year_visible  = (range_type == "year")
        date_visible  = (range_type == "date_range")
        self.excel_month_group.setVisible(month_visible)
        self.excel_year_group.setVisible(year_visible)
        if hasattr(self, 'excel_date_range_wrap'):
            self.excel_date_range_wrap.setVisible(date_visible)
        self.excel_date_selection_container.setVisible(month_visible or year_visible or date_visible)
        self.update_excel_preview()
    
    def update_preview(self):
        """Update the PDF preview text with correct order info"""
        self._refresh_export_filter_options("pdf")
        if self.export_range == "all":
            self.preview_label.setText("Will export ALL quote forms as PDF")
        elif self.export_range == "date_range":
            fd = self.from_date.date().toString("MM/dd/yyyy")
            td = self.to_date.date().toString("MM/dd/yyyy")
            self.preview_label.setText(f"Will export quote forms from {fd} to {td} as PDF")
        elif self.export_range == "month":
            self.preview_label.setText(
                f"Will export quote forms for {self.month_combo.currentText()} {self.year_edit_month.text()} as PDF")
        elif self.export_range == "year":
            self.preview_label.setText(
                f"Will export quote forms for the year {self.year_edit.text()} as PDF")
        # Update summary card scope label
        if hasattr(self, '_pdf_scope_lbl'):
            self._pdf_scope_lbl.setText(self._get_scope_label("pdf"))
        if hasattr(self, '_pdf_rec_lbl'):
            self._pdf_rec_lbl.setText(f"{self._get_total_count()} Quotes")
        self._refresh_reference_summary("pdf")

    def update_excel_preview(self):
        """Update the Excel preview text with correct order info"""
        self._refresh_export_filter_options("excel")
        range_type = getattr(self, 'excel_export_range', 'all')

        if range_type == "all":
            self.excel_preview_label.setText("Will export ALL quote forms as Excel")
        elif range_type == "date_range":
            fd = self.excel_from_date.date().toString("MM/dd/yyyy")
            td = self.excel_to_date.date().toString("MM/dd/yyyy")
            self.excel_preview_label.setText(f"Will export quote forms from {fd} to {td} as Excel")
        elif range_type == "month":
            self.excel_preview_label.setText(
                f"Will export quote forms for {self.excel_month_combo.currentText()} {self.excel_year_edit_month.text()} as Excel")
        elif range_type == "year":
            self.excel_preview_label.setText(
                f"Will export quote forms for the year {self.excel_year_edit.text()} as Excel")
        # Update summary card scope label
        if hasattr(self, '_excel_scope_lbl'):
            self._excel_scope_lbl.setText(self._get_scope_label("excel"))
        if hasattr(self, '_excel_rec_lbl'):
            self._excel_rec_lbl.setText(f"{self._get_total_count()} Quotes")
        self._refresh_reference_summary("excel")

    def _refresh_reference_summary(self, export_type):
        prefix = "excel_" if export_type == "excel" else ""
        records = self._export_filtered_records(export_type)
        filtered_count = len(records)
        clients = sorted({str(job.get("client", "")).strip() for job in records if str(job.get("client", "")).strip()})
        scope = self._get_scope_label(export_type)
        filters = self._export_filter_values(export_type)
        active_filters = []
        if filters["status"] != "All Status":
            active_filters.append(filters["status"])
        if filters["client"] != "All Clients":
            active_filters.append(filters["client"])
        if active_filters:
            scope = f"{scope} | " + " | ".join(active_filters)

        filtered_lbl = getattr(self, f"_{export_type}_filtered_summary_lbl", None)
        clients_lbl = getattr(self, f"_{export_type}_clients_summary_lbl", None)
        scope_lbl = getattr(self, f"_{export_type}_scope_summary_lbl", None)
        info_lbl = getattr(self, f"_{export_type}_info_summary_lbl", None)
        if filtered_lbl:
            filtered_lbl.setText(str(filtered_count))
        if clients_lbl:
            clients_lbl.setText(str(len(clients)) if clients else "—")
        if scope_lbl:
            scope_lbl.setText(scope)
        if info_lbl:
            info_lbl.setText(
                f"ⓘ  You are about to export {filtered_count} quote form(s) in {export_type.upper()} format."
            )
        
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_type == "pdf":
            filters = self._export_filter_values("pdf")
            if self.export_range == "all":
                return {"range": "all", "type": "pdf", **filters}
            
            elif self.export_range == "date_range":
                from_date = self.from_date.date().toPyDate()
                to_date = self.to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "pdf", **filters}
            
            elif self.export_range == "month":
                month = self.month_combo.currentIndex() + 1
                year = int(self.year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "pdf", **filters}
            
            elif self.export_range == "year":
                year = int(self.year_edit.text())
                return {"range": "year", "year": year, "type": "pdf", **filters}
        
        elif self.export_type == "excel":
            filters = self._export_filter_values("excel")
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            
            if range_type == "all":
                return {"range": "all", "type": "excel", **filters}
            
            elif range_type == "date_range":
                from_date = self.excel_from_date.date().toPyDate()
                to_date = self.excel_to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "excel", **filters}
            
            elif range_type == "month":
                month = self.excel_month_combo.currentIndex() + 1
                year = int(self.excel_year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "excel", **filters}
            
            elif range_type == "year":
                year = int(self.excel_year_edit.text())
                return {"range": "year", "year": year, "type": "excel", **filters}
    
    def start_export(self):
        """Start the export process based on selected type"""
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return

        self._export_in_progress = True
        active_btn = getattr(self, '_active_export_btn', None)

        try:
            if active_btn:
                active_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)

            export_params = self.get_export_parameters()

            for i in range(101):
                if not hasattr(self, '_export_in_progress'):
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)

            self._export_params = export_params
            self.accept()

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            if active_btn:
                active_btn.setEnabled(True)
            self._export_in_progress = False
# Update the JobFormDialog to ensure proper Firebase integration
class JobFormDialog(QtWidgets.QDialog):
    """Quote Form Creation Dialog - Enhanced Layout and Smart Behavior"""

    def __init__(self, main_window, parent=None, job_data=None, firebase_available=False):
        super().__init__(parent)
        self.allow_email_autofill = False  # Initialize as False
        self.main_window = main_window
        self.owner_tab = parent
        self.job_data = job_data  # For editing existing job
        self.is_editing = job_data is not None
        self.FIREBASE_AVAILABLE = firebase_available
        self.auto_generate_enabled = not self.is_editing  # Auto-generate only for new jobs
        self._last_expedite_btn = None  # Initialize last selected expedite button
        
        # Initialize client intelligence and template manager
        self.client_intelligence = ClientIntelligence()
        self.template_manager = TemplateManager()
        self.template_manager.initialize_default_templates()
        
        title = "Edit Quote Form" if self.is_editing else "Generate Professional Quote Form PDF"
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(950, 850)
        self.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        self.init_ui()
        
        # ⭐⭐ CRITICAL FIX: Enable email autofill ONLY after form is fully populated
        if self.is_editing:
            # After populating form, enable autofill for future client changes
            QtCore.QTimer.singleShot(100, self.enable_email_autofill)
        
        # Set initial focus
        if not self.is_editing:
            QtCore.QTimer.singleShot(50, lambda: self.project_site_edit.setFocus())

    def init_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        self.setStyleSheet("QDialog { background: #F8FAFC; }")

        # ── Hidden fields only used internally (not user-facing) ──
        _hidden = QtWidgets.QWidget()
        _hidden.setVisible(False)
        _hidden.setMaximumHeight(0)
        _hl = QtWidgets.QVBoxLayout(_hidden)
        _hl.setContentsMargins(0, 0, 0, 0)
        _hl.setSpacing(0)
        self.project_name_edit = QtWidgets.QLineEdit()
        self.timeline_info = QtWidgets.QLabel()
        for _w in [self.project_name_edit, self.timeline_info]:
            _hl.addWidget(_w)
        main_layout.addWidget(_hidden)

        # ── Shared styles ─────────────────────────────────────────
        _INP = """
            QLineEdit {
                background:white; border:1.5px solid #E2E8F0; border-radius:8px;
                padding:10px 14px; font-size:13px; color:#374151;
                font-family:'Inter','Segoe UI';
            }
            QLineEdit:focus { border-color:#0F766E; }
        """
        _COMBO = """
            QComboBox {
                background:white; border:1.5px solid #E2E8F0; border-radius:8px;
                padding:9px 14px; font-size:13px; color:#374151;
                font-family:'Inter','Segoe UI';
            }
            QComboBox:focus { border-color:#0F766E; }
            QComboBox::drop-down { border:none; width:28px; }
            QComboBox QAbstractItemView {
                border:1px solid #E2E8F0; border-radius:6px;
                selection-background-color:#0F766E;
            }
        """
        _LBL = (
            "font-size:12px; font-weight:700; color:#374151;"
            " background:transparent; border:none; font-family:'Inter','Segoe UI';"
        )
        _CARD = "QFrame { background:white; border:1px solid #E2E8F0; border-radius:10px; }"
        _CHK = """
            QCheckBox {
                spacing:8px; font-size:13px; color:#374151;
                font-family:'Inter','Segoe UI'; padding:4px 0; background:transparent;
            }
            QCheckBox::indicator {
                width:18px; height:18px; border:2px solid #CBD5E1;
                border-radius:4px; background:white;
            }
            QCheckBox::indicator:checked { background:#0F766E; border-color:#0F766E; }
            QCheckBox::indicator:hover { border-color:#0F766E; }
        """

        def _field(lbl_text, widget, hint=None, required=False):
            w = QtWidgets.QWidget()
            w.setStyleSheet("background:transparent;")
            vl = QtWidgets.QVBoxLayout(w)
            vl.setContentsMargins(0, 0, 0, 0)
            vl.setSpacing(5)
            lbl = QtWidgets.QLabel(lbl_text + (" *" if required else ""))
            lbl.setStyleSheet(_LBL)
            vl.addWidget(lbl)
            vl.addWidget(widget)
            if hint:
                h = QtWidgets.QLabel(hint)
                h.setStyleSheet(
                    "font-size:11px; color:#9CA3AF; background:transparent;"
                    " border:none; font-family:'Inter','Segoe UI';"
                )
                vl.addWidget(h)
            return w

        def _section(icon, title, subtitle):
            card = QtWidgets.QFrame()
            card.setStyleSheet(_CARD)
            cl = QtWidgets.QVBoxLayout(card)
            cl.setContentsMargins(24, 20, 24, 20)
            cl.setSpacing(16)
            hdr = QtWidgets.QHBoxLayout()
            hdr.setSpacing(10)
            ic = QtWidgets.QLabel()
            ic.setFixedSize(30, 30)
            ic.setAlignment(QtCore.Qt.AlignCenter)
            ic.setStyleSheet("background:transparent; border:none;")
            if isinstance(icon, QtGui.QIcon):
                ic.setPixmap(icon.pixmap(24, 24))
            else:
                ic.setText(icon)
                ic.setStyleSheet(
                    "font-size:22px; background:transparent; border:none; color:#0F766E;"
                )
            hc = QtWidgets.QVBoxLayout()
            hc.setSpacing(2)
            tl = QtWidgets.QLabel(title)
            tl.setStyleSheet(
                "font-size:20px; font-weight:900; color:#111827;"
                " background:transparent; border:none; font-family:'Inter','Segoe UI';"
            )
            sl = QtWidgets.QLabel(subtitle)
            sl.setStyleSheet(
                "font-size:12px; color:#6B7280; background:transparent;"
                " border:none; font-family:'Inter','Segoe UI';"
            )
            hc.addWidget(tl)
            hc.addWidget(sl)
            hdr.addWidget(ic)
            hdr.addLayout(hc)
            hdr.addStretch()
            cl.addLayout(hdr)
            div = QtWidgets.QFrame()
            div.setFrameShape(QtWidgets.QFrame.HLine)
            div.setStyleSheet("background:#E2E8F0; border:none; max-height:1px;")
            cl.addWidget(div)
            return card, cl

        # ── Scrollable content ────────────────────────────────────
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet(
            "QScrollArea { background:#F8FAFC; border:none; }"
            "QScrollArea > QWidget > QWidget { background:#F8FAFC; }"
        )
        content = QtWidgets.QWidget()
        content.setStyleSheet("background:#F8FAFC;")
        cly = QtWidgets.QVBoxLayout(content)
        cly.setContentsMargins(32, 24, 32, 24)
        cly.setSpacing(20)

        # ════════ PROJECT DETAILS ════════════════════════════════
        pd_card, pd = _section("📋", "Project Details", "Basic information about the quote")

        # Row 1: Quote Number | Client / Company
        r1 = QtWidgets.QHBoxLayout()
        r1.setSpacing(16)
        self.job_number_edit = QtWidgets.QLineEdit()
        self.job_number_edit.setReadOnly(True)
        self.job_number_edit.setPlaceholderText("Auto-generated")
        self.job_number_edit.setMinimumHeight(42)
        self.job_number_edit.setStyleSheet(
            "QLineEdit {"
            "  background:#EFF6FF; border:2px solid #2563EB; border-radius:8px;"
            "  padding:10px 14px; font-size:14px; color:#2563EB;"
            "  font-family:'Inter','Segoe UI'; font-weight:800;"
            "}"
        )
        self.client_combo = QtWidgets.QComboBox()
        self.client_combo.setEditable(True)
        self.client_combo.setMinimumHeight(42)
        self.client_combo.setStyleSheet(_COMBO)
        self.client_combo.lineEdit().setPlaceholderText("Select or type company name")
        r1.addWidget(_field("Quote Number", self.job_number_edit), 1)
        r1.addWidget(_field("Client / Company", self.client_combo, required=True), 1)
        pd.addLayout(r1)

        # Row 2: Project Site | Plant
        r2 = QtWidgets.QHBoxLayout()
        r2.setSpacing(16)
        self.project_site_edit = QtWidgets.QLineEdit()
        self.project_site_edit.setPlaceholderText("Enter complete project site address")
        self.project_site_edit.setMinimumHeight(42)
        self.project_site_edit.setStyleSheet(_INP)
        self.plant_edit = QtWidgets.QLineEdit()
        self.plant_edit.setPlaceholderText("Plant / facility name")
        self.plant_edit.setMinimumHeight(42)
        self.plant_edit.setStyleSheet(_INP)
        r2.addWidget(_field("Project Site", self.project_site_edit, required=True))
        r2.addWidget(_field("Plant", self.plant_edit))
        pd.addLayout(r2)

        # Row 3: Client Email | Client Address
        r3 = QtWidgets.QHBoxLayout()
        r3.setSpacing(16)
        self.client_email_edit = QtWidgets.QLineEdit()
        self.client_email_edit.setPlaceholderText("client@company.com")
        self.client_email_edit.setMinimumHeight(42)
        self.client_email_edit.setStyleSheet(_INP)
        self.client_address_edit = QtWidgets.QLineEdit()
        self.client_address_edit.setPlaceholderText("Mailing / billing address")
        self.client_address_edit.setMinimumHeight(42)
        self.client_address_edit.setStyleSheet(_INP)
        r3.addWidget(_field("Client Email", self.client_email_edit))
        r3.addWidget(_field("Client Address", self.client_address_edit))
        pd.addLayout(r3)

        # Row 4: Job Type | Sales Person
        r4 = QtWidgets.QHBoxLayout()
        r4.setSpacing(16)
        self.job_type_combo = QtWidgets.QComboBox()
        self.job_type_combo.addItems([
            "Engineering Design", "Construction", "Inspection", "Consultation",
            "Maintenance", "Research", "Drafting", "Peer Review", "Site Visit",
            "Report Preparation", "Permit Drawings", "Other",
        ])
        self.job_type_combo.setEditable(True)
        self.job_type_combo.setCurrentIndex(-1)
        self.job_type_combo.lineEdit().setPlaceholderText("Select or type job type")
        self.job_type_combo.setMinimumHeight(42)
        self.job_type_combo.setStyleSheet(_COMBO)
        self.sales_combo = QtWidgets.QComboBox()
        self.sales_combo.setEditable(True)
        self.sales_combo.setMinimumHeight(42)
        self.sales_combo.setStyleSheet(_COMBO)
        self.sales_combo.lineEdit().setPlaceholderText("Enter or select sales person")
        r4.addWidget(_field("Job Type", self.job_type_combo))
        r4.addWidget(_field("Sales Person", self.sales_combo))
        pd.addLayout(r4)

        # Row 5: Scope of Work (full width)
        self.scope_of_work_edit = QtWidgets.QLineEdit()
        self.scope_of_work_edit.setPlaceholderText("Enter detailed scope of work (comma-separated)...")
        self.scope_of_work_edit.setMinimumHeight(42)
        self.scope_of_work_edit.setStyleSheet(_INP)
        pd.addWidget(_field("Scope of Work", self.scope_of_work_edit))

        # Row 6: Agreed Cost | Expedite? | Extra %
        r6 = QtWidgets.QHBoxLayout()
        r6.setSpacing(16)

        cost_w = QtWidgets.QWidget()
        cost_w.setStyleSheet("background:transparent;")
        cost_vl = QtWidgets.QVBoxLayout(cost_w)
        cost_vl.setContentsMargins(0, 0, 0, 0)
        cost_vl.setSpacing(5)
        cost_lbl = QtWidgets.QLabel("Agreed Cost")
        cost_lbl.setStyleSheet(_LBL)
        cost_vl.addWidget(cost_lbl)
        self.engineering_costs_edit = QtWidgets.QLineEdit()
        self.engineering_costs_edit.setPlaceholderText("$0.00")
        self.engineering_costs_edit.setMinimumHeight(42)
        self.engineering_costs_edit.setStyleSheet(
            "QLineEdit {"
            "  background:white; border:1.5px solid #E2E8F0; border-radius:8px;"
            "  padding:10px 14px; font-size:13px; color:#374151;"
            "  font-family:'Inter','Segoe UI';"
            "}"
            "QLineEdit:focus { border-color:#0F766E; }"
        )
        cost_vl.addWidget(self.engineering_costs_edit)
        r6.addWidget(cost_w, 2)

        exp_w = QtWidgets.QWidget()
        exp_w.setStyleSheet("background:transparent;")
        exp_vl = QtWidgets.QVBoxLayout(exp_w)
        exp_vl.setContentsMargins(0, 0, 0, 0)
        exp_vl.setSpacing(5)
        exp_lbl = QtWidgets.QLabel("Expedite?")
        exp_lbl.setStyleSheet(_LBL)
        exp_vl.addWidget(exp_lbl)
        exp_rr = QtWidgets.QHBoxLayout()
        exp_rr.setSpacing(16)
        _RDO = "font-size:13px; color:#374151; font-family:'Inter','Segoe UI'; spacing:6px;"
        self.expedite_yes = QtWidgets.QRadioButton("Yes (50% Extra)")
        self.expedite_yes.setStyleSheet(_RDO)
        self.expedite_no = QtWidgets.QRadioButton("No")
        self.expedite_no.setStyleSheet(_RDO)
        self.expedite_group = QtWidgets.QButtonGroup()
        self.expedite_group.addButton(self.expedite_yes)
        self.expedite_group.addButton(self.expedite_no)
        self.expedite_group.setExclusive(True)
        exp_rr.addWidget(self.expedite_yes)
        exp_rr.addWidget(self.expedite_no)
        exp_rr.addStretch()
        exp_vl.addLayout(exp_rr)
        r6.addWidget(exp_w, 2)

        self.expedite_amount_edit = QtWidgets.QLineEdit()
        self.expedite_amount_edit.setPlaceholderText("0%")
        self.expedite_amount_edit.setMinimumHeight(42)
        self.expedite_amount_edit.setStyleSheet(_INP)
        r6.addWidget(_field("Extra %", self.expedite_amount_edit), 1)
        pd.addLayout(r6)

        # Total price summary (shown when expedite Yes is selected)
        self.total_price_lbl = QtWidgets.QLabel("")
        self.total_price_lbl.setStyleSheet(
            "font-size:12px; font-weight:700; color:#0F766E; background:#F0FDF4;"
            " border:1px solid #BBF7D0; border-radius:6px; padding:6px 14px;"
            " font-family:'Inter','Segoe UI';"
        )
        self.total_price_lbl.setVisible(False)
        pd.addWidget(self.total_price_lbl)

        # Row 7: Priority
        self.priority_combo = QtWidgets.QComboBox()
        self.priority_combo.addItems(["Low", "Medium", "High", "Urgent"])
        self.priority_combo.setCurrentIndex(0)
        self.priority_combo.setMinimumHeight(42)
        self.priority_combo.setStyleSheet(_COMBO)
        pd.addWidget(_field("Priority", self.priority_combo))
        cly.addWidget(pd_card)

        # ════════ TIMELINE ════════════════════════════════════════
        tl_card, tl = _section(
            JobFormsExportDialog._make_calendar_icon("#0F766E", 30),
            "Timeline",
            "Important dates for this quote",
        )
        tl_row = QtWidgets.QHBoxLayout()
        tl_row.setSpacing(16)

        _DATE_EDIT_SS = f"""
            QDateEdit {{
                background:white;
                border:1.5px solid #E2E8F0;
                border-radius:8px;
                font-size:13px;
                color:#374151;
                font-family:'Inter','Segoe UI';
                padding:0 34px 0 12px;
            }}
            QDateEdit:focus {{
                border-color:#0F766E;
            }}
            QDateEdit::drop-down {{
                border:none;
                width:28px;
                background:transparent;
            }}
            QDateEdit::down-arrow {{
                image:url("{CALENDAR_URL}");
                width:15px;
                height:15px;
                margin-right:7px;
            }}
            QDateEdit::up-button {{ width:0; }}
            QDateEdit::down-button {{ width:0; }}
        """

        def _date_widget(date_edit):
            date_edit.setStyleSheet(_DATE_EDIT_SS)
            date_edit.setFrame(False)
            date_edit.setMinimumHeight(42)
            date_edit.setMinimumWidth(168)
            date_edit.setMaximumWidth(190)
            return date_edit

        self.start_date_edit = _NoScrollDateEdit(QtCore.QDate.currentDate())
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.start_date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        _de_le = self.start_date_edit.lineEdit()
        if _de_le:
            _de_le.setPlaceholderText("MM-DD-YY")

        self.due_date_edit = _NoScrollDateEdit(QtCore.QDate.currentDate().addDays(30))
        self.due_date_edit.setCalendarPopup(True)
        self.due_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.due_date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        _de_le2 = self.due_date_edit.lineEdit()
        if _de_le2:
            _de_le2.setPlaceholderText("MM-DD-YY")

        self.deliverables_edit = QtWidgets.QLineEdit()
        self.deliverables_edit.setPlaceholderText("Enter deliverables like 1, 2, 3...")
        self.deliverables_edit.setMinimumHeight(42)
        self.deliverables_edit.setStyleSheet(_INP)
        self.deliverables_edit.installEventFilter(self)


        tl_row.setAlignment(QtCore.Qt.AlignTop)
        _sd_w = _field("Start Date", _date_widget(self.start_date_edit), required=True)
        _dd_w = _field("Due Date", _date_widget(self.due_date_edit), required=True)
        _del_w = _field(
            "Deliverables",
            self.deliverables_edit,
            hint="Separate multiple deliverables with commas",
        )
        tl_row.addWidget(_sd_w, 1, QtCore.Qt.AlignTop)
        tl_row.addWidget(_dd_w, 1, QtCore.Qt.AlignTop)
        tl_row.addWidget(_del_w, 2, QtCore.Qt.AlignTop)
        tl.addLayout(tl_row)

        # Visible timeline summary strip
        self.timeline_info_visible = QtWidgets.QLabel()
        self.timeline_info_visible.setStyleSheet(
            "QLabel {"
            "  background:#F0FDF4; border:1px solid #BBF7D0; border-radius:7px;"
            "  padding:7px 14px; font-size:12px; font-weight:700; color:#065F46;"
            "  font-family:'Inter','Segoe UI';"
            "}"
        )
        self.timeline_info_visible.setVisible(False)
        tl.addWidget(self.timeline_info_visible)
        cly.addWidget(tl_card)
        # ════════ SERVICES ════════════════════════════════════════
        sv_card, sv = _section("🔧", "Services", "Select applicable service categories")

        def _chk(label):
            c = QtWidgets.QCheckBox(label)
            c.setStyleSheet(_CHK)
            c.setFocusPolicy(QtCore.Qt.StrongFocus)
            return c

        self.structural_checkbox = _chk("Structural")
        self.civil_checkbox = _chk("Civil")
        self.electrical_checkbox = _chk("Electrical")
        self.mechanical_checkbox = _chk("Mechanical")
        self.plumbing_checkbox = _chk("Plumbing Design")
        self.anchor_calc_checkbox = _chk("Anchor Calculations")
        self.solidworks_checkbox = _chk("Solid Works")
        self.foundation_checkbox = _chk("Foundation")
        self.other_checkbox = _chk("Others")

        self.others_input = QtWidgets.QLineEdit()
        self.others_input.setPlaceholderText("Type other services (e.g., 1, 2, 3...)")
        self.others_input.setStyleSheet(_INP)
        self.others_input.setVisible(False)
        self.other_checkbox.stateChanged.connect(self.toggle_others_input)

        chk_grid = QtWidgets.QGridLayout()
        chk_grid.setVerticalSpacing(12)
        chk_grid.setHorizontalSpacing(20)
        for _i, _c in enumerate([
            self.structural_checkbox, self.civil_checkbox, self.electrical_checkbox,
            self.mechanical_checkbox, self.plumbing_checkbox, self.anchor_calc_checkbox,
            self.solidworks_checkbox, self.foundation_checkbox, self.other_checkbox,
        ]):
            chk_grid.addWidget(_c, _i // 3, _i % 3)
        chk_grid.addWidget(self.others_input, 3, 0, 1, 3)
        sv.addLayout(chk_grid)
        cly.addWidget(sv_card)

        scroll.setWidget(content)
        main_layout.addWidget(scroll, 1)

        # ── Footer bar ────────────────────────────────────────────
        footer = QtWidgets.QFrame()
        footer.setFixedHeight(68)
        footer.setStyleSheet("QFrame { background:white; border-top:1px solid #E2E8F0; }")
        fly = QtWidgets.QHBoxLayout(footer)
        fly.setContentsMargins(32, 12, 32, 12)
        fly.setSpacing(12)
        _GHOST = """
            QPushButton {
                background:white; color:#374151; border:1.5px solid #E2E8F0;
                border-radius:8px; font-size:13px; font-weight:700; padding:0 24px;
                font-family:'Inter','Segoe UI';
            }
            QPushButton:hover { background:#F8FAFC; }
        """
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedHeight(42)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet(_GHOST)
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setAutoDefault(True)

        _draft_btn = QtWidgets.QPushButton("Save Draft")
        _draft_btn.setFixedHeight(42)
        _draft_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        _draft_btn.setStyleSheet(_GHOST)
        _draft_btn.clicked.connect(self.save_as_draft)

        self.create_btn = QtWidgets.QPushButton("Create Quote  →")
        self.create_btn.setFixedHeight(42)
        self.create_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.create_btn.setStyleSheet("""
            QPushButton {
                background:#1E293B; color:white; border:none; border-radius:8px;
                font-size:13px; font-weight:800; padding:0 28px;
                font-family:'Inter','Segoe UI';
            }
            QPushButton:hover { background:#0F172A; }
        """)
        self.create_btn.clicked.connect(self.create_job_form)
        self.create_btn.setAutoDefault(True)
        self.create_btn.setDefault(False)

        fly.addStretch()
        fly.addWidget(self.cancel_btn)
        fly.addWidget(_draft_btn)
        fly.addWidget(self.create_btn)
        main_layout.addWidget(footer)

        # ── Signals ───────────────────────────────────────────────
        self.client_combo.currentTextChanged.connect(self.on_client_selected)
        self.engineering_costs_edit.textChanged.connect(self.validate_cost_input)
        self.expedite_yes.clicked.connect(self.on_expedite_clicked)
        self.expedite_no.clicked.connect(self.on_expedite_clicked)
        self.expedite_amount_edit.editingFinished.connect(self.update_expedite_amount)
        self.engineering_costs_edit.textChanged.connect(self.update_expedite_amount)
        self.expedite_yes.toggled.connect(self.update_expedite_amount)
        self.expedite_no.toggled.connect(self.update_expedite_amount)
        self.start_date_edit.dateChanged.connect(self.update_timeline_info)
        self.due_date_edit.dateChanged.connect(self.update_timeline_info)
        self.due_date_edit.dateChanged.connect(self._on_due_date_changed)
        self._priority_auto_set = False
        self.priority_combo.currentIndexChanged.connect(self._on_priority_manually_changed)
        self.update_timeline_info()

        if not self.is_editing:
            self.project_site_edit.editingFinished.connect(self.auto_generate_job_number)
            self.start_date_edit.dateChanged.connect(self.auto_generate_job_number)
            self.project_site_edit.textChanged.connect(
                lambda t: self.project_name_edit.setText(t)
            )

        # ── Data load ─────────────────────────────────────────────
        self.load_saved_clients()
        self.load_sales_persons()

        if self.is_editing:
            self.populate_form_data()
        else:
            self.auto_generate_job_number()
        self.allow_email_autofill = True
        self._enter_on_last_field = False
        self.last_logical_field = self.deliverables_edit

        # ── Enter navigation ──────────────────────────────────────
        self.setup_enter_key_navigation()

        # ── Ctrl+S ────────────────────────────────────────────────
        save_shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+S"), self)
        save_shortcut.activated.connect(self.create_job_form)

    def populate_form_data(self):
        """Populate form with existing job data for editing"""
        if not self.job_data:
            return
        
        # ⭐ TEMPORARILY disable autofill during initial population
        self.allow_email_autofill = False
        
        self.job_number_edit.setText(self.job_data.get('job_number', ''))
        
        # ==== CHANGED: Populate Project Name instead of Job Title ====
        self.project_name_edit.setText(self.job_data.get('project_name', ''))  # Changed from job_title
        
        client = self.job_data.get('client', '')
        index = self.client_combo.findText(client)
        if index >= 0:
            self.client_combo.setCurrentIndex(index)
        else:
            self.client_combo.setEditText(client)
        
        # Populate email and address if available
        self.client_email_edit.setText(self.job_data.get('client_email', ''))
        self.client_address_edit.setText(self.job_data.get('client_address', ''))
        
        # ==== REMOVED: Duplicate project_name_edit population ====
        # self.project_name_edit.setText(self.job_data.get('project_name', ''))
        
        self.plant_edit.setText(self.job_data.get('plant', ''))

        self.project_site_edit.setText(self.job_data.get('project_site_address', ''))
        sales_value = self.job_data.get('sales', '') or self.job_data.get('sales_person', '')

        index = self.sales_combo.findText(sales_value)
        if index >= 0:
            self.sales_combo.setCurrentIndex(index)
        else:
            self.sales_combo.setEditText(sales_value)

        self.scope_of_work_edit.setText(self.job_data.get('scope_of_work', ''))
        self.engineering_costs_edit.setText(self.job_data.get('engineering_costs', ''))
        
        job_type = self.job_data.get('job_type', '')
        index = self.job_type_combo.findText(job_type)
        if index >= 0:
            self.job_type_combo.setCurrentIndex(index)
        else:
            self.job_type_combo.setEditText(job_type)

        priority = self.job_data.get('status', '')
        index = self.priority_combo.findText(priority)
        if index >= 0:
            self.priority_combo.setCurrentIndex(index)
                    
        # Set dates — try multiple formats for backwards-compatibility
        def _parse_qdate(s):
            for fmt in ("MM-dd-yyyy", "MM/dd/yyyy", "yyyy-MM-dd", "M-d-yyyy", "M/d/yyyy"):
                d = QtCore.QDate.fromString(s, fmt)
                if d.isValid():
                    return d
            return QtCore.QDate()
        sd = _parse_qdate(self.job_data.get("start_date", ""))
        if sd.isValid():
            self.start_date_edit.setDate(sd)
        dd = _parse_qdate(self.job_data.get("due_date", ""))
        if dd.isValid():
            self.due_date_edit.setDate(dd)
        # Set deliverables — handle Firebase dict-vs-list quirk
        raw_del = self.job_data.get("deliverables", [])
        if isinstance(raw_del, dict):
            raw_del = [raw_del[k] for k in sorted(raw_del.keys(), key=lambda x: int(x) if x.isdigit() else x)]
        self.deliverables_edit.setText(", ".join(str(d) for d in raw_del if d))


        # Set services
        services = self.job_data.get('services', [])
        service_widgets = {
            'Structural': self.structural_checkbox,
            'Civil': self.civil_checkbox,
            'Electrical': self.electrical_checkbox,
            'Mechanical': self.mechanical_checkbox,
            'Plumbing Design': self.plumbing_checkbox,
            'Anchor Calculations': self.anchor_calc_checkbox,
            'Solid Works': self.solidworks_checkbox,
            'Foundation': self.foundation_checkbox
        }
        
        # Check if "Others:" is in services
        others_text = ""
        for service in services:
            if service.startswith("Others:"):
                others_text = service.replace("Others:", "").strip()
                if others_text:
                    self.other_checkbox.setChecked(True)
            elif service in service_widgets:
                service_widgets[service].setChecked(True)
        
        # Set the "Others" input text
        if others_text:
            self.others_input.setText(others_text)
            self.others_input.setVisible(True)
                
        for service in services:
            if service in service_widgets:
                service_widgets[service].setChecked(True)
        
        # ⭐⭐ IMPORTANT: Disable auto-generation for editing mode
        self.auto_generate_enabled = False
        # Set expedite data if it exists
        # --- FIX: Do NOT auto-select Yes/No unless user explicitly selected before ---
        expedite = self.job_data.get('expedite', None)
        expedite_amount = self.job_data.get('expedite_amount', '')

        # Reset everything first
        self.expedite_yes.setChecked(False)
        self.expedite_no.setChecked(False)
        self._last_expedite_btn = None
        self.expedite_amount_edit.clear()
        self.expedite_amount_edit.setPlaceholderText("Select Yes or No for expedite")

        # Apply ONLY if user had explicitly selected
        if expedite is True:
            self.expedite_yes.setChecked(True)
            self._last_expedite_btn = self.expedite_yes
            if expedite_amount:
                self.expedite_amount_edit.setText(expedite_amount)

        elif expedite is False and expedite_amount:
            # User explicitly chose NO earlier
            self.expedite_no.setChecked(True)
            self._last_expedite_btn = self.expedite_no
            self.expedite_amount_edit.setText(expedite_amount)

        # If no expedite data, ensure no button is selected
        if not expedite and not expedite_amount:
            self.expedite_yes.setChecked(False)
            self.expedite_no.setChecked(False)
            self._last_expedite_btn = None
            self.expedite_amount_edit.clear()
            self.expedite_amount_edit.setPlaceholderText("Select Yes or No for expedite")
    
            
    def setup_enter_key_navigation(self):
        """Set up Enter key to navigate between fields instead of submitting form"""
        # List of all input widgets in tab order
        self.input_widgets = []
        
        # Collect all widgets
        widgets_to_add = [
            self.project_name_edit,      # Project Name (hidden, skipped by navigator)
            self.client_combo,           # 1 - client/company
            self.project_site_edit,      # 2 - project site
            self.plant_edit,             # 3 - plant
            self.client_email_edit,      # 4 - client email
            self.client_address_edit,    # 5 - client address
            self.job_type_combo,         # 6 - job type
            self.sales_combo,            # 7 - sales
            self.scope_of_work_edit,     # 8 - scope of work
            self.engineering_costs_edit, # 9 - agreed cost
            self.priority_combo,         # 10 - priority
            self.start_date_edit,
            self.due_date_edit,
            self.deliverables_edit,
            self.structural_checkbox,
            self.civil_checkbox,
            self.electrical_checkbox,
            self.mechanical_checkbox,
            self.plumbing_checkbox,
            self.anchor_calc_checkbox,
            self.solidworks_checkbox,
            self.foundation_checkbox,
            self.other_checkbox,
            self.others_input,
            self.create_btn,
            self.cancel_btn
        ]
        
        # Only add widgets that exist
        for widget in widgets_to_add:
            if widget and widget.isWidgetType():
                self.input_widgets.append(widget)
        
        # Install event filters for all widgets
        for widget in self.input_widgets:
            if widget:
                widget.installEventFilter(self)
                
                # Special handling for date edit widgets
                if isinstance(widget, QtWidgets.QDateEdit):
                    # Also install event filter on the line edit portion
                    line_edit = widget.lineEdit()
                    if line_edit:
                        line_edit.installEventFilter(self)
        
        # Set tab order only for widgets in the same window
        for i in range(len(self.input_widgets) - 1):
            if (self.input_widgets[i] and 
                self.input_widgets[i + 1] and 
                self.input_widgets[i].window() == self.input_widgets[i + 1].window()):
                QtWidgets.QWidget.setTabOrder(self.input_widgets[i], self.input_widgets[i + 1])
        
        # Disable default button behavior for the create button
        self.create_btn.setAutoDefault(False)
        self.create_btn.setDefault(False)
        self.cancel_btn.setAutoDefault(False)
        self.cancel_btn.setDefault(False)
        
        # Setup date field UX
        self.setup_date_field_ux()
    
    def load_sales_persons(self):
        """Load sales persons into dropdown."""
        try:
            self.sales_combo.clear()
            sales_list = []
            
            if self.FIREBASE_AVAILABLE:
                from main import FirebaseManager
                sales_list = [
                    person.get("name", "")
                    for person in FirebaseManager.load_sales_people()
                    if person.get("name")
                ]
                _log.info("Loaded %s sales persons from Firebase", len(sales_list))
            else:
                sales_list = [
                    person.get("name", "")
                    for person in _load_local_sales_people()
                    if person.get("name")
                ]
                _log.info("Loaded %s sales persons from local data", len(sales_list))

            self.sales_combo.addItems(sorted(set(sales_list)))
            
            self.sales_combo.setCurrentIndex(-1)
            self.sales_combo.lineEdit().clear()
            
        except Exception as e:
            _log.error("Error loading sales persons: %s", e)
        
    def validate_cost_input(self):
        """Validate cost input to accept only numbers and automatically add $ prefix"""
        # Block signals to prevent recursive calls
        self.engineering_costs_edit.blockSignals(True)
        
        try:
            text = self.engineering_costs_edit.text().strip()
            cursor_pos = self.engineering_costs_edit.cursorPosition()
            
            # If text is empty, just return
            if not text:
                self.engineering_costs_edit.blockSignals(False)
                return
            
            # Remove any non-numeric characters except decimal point
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            
            # Remove any existing $ signs
            cleaned = cleaned.replace('$', '')
            
            # Ensure only one decimal point
            if cleaned.count('.') > 1:
                # Remove extra decimal points
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])
            
            # Add $ prefix if we have any valid number
            if cleaned:
                # Ensure it starts with $
                if not text.startswith('$'):
                    final_text = f"${cleaned}"
                else:
                    # Keep $ and clean the rest
                    final_text = f"${cleaned}"
            else:
                final_text = ""
            
            # Update the field if it changed
            if final_text != text:
                self.engineering_costs_edit.setText(final_text)
                
                # Adjust cursor position
                # If we added $ at the beginning, move cursor right by 1
                if not text.startswith('$') and final_text.startswith('$'):
                    new_pos = min(cursor_pos + 1, len(final_text))
                else:
                    new_pos = min(cursor_pos, len(final_text))
                
                self.engineering_costs_edit.setCursorPosition(new_pos)
                
            # Update expedite amount when cost changes
            self.update_expedite_amount()
                
        finally:
            self.engineering_costs_edit.blockSignals(False)
            
    def on_expedite_clicked(self, checked=None):
        """Handle expedite button clicks with toggle behavior"""
        sender = self.sender()  # Get which button was clicked
        
        # If clicking the already selected button, deselect it
        if checked and hasattr(self, '_last_expedite_btn') and self._last_expedite_btn == sender:
            # Temporarily make group non-exclusive to allow deselecting
            self.expedite_group.setExclusive(False)
            sender.setChecked(False)
            self.expedite_group.setExclusive(True)
            self._last_expedite_btn = None
            # Yes was toggled off — revert priority to Low
            if sender == self.expedite_yes and hasattr(self, 'priority_combo'):
                idx = self.priority_combo.findText("Low")
                if idx >= 0:
                    self.priority_combo.setCurrentIndex(idx)
        elif checked:
            if sender == self.expedite_yes:
                self.expedite_no.setChecked(False)
                # Auto-set priority to Urgent when expedite is Yes
                if hasattr(self, 'priority_combo'):
                    idx = self.priority_combo.findText("Urgent")
                    if idx >= 0:
                        self.priority_combo.setCurrentIndex(idx)
            else:
                # Expedite No selected — set priority to Low
                self.expedite_yes.setChecked(False)
                if hasattr(self, 'priority_combo'):
                    idx = self.priority_combo.findText("Low")
                    if idx >= 0:
                        self.priority_combo.setCurrentIndex(idx)

            self._last_expedite_btn = sender

        self.update_expedite_amount()
    def update_expedite_amount(self):
        """Update expedite amount based on selection and user input"""
        try:
            # Get base cost
            base_text = self.engineering_costs_edit.text().replace("$", "").replace(",", "").strip()
            base = float(base_text) if base_text else 0.0
            
            # Get current expedite amount text
            current_text = self.expedite_amount_edit.text().strip()
            
            # Check if YES is selected
            if self.expedite_yes.isChecked():
                # If field is empty or shows 0%, set to 50%
                if not current_text or current_text in ["0%", "$0.00", "0"]:
                    expedite_percent = 50
                    self.expedite_amount_edit.setText(f"{expedite_percent}%")
                    self.expedite_amount_edit.setCursorPosition(0)
                else:
                    # User has typed something - parse it
                    if "%" in current_text:
                        # Keep as percentage
                        pass
                    elif "$" in current_text:
                        # Convert dollar amount to percentage
                        dollar_amount = float(current_text.replace("$", "").replace(",", ""))
                        if base > 0:
                            percent = round((dollar_amount / base) * 100, 1)
                            self.expedite_amount_edit.setText(f"{percent}%")
            
            # Check if NO is selected
            elif self.expedite_no.isChecked():
                # If field is empty or shows 50%, set to 0%
                if not current_text or current_text in ["50%", "50"]:
                    self.expedite_amount_edit.setText("0%")
                    self.expedite_amount_edit.setCursorPosition(0)
                else:
                    # User has typed something - parse it
                    if "%" in current_text:
                        # Keep as percentage
                        pass
                    elif "$" in current_text:
                        # Convert dollar amount to percentage
                        dollar_amount = float(current_text.replace("$", "").replace(",", ""))
                        if base > 0:
                            percent = round((dollar_amount / base) * 100, 1)
                            self.expedite_amount_edit.setText(f"{percent}%")
            
            # If neither is selected (both unchecked)
            else:
                self.expedite_amount_edit.clear()
                self.expedite_amount_edit.setPlaceholderText("Select Yes or No for expedite")

            # Update total price label (visible summary strip)
            if hasattr(self, 'total_price_lbl'):
                if self.expedite_yes.isChecked() and base > 0:
                    total = base * 1.5
                    self.total_price_lbl.setText(
                        f"Total with 50% expedite:  ${base:,.2f} + ${base * 0.5:,.2f} = ${total:,.2f}"
                    )
                    self.total_price_lbl.setVisible(True)
                else:
                    self.total_price_lbl.setVisible(False)

        except Exception:
            if self.expedite_yes.isChecked():
                self.expedite_amount_edit.setText("50%")
            elif self.expedite_no.isChecked():
                self.expedite_amount_edit.setText("0%")
            else:
                self.expedite_amount_edit.clear()
                self.expedite_amount_edit.setPlaceholderText("Select Yes or No for expedite")
        
    def toggle_others_input(self, state):
        """Show/hide the 'Others' input field when checkbox is checked/unchecked"""
        if state == QtCore.Qt.Checked:
            self.others_input.setVisible(True)
            # UPDATED: Changed placeholder to indicate infinite entries
            self.others_input.setPlaceholderText("Type other services (separate with commas: 1, 2, 3, ...)")
            # Focus on the input field for convenience
            QtCore.QTimer.singleShot(100, self.others_input.setFocus)
        else:
            self.others_input.setVisible(False)
            self.others_input.clear()
        
    def toggle_client_intelligence(self):
        """Toggle client intelligence widget visibility"""
        if hasattr(self, 'client_suggestion_widget'):
            current_client = self.client_combo.currentText()
            self.client_suggestion_widget.update_client(current_client)
            self.client_suggestion_widget.setVisible(not self.client_suggestion_widget.isVisible())

    def show_template_dialog(self):
        """Show template selection dialog"""
        dialog = TemplateDialog(self.template_manager, 'quotes', self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            template = dialog.get_selected_template()
            if template:
                self.apply_template(template)

    def apply_template(self, template):
        """Apply selected template to current quote form"""
        # Apply scope of work
        if 'scope_of_work' in template:
            self.scope_of_work_edit.setText(template['scope_of_work'])
        
        # Apply services
        if 'services' in template:
            # Clear all service checkboxes first
            self.structural_checkbox.setChecked(False)
            self.civil_checkbox.setChecked(False)
            self.electrical_checkbox.setChecked(False)
            self.mechanical_checkbox.setChecked(False)
            self.plumbing_checkbox.setChecked(False)
            self.anchor_calc_checkbox.setChecked(False)
            self.solidworks_checkbox.setChecked(False)
            self.foundation_checkbox.setChecked(False)
            self.other_checkbox.setChecked(False)
            self.others_input.clear()
            self.others_input.setVisible(False)
            
            # Check services from template
            for service in template['services']:
                if service == 'Structural':
                    self.structural_checkbox.setChecked(True)
                elif service == 'Civil':
                    self.civil_checkbox.setChecked(True)
                elif service == 'Electrical':
                    self.electrical_checkbox.setChecked(True)
                elif service == 'Mechanical':
                    self.mechanical_checkbox.setChecked(True)
                elif service == 'Plumbing Design':
                    self.plumbing_checkbox.setChecked(True)
                elif service == 'Anchor Calculations':
                    self.anchor_calc_checkbox.setChecked(True)
                elif service == 'Solid Works':
                    self.solidworks_checkbox.setChecked(True)
                elif service == 'Foundation':
                    self.foundation_checkbox.setChecked(True)
                else:
                    # Handle other services
                    self.other_checkbox.setChecked(True)
                    self.others_input.setVisible(True)
                    current_others = self.others_input.text()
                    if current_others:
                        self.others_input.setText(f"{current_others}, {service}")
                    else:
                        self.others_input.setText(service)
        
        # Apply default price range as suggested cost
        if 'default_price_range' in template:
            # Extract the lower bound from price range
            price_range = template['default_price_range']
            if '$' in price_range and '-' in price_range:
                try:
                    lower_price = price_range.split('$')[1].split('-')[0].strip()
                    if lower_price.isdigit():
                        self.engineering_costs_edit.setText(f"${lower_price}")
                except:
                    pass
        
        # Update project name if template has a name and current name is empty
        if 'name' in template and not self.project_name_edit.text().strip():
            # Generate a project name based on template
            project_name = f"{template['name']} - {self.client_combo.currentText()}"
            self.project_name_edit.setText(project_name)
        
        # Show success message
        QtWidgets.QMessageBox.information(
            self, 
            "Template Applied", 
            f"Template '{template['name']}' has been applied successfully!"
        )

    def enable_email_autofill(self):
        """Enable email autofill for client changes"""
        self.allow_email_autofill = True
                    
    def on_client_selected(self, client_name):
        """Auto-fill email only; mailing/project address is project-specific."""
        if not getattr(self, "allow_email_autofill", True):
            return

        if isinstance(client_name, int):
            client_name = self.client_combo.itemText(client_name)
        client_name = str(client_name or "").strip()
        if not client_name:
            return

        clients = getattr(self, "saved_clients", {}) or {}
        client_data = clients.get(client_name)
        if not client_data:
            lowered = client_name.lower()
            client_data = next(
                (data for name, data in clients.items() if str(name).lower() == lowered),
                None,
            )
        if not isinstance(client_data, dict):
            return

        email = (
            client_data.get("company_email")
            or client_data.get("email")
            or client_data.get("primary_email")
            or ""
        )
        if email:
            self.client_email_edit.setText(str(email).strip())
        else:
            self.client_email_edit.clear()

        current_address = self.client_address_edit.text().strip()
        saved_addresses = {
            str(data.get(key, "")).strip()
            for data in clients.values()
            if isinstance(data, dict)
            for key in ("address", "mailing_address")
            if str(data.get(key, "")).strip()
        }
        if current_address in saved_addresses:
            self.client_address_edit.clear()

    def prefix_dollar_sign(self):
        """Automatically add $ prefix to engineering cost if missing"""
        text = self.engineering_costs_edit.text().strip()
        if text and not text.startswith("$"):
            cursor_pos = self.engineering_costs_edit.cursorPosition()
            self.engineering_costs_edit.blockSignals(True)
            self.engineering_costs_edit.setText(f"${text.replace('$', '')}")
            self.engineering_costs_edit.blockSignals(False)
            self.engineering_costs_edit.setCursorPosition(cursor_pos + 1 if cursor_pos > 0 else 1)

    def clean_placeholder(self, line_edit):
        if line_edit.placeholderText() and line_edit.text() == "":
            return  # placeholder OK

    def load_saved_clients(self):
        """Load saved companies from Firebase or local clients.json."""
        try:
            self.client_combo.clear()
            self.saved_clients = {}

            if self.FIREBASE_AVAILABLE:
                from main import db
                ref = db.reference('/clients')
                self.saved_clients = ref.get() or {}
            else:
                self.saved_clients = _load_local_clients()

            if self.saved_clients:
                clean_clients = [
                    c for c in sorted(self.saved_clients.keys(), key=str.lower)
                    if not str(c).startswith("-- Select")
                ]
                self.client_combo.addItems(clean_clients)

            self.client_combo.setCurrentIndex(-1)
            line_edit = self.client_combo.lineEdit()
            line_edit.setPlaceholderText("Enter or select company name")
            line_edit.clear()

        except Exception as e:
            _log.error("Error loading companies: %s", e)


    def set_focus_to_widget(self, widget):
        """Set focus to widget with appropriate selection"""
        if not widget or not widget.isEnabled() or not widget.isVisible():
            return False
        
        widget.setFocus()
        
        if isinstance(widget, QtWidgets.QLineEdit):
            widget.selectAll()
        elif isinstance(widget, QtWidgets.QComboBox):
            line_edit = widget.lineEdit()
            if line_edit:
                line_edit.selectAll()
        elif isinstance(widget, QtWidgets.QDateEdit):
            line_edit = widget.lineEdit()
            if line_edit:
                line_edit.selectAll()
        elif isinstance(widget, QtWidgets.QCheckBox):
            # Just focus, don't toggle
            pass
        
        return True
    
    def ensureWidgetVisible(self, widget):
        """Ensure the widget is visible in the scroll area"""
        try:
            # Find the scroll area in the widget's parent hierarchy
            scroll_area = None
            current_widget = widget
            
            while current_widget:
                if isinstance(current_widget, QtWidgets.QScrollArea):
                    scroll_area = current_widget
                    break
                current_widget = current_widget.parent()
            
            if scroll_area and widget:
                # Use Qt's built-in method to scroll to the widget
                scroll_area.ensureWidgetVisible(widget)
                
        except Exception as e:
            _log.warning("Scroll error: %s", e)

    def setup_date_field_ux(self):
        """Setup better UX for date fields for manual entry"""
        for date_edit in [self.start_date_edit, self.due_date_edit]:
            if date_edit:
                # Allow manual editing
                date_edit.setReadOnly(False)
                
                # Get the line edit portion
                line_edit = date_edit.lineEdit()
                if line_edit:
                    # Set placeholder to show expected format
                    line_edit.setPlaceholderText("MM-DD-YY")
                    
                    # Install event filter for the line edit too
                    line_edit.installEventFilter(self)
                    
                    # Connect to handle manual text changes
                    line_edit.textEdited.connect(
                        lambda text, de=date_edit: self.validate_date_input(de, text)
                    )

    def validate_date_input(self, date_edit, text):
        """Validate date input as user types"""
        if not text:
            return
        
        # Clean the text - remove any non-digit characters except dashes and slashes
        cleaned = ''.join(c for c in text if c.isdigit() or c in ['-', '/'])
        
        # Auto-format as user types
        if len(cleaned) >= 2 and '-' not in cleaned and '/' not in cleaned:
            # Add first separator after month
            if len(cleaned) == 2:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))
        
        elif len(cleaned) >= 5 and cleaned[2] in ['-', '/'] and cleaned[5:] == '':
            # Add second separator after day
            if len(cleaned) == 5:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))
            
    # ===== Event Filter for Deliverables Smart Entry =====
    def eventFilter(self, source, event):
        """Handle Enter key navigation with proper focus handling for date fields and backspace support"""
        # =====================================================
        # 1. Handle backspace/delete in date fields
        # =====================================================
        if isinstance(source, QtWidgets.QLineEdit):
            # Check if this line edit belongs to a date edit widget
            parent = source.parent()
            if parent and isinstance(parent, QtWidgets.QDateEdit) and parent in [self.start_date_edit, self.due_date_edit]:
                
                if event.type() == QtCore.QEvent.FocusIn:
                    # Select all text when date field gets focus
                    QtCore.QTimer.singleShot(10, lambda s=source: s.selectAll())
                    return False
                
                if event.type() == QtCore.QEvent.KeyPress:
                    # Handle Enter key - move to next field
                    if event.key() in [QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter]:
                        # Let the main Enter handling logic handle this
                        return False
                    
                    # Handle backspace/delete - clear selected text
                    elif event.key() in [QtCore.Qt.Key_Backspace, QtCore.Qt.Key_Delete]:
                        if source.hasSelectedText():
                            source.clear()
                            return True
                    
                    # Handle numeric keys and separators for manual date entry
                    elif event.key() in [
                        QtCore.Qt.Key_0, QtCore.Qt.Key_1, QtCore.Qt.Key_2, QtCore.Qt.Key_3,
                        QtCore.Qt.Key_4, QtCore.Qt.Key_5, QtCore.Qt.Key_6, QtCore.Qt.Key_7,
                        QtCore.Qt.Key_8, QtCore.Qt.Key_9, QtCore.Qt.Key_Minus, QtCore.Qt.Key_Slash
                    ]:
                        # If text is selected, clear it first so user can type fresh
                        if source.hasSelectedText():
                            source.clear()
                        # Continue normal processing
                        return False
        
        # =====================================================
        # 2. Handle Enter key for DELIVERABLES (last logical field)
        # =====================================================
        if source == self.deliverables_edit and event.type() == QtCore.QEvent.KeyPress:
            if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):

                # First Enter → auto comma (existing behavior)
                text = self.deliverables_edit.text().strip()
                if text and not text.endswith(","):
                    self.deliverables_edit.setText(text + ", ")

                # Scroll to Actions + focus Create button
                QtCore.QTimer.singleShot(30, lambda: self.ensureWidgetVisible(self.create_btn))
                QtCore.QTimer.singleShot(60, lambda: self.create_btn.setFocus())

                self._enter_on_last_field = True
                return True

        
        # =====================================================
        # 3. Handle Enter / Up / Down key navigation
        # =====================================================
        if event.type() == QtCore.QEvent.KeyPress:

            key = event.key()

            # Detect navigation keys
            is_enter = key in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter)
            is_down = key == QtCore.Qt.Key_Down
            is_up = key == QtCore.Qt.Key_Up

            if not (is_enter or is_down or is_up):
                return False

            # Find current widget index
            # 🔥 FIX: handle QComboBox lineEdit properly
            current_index = -1

            for i, widget in enumerate(self.input_widgets):
                if widget == source:
                    current_index = i
                    break
                
                # ✅ VERY IMPORTANT FIX (for client field)
                if isinstance(widget, QtWidgets.QComboBox) and source == widget.lineEdit():
                    current_index = i
                    break

            if current_index == -1:
                return False
            # Direction
            step = 1 if (is_enter or is_down) else -1
            next_index = current_index + step

            # ============================
            # END → ACTION BUTTONS
            # ============================
            if next_index >= len(self.input_widgets):

                # First Enter after last field → show action buttons
                if not self._enter_on_last_field:
                    self._enter_on_last_field = True
                    self.create_btn.setFocus()
                    self.ensureWidgetVisible(self.create_btn)
                    return True

                # Second Enter → loop to first field
                self._enter_on_last_field = False
                self.project_site_edit.setFocus()
                self.project_site_edit.selectAll()
                self.ensureWidgetVisible(self.project_site_edit)
                return True

            # Reset flag if user navigates normally
            self._enter_on_last_field = False
            # =====================================================
            # ACTION BUTTONS → LOOP BACK TO FIRST FIELD
            # =====================================================
            if source in [self.create_btn, self.cancel_btn] and is_enter:
                self._enter_on_last_field = False

                # Loop back to first field (Project Site)
                QtCore.QTimer.singleShot(20, lambda: self.project_site_edit.setFocus())
                QtCore.QTimer.singleShot(30, lambda: self.project_site_edit.selectAll())
                QtCore.QTimer.singleShot(40, lambda: self.ensureWidgetVisible(self.project_site_edit)
                )
                return True

            # ============================
            # NORMAL NAVIGATION
            # ============================
            attempts = 0
            while 0 <= next_index < len(self.input_widgets) and attempts < len(self.input_widgets):
                widget = self.input_widgets[next_index]

                if widget and widget.isEnabled() and widget.isVisible():

                    widget.setFocus()

                    # Select text smartly
                    if isinstance(widget, QtWidgets.QLineEdit):
                        QtCore.QTimer.singleShot(10, widget.selectAll)

                    elif isinstance(widget, QtWidgets.QComboBox):
                        le = widget.lineEdit()
                        if le:
                            QtCore.QTimer.singleShot(10, le.selectAll)

                    elif isinstance(widget, QtWidgets.QDateEdit):
                        le = widget.lineEdit()
                        if le:
                            QtCore.QTimer.singleShot(10, le.selectAll)

                    self.ensureWidgetVisible(widget)
                    return True

                next_index += step
                attempts += 1

            return True

                
        # =====================================================
        # 4. Handle Tab / Shift+Tab key navigation for all widgets
        # =====================================================
        if event.type() == QtCore.QEvent.KeyPress and event.key() in (
            QtCore.Qt.Key_Tab, QtCore.Qt.Key_Backtab
        ):
            is_back = (event.key() == QtCore.Qt.Key_Backtab or
                       bool(event.modifiers() & QtCore.Qt.ShiftModifier))
            step = -1 if is_back else 1

            current_index = -1
            for i, widget in enumerate(self.input_widgets):
                if widget and (widget == source or (
                    isinstance(widget, QtWidgets.QComboBox) and source == widget.lineEdit()
                )):
                    current_index = i
                    break

            if current_index >= 0:
                next_index = current_index + step
                attempts = 0
                while 0 <= next_index < len(self.input_widgets) and attempts < len(self.input_widgets):
                    nw = self.input_widgets[next_index]
                    if nw and nw.isEnabled() and nw.isVisible():
                        nw.setFocus()
                        if isinstance(nw, QtWidgets.QLineEdit):
                            QtCore.QTimer.singleShot(10, nw.selectAll)
                        elif isinstance(nw, (QtWidgets.QComboBox, QtWidgets.QDateEdit)):
                            le = nw.lineEdit()
                            if le:
                                QtCore.QTimer.singleShot(10, le.selectAll)
                        self.ensureWidgetVisible(nw)
                        return True
                    next_index += step
                    attempts += 1
                return True

        return super().eventFilter(source, event)

    # ===== Helper Methods =====
    def add_section_title(self, layout, text):
        label = QtWidgets.QLabel(text)
        label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                font-size: 16px;
                color: #2c3e50;
                border-bottom: 2px solid #dfe6e9;
                padding-bottom: 6px;
            }
        """)
        layout.addWidget(label)

    def add_field(self, layout, label_text, widget):
        field_layout = QtWidgets.QHBoxLayout()
        label = QtWidgets.QLabel(label_text)
        label.setStyleSheet("font-weight: 500; color: #2c3e50; min-width: 150px;")
        field_layout.addWidget(label)
        if isinstance(widget, (QtWidgets.QLineEdit, QtWidgets.QComboBox, QtWidgets.QDateEdit)):
            widget.setMaximumWidth(1150)
        field_layout.addWidget(widget, 1)
        field_layout.addStretch(1)
        layout.addLayout(field_layout)

    def create_styled_line_edit(self, placeholder="", read_only=False):
        edit = QtWidgets.QLineEdit()
        edit.setPlaceholderText(placeholder)
        edit.setReadOnly(read_only)
        edit.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QLineEdit:focus { 
                border-color: #3498db; 
                background: #f8f9fa; 
            }
            QLineEdit[readOnly="true"] { 
                background: #ecf0f1; 
                color: #7f8c8d; 
            }
        """)
        edit.setFocusPolicy(QtCore.Qt.StrongFocus)
        return edit
    def get_next_navigable_widget(self, current_index):
        """Get the next navigable widget, skipping buttons and looping back to start"""
        start_index = current_index
        next_index = (current_index + 1) % len(self.input_widgets)
        
        # Try to find next navigable widget (not a button)
        while next_index != start_index:
            widget = self.input_widgets[next_index]
            
            # Skip buttons
            if widget in [self.create_btn, self.cancel_btn]:
                next_index = (next_index + 1) % len(self.input_widgets)
                continue
            
            if widget and widget.isEnabled() and widget.isVisible():
                return widget, next_index
            
            next_index = (next_index + 1) % len(self.input_widgets)
        
        # If no other widget found, return first input field
        return self.project_site_edit, 0

    def create_styled_text_edit(self, placeholder="", height=100):
        edit = QtWidgets.QTextEdit()
        edit.setPlaceholderText(placeholder)
        edit.setMinimumHeight(height)
        edit.setStyleSheet("""
            QTextEdit {
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                padding: 8px;
                font-size: 13px;
            }
            QTextEdit:focus { border-color: #3498db; background: #f8f9fa; }
        """)
        return edit

    def create_styled_combo_box(self, items):
        """Create styled combo box with proper typing support - EXACTLY LIKE PROJECT DIALOG"""
        combo = QtWidgets.QComboBox()
        combo.addItems(items)
        combo.setEditable(True)

        line_edit = combo.lineEdit()
        line_edit.setPlaceholderText("Enter or select Company/Client name")

        # When user types → remove placeholder IMMEDIATELY
        def on_user_type(text, le=line_edit, cb=combo):
            if le.text().startswith("-- Select"):
                clean = le.text().replace("-- Select Company --", "").strip()
                le.blockSignals(True)
                le.setText(clean)
                le.blockSignals(False)

        line_edit.textEdited.connect(on_user_type)

        # When user focuses → clear placeholder
        def on_focus_in(event, le=line_edit):
            if le.text().startswith("-- Select"):
                le.blockSignals(True)
                le.clear()
                le.blockSignals(False)
            return QtWidgets.QLineEdit.focusInEvent(le, event)

        line_edit.focusInEvent = on_focus_in

        # Fix backspace and typing - ensure placeholder doesn't interfere
        def on_key_press(event):
            # Let the line edit handle the key press normally
            QtWidgets.QLineEdit.keyPressEvent(line_edit, event)
        
        line_edit.keyPressEvent = on_key_press

        # Styling
        combo.setStyleSheet("""
            QComboBox {
                padding: 9px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QComboBox:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
            QComboBox QAbstractItemView {
                selection-background-color: #3498db;
            }
        """)
        return combo


    def handle_focus_in_event(self, line_edit, event):
        """Handle focus in event - clear placeholder immediately"""
        if line_edit.text() == "-- Select Client --":
            line_edit.blockSignals(True)
            line_edit.clear()
            line_edit.blockSignals(False)
        
        # Call the original focusInEvent
        QtWidgets.QLineEdit.focusInEvent(line_edit, event)

    def clear_placeholder_on_type(self, line_edit, text):
        """Clear placeholder text when user starts typing"""
        if text and "-- Select Client --" in text:
            clean_text = text.replace("-- Select Client --", "").strip()
            line_edit.blockSignals(True)
            line_edit.setText(clean_text)
            line_edit.blockSignals(False)


    def create_styled_checkbox(self, text, checked=False):
        box = QtWidgets.QCheckBox(text)
        box.setChecked(checked)
        box.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-size: 13px;
                color: #2c3e50;
                padding: 4px;
            }
            QCheckBox:focus {
                background-color: #f0f8ff;
                border-radius: 4px;
                outline: 1px solid #3498db;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 2px solid #bdc3c7;
                border-radius: 4px;
                background: white;
            }
            QCheckBox::indicator:checked {
                background: #3498db;
                border-color: #3498db;
            }
            QCheckBox::indicator:checked:hover {
                background: #2980b9;
                border-color: #2980b9;
            }
        """)
        # Make checkbox focusable
        box.setFocusPolicy(QtCore.Qt.StrongFocus)
        return box

    def create_styled_date_edit(self, date):
        """Create styled date edit WITHOUT scroll, arrows, or auto increment"""
        d = QtWidgets.QDateEdit(date)
        d.setCalendarPopup(True)
        d.setDisplayFormat("MM-dd-yyyy")
        d.setReadOnly(False)

        # ✅ Disable mouse wheel
        d.wheelEvent = lambda event: None

        # ✅ Disable arrow key increment/decrement
        def keyPressEvent(event, original=d.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        d.keyPressEvent = keyPressEvent

        # ✅ MOST IMPORTANT: Disable internal stepping (THIS fixes your issue)
        d.stepBy = lambda x: None

        # ✅ Remove spin buttons completely
        d.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        # Keep placeholder for manual typing
        line_edit = d.lineEdit()
        if line_edit:
            line_edit.setPlaceholderText("MM-DD-YY")

        # Styling (UNCHANGED)
        d.setStyleSheet("""
            QDateEdit {
                padding: 9px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)

        return d

    def _on_priority_manually_changed(self):
        """User manually changed priority — stop auto-managing it."""
        self._priority_auto_set = False

    def _on_due_date_changed(self):
        """Auto-set priority to Urgent when due date is within 7 days; revert when it moves out."""
        today = QtCore.QDate.currentDate()
        due = self.due_date_edit.date()
        days_remaining = today.daysTo(due)
        if days_remaining <= 7:
            if self.priority_combo.currentText() != "Urgent":
                self._priority_auto_set = True
                idx = self.priority_combo.findText("Urgent")
                if idx >= 0:
                    self.priority_combo.blockSignals(True)
                    self.priority_combo.setCurrentIndex(idx)
                    self.priority_combo.blockSignals(False)
        else:
            if getattr(self, '_priority_auto_set', False):
                self._priority_auto_set = False
                idx = self.priority_combo.findText("Medium")
                if idx >= 0:
                    self.priority_combo.blockSignals(True)
                    self.priority_combo.setCurrentIndex(idx)
                    self.priority_combo.blockSignals(False)

    def update_timeline_info(self):
        start = self.start_date_edit.date().toPyDate()
        due = self.due_date_edit.date().toPyDate()
        days = (due - start).days
        info_text = (
            f"From: {start.strftime('%b %d, %Y')}   |   "
            f"To: {due.strftime('%b %d, %Y')}   |   "
            f"Duration: {days} day{'s' if days != 1 else ''}"
        )
        self.timeline_info.setText(info_text)
        if hasattr(self, 'timeline_info_visible'):
            self.timeline_info_visible.setText(info_text)
            self.timeline_info_visible.setVisible(True)

    def auto_generate_job_number(self):
        """Auto-generate quote number based ONLY on category and main sequence"""
        if self.auto_generate_enabled and (not self.job_number_edit.text() or self.job_number_edit.text() == "Auto-generated"):
            self.generate_job_number()
            # Initialize expedite buttons for new forms
            self.expedite_yes.setChecked(False)
            self.expedite_no.setChecked(False)
            self._last_expedite_btn = None
            self.expedite_amount_edit.clear()
            self.expedite_amount_edit.setPlaceholderText("Select Yes or No for expedite")

    def generate_job_number(self):
        """Generate professional quote number - automatically transition to next category at 999"""

        # Prefer owner_tab (reliable for both embedded and popup modes)
        parent = self.owner_tab if hasattr(self, 'owner_tab') and self.owner_tab else self.parent()
        if not hasattr(parent, 'job_forms') or not parent.job_forms:
            job_number = "QuoteA001"
            self.job_number_edit.setText(job_number)
            self.auto_generate_enabled = False
            return job_number
        
        # Dictionary to track highest sequence for each category
        category_sequences = {}
        
        # First pass: Find highest sequence for each category
        for job in parent.job_forms:
            job_num = job.get('job_number', '').upper()
            
            if 'QUOTE' not in job_num:
                continue
            
            # Extract category and sequence (ignoring suffixes)
            match = re.match(r'^QUOTE([A-Z]?)(\d+)', job_num)
            if match:
                category = match.group(1) or 'A'  # Default to 'A' if no category
                seq_str = match.group(2)
                
                try:
                    # Convert to integer (remove leading zeros)
                    seq_num = int(seq_str)
                    
                    # Update if this is higher than current max for this category
                    if category not in category_sequences or seq_num > category_sequences[category]:
                        category_sequences[category] = seq_num
                except ValueError:
                    continue
        
        # If no categories found, start with A001
        if not category_sequences:
            job_number = "QuoteA001"
            self.job_number_edit.setText(job_number)
            self.auto_generate_enabled = False
            return job_number
        
        # Get all used categories sorted alphabetically
        sorted_categories = sorted(category_sequences.keys())
        
        # Start with the last (highest) category
        current_category = sorted_categories[-1]
        current_max_seq = category_sequences[current_category]
        
        # Check if we need to roll over
        if current_max_seq >= 999:
            # Category is full (reached 999)
            
            if current_category == 'Z':
                # We've reached Z999 - wrap around to A but continue sequence
                # Find the absolute highest sequence across all categories
                all_max_seq = max(category_sequences.values())
                
                if all_max_seq >= 999:
                    # All categories are at 999 or more, start a new cycle
                    next_category = 'A'
                    next_sequence = all_max_seq + 1
                    job_number = f"Quote{next_category}{next_sequence:03d}"
                else:
                    # Find next available category that isn't at 999 yet
                    for cat in sorted(category_sequences.keys()):
                        if category_sequences[cat] < 999:
                            next_category = cat
                            next_sequence = category_sequences[cat] + 1
                            break
                    else:
                        # All categories at 999, start new with A
                        next_category = 'A'
                        next_sequence = all_max_seq + 1
                    
                    job_number = f"Quote{next_category}{next_sequence:03d}"
            else:
                # Move to next letter in alphabet
                next_category = chr(ord(current_category) + 1)
                
                # Check if next category already exists
                if next_category in category_sequences:
                    next_sequence = category_sequences[next_category] + 1
                    
                    # If next sequence would be 1000 or more, skip to next available category
                    if next_sequence >= 1000:
                        # Find next available category with sequence < 999
                        next_category = None
                        for i in range(ord(next_category), ord('Z') + 1):
                            cat = chr(i)
                            if cat not in category_sequences:
                                next_category = cat
                                next_sequence = 1
                                break
                            elif category_sequences[cat] < 999:
                                next_category = cat
                                next_sequence = category_sequences[cat] + 1
                                break
                        
                        # If no category found, wrap to A
                        if next_category is None:
                            next_category = 'A'
                            next_sequence = max(category_sequences.values()) + 1
                else:
                    # Start new category at 001
                    next_sequence = 1
                
                job_number = f"Quote{next_category}{next_sequence:03d}"
        else:
            # Continue in current category
            next_sequence = current_max_seq + 1
            
            # Check if we're hitting 1000 (shouldn't happen with our logic)
            if next_sequence >= 1000:
                # Move to next category
                if current_category == 'Z':
                    next_category = 'A'
                    next_sequence = max(category_sequences.values()) + 1
                else:
                    next_category = chr(ord(current_category) + 1)
                    next_sequence = 1
                
                job_number = f"Quote{next_category}{next_sequence:03d}"
            else:
                job_number = f"Quote{current_category}{next_sequence:03d}"
        
        # Set the quote number in the field
        self.job_number_edit.setText(job_number)
        
        # Disable auto-generation after manual generation
        self.auto_generate_enabled = False
        
        return job_number

    def get_next_category_sequence_ignore_suffixes(self, category):
        """Get next sequence number for the given category, IGNORING ALL SUFFIXES"""
        if not self.FIREBASE_AVAILABLE:
            return 1
        
        try:
            from main import db
            
            # Get all quote forms from Firebase
            ref = db.reference('/job_forms')
            job_forms_data = ref.get()
            
            sequences = []
            
            if job_forms_data:
                for job_id, job_data in job_forms_data.items():
                    job_number = job_data.get('job_number', '').upper()
                    
                    # Check if it's a QuoteCategoryXXX format
                    if job_number.startswith(f'QUOTE{category}'):
                        try:
                            # Extract ONLY the base part before any underscore
                            base_part = job_number
                            if '_' in job_number:
                                base_part = job_number.split('_')[0]
                            
                            # Extract sequence from base part (last digits)
                            # Find the number after the category letter
                            # Pattern: QUOTE + category letter + numbers
                            pattern = f'QUOTE{category}(\\d+)'  # Double backslash
                            match = re.match(pattern, base_part, re.IGNORECASE)
                            
                            if match:
                                seq_str = match.group(1)
                                # Remove leading zeros for proper numeric comparison
                                seq_num = int(seq_str)  # This keeps 026 as 26
                                sequences.append(seq_num)
                        except (ValueError, IndexError, AttributeError) as e:
                            continue
                        
            if sequences:
                # Get the highest sequence number
                max_seq = max(sequences)
                
                # Check owner_tab job_forms for any higher numbers not yet in Firebase
                parent = self.owner_tab if hasattr(self, 'owner_tab') and self.owner_tab else self.parent()
                if hasattr(parent, 'job_forms') and parent.job_forms:
                    for job in parent.job_forms:
                        job_num = job.get('job_number', '').upper()
                        if job_num.startswith(f'QUOTE{category}'):
                            try:
                                # Extract sequence
                                pattern = f'QUOTE{category}(\\d+)'  # Double backslash
                                match = re.match(pattern, job_num.split('_')[0], re.IGNORECASE)
                                if match:
                                    seq = int(match.group(1))
                                    if seq > max_seq:
                                        max_seq = seq
                            except:
                                continue
                
                next_seq = max_seq + 1
                return next_seq
            else:
                return 1  # First job of this category
                        
        except Exception as e:
            _log.warning("Error getting category sequence: %s", e)
            _log.exception("Traceback:")
            return 1
    
    def open_generated_pdf_and_close(self, job_data, dialog, pdf_path=None):
        """Open the generated PDF and close the confirmation dialog."""
        dialog.accept()
        if pdf_path and Path(pdf_path).exists():
            # Use the already-generated local file directly — fastest path
            try:
                import os, platform, subprocess
                if platform.system() == "Windows":
                    os.startfile(str(pdf_path))
                elif platform.system() == "Darwin":
                    subprocess.call(["open", str(pdf_path)])
                else:
                    subprocess.call(["xdg-open", str(pdf_path)])
            except Exception as e:
                _log.warning("Error opening PDF: %s", e)
                self.open_generated_pdf(job_data)
        else:
            self.open_generated_pdf(job_data)

    def maybe_sync_linked_project(self, job_data):
        """Ask whether quote edits should update the linked project."""
        project_number = (job_data.get("project_number") or "").strip()
        if not project_number:
            return "no_link"

        original = self.job_data or {}
        sync_fields = [
            "project_name",
            "client",
            "client_address",
            "project_site_address",
            "plant",
            "sales",
            "engineering_costs",
            "start_date",
            "due_date",
        ]
        changed_fields = [
            field for field in sync_fields
            if str(original.get(field, "") or "").strip() != str(job_data.get(field, "") or "").strip()
        ]
        if not changed_fields:
            return "no_changes"

        reply = QtWidgets.QMessageBox.question(
            self,
            "Update Linked Project?",
            f"This quote is linked to project {project_number}.\n\n"
            "Do you want to update the project with the quote changes too?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
            QtWidgets.QMessageBox.Yes,
        )
        if reply == QtWidgets.QMessageBox.Cancel:
            return "cancel"
        if reply != QtWidgets.QMessageBox.Yes:
            return "quote_only"

        return "synced" if self.sync_linked_project_from_quote(project_number, job_data) else "sync_failed"

    def sync_linked_project_from_quote(self, project_number, job_data):
        """Copy safe quote fields into the existing linked project."""
        try:
            updates = {
                "job_number": job_data.get("job_number", ""),
                "project_name": job_data.get("project_name", ""),
                "company": job_data.get("client", ""),
                "mail_address": job_data.get("client_address", ""),
                "site_address": job_data.get("project_site_address", "") or job_data.get("client_address", ""),
                "plant": job_data.get("plant", ""),
                "sales": job_data.get("sales", ""),
                "project_amount": self._parse_money_value(job_data.get("engineering_costs", 0)),
                "start_date": job_data.get("start_date", ""),
                "due_date": job_data.get("due_date", ""),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }

            if FIREBASE_AVAILABLE:
                ref = db.reference('/projects')
                existing = ref.order_by_child('project_number').equal_to(project_number).get() or {}
                if existing:
                    project_id = next(iter(existing.keys()))
                    ref.child(project_id).update(updates)
                    _log.info("Linked project %s synced from quote %s", project_number, job_data.get("job_number", ""))
                    return True

            main_window = getattr(self, "main_window", None)
            project_tab = getattr(main_window, "project_tab", None)
            if project_tab:
                for source_name in ("generated_projects", "cached_projects"):
                    for project in getattr(project_tab, source_name, []) or []:
                        if project.get("project_number") == project_number:
                            project.update(updates)
                if hasattr(project_tab, "load_projects"):
                    QtCore.QTimer.singleShot(0, project_tab.load_projects)

            return False
        except Exception as exc:
            _log.warning("Could not sync linked project %s from quote: %s", project_number, exc)
            QtWidgets.QMessageBox.warning(
                self,
                "Project Sync Failed",
                f"Quote was saved, but linked project {project_number} could not be updated.\n\n{exc}",
            )
            return False

    def _parse_money_value(self, value):
        try:
            return float(str(value).replace("$", "").replace(",", "").strip() or 0)
        except (ValueError, TypeError):
            return 0.0
    
    def create_job_form(self):
        """Create and save quote form to Firebase with duplicate validation"""
        if not self.validate_form():
            return
            
        try:
            # Collect all quote form data
            job_data = self.collect_job_form_data()
            
            # ⭐⭐ NEW: Validate quote number format
            job_number = job_data['job_number'].strip()
            if not job_number.upper().startswith('QUOTE'):
                QtWidgets.QMessageBox.warning(
                    self,
                    "Invalid Quote Number",
                    "Quote Number must start with 'Quote'!\n\n"
                    "Examples: QuoteA001, QuoteB002, QuoteC003\n"
                    f"You entered: {job_number}"
                )
                return
            
            # ⭐⭐ NEW: Check for duplicate quote number (only for new jobs)
            if not self.is_editing:
                
                # Check against parent's job_forms list
                parent = self.parent()
                if hasattr(parent, 'job_forms') and parent.job_forms:
                    for job in parent.job_forms:
                        if job.get('job_number', '').upper() == job_number.upper():
                            QtWidgets.QMessageBox.warning(
                                self,
                                "Duplicate Quote Number",
                                f"Quote Number '{job_number}' already exists!\n\n"
                                f"Please use a different quote number."
                            )
                            return
            
            # If editing, ensure we have the firebase_id from the original job data
            if self.is_editing and self.job_data and 'firebase_id' in self.job_data:
                job_data['firebase_id'] = self.job_data['firebase_id']
            if self.is_editing and self.job_data and self.job_data.get('project_number'):
                job_data['project_number'] = self.job_data.get('project_number')
            
            # Save to Firebase using the parent tab's method
            success = False
            owner = self.owner_tab
            if hasattr(owner, 'save_job_form_to_firebase'):
                success = owner.save_job_form_to_firebase(job_data)
            else:
                # Fallback: use direct Firebase save
                success = self.save_job_form_directly(job_data)

            # ADD THIS: If parent method doesn't work, try direct method
            if not success and hasattr(owner, 'save_job_form_to_firebase'):
                success = self.save_job_form_directly(job_data)

            # Final safety net for the embedded New Quote tab: never lose a quote because
            # Firebase/direct save returned False without raising an exception.
            if not success and hasattr(owner, 'save_job_form_locally'):
                _log.warning(
                    "Firebase/direct quote save returned False; using local backup for %s",
                    job_data.get('job_number', 'N/A')
                )
                success = owner.save_job_form_locally(job_data)
            
            
            if success:
                # Immediately add to in-memory list so the next generate_job_number()
                # call always sees the just-saved quote, regardless of Firebase refresh timing.
                if not self.is_editing:
                    owner_jobs = getattr(self.owner_tab, 'job_forms', None)
                    if isinstance(owner_jobs, list):
                        if not any(j.get('job_number') == job_data.get('job_number') for j in owner_jobs):
                            owner_jobs.append(job_data)

                if self.is_editing:
                    sync_result = self.maybe_sync_linked_project(job_data)
                    if sync_result == "cancel":
                        return

                # Generate PDF after successful save
                pdf_success, pdf_path = self.generate_job_form_pdf(job_data)
                
                if pdf_success:
                    # ✅ Save PDF to Firebase immediately
                    self.save_job_form_pdf_to_firebase(job_data['job_number'], pdf_path)
                
                action = "updated" if self.is_editing else "created"
                pdf_message = " and PDF generated" if pdf_success else " (PDF generation failed)"
                
                # ⭐⭐ NEW: Professional PDF Confirmation Dialog
                if pdf_success:
                    # Create a custom styled confirmation dialog
                    confirm_dialog = QtWidgets.QDialog(self)
                    confirm_dialog.setWindowTitle("✅ Success!")
                    confirm_dialog.setFixedSize(500, 400)
                    confirm_dialog.setStyleSheet("""
                        QDialog {
                            background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                stop:0 #f8fafc, stop:1 #e8f4f8);
                        }
                    """)
                    
                    layout = QtWidgets.QVBoxLayout(confirm_dialog)
                    layout.setSpacing(20)
                    layout.setContentsMargins(30, 30, 30, 30)
                    
                    # Success icon
                    icon_label = QtWidgets.QLabel("🎉")
                    icon_label.setStyleSheet("""
                        QLabel {
                            font-size: 60px;
                            text-align: center;
                        }
                    """)
                    icon_label.setAlignment(QtCore.Qt.AlignCenter)
                    layout.addWidget(icon_label)
                    
                    # Success message
                    success_label = QtWidgets.QLabel(
                        f"Quote Form {action.capitalize()} Successfully!"
                    )
                    success_label.setStyleSheet("""
                        QLabel {
                            font-size: 20px;
                            font-weight: bold;
                            color: #27ae60;
                            text-align: center;
                        }
                    """)
                    success_label.setAlignment(QtCore.Qt.AlignCenter)
                    success_label.setWordWrap(True)
                    layout.addWidget(success_label)
                    
                    # Details box
                    details_frame = QtWidgets.QFrame()
                    details_frame.setStyleSheet("""
                        QFrame {
                            background: white;
                            border: 2px solid #dfe6e9;
                            border-radius: 10px;
                            padding: 15px;
                        }
                    """)
                    details_layout = QtWidgets.QVBoxLayout(details_frame)
                    
                    # Quote Number
                    quote_label = QtWidgets.QLabel(f"📋 <b>Quote Number:</b> {job_data.get('job_number', 'N/A')}")
                    quote_label.setStyleSheet("font-size: 14px; color: #2c3e50;")
                    quote_label.setTextFormat(QtCore.Qt.RichText)
                    details_layout.addWidget(quote_label)
                    
                    
                    layout.addWidget(details_frame)
                    
                    # Action buttons
                    button_layout = QtWidgets.QHBoxLayout()
                    button_layout.setSpacing(15)
                    
                    # Open PDF button
                    open_pdf_btn = QtWidgets.QPushButton("📂 Open PDF")
                    open_pdf_btn.setFixedHeight(45)
                    open_pdf_btn.setStyleSheet("""
                        QPushButton {
                            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #3498db, stop:1 #2980b9);
                            color: white;
                            border: none;
                            border-radius: 8px;
                            font-weight: bold;
                            font-size: 14px;
                            padding: 10px 20px;
                        }
                        QPushButton:hover {
                            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                stop:0 #2980b9, stop:1 #21618c);
                        }
                        QPushButton:pressed {
                            background: #21618c;
                        }
                    """)
                    open_pdf_btn.clicked.connect(lambda checked=False, _p=pdf_path, _d=confirm_dialog: self.open_generated_pdf_and_close(job_data, _d, _p))
                    
                    # Close button
                    close_btn = QtWidgets.QPushButton("👌 Close")
                    close_btn.setFixedHeight(45)
                    close_btn.setStyleSheet("""
                        QPushButton {
                            background: #95a5a6;
                            color: white;
                            border: none;
                            border-radius: 8px;
                            font-weight: bold;
                            font-size: 14px;
                            padding: 10px 20px;
                        }
                        QPushButton:hover {
                            background: #7f8c8d;
                        }
                        QPushButton:pressed {
                            background: #6c757d;
                        }
                    """)
                    close_btn.clicked.connect(confirm_dialog.accept)
                    
                    button_layout.addWidget(open_pdf_btn)
                    button_layout.addWidget(close_btn)
                    
                    layout.addLayout(button_layout)
                    
                    # Show the confirmation dialog
                    confirm_dialog.exec_()
                else:
                    # If PDF generation failed, show regular message box
                    QtWidgets.QMessageBox.information(
                        self,
                        "Success",
                        f"""
                        <div style='text-align:center; font-size:14px;'>
                            <b>Quote Form {action.capitalize()} Successfully{pdf_message}!</b><br><br>

                            <table style='margin-left:auto; margin-right:auto; text-align:center;'>
                                <tr>
                                    <td><b>Quote Number:</b></td>
                                    <td>{job_data.get('job_number', 'N/A')}</td>
                                </tr>
                                <tr>
                                    <td><b>Client:</b></td>
                                    <td>{job_data.get('client') or job_data.get('client_name') or 'N/A'}</td>
                                </tr>
                                <tr><td colspan='2'><br></td></tr>
                            </table>
                        </div>
                        """
                    )
                                    
                # Force immediate refresh of the quote forms table
                if hasattr(owner, 'load_job_forms_from_firebase'):
                    owner.load_job_forms_from_firebase()
                    
                    # After refresh, re-apply any active filters
                    if hasattr(owner, 'filter_job_forms'):
                        owner.filter_job_forms()

                self.accept()
            else:
                parent_name = type(owner).__name__ if owner else "None"
                _log.warning(
                    "Quote form save failed after all save attempts. Parent=%s Quote=%s",
                    parent_name,
                    job_data.get('job_number', 'N/A')
                )
                QtWidgets.QMessageBox.critical(
                    self, "Error", 
                    "Failed to save quote form after Firebase and local backup attempts.\n\n"
                    f"Quote Number: {job_data.get('job_number', 'N/A')}\n"
                    f"Parent: {parent_name}\n\n"
                    "Please send this screen if it happens again."
                )
                    
        except Exception as e:
            _log.exception("Traceback:")
            QtWidgets.QMessageBox.critical(
                self, "Error", 
                f"An error occurred while {'updating' if self.is_editing else 'creating'} the quote form:\n{str(e)}"
            )
        
        
    def generate_job_form_pdf(self, job_data):
        """Generate quote Form PDF with professional header design - SAVE TO FIREBASE ONLY"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph,
                Spacer, Image, FrameBreak, PageBreak
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from pathlib import Path
            import tempfile
            
            def format_usd(value):
                """Format number/string as US currency with commas"""
                try:
                    if not value:
                        return ""
                    v = str(value).replace("$", "").replace(",", "").strip()
                    return f"${float(v):,.2f}"
                except:
                    return str(value)

            # -------------------------
            # COMPANY INFO FROM SETTINGS
            # -------------------------
            try:
                from main import Config as _Cfg
                company_name = _Cfg.COMPANY.get('name', 'MABS Engineering LLC')
                company_address = _Cfg.COMPANY.get('address', '').replace('\n', '<br/>')
                _logo = _Cfg.get_logo_path()
                logo_path = _logo if _logo is not None else resource_path("assets/logo.jpg")
            except Exception:
                company_name = "MABS Engineering LLC"
                company_address = "PO Box 1144, 15455 Manchester Rd, Ballwin, MO 63011"
                logo_path = resource_path("assets/logo.jpg")
            venmo_qr_path = resource_path("assets/venmo.png")

            # -------------------------
            # TEMPORARY FILE PATH
            # -------------------------
            temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
            temp_dir.mkdir(parents=True, exist_ok=True)
            filename = temp_dir / f"{job_data['job_number']}_job_form.pdf"
            
            # -------------------------
            # FOOTER FUNCTION - WILL BE ADDED TO EVERY PAGE
            # -------------------------
            def footer(canvas, doc):
                """Add footer to bottom of page"""
                canvas.saveState()
                
                footer_style = ParagraphStyle(
                    name="FooterStyle",
                    alignment=1,
                    fontName="Helvetica",
                    fontSize=8,  # Reduced from 8
                    textColor=colors.HexColor("#003366"),
                    leading=9    # Reduced from 10
                )
                
                # Footer line
                canvas.setStrokeColor(colors.HexColor("#003366"))
                canvas.setLineWidth(0.5)
                canvas.line(doc.leftMargin, 18 * mm, doc.width + doc.leftMargin, 20 * mm)  # Reduced from 20mm
                
                # Footer text
                try:
                    from main import Config as _FCfg
                    _addr = _FCfg.COMPANY.get('address', '').replace('\n', ', ')
                    _phone = _FCfg.COMPANY.get('phone', '')
                    _femail = _FCfg.COMPANY.get('email', '')
                    _fweb = _FCfg.COMPANY.get('website', '')
                    _fname = _FCfg.COMPANY.get('name', 'MABS Engineering LLC')
                except Exception:
                    _addr = "PO Box 1144, 15455 Manchester Rd, Ballwin, MO 63011"
                    _phone = "(314) 585-2003"
                    _femail = "info@mabs-engineering.com"
                    _fweb = "www.mabs-engineering.com"
                    _fname = "MABS Engineering LLC"
                footer_lines = [
                    f"Note: As the CEO of {_fname}, Dr. Ashiq reserves the right to change or cancel this policy at any time, at his discretion.",
                    f"Address: {_addr}",
                    f"Telephone: {_phone} • {_femail}",
                    _fweb
                ]
                
                # Start Y position for footer text
                y_position = 14 * mm  # Reduced from 15mm
                
                for line in footer_lines:
                    p = Paragraph(line, footer_style)
                    w, h = p.wrap(doc.width, 10 * mm)
                    p.drawOn(canvas, doc.leftMargin, y_position)
                    y_position -= 3.5 * mm  # Reduced from 4mm
                
                canvas.restoreState()
            
            # -------------------------
            # DOCUMENT - SINGLE PAGE WITH FOOTER
            # -------------------------
            doc = SimpleDocTemplate(
                str(filename),
                pagesize=A4,
                topMargin=1 * mm,
                bottomMargin=20 * mm,
                leftMargin=10 * mm,
                rightMargin=10 * mm
            )
            story = []
            styles = getSampleStyleSheet()
            
            # -------------------------
            # TEXT STYLES - REDUCED FONT SIZES
            # -------------------------
            styles.add(ParagraphStyle(
                name="HeaderCompany", 
                alignment=1, 
                fontName="Helvetica-Bold", 
                fontSize=26,  # Reduced from 26
                textColor=colors.HexColor("#2c3e50")
            ))
            styles.add(ParagraphStyle(
                name="DocumentTitle", 
                alignment=1, 
                fontName="Helvetica-Bold", 
                fontSize=14,  # Reduced from 14
                textColor=colors.HexColor("#2c3e50")
            ))
            styles.add(ParagraphStyle(name="Centered10", alignment=1, fontSize=9))  # Reduced from 10
            styles.add(ParagraphStyle(name="LabelBold", fontName="Helvetica-Bold", fontSize=10, alignment=0))  # Reduced from 11
            styles.add(ParagraphStyle(name="FieldValue", fontName="Helvetica", fontSize=9, alignment=0))  # Reduced from 10
            styles.add(ParagraphStyle(name="ProfessionalText", fontName="Helvetica", fontSize=8,  # Reduced from 9
                                    textColor=colors.HexColor("#333333"), alignment=0))

            # 🔹 Engineering Costs inline alignment
            engineering_line_style = ParagraphStyle(
                name="EngineeringLine",
                parent=styles["FieldValue"],
                leftIndent=8
            )

            # -------------------------
            # CLEAN PROFESSIONAL HEADER
            # -------------------------
            
            # Main header content
            header_table_data = []
            
            # Check if logo exists
            if logo_path.exists():
                logo = Image(str(logo_path), width=64, height=64)
                header_table_data = [
                    [logo, 
                    Paragraph(company_name, styles["HeaderCompany"]),
                    ""]
                ]
                col_widths = [50, doc.width - 100, 50]
            else:
                header_table_data = [
                    ["", 
                    Paragraph(company_name, styles["HeaderCompany"]),
                    ""]
                ]
                col_widths = [0, doc.width, 0]

            header_table = Table(header_table_data, colWidths=col_widths)
            header_table.setStyle(TableStyle([
                ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
                ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
                ('TOPPADDING', (0, 0), (0, 0), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), -1),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('LEFTPADDING', (0, 0), (0, 0), 12),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor("#3498db")),  # Reduced from 2
            ]))

            story.append(header_table)
            story.append(Spacer(1, 2 * mm))
            
            # Document title
            sales_style = ParagraphStyle(
                name="SalesBox",
                fontName="Helvetica-Bold",
                fontSize=9,  # Reduced from 9
                textColor=colors.black,
                alignment=1
            )

            sales_value = job_data.get("sales", "").strip()
            title_para = Paragraph("<u>New Job Request Form</u>", styles["DocumentTitle"])

            if sales_value:
                sales_text = f"Sales: {sales_value}"
                from reportlab.pdfbase.pdfmetrics import stringWidth
                text_width = stringWidth(sales_text, "Helvetica-Bold", 11)  # Reduced from 11
                sales_box_width = min(max(text_width + 10 * mm, 30 * mm), 65 * mm)  # Reduced dimensions

                sales_box = Table(
                    [[Paragraph(sales_text, sales_style)]],
                    colWidths=[sales_box_width],
                    rowHeights=[6 * mm],  # Reduced from 6mm
                )

                sales_box.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#E8F1FA")),
                    ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#6C8EBF")),  # Reduced from 1
                    ('ROUNDED', (0,0), (-1,-1), 4),  # Reduced from 4
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('LEFTPADDING', (0,0), (-1,-1), 1),
                    ('RIGHTPADDING', (0,0), (-1,-1), 1),
                    ('TOPPADDING', (0,0), (-1,-1), 0),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                ]))

                sales_title_row = Table(
                    [[sales_box, title_para, ""]],
                    colWidths=[None, None, None]
                )
            else:
                sales_title_row = Table(
                    [["", title_para, ""]],
                    colWidths=[None, None, None]
                )

            sales_title_row.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (1,0), (1,0), 'CENTER'),
                ('TOPPADDING', (1,0), (1,0), -3),  # Reduced from -4
                    ('BOTTOMPADDING', (1,0), (1,0), 2),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ]))

            story.append(sales_title_row)
            story.append(Spacer(1, 2 * mm))

            # -------------------------
            # HELPER: CLEAN FIELD BLOCK
            # -------------------------
            def add_field(label, value):
                table = Table(
                    [
                        [
                            Paragraph(f"<b>{label}</b>", styles["LabelBold"]),
                            Paragraph(":", styles["LabelBold"]),
                            Paragraph(f"<para>{value}</para>", styles["FieldValue"])
                        ]
                    ],
                    colWidths=[55 * mm, 5 * mm, doc.width - 60 * mm]
                )

                table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                ]))

                story.append(table)
                story.append(Spacer(1, 1.0 * mm))

            def add_title(label):
                title_para = Paragraph(f"<b>{label}</b>", styles["LabelBold"])
                title_table = Table([[title_para]], colWidths=[doc.width])
                title_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ]))

                story.append(title_table)
                story.append(Spacer(1, 0.3 * mm))  # Reduced from 0.5mm
            
            # -------------------------
            # MAIN FIELDS
            # -------------------------
            add_field("Quote Number", job_data.get("job_number", ""))
            add_field("Client / Company Name", job_data.get("client", ""))
            add_field("Project Site Address", job_data.get("project_site_address", ""))
            add_field("Project Name", job_data.get("project_name", ""))
            add_field("Scope of Work", job_data.get("scope_of_work", ""))
            
            # -------------------------
            # ENGINEERING COSTS
            # -------------------------
            engineering_rows = []
            engineering_rows.append(
                [Paragraph("<b>Engineering Costs:</b>", styles["LabelBold"])]
            )

            has_expedite_selection = (
                job_data.get("expedite") is True or
                job_data.get("expedite") is False and job_data.get("expedite_amount")
            )

            base_cost_text = job_data.get("engineering_costs", "").replace("$", "").replace(",", "").strip()
            try:
                base_cost_value = float(base_cost_text) if base_cost_text else 0.0
            except:
                base_cost_value = 0.0

            expedite_amount_text = job_data.get("expedite_amount", "").strip()
            expedite_dollar_amount = 0.0
            expedite_display_text = ""

            if expedite_amount_text:
                # Case 1: Percentage
                if "%" in expedite_amount_text:
                    try:
                        percent_value = float(expedite_amount_text.replace("%", "").strip())
                        expedite_dollar_amount = base_cost_value * (percent_value / 100)
                        expedite_display_text = f"${expedite_dollar_amount:,.2f}"
                    except:
                        expedite_display_text = expedite_amount_text

                # Case 2: Dollar amount with $
                elif "$" in expedite_amount_text:
                    try:
                        expedite_dollar_amount = float(
                            expedite_amount_text.replace("$", "").replace(",", "").strip()
                        )
                        expedite_display_text = f"${expedite_dollar_amount:,.2f}"
                    except:
                        expedite_display_text = expedite_amount_text

                # ✅ Case 3: Plain number (e.g. "5") → treat as $
                else:
                    try:
                        expedite_dollar_amount = float(expedite_amount_text)
                        expedite_display_text = f"${expedite_dollar_amount:,.2f}"
                    except:
                        expedite_display_text = expedite_amount_text

            if expedite_amount_text:
                if "%" in expedite_amount_text:
                    try:
                        percent_value = float(expedite_amount_text.replace("%", "").strip())
                        expedite_dollar_amount = base_cost_value * (percent_value / 100)
                        expedite_display_text = f"${expedite_dollar_amount:,.2f}"
                    except:
                        expedite_display_text = expedite_amount_text
                elif "$" in expedite_amount_text:
                    try:
                        expedite_dollar_amount = float(expedite_amount_text.replace("$", "").replace(",", "").strip())
                    except:
                        expedite_dollar_amount = 0.0

            if not has_expedite_selection:
                if base_cost_value > 0:
                    engineering_rows.append([
                        Paragraph(
                            f"Agreed Cost: ${base_cost_value:,.2f}"
                            "&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
                            "<b>Expedite?</b> ( [ ] Yes, 50% Extra [ ] No ): __________",
                            engineering_line_style
                        )
                    ])
                else:
                    engineering_rows.append([
                        Paragraph(
                            "Agreed Cost:&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
                            "<b>Expedite?</b> ( [ ] Yes, 50% Extra [ ] No ): __________",
                            engineering_line_style
                        )
                    ])
            else:
                expedite_status = job_data.get('expedite', False)
                engineering_rows.append([
                    Paragraph(
                        (
                            f"Agreed Cost: ${base_cost_value:,.2f}"
                            "&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;"
                            "<b>Expedite?</b> "
                            f"( [{'✔' if expedite_status else ' '}] Yes, 50% Extra "
                            f"[{'✔' if not expedite_status else ' '}] No ) : "
                            f"{expedite_display_text}"
                        ),
                        engineering_line_style
                    )
                ])

            if not has_expedite_selection:
                if base_cost_value > 0:
                    total_text = f"${base_cost_value:,.2f}"
                    engineering_rows.append([
                        Paragraph(
                            f"TOTAL: <b>{total_text}</b>",
                            engineering_line_style
                        )
                    ])
                else:
                    engineering_rows.append([
                        Paragraph(
                            "TOTAL:",
                            engineering_line_style
                        )
                    ])
            else:
                total = base_cost_value + expedite_dollar_amount
                if total > 0:
                    total_text = f"${total:,.2f}"
                    engineering_rows.append([
                        Paragraph(
                            f"TOTAL: <b>{total_text}</b>",
                            engineering_line_style
                        )
                    ])
                else:
                    engineering_rows.append([
                        Paragraph(
                            "TOTAL:",
                            engineering_line_style
                        )
                    ])

            engineering_table = Table(engineering_rows, colWidths=[doc.width])
            engineering_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 1.5),  # Reduced from 2
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),  # Reduced from 4
            ]))

            story.append(engineering_table)
            
            # -------------------------
            # FIXED: COURT APPEARANCE AND RATE SECTION - MATCHES ENGINEERING COSTS STYLING
            # -------------------------
            story.append(Spacer(1, 1.5 * mm))
            
            # Create a new style that matches engineering costs exactly
            court_rate_style = ParagraphStyle(
                name="CourtRateStyle",
                parent=engineering_line_style,  # Use engineering costs style as parent
                fontSize=10,  # Match field value size
                textColor=colors.black,
                alignment=0,
                leftIndent=0  # Match engineering costs indent
            )

            # Create table with same structure as engineering costs
            court_rate_rows = []
            court_rate_rows.append([
                Paragraph(
                    "<b>In the case of Court Appearance or Disposition:</b> N/A",
                    court_rate_style
                )
            ])
            court_rate_rows.append([
                Paragraph(
                    "Rate: $250/hour (portal-to-portal)",
                    court_rate_style
                )
            ])

            court_rate_table = Table(court_rate_rows, colWidths=[doc.width])
            
            court_rate_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 0.5),  # Minimal padding
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            
            story.append(court_rate_table)
            story.append(Spacer(1, 1.5 * mm))

            
            # -------------------------
            # FIXED: DELIVERABLES SECTION - COMPACT DESIGN
            # -------------------------
            add_title("Deliverables (Check all that apply):")
            
            deliverables = job_data.get("deliverables", [])
            
            if deliverables:
                # Clean and capitalize deliverables
                cleaned_deliverables = [d.strip().capitalize() for d in deliverables if d.strip()]
                
                # Always use 3 columns for consistent layout
                COLS = 3
                
                # Calculate how many rows we need
                total_items = len(cleaned_deliverables)
                ROWS = (total_items + COLS - 1) // COLS  # Ceiling division
                
                # Build grid rows
                grid_rows = []
                current_row = []
                
                for i, item in enumerate(cleaned_deliverables):
                    cell_text = f"[✔] {item}"
                    cell = Paragraph(f"<font size='9'>{cell_text}</font>", styles["ProfessionalText"])  # Reduced from 10
                    current_row.append(cell)
                    
                    # Complete row when we have enough items or reach the end
                    if (i + 1) % COLS == 0 or i == total_items - 1:
                        # Pad incomplete row with empty cells
                        while len(current_row) < COLS:
                            current_row.append(Paragraph("&nbsp;", styles["ProfessionalText"]))
                        
                        grid_rows.append(current_row)
                        current_row = []
                
                # Create table with dynamic row height based on number of rows
                available_width = doc.width
                col_width = available_width / COLS
                
                deliverables_table = Table(
                    grid_rows,
                    colWidths=[col_width] * COLS
                )
                
                deliverables_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                ]))

                story.append(deliverables_table)
            else:
                # No deliverables → show simple line
                empty_line = Table(
                    [[Paragraph("", styles["FieldValue"])]],
                    colWidths=[doc.width]
                )
                empty_line.setStyle(TableStyle([
                    ('TOPPADDING', (0,0), (-1,-1), 0),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 1.5),  # Reduced from 2
                ]))
                story.append(empty_line)
            
            story.append(Spacer(1, 0 * mm))
            
            # -------------------------
            # SERVICES SECTION - COMPACT DESIGN
            # -------------------------

            add_title("Services Required :")

            # UPDATED: Now 8 default services (was 6)
            standard_services = [
                "Structural",
                "Civil", 
                "Electrical",
                "Mechanical",
                "Plumbing Design",
                "Anchor Calculations",
                "Solid Works",       # NEW
                "Foundation"         # NEW
            ]

            selected_services = job_data.get("services", [])

            # Extract "Others" typed services
            others_services = []
            for service in selected_services:
                if service.startswith("Others:"):
                    others_text = service.replace("Others:", "").strip()
                    if others_text:
                        others_items = [item.strip() for item in others_text.split(',') if item.strip()]
                        others_services.extend(others_items)

            # Build display list: ALL defaults first, then ALL typed
            display_items = []

            # 1. Add ALL default services (all 8, checked/unchecked)
            for srv in standard_services:
                is_checked = srv in selected_services
                display_items.append((srv, is_checked, "default"))

            # 2. Add ALL typed services (all checked)
            for typed_srv in others_services:
                display_items.append((typed_srv.capitalize(), True, "typed"))

            # Calculate total items
            total_items = len(display_items)

            # Determine grid layout EXACTLY AS BEFORE
            if total_items <= 9:
                # 3x3 grid (max 9 items)
                COLS = 3
                MAX_ITEMS = 9
            else:
                # 4x3 grid (max 12 items)
                COLS = 4
                MAX_ITEMS = 12

            # Handle overflow if needed (8 defaults + typed services)
            if total_items > MAX_ITEMS:
                # Count how many to remove
                to_remove = total_items - MAX_ITEMS
                
                # Remove unchecked default services first
                # Find unchecked defaults (they're at the beginning of the list)
                unchecked_default_indices = []
                for i, (name, checked, type_tag) in enumerate(display_items):
                    if type_tag == "default" and not checked:
                        unchecked_default_indices.append(i)
                
                # Remove as many as needed (starting from first unchecked)
                removed = 0
                for idx in unchecked_default_indices:
                    if removed >= to_remove:
                        break
                    # Adjust index for items already removed
                    adjusted_idx = idx - removed
                    if adjusted_idx < len(display_items):
                        display_items.pop(adjusted_idx)
                        removed += 1
                
                # If still need to remove more, remove typed services from end
                if removed < to_remove:
                    remaining_to_remove = to_remove - removed
                    display_items = display_items[:-remaining_to_remove] if remaining_to_remove > 0 else display_items
                
                # Recalculate total
                total_items = len(display_items)

            # Create grid based on final count
            if total_items <= 9:
                COLS = 3
                ROWS = 3  # Max 9 items
            else:
                COLS = 4
                ROWS = 3  # Max 12 items

            # Ensure we don't exceed grid capacity (shouldn't happen but just in case)
            if total_items > COLS * ROWS:
                display_items = display_items[:COLS * ROWS]
                total_items = len(display_items)

            # Build the grid
            grid_rows = []
            current_row = []

            for i, (service_name, is_checked, service_type) in enumerate(display_items):
                box = "[✔]" if is_checked else "[ ]"
                
                cell_text = f"{box} {service_name}"
                
                cell = Paragraph(
                    f"<font size='9' color='#333333' fontname='Helvetica'>{cell_text}</font>",  # Reduced from 10
                    styles["ProfessionalText"]
                )
                current_row.append(cell)
                
                # Complete row when we have enough items or reach the end
                if (i + 1) % COLS == 0 or i == total_items - 1:
                    # Pad incomplete row
                    while len(current_row) < COLS:
                        current_row.append(Paragraph("&nbsp;", styles["ProfessionalText"]))
                    
                    grid_rows.append(current_row)
                    current_row = []

            # Create and style the table
            if grid_rows:
                # Calculate column width (account for spacing)
                available_width = doc.width - (8 * (COLS - 1))  # Reduced from 10mm
                col_width = available_width / COLS
                
                services_table = Table(
                    grid_rows,
                    colWidths=[col_width] * COLS
                )
                
                services_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),  # Reduced from 5
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                ]))
                
                story.append(services_table)

            story.append(Spacer(1, 1 * mm))

            # -------------------------
            # 50% DOWN PAYMENT BANNER
            # -------------------------
            banner_text = Paragraph(
                "<b>A 50% DOWN PAYMENT IS REQUIRED TO INITIATE</b>",
                ParagraphStyle(
                    name="BannerText",
                    alignment=1,
                    fontName="Helvetica-Bold",
                    fontSize=9,  # Reduced from 10
                    textColor=colors.HexColor("#A20000")
                )
            )

            banner_table = Table(
                [[banner_text]],
                colWidths=[doc.width],
            )

            banner_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Reduced from 1
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TOPPADDING', (0, 0), (-1, -1), 1.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story.append(Paragraph(
                "Payment Information",
                ParagraphStyle(name='Left12PDF', parent=styles['LabelBold'], alignment=0, fontSize=11)  # Reduced from 12
            ))
            story.append(Spacer(1, 2 * mm))

            story.append(banner_table)

            available_width = doc.width

            # Payment options
            left_section = [
                Table(
                    [[Paragraph("<para leftIndent='5'><b>Option 1: Check</b></para>", styles['LabelBold'])]],  # Reduced indent
                    colWidths=[available_width * 0.60],
                    rowHeights=[14],  # Reduced from 14
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#B6DDE8")),
                        ('BOX', (0, 0), (-1, -1), 0.8, colors.black),  # Reduced from 0.8
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ])
                ),
                Table(
                    [[Paragraph(
                        f"<para align='left' leading='13'>"
                        f"<b>Payable to:</b> {company_name}<br/>"
                        f"<b>Mailing Address:</b> {company_address}"
                        f"</para>",
                        styles['FieldValue']
                    )]],
                    colWidths=[available_width * 0.60],
                    rowHeights=[40],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 8),  # Reduced from 8
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),  # Reduced from 4
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ])
                ),
                Spacer(1, 1 * mm),
                Table(
                    [[Paragraph("<para leftIndent='5'><b>Option 3: Bank ACH Transfer</b></para>", styles['LabelBold'])]],
                    colWidths=[available_width * 0.60],
                    rowHeights=[14],  # Reduced from 14
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#EA9999")),
                        ('BOX', (0, 0), (-1, -1), 0.7, colors.black),  # Reduced from 0.8
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ])
                ),
                Table(
                    [[Paragraph(
                        "<para align='left' leading='12'>"  # Reduced from 12
                        "<font size='10'>"  # Reduced from 10
                        "<b>Account Type:</b> Checking<br/>"
                        "<b>Bank Name:</b> BMO Harris Bank<br/>"
                        "Routing Number: 071025661<br/>"
                        "Acct. Number: 4834994317"
                        "</font>"
                        "</para>",
                        styles['FieldValue']
                    )]],
                    colWidths=[available_width * 0.60],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('LEFTPADDING', (0, 0), (-1, -1), 8),  # Reduced from 8
                    ])
                ),
            ]

            right_section = [
                Table(
                    [[Paragraph("<para leftIndent='5'><b>Option 2: Zelle QR code</b></para>", styles['LabelBold'])]],
                    colWidths=[available_width * 0.40],
                    rowHeights=[14],  # Reduced from 14
                    style=TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#B6D7A8")),
                        ('BOX', (0, 0), (-1, -1), 0.8, colors.black),  # Reduced from 0.8
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ])
                ),
                Spacer(1, 1 * mm),
                Table(
                    [[Image(str(venmo_qr_path), width=98, height=98)]],
                    style=TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 2
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 2
                    ])
                ),
                Paragraph(
                    "<para align='center'><font size='8'>Scan to pay with Zelle</font></para>",  # Reduced from 9
                    styles['FieldValue']
                )
            ]

            payment_data = [[left_section, right_section]]

            payment_table = Table(
                payment_data,
                colWidths=[available_width * 0.60, available_width * 0.40]
            )

            payment_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),  # Reduced from 1
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),  # Reduced from 0.5
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))

            story.append(payment_table)
            story.append(Spacer(1, 2 * mm))

            # -------------------------
            # CLIENT AGREEMENT SECTION
            # -------------------------
            def add_client_agreement_title(label):
                title_para = Paragraph(f"<b>{label}</b>", styles["LabelBold"])
                title_table = Table([[title_para]], colWidths=[doc.width])
                title_table.setStyle(TableStyle([
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 2
                    ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 2
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ]))
                story.append(title_table)
                story.append(Spacer(1, 0.6 * mm))  # Reduced from 0.6mm

            add_client_agreement_title("Client Agreement :")

            agreement_style = ParagraphStyle(
                name="AgreementText",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=8,
                leading=10,
                textColor=colors.black,
                leftIndent=20,  # Reduced from 20
            )
            
            story.append(Spacer(1, 0.5 * mm))
            story.append(Paragraph(
                "By signing below, the client agrees to provide necessary documents, respond to RFIs within 3 business days, "
                "and acknowledges that deliverables will be considered final if no response is received within 10 business days.",
                agreement_style
            ))

            story.append(Spacer(1, 1.5 * mm))

            story.append(Table([
                [Paragraph("Client Signature :", styles["FieldValue"]),
                Paragraph("Date :", styles["FieldValue"])]
            ], colWidths=[doc.width * 0.75, doc.width * 0.2]))

            story.append(Spacer(1, -2 * mm))  # Reduced from -2mm

            # -------------------------
            # BUILD PDF WITH FOOTER
            # -------------------------
            doc.build(story, onFirstPage=footer, onLaterPages=footer)
            _log.info("Quote Form PDF generated to temp location for Firebase: %s", filename)
            
            # ✅ Return tuple (success, temp_path) - will be saved to Firebase only
            return True, filename
            
        except Exception as e:
            _log.error("PDF GENERATION FAILED: %s", e, exc_info=True)
            return False, None
        
    def open_generated_pdf(self, job_data):
        """Open quote form PDF directly from Firebase"""
        try:
            job_number = job_data.get('job_number', '')
            
            if not job_number:
                _log.warning("No quote number found")
                return False
            
            # ✅ ONLY use Firebase - no local directory lookup
            if self.FIREBASE_AVAILABLE:
                from main import FirebaseManager
                
                # Create temporary directory for PDF download
                temp_dir = Path(tempfile.gettempdir()) / "mabs_jobforms_temp"
                temp_dir.mkdir(parents=True, exist_ok=True)
                temp_pdf_path = temp_dir / f"{job_number}_job_form.pdf"
                
                # Load PDF from Firebase
                pdf_path = FirebaseManager.load_job_pdf_from_firebase(job_number, temp_pdf_path)
                
                if pdf_path and pdf_path.exists():
                    # ✅ Use parent's method to open the file
                    parent_tab = self.parent()
                    if hasattr(parent_tab, 'open_job_form_pdf_file'):
                        if parent_tab.open_job_form_pdf_file(pdf_path):
                            _log.info("Opened quote form PDF from Firebase: %s", job_number)
                            
                            # Clean up temporary file after a delay
                            QtCore.QTimer.singleShot(10000, lambda: self.cleanup_temp_pdf_file(pdf_path))
                            return True
                        else:
                            _log.warning("Failed to open PDF from Firebase: %s", job_number)
                            return False
                    else:
                        _log.warning("Parent tab doesn't have open_job_form_pdf_file method")
                        return False
                else:
                    _log.warning("PDF not found in Firebase for job: %s", job_number)
                    QtWidgets.QMessageBox.warning(
                        self,
                        "PDF Not Found",
                        f"No PDF found in Firebase for job: {job_number}\n\n"
                        f"Please generate the PDF first by clicking 'Generate quote Form PDF'."
                    )
                    return False
            else:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Firebase Not Available",
                    "Firebase is not available. Cannot open PDF."
                )
                return False
                
        except Exception as e:
            _log.warning("Error opening quote form PDF from Firebase: %s", e)
            QtWidgets.QMessageBox.critical(
                self, "Error Opening PDF", 
                f"Could not open PDF from Firebase:\n{str(e)}"
            )
            return False
        
    def save_job_form_directly(self, job_data):
        """Direct Firebase save method with duplicate validation"""
        
        if not self.FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - saving quote form directly to local backup")
            return self.owner_tab.save_job_form_locally(job_data) if hasattr(self.owner_tab, 'save_job_form_locally') else False
            
        try:
            from main import db
            ref = db.reference('/job_forms')
            
            job_number = job_data['job_number']
            
            # ⭐⭐ NEW: Check for duplicate (only for new jobs)
            if not self.is_editing:
                existing_jobs = ref.order_by_child('job_number').equal_to(job_number).get()
                
                if existing_jobs:
                    return False

            # Check if job already exists (by firebase_id if editing, otherwise by job_number)
            if 'firebase_id' in job_data and job_data['firebase_id']:
                # Update existing job using firebase_id
                job_id = job_data['firebase_id']
                job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(job_id).update(job_data)
                return True
            else:
                # Check if job with same job_number already exists
                existing_jobs = ref.order_by_child('job_number').equal_to(job_data['job_number']).get()
                
                if existing_jobs:
                    # Update existing job
                    job_id = list(existing_jobs.keys())[0]
                    job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    ref.child(job_id).update(job_data)
                    _log.info("Quote form UPDATED in Firebase: %s", job_data['job_number'])
                    return True
                else:
                    # Create new job
                    new_job_ref = ref.push()
                    job_data['firebase_id'] = new_job_ref.key
                    job_data['created_at'] = datetime.now(timezone.utc).isoformat()
                    job_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                    new_job_ref.set(job_data)
                    _log.info("Quote form CREATED in Firebase with ID: %s", new_job_ref.key)
                    return True
        except Exception as e:
            _log.warning("Error saving quote form to Firebase: %s", e)
            _log.exception("Traceback:")
            return self.owner_tab.save_job_form_locally(job_data) if hasattr(self.owner_tab, 'save_job_form_locally') else False

    # Add this method to the JobFormDialog class
    def cleanup_temp_pdf_file(self, file_path: Path):
        """Clean up temporary PDF file"""
        try:
            if file_path.exists():
                file_path.unlink()
                _log.info("Cleaned up temporary PDF file: %s", file_path)
        except Exception as e:
            _log.warning("Could not clean up temporary PDF file: %s", e)
            
    def save_job_form_pdf_to_firebase(self, job_number: str, pdf_path: Path):
        """Save quote form PDF to Firebase Realtime Database as Base64"""
        try:
            # Check if Firebase is available
            if not self.FIREBASE_AVAILABLE:
                _log.warning("Firebase not available - job PDF not saved")
                return False
            
            # Import main's FirebaseManager
            try:
                from main import FirebaseManager
            except ImportError:
                _log.warning("Could not import FirebaseManager")
                return False
            
            # Call the FirebaseManager method to save PDF to Firebase
            success = FirebaseManager.save_job_pdf_to_firebase(job_number, pdf_path)
            
            if success:
                _log.info("Quote form PDF saved to Firebase: %s", job_number)
            else:
                _log.warning("Failed to save quote form PDF to Firebase: %s", job_number)
                
            return success
            
        except Exception as e:
            _log.warning("Error saving quote form PDF to Firebase: %s", e)
            _log.exception("Traceback:")
            return False
        
    def save_as_draft(self):
        """Save current form data as a Draft quote without generating a PDF."""
        if not self.job_number_edit.text() or self.job_number_edit.text() == "Auto-generated":
            self.generate_job_number()
        job_data = self.collect_job_form_data()
        job_data['status'] = 'Draft'
        if self.is_editing and self.job_data and 'firebase_id' in self.job_data:
            job_data['firebase_id'] = self.job_data['firebase_id']
        owner = self.owner_tab
        success = False
        if hasattr(owner, 'save_job_form_to_firebase'):
            success = owner.save_job_form_to_firebase(job_data)
        if success:
            QtWidgets.QMessageBox.information(
                self, "Draft Saved",
                f"Quote {job_data.get('job_number', '')} saved as Draft."
            )
            self.accept()
        else:
            QtWidgets.QMessageBox.warning(
                self, "Save Failed",
                "Could not save draft. Please try again."
            )

    def collect_job_form_data(self):
        """Collect all quote form data into a dictionary"""
        # Get selected services
        selected_services = []
        
        if self.structural_checkbox.isChecked():
            selected_services.append("Structural")
        
        if self.civil_checkbox.isChecked():
            selected_services.append("Civil")
        
        if self.electrical_checkbox.isChecked():
            selected_services.append("Electrical")
        
        if self.mechanical_checkbox.isChecked():
            selected_services.append("Mechanical")
        
        if self.plumbing_checkbox.isChecked():
            selected_services.append("Plumbing Design")
        
        if self.anchor_calc_checkbox.isChecked():
            selected_services.append("Anchor Calculations")
        
        # NEW SERVICES
        if self.solidworks_checkbox.isChecked():
            selected_services.append("Solid Works")
        
        if self.foundation_checkbox.isChecked():
            selected_services.append("Foundation")
        
        # Handle "Others" checkbox and input
        if self.other_checkbox.isChecked():
            others_text = self.others_input.text().strip()
            if others_text:
                # Add "Others:" prefix followed by user's text
                selected_services.append(f"Others: {others_text}")
        
        # Parse deliverables (comma-separated)
        deliverables_text = self.deliverables_edit.text().strip()
        deliverables = [d.strip() for d in deliverables_text.split(',') if d.strip()] if deliverables_text else []
        
        job_data = {
            'job_number': self.job_number_edit.text().strip(),
            # ==== CHANGED: Use Project Name instead of Job Title ====
            'project_name': self.project_name_edit.text().strip(),  # Changed from job_title
            'client': self.client_combo.currentText().strip(),
            'client_email': self.client_email_edit.text().strip(),
            'client_address': self.client_address_edit.text().strip(),
            # ==== REMOVED: Duplicate project_name field ====
            # 'project_name': self.project_name_edit.text().strip(),
            'plant': self.plant_edit.text().strip(),
            'project_site_address': self.project_site_edit.text().strip(),
            'sales': self.sales_combo.currentText().strip(),
            'scope_of_work': self.scope_of_work_edit.text().strip(),
            'engineering_costs': self.engineering_costs_edit.text().strip(),
            'expedite': self.expedite_yes.isChecked(),
            'expedite_amount': self.expedite_amount_edit.text() if (self.expedite_yes.isChecked() or self.expedite_no.isChecked()) else "",
            'job_type': self.job_type_combo.currentText().strip(),
            'status': self.priority_combo.currentText().strip(),
            'start_date': self.start_date_edit.date().toString("MM-dd-yyyy"),
            'due_date': self.due_date_edit.date().toString("MM-dd-yyyy"),
            'deliverables': deliverables,
            'services': selected_services,
            'updated_at': datetime.now(timezone.utc).isoformat()
        }
        
        # Add this after creating the engineering_costs_edit
        if not self.is_editing:
            job_data['created_at'] = datetime.now(timezone.utc).isoformat()
        
        return job_data

    def validate_form(self):
        """No mandatory fields — users can generate a quote with any amount of data."""
        return True
    

class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year  # Minimum year (1 AD)
        self.end_year = end_year      # Maximum year (9999 AD)
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Navigation buttons
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton("◀◀")
        self.prev_block_btn.setFixedSize(40, 30)
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)
        
        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)
        
        self.next_block_btn = QtWidgets.QPushButton("▶▶")
        self.next_block_btn.setFixedSize(40, 30)
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        
        layout.addLayout(nav_layout)
        
        # Year grid container
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create 3x3 grid of year buttons
        self.year_buttons = []
        
        # Calculate current 9-year block start
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        # Current year display
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        
        layout.addLayout(current_layout)
        
        # Update the grid
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        """Calculate which 9-year block a year belongs to"""
        # Formula: ((year - 1) // 9) * 9 + 1
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        """Update the 3x3 grid with years from current 9-year block"""
        # Generate 9 consecutive years starting from current_block_start
        years = []
        
        for i in range(9):
            year = self.current_block_start + i
            years.append(year)
        
        # Update block label
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        
        # Update button texts and styles
        current_year = datetime.now().year
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            
            # Check if year is within valid range (1-9999)
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            # Style based on selection and current year
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #2980b9;
                    }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            # Connect button click
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        """Select a year"""
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        """Go to previous 9-year block (unlimited past)"""
        self.current_block_start -= 9
        
        # Unlimited past - no lower bound check
        # If we go below year 1, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        """Go to next 9-year block (unlimited future)"""
        self.current_block_start += 9
        
        # Unlimited future - no upper bound check
        # If we go above year 9999, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        """Set the selected year"""
        # Ensure year is within valid range
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        
    def get_selected_year(self):
        """Get the selected year"""
        return self.selected_year
    
    # Add signal for year selection
    year_selected = QtCore.pyqtSignal(int)

class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection with unlimited years"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📅 Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        # Create YearCalendarGrid with unlimited years
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        # Selected year display
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        
        self.selected_label = QtWidgets.QLabel(f"")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        
        layout.addLayout(selected_layout)
        
        # Action buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Connect signals
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        """Update selected year display when year is changed in calendar"""
        self.current_year = year
    
    def on_select_clicked(self):
        """Emit signal with selected year and close popup"""
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        """Get the selected year"""
        return self.current_year
    
class PDFExportDialog(QtWidgets.QDialog):
    """Professional PDF Export Dialog for Quote Forms"""
    
    def __init__(self, parent=None, available_dates=None):
        super().__init__(parent)
        self.available_dates = available_dates or []
        self.export_range = "all"  # Default export range
        self.selected_dates = []
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("📊 Export Quote Forms")
        self.setFixedSize(700, 600)
        self.setStyleSheet("""
            PDFExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📤 Export Quote Forms")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 3px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 1px;
                text-align: center;
            }
        """)
        header.setFixedHeight(70)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_card.setMaximumHeight(250)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Quote Forms")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)

        # From date section
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(200, 40)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)

        # To date section
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(200, 40)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 6px 8px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 13px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)

        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month Selector
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)

        # Month and Year selection in one row
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(40)

        # Month selection
        month_container = QtWidgets.QHBoxLayout()
        month_label = QtWidgets.QLabel("Select Month:")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_container.addWidget(month_label)
        self.month_combo = _NoScrollComboBox()
        self.month_combo.setFixedSize(200, 45)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months()
        month_container.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_container)

        # Year selection for month export
        year_container = QtWidgets.QHBoxLayout()
        year_label_month = QtWidgets.QLabel("Select Year:")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container.addWidget(year_label_month)

        # Year field with calendar button
        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.year_calendar_btn_month.setFixedSize(50, 45)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_container.addWidget(self.year_edit_month)
        year_container.addWidget(self.year_calendar_btn_month)
        month_year_row_layout.addLayout(year_container)

        # Add stretch to push everything to the left
        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        # Year Selector
        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        # Year selection row
        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Select Year:")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)
        
        # Year field with calendar button
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        
        # Calendar button
        self.year_calendar_btn = QtWidgets.QPushButton("📅")
        self.year_calendar_btn.setFixedSize(50, 45)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)

        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        # Initially hide all date selection components
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        # Preview Section
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_card.setMaximumHeight(150)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel("Ready to export all quote forms as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)
        
        # Progress Bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
        
        # Connect signals for live preview updates
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.from_date.wheelEvent = lambda e: e.ignore()
        self.from_date.stepBy = lambda x: None
        self.to_date.wheelEvent = lambda e: e.ignore()
        self.to_date.stepBy = lambda x: None
        self.month_combo.wheelEvent = lambda e: e.ignore()
        self.month_combo.currentTextChanged.connect(self.update_preview)

    def show_year_popup(self):
        """Show separate popup window for year selection (year export)"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        # Center the popup relative to main dialog
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        # Show as separate window
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show separate popup window for year selection (month+year export)"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        # Center the popup relative to main dialog
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        # Show as separate window
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def populate_months(self):
        """Populate months combo box"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_range_changed(self, range_type):
        """Handle export range changes"""
        self.export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_preview()
    
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_range == "all":
            return {"range": "all"}
        
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toPyDate()
            to_date = self.to_date.date().toPyDate()
            return {"range": "date_range", "from_date": from_date, "to_date": to_date}
        
        elif self.export_range == "month":
            month = self.month_combo.currentIndex() + 1
            year = int(self.year_edit_month.text())
            return {"range": "month", "month": month, "year": year}
        
        elif self.export_range == "year":
            year = int(self.year_edit.text())
            return {"range": "year", "year": year}
    
    def start_export(self):
        """Start the export process"""
        # Prevent multiple executions
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
            
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            # Simulate export process
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):  # Check if still valid
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            # Store export parameters for parent to use after dialog closes
            self._export_params = export_params
            
            # Simply accept the dialog - let parent handle the actual export
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False
            
class JobDetailsDialog(QtWidgets.QDialog):
    """Dialog to display job details"""
    
    def __init__(self, job_data, parent=None):
        super().__init__(parent)
        self.job_data = job_data
        self.setWindowTitle(f"Job Details - {job_data['job_number']}")
        self.setModal(True)
        self.resize(600, 500)
        self.init_ui()
    
    def format_services(self):
        """Return only selected services, formatted neatly."""
        services = self.job_data.get('services', [])
        
        # Clean blank entries
        services = [s.strip() for s in services if s.strip()]
        
        # Updated service widgets list
        standard_services = [
            'Structural', 'Civil', 'Electrical',
            'Mechanical', 'Plumbing Design', 'Anchor Calculations',
            'Solid Works', 'Foundation'  # NEW
        ]
        
        # Separate standard services and typed "Others" services
        all_services = []
        
        for s in services:
            if s == "Others":
                continue  # Skip standalone "Others"
            elif s.startswith("Others:"):
                others_text = s.replace("Others:", "").strip()
                if others_text:
                    # Split comma-separated services and capitalize
                    items = [item.strip().capitalize() for item in others_text.split(',') if item.strip()]
                    all_services.extend(items)
            else:
                all_services.append(s)
        
        return ", ".join(all_services)

    def format_deliverables(self):
        """Return deliverables exactly as user typed."""
        deliverables = self.job_data.get("deliverables", [])
        deliverables = [d.strip() for d in deliverables if d.strip()]

        return ", ".join(deliverables)

            
    def init_ui(self):
        self.setStyleSheet("""
            QDialog { background: #ffffff; }
            QScrollArea { background: #ffffff; border: none; }
            QWidget#scroll_content { background: #ffffff; }
        """)
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header
        header = QtWidgets.QLabel(f"Quote form: {self.job_data['job_number']}")
        header.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #1a2a3a;
            margin-bottom: 20px;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
            background: transparent;
        """)
        layout.addWidget(header)

        # Details in a scroll area
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { background: white; border: none; }")
        content = QtWidgets.QWidget()
        content.setObjectName("scroll_content")
        content.setStyleSheet("background: white;")
        scroll.setWidget(content)
        
        content_layout = QtWidgets.QVBoxLayout(content)
        
        # Create detail rows
        details = [
            ("Job Title:", self.job_data.get("job_title", "")),
            ("Client:", self.job_data.get("client", "")),
            ("Client Email:", self.job_data.get("client_email", "")),
            ("Mailing Address:", self.job_data.get("client_address", "")),  # NEW
            ("Project Name:", self.job_data.get("project_name", "")),
            ("Plant:", self.job_data.get("plant", "")),
            ("Project Site:", self.job_data.get("project_site_address", "")),

            # FIXED: Sales shows correct field even if key is different
            ("Sales:", self.job_data.get("sales", "") or self.job_data.get("sales_person", "")),

            # FIXED: Priority / Status
            ("Status:", self.job_data.get("status", "")),

            ("Start Date:", self.job_data.get("start_date", "")),
            ("Due Date:", self.job_data.get("due_date", "")),
            ("Agreed Cost:", self.job_data.get("engineering_costs", "N/A")),
            ("Scope of Work:", self.job_data.get("scope_of_work", "")),

            # FIXED: Clean & correct deliverables
            ("Deliverables:", self.format_deliverables()),

            # FIXED: Selected services ONLY
            ("Services:", self.format_services()),

            ("Created:", self.job_data.get("created_at", "")[:19]),
        ]

        
        for label, value in details:
            row_layout = QtWidgets.QHBoxLayout()
            
            label_widget = QtWidgets.QLabel(label)
            label_widget.setStyleSheet(
                "font-weight: 700; color: #1e293b; min-width: 130px;"
                " font-size: 13px; font-family: 'Inter','Segoe UI',sans-serif;"
                " background: transparent;"
            )
            row_layout.addWidget(label_widget)

            value_widget = QtWidgets.QLabel(str(value))
            value_widget.setStyleSheet(
                "color: #334155; font-size: 13px;"
                " font-family: 'Inter','Segoe UI',sans-serif; background: transparent;"
            )
            value_widget.setWordWrap(True)
            row_layout.addWidget(value_widget, 1)
            
            content_layout.addLayout(row_layout)
            content_layout.addSpacing(8)
        
        content_layout.addStretch()
        layout.addWidget(scroll)
        
        # Close button
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)




