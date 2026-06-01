from PyQt5 import QtWidgets, QtCore, QtGui
from datetime import datetime, timezone
import calendar
from decimal import Decimal
import os
import json
from pathlib import Path
import tempfile
# Add this import at the top with the other imports

from app_logger import get_logger
from app_theme import configure_filter_button
_log = get_logger(__name__)
try:
    from expenses_tab import ExpensesTab
except ImportError:
    ExpensesTab = None
    _log.warning("Warning: Could not import ExpensesTab")
    
try:
    from main import FirebaseManager, FIREBASE_AVAILABLE, Config, Currency, Invoice
except ImportError:
    # Expected during first-pass circular import; firebase_admin import below sets FIREBASE_AVAILABLE
    FirebaseManager = None
    FIREBASE_AVAILABLE = False
    Config = type('Config', (), {})
    Currency = type('Currency', (), {'format': lambda x: f"${x}"})
    Invoice = None
# Import Firebase configuration dynamically from the running main file

# Try to import Firebase modules
try:
    import firebase_admin
    from firebase_admin import credentials, db
    FIREBASE_AVAILABLE = True
except ImportError:
    FIREBASE_AVAILABLE = False
# Add to the top of balance_sheet_tab.py after imports
class InvoiceRevenueLink:
    """Helper class to track which revenue entries came from invoices"""
    
    @staticmethod
    def get_invoice_number_from_revenue(revenue_data):
        """Extract invoice number from revenue data if it's invoice-linked"""
        # Check if this revenue entry came from an invoice
        if revenue_data.get('is_invoice') and 'invoice_number' in revenue_data:
            return revenue_data['invoice_number']
        return None
    
    @staticmethod
    def is_invoice_linked(revenue_data):
        """Check if revenue entry is linked to an invoice"""
        return revenue_data.get('is_invoice', False) and 'invoice_number' in revenue_data

class BalanceSheetFirebaseManager:
    """Handles Firebase operations for balance sheet - SEPARATE NODE"""
    
    @staticmethod
    def save_expense(expense_data: dict) -> bool:
        """Save expense to Firebase - uses 'balance_sheet_expenses' node"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - expense not saved to Firebase")
            return False
            
        try:
            # USE SEPARATE NODE FOR BALANCE SHEET EXPENSES
            ref = db.reference('balance_sheet_expenses')
            
            if 'firebase_id' in expense_data and expense_data['firebase_id']:
                expense_id = expense_data['firebase_id']
                data_to_save = {k: v for k, v in expense_data.items() if k != 'firebase_id'}
                data_to_save['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(expense_id).update(data_to_save)
                _log.info("Updated balance sheet expense: %s", expense_id)
            else:
                expense_data['created_at'] = datetime.now(timezone.utc).isoformat()
                expense_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                new_ref = ref.push(expense_data)
                expense_data['firebase_id'] = new_ref.key
                _log.info("Created balance sheet expense: %s", new_ref.key)
            return True
        except Exception as e:
            _log.warning("Error saving balance sheet expense: %s", e)
            return False
    
    @staticmethod
    def load_expenses(year: int = None) -> list:
        """Load expenses from balance sheet Firebase node (NOT from expenses tab node)"""
        if not FIREBASE_AVAILABLE:
            return []
        try:
            # CRITICAL: Use 'balance_sheet_expenses' not 'expenses'
            ref = db.reference('balance_sheet_expenses')
            expenses_data = ref.get()
            expenses = []
            if expenses_data:
                for key, value in expenses_data.items():
                    if isinstance(value, dict):
                        value['firebase_id'] = key
                        expenses.append(value)
                _log.info("Loaded %s balance sheet expenses", len(expenses))
            else:
                _log.info("No balance sheet expenses found")
            if year is not None:
                expenses = [
                    exp for exp in expenses
                    if BalanceSheetFirebaseManager._entry_year(exp) == int(year)
                ]
            return expenses
        except Exception as e:
            _log.warning("Error loading balance sheet expenses: %s", e)
            return []
    
    @staticmethod
    def load_expenses_by_year(year: int) -> list:
        """Load balance sheet expenses filtered by year"""
        all_expenses = BalanceSheetFirebaseManager.load_expenses()
        return [exp for exp in all_expenses if exp.get('year') == year]
    
    @staticmethod
    def delete_entry(collection: str, firebase_id: str) -> bool:
        """Delete entry from Firebase"""
        if not FIREBASE_AVAILABLE:
            return False
        try:
            # Map collection to correct node
            node_map = {
                'expenses': 'balance_sheet_expenses',  # Balance sheet expenses use this node
                'revenue': 'revenue',
                'salary': 'salary'
            }
            node = node_map.get(collection, collection)
            ref = db.reference(f'{node}/{firebase_id}')
            ref.delete()
            _log.info("Deleted from %s: %s", node, firebase_id)
            return True
        except Exception as e:
            _log.warning("Error deleting: %s", e)
            return False
    
    @staticmethod
    def save_revenue(revenue_data: dict) -> bool:
        """Save revenue to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - revenue not saved to Firebase")
            return False
            
        try:
            ref = db.reference('revenue')
            
            if 'created_at' not in revenue_data:
                revenue_data['created_at'] = datetime.now(timezone.utc).isoformat()
            if 'updated_at' not in revenue_data:
                revenue_data['updated_at'] = datetime.now(timezone.utc).isoformat()
            
            if 'due_date' not in revenue_data:
                revenue_data['due_date'] = 'N/A'
            
            if 'amount' in revenue_data and isinstance(revenue_data['amount'], (int, float, Decimal)):
                revenue_data['amount'] = str(float(revenue_data['amount']))
            
            _log.info("(converted from print, see git history)")
            
            if 'firebase_id' in revenue_data and revenue_data['firebase_id']:
                revenue_id = revenue_data['firebase_id']
                data_to_save = {k: v for k, v in revenue_data.items() if k != 'firebase_id'}
                data_to_save['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(revenue_id).update(data_to_save)
                _log.info("Updated revenue in Firebase with ID: %s", revenue_id)
            else:
                new_ref = ref.push(revenue_data)
                revenue_data['firebase_id'] = new_ref.key
                _log.info("Created new revenue in Firebase with ID: %s", new_ref.key)
            
            return True
        except Exception as e:
            _log.warning("Error saving revenue to Firebase: %s", e)
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def load_revenue(year: int = None) -> list:
        """Load ALL revenue from Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot load revenue")
            return []
            
        try:
            ref = db.reference('revenue')
            revenue_data = ref.get()
            revenue = []
            if revenue_data:
                for key, value in revenue_data.items():
                    if isinstance(value, dict):
                        value['firebase_id'] = key
                        revenue.append(value)
                _log.info("Loaded %s revenue entries from Firebase", len(revenue))
            if year is not None:
                revenue = [
                    rev for rev in revenue
                    if BalanceSheetFirebaseManager._entry_year(rev) == int(year)
                ]
            return revenue
        except Exception as e:
            _log.warning("Error loading revenue from Firebase: %s", e)
            return []
    
    @staticmethod
    def save_salary(salary_data: dict) -> bool:
        """Save salary to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - salary not saved to Firebase")
            return False
            
        try:
            ref = db.reference('salary')
            
            if 'firebase_id' in salary_data and salary_data['firebase_id']:
                salary_id = salary_data['firebase_id']
                data_to_save = {k: v for k, v in salary_data.items() if k != 'firebase_id'}
                data_to_save['updated_at'] = datetime.now(timezone.utc).isoformat()
                ref.child(salary_id).update(data_to_save)
                _log.info("Updated salary in Firebase with ID: %s", salary_id)
            else:
                salary_data['created_at'] = datetime.now(timezone.utc).isoformat()
                salary_data['updated_at'] = datetime.now(timezone.utc).isoformat()
                new_ref = ref.push(salary_data)
                salary_data['firebase_id'] = new_ref.key
                _log.info("Created new salary in Firebase with ID: %s", new_ref.key)
            return True
        except Exception as e:
            _log.warning("Error saving salary to Firebase: %s", e)
            return False
    
    @staticmethod
    def load_salary(year: int = None) -> dict:
        """Load ALL salary from Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot load salary")
            return {"Inside America": [], "Outside America": []}
            
        try:
            ref = db.reference('salary')
            salary_data = ref.get()
            salary = {"Inside America": [], "Outside America": []}
            if salary_data:
                for key, value in salary_data.items():
                    if isinstance(value, dict):
                        value['firebase_id'] = key
                        region = value.get('region', 'Inside America')
                        salary[region].append(value)
                _log.info("Loaded %s salary entries from Firebase", sum(len(v) for v in salary.values()))
            if year is not None:
                salary = {
                    region: [
                        item for item in items
                        if BalanceSheetFirebaseManager._entry_year(item) == int(year)
                    ]
                    for region, items in salary.items()
                }
            return salary
        except Exception as e:
            _log.warning("Error loading salary from Firebase: %s", e)
            return {"Inside America": [], "Outside America": []}

    @staticmethod
    def update_balance_sheet_on_project_completion(project_number: str, project_data: dict) -> bool:
        """Update balance sheet when project is marked as completed due to full payment"""
        try:
            # Add revenue entry for completed project
            project_amount = float(project_data.get("project_amount", 0) or 0)
            if project_amount <= 0:
                return False
                
            revenue_data = {
                "source": f"Project {project_number}",
                "amount": project_amount,
                "date": datetime.now().strftime("%Y-%m-%d"),
                "description": f"Revenue from completed project {project_number}",
                "is_invoice": False,
                "project_number": project_number,
                "completion_date": project_data.get("completion_date", datetime.now(timezone.utc).isoformat()),
                "category": "Project Revenue"
            }
            
            # Save to Firebase
            BalanceSheetFirebaseManager.save_revenue(revenue_data)
            _log.info(f"Balance sheet updated with completed project revenue: {project_number}")
            return True
        except Exception as e:
            _log.error(f"Error updating balance sheet on project completion: {e}")
            return False

    @staticmethod
    def _entry_year(entry: dict) -> int:
        try:
            if entry.get("year"):
                return int(entry.get("year"))
            date_text = str(
                entry.get("date", "")
                or entry.get("received_date", "")
                or entry.get("paid_date", "")
                or entry.get("created_at", "")
            ).strip()
            if "T" in date_text:
                date_text = date_text.split("T", 1)[0]
            for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%y", "%m/%d/%y", "%B %d, %Y"):
                try:
                    return datetime.strptime(date_text, fmt).year
                except ValueError:
                    continue
        except Exception:
            pass
        return datetime.now().year


class _AnnualRefreshSignaler(QtCore.QObject):
    """Thread-safe bridge: background threads emit do_refresh to trigger
    _refresh_annual_revenue_background on the GUI thread via queued connection."""
    do_refresh = QtCore.pyqtSignal()


# Module-level singleton — set in BalanceSheetTab.__init__, used by payment_tracker.py
_annual_refresh_signaler: '_AnnualRefreshSignaler | None' = None


class _RealtimeRevenueSignaler(QtCore.QObject):
    """Thread-safe bridge: Firebase listener thread emits revenue_updated to push
    live /revenue/ changes to the GUI thread without polling."""
    revenue_updated = QtCore.pyqtSignal(object)  # passes current revenue list (list[dict])


# Module-level singleton — created in BalanceSheetTab.__init__
_revenue_signaler: '_RealtimeRevenueSignaler | None' = None


class _FinLoadSignaler(QtCore.QObject):
    """Thread-safe bridge: background load thread emits loaded to trigger
    UI update on the GUI thread via queued connection."""
    loaded = QtCore.pyqtSignal()


class BalanceSheetTab(QtWidgets.QWidget):
    """Annual Financial Summary Tab with Firebase integration and filters"""
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

        # Wire up the thread-safe annual-summary refresh signal
        global _annual_refresh_signaler
        _annual_refresh_signaler = _AnnualRefreshSignaler()
        _annual_refresh_signaler.do_refresh.connect(
            self._refresh_annual_revenue_background,
            QtCore.Qt.QueuedConnection
        )

        # Wire up the date-range background-load signal
        self._fin_load_signaler = _FinLoadSignaler()
        self._fin_load_signaler.loaded.connect(
            self._after_date_range_load,
            QtCore.Qt.QueuedConnection
        )

        # Wire up the real-time Firebase revenue listener signal
        global _revenue_signaler
        _revenue_signaler = _RealtimeRevenueSignaler()
        _revenue_signaler.revenue_updated.connect(
            self._on_realtime_revenue_update,
            QtCore.Qt.QueuedConnection
        )
        self._revenue_listener_handle = None
        # Debounce timer: coalesces rapid listener events (multiple field writes)
        # into a single UI update so editing one entry doesn't trigger 5 redraws.
        self._listener_debounce_timer = QtCore.QTimer(self)
        self._listener_debounce_timer.setSingleShot(True)
        self._listener_debounce_timer.setInterval(150)  # 150 ms quiet period
        self._listener_pending_revenue: list = []
        self._listener_debounce_timer.timeout.connect(self._flush_realtime_update)
        # Start listener after initial load so both don't race on startup
        QtCore.QTimer.singleShot(3000, self._start_revenue_listener)

        # Current year
        self.current_year = datetime.now().year
        self.annual_summary_year = self.current_year
        # Set of invoice_numbers that have an is_invoice entry in /revenue/ (all years).
        # Populated on every full Firebase refresh so cross-year invoices are covered.
        self._all_invoiced_numbers: set = set()

        # Firebase availability
        self.FIREBASE_AVAILABLE = getattr(self.main_window, 'FIREBASE_AVAILABLE', FIREBASE_AVAILABLE)
        self.db = getattr(self.main_window, 'db', None)

        _log.info("BalanceSheetTab: Firebase available = %s", self.FIREBASE_AVAILABLE)

        # Current selected category
        self.current_category = "Revenue"
        self._fin_page = 1
        self._fin_per_page = 10
        self._fin_all_items = []

        # Data storage for transaction table (remains independent)
        self.expenses_data = []
        self.revenue_data = []
        self.salary_data = {"Inside America": [], "Outside America": []}
        
        # NEW: Separate data stores for annual summary
        self.annual_expenses_data = []
        self.annual_revenue_data = []
        self.annual_salary_data = {"Inside America": [], "Outside America": []}
        
        # Filter state
        self.selected_category_filter = "All Categories"
        
        # Initialize UI
        self.init_ui()

        # Load ALL financial data (both transaction and annual summary)
        self.load_all_financial_data()

    def refresh_data(self, auto=False):
        """Reload Finance data while preserving the current year/category/filter view."""
        current_category = getattr(self, "current_category", "Revenue")
        search_text = self.search_edit.text() if hasattr(self, "search_edit") else ""
        category_text = self.category_combo.currentText() if hasattr(self, "category_combo") else current_category
        self.load_all_financial_data()
        if hasattr(self, "search_edit"):
            self.search_edit.setText(search_text)
        if hasattr(self, "category_combo"):
            index = self.category_combo.findText(category_text)
            if index >= 0:
                self.category_combo.setCurrentIndex(index)
        self.on_category_changed(current_category)
        
    def init_ui(self):
        """Initialize the UI with professional design matching JobFormTab"""
        # OUTER layout
        outer_layout = QtWidgets.QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        # SCROLL AREA covering the entire tab
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)

        # SCROLL CONTAINER
        container = QtWidgets.QWidget()
        scroll.setWidget(container)

        # MAIN LAYOUT inside scroll area
        main_layout = QtWidgets.QVBoxLayout(container)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(8)

        # Add scroll area to the tab
        outer_layout.addWidget(scroll)

        # ── Header (title + year + export + add) ────────────────────────
        self.create_header_section(main_layout)

        # ── KPI cards row ────────────────────────────────────────────────
        self.create_stats_section(main_layout)

        # ── Middle row: Annual Summary (full width) ──────────────────────
        mid_row = QtWidgets.QHBoxLayout()
        mid_row.setSpacing(12)

        annual_card = QtWidgets.QFrame()
        annual_card.setStyleSheet("""
            QFrame { background:#ffffff; border:1px solid #e2e8f0; border-radius:12px; }
        """)
        annual_inner = QtWidgets.QVBoxLayout(annual_card)
        annual_inner.setContentsMargins(0, 0, 0, 0)
        annual_inner.setSpacing(0)
        self._build_annual_card_header(annual_inner)
        self._build_annual_table(annual_inner)
        annual_card.setFixedHeight(240)

        mid_row.addWidget(annual_card, 1)
        main_layout.addLayout(mid_row)

        # ── Transactions sub-window ───────────────────────────────────────
        self.create_finance_table_section(main_layout)

        # Push all content to the top — prevents the layout from distributing
        # extra viewport height into mid_row when the transactions table is short
        # (e.g. Salary mode with few rows), which created a gap below annual_card.
        main_layout.addStretch(1)

        # Set initial year display
        self.update_year_display()
        
    def create_header_section(self, layout):
        header_frame = QtWidgets.QFrame()
        header_frame.setStyleSheet("""
            QFrame { background:#ffffff; border:none; border-radius:12px; }
        """)
        h = QtWidgets.QHBoxLayout(header_frame)
        h.setContentsMargins(20, 14, 20, 14)
        h.setSpacing(12)

        title_col = QtWidgets.QVBoxLayout()
        title_col.setSpacing(3)
        t = QtWidgets.QLabel("Financial Overview")
        t.setStyleSheet("font-size:20px; font-weight:900; color:#0f172a;")
        s = QtWidgets.QLabel("Track revenue, expenses and salaries — all in one place")
        s.setStyleSheet("font-size:12px; font-weight:600; color:#64748b;")
        title_col.addWidget(t)
        title_col.addWidget(s)
        h.addLayout(title_col, 1)

        self.year_btn = QtWidgets.QPushButton(str(self.current_year))
        self.year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_btn.setFixedHeight(34)
        self.year_btn.setStyleSheet("""
            QPushButton {
                background:#f0fdf4; color:#0f766e;
                border:1px solid #bbf7d0; border-radius:8px;
                font-size:13px; font-weight:900; padding:0 16px;
            }
            QPushButton:hover { background:#dcfce7; }
        """)
        self.year_btn.clicked.connect(self.show_year_calendar)
        h.addWidget(self.year_btn)

        self.export_btn = QtWidgets.QPushButton("⬇  Export")
        self.export_btn.setFixedHeight(34)
        self.export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.export_btn.setStyleSheet("""
            QPushButton { background:#475569; color:white; border:none;
                border-radius:8px; font-weight:800; font-size:13px; padding:0 16px; }
            QPushButton:hover { background:#334155; }
        """)
        self.export_btn.clicked.connect(self.show_export_dialog)
        h.addWidget(self.export_btn)

        self.add_btn = QtWidgets.QPushButton("+ Add Entry")
        self.add_btn.setFixedHeight(34)
        self.add_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.add_btn.setStyleSheet("""
            QPushButton { background:#0f766e; color:white; border:none;
                border-radius:8px; font-weight:900; font-size:13px; padding:0 16px; }
            QPushButton:hover { background:#0d625c; }
        """)
        self.add_btn.clicked.connect(self.open_add_dialog)
        self.add_btn.setVisible(False)  # hidden by default (Revenue is initial category)
        h.addWidget(self.add_btn)

        layout.addWidget(header_frame)

    def _fetch_data_background(self, _invoice_override=None):
        """Firebase reads only — NO Qt calls. Safe to run from a background thread.
        After this returns, call _apply_fetched_data_ui() on the main thread.
        Pass _invoice_override (captured on the main thread before spawning the worker)
        so a slow Firebase propagation never reverts a status that invoice_history already applied."""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            return
        try:
            _log.info("Loading all financial data for year %s...", self.current_year)

            all_expenses = BalanceSheetFirebaseManager.load_expenses()
            self.expenses_data = [exp for exp in all_expenses if exp.get('year') == self.current_year]
            self.annual_expenses_data = self.expenses_data.copy()
            _log.info("Loaded %s balance sheet expenses", len(self.expenses_data))

            all_revenue = BalanceSheetFirebaseManager.load_revenue()
            all_revenue = BalanceSheetTab._dedup_is_payment_entries(all_revenue)
            # Apply the locally-cached status overrides so switching to this tab never
            # overwrites a status change that invoice_history_tab already committed.
            _omap = _invoice_override if _invoice_override is not None \
                else dict(getattr(self, '_invoice_status_map', {}))
            inv_map = BalanceSheetTab._build_invoice_status_map()
            inv_map.update(_omap)      # invoice_history overrides always win
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue, inv_map)
            self.revenue_data = [
                rev for rev in all_revenue
                if BalanceSheetFirebaseManager._entry_year(rev) == self.current_year
            ]
            self.annual_revenue_data = [
                rev for rev in all_revenue
                if BalanceSheetFirebaseManager._entry_year(rev) == self.annual_summary_year
            ]
            _log.info("Loaded %s revenue entries", len(self.revenue_data))

            all_salary = BalanceSheetFirebaseManager.load_salary()
            self.salary_data = {"Inside America": [], "Outside America": []}
            self.annual_salary_data = {"Inside America": [], "Outside America": []}
            for region in ["Inside America", "Outside America"]:
                self.salary_data[region] = [sal for sal in all_salary[region] if sal.get('year') == self.current_year]
                self.annual_salary_data[region] = self.salary_data[region].copy()
            _log.info("Loaded %s salary entries", sum(len(v) for v in self.salary_data.values()))
        except Exception as e:
            _log.warning("Error fetching financial data in background: %s", e)

    def _apply_fetched_data_ui(self):
        """Qt widget updates after _fetch_data_background(). Must run on the main thread."""
        try:
            self.update_stats_cards()
            self.update_annual_summary()
            self.on_category_changed(self.current_category)
        except Exception as e:
            _log.warning("Error applying fetched financial data to UI: %s", e)

    def load_all_financial_data(self):
        """Load ALL financial data - both for transaction table and annual summary.
        Runs entirely on the calling thread (main thread only)."""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - using local data")
            self.load_local_data()
            return
        _omap = dict(getattr(self, '_invoice_status_map', {}))
        self._fetch_data_background(_omap)
        self._apply_fetched_data_ui()

    def refresh_on_tab_show(self):
        """Refresh all balance sheet data when the tab becomes visible.
        Runs the Firebase fetch in a background thread so the UI stays responsive."""
        import threading as _threading
        # Capture the override map on the main thread NOW so the background worker
        # uses the correct statuses even if Firebase hasn't propagated the latest write.
        _override_map = dict(getattr(self, '_invoice_status_map', {}))
        def _bg(_omap=_override_map):
            try:
                self._fetch_data_background(_omap)
            except Exception as _e:
                _log.warning("refresh_on_tab_show fetch failed: %s", _e)
            QtCore.QTimer.singleShot(0, self._apply_fetched_data_ui)
        _threading.Thread(target=_bg, daemon=True).start()

    def fix_date_edit(self, date_edit, set_today=True):
        """Make QDateEdit stable + optionally set today's date"""

        # ✅ Set today's date (only for new fields)
        if set_today:
            date_edit.setDate(QtCore.QDate.currentDate())

        # Disable scroll
        date_edit.wheelEvent = lambda event: None

        # Disable arrow keys
        def keyPressEvent(event, original=date_edit.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        date_edit.keyPressEvent = keyPressEvent

        # Disable stepping
        date_edit.stepBy = lambda x: None

        # Remove spin buttons
        date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
        
    def load_transaction_data_local(self):
        """Load transaction data from local JSON files"""
        try:
            data_dir = Path.home() / ".mabs_finance"
            
            expense_file = data_dir / f"expenses_{self.current_year}.json"
            if expense_file.exists():
                with open(expense_file, 'r') as f:
                    self.expenses_data = json.load(f)
            else:
                self.expenses_data = []
                
            revenue_file = data_dir / f"revenue_{self.current_year}.json"
            if revenue_file.exists():
                with open(revenue_file, 'r') as f:
                    self.revenue_data = json.load(f)
            else:
                self.revenue_data = []
                
            salary_file = data_dir / f"salary_{self.current_year}.json"
            if salary_file.exists():
                with open(salary_file, 'r') as f:
                    self.salary_data = json.load(f)
            else:
                self.salary_data = {"Inside America": [], "Outside America": []}
                
        except Exception as e:
            _log.warning("Error loading transaction local data: %s", e)
            self.expenses_data = []
            self.revenue_data = []
            self.salary_data = {"Inside America": [], "Outside America": []}
            
    def _create_breakdown_card(self, title, main_value, color, sub_rows):
        """Expanded stat card: title + main value + sub-value rows.

        sub_rows: list of (label_text, initial_value, text_color) tuples.
        Returns (card, main_value_label, [sub_value_labels...]).
        """
        BG_MAP = {
            "#27ae60": ("#f0fdf4", "#bbf7d0"),
            "#e74c3c": ("#fff7f7", "#fecaca"),
            "#3498db": ("#eff6ff", "#bfdbfe"),
        }
        ACCESSIBLE_TEXT = {
            "#27ae60": "#15803d",
            "#e74c3c": "#b91c1c",
            "#3498db": "#1d4ed8",
        }
        bg, border = BG_MAP.get(color, ("#f8fafc", "#e2e8f0"))
        main_color = ACCESSIBLE_TEXT.get(color, color)

        card = QtWidgets.QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {bg};
                border: 1px solid {border};
                border-radius: 10px;
            }}
        """)
        lay = QtWidgets.QVBoxLayout(card)
        lay.setContentsMargins(14, 8, 14, 8)
        lay.setSpacing(2)

        title_lbl = QtWidgets.QLabel(title.upper())
        title_lbl.setAlignment(QtCore.Qt.AlignCenter)
        title_lbl.setStyleSheet(
            "font-size:9px; font-weight:800; color:#64748b;"
            " letter-spacing:0.8px; background:transparent; border:none;")
        lay.addWidget(title_lbl)

        val_lbl = QtWidgets.QLabel(main_value)
        val_lbl.setAlignment(QtCore.Qt.AlignCenter)
        val_lbl.setObjectName("stat_value")
        val_lbl.setStyleSheet(
            f"font-size:19px; font-weight:900; color:{main_color};"
            " background:transparent; border:none;")
        lay.addWidget(val_lbl)

        # Divider
        div = QtWidgets.QFrame()
        div.setFrameShape(QtWidgets.QFrame.HLine)
        div.setStyleSheet(f"color:{border}; background:{border}; border:none; max-height:1px;")
        lay.addWidget(div)

        sub_labels = []
        for sub_text, sub_val, sub_color in sub_rows:
            sub_row = QtWidgets.QHBoxLayout()
            sub_row.setContentsMargins(0, 0, 0, 0)
            sub_row.setSpacing(4)
            lbl = QtWidgets.QLabel(sub_text)
            lbl.setStyleSheet(
                "font-size:9px; font-weight:700; color:#64748b;"
                " background:transparent; border:none;")
            val = QtWidgets.QLabel(sub_val)
            val.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            val.setStyleSheet(
                f"font-size:9px; font-weight:800; color:{sub_color};"
                " background:transparent; border:none;")
            sub_row.addWidget(lbl)
            sub_row.addWidget(val, 1)
            lay.addLayout(sub_row)
            sub_labels.append(val)

        return card, val_lbl, sub_labels

    def create_stats_section(self, layout):
        """Four KPI cards with hover tooltips showing breakdown details."""
        total_expenses = self.calculate_total_expenses()
        total_revenue  = self.calculate_total_revenue()
        total_salary   = self.calculate_total_salary()
        net_profit     = total_revenue - total_expenses - total_salary

        paid_rev   = sum(self._money_to_float(r.get('amount', 0))
                         for r in self.revenue_data
                         if str(r.get('status', '')).strip().lower() == 'paid')
        unpaid_rev = total_revenue - paid_rev
        sal_inside  = sum(self._money_to_float(s.get('amount', 0))
                          for s in self.salary_data.get('Inside America', []))
        sal_outside = sum(self._money_to_float(s.get('amount', 0))
                          for s in self.salary_data.get('Outside America', []))
        net_color = "#27ae60" if net_profit >= 0 else "#e74c3c"

        row = QtWidgets.QHBoxLayout()
        row.setSpacing(0)

        self.revenue_card, self.revenue_value_label = self.create_stat_card(
            "Total Revenue", f"${total_revenue:,.2f}", "#27ae60", icon="revenue")
        self.revenue_card.setToolTip(
            f"<b>Paid Revenue:</b> ${paid_rev:,.2f}<br>"
            f"<b>Unpaid Revenue:</b> ${unpaid_rev:,.2f}")

        self.expenses_card, self.expenses_value_label = self.create_stat_card(
            "Total Expenses", f"${total_expenses:,.2f}", "#e74c3c", icon="expenses")
        self.expenses_card.setToolTip(
            f"<b>Total:</b> ${total_expenses:,.2f}<br>"
            f"<b>Entries:</b> {len(self.expenses_data)}")

        self.salary_card, self.salary_value_label = self.create_stat_card(
            "Total Salaries", f"${total_salary:,.2f}", "#3498db", icon="salary")
        self.salary_card.setToolTip(
            f"<b>Inside America:</b> ${sal_inside:,.2f}<br>"
            f"<b>Outside America:</b> ${sal_outside:,.2f}")

        self.net_card, self.net_value_label = self.create_stat_card(
            "Net Profit / Loss", f"${net_profit:,.2f}", net_color, icon="netpl")
        self.net_value_label.setStyleSheet(
            f"font-size:19px; font-weight:900; color:"
            f"{'#15803d' if net_profit >= 0 else '#b91c1c'};"
            " background:transparent; border:none;")
        self.net_card.setToolTip(
            f"<b>Revenue:</b> ${total_revenue:,.2f}<br>"
            f"<b>Expenses:</b> -${total_expenses:,.2f}<br>"
            f"<b>Salaries:</b> -${total_salary:,.2f}")

        row.addStretch()
        for card in (self.revenue_card, self.expenses_card, self.salary_card, self.net_card):
            row.addWidget(card)
            row.addStretch()

        layout.addLayout(row)
    
    def refresh_invoice_revenues(self):
        """Refresh only invoice-linked revenue data from Firebase"""
        try:
            if not self.FIREBASE_AVAILABLE:
                return

            _log.info("Refreshing invoice revenues in balance sheet...")

            # Reload all revenue data from Firebase, applying any locally-cached
            # status overrides so a slow Firebase propagation doesn't revert a
            # status change that invoice_history_tab already applied.
            all_revenue = BalanceSheetFirebaseManager.load_revenue()
            _override_map = dict(getattr(self, '_invoice_status_map', {}))
            inv_map = BalanceSheetTab._build_invoice_status_map()
            inv_map.update(_override_map)   # signal overrides always win
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(
                all_revenue, inv_map)

            # Filter by current year for transaction table
            self.revenue_data = [
                rev for rev in all_revenue
                if BalanceSheetFirebaseManager._entry_year(rev) == self.current_year
            ]

            # Update annual summary data for current year
            self.annual_revenue_data = [
                rev for rev in all_revenue
                if BalanceSheetFirebaseManager._entry_year(rev) == self.annual_summary_year
            ]
            
            # Refresh the display based on current category
            if self.current_category == "Revenue":
                self.populate_revenue_data()
            
            # Update stats and annual summary
            self.update_stats_cards()
            self.update_annual_summary()
            
            _log.info("Refreshed invoice revenues: %s revenue entries", len(self.revenue_data))
            
        except Exception as e:
            _log.warning("Error refreshing invoice revenues: %s", e)
            import traceback
            traceback.print_exc()
            
            
    def load_financial_data_for_year(self, year):
        """Load data temporarily for export without changing UI year"""
        try:
            self.expenses_data = BalanceSheetFirebaseManager.load_expenses(year)
            all_revenue = BalanceSheetFirebaseManager.load_revenue(year)
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue)
            self.revenue_data = all_revenue
            self.salary_data = BalanceSheetFirebaseManager.load_salary(year)
        except Exception as e:
            _log.warning("Error loading data for export: %s", e)
            
    def show_export_dialog(self):
        """Show export options dialog"""
        dialog = ExportDialog(self)
        dialog.exec_()

    def create_stat_card(self, title, value, color, icon="●"):
        """Stat card — coloured icon circle left, title + value centred H & V on right."""
        BG_MAP = {
            "#e74c3c": ("#fff7f7", "#fecaca"),
            "#27ae60": ("#f0fdf4", "#bbf7d0"),
            "#3498db": ("#eff6ff", "#bfdbfe"),
        }
        ACCESSIBLE_TEXT = {
            "#27ae60": "#15803d",
            "#e74c3c": "#b91c1c",
            "#3498db": "#1d4ed8",
        }
        ICON_BG = {
            "#27ae60": ("#bbf7d0", "#16a34a"),   # green-200 tint
            "#e74c3c": ("#fecaca", "#dc2626"),   # red-200 tint
            "#3498db": ("#bfdbfe", "#2563eb"),   # blue-200 tint
        }
        bg, border       = BG_MAP.get(color, ("#f8fafc", "#e2e8f0"))
        text_color       = ACCESSIBLE_TEXT.get(color, color)
        icon_bg, icon_fg = ICON_BG.get(color, ("#f1f5f9", "#475569"))

        card = QtWidgets.QFrame()
        card.setFixedSize(220, 86)          # slightly wider than before
        card.setStyleSheet(f"""
            QFrame {{
                background: {bg};
                border: 1px solid {border};
                border-radius: 10px;
            }}
        """)
        card.setAccessibleName(f"{title} card")

        lay = QtWidgets.QHBoxLayout(card)
        lay.setContentsMargins(14, 0, 14, 0)   # 0 top/bottom — vertical centring done by stretch
        lay.setSpacing(12)
        lay.setAlignment(QtCore.Qt.AlignVCenter)

        # ── Painted icon (QPainter-drawn, matches design reference) ──────────
        icon_px  = BalanceSheetTab._paint_stat_icon(icon, icon_bg, icon_fg, size=44)
        icon_lbl = QtWidgets.QLabel()
        icon_lbl.setFixedSize(44, 44)
        icon_lbl.setAlignment(QtCore.Qt.AlignCenter)
        icon_lbl.setPixmap(icon_px)
        icon_lbl.setStyleSheet("background:transparent; border:none;")

        # ── Right column: title + value, both centred H & V ───────────────
        col = QtWidgets.QVBoxLayout()
        col.setSpacing(3)
        col.setContentsMargins(0, 0, 0, 0)
        col.addStretch(1)                      # push content to vertical centre

        title_label = QtWidgets.QLabel(title.upper())
        title_label.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        title_label.setWordWrap(True)
        title_label.setStyleSheet(
            "font-size:10px; font-weight:800; color:#64748b;"
            " letter-spacing:0.6px; background:transparent; border:none;")

        value_label = QtWidgets.QLabel(value)
        value_label.setAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        value_label.setStyleSheet(
            f"font-size:19px; font-weight:900; color:{text_color};"
            " background:transparent; border:none;")
        value_label.setAccessibleName(f"{title}: {value}")

        col.addWidget(title_label)
        col.addWidget(value_label)
        col.addStretch(1)                      # equal stretch below = perfectly centred

        lay.addWidget(icon_lbl)
        lay.addLayout(col, 1)
        return card, value_label

    # ------------------------------------------------------------------ #
    # Annual table icon helpers                                           #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _make_row_header_icon(color_hex: str, symbol: str, size: int = 22) -> "QtGui.QIcon":
        """Plain coloured symbol on transparent background — no filled box/circle."""
        px = QtGui.QPixmap(size, size)
        px.fill(QtCore.Qt.transparent)
        p = QtGui.QPainter(px)
        p.setRenderHint(QtGui.QPainter.Antialiasing)
        p.setPen(QtGui.QColor(color_hex))           # symbol drawn in the accent color
        p.setFont(QtGui.QFont("Segoe UI", max(9, size // 2), QtGui.QFont.Bold))
        p.drawText(QtCore.QRect(0, 0, size, size), QtCore.Qt.AlignCenter, symbol)
        p.end()
        return QtGui.QIcon(px)

    def _apply_annual_row_headers(self) -> None:
        """Set coloured icon + label for every row and stamp MONTH in the corner."""
        _rows = [
            ("REVENUE",  "#16a34a", "↗"),
            ("EXPENSES", "#dc2626", "⊖"),
            ("NET P/L",  "#16a34a", "◑"),
        ]
        vh = self.annual_table.verticalHeader()
        vh.setIconSize(QtCore.QSize(22, 22))        # ensure icons render at full symbol size

        for row, (label, color, sym) in enumerate(_rows):
            item = QtWidgets.QTableWidgetItem(f"  {label}")
            item.setIcon(self._make_row_header_icon(color, sym))
            self.annual_table.setVerticalHeaderItem(row, item)

        # Stamp "MONTH" in the top-left corner where the two headers meet
        corner = self.annual_table.findChild(QtWidgets.QAbstractButton)
        if corner:
            corner.setText("MONTH")
            corner.setStyleSheet(
                "QAbstractButton{"
                "background:#f8fafc; color:#475569;"
                "font-size:10px; font-weight:800; letter-spacing:0.5px;"
                "border:none;"
                "border-right:1px solid #e2e8f0;"
                "border-bottom:1px solid #e2e8f0;"
                "}")

    # ------------------------------------------------------------------ #
    # Painted icon helpers (stat cards + action buttons)                  #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _paint_stat_icon(icon_type: str, bg_hex: str, fg_hex: str, size: int = 50) -> "QtGui.QPixmap":
        """
        Paint a professional stat-card icon.
        icon_type: 'revenue' | 'expenses' | 'salary' | 'netpl'
        bg_hex : light-tint background (e.g. #dcfce7)
        fg_hex : accent color (e.g. #16a34a)
        """
        px = QtGui.QPixmap(size, size)
        px.fill(QtCore.Qt.transparent)
        p = QtGui.QPainter(px)
        p.setRenderHint(QtGui.QPainter.Antialiasing)

        # Perfect circle background
        p.setBrush(QtGui.QBrush(QtGui.QColor(bg_hex)))
        p.setPen(QtCore.Qt.NoPen)
        p.drawEllipse(0, 0, size, size)

        white = QtGui.QColor("#ffffff")
        fg    = QtGui.QColor(fg_hex)
        # Margin inside the circle — keep content away from the curved edge
        m = max(7, size // 5)

        if icon_type == "revenue":
            # Bold upward-right trending arrow drawn in fg (accent green)
            pw  = max(2, size // 7)
            pen = QtGui.QPen(fg, pw, QtCore.Qt.SolidLine,
                             QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin)
            p.setPen(pen)
            p.setBrush(QtCore.Qt.NoBrush)
            # Diagonal body of the arrow
            p.drawLine(m, size - m, size - m, m)
            # Arrowhead — two short wings at the top-right tip
            ah = max(5, size // 4)
            p.drawLine(size - m, m, size - m - ah, m)
            p.drawLine(size - m, m, size - m, m + ah)

        elif icon_type == "expenses":
            # Solid fg-coloured circle + white minus bar
            cx, cy = size // 2, size // 2
            cr = int(size * 0.36)
            p.setBrush(QtGui.QBrush(fg))
            p.setPen(QtCore.Qt.NoPen)
            p.drawEllipse(cx - cr, cy - cr, cr * 2, cr * 2)
            bar_len = int(cr * 1.05)
            bar_h   = max(2, size // 9)
            p.setBrush(QtGui.QBrush(white))
            p.drawRoundedRect(cx - bar_len, cy - bar_h // 2,
                              bar_len * 2, bar_h, bar_h // 2, bar_h // 2)

        elif icon_type == "salary":
            # People emoji 👥 rendered in the accent color, sized to fit the circle
            p.setPen(QtGui.QColor(fg_hex))
            p.setFont(QtGui.QFont("Segoe UI Emoji", max(14, int(size * 0.32)), QtGui.QFont.Bold))
            p.drawText(QtCore.QRect(0, 0, size, size), QtCore.Qt.AlignCenter, "👥")

        else:  # "netpl" — 3 ascending vertical bars in fg (accent green)
            p.setPen(QtCore.Qt.NoPen)
            p.setBrush(QtGui.QBrush(fg))
            bw     = max(3, size // 7)
            gap    = max(2, size // 9)
            bottom = int(size * 0.80)
            heights = [int(size * 0.28), int(size * 0.48), int(size * 0.68)]
            total_w = 3 * bw + 2 * gap
            sx = (size - total_w) // 2
            for i, h in enumerate(heights):
                p.drawRoundedRect(sx + i * (bw + gap), bottom - h,
                                  bw, h, bw // 3, bw // 3)
        p.end()
        return px

    @staticmethod
    def _make_action_icon(icon_type: str, color: str, size: int = 16) -> "QtGui.QIcon":
        """Paint a professional icon for table action buttons."""
        px = QtGui.QPixmap(size, size)
        px.fill(QtCore.Qt.transparent)
        p = QtGui.QPainter(px)
        p.setRenderHint(QtGui.QPainter.Antialiasing)

        c  = QtGui.QColor(color)
        pw = max(1, size // 7)

        if icon_type == "view":
            # Outlined eye: almond shape + filled pupil
            pen = QtGui.QPen(c, pw, QtCore.Qt.SolidLine,
                             QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin)
            p.setPen(pen)
            p.setBrush(QtCore.Qt.NoBrush)
            em = size // 5
            path = QtGui.QPainterPath()
            path.moveTo(em, size // 2)
            path.quadTo(size // 2, size // 5, size - em, size // 2)
            path.quadTo(size // 2, size * 4 // 5, em, size // 2)
            p.drawPath(path)
            p.setBrush(QtGui.QBrush(c))
            p.setPen(QtCore.Qt.NoPen)
            pr = max(2, size // 5)
            p.drawEllipse(size // 2 - pr, size // 2 - pr, pr * 2, pr * 2)

        elif icon_type == "edit":
            # Pencil: diagonal body + tip + cap line
            pen = QtGui.QPen(c, pw, QtCore.Qt.SolidLine,
                             QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin)
            p.setPen(pen)
            p.setBrush(QtCore.Qt.NoBrush)
            bw = size // 4   # body half-width
            m  = size // 6
            # Left & right edges of pencil body
            p.drawLine(m, size - m - bw, size - m - bw, m)
            p.drawLine(m + bw, size - m, size - m, m + bw)
            # Tip at bottom-left corner
            p.drawLine(m, size - m - bw, m + bw, size - m)
            # Cap line at top-right corner
            p.drawLine(size - m - bw, m, size - m, m + bw)
            # Small eraser/cap rule
            cap = bw // 2
            p.setPen(QtGui.QPen(c, pw + 1, QtCore.Qt.SolidLine,
                                QtCore.Qt.RoundCap))
            cx2, cy2 = size - m - cap, m + cap
            p.drawLine(cx2 - cap, cy2 - cap, cx2 + cap, cy2 + cap)

        else:  # "delete" — trash can
            pen = QtGui.QPen(c, pw, QtCore.Qt.SolidLine,
                             QtCore.Qt.RoundCap, QtCore.Qt.RoundJoin)
            p.setPen(pen)
            p.setBrush(QtCore.Qt.NoBrush)
            m  = max(2, size // 5)
            by = size // 3          # body top y
            bh = size - by - m // 2 # body height
            # Body
            p.drawRoundedRect(m, by, size - 2 * m, bh, 1, 1)
            # Lid
            p.drawLine(m - 1, by, size - m + 1, by)
            # Handle on lid
            hw = size // 3
            p.drawRoundedRect((size - hw) // 2, by - size // 5,
                              hw, size // 5, 2, 2)
            # Vertical stripes inside body
            gap3 = (size - 2 * m) // 3
            for i in range(1, 3):
                lx = m + i * gap3
                p.drawLine(lx, by + bh // 5, lx, by + bh * 4 // 5)

        p.end()
        return QtGui.QIcon(px)

    # ------------------------------------------------------------------ #
    #  Sub-window builders (Annual + Aging)                               #
    # ------------------------------------------------------------------ #

    def _sub_header(self, parent_layout, title, right_widgets=None):
        """Add a styled sub-window header bar with title and optional right-side widgets."""
        bar = QtWidgets.QFrame()
        bar.setStyleSheet("""
            QFrame {
                background: #f8fafc;
                border: none;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
            }
        """)
        h = QtWidgets.QHBoxLayout(bar)
        h.setContentsMargins(16, 10, 16, 10)
        h.setSpacing(10)
        lbl = QtWidgets.QLabel(title)
        lbl.setStyleSheet("font-size:13px; font-weight:900; color:#0f172a; background:transparent; border:none;")
        h.addWidget(lbl)
        h.addStretch()
        if right_widgets:
            for w in right_widgets:
                h.addWidget(w)
        parent_layout.addWidget(bar)

    def _build_annual_card_header(self, layout):
        # Year selector for annual table
        self.yearly_calendar_btn = QtWidgets.QPushButton(str(self.annual_summary_year))
        self.yearly_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.yearly_calendar_btn.setFixedHeight(28)
        self.yearly_calendar_btn.setStyleSheet("""
            QPushButton { background:#f0fdf4; color:#0f766e; border:1px solid #bbf7d0;
                border-radius:7px; font-size:12px; font-weight:900; padding:0 12px; }
            QPushButton:hover { background:#dcfce7; }
        """)
        self.yearly_calendar_btn.clicked.connect(self.show_annual_summary_year_calendar)

        self.expenses_breakdown_btn = QtWidgets.QPushButton("⊖  Expenses Breakdown")
        self.expenses_breakdown_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.expenses_breakdown_btn.setFixedHeight(28)
        self.expenses_breakdown_btn.setStyleSheet("""
            QPushButton { background:#fff7f7; color:#dc2626; border:1px solid #fecaca;
                border-radius:7px; font-size:12px; font-weight:800; padding:0 12px; }
            QPushButton:hover { background:#fee2e2; }
        """)
        self.expenses_breakdown_btn.clicked.connect(self.show_expenses_breakdown)

        self.salary_breakdown_btn = QtWidgets.QPushButton("◑  Salary Breakdown")
        self.salary_breakdown_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.salary_breakdown_btn.setFixedHeight(28)
        self.salary_breakdown_btn.setStyleSheet("""
            QPushButton { background:#eff6ff; color:#2563eb; border:1px solid #bfdbfe;
                border-radius:7px; font-size:12px; font-weight:800; padding:0 12px; }
            QPushButton:hover { background:#dbeafe; }
        """)
        self.salary_breakdown_btn.clicked.connect(self.show_salary_breakdown)

        self.annual_title = QtWidgets.QLabel("")  # kept for compat with update_year_display
        self.annual_title.hide()

        self._sub_header(layout, "Annual Financial Summary",
                         [self.yearly_calendar_btn, self.expenses_breakdown_btn, self.salary_breakdown_btn])

    def _build_annual_table(self, layout):
        import calendar as _cal
        self.annual_table = QtWidgets.QTableWidget()
        self.annual_table.setColumnCount(13)
        self.annual_table.setRowCount(3)
        self.annual_table.setHorizontalHeaderLabels(
            [_cal.month_abbr[i] for i in range(1, 13)] + ["Total"])
        self._apply_annual_row_headers()

        vh = self.annual_table.verticalHeader()
        vh.setDefaultAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
        vh.setFixedWidth(155)

        hh = self.annual_table.horizontalHeader()
        hh.setDefaultAlignment(QtCore.Qt.AlignCenter)

        self.annual_table.setStyleSheet("""
            QTableWidget {
                background:#ffffff; border:none;
                gridline-color:#f1f5f9; font-size:11px;
                alternate-background-color:#f8fafc;
            }
            QTableWidget:focus {
                border: 2px solid #0f766e;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding:8px 4px;
                border-bottom:1px solid #f1f5f9;
            }
            QTableWidget::item:selected { background:#e0f2fe; color:#0f172a; }
            QTableWidget::item:focus { border: 1px solid #0f766e; }
            QHeaderView::section {
                background:#f8fafc; color:#475569;
                font-weight:800; font-size:11px;
                padding:8px 4px; border:none;
                border-right:1px solid #e2e8f0;
                border-bottom:1px solid #e2e8f0;
            }
            QHeaderView::section:last {
                background:#f0fdf4; color:#0f766e; border-right:none;
            }
            QHeaderView::section:vertical {
                background:#f8fafc; color:#334155;
                font-weight:800; font-size:12px;
                padding:8px 12px; border:none;
                border-bottom:1px solid #e2e8f0;
            }
        """)

        self.annual_table.verticalHeader().setDefaultSectionSize(46)
        hdr = self.annual_table.horizontalHeader()
        hdr.setMinimumSectionSize(82)
        for i in range(12):
            hdr.setSectionResizeMode(i, QtWidgets.QHeaderView.Stretch)
        hdr.setSectionResizeMode(12, QtWidgets.QHeaderView.Fixed)
        self.annual_table.setColumnWidth(12, 145)
        self.annual_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.annual_table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.annual_table.setFont(QtGui.QFont("Inter", 9))
        # Exact fit: header ~30px + 3 rows × 46px = 168px; no scroll bar space needed
        self.annual_table.setFixedHeight(168)
        self.annual_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.annual_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.annual_table.setFocusPolicy(QtCore.Qt.TabFocus)
        self.annual_table.setAlternatingRowColors(True)
        self.annual_table.verticalHeader().setVisible(True)
        self.annual_table.setMouseTracking(True)
        self.annual_table.setAccessibleName("Annual Financial Summary table")
        self.annual_table.setAccessibleDescription(
            "Monthly revenue, expenses, and net profit/loss. Click a cell to view details.")

        # Clicking a Revenue-row month cell or the "Revenue" vertical header opens the detail popup
        self.annual_table.cellClicked.connect(self._on_annual_cell_clicked)
        self.annual_table.verticalHeader().sectionClicked.connect(
            self._on_annual_row_header_clicked)

        wrap = QtWidgets.QVBoxLayout()
        wrap.setContentsMargins(12, 10, 12, 12)
        wrap.addWidget(self.annual_table)
        layout.addLayout(wrap)

    def _build_aging_card(self, layout):
        self._sub_header(layout, "Invoice Aging")

        body = QtWidgets.QVBoxLayout()
        body.setContentsMargins(16, 14, 16, 14)
        body.setSpacing(10)

        self._aging_buckets = []
        bucket_defs = [
            ("0 – 30 days",  "Current",   "#10b981", "#f0fdf4", "#bbf7d0"),
            ("31 – 60 days", "Overdue",   "#f59e0b", "#fffbeb", "#fde68a"),
            ("61 + days",    "Past Due",  "#ef4444", "#fff7f7", "#fecaca"),
        ]
        for label_text, badge, color, bg, border in bucket_defs:
            cell = QtWidgets.QFrame()
            cell.setStyleSheet(f"""
                QFrame {{ background:{bg}; border:1px solid {border}; border-radius:8px; }}
            """)
            cell_lay = QtWidgets.QHBoxLayout(cell)
            cell_lay.setContentsMargins(12, 8, 12, 8)
            cell_lay.setSpacing(8)

            left = QtWidgets.QVBoxLayout()
            left.setSpacing(2)
            lbl = QtWidgets.QLabel(label_text)
            lbl.setStyleSheet(f"font-size:11px; font-weight:800; color:{color}; background:transparent; border:none;")
            badge_lbl = QtWidgets.QLabel(badge)
            badge_lbl.setStyleSheet(
                f"font-size:10px; font-weight:700; color:{color}; background:transparent; border:none;")
            left.addWidget(lbl)
            left.addWidget(badge_lbl)

            val_lbl = QtWidgets.QLabel("$0.00")
            val_lbl.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            val_lbl.setStyleSheet(
                f"font-size:16px; font-weight:900; color:{color}; background:transparent; border:none;")

            cell_lay.addLayout(left, 1)
            cell_lay.addWidget(val_lbl)
            body.addWidget(cell)
            self._aging_buckets.append(val_lbl)  # update_aging_section calls .setText() on each

        body.addStretch()
        layout.addLayout(body)

    def create_annual_summary_table(self, layout):
        """Create an attractive annual financial summary table with left calendar and right buttons"""
        # Group box for the annual summary
        summary_group = QtWidgets.QGroupBox()
        summary_group.setStyleSheet("""
            QGroupBox {
                font-weight: 900;
                font-size: 14px;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                margin-top: 0.5em;
                padding-top: 10px;
                background: #ffffff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 12px 0 12px;
                color: #0f766e;
                background: #f0fdf4;
                border-radius: 12px;
                font-size: 13px;
                font-weight: bold;
            }
        """)
        
        summary_layout = QtWidgets.QVBoxLayout(summary_group)
        summary_layout.setContentsMargins(10, 0, 10, 0)
        summary_layout.setSpacing(5)

        # Header row with calendar button on left, title in center, breakdown buttons on right
        header_widget = QtWidgets.QWidget()
        header_layout = QtWidgets.QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        header_layout.setSpacing(10)

        # Left side: Yearly Calendar Button
        self.yearly_calendar_btn = QtWidgets.QPushButton("📅 ")
        self.yearly_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.yearly_calendar_btn.setMinimumHeight(36)
        self.yearly_calendar_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                font-weight: bold;
                font-size: 16px;
                border-radius: 5px;
                border: 1px solid #21618c;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #5dade2, stop:1 #3498db);
            }
        """)
        self.yearly_calendar_btn.clicked.connect(self.show_annual_summary_year_calendar)

        # Center: Title
        self.annual_title = QtWidgets.QLabel(f"📈 ANNUAL FINANCIAL SUMMARY - {self.current_year}")
        self.annual_title.setAlignment(QtCore.Qt.AlignCenter)
        self.annual_title.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                padding: 0 20px;
            }
        """)

        # Right side: Breakdown Buttons
        buttons_layout = QtWidgets.QHBoxLayout()
        buttons_layout.setSpacing(8)

        self.expenses_breakdown_btn = QtWidgets.QPushButton("📊 Expenses Breakdown")
        self.expenses_breakdown_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.expenses_breakdown_btn.setMinimumHeight(36)
        self.expenses_breakdown_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e74c3c, stop:1 #c0392b);
                color: white;
                font-weight: bold;
                font-size: 12px;
                padding: 6px 15px;
                border-radius: 6px;
                border: 1px solid #922b21;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ec7063, stop:1 #e74c3c);
            }
        """)
        self.expenses_breakdown_btn.clicked.connect(self.show_expenses_breakdown)

        self.salary_breakdown_btn = QtWidgets.QPushButton("💰 Salary Breakdown")
        self.salary_breakdown_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.salary_breakdown_btn.setMinimumHeight(36)
        self.salary_breakdown_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                font-weight: bold;
                font-size: 12px;
                padding: 6px 15px;
                border-radius: 6px;
                border: 1px solid #1f618d;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #5dade2, stop:1 #3498db);
            }
        """)
        self.salary_breakdown_btn.clicked.connect(self.show_salary_breakdown)

        header_layout.addWidget(self.yearly_calendar_btn)
        header_layout.addStretch(1)
        header_layout.addWidget(self.annual_title)
        header_layout.addStretch(1)
        buttons_layout.addWidget(self.expenses_breakdown_btn)
        buttons_layout.addWidget(self.salary_breakdown_btn)
        header_layout.addLayout(buttons_layout)

        summary_layout.addWidget(header_widget)

        # Create table widget
        self.annual_table = QtWidgets.QTableWidget()
        self.annual_table.setColumnCount(13)  # 12 months + Total
        self.annual_table.setRowCount(3)  # Expenses, Revenue, and Net Profit/Loss
        
        month_headers = [calendar.month_abbr[i] for i in range(1, 13)] + ["Total"]
        self.annual_table.setHorizontalHeaderLabels(month_headers)

        # Set vertical headers
        self._apply_annual_row_headers()
        vertical_header = self.annual_table.verticalHeader()
        vertical_header.setDefaultAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
        vertical_header.setFixedWidth(155)
        horizontal_header = self.annual_table.horizontalHeader()
        horizontal_header.setDefaultAlignment(QtCore.Qt.AlignCenter)

        # Styling for the table
        self.annual_table.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;
                border: none;
                gridline-color: #e1e8ed;
                font-size: 12px;
                font-weight: 500;
                alternate-background-color: #f8f9fa;
            }
            QTableWidget::item {
                padding: 10px 8px;
                border-bottom: 1px solid #ecf0f1;
                border-right: 1px solid #ecf0f1;
            }
            QTableWidget::item:selected {
                background-color: transparent;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #34495e, stop:1 #2c3e50);
                color: white;
                font-weight: bold;
                font-size: 15px;
                padding: 8px 4px;
                border: none;
                border-right: 1px solid #3d566e;
            }
            QHeaderView::section:last {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #27ae60, stop:1 #229954);
                border-right: none;
            }
            QHeaderView::section:first {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3498db, stop:1 #2980b9);
            }
        """)
        
        # Fit the full year across common desktop widths without hiding the last months.
        self.annual_table.verticalHeader().setFixedWidth(155)
        for i in range(12):
            self.annual_table.setColumnWidth(i, 94)
        self.annual_table.setColumnWidth(12, 110)
        
        # Set row heights
        self.annual_table.verticalHeader().setDefaultSectionSize(45)
        
        # Make table read-only with keyboard-navigable selection
        self.annual_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.annual_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.annual_table.setFocusPolicy(QtCore.Qt.TabFocus)
        self.annual_table.setAccessibleName("Annual Financial Summary table")

        # Enable alternating row colors
        self.annual_table.setAlternatingRowColors(True)
        
        # Set font for better readability
        font = QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold)
        self.annual_table.setFont(font)
        
        # Set fixed height for the annual summary table
        self.annual_table.setFixedHeight(238)
        
        table_wrapper = QtWidgets.QHBoxLayout()
        table_wrapper.setContentsMargins(20, 0, 20, 0)

        table_wrapper.addWidget(self.annual_table)

        summary_layout.addSpacing(15)  # gap before table
        summary_layout.addLayout(table_wrapper)
        summary_layout.addSpacing(5)  # gap after table (optional)
        layout.addWidget(summary_group)

    def show_expenses_breakdown(self):
        """Show expenses breakdown dialog using ANNUAL data"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"📊 Expenses Breakdown - {self.current_year}")
        dialog.setModal(True)
        dialog.resize(920, 720)
        dialog.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QtWidgets.QLabel(f"📊 Expenses Breakdown - {self.current_year}")
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: white;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e74c3c, stop:1 #c0392b);
                border-radius: 8px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Create table for expenses breakdown
        table = QtWidgets.QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Expense Item", "Amount"])

        table.verticalHeader().setStyleSheet("""
        QHeaderView::section {
            background: transparent;
            color: #2c3e50;
        }
        """)
        table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1px solid #dde6e9;
                border-radius: 8px;
                gridline-color: #e1e8ed;
                font-size: 12px;
            }

            QHeaderView::section {
                background: transparent;
                color: #2c3e50;
                font-weight: bold;
                padding: 10px;
                border: none;
            }

            QTableCornerButton::section {
                background: transparent;
                border: none;
            }
            """)
        
        # Calculate expenses by name from ANNUAL data
        expense_totals = {}
        expense_display_names = {}
        total_expense_amount = 0

        for expense in self.annual_expenses_data:

            try:
                date_str = expense.get('date', '')
                if date_str:
                    date = datetime.strptime(date_str, "%m-%d-%Y")

                    # ✅ Only include selected annual summary year
                    if date.year != self.annual_summary_year:
                        continue

                raw_name = expense.get('name', 'Unknown')

                key = "".join(c for c in raw_name.lower() if c.isalnum())

                amount = float(expense.get('amount', '0').replace('$', '').replace(',', ''))

                expense_totals[key] = expense_totals.get(key, 0) + amount
                total_expense_amount += amount

                if key not in expense_display_names:
                    expense_display_names[key] = raw_name.strip().title()
            except:
                pass
        # Sort by amount descending
        sorted_expenses = sorted(expense_totals.items(), key=lambda x: x[1], reverse=True)

        table.setRowCount(len(sorted_expenses))

        for row, (key, amount) in enumerate(sorted_expenses):

            name = expense_display_names[key]
            # Expense Item
            item = QtWidgets.QTableWidgetItem(name)
            item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
            item.setForeground(QtGui.QColor("#2c3e50"))  # normal text color
            item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Normal))  # remove bold
            table.setItem(row, 0, item)
                        
            # Amount
            amount_item = QtWidgets.QTableWidgetItem(f"${amount:,.2f}")
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            amount_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            table.setItem(row, 1, amount_item)
            
            # Percentage
            table.setRowHeight(row, 40)
        
        # Adjust column widths
        table.setColumnWidth(0, 464)
        table.setColumnWidth(1, 250)
        
        layout.addWidget(table)
        
        # Total label
        total_label = QtWidgets.QLabel(f"Total Expenses: ${total_expense_amount:,.2f}")
        total_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px;
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 6px;
            }
        """)
        total_label.setAlignment(QtCore.Qt.AlignRight)
        layout.addWidget(total_label)
        
        # Close button
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setMinimumHeight(40)
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                font-weight: bold;
                border-radius: 6px;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()
    
    def show_salary_breakdown(self):
        """Show salary breakdown dialog using ANNUAL data"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"💰 Salary Breakdown - {self.current_year}")
        dialog.setModal(True)
        dialog.resize(800, 800)
        dialog.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QtWidgets.QLabel(f"💰 Salary Breakdown - {self.current_year}")
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: white;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2980b9);
                border-radius: 8px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Create tabs for Inside America and Outside America
        tab_widget = QtWidgets.QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background: white;
            }
            QTabBar::tab {
                background: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                font-weight: bold;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background: #d5dbdb;
            }
        """)
        
        # Inside America Tab
        inside_tab = QtWidgets.QWidget()
        inside_layout = QtWidgets.QVBoxLayout(inside_tab)
        
        inside_table = QtWidgets.QTableWidget()
        inside_table.setColumnCount(2)
        inside_table.setHorizontalHeaderLabels(["Employee Name", "Amount"])
        inside_table.verticalHeader().setFixedWidth(40)
        inside_table.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        inside_table.setAlternatingRowColors(True)
        inside_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        inside_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        inside_table.setStyleSheet("""
            QTableWidget {
                background: white;
                alternate-background-color: #f8fafc;
                border: 1px solid #dbe5ef;
                border-radius: 8px;
                gridline-color: #e6edf5;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 10px;
                color: #1e293b;
            }
            QHeaderView::section {
                background: #1f4e79;
                color: white;
                padding: 10px;
                font-weight: bold;
                border: none;
            }
        """)
        inside_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        inside_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
         
        # Populate Inside America salaries from ANNUAL data
        salary_totals = {}
        salary_display_names = {}

        for salary in self.annual_salary_data.get("Inside America", []):

            try:
                date_str = salary.get('date', '')
                if date_str:
                    date = datetime.strptime(date_str, "%m-%d-%Y")

                    # ✅ Filter by selected year
                    if date.year != self.annual_summary_year:
                        continue

                raw_name = salary.get('name', 'Unknown')

                key = "".join(c for c in raw_name.lower() if c.isalnum())

                amount = float(salary.get('amount', '0').replace('$', '').replace(',', ''))

                salary_totals[key] = salary_totals.get(key, 0) + amount

                if key not in salary_display_names:
                    salary_display_names[key] = raw_name.title()

            except:
                pass
        sorted_salaries = sorted(salary_totals.items(), key=lambda x: x[1], reverse=True)

        inside_table.setRowCount(len(sorted_salaries))

        total_inside = 0

        for row, (key, amount) in enumerate(sorted_salaries):

            name = salary_display_names[key]
            # Employee Name
            name_item = QtWidgets.QTableWidgetItem(name)
            name_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            inside_table.setItem(row, 0, name_item)
            
            # Amount
            total_inside += amount
            amount_item = QtWidgets.QTableWidgetItem(f"${amount:,.2f}")
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            amount_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            inside_table.setItem(row, 1, amount_item)
            
            inside_table.setRowHeight(row, 40)
        
        inside_table.setColumnWidth(1, 220)
        
        inside_layout.addWidget(inside_table)
        
        # Inside America total
        inside_total = QtWidgets.QLabel(f"Total Inside America: ${total_inside:,.2f}")
        inside_total.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px;
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }
        """)
        inside_total.setAlignment(QtCore.Qt.AlignRight)
        inside_layout.addWidget(inside_total)
        
        # Outside America Tab
        outside_tab = QtWidgets.QWidget()
        outside_layout = QtWidgets.QVBoxLayout(outside_tab)
        
        outside_table = QtWidgets.QTableWidget()
        outside_table.setColumnCount(2)
        outside_table.setHorizontalHeaderLabels(["Employee Name", "Amount"])
        outside_table.verticalHeader().setFixedWidth(40)
        outside_table.verticalHeader().setDefaultAlignment(QtCore.Qt.AlignCenter)
        outside_table.setAlternatingRowColors(True)
        outside_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        outside_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        outside_table.setStyleSheet(inside_table.styleSheet())
        outside_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        outside_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)

        # Populate Outside America salaries from ANNUAL data
        salary_totals = {}
        salary_display_names = {}

        for salary in self.annual_salary_data.get("Outside America", []):

            try:
                date_str = salary.get('date', '')
                if date_str:
                    date = datetime.strptime(date_str, "%m-%d-%Y")

                    # ✅ Filter by selected year
                    if date.year != self.annual_summary_year:
                        continue

                raw_name = salary.get('name', 'Unknown')

                key = "".join(c for c in raw_name.lower() if c.isalnum())

                amount = float(salary.get('amount', '0').replace('$', '').replace(',', ''))

                salary_totals[key] = salary_totals.get(key, 0) + amount

                if key not in salary_display_names:
                    salary_display_names[key] = raw_name.title()

            except:
                pass
            

        sorted_salaries = sorted(salary_totals.items(), key=lambda x: x[1], reverse=True)

        outside_table.setRowCount(len(sorted_salaries))

        total_outside = 0

        for row, (key, amount) in enumerate(sorted_salaries):

            name = salary_display_names[key]

            # Employee Name
            name_item = QtWidgets.QTableWidgetItem(name)
            name_item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft)
            outside_table.setItem(row, 0, name_item)

            # Amount
            total_outside += amount
            amount_item = QtWidgets.QTableWidgetItem(f"${amount:,.2f}")
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            amount_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            outside_table.setItem(row, 1, amount_item)

            outside_table.setRowHeight(row, 40)
        
        outside_table.setColumnWidth(1, 220)
        
        outside_layout.addWidget(outside_table)
        
        # Outside America total
        outside_total = QtWidgets.QLabel(f"Total Outside America: ${total_outside:,.2f}")
        outside_total.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                padding: 8px;
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }
        """)
        outside_total.setAlignment(QtCore.Qt.AlignRight)
        outside_layout.addWidget(outside_total)
        
        tab_widget.addTab(inside_tab, "Inside America")
        tab_widget.addTab(outside_tab, "🌍 Outside America")
        
        layout.addWidget(tab_widget)
        
        # Grand total
        grand_total = total_inside + total_outside
        grand_total_label = QtWidgets.QLabel(f"Total Salary: ${grand_total:,.2f}")
        grand_total_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px;
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 6px;
            }
        """)
        grand_total_label.setAlignment(QtCore.Qt.AlignRight)
        layout.addWidget(grand_total_label)
        
        # Close button
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setMinimumHeight(40)
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                font-weight: bold;
                border-radius: 6px;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    # Add this new method to BalanceSheetTab class to track invoice revenues

    def load_invoice_revenues(self):
        """Load invoice data and categorize as unpaid or paid revenue"""
        try:
            # Load all invoices from Firebase
            invoices_data = FirebaseManager.load_invoices()
            
            # Initialize revenue tracking
            self.unpaid_invoice_revenue = []
            self.paid_invoice_revenue = []
            
            for invoice_data in invoices_data:
                if not invoice_data or 'meta' not in invoice_data:
                    continue
                    
                meta = invoice_data['meta']
                invoice_number = meta.get('invoice_number', '')
                client_name = meta.get('client_name', '')
                invoice_date = meta.get('date', '')
                total_amount = meta.get('total', 0.0)
                status = meta.get('status', 'Pending')
                received_date = meta.get('received_date', '')
                
                # Create revenue entry from invoice
                revenue_entry = {
                    'source': f"Invoice - {invoice_number}",
                    'description': f"Invoice for {client_name}",
                    'amount': f"{total_amount:.2f}",
                    'date': invoice_date,
                    'invoice_number': invoice_number,
                    'status': status,
                    'received_date': received_date,
                    'is_invoice': True,
                    'firebase_id': invoice_data.get('firebase_id')
                }
                
                # Categorize based on status
                if status == "Paid":
                    self.paid_invoice_revenue.append(revenue_entry)
                else:
                    self.unpaid_invoice_revenue.append(revenue_entry)
                    
            _log.info("Loaded %s paid invoices and %s unpaid invoices", len(self.paid_invoice_revenue), len(self.unpaid_invoice_revenue))
            
        except Exception as e:
            _log.warning("Error loading invoice revenues: %s", e)
            self.paid_invoice_revenue = []
            self.unpaid_invoice_revenue = []

    def update_annual_summary(self):
        """Update the annual summary table using Firebase annual_revenue_data.

        Paid Revenue is summed from is_payment=True entries in annual_revenue_data —
        the same Firebase /revenue/ node the RevenueDetailDialog popup reads from,
        so the numbers always match the popup exactly.
        """
        monthly_paid     = [0] * 12
        monthly_expenses = [0] * 12

        # ── Paid Revenue: is_payment entries — only those whose invoice has an
        # is_invoice entry in the balance sheet (invoice was created and saved).
        invoiced_numbers = self._get_invoiced_numbers()
        for rev in self.annual_revenue_data:
            if not rev.get('is_payment'):
                continue
            inv_num = (rev.get('invoice_number') or '').strip()
            if not inv_num or inv_num not in invoiced_numbers:
                continue  # no matching invoice revenue entry in the balance sheet
            try:
                date_str = rev.get('date', rev.get('received_date', ''))
                dt = self._parse_finance_date(date_str)
                if dt and dt.year == self.annual_summary_year:
                    monthly_paid[dt.month - 1] += self._money_to_float(rev.get('amount', 0))
            except Exception as e:
                _log.warning("Error processing revenue entry for annual summary: %s", e)

        # ── Expenses: from Firebase annual data ───────────────────────────────
        for expense in self.annual_expenses_data:
            try:
                date = self._parse_finance_date(expense.get('date', ''))
                if date and date.year == self.annual_summary_year:
                    monthly_expenses[date.month - 1] += self._money_to_float(
                        expense.get('amount', 0)
                    )
            except Exception as e:
                _log.warning("Error processing expense for annual summary: %s", e)

        monthly_net    = [monthly_paid[i] - monthly_expenses[i] for i in range(12)]
        total_paid     = sum(monthly_paid)
        total_expenses = sum(monthly_expenses)
        total_net      = total_paid - total_expenses
        
        # Update table — 3 rows only (Unpaid Revenue is hidden per business rule)
        self.annual_table.setRowCount(3)
        self._apply_annual_row_headers()
        self.annual_table.clearContents()
        # Re-apply fixed row heights — setRowCount can reset them
        for _r in range(3):
            self.annual_table.setRowHeight(_r, 46)

        # Row 0 — Paid Revenue only
        import calendar as _cal
        for col in range(12):
            item = QtWidgets.QTableWidgetItem(f"${monthly_paid[col]:,.2f}")
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            if monthly_paid[col] > 0:
                item.setForeground(QtGui.QColor('#27ae60'))
                item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            else:
                item.setForeground(QtGui.QColor('#95a5a6'))
            item.setToolTip(
                f"Click to view paid revenue details for "
                f"{_cal.month_name[col + 1]} {self.annual_summary_year}")
            self.annual_table.setItem(0, col, item)
        total_paid_item = QtWidgets.QTableWidgetItem(f"${total_paid:,.2f}")
        total_paid_item.setTextAlignment(QtCore.Qt.AlignCenter)
        total_paid_item.setForeground(QtGui.QColor('#27ae60'))
        total_paid_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
        self.annual_table.setItem(0, 12, total_paid_item)
        # Make Revenue vertical header hint clickable
        self.annual_table.verticalHeader().setToolTip(
            "Click 'Revenue' to view all paid revenue for this year")

        # Row 1 — Expenses
        for col in range(12):
            item = QtWidgets.QTableWidgetItem(f"${monthly_expenses[col]:,.2f}")
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            if monthly_expenses[col] > 0:
                item.setForeground(QtGui.QColor('#e74c3c'))
                item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            else:
                item.setForeground(QtGui.QColor('#95a5a6'))
            self.annual_table.setItem(1, col, item)
        total_exp_item = QtWidgets.QTableWidgetItem(f"${total_expenses:,.2f}")
        total_exp_item.setTextAlignment(QtCore.Qt.AlignCenter)
        total_exp_item.setForeground(QtGui.QColor('#e74c3c'))
        total_exp_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
        self.annual_table.setItem(1, 12, total_exp_item)

        # Row 2 — Net Profit/Loss
        for col in range(12):
            item = QtWidgets.QTableWidgetItem(f"${monthly_net[col]:,.2f}")
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            if monthly_net[col] > 0:
                item.setForeground(QtGui.QColor('#27ae60'))
                item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            elif monthly_net[col] < 0:
                item.setForeground(QtGui.QColor('#e74c3c'))
                item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
            else:
                item.setForeground(QtGui.QColor('#95a5a6'))
            self.annual_table.setItem(2, col, item)

        total_net_item = QtWidgets.QTableWidgetItem(f"${total_net:,.2f}")
        total_net_item.setTextAlignment(QtCore.Qt.AlignCenter)
        if total_net > 0:
            total_net_item.setForeground(QtGui.QColor('#27ae60'))
            total_net_item.setBackground(QtGui.QColor('#e8f5e9'))
        elif total_net < 0:
            total_net_item.setForeground(QtGui.QColor('#e74c3c'))
            total_net_item.setBackground(QtGui.QColor('#fdedec'))
        else:
            total_net_item.setForeground(QtGui.QColor('#95a5a6'))
        total_net_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
        self.annual_table.setItem(2, 12, total_net_item)

        # Background highlights for first + total columns
        row_bg = {0: '#e8f5e9', 1: '#fdedec', 2: '#d4e6f1'}
        for row in range(3):
            for col in range(13):
                item = self.annual_table.item(row, col)
                if item and (col == 0 or col == 12):
                    item.setBackground(QtGui.QColor(row_bg[row]))

    @staticmethod
    def _dedup_is_payment_entries(all_revenue: list) -> list:
        """Deduplicate is_payment=True entries by payment_id, keeping the most recently
        updated one and deleting stale orphans from Firebase.  Non-payment entries are
        passed through unchanged.  Runs in a background thread — safe to call Firebase."""
        result = []          # non-payment entries pass through unchanged
        pid_map = {}         # payment_id -> best entry so far

        for rev in all_revenue:
            if not rev.get('is_payment'):
                result.append(rev)
                continue
            pid = rev.get('payment_id', '')
            if not pid:
                result.append(rev)   # no payment_id — include as-is (tax/legacy)
                continue
            existing = pid_map.get(pid)
            if existing is None:
                pid_map[pid] = rev
            else:
                rev_ts = rev.get('updated_at', rev.get('created_at', ''))
                ex_ts  = existing.get('updated_at', existing.get('created_at', ''))
                # Fallback to firebase_id comparison when timestamps are equal/missing
                # Firebase push IDs are lexicographically chronological: larger = newer
                if rev_ts == ex_ts:
                    rev_ts = rev.get('firebase_id', '')
                    ex_ts  = existing.get('firebase_id', '')
                if rev_ts > ex_ts:
                    # Current entry is newer — delete the old orphan
                    old_fid = existing.get('firebase_id', '')
                    if old_fid:
                        try:
                            BalanceSheetFirebaseManager.delete_entry('revenue', old_fid)
                        except Exception:
                            pass
                    pid_map[pid] = rev
                else:
                    # Existing entry is newer — delete the current orphan
                    fid = rev.get('firebase_id', '')
                    if fid:
                        try:
                            BalanceSheetFirebaseManager.delete_entry('revenue', fid)
                        except Exception:
                            pass

        result.extend(pid_map.values())
        return result

    @staticmethod
    def _build_invoice_status_map() -> dict:
        """Load all invoices from Firebase and return a map of
        {invoice_number: {status, received_date}} using the identical Overdue
        escalation logic as invoice_history_tab.get_invoice_status()."""
        from main import FirebaseManager as _FM
        from datetime import datetime as _dt, date as _date
        _DATE_FMTS = ("%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y")
        inv_map = {}
        for inv in (_FM.load_invoices() or []):
            meta = inv.get("meta") or {}
            inv_num = (meta.get("invoice_number") or "").strip()
            if not inv_num:
                continue
            raw_status = meta.get("status") or "Unpaid"
            # Overdue escalation: identical to invoice_history_tab.get_invoice_status()
            if raw_status in ("Unpaid", "Pending"):
                due_str = meta.get("due_date") or ""
                for _fmt in _DATE_FMTS:
                    try:
                        if _dt.strptime(due_str, _fmt).date() < _date.today():
                            raw_status = "Overdue"
                        break
                    except (ValueError, TypeError):
                        pass
            inv_map[inv_num] = {
                "status": raw_status,
                "received_date": meta.get("received_date") or "N/A",
            }
        return inv_map

    @staticmethod
    def _sync_revenue_statuses_from_invoices(all_revenue: list,
                                              inv_map: dict = None) -> list:
        """Patch status + received_date for every non-is_payment revenue entry so they
        exactly match the corresponding invoice — using the same source and the same
        Overdue escalation logic as invoice_history_tab.get_invoice_status().
        Pass a pre-built inv_map (from _build_invoice_status_map) to avoid an
        extra Firebase round-trip; omit it and one will be built here.
        Runs in a background thread.  Returns the same list (patched in-place)."""
        try:
            if inv_map is None:
                inv_map = BalanceSheetTab._build_invoice_status_map()
            # If Firebase load failed (empty map) skip patching entirely — better to
            # keep the stale value than to wipe every entry to "Unpaid".
            if not inv_map:
                return all_revenue
            for rev in all_revenue:
                if not isinstance(rev, dict) or rev.get("is_payment"):
                    continue
                inv_num = (rev.get("invoice_number") or "").strip()
                if not inv_num:
                    continue
                if inv_num in inv_map:
                    info = inv_map[inv_num]
                    rev["status"] = info["status"]
                    rev["received_date"] = info["received_date"]
                elif rev.get("is_invoice"):
                    # Invoice entry whose invoice was not found in Firebase (deleted or
                    # mismatch) — default to Unpaid so it never stays falsely "Paid".
                    rev["status"] = "Unpaid"
                    rev["received_date"] = "N/A"
        except Exception as _e:
            _log.warning("_sync_revenue_statuses_from_invoices: %s", _e)
        return all_revenue

    # ------------------------------------------------------------------ #
    # Real-time Firebase listener for /revenue/                            #
    # ------------------------------------------------------------------ #

    def _start_revenue_listener(self):
        """Attach a Firebase SSE listener to /revenue/ so any write from ANY device
        (this machine or another) is pushed to the GUI instantly — no polling."""
        if not self.FIREBASE_AVAILABLE:
            return
        if getattr(self, '_revenue_listener_handle', None) is not None:
            return  # already running
        try:
            from firebase_admin import db as _fdb
            _cache: dict = {}  # firebase_id → entry dict (maintained incrementally)

            def _on_event(event):
                if event.event_type == 'cancel':
                    _log.warning("Revenue listener cancelled: %s", event.data)
                    return
                path = event.path or '/'
                data = event.data

                if event.event_type == 'put':
                    if path == '/':
                        # Initial full load or complete replacement
                        _cache.clear()
                        if isinstance(data, dict):
                            for fid, entry in data.items():
                                if isinstance(entry, dict):
                                    _cache[fid] = {**entry, 'firebase_id': fid}
                    else:
                        # Single entry or nested field
                        parts = path.strip('/').split('/', 1)
                        key = parts[0]
                        if len(parts) == 1:
                            if data is None:
                                _cache.pop(key, None)
                            elif isinstance(data, dict):
                                _cache[key] = {**data, 'firebase_id': key}
                        else:
                            # Nested field update (e.g. /KEY/amount)
                            if key in _cache and data is not None:
                                _cache[key][parts[1].split('/')[0]] = data
                elif event.event_type == 'patch':
                    for key, val in (data or {}).items():
                        if val is None:
                            _cache.pop(key, None)
                        elif isinstance(val, dict):
                            if key in _cache:
                                _cache[key].update(val)
                                _cache[key]['firebase_id'] = key
                            else:
                                _cache[key] = {**val, 'firebase_id': key}

                # Push snapshot to GUI thread
                global _revenue_signaler
                if _revenue_signaler is not None:
                    _revenue_signaler.revenue_updated.emit(list(_cache.values()))

            self._revenue_listener_handle = _fdb.reference('revenue').listen(_on_event)
            _log.info("Real-time /revenue/ listener started")
        except Exception as e:
            _log.warning("Could not start revenue listener: %s", e)
            self._revenue_listener_handle = None

    def _stop_revenue_listener(self):
        """Stop the Firebase /revenue/ listener (called on tab close / app exit)."""
        timer = getattr(self, '_listener_debounce_timer', None)
        if timer is not None:
            timer.stop()
        handle = getattr(self, '_revenue_listener_handle', None)
        if handle is not None:
            try:
                handle.close()
            except Exception:
                pass
            self._revenue_listener_handle = None

    def update_revenue_entry_status(self, invoice_number: str, new_status: str, received_date: str):
        """Immediately sync status + received_date for an invoice revenue entry.

        Three-layer update (all on the GUI thread, no Firebase read):
          1. Patch revenue_data and annual_revenue_data in-memory.
          2. Patch the visible finance_table cells directly (like update_invoice_row_immediately).
          3. If no row is currently visible (different page/category) call populate_revenue_data
             so the next time Revenue tab is shown it reflects the correct values.
        """
        rd = received_date if new_status not in ('Unpaid', 'Overdue') else 'N/A'
        inv = (invoice_number or "").strip()

        # Keep the local override map in sync so the listener's next flush
        # uses the signal-provided status rather than stale Firebase data.
        if not hasattr(self, '_invoice_status_map'):
            self._invoice_status_map = {}
        if inv:
            self._invoice_status_map[inv] = {'status': new_status, 'received_date': rd}

        def _is_invoice_row(rev):
            if not isinstance(rev, dict):
                return False
            if rev.get('is_payment'):
                return False
            return (rev.get('invoice_number') or '').strip() == inv

        # ── 1. In-memory update ───────────────────────────────────────────────
        matched_rev = None
        for rev in self.revenue_data:
            if _is_invoice_row(rev):
                rev['status'] = new_status
                rev['received_date'] = rd
                rev['down_payment_received_date'] = rd
                matched_rev = rev
                break
        for rev in self.annual_revenue_data:
            if _is_invoice_row(rev):
                rev['status'] = new_status
                rev['received_date'] = rd
                rev['down_payment_received_date'] = rd
                break

        if matched_rev is None:
            return  # entry not loaded yet — Firebase listener will handle it

        # ── 2. Direct table-cell patch (instant, no full repopulate) ─────────
        if self.current_category == "Revenue" and hasattr(self, '_fin_all_items'):
            per_page = self._fin_per_page
            start_i = (self._fin_page - 1) * per_page
            page_slice = self._fin_all_items[start_i:start_i + per_page]
            patched = False
            for offset, page_rev in enumerate(page_slice):
                if page_rev is not matched_rev:
                    continue
                table_row = offset
                # Status cell (col 6) — pill badge
                self.finance_table.setCellWidget(
                    table_row, 6, self._make_revenue_status_pill(new_status))
                # Received Date cell (col 7)
                rd_item = QtWidgets.QTableWidgetItem(rd)
                rd_item.setTextAlignment(QtCore.Qt.AlignCenter)
                if rd != 'N/A':
                    if new_status == "Paid":
                        rd_item.setForeground(QtGui.QColor('#27ae60'))
                        rd_item.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold))
                    elif new_status == "Partially Paid":
                        rd_item.setForeground(QtGui.QColor('#8e44ad'))
                        rd_item.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold))
                self.finance_table.setItem(table_row, 7, rd_item)
                patched = True
                break

            # ── 3. If entry is off-page or filtered out, full repopulate ─────
            if not patched:
                self.populate_revenue_data()
        elif self.current_category == "Revenue":
            self.populate_revenue_data()

    def _on_realtime_revenue_update(self, all_revenue: list):
        """GUI-thread slot: called whenever Firebase pushes a /revenue/ change.
        Stores the latest snapshot and resets the debounce timer so that rapid
        field-level writes (5+ events for a single edit) collapse into one redraw."""
        self._listener_pending_revenue = all_revenue
        self._listener_debounce_timer.start()  # restarts the 150 ms countdown

    def _flush_realtime_update(self):
        """Debounce timer callback — runs on the GUI thread 150 ms after the last
        listener event. Syncs invoice statuses from /invoices/ in a background
        thread so the display always matches invoice history exactly."""
        all_revenue = self._listener_pending_revenue
        if not all_revenue:
            return
        try:
            year = self.annual_summary_year
            cur_year = self.current_year
            import threading as _threading

            # Capture the signal-based override map (built on main thread — safe)
            _override_map = dict(getattr(self, '_invoice_status_map', {}))

            def _bg(rev=list(all_revenue), omap=_override_map):
                # Build full invoice map from Firebase, then overlay signal overrides
                inv_map = BalanceSheetTab._build_invoice_status_map()
                inv_map.update(omap)  # signal updates always win over Firebase reads
                BalanceSheetTab._sync_revenue_statuses_from_invoices(rev, inv_map)
                new_annual = [
                    r for r in rev
                    if BalanceSheetFirebaseManager._entry_year(r) == year
                ]
                new_rev = [
                    r for r in rev
                    if BalanceSheetFirebaseManager._entry_year(r) == cur_year
                ]
                QtCore.QTimer.singleShot(
                    0, lambda a=new_annual, r=new_rev: self._apply_all_revenue(a, r)
                )

            _threading.Thread(target=_bg, daemon=True).start()
        except Exception as e:
            _log.warning("Real-time revenue flush failed: %s", e)

    def _refresh_annual_revenue_background(self):
        """Fetch fresh /revenue/ data in a background thread, deduplicate is_payment
        entries, sync statuses from /invoices/, then redraw the annual summary."""
        import threading as _threading
        # Capture signal overrides on the main thread before spawning the worker,
        # so a slow Firebase read never overwrites a status already set locally.
        _override_map = dict(getattr(self, '_invoice_status_map', {}))

        def _bg(omap=_override_map):
            try:
                all_revenue = BalanceSheetFirebaseManager.load_revenue()
                all_revenue = BalanceSheetTab._dedup_is_payment_entries(all_revenue)
                inv_map = BalanceSheetTab._build_invoice_status_map()
                inv_map.update(omap)          # signal overrides always win
                all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(
                    all_revenue, inv_map)
                new_data = [
                    rev for rev in all_revenue
                    if BalanceSheetFirebaseManager._entry_year(rev) == self.annual_summary_year
                ]
                QtCore.QTimer.singleShot(0, lambda d=new_data: self._apply_annual_revenue(d))
            except Exception as e:
                _log.warning("Annual revenue background refresh failed: %s", e)
        _threading.Thread(target=_bg, daemon=True).start()

    def _apply_annual_revenue(self, new_data):
        """Apply freshly fetched annual revenue data and redraw the summary table (main thread only).
        is_payment entries are always rebuilt from in-memory payments so a stale Firebase read
        (arriving before the write for an edit completes) never overwrites the correct values."""
        self.annual_revenue_data = new_data
        self.update_annual_summary_from_payments()  # rebuilds is_payment slice; calls update_annual_summary

    def _get_invoiced_numbers(self) -> set:
        """Return the set of invoice_numbers that have an is_invoice=True entry in the
        balance sheet (/revenue/).  Combines the full-year persistent set (updated on every
        Firebase refresh) with whatever is already in the currently-loaded annual/revenue
        data so the filter works correctly even before the first full refresh runs."""
        result = set(getattr(self, '_all_invoiced_numbers', ()))
        for r in list(self.annual_revenue_data or []) + list(self.revenue_data or []):
            if r.get('is_invoice') and r.get('invoice_number', '').strip():
                result.add(r['invoice_number'].strip())
        return result

    def update_annual_summary_from_payments(self):
        """Instantly rebuild the annual summary from in-memory payment data — no Firebase
        read needed. Called on the main thread right after any payment add/edit/delete so
        the REVENUE row reflects the change before the background Firebase refresh arrives.

        Strategy:
        - Non-payment (is_invoice, manual) entries are always preserved from annual_revenue_data.
        - is_payment entries whose payment_id IS in tracker.payments are replaced by the
          tracker version (so edits to date/amount show immediately).
        - is_payment entries whose payment_id is NOT in tracker.payments are kept from
          annual_revenue_data (legacy/external Firebase payments the local tracker doesn't know
          about — discarding them would make older months go blank)."""
        try:
            from payment_tracker import get_payment_tracker
            tracker = get_payment_tracker()
            year = self.annual_summary_year

            # Build set of payment_ids managed by the local tracker
            tracker_pids = {p.payment_id for p in tracker.payments if p.payment_id}

            # Keep non-payment entries (is_invoice, manual revenue, etc.) unchanged
            non_payment = [
                r for r in (self.annual_revenue_data or [])
                if not r.get('is_payment')
            ]

            # Invoice numbers the tracker already has payments for — used to
            # exclude legacy Firebase entries for those invoices so they don't
            # double-count alongside fresh tracker entries.
            tracker_inv_nos = {
                (p.invoice_number or '').strip()
                for p in tracker.payments
                if (p.invoice_number or '').strip()
            }

            # Keep is_payment entries that the local tracker does NOT know about
            # (legacy Firebase entries, entries from other devices, etc.).
            # Exclude entries whose invoice_number is already covered by the tracker —
            # those are superseded by tracker_entries and would cause double-counting.
            firebase_only = [
                r for r in (self.annual_revenue_data or [])
                if r.get('is_payment')
                and r.get('payment_id', '') not in tracker_pids
                and (r.get('invoice_number') or '').strip() not in tracker_inv_nos
            ]

            # Only include payments whose invoice has an is_invoice entry in the balance sheet.
            # This prevents orphaned payments (no matching invoice revenue) from inflating the
            # annual summary totals.
            invoiced_numbers = self._get_invoiced_numbers()

            # Rebuild the tracker-managed is_payment slice from live Payment objects
            tracker_entries = []
            for p in tracker.payments:
                if not p.invoice_number:
                    continue  # bare project payments don't appear in balance sheet
                if p.invoice_number.strip() not in invoiced_numbers:
                    continue  # invoice has no revenue entry in the balance sheet yet
                try:
                    dt = self._parse_finance_date(p.payment_date)
                    if not dt or dt.year != year:
                        continue
                except Exception:
                    continue
                tracker_entries.append({
                    'is_payment':     True,
                    'payment_id':     p.payment_id,
                    'amount':         float(p.amount),
                    'date':           p.payment_date,
                    'received_date':  p.payment_date,
                    'year':           year,
                    'source':         f'Project {p.project_number}',
                    'invoice_number': p.invoice_number,
                    'project_number': p.project_number,
                    'payment_stage':  p.payment_stage or '',
                    'firebase_id':    p.balance_sheet_id or '',
                    'status':         'Paid',
                })

            self.annual_revenue_data = non_payment + firebase_only + tracker_entries
            self.update_annual_summary()
            self.update_stats_cards()
        except Exception as e:
            _log.warning("Immediate annual summary update failed: %s", e)

    def _refresh_all_revenue_background(self):
        """Fetch fresh /revenue/ data in background, deduplicate is_payment entries,
        sync status/received_date from /invoices/ so balance sheet mirrors invoice history,
        then update BOTH the annual summary AND the paid revenues transaction table."""
        import threading as _threading
        _override_map = dict(getattr(self, '_invoice_status_map', {}))

        def _bg(omap=_override_map):
            try:
                all_revenue = BalanceSheetFirebaseManager.load_revenue()
                all_revenue = BalanceSheetTab._dedup_is_payment_entries(all_revenue)
                inv_map = BalanceSheetTab._build_invoice_status_map()
                inv_map.update(omap)
                all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue, inv_map)
                new_annual = [
                    rev for rev in all_revenue
                    if BalanceSheetFirebaseManager._entry_year(rev) == self.annual_summary_year
                ]
                new_revenue = [
                    rev for rev in all_revenue
                    if BalanceSheetFirebaseManager._entry_year(rev) == self.current_year
                ]
                # Build the set of invoice_numbers that have an is_invoice entry across
                # ALL years so the annual-summary filter works even for cross-year invoices.
                invoiced = {
                    r.get('invoice_number', '').strip()
                    for r in all_revenue
                    if r.get('is_invoice') and r.get('invoice_number', '').strip()
                }
                QtCore.QTimer.singleShot(
                    0, lambda a=new_annual, r=new_revenue, iv=invoiced:
                        self._apply_all_revenue(a, r, iv)
                )
            except Exception as e:
                _log.warning("All-revenue background refresh failed: %s", e)
        _threading.Thread(target=_bg, daemon=True).start()

    def _apply_all_revenue(self, new_annual, new_revenue, invoiced_numbers: set = None):
        """Apply fresh revenue data to both tables (main thread only).
        is_payment entries are always rebuilt from in-memory payments so a stale Firebase read
        (arriving before the write for an edit completes) never overwrites the correct values."""
        self.annual_revenue_data = new_annual
        self.revenue_data = new_revenue
        if invoiced_numbers is not None:
            self._all_invoiced_numbers = invoiced_numbers
        self.update_annual_summary_from_payments()  # rebuilds is_payment slice; calls update_annual_summary + update_stats_cards
        if self.current_category == "Revenue":
            self.populate_revenue_data()

    def create_finance_table_section(self, layout):
        """Create finance table section with integrated search and filters and full page scrolling"""
        table_group = QtWidgets.QGroupBox("Transactions")
        table_group.setStyleSheet("""
            QGroupBox {
                font-weight:900; font-size:13px; color:#0f766e;
                border: none;
                margin-top:0.5em; padding-top:10px; background: transparent;
            }
            QGroupBox::title {
                subcontrol-origin:margin; left:16px;
                padding:0 10px; background:#f0fdf4;
                border-radius:8px; font-size:12px;
            }
        """)
        table_layout = QtWidgets.QVBoxLayout(table_group)
        table_layout.setContentsMargins(5, 10, 5, 10)  # Left, Top, Right, Bottom
        
        # 🔍 Integrated Search + Filters Section
        search_filter_frame = QtWidgets.QFrame()
        search_filter_frame.setStyleSheet("QFrame { background: transparent; border: none; }")

        search_filter_layout = QtWidgets.QHBoxLayout(search_filter_frame)
        search_filter_layout.setSpacing(10)
        search_filter_layout.setContentsMargins(10, 0, 10, 0)

        # Date Range Filter Button
        self.date_range_button = configure_filter_button(QtWidgets.QPushButton(), height=36)
        self.date_range_button.clicked.connect(self.show_date_range_dialog)
        search_filter_layout.addWidget(self.date_range_button)

        # Search bar
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search by description, source, name...")
        self.search_edit.setMinimumHeight(36)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 10px;
                font-size: 13px;
                background: white;
                min-width: 300px;
            }
            QLineEdit:focus { border-color: #3498db; }
        """)
        self.search_edit.textChanged.connect(self.filter_finance_entries)
        search_filter_layout.addWidget(self.search_edit)

        # Flexible space
        search_filter_layout.addStretch(1)

        # Category filter dropdown (right side)
        filters_container = QtWidgets.QHBoxLayout()
        filters_container.setSpacing(8)

        category_label = QtWidgets.QLabel("Show:")
        category_label.setStyleSheet("font-weight: 600; color: #2c3e50; font-size: 13px;")
        filters_container.addWidget(category_label)

        self.category_combo = QtWidgets.QComboBox()
        self.category_combo.addItems(["Revenue", "Expenses", "Salary"])
        self.category_combo.setMinimumHeight(36)
        self.category_combo.setMinimumWidth(150)
        self.category_combo.setStyleSheet("""
            QComboBox {
                padding: 6px 32px 6px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-weight: 600;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                color: #1e293b;
            }
            QComboBox:hover  { border-color: #94a3b8; }
            QComboBox:focus  { border-color: #6366f1; }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 28px;
                border-left: 1px solid #e1e8ed;
                border-top-right-radius: 8px;
                border-bottom-right-radius: 8px;
                background: #f8fafc;
            }
            QComboBox::down-arrow {
                image: none;
                width: 0;
                height: 0;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #64748b;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #e1e8ed;
                border-radius: 6px;
                background: white;
                selection-background-color: #eff6ff;
                selection-color: #1d4ed8;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 13px;
                padding: 4px;
            }
        """)
        self.category_combo.currentTextChanged.connect(self.on_category_changed)
        self.category_combo.wheelEvent = lambda e: e.ignore()
        self.category_combo.keyPressEvent = lambda e, c=self.category_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.category_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.category_combo.clearFocus))
        filters_container.addWidget(self.category_combo)

        search_filter_layout.addLayout(filters_container)
        table_layout.addWidget(search_filter_frame)

        # Finance Table — no internal scroll; page-level scroll handles it
        self.finance_table = QtWidgets.QTableWidget()
        self.finance_table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.finance_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.finance_table.setWordWrap(True)
        self.finance_table.setAlternatingRowColors(True)
        
        # Set initial columns for Revenue (default)
        self.setup_revenue_table()
        
        # Professional table styling
        self.finance_table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                gridline-color: #e1e8ed;
                font-size: 13px;
                outline: none;
                selection-background-color: #e3f2fd;
                alternate-background-color: #f8fafc;
            }
            QTableWidget::item {
                padding: 10px 8px;
                border-bottom: 1px solid #f0f0f0;
                border-right: 1px solid #e1e8ed;
                color: #2c3e50;
            }
            QTableWidget::item:last {
                border-right: none;
            }
            QTableWidget::item:selected {
                background: #e3f2fd;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2c3e50, stop:1 #34495e);
                color: white;
                font-weight: bold;
                font-size: 12px;
                padding: 10px 6px;
                border: none;
                border-right: 1px solid #3a506b;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
            }
        """)
        
        self.finance_table.setShowGrid(True)
        self.finance_table.setGridStyle(QtCore.Qt.SolidLine)
        self.finance_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.finance_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.finance_table.setAlternatingRowColors(True)
        self.finance_table.verticalHeader().setVisible(False)
        self.finance_table.verticalHeader().setDefaultSectionSize(54)
        
        # Header properties
        header = self.finance_table.horizontalHeader()
        header.setDefaultAlignment(QtCore.Qt.AlignCenter)
        header.setHighlightSections(False)
        header.setFixedHeight(40)
        
        # Make columns sizeable
        for col in range(self.finance_table.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Interactive)

        table_layout.addWidget(self.finance_table)

        _pg_s = """QPushButton { background:#ffffff; color:#334155; border:1px solid #e2e8f0;
            border-radius:6px; font-size:12px; font-weight:700;
            min-width:32px; min-height:28px; padding:0 8px; }
            QPushButton:hover { background:#f1f5f9; border-color:#cbd5e1; }
            QPushButton:disabled { color:#cbd5e1; }"""
        fin_pg_frame = QtWidgets.QFrame()
        fin_pg_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        fin_pg_hbox = QtWidgets.QHBoxLayout(fin_pg_frame)
        fin_pg_hbox.setContentsMargins(4, 6, 4, 6)
        fin_pg_hbox.setSpacing(6)
        self._fin_info_lbl = QtWidgets.QLabel("")
        self._fin_info_lbl.setStyleSheet(
            "color:#94a3b8; font-size:11px; font-weight:600; background:transparent; border:none;")
        fin_pg_hbox.addWidget(self._fin_info_lbl)
        fin_pg_hbox.addStretch()
        self._fin_prev_btn = QtWidgets.QPushButton("‹")
        self._fin_prev_btn.setStyleSheet(_pg_s)
        self._fin_prev_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._fin_prev_btn.clicked.connect(self._fin_go_prev)
        fin_pg_hbox.addWidget(self._fin_prev_btn)
        self._fin_page_btns_layout = QtWidgets.QHBoxLayout()
        self._fin_page_btns_layout.setSpacing(4)
        fin_pg_hbox.addLayout(self._fin_page_btns_layout)
        self._fin_next_btn = QtWidgets.QPushButton("›")
        self._fin_next_btn.setStyleSheet(_pg_s)
        self._fin_next_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._fin_next_btn.clicked.connect(self._fin_go_next)
        fin_pg_hbox.addWidget(self._fin_next_btn)
        self._fin_pg_style = _pg_s
        table_layout.addWidget(fin_pg_frame)

        layout.addWidget(table_group)

    def on_category_changed(self, category):
        """Handle category change - update table headers and content"""
        self.current_category = category
        
        # Clear table
        self.finance_table.clear()
        self.finance_table.setRowCount(0)
        
        # Setup appropriate table based on selected category
        if category == "Revenue":
            self.setup_revenue_table()
            self.populate_revenue_data()
        elif category == "Expenses":
            self.setup_expenses_table()
            self.populate_expenses_data()
        else:  # Salary
            self.setup_salary_table()
            self.populate_salary_data()
        
        # Update add button — hide for Revenue (auto-created from invoices)
        if category == "Revenue":
            self.add_btn.setVisible(False)
        else:
            self.add_btn.setVisible(True)
            self.add_btn.setText(f"+ Add {category.rstrip('s')}")
        self.update_stats_cards()
        # Annual summary is NOT updated here - it stays constant

    def setup_revenue_table(self):
        """Setup table for Revenue entries with S.No, Date, Status, Due Date, Received Date, and Actions"""
        self.finance_table.setColumnCount(9)  # Added Received Date column
        self.finance_table.setHorizontalHeaderLabels([
            "S.No", "Date", "Revenue Source", "Description", "Amount ($)", "Due Date", "Status", "Payment Date", "Actions"
        ])
        
        # Set column widths
        self.finance_table.setColumnWidth(0, 54)    # S.No
        self.finance_table.setColumnWidth(1, 112)   # Date
        self.finance_table.setColumnWidth(2, 260)   # Revenue Source
        self.finance_table.setColumnWidth(3, 380)   # Description (wider — stretches)
        self.finance_table.setColumnWidth(4, 130)   # Amount
        self.finance_table.setColumnWidth(5, 112)   # Due Date
        self.finance_table.setColumnWidth(6, 140)   # Status
        self.finance_table.setColumnWidth(7, 120)   # Received Date
        self.finance_table.setColumnWidth(8, 100)   # Actions (narrower — 3 × 28px buttons)
        hdr = self.finance_table.horizontalHeader()
        hdr.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        hdr.setSectionResizeMode(8, QtWidgets.QHeaderView.Fixed)


    def setup_expenses_table(self):
        """Setup table for Expenses entries with S.No column"""
        self.finance_table.setColumnCount(6)
        self.finance_table.setHorizontalHeaderLabels([
            "S.No", "Date", "Expense Item", "Description", "Amount ($)", "Actions"
        ])
        self.finance_table.setColumnWidth(0, 54)    # S.No
        self.finance_table.setColumnWidth(1, 120)   # Date
        self.finance_table.setColumnWidth(2, 320)   # Expense Item
        self.finance_table.setColumnWidth(4, 150)   # Amount
        self.finance_table.setColumnWidth(5, 90)    # Actions (narrower — 2 × 28px buttons)
        hdr = self.finance_table.horizontalHeader()
        hdr.setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        hdr.setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)

    def sort_expenses_by_creation_date(self):
        """Sort expenses by creation date (newest first)"""
        try:
            if hasattr(self, 'annual_expenses_data') and self.annual_expenses_data:
                self.annual_expenses_data.sort(
                    key=lambda x: x.get('created_at', ''),
                    reverse=True
                )
        except Exception as e:
            _log.warning("Error sorting expenses: %s", e)
            
    def setup_salary_table(self):
        """Setup table for Salary entries with region column and S.No"""
        self.finance_table.setColumnCount(7)  # Added S.No column
        self.finance_table.setHorizontalHeaderLabels([
            "S.No", "Date", "Region", "Employee Name", "Description", "Amount ($)", "Actions"
        ])
        
        # Set column widths
        self.finance_table.setColumnWidth(0, 54)    # S.No
        self.finance_table.setColumnWidth(1, 120)   # Date
        self.finance_table.setColumnWidth(2, 170)   # Region
        self.finance_table.setColumnWidth(3, 260)   # Employee Name
        self.finance_table.setColumnWidth(5, 150)   # Amount
        self.finance_table.setColumnWidth(6, 90)    # Actions (narrower — 2 × 28px buttons)
        hdr = self.finance_table.horizontalHeader()
        hdr.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        hdr.setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)

    def populate_revenue_data(self):
        """Populate revenue table in chronological order (oldest first)."""
        filtered = self.filter_revenue_data()
        self._fin_all_items = filtered
        self._fin_page = 1
        self._fin_render_page()

    # Same palette as invoice_history_tab
    _REV_PILL_COLORS = {
        "Paid":           ("#d1fae5", "#065f46", "#6ee7b7"),
        "Unpaid":         ("#f8d7da", "#721c24", "#f5c6cb"),
        "Pending":        ("#fff3cd", "#856404", "#ffeaa7"),
        "Overdue":        ("#ffe5d9", "#a13700", "#ffb599"),
        "Partially Paid": ("#ede7f6", "#4a148c", "#d1c4e9"),
    }

    _STATUS_ICONS = {
        "Paid":           "✓",
        "Unpaid":         "✗",
        "Pending":        "⏳",
        "Overdue":        "⚠",
        "Partially Paid": "◐",
    }

    def _make_revenue_status_pill(self, status: str) -> QtWidgets.QWidget:
        """Centered static pill badge for the revenue status column."""
        bg, fg, border = self._REV_PILL_COLORS.get(
            status, ("#e2e8f0", "#64748b", "#cbd5e1"))
        icon = self._STATUS_ICONS.get(status, "")
        text = f"{icon}  {status}" if icon else status
        lbl = QtWidgets.QLabel(text)
        lbl.setSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        lbl.setAlignment(QtCore.Qt.AlignCenter)
        lbl.setStyleSheet(
            f"QLabel {{ background-color:{bg}; color:{fg}; "
            f"border:1px solid {border}; border-radius:12px; "
            f"padding:4px 12px; font-size:12px; font-weight:bold; "
            f"font-family:'Inter','Segoe UI',sans-serif; }}"
        )
        container = QtWidgets.QWidget()
        container.setStyleSheet("background:transparent; border:none;")
        lay = QtWidgets.QHBoxLayout(container)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.addStretch()
        lay.addWidget(lbl)
        lay.addStretch()
        return container

    def _render_revenue_rows(self, page_items, start_i):
        def format_amount(amount):
            if isinstance(amount, (int, float)):
                return f"${amount:,.2f}"
            elif isinstance(amount, str):
                clean = amount.replace('$', '').replace(',', '')
                try:
                    return f"${float(clean):,.2f}"
                except:
                    return "$0.00"
            return "$0.00"

        for row, revenue in enumerate(page_items):
            # S.No
            sno_item = QtWidgets.QTableWidgetItem(str(start_i + row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 0, sno_item)
            
            # Date (Invoice Date) - Always show the invoice date
            invoice_date = revenue.get('date', '')
            date_item = QtWidgets.QTableWidgetItem(invoice_date)
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            date_item.setToolTip(invoice_date)
            self.finance_table.setItem(row, 1, date_item)
            
            # Revenue Source
            source_text = revenue.get('source', '')
            source_item = QtWidgets.QTableWidgetItem(source_text)
            source_item.setTextAlignment(QtCore.Qt.AlignCenter)
            source_item.setToolTip(source_text)
            self.finance_table.setItem(row, 2, source_item)

            # Description
            desc_text = revenue.get('description', '')
            desc_item = QtWidgets.QTableWidgetItem(desc_text)
            desc_item.setTextAlignment(QtCore.Qt.AlignCenter)
            desc_item.setToolTip(desc_text)
            self.finance_table.setItem(row, 3, desc_item)
            
            # Amount
            amount = revenue.get('amount', '0')
            amount_item = QtWidgets.QTableWidgetItem(format_amount(amount))
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            self.finance_table.setItem(row, 4, amount_item)
            
            # Due Date - Show the due date for all statuses
            due_date = revenue.get('due_date', 'N/A')
            due_date_item = QtWidgets.QTableWidgetItem(due_date)
            due_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            
            # Highlight due date if overdue and unpaid
            status = revenue.get('status', 'Unpaid')
            if status == "Unpaid" and due_date != 'N/A':
                try:
                    due_date_obj = datetime.strptime(due_date, "%m-%d-%Y")
                    if due_date_obj < datetime.now():
                        due_date_item.setForeground(QtGui.QColor('#e74c3c'))
                        due_date_item.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                except:
                    pass
            self.finance_table.setItem(row, 5, due_date_item)
            
            # Status — colored pill badge
            self.finance_table.setCellWidget(row, 6, self._make_revenue_status_pill(status))
            
            # Received Date — always use received_date field (latest payment date for
            # Paid/Partially Paid; N/A for Unpaid/Pending).  The old fallback to
            # down_payment_received_date for Partially Paid is removed because
            # received_date is now always kept in sync with the latest payment.
            received_date = revenue.get('received_date', 'N/A') or 'N/A'
            received_date_item = QtWidgets.QTableWidgetItem(received_date)
            received_date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            
            # Highlight received date for Paid (green) and Partially Paid (purple)
            if received_date != 'N/A':
                if status == "Paid":
                    received_date_item.setForeground(QtGui.QColor('#27ae60'))
                    received_date_item.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold))
                elif status == "Partially Paid":
                    received_date_item.setForeground(QtGui.QColor('#8e44ad'))
                    received_date_item.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold))
            
            self.finance_table.setItem(row, 7, received_date_item)
            
            # Actions
            self.add_revenue_action_buttons(row, revenue, col_offset=8)

            self.finance_table.setRowHeight(row, 54)
        
    def populate_expenses_data(self):
        """Populate expenses table in chronological order (oldest first)."""
        filtered = self.filter_expenses_data()
        self._fin_all_items = filtered
        self._fin_page = 1
        self._fin_render_page()

    def _render_expenses_rows(self, page_items, start_i):
        for row, expense in enumerate(page_items):
            sno_item = QtWidgets.QTableWidgetItem(str(start_i + row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 0, sno_item)
            date_item = QtWidgets.QTableWidgetItem(expense.get('date', ''))
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 1, date_item)
            name_item = QtWidgets.QTableWidgetItem(expense.get('name', ''))
            name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 2, name_item)
            desc_item = QtWidgets.QTableWidgetItem(expense.get('description', ''))
            desc_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 3, desc_item)
            amount = expense.get('amount', '0')
            try:
                formatted_amount = f"${float(amount):,.2f}"
            except Exception:
                formatted_amount = f"${amount}"
            amount_item = QtWidgets.QTableWidgetItem(formatted_amount)
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            self.finance_table.setItem(row, 4, amount_item)
            self.add_action_buttons(row, expense, "Expenses", col_offset=5)
            self.finance_table.setRowHeight(row, 54)

    def populate_salary_data(self):
        """Populate salary table in chronological order (oldest first)."""
        filtered = self.filter_salary_data()
        self._fin_all_items = filtered
        self._fin_page = 1
        self._fin_render_page()

    def _render_salary_rows(self, page_items, start_i):
        for row, salary in enumerate(page_items):
            sno_item = QtWidgets.QTableWidgetItem(str(start_i + row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 0, sno_item)
            date_item = QtWidgets.QTableWidgetItem(salary.get('date', ''))
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 1, date_item)
            region_item = QtWidgets.QTableWidgetItem(salary.get('region', ''))
            region_item.setTextAlignment(QtCore.Qt.AlignCenter)
            if salary.get('region') == "Inside America":
                region_item.setForeground(QtGui.QColor('#2980b9'))
            else:
                region_item.setForeground(QtGui.QColor('#c0392b'))
            region_item.setFont(QtGui.QFont("Arial", 8, QtGui.QFont.Normal))
            self.finance_table.setItem(row, 2, region_item)
            name_item = QtWidgets.QTableWidgetItem(salary.get('name', ''))
            name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 3, name_item)
            desc_item = QtWidgets.QTableWidgetItem(salary.get('description', ''))
            desc_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.finance_table.setItem(row, 4, desc_item)
            amount = salary.get('amount', '0')
            try:
                formatted_amount = f"${float(amount):,.2f}"
            except Exception:
                formatted_amount = f"${amount}"
            amount_item = QtWidgets.QTableWidgetItem(formatted_amount)
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            amount_item.setForeground(QtGui.QColor('#27ae60'))
            self.finance_table.setItem(row, 5, amount_item)
            self.add_action_buttons(row, salary, "Salary", col_offset=6)
            self.finance_table.setRowHeight(row, 54)
            
    def _fin_render_page(self):
        import math
        total = len(self._fin_all_items)
        per_page = self._fin_per_page
        max_page = max(1, math.ceil(total / per_page))
        self._fin_page = max(1, min(self._fin_page, max_page))
        start_i = (self._fin_page - 1) * per_page
        end_i = min(start_i + per_page, total)
        page_items = self._fin_all_items[start_i:end_i]
        self.finance_table.setRowCount(len(page_items))
        if self.current_category == "Revenue":
            self._render_revenue_rows(page_items, start_i)
        elif self.current_category == "Expenses":
            self._render_expenses_rows(page_items, start_i)
        else:
            self._render_salary_rows(page_items, start_i)
        self._fin_resize_table()
        self._fin_rebuild_pagination(total, max_page)

    def _fin_resize_table(self):
        t = self.finance_table
        h = t.horizontalHeader().height()
        for i in range(t.rowCount()):
            h += t.rowHeight(i)
        t.setFixedHeight(h + 2)

    def _fin_rebuild_pagination(self, total, max_page):
        if not hasattr(self, '_fin_page_btns_layout'):
            return
        page_num = self._fin_page
        per_page = self._fin_per_page
        start = (page_num - 1) * per_page + 1 if total else 0
        end = min(page_num * per_page, total)
        if hasattr(self, '_fin_info_lbl'):
            category = self.current_category.lower() + "s" if not self.current_category.endswith("y") else self.current_category[:-1] + "ies"
            self._fin_info_lbl.setText(f"Showing {start}–{end} of {total} {self.current_category.lower()} records")
        while self._fin_page_btns_layout.count():
            item = self._fin_page_btns_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        _s = getattr(self, '_fin_pg_style', '')
        _win_start = max(1, min(page_num, max_page - 2))
        for p in range(_win_start, min(_win_start + 3, max_page + 1)):
            btn = QtWidgets.QPushButton(str(p))
            btn.setFixedSize(32, 28)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            if p == page_num:
                btn.setStyleSheet("""QPushButton {
                    background-color: #00756f; color: #ffffff;
                    border: 1px solid #00756f; border-radius: 6px;
                    font-size: 12px; font-weight: 700;
                    min-width: 32px; min-height: 28px; padding: 0 8px;
                }
                QPushButton:hover { background-color: #005f5a; color: #ffffff; }""")
            else:
                btn.setStyleSheet(_s)
                btn.clicked.connect(lambda _, pg=p: self._fin_go_to(pg))
            self._fin_page_btns_layout.addWidget(btn)
        if hasattr(self, '_fin_prev_btn'):
            self._fin_prev_btn.setEnabled(page_num > 1)
        if hasattr(self, '_fin_next_btn'):
            self._fin_next_btn.setEnabled(page_num < max_page)

    def _fin_go_prev(self):
        if self._fin_page > 1:
            self._fin_page -= 1
            self._fin_render_page()

    def _fin_go_next(self):
        self._fin_page += 1
        self._fin_render_page()

    def _fin_go_to(self, page):
        self._fin_page = page
        self._fin_render_page()

    def filter_finance_entries(self):
        if self.current_category == "Revenue":
            self.populate_revenue_data()
        elif self.current_category == "Expenses":
            self.populate_expenses_data()
        else:
            self.populate_salary_data()

        # 👇 ADD THIS LINE
        self.update_stats_cards()
        
    def filter_expenses_data(self):
        """Filter expenses data based on search and date filters"""
        search_text = self.search_edit.text().lower()
        
        # Helper to safely get amount as float for filtering
        def get_amount_value(expense):
            amount = expense.get('amount', 0)
            if isinstance(amount, (int, float)):
                return amount
            elif isinstance(amount, str):
                try:
                    return float(amount.replace('$', '').replace(',', ''))
                except:
                    return 0.0
            return 0.0
        
        # Check if date range filter is active
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        from_date = None
        to_date = None
        
        if date_range_active:
            from_date = self.current_from_date
            to_date = self.current_to_date
        
        filtered = []
        for expense in self.expenses_data:
            # ----- DATE RANGE FILTER (if active) -----
            matches_date = True
            if date_range_active and from_date and to_date:
                try:
                    expense_date = QtCore.QDate.fromString(expense.get('date', ''), "MM-dd-yyyy")
                    if expense_date.isValid():
                        matches_date = (from_date <= expense_date <= to_date)
                    else:
                        matches_date = False
                except:
                    matches_date = False
            else:
                # If no date range filter, filter by current year as before
                try:
                    date_obj = datetime.strptime(expense.get('date', ''), "%m-%d-%Y")
                    if date_obj.year != self.current_year:
                        continue
                except:
                    continue

            # ----- SEARCH FILTER -----
            amount_str = str(get_amount_value(expense))
            matches_search = (
                not search_text or
                search_text in expense.get('name', '').lower() or
                search_text in expense.get('description', '').lower() or
                search_text in amount_str.lower()
            )

            if matches_date and matches_search:
                filtered.append(expense)

        if date_range_active and from_date and to_date:
            # Date filter active → sort by date ascending (1 Jan → 31 Dec)
            def _exp_sort_key(e):
                try:
                    return datetime.strptime(e['date'], "%m-%d-%Y") if e.get('date') else datetime.min
                except (ValueError, TypeError):
                    return datetime.min
            filtered.sort(key=_exp_sort_key)
        else:
            # No filter → newest created_at first
            filtered.sort(key=lambda e: e.get('created_at') or '', reverse=True)
        return filtered

    def filter_revenue_data(self):
        """Filter revenue data based on search and date filters - respects status dates"""
        search_text = self.search_edit.text().lower()
        
        # Helper to safely get amount as string for searching
        def get_amount_str(revenue):
            amount = revenue.get('amount', 0)
            if isinstance(amount, (int, float)):
                return f"{amount:.2f}"
            elif isinstance(amount, str):
                return amount
            return "0"
        
        # Check if date range filter is active
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        from_date = None
        to_date = None
        
        if date_range_active:
            from_date = self.current_from_date
            to_date = self.current_to_date
        
        filtered = []
        for revenue in self.revenue_data:
            # Payment-tracker entries (is_payment=True) are internal records used only
            # for the per-month annual summary breakdown — hide them from the list view
            # so the invoice entry (is_invoice=True) remains the single visible record.
            if revenue.get('is_payment'):
                continue

            # Always filter by invoice date (the DATE column) for consistency
            status = revenue.get('status', 'Unpaid')
            date_str = revenue.get('date', '')

            # ----- DATE RANGE FILTER (if active) -----
            matches_date = True
            if date_range_active and from_date and to_date:
                try:
                    revenue_date = QtCore.QDate.fromString(date_str, "MM-dd-yyyy")
                    if revenue_date.isValid():
                        matches_date = (from_date <= revenue_date <= to_date)
                    else:
                        matches_date = False
                except:
                    matches_date = False
            else:
                # If no date range filter, filter by current year based on appropriate date
                try:
                    date_obj = datetime.strptime(date_str, "%m-%d-%Y")
                    if date_obj.year != self.current_year:
                        continue
                except:
                    continue

            # ----- SEARCH FILTER -----
            amount_str = get_amount_str(revenue)
            matches_search = (
                not search_text or
                search_text in revenue.get('source', '').lower() or
                search_text in revenue.get('description', '').lower() or
                search_text in amount_str.lower() or
                search_text in status.lower()
            )

            if matches_date and matches_search:
                filtered.append(revenue)

        if date_range_active and from_date and to_date:
            # Date filter active → sort by invoice date ascending (year → month → day)
            def _rev_sort_key(r):
                ds = r.get('date', '') or ''
                try:
                    return datetime.strptime(ds, "%m-%d-%Y")
                except (ValueError, TypeError):
                    return datetime.min
            filtered.sort(key=_rev_sort_key)
        else:
            # No filter → newest created_at first
            filtered.sort(key=lambda r: r.get('created_at') or '', reverse=True)
        return filtered

    def filter_salary_data(self):
        """Filter salary data based on search and date filters"""
        search_text = self.search_edit.text().lower()
        
        # Helper to safely get amount as string for searching
        def get_amount_str(salary):
            amount = salary.get('amount', 0)
            if isinstance(amount, (int, float)):
                return f"{amount:.2f}"
            elif isinstance(amount, str):
                return amount
            return "0"

        # Check if date range filter is active
        date_range_active = hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date')
        from_date = None
        to_date = None

        if date_range_active:
            from_date = self.current_from_date
            to_date = self.current_to_date

        filtered = []

        # -------- BUILD SALARY LIST --------
        all_salary = []
        for cat in ["Inside America", "Outside America"]:
            for salary in self.salary_data[cat]:
                salary_copy = salary.copy()
                salary_copy['region'] = cat
                all_salary.append(salary_copy)

        # -------- FILTER LOOP --------
        for salary in all_salary:
            # ----- DATE RANGE FILTER (if active) -----
            matches_date = True
            if date_range_active and from_date and to_date:
                try:
                    salary_date = QtCore.QDate.fromString(salary.get('date', ''), "MM-dd-yyyy")
                    if salary_date.isValid():
                        matches_date = (from_date <= salary_date <= to_date)
                    else:
                        matches_date = False
                except:
                    matches_date = False
            else:
                # If no date range filter, filter by current year as before
                try:
                    date_obj = datetime.strptime(salary.get('date', ''), "%m-%d-%Y")
                    if date_obj.year != self.current_year:
                        continue
                except:
                    continue

            # ----- SEARCH FILTER -----
            amount_str = get_amount_str(salary)
            matches_search = (
                not search_text or
                search_text in salary.get('name', '').lower() or
                search_text in salary.get('description', '').lower() or
                search_text in amount_str.lower() or
                search_text in salary.get('region', '').lower()
            )

            if matches_date and matches_search:
                filtered.append(salary)

        if date_range_active and from_date and to_date:
            # Date filter active → sort by date ascending (1 Jan → 31 Dec)
            def _sal_sort_key(s):
                try:
                    return datetime.strptime(s['date'], "%m-%d-%Y") if s.get('date') else datetime.min
                except (ValueError, TypeError):
                    return datetime.min
            filtered.sort(key=_sal_sort_key)
        else:
            # No filter → newest created_at first
            filtered.sort(key=lambda s: s.get('created_at') or '', reverse=True)
        return filtered

    def add_revenue_action_buttons(self, row, revenue: dict, col_offset: int):
        """Actions for revenue rows — Edit, Delete, plus View Payments for invoice entries."""
        actions_widget = QtWidgets.QWidget()
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(2, 2, 2, 2)
        actions_layout.setSpacing(3)

        _btn_ss = (
            "QPushButton{{background:{bg};border:1px solid {br};"
            "border-radius:6px;padding:0px;}}"
            "QPushButton:hover{{background:{hbg};border-color:{hbr};}}"
        )
        # View Payments button — only for invoice-linked entries
        if revenue.get('is_invoice') and revenue.get('invoice_number'):
            view_btn = QtWidgets.QPushButton()
            view_btn.setFixedSize(28, 28)
            view_btn.setIcon(BalanceSheetTab._make_action_icon("view", "#0f766e", 16))
            view_btn.setIconSize(QtCore.QSize(16, 16))
            view_btn.setToolTip("View payment history for this invoice")
            view_btn.setStyleSheet(_btn_ss.format(
                bg="white", br="#e2e8f0", hbg="#f0fdf4", hbr="#bbf7d0"))
            view_btn.clicked.connect(lambda checked=False, r=revenue: self.show_bs_payment_history(r))
            actions_layout.addWidget(view_btn)

        edit_btn = QtWidgets.QPushButton()
        edit_btn.setFixedSize(28, 28)
        edit_btn.setIcon(BalanceSheetTab._make_action_icon("edit", "#2563eb", 16))
        edit_btn.setIconSize(QtCore.QSize(16, 16))
        edit_btn.setToolTip("Edit entry")
        edit_btn.setStyleSheet(_btn_ss.format(
            bg="white", br="#e2e8f0", hbg="#eff6ff", hbr="#93c5fd"))
        edit_btn.clicked.connect(lambda: self.edit_entry(revenue, "Revenue"))

        delete_btn = QtWidgets.QPushButton()
        delete_btn.setFixedSize(28, 28)
        delete_btn.setIcon(BalanceSheetTab._make_action_icon("delete", "#f43f5e", 16))
        delete_btn.setIconSize(QtCore.QSize(16, 16))
        delete_btn.setToolTip("Delete entry")
        delete_btn.setStyleSheet(_btn_ss.format(
            bg="#fff1f2", br="#fecdd3", hbg="#ffe4e6", hbr="#fda4af"))
        delete_btn.clicked.connect(lambda: self.delete_entry(revenue, "Revenue"))

        actions_layout.addWidget(edit_btn)
        actions_layout.addWidget(delete_btn)
        actions_layout.addStretch()
        self.finance_table.setCellWidget(row, col_offset, actions_widget)
        self.finance_table.setRowHeight(row, 32)

    def show_bs_payment_history(self, revenue: dict):
        """Show payment history dialog for a balance-sheet invoice entry."""
        try:
            from payment_tracker import get_payment_tracker
            from tax_payment_tracker import get_tax_payment_tracker as _get_tt
            tracker = get_payment_tracker()
            tracker._load_payments()   # ensure fresh
            _bs_tax_tracker = _get_tt()
            _bs_tax_tracker._load_tax_payments()

            invoice_number = revenue.get('invoice_number', '')
            inv_total = self._money_to_float(revenue.get('amount', 0))

            # Project payments for this invoice (tax stored separately)
            all_payments = [
                p for p in tracker.payments
                if (p.invoice_number or '').strip() == invoice_number.strip()
                and (p.payment_stage or '').strip().lower() != 'tax'
            ]
            # Tax payments from the dedicated tax store
            _bs_inv_tax_pays = _bs_tax_tracker.get_invoice_taxes(invoice_number)

            total_paid = (
                sum(float(p.amount) for p in all_payments)
                + sum(float(t.amount) for t in _bs_inv_tax_pays)
            )
            remaining = max(inv_total - total_paid, 0.0)
            status = revenue.get('status', 'Unpaid')

            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle(f"Payment History — Invoice {invoice_number}")
            dialog.setWindowFlags(
                dialog.windowFlags()
                | QtCore.Qt.WindowMaximizeButtonHint
                | QtCore.Qt.WindowMinimizeButtonHint
            )
            dialog.setMinimumWidth(780)
            dialog.setMinimumHeight(500)
            layout = QtWidgets.QVBoxLayout(dialog)
            layout.setSpacing(10)
            layout.setContentsMargins(14, 14, 14, 14)

            # ── Header — load fresh invoice details from Firebase ─────────
            inv_client   = revenue.get('client_name', '') or ''
            inv_date     = revenue.get('invoice_date', '') or revenue.get('date', '') or ''
            inv_proj_names: list = []
            _inv_data = None   # initialise so it's always defined
            try:
                from main import FirebaseManager as _FM
                _raw_invs = _FM.load_invoices() or []
                _inv_data = next(
                    (i for i in _raw_invs
                     if (i.get('meta') or {}).get('invoice_number') == invoice_number),
                    None,
                )
                if _inv_data:
                    _meta = _inv_data.get('meta', {}) or {}
                    if not inv_client:
                        inv_client = _meta.get('client_name', '')
                    if not inv_date:
                        inv_date = _meta.get('date', '')
                    for _it in (_inv_data.get('items') or []):
                        _pn = str(_it.get('project_number', '') or '').strip()
                        _nm = str(_it.get('description', '') or _pn).strip()
                        if _nm and _nm not in inv_proj_names:
                            inv_proj_names.append(_nm)
            except Exception:
                pass

            if not inv_proj_names:
                _raw_desc = revenue.get('description', '') or ''
                inv_proj_names = [_raw_desc] if _raw_desc else []

            # Format project names: show all if ≤2, else first2 + "…"
            if len(inv_proj_names) == 0:
                proj_display = '—'
            elif len(inv_proj_names) <= 2:
                proj_display = ', '.join(inv_proj_names)
            else:
                proj_display = ', '.join(inv_proj_names[:2]) + f', +{len(inv_proj_names)-2} more…'

            hdr = QtWidgets.QFrame()
            hdr.setStyleSheet(
                "QFrame{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;}"
            )
            hg = QtWidgets.QGridLayout(hdr)
            hg.setContentsMargins(16, 12, 16, 12)
            hg.setHorizontalSpacing(24)
            hg.setVerticalSpacing(6)

            def _lbl(t, bold=False, color="#374151", wrap=False):
                l = QtWidgets.QLabel(str(t))
                l.setStyleSheet(
                    f"font-weight:{'700' if bold else '400'};"
                    f"color:{color};border:none;font-size:12px;"
                )
                if wrap:
                    l.setWordWrap(True)
                return l

            hg.addWidget(_lbl("Invoice #:",    True), 0, 0)
            hg.addWidget(_lbl(invoice_number),         0, 1)
            hg.addWidget(_lbl("Client:",       True), 0, 2)
            hg.addWidget(_lbl(inv_client or '—'),      0, 3)

            hg.addWidget(_lbl("Project:",      True), 1, 0)
            proj_lbl = _lbl(proj_display, wrap=True)
            proj_lbl.setMaximumWidth(340)
            hg.addWidget(proj_lbl,                     1, 1)
            hg.addWidget(_lbl("Invoice Date:", True), 1, 2)
            hg.addWidget(_lbl(inv_date or '—'),        1, 3)

            hg.addWidget(_lbl("Total Due:",    True), 2, 0)
            hg.addWidget(_lbl(f"${inv_total:,.2f}"),   2, 1)
            hg.addWidget(_lbl("Status:",       True), 2, 2)
            _st_color = {"Paid": "#15803d", "Partially Paid": "#1e40af",
                         "Overdue": "#b91c1c"}.get(status, "#78350f")
            sl = QtWidgets.QLabel(status)
            sl.setStyleSheet(
                f"font-weight:800;color:{_st_color};border:none;font-size:12px;"
            )
            hg.addWidget(sl, 2, 3)
            hg.setColumnStretch(1, 1)
            hg.setColumnStretch(3, 1)
            layout.addWidget(hdr)

            # ── Table — grouped by Project # ──────────────────────────────
            _tbl_hdr_lbl = QtWidgets.QLabel("Payment History by Project")
            _tbl_hdr_lbl.setStyleSheet(
                "font-weight:700;font-size:13px;color:#0f172a;border:none;"
                "padding:4px 0 2px 0;"
            )
            layout.addWidget(_tbl_hdr_lbl)
            tbl = QtWidgets.QTableWidget()
            COL_COUNT = 6
            tbl.setColumnCount(COL_COUNT)
            tbl.setHorizontalHeaderLabels(
                ["Project #", "Date", "Amount", "Method", "Stage", "Notes"]
            )
            tbl.horizontalHeader().setVisible(True)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            tbl.setAlternatingRowColors(False)
            tbl.verticalHeader().setVisible(False)
            tbl.setStyleSheet("""
                QTableWidget{background:white;border:1px solid #e2e8f0;
                             border-radius:6px;gridline-color:#f1f5f9;}
                QTableWidget::item{padding:6px 10px;color:#1e293b;}
                QHeaderView::section{background:#f8fafc;font-weight:700;padding:8px;
                    border:none;border-bottom:2px solid #e2e8f0;
                    min-height:34px;color:#374151;}
                QTableWidget::item:selected{background:#dbeafe;color:#1e40af;}
            """)

            import re as _re2

            def _fmt_d(raw):
                for fmt in ("%Y-%m-%d", "%m-%d-%Y"):
                    try:
                        return datetime.strptime(raw, fmt).strftime("%b %d, %Y")
                    except Exception:
                        pass
                return raw or "N/A"

            def _cell(t, al=QtCore.Qt.AlignCenter):
                it = QtWidgets.QTableWidgetItem(str(t))
                it.setTextAlignment(al)
                return it

            def _clean(s):
                return _re2.sub(r'\s*\(\d+%\)', '', s or '').strip() or '—'

            # ── Planned amount per project from loaded invoice items ───────
            planned_per_pn: dict = {}
            if _inv_data:
                try:
                    for _it in (_inv_data.get('items') or []):
                        _pn = str(_it.get('project_number', '') or '').strip()
                        if _pn:
                            _pd = float(_it.get('payment_due') or _it.get('unit_price') or 0)
                            planned_per_pn[_pn] = planned_per_pn.get(_pn, 0.0) + _pd
                except Exception:
                    pass

            # Determine project order: invoice items first (all projects), then
            # any extra project numbers found only in payments (edge case)
            pn_order: list = []
            seen_pns2: set = set()
            for _pn in planned_per_pn:
                if _pn and _pn not in seen_pns2:
                    pn_order.append(_pn)
                    seen_pns2.add(_pn)
            for _p in all_payments:
                _pn = (_p.project_number or "").strip()
                if _pn and _pn not in seen_pns2:
                    pn_order.append(_pn)
                    seen_pns2.add(_pn)

            _HDR_BG = QtGui.QColor("#1e3a5f")
            _HDR_FG = QtGui.QColor("#ffffff")
            _SUB_BG = QtGui.QColor("#f0f9ff")
            _SUB_FG = QtGui.QColor("#0369a1")
            _REM_GRN = QtGui.QColor("#15803d")
            _REM_RED = QtGui.QColor("#b91c1c")
            _TAX_HDR_BG = QtGui.QColor("#0f5a52")

            def _make_bg2(bg_color):
                it = QtWidgets.QTableWidgetItem("")
                it.setBackground(QtGui.QBrush(bg_color))
                return it

            # "sub" data = (paid, planned)
            rows_spec: list = []
            for pn in pn_order:
                pn_pays = sorted(
                    [p for p in all_payments
                     if (p.project_number or "").strip() == pn
                     and (p.payment_stage or "").strip().lower() != "tax"],
                    key=lambda p: p.payment_date or "",
                )
                pn_paid    = sum(float(p.amount) for p in pn_pays)
                pn_planned = planned_per_pn.get(pn, 0.0)
                rows_spec.append(("header", pn))
                for pay in pn_pays:
                    rows_spec.append(("pay", pay))
                rows_spec.append(("sub", (pn_paid, pn_planned)))

            if not all_payments and not pn_order:
                rows_spec = [("empty", None)]

            # ── TAX section (only when invoice has tax) ───────────────────
            # Reads from tax_payment_tracker (Firebase /tax_payments/)
            try:
                _bs_tax_amount = float(
                    ((_inv_data or {}).get('meta') or {}).get('tax_amount') or 0
                )
            except (TypeError, ValueError):
                _bs_tax_amount = 0.0
            if _bs_tax_amount > 0.005:
                tax_pays2 = sorted(_bs_inv_tax_pays, key=lambda t: t.payment_date or "")
                tax_paid_total2 = sum(float(t.amount) for t in tax_pays2)
                rows_spec.append(("tax_header", _bs_tax_amount))
                if tax_pays2:
                    for tp in tax_pays2:
                        rows_spec.append(("tax_pay", tp))
                else:
                    rows_spec.append(("tax_pending", (invoice_number, _bs_tax_amount)))
                rows_spec.append(("tax_sub", (tax_paid_total2, _bs_tax_amount)))

            tbl.setRowCount(len(rows_spec))
            for r, (kind, data) in enumerate(rows_spec):
                if kind == "header":
                    h = QtWidgets.QTableWidgetItem(f"  Project: {data}")
                    h.setBackground(QtGui.QBrush(_HDR_BG))
                    h.setForeground(QtGui.QBrush(_HDR_FG))
                    h.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                    h.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, h)
                    for c in range(1, COL_COUNT):
                        tbl.setItem(r, c, _make_bg2(_HDR_BG))
                    tbl.setSpan(r, 0, 1, COL_COUNT)
                    tbl.setRowHeight(r, 30)

                elif kind == "pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.project_number or "—").strip()))
                    tbl.setItem(r, 1, _cell(_fmt_d(pay.payment_date or "")))
                    amt_it = _cell(f"${float(pay.amount):,.2f}")
                    amt_it.setForeground(QtGui.QColor("#15803d"))
                    tbl.setItem(r, 2, amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or "—"))
                    tbl.setItem(r, 4, _cell(_clean(pay.payment_stage)))
                    tbl.setItem(r, 5, _cell(
                        pay.notes or "—",
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter,
                    ))
                    tbl.setRowHeight(r, 36)

                elif kind == "sub":
                    pn_paid, pn_planned = data
                    pn_remaining = max(pn_planned - pn_paid, 0.0)
                    _sub_font = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)

                    # Left: Paid (span 3)
                    paid_cell = QtWidgets.QTableWidgetItem(
                        f"  Paid: ${pn_paid:,.2f}"
                    )
                    paid_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    paid_cell.setForeground(QtGui.QBrush(_SUB_FG))
                    paid_cell.setFont(_sub_font)
                    paid_cell.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, paid_cell)
                    tbl.setItem(r, 1, _make_bg2(_SUB_BG))
                    tbl.setItem(r, 2, _make_bg2(_SUB_BG))
                    tbl.setSpan(r, 0, 1, 3)

                    # Right: Remaining (span 3)
                    _rem_color = _REM_GRN if pn_remaining <= 0 else _REM_RED
                    _rem_text = (
                        "Fully Paid ✓"
                        if pn_remaining <= 0
                        else f"Remaining: ${pn_remaining:,.2f}"
                    )
                    rem_cell = QtWidgets.QTableWidgetItem(f"{_rem_text}  ")
                    rem_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    rem_cell.setForeground(QtGui.QBrush(_rem_color))
                    rem_cell.setFont(_sub_font)
                    rem_cell.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, rem_cell)
                    tbl.setItem(r, 4, _make_bg2(_SUB_BG))
                    tbl.setItem(r, 5, _make_bg2(_SUB_BG))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                elif kind == "tax_header":
                    _th_item = QtWidgets.QTableWidgetItem("  TAX")
                    _th_item.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    _th_item.setForeground(QtGui.QBrush(QtGui.QColor("#ffffff")))
                    _th_item.setFont(QtGui.QFont("Consolas", 9, QtGui.QFont.Bold))
                    _th_item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, _th_item)
                    _th_note = QtWidgets.QTableWidgetItem(
                        f"Tax Amount: ${data:,.2f}  — Recorded when invoice is marked Paid  "
                    )
                    _th_note.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    _th_note.setForeground(QtGui.QBrush(QtGui.QColor("#a7f3d0")))
                    _th_note.setFont(QtGui.QFont("Segoe UI", 8))
                    _th_note.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 1, _th_note)
                    for c in range(2, COL_COUNT):
                        tbl.setItem(r, c, _make_bg2(_TAX_HDR_BG))
                    tbl.setSpan(r, 1, 1, COL_COUNT - 1)
                    tbl.setRowHeight(r, 30)

                elif kind == "tax_pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.invoice_number or "—").strip()))
                    tbl.setItem(r, 1, _cell(_fmt_d(pay.payment_date or "")))
                    _amt_it = _cell(f"${float(pay.amount):,.2f}")
                    _amt_it.setForeground(QtGui.QColor("#15803d"))
                    tbl.setItem(r, 2, _amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or "—"))
                    tbl.setItem(r, 4, _cell("Tax"))
                    tbl.setItem(r, 5, _cell(
                        pay.notes or "—",
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter,
                    ))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_pending":
                    inv_no2, tax_amt2 = data
                    tbl.setItem(r, 0, _cell(inv_no2))
                    tbl.setItem(r, 1, _cell("—"))
                    _pa = _cell(f"${tax_amt2:,.2f}")
                    _pa.setForeground(QtGui.QColor("#b45309"))
                    tbl.setItem(r, 2, _pa)
                    tbl.setItem(r, 3, _cell("—"))
                    _badge2 = QtWidgets.QTableWidgetItem("  Unpaid  ")
                    _badge2.setBackground(QtGui.QBrush(QtGui.QColor("#fef3c7")))
                    _badge2.setForeground(QtGui.QBrush(QtGui.QColor("#92400e")))
                    _badge2.setTextAlignment(QtCore.Qt.AlignCenter)
                    tbl.setItem(r, 4, _badge2)
                    tbl.setItem(r, 5, _cell(
                        "Pending — mark invoice as Paid to record",
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter,
                    ))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_sub":
                    _tp2, _tpl2 = data
                    _tr2 = max(_tpl2 - _tp2, 0.0)
                    _tsub_bg2 = QtGui.QColor("#f0fdf4") if _tr2 <= 0 else QtGui.QColor("#fef9c3")
                    _tsub_fg2 = QtGui.QColor("#15803d") if _tr2 <= 0 else QtGui.QColor("#92400e")
                    _tsub_font2 = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)
                    _tpc2 = QtWidgets.QTableWidgetItem(f"  Paid: ${_tp2:,.2f}")
                    _tpc2.setBackground(QtGui.QBrush(_tsub_bg2))
                    _tpc2.setForeground(QtGui.QBrush(_tsub_fg2))
                    _tpc2.setFont(_tsub_font2)
                    _tpc2.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, _tpc2)
                    tbl.setItem(r, 1, _make_bg2(_tsub_bg2))
                    tbl.setItem(r, 2, _make_bg2(_tsub_bg2))
                    tbl.setSpan(r, 0, 1, 3)
                    _trt2 = "Tax Paid ✓" if _tr2 <= 0 else f"Remaining: ${_tr2:,.2f}"
                    _trc2 = QtWidgets.QTableWidgetItem(f"{_trt2}  ")
                    _trc2.setBackground(QtGui.QBrush(_tsub_bg2))
                    _trc2.setForeground(QtGui.QBrush(_tsub_fg2))
                    _trc2.setFont(_tsub_font2)
                    _trc2.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, _trc2)
                    tbl.setItem(r, 4, _make_bg2(_tsub_bg2))
                    tbl.setItem(r, 5, _make_bg2(_tsub_bg2))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                else:
                    ni = QtWidgets.QTableWidgetItem("No payments recorded for this invoice")
                    ni.setTextAlignment(QtCore.Qt.AlignCenter)
                    tbl.setItem(r, 0, ni)
                    tbl.setSpan(r, 0, 1, COL_COUNT)
                    tbl.setRowHeight(r, 38)

            tbl.setColumnWidth(0, 165)   # Project #
            tbl.setColumnWidth(1, 115)   # Date
            tbl.setColumnWidth(2, 105)   # Amount
            tbl.setColumnWidth(3, 115)   # Method
            tbl.setColumnWidth(4, 130)   # Stage
            layout.addWidget(tbl)

            # ── Summary ───────────────────────────────────────────────────
            sf = QtWidgets.QFrame()
            sf.setFrameShape(QtWidgets.QFrame.NoFrame)
            sf.setStyleSheet(
                "QFrame{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;}"
                if remaining <= 0 else
                "QFrame{background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;}"
            )
            sl2 = QtWidgets.QHBoxLayout(sf)
            sl2.setContentsMargins(14, 8, 14, 8)

            def _sum_col(label, value, color="#1e293b"):
                col = QtWidgets.QVBoxLayout()
                l1 = QtWidgets.QLabel(label)
                l1.setStyleSheet("font-size:11px;color:#64748b;border:none;")
                l2 = QtWidgets.QLabel(value)
                l2.setStyleSheet(f"font-size:14px;font-weight:800;color:{color};border:none;")
                col.addWidget(l1); col.addWidget(l2)
                return col

            sl2.addLayout(_sum_col("Invoice Total", f"${inv_total:,.2f}"))
            sep = QtWidgets.QLabel("|")
            sep.setStyleSheet("color:#cbd5e1;font-size:18px;border:none;")
            sl2.addWidget(sep)
            sl2.addLayout(_sum_col("Total Paid", f"${total_paid:,.2f}", "#15803d"))
            sep2 = QtWidgets.QLabel("|")
            sep2.setStyleSheet("color:#cbd5e1;font-size:18px;border:none;")
            sl2.addWidget(sep2)
            sl2.addLayout(_sum_col("Remaining", f"${remaining:,.2f}",
                                   "#15803d" if remaining <= 0 else "#b45309"))
            sl2.addStretch()
            layout.addWidget(sf)

            close_btn = QtWidgets.QPushButton("Close")
            close_btn.setFixedHeight(34)
            close_btn.setStyleSheet("""
                QPushButton{background:#334155;color:white;border:none;border-radius:6px;
                            font-weight:bold;padding:0 22px;}
                QPushButton:hover{background:#1e293b;}
            """)
            close_btn.clicked.connect(dialog.accept)
            btn_row = QtWidgets.QHBoxLayout()
            btn_row.addStretch()
            btn_row.addWidget(close_btn)
            layout.addLayout(btn_row)
            dialog.exec_()

        except Exception as e:
            _log.warning("Error showing BS payment history: %s", e)
            QtWidgets.QMessageBox.critical(self, "Error", f"Could not load payment history:\n{e}")

    def add_action_buttons(self, row, data, category, col_offset):
        """Add Edit and Delete buttons to table"""
        actions_widget = QtWidgets.QWidget()
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(2, 2, 2, 2)  # Reduced margins
        actions_layout.setSpacing(3)  # Reduced spacing
        
        _btn_ss2 = (
            "QPushButton{{background:{bg};border:1px solid {br};"
            "border-radius:6px;padding:0px;}}"
            "QPushButton:hover{{background:{hbg};border-color:{hbr};}}"
        )
        edit_btn = QtWidgets.QPushButton()
        edit_btn.setFixedSize(28, 28)
        edit_btn.setIcon(BalanceSheetTab._make_action_icon("edit", "#2563eb", 16))
        edit_btn.setIconSize(QtCore.QSize(16, 16))
        edit_btn.setToolTip("Edit")
        edit_btn.setStyleSheet(_btn_ss2.format(
            bg="white", br="#e2e8f0", hbg="#eff6ff", hbr="#93c5fd"))
        edit_btn.clicked.connect(lambda: self.edit_entry(data, category))

        delete_btn = QtWidgets.QPushButton()
        delete_btn.setFixedSize(28, 28)
        delete_btn.setIcon(BalanceSheetTab._make_action_icon("delete", "#f43f5e", 16))
        delete_btn.setIconSize(QtCore.QSize(16, 16))
        delete_btn.setToolTip("Delete")
        delete_btn.setStyleSheet(_btn_ss2.format(
            bg="#fff1f2", br="#fecdd3", hbg="#ffe4e6", hbr="#fda4af"))
        delete_btn.clicked.connect(lambda: self.delete_entry(data, category))
        
        actions_layout.addWidget(edit_btn)
        actions_layout.addWidget(delete_btn)
        actions_layout.addStretch()
        
        self.finance_table.setCellWidget(row, col_offset, actions_widget)
        
        # Ensure the row height is sufficient
        self.finance_table.setRowHeight(row, 32)  # Slightly taller than buttons
    
    def show_date_range_dialog(self):
        """Show date range selection dialog"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📅 Select Date Range")
        dialog.setModal(True)
        dialog.resize(400, 200)
        dialog.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Title
        title = QtWidgets.QLabel("Select Date Range")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title)
        
        # Date inputs
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)
        
        # Get current dates
        current_from_date = QtCore.QDate.currentDate().addMonths(-1)
        current_to_date = QtCore.QDate.currentDate()
        
        if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date
        
        # From Date
        self.from_date_edit = QtWidgets.QDateEdit()
        self.from_date_edit.setDate(current_from_date)
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.from_date_edit.setMinimumWidth(148)
        self.from_date_edit.setFixedHeight(36)
        self.from_date_edit.wheelEvent = lambda e: e.ignore()
        self.from_date_edit.stepBy = lambda x: None
        self.from_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 6px 28px 6px 8px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
                font-weight: 600;
            }
            QDateEdit:focus { border-color: #00756f; }
        """)

        # To Date
        self.to_date_edit = QtWidgets.QDateEdit()
        self.to_date_edit.setDate(current_to_date)
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.to_date_edit.setMinimumWidth(148)
        self.to_date_edit.setFixedHeight(36)
        self.to_date_edit.wheelEvent = lambda e: e.ignore()
        self.to_date_edit.stepBy = lambda x: None
        self.to_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 6px 28px 6px 8px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
                font-weight: 600;
            }
            QDateEdit:focus { border-color: #00756f; }
        """)
        
        form_layout.addRow("From Date:", self.from_date_edit)
        form_layout.addRow("To Date:", self.to_date_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        clear_btn = QtWidgets.QPushButton("Clear Filter")
        clear_btn.setMinimumHeight(40)
        clear_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        
        apply_btn = QtWidgets.QPushButton("Apply Filter")
        apply_btn.setMinimumHeight(40)
        apply_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #2ecc71;
            }
        """)
        
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        button_layout.addWidget(apply_btn)
        
        layout.addLayout(button_layout)
        
        # Connect signals
        apply_btn.clicked.connect(lambda: self.apply_date_range_filter(dialog))
        clear_btn.clicked.connect(lambda: self.clear_date_range_filter(dialog))
        
        dialog.exec_()

    def load_all_years_data(self):
        """Load financial data from ALL years for date range filtering"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - loading all years from local data")
            self.load_all_years_local_data()
            return
            
        try:
            _log.info("Loading ALL years data from Firebase for date range filter...")
            
            # Clear existing data
            self.expenses_data = []
            self.revenue_data = []
            self.salary_data = {"Inside America": [], "Outside America": []}
            
            # Load all years data using BalanceSheetFirebaseManager
            # These now use the correct nodes
            all_expenses = BalanceSheetFirebaseManager.load_expenses()  # Now uses 'balance_sheet_expenses'
            all_revenue = BalanceSheetFirebaseManager.load_revenue()    # Uses 'revenue'
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue)
            all_salary = BalanceSheetFirebaseManager.load_salary()      # Uses 'salary'

            self.expenses_data = all_expenses
            self.revenue_data = all_revenue
            self.salary_data = all_salary
                
        except Exception as e:
            _log.warning("Error loading all years data from Firebase: %s", e)
            import traceback
            traceback.print_exc()
            self.load_all_years_local_data()
        
        self.update_stats_cards()
        self.on_category_changed(self.current_category)
    
    def load_all_years_local_data(self):
        """Load financial data from ALL years from local JSON files"""
        try:
            data_dir = Path.home() / ".mabs_finance"
            
            # Clear existing data
            self.expenses_data = []
            self.revenue_data = []
            self.salary_data = {"Inside America": [], "Outside America": []}
            
            # Find all expense files
            expense_files = list(data_dir.glob("expenses_*.json"))
            for expense_file in expense_files:
                year = expense_file.stem.replace("expenses_", "")
                if expense_file.exists():
                    with open(expense_file, 'r') as f:
                        year_expenses = json.load(f)
                        for expense in year_expenses:
                            expense['_year'] = year
                            self.expenses_data.append(expense)
            
            # Find all revenue files
            revenue_files = list(data_dir.glob("revenue_*.json"))
            for revenue_file in revenue_files:
                year = revenue_file.stem.replace("revenue_", "")
                if revenue_file.exists():
                    with open(revenue_file, 'r') as f:
                        year_revenue = json.load(f)
                        for revenue in year_revenue:
                            revenue['_year'] = year
                            self.revenue_data.append(revenue)
            
            # Find all salary files
            salary_files = list(data_dir.glob("salary_*.json"))
            for salary_file in salary_files:
                year = salary_file.stem.replace("salary_", "")
                if salary_file.exists():
                    with open(salary_file, 'r') as f:
                        year_salary = json.load(f)
                        for region, salaries in year_salary.items():
                            for salary in salaries:
                                salary['_year'] = year
                                self.salary_data[region].append(salary)
                            
        except Exception as e:
            _log.warning("Error loading all years local data: %s", e)
            self.expenses_data = []
            self.revenue_data = []
            self.salary_data = {"Inside America": [], "Outside America": []}
    
    def apply_date_range_filter(self, dialog):
        """Apply date range filter - Load ALL data across years in background thread."""
        from_date = self.from_date_edit.date()
        to_date = self.to_date_edit.date()

        self.current_from_date = from_date
        self.current_to_date = to_date

        from_date_formatted = from_date.toString("MM-dd-yy")
        to_date_formatted = to_date.toString("MM-dd-yy")

        configure_filter_button(
            self.date_range_button,
            f"{from_date_formatted} to {to_date_formatted}",
            active=True,
            height=36,
        )

        # Close dialog immediately so the UI stays responsive
        dialog.accept()

        # Load Firebase data in a background thread, then refresh the table
        import threading

        def _bg_load():
            try:
                if self.FIREBASE_AVAILABLE and self.db is not None:
                    all_expenses = BalanceSheetFirebaseManager.load_expenses()
                    all_revenue  = BalanceSheetFirebaseManager.load_revenue()
                    all_revenue  = BalanceSheetTab._dedup_is_payment_entries(all_revenue)
                    all_revenue  = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue)
                    all_salary   = BalanceSheetFirebaseManager.load_salary()
                    self.expenses_data = all_expenses
                    self.revenue_data  = all_revenue
                    self.salary_data   = all_salary
                else:
                    self.load_all_years_local_data()
            except Exception as _e:
                _log.warning("apply_date_range_filter bg load error: %s", _e)
            self._fin_load_signaler.loaded.emit()

        threading.Thread(target=_bg_load, daemon=True).start()

    def _after_date_range_load(self):
        """Called on the GUI thread after background data load for date-range filter."""
        self.update_stats_cards()
        self.filter_finance_entries()

    def clear_date_range_filter(self, dialog):
        """Clear date range filter and return to normal year-based view"""
        configure_filter_button(self.date_range_button, height=36)
        
        if hasattr(self, 'current_from_date'):
            del self.current_from_date
        if hasattr(self, 'current_to_date'):
            del self.current_to_date
        
        # Return to normal year-based view
        self.load_transaction_data()  # This loads only current year data
        self.filter_finance_entries()
        dialog.accept()

    def open_add_dialog(self):
        """Open appropriate dialog based on selected category"""
        if self.current_category == "Expenses":
            self.open_expense_dialog()
        elif self.current_category == "Revenue":
            self.open_revenue_dialog()
        else:  # Salary
            self.open_salary_dialog()

    def open_expense_dialog(self, edit_data=None):
        """Open dialog to add/edit expense with Firebase save - FIXED: All fields editable"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Add Expense" if not edit_data else "Edit Expense")
        dialog.setModal(True)
        dialog.resize(550, 400)
        dialog.setStyleSheet("""
            QLineEdit, QDateEdit, QComboBox, QTextEdit {
                font-size: 14px;
                padding: 8px 10px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus, QDateEdit:focus, QComboBox:focus, QTextEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
            QLabel {
                font-size: 13px;
                font-weight: 500;
            }
        """)
        
        # Set up keyboard shortcuts and navigation
        dialog.setFocusPolicy(QtCore.Qt.StrongFocus)
        
        layout = QtWidgets.QFormLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Title
        title = QtWidgets.QLabel("Expense Details")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addRow(title)

        # Date
        date_edit = QtWidgets.QDateEdit()
        date_edit.setCalendarPopup(True)
        date_edit.setDisplayFormat("MM-dd-yyyy")

        if edit_data:
            _raw = edit_data.get('date', '')
            _qd = QtCore.QDate()
            for _fmt in ("MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy", "M/d/yyyy"):
                _qd = QtCore.QDate.fromString(_raw, _fmt)
                if _qd.isValid():
                    break
            date_edit.setDate(_qd if _qd.isValid() else QtCore.QDate.currentDate())
            self.fix_date_edit(date_edit, set_today=False)
        else:
            self.fix_date_edit(date_edit, set_today=True)
        date_edit.setMinimumHeight(38)
        date_edit.setMinimumWidth(220)

        date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 10px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QDateEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)

        date_edit.calendarWidget().setMinimumWidth(380)
        layout.addRow("Date:", date_edit)
        
        # Expense Item - FIXED: Now editable
        name_edit = QtWidgets.QLineEdit()
        name_edit.setPlaceholderText("e.g., Office Rent, Utilities")
        if edit_data:
            name_edit.setText(edit_data.get('name', ''))
        name_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Expense Item:", name_edit)
        
        # Description - Using QTextEdit for multi-line, editable
        desc_edit = QtWidgets.QTextEdit()
        desc_edit.setPlaceholderText("Enter description...")
        desc_edit.setMaximumHeight(100)
        if edit_data:
            desc_edit.setText(edit_data.get('description', ''))
        desc_edit.setStyleSheet("""
            QTextEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QTextEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Description:", desc_edit)
        
        # Amount - with validation
        amount_edit = QtWidgets.QLineEdit()
        amount_edit.setPlaceholderText("0.00")
        if edit_data:
            amount_edit.setText(edit_data.get('amount', ''))
        amount_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Amount ($):", amount_edit)
        
        # Set up amount validation
        def validate_amount(text):
            # Allow only numbers, one decimal point, and digits
            if not text:
                return
            cursor_pos = amount_edit.cursorPosition()
            # Remove any non-digit or non-decimal characters
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            # Ensure only one decimal point
            if cleaned.count('.') > 1:
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])
            if cleaned != text:
                amount_edit.setText(cleaned)
                amount_edit.setCursorPosition(min(cursor_pos, len(cleaned)))
        
        amount_edit.textChanged.connect(validate_amount)
        
        # Setup navigation
        widgets = [date_edit, name_edit, desc_edit, amount_edit]
        for i in range(len(widgets) - 1):
            widgets[i].installEventFilter(self)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)  # Add spacing between buttons
        
        save_btn = QtWidgets.QPushButton("Save" if edit_data else "Add Expense")
        save_btn.setMinimumHeight(45)  # Increased height
        save_btn.setMinimumWidth(150)  # Set minimum width
        save_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #2ecc71; }
        """)
        
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setMinimumHeight(45)  # Increased height
        cancel_btn.setMinimumWidth(150)  # Set minimum width
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #c0392b; }
        """)
        cancel_btn.clicked.connect(dialog.reject)
        
        def save_expense():
            name = name_edit.text().strip() or "N/A"
            description = desc_edit.toPlainText().strip() or "N/A"
            amount = amount_edit.text().strip() or "0"

            expense_data = {
                'date': date_edit.date().toString("MM-dd-yyyy"),
                'name': name,
                'description': description,
                'amount': amount.replace('$', '').strip(),
                'year': self.current_year
            }

            if edit_data:
                if 'firebase_id' in edit_data:
                    expense_data['firebase_id'] = edit_data['firebase_id']

                if self.save_expense_to_firebase(expense_data, edit_data):
                    # FIX: Find by firebase_id instead of direct object comparison
                    found_index = -1
                    for i, exp in enumerate(self.expenses_data):
                        if exp.get('firebase_id') == expense_data.get('firebase_id'):
                            found_index = i
                            break
                    
                    if found_index >= 0:
                        self.expenses_data[found_index] = expense_data
                    else:
                        self.expenses_data.append(expense_data)
                    
                    # ALSO UPDATE ANNUAL DATA
                    annual_index = -1
                    for i, exp in enumerate(self.annual_expenses_data):
                        if exp.get('firebase_id') == expense_data.get('firebase_id'):
                            annual_index = i
                            break
                    if annual_index >= 0:
                        self.annual_expenses_data[annual_index] = expense_data
                    else:
                        self.annual_expenses_data.append(expense_data)
            else:
                if self.save_expense_to_firebase(expense_data):
                    self.expenses_data.append(expense_data)
                    # ADD TO ANNUAL DATA
                    self.annual_expenses_data.append(expense_data)

            self.filter_finance_entries()
            self.update_stats_cards()
            self.update_annual_summary()
            dialog.accept()
                
        save_btn.clicked.connect(save_expense)
        
        # Set up keyboard shortcuts
        save_btn.setShortcut(QtGui.QKeySequence("Ctrl+S"))
        save_btn.setToolTip("Save (Ctrl+S)")
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)
        
        # Set up Enter key navigation
        def handle_enter(event, current_widget):
            if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                if current_widget == date_edit:
                    name_edit.setFocus()
                elif current_widget == name_edit:
                    desc_edit.setFocus()
                elif current_widget == desc_edit:
                    amount_edit.setFocus()
                elif current_widget == amount_edit:
                    save_btn.setFocus()
                return True
            return False
        
        for widget in widgets:
            def create_handler(w):
                original = w.keyPressEvent
                def handler(event):
                    if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                        handle_enter(event, w)
                    else:
                        original(event)
                return handler
            widget.keyPressEvent = create_handler(widget)
    
        dialog.exec_()

    def refresh_invoice_history_tab(self, invoice_number=None):
        """Refresh invoice history tab to show updated invoice"""
        try:
            main_window = self.window()
            if hasattr(main_window, 'history_tab'):
                if invoice_number:
                    # If specific invoice, refresh that client's view
                    if hasattr(main_window.history_tab, 'refresh_invoices_immediately'):
                        main_window.history_tab.refresh_invoices_immediately()
                else:
                    # General refresh
                    if hasattr(main_window.history_tab, 'refresh_data'):
                        main_window.history_tab.refresh_data()
                _log.info("Invoice history tab refreshed")
        except Exception as e:
            _log.warning("Error refreshing invoice history tab: %s", e)
    
    def refresh_balance_sheet(self):
        """Force refresh balance sheet tab"""
        try:
            if hasattr(self, 'balance_sheet_tab'):
                self.balance_sheet_tab.load_all_financial_data()
                self.balance_sheet_tab.update_annual_summary()
                self.balance_sheet_tab.on_category_changed(self.balance_sheet_tab.current_category)
                self.balance_sheet_tab.update_stats_cards()
                _log.info("Balance sheet refreshed")
        except Exception as e:
            _log.warning("Error refreshing balance sheet: %s", e)
        
    def sync_revenue_to_invoice(self, revenue_data, old_revenue_data=None):
        """Sync changes from revenue entry back to the original invoice - handles all status types"""
        try:
            invoice_number = revenue_data.get('invoice_number')
            if not invoice_number:
                return
            
            _log.info("Syncing revenue changes to invoice: %s", invoice_number)
            
            if not self.FIREBASE_AVAILABLE:
                return
            
            from firebase_admin import db
            
            # Load the invoice from Firebase
            invoices_ref = db.reference('/invoices')
            all_invoices = invoices_ref.get()
            
            if not all_invoices:
                _log.info("⚠️ No invoices found in Firebase")
                return
            
            # Find the invoice
            invoice_id = None
            invoice_data = None
            for inv_id, inv_data in all_invoices.items():
                if inv_data and inv_data.get('meta', {}).get('invoice_number') == invoice_number:
                    invoice_id = inv_id
                    invoice_data = inv_data
                    break
            
            if not invoice_data:
                _log.warning("⚠️ Invoice %s not found in Firebase for sync", invoice_number)
                return
            
            # Check what changed and update invoice accordingly
            changed = False
            meta_updates = {}
            
            meta = invoice_data.get('meta', {})
            
            # Sync due date
            new_due_date = revenue_data.get('due_date', 'N/A')
            if new_due_date != meta.get('due_date', 'N/A'):
                meta_updates['due_date'] = new_due_date
                changed = True
                _log.info("(converted from print, see git history)")
            
            # Sync status - Now handles all 5 status types
            new_status = revenue_data.get('status', 'Pending')
            if new_status != meta.get('status', 'Pending'):
                meta_updates['status'] = new_status
                changed = True
                _log.info("(converted from print, see git history)")
            
            # Sync received date
            new_received_date = revenue_data.get('received_date', 'N/A')
            if new_received_date != meta.get('received_date', 'N/A'):
                meta_updates['received_date'] = new_received_date
                changed = True
                _log.info("(converted from print, see git history)")
            
            # Sync description (notes)
            new_description = revenue_data.get('description', '')
            if new_description != meta.get('notes', ''):
                meta_updates['notes'] = new_description
                changed = True
                _log.info("  - Notes updated")
            
            # Sync date
            new_date = revenue_data.get('date', '')
            if new_date and new_date != meta.get('date', ''):
                meta_updates['date'] = new_date
                changed = True
                _log.info("(converted from print, see git history)")
            
            if changed:
                meta_updates['updated_at'] = datetime.now(timezone.utc).isoformat()
                
                # Update the invoice in Firebase
                if meta_updates:
                    invoices_ref.child(invoice_id).child('meta').update(meta_updates)
                
                _log.info("Invoice %s updated from Balance Sheet changes", invoice_number)
                _log.info("   - Due Date: %s", new_due_date)
                _log.info("   - Status: %s", new_status)
                
                # Refresh invoice history if open
                self.refresh_invoice_history_tab(invoice_number)
            else:
                _log.info("ℹ️ No changes detected for invoice %s", invoice_number)
                    
        except Exception as e:
            _log.warning("Error syncing revenue to invoice: %s", e)
            import traceback
            traceback.print_exc()

    def refresh_invoice_history_tab(self, invoice_number=None):
        """Refresh invoice history tab to show updated invoice"""
        try:
            main_window = self.window()
            if hasattr(main_window, 'history_tab'):
                if invoice_number:
                    # If specific invoice, refresh that client's view
                    if hasattr(main_window.history_tab, 'refresh_invoices_immediately'):
                        main_window.history_tab.refresh_invoices_immediately()
                else:
                    # General refresh
                    if hasattr(main_window.history_tab, 'refresh_data'):
                        main_window.history_tab.refresh_data()
                _log.info("Invoice history tab refreshed")
        except Exception as e:
            _log.warning("Error refreshing invoice history tab: %s", e)
   
    def open_revenue_dialog(self, edit_data=None):
        """Open dialog to add/edit revenue with full status options (Paid, Unpaid, Pending, Overdue, Partially Paid)"""
        # Check if this revenue is linked to an invoice
        is_invoice_linked = edit_data and InvoiceRevenueLink.is_invoice_linked(edit_data)
        
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Add Revenue" if not edit_data else "Edit Revenue")
        dialog.setModal(True)
        dialog.resize(650, 520)
        dialog.setStyleSheet("""
            QLineEdit, QDateEdit, QComboBox, QTextEdit {
                font-size: 14px;
                padding: 8px 10px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus, QDateEdit:focus, QComboBox:focus, QTextEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
            QLabel {
                font-size: 13px;
                font-weight: 500;
            }
        """)
        
        dialog.setFocusPolicy(QtCore.Qt.StrongFocus)
        
        layout = QtWidgets.QFormLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Title
        title_text = "Revenue Details"
        if is_invoice_linked:
            invoice_num = edit_data.get('invoice_number', 'Unknown')
            title_text = f"💰 Revenue from Invoice: {invoice_num} (Linked to Invoice)"
        
        title = QtWidgets.QLabel(title_text)
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addRow(title)
        
        # Warning for invoice-linked revenue
        if is_invoice_linked:
            warning_label = QtWidgets.QLabel("⚠️ This revenue is linked to an invoice. Changes here will update the invoice in Invoice History.")
            warning_label.setStyleSheet("""
                QLabel {
                    color: #e67e22;
                    background-color: #fef5e7;
                    padding: 8px;
                    border-radius: 5px;
                    font-size: 11px;
                }
            """)
            warning_label.setWordWrap(True)
            layout.addRow(warning_label)
        
        # Revenue/Invoice Date
        date_edit = QtWidgets.QDateEdit()
        date_edit.setCalendarPopup(True)
        date_edit.setDisplayFormat("MM-dd-yyyy")

        self.fix_date_edit(date_edit)
        date_edit.wheelEvent = lambda event: None

        def keyPressEvent(event, original=date_edit.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        date_edit.keyPressEvent = keyPressEvent

        date_edit.stepBy = lambda x: None
        date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        date_edit.setMinimumHeight(38)
        date_edit.setMinimumWidth(220)
        
        # Set initial date if editing
        if edit_data:
            date_str = edit_data.get('date', '')
            if date_str:
                date_edit.setDate(QtCore.QDate.fromString(date_str, "MM-dd-yyyy"))
            else:
                date_edit.setDate(QtCore.QDate.currentDate())
        else:
            date_edit.setDate(QtCore.QDate.currentDate())
        
        if is_invoice_linked:
            date_edit.setEnabled(False)
            date_edit.setStyleSheet("""
                QDateEdit {
                    padding: 8px 10px;
                    border: 1px solid #bdc3c7;
                    border-radius: 5px;
                    background: #f0f0f0;
                    color: #999;
                }
            """)
        layout.addRow("Revenue Date:", date_edit)
        
        # Revenue Source
        source_edit = QtWidgets.QLineEdit()
        source_edit.setPlaceholderText("e.g., Client Payment, Service Revenue")
        if edit_data:
            source_edit.setText(edit_data.get('source', ''))
        if is_invoice_linked:
            source_edit.setEnabled(False)
            source_edit.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 5px;
                    background: #f0f0f0;
                    color: #999;
                }
            """)
        layout.addRow("Revenue Source:", source_edit)
        
        # Description
        desc_edit = QtWidgets.QTextEdit()
        desc_edit.setPlaceholderText("Enter description...")
        desc_edit.setMaximumHeight(100)
        if edit_data:
            desc_edit.setText(edit_data.get('description', ''))
        layout.addRow("Description:", desc_edit)
        
        # Amount
        amount_edit = QtWidgets.QLineEdit()
        amount_edit.setPlaceholderText("0.00")
        if edit_data:
            amount_edit.setText(edit_data.get('amount', ''))
        if is_invoice_linked:
            amount_edit.setEnabled(False)
            amount_edit.setStyleSheet("""
                QLineEdit {
                    padding: 8px;
                    border: 1px solid #bdc3c7;
                    border-radius: 5px;
                    background: #f0f0f0;
                    color: #999;
                }
            """)
        layout.addRow("Amount ($):", amount_edit)
        
        # Due Date
        due_date_edit = QtWidgets.QDateEdit()
        due_date_edit.setCalendarPopup(True)
        due_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.fix_date_edit(due_date_edit, set_today=False)
        due_date_edit.setMinimumHeight(38)
        due_date_edit.setMinimumWidth(220)
        
        due_date_label = QtWidgets.QLabel("Due Date:")
        
        # Status Dropdown
        status_combo = QtWidgets.QComboBox()
        status_combo.addItems(["Paid", "Unpaid", "Pending", "Overdue", "Partially Paid"])
        if edit_data:
            status_combo.setCurrentText(edit_data.get('status', 'Paid'))
        else:
            status_combo.setCurrentText("Paid")
        
        status_combo.setMinimumHeight(38)
        
        # Received Date
        received_date_edit = QtWidgets.QDateEdit()
        received_date_edit.setCalendarPopup(True)
        received_date_edit.setDisplayFormat("MM-dd-yyyy")
        
        # Set received date
        if edit_data:
            received_date_value = edit_data.get('received_date', '')
            if received_date_value and received_date_value != 'N/A':
                received_date_edit.setDate(QtCore.QDate.fromString(received_date_value, "MM-dd-yyyy"))
            else:
                received_date_edit.setDate(QtCore.QDate.currentDate())
        else:
            received_date_edit.setDate(QtCore.QDate.currentDate())
        
        received_date_edit.wheelEvent = lambda event: None

        def keyPressEvent_received(event, original=received_date_edit.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        received_date_edit.keyPressEvent = keyPressEvent_received

        received_date_edit.stepBy = lambda x: None
        received_date_edit.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        received_date_edit.setMinimumHeight(38)
        received_date_edit.setMinimumWidth(220)
        received_date_label = QtWidgets.QLabel("Received Date:")
        
        # Add widgets to layout
        layout.addRow(due_date_label, due_date_edit)
        layout.addRow("Status:", status_combo)
        layout.addRow(received_date_label, received_date_edit)
        
        # Store the actual due date value (for Unpaid statuses)
        actual_due_date = None
        
        # Initialize actual_due_date from edit data
        if edit_data:
            due_date_value = edit_data.get('due_date', 'N/A')
            if due_date_value and due_date_value != 'N/A':
                actual_due_date = due_date_value
            else:
                actual_due_date = None
        else:
            actual_due_date = None
        
        # Set initial due date value in the date picker if available
        if actual_due_date:
            try:
                due_date_edit.setDate(QtCore.QDate.fromString(actual_due_date, "MM-dd-yyyy"))
            except:
                due_date_edit.setDate(QtCore.QDate.currentDate().addDays(30))
        else:
            due_date_edit.setDate(QtCore.QDate.currentDate().addDays(30))
        
        # Function to update field visibility and display based on status
        def update_fields_and_display(status):
            is_paid = (status == "Paid")
            
            if is_paid:
                # For Paid status: Hide due date field, Show received date
                due_date_label.setVisible(False)
                due_date_edit.setVisible(False)
                received_date_label.setVisible(True)
                received_date_edit.setVisible(True)
            else:
                # For non-Paid status: Show due date field, Hide received date
                due_date_label.setVisible(True)
                due_date_edit.setVisible(True)
                received_date_label.setVisible(False)
                received_date_edit.setVisible(False)
                
                # Restore the actual due date value if it exists
                if actual_due_date:
                    try:
                        due_date_edit.setDate(QtCore.QDate.fromString(actual_due_date, "MM-dd-yyyy"))
                    except:
                        pass
        
        # Function to handle status changes
        def on_status_changed(new_status):
            nonlocal actual_due_date
            old_status = status_combo.currentText() if hasattr(status_combo, 'old_text') else None
            
            # When changing FROM non-Paid TO Paid: Store the current due date value
            if new_status == "Paid" and old_status != "Paid":
                if due_date_edit.isVisible():
                    actual_due_date = due_date_edit.date().toString("MM-dd-yyyy")
            
            # When changing FROM Paid TO non-Paid: Restore the stored due date
            if new_status != "Paid" and old_status == "Paid":
                if actual_due_date:
                    try:
                        due_date_edit.setDate(QtCore.QDate.fromString(actual_due_date, "MM-dd-yyyy"))
                    except:
                        pass
            
            # Update visibility and display
            update_fields_and_display(new_status)
            
            # Store old status for next change
            status_combo.old_text = new_status
        
        # Set initial visibility and display
        if edit_data:
            current_status = edit_data.get('status', 'Paid')
            update_fields_and_display(current_status)
            status_combo.old_text = current_status
        else:
            # New entry - default to Paid
            update_fields_and_display("Paid")
            status_combo.old_text = "Paid"
        
        # Connect status change
        status_combo.currentTextChanged.connect(on_status_changed)
        
        # Set up amount validation for non-linked revenues
        if not is_invoice_linked:
            def validate_amount(text):
                if not text:
                    return
                cursor_pos = amount_edit.cursorPosition()
                cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
                if cleaned.count('.') > 1:
                    parts = cleaned.split('.')
                    cleaned = parts[0] + '.' + ''.join(parts[1:])
                if cleaned != text:
                    amount_edit.setText(cleaned)
                    amount_edit.setCursorPosition(min(cursor_pos, len(cleaned)))
            amount_edit.textChanged.connect(validate_amount)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        save_btn = QtWidgets.QPushButton("Save" if edit_data else "Add Revenue")
        save_btn.setMinimumHeight(45)
        save_btn.setMinimumWidth(150)
        save_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #2ecc71; }
        """)
        
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setMinimumHeight(45)
        cancel_btn.setMinimumWidth(150)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #c0392b; }
        """)
        cancel_btn.clicked.connect(dialog.reject)
        
        def save_revenue():
            nonlocal actual_due_date
            status = status_combo.currentText()
            revenue_date = date_edit.date().toString("MM-dd-yyyy")
            
            # Determine due date based on status and user input
            if status == "Paid":
                # Check if this was originally Unpaid and changed to Paid
                # OR if this is a new Paid entry
                if actual_due_date:
                    # This came from Unpaid -> Paid, so keep the due date
                    due_date = actual_due_date
                else:
                    # This is/was Paid entry, so save as "N/A"
                    due_date = "N/A"
            else:
                # For non-Paid status: Get due date from date picker
                if due_date_edit.isVisible():
                    due_date = due_date_edit.date().toString("MM-dd-yyyy")
                    actual_due_date = due_date
                else:
                    due_date = actual_due_date if actual_due_date else "N/A"
            
            # Get received date
            received_date = "N/A"
            if status == "Paid":
                received_date = received_date_edit.date().toString("MM-dd-yyyy")
            
            # Determine year based on status
            if status == "Paid":
                date_parts = received_date.split('-')
                year = int(date_parts[2]) if len(date_parts) == 3 else self.current_year
            else:
                date_parts = revenue_date.split('-')
                year = int(date_parts[2]) if len(date_parts) == 3 else self.current_year
            
            revenue_data = {
                'date': revenue_date,
                'source': source_edit.text().strip() or "N/A",
                'description': desc_edit.toPlainText().strip() or "N/A",
                'amount': amount_edit.text().strip() or "0",
                'due_date': due_date,
                'status': status,
                'received_date': received_date,
                'year': year,
                'is_invoice': is_invoice_linked,
            }
            
            # Preserve invoice number if linked
            if is_invoice_linked and edit_data:
                revenue_data['invoice_number'] = edit_data.get('invoice_number', '')
            
            if edit_data:
                if 'firebase_id' in edit_data:
                    revenue_data['firebase_id'] = edit_data['firebase_id']
                
                if self.save_revenue_to_firebase(revenue_data, edit_data):
                    # Find and update in revenue_data
                    for i, rev in enumerate(self.revenue_data):
                        if rev.get('firebase_id') == revenue_data.get('firebase_id'):
                            self.revenue_data[i] = revenue_data
                            break
                    
                    # Update annual data
                    for i, rev in enumerate(self.annual_revenue_data):
                        if rev.get('firebase_id') == revenue_data.get('firebase_id'):
                            self.annual_revenue_data[i] = revenue_data
                            break
                    
                    # If this revenue is linked to an invoice, update the invoice
                    if is_invoice_linked and edit_data.get('invoice_number'):
                        self.sync_revenue_to_invoice(revenue_data, edit_data)
            else:
                if self.save_revenue_to_firebase(revenue_data):
                    self.revenue_data.append(revenue_data)
                    self.annual_revenue_data.append(revenue_data)
            
            self.filter_finance_entries()
            self.update_stats_cards()
            self.update_annual_summary()
            dialog.accept()
        
        save_btn.clicked.connect(save_revenue)
        save_btn.setShortcut(QtGui.QKeySequence("Ctrl+S"))
        save_btn.setToolTip("Save (Ctrl+S)")
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)
        
        # Simple Enter key navigation
        class EnterKeyFilter(QtCore.QObject):
            def __init__(self, parent, widgets, save_btn):
                super().__init__(parent)
                self.widgets = widgets
                self.save_btn = save_btn
            
            def eventFilter(self, obj, event):
                if event.type() == QtCore.QEvent.KeyPress:
                    if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                        if isinstance(obj, QtWidgets.QTextEdit):
                            if event.modifiers() == QtCore.Qt.ControlModifier:
                                return False
                            else:
                                for i, widget in enumerate(self.widgets):
                                    if widget == obj and widget.isVisible():
                                        for j in range(i + 1, len(self.widgets)):
                                            if self.widgets[j].isVisible():
                                                self.widgets[j].setFocus()
                                                return True
                                        self.save_btn.setFocus()
                                        return True
                        else:
                            for i, widget in enumerate(self.widgets):
                                if widget == obj and widget.isVisible():
                                    for j in range(i + 1, len(self.widgets)):
                                        if self.widgets[j].isVisible():
                                            self.widgets[j].setFocus()
                                            return True
                                    self.save_btn.setFocus()
                                    return True
                return False
        
        focus_widgets = [date_edit, source_edit, desc_edit, amount_edit, due_date_edit, status_combo, received_date_edit]
        key_filter = EnterKeyFilter(dialog, focus_widgets, save_btn)
        for widget in focus_widgets:
            widget.installEventFilter(key_filter)
        
        dialog.exec_()


    def open_salary_dialog(self, edit_data=None):
        """Open dialog to add/edit salary with Firebase save - FIXED: All fields editable"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Add Salary" if not edit_data else "Edit Salary")
        dialog.setModal(True)
        dialog.resize(550, 450)
        dialog.setStyleSheet("""
            QLineEdit, QDateEdit, QComboBox, QTextEdit {
                font-size: 14px;
                padding: 8px 10px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus, QDateEdit:focus, QComboBox:focus, QTextEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
            QLabel {
                font-size: 13px;
                font-weight: 500;
            }
        """)
        
        # Set up keyboard shortcuts and navigation
        dialog.setFocusPolicy(QtCore.Qt.StrongFocus)
        
        layout = QtWidgets.QFormLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Title
        title = QtWidgets.QLabel("Salary Details")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addRow(title)

        # Date
        date_edit = QtWidgets.QDateEdit()
        date_edit.setCalendarPopup(True)
        date_edit.setDisplayFormat("MM-dd-yyyy")

        if edit_data:
            _raw = edit_data.get('date', '')
            _qd = QtCore.QDate()
            for _fmt in ("MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy", "M/d/yyyy"):
                _qd = QtCore.QDate.fromString(_raw, _fmt)
                if _qd.isValid():
                    break
            date_edit.setDate(_qd if _qd.isValid() else QtCore.QDate.currentDate())
            self.fix_date_edit(date_edit, set_today=False)
        else:
            self.fix_date_edit(date_edit, set_today=True)
        date_edit.setMinimumHeight(38)
        date_edit.setMinimumWidth(220)

        date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 10px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QDateEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)

        date_edit.calendarWidget().setMinimumWidth(380)
        layout.addRow("Date:", date_edit)
        
        # Region/Category
        region_combo = QtWidgets.QComboBox()
        region_combo.addItems(["Inside America", "Outside America"])
        if edit_data:
            region_combo.setCurrentText(edit_data.get('region', 'Inside America'))
        region_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QComboBox:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Region:", region_combo)
        
        # Employee Name
        name_edit = QtWidgets.QLineEdit()
        name_edit.setPlaceholderText("Employee name")
        if edit_data:
            name_edit.setText(edit_data.get('name', ''))
        name_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Employee Name:", name_edit)
        
        # Description - Using QTextEdit for multi-line
        desc_edit = QtWidgets.QTextEdit()
        desc_edit.setPlaceholderText("Enter description (position, etc.)...")
        desc_edit.setMaximumHeight(100)
        if edit_data:
            desc_edit.setText(edit_data.get('description', ''))
        desc_edit.setStyleSheet("""
            QTextEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QTextEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Description:", desc_edit)
        
        # Amount - with validation
        amount_edit = QtWidgets.QLineEdit()
        amount_edit.setPlaceholderText("0.00")
        if edit_data:
            amount_edit.setText(edit_data.get('amount', ''))
        amount_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        layout.addRow("Amount ($):", amount_edit)
        
        # Set up amount validation
        def validate_amount(text):
            if not text:
                return
            cursor_pos = amount_edit.cursorPosition()
            # Remove any non-digit or non-decimal characters
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            # Ensure only one decimal point
            if cleaned.count('.') > 1:
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])
            if cleaned != text:
                amount_edit.setText(cleaned)
                amount_edit.setCursorPosition(min(cursor_pos, len(cleaned)))
        
        amount_edit.textChanged.connect(validate_amount)
        
        # Setup navigation
        widgets = [date_edit, region_combo, name_edit, desc_edit, amount_edit]
        for i in range(len(widgets) - 1):
            widgets[i].installEventFilter(self)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        save_btn = QtWidgets.QPushButton("Save" if edit_data else "Add Salary")
        save_btn.setMinimumHeight(45)
        save_btn.setMinimumWidth(150)
        save_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #2ecc71; }
        """)
        
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setMinimumHeight(45)
        cancel_btn.setMinimumWidth(150)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #c0392b; }
        """)
        cancel_btn.clicked.connect(dialog.reject)
        
        def save_salary():
            name = name_edit.text().strip() or "N/A"
            description = desc_edit.toPlainText().strip() or "N/A"
            amount = amount_edit.text().strip() or "0"
            region = region_combo.currentText()

            salary_data = {
                'date': date_edit.date().toString("MM-dd-yyyy"),
                'name': name,
                'description': description,
                'amount': amount.replace('$', '').strip(),
                'region': region,
                'year': self.current_year
            }

            if edit_data:
                if 'firebase_id' in edit_data:
                    salary_data['firebase_id'] = edit_data['firebase_id']

                if self.save_salary_to_firebase(salary_data, edit_data):
                    # Remove from old location
                    for cat in ["Inside America", "Outside America"]:
                        if edit_data in self.salary_data[cat]:
                            self.salary_data[cat].remove(edit_data)
                            break
                    
                    self.salary_data[region].append(salary_data)
                    
                    # ALSO UPDATE ANNUAL DATA
                    annual_index = -1
                    for i, sal in enumerate(self.annual_salary_data.get(region, [])):
                        if sal.get('firebase_id') == salary_data.get('firebase_id'):
                            annual_index = i
                            break
                    
                    # Remove from old location in annual data if found
                    if annual_index >= 0:
                        self.annual_salary_data[region][annual_index] = salary_data
                    else:
                        # Try to find in other region
                        for cat in ["Inside America", "Outside America"]:
                            for i, sal in enumerate(self.annual_salary_data.get(cat, [])):
                                if sal.get('firebase_id') == salary_data.get('firebase_id'):
                                    self.annual_salary_data[cat].pop(i)
                                    break
                        # Add to new region
                        self.annual_salary_data[region].append(salary_data)
            else:
                if self.save_salary_to_firebase(salary_data):
                    self.salary_data[region].append(salary_data)
                    # ADD TO ANNUAL DATA
                    self.annual_salary_data[region].append(salary_data)

            self.filter_finance_entries()
            self.update_stats_cards()
            self.update_annual_summary()  # This will now use updated annual data
            dialog.accept()
        
        save_btn.clicked.connect(save_salary)
        save_btn.setShortcut(QtGui.QKeySequence("Ctrl+S"))
        save_btn.setToolTip("Save (Ctrl+S)")
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addRow(button_layout)
        
        # Set up Enter key navigation
        def handle_enter(event, current_widget):
            if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                if current_widget == date_edit:
                    region_combo.setFocus()
                elif current_widget == region_combo:
                    name_edit.setFocus()
                elif current_widget == name_edit:
                    desc_edit.setFocus()
                elif current_widget == desc_edit:
                    amount_edit.setFocus()
                elif current_widget == amount_edit:
                    save_btn.setFocus()
                return True
            return False
        
        for widget in widgets:
            def create_handler(w):
                original = w.keyPressEvent
                def handler(event):
                    if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                        handle_enter(event, w)
                    else:
                        original(event)
                return handler
            widget.keyPressEvent = create_handler(widget)
    
        dialog.exec_()
    def set_future_date(self, date_edit, months=1):
        """Set date to today + given months"""
        today = QtCore.QDate.currentDate()
        future_date = today.addMonths(months)
        date_edit.setDate(future_date)
    
    def eventFilter(self, obj, event):
        """Handle Tab and Shift+Tab for navigation between fields"""
        if event.type() == QtCore.QEvent.KeyPress:
            if event.key() == QtCore.Qt.Key_Tab:
                # Find next widget
                parent = obj.parent()
                if parent and hasattr(parent, 'nextInFocusChain'):
                    parent.focusNextChild()
                    return True
            elif event.key() == QtCore.Qt.Key_Backtab:
                # Find previous widget (Shift+Tab)
                parent = obj.parent()
                if parent and hasattr(parent, 'focusPreviousChild'):
                    parent.focusPreviousChild()
                    return True
        return super().eventFilter(obj, event)

    def edit_entry(self, data, category):
        if category == "Expenses":
            if self.is_expense_from_expenses_tab(data):
                self.show_expense_edit_warning(data)
                return
            self.open_expense_dialog(edit_data=data)
        elif category == "Revenue":
            if InvoiceRevenueLink.is_invoice_linked(data):
                self.switch_to_invoice_management_for_revenue(data)
                return
            self.open_revenue_dialog(edit_data=data)
        else:  # Salary
            self.open_salary_dialog(edit_data=data)

    def switch_to_invoice_management_for_revenue(self, revenue_data):
        """Delegate to main_window.edit_invoice_by_number which owns Firebase access."""
        invoice_number = revenue_data.get("invoice_number", "")
        try:
            mw = self.main_window
            if hasattr(mw, "edit_invoice_by_number"):
                mw.edit_invoice_by_number(invoice_number)
            else:
                # Fallback: just navigate
                if hasattr(mw, "_nav_to"):
                    mw._nav_to(2)
                if hasattr(mw, "project_invoice_inner_tabs"):
                    mw.project_invoice_inner_tabs.setCurrentIndex(1)
        except Exception as exc:
            _log.warning("switch_to_invoice_management_for_revenue error: %s", exc)

    def is_expense_from_expenses_tab(self, expense_data):
        """Check if expense originated from Expenses Tab (not manually added in Balance Sheet)"""
        # Method 1: Check if it has expense_type field (Expenses Tab always has this)
        if expense_data.get('expense_type'):
            return True
        
        # Method 2: Check if it has Category field (Expenses Tab has Category)
        if expense_data.get('Category'):
            return True
        
        # Method 3: Check if it has expense_name field
        if expense_data.get('expense_name'):
            return True
        
        # Method 4: Check if it has vendor or payment_method fields
        if expense_data.get('vendor') or expense_data.get('payment_method'):
            return True
        
        # Method 5: Check if the expense exists in the main window's expenses_tab
        if hasattr(self.main_window, 'expenses_tab'):
            for exp in self.main_window.expenses_tab.expenses:
                if exp.get('firebase_id') == expense_data.get('firebase_id'):
                    # Found in expenses tab - definitely from Expenses Tab
                    return True
        
        return False

    def show_expense_edit_warning(self, expense_data):
        """Show professional warning dialog that expense should be edited in Expenses Tab"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📋 Expense Management")
        dialog.setModal(True)
        dialog.resize(500, 300)
        dialog.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(25, 25, 25, 25)
        
        # Icon and Header
        header_layout = QtWidgets.QHBoxLayout()
        
        info_icon = QtWidgets.QLabel("📋")
        info_icon.setStyleSheet("font-size: 48px;")
        header_layout.addWidget(info_icon)
        
        header_text = QtWidgets.QLabel("Expense Sync Information")
        header_text.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        header_layout.addWidget(header_text)
        header_layout.addStretch()
        
        layout.addLayout(header_layout)
        
        # Separator
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.HLine)
        separator.setStyleSheet("background-color: #bdc3c7;")
        layout.addWidget(separator)
        
        # Get the expense item name - FIXED: Use 'name' field which is what balance sheet uses
        expense_item = (
            expense_data.get('name') or           # Primary field from balance sheet
            expense_data.get('expense_name') or 
            expense_data.get('Expense Item') or 
            expense_data.get('type') or 
            expense_data.get('Category') or 
            'N/A'
        )
        
        # Safely get amount
        amount = expense_data.get('amount', 0)
        # Remove $ sign if present and format
        try:
            amount_value = float(str(amount).replace('$', '').replace(',', ''))
            amount_display = f"${amount_value:,.2f}"
        except:
            amount_display = f"${amount}"
        
        # Safely get date
        date_value = expense_data.get('date', 'N/A')
        
        # Message
        message = QtWidgets.QLabel(
            f"This expense was added through the <b>Expenses Tab</b>.\n\n"
            f"<b>Expense Details:</b><br>"
            f"• Expense Item: {expense_item}<br>"
            f"• Amount: {amount_display}<br>"
            f"• Date: {date_value}<br><br>"
            f"<span style='color: #e67e22;'>✏️ To edit this expense, please go to the <b>Expenses Tab</b>.</span><br>"
            f"<span style='color: #27ae60;'>🔄 Changes made in the Expenses Tab will automatically sync to the Balance Sheet.</span>"
        )
        message.setWordWrap(True)
        message.setStyleSheet("""
            QLabel {
                font-size: 13px;
                color: #2c3e50;
                line-height: 1.6;
                padding: 10px;
                background: white;
                border-radius: 8px;
            }
        """)
        layout.addWidget(message)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        # Go to Expenses Tab button
        go_to_expenses_btn = QtWidgets.QPushButton("📊 Go to Expenses Tab")
        go_to_expenses_btn.setMinimumHeight(45)
        go_to_expenses_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        go_to_expenses_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 13px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5dade2, stop:1 #3498db);
            }
        """)
        go_to_expenses_btn.clicked.connect(lambda: self.switch_to_expenses_tab_and_edit(dialog, expense_data))
        
        # Close button
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setMinimumHeight(45)
        close_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        close_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 13px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        
        button_layout.addWidget(go_to_expenses_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        dialog.exec_()
    
    def switch_to_expenses_tab_and_edit(self, dialog, expense_data):
        dialog.accept()

        try:
            if hasattr(self.main_window, "_nav_to"):
                self.main_window._nav_to(3)
            elif hasattr(self.main_window, "stack"):
                self.main_window.stack.setCurrentIndex(3)

            QtWidgets.QApplication.processEvents()

            if hasattr(self.main_window, "finance_inner_tabs"):
                for i in range(self.main_window.finance_inner_tabs.count()):
                    if "Expenses" in self.main_window.finance_inner_tabs.tabText(i):
                        self.main_window.finance_inner_tabs.setCurrentIndex(i)
                        break

            QtCore.QTimer.singleShot(
                400,
                lambda data=dict(expense_data): self.open_expense_in_expenses_tab(data)
            )
        except Exception as exc:
            _log.warning("Could not switch to Expenses tab for edit: %s", exc)
            QtWidgets.QMessageBox.warning(
                self,
                "Expense Edit",
                "Could not open the Expenses tab automatically. Please open Expenses and edit the item there.",
            )
        
    def open_expense_in_expenses_tab(self, expense_data):
        try:
            expenses_tab = getattr(self.main_window, "expenses_tab", None)
            if not expenses_tab:
                raise RuntimeError("Expenses tab is not available")

            if hasattr(expenses_tab, "load_expenses"):
                expenses_tab.load_expenses()

            target_ids = {
                str(expense_data.get("firebase_id", "")).strip(),
                str(expense_data.get("balance_sheet_firebase_id", "")).strip(),
            }
            target_ids.discard("")
            found_expense = None

            for exp in getattr(expenses_tab, "expenses", []) or []:
                exp_ids = {
                    str(exp.get("firebase_id", "")).strip(),
                    str(exp.get("balance_sheet_firebase_id", "")).strip(),
                }
                if target_ids and target_ids.intersection(exp_ids):
                    found_expense = exp
                    break

            if found_expense:
                expenses_tab.show_edit_expense_dialog(found_expense)
                return

            QtWidgets.QMessageBox.information(
                self,
                "Expense Not Found",
                "The expense could not be found in the Expenses tab. It may have been deleted or saved only in Balance Sheet.",
            )
        except Exception as exc:
            _log.warning("Could not open expense from Expenses tab: %s", exc)
            QtWidgets.QMessageBox.warning(
                self,
                "Expense Edit Error",
                f"Could not open the expense editor.\n\n{exc}",
            )
            
    def retry_find_expense(self, expense_data):
        """Retry finding expense after reload"""
        for exp in self.main_window.expenses_tab.expenses:
            if exp.get('firebase_id') == expense_data.get('firebase_id'):
                _log.info("Found expense after reload!")
                self.main_window.expenses_tab.edit_expense(exp)
                return
        
        # If still not found, show message
        QtWidgets.QMessageBox.information(
            self,
            "Expense Not Found",
            f"The expense could not be found in the Expenses Tab.\n\n"
            f"Firebase ID: {expense_data.get('firebase_id')}\n\n"
            f"This expense may have been deleted or may not exist in the Expenses Tab data."
        )
    
    def find_and_edit_expense_in_expenses_tab(self, expense_data):
        """Find and edit expense after reload"""
        for exp in self.main_window.expenses_tab.expenses:
            if exp.get('firebase_id') == expense_data.get('firebase_id'):
                self.main_window.expenses_tab.show_edit_expense_dialog(exp)
                return
        
        QtWidgets.QMessageBox.information(
            self,
            "Expense Not Found",
            "The expense could not be found in the Expenses Tab. It may have been deleted."
        )
    
    def save_expense_to_firebase(self, expense_data, old_data=None):
        """Save expense to Balance Sheet Firebase node"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - cannot save expense")
            QtWidgets.QMessageBox.warning(self, "Firebase Error", 
                                        "Firebase is not available. Data will not be saved.")
            return False
            
        try:
            entry_year = int(expense_data['date'].split("-")[2])
            expense_data['year'] = entry_year
            
            if old_data and 'firebase_id' in old_data:
                expense_data['firebase_id'] = old_data['firebase_id']
            
            # This uses BalanceSheetFirebaseManager which saves to 'balance_sheet_expenses'
            success = BalanceSheetFirebaseManager.save_expense(expense_data)
            
            if success:
                _log.info("Saved balance sheet expense to Firebase")
                # REMOVED: self.load_transaction_data() - this was causing duplication
                # REMOVED: self.update_annual_summary() - will be called by caller
            return success
            
        except Exception as e:
            _log.warning("Error saving balance sheet expense: %s", e)
            return False

    # In balance_sheet_tab.py, update the save_revenue_to_firebase method in BalanceSheetTab class:

    def save_revenue_to_firebase(self, revenue_data, old_data=None):
        """Save revenue to Firebase using flat structure - with status handling and due date"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - cannot save revenue")
            QtWidgets.QMessageBox.warning(self, "Firebase Error", 
                                        "Firebase is not available. Data will not be saved.")
            return False
            
        try:
            # The year is already determined in the dialog based on status
            entry_year = revenue_data.get('year', self.current_year)
            
            # If updating, pass firebase_id
            if old_data and 'firebase_id' in old_data:
                revenue_data['firebase_id'] = old_data['firebase_id']
            
            # Add timestamps
            revenue_data['updated_at'] = datetime.now(timezone.utc).isoformat()
            
            # Ensure due_date is present
            if 'due_date' not in revenue_data:
                revenue_data['due_date'] = 'N/A'
            
            success = BalanceSheetFirebaseManager.save_revenue(revenue_data)
            
            if success:
                _log.info("(converted from print, see git history)")
                _log.info("  - Year: %s", entry_year)
                _log.info("(converted from print, see git history)")
                
                # CRITICAL: If this revenue is linked to an invoice, sync to invoice
                if revenue_data.get('is_invoice') and revenue_data.get('invoice_number'):
                    self.sync_revenue_to_invoice(revenue_data, old_data)
                
            return success
            
        except Exception as e:
            _log.warning("Error saving revenue to Firebase: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Firebase Error", 
                                        f"Failed to save revenue: {str(e)}")
            return False

    def save_salary_to_firebase(self, salary_data, old_data=None):
        """Save salary to Firebase using flat structure"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - cannot save salary")
            QtWidgets.QMessageBox.warning(self, "Firebase Error", 
                                        "Firebase is not available. Data will not be saved.")
            return False
            
        try:
            entry_year = int(salary_data['date'].split("-")[2])
            salary_data['year'] = entry_year  # Store year in the data itself
            
            # If updating, pass firebase_id
            if old_data and 'firebase_id' in old_data:
                salary_data['firebase_id'] = old_data['firebase_id']
            
            success = BalanceSheetFirebaseManager.save_salary(salary_data)
            
            if success:
                _log.info("Saved salary to Firebase")
            return success
            
        except Exception as e:
            _log.warning("Error saving salary to Firebase: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Firebase Error", 
                                        f"Failed to save salary: {str(e)}")
            return False

    def delete_entry(self, data, category):
        """Delete a balance-sheet entry.

        Revenue entries are removed only from the balance-sheet tab (revenue_data,
        annual_revenue_data) and from Firebase /revenue.  Invoice history records
        and payment-tracker entries are intentionally NOT touched — they must
        remain intact after a balance-sheet-only deletion.
        """
        reply = QtWidgets.QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete this {category.rstrip('s')} entry?\n\n"
            f"This will remove the entry from the balance sheet and annual\n"
            f"summary only. Invoices and payment histories are not affected.",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )

        if reply == QtWidgets.QMessageBox.Yes:
            success = False

            if category == "Revenue":
                firebase_id = data.get('firebase_id')

                if firebase_id:
                    success = BalanceSheetFirebaseManager.delete_entry('revenue', firebase_id)
                    _log.info("Deleted from Firebase: revenue/%s", firebase_id)
                else:
                    _log.warning("No firebase_id found, can't delete from Firebase")
                    success = True

                if success:
                    # Remove from balance-sheet revenue cache
                    for i, rev in enumerate(self.revenue_data[:]):
                        if rev.get('firebase_id') == firebase_id or (not firebase_id and rev == data):
                            self.revenue_data.pop(i)
                            break

                    # Remove from annual summary cache
                    for i, ann_rev in enumerate(self.annual_revenue_data[:]):
                        if ann_rev.get('firebase_id') == firebase_id or (not firebase_id and ann_rev == data):
                            self.annual_revenue_data.pop(i)
                            break

                    # Also remove any linked balance-sheet payment-entries for this invoice
                    # (is_payment=True rows in /revenue that belong to the same invoice).
                    # These are balance-sheet-only rows; the actual payment-tracker records
                    # and invoice-history entries are NOT deleted.
                    if data.get('has_payment_entries') or data.get('is_invoice'):
                        inv_num = data.get('invoice_number', '')
                        linked_ids = [
                            r.get('firebase_id') for r in self.revenue_data
                            if r.get('is_payment') and r.get('invoice_number') == inv_num
                            and r.get('firebase_id')
                        ]
                        for lid in linked_ids:
                            try:
                                BalanceSheetFirebaseManager.delete_entry('revenue', lid)
                            except Exception as _e:
                                _log.warning("Could not delete linked payment entry %s: %s", lid, _e)
                        # Remove from local caches only
                        self.revenue_data = [
                            r for r in self.revenue_data
                            if not (r.get('is_payment') and r.get('invoice_number') == inv_num)
                        ]
                        self.annual_revenue_data = [
                            r for r in self.annual_revenue_data
                            if not (r.get('is_payment') and r.get('invoice_number') == inv_num)
                        ]

            elif category == "Expenses":
                firebase_id = data.get('firebase_id')
                if firebase_id:
                    success = BalanceSheetFirebaseManager.delete_entry('expenses', firebase_id)
                if success:
                    # Remove by firebase_id instead of direct comparison
                    for i, exp in enumerate(self.expenses_data[:]):
                        if exp.get('firebase_id') == firebase_id:
                            self.expenses_data.pop(i)
                            break
                    
                    for i, exp in enumerate(self.annual_expenses_data[:]):
                        if exp.get('firebase_id') == firebase_id:
                            self.annual_expenses_data.pop(i)
                            break
                            
            else:  # Salary
                firebase_id = data.get('firebase_id')
                if firebase_id:
                    success = BalanceSheetFirebaseManager.delete_entry('salary', firebase_id)
                if success:
                    for cat in ["Inside America", "Outside America"]:
                        for i, sal in enumerate(self.salary_data[cat][:]):
                            if sal.get('firebase_id') == firebase_id:
                                self.salary_data[cat].pop(i)
                                break
                    
                    for cat in ["Inside America", "Outside America"]:
                        for i, sal in enumerate(self.annual_salary_data.get(cat, [])[:]):
                            if sal.get('firebase_id') == firebase_id:
                                self.annual_salary_data[cat].pop(i)
                                break
            
            if success:
                self.filter_finance_entries()
                self.update_stats_cards()
                self.update_annual_summary()
                _log.info("Successfully deleted %s entry and updated all views", category)
            
    def load_transaction_data(self):
        """Load transaction table data from balance sheet Firebase node"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - using local data")
            self.load_transaction_data_local()
            return
            
        try:
            _log.info("Loading balance sheet data for year %s...", self.current_year)
            
            # Load from BALANCE SHEET node (separate from expenses tab)
            all_expenses = BalanceSheetFirebaseManager.load_expenses()
            self.expenses_data = [exp for exp in all_expenses if exp.get('year') == self.current_year]
            _log.info("Loaded %s balance sheet expenses for year %s", len(self.expenses_data), self.current_year)
            
            all_revenue = BalanceSheetFirebaseManager.load_revenue()
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue)
            self.revenue_data = [rev for rev in all_revenue if rev.get('year') == self.current_year]
            _log.info("Loaded %s revenue entries for year %s", len(self.revenue_data), self.current_year)
            
            all_salary = BalanceSheetFirebaseManager.load_salary()
            self.salary_data = {"Inside America": [], "Outside America": []}
            for region in ["Inside America", "Outside America"]:
                self.salary_data[region] = [sal for sal in all_salary[region] if sal.get('year') == self.current_year]
            _log.info("Loaded %s salary entries for year %s", sum(len(v) for v in self.salary_data.values()), self.current_year)
            
            # ALSO update annual summary data for current year
            self.annual_expenses_data = self.expenses_data.copy()
            self.annual_revenue_data = self.revenue_data.copy()
            self.annual_salary_data = self.salary_data.copy()
            
        except Exception as e:
            _log.warning("Error loading balance sheet data: %s", e)
            self.load_transaction_data_local()
        
        self.on_category_changed(self.current_category)
    
    def load_annual_summary_data_for_year(self, year):
        """Load annual summary data for a specific year from balance sheet node"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            self.load_annual_summary_local_for_year(year)
            return
            
        try:
            _log.info("Loading annual summary data for year %s from balance_sheet_expenses...", year)
            
            # Load from BALANCE SHEET node
            all_expenses = BalanceSheetFirebaseManager.load_expenses()
            self.annual_expenses_data = [exp for exp in all_expenses if exp.get('year') == year]
            _log.info("Loaded %s expenses for year %s", len(self.annual_expenses_data), year)
            
            all_revenue = BalanceSheetFirebaseManager.load_revenue()
            all_revenue = BalanceSheetTab._sync_revenue_statuses_from_invoices(all_revenue)
            self.annual_revenue_data = [rev for rev in all_revenue if rev.get('year') == year]
            _log.info("Loaded %s revenue entries for year %s", len(self.annual_revenue_data), year)
            
            all_salary = BalanceSheetFirebaseManager.load_salary()
            self.annual_salary_data = {"Inside America": [], "Outside America": []}
            for region in ["Inside America", "Outside America"]:
                self.annual_salary_data[region] = [sal for sal in all_salary[region] if sal.get('year') == year]
            _log.info("Loaded %s salary entries for year %s", sum(len(v) for v in self.annual_salary_data.values()), year)
            
        except Exception as e:
            _log.warning("Error loading annual summary: %s", e)
            import traceback
            traceback.print_exc()
            self.load_annual_summary_local_for_year(year)
        
        self.update_annual_summary()
    
    def load_financial_data(self):
        """Load financial data from Firebase - FIXED VERSION"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - using local data")
            self.load_local_data()
            return
            
        try:
            _log.info("Loading financial data for year %s from Firebase...", self.current_year)
            
            # Load expenses using Firebase manager
            self.expenses_data = BalanceSheetFirebaseManager.load_expenses()
            _log.info("Loaded %s expenses from Firebase", len(self.expenses_data))
            
            # Load revenue using Firebase manager
            self.revenue_data = BalanceSheetFirebaseManager.load_revenue(self.current_year)
            _log.info("Loaded %s revenue entries from Firebase", len(self.revenue_data))
            
            # Load salary using Firebase manager
            self.salary_data = BalanceSheetFirebaseManager.load_salary(self.current_year)
            _log.info("Loaded %s salary entries from Firebase", sum(len(v) for v in self.salary_data.values()))
            
            _log.info("Loaded financial data for year %s from Firebase", self.current_year)
            
        except Exception as e:
            _log.warning("Error loading from Firebase: %s", e)
            import traceback
            traceback.print_exc()
            self.load_local_data()
        
        self.update_stats_cards()
        self.update_annual_summary()
        self.on_category_changed(self.current_category)
    
    def load_local_data(self):
        """Load financial data from local JSON files (fallback)"""
        try:
            data_dir = Path.home() / ".mabs_finance"
            
            # Load expenses
            expense_file = data_dir / f"expenses_{self.current_year}.json"
            if expense_file.exists():
                with open(expense_file, 'r') as f:
                    self.expenses_data = json.load(f)
            else:
                self.expenses_data = []
                    
            # Load revenue
            revenue_file = data_dir / f"revenue_{self.current_year}.json"
            if revenue_file.exists():
                with open(revenue_file, 'r') as f:
                    self.revenue_data = json.load(f)
            else:
                self.revenue_data = []
                    
            # Load salary data
            salary_file = data_dir / f"salary_{self.current_year}.json"
            if salary_file.exists():
                with open(salary_file, 'r') as f:
                    self.salary_data = json.load(f)
            else:
                self.salary_data = {"Inside America": [], "Outside America": []}
                    
        except Exception as e:
            _log.warning("Error loading local data: %s", e)
            self.expenses_data = []
            self.revenue_data = []
            self.salary_data = {"Inside America": [], "Outside America": []}

    def calculate_total_expenses(self):
        """Calculate total expenses"""
        return sum(self._money_to_float(expense.get('amount', 0)) for expense in self.expenses_data)

    def calculate_total_revenue(self):
        """Calculate total revenue"""
        return sum(self._money_to_float(revenue.get('amount', 0)) for revenue in self.revenue_data)

    def calculate_total_salary(self):
        """Calculate total salary"""
        total = 0
        for cat in ["Inside America", "Outside America"]:
            for salary in self.salary_data.get(cat, []):
                total += self._money_to_float(salary.get('amount', 0))
        return total

    def _money_to_float(self, amount) -> float:
        """Parse currency values safely from Firebase/local JSON."""
        if isinstance(amount, (int, float)):
            return float(amount)
        if isinstance(amount, Decimal):
            return float(amount)
        if isinstance(amount, str):
            cleaned = amount.replace("$", "").replace(",", "").strip()
            try:
                return float(cleaned) if cleaned else 0.0
            except ValueError:
                return 0.0
        return 0.0

    def _parse_finance_date(self, date_text):
        """Accept the date formats used across invoices, expenses, and exports."""
        if not date_text or date_text == "N/A":
            return None
        text = str(date_text).strip()
        for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None

    def update_stats_cards(self):
        """Update stats cards (main values + sub-breakdown labels) from filtered data."""

        # ── Revenue ──────────────────────────────────────────────────────
        filtered_revenue = self.filter_revenue_data()
        total_revenue    = sum(self._money_to_float(r.get('amount', 0)) for r in filtered_revenue)
        paid_rev         = sum(self._money_to_float(r.get('amount', 0)) for r in filtered_revenue
                               if str(r.get('status', '')).strip().lower() == 'paid')
        unpaid_rev       = total_revenue - paid_rev

        # ── Expenses ─────────────────────────────────────────────────────
        filtered_expenses = self.filter_expenses_data()
        total_expenses    = sum(self._money_to_float(e.get('amount', 0)) for e in filtered_expenses)

        # ── Salary ───────────────────────────────────────────────────────
        filtered_salary = self.filter_salary_data()
        total_salary    = sum(self._money_to_float(s.get('amount', 0)) for s in filtered_salary)
        sal_inside      = sum(self._money_to_float(s.get('amount', 0)) for s in filtered_salary
                              if s.get('region') == 'Inside America')
        sal_outside     = sum(self._money_to_float(s.get('amount', 0)) for s in filtered_salary
                              if s.get('region') == 'Outside America')

        net       = total_revenue - total_expenses - total_salary
        net_color = "#15803d" if net >= 0 else "#b91c1c"

        # ── Update main value labels ──────────────────────────────────────
        self.revenue_value_label.setText(f"${total_revenue:,.2f}")
        self.expenses_value_label.setText(f"${total_expenses:,.2f}")
        self.salary_value_label.setText(f"${total_salary:,.2f}")

        if hasattr(self, "net_value_label"):
            self.net_value_label.setText(f"${net:,.2f}")
            self.net_value_label.setStyleSheet(
                f"font-size:22px; font-weight:900; color:{net_color};"
                " background:transparent; border:none;")
        if hasattr(self, "net_card"):
            _nbg, _nbrd = ("#f0fdf4", "#bbf7d0") if net >= 0 else ("#fff7f7", "#fecaca")
            self.net_card.setStyleSheet(
                f"QFrame {{ background:{_nbg}; border:1px solid {_nbrd}; border-radius:10px; }}")

        # ── Update hover tooltips ─────────────────────────────────────────
        if hasattr(self, "revenue_card"):
            self.revenue_card.setToolTip(
                f"<b>Paid Revenue:</b> ${paid_rev:,.2f}<br>"
                f"<b>Unpaid Revenue:</b> ${unpaid_rev:,.2f}")
        if hasattr(self, "expenses_card"):
            self.expenses_card.setToolTip(
                f"<b>Total:</b> ${total_expenses:,.2f}<br>"
                f"<b>Entries:</b> {len(filtered_expenses)}")
        if hasattr(self, "salary_card"):
            self.salary_card.setToolTip(
                f"<b>Inside America:</b> ${sal_inside:,.2f}<br>"
                f"<b>Outside America:</b> ${sal_outside:,.2f}")
        if hasattr(self, "net_card"):
            self.net_card.setToolTip(
                f"<b>Revenue:</b> ${total_revenue:,.2f}<br>"
                f"<b>Expenses:</b> -${total_expenses:,.2f}<br>"
                f"<b>Salaries:</b> -${total_salary:,.2f}")

        self.update_aging_section()
        
    # ------------------------------------------------------------------ #
    #  Invoice Aging Section                                               #
    # ------------------------------------------------------------------ #

    def create_aging_section(self, layout):
        """Horizontal bar showing outstanding invoice amounts by age bucket."""
        frame = QtWidgets.QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                margin: 0px 8px 4px 8px;
            }
        """)
        row = QtWidgets.QHBoxLayout(frame)
        row.setContentsMargins(20, 10, 20, 10)
        row.setSpacing(0)

        title = QtWidgets.QLabel("Invoice Aging (Outstanding)")
        title.setStyleSheet("font-size:12px; font-weight:700; color:#475569; border:none;")
        row.addWidget(title)
        row.addStretch()

        # Three bucket widgets — (label, value_label, color)
        self._aging_buckets = []
        for label_text, color in [
            ("0 – 30 days", "#10b981"),
            ("31 – 60 days", "#f59e0b"),
            ("61 + days",    "#ef4444"),
        ]:
            sep = QtWidgets.QFrame()
            sep.setFrameShape(QtWidgets.QFrame.VLine)
            sep.setStyleSheet("color: #e2e8f0; border: none; border-left: 1px solid #e2e8f0;")
            sep.setFixedWidth(1)
            row.addWidget(sep)

            cell = QtWidgets.QWidget()
            cell.setStyleSheet("border:none;")
            cell_lay = QtWidgets.QVBoxLayout(cell)
            cell_lay.setContentsMargins(24, 4, 24, 4)
            cell_lay.setSpacing(2)

            lbl = QtWidgets.QLabel(label_text)
            lbl.setStyleSheet(f"font-size:11px; color:#64748b; font-weight:600; border:none;")
            lbl.setAlignment(QtCore.Qt.AlignCenter)

            val = QtWidgets.QLabel("$0")
            val.setStyleSheet(f"font-size:16px; font-weight:800; color:{color}; border:none;")
            val.setAlignment(QtCore.Qt.AlignCenter)

            cell_lay.addWidget(lbl)
            cell_lay.addWidget(val)
            row.addWidget(cell)
            self._aging_buckets.append(val)

        layout.addWidget(frame)
        self.update_aging_section()

    def update_aging_section(self):
        """Recalculate aging buckets from current revenue data."""
        if not hasattr(self, '_aging_buckets'):
            return
        from status_enums import InvoiceStatus
        today = datetime.now().date()
        buckets = [0.0, 0.0, 0.0]   # 0-30, 31-60, 61+

        for rev in getattr(self, 'revenue_data', []):
            status = rev.get('status', '')
            if status not in InvoiceStatus.OPEN:
                continue
            due_raw = rev.get('due_date', '')
            if not due_raw or due_raw == 'N/A':
                continue
            try:
                due = datetime.strptime(due_raw, "%m-%d-%Y").date()
            except ValueError:
                continue
            days_past = (today - due).days
            amount = self._money_to_float(rev.get('amount', 0))
            if days_past <= 0:
                buckets[0] += amount      # not yet overdue — still in 0-30 bucket
            elif days_past <= 30:
                buckets[0] += amount
            elif days_past <= 60:
                buckets[1] += amount
            else:
                buckets[2] += amount

        for widget, value in zip(self._aging_buckets, buckets):
            widget.setText(f"${value:,.0f}")

    def get_group_box_style(self):
        """Return group box style matching JobFormTab"""
        return """
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                color: #2c3e50;
                border: 2px solid #dce4ec;
                border-radius: 6px;
                margin-top: 0.5em;
                padding-top: 7px;
                background: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 12px 0 12px;
                color: #2c3e50;
                font-weight: bold;
            }
        """

    # ------------------------------------------------------------------ #
    #  Annual table click handlers                                       #
    # ------------------------------------------------------------------ #

    def _on_annual_cell_clicked(self, row, col):
        """Open monthly revenue detail when a Revenue-row month cell is clicked."""
        if row == 0 and 0 <= col <= 11:
            dlg = self.RevenueDetailDialog(
                self, mode='month',
                year=self.annual_summary_year,
                month=col + 1,
            )
            dlg.exec_()

    def _on_annual_row_header_clicked(self, logical_index):
        """Open yearly revenue detail when the 'Revenue' row header is clicked."""
        if logical_index == 0:
            dlg = self.RevenueDetailDialog(
                self, mode='year',
                year=self.annual_summary_year,
            )
            dlg.exec_()

    # ------------------------------------------------------------------ #
    #  Revenue Detail Dialog                                              #
    # ------------------------------------------------------------------ #

    class RevenueDetailDialog(QtWidgets.QDialog):
        """Click-through popup: paid revenue by month (or full year) from the annual table."""

        def __init__(self, parent_tab, mode='month', year=None, month=None):
            super().__init__(parent_tab)
            self.parent_tab = parent_tab
            self.mode = mode          # 'month' | 'year'
            self._year = year or datetime.now().year
            self._month = month       # 1-12, only for mode='month'
            self._all_entries = []
            self._filtered_entries = []   # after date/search filter
            self._search_text = ''
            self._pg_page = 1
            self._pg_per_page = 15

            self.setWindowFlags(
                self.windowFlags() | QtCore.Qt.WindowMaximizeButtonHint)
            self.setModal(True)
            self.resize(1040, 640)
            self._init_ui()
            self._reload()

        # ── UI ─────────────────────────────────────────────────────────

        def _init_ui(self):
            root = QtWidgets.QVBoxLayout(self)
            root.setSpacing(0)
            root.setContentsMargins(0, 0, 0, 0)

            # ── Header bar ────────────────────────────────────────────
            hdr_w = QtWidgets.QWidget()
            hdr_w.setStyleSheet("background:#1e3a5f;")
            hdr_w.setFixedHeight(56)
            hdr_hl = QtWidgets.QHBoxLayout(hdr_w)
            hdr_hl.setContentsMargins(20, 0, 16, 0)
            hdr_hl.setSpacing(12)

            self._title_lbl = QtWidgets.QLabel()
            self._title_lbl.setStyleSheet(
                "font-size:16px;font-weight:bold;color:white;background:transparent;")
            hdr_hl.addWidget(self._title_lbl, 1)

            self._yr_btn = QtWidgets.QPushButton()
            self._yr_btn.setFixedSize(120, 32)
            self._yr_btn.setStyleSheet("""
                QPushButton{background:#2563eb;color:white;border:none;
                    border-radius:6px;font-size:12px;font-weight:bold;}
                QPushButton:hover{background:#1d4ed8;}
            """)
            self._yr_btn.clicked.connect(self._pick_year)
            hdr_hl.addWidget(self._yr_btn)
            root.addWidget(hdr_w)

            # ── Sub-toolbar (month nav / year date range) ──────────────
            sub_w = QtWidgets.QWidget()
            sub_w.setStyleSheet(
                "background:#f0fdf4;border-bottom:1px solid #bbf7d0;")
            sub_w.setFixedHeight(50)
            sub_hl = QtWidgets.QHBoxLayout(sub_w)
            sub_hl.setContentsMargins(16, 0, 16, 0)
            sub_hl.setSpacing(8)

            _nav_ss = """
                QPushButton{background:#dcfce7;color:#15803d;
                    border:1px solid #86efac;border-radius:6px;
                    font-size:13px;font-weight:bold;padding:0 12px;}
                QPushButton:hover{background:#bbf7d0;}
            """

            if self.mode == 'month':
                self._prev_btn = QtWidgets.QPushButton("◀  Prev")
                self._prev_btn.setFixedHeight(32)
                self._prev_btn.setStyleSheet(_nav_ss)
                self._prev_btn.clicked.connect(self._prev_month)

                self._period_lbl = QtWidgets.QLabel()
                self._period_lbl.setStyleSheet(
                    "font-size:14px;font-weight:bold;color:#166534;"
                    "background:transparent;padding:0 10px;")
                self._period_lbl.setAlignment(QtCore.Qt.AlignCenter)

                self._next_btn = QtWidgets.QPushButton("Next  ▶")
                self._next_btn.setFixedHeight(32)
                self._next_btn.setStyleSheet(_nav_ss)
                self._next_btn.clicked.connect(self._next_month)

                sub_hl.addWidget(self._prev_btn)
                sub_hl.addWidget(self._period_lbl, 1)
                sub_hl.addWidget(self._next_btn)
            else:
                # Year mode: date-range filter
                for lbl_txt in ["From:"]:
                    l = QtWidgets.QLabel(lbl_txt)
                    l.setStyleSheet(
                        "font-size:12px;font-weight:bold;color:#166534;"
                        "background:transparent;")
                    sub_hl.addWidget(l)

                self._from_de = QtWidgets.QDateEdit()
                self._from_de.setCalendarPopup(True)
                self._from_de.setDisplayFormat("MM-dd-yyyy")
                self._from_de.setMinimumHeight(32)
                self._from_de.setDate(QtCore.QDate(self._year, 1, 1))
                self._from_de.wheelEvent = lambda e: e.ignore()
                self._from_de.stepBy = lambda x: None
                sub_hl.addWidget(self._from_de)

                sub_hl.addSpacing(8)
                to_lbl = QtWidgets.QLabel("To:")
                to_lbl.setStyleSheet(
                    "font-size:12px;font-weight:bold;color:#166534;"
                    "background:transparent;")
                sub_hl.addWidget(to_lbl)

                self._to_de = QtWidgets.QDateEdit()
                self._to_de.setCalendarPopup(True)
                self._to_de.setDisplayFormat("MM-dd-yyyy")
                self._to_de.setMinimumHeight(32)
                self._to_de.setDate(QtCore.QDate(self._year, 12, 31))
                self._to_de.wheelEvent = lambda e: e.ignore()
                self._to_de.stepBy = lambda x: None
                sub_hl.addWidget(self._to_de)

                apply_btn = QtWidgets.QPushButton("Apply")
                apply_btn.setFixedHeight(32)
                apply_btn.setStyleSheet("""
                    QPushButton{background:#15803d;color:white;border:none;
                        border-radius:6px;font-size:12px;font-weight:bold;
                        padding:0 14px;}
                    QPushButton:hover{background:#166534;}
                """)
                apply_btn.clicked.connect(self._display)
                sub_hl.addSpacing(8)
                sub_hl.addWidget(apply_btn)
                sub_hl.addStretch()

            root.addWidget(sub_w)

            # ── Search + Export row ────────────────────────────────────
            tool_w = QtWidgets.QWidget()
            tool_w.setStyleSheet(
                "background:white;border-bottom:1px solid #e2e8f0;")
            tool_w.setFixedHeight(48)
            tool_hl = QtWidgets.QHBoxLayout(tool_w)
            tool_hl.setContentsMargins(16, 0, 16, 0)
            tool_hl.setSpacing(10)

            srch_lbl = QtWidgets.QLabel("🔍")
            srch_lbl.setStyleSheet("background:transparent;font-size:14px;")
            self._search_edit = QtWidgets.QLineEdit()
            self._search_edit.setPlaceholderText(
                "Search by source, description, amount, date…")
            self._search_edit.setMinimumHeight(32)
            self._search_edit.setStyleSheet("""
                QLineEdit{border:1px solid #d1d5db;border-radius:6px;
                    padding:0 10px;font-size:12px;background:white;}
                QLineEdit:focus{border:1px solid #15803d;}
            """)
            self._search_edit.textChanged.connect(self._on_search)

            exp_btn = QtWidgets.QPushButton("📄  Export PDF")
            exp_btn.setFixedHeight(32)
            exp_btn.setStyleSheet("""
                QPushButton{background:#1e3a5f;color:white;border:none;
                    border-radius:6px;font-size:12px;font-weight:bold;
                    padding:0 16px;}
                QPushButton:hover{background:#1d4ed8;}
            """)
            exp_btn.clicked.connect(self._export_pdf)

            tool_hl.addWidget(srch_lbl)
            tool_hl.addWidget(self._search_edit, 1)
            tool_hl.addSpacing(8)
            tool_hl.addWidget(exp_btn)
            root.addWidget(tool_w)

            # ── Table ──────────────────────────────────────────────────
            self._table = QtWidgets.QTableWidget()
            # month: S.No | Invoice Date | Revenue Source | Description | Invoice Total | Paid Amount | Paid Date | Actions
            # year:  S.No | Paid Month | Invoice Date | Revenue Source | Description | Invoice Total | Paid Amount | Paid Date | Actions
            if self.mode == 'month':
                cols = ["S.No", "Invoice Date", "Invoice #",
                        "Description", "Invoice Total",
                        "Paid Amount", "Paid Date", "Actions"]
                # Description=Stretch; money cols Fixed so they don't crowd Description
                col_modes = ['F','F','F','S','F','F','F','F']
                col_widths = [52, 115, 160, 0, 128, 128, 115, 95]
            else:
                cols = ["S.No", "Paid Month", "Invoice Date",
                        "Invoice #", "Description",
                        "Invoice Total", "Paid Amount", "Paid Date", "Actions"]
                # Description=Stretch; money cols Fixed
                col_modes = ['F','F','F','F','S','F','F','F','F']
                col_widths = [52, 125, 115, 165, 0, 128, 128, 115, 95]

            self._table.setColumnCount(len(cols))
            self._table.setHorizontalHeaderLabels(cols)
            self._table.setEditTriggers(
                QtWidgets.QAbstractItemView.NoEditTriggers)
            self._table.setSelectionBehavior(
                QtWidgets.QAbstractItemView.SelectRows)
            self._table.setSelectionMode(
                QtWidgets.QAbstractItemView.SingleSelection)
            self._table.setAlternatingRowColors(True)
            self._table.verticalHeader().setVisible(False)
            self._table.setShowGrid(True)
            self._table.setSortingEnabled(False)
            self._table.setHorizontalScrollBarPolicy(
                QtCore.Qt.ScrollBarAsNeeded)
            self._table.setStyleSheet("""
                QTableWidget{
                    background:white;
                    border:1px solid #e2e8f0;
                    border-radius:8px;
                    gridline-color:#e2e8f0;font-size:13px;
                    alternate-background-color:#f8fafc;
                }
                QTableWidget::item{padding:7px 10px;}
                QTableWidget::item:selected{
                    background:#dbeafe;color:#1e3a5f;}
                QHeaderView::section{
                    background:#f1f5f9;color:#374151;font-weight:700;
                    font-size:12px;padding:9px 10px;border:none;
                    border-bottom:2px solid #e2e8f0;
                    border-right:1px solid #e2e8f0;
                }
                QHeaderView::section:last{border-right:none;}
            """)
            hdr = self._table.horizontalHeader()
            for i, (mode, w) in enumerate(zip(col_modes, col_widths)):
                if mode == 'S':
                    hdr.setSectionResizeMode(
                        i, QtWidgets.QHeaderView.Stretch)
                else:
                    hdr.setSectionResizeMode(
                        i, QtWidgets.QHeaderView.Fixed)
                    self._table.setColumnWidth(i, w)
            self._table.verticalHeader().setDefaultSectionSize(48)
            root.addWidget(self._table, 1)

            # ── Footer (totals + pagination) ───────────────────────────
            foot_w = QtWidgets.QWidget()
            foot_w.setStyleSheet(
                "background:#f0fdf4;border-top:1px solid #bbf7d0;")
            foot_w.setFixedHeight(44)
            foot_hl = QtWidgets.QHBoxLayout(foot_w)
            foot_hl.setContentsMargins(16, 0, 16, 0)
            foot_hl.setSpacing(6)

            self._count_lbl = QtWidgets.QLabel()
            self._count_lbl.setStyleSheet(
                "font-size:11px;color:#6b7280;background:transparent;"
                "font-weight:600;")
            foot_hl.addWidget(self._count_lbl)
            foot_hl.addStretch()

            # Pagination controls (centre-right)
            _pg_s = (
                "QPushButton{background:#ffffff;color:#334155;"
                "border:1px solid #e2e8f0;border-radius:6px;"
                "font-size:12px;font-weight:700;"
                "min-width:30px;min-height:26px;padding:0 6px;}"
                "QPushButton:hover{background:#f1f5f9;border-color:#cbd5e1;}"
                "QPushButton:disabled{color:#cbd5e1;}")
            self._pg_prev_btn = QtWidgets.QPushButton("‹")
            self._pg_prev_btn.setStyleSheet(_pg_s)
            self._pg_prev_btn.setCursor(
                QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self._pg_prev_btn.clicked.connect(self._pg_go_prev)
            foot_hl.addWidget(self._pg_prev_btn)

            self._pg_btns_layout = QtWidgets.QHBoxLayout()
            self._pg_btns_layout.setSpacing(3)
            foot_hl.addLayout(self._pg_btns_layout)
            self._pg_style = _pg_s

            self._pg_next_btn = QtWidgets.QPushButton("›")
            self._pg_next_btn.setStyleSheet(_pg_s)
            self._pg_next_btn.setCursor(
                QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self._pg_next_btn.clicked.connect(self._pg_go_next)
            foot_hl.addWidget(self._pg_next_btn)
            foot_hl.addSpacing(12)

            self._total_lbl = QtWidgets.QLabel()
            self._total_lbl.setStyleSheet(
                "font-size:13px;font-weight:bold;color:#15803d;"
                "background:transparent;")
            foot_hl.addWidget(self._total_lbl)
            root.addWidget(foot_w)

        # ── Data helpers ────────────────────────────────────────────────

        def _fetch_revenue_for_year(self, year):
            pt = self.parent_tab
            if pt.annual_summary_year == year and pt.annual_revenue_data:
                return list(pt.annual_revenue_data)
            try:
                if pt.FIREBASE_AVAILABLE and pt.db is not None:
                    all_rev = BalanceSheetFirebaseManager.load_revenue()
                    return [r for r in all_rev if r.get('year') == year]
            except Exception as exc:
                _log.warning("RevenueDetailDialog: Firebase load: %s", exc)
            try:
                from pathlib import Path as _Path
                import json as _json
                fp = _Path.home() / ".mabs_finance" / f"revenue_{year}.json"
                if fp.exists():
                    with open(fp) as fh:
                        return _json.load(fh)
            except Exception as exc:
                _log.warning("RevenueDetailDialog: local load: %s", exc)
            return []

        def _extract_paid_entries(self, revenue_list, year, month=None):
            pt = self.parent_tab
            # Build invoice_number → invoice total lookup from parent records
            inv_totals = {}
            # Also build set of invoice numbers that have a real is_invoice entry
            # (same filter the annual summary uses — prevents orphaned is_payment
            # entries from appearing when their parent invoice was deleted)
            invoiced_numbers = set()
            for r in revenue_list:
                inv_no = r.get('invoice_number', '')
                if inv_no and not r.get('is_payment'):
                    inv_totals[inv_no] = pt._money_to_float(r.get('amount', 0))
                if r.get('is_invoice') and inv_no.strip():
                    invoiced_numbers.add(inv_no.strip())

            # Invoice numbers already covered by split-payment tracker records;
            # skip the parent Paid/PartiallyPaid row to prevent double-counting.
            split_inv_nos = {
                r.get('invoice_number', '') for r in revenue_list
                if r.get('is_payment') and r.get('invoice_number', '')
            }
            # Also include invoice numbers from the in-memory tracker — this covers the
            # window where tracker.add_payment() has run but the Firebase write is still
            # in flight, so the is_payment entry isn't in revenue_list yet.
            try:
                from payment_tracker import get_payment_tracker as _gpt
                _tk = _gpt()
                split_inv_nos |= {
                    (p.invoice_number or '').strip()
                    for p in _tk.payments
                    if (p.invoice_number or '').strip()
                }
            except Exception:
                pass

            results = []
            for rev in revenue_list:
                try:
                    if rev.get('has_payment_entries'):
                        continue
                    status = rev.get('status', '')
                    _created = rev.get('created_at', '')

                    # Payment-tracker individual entries and tax entries
                    if rev.get('is_payment'):
                        ds = rev.get('date', rev.get('received_date', ''))
                        dt = pt._parse_finance_date(ds)
                        if not dt or (year is not None and dt.year != year):
                            continue
                        if month and dt.month != month:
                            continue
                        inv_no   = (rev.get('invoice_number') or '').strip()
                        # Skip orphaned payment entries with no matching is_invoice record
                        # (matches the annual summary filter to keep totals consistent)
                        if inv_no and invoiced_numbers and inv_no not in invoiced_numbers:
                            continue
                        paid_amt = pt._money_to_float(rev.get('amount', 0))
                        inv_tot  = inv_totals.get(inv_no, paid_amt)
                        is_tax   = bool(rev.get('is_tax'))
                        stage    = 'TAX' if is_tax else rev.get('payment_stage', '')
                        results.append({
                            'invoice_date':   rev.get('date', 'N/A'),
                            'source':         rev.get('source', ''),
                            'description':    rev.get('description', ''),
                            'invoice_total':  inv_tot,
                            'paid_amount':    paid_amt,
                            'paid_date':      ds,
                            'paid_dt':        dt,
                            'created_at':     _created,
                            'status':         'Paid',
                            'invoice_number': inv_no,
                            'payment_stage':  stage,
                            '_is_payment':    True,
                            '_is_tax':        is_tax,
                        })
                        continue

                    if status == 'Paid':
                        _inv_chk = rev.get('invoice_number', '')
                        if _inv_chk and _inv_chk in split_inv_nos:
                            continue
                        paid_ds  = rev.get('received_date', rev.get('date', ''))
                        paid_dt  = pt._parse_finance_date(paid_ds)
                        if not paid_dt or (year is not None and paid_dt.year != year):
                            continue
                        if month and paid_dt.month != month:
                            continue
                        total = pt._money_to_float(rev.get('amount', 0))
                        results.append({
                            'invoice_date':   rev.get('date', 'N/A'),
                            'source':         rev.get('source', ''),
                            'description':    rev.get('description', ''),
                            'invoice_total':  total,
                            'paid_amount':    total,
                            'paid_date':      paid_ds,
                            'paid_dt':        paid_dt,
                            'created_at':     _created,
                            'status':         'Paid',
                            'invoice_number': rev.get('invoice_number', ''),
                            'payment_stage':  '',
                            '_is_payment':    False,
                            '_is_tax':        False,
                        })

                    elif status == 'Partially Paid':
                        _inv_chk = rev.get('invoice_number', '')
                        if _inv_chk and _inv_chk in split_inv_nos:
                            continue
                        dp_ds = rev.get('received_date',
                                        rev.get('down_payment_received_date', ''))
                        dp_dt = pt._parse_finance_date(dp_ds)
                        if not dp_dt or (year is not None and dp_dt.year != year):
                            continue
                        if month and dp_dt.month != month:
                            continue
                        inv_total  = pt._money_to_float(rev.get('amount', 0))
                        paid_split = pt._money_to_float(rev.get('paid_amount', 0))
                        results.append({
                            'invoice_date':   rev.get('date', 'N/A'),
                            'source':         rev.get('source', ''),
                            'description':    rev.get('description', ''),
                            'invoice_total':  inv_total,
                            'paid_amount':    paid_split if paid_split > 0 else inv_total,
                            'paid_date':      dp_ds,
                            'paid_dt':        dp_dt,
                            'created_at':     _created,
                            'status':         'Partially Paid',
                            'invoice_number': rev.get('invoice_number', ''),
                            'payment_stage':  'Down Payment',
                            '_is_payment':    False,
                            '_is_tax':        False,
                        })
                except Exception as exc:
                    _log.warning("RevenueDetailDialog skip entry: %s", exc)

            # Consolidate is_payment entries per invoice+date, keeping tax
            # entries separate from project payment entries so they display
            # as distinct rows (matching the balance-sheet payment-history view).
            payment_entries = [e for e in results if e.get('_is_payment') and not e.get('_is_tax')]
            tax_entries     = [e for e in results if e.get('_is_payment') and e.get('_is_tax')]
            other_entries   = [e for e in results if not e.get('_is_payment')]

            # Group project payments by invoice+date
            proj_groups: dict = {}
            for e in payment_entries:
                key = (e['invoice_number'], e['paid_date'])
                if key not in proj_groups:
                    proj_groups[key] = e.copy()
                    proj_groups[key]['payment_stage'] = e.get('payment_stage', '')
                else:
                    proj_groups[key]['paid_amount'] += e['paid_amount']
                    if e.get('payment_stage') and e['payment_stage'] not in proj_groups[key].get('description', ''):
                        proj_groups[key]['description'] = (
                            (proj_groups[key]['description'] + ', ' if proj_groups[key]['description'] else '')
                            + e['payment_stage']
                        )

            # Clean internal markers before returning
            for col in ('_is_payment', '_is_tax'):
                for e in list(proj_groups.values()) + tax_entries + other_entries:
                    e.pop(col, None)

            # Sort all entries together by paid_date (day 1→31) then created_at
            # so tax entries appear in chronological order alongside payments,
            # not forced to the end of each day.
            all_merged = other_entries + list(proj_groups.values()) + tax_entries
            all_merged.sort(key=lambda e: (e['paid_dt'], e.get('created_at', '')))
            return all_merged

        # ── Load & refresh ──────────────────────────────────────────────

        def _reload(self):
            if self.mode == 'month':
                rev_list = self._fetch_revenue_for_year(self._year)
                self._all_entries = self._extract_paid_entries(
                    rev_list, self._year, self._month)
            else:
                # Determine year span from date pickers
                if hasattr(self, '_from_de'):
                    from_year = self._from_de.date().year()
                    to_year   = self._to_de.date().year()
                else:
                    from_year = to_year = self._year
                if from_year == to_year:
                    rev_list = self._fetch_revenue_for_year(from_year)
                    self._all_entries = self._extract_paid_entries(
                        rev_list, from_year)
                else:
                    # Load all years in range with a single Firebase call when possible
                    try:
                        pt = self.parent_tab
                        if pt.FIREBASE_AVAILABLE and pt.db is not None:
                            all_rev = BalanceSheetFirebaseManager.load_revenue()
                            rev_list = [
                                r for r in all_rev
                                if from_year <= (r.get('year') or 0) <= to_year
                            ]
                        else:
                            raise RuntimeError("no db")
                    except Exception:
                        rev_list = []
                        for yr in range(from_year, to_year + 1):
                            rev_list.extend(self._fetch_revenue_for_year(yr))
                    # year=None → skip year filter; _display applies date-range
                    self._all_entries = self._extract_paid_entries(
                        rev_list, None)
            self._pg_page = 1
            self._update_labels()
            self._display()

        def _update_labels(self):
            import calendar as _cal
            if self.mode == 'month':
                mn = _cal.month_name[self._month]
                self.setWindowTitle(f"Paid Revenue — {mn} {self._year}")
                self._title_lbl.setText(
                    f"💰  Paid Revenue  ·  {mn} {self._year}")
                self._yr_btn.setText(f"📅  {self._year}")
                self._period_lbl.setText(f"{mn}  {self._year}")
            else:
                self.setWindowTitle(f"Annual Paid Revenue — {self._year}")
                self._title_lbl.setText(
                    f"💰  Annual Paid Revenue  ·  {self._year}")
                self._yr_btn.setText(f"📅  {self._year}")

        def _display(self):
            import math as _math
            search = self._search_text.lower().strip()
            entries = self._all_entries

            if self.mode == 'year' and hasattr(self, '_from_de'):
                from_py = self._from_de.date().toPyDate()
                to_py = self._to_de.date().toPyDate()
                entries = [e for e in entries
                           if from_py <= e['paid_dt'].date() <= to_py]
                # Sort ascending by date when filter applied (day 1→31, month/year low→high)
                entries = sorted(entries, key=lambda e: e['paid_dt'])

            if search:
                entries = [
                    e for e in entries
                    if (search in e['source'].lower()
                        or search in e['description'].lower()
                        or search in e['paid_date'].lower()
                        or search in e['invoice_date'].lower()
                        or search in e.get('invoice_number', '').lower()
                        or search in f"{e['paid_amount']:.2f}")
                ]

            # Cache filtered list for totals and pagination
            self._filtered_entries = entries
            total_all = len(entries)
            total_paid_all = sum(e['paid_amount'] for e in entries)

            # Pagination
            per_page = self._pg_per_page
            max_page = max(1, _math.ceil(total_all / per_page))
            self._pg_page = max(1, min(self._pg_page, max_page))
            start_i = (self._pg_page - 1) * per_page
            end_i = min(start_i + per_page, total_all)
            page_entries = entries[start_i:end_i]

            self._table.setRowCount(0)
            self._table.setRowCount(len(page_entries))

            for row, e in enumerate(page_entries):
                # global row number for S.No
                global_row = start_i + row
                # month: off=0  → Invoice Date = col 1
                # year:  off=1  → Paid Month = col 1, Invoice Date = col 2
                off = 0 if self.mode == 'month' else 1

                def _c(txt, align=QtCore.Qt.AlignCenter):
                    it = QtWidgets.QTableWidgetItem(str(txt))
                    it.setTextAlignment(align)
                    it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                    return it

                self._table.setItem(row, 0, _c(global_row + 1))

                if self.mode == 'year':
                    import calendar as _cal
                    self._table.setItem(
                        row, 1,
                        _c(f"{_cal.month_abbr[e['paid_dt'].month]} "
                           f"{e['paid_dt'].year}"))

                self._table.setItem(row, off + 1, _c(e['invoice_date']))
                _inv_display = e.get('invoice_number', '') or e['source']
                self._table.setItem(row, off + 2, _c(_inv_display))
                self._table.setItem(
                    row, off + 3,
                    _c(e['description'],
                       QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))

                tot_it = _c(f"${e['invoice_total']:,.2f}")
                tot_it.setForeground(QtGui.QColor('#64748b'))
                tot_it.setFont(QtGui.QFont("Segoe UI", 9))
                self._table.setItem(row, off + 4, tot_it)

                paid_it = _c(f"${e['paid_amount']:,.2f}")
                is_partial = e['paid_amount'] < e['invoice_total']
                paid_it.setForeground(
                    QtGui.QColor('#d97706' if is_partial else '#15803d'))
                paid_it.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                if is_partial:
                    paid_it.setToolTip(
                        f"Partial: ${e['paid_amount']:,.2f} of "
                        f"${e['invoice_total']:,.2f}")
                self._table.setItem(row, off + 5, paid_it)

                self._table.setItem(row, off + 6, _c(e['paid_date']))

                action_col = off + 7
                act_w = QtWidgets.QWidget()
                act_w.setStyleSheet("background:transparent;")
                act_hl = QtWidgets.QHBoxLayout(act_w)
                act_hl.setContentsMargins(6, 4, 6, 4)
                act_hl.setSpacing(0)
                act_hl.setAlignment(QtCore.Qt.AlignCenter)

                _rev_row = e
                view_btn = QtWidgets.QPushButton("View")
                view_btn.setFixedSize(62, 32)
                view_btn.setToolTip("")
                view_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #15803d; color: white;
                        border: none; border-radius: 6px;
                        padding: 0px; font-size: 11px; font-weight: bold;
                    }
                    QPushButton:hover { background-color: #166534; }
                    QPushButton:pressed { background-color: #14532d; }
                """)
                view_btn.clicked.connect(
                    lambda _, rv=_rev_row:
                        self._show_invoice_payment_history(rv))
                act_hl.addWidget(view_btn)
                self._table.setCellWidget(row, action_col, act_w)
                self._table.setRowHeight(row, 48)

            # Footer labels
            n = total_all
            self._count_lbl.setText(
                f"Showing {start_i + 1}–{end_i} of {n} "
                f"entr{'y' if n == 1 else 'ies'}" if n else "0 entries")
            self._total_lbl.setText(f"Total Paid:   ${total_paid_all:,.2f}")

            # Rebuild pagination buttons
            self._pg_rebuild(total_all, max_page)

        def _pg_rebuild(self, total, max_page):
            while self._pg_btns_layout.count():
                item = self._pg_btns_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            _s = self._pg_style
            page_num = self._pg_page
            win_start = max(1, min(page_num, max_page - 2))
            for p in range(win_start, min(win_start + 3, max_page + 1)):
                btn = QtWidgets.QPushButton(str(p))
                btn.setFixedSize(30, 26)
                btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                if p == page_num:
                    btn.setStyleSheet(
                        "QPushButton{background:#00756f;color:#fff;"
                        "border:1px solid #00756f;border-radius:6px;"
                        "font-size:12px;font-weight:700;"
                        "min-width:30px;min-height:26px;padding:0 6px;}"
                        "QPushButton:hover{background:#005f5a;color:#fff;}")
                else:
                    btn.setStyleSheet(_s)
                    btn.clicked.connect(lambda _, pg=p: self._pg_go_to(pg))
                self._pg_btns_layout.addWidget(btn)
            self._pg_prev_btn.setEnabled(page_num > 1)
            self._pg_next_btn.setEnabled(page_num < max_page)

        def _pg_go_prev(self):
            if self._pg_page > 1:
                self._pg_page -= 1
                self._display()

        def _pg_go_next(self):
            self._pg_page += 1
            self._display()

        def _pg_go_to(self, page):
            self._pg_page = page
            self._display()

        # ── Navigation ──────────────────────────────────────────────────

        def _prev_month(self):
            if self._month == 1:
                self._month = 12
                self._year -= 1
            else:
                self._month -= 1
            self._reload()

        def _next_month(self):
            if self._month == 12:
                self._month = 1
                self._year += 1
            else:
                self._month += 1
            self._reload()

        def _on_search(self, text):
            self._search_text = text
            self._pg_page = 1
            self._display()

        def _pick_year(self):
            dlg = self.parent_tab.YearPickerDialog(self, self._year)
            if dlg.exec_() == QtWidgets.QDialog.Accepted:
                self._year = dlg.selected_year
                if self.mode == 'year' and hasattr(self, '_from_de'):
                    self._from_de.setDate(QtCore.QDate(self._year, 1, 1))
                    self._to_de.setDate(QtCore.QDate(self._year, 12, 31))
                self._reload()

        def _show_invoice_payment_history(self, entry):
            """Show payment history for an invoice — identical layout to show_bs_payment_history."""
            import re as _re
            invoice_number = entry.get('invoice_number', '')
            source         = entry.get('source', '')
            inv_date       = entry.get('invoice_date', '')
            pt             = self.parent_tab

            # ── Load project payments (exclude tax stage) ──────────────
            try:
                from payment_tracker import get_payment_tracker as _get_pt
                _tracker = _get_pt()
                _tracker._load_payments()
                all_payments = [
                    p for p in _tracker.payments
                    if (p.invoice_number or '').strip() == invoice_number.strip()
                    and (p.payment_stage or '').strip().lower() != 'tax'
                ]
            except Exception:
                all_payments = []

            # ── Load tax payments ──────────────────────────────────────
            try:
                from tax_payment_tracker import get_tax_payment_tracker as _get_tt
                _tax_tracker = _get_tt()
                _tax_tracker._load_tax_payments()
                inv_tax_pays = _tax_tracker.get_invoice_taxes(invoice_number)
            except Exception:
                inv_tax_pays = []

            # ── Load invoice data from Firebase for correct totals ─────
            _inv_data   = None
            inv_client  = ''
            inv_proj_names: list = []
            planned_per_pn: dict = {}
            inv_total   = entry.get('invoice_total', 0.0)  # fallback
            status      = entry.get('status', '')
            try:
                from main import FirebaseManager as _FM
                _raw_invs = _FM.load_invoices() or []
                _inv_data = next(
                    (i for i in _raw_invs
                     if (i.get('meta') or {}).get('invoice_number') == invoice_number),
                    None)
                if _inv_data:
                    _meta = _inv_data.get('meta', {}) or {}
                    inv_client = _meta.get('client_name', '') or ''
                    if not inv_date:
                        inv_date = _meta.get('date', '')
                    # Use the authoritative invoice total from Firebase meta
                    _fb_total = pt._money_to_float(_meta.get('amount', 0))
                    if _fb_total > 0:
                        inv_total = _fb_total
                    if not status:
                        status = _meta.get('status', '')
                    for _it in (_inv_data.get('items') or []):
                        _pn = str(_it.get('project_number', '') or '').strip()
                        _nm = str(_it.get('description', '') or _pn).strip()
                        if _nm and _nm not in inv_proj_names:
                            inv_proj_names.append(_nm)
                        if _pn:
                            _pd = float(_it.get('payment_due') or
                                        _it.get('unit_price') or 0)
                            planned_per_pn[_pn] = planned_per_pn.get(_pn, 0.0) + _pd
            except Exception:
                pass

            if not inv_proj_names:
                inv_proj_names = [entry.get('description', '')] if entry.get('description') else []

            # Tax amount from invoice meta
            try:
                _tax_amount = float(
                    ((_inv_data or {}).get('meta') or {}).get('tax_amount') or 0
                )
            except (TypeError, ValueError):
                _tax_amount = 0.0

            tax_paid_total = sum(float(t.amount) for t in inv_tax_pays)
            proj_paid_total = sum(float(p.amount) for p in all_payments)
            total_paid = proj_paid_total + tax_paid_total
            remaining  = max(inv_total - total_paid, 0.0)

            def _clean_stage(s):
                return _re.sub(r'\s*\(\d+%\)', '', s or '').strip() or '—'

            def _fmt_d(raw):
                for fmt in ("%Y-%m-%d", "%m-%d-%Y"):
                    try:
                        return datetime.strptime(raw, fmt).strftime("%b %d, %Y")
                    except Exception:
                        pass
                return raw or '—'

            def _make_bg(bg_color):
                it = QtWidgets.QTableWidgetItem("")
                it.setBackground(QtGui.QBrush(bg_color))
                return it

            # ── Dialog ────────────────────────────────────────────────
            dlg = QtWidgets.QDialog(self)
            dlg.setWindowTitle(f"Payment History — {invoice_number or source}")
            dlg.setWindowFlags(
                dlg.windowFlags()
                | QtCore.Qt.WindowMaximizeButtonHint
                | QtCore.Qt.WindowMinimizeButtonHint)
            dlg.setMinimumWidth(800)
            dlg.setMinimumHeight(520)
            dlg.setStyleSheet("background:white;")

            root = QtWidgets.QVBoxLayout(dlg)
            root.setSpacing(10)
            root.setContentsMargins(14, 14, 14, 14)

            # ── Header grid ────────────────────────────────────────────
            hdr_frame = QtWidgets.QFrame()
            hdr_frame.setStyleSheet(
                "QFrame{background:#f8fafc;border:1px solid #e2e8f0;"
                "border-radius:8px;}")
            hg = QtWidgets.QGridLayout(hdr_frame)
            hg.setContentsMargins(16, 12, 16, 12)
            hg.setHorizontalSpacing(24)
            hg.setVerticalSpacing(6)

            def _lbl(t, bold=False, color="#374151", wrap=False):
                l = QtWidgets.QLabel(str(t))
                l.setStyleSheet(
                    f"font-weight:{'700' if bold else '400'};"
                    f"color:{color};border:none;font-size:12px;")
                if wrap:
                    l.setWordWrap(True)
                return l

            if len(inv_proj_names) == 0:
                proj_display = '—'
            elif len(inv_proj_names) <= 2:
                proj_display = ', '.join(inv_proj_names)
            else:
                proj_display = ', '.join(inv_proj_names[:2]) + f', +{len(inv_proj_names)-2} more…'

            _st_color = {"Paid": "#15803d", "Partially Paid": "#1e40af",
                         "Overdue": "#b91c1c"}.get(status, "#78350f")

            hg.addWidget(_lbl("Invoice #:", True),   0, 0)
            hg.addWidget(_lbl(invoice_number or '—'), 0, 1)
            hg.addWidget(_lbl("Client:", True),       0, 2)
            hg.addWidget(_lbl(inv_client or '—'),     0, 3)
            hg.addWidget(_lbl("Project:", True),      1, 0)
            pl = _lbl(proj_display, wrap=True)
            pl.setMaximumWidth(340)
            hg.addWidget(pl, 1, 1)
            hg.addWidget(_lbl("Invoice Date:", True), 1, 2)
            hg.addWidget(_lbl(inv_date or '—'),       1, 3)
            hg.addWidget(_lbl("Total Due:", True),    2, 0)
            hg.addWidget(_lbl(f"${inv_total:,.2f}"),  2, 1)
            hg.addWidget(_lbl("Status:", True),       2, 2)
            _sl = QtWidgets.QLabel(status or '—')
            _sl.setStyleSheet(
                f"font-weight:800;color:{_st_color};border:none;font-size:12px;")
            hg.addWidget(_sl, 2, 3)
            hg.setColumnStretch(1, 1)
            hg.setColumnStretch(3, 1)
            root.addWidget(hdr_frame)

            # ── Table ──────────────────────────────────────────────────
            hdr_lbl = QtWidgets.QLabel("Payment History by Project")
            hdr_lbl.setStyleSheet(
                "font-weight:700;font-size:13px;color:#0f172a;"
                "border:none;padding:4px 0 2px 0;")
            root.addWidget(hdr_lbl)

            COL = 6
            tbl = QtWidgets.QTableWidget()
            tbl.setColumnCount(COL)
            tbl.setHorizontalHeaderLabels(
                ["Project #", "Date", "Amount", "Method", "Stage", "Notes"])
            tbl.horizontalHeader().setVisible(True)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            tbl.setAlternatingRowColors(False)
            tbl.verticalHeader().setVisible(False)
            tbl.setStyleSheet("""
                QTableWidget{background:white;border:1px solid #e2e8f0;
                    border-radius:6px;gridline-color:#f1f5f9;}
                QTableWidget::item{padding:6px 10px;color:#1e293b;}
                QHeaderView::section{background:#f8fafc;font-weight:700;
                    padding:8px;border:none;
                    border-bottom:2px solid #e2e8f0;
                    min-height:34px;color:#374151;}
                QTableWidget::item:selected{background:#dbeafe;color:#1e40af;}
            """)
            tbl.setColumnWidth(0, 165)
            tbl.setColumnWidth(1, 115)
            tbl.setColumnWidth(2, 105)
            tbl.setColumnWidth(3, 115)
            tbl.setColumnWidth(4, 130)

            def _cell(t, al=QtCore.Qt.AlignCenter):
                it = QtWidgets.QTableWidgetItem(str(t))
                it.setTextAlignment(al)
                return it

            _HDR_BG    = QtGui.QColor("#1e3a5f")
            _HDR_FG    = QtGui.QColor("#ffffff")
            _SUB_BG    = QtGui.QColor("#f0f9ff")
            _SUB_FG    = QtGui.QColor("#0369a1")
            _REM_GRN   = QtGui.QColor("#15803d")
            _REM_RED   = QtGui.QColor("#b91c1c")
            _TAX_HDR_BG = QtGui.QColor("#0f5a52")

            # Build project order (invoice items first, then payment-only projects)
            pn_order, seen_pns = [], set()
            for _pn in planned_per_pn:
                if _pn and _pn not in seen_pns:
                    pn_order.append(_pn)
                    seen_pns.add(_pn)
            for _p in all_payments:
                _pn = (_p.project_number or '').strip()
                if _pn and _pn not in seen_pns:
                    pn_order.append(_pn)
                    seen_pns.add(_pn)

            rows_spec = []
            for pn in pn_order:
                pn_pays = sorted(
                    [p for p in all_payments
                     if (p.project_number or '').strip() == pn
                     and (p.payment_stage or '').strip().lower() != 'tax'],
                    key=lambda p: p.payment_date or '')
                pn_paid    = sum(float(p.amount) for p in pn_pays)
                pn_planned = planned_per_pn.get(pn, 0.0)
                rows_spec.append(("header", pn))
                for pay in pn_pays:
                    rows_spec.append(("pay", pay))
                rows_spec.append(("sub", (pn_paid, pn_planned)))

            if not all_payments and not pn_order:
                rows_spec = [("empty", None)]

            # TAX section — shown when invoice has a tax amount
            if _tax_amount > 0.005:
                tax_pays_sorted = sorted(inv_tax_pays, key=lambda t: t.payment_date or '')
                rows_spec.append(("tax_header", _tax_amount))
                if tax_pays_sorted:
                    for tp in tax_pays_sorted:
                        rows_spec.append(("tax_pay", tp))
                else:
                    rows_spec.append(("tax_pending", (invoice_number, _tax_amount)))
                rows_spec.append(("tax_sub", (tax_paid_total, _tax_amount)))

            tbl.setRowCount(len(rows_spec))
            for r, (kind, data) in enumerate(rows_spec):
                if kind == "header":
                    h = QtWidgets.QTableWidgetItem(f"  Project: {data}")
                    h.setBackground(QtGui.QBrush(_HDR_BG))
                    h.setForeground(QtGui.QBrush(_HDR_FG))
                    h.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                    h.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, h)
                    for c in range(1, COL):
                        tbl.setItem(r, c, _make_bg(_HDR_BG))
                    tbl.setSpan(r, 0, 1, COL)
                    tbl.setRowHeight(r, 30)

                elif kind == "pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.project_number or '—').strip()))
                    tbl.setItem(r, 1, _cell(_fmt_d(pay.payment_date or '')))
                    amt_it = _cell(f"${float(pay.amount):,.2f}")
                    amt_it.setForeground(QtGui.QColor("#15803d"))
                    amt_it.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))
                    tbl.setItem(r, 2, amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or '—'))
                    tbl.setItem(r, 4, _cell(_clean_stage(pay.payment_stage)))
                    tbl.setItem(r, 5, _cell(
                        pay.notes or '—',
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "sub":
                    pn_paid, pn_planned = data
                    pn_remaining = max(pn_planned - pn_paid, 0.0)
                    _sub_font = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)
                    # Left: Paid (span 3)
                    paid_cell = QtWidgets.QTableWidgetItem(f"  Paid: ${pn_paid:,.2f}")
                    paid_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    paid_cell.setForeground(QtGui.QBrush(_SUB_FG))
                    paid_cell.setFont(_sub_font)
                    paid_cell.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, paid_cell)
                    tbl.setItem(r, 1, _make_bg(_SUB_BG))
                    tbl.setItem(r, 2, _make_bg(_SUB_BG))
                    tbl.setSpan(r, 0, 1, 3)
                    # Right: Remaining (span 3)
                    _rem_color = _REM_GRN if pn_remaining <= 0 else _REM_RED
                    _rem_text  = "Fully Paid ✓" if pn_remaining <= 0 else f"Remaining: ${pn_remaining:,.2f}"
                    rem_cell = QtWidgets.QTableWidgetItem(f"{_rem_text}  ")
                    rem_cell.setBackground(QtGui.QBrush(_SUB_BG))
                    rem_cell.setForeground(QtGui.QBrush(_rem_color))
                    rem_cell.setFont(_sub_font)
                    rem_cell.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, rem_cell)
                    tbl.setItem(r, 4, _make_bg(_SUB_BG))
                    tbl.setItem(r, 5, _make_bg(_SUB_BG))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                elif kind == "tax_header":
                    _th = QtWidgets.QTableWidgetItem("  TAX")
                    _th.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    _th.setForeground(QtGui.QBrush(QtGui.QColor("#ffffff")))
                    _th.setFont(QtGui.QFont("Consolas", 9, QtGui.QFont.Bold))
                    _th.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, _th)
                    _th_note = QtWidgets.QTableWidgetItem(
                        f"Tax Amount: ${data:,.2f}  — Recorded when invoice is marked Paid  ")
                    _th_note.setBackground(QtGui.QBrush(_TAX_HDR_BG))
                    _th_note.setForeground(QtGui.QBrush(QtGui.QColor("#a7f3d0")))
                    _th_note.setFont(QtGui.QFont("Segoe UI", 8))
                    _th_note.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 1, _th_note)
                    for c in range(2, COL):
                        tbl.setItem(r, c, _make_bg(_TAX_HDR_BG))
                    tbl.setSpan(r, 1, 1, COL - 1)
                    tbl.setRowHeight(r, 30)

                elif kind == "tax_pay":
                    pay = data
                    tbl.setItem(r, 0, _cell((pay.invoice_number or '—').strip()))
                    tbl.setItem(r, 1, _cell(_fmt_d(pay.payment_date or '')))
                    _amt_it = _cell(f"${float(pay.amount):,.2f}")
                    _amt_it.setForeground(QtGui.QColor("#15803d"))
                    tbl.setItem(r, 2, _amt_it)
                    tbl.setItem(r, 3, _cell(pay.payment_method or '—'))
                    tbl.setItem(r, 4, _cell("Tax"))
                    tbl.setItem(r, 5, _cell(
                        pay.notes or '—',
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_pending":
                    inv_no2, tax_amt2 = data
                    tbl.setItem(r, 0, _cell(inv_no2))
                    tbl.setItem(r, 1, _cell('—'))
                    _pa = _cell(f"${tax_amt2:,.2f}")
                    _pa.setForeground(QtGui.QColor("#b45309"))
                    tbl.setItem(r, 2, _pa)
                    tbl.setItem(r, 3, _cell('—'))
                    _badge = QtWidgets.QTableWidgetItem("  Unpaid  ")
                    _badge.setBackground(QtGui.QBrush(QtGui.QColor("#fef3c7")))
                    _badge.setForeground(QtGui.QBrush(QtGui.QColor("#92400e")))
                    _badge.setTextAlignment(QtCore.Qt.AlignCenter)
                    tbl.setItem(r, 4, _badge)
                    tbl.setItem(r, 5, _cell(
                        "Pending — mark invoice as Paid to record",
                        QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter))
                    tbl.setRowHeight(r, 36)

                elif kind == "tax_sub":
                    _tp, _tpl = data
                    _tr = max(_tpl - _tp, 0.0)
                    _tsub_bg   = QtGui.QColor("#f0fdf4") if _tr <= 0 else QtGui.QColor("#fef9c3")
                    _tsub_fg   = QtGui.QColor("#15803d") if _tr <= 0 else QtGui.QColor("#92400e")
                    _tsub_font = QtGui.QFont("Segoe UI", 8, QtGui.QFont.Bold)
                    _tpc = QtWidgets.QTableWidgetItem(f"  Paid: ${_tp:,.2f}")
                    _tpc.setBackground(QtGui.QBrush(_tsub_bg))
                    _tpc.setForeground(QtGui.QBrush(_tsub_fg))
                    _tpc.setFont(_tsub_font)
                    _tpc.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 0, _tpc)
                    tbl.setItem(r, 1, _make_bg(_tsub_bg))
                    tbl.setItem(r, 2, _make_bg(_tsub_bg))
                    tbl.setSpan(r, 0, 1, 3)
                    _trt = "Tax Paid ✓" if _tr <= 0 else f"Remaining: ${_tr:,.2f}"
                    _trc = QtWidgets.QTableWidgetItem(f"{_trt}  ")
                    _trc.setBackground(QtGui.QBrush(_tsub_bg))
                    _trc.setForeground(QtGui.QBrush(_tsub_fg))
                    _trc.setFont(_tsub_font)
                    _trc.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                    tbl.setItem(r, 3, _trc)
                    tbl.setItem(r, 4, _make_bg(_tsub_bg))
                    tbl.setItem(r, 5, _make_bg(_tsub_bg))
                    tbl.setSpan(r, 3, 1, 3)
                    tbl.setRowHeight(r, 28)

                else:  # "empty"
                    no_it = QtWidgets.QTableWidgetItem(
                        "No payment records found for this invoice.")
                    no_it.setTextAlignment(QtCore.Qt.AlignCenter)
                    no_it.setForeground(QtGui.QColor('#9ca3af'))
                    tbl.setItem(r, 0, no_it)
                    tbl.setSpan(r, 0, 1, COL)

            root.addWidget(tbl)

            # ── Summary bar ────────────────────────────────────────────
            sf = QtWidgets.QFrame()
            sf.setFrameShape(QtWidgets.QFrame.NoFrame)
            sf.setStyleSheet(
                "QFrame{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;}"
                if remaining <= 0 else
                "QFrame{background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;}")
            sl2 = QtWidgets.QHBoxLayout(sf)
            sl2.setContentsMargins(14, 8, 14, 8)

            def _sum_col(label, value, color="#1e293b"):
                col = QtWidgets.QVBoxLayout()
                l1 = QtWidgets.QLabel(label)
                l1.setStyleSheet("font-size:11px;color:#64748b;border:none;")
                l2 = QtWidgets.QLabel(value)
                l2.setStyleSheet(
                    f"font-size:14px;font-weight:800;color:{color};border:none;")
                col.addWidget(l1)
                col.addWidget(l2)
                return col

            sl2.addLayout(_sum_col("Invoice Total", f"${inv_total:,.2f}"))
            _sep1 = QtWidgets.QLabel("|")
            _sep1.setStyleSheet("color:#cbd5e1;font-size:18px;border:none;")
            sl2.addWidget(_sep1)
            sl2.addLayout(_sum_col("Total Paid", f"${total_paid:,.2f}", "#15803d"))
            _sep2 = QtWidgets.QLabel("|")
            _sep2.setStyleSheet("color:#cbd5e1;font-size:18px;border:none;")
            sl2.addWidget(_sep2)
            _rem_color = "#15803d" if remaining <= 0 else "#b45309"
            sl2.addLayout(_sum_col("Remaining", f"${remaining:,.2f}", _rem_color))
            sl2.addStretch()
            root.addWidget(sf)

            # ── Close button ───────────────────────────────────────────
            _close_btn = QtWidgets.QPushButton("Close")
            _close_btn.setFixedHeight(34)
            _close_btn.setStyleSheet("""
                QPushButton{background:#334155;color:white;border:none;
                    border-radius:6px;font-weight:bold;padding:0 22px;}
                QPushButton:hover{background:#1e293b;}
            """)
            _close_btn.clicked.connect(dlg.accept)
            _btn_row = QtWidgets.QHBoxLayout()
            _btn_row.addStretch()
            _btn_row.addWidget(_close_btn)
            root.addLayout(_btn_row)
            dlg.exec_()

        # ── PDF export ──────────────────────────────────────────────────

        def _export_pdf(self):
            try:
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import (
                    SimpleDocTemplate, Table, TableStyle,
                    Paragraph, Spacer)
                from reportlab.lib.styles import (
                    getSampleStyleSheet, ParagraphStyle)
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from pathlib import Path as _Path
                import calendar as _cal

                # For Annual mode: show date-range picker before export
                export_from_py = None
                export_to_py = None
                if self.mode == 'year':
                    dlg = QtWidgets.QDialog(self)
                    dlg.setWindowTitle("Export Date Range")
                    dlg.setFixedSize(340, 230)
                    dlg.setStyleSheet(
                        "QDialog{background:#f8fafc;}"
                        "QLabel{font-size:13px;font-weight:600;color:#1e293b;"
                        "border:none;background:transparent;}")
                    dlg_lay = QtWidgets.QVBoxLayout(dlg)
                    dlg_lay.setContentsMargins(24, 20, 24, 20)
                    dlg_lay.setSpacing(12)

                    hdr_lbl = QtWidgets.QLabel("Select Export Date Range")
                    hdr_lbl.setStyleSheet(
                        "font-size:15px;font-weight:800;color:#0f172a;"
                        "border:none;background:transparent;")
                    dlg_lay.addWidget(hdr_lbl)

                    _de_style = (
                        "QDateEdit{padding:6px 28px 6px 8px;"
                        "border:2px solid #bdc3c7;border-radius:6px;"
                        "background:white;font-size:13px;font-weight:600;}"
                        "QDateEdit:focus{border-color:#00756f;}")
                    _lbl_ss = ("font-size:12px;font-weight:700;color:#374151;"
                               "border:none;background:transparent;")

                    # From row
                    from_row = QtWidgets.QHBoxLayout()
                    from_row.setSpacing(10)
                    from_lbl = QtWidgets.QLabel("From:")
                    from_lbl.setStyleSheet(_lbl_ss)
                    from_lbl.setFixedWidth(42)
                    _fd = QtWidgets.QDateEdit()
                    _fd.setCalendarPopup(True)
                    _fd.setDisplayFormat("MM-dd-yyyy")
                    _fd.setFixedHeight(36)
                    _fd.wheelEvent = lambda e: e.ignore()
                    _fd.stepBy = lambda x: None
                    _fd.setDate(QtCore.QDate(self._year, 1, 1))
                    _fd.setStyleSheet(_de_style)
                    from_row.addWidget(from_lbl)
                    from_row.addWidget(_fd, 1)
                    dlg_lay.addLayout(from_row)

                    # To row
                    to_row = QtWidgets.QHBoxLayout()
                    to_row.setSpacing(10)
                    to_lbl = QtWidgets.QLabel("To:")
                    to_lbl.setStyleSheet(_lbl_ss)
                    to_lbl.setFixedWidth(42)
                    _td = QtWidgets.QDateEdit()
                    _td.setCalendarPopup(True)
                    _td.setDisplayFormat("MM-dd-yyyy")
                    _td.setFixedHeight(36)
                    _td.wheelEvent = lambda e: e.ignore()
                    _td.stepBy = lambda x: None
                    _td.setDate(QtCore.QDate(self._year, 12, 31))
                    _td.setStyleSheet(_de_style)
                    to_row.addWidget(to_lbl)
                    to_row.addWidget(_td, 1)
                    dlg_lay.addLayout(to_row)

                    dlg_lay.addStretch()

                    btn_row = QtWidgets.QHBoxLayout()
                    btn_row.addStretch()
                    cancel_b = QtWidgets.QPushButton("Cancel")
                    cancel_b.setFixedSize(90, 36)
                    cancel_b.setStyleSheet(
                        "QPushButton{background:#f1f5f9;color:#334155;"
                        "border:1px solid #cbd5e1;border-radius:7px;"
                        "font-size:13px;font-weight:700;}"
                        "QPushButton:hover{background:#e2e8f0;}")
                    cancel_b.clicked.connect(dlg.reject)
                    export_b = QtWidgets.QPushButton("Export PDF")
                    export_b.setFixedSize(110, 36)
                    export_b.setStyleSheet(
                        "QPushButton{background:#0f766e;color:white;border:none;"
                        "border-radius:7px;font-size:13px;font-weight:800;}"
                        "QPushButton:hover{background:#0d625c;}")
                    export_b.clicked.connect(dlg.accept)
                    btn_row.addWidget(cancel_b)
                    btn_row.addSpacing(8)
                    btn_row.addWidget(export_b)
                    dlg_lay.addLayout(btn_row)

                    if dlg.exec_() != QtWidgets.QDialog.Accepted:
                        return
                    export_from_py = _fd.date().toPyDate()
                    export_to_py = _td.date().toPyDate()

                search = self._search_text.lower().strip()
                entries = self._all_entries

                # Apply date filter: custom export range (annual) or existing widget range
                if export_from_py and export_to_py:
                    entries = [e for e in entries
                               if export_from_py <= e['paid_dt'].date() <= export_to_py]
                elif self.mode == 'year' and hasattr(self, '_from_de'):
                    from_py = self._from_de.date().toPyDate()
                    to_py = self._to_de.date().toPyDate()
                    entries = [e for e in entries
                               if from_py <= e['paid_dt'].date() <= to_py]
                if search:
                    entries = [
                        e for e in entries
                        if (search in e['source'].lower()
                            or search in e['description'].lower()
                            or search in e['paid_date'].lower()
                            or search in e['invoice_date'].lower()
                            or search in e.get('invoice_number', '').lower()
                            or search in f"{e['paid_amount']:.2f}")
                    ]

                if not entries:
                    QtWidgets.QMessageBox.information(
                        self, "No Data", "No entries to export.")
                    return

                exp_dir = _Path.home() / "Downloads" / "Balance_Exports"
                exp_dir.mkdir(parents=True, exist_ok=True)
                if self.mode == 'month':
                    label = (f"{_cal.month_name[self._month]}"
                             f"_{self._year}")
                    period_str = (f"{_cal.month_name[self._month]} "
                                  f"{self._year}")
                else:
                    label = f"Annual_{self._year}"
                    if export_from_py and export_to_py:
                        period_str = (
                            f"{export_from_py.strftime('%m-%d-%Y')}"
                            f" — "
                            f"{export_to_py.strftime('%m-%d-%Y')}")
                    elif hasattr(self, '_from_de'):
                        period_str = (
                            f"{self._from_de.date().toString('MM-dd-yyyy')}"
                            f" — "
                            f"{self._to_de.date().toString('MM-dd-yyyy')}")
                    else:
                        period_str = f"Annual {self._year}"
                filepath = exp_dir / f"PaidRevenue_{label}.pdf"

                doc = SimpleDocTemplate(
                    str(filepath), pagesize=landscape(A4),
                    topMargin=0.5*inch, bottomMargin=0.5*inch,
                    leftMargin=0.5*inch, rightMargin=0.5*inch)

                styles = getSampleStyleSheet()
                title_s = ParagraphStyle(
                    'RDTitle', parent=styles['Heading1'],
                    fontSize=14, textColor=colors.white, alignment=1,
                    backColor=colors.HexColor('#1e3a5f'),
                    borderPadding=8, spaceAfter=2)
                # Left-aligned wrapping text for regular cells
                cell_s = ParagraphStyle(
                    'RDCell', parent=styles['Normal'],
                    fontSize=8, leading=10)
                # Centered wrapping text — used for Description column
                cell_c_s = ParagraphStyle(
                    'RDCellC', parent=styles['Normal'],
                    fontSize=8, leading=10, alignment=1)

                # Equal column widths across the full usable page width
                # Landscape A4 − 1 inch margins ≈ 10.69 in
                if self.mode == 'month':
                    col_hdr = ["S.No", "Invoice Date", "Invoice #",
                               "Description", "Invoice Total",
                               "Paid Amount", "Paid Date"]
                    _col_w_each = 10.5 / len(col_hdr)
                    col_w = [_col_w_each * inch] * len(col_hdr)
                else:
                    col_hdr = ["S.No", "Paid Month", "Invoice Date",
                               "Invoice #", "Description",
                               "Invoice Total", "Paid Amount", "Paid Date"]
                    _col_w_each = 10.56 / len(col_hdr)
                    col_w = [_col_w_each * inch] * len(col_hdr)

                try:
                    from main import Config as _Cfg
                    _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
                except Exception:
                    _co = 'MABS ENGINEERING SERVICES'
                elems = [
                    Paragraph(
                        f"{_co}  —  PAID REVENUE"
                        f"  ({period_str})", title_s),
                    Spacer(1, 0.12*inch),
                ]

                tbl_data = [col_hdr]
                row_styles = []
                for i, e in enumerate(entries, 1):
                    bg = (colors.HexColor('#f0fdf4')
                          if i % 2 == 0 else colors.white)
                    row_styles.append(
                        ('BACKGROUND', (0, i), (-1, i), bg))
                    desc_p = (Paragraph(e['description'], cell_c_s)
                              if e['description'] else '')
                    _inv_no = e.get('invoice_number', '') or e['source']
                    if self.mode == 'month':
                        tbl_data.append([
                            str(i), e['invoice_date'], _inv_no,
                            desc_p,
                            f"${e['invoice_total']:,.2f}",
                            f"${e['paid_amount']:,.2f}",
                            e['paid_date'],
                        ])
                    else:
                        tbl_data.append([
                            str(i),
                            f"{_cal.month_abbr[e['paid_dt'].month]} "
                            f"{e['paid_dt'].year}",
                            e['invoice_date'], _inv_no, desc_p,
                            f"${e['invoice_total']:,.2f}",
                            f"${e['paid_amount']:,.2f}",
                            e['paid_date'],
                        ])

                total_paid = sum(e['paid_amount'] for e in entries)
                total_row_idx = len(tbl_data)
                blank = [""] * len(col_hdr)
                total_row = blank[:]
                total_row[-3] = "TOTAL PAID"
                total_row[-2] = f"${total_paid:,.2f}"
                tbl_data.append(total_row)

                base_style = [
                    ('BACKGROUND', (0, 0), (-1, 0),
                     colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    # All cells centered — description uses cell_c_s Paragraph
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.4,
                     colors.HexColor('#e2e8f0')),
                    ('BACKGROUND', (0, total_row_idx), (-1, total_row_idx),
                     colors.HexColor('#d1fae5')),
                    ('FONTNAME', (0, total_row_idx), (-1, total_row_idx),
                     'Helvetica-Bold'),
                    ('ALIGN', (0, total_row_idx), (-1, total_row_idx),
                     'CENTER'),
                    ('LINEABOVE', (0, total_row_idx), (-1, total_row_idx),
                     1, colors.HexColor('#15803d')),
                ] + row_styles

                tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
                tbl.setStyle(TableStyle(base_style))
                elems.append(tbl)
                doc.build(elems)

                import platform, subprocess, os
                if platform.system() == "Windows":
                    os.startfile(filepath)
                elif platform.system() == "Darwin":
                    subprocess.run(["open", str(filepath)])
                else:
                    subprocess.run(["xdg-open", str(filepath)])

                QtWidgets.QMessageBox.information(
                    self, "Export Complete",
                    f"PDF saved!\n\nFile: {filepath.name}"
                    f"\nFolder: {exp_dir}")

            except ImportError:
                QtWidgets.QMessageBox.critical(
                    self, "Missing Library",
                    "reportlab is required.\n"
                    "Install: pip install reportlab")
            except Exception as exc:
                QtWidgets.QMessageBox.critical(
                    self, "Export Error", str(exc))
                _log.warning("RevenueDetailDialog PDF error: %s", exc)

    class YearPickerDialog(QtWidgets.QDialog):
        """Year picker dialog with 3x4 grid"""
        def __init__(self, parent=None, current_year=None):
            super().__init__(parent)
            self.setWindowTitle("Select Year")
            self.resize(560, 430)
            self.current_year = current_year or datetime.now().year
            self.start_year = (self.current_year // 12) * 12  # decade block start
            self.selected_year = self.current_year

            layout = QtWidgets.QVBoxLayout(self)
            layout.setContentsMargins(24, 24, 24, 24)
            layout.setSpacing(18)

            # Header with arrows
            header = QtWidgets.QHBoxLayout()

            self.prev_btn = QtWidgets.QPushButton("←")
            self.prev_btn.setFixedSize(58, 42)
            self.prev_btn.clicked.connect(self.prev_range)

            self.next_btn = QtWidgets.QPushButton("→")
            self.next_btn.setFixedSize(58, 42)
            self.next_btn.clicked.connect(self.next_range)

            self.title_lbl = QtWidgets.QLabel("", alignment=QtCore.Qt.AlignCenter)
            self.title_lbl.setStyleSheet("font-size: 24px; font-weight: bold;")

            header.addWidget(self.prev_btn)
            header.addWidget(self.title_lbl, 1)
            header.addWidget(self.next_btn)

            layout.addLayout(header)

            # Year Grid (3x4)
            self.grid_widget = QtWidgets.QWidget()
            self.grid = QtWidgets.QGridLayout(self.grid_widget)
            self.grid.setSpacing(16)
            layout.addWidget(self.grid_widget)

            # Build first view
            self.build_grid()

        def build_grid(self):
            # Clear previous buttons
            for i in reversed(range(self.grid.count())):
                w = self.grid.itemAt(i).widget()
                if w:
                    w.deleteLater()

            # Generate 12 continuous years
            years = list(range(self.start_year, self.start_year + 12))
            self.title_lbl.setText(f"{self.start_year} - {self.start_year + 11}")

            r, c = 0, 0
            for year in years:
                btn = QtWidgets.QPushButton(str(year))
                btn.setFixedSize(150, 72)
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background: {"#d1f5e0" if year == self.current_year else "#ffffff"};
                        border: 1px solid #cccccc;
                        border-radius: 10px;
                        font-size: 22px;
                        color: #333333;
                        font-weight: {"bold" if year == self.current_year else "500"};
                    }}
                    QPushButton:hover {{
                        background: #eef7ff;
                        border: 1px solid #7ab8ff;
                    }}
                """)
                btn.clicked.connect(lambda _, y=year: self.select_year(y))

                self.grid.addWidget(btn, r, c)

                c += 1
                if c == 3:
                    c = 0
                    r += 1

        def prev_range(self):
            self.start_year -= 12
            self.build_grid()

        def next_range(self):
            self.start_year += 12
            self.build_grid()

        def select_year(self, year):
            self.selected_year = year
            self.accept()
            
    def load_annual_summary_local_for_year(self, year):
        """Load annual summary data from local JSON files for a specific year"""
        try:
            data_dir = Path.home() / ".mabs_finance"
            
            # Load expenses for annual summary
            expense_file = data_dir / f"expenses_{year}.json"
            if expense_file.exists():
                with open(expense_file, 'r') as f:
                    self.annual_expenses_data = json.load(f)
            else:
                self.annual_expenses_data = []
                
            # Load revenue for annual summary
            revenue_file = data_dir / f"revenue_{year}.json"
            if revenue_file.exists():
                with open(revenue_file, 'r') as f:
                    self.annual_revenue_data = json.load(f)
            else:
                self.annual_revenue_data = []
                
            # Load salary for annual summary
            salary_file = data_dir / f"salary_{year}.json"
            if salary_file.exists():
                with open(salary_file, 'r') as f:
                    self.annual_salary_data = json.load(f)
            else:
                self.annual_salary_data = {"Inside America": [], "Outside America": []}
                
        except Exception as e:
            _log.warning("Error loading annual summary local data: %s", e)
            self.annual_expenses_data = []
            self.annual_revenue_data = []
            self.annual_salary_data = {"Inside America": [], "Outside America": []}
               
    def show_annual_summary_year_calendar(self):
        """Show year selection popup - ONLY updates annual summary and breakdowns, NOT transaction table or header year"""
        dialog = self.YearPickerDialog(self, self.annual_summary_year)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            # Store the selected year separately for annual summary
            self.annual_summary_year = dialog.selected_year
            if hasattr(self, 'yearly_calendar_btn'):
                self.yearly_calendar_btn.setText(str(self.annual_summary_year))
            # Update ONLY the annual summary title, NOT the header year
            self.annual_title.setText(f"📈 ANNUAL FINANCIAL SUMMARY - {self.annual_summary_year}")
            
            # Store current category (but don't reload transaction data)
            current_category = self.current_category
            
            # Load ONLY annual summary data for the selected year
            # We need to pass the selected year to the load method
            self.load_annual_summary_data_for_year(self.annual_summary_year)
            
            # Update the annual summary table
            self.update_annual_summary()
            
        
    def show_year_calendar(self):
        """Show year selection popup with 3x4 grid - updates everything for the new year"""
        dialog = self.YearPickerDialog(self, self.current_year)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.current_year = dialog.selected_year
            self.annual_summary_year = self.current_year
            if hasattr(self, 'yearly_calendar_btn'):
                self.yearly_calendar_btn.setText(str(self.annual_summary_year))
            
            self.year_btn.setText(f"{self.current_year} 📅")
            self.annual_title.setText(f"📈 ANNUAL FINANCIAL SUMMARY - {self.annual_summary_year}")
            
            # Store current category before reloading
            current_category = self.current_category
            
            # Load ALL financial data for the new year (both transaction and annual summary)
            self.load_all_financial_data()
            
            # Update the annual summary table
            self.update_annual_summary()
            
            # Ensure the transaction table shows the correct category with new year's data
            self.on_category_changed(current_category)
            
    def load_annual_summary_data(self):
        """Load ONLY annual summary data without affecting transaction table"""
        if not self.FIREBASE_AVAILABLE or self.db is None:
            _log.warning("Firebase not available - using local data for annual summary")
            self.load_annual_summary_local()
            return
            
        try:
            _log.info("Loading annual summary data for year %s from balance_sheet_expenses...", self.current_year)
            
            # Load from BALANCE SHEET node
            all_expenses = BalanceSheetFirebaseManager.load_expenses()
            self.annual_expenses_data = [exp for exp in all_expenses if exp.get('year') == self.current_year]
            _log.info("Loaded %s expenses for annual summary", len(self.annual_expenses_data))
            
            all_revenue = BalanceSheetFirebaseManager.load_revenue()
            self.annual_revenue_data = [rev for rev in all_revenue if rev.get('year') == self.current_year]
            _log.info("Loaded %s revenue for annual summary", len(self.annual_revenue_data))
            
            all_salary = BalanceSheetFirebaseManager.load_salary()
            self.annual_salary_data = {"Inside America": [], "Outside America": []}
            for region in ["Inside America", "Outside America"]:
                self.annual_salary_data[region] = [sal for sal in all_salary[region] if sal.get('year') == self.current_year]
            _log.info("Loaded %s salary entries for annual summary", sum(len(v) for v in self.annual_salary_data.values()))
            
        except Exception as e:
            _log.warning("Error loading annual summary from Firebase: %s", e)
            self.load_annual_summary_local()
        
        # Update the annual summary with the loaded data
        self.update_annual_summary()

    def load_annual_summary_local(self):
        """Load annual summary data from local JSON files"""
        try:
            data_dir = Path.home() / ".mabs_finance"
            
            # Load expenses for annual summary
            expense_file = data_dir / f"expenses_{self.current_year}.json"
            if expense_file.exists():
                with open(expense_file, 'r') as f:
                    self.annual_expenses_data = json.load(f)
            else:
                self.annual_expenses_data = []
                
            # Load revenue for annual summary
            revenue_file = data_dir / f"revenue_{self.current_year}.json"
            if revenue_file.exists():
                with open(revenue_file, 'r') as f:
                    self.annual_revenue_data = json.load(f)
            else:
                self.annual_revenue_data = []
                
            # Load salary for annual summary
            salary_file = data_dir / f"salary_{self.current_year}.json"
            if salary_file.exists():
                with open(salary_file, 'r') as f:
                    self.annual_salary_data = json.load(f)
            else:
                self.annual_salary_data = {"Inside America": [], "Outside America": []}
                
        except Exception as e:
            _log.warning("Error loading annual summary local data: %s", e)
            self.annual_expenses_data = []
            self.annual_revenue_data = []
            self.annual_salary_data = {"Inside America": [], "Outside America": []}
            

    def update_year_display(self):
        """Update the display for the selected year"""
        self.year_btn.setText(f"{self.current_year} 📅")


class _WheelBlocker(QtCore.QObject):
    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Wheel:
            return True
        return False

class _NoScrollComboBox(QtWidgets.QComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._wb = _WheelBlocker(self)
        self.installEventFilter(self._wb)

class _NoScrollDateEdit(QtWidgets.QDateEdit):
    def stepBy(self, steps):
        pass


class ExportDialog(QtWidgets.QDialog):
    """Export options dialog for Balance Sheet with PDF and Excel support"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_tab = parent
        self.setWindowTitle("📤 Export Options")
        self.setModal(True)
        self.setFixedSize(700, 750)

        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header
        header = QtWidgets.QLabel("📤 Export Manager - Balance Sheet")
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: white;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                border-radius: 8px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)

        # Export format selection (Professional layout)

        # Export Type Tabs
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 10px 26px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # Annual Financial Summary Tab
        self.annual_tab = QtWidgets.QWidget()
        self.setup_annual_tab()
        self.tab_widget.addTab(self.annual_tab, "📊 Annual Financial Summary")

        # Financial Overview Export Tab
        self.combined_tab = QtWidgets.QWidget()
        self.setup_combined_tab()
        self.tab_widget.addTab(self.combined_tab, "📋 Financial Activity Overview")

        layout.addWidget(self.tab_widget)

        # Preview Label
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
    
        self.export_btn = QtWidgets.QPushButton("🚀 Export")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.perform_export)

        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)

        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)

        layout.addLayout(button_layout)

        # Connect tab change to update preview
        self.tab_widget.currentChanged.connect(self.update_preview)
        
        # Initial preview update
        self.update_preview()
        
    def setup_annual_tab(self):
        """Setup Annual Financial Summary tab"""
        layout = QtWidgets.QVBoxLayout(self.annual_tab)
        layout.setSpacing(15)
        
        # Year selection
        year_frame = QtWidgets.QFrame()
        year_frame.setFixedHeight(150)
        year_frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 2px solid #3498db;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        year_layout = QtWidgets.QHBoxLayout(year_frame)
        
        year_label = QtWidgets.QLabel("Select Year:")
        year_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        
        self.year_edit = QtWidgets.QLineEdit(str(self.parent_tab.current_year))
        self.year_edit.setFixedSize(120, 40)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                padding: 5px;
                font-size: 14px;
                font-weight: bold;
                background: #f8f9fa;
            }
        """)
        
        self.year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn.setFixedHeight(40)
        self.year_calendar_btn.setMinimumWidth(72)
        self.year_calendar_btn.clicked.connect(self.show_year_picker)
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #2471a3;
            }
        """)
        
        year_layout.addWidget(year_label)
        year_layout.addWidget(self.year_edit)
        year_layout.addWidget(self.year_calendar_btn)
        year_layout.addStretch()
        
        layout.addWidget(year_frame)
        
        # Export format selection - USE DIFFERENT VARIABLE NAMES
        format_layout = QtWidgets.QHBoxLayout()

        format_label = QtWidgets.QLabel("Export Format:")
        format_label.setStyleSheet("font-size:13px;font-weight:bold;color:#2c3e50;")

        self.annual_excel_radio = QtWidgets.QRadioButton("Excel (.xlsx)")
        self.annual_excel_radio.setChecked(True)

        self.annual_pdf_radio = QtWidgets.QRadioButton("PDF")
        # Make them mutually exclusive
        annual_format_group = QtWidgets.QButtonGroup(self)
        annual_format_group.addButton(self.annual_excel_radio)
        annual_format_group.addButton(self.annual_pdf_radio)
        format_layout.addWidget(format_label)
        format_layout.addSpacing(15)
        format_layout.addWidget(self.annual_excel_radio)
        format_layout.addWidget(self.annual_pdf_radio)
        format_layout.addStretch()

        layout.addLayout(format_layout)

    def setup_combined_tab(self):
        """Setup Financial Overview Export tab"""
        layout = QtWidgets.QVBoxLayout(self.combined_tab)
        layout.setSpacing(15)
        
        # Date range selection
        range_frame = QtWidgets.QFrame()
        range_frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 2px solid #e67e22;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        range_layout = QtWidgets.QVBoxLayout(range_frame)
        
        # Date range radio buttons
        self.all_radio = QtWidgets.QRadioButton("All Data")
        self.all_radio.setChecked(True)
        self.date_range_radio = QtWidgets.QRadioButton("Date Range")
        self.month_radio = QtWidgets.QRadioButton("Month")
        self.year_radio = QtWidgets.QRadioButton("Year")
        
        for radio in [self.all_radio, self.date_range_radio, self.month_radio, self.year_radio]:
            radio.setStyleSheet("font-size: 13px; font-weight: bold; color: #2c3e50;")
            radio.toggled.connect(self.update_preview)
        
        range_layout.addWidget(self.all_radio)
        range_layout.addWidget(self.date_range_radio)
        range_layout.addWidget(self.month_radio)
        range_layout.addWidget(self.year_radio)
        
        # Date range inputs
        self.date_range_widget = QtWidgets.QWidget()
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_widget)
        date_range_layout.setContentsMargins(20, 10, 0, 0)
        date_range_layout.setSpacing(24)

        _date_style = """
            QDateEdit {
                padding: 10px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover { border-color: #3498db; }
        """
        _label_style = "font-weight: bold; color: #2c3e50; font-size: 13px; border: none; background: transparent;"

        from_col = QtWidgets.QVBoxLayout()
        from_col.setSpacing(6)
        from_lbl = QtWidgets.QLabel("From Date:")
        from_lbl.setStyleSheet(_label_style)
        from_col.addWidget(from_lbl)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setFixedSize(160, 45)
        self.from_date.setStyleSheet(_date_style)
        self.from_date.dateChanged.connect(self.update_preview)
        from_col.addWidget(self.from_date)
        date_range_layout.addLayout(from_col)

        to_col = QtWidgets.QVBoxLayout()
        to_col.setSpacing(6)
        to_lbl = QtWidgets.QLabel("To Date:")
        to_lbl.setStyleSheet(_label_style)
        to_col.addWidget(to_lbl)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setFixedSize(160, 45)
        self.to_date.setStyleSheet(_date_style)
        self.to_date.dateChanged.connect(self.update_preview)
        to_col.addWidget(self.to_date)
        date_range_layout.addLayout(to_col)
        date_range_layout.addStretch()

        range_layout.addWidget(self.date_range_widget)

        # Month selection
        self.month_widget = QtWidgets.QWidget()
        month_layout = QtWidgets.QHBoxLayout(self.month_widget)
        month_layout.setContentsMargins(20, 10, 0, 0)
        month_layout.setSpacing(24)

        month_col = QtWidgets.QVBoxLayout()
        month_col.setSpacing(6)
        month_lbl = QtWidgets.QLabel("Month")
        month_lbl.setStyleSheet(_label_style)
        month_col.addWidget(month_lbl)
        self.month_combo = _NoScrollComboBox()
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)
        self.month_combo.setFixedHeight(42)
        self.month_combo.setMinimumWidth(160)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover { border-color: #3498db; }
        """)
        self.month_combo.currentIndexChanged.connect(self.update_preview)
        month_col.addWidget(self.month_combo)
        month_layout.addLayout(month_col)

        month_year_col = QtWidgets.QVBoxLayout()
        month_year_col.setSpacing(6)
        month_year_lbl = QtWidgets.QLabel("Year")
        month_year_lbl.setStyleSheet(_label_style)
        month_year_col.addWidget(month_year_lbl)
        month_year_field_row = QtWidgets.QHBoxLayout()
        month_year_field_row.setSpacing(6)
        self.month_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.month_year_edit.setFixedSize(120, 45)
        self.month_year_edit.setReadOnly(True)
        self.month_year_edit.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        self.month_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.month_calendar_btn.setFixedHeight(45)
        self.month_calendar_btn.setMinimumWidth(72)
        self.month_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.month_calendar_btn.clicked.connect(lambda: self.show_year_picker_for_field(self.month_year_edit))
        self.month_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover { background: #2980b9; border-color: #21618c; }
            QPushButton:pressed { background: #21618c; }
        """)
        month_year_field_row.addWidget(self.month_year_edit)
        month_year_field_row.addWidget(self.month_calendar_btn)
        month_year_col.addLayout(month_year_field_row)
        month_layout.addLayout(month_year_col)
        month_layout.addStretch()

        range_layout.addWidget(self.month_widget)

        # Year selection
        self.year_widget = QtWidgets.QWidget()
        year_layout2 = QtWidgets.QHBoxLayout(self.year_widget)
        year_layout2.setContentsMargins(20, 10, 0, 0)
        year_layout2.setSpacing(10)

        year_lbl2 = QtWidgets.QLabel("Year:")
        year_lbl2.setStyleSheet(_label_style)
        year_layout2.addWidget(year_lbl2)

        self.year_edit2 = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit2.setFixedSize(120, 45)
        self.year_edit2.setReadOnly(True)
        self.year_edit2.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.year_calendar_btn2 = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn2.setFixedHeight(45)
        self.year_calendar_btn2.setMinimumWidth(72)
        self.year_calendar_btn2.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn2.clicked.connect(lambda: self.show_year_picker_for_field(self.year_edit2))
        self.year_calendar_btn2.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover { background: #2980b9; border-color: #21618c; }
            QPushButton:pressed { background: #21618c; }
        """)

        year_layout2.addWidget(self.year_edit2)
        year_layout2.addWidget(self.year_calendar_btn2)
        year_layout2.addStretch()

        range_layout.addWidget(self.year_widget)
        
        layout.addWidget(range_frame)
        
        # Category selection checkboxes
        categories_frame = QtWidgets.QFrame()
        categories_frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 2px solid #3498db;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        categories_layout = QtWidgets.QVBoxLayout(categories_frame)
        
        categories_label = QtWidgets.QLabel("Include in Export:")
        categories_label.setStyleSheet("""
        QLabel {
        font-size: 14px;
        font-weight: bold;
        color: #2c3e50;
        border: none;
        background: transparent;
        padding: 0px;
        }
        """)

        self.include_revenue = QtWidgets.QCheckBox("Revenue")
        self.include_expenses = QtWidgets.QCheckBox("Expenses")
        self.include_salary = QtWidgets.QCheckBox("Salaries")
        
        self.include_revenue.setChecked(True)
        self.include_expenses.setChecked(True)
        self.include_salary.setChecked(True)
        
        for cb in [self.include_revenue, self.include_expenses, self.include_salary]:
            cb.setStyleSheet("font-size: 13px; color: #2c3e50;")
            cb.stateChanged.connect(self.update_preview)
        
        categories_layout.addWidget(categories_label)
        categories_layout.addWidget(self.include_revenue)
        categories_layout.addWidget(self.include_expenses)
        categories_layout.addWidget(self.include_salary)
        
        layout.addWidget(categories_frame)
        # Export format selection - USE DIFFERENT VARIABLE NAMES
        format_layout = QtWidgets.QHBoxLayout()

        format_label = QtWidgets.QLabel("Export Format:")
        format_label.setStyleSheet("font-size:13px;font-weight:bold;color:#2c3e50;")

        self.combined_excel_radio = QtWidgets.QRadioButton("Excel (.xlsx)")
        self.combined_excel_radio.setChecked(True)

        self.combined_pdf_radio = QtWidgets.QRadioButton("PDF")

        combined_format_group = QtWidgets.QButtonGroup(self)
        combined_format_group.addButton(self.combined_excel_radio)
        combined_format_group.addButton(self.combined_pdf_radio)

        format_layout.addWidget(format_label)
        format_layout.addSpacing(15)
        format_layout.addWidget(self.combined_excel_radio)
        format_layout.addWidget(self.combined_pdf_radio)
        format_layout.addStretch()

        layout.addLayout(format_layout)
        # Initial visibility
        self.update_date_widgets_visibility()
    
    def update_date_widgets_visibility(self):
        """Update visibility of date range widgets"""
        self.date_range_widget.setVisible(self.date_range_radio.isChecked())
        self.month_widget.setVisible(self.month_radio.isChecked())
        self.year_widget.setVisible(self.year_radio.isChecked())
    
    def show_year_picker(self):
        """Show year picker for Annual tab"""
        dialog = self.parent_tab.YearPickerDialog(self, int(self.year_edit.text()))
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            self.year_edit.setText(str(dialog.selected_year))
            self.update_preview()
    
    def show_year_picker_for_field(self, field):
        """Show year picker for a specific field"""
        dialog = self.parent_tab.YearPickerDialog(self, int(field.text()))
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            field.setText(str(dialog.selected_year))
            self.update_preview()
    
    def update_preview(self):
        """Update preview text"""
        self.update_date_widgets_visibility()

        if self.tab_widget.currentIndex() == 0:
            # Annual tab
            date_str = datetime.now().strftime("%m-%d-%Y")
        else:
            # Balance Sheet tab
            if self.all_radio.isChecked():
                range_text = "All Data"
            elif self.date_range_radio.isChecked():
                range_text = f"{self.from_date.date().toString('MM-dd-yyyy')} to {self.to_date.date().toString('MM-dd-yyyy')}"
            elif self.month_radio.isChecked():
                range_text = f"{self.month_combo.currentText()} {self.month_year_edit.text()}"
            else:
                range_text = f"Year {self.year_edit2.text()}"

            categories = []
            if self.include_revenue.isChecked():
                categories.append("Revenue")
            if self.include_expenses.isChecked():
                categories.append("Expenses")
            if self.include_salary.isChecked():
                categories.append("Salaries")

            date_str = datetime.now().strftime("%m-%d-%Y")
    
    def perform_export(self):
        """Perform export based on selected options"""
        try:
            if self.tab_widget.currentIndex() == 0:
                # Annual tab
                year = int(self.year_edit.text())
                self.parent_tab.load_annual_summary_data_for_year(year)

                if self.annual_excel_radio.isChecked():
                    self.export_annual_balance_excel(year)
                else:
                    self.export_annual_balance_pdf(year)
            else:
                # Combined tab - LOAD ALL YEARS DATA FIRST
                if self.all_radio.isChecked() or self.date_range_radio.isChecked():
                    # Load ALL data across all years for complete filtering
                    self.parent_tab.load_all_years_data()
                elif self.year_radio.isChecked():
                    year = int(self.year_edit2.text())
                    self.parent_tab.load_financial_data_for_year(year)

                elif self.month_radio.isChecked():
                    year = int(self.month_year_edit.text())
                    self.parent_tab.load_financial_data_for_year(year)
                # Export based on format
                if self.combined_excel_radio.isChecked():
                    self.export_combined_excel()
                else:
                    self.export_combined_pdf()

            # Refresh annual summary after export
            self.parent_tab.load_annual_summary_data_for_year(
                self.parent_tab.annual_summary_year
            )
            self.parent_tab.update_annual_summary()
            self.accept()

        except ImportError as e:
            missing_package = str(e).split("'")[1] if "'" in str(e) else "required package"
            QtWidgets.QMessageBox.critical(
                self, "Missing Dependency",
                f"Error: {missing_package} is not installed.\n\n"
                f"Please install it using: pip install {missing_package}"
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Export Error",
                f"Error during export: {str(e)}"
            )
    def export_annual_balance_excel(self, year):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.pagebreak import Break
        from pathlib import Path
        from datetime import datetime

        export_dir = Path.home() / "Downloads" / "Balance_Exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        filepath = export_dir / f"Annual_Balance_Summary_{year}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Annual Summary"

        # Set print area for both pages
        ws.print_area = 'A1:O66'  # Entire print area covering both pages
        
        # Set page breaks
        ws.row_breaks.append(Break(id=33))  # Page break at row 33
        
        # Set print titles (repeat headers on each page)
        ws.print_title_rows = '1:4'  # Repeat first 4 rows on each page
        
        # Professional color scheme
        dark_blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        orange_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        
        # Fonts
        title_font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        subheader_font = Font(name='Calibri', size=11, bold=True, color="1F4E79")
        normal_font = Font(name='Calibri', size=10)
        bold_font = Font(name='Calibri', size=10, bold=True)
        
        # Alignments
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # Borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Hide gridlines for clean look
        ws.sheet_view.showGridLines = False
        ws.sheet_view.showRowColHeaders = True
        
        # Set column widths
        ws.column_dimensions['A'].width = 2  # Empty spacer column
        ws.column_dimensions['B'].width = 15  # Labels column
        ws.column_dimensions['C'].width = 12  # Jan
        ws.column_dimensions['D'].width = 12  # Feb
        ws.column_dimensions['E'].width = 12  # Mar
        ws.column_dimensions['F'].width = 12  # Apr
        ws.column_dimensions['G'].width = 12  # May
        ws.column_dimensions['H'].width = 12  # Jun
        ws.column_dimensions['I'].width = 12  # Jul
        ws.column_dimensions['J'].width = 12  # Aug
        ws.column_dimensions['K'].width = 12  # Sep
        ws.column_dimensions['L'].width = 12  # Oct
        ws.column_dimensions['M'].width = 12  # Nov
        ws.column_dimensions['N'].width = 12  # Dec
        ws.column_dimensions['O'].width = 18  # Summary column

        # ----------------------------------------------------
        # CALCULATE DATA - SEPARATE UNPAID AND PAID REVENUE
        # ----------------------------------------------------
        monthly_expenses = [0] * 12
        monthly_unpaid_revenue = [0] * 12
        monthly_paid_revenue = [0] * 12

        # Expenses
        for e in self.parent_tab.annual_expenses_data:
            try:
                d = datetime.strptime(e["date"], "%m-%d-%Y")
                if d.year == year:
                    m = d.month - 1
                    monthly_expenses[m] += float(e["amount"].replace("$", "").replace(",", ""))
            except:
                pass

        # Revenue — use is_payment=True entries to match Annual Summary table exactly
        _invoiced_nos = {
            r['invoice_number'].strip()
            for r in (self.parent_tab.annual_revenue_data or [])
            if r.get('is_invoice') and (r.get('invoice_number') or '').strip()
        }
        for r in self.parent_tab.annual_revenue_data:
            if not r.get('is_payment'):
                continue
            inv_num = (r.get('invoice_number') or '').strip()
            if not inv_num or inv_num not in _invoiced_nos:
                continue
            try:
                date_str = r.get('date', r.get('received_date', ''))
                d = self.parent_tab._parse_finance_date(date_str)
                if d and d.year == year:
                    m = d.month - 1
                    monthly_paid_revenue[m] += self.parent_tab._money_to_float(r.get('amount', 0))
            except Exception:
                pass

        total_expenses = sum(monthly_expenses)
        total_unpaid = sum(monthly_unpaid_revenue)
        total_paid = sum(monthly_paid_revenue)
        net_profit = total_paid - total_expenses

        # ----------------------------------------------------
        # PAGE 1 - ANNUAL FINANCIAL SUMMARY (Rows 1-32, Cols A-O)
        # ----------------------------------------------------
        
        # Main Title - MABS HEADING
        ws.merge_cells('B1:O1')
        title_cell = ws['B1']
        try:
            from main import Config as _Cfg
            _co_title = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
        except Exception:
            _co_title = 'MABS ENGINEERING LLC'
        title_cell.value = _co_title
        title_cell.font = Font(name='Calibri', size=24, bold=True, color="1F4E79")
        title_cell.alignment = center_align
        title_cell.fill = white_fill
        
        # Subtitle - Annual Financial Summary
        ws.merge_cells('B6:O6')
        subtitle_cell = ws['B6']
        subtitle_cell.value = f"ANNUAL FINANCIAL SUMMARY - {year}"
        subtitle_cell.font = Font(name='Calibri', size=16, bold=True, color="1F4E79")
        subtitle_cell.alignment = center_align
        subtitle_cell.fill = white_fill
        
        # Empty row for spacing
        ws.row_dimensions[3].height = 10
        
        # Month Headers (Row 10)
        ws['B10'] = "Months"
        ws['B10'].font = header_font
        ws['B10'].fill = dark_blue_fill
        ws['B10'].alignment = center_align
        ws['B10'].border = thin_border
                
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        
        for i, month in enumerate(months, 3):  # Columns C through N
            cell = ws.cell(row=10, column=i, value=month)
            cell.font = header_font
            cell.fill = dark_blue_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # REVENUE Row (Row 11)
        ws['B11'] = "Revenue"
        ws['B11'].font = bold_font
        ws['B11'].fill = light_gray_fill
        ws['B11'].alignment = center_align
        ws['B11'].border = thin_border

        for i, value in enumerate(monthly_paid_revenue, 3):
            cell = ws.cell(row=11, column=i, value=value)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = center_align
            cell.border = thin_border
            if value == 0:
                cell.font = Font(size=9)
            else:
                cell.font = normal_font

        # EXPENSE Row (Row 12)
        ws['B12'] = "Expense"
        ws['B12'].font = bold_font
        ws['B12'].fill = light_gray_fill
        ws['B12'].alignment = center_align
        ws['B12'].border = thin_border

        for i, value in enumerate(monthly_expenses, 3):
            cell = ws.cell(row=12, column=i, value=value)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = center_align
            cell.border = thin_border
            if value == 0:
                cell.font = Font(size=9)
            else:
                cell.font = normal_font
                
        # Empty rows for spacing
        ws.row_dimensions[7].height = 15
        ws.row_dimensions[8].height = 15
        
        # Summary Section
        # ------------------------------------------------
        # SUMMARY SECTION (ALIGNED FORMAT)
        # ------------------------------------------------

        # LABELS
        ws['G18'] = "Revenue"
        ws['G19'] = "Total Expense"
        ws['G20'] = "Net Profit"

        for r in range(18, 21):
            ws[f"G{r}"].font = bold_font
            ws[f"G{r}"].alignment = left_align

        # EQUAL SIGN COLUMN
        ws['H18'] = "="
        ws['H19'] = "="
        ws['H20'] = "="

        for r in range(18, 21):
            ws[f"H{r}"].alignment = center_align
            ws[f"H{r}"].font = bold_font

        # VALUES
        ws.merge_cells('I18:J18')
        rev_cell = ws['I18']
        rev_cell.value = total_paid
        rev_cell.number_format = '"$"#,##0.00'
        rev_cell.font = bold_font
        rev_cell.alignment = left_align

        ws.merge_cells('I19:J19')
        exp_cell = ws['I19']
        exp_cell.value = total_expenses
        exp_cell.number_format = '"$"#,##0.00'
        exp_cell.font = bold_font
        exp_cell.alignment = left_align

        ws.merge_cells('I20:J20')
        net_cell = ws['I20']
        net_cell.value = net_profit
        net_cell.number_format = '"$"#,##0.00'
        net_cell.font = bold_font
        net_cell.fill = green_fill if net_profit > 0 else orange_fill
        net_cell.alignment = left_align
        
        # Page 1 indicator at bottom
        ws['H32'] = "Page : 01"
        ws['H32'].font = Font(size=10, italic=True)
        ws['H32'].alignment = right_align
                
        # Fill empty cells in Page 1 with white
        for row in range(1, 33):
            for col in range(1, 16):  # Columns A to O
                if ws.cell(row=row, column=col).value is None:
                    ws.cell(row=row, column=col).fill = white_fill

        # ----------------------------------------------------
        # PAGE 2 - SALARY BREAKDOWN & EXPENSE BREAKDOWN
        # (Keep the rest of your existing code for Page 2 unchanged)
        # ----------------------------------------------------
        
        start_row = 35

        # -------------------------
        # SALARY TITLE
        # -------------------------
        ws.merge_cells("B35:E35")
        t = ws["B35"]
        t.value = "SALARY"
        t.font = Font(name='Calibri', size=14, bold=True)
        t.alignment = center_align

        # -------------------------
        # INSIDE AMERICA
        # -------------------------
        ws.merge_cells("B37:E37")
        c = ws["B37"]
        c.value = "Inside America"
        c.fill = light_blue_fill
        c.font = bold_font
        c.alignment = center_align
        for col in range(2,6):
            ws.cell(37,col).border = thin_border

        # HEADER
        ws.merge_cells("B38:C38")
        ws["B38"] = "Emp. Nm"

        ws.merge_cells("D38:E38")
        ws["D38"] = "Amount"

        for cell in ["B38","D38"]:
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = light_gray_fill

        for col in range(2,6):
            ws.cell(38,col).border = thin_border

        row = 39
        total_inside = 0

        inside_totals = {}

        for s in self.parent_tab.annual_salary_data.get("Inside America", []):
            try:
                amt = float(s["amount"].replace("$","").replace(",",""))
                name = s["name"].strip().title()
                inside_totals[name] = inside_totals.get(name, 0) + amt
            except:
                pass

        for name, amt in sorted(inside_totals.items()):
            
            ws.merge_cells(f"B{row}:C{row}")
            ws[f"B{row}"] = name
            ws[f"B{row}"].alignment = left_align

            ws.merge_cells(f"D{row}:E{row}")
            v = ws[f"D{row}"]
            v.value = amt
            v.number_format = '"$"#,##0.00'
            v.alignment = center_align

            for col in range(2,6):
                ws.cell(row,col).border = thin_border

            total_inside += amt
            row += 1

        # TOTAL
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"]="Total"
        ws[f"B{row}"].font=bold_font
        ws[f"B{row}"].alignment=center_align

        ws.merge_cells(f"D{row}:E{row}")
        v=ws[f"D{row}"]
        v.value=total_inside
        v.number_format='"$"#,##0.00'
        v.fill=green_fill
        v.font=bold_font
        v.alignment=center_align

        for col in range(2,6):
            ws.cell(row,col).border = thin_border

        row+=2

        # -------------------------
        # OUTSIDE AMERICA
        # -------------------------
        ws.merge_cells(f"B{row}:E{row}")
        c=ws[f"B{row}"]
        c.value="Outside America"
        c.font=bold_font
        c.fill=light_blue_fill
        c.alignment=center_align

        for col in range(2,6):
            ws.cell(row,col).border = thin_border

        row+=1

        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"]="Emp. Nm"

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"]="Amount"

        for cell in [f"B{row}",f"D{row}"]:
            ws[cell].font=bold_font
            ws[cell].alignment=center_align
            ws[cell].fill=light_gray_fill

        for col in range(2,6):
            ws.cell(row,col).border = thin_border

        row+=1

        total_outside = 0

        outside_totals = {}

        for s in self.parent_tab.annual_salary_data.get("Outside America", []):
            try:
                amt = float(s["amount"].replace("$","").replace(",",""))
                name = s["name"].strip().title()
                outside_totals[name] = outside_totals.get(name, 0) + amt
            except:
                pass

        for name, amt in sorted(outside_totals.items()):
            ws.merge_cells(f"B{row}:C{row}")
            ws[f"B{row}"] = name
            ws[f"B{row}"].alignment = left_align

            ws.merge_cells(f"D{row}:E{row}")
            v = ws[f"D{row}"]
            v.value = amt
            v.number_format = '"$"#,##0.00'
            v.alignment = center_align

            for col in range(2,6):
                ws.cell(row,col).border = thin_border

            total_outside += amt
            row += 1

        # TOTAL
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"]="Total"
        ws[f"B{row}"].font=bold_font
        ws[f"B{row}"].alignment = center_align

        ws.merge_cells(f"D{row}:E{row}")
        v=ws[f"D{row}"]
        v.value=total_outside
        v.number_format='"$"#,##0.00'
        v.fill=green_fill
        v.font=bold_font
        v.alignment=center_align

        for col in range(2,6):
            ws.cell(row,col).border = thin_border

        # -------------------------
        # EXPENSE BREAK-DOWN
        # -------------------------
        exp_start = 37

        ws.merge_cells("G35:K35")
        title = ws["G35"]
        title.value = "EXPENSE BREAK-DOWN"
        title.font = Font(name='Calibri', size=14, bold=True)
        title.alignment = center_align

        # HEADER
        ws.merge_cells("G37:I37")
        ws["G37"] = "Expense Item"

        ws.merge_cells("J37:K37")
        ws["J37"] = "Amount"

        for cell in ["G37","J37"]:
            ws[cell].font = bold_font
            ws[cell].alignment = center_align
            ws[cell].fill = light_blue_fill

        for col in range(7,12):
            ws.cell(37,col).border = thin_border

        # Calculate expense totals
        exp_totals = {}
        exp_display_names = {}

        for e in self.parent_tab.annual_expenses_data:
            try:
                d = datetime.strptime(e["date"], "%m-%d-%Y")
                if d.year == year:
                    raw_name = e["name"]
                    key = "".join(raw_name.lower().split())
                    amt = float(e["amount"].replace("$","").replace(",",""))

                    exp_totals[key] = exp_totals.get(key, 0) + amt

                    if key not in exp_display_names:
                        exp_display_names[key] = raw_name.title()
            except:
                pass

        row = 38
        total_exp = 0
        count = 0

        for key, amt in sorted(exp_totals.items(), key=lambda x:x[1], reverse=True):
            name = exp_display_names[key]

            ws.merge_cells(f"G{row}:I{row}")
            item = ws[f"G{row}"]
            item.value = name
            item.alignment = left_align

            ws.merge_cells(f"J{row}:K{row}")
            val = ws[f"J{row}"]
            val.value = amt
            val.number_format = '"$"#,##0.00'
            val.alignment = center_align

            for col in range(7,12):
                ws.cell(row,col).border = thin_border

            total_exp += amt
            row += 1
            count += 1

        # TOTAL ROW
        ws.merge_cells(f"G{row}:I{row}")
        ws[f"G{row}"] = "Total"
        ws[f"G{row}"].font = bold_font
        ws[f"G{row}"].alignment = center_align

        ws.merge_cells(f"J{row}:K{row}")
        val = ws[f"J{row}"]
        val.value = total_exp
        val.number_format = '"$"#,##0.00'
        val.fill = green_fill
        val.font = bold_font
        val.alignment = center_align

        for col in range(7,12):
            ws.cell(row,col).border = thin_border

        # GRAND TOTAL BOX
        mid_row = 37 + (row - 37) // 2

        ws[f"M{mid_row}"] = "Total ="
        ws[f"M{mid_row}"].font = bold_font
        ws[f"M{mid_row}"].alignment = right_align

        ws.merge_cells(start_row=mid_row, start_column=14, end_row=mid_row, end_column=15)

        grand = ws.cell(mid_row,14)

        grand_total = total_inside + total_outside + total_exp

        grand.value = grand_total
        grand.number_format = '"$"#,##0.00'
        grand.fill = green_fill
        grand.font = bold_font
        grand.alignment = center_align

        # Page 2 indicator
        ws['H66'] = "Page : 02"
        ws['H66'].font = Font(size=10, italic=True)
        ws['H66'].alignment = right_align

        # Page setup
        from openpyxl.worksheet.properties import PageSetupProperties

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False
        ws.page_setup.scale = 70

        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

        # Save the workbook
        wb.save(filepath)
        
        # Open the file
        self.open_file(filepath)
    
    
    def export_annual_balance_pdf(self, year):
        """Export Annual Financial Summary to PDF with separate Unpaid and Paid Revenue"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from pathlib import Path
            from datetime import datetime
            
            # Create export directory
            export_dir = Path.home() / "Downloads" / "Balance_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename
            filename = f"Annual_Balance_Summary_{year}.pdf"
            filepath = export_dir / filename
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(filepath), 
                pagesize=landscape(A4),
                topMargin=0.5*inch, 
                bottomMargin=0.5*inch,
                leftMargin=0.5*inch, 
                rightMargin=0.5*inch
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#2C3E50'),
                alignment=1,
                spaceAfter=10,
            )
            
            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#FFFFFF'),
                alignment=1,
                backColor=colors.HexColor('#34495E'),
                borderPadding=5
            )
            
            normal_style = styles['Normal']
            bold_style = ParagraphStyle(
                'BoldStyle',
                parent=styles['Normal'],
                fontSize=10,
                fontName='Helvetica-Bold'
            )
            
            section_header_style = ParagraphStyle(
                'SectionHeader',
                parent=styles['Heading3'],
                fontSize=14,
                textColor=colors.HexColor('#1F4E79'),
                alignment=1,
                fontName='Helvetica-Bold',
                spaceAfter=10
            )
            
            subheader_style = ParagraphStyle(
                'SubheaderStyle',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#1F4E79'),
                fontName='Helvetica-Bold',
                alignment=1,
                spaceAfter=5
            )
            
            # PAGE 1
            # =================================================================
            
            # Title
            try:
                from main import Config as _Cfg
                _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co = 'MABS ENGINEERING LLC'
            elements.append(Paragraph(f"{_co} - ANNUAL BALANCE SHEET - {year}", title_style))
            
            # Calculate monthly data - SEPARATE UNPAID AND PAID
            monthly_expenses = [0] * 12
            monthly_unpaid_revenue = [0] * 12
            monthly_paid_revenue = [0] * 12
            
            # Expenses
            for expense in self.parent_tab.annual_expenses_data:
                try:
                    date = datetime.strptime(expense.get('date', ''), "%m-%d-%Y")
                    if date.year == year:
                        month = date.month - 1
                        amount = float(expense.get('amount', '0').replace('$', '').replace(',', ''))
                        monthly_expenses[month] += amount
                except:
                    pass
            
            # Revenue — use is_payment=True entries to match Annual Summary table exactly
            _invoiced_nos = {
            r['invoice_number'].strip()
            for r in (self.parent_tab.annual_revenue_data or [])
            if r.get('is_invoice') and (r.get('invoice_number') or '').strip()
        }
            for revenue in self.parent_tab.annual_revenue_data:
                if not revenue.get('is_payment'):
                    continue
                inv_num = (revenue.get('invoice_number') or '').strip()
                if not inv_num or inv_num not in _invoiced_nos:
                    continue
                try:
                    date_str = revenue.get('date', revenue.get('received_date', ''))
                    date = self.parent_tab._parse_finance_date(date_str)
                    if date and date.year == year:
                        month = date.month - 1
                        amount = self.parent_tab._money_to_float(revenue.get('amount', 0))
                        monthly_paid_revenue[month] += amount
                except Exception:
                    pass

            total_expenses = sum(monthly_expenses)
            total_unpaid = sum(monthly_unpaid_revenue)
            total_paid = sum(monthly_paid_revenue)
            net_profit = total_paid - total_expenses
            
            # Monthly Summary Table with 3 rows: Revenue, Expenses, Net P/L
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"]

            table_data = []

            # Header row
            header_row = [''] + months
            table_data.append(header_row)

            gray_cells = []

            # Revenue row (row 1 in table) — paid revenue only, matches UI table
            revenue_row = ['Revenue']
            for i, value in enumerate(monthly_paid_revenue):
                revenue_row.append(f"${value:,.2f}")
                if value == 0:
                    gray_cells.append((1, i + 1))
            revenue_row.append(f"${total_paid:,.2f}")
            table_data.append(revenue_row)

            # Expenses row (row 2 in table)
            expenses_row = ['Expenses']
            for i, value in enumerate(monthly_expenses):
                expenses_row.append(f"${value:,.2f}")
                if value == 0:
                    gray_cells.append((2, i + 1))
            expenses_row.append(f"${total_expenses:,.2f}")
            table_data.append(expenses_row)

            # Net P/L row (row 3 in table)
            monthly_net = [monthly_paid_revenue[i] - monthly_expenses[i] for i in range(12)]
            net_row = ['Net P/L']
            for i, value in enumerate(monthly_net):
                net_row.append(f"${value:,.2f}")
            net_row.append(f"${net_profit:,.2f}")
            table_data.append(net_row)

            # Create table
            table = Table(table_data, colWidths=[1.2*inch] + [0.74*inch]*12 + [1*inch])

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#D5F5E3')),  # Revenue - Green
                ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#FADBD8')),  # Expenses - Red
                ('BACKGROUND', (-1, 1), (-1, 1), colors.HexColor('#D5F5E3')),
                ('BACKGROUND', (-1, 2), (-1, 2), colors.HexColor('#FADBD8')),
            ])

            for row, col in gray_cells:
                table_style.add('TEXTCOLOR', (col, row), (col, row), colors.HexColor('#A0A0A0'))
                table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor('#F5F5F5'))

            table.setStyle(table_style)

            elements.append(Spacer(1, 0.8*inch))
            elements.append(table)
            elements.append(Spacer(1, 0.4*inch))

            # Summary Section
            summary_data = [
                ["Revenue", ":", f"${total_paid:,.2f}"],
                ["Total Expense", ":", f"${total_expenses:,.2f}"],
                ["Net Profit", ":", f"${net_profit:,.2f}"]
            ]

            summary_table = Table(summary_data, colWidths=[1.2*inch, 0.2*inch, 1.2*inch])

            summary_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('ALIGN', (1,0), (1,-1), 'CENTER'),
                ('ALIGN', (2,0), (2,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 10),
                ('BACKGROUND', (2,2), (2,2), colors.HexColor('#D4E6F1')),
            ]))

            summary_table.hAlign = "CENTER"
            elements.append(summary_table)
            
            # Page 1 indicator
            elements.append(Spacer(1, 0.5*inch))
            
            # PAGE 2 - Keep your existing Page 2 code (Salary & Expense Breakdown)
            # =================================================================
            elements.append(PageBreak())
            
            # Create two-column layout for Page 2 using tables
            page2_data = []
            
            # Left Column - Salary Breakdown
            left_content = []
            
            # Salary Title
            left_content.append([Paragraph("SALARY", section_header_style)])
            left_content.append([Spacer(1, 0.1*inch)])
            
            # Inside America Section
            salary_data = []

            # INSIDE AMERICA HEADER
            salary_data.append(
                [Paragraph("<b>Inside America</b>", subheader_style), ""]
            )

            # COLUMN HEADERS
            salary_data.append(["Emp. Nm", "Amount"])

            # Collect Inside America totals
            inside_totals = {}
            for s in self.parent_tab.annual_salary_data.get("Inside America", []):
                try:
                    amt = float(s["amount"].replace("$","").replace(",",""))
                    name = s["name"].strip().title()
                    inside_totals[name] = inside_totals.get(name,0) + amt
                except:
                    pass

            total_inside = 0
            for name, amt in sorted(inside_totals.items()):
                salary_data.append([name, f"${amt:,.2f}"])
                total_inside += amt

            salary_data.append(["Total", f"${total_inside:,.2f}"])

            # OUTSIDE AMERICA HEADER
            salary_data.append(
                [Paragraph("<b>Outside America</b>", subheader_style), ""]
            )

            # COLUMN HEADERS
            salary_data.append(["Emp. Nm", "Amount"])

            # Collect Outside America totals
            outside_totals = {}
            for s in self.parent_tab.annual_salary_data.get("Outside America", []):
                try:
                    amt = float(s["amount"].replace("$","").replace(",",""))
                    name = s["name"].strip().title()
                    outside_totals[name] = outside_totals.get(name,0) + amt
                except:
                    pass

            total_outside = 0
            for name, amt in sorted(outside_totals.items()):
                salary_data.append([name, f"${amt:,.2f}"])
                total_outside += amt

            salary_data.append(["Total", f"${total_outside:,.2f}"])

            salary_table = Table(salary_data, colWidths=[2.8*inch,1*inch])

            salary_table.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#D9E1F2')),
                ('SPAN',(0,0),(1,0)),
                ('ALIGN',(0,0),(-1,0),'CENTER'),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#F2F2F2')),
                ('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),
                ('ALIGN',(0,1),(-1,1),'CENTER'),
                ('ALIGN',(0,len(inside_totals)+4),(-1,len(inside_totals)+4),'CENTER'),
                ('BACKGROUND',(1,len(inside_totals)+2),(1,len(inside_totals)+2),colors.HexColor('#C6EFCE')),
                ('BACKGROUND',(0,len(inside_totals)+3),(-1,len(inside_totals)+3),colors.HexColor('#D9E1F2')),
                ('SPAN',(0,len(inside_totals)+3),(1,len(inside_totals)+3)),
                ('ALIGN',(0,len(inside_totals)+3),(-1,len(inside_totals)+3),'CENTER'),
                ('FONTNAME',(0,len(inside_totals)+3),(-1,len(inside_totals)+3),'Helvetica-Bold'),
                ('BACKGROUND',(0,len(inside_totals)+4),(-1,len(inside_totals)+4),colors.HexColor('#F2F2F2')),
                ('FONTNAME',(0,len(inside_totals)+4),(-1,len(inside_totals)+4),'Helvetica-Bold'),
                ('BACKGROUND',(1,len(salary_data)-1),(1,len(salary_data)-1),colors.HexColor('#C6EFCE')),
                ('ALIGN',(0,1),(-1,1),'CENTER'),
                ('ALIGN',(1,2),(1,-1),'CENTER'),
                ('ALIGN',(0,2),(0,len(inside_totals)+1),'LEFT'),
                ('ALIGN',(0,len(inside_totals)+5),(0,len(salary_data)-2),'LEFT'),
                ('ALIGN',(0,len(inside_totals)+2),(0,len(inside_totals)+2),'CENTER'),
                ('ALIGN',(0,len(salary_data)-1),(0,len(salary_data)-1),'CENTER'),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ]))

            left_content.append([salary_table])
            left_table = Table(left_content)
            
            # Right Column - Expense Breakdown
            right_content = []
            
            # Expense Title
            right_content.append([Paragraph("EXPENSE BREAK-DOWN", section_header_style)])
            right_content.append([Spacer(1, 0.1*inch)])
            
            # Calculate consolidated expenses
            expense_totals = {}
            expense_display_names = {}

            for expense in self.parent_tab.annual_expenses_data:
                try:
                    date = datetime.strptime(expense.get('date', ''), "%m-%d-%Y")
                    if date.year == year:
                        raw_name = expense.get('name', 'Unknown')
                        key = "".join(raw_name.lower().split())
                        amount = float(expense.get('amount', '0').replace('$', '').replace(',', ''))
                        expense_totals[key] = expense_totals.get(key, 0) + amount
                        if key not in expense_display_names:
                            expense_display_names[key] = raw_name.strip().title()
                except:
                    pass
                        
            # Sort by amount descending
            expense_data = [["Expense Item", "Amount"]]

            for key, amount in sorted(expense_totals.items(), key=lambda x: x[1], reverse=True):
                name = expense_display_names[key]
                expense_data.append([name, f"${amount:,.2f}"])
            expense_data.append(["Total", f"${total_expenses:,.2f}"])
                        
            expense_table = Table(expense_data, colWidths=[3.5*inch, 1.4*inch])
            
            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -2), 'LEFT'),
                ('ALIGN', (0, -1), (0, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, -1), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (1, -1), (1, -1), colors.HexColor('#C6EFCE')),
            ]
            
            expense_table.setStyle(TableStyle(style_commands))
            right_content.append([expense_table])
            right_table = Table(right_content)
            
            # Create two-column layout for Page 2
            page2_layout = Table([[left_table, "", right_table]], colWidths=[4*inch, 0.6*inch, 5*inch])
            page2_layout.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (0, 0), -20),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            elements.append(page2_layout)
            elements.append(Spacer(1,0.35*inch))

            grand_total_value = total_inside + total_outside + total_expenses

            grand_total_data = [
                ["Total  =", f"${grand_total_value:,.2f}"]
            ]

            grand_total_table = Table(grand_total_data, colWidths=[0.5*inch,2*inch])

            grand_total_table.setStyle(TableStyle([
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,0),12),
                ('ALIGN',(0,0),(0,0),'RIGHT'),
                ('ALIGN',(1,0),(1,0),'CENTER'),
                ('BACKGROUND',(1,0),(1,0),colors.HexColor('#C6EFCE')),
                ('GRID',(1,0),(1,0),0.6,colors.black),
            ]))

            grand_total_table.hAlign = "CENTER"
            elements.append(grand_total_table)
            
            # Page 2 indicator
            elements.append(Spacer(1, 0.5*inch))
            
            def add_page_number(canvas, doc):
                page_num = canvas.getPageNumber()
                text = f"Page : {page_num:02d}"
                canvas.setFont("Helvetica", 10)
                width, height = landscape(A4)
                canvas.drawCentredString(width / 2.0, 20, text)
                
            # Build PDF
            doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
            
            # Open the file
            self.open_file(filepath)
            
            QtWidgets.QMessageBox.information(
                self, "Export Successful",
                f"✅ Annual Financial Summary exported successfully!\n\n"
                f"File saved to: {filepath}"
            )
            
        except ImportError:
            QtWidgets.QMessageBox.critical(
                self, "Missing Dependency",
                "ReportLab is not installed.\n\nPlease install it using: pip install reportlab"
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Export Error",
                f"Error exporting PDF: {str(e)}"
            )
            raise
    
    
    def export_combined_excel(self):
        """Export Financial Overview report to Excel with proper date range filtering"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from pathlib import Path
            from datetime import datetime
            
            # Helper function to safely convert amount to float
            def safe_float_amount(amount):
                if isinstance(amount, (int, float)):
                    return float(amount)
                elif isinstance(amount, str):
                    try:
                        return float(amount.replace('$', '').replace(',', ''))
                    except:
                        return 0.0
                return 0.0
            
            # Helper function to safely get string value
            def safe_str(value):
                if value is None:
                    return ''
                return str(value)
            
            # Create export directory
            export_dir = Path.home() / "Downloads" / "Balance_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename with date range info
            range_suffix = self.get_range_description().replace(" ", "_").replace("/", "-")
            filename = f"Financial_Overview_{range_suffix}.xlsx"
            filepath = export_dir / filename
            
            # Create workbook
            wb = Workbook()
            
            # Define styles
            title_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            normal_font = Font(name='Calibri', size=10)
            bold_font = Font(name='Calibri', size=10, bold=True)
            wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Get filtered data based on range
            range_description = self.get_range_description()
            
            # Revenue Section with columns: S.No, Date, Revenue Source, Description, Amount ($), Due Date, Status, Received Date
            if self.include_revenue.isChecked():
                _invoice_rev = [r for r in self.parent_tab.revenue_data
                    if not r.get('is_payment') and r.get('is_invoice', True)]
                revenue_data = self.filter_data_by_range(_invoice_rev)
                if revenue_data:
                    ws = wb.create_sheet("Revenue")
                    
                    # Title
                    ws.merge_cells('A1:H1')
                    cell = ws['A1']
                    try:
                        from main import Config as _Cfg
                        _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
                    except Exception:
                        _co = 'MABS ENGINEERING'
                    cell.value = f"{_co} - REVENUE REPORT ({range_description})"
                    cell.font = title_font
                    cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Headers: S.No, Date, Revenue Source, Description, Amount ($), Due Date, Status, Received Date
                    headers = ["S.No", "Date", "Revenue Source", "Description", "Amount ($)", "Due Date", "Status", "Received Date"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col, value=header)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Data
                    for row, rev in enumerate(revenue_data, 5):
                        ws.cell(row=row, column=1, value=row-4).alignment = Alignment(horizontal='center')
                        
                        # Date (Invoice Date)
                        date_cell = ws.cell(row=row, column=2, value=safe_str(rev.get('date', '')))
                        date_cell.alignment = Alignment(horizontal='center')
                        
                        # Revenue Source — strip "Invoice - " prefix if present
                        _src = safe_str(rev.get('source', ''))
                        if _src.startswith("Invoice - "):
                            _src = _src[len("Invoice - "):]
                        ws.cell(row=row, column=3, value=_src).alignment = Alignment(horizontal='center', vertical='center')

                        # Description
                        desc_cell = ws.cell(row=row, column=4, value=safe_str(rev.get('description', '')))
                        desc_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Amount
                        amount = safe_float_amount(rev.get('amount', 0))
                        amount_cell = ws.cell(row=row, column=5, value=amount)
                        amount_cell.number_format = '"$"#,##0.00'
                        amount_cell.alignment = Alignment(horizontal='center')
                        
                        # Due Date
                        due_date = safe_str(rev.get('due_date', 'N/A'))
                        due_date_cell = ws.cell(row=row, column=6, value=due_date)
                        due_date_cell.alignment = Alignment(horizontal='center')
                        
                        # Highlight overdue due dates
                        status = safe_str(rev.get('status', 'Unpaid'))
                        if status == "Unpaid" and due_date != 'N/A':
                            try:
                                due_date_obj = datetime.strptime(due_date, "%m-%d-%Y")
                                if due_date_obj < datetime.now():
                                    due_date_cell.font = Font(color="E74C3C", bold=True)
                            except:
                                pass
                        
                        # Status
                        status_cell = ws.cell(row=row, column=7, value=status)
                        status_cell.alignment = Alignment(horizontal='center')
                        
                        # Color code status
                        if status == "Paid":
                            status_cell.font = Font(color="27AE60", bold=True)
                        elif status == "Unpaid":
                            status_cell.font = Font(color="E74C3C", bold=True)
                        elif status == "Pending":
                            status_cell.font = Font(color="F39C12", bold=True)
                        elif status == "Overdue":
                            status_cell.font = Font(color="C0392B", bold=True)
                        elif status == "Partially Paid":
                            status_cell.font = Font(color="8E44AD", bold=True)
                        
                        # Received Date
                        received_date = safe_str(rev.get('received_date', 'N/A'))
                        received_cell = ws.cell(row=row, column=8, value=received_date)
                        received_cell.alignment = Alignment(horizontal='center')
                        
                        for col in range(1, 9):
                            ws.cell(row=row, column=col).border = thin_border
                        
                        # Auto-adjust row height for wrapped text
                        ws.row_dimensions[row].height = None
                    
                    # Total row
                    total_row = len(revenue_data) + 5
                    ws.cell(row=total_row, column=4, value="TOTAL").font = bold_font
                    total_amount = sum(safe_float_amount(r.get('amount', 0)) for r in revenue_data)
                    total_cell = ws.cell(row=total_row, column=5, value=total_amount)
                    total_cell.alignment = Alignment(horizontal='center')
                    total_cell.number_format = '"$"#,##0.00'
                    total_cell.font = bold_font
                    
                    # Adjust column widths
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 40
                    ws.column_dimensions['E'].width = 15
                    ws.column_dimensions['F'].width = 15
                    ws.column_dimensions['G'].width = 15
                    ws.column_dimensions['H'].width = 15
            
            # Expenses Section
            if self.include_expenses.isChecked():
                expenses_data = self.filter_data_by_range(self.parent_tab.expenses_data)
                if expenses_data:
                    ws = wb.create_sheet("Expenses")
                    
                    # Title
                    ws.merge_cells('A1:E1')
                    cell = ws['A1']
                    try:
                        from main import Config as _Cfg
                        _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
                    except Exception:
                        _co = 'MABS ENGINEERING'
                    cell.value = f"{_co} - EXPENSES REPORT ({range_description})"
                    cell.font = title_font
                    cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Headers
                    headers = ["S.No", "Date", "Expense Item", "Description", "Amount ($)"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col, value=header)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Data
                    for row, exp in enumerate(expenses_data, 5):
                        ws.cell(row=row, column=1, value=row-4).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=2, value=safe_str(exp.get('date', ''))).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=3, value=safe_str(exp.get('name', ''))).alignment = Alignment(horizontal='center', vertical='center')

                        # Description
                        desc_cell = ws.cell(row=row, column=4, value=safe_str(exp.get('description', '')))
                        desc_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        amount = safe_float_amount(exp.get('amount', 0))
                        amount_cell = ws.cell(row=row, column=5, value=amount)
                        amount_cell.number_format = '"$"#,##0.00'
                        amount_cell.alignment = Alignment(horizontal='center')
                        
                        for col in range(1, 6):
                            ws.cell(row=row, column=col).border = thin_border
                        
                        ws.row_dimensions[row].height = None
                    
                    # Total row
                    total_row = len(expenses_data) + 5
                    ws.cell(row=total_row, column=4, value="TOTAL").font = bold_font
                    total_amount = sum(safe_float_amount(e.get('amount', 0)) for e in expenses_data)
                    total_cell = ws.cell(row=total_row, column=5, value=total_amount)
                    total_cell.alignment = Alignment(horizontal='center')
                    total_cell.number_format = '"$"#,##0.00'
                    total_cell.font = bold_font
                    
                    # Adjust column widths
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 45
                    ws.column_dimensions['E'].width = 15
            
            # Salary Section
            if self.include_salary.isChecked():
                all_salary = []
                for cat in ["Inside America", "Outside America"]:
                    for sal in self.parent_tab.salary_data[cat]:
                        sal_copy = sal.copy()
                        sal_copy['region'] = cat
                        all_salary.append(sal_copy)
                
                salary_data = self.filter_data_by_range(all_salary)
                
                if salary_data:
                    ws = wb.create_sheet("Salaries")
                    
                    # Title
                    ws.merge_cells('A1:F1')
                    cell = ws['A1']
                    try:
                        from main import Config as _Cfg
                        _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
                    except Exception:
                        _co = 'MABS ENGINEERING'
                    cell.value = f"{_co} - SALARY REPORT ({range_description})"
                    cell.font = title_font
                    cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Headers
                    headers = ["S.No", "Date", "Region", "Employee", "Description", "Amount ($)"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col, value=header)
                        cell.font = header_font
                        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Data
                    for row, sal in enumerate(salary_data, 5):
                        ws.cell(row=row, column=1, value=row-4).alignment = Alignment(horizontal='center')
                        ws.cell(row=row, column=2, value=safe_str(sal.get('date', ''))).alignment = Alignment(horizontal='center')
                        
                        region = safe_str(sal.get('region', ''))
                        region_cell = ws.cell(row=row, column=3, value=region)
                        region_cell.alignment = Alignment(horizontal='center', vertical='center')
                        if region == "Inside America":
                            region_cell.font = Font(color="2980B9", bold=True)
                        else:
                            region_cell.font = Font(color="C0392B", bold=True)

                        ws.cell(row=row, column=4, value=safe_str(sal.get('name', ''))).alignment = Alignment(horizontal='center', vertical='center')

                        # Description
                        desc_cell = ws.cell(row=row, column=5, value=safe_str(sal.get('description', '')))
                        desc_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        amount = safe_float_amount(sal.get('amount', 0))
                        amount_cell = ws.cell(row=row, column=6, value=amount)
                        amount_cell.number_format = '"$"#,##0.00'
                        amount_cell.alignment = Alignment(horizontal='center')
                        
                        for col in range(1, 7):
                            ws.cell(row=row, column=col).border = thin_border
                        
                        ws.row_dimensions[row].height = None
                    
                    # Total row
                    total_row = len(salary_data) + 5
                    ws.cell(row=total_row, column=5, value="TOTAL").font = bold_font
                    total_amount = sum(safe_float_amount(s.get('amount', 0)) for s in salary_data)
                    total_cell = ws.cell(row=total_row, column=6, value=total_amount)
                    total_cell.alignment = Alignment(horizontal='center')
                    total_cell.number_format = '"$"#,##0.00'
                    total_cell.font = bold_font
                    
                    # Adjust column widths
                    ws.column_dimensions['A'].width = 10
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 18
                    ws.column_dimensions['D'].width = 20
                    ws.column_dimensions['E'].width = 35
                    ws.column_dimensions['F'].width = 15
             
            # Remove default sheet if empty
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Save file
            wb.save(str(filepath))
            
            # Open the file
            self.open_file(filepath)
            
            QtWidgets.QMessageBox.information(
                self, "Export Successful",
                f"✅ Financial Overview report exported successfully!\n\n"
                f"File saved to: {filepath}"
            )
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Export Error",
                f"Error exporting Excel: {str(e)}"
            )
            raise

    def export_combined_pdf(self):
        """Export Financial Overview report to PDF with proper date range filtering - Portrait mode"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from pathlib import Path
            from datetime import datetime
            
            # Helper function to safely convert amount to float
            def safe_float_amount(amount):
                if isinstance(amount, (int, float)):
                    return float(amount)
                elif isinstance(amount, str):
                    try:
                        return float(amount.replace('$', '').replace(',', ''))
                    except:
                        return 0.0
                return 0.0
            
            # Helper function to safely get string value
            def safe_str(value):
                if value is None:
                    return ''
                return str(value)
            
            # Create export directory
            export_dir = Path.home() / "Downloads" / "Balance_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename with date range info
            range_suffix = self.get_range_description().replace(" ", "_").replace("/", "-")
            filename = f"Financial_Overview_{range_suffix}.pdf"
            filepath = export_dir / filename
            
            # Create PDF document with portrait orientation
            doc = SimpleDocTemplate(
                str(filepath), 
                pagesize=A4,
                topMargin=0.5*inch, 
                bottomMargin=0.5*inch,
                leftMargin=0.5*inch, 
                rightMargin=0.5*inch
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#FFFFFF'),
                alignment=1,
                backColor=colors.HexColor('#2C3E50'),
                borderPadding=8,
                spaceAfter=10
            )
            
            section_style = ParagraphStyle(
                'SectionStyle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#2C3E50'),
                spaceBefore=12,
                spaceAfter=8,
                fontName='Helvetica-Bold'
            )
            
            # Style for normal cells with wrapping — centered
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8,
                leading=10,
                alignment=1
            )
            
            # Get range description
            range_description = self.get_range_description()
            
            # Title
            try:
                from main import Config as _Cfg
                _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co = 'MABS ENGINEERING'
            elements.append(Paragraph(f"{_co} - FINANCIAL OVERVIEW REPORT", title_style))
            elements.append(Paragraph(f"<b>Period:</b> {range_description}", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Revenue Section
            if self.include_revenue.isChecked():
                _invoice_rev = [r for r in self.parent_tab.revenue_data
                    if not r.get('is_payment') and r.get('is_invoice', True)]
                revenue_data = self.filter_data_by_range(_invoice_rev)
                if revenue_data:
                    elements.append(Paragraph("REVENUE", section_style))
                    
                    # Helper function to get status color
                    def get_status_color(status):
                        colors_map = {
                            "Paid": "#27AE60",
                            "Unpaid": "#E74C3C",
                            "Pending": "#F39C12",
                            "Overdue": "#C0392B",
                            "Partially Paid": "#8E44AD"
                        }
                        return colors_map.get(status, "#7F8C8D")
                    
                    hdr_s = ParagraphStyle('BsHdr', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, alignment=1, wordWrap='CJK')
                    # Headers: S.No, Date, Revenue Source, Description, Amount ($), Due Date, Status, Received Date
                    data = [[Paragraph(h, hdr_s) for h in ["S.No", "Date", "Revenue Source", "Description", "Amount ($)", "Due Date", "Status", "Received Date"]]]

                    for idx, rev in enumerate(revenue_data, 1):
                        amount = safe_float_amount(rev.get('amount', 0))
                        status = safe_str(rev.get('status', 'Unpaid'))
                        due_date = safe_str(rev.get('due_date', 'N/A'))
                        invoice_date = safe_str(rev.get('date', ''))
                        received_date = safe_str(rev.get('received_date', 'N/A'))
                        revenue_source = safe_str(rev.get('source', ''))
                        if revenue_source.startswith("Invoice - "):
                            revenue_source = revenue_source[len("Invoice - "):]
                        description = safe_str(rev.get('description', ''))
                        
                        # Create wrapped paragraphs
                        date_para = Paragraph(
                            f'<font size="8">{invoice_date}</font>',
                            cell_style
                        )
                        source_para = Paragraph(revenue_source, cell_style)
                        desc_para = Paragraph(description, cell_style)
                        received_date_para = Paragraph(
                            f'<font size="8">{received_date}</font>',
                            cell_style
                        )
                        
                        # Format status with color
                        status_para = Paragraph(f'<font color="{get_status_color(status)}"><b>{status}</b></font>', cell_style)
                        amount_para = Paragraph(
                            f'<para align="center"><font size="8">${amount:,.2f}</font></para>',
                            cell_style
                        )
                        # Format due date with highlight if overdue
                        due_date_para = Paragraph(due_date, cell_style)
                        if status == "Unpaid" and due_date != 'N/A':
                            try:
                                due_date_obj = datetime.strptime(due_date, "%m-%d-%Y")
                                if due_date_obj < datetime.now():
                                    due_date_para = Paragraph(f'<font color="#E74C3C"><b>{due_date}</b></font>', cell_style)
                            except:
                                pass
                        
                        data.append([
                            Paragraph(str(idx), cell_style),
                            date_para,
                            source_para,
                            desc_para,
                            amount_para,
                            due_date_para,
                            status_para,
                            received_date_para
                        ])
                    
                    # Add total row
                    total = sum(safe_float_amount(r.get('amount', 0)) for r in revenue_data)
                    total_label = Paragraph('<b>TOTAL</b>', cell_style)
                    total_amount = Paragraph(
                        f'<para align="center"><b>${total:,.2f}</b></para>',
                        cell_style
                    )

                    data.append(["", "", "", total_label, total_amount, "", "", ""])
                    
                    # Create table with appropriate column widths
                    table = Table(data, colWidths=[0.45*inch, 0.8*inch, 1.3*inch, 1.6*inch, 0.85*inch, 0.8*inch, 0.7*inch, 0.8*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 7),
                        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (3, -1), (3, -1), 'Helvetica-Bold'),
                        ('FONTNAME', (4, -1), (4, -1), 'Helvetica-Bold'),
                        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 0.2*inch))
            
            # Expenses Section
            if self.include_expenses.isChecked():
                expenses_data = self.filter_data_by_range(self.parent_tab.expenses_data)
                if expenses_data:
                    elements.append(Paragraph("EXPENSES", section_style))
                    hdr_s2 = ParagraphStyle('BsHdr2', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, alignment=1, wordWrap='CJK')
                    data = [[Paragraph(h, hdr_s2) for h in ["S.No", "Date", "Expense Item", "Description", "Amount ($)"]]]
                    for idx, exp in enumerate(expenses_data, 1):
                        amount = safe_float_amount(exp.get('amount', 0))
                        
                        date_para = Paragraph(
                            f'<font size="8">{safe_str(exp.get("date", ""))}</font>',
                            cell_style
                        )
                        
                        # Create wrapped paragraphs
                        item_para = Paragraph(safe_str(exp.get('name', '')), cell_style)
                        desc_para = Paragraph(safe_str(exp.get('description', '')), cell_style)
                        amount_para = Paragraph(
                            f'<para align="center"><font size="8">${amount:,.2f}</font></para>',
                            cell_style
                        )
                        data.append([
                            Paragraph(str(idx), cell_style),
                            date_para,
                            item_para,
                            desc_para,
                            amount_para
                        ])
                    
                    # Add total row
                    total = sum(safe_float_amount(e.get('amount', 0)) for e in expenses_data)
                    total_label = Paragraph('<b>TOTAL</b>', cell_style)
                    total_amount = Paragraph(
                        f'<para align="center"><b>${total:,.2f}</b></para>',
                        cell_style
                    )

                    data.append(["", "", "", total_label, total_amount])
                    table = Table(data, colWidths=[0.45*inch, 0.9*inch, 2.15*inch, 2.9*inch, 0.9*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (3, -1), (3, -1), 'Helvetica-Bold'),
                        ('FONTNAME', (4, -1), (4, -1), 'Helvetica-Bold'),
                        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 0.2*inch))
            
            # Salaries Section
            if self.include_salary.isChecked():
                all_salary = []
                for cat in ["Inside America", "Outside America"]:
                    for sal in self.parent_tab.salary_data[cat]:
                        sal_copy = sal.copy()
                        sal_copy['region'] = cat
                        all_salary.append(sal_copy)
                
                salary_data = self.filter_data_by_range(all_salary)
                
                if salary_data:
                    elements.append(Paragraph("SALARIES", section_style))
                    
                    hdr_s3 = ParagraphStyle('BsHdr3', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', textColor=colors.white, alignment=1, wordWrap='CJK')
                    data = [[Paragraph(h, hdr_s3) for h in ["S.No", "Date", "Region", "Employee", "Description", "Amount ($)"]]]
                    for idx, sal in enumerate(salary_data, 1):
                        amount = safe_float_amount(sal.get('amount', 0))
                        region = safe_str(sal.get('region', ''))
                        
                        date_para = Paragraph(
                            f'<font size="8">{safe_str(sal.get("date", ""))}</font>',
                            cell_style
                        )
                        
                        # Create wrapped paragraphs
                        emp_para = Paragraph(safe_str(sal.get('name', '')), cell_style)
                        desc_para = Paragraph(safe_str(sal.get('description', '')), cell_style)
                        
                        # Color region text
                        region_color = "#2980B9" if region == "Inside America" else "#C0392B"
                        region_para = Paragraph(f'<font color="{region_color}">{region}</font>', cell_style)
                        amount_para = Paragraph(
                            f'<para align="center"><font size="8">${amount:,.2f}</font></para>',
                            cell_style
                        )
                        
                        data.append([
                            Paragraph(str(idx), cell_style),
                            date_para,
                            region_para,
                            emp_para,
                            desc_para,
                            amount_para
                        ])
                    
                    # Add total row
                    total = sum(safe_float_amount(s.get('amount', 0)) for s in salary_data)
                    total_label = Paragraph('<b>TOTAL</b>', cell_style)
                    total_amount = Paragraph(
                        f'<para align="center"><b>${total:,.2f}</b></para>',
                        cell_style
                    )

                    data.append(["", "", "", total_label, "", total_amount])
                    
                    table = Table(data, colWidths=[0.45*inch, 0.9*inch, 1.3*inch, 1.5*inch, 2.25*inch, 0.9*inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (3, -1), (4, -1), 'Helvetica-Bold'),
                        ('FONTNAME', (5, -1), (5, -1), 'Helvetica-Bold'),
                        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                    ]))
                    elements.append(table)
            

            # Add page numbers
            def add_page_number(canvas, doc):
                page_num = canvas.getPageNumber()
                text = f"Page {page_num}"
                canvas.setFont("Helvetica", 9)
                canvas.drawRightString(A4[0] - 30, 20, text)
                canvas.drawString(30, 20, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            
            # Build PDF
            doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
            
            # Open the file
            self.open_file(filepath)
            
            QtWidgets.QMessageBox.information(
                self, "Export Successful",
                f"✅ Financial Overview report exported successfully!\n\n"
                f"File saved to: {filepath}"
            )
            
        except ImportError:
            QtWidgets.QMessageBox.critical(
                self, "Missing Dependency",
                "ReportLab is not installed.\n\nPlease install it using: pip install reportlab"
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Export Error",
                f"Error exporting PDF: {str(e)}"
            )
            raise
        
    
    def get_range_description(self):
        """Get description of selected range"""
        if self.all_radio.isChecked():
            return "All Data"
        elif self.date_range_radio.isChecked():
            return f"{self.from_date.date().toString('MM-dd-yyyy')} to {self.to_date.date().toString('MM-dd-yyyy')}"
        elif self.month_radio.isChecked():
            return f"{self.month_combo.currentText()} {self.month_year_edit.text()}"
        else:
            return f"Year {self.year_edit2.text()}"
    
    def _parse_export_item_date(self, item):
        """Parse finance dates from all records that can appear in exports."""
        date_fields = (
            "date",
            "received_date",
            "paid_date",
            "payment_date",
            "due_date",
            "created_at",
            "updated_at",
            "completion_date",
        )
        date_formats = (
            "%m-%d-%Y",
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%m-%d-%y",
            "%m/%d/%y",
            "%B %d, %Y",
        )

        for field in date_fields:
            raw_value = item.get(field)
            if not raw_value or raw_value == "N/A":
                continue

            text = str(raw_value).strip()
            if not text:
                continue

            if "T" in text:
                iso_text = text.replace("Z", "+00:00")
                try:
                    return datetime.fromisoformat(iso_text)
                except ValueError:
                    text = text.split("T", 1)[0]

            for date_format in date_formats:
                try:
                    return datetime.strptime(text, date_format)
                except ValueError:
                    continue

        return None
    
    def filter_data_by_range(self, data_list):
        """Filter data based on selected range"""
        if self.all_radio.isChecked():
            return self._sort_export_items(list(data_list))

        filtered = []
        for item in data_list:
            try:
                item_datetime = self._parse_export_item_date(item)
                item_year = None
                try:
                    if item.get("year"):
                        item_year = int(item.get("year"))
                except Exception:
                    item_year = None
                
                include = False
                
                if self.date_range_radio.isChecked():
                    if item_datetime is None:
                        continue
                    from_date = self.from_date.date().toPyDate()
                    to_date = self.to_date.date().toPyDate()
                    if from_date <= item_datetime.date() <= to_date:
                        include = True
                
                elif self.month_radio.isChecked():
                    year = int(self.month_year_edit.text())
                    month = self.month_combo.currentIndex() + 1
                    if item_datetime and item_datetime.year == year and item_datetime.month == month:
                        include = True
                
                elif self.year_radio.isChecked():
                    year = int(self.year_edit2.text())
                    if (item_datetime and item_datetime.year == year) or item_year == year:
                        include = True
                
                if include:
                    filtered.append(item)
                        
            except Exception as e:
                _log.warning("Error filtering item: %s", e)
                continue
        
        return self._sort_export_items(filtered)

    def _sort_export_items(self, items):
        """Sort export records chronologically while keeping undated rows."""
        def sort_key(item):
            parsed = self._parse_export_item_date(item)
            if parsed:
                return (0, parsed)
            try:
                return (1, datetime(int(item.get("year")), 1, 1))
            except Exception:
                return (2, datetime(1900, 1, 1))

        return sorted(items, key=sort_key)
            
    def open_file(self, filepath):
        """Open file with default application"""
        try:
            import platform
            import subprocess
            import os
            
            if platform.system() == "Windows":
                os.startfile(filepath)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", filepath])
            else:  # Linux
                subprocess.run(["xdg-open", filepath])
            return True
        except Exception as e:
            _log.warning("Error opening file: %s", e)
            return False
