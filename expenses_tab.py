from PyQt5 import QtWidgets, QtCore, QtGui
from pathlib import Path
from datetime import datetime, timezone
from decimal import Decimal
import json
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
from pathlib import Path
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from PyQt5.QtChart import (
    QChart, QChartView, QBarSet, QBarSeries,
    QBarCategoryAxis, QValueAxis, QPieSeries,
    QPieSlice, QLegend
)
from app_logger import get_logger
from app_theme import configure_filter_button
_log = get_logger(__name__)
# Add this function near the top of the file, after the imports
def format_amount_no_commas(amount):
    """Format amount without commas"""
    try:
        # Convert to integer if it's a whole number, otherwise keep 2 decimals
        amount_float = float(amount)
        if amount_float.is_integer():
            return f"${int(amount_float)}"
        else:
            return f"${amount_float:.2f}"
    except (ValueError, TypeError, AttributeError):
        return f"${amount}"
# Add this Currency class to expenses_tab.py
from decimal import Decimal, ROUND_HALF_UP

class Currency:
    """Currency formatting utilities - LOCAL COPY for expenses tab"""
    
    @staticmethod
    def format(value, symbol: str = "$") -> str:
        """Format decimal as currency"""
        return f"{symbol}{Currency.quantize(value)}"
    
    @staticmethod
    def format_whole(value, symbol: str = "$") -> str:
        """Format decimal as currency without decimal places and without thousands separators"""
        try:
            # Convert to integer (truncate decimal part)
            whole_value = int(float(value))
            # Return without commas: 12470310 not 12,470,310
            return f"{symbol}{whole_value}"
        except (ValueError, TypeError):
            # If conversion fails, fall back to regular format
            return Currency.format(value, symbol)
    
    @staticmethod
    def quantize(value) -> Decimal:
        """Quantize decimal to 2 decimal places"""
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

# Try to import Firebase modules
try:
    import firebase_admin
    from firebase_admin import credentials, db
    FIREBASE_AVAILABLE = True
except ImportError:
    FIREBASE_AVAILABLE = False

class ExpensesFirebaseManager:
    """Handles Firebase operations for expenses"""
    
    @staticmethod
    def save_expense(expense_data: dict) -> bool:
        """Save expense to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - expense not saved to Firebase")
            return False
            
        try:
            ref = db.reference('/expenses')

            # Create new expense reference first
            new_expense_ref = ref.push()

            firebase_id = new_expense_ref.key
            expense_data['firebase_id'] = firebase_id

            expense_data['created_at'] = datetime.now(timezone.utc).isoformat()
            expense_data['updated_at'] = datetime.now(timezone.utc).isoformat()

            new_expense_ref.set(expense_data)

            _log.info("Expense saved to Firebase with ID: %s", firebase_id)
            return True

        except Exception as e:
            _log.warning("Error saving expense to Firebase: %s", e)
            return False
    
    @staticmethod
    def load_expenses() -> list:
        """Load expenses from Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot load expenses")
            return []
            
        try:
            ref = db.reference('/expenses')
            expenses_data = ref.get()
            if expenses_data:
                expenses = []
                for expense_id, expense_data in expenses_data.items():
                    if expense_data:
                        expense_data['firebase_id'] = expense_id
                        expenses.append(expense_data)
                _log.info("Loaded %s expenses from Firebase", len(expenses))
                return expenses
            _log.info("No expenses found in Firebase")
            return []
        except Exception as e:
            _log.warning("Error loading expenses from Firebase: %s", e)
            return []

    @staticmethod
    def load_balance_sheet_expenses() -> list:
        """Load balance-sheet-only expenses so the Expenses tab can show a complete finance picture."""
        if not FIREBASE_AVAILABLE:
            return []
        try:
            ref = db.reference('/balance_sheet_expenses')
            expenses_data = ref.get()
            expenses = []
            if expenses_data:
                for expense_id, expense_data in expenses_data.items():
                    if isinstance(expense_data, dict):
                        expense_data = dict(expense_data)
                        expense_data['firebase_id'] = expense_id
                        expenses.append(expense_data)
            _log.info("Loaded %s balance sheet expense records for Expenses tab merge", len(expenses))
            return expenses
        except Exception as e:
            _log.warning("Error loading balance sheet expenses for merge: %s", e)
            return []
    
    @staticmethod
    def delete_expense(firebase_id: str) -> bool:
        """Delete expense from Firebase /expenses node"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot delete expense")
            return False
        try:
            db.reference(f'/expenses/{firebase_id}').delete()
            _log.info("Expense deleted from Firebase /expenses: %s", firebase_id)
            return True
        except Exception as e:
            _log.warning("Error deleting expense from Firebase: %s", e)
            return False

    @staticmethod
    def delete_balance_sheet_expense(firebase_id: str) -> bool:
        """Delete expense from Firebase /balance_sheet_expenses node"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot delete balance sheet expense")
            return False
        try:
            db.reference(f'/balance_sheet_expenses/{firebase_id}').delete()
            _log.info("Expense deleted from Firebase /balance_sheet_expenses: %s", firebase_id)
            return True
        except Exception as e:
            _log.warning("Error deleting balance sheet expense from Firebase: %s", e)
            return False

    # ===== NEW METHODS FOR CUSTOM CATEGORIES =====
    
    @staticmethod
    def save_custom_expense_type(category_name: str) -> bool:
        """Save custom expense type to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - category not saved")
            return False
            
        try:
            ref = db.reference('/custom_categories/expense_type')
            # Check if category already exists
            existing_categories = ref.get() or []
            if category_name not in existing_categories:
                existing_categories.append(category_name)
                ref.set(existing_categories)
                _log.info("Custom expense type saved: %s", category_name)
            return True
        except Exception as e:
            _log.warning("Error saving custom expense type: %s", e)
            return False
    
    @staticmethod
    def save_custom_category(expense_type: str, category_name: str) -> bool:
        """Save custom category to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - category not saved")
            return False
            
        try:
            ref = db.reference(f'/custom_categories/Categories/{expense_type}')
            existing_Categories = ref.get() or []
            if category_name not in existing_Categories:
                existing_Categories.append(category_name)
                ref.set(existing_Categories)
                _log.info("Custom category saved: %s under %s", category_name, expense_type)
            return True
        except Exception as e:
            _log.warning("Error saving custom category: %s", e)
            return False
    
    @staticmethod
    def save_custom_expense_name(category: str, expense_name: str) -> bool:
        """Save custom expense name to Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - expense name not saved")
            return False
            
        try:
            ref = db.reference(f'/custom_categories/expense_names/{category}')
            existing_expense_names = ref.get() or []
            if expense_name not in existing_expense_names:
                existing_expense_names.append(expense_name)
                ref.set(existing_expense_names)
                _log.info("Custom expense name saved: %s under %s", expense_name, category)
            return True
        except Exception as e:
            _log.warning("Error saving custom expense name: %s", e)
            return False
    
    @staticmethod
    def load_custom_categories() -> dict:
        """Load all custom categories from Firebase"""
        if not FIREBASE_AVAILABLE:
            _log.warning("Firebase not available - cannot load custom categories")
            return {}
            
        try:
            ref = db.reference('/custom_categories')
            custom_categories = ref.get() or {}
            _log.info("Loaded custom categories from Firebase")
            return custom_categories
        except Exception as e:
            _log.warning("Error loading custom categories: %s", e)
            return {}

    @staticmethod
    def save_vendor(vendor_name: str) -> bool:
        """Save a vendor name to Firebase for future dropdown use."""
        if not FIREBASE_AVAILABLE or not vendor_name:
            return False
        try:
            ref = db.reference('/vendors')
            existing = ref.get() or []
            if isinstance(existing, dict):
                existing = list(existing.values())
            if vendor_name not in existing:
                existing.append(vendor_name)
                ref.set(existing)
            return True
        except Exception as e:
            _log.warning("Error saving vendor: %s", e)
            return False

    @staticmethod
    def load_vendors() -> list:
        """Load saved vendor names from Firebase."""
        if not FIREBASE_AVAILABLE:
            return []
        try:
            ref = db.reference('/vendors')
            data = ref.get() or []
            if isinstance(data, dict):
                data = list(data.values())
            return sorted(str(v) for v in data if v)
        except Exception as e:
            _log.warning("Error loading vendors: %s", e)
            return []

class AddExpenseDialog(QtWidgets.QDialog):
    """Modal dialog for adding new expenses - Simplified categories with improved UX"""
    
    def __init__(self, parent=None, expense_data=None):
        super().__init__(parent)
        self.expense_data = expense_data
        self.is_editing = expense_data is not None
        
        # Add these flags to track balance sheet state
        self.was_in_balance_sheet = False  # Track if expense was in balance sheet
        self.initializing_form = True  # Track if we're initializing the form

        # Collect vendors from parent's loaded expenses + Firebase
        vendors_set = set()
        if parent and hasattr(parent, 'expenses'):
            for exp in (parent.expenses or []):
                v = str(exp.get('vendor', '') or '').strip()
                if v:
                    vendors_set.add(v)
        try:
            for v in ExpensesFirebaseManager.load_vendors():
                if v:
                    vendors_set.add(v)
        except Exception:
            pass
        self.vendors = sorted(vendors_set)

        title = "Edit Expense" if self.is_editing else "➕ Add New Expense"
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(950, 800)
        self.setStyleSheet("""
            QDialog {
                background: #f5f6fa;
            }
        """)
        
        # Initialize data structures for categories
        self.expense_type = [
            "O & M (Operations & Maintenance)",
            "Capital Expenses",
            "Other Expenses"
        ]
        
        # Categories for each expense type
        self.Categories = {
            "O & M (Operations & Maintenance)": [
                "Facilities & Utilities",
                "Office & Admin Overhead",
                "Engineering Software & IT",
                "Salaries, Labor & Related Costs",
                "Professional Services",
                "Insurance & Compliance",
                "Travel, Site Visits & Vehicles",
                "Marketing & Business Development",
                "Training, Licensure & Development",
                "Safety & Field Supplies",
                "Miscellaneous O & M"
            ],
            "Capital Expenses": [
                "Computer & Office Equipment",
                "Field & Inspection Equipment",
                "Furniture & Fixtures",
                "Vehicles",
                "Software (Capitalized)",
                "Leasehold Improvements",
                "Accumulated Depreciation"
            ],
            "Other Expenses": [
                "Other",
                "Salary/Bonuses",
                "Tax Expenses/Tax Deductions",
                "Medical/Benefits",
                "Meals & Entertainment",
                "Donations",
                "Bank Charges",
                "Contingency Funds",
                "Unexpected Costs"
            ]
        }
        
        # Expense names for each category
        self.expense_names = {
            "Other": [],
            "Facilities & Utilities": [
                "Office rent or co-working space fees",
                "Utilities (electricity, water, gas)",
                "Internet service",
                "Trash & cleaning services",
                "Property taxes (for office, if applicable)",
                "Office repairs & maintenance (HVAC, lights, minor repairs)"
            ],
            "Office & Admin Overhead": [
                "Office supplies (paper, pens, notebooks, printer ink)",
                "Printer/plotter maintenance & paper",
                "Postage & shipping (documents, contracts, samples)",
                "Bank fees & merchant processing fees",
                "Software: Microsoft 365 / Google Workspace",
                "Software: PDF tools (Bluebeam, Adobe, etc.)",
                "Software: Password manager",
                "Software: Others",
                "Cloud storage (Dropbox, Google Drive, OneDrive)"
            ],
            "Engineering Software & IT": [
                "Engineering software: SAP2000 / ETABS / STAAD / RAM / RISA",
                "Engineering software: Others",
                "CAD/BIM tools: AutoCAD, Civil 3D, Revit",
                "License/maintenance fees for all software",
                "IT support services",
                "Computer maintenance & small repairs",
                "Antivirus, backup services, security tools"
            ],
            "Salaries, Labor & Related Costs": [
                "Owner draw/salary",
                "Employee salaries & wages",
                "Overtime or temporary staff",
                "Payroll taxes paid by the company",
                "Employee benefits: Health insurance",
                "Employee benefits: Retirement plan contributions",
                "Employee benefits: Paid time off costs",
                "Payments to subcontract engineers, drafters"
            ],
            "Professional Services": [
                "Accounting & bookkeeping fees",
                "Tax preparation and consulting",
                "Legal services (contracts, company setup)",
                "Business consulting or coaching services",
                "Registered agent fees (if applicable)"
            ],
            "Insurance & Compliance": [
                "Professional liability / Errors & Omissions (E&O) insurance",
                "General liability insurance",
                "Business owner's policy (BOP)",
                "Workers' comp insurance",
                "Commercial auto insurance",
                "License renewals (PE license, SE license)",
                "Business license renewals",
                "Memberships"
            ],
            "Travel, Site Visits & Vehicles": [
                "Mileage (personal vehicle for business)",
                "Fuel costs (company vehicles)",
                "Parking fees & tolls",
                "Vehicle maintenance",
                "Airfare, hotels for out-of-town site visits",
                "Rental cars or rideshare for business trips",
                "Meals while traveling for business"
            ],
            "Marketing & Business Development": [
                "Website hosting and domain expenses",
                "Website maintenance & updates",
                "Graphic design (logo, templates, brochures)",
                "Online ads (Google, LinkedIn, Facebook)",
                "Printing of business cards, brochures, banners",
                "Sponsorships of events",
                "Client entertainment (dinners, coffee meetings)"
            ],
            "Training, Licensure & Development": [
                "Continuing education (PDH hours, webinars)",
                "Training courses (technical or business)",
                "Books, codes, and standards",
                "Exam fees for additional licenses"
            ],
            "Safety & Field Supplies": [
                "PPE: hard hats, safety vests, glasses, gloves, boots",
                "Field tools for inspections",
                "Calibration of field instruments",
                "First-aid kits and safety equipment"
            ],
            "Miscellaneous O & M": [
                "Subscriptions: LinkedIn Premium",
                "Subscriptions: Industry journals",
                "Project management tools",
                "Document management tools or e-signature services"
            ],
            "Computer & Office Equipment": [
                "Laptops",
                "Desktops",
                "Monitors",
                "Printers/Scanners",
                "Servers",
                "Networking Equipment"
            ],
            "Field & Inspection Equipment": [
                "Survey Equipment",
                "Testing Equipment",
                "Measurement Tools",
                "Safety Equipment",
                "Inspection Devices"
            ],
            "Furniture & Fixtures": [
                "Office Desks",
                "Chairs",
                "Filing Cabinets",
                "Shelving Units",
                "Conference Room Furniture"
            ],
            "Vehicles": [
                "Company Cars",
                "Trucks",
                "Vans",
                "Heavy Equipment",
                "Vehicle Accessories"
            ],
            "Software (Capitalized)": [
                "Engineering Software License",
                "ERP System",
                "CRM System",
                "Database Software",
                "Custom Software Development"
            ],
            "Leasehold Improvements": [
                "Office Renovations",
                "Electrical Work",
                "Plumbing Improvements",
                "HVAC Installation",
                "Security Systems"
            ],
            "Accumulated Depreciation": [
                "Depreciation Expense - Computers",
                "Depreciation Expense - Office Equipment",
                "Depreciation Expense - Vehicles",
                "Accumulated Depreciation"
            ],
            "Salary/Bonuses": [
                "Employee Salary",
                "Manager Salary",
                "Executive Salary",
                "Performance Bonus",
                "Year-end Bonus",
                "Commission Payments",
                "Incentive Payments"
            ],
            "Tax Expenses/Tax Deductions": [
                "Federal Income Tax",
                "Tax Deduction",
                "Payroll Tax",
                "Sales Tax",
                "Property Tax",
                "Business Tax"
            ],
            "Medical/Benefits": [
                "Health Insurance Premiums",
                "Dental Insurance",
                "Vision Insurance",
                "Retirement Contributions",
                "Life Insurance",
                "Disability Insurance",
                "Wellness Programs"
            ],
            "Meals & Entertainment": [
                "Client Meals",
                "Business Lunches",
                "Team Dinners",
                "Conference Meals",
                "Entertainment Expenses",
                "Team Building Events"
            ],
            "Donations": [
                "Charitable Donations",
                "Community Sponsorships",
                "Educational Donations",
                "Non-profit Contributions",
                "Event Sponsorships"
            ],
            "Bank Charges": [
                "Monthly Account Fees",
                "Transaction Fees",
                "Wire Transfer Fees",
                "Credit Card Processing Fees",
                "Check Printing Fees",
                "Overdraft Fees"
            ],
            "Contingency Funds": [
                "Emergency Funds",
                "Reserve Funds",
                "Project Contingency",
                "Operational Reserve",
                "Risk Management Fund"
            ],
            "Unexpected Costs": [
                "Emergency Repairs",
                "Unplanned Maintenance",
                "Price Increases",
                "Regulatory Changes",
                "Market Fluctuations"
            ]
        }
        
        self.init_ui()
        
        # Check if this expense was in balance sheet BEFORE populating
        if self.is_editing:
            self.was_in_balance_sheet = self.check_if_in_balance_sheet()
            self.populate_form_data()
        else:
            # For new expenses, show checkbox and check it by default
            self.checkbox_container.setVisible(True)
            self.balance_sheet_checkbox.setChecked(True)
        
        # Set initial focus
        if not self.is_editing:
            self.expense_date_edit.setFocus()
            # Select all text in date field
            line_edit = self.expense_date_edit.lineEdit()
            if line_edit:
                line_edit.selectAll()
        
        self.initializing_form = False
        # Evaluate button state now that all fields are populated
        QtCore.QTimer.singleShot(0, self._update_save_btn_state)
        # Snapshot for unsaved-changes detection (taken after form is fully populated)
        self._initial_snapshot = self._form_snapshot()

    def _form_snapshot(self):
        """Return a dict of current field values for change detection."""
        return {
            "date":         getattr(self, 'expense_date_edit', None) and self.expense_date_edit.date().toString("MM-dd-yyyy"),
            "type":         getattr(self, 'expense_type_combo', None) and self.expense_type_combo.currentText().strip(),
            "category":     getattr(self, 'Category_combo', None) and self.Category_combo.currentText().strip(),
            "name":         getattr(self, 'expense_name_combo', None) and self.expense_name_combo.currentText().strip(),
            "vendor":       getattr(self, 'vendor_combo', None) and self.vendor_combo.currentText().strip(),
            "description":  getattr(self, 'description_edit', None) and self.description_edit.text().strip(),
            "amount":       getattr(self, 'amount_edit', None) and self.amount_edit.text().strip(),
            "project":      getattr(self, 'project_combo', None) and self.project_combo.currentText().strip(),
        }

    def _has_unsaved_changes(self):
        return self._form_snapshot() != self._initial_snapshot

    def reject(self):
        if self._has_unsaved_changes():
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes.\nAre you sure you want to discard them and close?",
                QtWidgets.QMessageBox.Discard | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel,
            )
            if reply != QtWidgets.QMessageBox.Discard:
                return
        super().reject()

    def closeEvent(self, event):
        if self._has_unsaved_changes():
            reply = QtWidgets.QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved changes.\nAre you sure you want to discard them and close?",
                QtWidgets.QMessageBox.Discard | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel,
            )
            if reply != QtWidgets.QMessageBox.Discard:
                event.ignore()
                return
        event.accept()

    def _active_project_options(self):
        """Return active projects as (display text, project number) pairs."""
        inactive_statuses = {
            "Completed & Invoiced",
            "Paid",
            "Cancelled",
            "Cancel",
            "Completed",
        }
        projects = []
        parent = self.parent()
        main_window = getattr(parent, "main_window", None)
        project_tab = getattr(main_window, "project_tab", None)

        for attr in ("cached_projects", "generated_projects"):
            projects.extend(getattr(project_tab, attr, []) or [])

        if not projects:
            try:
                from main import FirebaseManager
                projects = FirebaseManager.load_projects() or []
            except Exception as exc:
                _log.warning("Could not load active projects for expense dialog: %s", exc)
                projects = []

        seen = set()
        options = [("General / Overhead", "")]
        for project in projects:
            if not isinstance(project, dict):
                continue
            status = str(project.get("status", "") or "").strip()
            if status in inactive_statuses:
                continue
            project_number = str(project.get("project_number", "") or "").strip()
            if not project_number or project_number in seen:
                continue
            seen.add(project_number)
            project_name = str(project.get("project_name", "") or "").strip()
            display = project_number if not project_name else f"{project_number} - {project_name}"
            options.append((display, project_number))
        return options

    def _populate_project_combo(self):
        self.project_combo.clear()
        for display, project_number in self._active_project_options():
            self.project_combo.addItem(display, project_number)
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # ===== Header (Matching Job Form style) =====
        header = QtWidgets.QFrame()
        header.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                color: white;
                padding: 12px 24px;
            }
        """)
        header_layout = QtWidgets.QVBoxLayout(header)
        header_layout.setSpacing(2)

        title = QtWidgets.QLabel("Edit Expense" if self.is_editing else "➕ Add New Expense")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: white;")
        subtitle = QtWidgets.QLabel("Update expense details carefully." if self.is_editing else "Enter all expense details carefully.")
        subtitle.setStyleSheet("font-size: 12px; color: #ecf0f1;")

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        layout.addWidget(header)

        # ===== Scrollable Content =====
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll_widget = QtWidgets.QWidget()
        scroll.setWidget(scroll_widget)

        form_layout = QtWidgets.QVBoxLayout(scroll_widget)
        form_layout.setContentsMargins(50, 30, 50, 30)
        form_layout.setSpacing(25)

        # ===== Basic Information =====
        self.add_section_title(form_layout, "📝 Basic Information")
        
        # Expense Date
        self.expense_date_edit = self.create_fixed_date_edit(QtCore.QDate.currentDate())
        self.expense_date_edit.setCalendarPopup(True)
        self.expense_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.expense_date_edit.setReadOnly(False)
        
        date_line_edit = self.expense_date_edit.lineEdit()
        if date_line_edit:
            date_line_edit.setPlaceholderText("MM-DD-YYYY")
        
        self.expense_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; background: #f8f9fa; }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #3498db;
                color: white;
            }
            QCalendarWidget QToolButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
            }
            QCalendarWidget QMenu {
                background-color: white;
            }
        """)
        self.add_field(form_layout, "📅 Expense Date:", self.expense_date_edit)

        # Main Expense Type
        self.expense_type_combo = self.create_styled_combo_box(
            self.expense_type,
            "Select expense type *"
        )
        self.expense_type_combo.currentTextChanged.connect(self.on_expense_type_changed)
        self.expense_type_combo.currentTextChanged.connect(self._update_save_btn_state)
        self.add_field(form_layout, "📂 Expense Type *:", self.expense_type_combo)

        # Category
        self.Category_combo = self.create_styled_combo_box(
            [],
            "Select or enter Category *"
        )
        self.Category_combo.currentTextChanged.connect(self.on_Category_changed)
        self.Category_combo.currentTextChanged.connect(self._update_save_btn_state)
        self.add_field(form_layout, "📋 Category *:", self.Category_combo)

        # Expense Name/Description
        self.expense_name_combo = self.create_styled_combo_box(
            [],
            "Select or enter specific expense name *"
        )
        self.expense_name_combo.currentTextChanged.connect(self._update_save_btn_state)
        self.add_field(form_layout, "📝 Expense Name *:", self.expense_name_combo)

        # Vendor (editable dropdown — saves new vendors automatically)
        self.vendor_combo = self.create_styled_combo_box(
            self.vendors,
            "Type or select vendor/supplier"
        )
        self.add_field(form_layout, "🏢 Vendor/Supplier:", self.vendor_combo)

        # Description
        self.description_edit = QtWidgets.QLineEdit()
        self.description_edit.setPlaceholderText("Enter additional expense description...")
        self.description_edit.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QLineEdit:focus { border-color: #3498db; background: #f8f9fa; }
        """)
        self.add_field(form_layout, "📝 Additional Description:", self.description_edit)

        # ===== Financial Information =====
        self.add_section_title(form_layout, "💰 Financial Information")

        # Amount
        self.amount_edit = self.create_styled_line_edit("$0.00")
        self.add_field(form_layout, "💰 Amount *:", self.amount_edit)

        self.amount_edit.textChanged.connect(self.validate_amount_input)
        self.amount_edit.textChanged.connect(self._update_save_btn_state)

        # Project/Client
        self.project_combo = self.create_styled_combo_box([], "Enter or select project")
        self._populate_project_combo()
        self.add_field(form_layout, "🎯 Project:", self.project_combo)

        # ===== Actions Header with Balance Sheet Checkbox (same line) =====
        actions_header_layout = QtWidgets.QHBoxLayout()
        actions_header_layout.setContentsMargins(0, 10, 0, 10)

        # Left side: Actions title
        actions_title = QtWidgets.QLabel("🚀 Actions")
        actions_title.setStyleSheet("""
            QLabel {
                font-weight: bold;
                font-size: 16px;
                color: #2c3e50;
                border-bottom: 2px solid #dfe6e9;
                padding-bottom: 6px;
            }
        """)
        actions_header_layout.addWidget(actions_title)

        # Add stretch to push checkbox to the right
        actions_header_layout.addStretch()

        # Right side: Balance Sheet Checkbox container
        self.checkbox_container = QtWidgets.QWidget()
        checkbox_layout = QtWidgets.QHBoxLayout(self.checkbox_container)
        checkbox_layout.setContentsMargins(0, 0, 0, 0)

        self.balance_sheet_checkbox = QtWidgets.QCheckBox("✓ Save this expense to Balance Sheet")
        self.balance_sheet_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 12px;
                font-weight: bold;
                color: #2c3e50;
                spacing: 8px;
                padding: 8px 12px;
                background-color: #e8f4fd;
                border: 1px solid #3498db;
                border-radius: 6px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #bdc3c7;
                background: white;
                border-radius: 4px;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #27ae60;
                background: #27ae60;
                border-radius: 4px;
                image: url(data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='white'%3E%3Cpath d='M20 6L9 17l-5-5-1.5 1.5L9 20 21 8z'/%3E%3C/svg%3E);
            }
            QCheckBox:hover {
                background-color: #d4e6f1;
                border-color: #2980b9;
            }
        """)

        checkbox_layout.addWidget(self.balance_sheet_checkbox)
        actions_header_layout.addWidget(self.checkbox_container)

        # Add the header layout to form
        form_layout.addLayout(actions_header_layout)

        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.setSpacing(20)

        self.save_btn = QtWidgets.QPushButton("💾 Update Expense" if self.is_editing else "💾 Save Expense")
        self.save_btn.setMinimumHeight(48)
        self.save_btn.setEnabled(self.is_editing)
        if self.is_editing:
            self.save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    font-weight: bold;
                    font-size: 14px;
                    border-radius: 8px;
                    padding: 10px 20px;
                }
                QPushButton:hover { background-color: #34ce57; }
            """)
        else:
            self.save_btn.setStyleSheet("""
                QPushButton {
                    background-color: #95a5a6;
                    color: white;
                    font-weight: bold;
                    font-size: 14px;
                    border-radius: 8px;
                    padding: 10px 20px;
                }
                QPushButton:hover { background-color: #95a5a6; }
            """)
        self.save_btn.clicked.connect(self.accept)
        self.save_btn.setAutoDefault(False)
        self.save_btn.setDefault(False)

        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setMinimumHeight(48)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                padding: 10px 20px;
            }
            QPushButton:hover { background: #5a6268; }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        self.cancel_btn.setAutoDefault(False)

        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.cancel_btn)
        form_layout.addLayout(btn_layout)

        layout.addWidget(scroll)
        
        # Set up Enter key navigation and auto-scrolling
        self.setup_enter_key_navigation()
        
        # Setup date field UX for manual entry
        self.setup_date_field_ux()
        self._patch_date_enter()
        for w in [
            self.expense_date_edit,
            self.expense_type_combo,
            self.Category_combo,
            self.expense_name_combo,
            self.vendor_combo,
            self.description_edit,
            self.amount_edit,
            self.project_combo,
        ]:
            self._install_arrow_navigation(w)

    # Fixed base styles for required-field highlighting — avoids fragile string replacement
    _COMBO_STYLE_OK  = ("QComboBox { padding:10px 12px; border:1px solid #bdc3c7; border-radius:6px;"
                        " background:white; font-size:13px; }"
                        " QComboBox:focus { border-color:#3498db; background:#f8f9fa; }"
                        " QComboBox QAbstractItemView { selection-background-color:#3498db; }")
    _COMBO_STYLE_ERR = ("QComboBox { padding:10px 12px; border:2px solid #e74c3c; border-radius:6px;"
                        " background:#fff8f8; font-size:13px; }"
                        " QComboBox:focus { border-color:#e74c3c; background:#fff8f8; }"
                        " QComboBox QAbstractItemView { selection-background-color:#3498db; }")
    _AMOUNT_STYLE_OK  = ("QLineEdit { padding:10px 12px; border:1px solid #bdc3c7; border-radius:6px;"
                         " background:white; font-size:13px; }"
                         " QLineEdit:focus { border-color:#3498db; background:#f8f9fa; }")
    _AMOUNT_STYLE_ERR = ("QLineEdit { padding:10px 12px; border:2px solid #e74c3c; border-radius:6px;"
                         " background:#fff8f8; font-size:13px; }"
                         " QLineEdit:focus { border-color:#e74c3c; background:#fff8f8; }")

    def _update_save_btn_state(self, *_):
        """Enable the save/update button only when all required fields have values."""
        if self.initializing_form:
            return

        # In edit mode the existing data is already valid — keep button always enabled.
        if self.is_editing:
            if hasattr(self, 'save_btn'):
                self.save_btn.setEnabled(True)
                self.save_btn.setStyleSheet(
                    "QPushButton { background-color:#28a745; color:white; font-weight:bold;"
                    " font-size:14px; border-radius:8px; padding:10px 20px; }"
                    " QPushButton:hover { background-color:#34ce57; }"
                )
            return

        expense_type = getattr(self, 'expense_type_combo', None) and self.expense_type_combo.currentText().strip()
        category     = getattr(self, 'Category_combo', None) and self.Category_combo.currentText().strip()
        expense_name = getattr(self, 'expense_name_combo', None) and self.expense_name_combo.currentText().strip()
        amount_raw   = getattr(self, 'amount_edit', None) and self.amount_edit.text().replace("$", "").strip()
        try:
            amount_ok = float(amount_raw) > 0 if amount_raw else False
        except ValueError:
            amount_ok = False
        all_ok = bool(expense_type and category and expense_name and amount_ok)

        if hasattr(self, 'save_btn'):
            self.save_btn.setEnabled(all_ok)
            bg, hover = ("#28a745", "#34ce57") if all_ok else ("#95a5a6", "#95a5a6")
            self.save_btn.setStyleSheet(f"""
                QPushButton {{ background-color:{bg}; color:white; font-weight:bold;
                               font-size:14px; border-radius:8px; padding:10px 20px; }}
                QPushButton:hover {{ background-color:{hover}; }}
            """)

        # Highlight empty required combo fields with red border
        for attr, val in (("expense_type_combo", expense_type),
                          ("Category_combo",      category),
                          ("expense_name_combo",  expense_name)):
            w = getattr(self, attr, None)
            if w:
                w.setStyleSheet(self._COMBO_STYLE_ERR if not val else self._COMBO_STYLE_OK)

        # Highlight empty/invalid amount field
        if hasattr(self, 'amount_edit'):
            self.amount_edit.setStyleSheet(
                self._AMOUNT_STYLE_ERR if not amount_ok else self._AMOUNT_STYLE_OK
            )

    def validate_amount_input(self):
        """Validate amount input to accept only numbers and auto-add $ prefix"""
        self.amount_edit.blockSignals(True)
        try:
            text = self.amount_edit.text().strip()
            cursor_pos = self.amount_edit.cursorPosition()

            if not text:
                self.amount_edit.blockSignals(False)
                return

            # Remove everything except digits and dot
            cleaned = ''.join(c for c in text if c.isdigit() or c == '.')
            cleaned = cleaned.replace('$', '')

            # Allow only one decimal
            if cleaned.count('.') > 1:
                parts = cleaned.split('.')
                cleaned = parts[0] + '.' + ''.join(parts[1:])

            final_text = f"${cleaned}" if cleaned else ""

            if final_text != text:
                self.amount_edit.setText(final_text)
                new_pos = min(cursor_pos + 1, len(final_text)) if not text.startswith("$") else cursor_pos
                self.amount_edit.setCursorPosition(new_pos)

        finally:
            self.amount_edit.blockSignals(False)

    def setup_enter_key_navigation(self):
        """Enter key moves focus to next field for ALL widgets (including editable combo boxes)"""

        self.input_widgets = [
            self.expense_date_edit,
            self.expense_type_combo,
            self.Category_combo,
            self.expense_name_combo,
            self.vendor_combo,
            self.description_edit,
            self.amount_edit,
            self.project_combo,
            self.save_btn,
            self.cancel_btn
        ]

        # Install event filters
        for widget in self.input_widgets:
            if not widget:
                continue

            widget.installEventFilter(self)

            if isinstance(widget, QtWidgets.QDateEdit):
                if widget.lineEdit():
                    widget.lineEdit().installEventFilter(self)
            elif isinstance(widget, QtWidgets.QDoubleSpinBox):
                if widget.lineEdit():
                    widget.lineEdit().installEventFilter(self)
            elif isinstance(widget, QtWidgets.QComboBox):
                if widget.lineEdit():
                    widget.lineEdit().installEventFilter(self)

        # Set correct tab order
        for i in range(len(self.input_widgets) - 1):
            w1 = self.input_widgets[i]
            w2 = self.input_widgets[i + 1]
            if w1 and w2:
                QtWidgets.QWidget.setTabOrder(w1, w2)

        # Disable default button behavior
        self.save_btn.setAutoDefault(False)
        self.save_btn.setDefault(False)
        self.cancel_btn.setAutoDefault(False)
        self.cancel_btn.setDefault(False)

    def create_fixed_date_edit(self, date=None):
        """Create QDateEdit with NO scroll, NO arrow increment, NO auto change"""
        d = QtWidgets.QDateEdit(date if date else QtCore.QDate.currentDate())
        d.setCalendarPopup(True)
        d.setDisplayFormat("MM-dd-yyyy")
        d.setReadOnly(False)

        # ❌ Disable mouse wheel
        d.wheelEvent = lambda event: None

        # ❌ Disable arrow keys increment
        def keyPressEvent(event, original=d.keyPressEvent):
            if event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
                return
            original(event)
        d.keyPressEvent = keyPressEvent

        # ❌ Disable internal stepping (MOST IMPORTANT)
        d.stepBy = lambda x: None

        # ❌ Remove spin buttons
        d.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

        # Keep your styling
        d.setStyleSheet("""
            QDateEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; background: #f8f9fa; }
        """)

        return d

    def _patch_date_enter(self):
        le = self.expense_date_edit.lineEdit()
        if not le:
            return

        original = le.keyPressEvent

        def keyPressEvent(event):
            if event.key() in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
                # Send TAB to dialog
                QtWidgets.QApplication.sendEvent(
                    self,
                    QtGui.QKeyEvent(
                        QtCore.QEvent.KeyPress,
                        QtCore.Qt.Key_Tab,
                        QtCore.Qt.NoModifier
                    )
                )
                return
            original(event)

        le.keyPressEvent = keyPressEvent

    def _install_arrow_navigation(self, widget):
        if not widget:
            return

        target = widget
        if isinstance(widget, (QtWidgets.QComboBox, QtWidgets.QDateEdit, QtWidgets.QDoubleSpinBox)):
            target = widget.lineEdit()

        if not target:
            return

        original = target.keyPressEvent

        def keyPressEvent(event):
            if event.key() == QtCore.Qt.Key_Down:
                self._move_focus(direction=1)
                return
            if event.key() == QtCore.Qt.Key_Up:
                self._move_focus(direction=-1)
                return
            original(event)

        target.keyPressEvent = keyPressEvent

    def keyPressEvent(self, event):
        key = event.key()
        modifiers = event.modifiers()

        # Ctrl + S → Save
        if key == QtCore.Qt.Key_S and modifiers & QtCore.Qt.ControlModifier:
            if self.save_btn.isEnabled():
                self.save_btn.click()
            return

        # Enter → forward
        if key in (QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter):
            self._move_focus(direction=1)
            return

        # Down Arrow → forward
        if key == QtCore.Qt.Key_Down:
            self._move_focus(direction=1)
            return

        # Up Arrow → backward
        if key == QtCore.Qt.Key_Up:
            self._move_focus(direction=-1)
            return

        super().keyPressEvent(event)

    def setup_date_field_ux(self):
        """Setup better UX for date fields for manual entry"""
        if hasattr(self, 'expense_date_edit') and self.expense_date_edit:
            self.expense_date_edit.setReadOnly(False)
            line_edit = self.expense_date_edit.lineEdit()
            if line_edit:
                line_edit.setPlaceholderText("MM-DD-YYYY")
                line_edit.installEventFilter(self)
                
                original_keyPressEvent = line_edit.keyPressEvent
                def custom_keyPressEvent(event, le=line_edit, w=self.expense_date_edit):
                    if event.key() in [QtCore.Qt.Key_Return, QtCore.Qt.Key_Enter]:
                        self.move_to_next_widget(w)
                        return
                    original_keyPressEvent(event)
                line_edit.keyPressEvent = lambda event, le=line_edit, w=self.expense_date_edit: custom_keyPressEvent(event)
                
                line_edit.textEdited.connect(
                    lambda text, de=self.expense_date_edit: self.validate_date_input(de, text)
                )
    
    def validate_date_input(self, date_edit, text):
        """Validate date input as user types"""
        if not text:
            return
        
        # Clean the text - remove any non-digit characters except dashes and slashes
        cleaned = ''.join(c for c in text if c.isdigit() or c in ['-', '/'])
        
        # Auto-format as user types
        if len(cleaned) >= 2 and '-' not in cleaned and '/' not in cleaned:
            # Add first separator after month
            if len(cleaned) == 2:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))
        
        elif len(cleaned) >= 5 and cleaned[2] in ['-', '/'] and cleaned[5:] == '':
            # Add second separator after day
            if len(cleaned) == 5:
                formatted = f"{cleaned}-"
                date_edit.lineEdit().setText(formatted)
                date_edit.lineEdit().setCursorPosition(len(formatted))
    
    def eventFilter(self, source, event):
        """Handle Enter key navigation with proper focus handling for date fields and backspace support"""
        if isinstance(source, QtWidgets.QLineEdit):
            parent = source.parent()
            if parent and isinstance(parent, QtWidgets.QDateEdit) and parent == self.expense_date_edit:
                
                if event.type() == QtCore.QEvent.FocusIn:
                    QtCore.QTimer.singleShot(10, lambda s=source: s.selectAll())
                    return False
                
                if event.type() == QtCore.QEvent.KeyPress:
                    if event.key() in [QtCore.Qt.Key_Backspace, QtCore.Qt.Key_Delete]:
                        if source.hasSelectedText():
                            source.clear()
                            return True
                    
                    elif event.key() in [
                        QtCore.Qt.Key_0, QtCore.Qt.Key_1, QtCore.Qt.Key_2, QtCore.Qt.Key_3,
                        QtCore.Qt.Key_4, QtCore.Qt.Key_5, QtCore.Qt.Key_6, QtCore.Qt.Key_7,
                        QtCore.Qt.Key_8, QtCore.Qt.Key_9, QtCore.Qt.Key_Minus, QtCore.Qt.Key_Slash
                    ]:
                        if source.hasSelectedText():
                            source.clear()
                        return False
        
        return super().eventFilter(source, event)
    
    def move_to_next_widget(self, current_widget):
        """Move focus to the next widget (combo-aware)"""
        if isinstance(current_widget, QtWidgets.QLineEdit):
            parent = current_widget.parent()
            if isinstance(parent, QtWidgets.QComboBox):
                current_widget = parent
            elif isinstance(parent, QtWidgets.QDateEdit):
                current_widget = parent
            elif isinstance(parent, QtWidgets.QDoubleSpinBox):
                current_widget = parent

        try:
            current_index = self.input_widgets.index(current_widget)
        except ValueError:
            return

        for i in range(current_index + 1, len(self.input_widgets)):
            next_widget = self.input_widgets[i]

            if next_widget in (self.save_btn, self.cancel_btn):
                continue

            if next_widget.isEnabled() and next_widget.isVisible():
                next_widget.setFocus()

                if isinstance(next_widget, QtWidgets.QComboBox):
                    if next_widget.lineEdit():
                        QtCore.QTimer.singleShot(10, next_widget.lineEdit().selectAll)
                elif isinstance(next_widget, QtWidgets.QLineEdit):
                    QtCore.QTimer.singleShot(10, next_widget.selectAll)
                elif isinstance(next_widget, QtWidgets.QDoubleSpinBox):
                    QtCore.QTimer.singleShot(10, next_widget.lineEdit().selectAll)
                elif isinstance(next_widget, QtWidgets.QDateEdit):
                    QtCore.QTimer.singleShot(10, next_widget.lineEdit().selectAll)

                self.ensureWidgetVisible(next_widget)
                return
    
    def ensureWidgetVisible(self, widget):
        """Ensure the widget is visible in the scroll area"""
        try:
            scroll_area = None
            current_widget = widget
            
            while current_widget:
                if isinstance(current_widget, QtWidgets.QScrollArea):
                    scroll_area = current_widget
                    break
                current_widget = current_widget.parent()
            
            if scroll_area and widget:
                scroll_area.ensureWidgetVisible(widget)
                
        except Exception as e:
            _log.warning("Scroll error: %s", e)
    
    def on_expense_type_changed(self, expense_type):
        """Update Category dropdown when expense type changes"""
        self.Category_combo.clear()
        
        if expense_type in self.Categories:
            subcats = self.Categories[expense_type]
            self.Category_combo.addItems(subcats)
        
        self.Category_combo.setCurrentIndex(-1)
        self.Category_combo.lineEdit().clear()
        self.Category_combo.lineEdit().setPlaceholderText("Select or enter Category *")
        
        # Clear expense name dropdown
        self.expense_name_combo.clear()
        self.expense_name_combo.setCurrentIndex(-1)
        self.expense_name_combo.lineEdit().clear()
        self.expense_name_combo.lineEdit().setPlaceholderText("Select or enter specific expense name *")
    
    def on_Category_changed(self, Category):
        """Update expense name dropdown when Category changes"""
        self.expense_name_combo.clear()
        
        if Category in self.expense_names:
            expense_names = self.expense_names[Category]
            self.expense_name_combo.addItems(expense_names)
        
        self.expense_name_combo.setCurrentIndex(-1)
        self.expense_name_combo.lineEdit().clear()
        self.expense_name_combo.lineEdit().setPlaceholderText("Select or enter specific expense name *")
    
    def _move_focus(self, direction=1):
        widgets = self._get_focus_chain()
        fw = self.focusWidget()

        if isinstance(fw, QtWidgets.QLineEdit):
            parent = fw.parent()
            if isinstance(parent, (QtWidgets.QComboBox, QtWidgets.QDateEdit, QtWidgets.QDoubleSpinBox)):
                fw = parent

        if fw not in widgets:
            widgets[0].setFocus()
            self.ensureWidgetVisible(widgets[0])
            return

        idx = widgets.index(fw)
        next_idx = idx + direction

        if next_idx >= len(widgets):
            widgets[0].setFocus()
            self.ensureWidgetVisible(widgets[0])
            return

        if next_idx < 0:
            widgets[0].setFocus()
            self.ensureWidgetVisible(widgets[0])
            return

        target = widgets[next_idx]
        target.setFocus()

        if isinstance(target, QtWidgets.QComboBox) and target.lineEdit():
            target.lineEdit().selectAll()
        elif isinstance(target, QtWidgets.QLineEdit):
            target.selectAll()
        elif isinstance(target, QtWidgets.QDoubleSpinBox):
            target.lineEdit().selectAll()
        elif isinstance(target, QtWidgets.QDateEdit):
            target.lineEdit().selectAll()

        self.ensureWidgetVisible(target)

    def _get_focus_chain(self):
        return [
            self.expense_date_edit,
            self.expense_type_combo,
            self.Category_combo,
            self.expense_name_combo,
            self.vendor_combo,
            self.description_edit,
            self.amount_edit,
            self.project_combo,
            self.save_btn,
            self.cancel_btn
        ]

    def add_section_title(self, layout, text):
        """Add a section title to the form"""
        label = QtWidgets.QLabel(text)
        label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                font-size: 16px;
                color: #2c3e50;
                border-bottom: 2px solid #dfe6e9;
                padding-bottom: 6px;
            }
        """)
        layout.addWidget(label)
    
    def add_field(self, layout, label_text, widget):
        """Add a form field with label and widget"""
        field_layout = QtWidgets.QHBoxLayout()
        label = QtWidgets.QLabel(label_text)
        label.setStyleSheet("font-weight: 500; color: #2c3e50; min-width: 150px;")
        field_layout.addWidget(label)
        field_layout.addWidget(widget, 1)
        layout.addLayout(field_layout)
    
    def create_styled_combo_box(self, items, placeholder=""):
        """Create editable combo box with placeholder behavior like job form"""
        combo = QtWidgets.QComboBox()
        combo.addItems(items)
        combo.setEditable(True)
        # Prevent accidental value changes via scroll wheel (issue #12)
        combo.wheelEvent = lambda e: e.ignore()
        _orig_key = combo.keyPressEvent
        def _no_scroll_keys(e, _ok=_orig_key, _cb=combo):
            if e.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) and not _cb.view().isVisible():
                e.ignore()
                return
            _ok(e)
        combo.keyPressEvent = _no_scroll_keys

        line_edit = combo.lineEdit()

        def on_user_type(text, le=line_edit, cb=combo):
            if le.text().startswith("-- Select"):
                clean = le.text().replace("-- Select Client --", "")
                clean = clean.replace("-- Select Job Type --", "").strip()
                le.blockSignals(True)
                le.setText(clean)
                le.blockSignals(False)

        line_edit.textEdited.connect(on_user_type)

        def on_focus_in(event, le=line_edit):
            if le.text().startswith("-- Select"):
                le.blockSignals(True)
                le.clear()
                le.blockSignals(False)
            return QtWidgets.QLineEdit.focusInEvent(le, event)

        line_edit.focusInEvent = on_focus_in

        line_edit.setPlaceholderText(placeholder)

        line_edit.textEdited.connect(lambda t: line_edit.setPlaceholderText(""))

        old_mouse = line_edit.mousePressEvent
        def new_mouse(event):
            line_edit.setPlaceholderText("")
            old_mouse(event)

        line_edit.mousePressEvent = new_mouse

        combo.setCurrentIndex(-1)
        line_edit.clear()

        combo.setStyleSheet("""
            QComboBox {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QComboBox:focus { border-color: #3498db; background: #f8f9fa; }
            QComboBox QAbstractItemView {
                selection-background-color: #3498db;
            }
        """)
        return combo
    
    def populate_form_data(self):
        """Populate form with existing expense data for editing"""
        if not self.expense_data:
            return
            
        # Set basic information
        date_str = self.expense_data.get('date', '')
        if date_str:
            try:
                date = QtCore.QDate.fromString(date_str, "MM-dd-yyyy")
                if date.isValid():
                    self.expense_date_edit.setDate(date)
            except:
                try:
                    date = QtCore.QDate.fromString(date_str, "MMMM d, yyyy")
                    if date.isValid():
                        self.expense_date_edit.setDate(date)
                except:
                    self.expense_date_edit.setDate(QtCore.QDate.currentDate())
        
        # Populate expense type
        expense_type = self.expense_data.get('expense_type', '')
        if expense_type:
            index = self.expense_type_combo.findText(expense_type)
            if index >= 0:
                self.expense_type_combo.setCurrentIndex(index)
            else:
                self.expense_type_combo.setEditText(expense_type)
        else:
            self.expense_type_combo.setEditText("Other Expenses")
        
        # Populate Category
        Category = self.expense_data.get('Category', '')
        if Category:
            if expense_type:
                self.on_expense_type_changed(expense_type)
            
            index = self.Category_combo.findText(Category)
            if index >= 0:
                self.Category_combo.setCurrentIndex(index)
            else:
                self.Category_combo.setEditText(Category)
        
        # Populate expense name
        expense_name = self.expense_data.get('expense_name', '')
        if expense_name:
            if Category:
                self.on_Category_changed(Category)
            
            index = self.expense_name_combo.findText(expense_name)
            if index >= 0:
                self.expense_name_combo.setCurrentIndex(index)
            else:
                self.expense_name_combo.setEditText(expense_name)
            
        vendor = self.expense_data.get('vendor', '')
        if vendor:
            index = self.vendor_combo.findText(vendor)
            if index >= 0:
                self.vendor_combo.setCurrentIndex(index)
            else:
                self.vendor_combo.setEditText(vendor)
        self.description_edit.setText(self.expense_data.get('description', ''))

        # Set financial information
        amount = self.expense_data.get('amount', 0)
        self.amount_edit.setText(f"${amount:.2f}" if amount else "$0.00")

        project = self.expense_data.get('project', '')
        if project:
            index = self.project_combo.findData(project)
            if index < 0:
                index = self.project_combo.findText(project)
            if index >= 0:
                self.project_combo.setCurrentIndex(index)
            else:
                self.project_combo.setEditText(project)
        
        # ===== CRITICAL FIX: Set checkbox visibility and state =====
        # Check if this expense was previously saved to balance sheet
        self.was_in_balance_sheet = self.check_if_in_balance_sheet()
        
        if self.was_in_balance_sheet:
            # Show checkbox and check it
            self.checkbox_container.setVisible(True)
            self.balance_sheet_checkbox.setChecked(True)
        else:
            # Hide checkbox completely
            self.checkbox_container.setVisible(False)
            # Checkbox state doesn't matter when hidden
    
    def check_if_in_balance_sheet(self):
        """Check if this expense exists in balance sheet expenses data"""
        if not self.expense_data:
            return False
        
        # Try to find the balance sheet tab from the parent
        parent = self.parent()
        main_window = None
        
        while parent:
            if hasattr(parent, 'balance_sheet_tab'):
                main_window = parent
                break
            parent = parent.parent()
        
        if not main_window or not hasattr(main_window, 'balance_sheet_tab'):
            return False
        
        balance_tab = main_window.balance_sheet_tab
        if not balance_tab or not hasattr(balance_tab, 'expenses_data'):
            return False
        
        # Check if there's a matching expense entry
        for expense in balance_tab.expenses_data:
            if expense.get('firebase_id') == self.expense_data.get('firebase_id'):
                return True
                
        return False
    
    def create_styled_line_edit(self, placeholder=""):
        """Create styled QLineEdit (same UX as Job Form)"""
        line_edit = QtWidgets.QLineEdit()
        line_edit.setPlaceholderText(placeholder)

        line_edit.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 1px solid #bdc3c7;
                border-radius: 6px;
                background: white;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)

        return line_edit

    def get_expense_data(self):
        """Return expense data from form with validation"""
        expense_type = self.expense_type_combo.currentText().strip()
        Category = self.Category_combo.currentText().strip()
        expense_name = self.expense_name_combo.currentText().strip()
        amount_raw = self.amount_edit.text().replace("$", "").strip()

        # Validate all required fields
        missing = []
        if not expense_type:
            missing.append("Expense Type")
        if not Category:
            missing.append("Category")
        if not expense_name:
            missing.append("Expense Name")
        try:
            if not amount_raw or float(amount_raw) <= 0:
                missing.append("Amount (must be > 0)")
        except ValueError:
            missing.append("Amount (invalid number)")

        if missing:
            QtWidgets.QMessageBox.warning(
                self, "Required Fields Missing",
                "The following required fields must be filled:\n\n• " + "\n• ".join(missing)
            )
            self._update_save_btn_state()
            return None

        stored_date = self.expense_date_edit.date().toString("MM-dd-yyyy")

        amount_text = self.amount_edit.text().replace("$", "").strip()

        selected_project = self.project_combo.currentData()
        if selected_project is None:
            selected_project = self.project_combo.currentText().strip()

        expense_data = {
            "date": stored_date,
            "expense_type": expense_type,
            "Category": Category,
            "expense_name": expense_name,
            "vendor": self.vendor_combo.currentText().strip(),
            "description": self.description_edit.text(),
            "amount": float(amount_text) if amount_text else 0.0,
            "project": selected_project,
            "project_number": selected_project,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "save_to_balance_sheet": self.balance_sheet_checkbox.isChecked()
        }

        expense_data["type"] = Category

        if self.is_editing and self.expense_data and 'firebase_id' in self.expense_data:
            expense_data['firebase_id'] = self.expense_data['firebase_id']

        return expense_data

    def accept(self):
        """Override accept to validate required fields"""
        expense_data = self.get_expense_data()
        if expense_data is not None:
            super().accept()
            
class CategoryExpenseDialog(QtWidgets.QDialog):
    """Dialog to show category-specific expenses"""
    
    def __init__(self, category_name, expenses, parent=None):
        super().__init__(parent)
        self.category_name = category_name
        self.expenses = expenses
        self.setWindowTitle(f"📊 {category_name} Expenses")
        self.setModal(True)
        self.resize(900, 600)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QtWidgets.QLabel(f"📊 {self.category_name} - Expense Details")
        header.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                border-radius: 10px;
                color: white;
                text-align: center;
            }
        """)
        layout.addWidget(header)
        
        # Summary card
        total_amount = sum(expense.get('amount', 0) for expense in self.expenses)
        count = len(self.expenses)
        
        summary_frame = QtWidgets.QFrame()
        summary_frame.setStyleSheet("""
            QFrame {
                background: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        summary_layout = QtWidgets.QHBoxLayout(summary_frame)
        
        total_label = QtWidgets.QLabel(f"💰 Total: ${total_amount:,.2f}")
        total_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #27ae60;")
        
        count_label = QtWidgets.QLabel(f"📋 Count: {count} expenses")
        count_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #3498db;")
        
        summary_layout.addWidget(total_label)
        summary_layout.addStretch()
        summary_layout.addWidget(count_label)
        
        layout.addWidget(summary_frame)
        
        # Table
        table_frame = QtWidgets.QFrame()
        table_layout = QtWidgets.QVBoxLayout(table_frame)
        
        self.expenses_table = QtWidgets.QTableWidget()
        self.expenses_table.setColumnCount(7)
        self.expenses_table.setHorizontalHeaderLabels([
            "Date", "Vendor", "Description", "Amount",
            "Project", "Method", "Status"
        ])
        
        # Table styling
        self.expenses_table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                gridline-color: #e1e8ed;
                font-size: 9px;
            }
            QTableWidget::item {
                padding: 8px 4px;
                border-bottom: 1px solid #f8f9fa;
                border-right: 1px solid #e1e8ed;
                color: #2c3e50;
            }
            QTableWidget::item:selected {
                background: #e3f2fd;
                color: #2c3e50;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2c3e50, stop:1 #34495e);
                color: white;
                font-weight: bold;
                font-size: 9px;
                padding: 8px 6px;
                border: none;
                border-right: 1px solid #3a506b;
            }
        """)
        
        self.expenses_table.setShowGrid(True)
        self.expenses_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.expenses_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.expenses_table.setAlternatingRowColors(True)
        self.expenses_table.verticalHeader().setVisible(False)
        
        # Set column widths
        header = self.expenses_table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Interactive)  # Vendor
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)  # Description
        self.expenses_table.setColumnWidth(1, 200)
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)  # Amount
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)  # Project
        header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)  # Method
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)  # Status
        
        self.populate_table()
        table_layout.addWidget(self.expenses_table)
        layout.addWidget(table_frame)
        
        # Close button
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background: #6c757d;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 6px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #5a6268;
            }
        """)
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
    
    def populate_table(self):
        """Populate table with category expenses"""
        self.expenses_table.setRowCount(len(self.expenses))
        
        for row, expense in enumerate(reversed(self.expenses)):
            # Date
            date_item = QtWidgets.QTableWidgetItem(expense.get('date', ''))
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 0, date_item)
            
            # Vendor
            vendor_item = QtWidgets.QTableWidgetItem(expense.get('vendor', ''))
            self.expenses_table.setItem(row, 1, vendor_item)
            
            # Description
            desc_item = QtWidgets.QTableWidgetItem(expense.get('description', ''))
            self.expenses_table.setItem(row, 2, desc_item)
            
            # Amount
            amount = expense.get('amount', 0)
            # Use format_whole instead of format to remove commas
            amount_display = format_amount_no_commas(amount) if hasattr(Currency, 'format_whole') else format_amount_no_commas(amount)
            amount_item = QtWidgets.QTableWidgetItem(amount_display)
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 7, amount_item)
            # Project
            project_item = QtWidgets.QTableWidgetItem(expense.get('project', ''))
            project_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 4, project_item)
            
            # Payment Method
            method_item = QtWidgets.QTableWidgetItem(expense.get('payment_method', ''))
            method_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 5, method_item)
            
            # Status
            status_item = QtWidgets.QTableWidgetItem(expense.get('status', ''))
            status_item.setTextAlignment(QtCore.Qt.AlignCenter)
            
            # Color code status
            status = expense.get('status', '')
            if 'Approved' in status:
                status_item.setForeground(QtGui.QColor('#27ae60'))
            elif 'Pending' in status:
                status_item.setForeground(QtGui.QColor('#f39c12'))
            elif 'Denied' in status:
                status_item.setForeground(QtGui.QColor('#e74c3c'))
            elif 'Reimbursed' in status:
                status_item.setForeground(QtGui.QColor('#3498db'))
                
            self.expenses_table.setItem(row, 6, status_item)

class YearPickerDialog(QtWidgets.QDialog):
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.setWindowTitle("Select Year")
        self.resize(420, 320)
        self.current_year = current_year or datetime.now().year

        self.start_year = (self.current_year // 12) * 12    # decade block start

        layout = QtWidgets.QVBoxLayout(self)

        # ===== Header with arrows =====
        header = QtWidgets.QHBoxLayout()

        self.prev_btn = QtWidgets.QPushButton("←")
        self.prev_btn.setFixedSize(40, 30)
        self.prev_btn.clicked.connect(self.prev_range)

        self.next_btn = QtWidgets.QPushButton("→")
        self.next_btn.setFixedSize(40, 30)
        self.next_btn.clicked.connect(self.next_range)

        self.title_lbl = QtWidgets.QLabel("", alignment=QtCore.Qt.AlignCenter)
        self.title_lbl.setStyleSheet("font-size: 18px; font-weight: bold;")

        header.addWidget(self.prev_btn)
        header.addWidget(self.title_lbl, 1)
        header.addWidget(self.next_btn)

        layout.addLayout(header)

        # ===== Year Grid (3x4) =====
        self.grid_widget = QtWidgets.QWidget()
        self.grid = QtWidgets.QGridLayout(self.grid_widget)
        self.grid.setSpacing(12)
        layout.addWidget(self.grid_widget)

        # Build first view
        self.build_grid()

    def build_grid(self):
        # Clear previous buttons
        for i in reversed(range(self.grid.count())):
            w = self.grid.itemAt(i).widget()
            if w:
                w.deleteLater()

        # Generate 12 continuous years
        years = list(range(self.start_year, self.start_year + 12))
        self.title_lbl.setText(f"{self.start_year} - {self.start_year + 11}")

        r, c = 0, 0
        for year in years:
            btn = QtWidgets.QPushButton(str(year))
            btn.setFixedSize(110, 55)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {"#d1f5e0" if year == self.current_year else "#ffffff"};
                    border: 1px solid #cccccc;
                    border-radius: 8px;
                    font-size: 17px;
                    color: #333333;
                    font-weight: {"bold" if year == self.current_year else "500"};
                }}
                QPushButton:hover {{
                    background: #eef7ff;
                    border: 1px solid #7ab8ff;
                }}
            """)
            btn.clicked.connect(lambda _, y=year: self.select_year(y))

            self.grid.addWidget(btn, r, c)

            c += 1
            if c == 3:
                c = 0
                r += 1


    def prev_range(self):
        self.start_year -= 12
        self.build_grid()

    def next_range(self):
        self.start_year += 12
        self.build_grid()

    def select_year(self, year):
        self.selected_year = year
        self.accept()

class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year  # Minimum year (1 AD)
        self.end_year = end_year      # Maximum year (9999 AD)
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Navigation buttons
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton("◀◀")
        self.prev_block_btn.setFixedSize(40, 30)
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)
        
        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)
        
        self.next_block_btn = QtWidgets.QPushButton("▶▶")
        self.next_block_btn.setFixedSize(40, 30)
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        
        layout.addLayout(nav_layout)
        
        # Year grid container
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create 3x3 grid of year buttons
        self.year_buttons = []
        
        # Calculate current 9-year block start
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        # Current year display
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        
        layout.addLayout(current_layout)
        
        # Update the grid
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        """Calculate which 9-year block a year belongs to"""
        # Formula: ((year - 1) // 9) * 9 + 1
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        """Update the 3x3 grid with years from current 9-year block"""
        # Generate 9 consecutive years starting from current_block_start
        years = []
        
        for i in range(9):
            year = self.current_block_start + i
            years.append(year)
        
        # Update block label
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        
        # Update button texts and styles
        current_year = datetime.now().year
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            
            # Check if year is within valid range (1-9999)
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            # Style based on selection and current year
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #2980b9;
                    }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            # Connect button click
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        """Select a year"""
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        """Go to previous 9-year block (unlimited past)"""
        self.current_block_start -= 9
        
        # Unlimited past - no lower bound check
        # If we go below year 1, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        """Go to next 9-year block (unlimited future)"""
        self.current_block_start += 9
        
        # Unlimited future - no upper bound check
        # If we go above year 9999, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        """Set the selected year"""
        # Ensure year is within valid range
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        
    def get_selected_year(self):
        """Get the selected year"""
        return self.selected_year
    
    # Add signal for year selection
    year_selected = QtCore.pyqtSignal(int)

class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection with unlimited years"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📅 Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        # Create YearCalendarGrid with unlimited years
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        # Selected year display
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        
        self.selected_label = QtWidgets.QLabel(f"")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        
        layout.addLayout(selected_layout)
        
        # Action buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Connect signals
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        """Update selected year display when year is changed in calendar"""
        self.current_year = year
    
    def on_select_clicked(self):
        """Emit signal with selected year and close popup"""
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        """Get the selected year"""
        return self.current_year
    
class _WheelBlocker(QtCore.QObject):
    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Wheel:
            return True
        return False

class _NoScrollComboBox(QtWidgets.QComboBox):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._wb = _WheelBlocker(self)
        self.installEventFilter(self._wb)

class _NoScrollDateEdit(QtWidgets.QDateEdit):
    def stepBy(self, steps):
        pass


class ExpensesExportDialog(QtWidgets.QDialog):
    """Professional PDF/Excel Export Dialog for Expenses with Tabs"""
    
    def __init__(self, parent=None, available_dates=None):
        super().__init__(parent)
        self.available_dates = available_dates or []
        self.export_range = "all"  # Default export range
        self.selected_dates = []
        self.export_type = "pdf"  # Default export type
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("📊 Export Expenses")
        self.setFixedSize(700, 750)
        self.setStyleSheet("""
            ExpensesExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📤 Export Manager - Expenses")
        header.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 10px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Export Type Tabs
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #d5dbdb;
            }
        """)
        
        # PDF Export Tab
        self.pdf_tab = QtWidgets.QWidget()
        self.setup_pdf_tab()
        self.tab_widget.addTab(self.pdf_tab, "📄 PDF Export")
        
        # Excel Export Tab
        self.excel_tab = QtWidgets.QWidget()
        self.setup_excel_tab()
        self.tab_widget.addTab(self.excel_tab, "📊 Excel Export")
        
        layout.addWidget(self.tab_widget)
        
        # Connect tab change signal
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
        # Progress Bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
    
    def setup_pdf_tab(self):
        """Setup the PDF export tab"""
        layout = QtWidgets.QVBoxLayout(self.pdf_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Expenses")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)

        # From date section
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(160, 45)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)

        # To date section
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(160, 45)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)

        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month Selector
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)

        # Month and Year selection in one row
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(24)

        month_col = QtWidgets.QVBoxLayout()
        month_col.setSpacing(6)
        month_label = QtWidgets.QLabel("Month")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_col.addWidget(month_label)
        self.month_combo = _NoScrollComboBox()
        self.month_combo.setFixedHeight(42)
        self.month_combo.setMinimumWidth(160)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months()
        month_col.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_col)

        year_container_col = QtWidgets.QVBoxLayout()
        year_container_col.setSpacing(6)
        year_label_month = QtWidgets.QLabel("Year")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container_col.addWidget(year_label_month)

        year_field_row = QtWidgets.QHBoxLayout()
        year_field_row.setSpacing(6)

        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn_month.setFixedHeight(45)
        self.year_calendar_btn_month.setMinimumWidth(72)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_field_row.addWidget(self.year_edit_month)
        year_field_row.addWidget(self.year_calendar_btn_month)
        year_container_col.addLayout(year_field_row)
        month_year_row_layout.addLayout(year_container_col)
        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        # Year Selector
        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        # Year selection row
        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Year")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)

        # Year field
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.year_calendar_btn.setFixedHeight(45)
        self.year_calendar_btn.setMinimumWidth(72)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)

        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        # Initially hide all date selection components
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        # Preview Section
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel("Ready to export all expenses as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)
        
        # Connect signals for live preview updates
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.month_combo.currentTextChanged.connect(self.update_preview)
    
    def setup_excel_tab(self):
        """Setup the Excel export tab"""
        layout = QtWidgets.QVBoxLayout(self.excel_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 Excel Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.excel_all_radio = QtWidgets.QRadioButton("📋 Export All Expenses")
        self.excel_all_radio.setChecked(True)
        self.excel_all_radio.toggled.connect(lambda: self.on_excel_range_changed("all"))
        
        self.excel_date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.excel_date_range_radio.toggled.connect(lambda: self.on_excel_range_changed("date_range"))
        
        self.excel_month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.excel_month_radio.toggled.connect(lambda: self.on_excel_range_changed("month"))
        
        self.excel_year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.excel_year_radio.toggled.connect(lambda: self.on_excel_range_changed("year"))
        
        options_layout.addWidget(self.excel_all_radio)
        options_layout.addWidget(self.excel_date_range_radio)
        options_layout.addWidget(self.excel_month_radio)
        options_layout.addWidget(self.excel_year_radio)
        
        range_group.addButton(self.excel_all_radio)
        range_group.addButton(self.excel_date_range_radio)
        range_group.addButton(self.excel_month_radio)
        range_group.addButton(self.excel_year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container for Excel
        self.excel_date_selection_container = QtWidgets.QWidget()
        self.excel_date_selection_layout = QtWidgets.QVBoxLayout(self.excel_date_selection_container)
        self.excel_date_selection_layout.setSpacing(15)
        self.excel_date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector for Excel
        self.excel_date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.excel_date_range_group.setMinimumHeight(120)
        self.excel_date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_date_range_layout = QtWidgets.QHBoxLayout(self.excel_date_range_group)
        excel_date_range_layout.setSpacing(20)

        # From date section
        excel_from_layout = QtWidgets.QVBoxLayout()
        excel_from_label = QtWidgets.QLabel("From Date:")
        excel_from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_from_layout.addWidget(excel_from_label)
        self.excel_from_date = _NoScrollDateEdit()
        self.excel_from_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.excel_from_date.setCalendarPopup(True)
        self.excel_from_date.setFixedSize(160, 45)
        self.excel_from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_from_layout.addWidget(self.excel_from_date)
        excel_date_range_layout.addLayout(excel_from_layout)

        # To date section
        excel_to_layout = QtWidgets.QVBoxLayout()
        excel_to_label = QtWidgets.QLabel("To Date:")
        excel_to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_to_layout.addWidget(excel_to_label)
        self.excel_to_date = _NoScrollDateEdit()
        self.excel_to_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_to_date.setDate(QtCore.QDate.currentDate())
        self.excel_to_date.setCalendarPopup(True)
        self.excel_to_date.setFixedSize(160, 45)
        self.excel_to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_to_layout.addWidget(self.excel_to_date)
        excel_date_range_layout.addLayout(excel_to_layout)

        excel_date_range_layout.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_date_range_group)

        # Month Selector for Excel
        self.excel_month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.excel_month_group.setMinimumHeight(150)
        self.excel_month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_month_layout = QtWidgets.QVBoxLayout(self.excel_month_group)
        excel_month_layout.setSpacing(15)

        # Month and Year selection in one row
        excel_month_year_row_layout = QtWidgets.QHBoxLayout()
        excel_month_year_row_layout.setSpacing(24)

        excel_month_col = QtWidgets.QVBoxLayout()
        excel_month_col.setSpacing(6)
        excel_month_label = QtWidgets.QLabel("Month")
        excel_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_month_col.addWidget(excel_month_label)
        self.excel_month_combo = _NoScrollComboBox()
        self.excel_month_combo.setFixedHeight(42)
        self.excel_month_combo.setMinimumWidth(160)
        self.excel_month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months_excel()
        excel_month_col.addWidget(self.excel_month_combo)
        excel_month_year_row_layout.addLayout(excel_month_col)

        excel_year_container_col = QtWidgets.QVBoxLayout()
        excel_year_container_col.setSpacing(6)
        excel_year_month_label = QtWidgets.QLabel("Year")
        excel_year_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_container_col.addWidget(excel_year_month_label)

        excel_year_field_row = QtWidgets.QHBoxLayout()
        excel_year_field_row.setSpacing(6)

        self.excel_year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit_month.setFixedSize(150, 45)
        self.excel_year_edit_month.setReadOnly(True)
        self.excel_year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        self.excel_year_calendar_btn_month = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn_month.setFixedHeight(45)
        self.excel_year_calendar_btn_month.setMinimumWidth(72)
        self.excel_year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month_excel)

        excel_year_field_row.addWidget(self.excel_year_edit_month)
        excel_year_field_row.addWidget(self.excel_year_calendar_btn_month)
        excel_year_container_col.addLayout(excel_year_field_row)
        excel_month_year_row_layout.addLayout(excel_year_container_col)
        excel_month_year_row_layout.addStretch()
        excel_month_layout.addLayout(excel_month_year_row_layout)
        self.excel_date_selection_layout.addWidget(self.excel_month_group)

        # Year Selector for Excel
        self.excel_year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.excel_year_group.setMinimumHeight(120)
        self.excel_year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_year_layout = QtWidgets.QVBoxLayout(self.excel_year_group)
        excel_year_layout.setSpacing(15)

        # Year selection row
        excel_year_row_layout = QtWidgets.QHBoxLayout()
        excel_year_label = QtWidgets.QLabel("Year")
        excel_year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_row_layout.addWidget(excel_year_label)

        # Year field
        self.excel_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit.setFixedSize(150, 45)
        self.excel_year_edit.setReadOnly(True)
        self.excel_year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.excel_year_calendar_btn = QtWidgets.QPushButton("▼ Year")
        self.excel_year_calendar_btn.setFixedHeight(45)
        self.excel_year_calendar_btn.setMinimumWidth(72)
        self.excel_year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 13px;
                font-weight: 700;
                padding: 0 12px;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn.clicked.connect(self.show_year_popup_excel)
        
        excel_year_row_layout.addWidget(self.excel_year_edit)
        excel_year_row_layout.addWidget(self.excel_year_calendar_btn)
        excel_year_row_layout.addStretch()
        excel_year_layout.addLayout(excel_year_row_layout)

        self.excel_date_selection_layout.addWidget(self.excel_year_group)

        layout.addWidget(self.excel_date_selection_container)

        # Initially hide all date selection components for Excel
        self.excel_date_selection_container.setVisible(False)
        self.excel_date_range_group.setVisible(False)
        self.excel_month_group.setVisible(False)
        self.excel_year_group.setVisible(False)
        
        # Preview Section for Excel
        excel_preview_card = QtWidgets.QGroupBox("👁️ Excel Export Preview")
        excel_preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        excel_preview_layout = QtWidgets.QVBoxLayout(excel_preview_card)
        
        self.excel_preview_label = QtWidgets.QLabel("Ready to export all expenses as Excel")
        self.excel_preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.excel_preview_label.setWordWrap(True)
        excel_preview_layout.addWidget(self.excel_preview_label)
        
        layout.addWidget(excel_preview_card)
        
        # Connect signals for live preview updates for Excel
        self.excel_from_date.dateChanged.connect(self.update_excel_preview)
        self.excel_to_date.dateChanged.connect(self.update_excel_preview)
        self.excel_month_combo.currentTextChanged.connect(self.update_excel_preview)
    
    def show_year_popup(self):
        """Show year calendar popup for PDF year selection"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        # Center the popup
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show year calendar popup for PDF month+year selection"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_excel(self):
        """Show year calendar popup for Excel year selection"""
        try:
            current_year = int(self.excel_year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month_excel(self):
        """Show year calendar popup for Excel month+year selection"""
        try:
            current_year = int(self.excel_year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for PDF year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for PDF month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_excel(self, year):
        """Handle year selection from popup for Excel year export"""
        self.excel_year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def on_year_selected_for_month_excel(self, year):
        """Handle year selection from popup for Excel month+year export"""
        self.excel_year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def populate_months(self):
        """Populate months combo box for PDF"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)

    def populate_months_excel(self):
        """Populate months combo box for Excel"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.excel_month_combo.addItems(months)
        self.excel_month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_tab_changed(self, index):
        """Handle tab changes"""
        if index == 0:  # PDF tab
            self.export_type = "pdf"
            self.export_btn.setText("🚀 Export PDF")
            self.update_preview()
        elif index == 1:  # Excel tab
            self.export_type = "excel"
            self.export_btn.setText("🚀 Export Excel")
            self.update_excel_preview()
    
    def on_range_changed(self, range_type):
        """Handle export range changes for PDF"""
        self.export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_preview()

    def on_excel_range_changed(self, range_type):
        """Handle export range changes for Excel"""
        self.excel_export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.excel_date_range_group.setVisible(date_range_visible)
        self.excel_month_group.setVisible(month_visible)
        self.excel_year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.excel_date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_excel_preview()
    
    def update_preview(self):
        """Update the PDF preview text"""
        if self.export_range == "all":
            self.preview_label.setText("📋 Will export ALL expenses as PDF")
        
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toString("MM/dd/yyyy")
            to_date = self.to_date.date().toString("MM/dd/yyyy")
            self.preview_label.setText(f"📅 Will export expenses from {from_date} to {to_date} as PDF")
        
        elif self.export_range == "month":
            month = self.month_combo.currentText()
            year = self.year_edit_month.text()
            self.preview_label.setText(f"🗓️ Will export expenses for {month} {year} as PDF")
        
        elif self.export_range == "year":
            year = self.year_edit.text()
            self.preview_label.setText(f"📊 Will export expenses for the year {year} as PDF")

    def update_excel_preview(self):
        """Update the Excel preview text"""
        if hasattr(self, 'excel_export_range'):
            range_type = self.excel_export_range
        else:
            range_type = "all"
        
        if range_type == "all":
            self.excel_preview_label.setText("📋 Will export ALL expenses as Excel")
        
        elif range_type == "date_range":
            from_date = self.excel_from_date.date().toString("MM/dd/yyyy")
            to_date = self.excel_to_date.date().toString("MM/dd/yyyy")
            self.excel_preview_label.setText(f"📅 Will export expenses from {from_date} to {to_date} as Excel")
        
        elif range_type == "month":
            month = self.excel_month_combo.currentText()
            year = self.excel_year_edit_month.text()
            self.excel_preview_label.setText(f"🗓️ Will export expenses for {month} {year} as Excel")
        
        elif range_type == "year":
            year = self.excel_year_edit.text()
            self.excel_preview_label.setText(f"📊 Will export expenses for the year {year} as Excel")
    
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_type == "pdf":
            if self.export_range == "all":
                return {"range": "all", "type": "pdf"}
            
            elif self.export_range == "date_range":
                from_date = self.from_date.date().toPyDate()
                to_date = self.to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "pdf"}
            
            elif self.export_range == "month":
                month = self.month_combo.currentIndex() + 1
                year = int(self.year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "pdf"}
            
            elif self.export_range == "year":
                year = int(self.year_edit.text())
                return {"range": "year", "year": year, "type": "pdf"}
        
        elif self.export_type == "excel":
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            
            if range_type == "all":
                return {"range": "all", "type": "excel"}
            
            elif range_type == "date_range":
                from_date = self.excel_from_date.date().toPyDate()
                to_date = self.excel_to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "excel"}
            
            elif range_type == "month":
                month = self.excel_month_combo.currentIndex() + 1
                year = int(self.excel_year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "excel"}
            
            elif range_type == "year":
                year = int(self.excel_year_edit.text())
                return {"range": "year", "year": year, "type": "excel"}
    
    def start_export(self):
        """Start the export process based on selected type"""
        # Prevent multiple executions
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
            
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            # Simulate export process
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):  # Check if still valid
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            # Store export parameters for parent to use after dialog closes
            self._export_params = export_params
            
            # Simply accept the dialog - let parent handle the actual export
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False
            
# Replace your existing PDFExportDialog class with this one (which uses the YearCalendarPopup)
class PDFExportDialog(QtWidgets.QDialog):
    """Professional PDF Export Dialog for Expenses - Now with YearCalendarPopup"""
    
    def __init__(self, parent=None, available_dates=None):
        super().__init__(parent)
        self.available_dates = available_dates or []
        self.export_range = "all"  # Default export range
        self.selected_dates = []
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("📊 Export Expenses")
        self.setFixedSize(700, 600)
        self.setStyleSheet("""
            PDFExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📤 Export Expenses")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 3px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 1px;
                text-align: center;
            }
        """)
        header.setFixedHeight(70)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_card.setMaximumHeight(250)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Expenses")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)

        # From date section
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = _NoScrollDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(160, 45)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)

        # To date section
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = _NoScrollDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(160, 45)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)

        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month Selector
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)

        # Month and Year selection in one row
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(15)

        # Month selection
        month_container = QtWidgets.QHBoxLayout()
        month_label = QtWidgets.QLabel("Select Month:")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_container.addWidget(month_label)
        self.month_combo = QtWidgets.QComboBox()
        self.month_combo.setFixedSize(200, 45)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months()
        month_container.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_container)

        # Year selection for month export
        year_container = QtWidgets.QHBoxLayout()
        year_label_month = QtWidgets.QLabel("Select Year:")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container.addWidget(year_label_month)

        # Year field with calendar button
        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.year_calendar_btn_month.setFixedSize(50, 45)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_container.addWidget(self.year_edit_month)
        year_container.addWidget(self.year_calendar_btn_month)
        month_year_row_layout.addLayout(year_container)

        # Add stretch to push everything to the left
        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        # Year Selector
        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        # Year selection row
        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Select Year:")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)
        
        # Year field with calendar button
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        
        # Calendar button
        self.year_calendar_btn = QtWidgets.QPushButton("📅")
        self.year_calendar_btn.setFixedSize(50, 45)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)

        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        # Initially hide all date selection components
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        # Preview Section
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_card.setMaximumHeight(150)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel("Ready to export all expenses as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)
        
        # Progress Bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
        
        # Connect signals for live preview updates
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.month_combo.currentTextChanged.connect(self.update_preview)
    
    def show_year_popup(self):
        """Show separate popup window for year selection (year export)"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        # Center the popup relative to main dialog
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        # Show as separate window
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show separate popup window for year selection (month+year export)"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        # Center the popup relative to main dialog
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        # Show as separate window
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def populate_months(self):
        """Populate months combo box"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_range_changed(self, range_type):
        """Handle export range changes"""
        self.export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_preview()
    
    def update_preview(self):
        """Update the PDF preview text with MM-dd-yyyy format"""
        if self.export_range == "all":
            self.preview_label.setText("📋 Will export ALL expenses as PDF")
        
        elif self.export_range == "date_range":
            # Use MM-dd-yyyy format
            from_date = self.from_date.date().toString("MM-dd-yyyy")
            to_date = self.to_date.date().toString("MM-dd-yyyy")
            self.preview_label.setText(f"📅 Will export expenses from {from_date} to {to_date} as PDF")
        
        elif self.export_range == "month":
            month = self.month_combo.currentText()
            year = self.year_edit_month.text()
            self.preview_label.setText(f"🗓️ Will export expenses for {month} {year} as PDF")
        
        elif self.export_range == "year":
            year = self.year_edit.text()
            self.preview_label.setText(f"📊 Will export expenses for the year {year} as PDF")
            
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_range == "all":
            return {"range": "all"}
        
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toPyDate()
            to_date = self.to_date.date().toPyDate()
            return {"range": "date_range", "from_date": from_date, "to_date": to_date}
        
        elif self.export_range == "month":
            month = self.month_combo.currentIndex() + 1
            year = int(self.year_edit_month.text())
            return {"range": "month", "month": month, "year": year}
        
        elif self.export_range == "year":
            year = int(self.year_edit.text())
            return {"range": "year", "year": year}
    
    def start_export(self):
        """Start the export process"""
        # Prevent multiple executions
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
            
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            # Simulate export process
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):  # Check if still valid
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            # Store export parameters for parent to use after dialog closes
            self._export_params = export_params
            
            # Simply accept the dialog - let parent handle the actual export
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False
            

class MonthYearPickerDialog(QtWidgets.QDialog):
    """A popup dialog to pick month & year with a grid layout."""
    
    MONTH_NAMES = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    def __init__(self, parent=None, selected_year=None, selected_month=None):
        super().__init__(parent)
        self.setWindowTitle("Select Month & Year")
        self.resize(420, 370)

        self.selected_year = selected_year or datetime.now().year
        self.selected_month = selected_month or datetime.now().month

        layout = QtWidgets.QVBoxLayout(self)

        # ===== HEADER WITH LEFT / RIGHT ARROWS =====
        header = QtWidgets.QHBoxLayout()

        prev_btn = QtWidgets.QPushButton("←")
        prev_btn.setFixedSize(40, 30)
        prev_btn.clicked.connect(self.prev_year)

        next_btn = QtWidgets.QPushButton("→")
        next_btn.setFixedSize(40, 30)
        next_btn.clicked.connect(self.next_year)

        self.title_lbl = QtWidgets.QLabel(str(self.selected_year))
        self.title_lbl.setAlignment(QtCore.Qt.AlignCenter)
        self.title_lbl.setStyleSheet("font-size: 18px; font-weight: bold;")

        header.addWidget(prev_btn)
        header.addWidget(self.title_lbl, 1)
        header.addWidget(next_btn)

        layout.addLayout(header)

        # ===== MONTH GRID (12 MONTHS) =====
        self.grid_widget = QtWidgets.QWidget()
        self.grid = QtWidgets.QGridLayout(self.grid_widget)
        self.grid.setSpacing(12)
        layout.addWidget(self.grid_widget)

        self.build_grid()

    def build_grid(self):
        # Clear old buttons
        for i in reversed(range(self.grid.count())):
            w = self.grid.itemAt(i).widget()
            if w:
                w.deleteLater()

        r, c = 0, 0
        for i, name in enumerate(self.MONTH_NAMES):
            month_number = i + 1

            btn = QtWidgets.QPushButton(name)
            btn.setFixedSize(120, 45)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {"#d1f5e0" if month_number == self.selected_month else "#ffffff"};
                    border: 1px solid #cccccc;
                    border-radius: 8px;
                    font-size: 15px;
                    color: #333;
                    font-weight: {"bold" if month_number == self.selected_month else "500"};
                }}
                QPushButton:hover {{
                    background: #eef7ff;
                    border: 1px solid #7ab8ff;
                }}
            """)
            btn.clicked.connect(lambda _, m=month_number: self.select_month(m))

            self.grid.addWidget(btn, r, c)

            c += 1
            if c == 3:
                c = 0
                r += 1

    def prev_year(self):
        self.selected_year -= 1
        self.title_lbl.setText(str(self.selected_year))

    def next_year(self):
        self.selected_year += 1
        self.title_lbl.setText(str(self.selected_year))

    def select_month(self, m):
        self.selected_month = m
        self.accept()


class PieChartView(QChartView):
    """Custom pie chart view with hover tooltips"""
    def __init__(self, chart, parent=None):
        super().__init__(chart, parent)
        self.setMouseTracking(True)
        self.pie_series = None
        self.pie_slice_categories = {}
        self.current_tooltip = None

    def mouseMoveEvent(self, event):
        """Show tooltip on hover"""
        if self.pie_series and self.pie_slice_categories:
            slices = self.pie_series.slices()
            for idx, slice_obj in enumerate(slices):
                if idx < len(self.pie_slice_categories):
                    category = self.pie_slice_categories[idx]
                    amount = slice_obj.value()
                    tooltip_text = f"{category} - ${amount:,.2f}"
                    QtWidgets.QToolTip.showText(event.globalPos(), tooltip_text, self)
                    return

        QtWidgets.QToolTip.hideText()
        return super().mouseMoveEvent(event)


class SliceTooltip(QtWidgets.QWidget):
    """Clean tooltip for pie slices"""
    def __init__(self, text, parent=None):
        super().__init__(parent)
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint | QtCore.Qt.ToolTip)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(0)

        label = QtWidgets.QLabel(text)
        label.setWordWrap(False)
        label.setStyleSheet("""
            QLabel {
                background-color: #ffffff;
                color: #1e293b;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 8px 12px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                font-size: 13px;
                font-weight: 500;
            }
        """)
        layout.addWidget(label)
        self.adjustSize()


class ExpensesTab(QtWidgets.QWidget):
    """Professional Expenses Management Tab with Full Page Scrolling - Firebase Integrated"""
    
    # Add this to your ExpensesTab class initialization (in __init__ method)
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self.expenses = []
        self.slice_tooltip = None  # For pie slice tooltips
        self.pie_slice_categories = []  # Stores category names in order of slices
        
        # SEPARATE YEAR SELECTION FOR EACH CHART
        self.bar_chart_year = datetime.now().year   # For bar chart only
        self.pie_chart_year = datetime.now().year   # For pie chart only  
        self.pie_chart_month = datetime.now().month  # For pie chart only

        self.cached_expenses = []  # Cache for Firebase data
        self._exp_page = 1
        self._exp_per_page = 10
        self._exp_all_items = []

        # ===== EXTENSIVE COLOR PALETTE FOR PIE CHART =====
        # 30+ distinct colors for default Categories (one unique color for each)
        self.default_category_colors = [
            # Primary Colors (10)
            "#3498db",  # Blue
            "#e74c3c",  # Red
            "#2ecc71",  # Green
            "#f39c12",  # Orange
            "#9b59b6",  # Purple
            "#1abc9c",  # Teal
            "#d35400",  # Pumpkin
            "#c0392b",  # Dark Red
            "#16a085",  # Sea Green
            "#8e44ad",  # Dark Purple
            
            # Secondary Colors (10)
            "#27ae60",  # Nephritis
            "#2980b9",  # Belize Hole
            "#e67e22",  # Carrot
            "#95a5a6",  # Concrete
            "#34495e",  # Wet Asphalt
            "#f1c40f",  # Sun Flower
            "#e84393",  # Pink
            "#00cec9",  # Robin's Egg Blue
            "#fd79a8",  # Pink 2
            "#6c5ce7",  # Light Purple
            
            # Tertiary Colors (10+)
            "#00b894",  # Mint Green
            "#00cec9",  # Turquoise
            "#0984e3",  # Light Blue
            "#6c5ce7",  # Lavender
            "#a29bfe",  # Light Lavender
            "#dfe6e9",  # Silver
            "#636e72",  # Gray Blue
            "#b2bec3",  # Light Gray Blue
            "#fd79a8",  # Coral Pink
            "#fab1a0",  # Peach
            "#ff7675",  # Light Red
            "#fd9644",  # Light Orange
            "#26de81",  # Light Green
            "#20bf6b",  # Green Blue
            "#0fb9b1",  # Cyan
            "#45aaf2",  # Sky Blue
            "#4b7bec",  # Royal Blue
            "#a55eea",  # Bright Purple
            "#d1d8e0",  # Light Silver
            "#778ca3",  # Slate Gray
            "#3867d6",  # Cobalt Blue
            "#8854d0",  # Amethyst
            "#eb3b5a",  # Crimson
            "#fa8231",  # Flame
            "#fed330",  # Mustard
            "#2d98da",  # Summer Sky
            "#fc5c65",  # Sunset Orange
            "#a5b1c2",  # Cool Gray
            "#4b6584",  # Dark Blue Gray
            "#0c2461",  # Navy Blue
            "#3c6382",  # Steel Blue
            "#60a3bc",  # Cadet Blue
            "#82ccdd",  # Pale Cyan
            "#b8e994",  # Light Lime
            "#78e08f",  # Emerald
            "#38ada9",  # Teal Green
            "#079992",  # Dark Teal
            "#1e3799",  # Dark Blue
            "#0a3d62",  # Midnight Blue
            "#3c40c6",  # Indigo
            "#575fcf",  # Periwinkle
            "#4bcffa",  # Light Sky Blue
            "#0fbcf9",  # Electric Blue
            "#00d8d6",  # Turquoise Blue
            "#34e7e4",  # Aqua
            "#ff3f34",  # Tomato
            "#ff5e57",  # Coral
            "#ffd32a",  # Lemon
            "#ffdd59",  # Bright Yellow
            "#ffa801",  # Amber
            "#ffc048",  # Gold
            "#ff9f1a",  # Orange Peel
            "#cd6133",  # Bronze
            "#cc8e35",  # Ochre
            "#ccae62",  # Tan
        ]
        
        # 30+ distinct colors for user-added Categories
        self.user_category_colors = [
            # Vibrant Colors (15)
            "#ff6b6b",  # Watermelon
            "#48dbfb",  # Electric Blue
            "#1dd1a1",  # Caribbean Green
            "#feca57",  # Casandora Yellow
            "#ff9ff3",  # Jigglypuff
            "#54a0ff",  # Bleu de France
            "#5f27cd",  # Bluebell
            "#00d2d3",  # Cyanite
            "#c8d6e5",  # Hint of Ice Pack
            "#ff9f43",  # Orange Hibiscus
            "#ee5a24",  # Pomegranate
            "#f368e0",  # Lighter Purple
            "#0abde3",  # Bright Yarrow
            "#10ac84",  # Dark Mountain Meadow
            "#222f3e",  # Imperial Primer
            
            # Pastel Colors (15+)
            "#ffcccc",  # Light Pink
            "#ccffcc",  # Light Mint
            "#ccccff",  # Light Lavender
            "#ffffcc",  # Light Yellow
            "#ffcc99",  # Light Peach
            "#ccffff",  # Light Cyan
            "#ffccff",  # Light Magenta
            "#e6ccff",  # Light Lilac
            "#ccff99",  # Light Lime
            "#99ffcc",  # Light Seafoam
            "#99ffff",  # Light Aqua
            "#ff9966",  # Light Coral
            "#ffcc66",  # Light Gold
            "#cc9966",  # Light Bronze
            "#6699cc",  # Light Steel Blue
            "#9966cc",  # Light Amethyst
            "#66cccc",  # Light Teal
            "#99cc66",  # Light Olive
            "#cc6699",  # Light Rose
            "#669999",  # Light Slate
            "#999966",  # Light Khaki
            "#cc9999",  # Light Salmon
            "#99cc99",  # Light Sage
            "#cccc99",  # Light Beige
            "#9999cc",  # Light Periwinkle
            "#cc99cc",  # Light Orchid
            "#ff99cc",  # Light Bubblegum
            "#99ff99",  # Light Chartreuse
            "#ccff66",  # Light Green Yellow
            "#ffcc33",  # Light Saffron
            "#cc9933",  # Light Honey
            "#996633",  # Light Brown
            "#663399",  # Rebecca Purple
            "#339966",  # Sea Green
            "#336699",  # Lapis Lazuli
            "#993366",  # Raspberry
            "#669966",  # Moss Green
            "#996699",  # Antique Fuchsia
            "#336666",  # Deep Sea Green
            "#663366",  # Imperial Purple
            "#666699",  # Dark Blue Gray
            "#996666",  # Copper Rose
            "#669966",  # Oxley
            "#666633",  # Dark Olive
            "#339999",  # Persian Green
            "#993333",  # Red Oxide
            "#336633",  # Hunter Green
            "#663333",  # Persian Plum
            "#333366",  # Midnight Blue
            "#666666",  # Dim Gray
        ]
        
        # Combined color palette for all categories
        self.category_colors = {}
        
        # Track which colors are already assigned
        self.assigned_colors = set()
        
        # Map default Categories to specific colors
        self.default_category_color_map = {
            # Facilities & Utilities
            "Facilities & Utilities": "#3498db",  # Blue
            
            # Office & Admin Overhead
            "Office & Admin Overhead": "#e74c3c",  # Red
            
            # Engineering Software & IT
            "Engineering Software & IT": "#2ecc71",  # Green
            
            # Salaries, Labor & Related Costs
            "Salaries, Labor & Related Costs": "#f39c12",  # Orange
            
            # Professional Services
            "Professional Services": "#9b59b6",  # Purple
            
            # Insurance & Compliance
            "Insurance & Compliance": "#1abc9c",  # Teal
            
            # Travel, Site Visits & Vehicles
            "Travel, Site Visits & Vehicles": "#d35400",  # Pumpkin
            
            # Marketing & Business Development
            "Marketing & Business Development": "#c0392b",  # Dark Red
            
            # Training, Licensure & Development
            "Training, Licensure & Development": "#16a085",  # Sea Green
            
            # Safety & Field Supplies
            "Safety & Field Supplies": "#8e44ad",  # Dark Purple
            
            # Miscellaneous O & M
            "Miscellaneous O & M": "#27ae60",  # Nephritis
            
            # Capital Expenses Categories
            "Computer & Office Equipment": "#2980b9",  # Belize Hole
            "Field & Inspection Equipment": "#e67e22",  # Carrot
            "Furniture & Fixtures": "#95a5a6",  # Concrete
            "Vehicles": "#34495e",  # Wet Asphalt
            "Software (Capitalized)": "#f1c40f",  # Sun Flower
            "Leasehold Improvements": "#e84393",  # Pink
            "Accumulated Depreciation": "#00cec9",  # Robin's Egg Blue
            
            # Other Expenses Categories
            "Salary/Bonuses": "#fd79a8",  # Pink 2
            "Tax Expenses/Tax Deductions": "#6c5ce7",  # Light Purple
            "Medical/Benefits": "#00b894",  # Mint Green
            "Meals & Entertainment": "#0984e3",  # Light Blue
            "Donations": "#a29bfe",  # Light Lavender
            "Bank Charges": "#636e72",  # Gray Blue
            "Contingency Funds": "#fd9644",  # Light Orange
            "Unexpected Costs": "#26de81",  # Light Green
        }
        
        # Track used user colors
        self.user_color_index = 0
        
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month
        
        # Filter variables
        self.category_filter_menu = QtWidgets.QMenu()
        self.selected_category_filter = "All Categories"
        self.status_filter_menu = QtWidgets.QMenu()
        self.selected_status_filter = "All Status"
        
        self.init_ui()
        
        # Load initial data with a small delay to ensure main window is fully initialized
        QtCore.QTimer.singleShot(100, self.initial_data_load)
    
    def initial_data_load(self):
        """Load initial data after UI is fully initialized"""
        self.load_expenses()

    def refresh_data(self, auto=False):
        """Reload expenses from Firebase while preserving filters and search."""
        if getattr(self, "_finance_refreshing", False):
            return
        self._finance_refreshing = True
        try:
            search_text = self.search_edit.text() if hasattr(self, "search_edit") else ""
            category_text = self.categories_filter_combo.currentText() if hasattr(self, "categories_filter_combo") else "All Categories"
            date_text = self.date_range_button.text() if hasattr(self, "date_range_button") else ""
            self.load_expenses()
            if hasattr(self, "expense_sync_label"):
                self.expense_sync_label.setText(
                    datetime.now().strftime(("Auto-synced" if auto else "Synced") + " %I:%M %p")
                )
            if hasattr(self, "search_edit"):
                self.search_edit.setText(search_text)
            if hasattr(self, "categories_filter_combo"):
                index = self.categories_filter_combo.findText(category_text)
                self.categories_filter_combo.setCurrentIndex(index if index >= 0 else 0)
            if date_text and hasattr(self, "date_range_button"):
                self.date_range_button.setText(date_text)
            self.filter_expenses()
        finally:
            self._finance_refreshing = False

    def _safe_amount(self, value) -> float:
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace("$", "").replace(",", "").strip() or 0)
            except ValueError:
                return 0.0
        return 0.0

    def _expense_year(self, expense_data) -> int:
        date_text = str(expense_data.get("date", "") or "")
        for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(date_text, fmt).year
            except ValueError:
                continue
        return datetime.now().year

    def _normalize_balance_sheet_expense(self, expense_data: dict) -> dict:
        """Convert a balance sheet expense record into the Expenses tab display schema."""
        # --- Preferred: fields preserved by the updated save_to_balance_sheet_expenses ---
        expense_type   = str(expense_data.get("expense_type", "") or "").strip()
        category_field = str(expense_data.get("Category",      "") or "").strip()
        expense_name_f = str(expense_data.get("expense_name",  "") or "").strip()
        description_f  = str(expense_data.get("description",   "") or "").strip()
        name_field     = str(expense_data.get("name",          "") or "Finance Expense").strip()

        # --- Backward-compat fallback for older records that only had name+description ---
        # Old format stored the original expense_type inside the 'description' field.
        if not expense_type or expense_type == "Balance Sheet Expense":
            # If description looks like a stored expense_type use it; else fall back
            expense_type  = description_f if description_f else "Balance Sheet Expense"
            display_descr = ""          # already promoted to expense_type, nothing left to show
        else:
            display_descr = description_f  # genuine user description

        # Category: prefer stored Category, fall back to name
        category = category_field or name_field or "Finance Expense"
        # Expense Name: prefer stored expense_name, fall back to name
        expense_name = expense_name_f or name_field or "Finance Expense"

        return {
            "date":         expense_data.get("date", "") or expense_data.get("expense_date", ""),
            "expense_type": expense_type,
            "type":         category,
            "Category":     category,
            "expense_name": expense_name,
            "vendor":       expense_data.get("vendor", ""),
            "description":  display_descr,
            "amount":       self._safe_amount(expense_data.get("amount", 0)),
            "project":      expense_data.get("project", ""),
            "project_number": expense_data.get("project_number", expense_data.get("project_no", "")),
            "payment_method": expense_data.get("payment_method", ""),
            "reference":    expense_data.get("reference", ""),
            "notes":        expense_data.get("notes", ""),
            "created_at":   expense_data.get("created_at", ""),
            "updated_at":   expense_data.get("updated_at", ""),
            "firebase_id":  expense_data.get("firebase_id", ""),
            "balance_sheet_firebase_id": expense_data.get("firebase_id", ""),
            "finance_source":   "balance_sheet",
            "read_only_expense": True,
        }

    def _merge_balance_sheet_expenses(self, firebase_expenses: list, balance_sheet_expenses: list) -> list:
        """Merge the two expense stores without duplicating records with the same Firebase id."""
        merged = [dict(expense) for expense in firebase_expenses]
        seen_ids = {
            str(expense.get("firebase_id", ""))
            for expense in merged
            if expense.get("firebase_id")
        }

        for expense in balance_sheet_expenses:
            expense_id = str(expense.get("firebase_id", ""))
            if expense_id and expense_id in seen_ids:
                continue
            normalized = self._normalize_balance_sheet_expense(expense)
            merged.append(normalized)
            if expense_id:
                seen_ids.add(expense_id)

        _log.info(
            "Expenses tab merged %s regular expenses + %s balance sheet-only expenses = %s total",
            len(firebase_expenses),
            max(len(merged) - len(firebase_expenses), 0),
            len(merged),
        )
        return merged

    def _expense_category_map(self) -> dict:
        return {
            "O & M (Operations & Maintenance)": [
                "Facilities & Utilities",
                "Office & Admin Overhead",
                "Engineering Software & IT",
                "Salaries, Labor & Related Costs",
                "Professional Services",
                "Insurance & Compliance",
                "Travel, Site Visits & Vehicles",
                "Marketing & Business Development",
                "Training, Licensure & Development",
                "Safety & Field Supplies",
                "Miscellaneous O & M",
            ],
            "Capital Expenses": [
                "Computer & Office Equipment",
                "Field & Inspection Equipment",
                "Furniture & Fixtures",
                "Vehicles",
                "Software (Capitalized)",
                "Leasehold Improvements",
                "Accumulated Depreciation",
            ],
            "Other Expenses": [
                "Other",
                "Salary/Bonuses",
                "Tax Expenses/Tax Deductions",
                "Medical/Benefits",
                "Meals & Entertainment",
                "Donations",
                "Bank Charges",
                "Contingency Funds",
                "Unexpected Costs",
            ],
        }

    def _expense_name_map(self) -> dict:
        return {
            "Other": [],
            "Facilities & Utilities": [
                "Office rent or co-working space fees",
                "Utilities (electricity, water, gas)",
                "Internet service",
                "Trash & cleaning services",
                "Property taxes (for office, if applicable)",
                "Office repairs & maintenance (HVAC, lights, minor repairs)",
            ],
            "Office & Admin Overhead": [
                "Office supplies (paper, pens, notebooks, printer ink)",
                "Printer/plotter maintenance & paper",
                "Postage & shipping (documents, contracts, samples)",
                "Bank fees & merchant processing fees (credit card, PayPal, Stripe)",
                "Software for admin: Microsoft 365 / Google Workspace",
                "Software for admin: PDF tools (Bluebeam, Adobe, etc.)",
                "Software for admin: Password manager",
                "Software for admin: Others",
                "Cloud storage (Dropbox, Google Drive, OneDrive)",
            ],
            "Engineering Software & IT": [
                "Engineering software subscriptions: SAP2000 / ETABS / STAAD / RAM / RISA, etc.",
                "Engineering software subscriptions: Others",
                "CAD/BIM tools: AutoCAD, Civil 3D, Revit",
                "License/maintenance fees for all software",
                "IT support services",
                "Computer maintenance & small repairs",
                "Antivirus, backup services, and other security tools",
            ],
            "Salaries, Labor & Related Costs": [
                "Owner draw/salary",
                "Employee salaries & wages",
                "Overtime or temporary staff",
                "Payroll taxes paid by the company",
                "Employee benefits: Health insurance contributions",
                "Employee benefits: Retirement plan contributions",
                "Employee benefits: Paid time off costs",
                "Payments to subcontract engineers, drafters, or reviewers",
            ],
            "Professional Services": [
                "Accounting & bookkeeping fees",
                "Tax preparation and consulting",
                "Legal services (contracts, company setup, trademark, disputes)",
                "Business consulting or coaching services",
                "Registered agent fees (if applicable)",
            ],
            "Insurance & Compliance": [
                "Professional liability / Errors & Omissions (E&O) insurance",
                "General liability insurance",
                "Business owner's policy (BOP)",
                "Workers' comp insurance",
                "Commercial auto insurance",
                "License renewals (PE license, SE license, etc.)",
                "Business license renewals",
                "Memberships",
            ],
            "Travel, Site Visits & Vehicles": [
                "Mileage",
                "Fuel costs",
                "Parking fees & tolls",
                "Vehicle maintenance",
                "Airfare, hotels, per diem for out-of-town site visits or client meetings",
                "Rental cars or rideshare for business trips",
                "Meals while traveling for business",
            ],
            "Marketing & Business Development": [
                "Website hosting and domain expenses",
                "Website maintenance & small updates",
                "Graphic design (logo, templates, brochures)",
                "Online ads (Google, LinkedIn, Facebook)",
                "Printing of business cards, brochures, banners",
                "Sponsorships of events",
                "Client entertainment",
            ],
            "Training, Licensure & Development": [
                "Continuing education (PDH hours, webinars, seminars, conferences)",
                "Training courses (technical or business/marketing)",
                "Books, codes, and standards",
                "Exam fees for additional licenses",
            ],
            "Safety & Field Supplies": [
                "PPE: hard hats, safety vests, safety glasses, gloves, boots",
                "Field tools for inspections",
                "Calibration and maintenance of field instruments",
                "First-aid kits and basic safety equipment",
            ],
            "Miscellaneous O & M": [
                "Subscriptions: LinkedIn Premium",
                "Subscriptions: Industry journals or magazines",
                "Software for project management and CRM",
                "Document management tools or e-signature services",
            ],
            "Computer & Office Equipment": [
                "Laptops",
                "Desktops",
                "Monitors",
                "Printers/Scanners",
                "Servers",
                "Networking Equipment",
            ],
            "Field & Inspection Equipment": [
                "Survey Equipment",
                "Testing Equipment",
                "Measurement Tools",
                "Safety Equipment",
                "Inspection Devices",
            ],
            "Furniture & Fixtures": [
                "Office Desks",
                "Chairs",
                "Filing Cabinets",
                "Shelving Units",
                "Conference Room Furniture",
            ],
            "Vehicles": [
                "Company Cars",
                "Trucks",
                "Vans",
                "Heavy Equipment",
                "Vehicle Accessories",
            ],
            "Software (Capitalized)": [
                "Engineering Software License",
                "ERP System",
                "CRM System",
                "Database Software",
                "Custom Software Development",
            ],
            "Leasehold Improvements": [
                "Office Renovations",
                "Electrical Work",
                "Plumbing Improvements",
                "HVAC Installation",
                "Security Systems",
            ],
            "Accumulated Depreciation": [
                "Depreciation Expense - Computers",
                "Depreciation Expense - Office Equipment",
                "Depreciation Expense - Vehicles",
                "Accumulated Depreciation",
            ],
            "Salary/Bonuses": [
                "Employee Salary",
                "Manager Salary",
                "Executive Salary",
                "Performance Bonus",
                "Year-end Bonus",
                "Commission Payments",
                "Incentive Payments",
            ],
            "Tax Expenses/Tax Deductions": [
                "Federal Income Tax",
                "Tax Deduction",
                "Payroll Tax",
                "Sales Tax",
                "Property Tax",
                "Business Tax",
            ],
            "Medical/Benefits": [
                "Health Insurance Premiums",
                "Dental Insurance",
                "Vision Insurance",
                "Retirement Contributions",
                "Life Insurance",
                "Disability Insurance",
                "Wellness Programs",
            ],
            "Meals & Entertainment": [
                "Client Meals",
                "Business Lunches",
                "Team Dinners",
                "Conference Meals",
                "Entertainment Expenses",
                "Team Building Events",
            ],
            "Donations": [
                "Charitable Donations",
                "Community Sponsorships",
                "Educational Donations",
                "Non-profit Contributions",
                "Event Sponsorships",
            ],
            "Bank Charges": [
                "Monthly Account Fees",
                "Transaction Fees",
                "Wire Transfer Fees",
                "Credit Card Processing Fees",
                "Check Printing Fees",
                "Overdraft Fees",
            ],
            "Contingency Funds": [
                "Emergency Funds",
                "Reserve Funds",
                "Project Contingency",
                "Operational Reserve",
                "Risk Management Fund",
            ],
            "Unexpected Costs": [
                "Emergency Repairs",
                "Unplanned Maintenance",
                "Price Increases",
                "Regulatory Changes",
                "Market Fluctuations",
            ],
        }
     
    def init_ui(self):
        # Main layout with scroll area for full page scrolling
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Create scroll area
        scroll_area = QtWidgets.QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QtWidgets.QFrame.NoFrame)
        scroll_area.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        
        # Create scroll content widget
        scroll_content = QtWidgets.QWidget()
        scroll_layout = QtWidgets.QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(20, 30, 20, 20)
        scroll_layout.setSpacing(12)
        
        # ===== HEADER & QUICK ACTIONS =====
        header_frame = self.create_header_section()
        scroll_layout.addWidget(header_frame)
        
        # ===== STATISTICS CARDS =====
        stats_frame = self.create_statistics_section()
        scroll_layout.addWidget(stats_frame)
        
        # ===== QUICK INLINE ENTRY =====
        self.quick_entry_frame = self.create_quick_entry_section()
        self.quick_entry_frame.setVisible(False)
        scroll_layout.addWidget(self.quick_entry_frame)
        
        # ===== ANALYTICS SECTION =====
        analytics_frame = self.create_analytics_section()
        analytics_frame.setMaximumHeight(470)
        scroll_layout.addWidget(analytics_frame)

        # ===== EXPENSE TABLE SECTION =====
        table_frame = self.create_table_section()
        scroll_layout.addWidget(table_frame)
        
        scroll_layout.addStretch()
        
        # Set scroll content
        scroll_area.setWidget(scroll_content)
        main_layout.addWidget(scroll_area)
    
    def create_header_section(self):
        header_frame = QtWidgets.QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """)
        h = QtWidgets.QHBoxLayout(header_frame)
        h.setContentsMargins(20, 14, 20, 14)
        h.setSpacing(12)

        col = QtWidgets.QVBoxLayout()
        col.setSpacing(3)
        t = QtWidgets.QLabel("Expense Management")
        t.setStyleSheet(
            "font-size:20px; font-weight:900; color:#0f172a;"
            " font-family:'Inter','Segoe UI'; background:transparent; border:none;")
        s = QtWidgets.QLabel("Capture, categorize and track company expenses — synced to Finance automatically")
        s.setStyleSheet(
            "font-size:12px; font-weight:600; color:#64748b;"
            " font-family:'Inter','Segoe UI'; background:transparent; border:none;")
        col.addWidget(t)
        col.addWidget(s)
        h.addLayout(col, 1)

        self.expense_sync_label = QtWidgets.QLabel("Auto-sync enabled")
        self.expense_sync_label.setAlignment(QtCore.Qt.AlignCenter)
        self.expense_sync_label.setStyleSheet("""
            QLabel {
                background: #f0fdf4; color: #065f46;
                border: 1px solid #bbf7d0; border-radius: 12px;
                padding: 5px 14px; font-size: 12px; font-weight: 900;
                font-family: 'Inter', 'Segoe UI';
            }
        """)
        h.addWidget(self.expense_sync_label)
        return header_frame

        
    def create_statistics_section(self):
        frame = QtWidgets.QFrame()
        main_layout = QtWidgets.QGridLayout(frame)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setColumnStretch(0, 0)   # LEFT: Export button
        main_layout.setColumnStretch(1, 1)   # CENTER: stat cards
        main_layout.setColumnStretch(2, 0)   # RIGHT: Add expense button

        # -------------------------
        # LEFT SIDE: EXPORT BUTTON
        # -------------------------
        export_btn = QtWidgets.QPushButton("Export")
        export_btn.setFixedSize(120, 44)
        export_btn.setStyleSheet("""
            QPushButton {
                background: #475569; color: white; border: none;
                border-radius: 8px; font-family: 'Inter', 'Segoe UI';
                font-weight: 800; font-size: 13px;
            }
            QPushButton:hover { background: #334155; }
        """)
        export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        export_btn.clicked.connect(self.open_pdf_export_dialog)
        main_layout.addWidget(export_btn, 0, 0, QtCore.Qt.AlignLeft)

        # -------------------------
        # CENTER: STAT CARDS
        # -------------------------
        self.cards_container = QtWidgets.QWidget()
        self.cards_layout = QtWidgets.QHBoxLayout(self.cards_container)
        self.cards_layout.setContentsMargins(0, 0, 0, 0)
        self.cards_layout.setSpacing(16)
        self.cards_layout.setAlignment(QtCore.Qt.AlignCenter)

        # Create stat cards with initial values
        self.update_statistics_cards()

        main_layout.addWidget(self.cards_container, 0, 1, QtCore.Qt.AlignCenter)

        # -------------------------
        # RIGHT: ADD EXPENSE BUTTON
        # -------------------------
        self.add_expense_btn = QtWidgets.QPushButton("+ Add Expense")
        self.add_expense_btn.setFixedSize(180, 44)
        self.add_expense_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.add_expense_btn.setStyleSheet("""
            QPushButton {
                background: #0f766e; color: #ffffff; border: none;
                border-radius: 8px; font-family: 'Inter', 'Segoe UI';
                font-size: 14px; font-weight: 900; padding: 0 18px;
            }
            QPushButton:hover { background: #0d625c; }
            QPushButton:pressed { background: #0a4f49; }
        """)
        self.add_expense_btn.clicked.connect(self.show_quick_entry)

        btn_container = QtWidgets.QWidget()
        btn_layout = QtWidgets.QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.addStretch()
        btn_layout.addWidget(self.add_expense_btn)

        main_layout.addWidget(btn_container, 0, 2, QtCore.Qt.AlignRight)

        return frame


    def create_quick_entry_section(self):
        frame = QtWidgets.QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
            }
            QLabel {
                background: transparent;
                border: none;
                color: #334155;
                font-family: 'Inter', 'Segoe UI';
                font-size: 12px;
                font-weight: 800;
            }
            QLineEdit, QComboBox, QDateEdit {
                background: white;
                border: 1.5px solid #d8e2ec;
                border-radius: 8px;
                padding: 7px 10px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 13px;
                min-height: 28px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border-color: #00756f;
            }
            QComboBox::drop-down, QDateEdit::drop-down {
                border: none;
                width: 24px;
                background: transparent;
            }
            QComboBox::down-arrow, QDateEdit::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #64748b;
                margin-right: 8px;
            }
            QCalendarWidget QWidget {
                background: white;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
                font-size: 12px;
            }
            QCalendarWidget QToolButton {
                background: #f8fafc;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                border-radius: 6px;
                margin: 2px;
                padding: 4px 8px;
            }
        """)
        layout = QtWidgets.QVBoxLayout(frame)
        layout.setContentsMargins(18, 14, 18, 16)
        layout.setSpacing(12)

        header = QtWidgets.QHBoxLayout()
        title = QtWidgets.QLabel("Quick Expense Entry")
        title.setStyleSheet("""
            QLabel {
                color: #0f172a;
                font-size: 17px;
                font-weight: 900;
                font-family: 'Inter', 'Segoe UI';
            }
        """)
        close_btn = QtWidgets.QPushButton("Cancel")
        close_btn.setFixedSize(100, 36)
        close_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        close_btn.setStyleSheet("""
            QPushButton {
                background: #f8fafc;
                color: #334155;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                font-weight: 800;
            }
            QPushButton:hover { background: #eef2f7; }
        """)
        close_btn.clicked.connect(lambda: self.quick_entry_frame.setVisible(False))
        header.addWidget(title)
        header.addStretch()
        header.addWidget(close_btn)
        layout.addLayout(header)

        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)

        self.quick_date_edit = QtWidgets.QDateEdit()
        self.quick_date_edit.setCalendarPopup(True)
        self.quick_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.quick_date_edit.setDate(QtCore.QDate.currentDate())
        self.quick_date_edit.setFixedHeight(44)

        self.quick_type_combo = QtWidgets.QComboBox()
        self.quick_type_combo.setEditable(True)
        self.quick_type_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.quick_type_combo.addItems([
            "O & M (Operations & Maintenance)",
            "Capital Expenses",
            "Other Expenses",
        ])
        self.quick_type_combo.setCurrentIndex(-1)
        if self.quick_type_combo.lineEdit():
            self.quick_type_combo.lineEdit().setPlaceholderText("Select or type expense type")
        self.quick_type_combo.setFixedHeight(44)
        self.quick_type_combo.currentTextChanged.connect(self.update_quick_categories)

        self.quick_category_combo = QtWidgets.QComboBox()
        self.quick_category_combo.setEditable(True)
        self.quick_category_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.quick_category_combo.setFixedHeight(44)
        self.quick_category_combo.setMaxVisibleItems(12)
        self.quick_category_combo.currentTextChanged.connect(self.update_quick_expenses)
        self.quick_expense_combo = QtWidgets.QComboBox()
        self.quick_expense_combo.setEditable(True)
        self.quick_expense_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.quick_expense_combo.setFixedHeight(44)
        self.quick_expense_combo.setMaxVisibleItems(14)
        self.quick_expense_combo.lineEdit().setPlaceholderText("Select or enter expense")
        # auto_classify intentionally not connected — users select type/category/name manually
        self.quick_vendor_combo = QtWidgets.QComboBox()
        self.quick_vendor_combo.setEditable(True)
        self.quick_vendor_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.quick_vendor_combo.setFixedHeight(44)
        self.quick_vendor_combo.setMaxVisibleItems(14)
        if self.quick_vendor_combo.lineEdit():
            self.quick_vendor_combo.lineEdit().setPlaceholderText("Type or select vendor")
        self.quick_amount_edit = QtWidgets.QLineEdit()
        self.quick_amount_edit.setPlaceholderText("0.00")
        self.quick_project_combo = QtWidgets.QComboBox()
        self.quick_project_combo.setEditable(True)
        self.quick_project_combo.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        self.quick_project_combo.setFixedHeight(44)
        self.quick_project_combo.setMaxVisibleItems(14)
        self.refresh_quick_project_combo()
        self.quick_description_edit = QtWidgets.QLineEdit()
        self.quick_description_edit.setPlaceholderText("Description")
        self.quick_balance_sheet_check = QtWidgets.QCheckBox("Save to Balance Sheet")
        self.quick_balance_sheet_check.setChecked(True)
        self.quick_balance_sheet_check.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.quick_balance_sheet_check.setStyleSheet("""
            QCheckBox {
                background: transparent;
                border: none;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
                font-size: 13px;
                font-weight: 900;
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 1.5px solid #99b8c9;
                border-radius: 5px;
                background: white;
            }
            QCheckBox::indicator:checked {
                background: #00756f;
                border-color: #00756f;
            }
        """)

        fields = [
            ("Date", self.quick_date_edit, 1),
            ("Type", self.quick_type_combo, 2),
            ("Category", self.quick_category_combo, 2),
            ("Expense", self.quick_expense_combo, 2),
            ("Amount", self.quick_amount_edit, 1),
            ("Vendor", self.quick_vendor_combo, 2),
            ("Project", self.quick_project_combo, 2),
            ("Description", self.quick_description_edit, 3),
        ]
        col = 0
        row = 0
        for label_text, widget, span in fields:
            if col + span > 7:
                row += 2
                col = 0
            label = QtWidgets.QLabel(label_text)
            grid.addWidget(label, row, col, 1, span)
            grid.addWidget(widget, row + 1, col, 1, span)
            col += span
        for col_index in range(7):
            grid.setColumnStretch(col_index, 1)
        layout.addLayout(grid)

        action_row = QtWidgets.QHBoxLayout()
        hint = QtWidgets.QLabel("Checked expenses sync into Balance Sheet.")
        hint.setStyleSheet("color:#64748b; font-size:12px; font-weight:700; background:transparent; border:none;")
        self.quick_save_btn = QtWidgets.QPushButton("Save Expense")
        self.quick_save_btn.setFixedSize(180, 42)
        self.quick_save_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.quick_save_btn.setEnabled(False)
        self.quick_save_btn.setStyleSheet("""
            QPushButton {
                background: #9ca3af;
                color: white;
                border: none;
                border-radius: 8px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 15px;
                font-weight: 900;
            }
            QPushButton:hover { background: #9ca3af; }
        """)
        self.quick_save_btn.clicked.connect(self.save_quick_expense)
        action_row.addWidget(self.quick_balance_sheet_check)
        action_row.addSpacing(12)
        action_row.addWidget(hint)
        action_row.addStretch()
        action_row.addWidget(self.quick_save_btn)
        layout.addLayout(action_row)

        # Connect all 4 required fields to validation
        self.quick_type_combo.currentTextChanged.connect(self._update_quick_save_btn)
        self.quick_category_combo.currentTextChanged.connect(self._update_quick_save_btn)
        self.quick_expense_combo.currentTextChanged.connect(self._update_quick_save_btn)
        self.quick_amount_edit.textChanged.connect(self._update_quick_save_btn)

        self.update_quick_categories("")
        return frame

    def show_quick_entry(self):
        self.refresh_quick_project_combo()
        self.refresh_quick_vendor_combo()
        self.quick_entry_frame.setVisible(True)
        self.quick_date_edit.setDate(QtCore.QDate.currentDate())
        self.quick_amount_edit.setFocus()
        self._update_quick_save_btn()

    def _active_project_options_for_tab(self):
        inactive_statuses = {
            "Completed & Invoiced",
            "Paid",
            "Cancelled",
            "Cancel",
            "Completed",
        }
        projects = []
        project_tab = getattr(self.main_window, "project_tab", None)
        for attr in ("cached_projects", "generated_projects"):
            projects.extend(getattr(project_tab, attr, []) or [])

        if not projects:
            try:
                from main import FirebaseManager
                projects = FirebaseManager.load_projects() or []
            except Exception as exc:
                _log.warning("Could not load active projects for quick expense: %s", exc)
                projects = []

        options = [("General / Overhead", "")]
        seen = set()
        for project in projects:
            if not isinstance(project, dict):
                continue
            status = str(project.get("status", "") or "").strip()
            if status in inactive_statuses:
                continue
            project_number = str(project.get("project_number", "") or "").strip()
            if not project_number or project_number in seen:
                continue
            seen.add(project_number)
            project_name = str(project.get("project_name", "") or "").strip()
            display = project_number if not project_name else f"{project_number} - {project_name}"
            options.append((display, project_number))
        return options

    def refresh_quick_project_combo(self):
        if not hasattr(self, "quick_project_combo"):
            return
        current_data = self.quick_project_combo.currentData()
        current_text = self.quick_project_combo.currentText().strip()
        self.quick_project_combo.blockSignals(True)
        self.quick_project_combo.clear()
        for display, project_number in self._active_project_options_for_tab():
            self.quick_project_combo.addItem(display, project_number)
        restore_index = self.quick_project_combo.findData(current_data)
        if restore_index < 0 and current_text:
            restore_index = self.quick_project_combo.findText(current_text)
        if restore_index >= 0:
            self.quick_project_combo.setCurrentIndex(restore_index)
        else:
            self.quick_project_combo.setCurrentIndex(-1)
            self.quick_project_combo.setEditText("")
        if self.quick_project_combo.lineEdit():
            self.quick_project_combo.lineEdit().setPlaceholderText("Project #")
        self.quick_project_combo.blockSignals(False)

    def refresh_quick_vendor_combo(self):
        if not hasattr(self, "quick_vendor_combo"):
            return
        current_text = self.quick_vendor_combo.currentText().strip() if self.quick_vendor_combo.lineEdit() else ""
        vendors_set = set()
        for exp in (getattr(self, "expenses", []) or []):
            v = str(exp.get('vendor', '') or '').strip()
            if v:
                vendors_set.add(v)
        try:
            for v in ExpensesFirebaseManager.load_vendors():
                if v:
                    vendors_set.add(v)
        except Exception:
            pass
        vendors = sorted(vendors_set)
        self.quick_vendor_combo.blockSignals(True)
        self.quick_vendor_combo.clear()
        self.quick_vendor_combo.addItems(vendors)
        if current_text:
            idx = self.quick_vendor_combo.findText(current_text)
            if idx >= 0:
                self.quick_vendor_combo.setCurrentIndex(idx)
            else:
                self.quick_vendor_combo.setEditText(current_text)
        else:
            self.quick_vendor_combo.setCurrentIndex(-1)
            if self.quick_vendor_combo.lineEdit():
                self.quick_vendor_combo.lineEdit().clear()
                self.quick_vendor_combo.lineEdit().setPlaceholderText("Type or select vendor")
        self.quick_vendor_combo.blockSignals(False)

    def _quick_categories(self):
        return {
            "O & M (Operations & Maintenance)": [
                "Facilities & Utilities",
                "Office & Admin Overhead",
                "Engineering Software & IT",
                "Salaries, Labor & Related Costs",
                "Professional Services",
                "Insurance & Compliance",
                "Travel, Site Visits & Vehicles",
                "Marketing & Business Development",
                "Training, Licensure & Development",
                "Safety & Field Supplies",
                "Miscellaneous O & M",
            ],
            "Capital Expenses": [
                "Computer & Office Equipment",
                "Field & Inspection Equipment",
                "Furniture & Fixtures",
                "Vehicles",
                "Software (Capitalized)",
                "Leasehold Improvements",
                "Accumulated Depreciation",
            ],
            "Other Expenses": [
                "Other",
                "Salary/Bonuses",
                "Tax Expenses/Tax Deductions",
                "Medical/Benefits",
                "Meals & Entertainment",
                "Donations",
                "Bank Charges",
                "Contingency Funds",
                "Unexpected Costs",
            ],
        }

    def _quick_expense_names(self):
        return {
            "Other": [],
            "Facilities & Utilities": [
                "Office rent or co-working space fees",
                "Utilities (electricity, water, gas)",
                "Internet service",
                "Trash & cleaning services",
                "Property taxes (for office, if applicable)",
                "Office repairs & maintenance (HVAC, lights, minor repairs)",
            ],
            "Office & Admin Overhead": [
                "Office supplies (paper, pens, notebooks, printer ink)",
                "Printer/plotter maintenance & paper",
                "Postage & shipping (documents, contracts, samples)",
                "Bank fees & merchant processing fees (credit card, PayPal, Stripe)",
                "Software for admin: Microsoft 365 / Google Workspace",
                "Software for admin: PDF tools (Bluebeam, Adobe, etc.)",
                "Software for admin: Password manager",
                "Software for admin: Others",
                "Cloud storage (Dropbox, Google Drive, OneDrive)",
            ],
            "Engineering Software & IT": [
                "Engineering software subscriptions: SAP2000 / ETABS / STAAD / RAM / RISA, etc.",
                "Engineering software subscriptions: Others",
                "CAD/BIM tools: AutoCAD, Civil 3D, Revit",
                "License/maintenance fees for all software",
                "IT support services",
                "Computer maintenance & small repairs",
                "Antivirus, backup services, and other security tools",
            ],
            "Salaries, Labor & Related Costs": [
                "Owner draw/salary",
                "Employee salaries & wages",
                "Overtime or temporary staff",
                "Payroll taxes paid by the company",
                "Employee benefits: Health insurance contributions",
                "Employee benefits: Retirement plan contributions",
                "Employee benefits: Paid time off costs",
                "Payments to subcontract engineers, drafters, or reviewers",
            ],
            "Professional Services": [
                "Accounting & bookkeeping fees",
                "Tax preparation and consulting",
                "Legal services (contracts, company setup, trademark, disputes)",
                "Business consulting or coaching services",
                "Registered agent fees (if applicable)",
            ],
            "Insurance & Compliance": [
                "Professional liability / Errors & Omissions (E&O) insurance",
                "General liability insurance",
                "Business owner's policy (BOP)",
                "Workers' comp insurance",
                "Commercial auto insurance",
                "License renewals (PE license, SE license, etc.)",
                "Business license renewals",
                "Memberships",
            ],
            "Travel, Site Visits & Vehicles": [
                "Mileage (personal vehicle for business)",
                "Fuel costs (company vehicles)",
                "Parking fees & tolls",
                "Vehicle maintenance for company-owned vehicle",
                "Airfare, hotels, per diem for site visits or client meetings",
                "Rental cars or rideshare for business trips",
                "Meals while traveling for business",
            ],
            "Marketing & Business Development": [
                "Website hosting and domain expenses",
                "Website maintenance & small updates",
                "Graphic design (logo, templates, brochures)",
                "Online ads (Google, LinkedIn, Facebook)",
                "Printing of business cards, brochures, banners",
                "Sponsorships of events",
                "Client entertainment",
            ],
            "Training, Licensure & Development": [
                "Continuing education (PDH hours, webinars, seminars, conferences)",
                "Training courses (technical or business/marketing)",
                "Books, codes, and standards",
                "Exam fees for additional licenses",
            ],
            "Safety & Field Supplies": [
                "PPE: hard hats, safety vests, safety glasses, gloves, boots",
                "Field tools for inspections",
                "Calibration and maintenance of field instruments",
                "First-aid kits and basic safety equipment",
            ],
            "Miscellaneous O & M": [
                "LinkedIn Premium",
                "Industry journals or magazines",
                "Project management and CRM tools",
                "Document management tools or e-signature services",
            ],
            "Computer & Office Equipment": [
                "Computer & Office Equipment",
                "Laptops",
                "Desktops",
                "Monitors",
                "Printers/Scanners",
                "Servers",
                "Networking Equipment",
            ],
            "Field & Inspection Equipment": [
                "Field & Inspection Equipment",
                "Survey Equipment",
                "Testing Equipment",
                "Measurement Tools",
                "Safety Equipment",
                "Inspection Devices",
            ],
            "Furniture & Fixtures": [
                "Furniture & Fixtures",
                "Office Desks",
                "Chairs",
                "Filing Cabinets",
                "Shelving Units",
                "Conference Room Furniture",
            ],
            "Vehicles": [
                "Vehicles",
                "Company Cars",
                "Trucks",
                "Vans",
                "Heavy Equipment",
                "Vehicle Accessories",
            ],
            "Software (Capitalized)": [
                "Software (Capitalized)",
                "Engineering Software License",
                "ERP System",
                "CRM System",
                "Database Software",
                "Custom Software Development",
            ],
            "Leasehold Improvements": [
                "Leasehold Improvements",
                "Office Renovations",
                "Electrical Work",
                "Plumbing Improvements",
                "HVAC Installation",
                "Security Systems",
            ],
            "Accumulated Depreciation": [
                "Depreciation Expense - Computers",
                "Depreciation Expense - Office Equipment",
                "Depreciation Expense - Vehicles",
                "Accumulated Depreciation",
            ],
            "Salary/Bonuses": [
                "Employee Salary",
                "Manager Salary",
                "Executive Salary",
                "Performance Bonus",
                "Year-end Bonus",
                "Commission Payments",
                "Incentive Payments",
            ],
            "Tax Expenses/Tax Deductions": [
                "Federal Income Tax",
                "Tax Deduction",
                "Payroll Tax",
                "Sales Tax",
                "Property Tax",
                "Business Tax",
            ],
            "Medical/Benefits": [
                "Health Insurance Premiums",
                "Dental Insurance",
                "Vision Insurance",
                "Retirement Contributions",
                "Life Insurance",
                "Disability Insurance",
                "Wellness Programs",
            ],
            "Meals & Entertainment": [
                "Client Meals",
                "Business Lunches",
                "Team Dinners",
                "Conference Meals",
                "Entertainment Expenses",
                "Team Building Events",
            ],
            "Donations": [
                "Charitable Donations",
                "Community Sponsorships",
                "Educational Donations",
                "Non-profit Contributions",
                "Event Sponsorships",
            ],
            "Bank Charges": [
                "Monthly Account Fees",
                "Transaction Fees",
                "Wire Transfer Fees",
                "Credit Card Processing Fees",
                "Check Printing Fees",
                "Overdraft Fees",
            ],
            "Contingency Funds": [
                "Emergency Funds",
                "Reserve Funds",
                "Project Contingency",
                "Operational Reserve",
                "Risk Management Fund",
            ],
            "Unexpected Costs": [
                "Emergency Repairs",
                "Unplanned Maintenance",
                "Price Increases",
                "Regulatory Changes",
                "Market Fluctuations",
            ],
        }

    def update_quick_categories(self, expense_type):
        categories = self._quick_categories()
        guided_type = expense_type in (
            "O & M (Operations & Maintenance)",
            "Capital Expenses",
        )
        current = self.quick_category_combo.currentText().strip() if guided_type else ""
        self.quick_category_combo.blockSignals(True)
        self.quick_category_combo.clear()
        self.quick_category_combo.addItems(categories.get(expense_type, []))
        if current:
            self.quick_category_combo.setEditText(current)
        else:
            self.quick_category_combo.setCurrentIndex(-1)
            if self.quick_category_combo.lineEdit():
                self.quick_category_combo.lineEdit().clear()
                self.quick_category_combo.lineEdit().setPlaceholderText("Select or type category")
        self.quick_category_combo.blockSignals(False)
        self.update_quick_expenses(self.quick_category_combo.currentText())

    def update_quick_expenses(self, category):
        if not hasattr(self, "quick_expense_combo"):
            return
        items = self._quick_expense_names().get(category, [])
        manual_expense = (
            self.quick_type_combo.currentText() == "Other Expenses"
            and category == "Other"
        )
        current = self.quick_expense_combo.currentText().strip()
        if items and current not in items:
            current = ""
        self.quick_expense_combo.blockSignals(True)
        self.quick_expense_combo.clear()
        self.quick_expense_combo.addItems(items)
        if manual_expense:
            self.quick_expense_combo.setCurrentIndex(-1)
            self.quick_expense_combo.setEditText("")
            if self.quick_expense_combo.lineEdit():
                self.quick_expense_combo.lineEdit().setPlaceholderText("Type expense manually")
                self.quick_expense_combo.lineEdit().setReadOnly(False)
        elif current:
            self.quick_expense_combo.setEditText(current)
        else:
            self.quick_expense_combo.setCurrentIndex(-1)
            self.quick_expense_combo.setEditText("")
            if self.quick_expense_combo.lineEdit():
                self.quick_expense_combo.lineEdit().setPlaceholderText("Select or enter expense")
        self.quick_expense_combo.blockSignals(False)

    def _combo_entry_text(self, combo):
        if combo.isEditable() and combo.lineEdit():
            typed_text = combo.lineEdit().text().strip()
            if typed_text:
                return typed_text
        return combo.currentText().strip()

    def auto_classify_quick_expense(self, text):
        try:
            if (
                self.quick_type_combo.currentText() == "Other Expenses"
                and self._combo_entry_text(self.quick_category_combo) == "Other"
            ):
                return
            text = (text or "").strip().lower()
            if len(text) < 3 or getattr(self, "_quick_auto_classifying", False):
                return
            for category, expense_items in self._quick_expense_names().items():
                for expense_item in expense_items:
                    item_text = expense_item.lower()
                    if text == item_text or text in item_text:
                        expense_type = next(
                            (
                                type_name
                                for type_name, categories in self._quick_categories().items()
                                if category in categories
                            ),
                            "",
                        )
                        if not expense_type:
                            return
                        self._quick_auto_classifying = True
                        self.quick_type_combo.setCurrentText(expense_type)
                        self.update_quick_categories(expense_type)
                        self.quick_category_combo.setCurrentText(category)
                        self.update_quick_expenses(category)
                        self.quick_expense_combo.setEditText(expense_item)
                        self._quick_auto_classifying = False
                        return
        except Exception as e:
            self._quick_auto_classifying = False
            _log.warning("Quick expense auto-classification skipped: %s", e)

    def _update_quick_save_btn(self, *_):
        """Enable quick-entry Save button only when all 4 required fields are filled."""
        if not hasattr(self, 'quick_save_btn'):
            return
        expense_type = self.quick_type_combo.currentText().strip()
        category     = self.quick_category_combo.currentText().strip()
        expense_name = self.quick_expense_combo.currentText().strip()
        amount_raw   = self.quick_amount_edit.text().replace("$", "").strip()
        try:
            amount_ok = float(amount_raw) > 0 if amount_raw else False
        except ValueError:
            amount_ok = False
        all_ok = bool(expense_type and category and expense_name and amount_ok)
        self.quick_save_btn.setEnabled(all_ok)
        if all_ok:
            self.quick_save_btn.setStyleSheet("""
                QPushButton {
                    background: #00756f; color: white; border: none;
                    border-radius: 8px; font-family: 'Inter', 'Segoe UI';
                    font-size: 15px; font-weight: 900;
                }
                QPushButton:hover { background: #00645f; }
            """)
        else:
            self.quick_save_btn.setStyleSheet("""
                QPushButton {
                    background: #9ca3af; color: white; border: none;
                    border-radius: 8px; font-family: 'Inter', 'Segoe UI';
                    font-size: 15px; font-weight: 900;
                }
                QPushButton:hover { background: #9ca3af; }
            """)

    def save_quick_expense(self):
        amount = self._safe_amount(self.quick_amount_edit.text())
        if amount <= 0:
            QtWidgets.QMessageBox.warning(self, "Amount Required", "Enter a valid expense amount.")
            return
        category = self._combo_entry_text(self.quick_category_combo) or "Other"
        expense_name = self._combo_entry_text(self.quick_expense_combo) or category
        selected_project = self.quick_project_combo.currentData()
        if selected_project is None:
            selected_project = self.quick_project_combo.currentText().strip()
        if self.quick_type_combo.currentText() == "Other Expenses" and category == "Other" and not expense_name:
            QtWidgets.QMessageBox.warning(self, "Expense Required", "Please type the expense name.")
            self.quick_expense_combo.setFocus()
            return
        expense_data = {
            "date": self.quick_date_edit.date().toString("MM-dd-yyyy"),
            "expense_type": self.quick_type_combo.currentText(),
            "type": category,
            "Category": category,
            "expense_name": expense_name,
            "vendor": self.quick_vendor_combo.currentText().strip(),
            "description": self.quick_description_edit.text().strip(),
            "amount": amount,
            "project": selected_project,
            "project_number": selected_project,
            "method": "",
            "save_to_balance_sheet": self.quick_balance_sheet_check.isChecked(),
        }
        self.save_expense(expense_data)
        self.quick_expense_combo.setEditText("")
        self.quick_amount_edit.clear()
        self.quick_description_edit.clear()
        if self.quick_vendor_combo.lineEdit():
            self.quick_vendor_combo.lineEdit().clear()
        self.quick_project_combo.setCurrentIndex(-1)
        self.quick_project_combo.setEditText("")
        self.update_quick_categories(self.quick_type_combo.currentText())
        self.quick_entry_frame.setVisible(False)


    def create_stat_card(self, title, value, color, icon):
        # Map old pastel background strings to modern tinted card styles
        COLOR_MAP = {
            "background: #DFF0FA;": ("#eff6ff", "#bfdbfe", "#2563eb"),
            "background: #F7DDE2;": ("#fff7f7", "#fecaca", "#dc2626"),
            "background: #EBDDFA;": ("#faf5ff", "#e9d5ff", "#7c3aed"),
        }
        bg, border, val_color = COLOR_MAP.get(color.strip(), ("#f8fafc", "#e2e8f0", "#0f172a"))

        card = QtWidgets.QFrame()
        card.setFixedSize(220, 86)
        card.setStyleSheet(f"""
            QFrame {{
                background: {bg};
                border: 1px solid {border};
                border-radius: 10px;
            }}
        """)

        layout = QtWidgets.QVBoxLayout(card)
        layout.setContentsMargins(16, 10, 16, 10)
        layout.setSpacing(4)

        title_label = QtWidgets.QLabel(title.upper())
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        title_label.setObjectName("stat_title")
        title_label.setStyleSheet(
            "background:transparent; border:none;"
            " font-size:10px; font-weight:800; color:#94a3b8; letter-spacing:0.8px;")

        value_label = QtWidgets.QLabel(value)
        value_label.setAlignment(QtCore.Qt.AlignCenter)
        value_label.setObjectName("stat_value")
        value_label.setStyleSheet(
            f"background:transparent; border:none;"
            f" font-size:18px; font-weight:900; color:{val_color};")

        layout.addWidget(title_label)
        layout.addWidget(value_label)
        return card

    def create_analytics_section(self):
        """Create analytics section with bar chart and pie chart"""
        frame = QtWidgets.QFrame()
        layout = QtWidgets.QHBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)
        
        # Left side - Bar Chart with controls
        chart_card = self.create_chart_card()
        layout.addWidget(chart_card, 2)
        
        # Right side - Pie Chart with controls
        pie_card = self.create_pie_chart_card()
        layout.addWidget(pie_card, 1)
        
        return frame

    def open_year_picker(self):
        """Year picker for BAR CHART only"""
        dlg = YearPickerDialog(self, self.bar_chart_year)  # Use bar_chart_year
        if dlg.exec_() == QtWidgets.QDialog.Accepted:  # ✅ FIXED: Changed 'Acepted' to 'Accepted'
            self.bar_chart_year = dlg.selected_year  # Update only bar chart year
            self.update_bar_chart()  # Only update bar chart

    def create_chart_card(self):
        """Create bar chart card with controls"""
        card = QtWidgets.QGroupBox("")
        card.setMaximumHeight(500)
        card.setStyleSheet("""
            QGroupBox {
                color: #2c3e50;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                margin-top: 0;
                padding-top: 0;
                background: white;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        
        # Year selection controls
        controls_layout = QtWidgets.QHBoxLayout()
        controls_layout.setSpacing(15)

        analytics_title = QtWidgets.QLabel("Yearly Expense Trend")
        analytics_title.setStyleSheet("""
            QLabel {
                background: transparent;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
                font-size: 16px;
                font-weight: 900;
            }
        """)
        controls_layout.addWidget(analytics_title)

        # *** NEW CHART TITLE (centered) ***
        current_year = datetime.now().year
        self.bar_title = QtWidgets.QLabel(f"Yearly Expenses - {current_year}")
        self.bar_title.setStyleSheet("""
            QLabel {
                background: #f8fafc;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                padding: 5px 12px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 13px;
                font-weight: 800;
            }
        """)
        self.bar_title.setAlignment(QtCore.Qt.AlignCenter)

        # Add title to center with stretch on both sides
        controls_layout.addStretch()
        controls_layout.addWidget(self.bar_title)

        year_btn = QtWidgets.QPushButton("Change Year")
        year_btn.setFixedHeight(32)
        year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        year_btn.setStyleSheet("""
            QPushButton {
                background: #f8fafc;
                color: #0f766e;
                border: 1px solid #99f6e4;
                border-radius: 8px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 12px;
                font-weight: 900;
                padding: 0 12px;
            }
            QPushButton:hover { background: #ecfdf5; }
        """)
        year_btn.clicked.connect(self.open_year_picker)
        controls_layout.addWidget(year_btn)

        layout.addLayout(controls_layout)
        
        # Create bar chart
        self.bar_chart_widget = self.create_bar_chart()
        QtCore.QTimer.singleShot(0, lambda: self.position_axis_titles(None))
        layout.addWidget(self.bar_chart_widget)
        
        return card
    
    def get_color_for_category(self, category):
        """Get a unique color for a category"""
        # If we already assigned a color for this category, return it
        if category in self.category_colors:
            return self.category_colors[category]
        
        # Check if it's a default category with pre-assigned color
        if category in self.default_category_color_map:
            color = self.default_category_color_map[category]
            if color not in self.assigned_colors:
                self.category_colors[category] = color
                self.assigned_colors.add(color)
                return color
        
        # Check if it matches any default category pattern
        is_default = False
        default_patterns = [
            "Facilities & Utilities",
            "Office & Admin Overhead",
            "Engineering Software & IT",
            "Salaries, Labor & Related Costs",
            "Professional Services",
            "Insurance & Compliance",
            "Travel, Site Visits & Vehicles",
            "Marketing & Business Development",
            "Training, Licensure & Development",
            "Safety & Field Supplies",
            "Miscellaneous O & M",
            "Computer & Office Equipment",
            "Field & Inspection Equipment",
            "Furniture & Fixtures",
            "Vehicles",
            "Software (Capitalized)",
            "Leasehold Improvements",
            "Accumulated Depreciation",
            "Salary/Bonuses",
            "Tax Expenses/Tax Deductions",
            "Medical/Benefits",
            "Meals & Entertainment",
            "Donations",
            "Bank Charges",
            "Contingency Funds",
            "Unexpected Costs"
        ]
        
        # Check if category starts with any default pattern
        for pattern in default_patterns:
            if category == pattern:
                is_default = True
                break
        
        # Assign color based on category type
        if is_default:
            # Assign from default colors (excluding already used ones)
            for color in self.default_category_colors:
                if color not in self.assigned_colors:
                    self.category_colors[category] = color
                    self.assigned_colors.add(color)
                    return color
        else:
            # User-added category - use sequential assignment from user colors
            if self.user_color_index < len(self.user_category_colors):
                color = self.user_category_colors[self.user_color_index]
                self.user_color_index += 1
            else:
                # If we run out of predefined user colors, start from beginning
                color = self.user_category_colors[0]
                self.user_color_index = 1
                
            # Make sure color is not already assigned
            while color in self.assigned_colors and len(self.assigned_colors) < len(self.user_category_colors):
                color = self.user_category_colors[self.user_color_index % len(self.user_category_colors)]
                self.user_color_index += 1
            
            self.category_colors[category] = color
            self.assigned_colors.add(color)
            return color
        
        # Fallback: generate a random color that doesn't conflict
        import random
        def random_color():
            # Generate vibrant colors, not too dark or too light
            r = random.randint(50, 200)
            g = random.randint(50, 200)
            b = random.randint(50, 200)
            return "#{:02x}{:02x}{:02x}".format(r, g, b)
        
        color = random_color()
        # Ensure color is not too similar to existing colors
        max_attempts = 100
        attempt = 0
        while color in self.assigned_colors and attempt < max_attempts:
            color = random_color()
            attempt += 1
        
        self.category_colors[category] = color
        self.assigned_colors.add(color)
        return color

    def reset_color_assignments(self):
        """Reset color assignments to handle new data"""
        # Keep the pre-assigned default colors
        for category, color in self.default_category_color_map.items():
            self.category_colors[category] = color
            self.assigned_colors.add(color)
        
        # Reset user color index
        self.user_color_index = 0
    def regenerate_category_colors(self):
        """Regenerate colors for all categories (use with caution)"""
        old_assignments = self.category_colors.copy()
        self.category_colors = {}
        self.assigned_colors = set()
        self.user_color_index = 0
        
        # Reassign colors to all categories
        for category in old_assignments.keys():
            self.get_color_for_category(category)
        
        _log.info("Regenerated colors for %s categories", len(self.category_colors))
        self.save_custom_category_colors()
        self.update_pie_chart()
        # Note: We don't clear all colors, only reset the index
        # Existing assignments remain to maintain consistency
    def show_color_assignments(self):
        """Show current color assignments (for debugging)"""
        _log.info("\n=== Color Assignments ===")
        _log.info("Total categories: %s", len(self.category_colors))
        _log.info("Assigned colors: %s", len(self.assigned_colors))
        _log.info("User color index: %s", self.user_color_index)
        
        # Group by category type
        default_categories = []
        user_categories = []
        
        for category in self.category_colors.keys():
            is_default = False
            for pattern in self.default_category_color_map.keys():
                if category == pattern:
                    is_default = True
                    break
            
            if is_default or category in self.default_category_color_map:
                default_categories.append(category)
            else:
                user_categories.append(category)
        
        _log.info("\nDefault categories (%s):", len(default_categories))
        for cat in sorted(default_categories):
            color = self.category_colors.get(cat, "N/A")
            _log.info("  %s: %s", cat, color)
        
        _log.info("\nUser categories (%s):", len(user_categories))
        for cat in sorted(user_categories):
            color = self.category_colors.get(cat, "N/A")
            _log.info("  %s: %s", cat, color)
        
    def open_expenses_viewer(self):
        """Open the expenses viewer in a new window"""
        from expenses_viewer import ExpensesViewerWindow
        self.expenses_viewer = ExpensesViewerWindow(self.expenses, self)
        self.expenses_viewer.show()
        
    def create_bar_chart(self):
        """Create bar chart with proper bar and label alignment"""
        chart = QChart()
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QtGui.QBrush(QtGui.QColor("#FFFFFF")))
        chart.setPlotAreaBackgroundVisible(True)
        chart.setPlotAreaBackgroundBrush(QtGui.QBrush(QtGui.QColor("#f8fafc")))
        chart.setMargins(QtCore.QMargins(16, 12, 10, 18))

        # ---- BAR SERIES ----
        series = QBarSeries()
        series.setLabelsVisible(True)
        series.setLabelsFormat("$@value")
        series.setLabelsPosition(QBarSeries.LabelsOutsideEnd)

        # ---- BAR SET ----
        self.bar_set = QBarSet("Expenses")
        self.bar_set.setColor(QtGui.QColor("#0f766e"))
        self.bar_set.setBorderColor(QtGui.QColor("#115e59"))
        self.bar_set.setLabelColor(QtGui.QColor("#334155"))
        series.append(self.bar_set)

        chart.addSeries(series)

        # ---- X-AXIS ----
        self.axis_x = QBarCategoryAxis()
        font = QtGui.QFont("Inter", 8)
        font.setLetterSpacing(QtGui.QFont.PercentageSpacing, 100)

        self.axis_x.setLabelsFont(font)
        self.axis_x.setLabelsAngle(-45)  # Default angle
        self.axis_x.setGridLineVisible(False)
        self.axis_x.setLabelsColor(QtGui.QColor("#475569"))
        self.axis_x.setLinePen(QtGui.QPen(QtGui.QColor("#cbd5e1")))
        chart.addAxis(self.axis_x, QtCore.Qt.AlignBottom)
        series.attachAxis(self.axis_x)

        # ---- Y-AXIS ----
        self.axis_y = QValueAxis()
        self.axis_y.setLabelFormat("$%d")
        self.axis_y.setTickCount(6)
        self.axis_y.setGridLineVisible(True)
        self.axis_y.setMinorGridLineVisible(False)
        self.axis_y.setLabelsColor(QtGui.QColor("#475569"))
        self.axis_y.setGridLineColor(QtGui.QColor("#e5edf5"))
        self.axis_y.setLinePen(QtGui.QPen(QtGui.QColor("#cbd5e1")))
        chart.addAxis(self.axis_y, QtCore.Qt.AlignLeft)
        series.attachAxis(self.axis_y)

        # ---- CHART VIEW ----
        chart_view = QChartView(chart)
        chart_view.setRenderHint(QtGui.QPainter.Antialiasing)
        
        # Create axis titles
        self.x_axis_title = QtWidgets.QGraphicsTextItem("Months")
        self.x_axis_title.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
        self.x_axis_title.setDefaultTextColor(QtGui.QColor("#334155"))
        
        self.y_axis_title = QtWidgets.QGraphicsTextItem("Expenses")
        self.y_axis_title.setFont(QtGui.QFont("Inter", 9, QtGui.QFont.DemiBold))
        self.y_axis_title.setDefaultTextColor(QtGui.QColor("#334155"))
        
        # Now we can safely add to scene
        scene = chart_view.scene()
        scene.addItem(self.x_axis_title)
        scene.addItem(self.y_axis_title)
        
        # Override resizeEvent to position titles correctly
        def custom_resize_event(event):
            # Call parent resize event first
            QChartView.resizeEvent(chart_view, event)
            
            # Position axis titles after resize
            self.position_axis_titles(event)
        
        chart_view.resizeEvent = custom_resize_event

        chart_view.setMinimumHeight(280)
        chart_view.setMaximumHeight(320)
        chart.layout().invalidate()

        chart_view.setStyleSheet("border: none; background: transparent;")
        return chart_view

    def position_axis_titles(self, event):
        """Position axis titles correctly"""
        if not hasattr(self, 'bar_chart_widget') or not self.bar_chart_widget:
            return
        
        chart = self.bar_chart_widget.chart()
        if not chart:
            return
        
        plot_area = chart.plotArea()
        
        # Position X-Axis title: Bottom center, below x-axis labels
        if hasattr(self, 'x_axis_title'):
            rect = self.x_axis_title.boundingRect()
            
            # Calculate position - center horizontally, below x-axis
            x = plot_area.left() + (plot_area.width() / 2) - (rect.width() / 2)
            y = plot_area.bottom() + 50  # Position below x-axis labels
            
            self.x_axis_title.setPos(x, y)
            self.x_axis_title.setRotation(0)  # Horizontal text
        
        # Position Y-Axis title: Left side, vertical text, left of y-axis labels
        if hasattr(self, 'y_axis_title'):
            rect = self.y_axis_title.boundingRect()
            
            # Rotate 90 degrees counter-clockwise for vertical text
            self.y_axis_title.setRotation(-90)
            
            # Calculate position - left of y-axis, centered vertically
            x = plot_area.left() - 70 # Further left to be left of y-axis labels
            y = plot_area.top() + (plot_area.height() / 2) + (rect.width() / 2)
            
            self.y_axis_title.setPos(x, y)
        
    def create_pie_chart_card(self):
        """Create pie chart card with month selection"""
        card = QtWidgets.QGroupBox("")
        card.setStyleSheet("""
            QGroupBox {
                color: #2c3e50;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                margin-top: 0;
                padding-top: 0;
                background: white;
            }
        """)

        layout = QtWidgets.QVBoxLayout(card)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(6)
        
        # --- HEADER ROW: Title + Calendar icon ---

        header_row = QtWidgets.QHBoxLayout()

        # Title Label (same text as GroupBox but inside layout)
        section_title = QtWidgets.QLabel("Monthly Category Mix")
        section_title.setStyleSheet("""
            QLabel {
                background: transparent;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
                font-size: 16px;
                font-weight: 900;
            }
        """)
        header_row.addWidget(section_title)

        self.pie_title_label = QtWidgets.QLabel("Current Month")
        self.pie_title_label.setStyleSheet("""
            QLabel {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
                font-size: 12px;
                color: #334155;
                font-family: 'Inter', 'Segoe UI';
                font-weight: 800;
                padding: 4px 10px;
            }
        """)

        header_row.addWidget(self.pie_title_label)

        header_row.addStretch()

        month_btn = QtWidgets.QPushButton("Change Month")
        month_btn.setFixedHeight(32)
        month_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        month_btn.setStyleSheet("""
            QPushButton {
                background: #f8fafc;
                color: #0f766e;
                border: 1px solid #99f6e4;
                border-radius: 8px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 12px;
                font-weight: 900;
                padding: 0 12px;
            }
            QPushButton:hover { background: #ecfdf5; }
        """)
        month_btn.clicked.connect(self.open_month_year_picker)
        header_row.addWidget(month_btn)

        layout.addLayout(header_row)

                
        # Create pie chart
        self.pie_chart_widget = self.create_pie_chart()
        layout.addWidget(self.pie_chart_widget)
        # === CLICK RESULT TABLE ===
        
        self.pie_click_table = QtWidgets.QTableWidget()
        self.pie_click_table.setColumnCount(5)
        self.pie_click_table.setHorizontalHeaderLabels(
            ["Category", "Expenses", "Amount", "Frequency", "% of total"]
        )
        self.pie_click_table.verticalHeader().setVisible(False)

        # Increase width of Category and Expenses columns
        self.pie_click_table.setColumnWidth(0, 180)   # Category
        self.pie_click_table.setColumnWidth(1, 151)   # Expenses


        header = self.pie_click_table.horizontalHeader()

        # Remove all default Qt sorting arrows
        header.setSortIndicatorShown(False)
        self.pie_click_table.setSortingEnabled(True)
        header.setSectionsClickable(True)

        # Hide Qt arrow graphics for all columns
        header.setStyleSheet("""
            QHeaderView::down-arrow { width: 0; height: 0; }
            QHeaderView::up-arrow { width: 0; height: 0; }
        """)

        # Add arrow ONLY in the "Expenses" column (index 1)

        # Allow clicking only the Expenses column
        header.sectionClicked.connect(self.on_expense_header_clicked)


        self.pie_click_table.setRowCount(0)
        self.pie_click_table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                border: 1px solid #e5edf5;
                border-radius: 6px;
                color: #0f172a;
                font-family: 'Inter', 'Segoe UI';
                font-size: 11px;
            }
            QTableWidget::item {
                background: #ffffff;
                padding: 2px 4px;
            }
            QHeaderView::section {
                background: #172033;
                color: white;
                font-weight: 800;
                padding: 0px;
            }
        """)

        # Table container kept in memory for click-handler compatibility but not shown
        self.table_container = QtWidgets.QFrame()
        self.table_container.setFixedHeight(0)
        table_container_layout = QtWidgets.QVBoxLayout(self.table_container)
        table_container_layout.setContentsMargins(0, 0, 0, 0)
        table_container_layout.addWidget(self.pie_click_table)
        self.table_container.hide()
        self.pie_click_table.hide()



        # Summary label
        self.pie_summary_label = QtWidgets.QLabel("Total: $0.00 | Categories: 0")
        self.pie_summary_label.setStyleSheet("""
            QLabel {
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 8px 10px;
                font-family: 'Inter', 'Segoe UI';
                font-size: 12px;
                font-weight: 800;
                color: #334155;
            }
        """)
        self.pie_summary_label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(self.pie_summary_label)
        
        return card
    
    def open_month_picker_dialog(self):
        """Open dialog where user selects only month & year"""

        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Select Month & Year")
        dialog.setModal(True)
        dialog.resize(300, 150)

        layout = QtWidgets.QVBoxLayout(dialog)

        # Year selector
        year_combo = QtWidgets.QComboBox()
        current_year = datetime.now().year
        for y in range(current_year - 5, current_year + 1):
            year_combo.addItem(str(y))
        year_combo.setCurrentText(str(current_year))

        # Month selector
        month_combo = QtWidgets.QComboBox()
        months = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ]
        month_combo.addItems(months)
        month_combo.setCurrentIndex(datetime.now().month - 1)

        layout.addWidget(QtWidgets.QLabel("Select Year:"))
        layout.addWidget(year_combo)
        layout.addWidget(QtWidgets.QLabel("Select Month:"))
        layout.addWidget(month_combo)

        # Buttons
        btns = QtWidgets.QHBoxLayout()
        ok_btn = QtWidgets.QPushButton("OK")
        cancel_btn = QtWidgets.QPushButton("Cancel")
        
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        layout.addLayout(btns)

        def apply_selection():
            self.selected_year = int(year_combo.currentText())
            self.selected_month = month_combo.currentIndex() + 1
            dialog.accept()
            self.update_pie_chart()

        def cancel():
            dialog.reject()

        ok_btn.clicked.connect(apply_selection)
        cancel_btn.clicked.connect(cancel)

        dialog.exec_()

    # Add this method to your ExpensesTab class
    def open_pdf_export_dialog(self):
        """Open the professional PDF/Excel export dialog for expenses"""
        try:
            # Collect available dates from expenses for the preview
            available_dates = []
            for expense in self.expenses:
                try:
                    expense_date = datetime.strptime(expense.get('date', ''), "%m-%d-%Y")
                    available_dates.append(expense_date)
                except ValueError:
                    continue
            
            dialog = ExpensesExportDialog(self, available_dates)  # Use new dialog
            result = dialog.exec_()
            
            # Only perform export if dialog was accepted AND has export parameters
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                # Get export parameters and perform actual export
                export_params = dialog._export_params
                if export_params["type"] == "pdf":
                    self.perform_pdf_export(export_params)
                elif export_params["type"] == "excel":
                    self.perform_excel_export(export_params)
                        
        except Exception as e:
            _log.warning("Error opening export dialog: %s", e)
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error opening export dialog: {str(e)}")
    
    def perform_excel_export(self, export_params):
        """Perform Excel export for expenses based on parameters"""
        try:
            # Filter expenses based on export parameters
            expenses_to_export = []
            
            for expense in self.expenses:
                try:
                    # Parse expense date - handle MM-dd-yyyy format
                    expense_datetime = None
                    date_str = expense.get('date', '')
                    
                    if not date_str:
                        continue
                    
                    # Try MM-dd-yyyy format first
                    try:
                        expense_datetime = datetime.strptime(date_str, "%m-%d-%Y")
                    except ValueError:
                        # Try alternative formats if needed
                        date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y"]
                        for date_format in date_formats:
                            try:
                                expense_datetime = datetime.strptime(date_str, date_format)
                                break
                            except ValueError:
                                continue
                    
                    # If we still couldn't parse the date, EXCLUDE the expense
                    if expense_datetime is None:
                        _log.info("(converted from print, see git history)")
                        continue
                    
                    include_expense = False
                    
                    if export_params["range"] == "all":
                        include_expense = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        # Convert to date objects for comparison
                        expense_date_only = expense_datetime.date()
                        
                        # Ensure both from_date and to_date are date objects
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        # Check if expense date is within the range (inclusive)
                        if from_date_only <= expense_date_only <= to_date_only:
                            include_expense = True
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if expense_datetime.month == month and expense_datetime.year == year:
                            include_expense = True
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if expense_datetime.year == year:
                            include_expense = True
                    
                    if include_expense:
                        expenses_to_export.append(expense)
                        
                except Exception as e:
                    _log.warning("Error processing expense: %s", e)
                    continue
            
            if not expenses_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No expenses found matching the selected criteria.")
                return
            
            _log.info("Excel Export: Found %s expenses to export", len(expenses_to_export))
            
            # Generate the combined Excel
            self.generate_combined_excel(expenses_to_export, export_params)
                
        except Exception as e:
            _log.warning("Error performing Excel export: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during Excel export: {str(e)}")

    def generate_combined_excel(self, expenses, export_params):
        """Generate a professional combined Excel report for expenses"""
        try:
            # Create export directory if it doesn't exist
            export_dir = Path.home() / "Downloads" / "Expense_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Expenses_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Expenses_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Expenses_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Expenses_{year}.xlsx"

            excel_path = export_dir / filename

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Header information
            # Header information
            ws.merge_cells('A1:H1')
            try:
                from main import Config as _Cfg
                _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co = 'MABS ENGINEERING LLC'
            ws['A1'] = f"{_co} - EXPENSES REPORT"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Date info
            generated_date = datetime.now().strftime("%m-%d-%Y")
            ws['A2'] = f"Generated: {generated_date}"
            
            # Export range info
            if export_params["range"] == "all":
                export_range_text = "All Expenses"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m-%d-%Y")
                to_date = export_params["to_date"].strftime("%m-%d-%Y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws['A3'] = f"Period: {export_range_text}"

            # Table headers
            headers = ["S.No.", "Date", "Expense Type", "Category", "Expense", "Vendor", "Amount", "Project"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Expense data
            for row_idx, expense in enumerate(expenses, 6):  # Start at row 6
                # Convert dates to MM-dd-YYYY format
                date_str = expense.get('date', '')
                try:
                    # Try to parse MM-dd-yyyy format
                    expense_date = datetime.strptime(date_str, "%m-%d-%Y")
                    display_date = expense_date.strftime("%m-%d-%Y")
                except:
                    try:
                        expense_date = datetime.strptime(date_str, "%m/%d/%Y")
                        display_date = expense_date.strftime("%m-%d-%Y")
                    except:
                        display_date = date_str
                
                # Get Expense Type
                expense_type = expense.get('expense_type', '')
                if not expense_type:
                    expense_type = "Other Expenses"
                
                # Get Category
                sub_category = expense.get('Category', expense.get('type', ''))
                
                # Get expense name
                expense_name = expense.get('expense_name', '')
                if not expense_name:
                    # Auto-generate if not provided
                    category = expense.get('type', '')
                    category_expenses = [e for e in expenses if e.get('type') == category]
                    expense_number = category_expenses.index(expense) + 1
                    expense_name = f"{expense_number}"
                
                data = [
                    row_idx - 5,
                    display_date,
                    self.remove_emojis(expense_type),
                    self.remove_emojis(sub_category),
                    self.remove_emojis(
                        expense_name[:25] + "..." if len(expense_name) > 25 else expense_name
                    ),
                    self.remove_emojis(expense.get('vendor', '')),
                    expense.get('amount', 0),
                    self.remove_emojis(expense.get('project', '')),
                ]

                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Style for sequential number column
                    if col == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    
                    # Style for amount column
                    if col == 7:  # Amount column
                        if isinstance(value, (int, float)):
                            cell.number_format = '"$"#,##0.00'
                            if value > 1000:
                                cell.font = Font(color="FF0000", bold=True)
                            elif value > 500:
                                cell.font = Font(color="FFA500")

            # Auto-adjust column widths
            column_widths = {
                1: 8,   # S.No.
                2: 12,  # Date
                3: 33,  # Expense Type
                4: 30,  # Category
                5: 30,  # Expense
                6: 22,  # Vendor
                7: 15,  # Amount
                8: 30,  # Project
            }
            
            for col_idx in range(1, len(headers) + 1):
                if col_idx in column_widths:
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = column_widths[col_idx]
                else:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[column_letter]:
                        if cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Add alternating row colors
            for row in range(6, ws.max_row + 1):
                if row % 2 == 0:  # Even rows
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.fill.start_color.index == '00000000':  # Default fill
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

            # Save the workbook
            wb.save(str(excel_path))

            # Open the Excel file
            if self.open_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"The Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            _log.warning("Error generating combined Excel: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error", 
                                        f"Error generating Excel: {str(e)}")
            
    # Add this method to your ExpensesTab class
    def perform_pdf_export(self, export_params):
        """Perform the actual PDF export based on parameters"""
        try:
            # Filter expenses based on export parameters
            expenses_to_export = []
            
            for expense in self.expenses:
                try:
                    # Parse expense date - handle MM-dd-yyyy format (your actual format)
                    expense_datetime = None
                    date_str = expense.get('date', '')
                    
                    if not date_str:
                        continue
                    
                    # DEBUG: Print what we're trying to parse
                    _log.debug("(converted from print, see git history)")
                    
                    # Try MM-dd-yyyy format first (your stored format)
                    try:
                        expense_datetime = datetime.strptime(date_str, "%m-%d-%Y")
                        _log.debug("DEBUG: Successfully parsed as MM-dd-yyyy: %s", expense_datetime)
                    except ValueError as e1:
                        _log.debug("DEBUG: Failed to parse as MM-dd-yyyy: %s", e1)
                        
                        # Try alternative formats if needed
                        date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%d/%m/%Y"]
                        for date_format in date_formats:
                            try:
                                expense_datetime = datetime.strptime(date_str, date_format)
                                _log.debug("DEBUG: Successfully parsed as %s: %s", date_format, expense_datetime)
                                break
                            except ValueError:
                                continue
                    
                    # If we still couldn't parse the date, EXCLUDE the expense
                    if expense_datetime is None:
                        _log.info("(converted from print, see git history)")
                        continue
                    
                    include_expense = False
                    
                    if export_params["range"] == "all":
                        include_expense = True
                        _log.debug('(debug log, see original)')
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        # Convert to date objects for comparison
                        expense_date_only = expense_datetime.date()
                        
                        # Ensure both from_date and to_date are date objects (not datetime)
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        # Check if expense date is within the range (inclusive)
                        if from_date_only <= expense_date_only <= to_date_only:
                            include_expense = True
                            _log.debug("DEBUG: Including expense %s in range %s to %s", expense_date_only, from_date_only, to_date_only)
                        else:
                            _log.debug("DEBUG: Excluding expense with date %s - not in export range %s to %s", expense_date_only, from_date_only, to_date_only)
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if expense_datetime.month == month and expense_datetime.year == year:
                            include_expense = True
                            _log.debug("DEBUG: Including expense for month %s/%s", month, year)
                        else:
                            _log.debug("DEBUG: Excluding expense with date %s - not in month %s/%s", expense_datetime, month, year)
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if expense_datetime.year == year:
                            include_expense = True
                            _log.debug("DEBUG: Including expense for year %s", year)
                        else:
                            _log.debug("DEBUG: Excluding expense with date %s - not in year %s", expense_datetime, year)
                    
                    if include_expense:
                        expenses_to_export.append(expense)
                        
                except Exception as e:
                    _log.warning("Error processing expense: %s", e)
                    import traceback
                    traceback.print_exc()
                    continue
            
            _log.debug("DEBUG: Found %s expenses to export out of %s total", len(expenses_to_export), len(self.expenses))
            
            if not expenses_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No expenses found matching the selected criteria.")
                return
            
            _log.info("PDF Export: Found %s expenses to export", len(expenses_to_export))
            
            # Generate the combined PDF
            self.generate_combined_pdf(expenses_to_export, export_params)
                
        except Exception as e:
            _log.warning("Error performing PDF export: %s", e)
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during PDF export: {str(e)}")

    def generate_combined_pdf(self, expenses, export_params):
        """Generate a professional combined PDF report for expenses"""
        try:
            # Create export directory if it doesn't exist
            export_dir = Path.home() / "Downloads" / "Expense_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if export_params["range"] == "all":
                filename = f"All_Expenses_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"Expenses_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"Expenses_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"Expenses_{year}.pdf"

            pdf_path = export_dir / filename

            # Create PDF document with reduced margins
            doc = SimpleDocTemplate(str(pdf_path), pagesize=A4, 
                                topMargin=0.2*inch, bottomMargin=0.3*inch,
                                leftMargin=0.4*inch, rightMargin=0.4*inch)
            elements = []

            # Get styles
            styles = getSampleStyleSheet()
            
            # Header table style for generated date
            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#7f8c8d'),
                fontName='Helvetica'
            )
            
            # Main Title Style
            main_title_style = ParagraphStyle(
                'MainTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=20,  # Reduced spacing
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,  # Center aligned
                fontName='Helvetica-Bold'
            )
            
            # Statistics Style
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=15,  # Reduced spacing
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,  # Center aligned
                fontName='Helvetica-Bold'
            )

            # 1. Header with generated date on right
            generated_date = datetime.now().strftime("%m/%d/%Y")

            header_data = [
                ['', Paragraph(f"{generated_date}", header_style)]
            ]

            header_table = Table(header_data, colWidths=[6.2*inch, 0.9*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))

            elements.append(header_table)
            
            # 2. Main Title: "MABS ENGINEERING EXPENSE REPORT"
            try:
                from main import Config as _Cfg
                _co = _Cfg.COMPANY.get('name', 'MABS Engineering LLC').upper()
            except Exception:
                _co = 'MABS ENGINEERING LLC'
            main_title = Paragraph(f"{_co} EXPENSES REPORT", main_title_style)
            elements.append(main_title)
            
            # Calculate statistics
            total_expenses = len(expenses)
            total_amount = sum(expense.get('amount', 0) for expense in expenses)
            
            # 3. Statistics (centered)
            stats_text = f"Expense Entries: {total_expenses}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Total Expense: ${total_amount:,.2f}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            # 4. Export range info (left-aligned)
            # compute export_range_text (needed for left aligned display)
            if export_params["range"] == "all":
                export_range_text = "All Expenses"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%Y")
                to_date = export_params["to_date"].strftime("%m/%d/%Y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"

            export_para = Paragraph(
                f"{export_range_text}",
                ParagraphStyle(
                    'ExportLeft',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor=colors.HexColor('#2c3e50'),
                    fontName='Helvetica-Bold',
                    alignment=0,
                    leftIndent=10,
                    firstLineIndent=0,
                    spaceBefore=2,
                    spaceAfter=12
                )
            )

            export_table = Table(
                [[export_para]],
                colWidths=[7.8*inch]  # full width, touches left margin
            )

            export_table.setStyle(TableStyle([
                ('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),
                ('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),
                ('BOTTOMPADDING',(0,0),(-1,-1),0),
            ]))

            elements.append(export_table)
            elements.append(Spacer(1,10))


            # 5. Expenses Table
            if expenses:
                cell_s = ParagraphStyle(
                    'CellStyle',
                    parent=styles['Normal'],
                    fontSize=7,
                    fontName='Helvetica',
                    textColor=colors.HexColor('#2c3e50'),
                    alignment=1,
                    wordWrap='CJK',
                )
                hdr_s = ParagraphStyle(
                    'HdrStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    fontName='Helvetica-Bold',
                    textColor=colors.whitesmoke,
                    alignment=1,
                    wordWrap='CJK',
                )
                # Prepare table data - UPDATED to include Expense Type and Category
                table_data = [[
                    Paragraph("Date", hdr_s),
                    Paragraph("Expense Type", hdr_s),
                    Paragraph("Category", hdr_s),
                    Paragraph("Expense", hdr_s),
                    Paragraph("Vendor", hdr_s),
                    Paragraph("Amount", hdr_s),
                ]]

                for expense in expenses:
                    # Get date in MM-dd-yyyy format (as stored in your expenses)
                    date_str = expense.get('date', '')

                    # Try to format the date properly for display
                    try:
                        # Try to parse MM-dd-yyyy format
                        expense_date = datetime.strptime(date_str, "%m-%d-%Y")
                        # Display as MM/dd/yyyy in PDF
                        display_date = expense_date.strftime("%m/%d/%Y")
                    except ValueError:
                        # Try alternative formats if the first fails
                        try:
                            # Try M/d/yyyy format
                            expense_date = datetime.strptime(date_str, "%m/%d/%Y")
                            display_date = expense_date.strftime("%m/%d/%Y")
                        except ValueError:
                            # If all parsing fails, use the original string
                            display_date = date_str

                    # Get Expense Type - AUTO-SET TO "Other Expenses" IF EMPTY
                    expense_type = expense.get('expense_type', '')
                    if not expense_type:
                        expense_type = "Other Expenses"

                    # Get Category
                    sub_category = expense.get('sub_category', expense.get('type', ''))

                    # Get expense name or use auto-generated number
                    expense_name = expense.get('expense_name', '')
                    if not expense_name:
                        # Auto-generate if not provided
                        category = expense.get('type', '')
                        category_expenses = [e for e in expenses if e.get('type') == category]
                        expense_number = category_expenses.index(expense) + 1
                        expense_name = f"{expense_number}"

                    # Remove emojis from all text fields
                    clean_expense_type = self.remove_emojis(expense_type)
                    clean_sub_category = self.remove_emojis(sub_category)
                    clean_expense_name = self.remove_emojis(expense_name)
                    vendor = self.remove_emojis(expense.get('vendor', ''))

                    # Get amount
                    amount = expense.get('amount', 0)

                    table_data.append([
                        Paragraph(display_date, cell_s),
                        Paragraph(clean_expense_type, cell_s),
                        Paragraph(clean_sub_category, cell_s),
                        Paragraph(clean_expense_name, cell_s),
                        Paragraph(vendor, cell_s),
                        Paragraph(f"${int(amount)}" if hasattr(Currency, 'format_whole') else format_amount_no_commas(amount), cell_s),
                    ])
                                    
                # Create table with adjusted column widths for 6 columns
                col_widths = [
                    0.9*inch,   # Date
                    1.8*inch,   # Expense Type
                    1.3*inch,   # Category
                    1.5*inch,   # Expense
                    1.2*inch,   # Vendor
                    0.9*inch,   # Amount
                ]
                
                expense_table = Table(table_data, colWidths=col_widths)
                expense_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (5, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (5, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (5, 0), 'CENTER'),  # Header centered
                ('FONTNAME', (0, 0), (5, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (5, 0), 8),
                ('BOTTOMPADDING', (0, 0), (5, 0), 8),
                
                # Data row styling
                ('BACKGROUND', (0, 1), (5, -1), colors.HexColor('#ffffff')),
                ('TEXTCOLOR', (0, 1), (5, -1), colors.HexColor('#2c3e50')),
                ('FONTNAME', (0, 1), (5, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (5, -1), 7),
                ('GRID', (0, 0), (5, -1), 0.5, colors.HexColor('#bdc3c7')),
                ('ROWBACKGROUNDS', (0, 1), (5, -1), [colors.HexColor('#f8f9fa'), colors.white]),
                
                # Row padding
                ('TOPPADDING', (0, 0), (5, -1), 4),
                ('BOTTOMPADDING', (0, 0), (5, -1), 4),
                
                # CENTER ALL COLUMNS - REMOVED COLUMN-SPECIFIC ALIGNMENTS
                ('ALIGN', (0, 1), (5, -1), 'CENTER'),  # Center ALL data cells
                
                # Enable word wrap for all columns (optional but recommended for centered text)
                ('WORDWRAP', (0, 1), (5, -1), True),
                
                # Vertical alignment - center vertically as well
                ('VALIGN', (0, 0), (5, -1), 'MIDDLE'),
            ]))
                
                elements.append(expense_table)
            else:
                no_data_style = ParagraphStyle(
                    'NoData',
                    parent=styles['Normal'],
                    fontSize=12,
                    textColor=colors.HexColor('#7f8c8d'),
                    alignment=1
                )
                elements.append(Paragraph("No expenses found for the selected criteria.", no_data_style))

            # Build PDF
            doc.build(elements)

            # Open the PDF
            if self.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"The PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            _log.warning("Error generating combined PDF: %s", e)
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", 
                                        f"Error generating PDF: {str(e)}")
            
    # Add this helper method to remove emojis
    def remove_emojis(self, text):
        """Remove emojis and special characters from text"""
        if not text:
            return text
        
        # Common emoji replacements
        emoji_replacements = {
            '🏢': '', '🎯': '', '💳': '', '💻': '', '✈️': '', 
            '🖥️': '', '📄': '', '🏦': '', '💵': '', '📱': '',
            '🔗': '', '✅': '', '💰': '', '📋': '', '📊': '',
            '⚡': '', '🗑️': '', '✏️': '', '👁️': '', '📅': '',
            '📂': '', '📝': '', '🥧': '', '🔍': '', '❌': '',
            '➕': '', '📤': '', '🚀': ''
        }
        
        # Replace emojis
        cleaned_text = text
        for emoji, replacement in emoji_replacements.items():
            cleaned_text = cleaned_text.replace(emoji, replacement)
        
        # Remove any remaining special characters and extra spaces
        cleaned_text = ' '.join(cleaned_text.split())
        
        return cleaned_text.strip()

    # Add this helper method to your ExpensesTab class
    def open_file(self, file_path):
        """Open file with default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == "Darwin":  # macOS
                subprocess.call(("open", file_path))
            elif platform.system() == "Windows":  # Windows
                os.startfile(file_path)
            else:  # linux variants
                subprocess.call(("xdg-open", file_path))
            return True
        except Exception as e:
            _log.warning("Error opening file: %s", e)
            return False
    
    def display_expense_in_table(self, exp):
        """Update table row based on selected expense number"""
        category = exp.get("type")
        
        # Get current pie chart month/year
        selected_month = self.pie_chart_month
        selected_year = self.pie_chart_year

        # Get all category expenses from CURRENT MONTH ONLY
        matched = []
        for expense in self.expenses:
            date_str = expense.get('date', '')
            if not date_str:
                continue
            
            try:
                expense_date = datetime.strptime(date_str, '%m-%d-%Y')  # Changed format
                if (expense.get('type') == category and 
                    expense_date.year == selected_year and 
                    expense_date.month == selected_month):
                    matched.append(expense)
            except:
                continue

        # Find selected expense index within filtered list
        selected_number = matched.index(exp) + 1 if exp in matched else 1
        total_freq = len(matched)

        self.pie_click_table.setRowCount(1)

        # Column 0 → Category
        self.pie_click_table.setItem(0, 0, QtWidgets.QTableWidgetItem(category))

        # Column 1 → Expenses (the number)
        exp_name = exp.get("expense_name", "")
        if exp_name:
            self.pie_click_table.setItem(0, 1, QtWidgets.QTableWidgetItem(exp_name))
        else:
            self.pie_click_table.setItem(0, 1, QtWidgets.QTableWidgetItem(str(selected_number)))

        # Column 2 → Amount - USE FORMATTING WITHOUT COMMAS
        amount = exp.get('amount', 0)
        # Format amount without commas
        try:
            amount_float = float(amount)
            if amount_float.is_integer():
                amount_display = f"${int(amount_float)}"
            else:
                amount_display = f"${amount_float:.2f}"
        except:
            amount_display = f"${amount}"
        
        amount_item = QtWidgets.QTableWidgetItem(amount_display)
        amount_item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.pie_click_table.setItem(0, 2, amount_item)

        # Column 3 → Frequency (from current month only)
        freq_item = QtWidgets.QTableWidgetItem(str(total_freq))
        freq_item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.pie_click_table.setItem(0, 3, freq_item)

        # Column 4 → % of total category (from current month only)
        total_amount = sum(e.get("amount", 0) for e in matched)
        percent = (exp.get("amount") / total_amount) * 100 if total_amount else 0
        percent_item = QtWidgets.QTableWidgetItem(f"{percent:.1f}%")
        percent_item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.pie_click_table.setItem(0, 4, percent_item)
        
        self.center_pie_table()
        
    def on_expense_header_clicked(self, col):
        """Show dropdown containing only expense numbers from current pie chart month"""

        # Only for the "Expenses" column
        if col != 1:
            return
            
        if self.pie_click_table.rowCount() == 0:
            return
        
        category = self.pie_click_table.item(0, 0).text()
        
        # Get current pie chart month/year
        selected_month = self.pie_chart_month
        selected_year = self.pie_chart_year

        # Filter expenses by category AND current pie chart month/year
        matched = []
        for expense in self.expenses:
            date_str = expense.get('date', '')
            if not date_str:
                continue
            
            try:
                expense_date = datetime.strptime(date_str, '%m-%d-%Y')
                if (expense.get('type') == category and 
                    expense_date.year == selected_year and 
                    expense_date.month == selected_month):
                    matched.append(expense)
            except:
                continue

        if len(matched) <= 1:
            return

        # Create dropdown menu
        menu = QtWidgets.QMenu(self)

        for exp in matched:
            exp_name = exp.get("expense_name", "").strip()

            # If user manually entered an expense name → show it
            if exp_name:
                action = menu.addAction(exp_name)
            else:
                # Use index within this filtered list (1, 2, 3 for current month only)
                number = matched.index(exp) + 1
                action = menu.addAction(str(number))

            action.expense = exp

        def on_selected(action):
            self.display_expense_in_table(action.expense)

        menu.triggered.connect(on_selected)

        header = self.pie_click_table.horizontalHeader()

        # Position dropdown under column 1
        x = header.sectionPosition(1)
        y = header.height()

        menu.exec_(header.mapToGlobal(QtCore.QPoint(x, y)))
        
    def create_pie_chart(self):
        """Create a professional pie chart"""
        chart = QChart()
        chart.setMargins(QtCore.QMargins(8, 8, 8, 8))
        chart.layout().setContentsMargins(0, 0, 0, 0)

        chart.setTheme(QChart.ChartThemeLight)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QtGui.QBrush(QtGui.QColor("#ffffff")))
        
        # Create pie series
        self.pie_series = QPieSeries()
        self.pie_series.clicked.connect(self.on_pie_slice_clicked)
        self.pie_series.setPieSize(0.82)
        self.pie_series.setHoleSize(0.40)
        self.pie_series.setHorizontalPosition(0.5)
        self.pie_series.setVerticalPosition(0.5)

        self.pie_series.setLabelsVisible(True)

        # Add sample slice
        slice = self.pie_series.append("No Data", 1)
        slice.setColor(QtGui.QColor("#bdc3c7"))
        slice.setLabelVisible(False)

        chart.addSeries(self.pie_series)

        # Legend — right side, vertical list
        legend = chart.legend()
        legend.setVisible(True)
        legend.setAlignment(QtCore.Qt.AlignRight)
        legend.setMarkerShape(QLegend.MarkerShapeCircle)
        legend.setFont(QtGui.QFont("Inter", 8, QtGui.QFont.Medium))
        legend.setLabelColor(QtGui.QColor("#1E293B"))
        legend.setBorderColor(QtGui.QColor("#ffffff"))
        legend.setBackgroundVisible(True)
        legend.setColor(QtGui.QColor("#ffffff"))

        chart_view = PieChartView(chart)
        chart_view.pie_series = self.pie_series
        chart_view.pie_slice_categories = self.pie_slice_categories
        chart_view.setRenderHint(QtGui.QPainter.Antialiasing)
        chart_view.setMinimumHeight(420)
        chart_view.setMaximumHeight(520)
        chart_view.setStyleSheet("background: transparent; border: none;")

        return chart_view
    
    def show_date_range_dialog(self):
        """Show date range selection dialog - styled like balance sheet tab"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📅 Select Date Range")
        dialog.setModal(True)
        dialog.resize(400, 200)
        dialog.setStyleSheet("QDialog { background: #f5f6fa; }")

        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Title
        title = QtWidgets.QLabel("Select Date Range")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title)

        # Date inputs
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)

        # Restore previously chosen dates if available
        current_from_date = QtCore.QDate.currentDate().addMonths(-1)
        current_to_date = QtCore.QDate.currentDate()
        if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date

        date_style = """
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """

        self.from_date_edit = QtWidgets.QDateEdit()
        self.from_date_edit.setDate(current_from_date)
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.from_date_edit.setStyleSheet(date_style)
        self.from_date_edit.wheelEvent = lambda e: e.ignore()
        self.from_date_edit.stepBy = lambda x: None

        self.to_date_edit = QtWidgets.QDateEdit()
        self.to_date_edit.setDate(current_to_date)
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("MM-dd-yyyy")
        self.to_date_edit.setStyleSheet(date_style)
        self.to_date_edit.wheelEvent = lambda e: e.ignore()
        self.to_date_edit.stepBy = lambda x: None

        form_layout.addRow("From Date:", self.from_date_edit)
        form_layout.addRow("To Date:", self.to_date_edit)
        layout.addLayout(form_layout)

        btn_style_clear = """
            QPushButton {
                background: #95a5a6; color: white; font-weight: bold;
                border-radius: 8px; border: none; font-size: 14px;
            }
            QPushButton:hover { background: #7f8c8d; }
        """
        btn_style_apply = """
            QPushButton {
                background: #27ae60; color: white; font-weight: bold;
                border-radius: 8px; border: none; font-size: 14px;
            }
            QPushButton:hover { background: #2ecc71; }
        """

        clear_btn = QtWidgets.QPushButton("Clear Filter")
        clear_btn.setMinimumHeight(40)
        clear_btn.setStyleSheet(btn_style_clear)

        apply_btn = QtWidgets.QPushButton("Apply Filter")
        apply_btn.setMinimumHeight(40)
        apply_btn.setStyleSheet(btn_style_apply)

        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        button_layout.addWidget(apply_btn)
        layout.addLayout(button_layout)

        apply_btn.clicked.connect(lambda: self.apply_date_range_filter(dialog))
        clear_btn.clicked.connect(lambda: self.clear_date_range_filter(dialog))

        dialog.exec_()

    def apply_date_range_filter(self, dialog=None):
        """Apply date range filter"""
        from_date_qdate = self.from_date_edit.date()
        to_date_qdate = self.to_date_edit.date()

        self.current_from_date = from_date_qdate
        self.current_to_date = to_date_qdate

        from_label = from_date_qdate.toString("MM-dd-yy")
        to_label = to_date_qdate.toString("MM-dd-yy")
        configure_filter_button(
            self.date_range_button,
            f"{from_label} to {to_label}",
            active=True,
            height=40,
        )

        self.filter_expenses()
        if dialog:
            dialog.accept()

    def clear_date_range_filter(self, dialog=None):
        """Clear date range filter AND reset all other filters"""
        configure_filter_button(self.date_range_button, height=40)
        
        # Clear the stored date objects
        if hasattr(self, 'current_from_date'):
            del self.current_from_date
        if hasattr(self, 'current_to_date'):
            del self.current_to_date
        
        # ✅ CLEAR ALL OTHER FILTERS TOO
        # Block signals to prevent multiple filter calls
        self.categories_filter_combo.blockSignals(True)
        self.search_edit.blockSignals(True)
        
        # Reset category filter to "All Categories"
        self.categories_filter_combo.setCurrentText("All Categories")
        
        # Clear search text
        self.search_edit.clear()
        
        # Restore signals
        self.categories_filter_combo.blockSignals(False)
        self.search_edit.blockSignals(False)
        
        # ✅ Now trigger filtering with all filters cleared
        # Update categories dropdown to show ALL categories
        self.categories_filter_combo.blockSignals(True)
        self.update_categories_filter(self.expenses)  # Show all categories
        self.categories_filter_combo.blockSignals(False)
        
        # Display all expenses
        self.display_filtered_expenses(self.expenses)
        self.update_statistics(self.expenses)
        if dialog:
            dialog.accept()

    def create_table_section(self):
        """Create expense table section with search and filters - UPDATED with date range and categories filter"""
        frame = QtWidgets.QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
        """)
        layout = QtWidgets.QVBoxLayout(frame)
        layout.setSpacing(10)
        layout.setContentsMargins(12, 12, 12, 12)

        # alias so the rest of the method doesn't change
        table_layout = layout
        
        # 🔍 Integrated Search and Filter Section
        search_filter_frame = QtWidgets.QFrame()
        search_filter_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        
        search_filter_layout = QtWidgets.QHBoxLayout(search_filter_frame)
        search_filter_layout.setSpacing(10)
        search_filter_layout.setContentsMargins(0, 0, 0, 0)
        
        # --------------------------------------
        # LEFT SIDE → Date Range Filter + Search
        # --------------------------------------
        left_section = QtWidgets.QHBoxLayout()
        left_section.setSpacing(10)
        left_section.setContentsMargins(0, 0, 0, 0)
        
        # Calendar Date Range Filter Button
        self.date_range_button = configure_filter_button(QtWidgets.QPushButton(), height=40)
        self.date_range_button.clicked.connect(self.show_date_range_dialog)
        
        # Search Bar
        self.search_edit = QtWidgets.QLineEdit()
        # In create_table_section() method, update the search_edit placeholder text:
        self.search_edit.setPlaceholderText("Search by category, description, project, date, vendor, or amount...")
        self.search_edit.setMinimumHeight(40)
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
                background: white;
                min-width: 430px;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #fafbfc;
            }
        """)
        self.search_edit.textChanged.connect(self.on_search_text_changed)  # Changed connection
        
        left_section.addWidget(self.date_range_button)
        left_section.addWidget(self.search_edit)
        
        search_filter_layout.addLayout(left_section)
        
        # --------------------------------------
        # RIGHT SIDE → Categories Filter
        # --------------------------------------
        search_filter_layout.addStretch(1)
        
        filter_right = QtWidgets.QHBoxLayout()
        
        filter_label = QtWidgets.QLabel("Category:")
        filter_label.setStyleSheet("font-weight: 800; color: #334155; font-size: 13px; font-family: 'Inter', 'Segoe UI'; padding: 5px 0px;")
        filter_right.addWidget(filter_label)
        
        self.categories_filter_combo = QtWidgets.QComboBox()
        # Initialize with "All Categories" only - actual categories will be populated when data loads
        self.categories_filter_combo.addItem("All Categories")
        self.categories_filter_combo.setMinimumHeight(40)
        
        # Set to show 6 items with vertical scrolling for more (compact size)
        self.categories_filter_combo.setMaxVisibleItems(6)
        self.categories_filter_combo.view().setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        
        # Compact height for dropdown (6 items)
        item_height = 28  # Smaller item height
        dropdown_height = (item_height * 6) + 4  # 6 items + small padding
        
        self.categories_filter_combo.setStyleSheet("""
            QComboBox {
                padding: 6px 10px;
                border: 1px solid #c8c8c8;
                border-radius: 8px;
                background: white;
                font-size: 13px;
                font-family: 'Inter', 'Segoe UI';
                min-width: 170px;
            }

            QComboBox::drop-down {
                width: 20px;
                border-left: 1px solid #c8c8c8;
                background: #f2f2f2;
            }

            QComboBox::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }

            /* CLEAN DROP DOWN */
            QComboBox QAbstractItemView {
                border: 1px solid #c8c8c8;
                background: white;
                outline: none;
                padding: 0px;
            }

            QComboBox QAbstractItemView::item {
                padding: 6px 10px;
                border: none;
                font-size: 12px;
            }

            QComboBox QAbstractItemView::item:hover {
                background: #f5f5f5;
            }
        """)

        self.categories_filter_combo.currentTextChanged.connect(self.filter_expenses)
        self.categories_filter_combo.wheelEvent = lambda e: e.ignore()
        self.categories_filter_combo.keyPressEvent = lambda e, c=self.categories_filter_combo: (
            QtWidgets.QComboBox.keyPressEvent(c, e)
            if e.key() not in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down) or c.view().isVisible()
            else e.ignore()
        )
        self.categories_filter_combo.currentIndexChanged.connect(
            lambda: QtCore.QTimer.singleShot(0, self.categories_filter_combo.clearFocus))

        filter_right.addWidget(self.categories_filter_combo)
        
        search_filter_layout.addLayout(filter_right)
        
        # Attach to table section
        table_layout.addWidget(search_filter_frame)
        table_layout.setSpacing(1)   # reduce vertical gap

        # Results Counter
        results_frame = QtWidgets.QFrame()
        results_frame.setStyleSheet("background: transparent;")
        results_layout = QtWidgets.QHBoxLayout(results_frame)
        results_layout.addStretch()
        table_layout.addWidget(results_frame)
        
        # Expenses Table with UPDATED columns (REMOVED STATUS COLUMN)
        self.expenses_table = QtWidgets.QTableWidget()
        self.expenses_table.setColumnCount(8)
        self.expenses_table.setHorizontalHeaderLabels([
            "S.No", "Date", "Expense Type", "Category",
            "Expense Name", "Vendor", "Amount", "Actions"
        ])

        # Professional table styling with vertical grid lines - UPDATED
        self.expenses_table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                alternate-background-color: #f8fafc;
                border: none;
                border-radius: 8px;
                gridline-color: #e5edf5;
                font-size: 12px;
                font-family: 'Inter', 'Segoe UI', sans-serif;
                outline: none;
                selection-background-color: transparent;
            }
            QTableWidget::item {
                padding: 6px 8px;
                border-bottom: 1px solid #e5edf5;
                border-right: 1px solid #e5edf5;
                color: #0f172a;
                font-size: 11px;
                background: transparent;
            }
            QTableWidget::item:last {
                border-right: none;
            }
            QTableWidget::item:selected {
                background-color: #e8f7f5;
                color: #0f172a;
            }
            QTableWidget::item:selected:hover {
                background-color: #e8f7f5;
                color: #0f172a;
            }
            QTableWidget::item:hover {
                background: #f3faf9;
            }
            QHeaderView::section {
                background: #172033;
                color: white;
                font-weight: 800;
                font-size: 11px;
                padding: 9px 3px;
                border: none;
                border-right: 1px solid #374151;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QTableWidget QTableCornerButton::section {
                background: #172033;
                border: none;
                border-bottom: 1px solid #374151;
                border-right: 1px solid #374151;
            }
        """)
        
        # Match Project Dashboard ledger grid.
        self.expenses_table.setShowGrid(True)
        self.expenses_table.setGridStyle(QtCore.Qt.SolidLine)
        
        # Table properties for compact look
        self.expenses_table.horizontalHeader().setStretchLastSection(False)
        self.expenses_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.expenses_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.expenses_table.setAlternatingRowColors(True)
        self.expenses_table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.expenses_table.setFixedHeight(160)
        self.expenses_table.viewport().setAutoFillBackground(True)
        self.expenses_table.verticalHeader().setVisible(False)
        self.expenses_table.setSortingEnabled(True)
        self.expenses_table.setWordWrap(True)
        self.expenses_table.setAccessibleName("Expenses table")
        self.expenses_table.setAccessibleDescription(
            "List of expense records. Hover over a cell to see the full text.")
        self.expenses_table.setTextElideMode(QtCore.Qt.ElideNone)
        
        # Header properties
        header = self.expenses_table.horizontalHeader()
        header.setDefaultAlignment(QtCore.Qt.AlignCenter)
        header.setHighlightSections(False)
        header.setFixedHeight(42)
        header.setSortIndicatorShown(True)
        
        # col 0 S.No, 1 Date, 2 Expense Type, 3 Category,
        # 4 Expense Name, 5 Vendor, 6 Amount, 7 Actions
        self.expenses_table.setColumnWidth(0, 55)
        self.expenses_table.setColumnWidth(1, 110)
        self.expenses_table.setColumnWidth(2, 290)   # Expense Type
        self.expenses_table.setColumnWidth(3, 250)   # Category
        self.expenses_table.setColumnWidth(5, 200)   # Vendor
        self.expenses_table.setColumnWidth(6, 100)
        self.expenses_table.setColumnWidth(7, 190)
        # Fixed columns
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)
        header.setSectionResizeMode(7, QtWidgets.QHeaderView.Fixed)
        # Expense Name stretches to fill remaining horizontal space
        header.setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        # Type, Category, Vendor are interactive (user can resize)
        for col in (2, 3, 5):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Interactive)

        table_layout.addWidget(self.expenses_table)

        _pg_s = """QPushButton { background:#ffffff; color:#334155; border:1px solid #e2e8f0;
            border-radius:6px; font-size:12px; font-weight:700;
            min-width:32px; min-height:28px; padding:0 8px; }
            QPushButton:hover { background:#f1f5f9; border-color:#cbd5e1; }
            QPushButton:disabled { color:#cbd5e1; }"""
        exp_pg_frame = QtWidgets.QFrame()
        exp_pg_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        exp_pg_hbox = QtWidgets.QHBoxLayout(exp_pg_frame)
        exp_pg_hbox.setContentsMargins(4, 6, 4, 6)
        exp_pg_hbox.setSpacing(6)
        self._exp_info_lbl = QtWidgets.QLabel("")
        self._exp_info_lbl.setStyleSheet(
            "color:#94a3b8; font-size:11px; font-weight:600; background:transparent; border:none;")
        exp_pg_hbox.addWidget(self._exp_info_lbl)
        exp_pg_hbox.addStretch()
        self._exp_prev_btn = QtWidgets.QPushButton("‹")
        self._exp_prev_btn.setStyleSheet(_pg_s)
        self._exp_prev_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._exp_prev_btn.clicked.connect(self._exp_go_prev)
        exp_pg_hbox.addWidget(self._exp_prev_btn)
        self._exp_page_btns_layout = QtWidgets.QHBoxLayout()
        self._exp_page_btns_layout.setSpacing(4)
        exp_pg_hbox.addLayout(self._exp_page_btns_layout)
        self._exp_next_btn = QtWidgets.QPushButton("›")
        self._exp_next_btn.setStyleSheet(_pg_s)
        self._exp_next_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self._exp_next_btn.clicked.connect(self._exp_go_next)
        exp_pg_hbox.addWidget(self._exp_next_btn)
        self._exp_pg_style = _pg_s
        table_layout.addWidget(exp_pg_frame)

        return frame

    def _ledger_item(self, text="", alignment=QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft):
        item = QtWidgets.QTableWidgetItem(str(text or ""))
        item.setTextAlignment(alignment)
        item.setToolTip(str(text or ""))
        return item

    def _tint_color(self, color, alpha=24):
        qcolor = QtGui.QColor(color or "#64748b")
        qcolor.setAlpha(alpha)
        return qcolor

    def _style_ledger_row(self, row, expense):
        table_font = QtGui.QFont("Inter", 9)
        text_brush = QtGui.QBrush(QtGui.QColor("#0f172a"))
        white_brush = QtGui.QBrush(QtGui.QColor("#ffffff"))

        for col in range(7):
            item = self.expenses_table.item(row, col)
            if not item:
                continue
            item.setFont(table_font)
            item.setForeground(text_brush)
            item.setBackground(white_brush)
            # Preserve per-column alignment already set by _populate_ledger_row

    def _format_expense_date(self, date_str):
        try:
            qd = QtCore.QDate.fromString(str(date_str or ""), "MM-dd-yyyy")
            return qd.toString("MM-dd-yyyy") if qd.isValid() else str(date_str or "")
        except Exception:
            return str(date_str or "")

    def _expense_project_number(self, expense):
        for key in ("project_number", "project_no", "project_id"):
            value = str(expense.get(key, "") or "").strip()
            if value:
                return value

        raw_project = str(expense.get("project", "") or "").strip()
        if not raw_project:
            return ""

        project_tab = getattr(self.main_window, "project_tab", None)
        projects = []
        for attr in ("cached_projects", "generated_projects"):
            projects.extend(getattr(project_tab, attr, []) or [])

        raw_lower = raw_project.lower()
        for project in projects:
            project_number = str(project.get("project_number", "") or "").strip()
            project_name = str(project.get("project_name", "") or "").strip()
            if raw_project == project_number:
                return project_number
            if project_number and raw_project.startswith(f"{project_number} - "):
                return project_number
            if project_name and raw_lower == project_name.lower():
                return project_number or raw_project

        return raw_project

    def _populate_ledger_row(self, row, expense, sno=None):
        category = expense.get('Category', expense.get('type', '')) or "Uncategorized"
        expense_type = expense.get('expense_type', '') or expense.get('type', '')
        expense_name = expense.get('expense_name', '') or ''
        amount = expense.get('amount', 0)

        # col: 0=S.No, 1=Date, 2=Expense Type, 3=Category, 4=Expense Name, 5=Vendor, 6=Amount, 7=Actions
        values = [
            str(sno if sno is not None else row + 1),
            self._format_expense_date(expense.get('date', '')),
            expense_type,
            category,
            expense_name,
            expense.get('vendor', ''),
            format_amount_no_commas(amount),
        ]

        for col, value in enumerate(values):
            alignment = QtCore.Qt.AlignVCenter | QtCore.Qt.AlignCenter
            self.expenses_table.setItem(row, col, self._ledger_item(value, alignment))

        self.expenses_table.setCellWidget(row, 7, self.create_enhanced_action_buttons(expense))
        self._style_ledger_row(row, expense)

    def on_search_text_changed(self, search_text):
        """Handle search text changes - intelligent filtering"""
        # Simply call filter_expenses which handles all filter logic
        # This will keep the current category filter and date range filter active
        self.filter_expenses()
        
    def update_categories_filter(self, filtered_expenses=None):
        """Update categories filter combo box with categories from filtered expenses"""
        # Use filtered expenses if provided, otherwise all expenses
        expenses_to_use = filtered_expenses if filtered_expenses is not None else self.expenses
        
        # Get unique categories from the provided expenses
        current_categories = sorted(set(
            expense.get('Category', expense.get('type', '')) 
            for expense in expenses_to_use 
            if expense.get('Category') or expense.get('type')
        ))
        
        # Store current selection
        current_selection = self.categories_filter_combo.currentText()
        
        # Check if the category list would actually change
        current_items = [self.categories_filter_combo.itemText(i) for i in range(self.categories_filter_combo.count())]
        new_items = ["All Categories"] + [cat for cat in current_categories if cat]
        
        # Only update if the list actually changed
        if set(current_items) != set(new_items):
            # Block signals to prevent recursion
            self.categories_filter_combo.blockSignals(True)
            
            self.categories_filter_combo.clear()
            self.categories_filter_combo.addItem("All Categories")
            
            # Add only categories that exist in the filtered data
            for category in current_categories:
                if category:  # Only add non-empty categories
                    self.categories_filter_combo.addItem(category)
            
            # Restore previous selection if it still exists, otherwise default to "All Categories"
            if current_selection in [self.categories_filter_combo.itemText(i) for i in range(self.categories_filter_combo.count())]:
                self.categories_filter_combo.setCurrentText(current_selection)
            else:
                self.categories_filter_combo.setCurrentText("All Categories")
            
            # Restore signals
            self.categories_filter_combo.blockSignals(False)
    def get_group_box_style(self):
        return """
            QGroupBox {
                font-family: 'Inter', 'Segoe UI';
                font-weight: 900;
                font-size: 16px;
                color: #0f172a;
                border: 1px solid #d8e2ec;
                border-radius: 8px;
                margin-top: 0px;
                padding-top: 12px;
                background: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 0 10px;
                color: #0f172a;
                background: white;
                font-weight: 900;
            }
        """
    
    def show_add_expense_dialog(self):
        """Show the add expense dialog"""
        dialog = AddExpenseDialog(self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            expense_data = dialog.get_expense_data()
            if expense_data:  # Only save if data is valid (not None)
                self.save_expense(expense_data)
        
    def show_edit_expense_dialog(self, expense_data):
        """Show the edit expense dialog"""
        dialog = AddExpenseDialog(self, expense_data)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            updated_data = dialog.get_expense_data()
            if updated_data:  # Only update if data is valid (not None)
                self.update_expense(expense_data, updated_data)
                
    def save_expense(self, expense_data):
        """Save expense data to Firebase with auto-generated expense numbers"""
        
        # Auto-generate expense number if expense_name is empty
        if not expense_data.get('expense_name'):
            category = expense_data.get('type', '')
            # Count existing expenses in the same category
            category_expenses = [e for e in self.expenses if e.get('type') == category]
            expense_number = len(category_expenses) + 1
            expense_data['expense_name'] = f"{expense_number}"
        
        # Ensure expense_type is set to "Other Expenses" if empty
        if not expense_data.get('expense_type'):
            expense_data['expense_type'] = "Other Expenses"
        
        # Get the category to ensure it has a color
        category = expense_data.get('Category', '')
        if category:
            # This will assign a color if not already assigned
            self.get_color_for_category(category)
        
        # Save to Firebase
        success = ExpensesFirebaseManager.save_expense(expense_data)
        if success:
            vendor = str(expense_data.get('vendor', '') or '').strip()
            if vendor:
                ExpensesFirebaseManager.save_vendor(vendor)
            # Save custom category colors
            self.save_custom_category_colors()
            self.show_success_message("Expense added successfully!")
            
            # Instead of reloading everything, add the new expense to current data
            # and update UI while preserving filters
            self.expenses.append(expense_data)
            
            # Store current filter state
            current_category_filter = self.categories_filter_combo.currentText()
            current_search_text = self.search_edit.text()
            date_range_active = "to" in self.date_range_button.text()
            
            # Update all UI components
            self.update_expenses_table()
            self.update_charts()
            
            # Apply filters again to include the new expense
            self.filter_expenses()
            
            # Restore search text if it was set
            self.search_edit.setText(current_search_text)
            
            # CHECK IF WE SHOULD SAVE TO BALANCE SHEET
            if expense_data.get('save_to_balance_sheet', True):
                self.save_to_balance_sheet_expenses(expense_data)
            
        else:
            # Add to local expenses
            self.expenses.append(expense_data)
            
            # Update all UI components
            self.update_expenses_table()
            self.update_charts()  # This will update both bar and pie charts
            self.update_statistics()
            
            # Apply filters again to include the new expense
            self.filter_expenses()
            
            # CHECK IF WE SHOULD SAVE TO BALANCE SHEET
            if expense_data.get('save_to_balance_sheet', True):
                self.save_to_balance_sheet_expenses(expense_data)
                        
    def save_to_balance_sheet_expenses(self, expense_data):
        """Save expense to balance sheet expenses list and Firebase"""
        balance_sheet_expense = {
            'date':          expense_data.get('date', ''),
            'expense_date':  expense_data.get('date', ''),
            # Preserve every original Expenses-tab field so the columns are correct on reload
            'expense_type':  expense_data.get('expense_type', ''),
            'Category':      expense_data.get('Category', ''),
            'expense_name':  expense_data.get('expense_name', '') or expense_data.get('Category', 'Expense'),
            'name':          expense_data.get('expense_name', '') or expense_data.get('Category', 'Expense'),
            'description':   expense_data.get('description', ''),
            'vendor':        expense_data.get('vendor', ''),
            'amount':        f"{self._safe_amount(expense_data.get('amount', 0)):.2f}",
            'year':          self._expense_year(expense_data),
            'created_at':    datetime.now(timezone.utc).isoformat(),
            'project':       expense_data.get('project', ''),
            'project_number': expense_data.get('project_number', ''),
            'firebase_id':   expense_data.get('firebase_id', ''),
        }
        
        # Try to find the balance sheet tab
        balance_tab = None
        if hasattr(self, 'main_window') and hasattr(self.main_window, 'balance_sheet_tab'):
            balance_tab = self.main_window.balance_sheet_tab
        
        if balance_tab and hasattr(balance_tab, 'expenses_data'):
            # Add to local data (transaction table)
            found_index = -1

            for i, exp in enumerate(balance_tab.expenses_data):
                if exp.get('firebase_id') == expense_data.get('firebase_id'):
                    found_index = i
                    break

            if found_index >= 0:
                # Update existing expense
                balance_tab.expenses_data[found_index] = balance_sheet_expense.copy()
                _log.info("Updated existing expense in balance sheet")
            else:
                # Add new expense
                balance_tab.expenses_data.append(balance_sheet_expense.copy())
                _log.info("Added new expense to balance sheet")
    
            # ===== CRITICAL FIX: Also add to annual data =====
            # ===== FIX: Update or insert in annual_expenses_data =====
            if hasattr(balance_tab, 'annual_expenses_data'):

                found_index = -1

                for i, exp in enumerate(balance_tab.annual_expenses_data):
                    if exp.get('firebase_id') == expense_data.get('firebase_id'):
                        found_index = i
                        break

                if found_index >= 0:
                    # Update existing
                    balance_tab.annual_expenses_data[found_index] = balance_sheet_expense.copy()
                    _log.info("Updated annual balance sheet expense")
                else:
                    # Add new
                    balance_tab.annual_expenses_data.append(balance_sheet_expense.copy())
                    _log.info("Added annual balance sheet expense")
                    
            # Also save to Firebase using balance sheet's save method
            if hasattr(balance_tab, 'save_expense_to_firebase'):
                # Pass without old_data to create new entry
                balance_tab.save_expense_to_firebase(balance_sheet_expense)
            
            # Sort by creation date (newest first) for the default view
            balance_tab.sort_expenses_by_creation_date()
            
            # Update UI - update both annual summary and stats cards
            balance_tab.update_annual_summary()
            balance_tab.update_stats_cards()
            
            # If currently viewing expenses in balance sheet, refresh the table
            if hasattr(balance_tab, 'current_category') and balance_tab.current_category == "Expenses":
                balance_tab.filter_finance_entries()
                
            _log.info("Saved expense to balance sheet and Firebase: %s", balance_sheet_expense)
        else:
            _log.info("Balance sheet tab not accessible - expense not saved to balance sheet")
            
    def update_expense(self, old_expense_data, updated_expense_data):
        """Update existing expense"""
        # Ensure expense_type is set to "Other Expenses" if empty
        if not updated_expense_data.get('expense_type'):
            updated_expense_data['expense_type'] = "Other Expenses"
        
        # Auto-generate expense number if expense_name is empty
        if not updated_expense_data.get('expense_name'):
            category = updated_expense_data.get('type', '')
            # Count existing expenses in the same category
            category_expenses = [e for e in self.expenses if e != old_expense_data and e.get('type') == category]
            expense_number = len(category_expenses) + 1
            updated_expense_data['expense_name'] = f"{expense_number}"
        
        # Save vendor for future dropdown use
        vendor = str(updated_expense_data.get('vendor', '') or '').strip()
        if vendor:
            ExpensesFirebaseManager.save_vendor(vendor)

        # Update Firebase if available — route to the correct node
        if 'firebase_id' in old_expense_data and FIREBASE_AVAILABLE:
            try:
                _fid = old_expense_data['firebase_id']
                if old_expense_data.get('finance_source') == 'balance_sheet':
                    node = f'/balance_sheet_expenses/{_fid}'
                else:
                    node = f'/expenses/{_fid}'
                db.reference(node).update(updated_expense_data)
            except Exception as e:
                _log.warning("Error updating expense in Firebase: %s", e)
                self.show_warning_message("Failed to update in Firebase. Updated locally only.")
        
        # Update local data
        if old_expense_data in self.expenses:
            index = self.expenses.index(old_expense_data)
            self.expenses[index] = updated_expense_data
        
        # Only update balance sheet if checkbox is ticked
        if updated_expense_data.get('save_to_balance_sheet', True):
            self.update_balance_sheet_expense(old_expense_data, updated_expense_data)
        else:
            # If checkbox is NOT ticked, do NOTHING to balance sheet
            # The old entry (if any) remains completely unchanged
            _log.info("Checkbox not ticked - balance sheet entry unchanged")
        
        self._refresh_table_keep_page()
        self.update_charts()
        self.update_statistics()
        self.update_categories_filter()
    
    def update_balance_sheet_expense(self, old_expense_data, updated_expense_data):

        balance_tab = None
        if hasattr(self.main_window, 'balance_sheet_tab'):
            balance_tab = self.main_window.balance_sheet_tab

        if not balance_tab:
            _log.info("Balance sheet tab not found")
            return

        firebase_id = old_expense_data.get("firebase_id")

        updated_record = {
            "date": updated_expense_data.get("date", ""),
            "name": updated_expense_data.get("expense_name", "") or updated_expense_data.get("Category", "Expense"),
            "description": updated_expense_data.get("description", "") or updated_expense_data.get("expense_type", ""),
            "amount": f"{self._safe_amount(updated_expense_data.get('amount', 0)):.2f}",
            "year": self._expense_year(updated_expense_data),
            "firebase_id": firebase_id
        }

        found_index = -1

        for i, exp in enumerate(balance_tab.expenses_data):
            if exp.get("firebase_id") == firebase_id:
                found_index = i
                break

        if found_index >= 0:

            old_data = balance_tab.expenses_data[found_index]

            balance_tab.expenses_data[found_index] = updated_record.copy()

            # Update annual data
            annual_index = -1

            for i, exp in enumerate(balance_tab.annual_expenses_data):
                if exp.get("firebase_id") == firebase_id:
                    annual_index = i
                    break

            if annual_index >= 0:
                balance_tab.annual_expenses_data[annual_index] = updated_record.copy()

            # ⭐ Firebase update (same as revenue logic)
            if hasattr(balance_tab, "save_expense_to_firebase"):
                balance_tab.save_expense_to_firebase(updated_record, old_data)

            _log.info("Updated existing expense in balance sheet")

        else:

            balance_tab.expenses_data.append(updated_record.copy())
            balance_tab.annual_expenses_data.append(updated_record.copy())

            if hasattr(balance_tab, "save_expense_to_firebase"):
                balance_tab.save_expense_to_firebase(updated_record)

            _log.info("Added new expense to balance sheet")

        balance_tab.update_annual_summary()
        balance_tab.update_stats_cards()

        if hasattr(balance_tab, "current_category") and balance_tab.current_category == "Expenses":
            balance_tab.filter_finance_entries()
            
    def load_expenses(self):
        """Load expenses from Firebase and include balance-sheet-only expense records."""
        firebase_expenses = ExpensesFirebaseManager.load_expenses()
        balance_sheet_expenses = ExpensesFirebaseManager.load_balance_sheet_expenses()
        self.expenses = self._merge_balance_sheet_expenses(firebase_expenses, balance_sheet_expenses)
        self.cached_expenses = self.expenses.copy()
        
        # Load custom category colors
        self.load_custom_category_colors()
        
        self.update_expenses_table()
        self.update_charts()
        self.update_statistics()  # This will now update the cards properly
        self.update_categories_filter()  # Update the categories filter with actual data
        self.update_filter_menus()
        
        # NEW: Apply existing filters after loading data
        self.filter_expenses()  # Add this line
    
    def _fit_table_height(self):
        """Resize table to show all rows without internal scrolling."""
        row_count = self.expenses_table.rowCount()
        header_h = self.expenses_table.horizontalHeader().height()
        rows_h = sum(self.expenses_table.rowHeight(r) for r in range(row_count))
        total_h = header_h + rows_h + 4
        self.expenses_table.setFixedHeight(max(160, total_h))
        self.expenses_table.updateGeometry()

    def update_expenses_table(self):
        """Update the professional expense ledger from current data."""
        self._exp_all_items = list(reversed(self.expenses))
        self._exp_page = 1
        self._exp_render_page()
            
    def on_pie_slice_clicked(self, slice):
        self.pie_click_table.show()

        # Find slice index in series
        slices = self.pie_series.slices()
        slice_index = -1
        for idx, s in enumerate(slices):
            if s is slice:
                slice_index = idx
                break

        if slice_index < 0 or slice_index >= len(self.pie_slice_categories):
            return

        category_name = self.pie_slice_categories[slice_index]
        amount_value = slice.value()

        # Show click tooltip
        tooltip_text = f"{category_name} - ${amount_value:,.2f}"
        self._show_slice_tooltip(tooltip_text)

        # Filter expenses by CURRENT PIE CHART month/year AND category
        selected_month = self.pie_chart_month  # Use current pie chart month
        selected_year = self.pie_chart_year    # Use current pie chart year
        
        matched = []
        for expense in self.expenses:
            date_str = expense.get('date', '')
            if not date_str:
                continue
            
            try:
                expense_date = datetime.strptime(date_str, '%m-%d-%Y')  # Changed format
                # Check if expense matches both category AND selected month/year
                if (expense.get('type') == category_name and 
                    expense_date.year == selected_year and 
                    expense_date.month == selected_month):
                    matched.append(expense)
            except:
                continue

        if not matched:
            return

        # Always show FIRST expense created
        first_expense = matched[0]
        self.display_expense_in_table(first_expense)

        # Update table header
        self.update_pie_header(len(matched))

    def _show_slice_tooltip(self, text):
        """Show tooltip for clicked pie slice"""
        cursor_pos = QtGui.QCursor.pos()
        QtWidgets.QToolTip.showText(cursor_pos, text, self, self.pie_chart_widget.rect())

    def create_enhanced_action_buttons(self, expense):
        """Create enhanced action buttons — original style, vertically centered in row."""
        outer = QtWidgets.QWidget()
        outer_layout = QtWidgets.QVBoxLayout(outer)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setAlignment(QtCore.Qt.AlignCenter)

        actions_widget = QtWidgets.QWidget()
        actions_widget.setMinimumWidth(166)
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(4, 3, 4, 3)
        actions_layout.setSpacing(4)
        actions_layout.setAlignment(QtCore.Qt.AlignCenter)

        view_btn = QtWidgets.QPushButton("View")
        view_btn.setToolTip("View Details")
        view_btn.setFixedSize(48, 30)
        view_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                font-size: 11px;
                font-weight: bold;
                padding: 2px;
            }
            QPushButton:hover { background-color: #e9ecef; border-color: #3498db; }
            QPushButton:pressed { background-color: #dee2e6; }
        """)
        view_btn.clicked.connect(lambda: self.view_expense_details(expense))

        edit_btn = QtWidgets.QPushButton("Edit")
        edit_btn.setToolTip("Edit Expense")
        edit_btn.setFixedSize(46, 30)
        edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                font-size: 11px;
                font-weight: bold;
                padding: 2px;
            }
            QPushButton:hover { background-color: #e9ecef; border-color: #f39c12; }
            QPushButton:pressed { background-color: #dee2e6; }
        """)
        edit_btn.clicked.connect(lambda: self.edit_expense(expense))

        delete_btn = QtWidgets.QPushButton("Del")
        delete_btn.setToolTip("Delete Expense")
        delete_btn.setFixedSize(42, 30)
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                font-size: 11px;
                font-weight: bold;
                padding: 2px;
            }
            QPushButton:hover { background-color: #e9ecef; border-color: #e74c3c; }
            QPushButton:pressed { background-color: #dee2e6; }
        """)
        delete_btn.clicked.connect(lambda: self.delete_expense(expense))

        actions_layout.addWidget(view_btn)
        actions_layout.addWidget(edit_btn)
        actions_layout.addWidget(delete_btn)

        outer_layout.addWidget(actions_widget)
        return outer
    
    def update_filter_menus(self):
        """Update filter menus with current data"""
        # Populate category filter options
        categories = sorted({expense.get('type', '') for expense in self.expenses if expense.get('type')})
        self.category_filter_menu.clear()

        # Add ALL option
        all_action = self.category_filter_menu.addAction("All Categories")
        all_action.triggered.connect(lambda: self.apply_category_filter("All Categories"))

        # Add category names
        for category in categories:
            act = self.category_filter_menu.addAction(category)
            act.triggered.connect(lambda checked, c=category: self.apply_category_filter(c))
    
    def apply_category_filter(self, value):
        """Apply category filter"""
        self.selected_category_filter = value
        self.filter_expenses()
    
    def filter_expenses(self):
        """Filter expenses based on active search, category, and date filters"""
        category_filter = self.categories_filter_combo.currentText()
        search_text = self.search_edit.text().lower().strip()
        
        # Check if date range filter is active
        date_range_active = "to" in self.date_range_button.text()
        from_date = None
        to_date = None
        
        if date_range_active:
            # Use stored QDate objects if available
            if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
                from_date = self.current_from_date
                to_date = self.current_to_date
            else:
                try:
                    date_text = self.date_range_button.text().replace("📅 ", "")
                    from_str, to_str = date_text.split(" to ")
                    # Parse from US format
                    from_date = QtCore.QDate.fromString(from_str, "MMMM d, yyyy")
                    to_date = QtCore.QDate.fromString(to_str, "MMMM d, yyyy")
                except:
                    date_range_active = False
        
        # STEP 1: First get ALL categories from expenses that pass date filter only
        # This will be used to populate the categories dropdown
        categories_for_dropdown = set()
        
        def _parse_expense_qdate(exp):
            """Parse expense date field trying multiple formats; returns QDate."""
            raw = exp.get('date', '') or ''
            for _fmt in ("MM-dd-yyyy", "yyyy-MM-dd", "MM/dd/yyyy", "M/d/yyyy"):
                qd = QtCore.QDate.fromString(raw, _fmt)
                if qd.isValid():
                    return qd
            return QtCore.QDate()

        for expense in self.expenses:
            # Check date filter — always uses expense date field, not created_at
            matches_date = True
            if date_range_active and from_date and to_date:
                expense_date = _parse_expense_qdate(expense)
                matches_date = expense_date.isValid() and (from_date <= expense_date <= to_date)

            # Collect categories from expenses that pass date filter only
            if matches_date:
                cat = expense.get('Category', expense.get('type', ''))
                if cat:
                    categories_for_dropdown.add(cat)

        # STEP 2: Update categories dropdown with ALL categories from date-filtered data
        # (not search-filtered data)
        self.update_categories_filter_based_on_active_filters(categories_for_dropdown)

        # STEP 3: Now filter expenses based on ALL active filters
        filtered_expenses = []

        for expense in self.expenses:
            # Check date filter — always uses expense date field, not created_at
            matches_date = True
            if date_range_active and from_date and to_date:
                expense_date = _parse_expense_qdate(expense)
                matches_date = expense_date.isValid() and (from_date <= expense_date <= to_date)
            
            # Check search filter (if active)
            matches_search = True
            if search_text:
                # Try to parse and format date for searching
                date_str = expense.get('date', '')
                display_date = ""
                try:
                    if date_str:
                        qd = QtCore.QDate.fromString(date_str, "MM-dd-yyyy")
                        if qd.isValid():
                            # Multiple date formats for flexible searching
                            display_date = qd.toString("MM-dd-yyyy").lower()
                            alt_date1 = qd.toString("MM/dd/yyyy").lower()
                            alt_date2 = qd.toString("MMMM d, yyyy").lower()
                            alt_date3 = qd.toString("MMM d, yyyy").lower()
                            alt_date4 = str(qd.year())
                            alt_date5 = qd.toString("MMMM").lower()
                            alt_date6 = qd.toString("MMM").lower()
                            alt_date7 = qd.toString("dd/MM/yyyy").lower()
                            alt_date8 = qd.toString("dd-MM-yyyy").lower()
                except:
                    pass
                
                # Get all searchable fields
                expense_type = expense.get('expense_type', '').lower()
                category = expense.get('Category', expense.get('type', '')).lower()
                vendor = expense.get('vendor', '').lower()
                description = expense.get('description', '').lower()
                project = expense.get('project', '').lower()
                payment_method = expense.get('payment_method', '').lower()
                expense_name = expense.get('expense_name', '').lower()
                reference = expense.get('reference', '').lower()
                notes = expense.get('notes', '').lower()
                
                # Try to parse amount for searching
                amount = expense.get('amount', 0)
                amount_str = str(amount).lower()
                
                # Check if search is for amount (dollar amounts, numbers)
                search_amount_match = False
                try:
                    # Remove currency symbols and commas for amount comparison
                    clean_search = search_text.replace('$', '').replace(',', '').strip()
                    if clean_search:
                        # Try to match exact amount
                        if float(clean_search) == float(amount):
                            search_amount_match = True
                        # Try to match partial amount (e.g., "249" matches "2499.99")
                        elif str(int(float(amount))).startswith(clean_search.replace('.', '')):
                            search_amount_match = True
                except:
                    pass
                
                # Check all searchable fields (including amount)
                matches_search = (
                    search_text in expense_type or
                    search_text in category or
                    search_text in vendor or
                    search_text in description or
                    search_text in project or
                    search_text in payment_method or
                    search_text in expense_name or
                    search_text in reference or
                    search_text in notes or
                    search_amount_match or
                    (date_str and (
                        search_text in display_date or
                        search_text in alt_date1 or
                        search_text in alt_date2 or
                        search_text in alt_date3 or
                        search_text in alt_date4 or
                        search_text in alt_date5 or
                        search_text in alt_date6 or
                        search_text in alt_date7 or
                        search_text in alt_date8 or
                        search_text in date_str.lower() or
                        search_text in str(expense.get('date', '')).lower()
                    ))
                )
                
                # Additional fuzzy search for partial matches in vendor/description
                if not matches_search and len(search_text) > 2:
                    # Check if any word in the search text matches
                    search_words = search_text.split()
                    for word in search_words:
                        if (word in vendor or 
                            word in description or 
                            word in expense_name or
                            word in project):
                            matches_search = True
                            break
            
            # Check category filter (if active)
            matches_category = True
            if category_filter != "All Categories":
                matches_category = (
                    expense.get('Category', expense.get('type', '')) == category_filter
                )
            
            # Check if expense passes ALL active filters
            if matches_date and matches_search and matches_category:
                filtered_expenses.append(expense)
        
        # STEP 4: Sort by date ascending only when date filter is active, else newest first
        if date_range_active and from_date and to_date:
            def _date_key(exp):
                try:
                    from datetime import datetime as _dt
                    return _dt.strptime(exp.get('date', ''), "%m-%d-%Y")
                except Exception:
                    from datetime import datetime as _dt
                    return _dt.min
            filtered_expenses.sort(key=_date_key)
            self.display_filtered_expenses(filtered_expenses, date_asc=True)
        else:
            self.display_filtered_expenses(filtered_expenses, date_asc=False)
        self.update_statistics(filtered_expenses)
    
    def update_categories_filter_based_on_active_filters(self, categories_set):
        """Update categories filter combo box based on currently active filters"""
        current_selection = self.categories_filter_combo.currentText()
        # A pinned category (set during delete) must stay selected regardless of
        # whether it still has expenses — it may have become empty after deletion.
        pinned = getattr(self, '_pinned_filter_category', None)
        target = pinned if pinned else current_selection

        self.categories_filter_combo.blockSignals(True)

        self.categories_filter_combo.clear()
        self.categories_filter_combo.addItem("All Categories")

        for category in sorted(categories_set):
            if category:
                self.categories_filter_combo.addItem(category)

        # If the target category is no longer in the list (e.g. all its expenses were
        # deleted), add it back so the user stays on their chosen filter.
        existing = [self.categories_filter_combo.itemText(i)
                    for i in range(self.categories_filter_combo.count())]
        if target and target != "All Categories" and target not in existing:
            self.categories_filter_combo.addItem(target)

        if target and target in [self.categories_filter_combo.itemText(i)
                                  for i in range(self.categories_filter_combo.count())]:
            self.categories_filter_combo.setCurrentText(target)
        else:
            self.categories_filter_combo.setCurrentText("All Categories")

        self.categories_filter_combo.blockSignals(False)


    
    def open_month_year_picker(self):
        """Month/Year picker for PIE CHART only"""
        dlg = MonthYearPickerDialog(
            self,
            selected_year=self.pie_chart_year,  # Use pie_chart_year
            selected_month=self.pie_chart_month  # Use pie_chart_month
        )

        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.pie_chart_year = dlg.selected_year  # Update only pie chart year
            self.pie_chart_month = dlg.selected_month  # Update only pie chart month
            self.update_pie_chart()  # Only update pie chart

    
    def display_filtered_expenses(self, expenses, date_asc=False):
        """Display filtered expenses in the professional ledger."""
        # date_asc=True → oldest first (date filter active)
        # date_asc=False → newest first (default)
        self._exp_all_items = list(expenses) if date_asc else list(reversed(expenses))
        self._exp_page = 1
        self._exp_render_page()

    def _exp_render_page(self):
        import math
        total = len(self._exp_all_items)
        per_page = self._exp_per_page
        max_page = max(1, math.ceil(total / per_page))
        self._exp_page = max(1, min(self._exp_page, max_page))
        start_i = (self._exp_page - 1) * per_page
        end_i = min(start_i + per_page, total)
        page_items = self._exp_all_items[start_i:end_i]
        self.expenses_table.setSortingEnabled(False)
        self.expenses_table.setRowCount(len(page_items))
        for row, expense in enumerate(page_items):
            self._populate_ledger_row(row, expense, sno=start_i + row + 1)
        # Let each row grow to fit its content, then enforce a minimum so
        # action buttons always have enough height regardless of text length
        self.expenses_table.resizeRowsToContents()
        for r in range(self.expenses_table.rowCount()):
            if self.expenses_table.rowHeight(r) < 50:
                self.expenses_table.setRowHeight(r, 50)
        self.expenses_table.setSortingEnabled(True)
        self._fit_table_height()
        self._exp_rebuild_pagination(total, max_page)

    def _exp_rebuild_pagination(self, total, max_page):
        if not hasattr(self, '_exp_page_btns_layout'):
            return
        page_num = self._exp_page
        per_page = self._exp_per_page
        start = (page_num - 1) * per_page + 1 if total else 0
        end = min(page_num * per_page, total)
        if hasattr(self, '_exp_info_lbl'):
            self._exp_info_lbl.setText(f"Showing {start}–{end} of {total} expenses")
        while self._exp_page_btns_layout.count():
            item = self._exp_page_btns_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        _s = getattr(self, '_exp_pg_style', '')
        _win_start = max(1, min(page_num, max_page - 2))
        for p in range(_win_start, min(_win_start + 3, max_page + 1)):
            btn = QtWidgets.QPushButton(str(p))
            btn.setFixedSize(32, 28)
            btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            if p == page_num:
                btn.setStyleSheet("""QPushButton {
                    background-color: #00756f; color: #ffffff;
                    border: 1px solid #00756f; border-radius: 6px;
                    font-size: 12px; font-weight: 700;
                    min-width: 32px; min-height: 28px; padding: 0 8px;
                }
                QPushButton:hover { background-color: #005f5a; color: #ffffff; }""")
            else:
                btn.setStyleSheet(_s)
                btn.clicked.connect(lambda _, pg=p: self._exp_go_to(pg))
            self._exp_page_btns_layout.addWidget(btn)
        if hasattr(self, '_exp_prev_btn'):
            self._exp_prev_btn.setEnabled(page_num > 1)
        if hasattr(self, '_exp_next_btn'):
            self._exp_next_btn.setEnabled(page_num < max_page)

    def _exp_go_prev(self):
        if self._exp_page > 1:
            self._exp_page -= 1
            self._exp_render_page()

    def _exp_go_next(self):
        self._exp_page += 1
        self._exp_render_page()

    def _exp_go_to(self, page):
        self._exp_page = page
        self._exp_render_page()

    def _refresh_table_keep_page(self):
        """Re-apply active filters and re-render without resetting the current page."""
        saved_page = getattr(self, '_exp_page', 1)
        self.filter_expenses()
        # filter_expenses resets _exp_page → restore it (clamping handled by _exp_render_page)
        self._exp_page = saved_page
        self._exp_render_page()

    def update_charts(self):
        """Update both bar chart and pie chart with current data"""
        self.update_bar_chart()
        self.update_pie_chart()
    
    def show_no_data_in_bar_chart(self):
        """Show "No Data" message in bar chart when there's no data"""
        chart = self.bar_chart_widget.chart()
        
        # Clear everything
        chart.removeAllSeries()
        
        # Add a text item or empty series
        # Create a simple empty bar series with "No Data" label
        new_bars = QBarSet("No Data")
        new_bars.append(1)  # Minimal value
        
        # Make it invisible or very light
        new_bars.setColor(QtGui.QColor("#f0f0f0"))  # Very light gray
        new_bars.setBorderColor(QtGui.QColor("#f0f0f0"))
        
        new_series = QBarSeries()
        new_series.append(new_bars)
        new_series.setLabelsVisible(True)
        new_series.setLabelsFormat("No Expenses")
        new_series.setLabelsPosition(QBarSeries.LabelsCenter)
        
        chart.addSeries(new_series)
        
        # Create minimal axes
        axis_x = QBarCategoryAxis()
        axis_x.append(["No Data"])
        axis_x.setLabelsFont(QtGui.QFont("Segoe UI", 9))
        axis_x.setLabelsColor(QtGui.QColor("#999999"))
        
        axis_y = QValueAxis()
        axis_y.setRange(0, 2)
        axis_y.setTickCount(3)
        axis_y.setLabelsVisible(False)
        axis_y.setGridLineVisible(False)
        
        chart.addAxis(axis_x, QtCore.Qt.AlignBottom)
        chart.addAxis(axis_y, QtCore.Qt.AlignLeft)
        
        new_series.attachAxis(axis_x)
        new_series.attachAxis(axis_y)
        
        # Hide legend
        chart.legend().hide()
        
        # Update the bar title
        selected_year = self.bar_chart_year
        self.bar_title.setText(f"{selected_year} - No expenses")
        
        chart.update()
        self.bar_chart_widget.repaint()
    
    def update_bar_chart(self):
        chart = self.bar_chart_widget.chart()
        scene = self.bar_chart_widget.scene()
        """Update BAR CHART with BAR CHART year only"""
        try:
            selected_year = self.bar_chart_year
            self.bar_title.setText(f"{selected_year} Yearly")

            # Clear the chart completely first
            chart = self.bar_chart_widget.chart()
            
            # Remove all existing series
            chart = self.bar_chart_widget.chart()

            # 🔥 FULL HARD RESET
            chart.removeAllSeries()

            for ax in chart.axes():
                chart.removeAxis(ax)

            # Calculate monthly expenses for the selected year
            monthly_expenses = {month: 0 for month in range(1, 13)}
            for expense in self.expenses:
                try:
                    d = expense.get("date", "")
                    if d:
                        dt = datetime.strptime(d, "%m-%d-%Y")
                        if dt.year == selected_year:
                            monthly_expenses[dt.month] += self._safe_amount(expense.get("amount", 0))
                except:
                    pass

            categories = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            amounts = [monthly_expenses[month] for month in range(1, 13)]
            
            # FIX: Check if we have real data (not just empty categories)
            has_real_data = any(amount > 0 for amount in amounts) and len(categories) > 0
            
            if not has_real_data:
                # Show "No Data" message instead of showing $1
                self.show_no_data_in_bar_chart()
                return
                
            # Create new bar set
            new_bars = QBarSet("")
            new_bars.setColor(QtGui.QColor("#0f766e"))
            new_bars.setBorderColor(QtGui.QColor("#115e59"))
            new_bars.setLabelColor(QtGui.QColor("#334155"))

            for amt in amounts:
                new_bars.append(amt)

            # Create new series
            new_series = QBarSeries()
            new_series.append(new_bars)
            new_series.setLabelsVisible(True)
            new_series.setLabelsFormat("$@value")
            new_series.setLabelsPosition(QBarSeries.LabelsOutsideEnd)
            
            # Bar width control
            count = len(categories)
            if count <= 1:
                new_series.setBarWidth(0.15)
            elif count == 2:
                new_series.setBarWidth(0.10)
            elif count <= 6:
                new_series.setBarWidth(0.20)
            elif count <= 14:
                    new_series.setBarWidth(0.30)
            else:
                new_series.setBarWidth(0.70)

            # Add series to chart
            chart.addSeries(new_series)

            # Create X-axis
            self.axis_x = QBarCategoryAxis()
            
            self.axis_x.setLabelsAngle(0)
            
            font_x = QtGui.QFont("Inter", 8)
            self.axis_x.setLabelsFont(font_x)
            self.axis_x.setLabelsColor(QtGui.QColor("#475569"))
            self.axis_x.setGridLineVisible(False)
            self.axis_x.clear()

            if categories:
                self.axis_x.append(categories)
            else:
                self.axis_x.append([""])
                
            chart.addAxis(self.axis_x, QtCore.Qt.AlignBottom)
            new_series.attachAxis(self.axis_x)
            
            # Create Y-axis
            self.axis_y = QValueAxis()
            self.axis_y.setTickCount(6)
            self.axis_y.setMinorTickCount(0)

            max_val = max(amounts) if amounts else 100
            self.axis_y.setRange(0, max_val * 1.2)
            self.axis_y.setLabelFormat("$%d")  # Changed to remove commas
            font_y = QtGui.QFont("Inter", 8)
            self.axis_y.setLabelsFont(font_y)
            self.axis_y.setLabelsColor(QtGui.QColor("#475569"))
            self.axis_y.setGridLineVisible(True)
            self.axis_y.setGridLineColor(QtGui.QColor("#e5edf5"))
            self.axis_y.setMinorGridLineVisible(False)

            chart.addAxis(self.axis_y, QtCore.Qt.AlignLeft)
            new_series.attachAxis(self.axis_y)

            # Hide legend
            chart.legend().hide()

            # Adjust margins based on category count
            if count > 6:
                # More space needed for angled labels
                chart.setMargins(QtCore.QMargins(16, 12, 10, 8))
            else:
                chart.setMargins(QtCore.QMargins(16, 12, 10, 48))

            chart.setBackgroundBrush(QtGui.QBrush(QtGui.QColor("#ffffff")))
            chart.setPlotAreaBackgroundVisible(True)
            chart.setPlotAreaBackgroundBrush(QtGui.QBrush(QtGui.QColor("#f8fafc")))

            # Refresh chart
            chart.update()
            self.bar_chart_widget.repaint()
                        
        except Exception as e:
            _log.warning("Error updating bar chart: %s", e)
            import traceback
            traceback.print_exc()
            
    def center_pie_table(self):
        rows = self.pie_click_table.rowCount()
        cols = self.pie_click_table.columnCount()

        for r in range(rows):
            for c in range(cols):
                item = self.pie_click_table.item(r, c)
                if item:
                    item.setTextAlignment(QtCore.Qt.AlignCenter)

    def update_pie_header(self, freq):
        """Correct the header after clicking a pie slice"""
        header = self.pie_click_table.horizontalHeader()

        # Column 0 must always be Category
        header_item = self.pie_click_table.horizontalHeaderItem(0)
        if header_item:
            header_item.setText("Category")

        # Column 1 must always be Expenses (number dropdown)
        header_item = self.pie_click_table.horizontalHeaderItem(1)
        if header_item:
            header_item.setText("Expenses")
    
    def save_custom_category_colors(self):
        """Save custom category colors to Firebase"""
        if not FIREBASE_AVAILABLE:
            return False
            
        try:
            # Save with timestamp and metadata
            color_data = {
                'colors': self.category_colors,
                'assigned_colors': list(self.assigned_colors),
                'user_color_index': self.user_color_index,
                'last_updated': datetime.now(timezone.utc).isoformat(),
                'total_categories': len(self.category_colors)
            }
            
            ref = db.reference('/custom_category_colors')
            ref.set(color_data)
            _log.info("Custom category colors saved: %s colors", len(self.category_colors))
            return True
        except Exception as e:
            _log.warning("Error saving custom category colors: %s", e)
            return False

    def load_custom_category_colors(self):
        """Load custom category colors from Firebase"""
        if not FIREBASE_AVAILABLE:
            return False
            
        try:
            ref = db.reference('/custom_category_colors')
            saved_data = ref.get() or {}
            
            if 'colors' in saved_data:
                # Load colors
                loaded_colors = saved_data['colors']
                
                # Merge with current assignments (prioritize loaded colors)
                for category, color in loaded_colors.items():
                    if color not in self.assigned_colors:
                        self.category_colors[category] = color
                        self.assigned_colors.add(color)
                
                # Load other metadata
                if 'user_color_index' in saved_data:
                    self.user_color_index = saved_data['user_color_index']
                
                _log.info("Loaded %s custom category colors", len(loaded_colors))
                return True
            
            return False
        except Exception as e:
            _log.warning("Error loading custom category colors: %s", e)
            return False

    def update_pie_chart(self):
        """Update PIE CHART with PIE CHART month/year only"""
        self.pie_click_table.hide()
        
        # Reset color assignments to ensure consistency
        self.reset_color_assignments()
        
        # Use PIE CHART specific attributes
        selected_month = self.pie_chart_month
        selected_year = self.pie_chart_year
        
        # --- AUTO-HIDE TITLE WHEN NOT CURRENT MONTH ---
        current_month = datetime.now().month
        current_year = datetime.now().year

        # ===== Update Pie Chart Title =====
        if selected_month == current_month and selected_year == current_year:
            # Current running month
            self.pie_title_label.setText("Current Month")
        else:
            # Show "March 2025 Expenses"
            month_name = datetime.strptime(str(selected_month), "%m").strftime("%B")
            self.pie_title_label.setText(f"{month_name} {selected_year}")

        self.pie_title_label.show()

        category_expenses = {}
        total_amount = 0

        
        # Collect category totals using PIE CHART filters
        for expense in self.expenses:
            date_str = expense.get('date', '')
            if not date_str:
                continue
            
            try:
                expense_date = datetime.strptime(date_str, '%m-%d-%Y')
                # Filter by pie_chart_year and pie_chart_month
                if expense_date.year == selected_year and expense_date.month == selected_month:
                    category = expense.get("Category", expense.get("type", "Unknown"))
                    amount = expense.get('amount', 0)
                    category_expenses[category] = category_expenses.get(category, 0) + amount
                    total_amount += amount
            except:
                continue

        # Clear old slices
        self.pie_series.clear()
        self.pie_slice_categories = []  # Clear category list

        # Add slices
        if total_amount == 0:
            slice = self.pie_series.append("No Data", 1)
            slice.setColor(QtGui.QColor("#bdc3c7"))
            slice.setLabelVisible(False)
        else:
            # Sort categories by amount for consistent ordering
            sorted_categories = sorted(
                category_expenses.items(),
                key=lambda x: x[1],
                reverse=True
            )
            
            for category, amount in sorted_categories:
                percent = (amount / total_amount) * 100

                label_text = f"{percent:.1f}%"
                s = self.pie_series.append(label_text, amount)

                # Store full category name for tooltips
                self.pie_slice_categories.append(category)

                # Get unique color for this category
                color = self.get_color_for_category(category)
                s.setColor(QtGui.QColor(color))
                # Large slices: follow the arc; small slices: stay horizontal so text fits
                if percent >= 12:
                    s.setLabelPosition(QPieSlice.LabelInsideNormal)
                else:
                    s.setLabelPosition(QPieSlice.LabelInsideHorizontal)
                s.setLabelFont(QtGui.QFont("Inter", 7, QtGui.QFont.Bold))
                s.setLabelColor(QtGui.QColor("#FFFFFF"))
                s.setLabelVisible(True)
            # Update legend markers — use series-scoped call for correct order
            chart = self.pie_chart_widget.chart()
            legend = chart.legend()

            try:
                all_markers = legend.markers(self.pie_series)
            except TypeError:
                all_markers = [m for m in legend.markers()
                               if m.series() == self.pie_series]

            # Hide all first
            for m in all_markers:
                m.setVisible(False)

            # Re-label and explicitly re-colour each visible marker
            for marker, (category, amount) in zip(all_markers, sorted_categories[:12]):
                amount_display = format_amount_no_commas(amount)
                short_cat = (category[:25] + "…") if len(category) > 25 else category
                slice_color = self.get_color_for_category(category)
                # setLabel() can break Qt's automatic colour sync — restore it manually
                marker.setBrush(QtGui.QBrush(QtGui.QColor(slice_color)))
                marker.setPen(QtGui.QPen(QtGui.QColor(slice_color)))
                marker.setLabel(f"{short_cat} - {amount_display}")
                marker.setFont(QtGui.QFont("Inter", 8))
                marker.setLabelBrush(QtGui.QBrush(QtGui.QColor("#1E293B")))
                marker.setVisible(True)

        # Update summary text using PIE CHART attributes
        month_name = datetime(selected_year, selected_month, 1).strftime("%B")

        # Format without commas
        total_display = format_amount_no_commas(total_amount)

        self.pie_summary_label.setText(
            f"📅 {month_name} {selected_year} | "
            f"💰 Monthly Expenses: {total_display} | "
            f"📊 Categories: {len(category_expenses)}"
        )
    
    def update_statistics_cards(self, filtered_expenses=None):
        """Update the statistics cards with current expense data"""
        # Use filtered expenses if provided, otherwise all expenses
        expenses_to_analyze = filtered_expenses if filtered_expenses is not None else self.expenses
        
        def safe_float_convert(value):
            """Safely convert amount to float, handling various formats"""
            if isinstance(value, (int, float)):
                return float(value)
            try:
                # Remove $ sign, commas, and any whitespace
                cleaned = str(value).replace('$', '').replace(',', '').strip()
                return float(cleaned) if cleaned else 0.0
            except (ValueError, TypeError):
                return 0.0
        
        # Calculate current stats from actual expenses data
        total_categories = len(set(
            expense.get('type', '') 
            for expense in expenses_to_analyze 
            if expense.get('type')
        ))
        total_expenses = len(expenses_to_analyze)
        total_amount = sum(safe_float_convert(expense.get('amount', 0)) for expense in expenses_to_analyze)

        # Format total amount without commas
        total_display = Currency.format_whole(total_amount) if hasattr(Currency, 'format_whole') else f"${total_amount:,.2f}"

        stats_cards = [
            ("Categories", f"{total_categories} types", "background: #DFF0FA;"),
            ("Expense Records", f"{total_expenses} entries", "background: #F7DDE2;"),
            ("Total Spent", total_display, "background: #EBDDFA;")
        ]
        
        for i, (title, value, color) in enumerate(stats_cards):
            if i < self.cards_layout.count():
                card = self.cards_layout.itemAt(i).widget()
                if card:
                    # Find value label by objectName — avoids text-matching issues
                    for child in card.findChildren(QtWidgets.QLabel):
                        if child.objectName() == "stat_value":
                            child.setText(value)
                            break
            else:
                card = self.create_stat_card(title, value, color, None)
                self.cards_layout.addWidget(card)
                
    def update_statistics(self, filtered_expenses=None):
        """Update statistics cards with current data"""
        # This method now properly updates the cards
        self.update_statistics_cards(filtered_expenses)
        
    def view_expense_details(self, expense):
        """View expense details - styled like invoice history view."""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Expense Details")
        dialog.setMinimumSize(640, 420)
        dialog.resize(700, 480)
        dialog.setWindowFlags(
            dialog.windowFlags()
            | QtCore.Qt.WindowMaximizeButtonHint
            | QtCore.Qt.WindowMinimizeButtonHint
        )
        layout = QtWidgets.QVBoxLayout(dialog)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        formatted_date = QtCore.QDate.fromString(
            expense.get('date', ''), "MM-dd-yyyy"
        ).toString("MMMM d, yyyy") or expense.get('date', 'N/A')
        try:
            amt_display = f"${float(expense.get('amount', 0)):,.2f}"
        except (ValueError, TypeError):
            amt_display = str(expense.get('amount', 'N/A'))

        exp_type = expense.get('expense_type', expense.get('type', '')) or 'N/A'
        category = expense.get('Category', expense.get('type', '')) or 'N/A'
        exp_name = expense.get('expense_name', '') or 'N/A'
        vendor   = expense.get('vendor', '') or 'N/A'
        project  = expense.get('project', '') or 'N/A'
        method   = expense.get('payment_method', '') or 'N/A'
        descr    = expense.get('description', '') or 'N/A'

        html = f"""
        <h2 style="margin:0 0 10px 0; color:#0f172a;">Expense Details</h2>
        <table border="0" cellspacing="4" cellpadding="5" style="font-size:13px;">
            <tr><td><b>Date:</b></td><td>{formatted_date}</td></tr>
            <tr><td><b>Expense Type:</b></td><td>{exp_type}</td></tr>
            <tr><td><b>Category:</b></td><td>{category}</td></tr>
            <tr><td><b>Expense Name:</b></td><td>{exp_name}</td></tr>
            <tr><td><b>Vendor:</b></td><td>{vendor}</td></tr>
            <tr><td><b>Amount:</b></td><td><b style="color:#0f766e;">{amt_display}</b></td></tr>
            <tr><td><b>Project:</b></td><td>{project}</td></tr>
            <tr><td><b>Payment Method:</b></td><td>{method}</td></tr>
            <tr><td><b>Description:</b></td><td>{descr}</td></tr>
        </table>
        """

        text_edit = QtWidgets.QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(html)
        layout.addWidget(text_edit)

        close_btn = QtWidgets.QPushButton("Close")
        close_btn.setFixedSize(100, 36)
        close_btn.setStyleSheet(
            "QPushButton { background: #0f766e; color: white; border: none; "
            "border-radius: 6px; font-size: 12px; font-weight: bold; }"
            "QPushButton:hover { background: #115e59; }"
        )
        close_btn.clicked.connect(dialog.accept)
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        dialog.exec_()
    
    def edit_expense(self, expense):
        """Edit expense - opens the form with existing data"""
        self.show_edit_expense_dialog(expense)
    
    def delete_expense(self, expense):
        reply = QtWidgets.QMessageBox.question(
            self, "Delete Expense",
            "Are you sure you want to delete this expense?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )

        if reply == QtWidgets.QMessageBox.Yes:
            current_search_text = self.search_edit.text()
            current_category = self.categories_filter_combo.currentText()

            if expense in self.expenses:
                self.expenses.remove(expense)

            self.expenses_table.setUpdatesEnabled(False)

            # Pin the category so filter_expenses/update_categories_filter_based_on_active_filters
            # won't reset the combo even when the category becomes empty after deletion.
            self._pinned_filter_category = current_category
            self._refresh_table_keep_page()
            self._pinned_filter_category = None

            QtCore.QTimer.singleShot(0, self.update_charts)
            self.update_statistics_cards()

            # Restore search text if it was set
            if current_search_text:
                self.search_edit.setText(current_search_text)

            self.expenses_table.setUpdatesEnabled(True)

            if 'firebase_id' in expense:
                _fid = expense['firebase_id']
                _is_bs = expense.get('finance_source') == 'balance_sheet'
                if _is_bs:
                    QtCore.QTimer.singleShot(
                        100, lambda fid=_fid: ExpensesFirebaseManager.delete_balance_sheet_expense(fid)
                    )
                else:
                    QtCore.QTimer.singleShot(
                        100, lambda fid=_fid: ExpensesFirebaseManager.delete_expense(fid)
                    )

            self.show_success_message("Expense deleted successfully!")
            
    def show_category_expense_dialog(self):
        """Show dialog to select category for detailed view"""
        if not self.expenses:
            QtWidgets.QMessageBox.information(self, "No Data", "No expenses available to view.")
            return
        
        # Get unique categories
        categories = sorted(set(expense.get('type', 'Unknown') for expense in self.expenses))
        
        if not categories:
            QtWidgets.QMessageBox.information(self, "No Categories", "No expense categories found.")
            return
        
        # Create category selection dialog
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("🔍 Select Category")
        dialog.setModal(True)
        dialog.resize(400, 300)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        # Title
        title = QtWidgets.QLabel("Select Expense Type")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(title)
        
        # Category list
        category_list = QtWidgets.QListWidget()
        category_list.addItems(categories)
        category_list.setStyleSheet("""
            QListWidget {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                padding: 5px;
                background: white;
            }
            QListWidget::item {
                padding: 10px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background: #3498db;
                color: white;
                border-radius: 4px;
            }
            QListWidget::item:hover {
                background: #f8f9fa;
            }
        """)
        layout.addWidget(category_list)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        view_btn = QtWidgets.QPushButton("View Expenses")
        view_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 6px;
                border: none;
            }
            QPushButton:hover {
                background: #2ecc71;
            }
        """)
        
        cancel_btn = QtWidgets.QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 6px;
                border: none;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        
        button_layout.addWidget(view_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        # Connect signals
        def view_selected():
            selected_items = category_list.selectedItems()
            if selected_items:
                category = selected_items[0].text()
                dialog.accept()
                self.show_category_expenses(category)
        
        view_btn.clicked.connect(view_selected)
        cancel_btn.clicked.connect(dialog.reject)
        category_list.itemDoubleClicked.connect(lambda item: view_selected())
        
        dialog.exec_()
    
    def show_category_expenses(self, category_name):
        """Show expenses for a specific category"""
        category_expenses = [exp for exp in self.expenses if exp.get('type') == category_name]
        
        if not category_expenses:
            QtWidgets.QMessageBox.information(self, "No Expenses", f"No expenses found for category: {category_name}")
            return
        
        dialog = CategoryExpenseDialog(category_name, category_expenses, self)
        dialog.exec_()
    
    # Message display methods
    def show_success_message(self, message):
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Success")
        msg.setText(message)
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.exec_()
    
    def show_error_message(self, message):
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Error")
        msg.setText(message)
        msg.setIcon(QtWidgets.QMessageBox.Critical)
        msg.exec_()
    
    def show_warning_message(self, message):
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Warning")
        msg.setText(message)
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        msg.exec_()
    
    def show_info_message(self, message):
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Information")
        msg.setText(message)
        msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.exec_()
