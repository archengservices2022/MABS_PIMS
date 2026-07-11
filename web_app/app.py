"""MABS PIMS - Flask Web Application"""
import os
import re
import json
import base64
import tempfile
import logging
import secrets
import hashlib
from datetime import datetime, timezone, timedelta
from pathlib import Path
from functools import wraps
from typing import Dict, List, Optional

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file, abort
)

try:
    import anthropic as _anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

try:
    from pypdf import PdfReader as _PdfReader
    PYPDF_AVAILABLE = True
except ImportError:
    PYPDF_AVAILABLE = False

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("pims.web")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent   # project root
DATA_DIR = BASE_DIR / "data"
ASSETS_DIR = BASE_DIR / "assets"

# ── Firebase config ───────────────────────────────────────────────────────────
FIREBASE_API_KEY = "AIzaSyD6F6T_KIZ90TkCOL03-jSXTeuPM5WVwJY"
FIREBASE_DB_URL  = "https://invoice-7fe93-default-rtdb.firebaseio.com"

FIREBASE_AVAILABLE = False
db = None

try:
    import firebase_admin
    from firebase_admin import credentials, db as firebase_db
    from firebase_admin.exceptions import FirebaseError

    # Cloud deployment: FIREBASE_SERVICE_ACCOUNT_JSON env var (full JSON string)
    _sa_json_str = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
    if _sa_json_str:
        if not firebase_admin._apps:
            _sa_dict = json.loads(_sa_json_str)
            cred = credentials.Certificate(_sa_dict)
            firebase_admin.initialize_app(cred, {"databaseURL": FIREBASE_DB_URL})
        db = firebase_db
        FIREBASE_AVAILABLE = True
        log.info("Firebase initialised from environment variable")
    else:
        _service_key_candidates = [
            Path.home() / ".mabs" / "servicekey.json",
            DATA_DIR / "servicekey.json",
            BASE_DIR / "servicekey.json",
        ]
        _key_path = next((p for p in _service_key_candidates if p.exists()), None)

        if _key_path:
            if not firebase_admin._apps:
                cred = credentials.Certificate(str(_key_path))
                firebase_admin.initialize_app(cred, {"databaseURL": FIREBASE_DB_URL})
            db = firebase_db
            FIREBASE_AVAILABLE = True
            log.info("Firebase initialised from %s", _key_path)
        else:
            log.warning("No Firebase service key found — Firebase disabled")
except ImportError:
    log.warning("firebase-admin not installed — Firebase disabled")
except Exception as exc:
    log.error("Firebase init error: %s", exc)

# ── Role helpers ──────────────────────────────────────────────────────────────
ROLE_PAGES = {
    "admin":    ["dashboard", "quotes", "projects", "invoicing", "payroll", "financial", "settings", "employees", "sales_dashboard", "timesheets"],
    "sales":    ["sales_dashboard", "quotes", "employees", "timesheets"],
    "projects": ["projects", "invoicing", "employees", "timesheets"],
    "finance":  ["financial", "employees", "timesheets"],
    "engineer": ["employees", "timesheets"],
}

ALL_PAGES = ["dashboard", "sales_dashboard", "quotes", "projects", "invoicing", "payroll",
             "financial", "settings", "employees", "timesheets"]

PAGE_LABELS = {
    "dashboard":      "Dashboard",
    "quotes":         "Quotes",
    "projects":       "Projects",
    "invoicing":      "Invoicing",
    "payroll":        "Payroll",
    "financial":      "Financial",
    "settings":       "Settings",
    "employees":      "Employees",
    "sales_dashboard":"Sales Dashboard",
    "timesheets":     "Timesheets",
}

def normalize_role(role: str) -> str:
    r = str(role or "sales").strip().lower()
    return r if r in ROLE_PAGES else "sales"

def get_allowed_pages(role: str, custom_pages=None) -> list:
    """Return the effective page list: custom override if set, else role default."""
    if custom_pages and isinstance(custom_pages, list) and len(custom_pages) > 0:
        return custom_pages
    return ROLE_PAGES.get(normalize_role(role), [])

def can_access(role: str, page: str, custom_pages=None) -> bool:
    return page in get_allowed_pages(role, custom_pages)

def first_page(role: str, custom_pages=None) -> str:
    pages = get_allowed_pages(role, custom_pages)
    return pages[0] if pages else "quotes"

# ── Flask app ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "mabs-pims-secret-2025-change-in-prod")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

# Register Jinja2 filters (will be added after _format_date_display is defined)
# We'll add these filters later in the code once the functions are defined

# ── Auth decorators ───────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def role_required(page_key):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user_email" not in session:
                return redirect(url_for("login"))
            role         = session.get("user_role", "")
            custom_pages = session.get("custom_pages") or None
            if not can_access(role, page_key, custom_pages):
                flash("You don't have permission to access this page.", "danger")
                return redirect(url_for(first_page(role, custom_pages)))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ── Firebase helpers ──────────────────────────────────────────────────────────
def fb_ref(path: str):
    if not FIREBASE_AVAILABLE:
        return None
    return db.reference(path)

def fb_get(path: str):
    ref = fb_ref(path)
    return ref.get() if ref else None

def fb_push(path: str, data: dict) -> Optional[str]:
    ref = fb_ref(path)
    if not ref:
        return None
    new_ref = ref.push()
    data["firebase_id"] = new_ref.key
    new_ref.set(data)
    return new_ref.key

def fb_update(path: str, data: dict) -> bool:
    ref = fb_ref(path)
    if not ref:
        return False
    ref.update(data)
    return True

def fb_delete(path: str) -> bool:
    ref = fb_ref(path)
    if not ref:
        return False
    ref.delete()
    return True

def load_settings() -> dict:
    try:
        data = fb_get("/settings") or {}
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    try:
        sf = DATA_DIR / "settings.json"
        if sf.exists():
            with open(sf, encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _get_ai_client():
    """Return an Anthropic client using the API key stored in settings, or None."""
    if not ANTHROPIC_AVAILABLE:
        return None
    key = load_settings().get("ai", {}).get("anthropic_key", "").strip()
    if not key:
        key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not key:
        return None
    return _anthropic.Anthropic(api_key=key)


def _ai_call(prompt: str, system: str = "", max_tokens: int = 1024) -> str:
    """Run a single Claude call; returns the text or raises RuntimeError."""
    client = _get_ai_client()
    if not client:
        raise RuntimeError("No Anthropic API key configured. Add it in Settings → AI.")
    msgs = [{"role": "user", "content": prompt}]
    kwargs = {"model": "claude-haiku-4-5-20251001", "max_tokens": max_tokens, "messages": msgs}
    if system:
        kwargs["system"] = system
    resp = client.messages.create(**kwargs)
    return resp.content[0].text.strip()


def _ai_call_with_image(prompt: str, image_b64: str, media_type: str, max_tokens: int = 1024) -> str:
    """Run a single Claude vision call with an image; returns the text or raises RuntimeError."""
    client = _get_ai_client()
    if not client:
        raise RuntimeError("No Anthropic API key configured. Add it in Settings → AI.")
    msgs = [{
        "role": "user",
        "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
            {"type": "text", "text": prompt},
        ],
    }]
    resp = client.messages.create(model="claude-haiku-4-5-20251001", max_tokens=max_tokens, messages=msgs)
    return resp.content[0].text.strip()


def company_info() -> dict:
    settings = load_settings()
    defaults = {
        "name":    "MABS Engineering LLC",
        "address": "15455 Manchester Rd, PO Box 1144\nManchester, MO 63011",
        "email":   "admin@habbengineering.com",
        "phone":   "314-303-0004",
        "website": "www.mabs-engineeringg.com",
    }
    defaults.update(settings.get("company", {}))
    return defaults

# ── Authentication ────────────────────────────────────────────────────────────
def firebase_sign_in(email: str, password: str):
    """Returns (ok, uid, error_msg)"""
    url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"
    try:
        resp = requests.post(url, json={"email": email, "password": password,
                                        "returnSecureToken": True}, timeout=10)
        if resp.status_code == 200:
            return True, resp.json().get("localId", ""), ""
        err = resp.json().get("error", {}).get("message", "Invalid credentials")
        return False, "", err
    except requests.exceptions.Timeout:
        return False, "", "Connection timed out. Please try again."
    except Exception as exc:
        return False, "", str(exc)

def load_user_profile(uid: str) -> Optional[dict]:
    if not FIREBASE_AVAILABLE:
        return None
    data = fb_get(f"/users/{uid}")
    if data and isinstance(data, dict):
        data["firebase_uid"] = uid
        cp = data.get("custom_pages")
        if isinstance(cp, dict):
            data["custom_pages"] = [cp[k] for k in sorted(cp.keys(), key=lambda x: int(x) if str(x).isdigit() else 0)]
        elif not isinstance(cp, list):
            data["custom_pages"] = None
        return data
    return None

# ── Context processor ─────────────────────────────────────────────────────────
@app.context_processor
def inject_globals():
    role = session.get("user_role", "")
    uid  = session.get("user_uid", "")

    # Global "My Time Clock" widget (topbar) — every role has "employees" access,
    # so this is computed for every logged-in page view, not just /dashboard.
    clock_widget = {}
    if uid:
        my_entries = [e for e in _load_time_entries() if e.get("employee_uid") == uid]
        clock_widget = {
            "my_open_entry":         next((e for e in my_entries if e.get("status") == "open"), None),
            "last_project_number":   my_entries[0].get("project_number", "") if my_entries else "",
            "clock_active_projects": [p for p in _load_projects_list()
                                       if isinstance(p, dict) and p.get("status", "") not in ("Completed", "invoiced_Fully paid", "Cancelled")],
            "today_str":             datetime.now().strftime("%Y-%m-%d"),
        }

    return {
        "user_name":   session.get("user_name", ""),
        "user_email":  session.get("user_email", ""),
        "user_role":   role,
        "user_uid":    uid,
        "allowed_pages": get_allowed_pages(role, session.get("custom_pages") or None),
        "company":     company_info(),
        "now":         datetime.now(),
        "timedelta":   timedelta,
        "format_date": _format_date_display,  # Make date formatter available in templates (MM-DD-YYYY)
        "format_date_invoice": _format_date_invoice,  # Invoice-specific formatter (MM-DD-YY)
        **clock_widget,
    }

# Register Jinja2 filters for date formatting
@app.template_filter('format_date')
def format_date_filter(date_str):
    """Jinja filter to format dates as MM-DD-YYYY."""
    return _format_date_display(date_str)

@app.template_filter('format_date_invoice')
def format_date_invoice_filter(date_str):
    """Jinja filter to format dates as MM-DD-YY (invoice format)."""
    return _format_date_invoice(date_str)

# ── Routes: Auth ──────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_email" in session:
        return redirect(url_for(first_page(session.get("user_role", "sales"))))

    error = None
    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if not email or not password:
            error = "Please enter your email and password."
        elif not FIREBASE_AVAILABLE:
            error = "Authentication service unavailable. Please check server configuration."
        else:
            ok, uid, err_msg = firebase_sign_in(email, password)
            if not ok:
                error = "Invalid email or password. Please try again."
                log.warning("Login failed for %s: %s", email, err_msg)
            else:
                profile = load_user_profile(uid)
                if not profile:
                    error = "No app profile found. Contact your administrator."
                elif not profile.get("active", True):
                    error = "Your account is inactive. Contact your administrator."
                else:
                    role         = normalize_role(profile.get("role", "sales"))
                    custom_pages = profile.get("custom_pages") or None
                    session.permanent        = True
                    session["user_email"]    = email
                    session["user_uid"]      = uid
                    session["user_name"]     = profile.get("username", email.split("@")[0])
                    session["user_role"]     = role
                    session["custom_pages"]  = custom_pages
                    log.info("Login: %s (%s) pages=%s", email, role, custom_pages or "role-default")
                    return redirect(url_for(first_page(role, custom_pages)))

    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

def firebase_send_password_reset(email: str):
    """Send password reset email using Firebase. Returns (success, error_msg)"""
    try:
        url = f"https://identitytoolkit.googleapis.com/v1/accounts:sendOobCode?key={FIREBASE_API_KEY}"
        resp = requests.post(url, json={"email": email, "requestType": "PASSWORD_RESET"}, timeout=10)

        if resp.status_code == 200:
            return True, ""

        err = resp.json().get("error", {}).get("message", "Failed to send reset email")
        return False, err
    except requests.exceptions.Timeout:
        return False, "Connection timed out"
    except Exception as exc:
        return False, str(exc)

@app.route("/api/forgot-password", methods=["POST"])
def api_forgot_password():
    """AJAX endpoint for forgot password"""
    try:
        data = request.get_json() or {}
        email = data.get("email", "").strip().lower()

        if not email:
            return jsonify({"success": False, "error": "Please enter your email address."})

        if not FIREBASE_AVAILABLE:
            return jsonify({"success": False, "error": "Service unavailable. Please contact your administrator."})

        user_found = False
        user_data = None
        users = fb_get("/users") or {}

        for uid, u_data in users.items():
            if isinstance(u_data, dict) and u_data.get("email", "").lower() == email:
                user_found = True
                user_data = u_data
                break

        if not user_found:
            return jsonify({"success": False, "error": "Account does not exist. Please contact your administrator."})

        if not user_data.get("active", True):
            return jsonify({"success": False, "error": "Account does not exist. Please contact your administrator."})

        ok, err_msg = firebase_send_password_reset(email)

        if ok:
            log.info("Password reset email sent via Firebase for %s", email)
            return jsonify({"success": True, "message": f"Reset link sent to {email}. Please check your email to continue."})
        else:
            log.error("Firebase password reset failed for %s: %s", email, err_msg)
            return jsonify({"success": False, "error": "Failed to send reset email. Please try again later."})

    except Exception as e:
        log.error("API forgot password error: %s", str(e))
        return jsonify({"success": False, "error": "An error occurred. Please try again later."})

def firebase_reset_password_with_code(email: str, password: str, oob_code: str):
    """Reset password using Firebase OOB code. Returns (ok, error_msg)"""
    try:
        url = f"https://identitytoolkit.googleapis.com/v1/accounts:resetPassword?key={FIREBASE_API_KEY}"
        resp = requests.post(url, json={
            "oobCode": oob_code,
            "newPassword": password
        }, timeout=10)

        if resp.status_code == 200:
            return True, ""

        err = resp.json().get("error", {}).get("message", "Failed to reset password")
        return False, err
    except requests.exceptions.Timeout:
        return False, "Connection timed out"
    except Exception as exc:
        return False, str(exc)

@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    email = request.args.get("email", "").strip().lower()
    oob_code = request.args.get("oobCode", "").strip()
    error = None
    message = None
    valid_token = False

    if not email or not oob_code:
        error = "Invalid reset link. Please request a new one."
    else:
        if not FIREBASE_AVAILABLE:
            error = "Password reset service unavailable. Please contact your administrator."
        else:
            try:
                valid_token = True

                if request.method == "POST":
                    password = request.form.get("password", "")
                    confirm = request.form.get("confirm_password", "")

                    if not password or not confirm:
                        error = "Please enter both password fields."
                    elif len(password) < 8:
                        error = "Password must be at least 8 characters long."
                    elif password != confirm:
                        error = "Passwords do not match."
                    else:
                        try:
                            ok, err_msg = firebase_reset_password_with_code(email, password, oob_code)
                            if not ok:
                                error = "Failed to reset password. " + err_msg
                                log.error("Password reset failed for %s: %s", email, err_msg)
                            else:
                                ok, uid, err_msg = firebase_sign_in(email, password)
                                if ok:
                                    profile = load_user_profile(uid)
                                    if profile and profile.get("active", True):
                                        role = normalize_role(profile.get("role", "sales"))
                                        session.permanent = True
                                        session["user_email"] = email
                                        session["user_uid"] = uid
                                        session["user_name"] = profile.get("username", email.split("@")[0])
                                        session["user_role"] = role
                                        log.info("Password reset and auto-login for %s", email)
                                        return redirect(url_for(first_page(role)))

                                message = "Password reset successfully. You can now sign in with your new password."
                                valid_token = False
                                log.info("Password reset successful for %s", email)
                        except Exception as e:
                            error = "An error occurred. Please try again later."
                            log.error("Password reset error: %s", str(e))
            except Exception as e:
                error = "An error occurred. Please try again later."
                log.error("Reset password error: %s", str(e))

    return render_template("reset_password.html", email=email, error=error, message=message, valid_token=valid_token)

# ── Routes: Dashboard ─────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return redirect(url_for(first_page(session.get("user_role", "sales"))))

@app.route("/portfolio")
def portfolio():
    github_username = "archengservices2022"
    featured_repo = "MABS_PIMS"
    return render_template(
        "portfolio.html",
        github_username=github_username,
        github_profile_url=f"https://github.com/{github_username}",
        github_repo_url=f"https://github.com/{github_username}/{featured_repo}",
        featured_repo=featured_repo,
    )

@app.route("/dashboard")
@role_required("dashboard")
def dashboard():
    _auto_flag_overdue()
    invoices = fb_get("/invoices") or {}
    projects = fb_get("/projects") or {}
    quotes   = fb_get("/job_forms") or {}
    expenses = fb_get("/balance_sheet_expenses") or {}

    # ── Stats ──────────────────────────────────────────────────────────────
    inv_list  = [dict(v, firebase_id=k) for k, v in invoices.items()
                 if isinstance(v, dict)] if isinstance(invoices, dict) else []
    proj_list = [dict(v, firebase_id=k) for k, v in projects.items()
                 if isinstance(v, dict) and v.get("project_number")] if isinstance(projects, dict) else []
    quot_list = list(quotes.values())   if isinstance(quotes, dict)   else []

    # Dashboard totals — current year only
    cur_year = str(datetime.now().year)

    cur_year_invs = [i for i in inv_list if isinstance(i, dict)
                     and (i.get("meta", {}).get("invoice_date", "") or "").startswith(cur_year)]
    cur_year_projs = [p for p in proj_list if isinstance(p, dict)
                      and (p.get("created_at", "") or "").startswith(cur_year)]
    cur_year_quots = [q for q in quot_list if isinstance(q, dict)
                      and (q.get("date", "") or q.get("created_at", "") or "").startswith(cur_year)]

    total_invoiced = sum(_safe_float(i.get("meta", {}).get("total", 0)) for i in cur_year_invs)

    # Collected = payments received this year by payment date (across all invoices)
    total_paid = 0.0
    for _i in inv_list:
        if not isinstance(_i, dict): continue
        for _pay in (_i.get("payment_log", []) or []):
            if (_pay.get("date", "") or "").startswith(cur_year):
                total_paid += _safe_float(_pay.get("amount", 0))
        for _tp in (_i.get("tax_payments", []) or []):
            if (_tp.get("date", "") or "").startswith(cur_year):
                total_paid += _safe_float(_tp.get("amount", 0))

    # Outstanding = unpaid balance across ALL years (money still owed)
    total_outstanding = sum(
        max(0.0, _safe_float(i.get("meta", {}).get("total", 0))
            - _safe_float(i.get("meta", {}).get("amount_paid", 0))
            - sum(_safe_float(tp.get("amount", 0)) for tp in (i.get("tax_payments", []) or [])))
        for i in inv_list if isinstance(i, dict)
        and (i.get("meta", {}).get("status", "") not in ("Cancelled",))
    )

    # Current year counts only
    active_projects = sum(1 for p in cur_year_projs
                          if isinstance(p, dict) and p.get("status", "") not in ("Completed", "invoiced_Fully paid", "Cancelled"))
    open_quotes     = sum(1 for q in cur_year_quots
                          if isinstance(q, dict) and q.get("status", "Not Started") not in ("Completed", "Cancelled", "Invoiced"))

    # ── Recent invoices ────────────────────────────────────────────────────
    recent_invoices = sorted(
        [i for i in inv_list if isinstance(i, dict)],
        key=lambda x: x.get("meta", {}).get("created_at", ""),
        reverse=True
    )[:5]

    # ── Recent projects ────────────────────────────────────────────────────
    recent_projects = sorted(
        [p for p in proj_list if isinstance(p, dict)],
        key=lambda x: x.get("created_at", ""),
        reverse=True
    )[:5]

    # ── Monthly revenue for chart — always show last 6 calendar months ────────
    monthly = {}
    for inv in inv_list:
        if not isinstance(inv, dict):
            continue
        date_str = inv.get("meta", {}).get("invoice_date", "") or ""
        try:
            dt  = datetime.fromisoformat(date_str[:10])
            key = dt.strftime("%b %Y")
            amt = float(str(inv.get("meta", {}).get("total", 0) or 0).replace(",", ""))
            monthly[key] = monthly.get(key, 0) + amt
        except Exception:
            pass

    # Build a full 6-month window ending this month (zero-fill gaps)
    from dateutil.relativedelta import relativedelta
    now = datetime.now()
    chart_labels = [(now - relativedelta(months=i)).strftime("%b %Y") for i in range(5, -1, -1)]
    chart_data   = [monthly.get(m, 0) for m in chart_labels]

    # ── Status distribution for donut charts — current year only ──────────────────────────────────
    inv_status_counts = {}
    for i in cur_year_invs:
        if isinstance(i, dict):
            st = i.get("meta", {}).get("status") or "Draft"
            inv_status_counts[st] = inv_status_counts.get(st, 0) + 1

    proj_status_counts = {}
    for p in cur_year_projs:
        if isinstance(p, dict):
            st = p.get("status") or "Not Started"
            proj_status_counts[st] = proj_status_counts.get(st, 0) + 1

    # ── Alert counts — show all warnings (not filtered by year) ──────────────────────────────────────
    today_str     = datetime.now().strftime("%Y-%m-%d")
    week_str      = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    three_day_str = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")
    overdue_count = sum(1 for i in inv_list
                        if isinstance(i, dict) and i.get("meta", {}).get("status", "") == "Overdue")
    _QTERMINAL = {"Approved", "Converted", "Invoiced", "Rejected", "Cancelled", "Expired"}
    expiring_count = sum(1 for q in quot_list
                         if isinstance(q, dict)
                         and q.get("status", "Not Started") not in _QTERMINAL
                         and q.get("valid_until", "")
                         and today_str <= q.get("valid_until", "") <= week_str)
    expiring_soon_quotes = sorted(
        [q for q in quot_list
         if isinstance(q, dict)
         and q.get("status", "Not Started") not in _QTERMINAL
         and q.get("valid_until", "")
         and today_str <= q.get("valid_until", "") <= three_day_str],
        key=lambda x: x.get("valid_until", "")
    )

    # ── Action queue ──────────────────────────────────────────────────────────
    followup_quotes = sorted(
        [q for q in quot_list
         if isinstance(q, dict)
         and q.get("status", "") not in _QTERMINAL
         and q.get("follow_up_date", "")
         and q.get("follow_up_date", "") <= today_str],
        key=lambda x: x.get("follow_up_date", "")
    )[:8]

    approved_quotes = sorted(
        [q for q in quot_list
         if isinstance(q, dict)
         and q.get("status", "") == "Approved"],
        key=lambda x: x.get("date", x.get("created_at", "")),
        reverse=True
    )[:6]

    # Active projects by status for pipeline view
    # Map all status variants to display buckets
    _IN_PROGRESS = {"In Progress", "Active", "invoiced_Not paid yet",
                    "invoiced_Partially paid", "Invoiced", "Sent",
                    "Sent out_Invoiced"}
    _NOT_STARTED = {"Not Started"}
    _ON_HOLD     = {"On Hold"}

    def _pipeline_bucket(status):
        s = (status or "Not Started").strip()
        if s in _IN_PROGRESS:  return "In Progress"
        if s in _ON_HOLD:      return "On Hold"
        if s in _NOT_STARTED:  return "Not Started"
        return None  # Completed / Cancelled / invoiced_Fully paid — excluded

    pipeline = {"Not Started": [], "In Progress": [], "On Hold": []}
    for _p in cur_year_projs:
        if not isinstance(_p, dict): continue
        _bucket = _pipeline_bucket(_p.get("status", "Not Started"))
        if _bucket:
            pipeline[_bucket].append(_p)

    # ── Urgent alerts (overdue + due within 3 days only) ─────────────────────
    three_day_str = (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")

    # 1. Overdue invoices — ALL years (show all warnings)
    reminder_overdue_invoices = sorted(
        [i for i in inv_list if isinstance(i, dict)
         and i.get("meta", {}).get("status", "") == "Overdue"],
        key=lambda x: x.get("meta", {}).get("due_date", "")
    )[:5]

    # 2. Invoices due within 3 days (not yet overdue) — ALL years (show all warnings)
    reminder_due_soon = sorted(
        [i for i in inv_list if isinstance(i, dict)
         and i.get("meta", {}).get("status", "") not in ("Paid", "Overdue", "Cancelled")
         and i.get("meta", {}).get("due_date", "")
         and today_str <= i.get("meta", {}).get("due_date", "") <= three_day_str],
        key=lambda x: x.get("meta", {}).get("due_date", "")
    )[:5]

    reminder_proj_due = []
    reminder_stalled  = []
    reminder_total = len(reminder_overdue_invoices) + len(reminder_due_soon)

    # ── Projects ready to invoice (have Pending Invoice stages) — current year only ──────────────
    projects_ready_to_invoice = []
    for p in cur_year_projs:
        if not isinstance(p, dict):
            continue
        if p.get("status", "") in ("Completed", "invoiced_Fully paid", "Cancelled"):
            continue
        stages = p.get("payment_stages", []) or []
        pending_stages = [s for s in stages if isinstance(s, dict) and s.get("status", "") == "Pending Invoice"]
        if pending_stages:
            pending_amt = sum(_safe_float(s.get("amount", 0)) for s in pending_stages)
            projects_ready_to_invoice.append({
                "firebase_id":    p.get("firebase_id", ""),
                "project_number": p.get("project_number", ""),
                "project_name":   p.get("project_name", ""),
                "client_name":    p.get("client_name", ""),
                "pending_stages": len(pending_stages),
                "pending_amount": pending_amt,
            })
    projects_ready_to_invoice = sorted(projects_ready_to_invoice, key=lambda x: x["project_number"], reverse=True)[:8]

    # ── This month vs last month collected ────────────────────────────────────
    this_month_str = datetime.now().strftime("%Y-%m")
    last_month_str = (datetime.now() - timedelta(days=30)).strftime("%Y-%m")
    this_month_collected = 0.0
    last_month_collected = 0.0
    for inv in inv_list:
        if not isinstance(inv, dict):
            continue
        for pay in (inv.get("payment_log", []) or []):
            ds = pay.get("date", "") or ""
            if ds[:7] == this_month_str:
                this_month_collected += _safe_float(pay.get("amount", 0))
            elif ds[:7] == last_month_str:
                last_month_collected += _safe_float(pay.get("amount", 0))
        for tp in (inv.get("tax_payments", []) or []):
            ds = tp.get("date", "") or ""
            if ds[:7] == this_month_str:
                this_month_collected += _safe_float(tp.get("amount", 0))
            elif ds[:7] == last_month_str:
                last_month_collected += _safe_float(tp.get("amount", 0))

    # ── Module overview stats — current year only ─────────────────────────────────────────────────
    _QTERMINAL_ALL = {"Approved", "Converted", "Invoiced", "Rejected", "Cancelled", "Expired"}
    quot_status_counts: Dict[str, int] = {}
    quotes_pipeline_value = 0.0
    quotes_converted = 0
    for _q in cur_year_quots:
        if not _q or not isinstance(_q, dict): continue
        _st = _q.get("status", "Not Started")
        quot_status_counts[_st] = quot_status_counts.get(_st, 0) + 1
        if _st not in {"Rejected", "Cancelled", "Expired"}:
            quotes_pipeline_value += _safe_float(_q.get("total", 0))
        # Count as converted if: status is Converted/Invoiced OR has linked_project_id (was converted)
        if _st in {"Converted", "Invoiced"} or _q.get("linked_project_id"):
            quotes_converted += 1
    _total_quotes = len(cur_year_quots)
    quotes_conversion_rate = int(quotes_converted / _total_quotes * 100) if _total_quotes > 0 else 0

    # Active Contract Value = sum of contract values for active projects created in current year
    proj_contract_active = sum(
        _safe_float(p.get("contract_value", 0)) for p in cur_year_projs
        if isinstance(p, dict) and p.get("status", "") not in ("Completed", "invoiced_Fully paid", "Cancelled")
    )
    proj_contract_total  = sum(_safe_float(p.get("contract_value", 0)) for p in cur_year_projs if isinstance(p, dict))
    proj_completed_count = sum(1 for p in cur_year_projs if isinstance(p, dict) and p.get("status", "") in ("Completed", "invoiced_Fully paid"))

    inv_overdue_amt = sum(
        _safe_float(i.get("meta", {}).get("total", 0)) - _safe_float(i.get("meta", {}).get("amount_paid", 0))
        for i in cur_year_invs if isinstance(i, dict) and i.get("meta", {}).get("status", "") == "Overdue"
    )

    # ── Team status (Employees module) ─────────────────────────────────────
    all_time_entries = _load_time_entries()
    all_time_off     = _load_time_off_requests()
    clocked_in_now   = [e for e in all_time_entries if e.get("status") == "open"]
    pending_time_off = [r for r in all_time_off if r.get("status") == "Pending"]

    # ── Commission widget ──────────────────────────────────────────────────────
    _dash_role = normalize_role(session.get("user_role", ""))
    _dash_name = session.get("user_name", "")
    _cur_month = datetime.now().strftime("%Y-%m")
    _dash_commission: Dict[str, object] = {}

    # Project status lookup for cancellation check
    _dash_proj_status: Dict[str, str] = {}
    if isinstance(projects, dict):
        for _pid, _pd in projects.items():
            if _pd and isinstance(_pd, dict):
                _dash_proj_status[_pid] = _pd.get("status", "")

    _CONV_DASH = {"Converted", "Invoiced"}

    if _dash_role == "sales":
        # Personal commission for logged-in salesperson
        _uid = session.get("user_uid", "")
        _u_data = fb_get(f"/users/{_uid}") or {} if _uid else {}
        _rate = _safe_float(_u_data.get("commission_rate", 0))
        _emp_type = _u_data.get("employee_type", "")
        _my_quotes = [q for q in quot_list if isinstance(q, dict) and
                      (q.get("salesperson") or "").strip() == _dash_name]
        _my_total = len(_my_quotes)
        _my_conv = 0
        _my_rev = 0.0
        _my_rev_month = 0.0
        for _q in _my_quotes:
            _linked = _q.get("linked_project_id", "")
            _is_conv = _q.get("status", "") in _CONV_DASH or bool(_linked)
            if _is_conv and not (_linked and _dash_proj_status.get(_linked) == "Cancelled"):
                _my_conv += 1
                _qval = _safe_float(_q.get("total", 0))
                _my_rev += _qval
                if (_q.get("date", "") or "").startswith(_cur_month):
                    _my_rev_month += _qval
        _dash_commission = {
            "role":             "sales",
            "employee_type":    _emp_type,
            "commission_rate":  _rate,
            "quotes_total":     _my_total,
            "quotes_converted": _my_conv,
            "win_rate":         round(_my_conv / _my_total * 100) if _my_total else 0,
            "revenue_alltime":  _my_rev,
            "revenue_month":    _my_rev_month,
            "earned_alltime":   _my_rev * _rate / 100,
            "earned_month":     _my_rev_month * _rate / 100,
            "rate_set":         _rate > 0,
        }

    elif _dash_role == "admin":
        # Admin summary: total commissions across all salespeople
        _all_users = _load_all_users()
        _comm_map: Dict[str, float] = {}
        for _u in _all_users:
            if normalize_role(_u.get("role", "")) == "sales":
                _uname = (_u.get("username") or "").strip()
                if _uname:
                    _comm_map[_uname] = _safe_float(_u.get("commission_rate", 0))
        _admin_comm_total = 0.0
        _admin_comm_month = 0.0
        for _q in quot_list:
            if not isinstance(_q, dict): continue
            _sp = (_q.get("salesperson") or "").strip()
            _rate = _comm_map.get(_sp, 0)
            if not _rate: continue
            _linked = _q.get("linked_project_id", "")
            _is_conv = _q.get("status", "") in _CONV_DASH or bool(_linked)
            if _is_conv and not (_linked and _dash_proj_status.get(_linked) == "Cancelled"):
                _qval = _safe_float(_q.get("total", 0))
                _admin_comm_total += _qval * _rate / 100
                if (_q.get("date", "") or "").startswith(_cur_month):
                    _admin_comm_month += _qval * _rate / 100
        _dash_commission = {
            "role":        "admin",
            "total":       _admin_comm_total,
            "this_month":  _admin_comm_month,
        }

    # Pending employee expense approvals (admin only)
    pending_expenses_dash = []
    if normalize_role(session.get("user_role", "")) == "admin":
        raw_dash_exp = fb_get("/expenses") or {}
        if isinstance(raw_dash_exp, dict):
            for eid, edata in raw_dash_exp.items():
                if isinstance(edata, dict) and edata.get("status") == "Pending":
                    edata["firebase_id"] = eid
                    pending_expenses_dash.append(edata)
        pending_expenses_dash.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    return render_template("dashboard.html",
        clocked_in_now=clocked_in_now,
        pending_time_off=pending_time_off,
        pending_expenses_dash=pending_expenses_dash,
        total_invoiced=total_invoiced,
        total_paid=total_paid,
        total_outstanding=total_outstanding,
        active_projects=active_projects,
        open_quotes=open_quotes,
        total_invoices=len(cur_year_invs),
        recent_invoices=recent_invoices,
        recent_projects=recent_projects,
        chart_labels=json.dumps(chart_labels),
        chart_data=json.dumps(chart_data),
        overdue_count=overdue_count,
        expiring_count=expiring_count,
        expiring_soon_quotes=expiring_soon_quotes,
        followup_quotes=followup_quotes,
        approved_quotes=approved_quotes,
        pipeline=pipeline,
        inv_status_labels=json.dumps(list(inv_status_counts.keys())),
        inv_status_data=json.dumps(list(inv_status_counts.values())),
        proj_status_labels=json.dumps(list(proj_status_counts.keys())),
        proj_status_data=json.dumps(list(proj_status_counts.values())),
        ai_enabled=bool(_get_ai_client()),
        reminder_overdue_invoices=reminder_overdue_invoices,
        reminder_due_soon=reminder_due_soon,
        reminder_proj_due=reminder_proj_due,
        reminder_stalled=reminder_stalled,
        reminder_total=reminder_total,
        projects_ready_to_invoice=projects_ready_to_invoice,
        this_month_collected=this_month_collected,
        last_month_collected=last_month_collected,
        inv_status_counts=inv_status_counts,
        proj_status_counts=proj_status_counts,
        proj_list=proj_list,
        quot_status_counts=quot_status_counts,
        quotes_approved_count=sum(quot_status_counts.get(s, 0) for s in ('Approved', 'Converted', 'Invoiced', 'Completed')),
        quotes_pipeline_value=quotes_pipeline_value,
        quotes_conversion_rate=quotes_conversion_rate,
        quotes_converted=quotes_converted,
        proj_contract_total=proj_contract_total,
        proj_contract_active=proj_contract_active,
        proj_completed_count=proj_completed_count,
        inv_overdue_amt=inv_overdue_amt,
        dash_commission=_dash_commission,
    )

# ── Routes: Sales Dashboard ───────────────────────────────────────────────────
@app.route("/sales-dashboard")
@role_required("sales_dashboard")
def sales_dashboard():
    user_name  = session.get("user_name", "")
    raw        = fb_get("/job_forms") or {}
    today_str  = datetime.now().strftime("%Y-%m-%d")
    week_str   = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    _TERMINAL  = {"Approved", "Converted", "Invoiced", "Rejected", "Cancelled", "Expired"}

    all_quotes: list = []
    for fid, fdata in (raw.items() if isinstance(raw, dict) else []):
        if fdata and isinstance(fdata, dict):
            fdata["firebase_id"] = fid
            fdata.setdefault("status", "Not Started")
            all_quotes.append(fdata)

    # Filter to current salesperson's quotes when possible
    my_quotes = ([q for q in all_quotes if q.get("salesperson", "") == user_name]
                 if user_name else all_quotes) or all_quotes

    open_quotes     = [q for q in my_quotes if q.get("status") not in _TERMINAL]
    converted       = [q for q in my_quotes if q.get("status") in ("Converted", "Invoiced", "Approved")]
    pipeline_value  = sum(_safe_float(q.get("total", 0)) for q in open_quotes)
    total_value     = sum(_safe_float(q.get("total", 0)) for q in my_quotes)
    win_rate        = round(len(converted) / len(my_quotes) * 100) if my_quotes else 0

    next14_str = (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")
    followups_due = sorted(
        [q for q in my_quotes
         if q.get("status") not in _TERMINAL
         and q.get("follow_up_date", "")
         and q.get("follow_up_date", "") <= next14_str],
        key=lambda x: x.get("follow_up_date", "")
    )

    expiring_soon = sorted(
        [q for q in my_quotes
         if q.get("status") not in _TERMINAL
         and q.get("valid_until", "")
         and today_str <= q.get("valid_until", "") <= week_str],
        key=lambda x: x.get("valid_until", "")
    )

    recent_quotes = sorted(
        my_quotes,
        key=lambda x: x.get("created_at", x.get("date", "")),
        reverse=True
    )[:8]

    # Commission stats
    _uid = session.get("user_uid", "")
    _u_data = fb_get(f"/users/{_uid}") or {} if _uid else {}
    _comm_rate = _safe_float(_u_data.get("commission_rate", 0))
    _emp_type = _u_data.get("employee_type", "")
    _cur_month = datetime.now().strftime("%Y-%m")
    _proj_raw = fb_get("/projects") or {}
    _proj_status_sd: Dict[str, str] = {}
    if isinstance(_proj_raw, dict):
        for _pid, _pd in _proj_raw.items():
            if _pd and isinstance(_pd, dict):
                _proj_status_sd[_pid] = _pd.get("status", "")
    _CONV_SD = {"Converted", "Invoiced"}
    _comm_rev_alltime = 0.0
    _comm_rev_month = 0.0
    _conv_count = 0
    for _q in my_quotes:
        _linked = _q.get("linked_project_id", "")
        _is_conv = _q.get("status", "") in _CONV_SD or bool(_linked)
        if _is_conv and not (_linked and _proj_status_sd.get(_linked) == "Cancelled"):
            _qval = _safe_float(_q.get("total", 0))
            _conv_count += 1
            _comm_rev_alltime += _qval
            if (_q.get("date", "") or "").startswith(_cur_month):
                _comm_rev_month += _qval
    commission = {
        "rate":           _comm_rate,
        "rate_set":       _comm_rate > 0,
        "employee_type":  _emp_type,
        "earned_month":   _comm_rev_month * _comm_rate / 100,
        "earned_alltime": _comm_rev_alltime * _comm_rate / 100,
        "revenue_month":  _comm_rev_month,
        "revenue_alltime": _comm_rev_alltime,
        "conv_count":     _conv_count,
    }

    return render_template("sales_dashboard.html",
        my_quotes=my_quotes,
        open_quotes=open_quotes,
        converted=converted,
        pipeline_value=pipeline_value,
        total_value=total_value,
        win_rate=win_rate,
        followups_due=followups_due,
        expiring_soon=expiring_soon,
        recent_quotes=recent_quotes,
        today_str=today_str,
        commission=commission,
    )

# ── Routes: Quotes ────────────────────────────────────────────────────────────
@app.route("/quotes")
@role_required("quotes")
def quotes():
    raw = fb_get("/job_forms") or {}
    _QUOTE_TERMINAL = {"Approved", "Converted", "Invoiced", "Rejected", "Cancelled", "Expired"}
    today_str = datetime.now().strftime("%Y-%m-%d")
    items = []
    for fid, fdata in (raw.items() if isinstance(raw, dict) else []):
        if fdata and isinstance(fdata, dict):
            fdata["firebase_id"] = fid
            fdata.setdefault("status", "Not Started")
            # Auto-sync: non-standard statuses → standard equivalents
            if fdata["status"] in {"Draft", "In Review", "On Hold", "Completed"}:
                fb_update(f"/job_forms/{fid}", {
                    "status":     "In Progress",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                })
                fdata["status"] = "In Progress"
            if fdata["status"] == "Invoiced":
                fb_update(f"/job_forms/{fid}", {
                    "status":     "Converted",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                })
                fdata["status"] = "Converted"
            # Auto-sync: quote with a linked project should be Converted
            if (fdata.get("linked_project_id")
                    and fdata["status"] not in {"Converted", "Rejected", "Cancelled"}):
                fb_update(f"/job_forms/{fid}", {
                    "status":     "Converted",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                })
                fdata["status"] = "Converted"
            # Auto-expire quotes past valid_until date
            valid_until = fdata.get("valid_until", "")
            if (valid_until and valid_until < today_str
                    and fdata["status"] not in _QUOTE_TERMINAL):
                fb_update(f"/job_forms/{fid}", {
                    "status": "Expired",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
                fdata["status"] = "Expired"
            # Expiry warning: days remaining for active quotes
            if valid_until and fdata["status"] not in _QUOTE_TERMINAL:
                try:
                    from datetime import date as _date
                    delta = (_date.fromisoformat(valid_until[:10]) - _date.today()).days
                    fdata["_days_until_expiry"] = delta
                except Exception:
                    pass
            items.append(fdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    search        = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "")
    year_filter   = request.args.get("year", "")
    month_filter  = request.args.get("month", "")
    date_from     = request.args.get("from", "")
    date_to       = request.args.get("to", "")

    if search:
        items = [i for i in items if search in str(i).lower()]
    if status_filter:
        items = [i for i in items if i.get("status", "") == status_filter]
    if year_filter:
        items = [i for i in items if (i.get("date") or "").startswith(year_filter)]
    if year_filter and month_filter:
        prefix = f"{year_filter}-{month_filter.zfill(2)}"
        items = [i for i in items if (i.get("date") or "").startswith(prefix)]
    if date_from:
        items = [i for i in items if (i.get("date") or "") >= date_from]
    if date_to:
        items = [i for i in items if (i.get("date") or "") <= date_to]

    # Build available years from all quote dates (before filtering)
    all_items_raw = []
    for fid, fdata in (fb_get("/job_forms") or {}).items() if isinstance(fb_get("/job_forms") or {}, dict) else []:
        if fdata and isinstance(fdata, dict):
            all_items_raw.append(fdata)
    available_years = sorted(
        {(d.get("date") or "")[:4] for d in all_items_raw if len((d.get("date") or "")) >= 4},
        reverse=True
    )

    # Build follow-ups lists from ALL items (unaffected by list filters)
    from datetime import date as _date, timedelta as _td
    _today     = datetime.now().strftime("%Y-%m-%d")
    _today_d   = _date.today()
    _in14      = (_today_d + _td(days=14)).isoformat()
    follow_ups = []       # due today or overdue
    upcoming_followups = []  # due tomorrow → +14 days
    for q in all_items_raw:
        fu = q.get("follow_up_date", "")
        if not fu or q.get("status", "") in _QUOTE_TERMINAL:
            continue
        q_copy = dict(q)
        if fu <= _today:
            try:
                q_copy["_fu_days_overdue"] = (_today_d - _date.fromisoformat(fu[:10])).days
            except Exception:
                q_copy["_fu_days_overdue"] = 0
            follow_ups.append(q_copy)
        elif fu <= _in14:
            try:
                q_copy["_fu_days_ahead"] = (_date.fromisoformat(fu[:10]) - _today_d).days
            except Exception:
                q_copy["_fu_days_ahead"] = 0
            upcoming_followups.append(q_copy)
    follow_ups.sort(key=lambda x: x.get("follow_up_date", ""))
    upcoming_followups.sort(key=lambda x: x.get("follow_up_date", ""))

    statuses   = ["Not Started", "In Progress", "Sent", "Approved", "Converted", "Rejected", "Expired", "Cancelled"]
    active_tab = request.args.get("tab", "all")
    today_date = datetime.now().strftime("%Y-%m-%d")

    # KPI stats computed from filtered items (matches table rows)
    _OPEN_STATUSES      = {"Not Started", "In Progress"}
    _APPROVED_STATUSES  = {"Approved", "Completed"}
    _CONVERTED_STATUSES = {"Converted", "Invoiced"}
    q_total     = len(items)
    q_open      = sum(1 for q in items if q.get("status", "Not Started") in _OPEN_STATUSES)
    q_approved  = sum(1 for q in items if q.get("status", "") in _APPROVED_STATUSES)
    # Count as converted if: status is Converted/Invoiced OR has linked_project_id (was converted)
    q_converted = sum(1 for q in items if q.get("status", "") in _CONVERTED_STATUSES or q.get("linked_project_id"))
    q_conv_rate = round(q_converted / q_total * 100) if q_total else 0
    q_pipeline  = sum(_safe_float(q.get("total", 0)) for q in items if q.get("status", "Not Started") in _OPEN_STATUSES | _APPROVED_STATUSES)
    # Count won value if: status is Converted/Invoiced OR has linked_project_id (was converted)
    q_won_val   = sum(_safe_float(q.get("total", 0)) for q in items if q.get("status", "") in _CONVERTED_STATUSES or q.get("linked_project_id"))

    # ── Commission calculation (Sales People tab) ─────────────────────────────
    # Build commission_rate lookup keyed by username (sales users only)
    _comm_by_name: Dict[str, dict] = {}
    for _u in _load_all_users():
        if normalize_role(_u.get("role", "")) == "sales":
            _uname = (_u.get("username") or "").strip()
            if _uname:
                _comm_by_name[_uname] = {
                    "employee_type":   _u.get("employee_type", ""),
                    "commission_rate": _safe_float(_u.get("commission_rate", 0)),
                    "email":           _u.get("email", ""),
                }

    # Load project statuses to detect cancelled linked projects
    _proj_raw = fb_get("/projects") or {}
    _proj_status: Dict[str, str] = {}
    if isinstance(_proj_raw, dict):
        for _pid, _pd in _proj_raw.items():
            if _pd and isinstance(_pd, dict):
                _proj_status[_pid] = _pd.get("status", "")

    # Aggregate per-salesperson stats from ALL quotes (unfiltered)
    _sp_stats: Dict[str, dict] = {}
    _CONV_ST = {"Converted", "Invoiced"}
    for _q in all_items_raw:
        _sp = (_q.get("salesperson") or "").strip()
        if not _sp:
            continue
        if _sp not in _sp_stats:
            _ci = _comm_by_name.get(_sp, {})
            _sp_stats[_sp] = {
                "name":              _sp,
                "email":             _ci.get("email", ""),
                "employee_type":     _ci.get("employee_type", ""),
                "commission_rate":   _ci.get("commission_rate", 0.0),
                "quotes_total":      0,
                "quotes_converted":  0,
                "revenue_generated": 0.0,
                "commission_earned": 0.0,
            }
        _sp_stats[_sp]["quotes_total"] += 1
        _linked = _q.get("linked_project_id", "")
        _is_conv = _q.get("status", "") in _CONV_ST or bool(_linked)
        if _is_conv:
            _proj_cancelled = _linked and _proj_status.get(_linked, "") == "Cancelled"
            if not _proj_cancelled:
                _sp_stats[_sp]["quotes_converted"] += 1
                _sp_stats[_sp]["revenue_generated"] += _safe_float(_q.get("total", 0))
                _rate = _sp_stats[_sp]["commission_rate"]
                _sp_stats[_sp]["commission_earned"] += _safe_float(_q.get("total", 0)) * _rate / 100

    for _s in _sp_stats.values():
        _t = _s["quotes_total"]
        _s["win_rate"] = round(_s["quotes_converted"] / _t * 100) if _t else 0

    salesperson_stats = sorted(_sp_stats.values(), key=lambda x: x["commission_earned"], reverse=True)

    # ── Monthly commission breakdown for Payroll tab (admin only) ─────────────
    _monthly_comm: Dict[str, Dict[str, dict]] = {}  # {YYYY-MM: {sp_name: {earned, revenue}}}
    for _q in all_items_raw:
        _sp = (_q.get("salesperson") or "").strip()
        if not _sp or _sp not in _sp_stats:
            continue
        _rate = _sp_stats[_sp]["commission_rate"]
        if not _rate:
            continue
        _linked = _q.get("linked_project_id", "")
        _is_conv = _q.get("status", "") in _CONV_ST or bool(_linked)
        if not _is_conv:
            continue
        if _linked and _proj_status.get(_linked, "") == "Cancelled":
            continue
        _qdate = (_q.get("date") or "")[:7]  # YYYY-MM
        if not _qdate:
            continue
        if _qdate not in _monthly_comm:
            _monthly_comm[_qdate] = {}
        if _sp not in _monthly_comm[_qdate]:
            _monthly_comm[_qdate][_sp] = {"earned": 0.0, "revenue": 0.0}
        _qval = _safe_float(_q.get("total", 0))
        _monthly_comm[_qdate][_sp]["earned"]  += _qval * _rate / 100
        _monthly_comm[_qdate][_sp]["revenue"] += _qval

    # Load existing commission payments
    _comm_payments_raw = fb_get("/commission_payments") or {}
    _paid_set: set = set()  # (period, sp_name)
    _comm_payments_list = []
    if isinstance(_comm_payments_raw, dict):
        for _cpid, _cp in _comm_payments_raw.items():
            if _cp and isinstance(_cp, dict):
                _paid_set.add((_cp.get("period", ""), _cp.get("salesperson", "")))
                _comm_payments_list.append({**_cp, "id": _cpid})

    # Build sorted monthly rows for template
    monthly_payroll = []
    for _period in sorted(_monthly_comm.keys(), reverse=True):
        for _sp_name, _vals in sorted(_monthly_comm[_period].items()):
            monthly_payroll.append({
                "period":    _period,
                "sp_name":   _sp_name,
                "earned":    _vals["earned"],
                "revenue":   _vals["revenue"],
                "paid":      (_period, _sp_name) in _paid_set,
                "rate":      _sp_stats[_sp_name]["commission_rate"],
                "emp_type":  _sp_stats[_sp_name]["employee_type"],
            })

    return render_template("quotes.html", quotes=items, statuses=statuses,
                           search=search, status_filter=status_filter,
                           year_filter=year_filter, month_filter=month_filter,
                           date_from=date_from, date_to=date_to,
                           available_years=available_years,
                           follow_ups=follow_ups,
                           upcoming_followups=upcoming_followups,
                           clients=_load_clients(), sales_people=_load_sales_people(),
                           active_tab=active_tab, today_date=today_date,
                           next_num=_next_quote_number(),
                           q_total=q_total, q_open=q_open, q_approved=q_approved,
                           q_converted=q_converted, q_conv_rate=q_conv_rate,
                           q_pipeline=q_pipeline, q_won_val=q_won_val,
                           salesperson_stats=salesperson_stats,
                           monthly_payroll=monthly_payroll,
                           comm_payments=_comm_payments_list)

@app.route("/quotes/export")
@role_required("quotes")
def quotes_export():
    import io

    raw = fb_get("/job_forms") or {}
    items = []
    for fid, fdata in (raw.items() if isinstance(raw, dict) else []):
        if fdata and isinstance(fdata, dict):
            fdata["firebase_id"] = fid
            items.append(fdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    status_filter = request.args.get("status", "")
    date_from     = request.args.get("from", "")
    date_to       = request.args.get("to", "")
    if status_filter:
        items = [i for i in items if i.get("status", "") == status_filter]
    if date_from:
        items = [i for i in items if (i.get("date") or "") >= date_from]
    if date_to:
        items = [i for i in items if (i.get("date") or "") <= date_to]

    import csv
    output = io.StringIO()
    w = csv.writer(output)
    co = company_info()

    def fmt_csv_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    # Add company header and blank row
    w.writerow([f"{co.get('name','')} - Quotes Report"])
    w.writerow([])

    headers = ["Quote Number", "Client", "Project / Scope", "Salesperson", "Date", "Valid Until", "Status", "Subtotal", "Tax", "Total", "Notes"]
    w.writerow(headers)

    for q in items:
        total = _safe_float(q.get("total", 0))
        subtotal = _safe_float(q.get("subtotal", 0))
        tax = total - subtotal
        w.writerow([
            q.get("job_number",""),
            q.get("client_name",""),
            q.get("project_name",""),
            q.get("salesperson",""),
            fmt_csv_date(q.get("date","")),
            fmt_csv_date(q.get("valid_until","")),
            q.get("status",""),
            f"{subtotal:.2f}",
            f"{tax:.2f}",
            f"{total:.2f}",
            q.get("notes","")
        ])

    output.seek(0)
    from flask import Response
    fname = f"quotes_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/quotes/export/excel")
@role_required("quotes")
def quotes_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    raw = fb_get("/job_forms") or {}
    items = []
    for fid, fdata in (raw.items() if isinstance(raw, dict) else []):
        if fdata and isinstance(fdata, dict):
            items.append(fdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    if request.args.get("status"):
        items = [i for i in items if i.get("status","") == request.args["status"]]
    if request.args.get("from"):
        items = [i for i in items if (i.get("date") or "") >= request.args["from"]]
    if request.args.get("to"):
        items = [i for i in items if (i.get("date") or "") <= request.args["to"]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes"
    co = company_info()

    hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    hdr_font = Font(color="FFFFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13, color="FF0F766E")
    alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add title row
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=f"{co.get('name','')} - Quotes Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Add date range row
    _pdf_from = request.args.get("from", "")
    _pdf_to   = request.args.get("to", "")
    _date_range = ""
    if _pdf_from and _pdf_to:
        _date_range = f"{_pdf_from} to {_pdf_to}"
    elif _pdf_from:
        _date_range = f"From {_pdf_from}"
    elif _pdf_to:
        _date_range = f"Up to {_pdf_to}"

    if _date_range:
        ws.merge_cells('A2:K2')
        date_cell = ws.cell(row=2, column=1, value=_date_range)
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16

    headers = ["Quote Number","Client","Project / Scope","Salesperson","Date","Valid Until",
               "Status","Subtotal ($)","Tax ($)","Total ($)","Notes"]
    header_row = 3 if _date_range else 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr

    def fmt_date(d):
        if not d or d == "—":
            return "—"
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for q in items:
        date_str = q.get("date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(q)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda q: q.get("created_at", ""))

    ri = header_row + 1

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for q in grouped[year][month][day]:
                    row = [q.get("job_number",""), q.get("client_name",""), q.get("project_name",""),
                           q.get("salesperson",""), fmt_date(q.get("date","")), fmt_date(q.get("valid_until","")),
                           q.get("status",""), _safe_float(q.get("subtotal",0)),
                           _safe_float(q.get("tax_amount",0)), _safe_float(q.get("total",0)),
                           q.get("notes","")]
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if ri % 2 == 0:
                            cell.fill = alt_fill
                        if ci in (8, 9, 10):
                            cell.number_format = '"$"#,##0.00'
                        cell.alignment = ctr
                    ri += 1

    # Increase column widths
    col_widths = [18, 25, 35, 20, 14, 14, 14, 16, 14, 16, 30]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{header_row + 1}"

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    fname = f"quotes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/quotes/export/pdf")
@role_required("quotes")
def quotes_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab is not installed. Run: pip install reportlab", "danger")
        return redirect(url_for("quotes", tab="export"))

    import io as _io
    raw = fb_get("/job_forms") or {}
    items = []
    for fid, fdata in (raw.items() if isinstance(raw, dict) else []):
        if fdata and isinstance(fdata, dict):
            items.append(fdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    if request.args.get("status"):
        items = [i for i in items if i.get("status","") == request.args["status"]]
    if request.args.get("from"):
        items = [i for i in items if (i.get("date") or "") >= request.args["from"]]
    if request.args.get("to"):
        items = [i for i in items if (i.get("date") or "") <= request.args["to"]]

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    co = company_info()
    elems = []

    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=15,
                              fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                              alignment=1)  # 1 = CENTER
    sub_s   = ParagraphStyle("S", parent=styles["Normal"], fontSize=9,
                              textColor=colors.HexColor("#64748B"), spaceAfter=14,
                              alignment=0)  # 0 = LEFT

    elems.append(Paragraph(f"{co.get('name','')} — Quotes Report", title_s))
    elems.append(Spacer(1, 0.2*inch))
    _pdf_from = request.args.get("from", "")
    _pdf_to   = request.args.get("to", "")
    _date_range = ""
    if _pdf_from and _pdf_to:
        _date_range = f"{_pdf_from} to {_pdf_to}"
    elif _pdf_from:
        _date_range = f"From {_pdf_from}"
    elif _pdf_to:
        _date_range = f"Up to {_pdf_to}"
    if _date_range:
        elems.append(Paragraph(_date_range, sub_s))
        elems.append(Spacer(1, 0.15*inch))

    hdrs = ["Quote Number","Client","Project / Scope","Salesperson","Date","Status","Total"]
    data = [hdrs]
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=10, wordWrap='CJK')
    group_style = ParagraphStyle("group", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", alignment=1, leading=10, textColor=colors.HexColor("#0F172A"))

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for q in items:
        date_str = q.get("date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(q)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda q: q.get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for q in grouped[year][month][day]:
                    date_str = q.get("date","—")
                    if date_str and date_str != "—" and len(str(date_str)) >= 10:
                        parts = str(date_str)[:10].split("-")
                        if len(parts) == 3:
                            date_str = f"{parts[1]}-{parts[2]}-{parts[0]}"
                    total = _safe_float(q.get('total',0))
                    data.append([
                        Paragraph(q.get("job_number","—"), cell_style),
                        Paragraph(q.get("client_name","—"), cell_style),
                        Paragraph(q.get("project_name") or "—", cell_style),
                        Paragraph(q.get("salesperson","—"), cell_style),
                        Paragraph(date_str, cell_style),
                        Paragraph(q.get("status","—"), cell_style),
                        Paragraph(f"${total:,.2f}", cell_style),
                    ])

    cw = [1.2*inch, 2.0*inch, 3.0*inch, 1.8*inch, 1.1*inch, 1.2*inch, 1.1*inch]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)

    from flask import Response
    fname = f"quotes_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/sales-people/new", methods=["POST"])
@role_required("quotes")
def sales_person_new():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Name is required.", "danger")
    else:
        person = {
            "name":       name,
            "phone":      request.form.get("phone", "").strip(),
            "email":      request.form.get("email", "").strip(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        fb_push("/sales_persons", person)
        # Keep local file in sync for desktop compatibility
        try:
            lp = DATA_DIR / "sales_persons.json"
            existing = []
            if lp.exists():
                with open(lp, encoding="utf-8") as f:
                    existing = json.load(f)
            if not isinstance(existing, list):
                existing = []
            existing.append(person)
            with open(lp, "w", encoding="utf-8") as f:
                json.dump(existing, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            log.warning("Could not update local sales_persons.json: %s", exc)
        flash(f"Sales person '{name}' added.", "success")
    return redirect(url_for("quotes", tab="salespeople"))

@app.route("/sales-people/<person_id>/delete", methods=["POST"])
@role_required("quotes")
def sales_person_delete(person_id):
    fb_delete(f"/sales_persons/{person_id}")
    flash("Sales person removed.", "success")
    return redirect(url_for("quotes", tab="salespeople"))

@app.route("/quotes/new", methods=["GET", "POST"])
@role_required("quotes")
def quotes_new():
    clients   = _load_clients()
    sales_ppl = _load_sales_people()
    if request.method == "POST":
        data = _parse_quote_form(request.form)

        # Validate quote number is not duplicate
        job_num = data.get("job_number", "").strip()
        if job_num:
            existing_quote, _ = _find_quote_by_number(job_num)
            if existing_quote:
                flash(f"Quote number {job_num} already exists. Please use a different number.", "danger")
                return render_template("quote_form.html", quote=data, clients=clients,
                                     sales_people=sales_ppl, is_new=True,
                                     next_num=_next_quote_number())

        data["created_at"] = datetime.now(timezone.utc).isoformat()
        data["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["created_by"] = session.get("user_email", "")
        key = fb_push("/job_forms", data)
        if key:
            flash("Quote created successfully.", "success")
        else:
            flash("Quote saved locally (Firebase offline).", "warning")
        return redirect(url_for("quotes"))
    return render_template("quote_form.html", quote=None, clients=clients,
                           sales_people=sales_ppl, is_new=True,
                           next_num=_next_quote_number())

@app.route("/quotes/<quote_id>", methods=["GET"])
@role_required("quotes")
def quote_detail(quote_id):
    data = fb_get(f"/job_forms/{quote_id}")
    if not data:
        abort(404)
    data["firebase_id"] = quote_id

    # Linked project — via linked_project_id, or fall back to searching by source_quote
    linked_project = None
    lpid = data.get("linked_project_id")
    if lpid:
        linked_project = fb_get(f"/projects/{lpid}") or None
        if linked_project:
            linked_project["firebase_id"] = lpid
    if not linked_project:
        raw_proj = fb_get("/projects") or {}
        for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
            if isinstance(pdata, dict) and pdata.get("source_quote") == quote_id:
                pdata["firebase_id"] = pid
                linked_project = pdata
                break

    # Linked invoice
    linked_invoice = None
    liid = data.get("linked_invoice_id")
    if liid:
        linked_invoice = fb_get(f"/invoices/{liid}") or None
        if linked_invoice:
            linked_invoice["firebase_id"] = liid
    if not linked_invoice:
        raw_inv = fb_get("/invoices") or {}
        for iid, idata in (raw_inv.items() if isinstance(raw_inv, dict) else []):
            if isinstance(idata, dict) and idata.get("meta", {}).get("source_quote") == quote_id:
                idata["firebase_id"] = iid
                linked_invoice = idata
                break

    return render_template("quote_detail.html", quote=data,
                           linked_project=linked_project, linked_invoice=linked_invoice,
                           ai_enabled=bool(_get_ai_client()))

@app.route("/quotes/<quote_id>/followup-done", methods=["POST"])
@role_required("quotes")
def quote_followup_done(quote_id):
    fb_update(f"/job_forms/{quote_id}", {
        "follow_up_date": "",
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    flash("Follow-up marked as done.", "success")
    return redirect(url_for("quotes", tab="followups"))

@app.route("/quotes/<quote_id>/followup-snooze", methods=["POST"])
@role_required("quotes")
def quote_followup_snooze(quote_id):
    from datetime import timedelta
    days = int(request.form.get("days", 7))
    new_date = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    fb_update(f"/job_forms/{quote_id}", {
        "follow_up_date": new_date,
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    flash(f"Follow-up snoozed — rescheduled to {new_date}.", "success")
    return redirect(url_for("quotes", tab="followups"))

@app.route("/quotes/<quote_id>/duplicate", methods=["POST"])
@role_required("quotes")
def quote_duplicate(quote_id):
    original = fb_get(f"/job_forms/{quote_id}")
    if not original or not isinstance(original, dict):
        flash("Quote not found.", "danger")
        return redirect(url_for("quotes"))
    new_q = {k: v for k, v in original.items() if not k.startswith("firebase_")}
    new_q["job_number"]   = _next_quote_number()
    new_q["status"]       = "Not Started"
    new_q["date"]         = datetime.now().strftime("%Y-%m-%d")
    new_q["valid_until"]  = ""
    new_q["created_at"]   = datetime.now(timezone.utc).isoformat()
    new_q["updated_at"]   = datetime.now(timezone.utc).isoformat()
    new_q["created_by"]   = session.get("user_email", "")
    # Clear linked records — duplicate is a fresh quote
    new_q.pop("linked_project_id", None)
    new_q.pop("linked_invoice_id", None)
    key = fb_push("/job_forms", new_q)
    if key:
        flash(f"Quote duplicated as {new_q['job_number']}.", "success")
        return redirect(url_for("quote_detail", quote_id=key))
    flash("Failed to duplicate quote.", "danger")
    return redirect(url_for("quote_detail", quote_id=quote_id))

@app.route("/quotes/<quote_id>/edit", methods=["GET", "POST"])
@role_required("quotes")
def quote_edit(quote_id):
    data = fb_get(f"/job_forms/{quote_id}") or {}
    data["firebase_id"] = quote_id
    clients   = _load_clients()
    sales_ppl = _load_sales_people()
    if request.method == "POST":
        updated = _parse_quote_form(request.form)
        updated["updated_at"] = datetime.now(timezone.utc).isoformat()
        fb_update(f"/job_forms/{quote_id}", updated)
        flash("Quote updated.", "success")
        return redirect(url_for("quote_detail", quote_id=quote_id))
    return render_template("quote_form.html", quote=data, clients=clients,
                           sales_people=sales_ppl, is_new=False)

@app.route("/quotes/<quote_id>/delete", methods=["POST"])
@role_required("quotes")
def quote_delete(quote_id):
    fb_delete(f"/job_forms/{quote_id}")
    flash("Quote deleted.", "success")
    return redirect(url_for("quotes"))

@app.route("/quotes/<quote_id>/status", methods=["POST"])
@role_required("quotes")
def quote_status(quote_id):
    new_status = request.form.get("status", "Not Started")
    fb_update(f"/job_forms/{quote_id}", {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    flash(f"Status updated to {new_status}.", "success")
    return redirect(url_for("quotes"))

@app.route("/projects/<project_id>/quote", methods=["GET", "POST"])
@role_required("quotes")
def project_to_quote(project_id):
    project = fb_get(f"/projects/{project_id}") or {}
    project["firebase_id"] = project_id
    clients   = _load_clients()
    sales_ppl = _load_sales_people()
    if request.method == "POST":
        data = _parse_quote_form(request.form)
        data["source_project"]     = project_id
        data["source_project_num"] = project.get("project_number", "")
        data["created_at"] = datetime.now(timezone.utc).isoformat()
        data["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["created_by"] = session.get("user_email", "")
        key = fb_push("/job_forms", data)
        flash("Quote created from project.", "success")
        return redirect(url_for("quote_detail", quote_id=key))
    prefill = {
        "client_name":  project.get("client_name", ""),
        "project_name": project.get("project_name", ""),
        "description":  project.get("description", ""),
        "salesperson":  project.get("assigned_to", ""),
        "job_number":   "",
    }
    return render_template("quote_form.html", quote=prefill, clients=clients,
                           sales_people=sales_ppl, is_new=True,
                           next_num=_next_quote_number(),
                           from_project=project)

# ── Routes: Projects ──────────────────────────────────────────────────────────
@app.route("/projects")
@role_required("projects")
def projects():
    raw = fb_get("/projects") or {}
    raw_inv = fb_get("/invoices") or {}
    items = []
    _now_iso = datetime.now(timezone.utc).isoformat()
    for pid, pdata in (raw.items() if isinstance(raw, dict) else []):
        if pdata and isinstance(pdata, dict):
            pdata["firebase_id"] = pid
            pdata["_has_overdue"] = _project_has_overdue_stage(pdata.get("payment_stages"), raw_inv)
            # Repair status if amount_paid contradicts stored status
            _amt   = _safe_float(pdata.get("amount_paid", 0))
            _cv    = _safe_float(pdata.get("contract_value", 0))
            _st    = pdata.get("status") or "Not Started"
            _DONE_ST   = {"invoiced_Fully paid", "Cancelled"}
            _MANUAL_ST = {"Cancelled", "On Hold", "Ready to Sent",
                          "Sent out_Invoiced", "Sent out_Not Invoiced"}
            if _st not in _DONE_ST:
                if _cv > 0 and _amt >= _cv - 0.01:
                    # Fully paid — also migrates legacy "Completed" → "invoiced_Fully paid"
                    pdata["status"] = "invoiced_Fully paid"
                    fb_update(f"/projects/{pid}", {"status": "invoiced_Fully paid", "updated_at": _now_iso})
                elif _cv > 0 and 0 < _amt < _cv - 0.01 and _st not in (_MANUAL_ST | {"invoiced_Partially paid"}):
                    # Partial payment — upgrades from any status including "invoiced_Not paid yet"
                    pdata["status"] = "invoiced_Partially paid"
                    fb_update(f"/projects/{pid}", {"status": "invoiced_Partially paid", "updated_at": _now_iso})
                elif _amt == 0 and _st in ("Not Started", "In Progress", "Active"):
                    # Invoice created but $0 collected yet
                    _stages = pdata.get("payment_stages") or []
                    if any(isinstance(s, dict) and s.get("status") == "Invoiced" for s in _stages):
                        pdata["status"] = "invoiced_Not paid yet"
                        fb_update(f"/projects/{pid}", {"status": "invoiced_Not paid yet", "updated_at": _now_iso})
                elif _amt > 0 and _st == "Not Started":
                    pdata["status"] = "In Progress"
                    fb_update(f"/projects/{pid}", {"status": "In Progress", "updated_at": _now_iso})
            items.append(pdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    search        = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "")
    overdue_filter = request.args.get("overdue", "")
    date_from     = request.args.get("from", "")
    date_to       = request.args.get("to", "")
    client_filter = request.args.get("client", "")
    plant_filter  = request.args.get("plant", "").strip().upper()

    # Annotate change orders and normalise plant
    all_plants_set = set()
    for i in items:
        cos = i.get("change_orders") or []
        i["_has_co"] = isinstance(cos, list) and len(cos) > 0
        pl = (i.get("plant") or "").strip().upper()
        i["_plant_display"] = pl
        if pl:
            all_plants_set.add(pl)
    all_plants = sorted(all_plants_set)

    if search:
        items = [i for i in items if search in str(i).lower()]
    if status_filter:
        items = [i for i in items if i.get("status", "") == status_filter]
    if overdue_filter:
        items = [i for i in items if i.get("_has_overdue")]
    if client_filter:
        items = [i for i in items if i.get("client_name", "") == client_filter]
    if plant_filter:
        items = [i for i in items if i.get("_plant_display", "") == plant_filter]
    if date_from:
        items = [i for i in items if (i.get("start_date") or i.get("created_at","")[:10]) >= date_from]
    if date_to:
        items = [i for i in items if (i.get("start_date") or i.get("created_at","")[:10]) <= date_to]

    status_counts = {}
    for i in items:
        st = i.get("status") or "Not Started"
        status_counts[st] = status_counts.get(st, 0) + 1
    overdue_count = sum(1 for i in items if i.get("_has_overdue"))

    statuses = ["Sent out_Invoiced", "Sent out_Not Invoiced",
                "invoiced_Not paid yet", "invoiced_Partially paid", "invoiced_Fully paid"]
    clients = _load_clients()
    next_project_num = _next_project_number()
    active_tab = request.args.get("tab", "all-projects")

    # KPI stats from filtered projects
    _EXCLUDED_STATUSES = {"invoiced_Fully paid", "Cancelled"}
    p_total_count = len(items)
    p_total_cv    = sum(_safe_float(p.get("contract_value", 0)) for p in items)
    p_active_cv   = sum(_safe_float(p.get("contract_value", 0)) for p in items if p.get("status", "") not in _EXCLUDED_STATUSES)

    # Calculate collected amount from projects' amount_paid (already synced from invoices)
    # Using project.amount_paid avoids double-counting in multi-project invoices
    p_total_paid = sum(_safe_float(p.get("amount_paid", 0)) for p in items)

    p_outstanding = p_total_cv - p_total_paid

    return render_template("projects.html", projects=items, statuses=statuses,
                           search=search, status_filter=status_filter,
                           overdue_filter=overdue_filter, overdue_count=overdue_count,
                           date_from=date_from, date_to=date_to,
                           client_filter=client_filter, plant_filter=plant_filter,
                           all_plants=all_plants,
                           clients=clients, next_project_num=next_project_num,
                           active_tab=active_tab, status_counts=status_counts,
                           p_total_count=p_total_count, p_total_cv=p_total_cv,
                           p_active_cv=p_active_cv, p_total_paid=p_total_paid,
                           p_outstanding=p_outstanding)

@app.route("/projects/new", methods=["GET", "POST"])
@role_required("projects")
def project_new():
    clients = _load_clients()
    if request.method == "POST":
        data = _parse_project_form(request.form)

        # If arriving from "Convert to Project", resolve the source quote and
        # guard against converting the same quote into a second project.
        source_quote_id = request.form.get("source_quote_id", "").strip()
        source_quote = fb_get(f"/job_forms/{source_quote_id}") if source_quote_id else None
        if source_quote and source_quote.get("linked_project_id"):
            flash(f"This quote was already converted to Project "
                  f"{source_quote.get('linked_project_num','')}.", "info")
            return redirect(url_for("project_detail", project_id=source_quote["linked_project_id"]))

        # Check for custom stage amounts from frontend
        custom_stage_amounts = None
        custom_stage_amounts_json = request.form.get("custom_stage_amounts", "")
        if custom_stage_amounts_json:
            try:
                import json
                custom_stage_amounts = json.loads(custom_stage_amounts_json)
                total_amount = sum(_safe_float(s.get("amount", 0)) for s in custom_stage_amounts)
                contract_value = _safe_float(data.get("contract_value", 0))

                if abs(total_amount - contract_value) > 0.01:
                    flash(f"❌ Error: Payment plan total (${total_amount:.2f}) does not match contract value (${contract_value:.2f}). Please adjust amounts and try again.", "danger")
                    return redirect(url_for("project_new"))
            except (json.JSONDecodeError, ValueError):
                custom_stage_amounts = None

        # Validate PO/WO number is not duplicate
        po_wo_num = data.get("po_wo_number", "").strip()
        if po_wo_num:
            all_projects = fb_get("/projects") or {}
            if isinstance(all_projects, dict):
                for proj_data in all_projects.values():
                    if isinstance(proj_data, dict) and proj_data.get("po_wo_number", "").strip() == po_wo_num:
                        flash(f"PO/WO number {po_wo_num} already exists. Please use a different number.", "danger")
                        return redirect(url_for("project_new"))

        # Always generate project number server-side to prevent duplicates
        data["project_number"] = _next_project_number()
        down_pct = _safe_float(data.get("down_payment_percent", 0))
        mode, installments, custom_amounts = _resolve_installment_plan(data)
        data["down_payment_percent"]       = down_pct
        data["installment_count"]          = installments
        data["installment_mode"]           = mode
        data["custom_installment_amounts"] = custom_amounts or []

        # If custom amounts provided from frontend, use them directly
        if custom_stage_amounts:
            # Create payment stages from custom amounts
            payment_stages = []
            for idx, amount_data in enumerate(custom_stage_amounts):
                payment_stages.append({
                    "name": amount_data.get("name", f"Stage {idx+1}"),
                    "amount": _safe_float(amount_data.get("amount", 0)),
                    "status": "Pending Invoice",
                    "invoice_id": "",
                    "invoice_number": ""
                })
            data["payment_stages"] = payment_stages
        else:
            # Otherwise compute stages based on down payment and installments
            data["payment_stages"] = _compute_payment_stages(
                _safe_float(data["contract_value"]), down_pct, installments, custom_amounts=custom_amounts)
        data["created_at"] = datetime.now(timezone.utc).isoformat()
        data["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["created_by"] = session.get("user_email", "")

        if source_quote:
            data["source_quote_id"]   = source_quote_id
            data["source_quote"]      = source_quote_id
            data["source_quote_num"]  = source_quote.get("job_number", "")

        pid = fb_push("/projects", data)

        if source_quote:
            fb_update(f"/job_forms/{source_quote_id}", {
                "status":             "Converted",
                "linked_project_id":  pid,
                "linked_project_num": data["project_number"],
                "updated_at":         datetime.now(timezone.utc).isoformat(),
            })
            flash(f"Project {data['project_number']} created and linked to quote "
                  f"{source_quote.get('job_number','')}.", "success")
            return redirect(url_for("project_detail", project_id=pid))

        flash(f"Project {data['project_number']} created successfully.", "success")
        return redirect(url_for("projects", tab="all-projects"))
    sales_people = [p.get("name","") for p in _load_sales_people() if p.get("name","")]
    prefill_quote = request.args.get("from_quote", "")
    prefill_quote_id, prefill_quote_data = _find_quote_by_number(prefill_quote)
    if prefill_quote_data and prefill_quote_data.get("linked_project_id"):
        flash(f"Quote {prefill_quote} was already converted to Project "
              f"{prefill_quote_data.get('linked_project_num','')}.", "info")
        return redirect(url_for("project_detail", project_id=prefill_quote_data["linked_project_id"]))
    next_proj_num = _next_project_number()
    return render_template("project_form.html", project=None, clients=clients,
                           sales_people=sales_people, prefill_quote=prefill_quote,
                           prefill_quote_id=prefill_quote_id or "",
                           is_new=True, next_proj_num=next_proj_num)

@app.route("/projects/<project_id>", methods=["GET"])
@role_required("projects")
def project_detail(project_id):
    data = fb_get(f"/projects/{project_id}")
    if not data:
        abort(404)
    data["firebase_id"] = project_id
    proj_num = data.get("project_number", "")

    # Debug: Log what we loaded from Firebase
    stages = data.get("payment_stages", [])
    if isinstance(stages, list):
        for idx, stage in enumerate(stages):
            if isinstance(stage, dict):
                print(f"[LOAD_STAGE] Idx {idx}: status={stage.get('status')}, paid={stage.get('amount_paid')}, inv_id={stage.get('invoice_id')}, inv_num={stage.get('invoice_number')}", flush=True)

    # Older projects stored "payment_stages" as a flat list of stage-name
    # strings (no per-stage amount/status tracking). Only the structured
    # list-of-dicts format produced by _compute_payment_stages should drive
    # the Payment Plan card — otherwise hide it rather than erroring.
    raw_stages = data.get("payment_stages")
    if not (isinstance(raw_stages, list) and raw_stages and all(isinstance(s, dict) for s in raw_stages)):
        data["payment_stages"] = []
    else:
        # Normalize old "Not Invoiced" status to "Pending Invoice"
        for stage in data["payment_stages"]:
            if stage.get("status") == "Not Invoiced":
                stage["status"] = "Pending Invoice"

        # Recalculate payment stage statuses based on actual paid amounts from invoices
        # First, build a map of invoices by stage (supporting multi-project invoices)
        raw_inv = fb_get("/invoices") or {}
        stage_invoices = {}
        if isinstance(raw_inv, dict):
            for iid, inv in raw_inv.items():
                if not isinstance(inv, dict):
                    continue
                inv_meta = inv.get("meta", {}) or {}

                # Check if this project is linked to the invoice (supports multi-project)
                linked_projects = _invoice_linked_projects(inv)
                if proj_num not in linked_projects:
                    continue

                # For multi-project invoices, find the stage index from linked_projects
                stage_idx = -1
                lp_list = inv_meta.get("linked_projects") or []
                if isinstance(lp_list, list):
                    for lp in lp_list:
                        if isinstance(lp, dict) and lp.get("project_number") == proj_num:
                            stage_idx = lp.get("payment_stage_index", -1)
                            break

                # Fallback to old single-project format
                if stage_idx < 0:
                    stage_idx = inv_meta.get("payment_stage_index", -1)

                if stage_idx >= 0:
                    if stage_idx not in stage_invoices:
                        stage_invoices[stage_idx] = []
                    stage_invoices[stage_idx].append(inv)

        today_str = datetime.now().strftime("%Y-%m-%d")
        for idx, stage in enumerate(data["payment_stages"]):
            stage_amount = _safe_float(stage.get("amount", 0))

            # Calculate amount paid from all invoices for this stage
            amount_paid = 0
            due_date = ""
            invoice_id = ""
            invoice_number = ""
            if idx in stage_invoices:
                for inv in stage_invoices[idx]:
                    # Sum invoice payments using payment_log filtered by project_number (not share-based)
                    # Don't include tax payments in amount_paid - tax is tracked separately
                    inv_meta = inv.get("meta", {}) or {}
                    proj_payments = sum(_safe_float(p.get("amount", 0)) for p in (inv.get("payment_log", []) or []) if p.get("project_number") == proj_num)

                    # Fallback: if payment_log entries don't have project_number (legacy data),
                    # use the stage's amount_paid from allocation
                    if proj_payments == 0 and proj_num in _invoice_linked_projects(inv):
                        proj_payments = _safe_float(stage.get("amount_paid", 0))

                    amount_paid += proj_payments
                    due_date = due_date or inv_meta.get("due_date", "")
                    # Capture invoice details (use first invoice for this stage)
                    if not invoice_id:
                        invoice_id = inv.get("firebase_id", "")
                        invoice_number = inv_meta.get("invoice_number", "")

            is_overdue = bool(due_date) and due_date < today_str

            # Store calculated amounts and invoice info on stage for template display
            stage["amount_paid"] = amount_paid
            stage["invoice_id"] = invoice_id
            stage["invoice_number"] = invoice_number

            # Calculate status based on actual paid vs total
            if amount_paid >= (stage_amount - 0.01):
                stage["_display_status"] = "Paid"
            elif amount_paid > 0:
                stage["_display_status"] = "Overdue" if is_overdue else "Partially Paid"
            else:
                # No payment yet - keep the stage's real status (e.g. "Pending
                # Invoice" if no invoice has been generated yet, or "Invoiced"
                # if it has but remains unpaid) - unless that invoice is now
                # past its due date, in which case flag it as Overdue.
                stage_status = stage.get("status") or "Pending Invoice"
                if stage_status == "Invoiced" and is_overdue:
                    stage["_display_status"] = "Overdue"
                else:
                    stage["_display_status"] = stage_status

    # Load invoices linked to this project (directly, or via per-line-item
    # project overrides on invoices that span multiple projects)
    raw_inv = fb_get("/invoices") or {}
    project_invoices = []
    if isinstance(raw_inv, dict):
        for iid, idata in raw_inv.items():
            if isinstance(idata, dict) and proj_num in _invoice_linked_projects(idata):
                idata["firebase_id"] = iid
                idata["_project_share"] = _invoice_project_share(idata, proj_num)
                project_invoices.append(idata)
    project_invoices.sort(key=lambda x: x.get("meta", {}).get("invoice_date", ""), reverse=True)

    # Recalculate fresh statuses for display (based on actual payments)
    for invoice in project_invoices:
        # Calculate project-specific paid amount from payment_log (filtered by project_number)
        proj_payments = sum(_safe_float(p.get("amount", 0)) for p in (invoice.get("payment_log", []) or []) if p.get("project_number") == proj_num)

        # Fallback: if payment_log entries don't have project_number (legacy data),
        # use the invoice's total amount_paid and allocate by share
        if proj_payments == 0 and proj_num in _invoice_linked_projects(invoice):
            project_share = _invoice_project_share(invoice, proj_num)
            total_amount_paid = _safe_float(invoice.get("meta", {}).get("amount_paid", 0))
            proj_payments = total_amount_paid * project_share

        # Calculate tax paid - allocate proportionally if not per-project
        tax_payments = invoice.get("tax_payments", []) or []
        total_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments)

        # If tax_payments have project_number, filter by it; otherwise use share
        project_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments if tp.get("project_number") == proj_num)
        if project_tax_paid == 0 and total_tax_paid > 0:
            # Tax payments don't have project_number, allocate by share
            project_share = invoice.get("_project_share", 1.0)
            project_tax_paid = project_share * total_tax_paid

        invoice["_project_paid"] = proj_payments + project_tax_paid

        # Calculate status based on this project's paid vs this project's total
        meta = invoice.get("meta", {}) or {}
        project_share = invoice.get("_project_share", 1.0)
        invoice_subtotal = _safe_float(meta.get("subtotal", 0)) or (_safe_float(meta.get("total", 0)) - _safe_float(meta.get("tax_amount", 0)))
        invoice_tax = _safe_float(meta.get("tax_amount", 0))

        # This project's portion of the invoice
        project_invoice_due = (invoice_subtotal + invoice_tax) * project_share
        project_paid = invoice["_project_paid"]

        # Status based on this project's amounts
        if project_paid >= (project_invoice_due - 0.01):
            invoice["_display_status"] = "Paid"
        elif project_paid > 0:
            invoice["_display_status"] = "Partial"
        else:
            # Check if overdue
            due_date = meta.get("due_date", "")
            today = datetime.now().strftime("%Y-%m-%d")
            if due_date and due_date < today:
                invoice["_display_status"] = "Overdue"
            else:
                invoice["_display_status"] = "Sent"

    # Load expenses linked to this project
    raw_exp = fb_get("/balance_sheet_expenses") or {}
    project_expenses = []
    if isinstance(raw_exp, dict):
        for eid, edata in raw_exp.items():
            if isinstance(edata, dict) and edata.get("project_number", "") == proj_num:
                edata["firebase_id"] = eid
                project_expenses.append(edata)
    project_expenses.sort(key=lambda x: x.get("date", ""), reverse=True)

    # P&L totals — invoices spanning multiple projects only count their prorated share here
    # Include both invoice amount and tax in total invoiced (to match collected which includes tax payments)
    inv_total   = sum((_safe_float(i.get("meta",{}).get("total", 0)) or _safe_float(i.get("meta",{}).get("subtotal", 0)) + _safe_float(i.get("meta",{}).get("tax_amount", 0))) * i.get("_project_share", 1.0) for i in project_invoices)

    # For paid amount, use actual payment_log entries filtered by project_number (not share-based)
    inv_paid = 0
    for invoice in project_invoices:
        # Get actual payments from payment_log for this project
        proj_payments = sum(_safe_float(p.get("amount", 0)) for p in (invoice.get("payment_log", []) or []) if p.get("project_number") == proj_num)

        # Get tax payments - allocate proportionally to project's share if not per-project
        tax_payments = invoice.get("tax_payments", []) or []
        total_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments)

        # If tax_payments have project_number, filter by it; otherwise use share
        project_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments if tp.get("project_number") == proj_num)
        if project_tax_paid == 0 and total_tax_paid > 0:
            # Tax payments don't have project_number, allocate by share
            project_share = invoice.get("_project_share", 1.0)
            project_tax_paid = project_share * total_tax_paid

        inv_paid += proj_payments + project_tax_paid
    exp_total   = sum(_safe_float(e.get("amount", 0))                     for e in project_expenses)
    gross_profit = inv_paid - exp_total

    # Labor hours & cost logged against this project (Employees module)
    rate_by_uid = {u.get("firebase_uid"): _safe_float(u.get("hourly_rate", 0)) for u in _load_all_users()}
    labor_by_employee: Dict[str, dict] = {}
    labor_total_minutes = 0.0
    labor_total_cost = 0.0
    for e in _load_time_entries():
        if not proj_num or e.get("status") != "closed" or e.get("project_number") != proj_num:
            continue
        minutes = _safe_float(e.get("duration_minutes", 0))
        rate = rate_by_uid.get(e.get("employee_uid"), 0.0)
        bucket = labor_by_employee.setdefault(e.get("employee_name", "Unknown"), {"minutes": 0.0, "rate": rate, "cost": 0.0})
        bucket["minutes"] += minutes
        bucket["cost"]    += (minutes / 60.0) * rate
        labor_total_minutes += minutes
        labor_total_cost    += (minutes / 60.0) * rate
    # Allow manual actual_labor_cost override when no time entries are tracked
    manual_labor_cost = _safe_float(data.get("actual_labor_cost", 0))
    display_labor_cost = labor_total_cost if labor_total_cost > 0 else manual_labor_cost
    net_profit = gross_profit - display_labor_cost

    # Source quote that generated this project
    source_quote = None
    sq_id = data.get("source_quote")
    if sq_id:
        source_quote = fb_get(f"/job_forms/{sq_id}") or None
        if source_quote:
            source_quote["firebase_id"] = sq_id

    # Check if project has pending stages (using desktop workflow logic)
    all_invoices = fb_get("/invoices") or {}
    detection = _get_next_payment_stage(data, all_invoices)
    has_pending_stage = detection.get("stage_name") is not None
    next_stage_idx = detection.get("stage_idx")
    next_stage_name = detection.get("stage_name", "")
    next_stage_amount = detection.get("amount", 0)

    # Change orders
    change_orders = data.get("change_orders") or []
    if not isinstance(change_orders, list):
        change_orders = list(change_orders.values()) if isinstance(change_orders, dict) else []
    # Actual sum of approved CO amounts — used in the CO table and KPI card.
    # Include Approved, Invoiced, and Paid statuses (all active/completed COs)
    co_approved_total = sum(_safe_float(co.get("amount", 0)) for co in change_orders
                            if co.get("status") in ("Approved", "Invoiced", "Paid"))
    # Base contract value = Contract Value - Sum(Approved Change Orders)
    # Always recalculate to ensure consistency with current CO amounts
    base_contract = _safe_float(data.get("contract_value", 0)) - co_approved_total
    # Actual contract increase from base — used in Financial Summary bar so
    # Base + co_contract_increase always equals contract_value exactly.
    co_contract_increase = max(0.0, _safe_float(data.get("contract_value", 0)) - base_contract)

    return render_template("project_detail.html", project=data,
                           project_invoices=project_invoices,
                           project_expenses=project_expenses,
                           inv_total=inv_total, inv_paid=inv_paid,
                           exp_total=exp_total, gross_profit=gross_profit,
                           labor_by_employee=labor_by_employee,
                           labor_total_minutes=labor_total_minutes,
                           labor_total_cost=display_labor_cost,
                           manual_labor_cost=manual_labor_cost,
                           net_profit=net_profit,
                           source_quote=source_quote,
                           has_pending_stage=has_pending_stage,
                           next_stage_idx=next_stage_idx,
                           next_stage_name=next_stage_name,
                           next_stage_amount=next_stage_amount,
                           change_orders=change_orders,
                           co_approved_total=co_approved_total,
                           co_contract_increase=co_contract_increase,
                           base_contract=base_contract)

# ── Change Order Routes ───────────────────────────────────────────────────────

@app.route("/projects/<project_id>/change-orders/new", methods=["POST"])
@role_required("projects")
def co_new(project_id):
    project = fb_get(f"/projects/{project_id}") or {}
    if not project:
        abort(404)
    cos = project.get("change_orders") or []
    if not isinstance(cos, list):
        cos = list(cos.values()) if isinstance(cos, dict) else []
    co_num = f"CO-{len(cos)+1:03d}"
    now_str = datetime.now(timezone.utc).isoformat()
    new_co = {
        "co_number":    co_num,
        "title":        request.form.get("title", "").strip(),
        "description":  request.form.get("description", "").strip(),
        "amount":       _safe_float(request.form.get("amount", 0)),
        "status":       "Draft",
        "created_at":   now_str,
        "created_by":   session.get("user_email", ""),
        "submitted_at": "",
        "approved_at":  "",
        "notes":        request.form.get("notes", "").strip(),
    }
    cos.append(new_co)
    # Save base contract value on first CO creation
    if not project.get("base_contract_value"):
        fb_update(f"/projects/{project_id}", {"base_contract_value": project.get("contract_value", 0)})
    fb_update(f"/projects/{project_id}", {"change_orders": cos})
    flash(f"Change Order {co_num} created.", "success")
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-change-orders")

@app.route("/projects/<project_id>/change-orders/<int:co_idx>/status", methods=["POST"])
@role_required("projects")
def co_status(project_id, co_idx):
    project = fb_get(f"/projects/{project_id}") or {}
    if not project:
        abort(404)
    cos = project.get("change_orders") or []
    if not isinstance(cos, list):
        cos = list(cos.values()) if isinstance(cos, dict) else []
    if co_idx >= len(cos):
        abort(404)
    new_status = request.form.get("status", "")
    valid = {"Draft", "Submitted", "Approved", "Rejected"}
    if new_status not in valid:
        abort(400)
    now_str = datetime.now(timezone.utc).isoformat()
    cos[co_idx]["status"] = new_status
    if new_status == "Submitted":
        cos[co_idx]["submitted_at"] = now_str
    if new_status == "Approved":
        cos[co_idx]["approved_at"] = now_str
        # Increase contract value and add payment stage
        co_amount = _safe_float(cos[co_idx].get("amount", 0))
        co_title  = cos[co_idx].get("title", cos[co_idx].get("co_number", ""))
        old_value = _safe_float(project.get("contract_value", 0))
        new_value = old_value + co_amount
        stages = project.get("payment_stages") or []
        if not isinstance(stages, list):
            stages = []
        stages.append({
            "name":   f"{cos[co_idx]['co_number']} – {co_title}",
            "amount": co_amount,
            "amount_paid": "0",  # Track payments same as regular installments
            "status": "Pending Invoice",
            "co_number": cos[co_idx]['co_number'],  # Link back to change order
            "co_index": co_idx,  # Store CO index for reference
        })
        update_data = {
            "change_orders":  cos,
            "contract_value": new_value,
            "payment_stages": stages,
        }
        # Auto-update project status to "In Progress" if project was fully paid and new CO payment is unpaid
        if project.get("status") in ("Completed", "invoiced_Fully paid") and co_amount > 0:
            update_data["status"] = "In Progress"
            update_data["updated_at"] = now_str
        fb_update(f"/projects/{project_id}", update_data)
        flash(f"{cos[co_idx]['co_number']} approved — contract updated to ${new_value:,.0f} and new payment stage added.", "success")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-change-orders")
    fb_update(f"/projects/{project_id}", {"change_orders": cos})
    flash(f"{cos[co_idx]['co_number']} status updated to {new_status}.", "success")
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-change-orders")

@app.route("/projects/<project_id>/edit", methods=["GET", "POST"])
@role_required("projects")
def project_edit(project_id):
    data = fb_get(f"/projects/{project_id}") or {}
    data["firebase_id"] = project_id
    clients = _load_clients()

    # Older projects stored "payment_stages" as a flat list of stage-name
    # strings (no per-stage amount/status tracking) — normalize so the form's
    # saved-plan preview and lock check (which expect dicts) don't choke on them.
    raw_stages = data.get("payment_stages")
    if not (isinstance(raw_stages, list) and raw_stages and all(isinstance(s, dict) for s in raw_stages)):
        data["payment_stages"] = []
    else:
        # Normalize old "Not Invoiced" status to "Pending Invoice"
        for stage in data["payment_stages"]:
            if stage.get("status") == "Not Invoiced":
                stage["status"] = "Pending Invoice"
    if request.method == "POST":
        updated = _parse_project_form(request.form)
        down_pct = _safe_float(updated.get("down_payment_percent", 0))
        mode, installments, custom_amounts = _resolve_installment_plan(updated)
        updated["down_payment_percent"]       = down_pct
        updated["installment_count"]          = installments
        updated["installment_mode"]           = mode
        updated["custom_installment_amounts"] = custom_amounts or []

        # Handle updated stage amounts from contract value auto-adjustment
        updated_stage_amounts_json = request.form.get("updated_stage_amounts", "")
        amounts_updated = False
        if updated_stage_amounts_json:
            try:
                import json
                updated_stage_amounts = json.loads(updated_stage_amounts_json)
                # Update payment stages with updated amounts (preserving status and other fields)
                existing_stages = data.get("payment_stages") or []
                for i, amount_data in enumerate(updated_stage_amounts):
                    if i < len(existing_stages) and isinstance(existing_stages[i], dict):
                        existing_stages[i]["amount"] = _safe_float(amount_data.get("amount", 0))
                        # Preserve status and other fields
                        if "status" in amount_data:
                            existing_stages[i]["status"] = amount_data.get("status")

                updated["payment_stages"] = existing_stages
                amounts_updated = True
            except (json.JSONDecodeError, ValueError):
                flash("Error processing updated payment amounts.", "warning")

        # Check if down payment or installment count changed
        old_down_pct = _safe_float(data.get("down_payment_percent", 0))
        old_installments = _safe_float(data.get("installment_count", 1))
        payment_plan_changed = (down_pct != old_down_pct) or (installments != old_installments)

        # Handle custom stage amounts from frontend (for new customizations or preview stages)
        if not amounts_updated:
            custom_stage_amounts_json = request.form.get("custom_stage_amounts", "")
            if custom_stage_amounts_json and custom_stage_amounts_json != "[]":
                try:
                    import json
                    custom_stage_amounts = json.loads(custom_stage_amounts_json)
                    if custom_stage_amounts:  # Only process if list is not empty
                        # If custom amounts has name and amount fields (from preview),
                        # ensure they include status and invoice_id fields
                        if custom_stage_amounts and isinstance(custom_stage_amounts[0], dict) and "name" in custom_stage_amounts[0]:
                            # Ensure all required fields are present
                            enriched_stages = []
                            for stage in custom_stage_amounts:
                                enriched = {
                                    "name": stage.get("name", ""),
                                    "amount": _safe_float(stage.get("amount", 0)),
                                    "status": stage.get("status", "Pending Invoice"),
                                    "invoice_id": stage.get("invoice_id", "")
                                }
                                enriched_stages.append(enriched)
                            updated["payment_stages"] = enriched_stages
                        else:
                            # If payment plan changed, recalculate stages first, then apply custom amounts
                            existing_stages = data.get("payment_stages") or []
                            if payment_plan_changed:
                                # Recalculate stages based on new installment count
                                new_stages = _compute_payment_stages(
                                    _safe_float(updated["contract_value"]), down_pct, installments, custom_amounts=custom_amounts)
                                # Then apply custom amounts to the new stages
                                for i, amount_data in enumerate(custom_stage_amounts):
                                    if i < len(new_stages) and isinstance(new_stages[i], dict):
                                        new_stages[i]["amount"] = _safe_float(amount_data.get("amount", 0))
                                updated["payment_stages"] = new_stages
                            else:
                                # Payment plan didn't change, just update existing stage amounts
                                for i, amount_data in enumerate(custom_stage_amounts):
                                    if i < len(existing_stages) and isinstance(existing_stages[i], dict):
                                        existing_stages[i]["amount"] = _safe_float(amount_data.get("amount", 0))
                                updated["payment_stages"] = existing_stages
                        amounts_updated = True
                except (json.JSONDecodeError, ValueError):
                    flash("Error processing custom payment amounts. Using default distribution.", "warning")

        # If amounts were not customized, use standard distribution logic
        if not amounts_updated:
            existing_stages = data.get("payment_stages") or []
            plan_in_progress = any(s.get("status") != "Pending Invoice" for s in existing_stages if isinstance(s, dict))

            if plan_in_progress:
                # Stages already have invoices/payments against them — keep the plan intact
                flash("Payment plan kept as-is because one or more stages are already invoiced.", "info")
            else:
                # Always recalculate if payment plan changed or if no existing stages
                updated["payment_stages"] = _compute_payment_stages(
                    _safe_float(updated["contract_value"]), down_pct, installments, custom_amounts=custom_amounts)

        # Sync payment stages with change order amounts
        # If a payment stage is linked to a CO (stage name contains "CO-"), update that CO's amount
        updated_stages = updated.get("payment_stages") or []
        change_orders_to_sync = data.get("change_orders") or []
        if not isinstance(change_orders_to_sync, list):
            change_orders_to_sync = list(change_orders_to_sync.values()) if isinstance(change_orders_to_sync, dict) else []

        for stage in updated_stages:
            if isinstance(stage, dict):
                stage_name = stage.get("name", "")
                if "CO-" in stage_name:
                    # Extract CO number from stage name
                    co_num = stage_name.split(" ")[0] if " " in stage_name else stage_name
                    stage_amount = _safe_float(stage.get("amount", 0))
                    # Find and update matching change order
                    for co in change_orders_to_sync:
                        if isinstance(co, dict) and co.get("co_number", "") == co_num:
                            co["amount"] = str(stage_amount)
                            break

        if change_orders_to_sync:
            updated["change_orders"] = change_orders_to_sync

        updated["updated_at"] = datetime.now(timezone.utc).isoformat()

        # Keep base_contract_value in sync with the edited contract_value and change orders
        # Formula: base_contract_value = contract_value - sum(approved change order amounts)
        # This ensures Financial Summary shows correct Base + COs = Total breakdown
        # Use the synced change_orders (which now have updated amounts from payment stages)
        cos_for_calculation = change_orders_to_sync if change_orders_to_sync else (data.get("change_orders") or [])
        if not isinstance(cos_for_calculation, list):
            cos_for_calculation = list(cos_for_calculation.values()) if isinstance(cos_for_calculation, dict) else []
        # Sum ALL approved change orders (status = "Approved" or any status indicating approved/active)
        co_approved_sum = sum(_safe_float(co.get("amount", 0)) for co in cos_for_calculation
                             if co.get("status") in ("Approved", "Invoiced", "Paid"))
        # Always update base_contract_value to maintain correct breakdown
        updated["base_contract_value"] = _safe_float(updated.get("contract_value", 0)) - co_approved_sum

        # Guard: if the project already has payments, don't let a form submission silently
        # downgrade its status to "Not Started". Preserve any manually-set status above "Not Started"
        # for paid/partially-paid projects; _sync_project_payment below will then correct it upward.
        existing_status = data.get("status", "Not Started")
        existing_paid   = _safe_float(data.get("amount_paid", 0))
        new_status       = updated.get("status", "Not Started")
        if existing_paid > 0 and new_status == "Not Started" and existing_status not in ("Not Started", "Cancelled"):
            updated["status"] = existing_status

        fb_update(f"/projects/{project_id}", updated)

        # Re-sync payment-derived status so that a fully-paid project always shows Completed.
        proj_num = updated.get("project_number", data.get("project_number", ""))
        if proj_num:
            _sync_project_payment(proj_num)
            _auto_complete_project_if_paid(proj_num)

        flash("Project updated successfully.", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    sales_people = [p.get("name","") for p in _load_sales_people() if p.get("name","")]
    return render_template("project_form.html", project=data, clients=clients,
                           sales_people=sales_people, prefill_quote="", is_new=False)

@app.route("/projects/<project_id>/status", methods=["POST"])
@role_required("projects")
def project_status(project_id):
    new_status = request.form.get("status", "In Progress")
    fb_update(f"/projects/{project_id}", {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    flash(f"Status updated to {new_status}.", "success")
    return redirect(url_for("project_detail", project_id=project_id))

@app.route("/projects/<project_id>/stage/<int:stage_idx>/invoice", methods=["GET"])
@role_required("projects")
def project_stage_invoice(project_id, stage_idx):
    """Kept for backwards compatibility — redirects to the old form flow."""
    project = fb_get(f"/projects/{project_id}") or {}
    stages = project.get("payment_stages") or []
    if not (0 <= stage_idx < len(stages)) or not isinstance(stages[stage_idx], dict):
        abort(404)
    stage = stages[stage_idx]
    return redirect(url_for("invoice_new",
                            project=project.get("project_number", ""),
                            client=project.get("client_name", ""),
                            stage_idx=stage_idx,
                            stage_name=stage.get("name", ""),
                            stage_amount=stage.get("amount", 0)))

def _create_stage_invoice(project_id: str, stage_idx: int, mark_paid: bool = False):
    """Create an invoice for a payment stage instantly — no form needed.
    Returns (invoice_id, error_message). If mark_paid=True, also marks it Paid
    and syncs the project payment totals.
    """
    project = fb_get(f"/projects/{project_id}") or {}
    stages  = project.get("payment_stages") or []
    if not (0 <= stage_idx < len(stages)) or not isinstance(stages[stage_idx], dict):
        return None, "Stage not found."
    first_pending = next((i for i, s in enumerate(stages) if s.get("status") == "Pending Invoice"), None)
    if first_pending is None or stage_idx != first_pending:
        return None, "That stage isn't ready yet — complete earlier stages first."

    stage      = stages[stage_idx]
    amount     = _safe_float(stage.get("amount", 0))
    proj_num   = project.get("project_number", "")
    client     = project.get("client_name", "")
    now_str    = datetime.now(timezone.utc).isoformat()
    today      = datetime.now().strftime("%Y-%m-%d")
    due_date   = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
    inv_num    = _next_invoice_number()
    inv_status = "Paid" if mark_paid else "Draft"
    amt_paid   = str(amount) if mark_paid else "0"

    invoice_data = {
        "meta": {
            "invoice_number":      inv_num,
            "invoice_date":        today,
            "due_date":            due_date,
            "client_name":         client,
            "project_number":      proj_num,
            "status":              inv_status,
            "subtotal":            str(amount),
            "tax_rate":            "0",
            "tax_amount":          "0",
            "total":               str(amount),
            "amount_paid":         amt_paid,
            "notes":               "",
            "terms":               "",
            "payment_method":      "",
            "payment_stage_index": stage_idx,
            "payment_stage":       stage.get("name", ""),
            "linked_projects":     [{"project_number": proj_num, "payment_stage_index": stage_idx}],
            "created_at":          now_str,
            "updated_at":          now_str,
            "created_by":          session.get("user_email", ""),
        },
        "line_items": [{
            "description": stage.get("name", f"Payment Stage {stage_idx + 1}"),
            "quantity":    "1",
            "unit_price":  str(amount),
            "amount":      str(amount),
            "project_number": proj_num,
        }],
    }
    iid = fb_push("/invoices", invoice_data)
    stage_status = "Paid" if mark_paid else "Invoiced"
    _mark_project_stage(proj_num, stage_idx, stage_status, invoice_id=iid, amount=_safe_float(amount))
    if project.get("status", "Not Started") == "Not Started":
        fb_update(f"/projects/{project_id}", {"status": "In Progress"})
    if mark_paid:
        _sync_project_payment(proj_num)
        _auto_complete_project_if_paid(proj_num)
        _upsert_revenue_entry(iid, invoice_data["meta"])
    return iid, None

@app.route("/projects/<project_id>/stage/<int:stage_idx>/create-invoice", methods=["POST"])
@role_required("projects")
def project_stage_create_invoice(project_id, stage_idx):
    """One-click: create invoice for a stage instantly and go straight to it."""
    iid, err = _create_stage_invoice(project_id, stage_idx, mark_paid=False)
    if err:
        flash(err, "warning")
        return redirect(url_for("project_detail", project_id=project_id))
    flash("Invoice created.", "success")
    return redirect(url_for("invoice_detail", invoice_id=iid))

@app.route("/projects/<project_id>/stage/<int:stage_idx>/mark-paid", methods=["POST"])
@role_required("projects")
def project_stage_mark_paid(project_id, stage_idx):
    """One-click: create invoice + mark Paid immediately."""
    iid, err = _create_stage_invoice(project_id, stage_idx, mark_paid=True)
    if err:
        flash(err, "warning")
        return redirect(url_for("project_detail", project_id=project_id))
    flash("Stage marked as paid and invoice created.", "success")
    return redirect(url_for("project_detail", project_id=project_id))

@app.route("/projects/<project_id>/stage/<int:stage_idx>/set-amount", methods=["POST"])
@role_required("projects")
def project_stage_set_amount(project_id, stage_idx):
    data = fb_get(f"/projects/{project_id}")
    if not data:
        abort(404)
    stages = data.get("payment_stages") or []
    if not (0 <= stage_idx < len(stages)) or not isinstance(stages[stage_idx], dict):
        flash("Stage not found.", "danger")
        return redirect(url_for("project_detail", project_id=project_id))
    if stages[stage_idx].get("status") != "Pending":
        flash("Cannot edit amount on an already-invoiced stage.", "warning")
        return redirect(url_for("project_detail", project_id=project_id))
    try:
        new_amount = round(float(str(request.form.get("amount", "0")).replace(",", "")), 2)
    except (ValueError, TypeError):
        flash("Invalid amount.", "danger")
        return redirect(url_for("project_detail", project_id=project_id))
    stages[stage_idx]["amount"] = new_amount

    # Auto-balance: spread remaining balance equally across all OTHER pending stages
    contract_value = _safe_float(data.get("contract_value", 0))
    locked_sum = sum(
        _safe_float(s.get("amount", 0))
        for i, s in enumerate(stages)
        if i != stage_idx and isinstance(s, dict) and s.get("status") != "Pending"
    )
    other_pending = [
        i for i, s in enumerate(stages)
        if i != stage_idx and isinstance(s, dict) and s.get("status") == "Pending Invoice"
    ]
    if other_pending and contract_value > 0:
        remaining = round(contract_value - locked_sum - new_amount, 2)
        per = round(remaining / len(other_pending), 2)
        allocated = 0
        for j, idx in enumerate(other_pending):
            if j < len(other_pending) - 1:
                stages[idx]["amount"] = per
                allocated += per
            else:
                stages[idx]["amount"] = round(remaining - allocated, 2)

    fb_update(f"/projects/{project_id}", {
        "payment_stages": stages,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    flash(f"Stage updated to ${new_amount:,.2f} — remaining stages adjusted to balance.", "success")
    return redirect(url_for("project_detail", project_id=project_id))

@app.route("/projects/<project_id>/delete", methods=["POST"])
@role_required("projects")
def project_delete(project_id):
    project = fb_get(f"/projects/{project_id}") or {}
    source_quote_id = project.get("source_quote_id")

    fb_delete(f"/projects/{project_id}")

    # If project was created from a quote, reset quote status to "Approved"
    if source_quote_id:
        fb_update(f"/job_forms/{source_quote_id}", {
            "status": "Approved",
            "linked_project_id": None,
            "linked_project_num": None,
            "updated_at": datetime.now(timezone.utc).isoformat()
        })

    flash("Project deleted.", "success")
    return redirect(url_for("projects"))

@app.route("/projects/<project_id>/notes/add", methods=["POST"])
@role_required("projects")
def project_note_add(project_id):
    text = request.form.get("note_text", "").strip()
    if not text:
        flash("Note cannot be empty.", "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-notes")
    project = fb_get(f"/projects/{project_id}") or {}
    log = project.get("activity_log") or []
    if not isinstance(log, list):
        log = []
    log.append({
        "text":       text,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": session.get("user_name") or session.get("user_email", ""),
    })
    fb_update(f"/projects/{project_id}", {
        "activity_log": log,
        "updated_at":   datetime.now(timezone.utc).isoformat(),
    })
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-notes")

@app.route("/projects/<project_id>/notes/<int:idx>/delete", methods=["POST"])
@role_required("projects")
def project_note_delete(project_id, idx):
    project = fb_get(f"/projects/{project_id}") or {}
    log = project.get("activity_log") or []
    if isinstance(log, list) and 0 <= idx < len(log):
        log.pop(idx)
        fb_update(f"/projects/{project_id}", {
            "activity_log": log,
            "updated_at":   datetime.now(timezone.utc).isoformat(),
        })
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-notes")

# ── Routes: Projects Export ───────────────────────────────────────────────────
def _filter_projects_export(items):
    if request.args.get("status"):
        items = [i for i in items if i.get("status","") == request.args["status"]]
    if request.args.get("client"):
        items = [i for i in items if i.get("client_name","") == request.args["client"]]
    date_from = request.args.get("from","")
    date_to   = request.args.get("to","")
    if date_from:
        items = [i for i in items if (i.get("start_date") or i.get("created_at","")[:10]) >= date_from]
    if date_to:
        items = [i for i in items if (i.get("start_date") or i.get("created_at","")[:10]) <= date_to]
    return items

@app.route("/projects/export/csv")
@role_required("projects")
def projects_export_csv():
    import csv
    import io
    raw = fb_get("/projects") or {}
    items = []
    for pid, pdata in (raw.items() if isinstance(raw, dict) else []):
        if pdata and isinstance(pdata, dict):
            items.append(pdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    items = _filter_projects_export(items)
    output = io.StringIO()
    w = csv.writer(output)
    co = company_info()

    def fmt_csv_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    # Add company header and blank row
    w.writerow([f"{co.get('name','')} - Projects Report"])
    w.writerow([])

    headers = ["Project Number", "Project Name", "Client", "Start Date", "End Date", "Contract Value", "Amount Paid", "Outstanding", "Status"]
    w.writerow(headers)

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for p in items:
        date_str = p.get("start_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(p)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda p: p.get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for p in grouped[year][month][day]:
                    cv = _safe_float(p.get("contract_value", 0))
                    paid = _safe_float(p.get("amount_paid", 0))
                    row = [
                        p.get("project_number",""),
                        p.get("project_name",""),
                        p.get("client_name",""),
                        fmt_csv_date(p.get("start_date","")),
                        fmt_csv_date(p.get("end_date","")),
                        f"{cv:.2f}",
                        f"{paid:.2f}",
                        f"{cv-paid:.2f}",
                        p.get("status","")
                    ]
                    w.writerow(row)

    output.seek(0)
    from flask import Response
    fname = f"projects_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/projects/export/excel")
@role_required("projects")
def projects_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io
    raw = fb_get("/projects") or {}
    items = []
    for pid, pdata in (raw.items() if isinstance(raw, dict) else []):
        if pdata and isinstance(pdata, dict):
            items.append(pdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    items = _filter_projects_export(items)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    hdr_font = Font(color="FFFFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13, color="FF0F766E")
    alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add title row
    co = company_info()
    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value=f"{co.get('name','')} - Projects Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # Add headers
    headers = ["Project Number","Project Name","Client","Start Date","End Date",
               "Contract Value ($)","Amount Paid ($)","Outstanding ($)","Status"]
    header_row = 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr

    def fmt_proj_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for p in items:
        date_str = p.get("start_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(p)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda p: p.get("created_at", ""))

    ri = header_row + 1

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for p in grouped[year][month][day]:
                    cv   = _safe_float(p.get("contract_value", 0))
                    paid = _safe_float(p.get("amount_paid", 0))
                    row = [p.get("project_number",""), p.get("project_name",""),
                           p.get("client_name",""), fmt_proj_date(p.get("start_date","")), fmt_proj_date(p.get("end_date","")),
                           cv, paid, cv - paid, p.get("status","")]
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if ri % 2 == 0:
                            cell.fill = alt_fill
                        if ci in (6, 7, 8):
                            cell.number_format = '"$"#,##0.00'
                        cell.alignment = ctr
                    ri += 1

    # Increase column widths
    col_widths = [22, 28, 26, 16, 16, 18, 18, 18, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    fname = f"projects_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/projects/export/pdf")
@role_required("projects")
def projects_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("projects", tab="export"))
    import io as _io
    raw = fb_get("/projects") or {}
    items = []
    for pid, pdata in (raw.items() if isinstance(raw, dict) else []):
        if pdata and isinstance(pdata, dict):
            items.append(pdata)
    items.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    items = _filter_projects_export(items)
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=0.9*inch, rightMargin=0.9*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    co = company_info()
    elems = []
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=15,
                              fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                              alignment=1)  # CENTER
    elems.append(Paragraph(f"{co.get('name','')} — Projects Report", title_s))
    elems.append(Spacer(1, 0.2*inch))
    hdrs = ["Project Number", "Project Name", "Client", "Start Date", "End Date", "Contract Value", "Paid", "Outstanding", "Status"]
    data = [hdrs]
    def fmt_date_pdf(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=10, wordWrap='CJK')
    group_style = ParagraphStyle("group", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", alignment=1, leading=10, textColor=colors.HexColor("#0F172A"))

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for p in items:
        date_str = p.get("start_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(p)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda p: p.get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for p in grouped[year][month][day]:
                    cv   = _safe_float(p.get("contract_value", 0))
                    paid = _safe_float(p.get("amount_paid", 0))
                    data.append([
                        Paragraph(p.get("project_number","—"), cell_style),
                        Paragraph(p.get("project_name","—") or "—", cell_style),
                        Paragraph(p.get("client_name","—") or "—", cell_style),
                        Paragraph(fmt_date_pdf(p.get("start_date","")), cell_style),
                        Paragraph(fmt_date_pdf(p.get("end_date","")), cell_style),
                        Paragraph(f"${cv:,.0f}", cell_style),
                        Paragraph(f"${paid:,.0f}", cell_style),
                        Paragraph(f"${cv-paid:,.0f}", cell_style),
                        Paragraph(p.get("status","—"), cell_style),
                    ])
    cw = [1.5*inch, 2.0*inch, 1.6*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    from flask import Response
    fname = f"projects_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

# ── Routes: Invoicing ─────────────────────────────────────────────────────────
@app.route("/invoicing")
@role_required("invoicing")
def invoicing():
    _auto_flag_overdue()
    raw = fb_get("/invoices") or {}

    # Build project_number → plant lookup from projects
    raw_proj = fb_get("/projects") or {}
    proj_plant_map = {}
    if isinstance(raw_proj, dict):
        for pdata in raw_proj.values():
            if isinstance(pdata, dict):
                pnum = pdata.get("project_number", "")
                plt  = (pdata.get("plant") or "").strip().upper()
                if pnum and plt:
                    proj_plant_map[pnum] = plt

    items = []
    for iid, idata in (raw.items() if isinstance(raw, dict) else []):
        if idata and isinstance(idata, dict):
            idata["firebase_id"] = iid
            # Attach plant from the linked project
            meta = idata.get("meta", {}) or {}
            proj_num = meta.get("project_number", "")
            if not proj_num:
                # Try to get from first line item
                for li in (idata.get("line_items") or []):
                    if isinstance(li, dict) and li.get("project_number"):
                        proj_num = li["project_number"]
                        break
            idata["plant_state"] = proj_plant_map.get(proj_num, "")
            items.append(idata)
    items.sort(key=lambda x: x.get("meta", {}).get("created_at", ""), reverse=True)
    all_invoices_raw = list(items)

    # Calculate status for all invoices BEFORE filtering (so filter uses calculated status, not stored status)
    today_str = datetime.now().strftime("%Y-%m-%d")
    for inv in items:
        m = inv.get("meta", {})
        calculated_status = _calculate_invoice_status(inv)
        m["status"] = calculated_status
        due = m.get("due_date", "") or ""
        if calculated_status in ("Sent", "Viewed", "Partial") and due and due < today_str:
            m["status"] = "Overdue"

    search        = request.args.get("q", "").strip().lower()
    status_filter = request.args.get("status", "")
    date_from     = request.args.get("from", "")
    date_to       = request.args.get("to", "")
    client_filter = request.args.get("client", "")
    plant_filter  = request.args.get("plant", "").strip().upper()

    # Recalculate status for every invoice BEFORE filtering so status_filter works correctly
    today_str = datetime.now().strftime("%Y-%m-%d")
    for inv in items:
        m = inv.get("meta", {})
        calculated_status = _calculate_invoice_status(inv)
        due = m.get("due_date", "") or ""
        if calculated_status in ("Sent", "Viewed", "Partial") and due and due < today_str:
            calculated_status = "Overdue"
        m["status"] = calculated_status

    # IMPORTANT: Create items_for_collected from ALL invoices first
    # Then apply ONLY status/client/plant/search filters (NO invoice date filter)
    # This allows invoices from ANY year to contribute payments to collected amount
    items_for_collected = []
    for inv in all_invoices_raw:
        if not inv or not isinstance(inv, dict):
            continue
        # Apply only: status, client, plant, search filters
        if search and search not in str(inv).lower():
            continue
        if status_filter and inv.get("meta", {}).get("status", "") != status_filter:
            continue
        if client_filter and inv.get("meta", {}).get("client_name", "") != client_filter:
            continue
        if plant_filter and inv.get("plant_state", "") != plant_filter:
            continue
        # Add to collected list (NO invoice date filter!)
        items_for_collected.append(inv)

    # Now apply all filters to items for display (including search and invoice date)
    if search:
        items = [i for i in items if search in str(i).lower()]
    if status_filter:
        items = [i for i in items if i.get("meta", {}).get("status", "") == status_filter]
    if client_filter:
        items = [i for i in items if i.get("meta", {}).get("client_name", "") == client_filter]
    if plant_filter:
        items = [i for i in items if i.get("plant_state", "") == plant_filter]

    if date_from:
        items = [i for i in items if (i.get("meta", {}).get("invoice_date") or "") >= date_from]
    if date_to:
        items = [i for i in items if (i.get("meta", {}).get("invoice_date") or "") <= date_to]

    # Build filter dropdown lists
    inv_clients = _load_clients()
    all_plants = sorted({i.get("plant_state", "") for i in all_invoices_raw if i.get("plant_state", "")})

    statuses = ["Draft", "Sent", "Viewed", "Paid", "Partial", "Overdue", "Cancelled"]
    active_tab = request.args.get("tab", "all-invoices")

    # KPI stats — invoices filtered by invoice_date, collected by payment_date
    _kpi_rows = []

    for inv in items:
        m  = inv.get("meta", {}) or {}
        st = m.get("status", "Draft")
        total_val = _safe_float(m.get("total", 0))
        # Calculate total paid: sum of all payments (not filtered by date)
        # For Collected card, we'll separately calculate payments within date range
        total_all_paid = sum(_safe_float(p.get("amount", 0)) for p in (inv.get("payment_log", []) or []))
        total_all_paid += sum(_safe_float(p.get("amount", 0)) for p in (inv.get("tax_payments", []) or []))
        _kpi_rows.append((st, total_val, total_all_paid))

    i_total       = len(_kpi_rows)
    i_draft_count = sum(1 for st, _, __ in _kpi_rows if st == "Draft")
    i_sent_count  = sum(1 for st, _, __ in _kpi_rows if st in ("Sent", "Viewed"))
    i_paid_count  = sum(1 for st, _, __ in _kpi_rows if st == "Paid")
    i_over_count  = sum(1 for st, _, __ in _kpi_rows if st == "Overdue")
    i_total_val   = sum(total for _, total, __ in _kpi_rows)
    i_total_paid  = sum(paid for _, __, paid in _kpi_rows)
    i_outstanding = i_total_val - i_total_paid
    i_overdue_amt = sum(total for st, total, __ in _kpi_rows if st == "Overdue")

    # Collected amount: ALL payments from filtered invoices
    # Based ONLY on payment received date from payment history
    # RESPECTS: status, client, plant, search filters
    # IGNORES: invoice creation date
    # Shows ALL payments (past, present, future) from filtered invoices
    # If date range is provided, filters payments by that range
    i_total_paid_in_range = 0.0

    def _normalize_date(d):
        """Convert various date formats to YYYY-MM-DD for comparison"""
        if not d or not isinstance(d, str):
            return ""
        d = d.strip()
        # Already in YYYY-MM-DD format (e.g., "2027-03-17")
        if d.count('-') == 2 and len(d) == 10:
            if d[0:4].isdigit() and d[4] == '-' and d[7] == '-':
                return d
        # MM-DD-YYYY format (e.g., "03-17-2027")
        if len(d) == 10 and d[2] == '-' and d[5] == '-':
            parts = d.split('-')
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                return f"{parts[2]}-{parts[0]}-{parts[1]}"
        # Try DD-MM-YYYY format (e.g., "17-03-2027")
        if len(d) == 10 and d[2] == '-' and d[5] == '-':
            parts = d.split('-')
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                day = int(parts[0]) if int(parts[0]) > 12 else int(parts[1])
                month = int(parts[1]) if int(parts[1]) <= 12 else int(parts[0])
                year = parts[2]
                return f"{year}-{month:02d}-{day:02d}"
        return d

    def _pay_in_date_range(d):
        """Check if payment date falls within the selected date range"""
        # If no date filters, include all payments (past, present, future)
        if not date_from and not date_to:
            return True
        norm_d = _normalize_date(d)
        from_ok = (not date_from or norm_d >= date_from)
        to_ok = (not date_to or norm_d <= date_to)
        return from_ok and to_ok

    # Iterate through filtered invoices (respects status/client/plant/search filters)
    # Count ALL payments from those invoices (past, present, future)
    # If date range is set, only count payments within that range
    for inv in items_for_collected:
        if not inv or not isinstance(inv, dict):
            continue
        # Count all payments in payment_log based on payment date
        for p in (inv.get("payment_log", []) or []):
            if not p or not isinstance(p, dict):
                continue
            pay_date = p.get("date", "")
            pay_amt = p.get("amount", 0)
            if _pay_in_date_range(pay_date):
                i_total_paid_in_range += _safe_float(pay_amt)
        # Count all tax payments based on payment date
        for tp in (inv.get("tax_payments", []) or []):
            if not tp or not isinstance(tp, dict):
                continue
            tax_date = tp.get("date", "")
            tax_amt = tp.get("amount", 0)
            if _pay_in_date_range(tax_date):
                i_total_paid_in_range += _safe_float(tax_amt)

    # Collection rate based on payments in date range vs total invoice amount of filtered invoices
    i_coll_rate   = round(i_total_paid_in_range / i_total_val * 100) if i_total_val else 0

    # Ensure all invoices have amount_paid and tax_paid in meta for template compatibility
    for inv in items:
        if "meta" not in inv:
            inv["meta"] = {}
        if "amount_paid" not in inv["meta"]:
            # Calculate from payment_log if missing
            payment_log = inv.get("payment_log", []) or []
            total_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log)
            inv["meta"]["amount_paid"] = str(total_paid) if total_paid > 0 else "0"
        if "tax_paid" not in inv["meta"]:
            # Calculate from tax_payments if missing
            tax_log = inv.get("tax_payments", []) or []
            tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
            inv["meta"]["tax_paid"] = str(tax_paid) if tax_paid > 0 else "0"

    settings = load_settings()
    default_tax_rate = settings.get("company", {}).get("default_tax_rate", 0)

    return render_template("invoicing.html", invoices=items, statuses=statuses,
                           search=search, status_filter=status_filter,
                           date_from=date_from, date_to=date_to,
                           client_filter=client_filter, inv_clients=inv_clients,
                           plant_filter=plant_filter, inv_plants=all_plants,
                           active_tab=active_tab,
                           i_total=i_total, i_draft_count=i_draft_count,
                           i_sent_count=i_sent_count, i_paid_count=i_paid_count,
                           i_over_count=i_over_count, i_total_val=i_total_val,
                           i_total_paid=i_total_paid, i_total_paid_in_range=i_total_paid_in_range,
                           i_outstanding=i_outstanding,
                           i_coll_rate=i_coll_rate, i_overdue_amt=i_overdue_amt,
                           default_tax_rate=default_tax_rate)

@app.route("/api/projects/<project_ids>", methods=["GET"])
@role_required("projects")
def api_get_projects(project_ids):
    """API endpoint to get project data with payment stage detection (desktop workflow logic)"""
    ids = [pid.strip() for pid in project_ids.split(",") if pid.strip()]
    all_projects = fb_get("/projects") or {}
    all_invoices = fb_get("/invoices") or {}

    projects = []
    for proj_id in ids:
        if proj_id in all_projects:
            proj = all_projects[proj_id]
            if isinstance(proj, dict):
                proj["firebase_id"] = proj_id

                # Use new detection logic (works with payment_stages, applies desktop workflow)
                try:
                    detection = _get_next_payment_stage(proj, all_invoices)
                    if detection.get("stage_name"):
                        proj["next_stage"] = detection.get("stage_name")
                        proj["next_stage_amount"] = detection.get("amount", 0)
                    else:
                        proj["next_stage"] = "Fully Invoiced"
                        proj["next_stage_amount"] = 0
                    proj["stage_blocked"] = detection.get("blocked", False)
                    proj["stage_reason"] = detection.get("reason", "")
                except Exception as e:
                    log.error(f"Error detecting stage for {proj_id}: {e}")
                    proj["next_stage"] = "Error"
                    proj["next_stage_amount"] = 0
                    proj["stage_blocked"] = True
                    proj["stage_reason"] = str(e)

                projects.append(proj)

    return jsonify(projects)

@app.route("/invoicing/create-bulk", methods=["POST"])
@role_required("invoicing")
def create_bulk_invoices():
    """Create invoices for multiple projects' next stages."""
    project_ids = request.form.get("project_ids", "").split(",")
    project_ids = [pid.strip() for pid in project_ids if pid.strip()]

    if not project_ids:
        return jsonify({"success": False, "error": "No projects selected"}), 400

    all_projects = fb_get("/projects") or {}
    all_invoices = fb_get("/invoices") or {}
    created_invoice_ids = []

    try:
        for proj_id in project_ids:
            if proj_id not in all_projects:
                continue

            proj_data = all_projects[proj_id]
            if not isinstance(proj_data, dict):
                continue

            proj_num = proj_data.get("project_number", "")
            proj_name = proj_data.get("project_name", "")
            client_name = proj_data.get("client_name", "") or proj_data.get("company", "")
            stages = proj_data.get("payment_stages", [])

            if not isinstance(stages, list) or not stages:
                continue

            # Find which stages have been invoiced
            invoiced_stages = set()
            if isinstance(all_invoices, dict):
                for inv_data in all_invoices.values():
                    if isinstance(inv_data, dict) and inv_data.get("meta", {}).get("project_number", "") == proj_num:
                        stage_idx = inv_data.get("meta", {}).get("payment_stage_index")
                        if stage_idx is not None:
                            invoiced_stages.add(int(stage_idx))

            # Find first stage NOT invoiced
            next_stage_idx = None
            next_stage = None
            for idx, stage in enumerate(stages):
                if isinstance(stage, dict) and idx not in invoiced_stages:
                    next_stage_idx = idx
                    next_stage = stage
                    break

            if next_stage_idx is None or next_stage is None:
                # All stages already invoiced, skip this project
                continue

            # Build invoice data
            stage_name = next_stage.get("name", f"Stage {next_stage_idx + 1}")
            stage_amount = _safe_float(next_stage.get("amount", 0))

            invoice_data = {
                "meta": {
                    "invoice_number": _next_invoice_number(),
                    "project_number": proj_num,
                    "client_name": client_name,
                    "invoice_date": datetime.now().strftime("%Y-%m-%d"),
                    "due_date": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
                    "status": "Draft",
                    "payment_stage_index": next_stage_idx,
                    "payment_stage": stage_name,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": session.get("user_email", "")
                },
                "line_items": [
                    {
                        "description": f"{proj_name} — {stage_name}",
                        "project_number": proj_num,
                        "quantity": 1,
                        "unit_price": stage_amount,
                        "amount": stage_amount
                    }
                ]
            }

            # Create the invoice
            inv_id = fb_push("/invoices", invoice_data)
            created_invoice_ids.append(inv_id)

            # Mark stage as Invoiced with the actual stage amount
            _mark_project_stage(proj_num, next_stage_idx, "Invoiced", invoice_id=inv_id, amount=stage_amount)

        if created_invoice_ids:
            flash(f"Created {len(created_invoice_ids)} invoice(s) successfully.", "success")
            return jsonify({"success": True, "invoice_ids": created_invoice_ids})
        else:
            return jsonify({"success": False, "error": "No invoices created. All selected projects may be fully invoiced."}), 400

    except Exception as e:
        import traceback
        log.error("Bulk invoice creation error: %s", traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/projects/<project_id>/change-orders/<int:co_idx>/invoice", methods=["GET"])
@role_required("invoicing")
def co_create_invoice(project_id, co_idx):
    """Create invoice from approved change order - pre-fills with CO details."""
    project = fb_get(f"/projects/{project_id}") or {}
    if not project:
        abort(404)

    cos = project.get("change_orders") or []
    if not isinstance(cos, list):
        cos = list(cos.values()) if isinstance(cos, dict) else []
    if co_idx >= len(cos):
        abort(404)

    change_order = cos[co_idx]

    # Only allow invoice creation from approved COs
    if change_order.get("status") != "Approved":
        flash("Can only create invoice from approved change orders.", "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-change-orders")

    # Redirect to invoice creation form with CO data pre-filled
    # The invoice_new route will handle rendering the form with these parameters
    co_title = change_order.get("title", change_order.get("co_number", ""))
    co_amount = _safe_float(change_order.get("amount", 0))
    co_number = change_order.get("co_number", "")
    project_number = project.get("project_number", "")
    client_name = project.get("client_name", "")

    # Find the payment stage for this CO
    stages = project.get("payment_stages") or []
    stage_idx = None
    for idx, stage in enumerate(stages):
        if isinstance(stage, dict) and co_number in stage.get("name", ""):
            stage_idx = idx
            break

    # Build query parameters for pre-filling the form
    params = {
        "project": project_number,
        "client": client_name,
        "co_number": co_number,
        "co_title": co_title,
        "co_amount": str(co_amount),
    }

    if stage_idx is not None:
        params["stage_idx"] = stage_idx
        params["stage_name"] = stages[stage_idx].get("name", "")
        params["stage_amount"] = str(co_amount)

    # Redirect to invoice_new with CO parameters
    query_string = "&".join(f"{k}={v}" for k, v in params.items())
    return redirect(url_for("invoice_new") + f"?{query_string}")

@app.route("/invoicing/new", methods=["GET", "POST"])
@role_required("invoicing")
def invoice_new():
    clients  = _load_clients()
    projects = _load_projects_list()
    if request.method == "POST":
        data = _parse_invoice_form(request.form)
        project_number = data["meta"].get("project_number", "")

        # Get all projects and invoices for validation
        all_projects = fb_get("/projects") or {}
        raw_invoices = fb_get("/invoices") or {}

        # Check line items for duplicate stage invoicing
        item_projects = request.form.getlist("item_project[]")
        item_stage_indices = request.form.getlist("item_stage_index[]")

        # Find which stages are already invoiced
        invoiced_stages_map = {}  # {project_number: {stage_idx}}
        if isinstance(raw_invoices, dict):
            for inv_data in raw_invoices.values():
                if isinstance(inv_data, dict):
                    inv_proj = inv_data.get("meta", {}).get("project_number", "")
                    stage_idx = inv_data.get("meta", {}).get("payment_stage_index")
                    if inv_proj and stage_idx is not None:
                        if inv_proj not in invoiced_stages_map:
                            invoiced_stages_map[inv_proj] = set()
                        invoiced_stages_map[inv_proj].add(int(stage_idx))

        # Check if any line item's stage is already invoiced
        duplicate_stages = []
        for i, proj_num in enumerate(item_projects):
            if i < len(item_stage_indices) and item_stage_indices[i]:
                try:
                    stage_idx = int(item_stage_indices[i])
                    if proj_num in invoiced_stages_map and stage_idx in invoiced_stages_map[proj_num]:
                        duplicate_stages.append(f"{proj_num} - Stage {stage_idx + 1}")
                except (ValueError, IndexError):
                    pass

        if duplicate_stages:
            flash(f"Cannot create invoice. The following stages are already invoiced: {', '.join(duplicate_stages)}", "danger")
            return redirect(url_for("invoice_new"))

        # Validate invoice number is not duplicate
        invoice_num = data.get("meta", {}).get("invoice_number", "").strip()
        if invoice_num:
            if isinstance(raw_invoices, dict):
                for inv_data in raw_invoices.values():
                    if isinstance(inv_data, dict) and inv_data.get("meta", {}).get("invoice_number", "").strip() == invoice_num:
                        flash(f"Invoice number {invoice_num} already exists. Please use a different number.", "danger")
                        return redirect(url_for("invoice_new"))

        # Check if project is fully invoiced
        if project_number:
            for pid, pdata in (all_projects.items() if isinstance(all_projects, dict) else []):
                if isinstance(pdata, dict) and pdata.get("project_number", "") == project_number:
                    stages = pdata.get("payment_stages", [])
                    if isinstance(stages, list) and stages:
                        # Count how many stages have invoices
                        stages_with_invoices = set()
                        if isinstance(raw_invoices, dict):
                            for iid, idata in raw_invoices.items():
                                if isinstance(idata, dict):
                                    inv_proj = idata.get("meta", {}).get("project_number", "")
                                    if inv_proj == project_number:
                                        stage_idx = idata.get("meta", {}).get("payment_stage_index")
                                        if stage_idx is not None:
                                            stages_with_invoices.add(int(stage_idx))

                        # Check if all stages have invoices
                        total_stages = len([s for s in stages if isinstance(s, dict)])
                        if total_stages > 0 and len(stages_with_invoices) >= total_stages:
                            flash("This project has already been fully invoiced. All payment stages have invoices created.", "warning")
                            return redirect(url_for("project_detail", project_id=pid))
                    break

        data["meta"]["created_at"] = datetime.now(timezone.utc).isoformat()
        data["meta"]["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["meta"]["created_by"] = session.get("user_email", "")
        # Use status from form if provided, otherwise default to "Draft"
        form_status = request.form.get("status", "Draft").strip()
        valid_statuses = {"Draft", "Sent", "Viewed", "Paid", "Partial", "Overdue", "Cancelled"}
        data["meta"]["status"] = form_status if form_status in valid_statuses else "Draft"
        # Initialize amount_paid and tax_paid to 0 for new invoices
        data["meta"]["amount_paid"] = "0"
        data["meta"]["tax_paid"] = "0"

        # Store change order number if this invoice is from a CO
        co_number = request.form.get("co_number", "").strip()
        if co_number:
            data["meta"]["co_number"] = co_number

        stage_idx_raw = request.form.get("payment_stage_index", "")
        stage_name    = request.form.get("payment_stage", "")

        # If no stage was explicitly selected, auto-detect the first pending stage
        if stage_idx_raw == "":
            proj_num = data["meta"].get("project_number", "")
            if proj_num:
                all_projects = fb_get("/projects") or {}
                for pid, pdata in (all_projects.items() if isinstance(all_projects, dict) else []):
                    if isinstance(pdata, dict) and pdata.get("project_number", "") == proj_num:
                        stages = pdata.get("payment_stages", [])
                        if isinstance(stages, list) and stages:
                            # Find first pending stage
                            found_pending = False
                            for idx, stage in enumerate(stages):
                                if isinstance(stage, dict) and stage.get("status") == "Pending Invoice":
                                    stage_idx_raw = str(idx)
                                    stage_name = stage.get("name", f"Stage {idx + 1}")
                                    found_pending = True
                                    break

                            # If no pending stage found, all stages are already invoiced
                            if not found_pending:
                                flash("This project has already been fully invoiced. All payment stages have been invoiced.", "warning")
                                return redirect(url_for("project_detail", project_id=pid))
                        break

        if stage_idx_raw != "":
            data["meta"]["payment_stage_index"] = int(stage_idx_raw)
            data["meta"]["payment_stage"]       = stage_name

        inv_id = fb_push("/invoices", data)
        invoice_number = data["meta"].get("invoice_number", "")

        # Mark stages as invoiced
        if stage_idx_raw != "":
            # Single project with single stage - use only the subtotal (without tax)
            # Payment stages should show only the project amount, not including tax
            invoice_subtotal = _safe_float(data["meta"].get("subtotal", 0))
            _mark_project_stage(data["meta"].get("project_number", ""),
                                int(stage_idx_raw), "Invoiced", invoice_id=inv_id, invoice_number=invoice_number, amount=invoice_subtotal)
        else:
            # Multiple projects - check if line items have stage indices
            item_projects = request.form.getlist("item_project[]")
            item_stage_indices = request.form.getlist("item_stage_index[]")

            # Store linked projects info for multi-project invoice detection
            linked_projects = []

            # Mark each stage from line items - use project's actual line item amount (without tax)
            invoice_subtotal = _safe_float(data["meta"].get("subtotal", 0))
            for i, proj_num in enumerate(item_projects):
                if i < len(item_stage_indices):
                    stage_idx_str = item_stage_indices[i].strip() if item_stage_indices[i] else ""
                    if stage_idx_str:
                        try:
                            stage_idx = int(stage_idx_str)
                            # Get project's actual line item amount (not including tax)
                            line_items = data.get("line_items", []) or []
                            project_line_amount = sum(_safe_float(item.get("amount", 0)) for item in line_items if isinstance(item, dict) and item.get("project_number", "") == proj_num)
                            _mark_project_stage(proj_num, stage_idx, "Invoiced", invoice_id=inv_id, invoice_number=invoice_number, amount=project_line_amount)
                            linked_projects.append({"project_number": proj_num, "payment_stage_index": stage_idx})
                        except (ValueError, IndexError):
                            pass

            # Update invoice metadata with linked projects for multi-project invoices
            # SORT by project number (extract last digits, sort numerically) so 005 comes before 006
            if linked_projects:
                linked_projects.sort(key=lambda x: int(x.get("project_number", "")[-3:]) if x.get("project_number", "")[-3:].isdigit() else x.get("project_number", ""))
                fb_update(f"/invoices/{inv_id}", {"meta/linked_projects": linked_projects})

        # Auto-advance project status: Not Started → In Progress when first invoice created
        proj_nums_to_update = set()
        if stage_idx_raw != "":
            pn = data["meta"].get("project_number", "")
            if pn:
                proj_nums_to_update.add(pn)
        else:
            for pn in request.form.getlist("item_project[]"):
                if pn:
                    proj_nums_to_update.add(pn)
        if proj_nums_to_update:
            all_proj = fb_get("/projects") or {}
            for pid, pdata in (all_proj.items() if isinstance(all_proj, dict) else []):
                if isinstance(pdata, dict) and pdata.get("project_number", "") in proj_nums_to_update:
                    if pdata.get("status", "Not Started") == "Not Started":
                        fb_update(f"/projects/{pid}", {"status": "In Progress"})

        # Handle payment entry if provided when creating invoice
        payment_amount = data.pop("_payment_amount", "").strip()
        payment_date = data.pop("_payment_date", "").strip()
        payment_reference = data.pop("_payment_reference", "").strip()

        if payment_amount and _safe_float(payment_amount) > 0:
            # Use sequential distribution: allocate to projects first, then tax
            amount = _safe_float(payment_amount)
            payment_log = data.get("payment_log", []) or []
            if not isinstance(payment_log, list):
                payment_log = []

            tax_log = data.get("tax_payments", []) or []
            if not isinstance(tax_log, list):
                tax_log = []

            main_project = data["meta"].get("project_number", "")
            line_items = data.get("line_items", []) or []
            tax_amount = _safe_float(data["meta"].get("tax_amount", 0))
            linked_projects = data["meta"].get("linked_projects", [])

            remaining = amount

            # Step 1: Allocate to project(s)
            if not linked_projects and main_project:
                linked_projects = [main_project]

            for proj_num in linked_projects:
                if remaining <= 0:
                    break

                # Calculate project's invoice amount from line items
                proj_amount = sum(_safe_float(item.get("amount", 0))
                                for item in line_items
                                if isinstance(item, dict) and
                                (item.get("project_number") == proj_num or not item.get("project_number")))

                # Calculate already received
                proj_received = sum(_safe_float(p.get("amount", 0))
                                  for p in payment_log
                                  if p.get("project_number") == proj_num)

                proj_needs = max(0, proj_amount - proj_received)

                if proj_needs > 0:
                    allocate = min(proj_needs, remaining)
                    # Get stage name and index for this project
                    _stage_name = data["meta"].get("payment_stage", "")
                    _stage_idx = data["meta"].get("payment_stage_index")
                    if _stage_idx is not None:
                        try:
                            _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                        except (ValueError, TypeError):
                            _stage_idx = None
                    if not _stage_name and _stage_idx is not None:
                        _stage_name = f"Stage {_stage_idx + 1}"

                    payment_log.append({
                        "amount": str(allocate),
                        "date": payment_date or datetime.now().strftime("%Y-%m-%d"),
                        "method": data["meta"].get("payment_method", ""),
                        "reference": payment_reference,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "project_number": proj_num,
                        "invoice_number": data["meta"].get("invoice_number", ""),
                        "stage_name": _stage_name,
                        "stage_index": _stage_idx or "",
                    })
                    remaining -= allocate

            # Step 2: Allocate remaining to tax
            if remaining > 0 and tax_amount > 0:
                tax_received = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
                tax_needs = max(0, tax_amount - tax_received)

                if tax_needs > 0:
                    allocate_tax = min(tax_needs, remaining)
                    tax_log.append({
                        "amount": str(allocate_tax),
                        "date": payment_date or datetime.now().strftime("%Y-%m-%d"),
                        "method": data["meta"].get("payment_method", ""),
                        "reference": payment_reference,
                    })
                    remaining -= allocate_tax

            # Update invoice with payment data
            payment_log_data = payment_log
            tax_log_data = tax_log

            # Calculate total amount_paid and tax_paid
            total_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log_data)
            tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log_data)

            fb_update(f"/invoices/{inv_id}", {
                "payment_log": payment_log_data,
                "tax_payments": tax_log_data,
                "meta/amount_paid": str(total_paid),
                "meta/tax_paid": str(tax_paid),
            })

            # Update project stage payment amounts and status
            _update_project_stage_payment_status(inv_id)

            # Sync project-level payments
            for proj_num in proj_nums_to_update:
                _sync_project_payment(proj_num)
                _auto_complete_project_if_paid(proj_num)

            # Update financial entries
            fresh_meta = (fb_get(f"/invoices/{inv_id}") or {}).get("meta", {})
            _upsert_revenue_entry(inv_id, fresh_meta)

        flash("Invoice created successfully.", "success")
        return redirect(url_for("invoicing", tab="all-invoices"))
    next_num     = _next_invoice_number()
    prefill_proj = request.args.get("project", "")
    prefill_client = request.args.get("client", "")
    stage_idx    = request.args.get("stage_idx", "")
    stage_name   = request.args.get("stage_name", "")
    stage_amount = request.args.get("stage_amount", "")
    multiple_projects = request.args.get("projects", "")  # Comma-separated firebase IDs

    print(f"\n=== INVOICE_NEW GET REQUEST ===", flush=True)
    print(f"All args: {dict(request.args)}", flush=True)
    print(f"multiple_projects value: '{multiple_projects}'", flush=True)

    prefill_name   = ""
    prefill_amount = ""
    prefill_items  = []

    # Handle change order pre-filling
    co_number = request.args.get("co_number", "")
    co_title = request.args.get("co_title", "")
    co_amount = request.args.get("co_amount", "")
    if co_number and co_title and co_amount:
        # This is a change order invoice - use CO details as line item
        prefill_name = co_title
        prefill_amount = _safe_float(co_amount)
        prefill_items = [{
            "description": f"{co_title} ({co_number})",
            "project": prefill_proj,
            "amount": str(prefill_amount),
            "stage_index": request.args.get("stage_idx", "")
        }]

    # Handle multiple projects from modal (one line item per project, matching desktop software)
    if multiple_projects:
        project_ids = [pid.strip() for pid in multiple_projects.split(",") if pid.strip()]
        all_projects_data = fb_get("/projects") or {}
        raw_invoices = fb_get("/invoices") or {}

        # Auto-populate Project Number field if only one project selected
        if len(project_ids) == 1 and project_ids[0] in all_projects_data:
            single_proj_data = all_projects_data[project_ids[0]]
            if isinstance(single_proj_data, dict):
                prefill_proj = single_proj_data.get("project_number", "")

        for proj_id in project_ids:
            if proj_id in all_projects_data:
                proj_data = all_projects_data[proj_id]
                if isinstance(proj_data, dict):
                    proj_num = proj_data.get("project_number", "")
                    proj_name = proj_data.get("project_name", "")

                    # Detect next payment stage for this project
                    detection = _get_next_payment_stage(proj_data, raw_invoices)
                    next_stage_idx = detection.get("stage_idx")
                    next_stage_name = detection.get("stage_name")
                    stage_amount = detection.get("amount", 0)

                    # Use stage amount if available, otherwise use outstanding balance
                    if stage_amount > 0 and next_stage_idx is not None:
                        amount_to_invoice = stage_amount
                    else:
                        contract_value = _safe_float(proj_data.get("contract_value", 0))
                        amount_paid = _safe_float(proj_data.get("amount_paid", 0))
                        outstanding = contract_value - amount_paid
                        amount_to_invoice = outstanding if outstanding > 0 else contract_value

                    if amount_to_invoice > 0:
                        # Include stage name in description if available
                        description = proj_name
                        if next_stage_name:
                            description = f"{proj_name} — {next_stage_name}"

                        prefill_items.append({
                            "description": description,
                            "project": proj_num,
                            "amount": f"{amount_to_invoice:.2f}",
                            "stage_index": next_stage_idx  # Store the detected stage index
                        })

                    # Use first project for client field
                    if not prefill_client:
                        prefill_client = proj_data.get("client_name", "") or proj_data.get("company", "")

    # Set prefill_name and prefill_amount when single project is selected (with or without stage_idx)
    if prefill_proj:
        for p in projects:
            if p.get("project_number", "") == prefill_proj:
                prefill_name = p.get("project_name", "")
                if not stage_idx:  # Only set amount if no stage is specified
                    outstanding  = _safe_float(p.get("contract_value", 0)) - _safe_float(p.get("amount_paid", 0))
                    prefill_amount = f"{outstanding:.2f}" if outstanding > 0 else f"{_safe_float(p.get('contract_value', 0)):.2f}"
                break

    # Build invoiced stages map for display
    invoiced_stages_map = {}
    raw_inv = fb_get("/invoices") or {}
    if isinstance(raw_inv, dict):
        for inv_data in raw_inv.values():
            if isinstance(inv_data, dict):
                inv_proj = inv_data.get("meta", {}).get("project_number", "")
                inv_stage_idx = inv_data.get("meta", {}).get("payment_stage_index")
                if inv_proj and inv_stage_idx is not None:
                    if inv_proj not in invoiced_stages_map:
                        invoiced_stages_map[inv_proj] = set()
                    invoiced_stages_map[inv_proj].add(int(inv_stage_idx))

    projects = _enrich_projects_with_next_stage(projects, raw_inv)

    print(f"\n=== RENDERING TEMPLATE ===", flush=True)
    print(f"stage_idx='{stage_idx}', stage_idx != '' = {stage_idx != ''}", flush=True)
    print(f"prefill_items type: {type(prefill_items)}, length: {len(prefill_items) if isinstance(prefill_items, list) else 'N/A'}", flush=True)
    print(f"prefill_proj='{prefill_proj}'", flush=True)
    print(f"Condition check: stage_idx != '' = {stage_idx != ''}, prefill_items length = {len(prefill_items) if isinstance(prefill_items, list) else 0}", flush=True)

    # Lock unit price if loading from project payment stages
    lock_unit_price = isinstance(prefill_items, list) and len(prefill_items) > 0

    settings = load_settings()
    default_tax_rate = settings.get("company", {}).get("default_tax_rate", 0)

    return render_template("invoice_form.html", invoice=None, clients=clients,
                           projects=projects, next_num=next_num, is_new=True,
                           prefill_proj=prefill_proj, prefill_client=prefill_client,
                           prefill_name=prefill_name, prefill_amount=prefill_amount,
                           prefill_items=prefill_items,
                           stage_idx=stage_idx, stage_name=stage_name, stage_amount=stage_amount,
                           invoiced_stages_map=invoiced_stages_map, lock_unit_price=lock_unit_price,
                           default_tax_rate=default_tax_rate)

@app.route("/invoicing/<invoice_id>", methods=["GET"])
@role_required("invoicing")
def invoice_detail(invoice_id):
    data = fb_get(f"/invoices/{invoice_id}")
    if not data:
        abort(404)
    data["firebase_id"] = invoice_id

    # Linked project(s) — an invoice can bill multiple projects via per-line-item overrides
    linked_project = None
    linked_projects = []
    proj_num = data.get("meta", {}).get("project_number", "")
    all_proj_nums = _invoice_linked_projects(data)

    # Calculate amount paid per project from payment_log
    payment_log = data.get("payment_log", [])
    if not isinstance(payment_log, list):
        payment_log = []

    if proj_num or all_proj_nums:
        raw_proj = fb_get("/projects") or {}
        for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
            if not isinstance(pdata, dict):
                continue
            num = pdata.get("project_number")
            if num == proj_num or num in all_proj_nums:
                pdata = dict(pdata)
                pdata["firebase_id"] = pid
                pdata["_share"] = _invoice_project_share(data, num)
                # Calculate amount paid to this specific project
                proj_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log if p.get("project_number") == num)
                pdata["_paid"] = proj_paid
                if num == proj_num:
                    linked_project = pdata
                linked_projects.append(pdata)

    # Enrich payment_log with project names, firebase_id, and stage names
    raw_proj = fb_get("/projects") or {}
    enriched_payment_log = []
    for payment in payment_log:
        payment_copy = dict(payment)
        proj_num = payment_copy.get("project_number", "")

        if proj_num:
            # Find project data to get name, firebase_id and stage
            for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
                if not isinstance(pdata, dict):
                    continue
                if pdata.get("project_number") == proj_num:
                    payment_copy["project_name"] = pdata.get("project_name", "")
                    payment_copy["project_firebase_id"] = pid  # Add firebase_id for link
                    # Find the stage name and status from payment_stages
                    stages = pdata.get("payment_stages", [])
                    if isinstance(stages, list):
                        for stage in stages:
                            if isinstance(stage, dict) and stage.get("invoice_id") == invoice_id:
                                payment_copy["stage_name"] = stage.get("name", "Invoice Payment")
                                payment_copy["stage_status"] = stage.get("status", "Invoiced")
                                break
                    break
        enriched_payment_log.append(payment_copy)

    # Tax payments kept separate from projects (no enrichment with project data)
    tax_log = data.get("tax_payments", [])
    if not isinstance(tax_log, list):
        tax_log = []
    enriched_tax_payments = [dict(payment) for payment in tax_log]

    # Calculate current status based on actual payments (not stored status)
    calculated_status = _calculate_invoice_status(data)
    # Update the invoice data with calculated status for display
    data["meta"]["status"] = calculated_status

    # Ensure tax_paid is set in meta for template compatibility
    if "tax_paid" not in data["meta"]:
        tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
        data["meta"]["tax_paid"] = str(tax_paid) if tax_paid > 0 else "0"

    # Refresh project payment stage amounts to ensure they're always current
    _update_project_stage_payment_status(invoice_id)

    # Source quote — via the linked project's source_quote field
    source_quote = None
    if linked_project:
        sq_id = linked_project.get("source_quote")
        if sq_id:
            source_quote = fb_get(f"/job_forms/{sq_id}") or None
            if source_quote:
                source_quote["firebase_id"] = sq_id

    return render_template("invoice_detail.html", invoice=data, company=company_info(),
                           today_date=datetime.now().strftime("%Y-%m-%d"),
                           linked_project=linked_project, linked_projects=linked_projects,
                           source_quote=source_quote, enriched_payment_log=enriched_payment_log,
                           enriched_tax_payments=enriched_tax_payments)

@app.route("/invoicing/<invoice_id>/edit", methods=["GET", "POST"])
@role_required("invoicing")
def invoice_edit(invoice_id):
    invoice_data = fb_get(f"/invoices/{invoice_id}")
    if not invoice_data:
        abort(404)

    clients = _load_clients()
    projects = _load_projects_list()

    if request.method == "POST":
        # Handle invoice update
        data = _parse_invoice_form(request.form)

        # Update metadata with current timestamp
        data["meta"]["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["meta"]["updated_by"] = session.get("user_email", "")

        # Keep the original created_at and other preserved fields
        original_meta = invoice_data.get("meta", {})
        if "created_at" in original_meta:
            data["meta"]["created_at"] = original_meta["created_at"]
        # Ensure invoice_number is not accidentally changed
        if "invoice_number" in original_meta:
            data["meta"]["invoice_number"] = original_meta["invoice_number"]
        # Preserve payment stage linkage
        if "payment_stage_index" in original_meta:
            data["meta"]["payment_stage_index"] = original_meta["payment_stage_index"]
        if "payment_stage" in original_meta:
            data["meta"]["payment_stage"] = original_meta["payment_stage"]
        # Preserve linked_projects for multi-project invoices
        if "linked_projects" in original_meta:
            data["meta"]["linked_projects"] = original_meta["linked_projects"]
        # Preserve change order number if this is a CO invoice
        if "co_number" in original_meta:
            data["meta"]["co_number"] = original_meta["co_number"]

        # Preserve existing payment history when updating invoice details
        if "payment_log" in invoice_data:
            data["payment_log"] = invoice_data["payment_log"]
        if "tax_payments" in invoice_data:
            data["tax_payments"] = invoice_data["tax_payments"]

        # Get payment information from form
        payment_amount_str = request.form.get("payment_amount", "").strip()
        payment_date = request.form.get("payment_date", "").strip()
        payment_method = request.form.get("payment_method", "").strip()
        payment_reference = request.form.get("payment_reference", "").strip()

        print(f"\n[INVOICE_EDIT] START - invoice_id={invoice_id}", flush=True)
        print(f"[INVOICE_EDIT] payment_amount_str='{payment_amount_str}', payment_date='{payment_date}', payment_method='{payment_method}'", flush=True)
        print(f"[INVOICE_EDIT] All form keys: {list(request.form.keys())}", flush=True)

        # When editing, calculate the difference between new and old amount_paid
        # Only distribute the DIFFERENCE as a new payment (don't double-count)
        old_amount_paid = _safe_float(invoice_data.get("meta", {}).get("amount_paid", 0))
        new_amount_paid = _safe_float(payment_amount_str) if payment_amount_str else old_amount_paid
        payment_difference = new_amount_paid - old_amount_paid

        print(f"[INVOICE_EDIT] old_amount={old_amount_paid}, new_amount={new_amount_paid}, diff={payment_difference}", flush=True)
        print(f"[INVOICE_EDIT] Current meta: {invoice_data.get('meta', {})}", flush=True)

        # Preserve amount_paid and tax_paid in meta if not changing payment
        if payment_difference <= 0:
            # No new payment being added, so preserve existing values
            if "amount_paid" in invoice_data.get("meta", {}):
                data["meta"]["amount_paid"] = invoice_data["meta"]["amount_paid"]
            if "tax_paid" in invoice_data.get("meta", {}):
                data["meta"]["tax_paid"] = invoice_data["meta"]["tax_paid"]

        # Update invoice in Firebase (meta and line_items, preserving payments)
        fb_update(f"/invoices/{invoice_id}", data)

        if payment_difference > 0:
            print(f"[INVOICE_EDIT] Processing payment_difference={payment_difference}", flush=True)
            # Use sequential distribution: allocate to projects first, then tax
            amount = payment_difference
            payment_log = data.get("payment_log", []) or []
            if not isinstance(payment_log, list):
                payment_log = []

            tax_log = data.get("tax_payments", []) or []
            if not isinstance(tax_log, list):
                tax_log = []

            main_project = data["meta"].get("project_number", "")
            line_items = data.get("line_items", []) or []
            tax_amount = _safe_float(data["meta"].get("tax_amount", 0))
            linked_projects = data["meta"].get("linked_projects", [])

            print(f"[INVOICE_EDIT] main_project={main_project}, tax_amount={tax_amount}, linked_projects={linked_projects}", flush=True)

            remaining = amount

            # Step 1: Allocate to project(s)
            if not linked_projects and main_project:
                linked_projects = [main_project]

            for proj_item in linked_projects:
                if remaining <= 0:
                    break

                # Extract project number from dict or string
                if isinstance(proj_item, dict):
                    proj_num = proj_item.get("project_number", "")
                else:
                    proj_num = proj_item

                if not proj_num:
                    continue

                print(f"[INVOICE_EDIT] Processing project: {proj_num}", flush=True)

                # Calculate project's invoice amount from line items
                proj_amount = sum(_safe_float(item.get("amount", 0))
                                for item in line_items
                                if isinstance(item, dict) and
                                (item.get("project_number") == proj_num or not item.get("project_number")))

                # Calculate already received
                proj_received = sum(_safe_float(p.get("amount", 0))
                                  for p in payment_log
                                  if p.get("project_number") == proj_num)

                proj_needs = max(0, proj_amount - proj_received)

                if proj_needs > 0:
                    allocate = min(proj_needs, remaining)
                    # Get stage name and index for this project
                    _stage_name = data["meta"].get("payment_stage", "")
                    _stage_idx = data["meta"].get("payment_stage_index")
                    if _stage_idx is not None:
                        try:
                            _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                        except (ValueError, TypeError):
                            _stage_idx = None
                    if not _stage_name and _stage_idx is not None:
                        _stage_name = f"Stage {_stage_idx + 1}"

                    payment_log.append({
                        "amount": str(allocate),
                        "date": payment_date or datetime.now().strftime("%Y-%m-%d"),
                        "method": payment_method,
                        "reference": payment_reference,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "project_number": proj_num,
                        "invoice_number": data["meta"].get("invoice_number", ""),
                        "stage_name": _stage_name,
                        "stage_index": _stage_idx or "",
                    })
                    remaining -= allocate

            # Step 2: Allocate remaining to tax
            if remaining > 0 and tax_amount > 0:
                tax_received = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
                tax_needs = max(0, tax_amount - tax_received)

                if tax_needs > 0:
                    allocate_tax = min(tax_needs, remaining)
                    tax_log.append({
                        "amount": str(allocate_tax),
                        "date": payment_date or datetime.now().strftime("%Y-%m-%d"),
                        "method": payment_method,
                        "reference": payment_reference,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                    remaining -= allocate_tax

            # Update invoice with payment data
            total_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log)
            tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)

            print("\n[INVOICE_EDIT] PAYMENT DISTRIBUTION COMPLETE", flush=True)
            print(f"[INVOICE_EDIT] payment_log entries: {len(payment_log)}, total_paid={total_paid}, tax_paid={tax_paid}", flush=True)
            print(f"[INVOICE_EDIT] payment_log={payment_log}", flush=True)
            print(f"[INVOICE_EDIT] tax_log={tax_log}", flush=True)

            # Update invoice with payment log and totals
            update_dict = {
                "payment_log": payment_log,
                "tax_payments": tax_log,
                "meta/amount_paid": str(total_paid),
                "meta/tax_paid": str(tax_paid),
            }
            print(f"[INVOICE_EDIT] Updating Firebase with: {update_dict}", flush=True)
            fb_update(f"/invoices/{invoice_id}", update_dict)

            # Verify the update by reading back
            updated_invoice = fb_get(f"/invoices/{invoice_id}")
            updated_meta = updated_invoice.get("meta", {}) if updated_invoice else {}
            print(f"[INVOICE_EDIT] After update - amount_paid={updated_meta.get('amount_paid')}, tax_paid={updated_meta.get('tax_paid')}", flush=True)
            print("[INVOICE_EDIT] Firebase update complete", flush=True)

            # Update project stage payment amounts and status
            _update_project_stage_payment_status(invoice_id)

            # Sync project-level payment and financial summary (same as Add Payment dialog)
            fresh_inv = fb_get(f"/invoices/{invoice_id}") or data
            linked_projects = _invoice_linked_projects(fresh_inv)
            for proj_num in linked_projects:
                _sync_project_payment(proj_num)
                _auto_complete_project_if_paid(proj_num)

            # Update financial entries
            fresh_meta = (fb_get(f"/invoices/{invoice_id}") or {}).get("meta", {})
            _upsert_revenue_entry(invoice_id, fresh_meta)
        else:
            # Even if no new payment is being added, recalculate stage status based on actual payments
            # This ensures the stage shows Paid/Partial/Invoiced correctly
            _update_project_stage_payment_status(invoice_id)

            # Also link invoice to its stage
            invoice_number = data["meta"].get("invoice_number", "")
            main_project = data["meta"].get("project_number", "")

            # Handle single-project invoices with payment_stage_index
            stage_idx_raw = data["meta"].get("payment_stage_index")
            if stage_idx_raw is not None and main_project and invoice_number:
                try:
                    _mark_project_stage(main_project, int(stage_idx_raw), "Invoiced",
                                      invoice_id=invoice_id, invoice_number=invoice_number)
                except (ValueError, TypeError):
                    pass

            # Handle multi-project invoices with linked_projects
            linked_projects = data["meta"].get("linked_projects", [])
            if linked_projects and invoice_number:
                for proj_info in linked_projects:
                    if isinstance(proj_info, dict):
                        proj_num = proj_info.get("project_number", "")
                        stage_idx = proj_info.get("payment_stage_index")
                        if proj_num and stage_idx is not None:
                            try:
                                _mark_project_stage(proj_num, int(stage_idx), "Invoiced",
                                                  invoice_id=invoice_id, invoice_number=invoice_number)
                            except (ValueError, TypeError):
                                pass

        flash("Invoice updated successfully.", "success")
        return redirect(url_for("invoice_detail", invoice_id=invoice_id))

    # GET request - load invoice data for editing
    meta = invoice_data.get("meta", {})

    # Pass invoice data to form with is_new=False to indicate editing
    return render_template("invoice_form.html",
                         invoice=invoice_data,
                         clients=clients,
                         projects=projects,
                         next_num=meta.get("invoice_number", ""),
                         is_new=False,
                         invoice_id=invoice_id)

@app.route("/invoicing/<invoice_id>/send-reminder", methods=["POST"])
@role_required("invoicing")
def invoice_send_reminder(invoice_id):
    """Send a payment reminder email directly to the client."""
    try:
        inv = fb_get(f"/invoices/{invoice_id}")
        if not inv:
            return jsonify({"ok": False, "error": "Invoice not found"}), 404

        meta         = inv.get("meta", {}) or {}
        status       = meta.get("status", "")
        if status in ("Paid", "Cancelled", "Draft"):
            return jsonify({"ok": False, "error": "No reminder needed for this status"}), 400

        settings     = fb_get("/settings") or {}
        em           = settings.get("email", {}) or {}
        co           = settings.get("company", {}) or {}
        if not em.get("enabled") or not em.get("smtp_user"):
            return jsonify({"ok": False, "error": "Email sending is not configured in Settings"}), 400

        client_name  = (meta.get("client_name") or "").strip()
        client_email = ""
        raw_clients  = fb_get("/clients") or {}
        if isinstance(raw_clients, dict):
            for ckey, cd in raw_clients.items():
                if not isinstance(cd, dict):
                    continue
                # Client name is the Firebase key; company is a separate field
                if ckey.strip() == client_name or cd.get("company", "").strip() == client_name:
                    client_email = cd.get("email", "").strip()
                    break
        if not client_email:
            return jsonify({"ok": False, "error": f"No email address found for client '{client_name}'. Please add it in the Clients section."}), 400

        inv_num      = meta.get("invoice_number", invoice_id)
        inv_date     = meta.get("invoice_date", "")
        due_date     = meta.get("due_date", "")
        total        = _safe_float(meta.get("total", 0))
        amt_paid     = _safe_float(meta.get("amount_paid", 0))
        tax_paid     = _safe_float(meta.get("tax_paid", 0))
        outstanding  = max(0.0, total - amt_paid - tax_paid)
        proj_num     = meta.get("project_number", "")
        proj_name    = meta.get("project_name", proj_num)
        today_str    = datetime.now().strftime("%Y-%m-%d")
        is_overdue   = bool(due_date and due_date < today_str)
        company_name = co.get("name", "Our Company")
        from_name    = em.get("from_name", company_name)

        if is_overdue:
            subject = f"OVERDUE: Invoice {inv_num} — ${outstanding:,.2f} Outstanding"
            tone_line = f'<p style="color:#dc2626;font-weight:bold;">This invoice is <u>past due</u>. Immediate payment is appreciated.</p>'
            header_color = "#dc2626"
            header_label = "OVERDUE NOTICE"
        else:
            subject = f"Payment Reminder: Invoice {inv_num} — ${outstanding:,.2f} Outstanding"
            tone_line = "<p>We hope this message finds you well. This is a friendly reminder that the invoice below is due for payment.</p>"
            header_color = "#0D9488"
            header_label = "PAYMENT REMINDER"

        html_body = f"""<html><body style="margin:0;padding:0;background:#f1f5f9;font-family:Arial,sans-serif;color:#1a1a1a;">
<div style="max-width:600px;margin:32px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08);">
  <div style="background:{header_color};padding:28px 32px;">
    <div style="font-size:1.4rem;font-weight:800;color:#fff;">{company_name}</div>
    <div style="font-size:0.95rem;color:rgba(255,255,255,.85);margin-top:4px;">{header_label}</div>
  </div>
  <div style="padding:28px 32px;">
    <p>Dear {client_name},</p>
    {tone_line}
    <table style="width:100%;border-collapse:collapse;margin:20px 0;font-size:0.9rem;">
      <tr style="background:#f8fafc;"><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Invoice #</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;font-weight:700;">{inv_num}</td></tr>
      <tr><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Invoice Date</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;">{inv_date}</td></tr>
      <tr style="background:#f8fafc;"><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Due Date</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;{'color:#dc2626;font-weight:bold;' if is_overdue else ''}">{due_date or '—'}</td></tr>
      {'<tr><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Project</td><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;">' + proj_name + '</td></tr>' if proj_name else ''}
      <tr style="background:#f8fafc;"><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Invoice Total</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;">${total:,.2f}</td></tr>
      {'<tr><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#64748b;">Amount Paid</td><td style="padding:10px 14px;border-bottom:1px solid #e2e8f0;color:#10b981;">${:,.2f}</td></tr>'.format(amt_paid + tax_paid) if (amt_paid + tax_paid) > 0 else ''}
      <tr><td style="padding:12px 14px;font-weight:700;font-size:1rem;">Outstanding Balance</td>
          <td style="padding:12px 14px;font-weight:700;font-size:1.1rem;color:{header_color};">${outstanding:,.2f}</td></tr>
    </table>
    <p>Please arrange payment at your earliest convenience. If you have already made this payment, please disregard this message.</p>
    <p>If you have any questions, please reply to this email and we will be happy to assist.</p>
    <p style="margin-top:28px;">Warm regards,<br><strong>{from_name}</strong><br>
       <span style="color:#64748b;font-size:0.85rem;">{co.get('phone','')}{'&nbsp;&nbsp;·&nbsp;&nbsp;' + co.get('email','') if co.get('email') else ''}</span></p>
  </div>
  <div style="background:#f8fafc;padding:14px 32px;font-size:0.75rem;color:#94a3b8;text-align:center;">
    This is an automated reminder from {company_name}'s billing system.
  </div>
</div></body></html>"""

        msg            = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{from_name} <{em['smtp_user']}>"
        msg["To"]      = client_email
        msg["Reply-To"] = em["smtp_user"]
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(em.get("smtp_host", "smtp.gmail.com"), int(em.get("smtp_port", 587)), timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(em["smtp_user"], em.get("smtp_password", ""))
            srv.sendmail(em["smtp_user"], [client_email], msg.as_string())

        fb_update(f"/invoices/{invoice_id}", {
            "meta/last_reminder_sent": today_str,
            "meta/last_reminder_type": "Overdue Notice" if is_overdue else "Payment Reminder",
        })

        return jsonify({"ok": True, "sent_to": client_email, "type": "overdue" if is_overdue else "reminder"})

    except Exception as exc:
        log.error("invoice_send_reminder %s: %s", invoice_id, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/invoicing/<invoice_id>/status", methods=["POST"])
@role_required("invoicing")
def invoice_status(invoice_id):
    new_status       = request.form.get("status", "Draft")
    amount_paid      = request.form.get("amount_paid", "")
    payment_method   = request.form.get("payment_method", "")
    payment_ref      = request.form.get("payment_reference", "")
    updates = {
        "meta/status":     new_status,
        "meta/updated_at": datetime.now(timezone.utc).isoformat()
    }
    if amount_paid:
        updates["meta/amount_paid"] = amount_paid
    if payment_method:
        updates["meta/payment_method"] = payment_method
    if payment_ref:
        updates["meta/payment_reference"] = payment_ref
    fb_update(f"/invoices/{invoice_id}", updates)

    # ── Auto-sync: re-read fresh meta then sync project + balance sheet ──────
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    m = inv_data.get("meta", {})
    main_proj_num = m.get("project_number", "")
    for proj_num in _invoice_linked_projects(inv_data):
        _sync_project_payment(proj_num)
        if new_status == "Paid":
            _auto_complete_project_if_paid(proj_num)
        elif new_status in ("Sent", "Viewed", "Overdue"):
            _advance_project_to_in_progress(proj_num)
    if main_proj_num:
        # Roll the linked payment-plan stage's status forward with the invoice
        # (stages live on the invoice's main project only)
        stage_idx_meta = m.get("payment_stage_index")
        if stage_idx_meta is not None and stage_idx_meta != "":
            stage_status = {
                "Paid": "Paid",
                "Partial": "Partially Paid",
                "Draft": "Invoiced",
                "Sent": "Invoiced",
                "Viewed": "Invoiced",
                "Overdue": "Invoiced",
            }.get(new_status)
            if stage_status:
                # Update the stage with invoice subtotal (without tax)
                invoice_subtotal = _safe_float(m.get("subtotal", 0)) or (_safe_float(m.get("total", 0)) - _safe_float(m.get("tax_amount", 0)))
                _mark_project_stage(main_proj_num, int(stage_idx_meta), stage_status, invoice_id=invoice_id, amount=invoice_subtotal)
    if new_status in ("Paid", "Partial"):
        _upsert_revenue_entry(invoice_id, m)

    flash(f"Invoice updated to {new_status}. Project & balance sheet synced.", "success")
    return redirect(url_for("invoice_detail", invoice_id=invoice_id))

@app.route("/api/invoices/<invoice_id>", methods=["GET"])
@role_required("invoicing")
def api_get_invoice(invoice_id):
    """Get invoice details as JSON for frontend"""
    invoice = fb_get(f"/invoices/{invoice_id}")
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404
    return jsonify(invoice)

@app.route("/api/invoices/<invoice_id>/update-amount", methods=["POST"])
@role_required("invoicing")
def invoice_update_amount(invoice_id):
    """Update invoice line item amount from payment stage edit - handles multi-project invoices"""
    try:
        new_amount = _safe_float(request.form.get("new_amount", 0))
        project_number = request.form.get("project_number", "").strip()
        invoice = fb_get(f"/invoices/{invoice_id}") or {}
        meta = invoice.get("meta", {})
        tax_rate = _safe_float(meta.get("tax_rate", 0))

        # Update line items if they exist
        line_items = invoice.get("line_items", [])
        if line_items:
            # Find the correct line item to update
            line_item_idx = 0

            if project_number:
                # For multi-project invoices: find line item by project_number
                # This matches the payment_log approach where payments are tracked by project_number
                for idx, item in enumerate(line_items):
                    if isinstance(item, dict) and item.get("project_number", "").strip() == project_number:
                        line_item_idx = idx
                        break

            # Update only the identified line item's amount and unit_price
            line_items[line_item_idx]["amount"] = str(new_amount)
            line_items[line_item_idx]["unit_price"] = str(new_amount)

            # Recalculate Down Payment percentage if it's a down payment item
            if "Down Payment" in line_items[line_item_idx].get("description", ""):
                # Get contract value from project data
                contract_value = 0
                if project_number:
                    proj_data = fb_get(f"/projects/{project_number}") or {}
                    contract_value = _safe_float(proj_data.get("contract_value", 0))
                else:
                    # Fallback: get from main project_number
                    proj_num = meta.get("project_number", "")
                    if proj_num:
                        proj_data = fb_get(f"/projects/{proj_num}") or {}
                        contract_value = _safe_float(proj_data.get("contract_value", 0))

                # Calculate correct percentage
                if contract_value > 0:
                    dp_pct = int(round((new_amount / contract_value) * 100))
                    # Update description with correct percentage
                    desc = line_items[line_item_idx].get("description", "Down Payment")
                    base_desc = desc.split("(")[0].strip() if "(" in desc else desc
                    line_items[line_item_idx]["description"] = f"{base_desc} ({dp_pct}%)"

            invoice["line_items"] = line_items

            # Recalculate subtotal and total as sum of ALL line items (critical for multi-project!)
            # Same approach as payment_sequential: sum actual line item amounts
            invoice_subtotal = sum(_safe_float(item.get("amount", 0)) for item in line_items if isinstance(item, dict))
            meta["subtotal"] = str(invoice_subtotal)

            # Recalculate tax based on new subtotal and tax rate (same as payment_sequential process)
            new_tax_amount = invoice_subtotal * (tax_rate / 100.0) if tax_rate > 0 else 0
            meta["tax_amount"] = str(new_tax_amount)
            meta["total"] = str(invoice_subtotal + new_tax_amount)

            # For multi-project invoices: update linked_projects metadata to match current line items
            # This ensures _allocate_invoice_payment_sequential knows about all projects
            projects_in_items = set()
            for item in line_items:
                if isinstance(item, dict):
                    proj_num = item.get("project_number", "")
                    if proj_num:
                        projects_in_items.add(proj_num)

            # If multiple projects in line items, update linked_projects with CORRECT stage_index for each
            if len(projects_in_items) > 1:
                linked_projects_list = []
                for proj_num in sorted(projects_in_items):
                    # Find the correct stage_index for this project by matching line item descriptions
                    detected_stage_idx = meta.get("payment_stage_index", 0)  # Default fallback

                    for item in line_items:
                        if isinstance(item, dict) and item.get("project_number", "").strip() == proj_num:
                            # Found a line item for this project
                            item_desc = item.get("description", "")
                            # Look for stage name in the line item description
                            proj_id, proj_data = _find_project_by_number(proj_num)
                            if proj_data:
                                proj_stages = proj_data.get("payment_stages", [])
                                for pidx, pstage in enumerate(proj_stages):
                                    if isinstance(pstage, dict):
                                        pstage_name = pstage.get("name", "")
                                        if pstage_name and pstage_name in item_desc:
                                            detected_stage_idx = pidx
                                            break
                            break

                    linked_projects_list.append({
                        "project_number": proj_num,
                        "payment_stage_index": detected_stage_idx
                    })

                meta["linked_projects"] = linked_projects_list

        meta["updated_at"] = datetime.now(timezone.utc).isoformat()
        invoice["meta"] = meta
        invoice["updated_at"] = datetime.now(timezone.utc).isoformat()
        fb_update(f"/invoices/{invoice_id}", invoice)

        # When editing amounts, redistribute payments the SAME WAY as payment_sequential
        # This ensures both work identically - payments distributed sequentially across projects then tax
        # Use the locally updated invoice object (not fresh from DB) to get updated line_items and metadata
        amount_paid = _safe_float(meta.get("amount_paid", 0))
        tax_amount = _safe_float(meta.get("tax_amount", 0))
        line_items = invoice.get("line_items", []) or []  # Use local updated line_items
        main_project = meta.get("project_number", "")
        linked_projects = meta.get("linked_projects", [])  # Use updated linked_projects from meta

        # Only redistribute if there's an amount paid (otherwise no payment history to redistribute)
        if amount_paid > 0:
            # Clear existing payment logs and rebuild them sequentially (same as payment_sequential)
            new_payment_log = []
            new_tax_log = []
            remaining_to_distribute = amount_paid

            # Extract ALL projects from BOTH line_items AND linked_projects metadata
            # This handles invoices where some projects might not have line items
            projects_from_items = set()
            for item in line_items:
                if isinstance(item, dict):
                    proj_num = item.get("project_number", "")
                    if proj_num:
                        projects_from_items.add(proj_num)

            # Also get projects from linked_projects metadata (might have projects without line items)
            projects_from_meta = set()
            for proj_info in (meta.get("linked_projects") or []):
                if isinstance(proj_info, dict):
                    proj_num = proj_info.get("project_number", "")
                    if proj_num:
                        projects_from_meta.add(proj_num)

            # Merge both sets - include projects from BOTH line_items AND metadata
            all_projects = projects_from_items | projects_from_meta

            # Build linked_projects from merged project list
            if all_projects:
                linked_projects = [
                    {"project_number": proj_num, "payment_stage_index": meta.get("payment_stage_index", 0)}
                    for proj_num in sorted(all_projects)
                ]
            elif not linked_projects and main_project:
                linked_projects = [{"project_number": main_project, "payment_stage_index": meta.get("payment_stage_index", 0)}]

            # Step 1: Distribute to projects sequentially (sorted by project number)
            if linked_projects:
                # Sort projects by number
                def get_sort_key(x):
                    proj_num = x.get("project_number", "") if isinstance(x, dict) else x
                    if proj_num and proj_num[-3:].isdigit():
                        return int(proj_num[-3:])
                    return proj_num
                sorted_projects = sorted(linked_projects, key=get_sort_key)

                for proj_info in sorted_projects:
                    if remaining_to_distribute <= 0:
                        break

                    proj_num = proj_info.get("project_number", "") if isinstance(proj_info, dict) else proj_info
                    if not proj_num:
                        continue

                    # Get this project's line item amount
                    proj_amount = sum(_safe_float(item.get("amount", 0)) for item in line_items
                                    if isinstance(item, dict) and item.get("project_number", "").strip() == proj_num)

                    if proj_amount > 0:
                        # Allocate up to this project's amount
                        distribute_to_proj = min(proj_amount, remaining_to_distribute)

                        # Get stage info for payment entry
                        _stage_name = meta.get("payment_stage", "")
                        _stage_idx = meta.get("payment_stage_index")
                        if _stage_idx is not None:
                            try:
                                _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                            except (ValueError, TypeError):
                                _stage_idx = None

                        if not _stage_name and _stage_idx is not None:
                            _stage_name = f"Stage {_stage_idx + 1}"

                        new_payment_log.append({
                            "amount": str(distribute_to_proj),
                            "date": datetime.now().strftime("%Y-%m-%d"),
                            "created_at": datetime.now(timezone.utc).isoformat(),
                            "project_number": proj_num,
                            "invoice_number": meta.get("invoice_number", ""),
                            "stage_name": _stage_name,
                            "stage_index": _stage_idx or "",
                        })
                        remaining_to_distribute -= distribute_to_proj

            # Step 2: Distribute remaining to tax
            if remaining_to_distribute > 0 and tax_amount > 0:
                tax_needs = min(tax_amount, remaining_to_distribute)
                new_tax_log.append({
                    "amount": str(tax_needs),
                    "date": datetime.now().strftime("%Y-%m-%d"),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                remaining_to_distribute -= tax_needs

            # Update Firebase with redistributed payment logs (same structure as payment_sequential)
            fb_update(f"/invoices/{invoice_id}", {
                "payment_log": new_payment_log,
                "tax_payments": new_tax_log if new_tax_log else []
            })

        # Resync project stage payment status to ensure invoice_number is properly linked
        _update_project_stage_payment_status(invoice_id)

        return jsonify({"success": True, "message": "Invoice updated"})
    except Exception as e:
        log.error("Invoice update error: %s", str(e))
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/invoicing/<invoice_id>/delete", methods=["POST"])
@role_required("invoicing")
def invoice_delete(invoice_id):
    # Get invoice data before deleting
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    meta = inv_data.get("meta", {})

    print(f"\n=== INVOICE DELETE DEBUG ===", flush=True)
    print(f"Invoice ID: {invoice_id}", flush=True)
    print(f"Meta data: {meta}", flush=True)

    # Revert payment stages back to "Pending" and collect projects to sync
    project_numbers_to_sync = set()

    project_number = meta.get("project_number", "")
    payment_stage_index = meta.get("payment_stage_index")

    print(f"project_number: '{project_number}', payment_stage_index: {payment_stage_index}", flush=True)

    # Get payment amounts to subtract from stages
    payment_log = inv_data.get("payment_log", []) or []
    total_payments_deleted = sum(_safe_float(p.get("amount", 0)) for p in payment_log if isinstance(p, dict))
    print(f"Total payments in deleted invoice: {total_payments_deleted}", flush=True)

    if project_number and payment_stage_index is not None:
        # Single project with single stage - clear its amount_paid since this invoice is deleted
        print(f"Reverting single stage: {project_number} stage {payment_stage_index}", flush=True)
        _mark_project_stage(project_number, payment_stage_index, "Pending Invoice", amount_paid=0)
        project_numbers_to_sync.add(project_number)
    elif project_number and not payment_stage_index:
        # Invoice without explicit stage - find the stage by invoice_id
        print("Invoice without payment_stage_index - finding by invoice_id", flush=True)
        all_proj = fb_get("/projects") or {}
        for pid, pdata in (all_proj.items() if isinstance(all_proj, dict) else []):
            if isinstance(pdata, dict) and pdata.get("project_number", "") == project_number:
                stages = pdata.get("payment_stages", [])
                if isinstance(stages, list):
                    for idx, stage in enumerate(stages):
                        if isinstance(stage, dict) and stage.get("invoice_id") == invoice_id:
                            print(f"Found stage {idx} with this invoice_id, reverting", flush=True)
                            _mark_project_stage(project_number, idx, "Pending Invoice", amount_paid=0)
                            project_numbers_to_sync.add(project_number)
                            break
                break
    else:
        # Multiple projects - check linked_projects first, then line items
        linked_projects = meta.get("linked_projects", [])
        print(f"Linked projects: {linked_projects}", flush=True)
        if isinstance(linked_projects, list) and linked_projects:
            for lp in linked_projects:
                if isinstance(lp, dict):
                    proj_num = lp.get("project_number", "")
                    stage_idx = lp.get("payment_stage_index")
                    print(f"Processing linked project: proj_num={proj_num}, stage_idx={stage_idx}", flush=True)
                    if proj_num and stage_idx is not None:
                        _mark_project_stage(proj_num, stage_idx, "Pending Invoice", amount_paid=0)
                        project_numbers_to_sync.add(proj_num)
        else:
            # Fallback: check line items for older invoices
            line_items = inv_data.get("line_items", [])
            print(f"Line items: {line_items}", flush=True)
            if isinstance(line_items, list):
                for item in line_items:
                    if isinstance(item, dict):
                        proj_num = item.get("project_number", "") or item.get("project", "")
                        stage_idx = item.get("stage_index")
                        print(f"Processing line item: proj_num={proj_num}, stage_idx={stage_idx}", flush=True)
                        if proj_num and stage_idx is not None:
                            _mark_project_stage(proj_num, stage_idx, "Pending Invoice", amount_paid=0)
                            project_numbers_to_sync.add(proj_num)

    # Delete associated revenue entries
    all_revenue = fb_get("/balance_sheet_revenue") or {}
    if isinstance(all_revenue, dict):
        for rev_id, rev_data in list(all_revenue.items()):
            if isinstance(rev_data, dict) and rev_data.get("invoice_id") == invoice_id:
                fb_delete(f"/balance_sheet_revenue/{rev_id}")
                print(f"Deleted revenue entry: {rev_id}", flush=True)

    # Delete the invoice
    fb_delete(f"/invoices/{invoice_id}")
    print(f"=== INVOICE DELETE COMPLETE ===\n", flush=True)

    # Recalculate stage amounts_paid for all affected projects based on remaining invoices
    for proj_num in project_numbers_to_sync:
        if proj_num:
            # Find all invoices linked to this project and update their stage amounts_paid
            all_invoices = fb_get("/invoices") or {}
            for inv_id, inv_data in (all_invoices.items() if isinstance(all_invoices, dict) else []):
                if isinstance(inv_data, dict):
                    inv_meta = inv_data.get("meta", {}) or {}
                    # Check if this invoice is linked to the project we're updating
                    linked_projects = inv_meta.get("linked_projects", [])
                    is_linked = False

                    # Check if project is in linked_projects
                    if isinstance(linked_projects, list):
                        for lp in linked_projects:
                            if isinstance(lp, dict) and lp.get("project_number") == proj_num:
                                is_linked = True
                                break
                            elif isinstance(lp, str) and lp == proj_num:
                                is_linked = True
                                break

                    # Also check if it's the main project_number
                    if inv_meta.get("project_number") == proj_num:
                        is_linked = True

                    if is_linked:
                        _update_project_stage_payment_status(inv_id)
                        print(f"Updated stage amounts_paid for invoice {inv_id}", flush=True)

    # Sync payment amounts for all affected projects
    for proj_num in project_numbers_to_sync:
        if proj_num:
            # Recalculate all stages for this project from remaining invoices
            _sync_project_payment(proj_num)
            print(f"Synced payment for project: {proj_num}", flush=True)

            # Update project status based on remaining payments
            proj_id, pdata = _find_project_by_number(proj_num)
            if proj_id and pdata:
                amount_paid = _safe_float(pdata.get("amount_paid", 0))
                current_status = pdata.get("status", "Not Started")

                # Invoice-driven statuses that should revert when the invoice is removed
                _invoice_statuses = {
                    "Invoiced", "invoiced_Not paid yet", "invoiced_Partially paid",
                    "invoiced_Fully paid", "Sent", "Sent out_Invoiced",
                }
                # If still has payments, change to In Progress
                if amount_paid > 0 and current_status in ("Completed", "invoiced_Fully paid"):
                    fb_update(f"/projects/{proj_id}", {
                        "status": "In Progress",
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    })
                # If no payments left AND status was invoice-driven, revert to In Progress
                elif amount_paid == 0 and current_status in _invoice_statuses:
                    fb_update(f"/projects/{proj_id}", {
                        "status": "In Progress",
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    })

                # Recalculate all payment stage amounts for this project from remaining invoices
                stages = pdata.get("payment_stages", [])
                if isinstance(stages, list):
                    for stage_idx, stage in enumerate(stages):
                        if isinstance(stage, dict):
                            # Recalculate this stage's amount_paid from remaining invoices
                            all_invoices = fb_get("/invoices") or {}
                            stage_paid = 0.0
                            if isinstance(all_invoices, dict):
                                for inv_id, inv_data in all_invoices.items():
                                    if isinstance(inv_data, dict):
                                        inv_meta = inv_data.get("meta", {}) or {}
                                        # Check if this invoice covers this project and stage
                                        if (inv_meta.get("project_number") == proj_num and
                                            inv_meta.get("payment_stage_index") == stage_idx):
                                            # Sum payments for this invoice
                                            payment_log = inv_data.get("payment_log", [])
                                            if isinstance(payment_log, list):
                                                stage_paid += sum(_safe_float(p.get("amount", 0)) for p in payment_log)

                            # Update stage with recalculated amount_paid
                            stage["amount_paid"] = str(stage_paid)
                            print(f"[DELETE] Recalculated stage {stage_idx} amount_paid: {stage_paid}")

                    # Save updated stages
                    fb_update(f"/projects/{proj_id}", {
                        "payment_stages": stages,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    })

    flash("Invoice deleted successfully", "success")
    return redirect(url_for("invoicing"))

@app.route("/invoicing/<invoice_id>/pdf")
@role_required("invoicing")
def invoice_pdf(invoice_id):
    try:
        from reportlab.lib.pagesizes import A4
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("invoice_detail", invoice_id=invoice_id))

    pdf_bytes = _generate_invoice_pdf_bytes(invoice_id)
    if not pdf_bytes:
        abort(404)

    invoice = fb_get(f"/invoices/{invoice_id}")
    meta = invoice.get("meta", {}) if invoice else {}

    from flask import Response
    fname = f"Invoice_{meta.get('invoice_number','')}.pdf"
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={"Content-Disposition": f"inline;filename={fname}"})

# ── Routes: Invoicing Export ──────────────────────────────────────────────────
def _filter_invoices_export(items):
    if request.args.get("status"):
        items = [i for i in items if i.get("meta",{}).get("status","") == request.args["status"]]
    if request.args.get("client"):
        items = [i for i in items if i.get("meta",{}).get("client_name","") == request.args["client"]]
    date_from = request.args.get("from","")
    date_to   = request.args.get("to","")
    if date_from:
        items = [i for i in items if (i.get("meta",{}).get("invoice_date") or "") >= date_from]
    if date_to:
        items = [i for i in items if (i.get("meta",{}).get("invoice_date") or "") <= date_to]
    return items

@app.route("/invoicing/export/csv")
@role_required("invoicing")
def invoicing_export_csv():
    import csv
    import io
    raw = fb_get("/invoices") or {}
    items = []
    for iid, idata in (raw.items() if isinstance(raw, dict) else []):
        if idata and isinstance(idata, dict):
            items.append(idata)
    items.sort(key=lambda x: x.get("meta", {}).get("created_at", ""), reverse=True)
    items = _filter_invoices_export(items)
    output = io.StringIO()
    w = csv.writer(output)
    co = company_info()

    def fmt_csv_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    # Add company header and blank row
    w.writerow([f"{co.get('name','')} - Invoices Report"])
    w.writerow([])

    headers = ["Invoice Number", "Client", "Project", "Date", "Due Date", "Status", "Subtotal", "Tax", "Total", "Amount Paid", "Outstanding"]
    w.writerow(headers)

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for inv in items:
        m = inv.get("meta", {})
        date_str = m.get("invoice_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append((inv, m))

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda x: x[1].get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for inv, m in grouped[year][month][day]:
                    total = _safe_float(m.get("total", 0))
                    paid = _safe_float(m.get("amount_paid", 0))
                    tax_paid = _safe_float(m.get("tax_paid", 0))
                    total_paid = paid + tax_paid
                    subtotal = _safe_float(m.get("subtotal", 0))
                    tax = _safe_float(m.get("tax_amount", 0))
                    linked_projects = _invoice_linked_projects(inv)
                    projects_str = ", ".join(sorted(linked_projects)) if linked_projects else ""
                    row = [
                        m.get("invoice_number",""),
                        m.get("company_name","") or m.get("client_name",""),
                        projects_str,
                        fmt_csv_date(m.get("invoice_date","")),
                        fmt_csv_date(m.get("due_date","")),
                        m.get("status",""),
                        f"{subtotal:.2f}",
                        f"{tax:.2f}",
                        f"{total:.2f}",
                        f"{total_paid:.2f}",
                        f"{total-total_paid:.2f}"
                    ]
                    # Clean any "—" placeholders (replace with empty string)
                    row = ["" if str(cell).strip() == "—" else cell for cell in row]
                    w.writerow(row)

    output.seek(0)
    from flask import Response
    fname = f"invoices_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/invoicing/export/excel")
@role_required("invoicing")
def invoicing_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io
    raw = fb_get("/invoices") or {}
    items = []
    for iid, idata in (raw.items() if isinstance(raw, dict) else []):
        if idata and isinstance(idata, dict):
            items.append(idata)
    items.sort(key=lambda x: x.get("meta", {}).get("created_at", ""), reverse=True)
    items = _filter_invoices_export(items)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    hdr_font = Font(color="FFFFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13, color="FF0F766E")
    alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add title row
    co = company_info()
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=f"{co.get('name','')} - Invoices Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    headers = ["Invoice Number","Client","Project","Date","Due Date",
               "Subtotal ($)","Tax ($)","Total ($)","Paid ($)","Outstanding ($)","Status"]
    header_row = 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr

    def fmt_inv_excel_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for inv in items:
        m = inv.get("meta", {})
        date_str = m.get("invoice_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(inv)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda x: x.get("meta", {}).get("created_at", ""))

    ri = header_row + 1

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for inv in grouped[year][month][day]:
                    m = inv.get("meta", {})
                    total = _safe_float(m.get("total", 0))
                    paid  = _safe_float(m.get("amount_paid", 0))
                    tax_paid = _safe_float(m.get("tax_paid", 0))
                    total_paid = paid + tax_paid
                    subtotal = _safe_float(m.get("subtotal", 0))
                    tax = _safe_float(m.get("tax_amount", 0))
                    linked_projects = _invoice_linked_projects(inv)
                    projects_str = ", ".join(sorted(linked_projects)) if linked_projects else ""
                    row = [m.get("invoice_number",""), m.get("client_name",""),
                           projects_str, fmt_inv_excel_date(m.get("invoice_date","")), fmt_inv_excel_date(m.get("due_date","")),
                           subtotal, tax, total, total_paid, total - total_paid, m.get("status","")]
                    # Clean any "—" placeholders (replace with empty string)
                    row = ["" if str(cell).strip() == "—" else cell for cell in row]
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if ri % 2 == 0:
                            cell.fill = alt_fill
                        if ci in (7, 8, 9, 10, 11):
                            cell.number_format = '"$"#,##0.00'
                        cell.alignment = ctr
                    ri += 1

    # Increase column widths
    col_widths = [22, 25, 44, 14, 14, 14, 16, 14, 16, 14, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import Response
    fname = f"invoices_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/invoicing/export/pdf")
@role_required("invoicing")
def invoicing_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("invoicing", tab="export"))
    import io as _io
    raw = fb_get("/invoices") or {}
    items = []
    for iid, idata in (raw.items() if isinstance(raw, dict) else []):
        if idata and isinstance(idata, dict):
            items.append(idata)
    items.sort(key=lambda x: x.get("meta", {}).get("created_at", ""), reverse=True)
    items = _filter_invoices_export(items)
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=0.2*inch, rightMargin=0.2*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    co = company_info()
    elems = []
    from reportlab.lib.units import inch
    from reportlab.platypus import Spacer

    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=15,
                              fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                              alignment=1)  # CENTER

    elems.append(Paragraph(f"{co.get('name','')} — Invoices Report", title_s))
    elems.append(Spacer(1, 0.2*inch))
    hdrs = ["Invoice Number", "Client", "Project", "Date", "Due Date", "Subtotal", "Tax", "Total", "Paid", "Outstanding", "Status"]
    data = [hdrs]
    def fmt_inv_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=10, wordWrap='CJK')
    group_style = ParagraphStyle("group", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", alignment=1, leading=10, textColor=colors.HexColor("#0F172A"))

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for inv in items:
        m = inv.get("meta", {})
        date_str = m.get("invoice_date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(inv)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda x: x.get("meta", {}).get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for inv in grouped[year][month][day]:
                    m = inv.get("meta", {})
                    total = _safe_float(m.get("total", 0))
                    paid  = _safe_float(m.get("amount_paid", 0))
                    tax_paid = _safe_float(m.get("tax_paid", 0))
                    total_paid = paid + tax_paid
                    subtotal = _safe_float(m.get("subtotal", 0))
                    tax = _safe_float(m.get("tax_amount", 0))
                    linked_projects = _invoice_linked_projects(inv)
                    projects_str = "\n".join(sorted(linked_projects)) if linked_projects else "—"
                    data.append([
                        Paragraph(m.get("invoice_number","—"), cell_style),
                        Paragraph(m.get("client_name","—") or "—", cell_style),
                        Paragraph(projects_str, cell_style),
                        Paragraph(fmt_inv_date(m.get("invoice_date","")), cell_style),
                        Paragraph(fmt_inv_date(m.get("due_date","")), cell_style),
                        Paragraph(f"${subtotal:,.0f}", cell_style),
                        Paragraph(f"${tax:,.0f}", cell_style),
                        Paragraph(f"${total:,.0f}", cell_style),
                        Paragraph(f"${total_paid:,.0f}", cell_style),
                        Paragraph(f"${total-total_paid:,.0f}", cell_style),
                        Paragraph(m.get("status","—"), cell_style),
                    ])
    cw = [1.3*inch, 1.5*inch, 1.7*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.65*inch, 0.7*inch, 0.7*inch, 0.95*inch, 0.8*inch]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    from flask import Response
    fname = f"invoices_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

# ── Routes: Clients ───────────────────────────────────────────────────────────
@app.route("/clients")
@role_required("invoicing")
def clients():
    raw = fb_get("/clients") or {}
    items = []
    if isinstance(raw, dict):
        for company_name, cdata in raw.items():
            if cdata and isinstance(cdata, dict):
                # Ensure company_name is set from the Firebase key (primary identifier)
                cdata["company_name"] = company_name
                items.append(cdata)
    items.sort(key=lambda x: x.get("client_name", "").lower())
    search = request.args.get("q", "").strip().lower()
    tag_filter = request.args.get("tag", "")
    # Collect all tags from full list before filtering so chips always show
    all_tags = sorted({t for i in items for t in (i.get("tags") or []) if t})
    if search:
        items = [i for i in items if search in (
            (i.get("client_name","") + " " + i.get("company_name","") + " " +
             i.get("email","") + " " + i.get("phone",""))).lower()]
    if tag_filter:
        items = [i for i in items if tag_filter in (i.get("tags") or [])]
    active_tab = request.args.get("tab", "all-clients")
    return render_template("clients.html", clients=items, active_tab=active_tab,
                           search=search, tag_filter=tag_filter, all_tags=all_tags)

def _sync_client_changes(old_company_name, new_company_name, new_client_name):
    """Sync client changes to all related invoices, quotes, and projects by client_id.
    Also migrates legacy records (without client_id) to use client_id."""
    # Get the client_id from the new client record
    client_data = fb_get(f"/clients/{new_company_name}") or {}
    client_id = client_data.get("client_id")

    if not client_id:
        print(f"[SYNC] No client_id found for '{new_company_name}', falling back to name-based sync", flush=True)
        # Fallback to old method for legacy records without client_id
        _sync_client_changes_by_name(old_company_name, new_company_name, new_client_name)
        return

    print(f"[SYNC] Starting sync by client_id: '{client_id}'", flush=True)

    # Update invoices by client_id
    invoices = fb_get("/invoices") or {}
    if isinstance(invoices, dict):
        for inv_id, inv_data in invoices.items():
            if isinstance(inv_data, dict):
                meta = inv_data.get("meta", {})
                if isinstance(meta, dict):
                    # Match by client_id OR by legacy name matching (and add client_id to legacy records)
                    if meta.get("client_id") == client_id:
                        print(f"[SYNC] Updating invoice {inv_id} with client_id={client_id}", flush=True)
                        meta["company_name"] = new_company_name
                        meta["client_name"] = new_client_name
                        inv_data["meta"] = meta
                        fb_update(f"/invoices/{inv_id}", inv_data)
                    elif not meta.get("client_id") and (meta.get("company_name", "") == old_company_name or meta.get("client_name", "") == old_company_name):
                        print(f"[SYNC] Migrating legacy invoice {inv_id}: '{meta.get('company_name') or meta.get('client_name')}' → '{new_company_name}' (adding client_id)", flush=True)
                        meta["client_id"] = client_id
                        meta["company_name"] = new_company_name
                        meta["client_name"] = new_client_name
                        inv_data["meta"] = meta
                        fb_update(f"/invoices/{inv_id}", inv_data)

    # Update quotes by client_id (stored in /job_forms)
    quotes = fb_get("/job_forms") or {}
    if isinstance(quotes, dict):
        print(f"[SYNC] Checking {len(quotes)} quotes for client_id={client_id}", flush=True)
        for quote_id, quote_data in quotes.items():
            if isinstance(quote_data, dict):
                quote_cid = quote_data.get("client_id", "")
                quote_company = quote_data.get("company_name", "")
                print(f"[SYNC] Quote {quote_id}: client_id='{quote_cid}' company_name='{quote_company}'", flush=True)
                # Match by client_id OR by legacy name matching (and add client_id to legacy records)
                if quote_data.get("client_id") == client_id:
                    print(f"[SYNC] Updating quote {quote_id} with client_id={client_id}: '{quote_data.get('company_name')}' → '{new_company_name}'", flush=True)
                    quote_data["company_name"] = new_company_name
                    quote_data["client_name"] = new_client_name
                    fb_update(f"/job_forms/{quote_id}", quote_data)
                    print(f"[SYNC] Quote {quote_id} updated: company_name='{new_company_name}' client_name='{new_client_name}'", flush=True)
                elif not quote_data.get("client_id") and (quote_data.get("company_name", "") == old_company_name or quote_data.get("client_name", "") == old_company_name):
                    print(f"[SYNC] Migrating legacy quote {quote_id}: '{quote_data.get('company_name') or quote_data.get('client_name')}' → '{new_company_name}' (adding client_id)", flush=True)
                    quote_data["client_id"] = client_id
                    quote_data["company_name"] = new_company_name
                    quote_data["client_name"] = new_client_name
                    fb_update(f"/job_forms/{quote_id}", quote_data)

    # Update projects by client_id
    projects = fb_get("/projects") or {}
    if isinstance(projects, dict):
        for proj_id, proj_data in projects.items():
            if isinstance(proj_data, dict):
                # Match by client_id OR by legacy name matching (and add client_id to legacy records)
                if proj_data.get("client_id") == client_id:
                    print(f"[SYNC] Updating project {proj_id} with client_id={client_id}", flush=True)
                    proj_data["company_name"] = new_company_name
                    proj_data["client_name"] = new_client_name
                    fb_update(f"/projects/{proj_id}", proj_data)
                elif not proj_data.get("client_id") and (proj_data.get("company_name", "") == old_company_name or proj_data.get("client_name", "") == old_company_name):
                    print(f"[SYNC] Migrating legacy project {proj_id}: '{proj_data.get('company_name') or proj_data.get('client_name')}' → '{new_company_name}' (adding client_id)", flush=True)
                    proj_data["client_id"] = client_id
                    proj_data["company_name"] = new_company_name
                    proj_data["client_name"] = new_client_name
                    fb_update(f"/projects/{proj_id}", proj_data)

def _sync_client_changes_by_name(old_company_name, new_company_name, new_client_name):
    """Legacy fallback: Sync client changes by name matching (for records without client_id)."""
    print(f"[SYNC] Fallback sync by name: old='{old_company_name}' → new='{new_company_name}'", flush=True)

    # Update invoices
    invoices = fb_get("/invoices") or {}
    if isinstance(invoices, dict):
        for inv_id, inv_data in invoices.items():
            if isinstance(inv_data, dict):
                meta = inv_data.get("meta", {})
                if isinstance(meta, dict):
                    if meta.get("company_name", "") == old_company_name or meta.get("client_name", "") == old_company_name:
                        meta["company_name"] = new_company_name
                        meta["client_name"] = new_client_name
                        inv_data["meta"] = meta
                        fb_update(f"/invoices/{inv_id}", inv_data)

    # Update quotes (stored in /job_forms)
    quotes = fb_get("/job_forms") or {}
    if isinstance(quotes, dict):
        for quote_id, quote_data in quotes.items():
            if isinstance(quote_data, dict):
                if quote_data.get("company_name", "") == old_company_name or quote_data.get("client_name", "") == old_company_name:
                    quote_data["company_name"] = new_company_name
                    quote_data["client_name"] = new_client_name
                    fb_update(f"/job_forms/{quote_id}", quote_data)

    # Update projects
    projects = fb_get("/projects") or {}
    if isinstance(projects, dict):
        for proj_id, proj_data in projects.items():
            if isinstance(proj_data, dict):
                if proj_data.get("company_name", "") == old_company_name or proj_data.get("client_name", "") == old_company_name:
                    proj_data["company_name"] = new_company_name
                    proj_data["client_name"] = new_client_name
                    fb_update(f"/projects/{proj_id}", proj_data)

@app.route("/clients/new", methods=["GET", "POST"])
@role_required("invoicing")
def client_new():
    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        client_name = request.form.get("client_name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        notes = request.form.get("notes", "").strip()
        tags = request.form.get("tags", "").strip()

        # At least one of company_name or client_name must be provided
        if not company_name and not client_name:
            flash("Either Company Name or Client Name is required.", "danger")
            return render_template("client_form.html", client=None, is_new=True)

        # If company_name is empty, use client_name as the company_name
        if not company_name:
            company_name = client_name

        # Use company_name as primary identifier
        primary_id = company_name

        # Build form data for re-rendering on error
        form_data = {
            "company_name": company_name,
            "client_name": client_name,
            "email": email,
            "phone": phone,
            "address": address,
            "notes": notes,
            "tags": tags,
        }

        # Check for duplicate email
        if email:
            all_clients = fb_get("/clients") or {}
            for existing_id, existing_data in all_clients.items():
                if isinstance(existing_data, dict) and existing_data.get("email", "").strip().lower() == email.lower():
                    flash(f"Email address '{email}' is already in use by another client.", "danger")
                    form_data["email"] = ""
                    return render_template("client_form.html", client=form_data, is_new=True)

        # Check for duplicate phone
        if phone:
            all_clients = fb_get("/clients") or {}
            for existing_id, existing_data in all_clients.items():
                if isinstance(existing_data, dict) and existing_data.get("phone", "").strip() == phone:
                    flash(f"Phone number '{phone}' is already in use by another client.", "danger")
                    form_data["phone"] = ""
                    return render_template("client_form.html", client=form_data, is_new=True)

        raw_tags = request.form.get("tags", "")
        tags = [t.strip() for t in raw_tags.split(",") if t.strip()]

        # Generate unique client_id for this client
        client_id = secrets.token_hex(8)

        data = {
            "client_id":    client_id,
            "company_name": company_name,
            "client_name":  client_name,
            "email":        email,
            "phone":        phone,
            "address":      request.form.get("address", ""),
            "notes":        request.form.get("notes", ""),
            "tags":         tags,
            "updated_at":   datetime.now(timezone.utc).isoformat(),
        }
        fb_update(f"/clients/{primary_id}", data)
        flash("Client saved.", "success")
        return redirect(url_for("clients", tab="all-clients"))
    return render_template("client_form.html", client=None, is_new=True)

@app.route("/clients/<company_name>/edit", methods=["GET", "POST"])
@role_required("invoicing")
def client_edit(company_name):
    data = fb_get(f"/clients/{company_name}") or {}
    # Get the ACTUAL company name from database (in case it's different)
    original_company_name = data.get("company_name", company_name)
    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        new_client_name = request.form.get("client_name", data.get("client_name", "")).strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        notes = request.form.get("notes", "").strip()
        tags = request.form.get("tags", "").strip()

        # At least one of company_name or client_name must be provided
        if not company_name and not new_client_name:
            flash("Either Company Name or Client Name is required.", "danger")
            return render_template("client_form.html", client=data, is_new=False)

        # If company_name is empty, use client_name as the company_name
        if not company_name:
            company_name = new_client_name

        # Use company_name as primary identifier
        new_primary_id = company_name

        # Build form data for re-rendering on error
        form_data = {
            "company_name": company_name,
            "client_name": new_client_name,
            "email": email,
            "phone": phone,
            "address": address,
            "notes": notes,
            "tags": tags,
        }

        # Check for duplicate email (excluding current client)
        if email:
            all_clients = fb_get("/clients") or {}
            for existing_id, existing_data in all_clients.items():
                if existing_id != original_company_name and isinstance(existing_data, dict):
                    if existing_data.get("email", "").strip().lower() == email.lower():
                        flash(f"Email address '{email}' is already in use by another client.", "danger")
                        form_data["email"] = ""
                        return render_template("client_form.html", client=form_data, is_new=False)

        # Check for duplicate phone (excluding current client)
        if phone:
            all_clients = fb_get("/clients") or {}
            for existing_id, existing_data in all_clients.items():
                if existing_id != original_company_name and isinstance(existing_data, dict):
                    if existing_data.get("phone", "").strip() == phone:
                        flash(f"Phone number '{phone}' is already in use by another client.", "danger")
                        form_data["phone"] = ""
                        return render_template("client_form.html", client=form_data, is_new=False)

        raw_tags = request.form.get("tags", "")
        tags = [t.strip() for t in raw_tags.split(",") if t.strip()]

        # Get the client_id from existing data (don't generate a new one)
        client_id = data.get("client_id", secrets.token_hex(8))

        updated = {
            "client_id":    client_id,
            "company_name": company_name,
            "client_name":  new_client_name,
            "email":        email,
            "phone":        phone,
            "address":      request.form.get("address", ""),
            "notes":        request.form.get("notes", ""),
            "tags":         tags,
            "updated_at":   datetime.now(timezone.utc).isoformat(),
        }
        # If primary ID changed, delete old entry
        if new_primary_id != original_company_name:
            fb_delete(f"/clients/{original_company_name}")
        fb_update(f"/clients/{new_primary_id}", updated)

        # Sync all related invoices, quotes, and projects with new company name
        _sync_client_changes(original_company_name, new_primary_id, new_client_name)

        flash("Client updated.", "success")
        return redirect(url_for("clients"))
    return render_template("client_form.html", client=data, is_new=False)

@app.route("/clients/<company_name>/delete", methods=["POST"])
@role_required("invoicing")
def delete_client(company_name):
    # Only delete the client record - do NOT cascade delete quotes/projects/invoices
    # They keep the client information (company_name, client_name, client_id) for historical reference
    fb_delete(f"/clients/{company_name}")
    flash(f"Client '{company_name}' deleted.", "success")
    return redirect(url_for("clients"))

# ── Client Statement PDF ──────────────────────────────────────────────────────
@app.route("/clients/<company_name>/statement")
@role_required("invoicing")
def client_statement(company_name):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("clients"))
    import io as _io

    co = company_info()
    client_data = fb_get(f"/clients/{company_name}") or {}
    client_name = client_data.get("client_name", "")

    # Load all invoices for this client
    raw_inv = fb_get("/invoices") or {}
    inv_list = []
    if isinstance(raw_inv, dict):
        for iid, idata in raw_inv.items():
            if isinstance(idata, dict) and idata.get("meta", {}).get("client_name", "") == client_name:
                idata["firebase_id"] = iid
                inv_list.append(idata)
    inv_list.sort(key=lambda x: x.get("meta", {}).get("invoice_date", ""))

    total_invoiced = sum(_safe_float(i.get("meta",{}).get("total", 0)) for i in inv_list)
    total_paid     = sum(_safe_float(i.get("meta",{}).get("amount_paid", 0)) for i in inv_list)
    total_paid    += sum(_safe_float(p.get("amount", 0)) for i in inv_list for p in i.get("tax_payments", []))
    balance_due    = total_invoiced - total_paid

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    teal   = colors.HexColor("#0F766E")
    dark   = colors.HexColor("#0F172A")
    muted  = colors.HexColor("#64748B")
    light  = colors.HexColor("#F8FAFC")
    border = colors.HexColor("#E2E8F0")
    lbl = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=8,  fontName="Helvetica-Bold", textColor=muted)
    val = ParagraphStyle("val", parent=styles["Normal"], fontSize=10, fontName="Helvetica",       textColor=dark, spaceAfter=6)
    sm  = ParagraphStyle("sm",  parent=styles["Normal"], fontSize=9,  fontName="Helvetica",       textColor=muted)
    h2  = ParagraphStyle("h2",  parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold",  textColor=dark, spaceBefore=10, spaceAfter=6)
    elems = []

    # Header
    hdr_data = [[
        Paragraph(f"<b>{co.get('name','')}</b>",
                  ParagraphStyle("cn", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", textColor=dark)),
        Paragraph("CLIENT STATEMENT",
                  ParagraphStyle("cs", parent=styles["Normal"], fontSize=18, fontName="Helvetica-Bold", textColor=teal, alignment=2)),
    ],[
        Paragraph(f"{co.get('address','').replace(chr(10),' | ')}  |  {co.get('phone','')}  |  {co.get('email','')}", sm),
        Paragraph(f"As of {datetime.now().strftime('%B %d, %Y')}",
                  ParagraphStyle("dt", parent=styles["Normal"], fontSize=10, textColor=muted, alignment=2)),
    ]]
    hdr = Table(hdr_data, colWidths=[3.5*inch, 3.5*inch])
    hdr.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    elems.extend([hdr, HRFlowable(width="100%", thickness=2, color=teal, spaceAfter=12)])

    # Client info
    elems.append(Paragraph("BILL TO", lbl))
    elems.append(Paragraph(f"<b>{client_name}</b>", val))
    if client_data.get("company"):
        elems.append(Paragraph(client_data["company"], sm))
    if client_data.get("email"):
        elems.append(Paragraph(client_data["email"], sm))
    if client_data.get("phone"):
        elems.append(Paragraph(client_data["phone"], sm))
    elems.append(Spacer(1, 12))

    # Invoice table - one row per invoice with all projects listed vertically
    elems.append(Paragraph("INVOICE HISTORY", h2))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=border, spaceAfter=6))
    tbl_data = [["Invoice #", "Project", "Date", "Due Date", "Status", "Total", "Paid", "Balance"]]

    for inv in inv_list:
        m = inv.get("meta", {})
        total = _safe_float(m.get("total", 0))
        paid  = _safe_float(m.get("amount_paid", 0))
        tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in inv.get("tax_payments", []))
        total_paid_with_tax = paid + tax_paid
        balance = total - total_paid_with_tax

        # Get all linked projects
        linked_projects = _invoice_linked_projects(inv)
        projects_list = sorted(list(linked_projects)) if linked_projects else ["—"]

        # Create project column with all projects listed vertically (separated by newlines)
        projects_text = "\n".join(projects_list)

        # Add single row with invoice info and all projects stacked in one cell
        tbl_data.append([
            m.get("invoice_number", "—"),
            projects_text,
            m.get("invoice_date", "—") or "—",
            m.get("due_date", "—") or "—",
            m.get("status", "—"),
            f"${total:,.2f}",
            f"${total_paid_with_tax:,.2f}",
            f"${balance:,.2f}",
        ])

    if not inv_list:
        tbl_data.append(["No invoices found.", "", "", "", "", "", "", ""])

    cw = [1.0*inch, 1.2*inch, 0.85*inch, 0.85*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch]
    tbl = Table(tbl_data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), dark),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 8),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 3), ("BOTTOMPADDING",(0,0),(-1,0),3),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, light]),
        ("GRID",          (0,0), (-1,-1), 0.4, border),
        ("TOPPADDING",    (0,1), (-1,-1), 6), ("BOTTOMPADDING",(0,1),(-1,-1),6),
        ("LEFTPADDING",   (0,1), (-1,-1), 4), ("RIGHTPADDING",(0,1),(-1,-1),4),
        ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("ALIGN",         (-3,1),(-1,-1), "RIGHT"),
    ]))
    elems.append(tbl)

    # Summary
    elems.append(Spacer(1, 16))
    sum_data = [
        ["Total Invoiced",  f"${total_invoiced:,.2f}"],
        ["Total Paid",      f"${total_paid:,.2f}"],
        ["Balance Due",     f"${balance_due:,.2f}"],
    ]
    sum_tbl = Table(sum_data, colWidths=[2.5*inch, 1.5*inch])
    sum_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("FONTNAME",  (0,2), (-1,2),  "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1),  muted),
        ("FONTNAME",  (1,0), (1,-1),  "Helvetica-Bold"),
        ("TEXTCOLOR", (1,2), (1,2),   colors.HexColor("#DC2626") if balance_due > 0 else teal),
        ("TEXTCOLOR", (1,0), (1,0),   teal),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white, light]),
        ("GRID",      (0,0), (-1,-1), 0.4, border),
        ("TOPPADDING",(0,0), (-1,-1), 6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("ALIGN",     (1,0), (1,-1),  "RIGHT"),
    ]))
    elems.append(sum_tbl)

    doc.build(elems)
    buf.seek(0)
    from flask import Response
    import re
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '', client_name.replace(" ", "_"))
    fname = f"statement_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment;filename="{fname}"'})

# ── Routes: Payroll ───────────────────────────────────────────────────────────
def _load_employee_profiles() -> list:
    raw = fb_get("/employee_profiles") or {}
    profiles = []
    if isinstance(raw, dict):
        for pid, pdata in raw.items():
            if isinstance(pdata, dict):
                pdata["firebase_id"] = pid
                profiles.append(pdata)
    profiles.sort(key=lambda x: x.get("name", "").lower())
    return profiles

def _auto_generate_monthly_salaries() -> int:
    """Create payroll entries for users with monthly_salary > 0 (Settings-based employees).
    Runs once per calendar month per employee — skipped if an entry for the
    current YYYY-MM already exists in /balance_sheet_salary.
    Returns the number of new entries created.
    """
    users = _load_all_users()
    if not users:
        return 0
    now         = datetime.now()
    month_key   = now.strftime("%Y-%m")
    month_label = now.strftime("%B %Y")
    raw_sal     = fb_get("/balance_sheet_salary") or {}
    existing    = set()
    if isinstance(raw_sal, dict):
        for entry in raw_sal.values():
            if isinstance(entry, dict):
                name = (entry.get("employee_name") or entry.get("employee", "")).strip().lower()
                date = (entry.get("date") or "")[:7]
                if name and date == month_key:
                    existing.add(name)
    created = 0
    for u in users:
        monthly = _safe_float(u.get("monthly_salary", 0))
        if monthly <= 0:
            continue
        name = (u.get("username") or "").strip()
        if not name or name.lower() in existing:
            continue
        fb_push("/balance_sheet_salary", {
            "employee_name":  name,
            "employee":       name,
            "amount":         monthly,
            "date":           now.strftime("%Y-%m-01"),
            "description":    f"Monthly salary — {month_label}",
            "type":           "salary",
            "region":         u.get("region", ""),
            "auto_generated": True,
            "created_at":     datetime.now(timezone.utc).isoformat(),
        })
        created += 1
    return created

@app.route("/payroll")
@role_required("payroll")
def payroll():
    _auto_generate_monthly_salaries()
    employee_filter = request.args.get("employee", "")
    year_filter     = request.args.get("year", "")
    region_filter   = request.args.get("region", "")

    # Load salary records server-side so the table renders without relying on JS fetch
    raw_sal = fb_get("/balance_sheet_salary") or {}
    salaries = []
    if isinstance(raw_sal, dict):
        for sid, sdata in raw_sal.items():
            if isinstance(sdata, dict):
                sdata["firebase_id"] = sid
                salaries.append(sdata)
    salaries.sort(key=lambda s: s.get("date", ""), reverse=True)

    # Build employee list from /users (Settings) — normalize 'username' → 'name' for template compatibility
    raw_users = _load_all_users()
    employee_profiles = [
        dict(u, name=u.get("username", ""))
        for u in raw_users
        if u.get("active", True)
    ]

    # ── Commission per salesperson per period ─────────────────────────────────
    _sales_comm_map: Dict[str, float] = {}
    for _u in raw_users:
        if normalize_role(_u.get("role", "")) == "sales":
            _uname = (_u.get("username") or "").strip()
            if _uname:
                _sales_comm_map[_uname] = _safe_float(_u.get("commission_rate", 0))

    _qraw_pay = fb_get("/job_forms") or {}
    _praw_pay = fb_get("/projects") or {}
    _pst_pay: Dict[str, str] = {}
    if isinstance(_praw_pay, dict):
        for _pid, _pd in _praw_pay.items():
            if _pd and isinstance(_pd, dict):
                _pst_pay[_pid] = _pd.get("status", "")

    # commission_by_period[sp_name][YYYY-MM] = earned amount
    commission_by_period: Dict[str, Dict[str, float]] = {}
    _CONV_PAY = {"Converted", "Invoiced"}
    if isinstance(_qraw_pay, dict):
        for _fid, _fdata in _qraw_pay.items():
            if not _fdata or not isinstance(_fdata, dict):
                continue
            _sp = (_fdata.get("salesperson") or "").strip()
            _rate = _sales_comm_map.get(_sp, 0)
            if not _sp or not _rate:
                continue
            _linked = _fdata.get("linked_project_id", "")
            _is_conv = _fdata.get("status", "") in _CONV_PAY or bool(_linked)
            if not _is_conv:
                continue
            if _linked and _pst_pay.get(_linked) == "Cancelled":
                continue
            _period = (_fdata.get("date") or "")[:7]
            if not _period:
                continue
            _earned = _safe_float(_fdata.get("total", 0)) * _rate / 100
            if _sp not in commission_by_period:
                commission_by_period[_sp] = {}
            commission_by_period[_sp][_period] = \
                commission_by_period[_sp].get(_period, 0.0) + _earned

    # Load which periods are marked paid
    _cpay_raw = fb_get("/commission_payments") or {}
    comm_paid_set: set = set()
    if isinstance(_cpay_raw, dict):
        for _cpid, _cp in _cpay_raw.items():
            if _cp and isinstance(_cp, dict):
                _name   = _cp.get("salesperson", "").strip().lower()
                _period = _cp.get("period", "").strip()
                if _name and _period:
                    comm_paid_set.add(f"{_name}|{_period}")

    return render_template("payroll.html",
        employee_filter=employee_filter,
        year_filter=year_filter,
        region_filter=region_filter,
        employee_profiles=employee_profiles,
        salaries=salaries,
        commission_by_period=commission_by_period,
        comm_paid_set=list(comm_paid_set))

# ── Payroll Export Routes ─────────────────────────────────────────────────────
@app.route("/payroll/export/csv")
@login_required
def payroll_export_csv():
    import csv
    import io
    raw_sal = fb_get("/balance_sheet_salary") or {}
    salaries = []
    for sid, sdata in (raw_sal.items() if isinstance(raw_sal, dict) else []):
        if isinstance(sdata, dict):
            sdata["firebase_id"] = sid
            salaries.append(sdata)

    # Apply filters
    region_filter = request.args.get("region", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if region_filter:
        salaries = [s for s in salaries if s.get("region") == region_filter]
    if date_from:
        salaries = [s for s in salaries if (s.get("date") or "") >= date_from]
    if date_to:
        salaries = [s for s in salaries if (s.get("date") or "") <= date_to]

    salaries.sort(key=lambda s: s.get("date", ""), reverse=False)

    output = io.StringIO()
    w = csv.writer(output)
    co = company_info()

    # Add company header and blank row
    w.writerow([f"{co.get('name','')} - Payroll Report"])
    w.writerow([])

    def fmt_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    headers = ["Employee", "Date", "Amount", "Region", "Notes", "Status"]
    w.writerow(headers)

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for sal in salaries:
        date_str = sal.get("date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(sal)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda s: s.get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for sal in grouped[year][month][day]:
                    w.writerow([
                        sal.get("employee_name", ""),
                        fmt_date(sal.get("date", "")),
                        f"{_safe_float(sal.get('amount', 0)):.2f}",
                        sal.get("region", ""),
                        sal.get("notes", ""),
                        sal.get("salary_status", "")
                    ])

    from flask import Response
    fname = f"payroll_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/payroll/export/excel")
@login_required
def payroll_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io as _io

    raw_sal = fb_get("/balance_sheet_salary") or {}
    salaries = []
    for sid, sdata in (raw_sal.items() if isinstance(raw_sal, dict) else []):
        if isinstance(sdata, dict):
            sdata["firebase_id"] = sid
            salaries.append(sdata)

    # Apply filters
    region_filter = request.args.get("region", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if region_filter:
        salaries = [s for s in salaries if s.get("region") == region_filter]
    if date_from:
        salaries = [s for s in salaries if (s.get("date") or "") >= date_from]
    if date_to:
        salaries = [s for s in salaries if (s.get("date") or "") <= date_to]

    salaries.sort(key=lambda s: s.get("date", ""), reverse=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    hdr_font = Font(color="FFFFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=13, color="FF0F766E")
    alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    co = company_info()
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value=f"{co.get('name','')} - Payroll Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    def fmt_excel_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    headers = ["Employee", "Date", "Amount", "Region", "Notes", "Status"]
    header_row = 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = ctr

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for sal in salaries:
        date_str = sal.get("date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(sal)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda s: s.get("created_at", ""))

    ri = header_row + 1

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for sal in grouped[year][month][day]:
                    row = [sal.get("employee_name", ""), fmt_excel_date(sal.get("date", "")),
                           _safe_float(sal.get("amount", 0)), sal.get("region", ""), sal.get("notes", ""), sal.get("salary_status", "")]
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        if ri % 2 == 0:
                            cell.fill = alt_fill
                        if ci == 3:
                            cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ri += 1

    col_widths = [25, 14, 14, 18, 30, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import Response
    fname = f"payroll_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={fname}"})

@app.route("/payroll/export/pdf")
@login_required
def payroll_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("payroll"))
    import io as _io

    raw_sal = fb_get("/balance_sheet_salary") or {}
    salaries = []
    for sid, sdata in (raw_sal.items() if isinstance(raw_sal, dict) else []):
        if isinstance(sdata, dict):
            sdata["firebase_id"] = sid
            salaries.append(sdata)

    # Apply filters
    region_filter = request.args.get("region", "")
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")

    if region_filter:
        salaries = [s for s in salaries if s.get("region") == region_filter]
    if date_from:
        salaries = [s for s in salaries if (s.get("date") or "") >= date_from]
    if date_to:
        salaries = [s for s in salaries if (s.get("date") or "") <= date_to]

    salaries.sort(key=lambda s: s.get("date", ""), reverse=False)

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.8*inch, rightMargin=0.8*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    co = company_info()
    elems = []

    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=15,
                              fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                              alignment=1)
    elems.append(Paragraph(f"{co.get('name','')} - Payroll Report", title_s))
    from reportlab.platypus import Spacer
    elems.append(Spacer(1, 0.2*inch))

    def fmt_pdf_date(d):
        if not d or d == "—":
            return ""
        d = str(d)[:10]
        parts = d.split("-")
        return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

    hdrs = ["Employee", "Date", "Amount", "Region", "Notes", "Status"]
    data = [hdrs]

    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=10, wordWrap='CJK')
    group_style = ParagraphStyle("group", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", alignment=1, leading=10, textColor=colors.HexColor("#0F172A"))

    from collections import defaultdict
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for sal in salaries:
        date_str = sal.get("date", "")
        if date_str:
            year = date_str[:4]
            month = date_str[5:7]
            day = date_str[8:10]
            grouped[year][month][day].append(sal)

    # Sort by created_at within each day (first created at top)
    for year in grouped:
        for month in grouped[year]:
            for day in grouped[year][month]:
                grouped[year][month][day].sort(key=lambda s: s.get("created_at", ""))

    for year in sorted(grouped.keys()):
        for month in sorted(grouped[year].keys()):
            for day in sorted(grouped[year][month].keys()):
                for sal in grouped[year][month][day]:
                    data.append([
                        Paragraph(sal.get("employee_name", "—"), cell_style),
                        Paragraph(fmt_pdf_date(sal.get("date", "")), cell_style),
                        Paragraph(f"${_safe_float(sal.get('amount', 0)):,.2f}", cell_style),
                        Paragraph(sal.get("region", "—"), cell_style),
                        Paragraph(sal.get("notes", "—"), cell_style),
                        Paragraph(sal.get("salary_status", "—"), cell_style),
                    ])

    cw = [1.8*inch, 1.0*inch, 1.0*inch, 1.2*inch, 1.5*inch, 1.0*inch]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 9),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,1), (-1,-1), 5),
        ("BOTTOMPADDING", (0,1), (-1,-1), 5),
        ("ALIGN",         (0,1), (-1,-1), "CENTER"),
        ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)

    from flask import Response
    fname = f"payroll_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment;filename={fname}"})

# ── Financial Export Routes ───────────────────────────────────────────────────

def _fmt_date_export(d):
    if not d or d == "—":
        return ""
    d = str(d)[:10]
    parts = d.split("-")
    return f"{parts[1]}-{parts[2]}-{parts[0]}" if len(parts) == 3 else d

def _sanitize_for_export(cell):
    """Remove any "—" or similar placeholder characters from cell values"""
    if cell is None or cell == "":
        return ""
    cell_str = str(cell).strip()
    # Remove en-dash, em-dash, and other dash-like characters
    if cell_str in ("—", "–", "-", "―", ""):
        return ""
    # Also handle if the entire string is just dashes
    if all(c in "—–-―" for c in cell_str):
        return ""
    return cell

def _export_response(buf_or_str, fmt, name_prefix):
    from flask import Response
    ts = datetime.now().strftime("%Y%m%d")
    if fmt == "pdf":
        return Response(buf_or_str.getvalue(), mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment;filename={name_prefix}_{ts}.pdf"})
    if fmt == "excel":
        return Response(buf_or_str.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment;filename={name_prefix}_{ts}.xlsx"})
    return Response(buf_or_str.getvalue(), mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f"attachment;filename={name_prefix}_{ts}.csv"})

def _make_excel(title, headers, rows, col_widths, num_cols=None):
    import openpyxl, io as _io
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = title
    hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    hdr_font = Font(color="FFFFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=12, color="FF0F766E")
    alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=False)
    co = company_info()
    ncols = len(headers)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    tc = ws.cell(row=1, column=1, value=f"{co.get('name','')} – {title}"); tc.font = title_font
    tc.alignment = Alignment(horizontal="center", vertical="center")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr
    num_cols = num_cols or []
    for ri, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0: cell.fill = alt_fill
            if ci in num_cols: cell.number_format = '"$"#,##0.00'
            cell.alignment = ctr
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def _make_pdf(title, headers, data_rows, col_widths_inch):
    import io as _io
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        return None
    buf = _io.BytesIO()
    use_landscape = len(headers) > 7
    pagesize = landscape(A4) if use_landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.4*inch, bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    co = company_info()
    title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=13,
                              fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0F766E"), spaceAfter=4, alignment=1)
    cell_s = ParagraphStyle("C", parent=styles["Normal"], fontSize=7.5, alignment=1, leading=9)
    elems = [Paragraph(f"{co.get('name','')} – {title}", title_s), Spacer(1, 0.15*inch)]
    tdata = [[Paragraph(str(h), ParagraphStyle("H", parent=styles["Normal"],
              fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=1)) for h in headers]]
    for row in data_rows:
        tdata.append([Paragraph(str(v) if v is not None else "—", cell_s) for v in row])
    tbl = Table(tdata, colWidths=[w*inch for w in col_widths_inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 8),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#E2E8F0")),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    return buf

# ── Income exports ────────────────────────────────────────────────────────────
@app.route("/financial/income/export/<fmt>")
@login_required
def financial_income_export(fmt):
    import csv, io as _io
    cur_year  = datetime.now().year
    prev_year = cur_year - 1
    revenue      = fb_get("/balance_sheet_revenue") or {}
    invoices_raw = fb_get("/invoices") or {}
    invoices     = invoices_raw if isinstance(invoices_raw, dict) else {}

    rows = []
    if isinstance(revenue, dict):
        for rid, rdata in revenue.items():
            if not isinstance(rdata, dict): continue
            inv_id = rdata.get("invoice_id")
            if not inv_id or inv_id not in invoices: continue
            inv_data = invoices[inv_id]
            if not isinstance(inv_data, dict): continue
            inv_meta = inv_data.get("meta", {}) or {}
            inv_date = inv_meta.get("invoice_date", "") or rdata.get("date", "")
            # Match financial() — include current and previous year
            try:
                inv_year = int(inv_date[:4])
            except (ValueError, TypeError):
                continue
            if inv_year not in (cur_year, prev_year): continue
            # Recalculate status from actual payment data (same as financial())
            pay_log  = inv_data.get("payment_log", []) or []
            amt_paid = sum(_safe_float(p.get("amount", 0)) for p in pay_log)
            tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in (inv_data.get("tax_payments", []) or []))
            inv_total = _safe_float(inv_meta.get("total", 0)) or (
                _safe_float(inv_meta.get("subtotal", 0)) + _safe_float(inv_meta.get("tax_amount", 0)))
            total_paid = amt_paid + tax_paid
            if total_paid >= (inv_total - 0.01) and inv_total > 0:
                status = "Paid"
            elif total_paid > 0:
                status = "Partial"
            else:
                status = "Unpaid"
            if status not in ("Paid", "Partial"): continue
            coll_date = (max(pay_log, key=lambda p: p.get("date", "")).get("date", rdata.get("date", ""))
                         if pay_log else rdata.get("date", ""))
            method    = ((max(pay_log, key=lambda p: p.get("date", "")).get("method", "") if pay_log else "")
                         or inv_meta.get("payment_method", ""))
            # Convert "—" placeholder to empty string
            if method == "—":
                method = ""
            # Get all project numbers for multi-project invoices
            linked_projects = inv_meta.get("linked_projects", [])
            if isinstance(linked_projects, list) and len(linked_projects) > 0:
                project_nums = []
                for lp in linked_projects:
                    if isinstance(lp, dict):
                        proj_num = lp.get("project_number", "")
                    elif isinstance(lp, str):
                        proj_num = lp
                    else:
                        proj_num = ""
                    if proj_num:
                        project_nums.append(proj_num)
                project_str = ", ".join(project_nums) if project_nums else rdata.get("project_number", "")
            else:
                project_str = rdata.get("project_number", "")
            # Clean method and project_str of any "—" characters
            clean_method = "" if (method == "—" or str(method).strip() == "—") else method
            clean_project = "" if (project_str == "—" or str(project_str).strip() == "—") else project_str
            rows.append([
                _fmt_date_export(inv_date),
                _fmt_date_export(coll_date),
                inv_meta.get("invoice_number", rdata.get("invoice_number", "")),
                rdata.get("client_name", ""),
                clean_project,
                status,
                clean_method,
                _safe_float(inv_meta.get("subtotal", 0)),
                _safe_float(inv_meta.get("tax_amount", 0)),
                inv_total,
                total_paid,
            ])
    rows.sort(key=lambda r: r[0])
    # Clean any "—" placeholders from rows (replace with empty string) using robust sanitization
    rows = [[_sanitize_for_export(cell) for cell in row] for row in rows]
    label   = f"income_{cur_year}_{prev_year}"
    title   = f"Income Report {prev_year}–{cur_year}"
    headers = ["Invoice Date", "Collection Date", "Invoice #", "Client", "Project",
               "Status", "Payment Method", "Subtotal", "Tax", "Invoice Total", "Amount Paid"]
    if fmt == "excel":
        buf = _make_excel(title, headers, rows,
                          [18,18,18,26,20,10,16,12,10,13,13], num_cols=[8,9,10,11])
        return _export_response(buf, "excel", label)
    if fmt == "pdf":
        fmt_rows = [[r[0],r[1],r[2],r[3],r[4],r[5],r[6],
                     f"${r[7]:,.2f}",f"${r[8]:,.2f}",f"${r[9]:,.2f}",f"${r[10]:,.2f}"] for r in rows]
        buf = _make_pdf(title, headers, fmt_rows,
                        [1.1,1.1,1.2,1.6,1.3,0.7,0.9,0.9,0.7,1.0,1.0])
        if not buf: return redirect(url_for("financial"))
        return _export_response(buf, "pdf", label)
    out = _io.StringIO()
    out.write(f"{company_info().get('name','')} – {title}\n\n")
    col_widths = [20, 20, 25, 40, 30, 15, 25, 18, 15, 20, 20]
    header_line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    out.write(header_line + "\n")
    for r in rows:
        row_data = [
            _sanitize_for_export(r[0]).ljust(13), _sanitize_for_export(r[1]).ljust(13),
            _sanitize_for_export(r[2]).ljust(18), _sanitize_for_export(r[3]).ljust(26),
            _sanitize_for_export(r[4]).ljust(20), _sanitize_for_export(r[5]).ljust(10),
            _sanitize_for_export(r[6]).ljust(16),
            f"{_safe_float(r[7]):.2f}".ljust(12), f"{_safe_float(r[8]):.2f}".ljust(10),
            f"{_safe_float(r[9]):.2f}".ljust(13), f"{_safe_float(r[10]):.2f}".ljust(13)
        ]
        out.write("  ".join(row_data) + "\n")
    return _export_response(out, "csv", label)

# ── Expenses exports ──────────────────────────────────────────────────────────
@app.route("/financial/expenses/export/<fmt>")
@login_required
def financial_expenses_export(fmt):
    import csv, io as _io
    search_f  = (request.args.get("search",    "") or "").strip().lower()
    type_f    = (request.args.get("type",      "") or "").strip().lower()
    vendor_f  = (request.args.get("vendor",    "") or "").strip().lower()
    date_from = (request.args.get("date_from", "") or "").strip()
    date_to   = (request.args.get("date_to",   "") or "").strip()

    raw = fb_get("/balance_sheet_expenses") or {}
    expenses = []
    if isinstance(raw, dict):
        for eid, edata in raw.items():
            if isinstance(edata, dict):
                edata["firebase_id"] = eid
                expenses.append(edata)
    expenses.sort(key=lambda e: e.get("date", ""))

    filtered = []
    for e in expenses:
        edate   = e.get("date", "") or ""
        etype   = (e.get("expense_type", "") or "").lower()
        evend   = (e.get("vendor", "") or "").lower()
        ename   = (e.get("expense_name", "") or e.get("description", "") or "").lower()
        ecat    = (e.get("category", "") or "").lower()
        eproj   = (e.get("project_number", "") or "").lower()
        eby     = (e.get("submitted_by_name", "") or e.get("created_by", "") or "").lower()
        if type_f   and etype != type_f:  continue
        if vendor_f and evend != vendor_f: continue
        if date_from and edate < date_from: continue
        if date_to   and edate > date_to:   continue
        if search_f  and not any(search_f in s for s in [ename, evend, etype, ecat, eproj, eby]): continue
        filtered.append(e)

    headers = ["Date", "Expense Type", "Expense Name", "Category", "Vendor", "Project", "Amount", "Submitted By"]
    rows = []
    for e in filtered:
        row = [_fmt_date_export(e.get("date","")),
               _sanitize_for_export(e.get("expense_type","")),
               _sanitize_for_export(e.get("expense_name","") or e.get("description","")),
               _sanitize_for_export(e.get("category","")),
               _sanitize_for_export(e.get("vendor","")),
               _sanitize_for_export(e.get("project_number","")),
               _safe_float(e.get("amount",0)),
               _sanitize_for_export(e.get("submitted_by_name","") or e.get("created_by",""))]
        rows.append(row)

    # Clean any remaining "—" placeholders from rows (replace with empty string) using robust sanitization
    rows = [[_sanitize_for_export(cell) for cell in row] for row in rows]

    label = "expenses"
    if type_f:    label += f"_{type_f}"
    if vendor_f:  label += f"_{vendor_f}"
    if date_from: label += f"_from{date_from}"
    if date_to:   label += f"_to{date_to}"

    if fmt == "excel":
        buf = _make_excel("Expenses", headers, rows, [18,38,55,35,42,28,12,32], num_cols=[7])
        return _export_response(buf, "excel", label)
    if fmt == "pdf":
        fmt_rows = [[r[0],r[1],r[2],r[3],r[4],r[5],f"${r[6]:,.2f}",r[7]] for r in rows]
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io as _io_pdf
        buf = _io_pdf.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=0.3*inch, rightMargin=0.3*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        co = company_info()
        elems = []
        title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=15,
                                  fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                                  alignment=1)
        elems.append(Paragraph(f"{co.get('name','')} — Expenses Report", title_s))
        from reportlab.platypus import Spacer
        elems.append(Spacer(1, 0.2*inch))
        data = [headers]
        cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=10, wordWrap='CJK')
        for fmt_row in fmt_rows:
            data.append([Paragraph(str(cell), cell_style) for cell in fmt_row])
        cw = [1.0*inch, 1.2*inch, 1.8*inch, 1.0*inch, 1.6*inch, 1.2*inch, 0.9*inch, 1.6*inch]
        tbl = Table(data, colWidths=cw, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 9),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,0), 8),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ("TOPPADDING",    (0,1), (-1,-1), 5),
            ("BOTTOMPADDING", (0,1), (-1,-1), 5),
            ("ALIGN",         (0,1), (-1,-1), "CENTER"),
            ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
        ]))
        elems.append(tbl)
        doc.build(elems)
        buf.seek(0)
        if not buf: return redirect(url_for("financial"))
        return _export_response(buf, "pdf", label)
    out = _io.StringIO()
    out.write(f"{company_info().get('name','')} – Expenses Report\n\n")
    col_widths = [18, 25, 40, 25, 40, 25, 18, 30]
    header_line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    out.write(header_line + "\n")
    for r in rows:
        row_data = [
            _sanitize_for_export(r[0]).ljust(12), _sanitize_for_export(r[1]).ljust(18),
            _sanitize_for_export(r[2]).ljust(26), _sanitize_for_export(r[3]).ljust(16),
            _sanitize_for_export(r[4]).ljust(28), _sanitize_for_export(r[5]).ljust(18),
            f"{_safe_float(r[6]):.2f}".ljust(12), _sanitize_for_export(r[7]).ljust(22)
        ]
        out.write("  ".join(row_data) + "\n")
    return _export_response(out, "csv", label)

# ── By Project exports ────────────────────────────────────────────────────────
@app.route("/financial/by-project/export/<fmt>")
@login_required
def financial_byproject_export(fmt):
    import csv, io as _io
    year = int(request.args.get("year", datetime.now().year))
    projects = _load_projects_list()
    inv_raw = fb_get("/invoices") or {}
    invoices = inv_raw if isinstance(inv_raw, dict) else {}
    exp_raw = fb_get("/expenses") or {}
    exp_list = [v for v in (exp_raw.values() if isinstance(exp_raw, dict) else []) if isinstance(v, dict)]

    rows = []
    for p in projects:
        pnum = p.get("project_number", "")
        p_invoiced = p_collected = 0.0
        for inv_id, inv_data in invoices.items():
            if not isinstance(inv_data, dict): continue
            if pnum not in _invoice_linked_projects(inv_data): continue
            inv_meta = inv_data.get("meta", {}) or {}
            inv_date = inv_meta.get("invoice_date", "")
            if not inv_date or inv_date[:4] != str(year): continue
            line_items = inv_data.get("line_items", []) or []
            proj_line = sum(_safe_float(li.get("amount", 0)) for li in line_items
                            if isinstance(li, dict) and str(li.get("project_number","")).strip() == pnum)
            inv_subtotal = _safe_float(inv_meta.get("subtotal", 0))
            inv_tax = _safe_float(inv_meta.get("tax_amount", 0))
            share = (proj_line / inv_subtotal) if inv_subtotal > 0 and proj_line > 0 else (
                1.0 / max(1, len(_invoice_linked_projects(inv_data))))
            pay_log = inv_data.get("payment_log", []) or []
            proj_paid = sum(_safe_float(pl.get("amount", 0)) for pl in pay_log if pl.get("project_number","") == pnum)
            tax_payments = inv_data.get("tax_payments", []) or []
            total_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments)
            proj_tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in tax_payments if tp.get("project_number") == pnum) or share * total_tax_paid
            p_invoiced += proj_line + share * inv_tax
            p_collected += proj_paid + proj_tax_paid
        p_cos = p.get("change_orders") or []
        if not isinstance(p_cos, list): p_cos = list(p_cos.values()) if isinstance(p_cos, dict) else []
        p_contract = _safe_float(p.get("contract_value", 0))
        p_expenses = sum(_safe_float(e.get("amount", 0)) for e in exp_list if e.get("project_number", "") == pnum)
        p_gp = p_collected - p_expenses
        outstanding = max(0.0, p_contract - p_collected)
        margin = min(100.0, (p_collected / p_contract * 100) if p_contract > 0 else 0.0)
        if p_invoiced == 0 and p_collected == 0 and p_expenses == 0: continue
        rows.append([pnum, p.get("project_name",""), p.get("client_name",""), p.get("status",""),
                     p_contract, p_invoiced, p_collected, outstanding, p_expenses, p_gp, f"{margin:.0f}%"])
    rows.sort(key=lambda r: r[0], reverse=True)
    # Clean any "—" placeholders from rows (replace with empty string) using robust sanitization
    rows = [[_sanitize_for_export(cell) for cell in row] for row in rows]
    headers = ["Project #","Project Name","Client","Status","Contract","Invoiced","Collected","Outstanding","Expenses","Gross Profit","Margin"]
    if fmt == "excel":
        num_rows = [[r[0],r[1],r[2],r[3],r[4],r[5],r[6],r[7],r[8],r[9],r[10]] for r in rows]
        buf = _make_excel(f"By Project {year}", headers, num_rows,
                          [20,40,36,20,13,13,13,13,13,13,9], num_cols=[5,6,7,8,9,10])
        return _export_response(buf, "excel", f"by_project_{year}")
    if fmt == "pdf":
        fmt_rows = [[r[0],r[1],r[2],r[3],f"${r[4]:,.2f}",f"${r[5]:,.2f}",
                     f"${r[6]:,.2f}",f"${r[7]:,.2f}",f"${r[8]:,.2f}",f"${r[9]:,.2f}",r[10]] for r in rows]
        buf = _make_pdf(f"By Project {year}", headers, fmt_rows,
                        [0.8,1.6,1.4,0.85,1.0,1.0,1.0,1.0,1.0,1.0,0.7])
        if not buf: return redirect(url_for("financial"))
        return _export_response(buf, "pdf", f"by_project_{year}")
    out = _io.StringIO()
    out.write(f"{company_info().get('name','')} – By Project {year}\n\n")
    col_widths = [20, 40, 35, 20, 18, 18, 18, 18, 18, 18, 14]
    header_line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    out.write(header_line + "\n")
    for r in rows:
        row_data = [
            str(r[0]).ljust(14), str(r[1]).ljust(26), str(r[2]).ljust(22), str(r[3]).ljust(14),
            f"{r[4]:.2f}".ljust(13), f"{r[5]:.2f}".ljust(13), f"{r[6]:.2f}".ljust(13),
            f"{r[7]:.2f}".ljust(13), f"{r[8]:.2f}".ljust(13), f"{r[9]:.2f}".ljust(13), str(r[10]).ljust(9)
        ]
        out.write("  ".join(row_data) + "\n")
    return _export_response(out, "csv", f"by_project_{year}")

# ── A/R Aging exports ─────────────────────────────────────────────────────────
@app.route("/financial/aging/export/<fmt>")
@login_required
def financial_aging_export(fmt):
    import csv, io as _io
    today_d = datetime.now().date()
    inv_raw = fb_get("/invoices") or {}
    BUCKET_LABELS = {"current":"Current","1_30":"1-30 Days","31_60":"31-60 Days","61_90":"61-90 Days","90plus":"90+ Days"}
    buckets = {"current":[],"1_30":[],"31_60":[],"61_90":[],"90plus":[]}
    for inv_id, inv in (inv_raw.items() if isinstance(inv_raw, dict) else []):
        if not isinstance(inv, dict): continue
        meta = inv.get("meta", {}) or {}
        status = meta.get("status", "")
        if status in ("Paid", "Cancelled", "Draft"): continue
        amt_paid = sum(_safe_float(p.get("amount",0)) for p in (inv.get("payment_log",[]) or []))
        tax_paid = sum(_safe_float(tp.get("amount",0)) for tp in (inv.get("tax_payments",[]) or []))
        total = _safe_float(meta.get("subtotal",0)) + _safe_float(meta.get("tax_amount",0))
        balance = total - amt_paid - tax_paid
        if balance <= 0: continue
        inv_date_str = meta.get("invoice_date","") or meta.get("date","")
        net_terms = int(meta.get("net_terms", 30) or 30)
        try:
            inv_date_d = datetime.strptime(str(inv_date_str)[:10], "%Y-%m-%d").date()
            due_date_d = inv_date_d + timedelta(days=net_terms)
            days_overdue = (today_d - due_date_d).days
        except Exception:
            days_overdue = 0; due_date_d = today_d; inv_date_d = today_d
        entry = {"invoice_number": meta.get("invoice_number",""), "client_name": meta.get("client_name",""),
                 "invoice_date": str(inv_date_d), "due_date": str(due_date_d),
                 "net_terms": net_terms, "days_overdue": days_overdue, "balance": balance,
                 "firebase_id": inv_id}
        if days_overdue <= 0: buckets["current"].append(entry)
        elif days_overdue <= 30: buckets["1_30"].append(entry)
        elif days_overdue <= 60: buckets["31_60"].append(entry)
        elif days_overdue <= 90: buckets["61_90"].append(entry)
        else: buckets["90plus"].append(entry)

    all_rows = []
    for key in ["current","1_30","31_60","61_90","90plus"]:
        for e in sorted(buckets[key], key=lambda x: x["days_overdue"], reverse=True):
            days_label = f"Due in {-e['days_overdue']} days" if e["days_overdue"] <= 0 else f"{e['days_overdue']} days"
            net_terms_val = e.get("net_terms", 0)
            net_terms_str = f"Net {net_terms_val}" if net_terms_val and net_terms_val != "—" else ""
            all_rows.append([BUCKET_LABELS[key], e["invoice_number"], e["client_name"],
                             _fmt_date_export(e["invoice_date"]), _fmt_date_export(e["due_date"]),
                             net_terms_str, days_label, e["balance"]])
    # Clean any "—" placeholders from rows (replace with empty string) using robust sanitization
    all_rows = [[_sanitize_for_export(cell) for cell in row] for row in all_rows]
    headers = ["Bucket","Invoice","Client","Invoice Date","Due Date","Net Terms","Days Overdue","Balance Due"]
    if fmt == "excel":
        buf = _make_excel("AR Aging", headers, all_rows, [14,14,32,16,16,10,12,13], num_cols=[7])
        return _export_response(buf, "excel", "ar_aging")
    if fmt == "pdf":
        fmt_rows = [[r[0],r[1],r[2],r[3],r[4],r[5],r[6],f"${r[7]:,.2f}"] for r in all_rows]
        buf = _make_pdf("A/R Aging Report", headers, fmt_rows, [1.0,1.1,2.2,1.0,1.0,0.8,1.0,1.0])
        if not buf: return redirect(url_for("financial"))
        return _export_response(buf, "pdf", "ar_aging")
    out = _io.StringIO()
    out.write(f"{company_info().get('name','')} – A/R Aging Report\n\n")
    col_widths = [18, 20, 45, 18, 18, 15, 18, 20]
    header_line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    out.write(header_line + "\n")
    for r in all_rows:
        row_data = [
            str(r[0]).ljust(12), str(r[1]).ljust(14), str(r[2]).ljust(32), str(r[3]).ljust(12),
            str(r[4]).ljust(12), str(r[5]).ljust(10), str(r[6]).ljust(12), f"{r[7]:.2f}".ljust(13)
        ]
        out.write("  ".join(row_data) + "\n")
    return _export_response(out, "csv", "ar_aging")

# ── Employee Profile API ──────────────────────────────────────────────────────
@app.route("/api/employee-profiles", methods=["GET"])
@login_required
def api_employee_profiles_get():
    return jsonify({"profiles": _load_employee_profiles()})

@app.route("/api/employee-profiles", methods=["POST"])
@login_required
def api_employee_profiles_post():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name is required"}), 400
    profile = {
        "name":         name,
        "title":        (data.get("title") or "").strip(),
        "region":       data.get("region", "Inside America"),
        "hourly_rate":  float(data.get("hourly_rate") or 0),
        "email":        (data.get("email") or "").strip(),
        "created_at":   datetime.now(timezone.utc).isoformat(),
    }
    pid = fb_push("/employee_profiles", profile)
    profile["firebase_id"] = pid
    return jsonify({"success": True, "profile": profile}), 201

@app.route("/api/employee-profiles/<profile_id>", methods=["PATCH"])
@login_required
def api_employee_profiles_patch(profile_id):
    data = request.get_json() or {}
    updates = {}
    for field in ("name", "title", "region", "email"):
        if field in data:
            updates[field] = (data[field] or "").strip()
    if "hourly_rate" in data:
        updates["hourly_rate"] = float(data.get("hourly_rate") or 0)
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    fb_update(f"/employee_profiles/{profile_id}", updates)
    return jsonify({"success": True})

@app.route("/api/employee-profiles/<profile_id>", methods=["DELETE"])
@login_required
def api_employee_profiles_delete(profile_id):
    fb_delete(f"/employee_profiles/{profile_id}")
    return jsonify({"success": True})

@app.route("/api/payroll/salaries", methods=["GET"])
@login_required
def get_salaries():
    salaries_data = fb_get("/balance_sheet_salary") or {}
    salaries = []
    if isinstance(salaries_data, dict):
        for sid, sdata in salaries_data.items():
            if isinstance(sdata, dict):
                sdata["firebase_id"] = sid
                salaries.append(sdata)

    search    = request.args.get("search", "").lower()
    region    = request.args.get("region", "")
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")

    filtered = salaries
    if search:
        filtered = [s for s in filtered if search in s.get("employee_name", "").lower()]
    if region:
        filtered = [s for s in filtered if s.get("region", "") == region]
    if date_from:
        filtered = [s for s in filtered if s.get("date", "") >= date_from]
    if date_to:
        filtered = [s for s in filtered if s.get("date", "") <= date_to]

    return jsonify({"salaries": filtered})

@app.route("/api/payroll/salaries", methods=["POST"])
@login_required
def create_salary():
    data = request.json
    date_str = data.get("date", "")
    try:
        year = int(date_str[:4]) if date_str else datetime.now().year
    except (ValueError, IndexError):
        year = datetime.now().year

    region = data.get("region", "Inside America")
    if region == "USA":
        region = "Inside America"
    elif region == "International":
        region = "Outside America"

    sal_data = {
        "employee_name": data.get("employee_name"),
        "date":          date_str,
        "amount":        float(data.get("amount", 0)),
        "notes":         data.get("notes", ""),
        "region":        region,
        "year":          year,
        "salary_status": data.get("salary_status", "Paid"),
        "created_at":    datetime.now(timezone.utc).isoformat(),
    }
    fb_push("/balance_sheet_salary", sal_data)
    return jsonify({"success": True})

@app.route("/api/payroll/salaries/<sal_id>", methods=["PUT"])
@login_required
def update_salary(sal_id):
    data = request.json
    date_str = data.get("date", "")
    try:
        year = int(date_str[:4]) if date_str else datetime.now().year
    except (ValueError, IndexError):
        year = datetime.now().year
    region = data.get("region", "Inside America")
    if region == "USA":
        region = "Inside America"
    elif region == "International":
        region = "Outside America"
    sal_data = {
        "employee_name": data.get("employee_name"),
        "date":          date_str,
        "amount":        float(data.get("amount", 0)),
        "notes":         data.get("notes", ""),
        "region":        region,
        "year":          year,
        "salary_status": data.get("salary_status", "Paid"),
        "updated_at":    datetime.now(timezone.utc).isoformat(),
    }
    fb_update(f"/balance_sheet_salary/{sal_id}", sal_data)
    return jsonify({"success": True})

@app.route("/api/payroll/salaries/<sal_id>", methods=["DELETE"])
@login_required
def delete_salary(sal_id):
    fb_delete(f"/balance_sheet_salary/{sal_id}")
    return jsonify({"success": True})

# ── Routes: Financial ─────────────────────────────────────────────────────────
@app.route("/financial")
@role_required("financial")
def financial():
    invoices = fb_get("/invoices") or {}
    expenses = fb_get("/balance_sheet_expenses") or {}
    revenue  = fb_get("/balance_sheet_revenue") or {}

    # Get filter parameters from URL
    filter_expense = request.args.get("filter_expense", "")
    selected_year = request.args.get("year", str(datetime.now().year))
    try:
        selected_year = int(selected_year)
    except (ValueError, TypeError):
        selected_year = datetime.now().year

    inv_list  = [v for v in invoices.values() if isinstance(v, dict)] if isinstance(invoices, dict) else []
    exp_list_raw  = []
    if isinstance(expenses, dict):
        for eid, edata in expenses.items():
            if isinstance(edata, dict):
                edata["firebase_id"] = eid
                # Ensure amount is always a float (some old entries might be strings)
                edata["amount"] = _safe_float(edata.get("amount", 0))
                exp_list_raw.append(edata)

    # Group expenses by name and sum amounts
    def group_expenses_by_name(items):
        """Group expenses by name and sum amounts, preserving expense_type and category"""
        grouped = {}
        for item in items:
            exp_name = item.get("expense_name", "") or item.get("description", "—")
            amount = _safe_float(item.get("amount", 0))
            if exp_name not in grouped:
                grouped[exp_name] = {
                    "expense_name": exp_name,
                    "description": exp_name,
                    "expense_type": item.get("expense_type", ""),
                    "category": item.get("category", ""),
                    "amount": 0,
                    "firebase_id": item.get("firebase_id", ""),
                    "date": item.get("date", ""),
                    "vendor": item.get("vendor", ""),
                    "project_number": item.get("project_number", ""),
                    "notes": item.get("notes", ""),
                    "receipt_filename": item.get("receipt_filename", "")
                }
            grouped[exp_name]["amount"] += amount
        return list(grouped.values())

    # Keep BOTH versions: raw for expense tab (all individual entries), grouped for balance sheet
    exp_list_all = sorted(exp_list_raw, key=lambda x: x.get("created_at", "") or x.get("date", ""), reverse=True)  # Sort by creation date (newest first)
    exp_list_grouped_all = group_expenses_by_name(exp_list_raw)  # Consolidated for reference

    # Filter expenses by selected year for balance sheet
    def filter_by_year(items, year_val):
        """Filter items by year from date field"""
        filtered = []
        for item in items:
            date_str = item.get("date", "")
            if date_str:
                try:
                    year = int(date_str[:4])
                    if year == year_val:
                        filtered.append(item)
                except (ValueError, IndexError):
                    pass
        return filtered

    # Create a list for Overview tab KPI cards that filters by present running year (stat_card_year)
    # This is separate from the Balance Sheet view which uses selected_year
    exp_list_for_overview = filter_by_year(exp_list_raw, datetime.now().year)

    # Filter for balance sheet (only selected year)
    exp_list_raw_filtered = filter_by_year(exp_list_raw, selected_year)
    exp_list_filtered = group_expenses_by_name(exp_list_raw_filtered)

    # Filter expenses if filter_expense parameter provided (apply to both raw and grouped lists)
    if filter_expense:
        exp_list_all = [e for e in exp_list_all if (e.get("expense_name", "") or e.get("description", "—")).lower() == filter_expense.lower()]
        # Also filter the grouped list for summary cards
        exp_list_filtered = [e for e in exp_list_filtered if (e.get("expense_name", "") or e.get("description", "—")).lower() == filter_expense.lower()]

    # For balance sheet, use filtered list
    exp_list = exp_list_filtered

    # Collect all unique vendors for dropdown
    all_vendors = sorted(set(e.get("vendor", "") for e in exp_list_all if e.get("vendor", "").strip()))

    rev_list = []
    if isinstance(revenue, dict):
        for rid, rdata in revenue.items():
            if isinstance(rdata, dict):
                # Check if the invoice still exists (not deleted)
                invoice_id = rdata.get("invoice_id")
                if invoice_id and invoice_id not in invoices:
                    # Skip deleted invoices - their revenue entries should not display
                    continue

                rdata["firebase_id"] = rid
                # Older entries (and ones written by the desktop app) only carry
                # 'amount'/'client' rather than 'amount_paid'/'total'/'client_name' —
                # normalize so the template can rely on a consistent field set.
                rdata.setdefault("amount_paid", rdata.get("amount", 0))
                rdata.setdefault("total", rdata.get("amount", 0))
                rdata.setdefault("client_name", rdata.get("client", ""))
                rdata.setdefault("status", "Paid")
                rdata.setdefault("tax_amount", rdata.get("tax_amount", 0))

                # Get linked projects and tax info from actual invoice
                if invoice_id and invoice_id in invoices:
                    inv_data = invoices[invoice_id]
                    if isinstance(inv_data, dict):
                        inv_meta = inv_data.get("meta", {}) or {}
                        # Get all linked projects for multi-project invoices
                        linked = _invoice_linked_projects(inv_data)
                        rdata["linked_projects"] = linked
                        # Update tax_amount and total from invoice meta
                        rdata["tax_amount"] = _safe_float(inv_meta.get("tax_amount", 0))
                        rdata["total"] = _safe_float(inv_meta.get("total", 0))

                rev_list.append(rdata)
    rev_list.sort(key=lambda x: x.get("date", ""), reverse=True)

    # Filter to show only Paid and Partial invoices (present/active invoices only)
    # This matches the Invoicing tab display
    rev_list = [r for r in rev_list if r.get("status") in ["Paid", "Partial"]]

    # Enrich rev_list with current amount_paid from invoice meta (fixes stale revenue entries)
    updated_rev_list = []
    for r in rev_list:
        inv_id = r.get("invoice_id")
        if not inv_id or inv_id not in invoices:
            continue
        inv_data = invoices[inv_id]
        if not isinstance(inv_data, dict):
            continue
        inv_meta = inv_data.get("meta", {}) or {}
        inv_total = _safe_float(inv_meta.get("total", 0))
        amount_paid = _safe_float(inv_meta.get("amount_paid", 0))
        tax_paid = sum(_safe_float(tp.get("amount", 0)) for tp in (inv_data.get("tax_payments", []) or []))
        total_paid_for_inv = amount_paid + tax_paid

        r["amount_paid"]   = amount_paid
        r["tax_paid"]      = tax_paid
        r["total"]         = inv_total
        r["tax_amount"]    = _safe_float(inv_meta.get("tax_amount", 0))
        r["invoice_date"]  = inv_meta.get("invoice_date", "") or r.get("date", "")
        # Collection date = latest payment_log entry date, fallback to revenue record date
        pay_log = inv_data.get("payment_log", []) or []
        if pay_log:
            latest = max(pay_log, key=lambda p: p.get("date", ""))
            r["collection_date"]  = latest.get("date", r.get("date", ""))
            r["payment_method"]   = latest.get("method", "") or inv_meta.get("payment_method", "")
        else:
            r["collection_date"]  = r.get("date", "")
            r["payment_method"]   = inv_meta.get("payment_method", "")

        # Calculate status based on total vs amount_paid for this P&L invoice
        if total_paid_for_inv >= (inv_total - 0.01):
            r["status"] = "Paid"
        elif total_paid_for_inv > 0:
            r["status"] = "Partial"
        else:
            r["status"] = "Unpaid"

        # Only include in P&L if invoice has been paid or is partially paid (not Unpaid)
        if r["status"] in ["Paid", "Partial"]:
            updated_rev_list.append(r)
    rev_list = updated_rev_list
    rev_list.sort(key=lambda x: x.get("invoice_date", "") or x.get("date", ""), reverse=True)

    # Helper to extract year from date string
    def _extract_year_from_date(date_str):
        """Extract year from date string"""
        try:
            return int(date_str[:4])
        except (ValueError, IndexError, TypeError):
            return None

    # Get actual current year for stat cards (always use current system year, not URL filter)
    stat_card_year = datetime.now().year
    prev_year = stat_card_year - 1

    # Get selected year for Balance Sheet filtering (can be different from current year)
    selected_year = request.args.get("year", str(datetime.now().year))
    try:
        selected_year = int(selected_year)
    except (ValueError, TypeError):
        selected_year = datetime.now().year

    # Total collected = sum of payment_log entries by PAYMENT DATE in past & present years
    total_collected = 0.0
    for inv_id, inv_data_r in invoices.items():
        if not isinstance(inv_data_r, dict):
            continue
        for pay in (inv_data_r.get("payment_log", []) or []):
            pay_year = _extract_year_from_date(pay.get("date", ""))
            if pay_year in [stat_card_year, prev_year]:
                total_collected += _safe_float(pay.get("amount", 0))
        for tp in (inv_data_r.get("tax_payments", []) or []):
            tax_year = _extract_year_from_date(tp.get("date", ""))
            if tax_year in [stat_card_year, prev_year]:
                total_collected += _safe_float(tp.get("amount", 0))

    # Filter Income tab: include payment records where EITHER the payment date OR the invoice date
    # falls in current/prev year. This ensures a 2026 invoice paid in 2028 still appears in 2028.
    rev_list = [r for r in rev_list if
        _extract_year_from_date(r.get("date", "")) in [stat_card_year, prev_year]
        or _extract_year_from_date(invoices.get(r.get("invoice_id"), {}).get("meta", {}).get("invoice_date", "")) in [stat_card_year, prev_year]]

    # Sort by date ascending (oldest to newest)
    rev_list = sorted(rev_list, key=lambda r: r.get('invoice_date', '') or r.get('date', ''), reverse=True)

    # Recalculate statuses based on actual payments
    for inv in inv_list:
        inv["meta"]["status"] = _calculate_invoice_status(inv)

    # Build prior-year outstanding invoices list: invoices older than prev_year that still have a balance.
    # These show as a separate section in the Income tab so users can record late payments.
    prior_outstanding_invs = []
    for inv in inv_list:
        meta = inv.get("meta", {}) or {}
        inv_year = _extract_year_from_date(meta.get("invoice_date", ""))
        if inv_year is None or inv_year >= prev_year:
            continue
        inv_total  = _safe_float(meta.get("total", 0))
        inv_paid   = _safe_float(meta.get("amount_paid", 0))
        inv_tax_pd = sum(_safe_float(tp.get("amount", 0)) for tp in (inv.get("tax_payments", []) or []))
        if inv_total > 0 and (inv_paid + inv_tax_pd) < (inv_total - 0.01):
            prior_outstanding_invs.append({
                "invoice_number": meta.get("invoice_number", ""),
                "invoice_date":   meta.get("invoice_date", ""),
                "client":         meta.get("client_name", ""),
                "project_number": meta.get("project_number", ""),
                "total":          inv_total,
                "amount_paid":    inv_paid + inv_tax_pd,
                "outstanding":    inv_total - inv_paid - inv_tax_pd,
                "status":         meta.get("status", "Unpaid"),
                "firebase_id":    inv.get("firebase_id", ""),
            })

    # Filter invoices by current and previous years for stat cards (Overview tab)
    prev_year = stat_card_year - 1
    inv_list_filtered = [i for i in inv_list
                        if _extract_year_from_date(i.get("meta", {}).get("invoice_date", "")) in [stat_card_year, prev_year]]

    total_invoiced    = sum(_safe_float(i.get("meta", {}).get("total", 0)) for i in inv_list_filtered)
    invoiced_count    = len(inv_list_filtered)  # Count of invoices created in current & previous years
    # For Overview KPI: use total_collected (based on payment date) instead of invoice date filtering
    # total_paid is only used for Balance Sheet and Income tab calculations below
    total_tax_paid    = sum(_safe_float(p.get("amount", 0)) for inv in inv_list_filtered for p in inv.get("tax_payments", []))
    # Overview KPI uses total_collected (filtered by payment date, not invoice date)
    total_outstanding = total_invoiced - total_collected
    # Use Overview-filtered expenses (filtered by present running year, not selected year)
    exp_list_year_filtered = group_expenses_by_name(exp_list_for_overview)
    total_expenses    = sum(_safe_float(e.get("amount", 0)) for e in exp_list_year_filtered)
    exp_list_year_filtered_count = len(exp_list_year_filtered)
    # Net profit based on actual collected payments, not invoice dates
    net_profit        = total_collected - total_expenses

    # Calculate previous year data for year-over-year comparison (prev_year already defined above)
    inv_list_prev_year = [i for i in inv_list
                          if _extract_year_from_date(i.get("meta", {}).get("invoice_date", "")) == prev_year]
    prev_year_total_invoiced = sum(_safe_float(i.get("meta", {}).get("total", 0)) for i in inv_list_prev_year)
    prev_year_total_paid = sum(_safe_float(i.get("meta", {}).get("amount_paid", 0)) for i in inv_list_prev_year)
    prev_year_total_tax_paid = sum(_safe_float(p.get("amount", 0)) for inv in inv_list_prev_year for p in inv.get("tax_payments", []))
    prev_year_total_paid += prev_year_total_tax_paid
    prev_year_total_outstanding = prev_year_total_invoiced - prev_year_total_paid
    exp_list_prev_year = [e for e in exp_list if _extract_year_from_date(e.get("date", "")) == prev_year]
    prev_year_total_expenses = sum(_safe_float(e.get("amount", 0)) for e in exp_list_prev_year)

    # Monthly breakdown for chart (last 6 months)
    monthly_revenue  = {}
    monthly_expenses = {}
    for inv in inv_list:
        ds = inv.get("meta", {}).get("invoice_date", "") or ""
        try:
            key = datetime.fromisoformat(ds[:10]).strftime("%b %Y")
            line_paid = _safe_float(inv.get("meta", {}).get("amount_paid", 0))
            tax_paid = sum(_safe_float(p.get("amount", 0)) for p in inv.get("tax_payments", []))
            monthly_revenue[key] = monthly_revenue.get(key, 0) + line_paid + tax_paid
        except Exception:
            pass
    for exp in exp_list:
        ds = exp.get("date", "") or ""
        try:
            key = datetime.fromisoformat(ds[:10]).strftime("%b %Y")
            monthly_expenses[key] = monthly_expenses.get(key, 0) + _safe_float(exp.get("amount", 0))
        except Exception:
            pass

    from dateutil.relativedelta import relativedelta as _rd
    _now = datetime.now()
    all_months = [(_now - _rd(months=i)).strftime("%b %Y") for i in range(5, -1, -1)]
    rev_data   = [monthly_revenue.get(m, 0) for m in all_months]
    exp_data   = [monthly_expenses.get(m, 0) for m in all_months]

    # Annual breakdown for Balance Sheet (selected year or current year)
    selected_year = request.args.get("year", None)
    if selected_year:
        try:
            current_year = int(selected_year)
        except (ValueError, TypeError):
            current_year = _now.year
    else:
        current_year = _now.year
    selected_year = current_year
    annual_revenue = {i: 0.0 for i in range(1, 13)}  # Months 1-12
    annual_expenses = {i: 0.0 for i in range(1, 13)}

    for inv in inv_list:
        for pay in (inv.get("payment_log", []) or []):
            pay_ds = pay.get("date", "") or ""
            try:
                pay_date = datetime.fromisoformat(pay_ds[:10])
                if pay_date.year == current_year:
                    annual_revenue[pay_date.month] += _safe_float(pay.get("amount", 0))
            except Exception:
                pass
        for tax_pay in (inv.get("tax_payments", []) or []):
            tax_ds = tax_pay.get("date", "") or ""
            try:
                tax_date = datetime.fromisoformat(tax_ds[:10])
                if tax_date.year == current_year:
                    annual_revenue[tax_date.month] += _safe_float(tax_pay.get("amount", 0))
            except Exception:
                pass

    for exp in exp_list:
        ds = exp.get("date", "") or ""
        try:
            exp_date = datetime.fromisoformat(ds[:10])
            if exp_date.year == current_year:
                month = exp_date.month
                annual_expenses[month] += _safe_float(exp.get("amount", 0))
        except Exception:
            pass

    # Calculate annual net P/L per month
    annual_netpl = {i: annual_revenue[i] - annual_expenses[i] for i in range(1, 13)}

    # Labor cost (Employees module) — additive figures, do not affect total_expenses/net_profit above
    # Build uid→rate map from /users; Employee Directory rates override when names match
    _all_users        = _load_all_users()
    _dir_rate_by_name = {p.get("name", "").strip().lower(): _safe_float(p.get("hourly_rate", 0))
                         for p in _load_employee_profiles()
                         if _safe_float(p.get("hourly_rate", 0)) > 0}
    rate_by_uid: Dict[str, float] = {}
    for _u in _all_users:
        _uid   = _u.get("firebase_uid")
        _uname = (_u.get("name") or _u.get("display_name") or "").strip().lower()
        _rate  = _safe_float(_u.get("hourly_rate", 0))
        if _uid:
            rate_by_uid[_uid] = _dir_rate_by_name.get(_uname, _rate)
    labor_cost_by_project: Dict[str, float] = {}
    total_labor_cost = 0.0
    for e in _load_time_entries():
        if e.get("status") != "closed":
            continue
        minutes = _safe_float(e.get("duration_minutes", 0))
        cost = (minutes / 60.0) * rate_by_uid.get(e.get("employee_uid"), 0.0)
        total_labor_cost += cost
        pnum_e = e.get("project_number", "")
        if pnum_e:
            labor_cost_by_project[pnum_e] = labor_cost_by_project.get(pnum_e, 0.0) + cost
    net_profit_after_labor = net_profit - total_labor_cost

    # Per-project P&L
    projects_list = _load_projects_list()
    project_pnl = []
    for p in projects_list:
        pnum = p.get("project_number", "")

        # Calculate INVOICED and COLLECTED based on line items and payment_log (not equal split!)
        p_invoiced = 0
        p_collected = 0

        raw_inv = fb_get("/invoices") or {}
        if isinstance(raw_inv, dict):
            for inv_id, inv_data in raw_inv.items():
                if not isinstance(inv_data, dict):
                    continue
                if pnum not in _invoice_linked_projects(inv_data):
                    continue
                # Filter by current running year based on invoice_date
                inv_meta = inv_data.get("meta", {}) or {}
                inv_date_str = inv_meta.get("invoice_date", "")
                inv_year = _extract_year_from_date(inv_date_str)
                if inv_year != stat_card_year:
                    continue

                # Calculate project's portion from actual line items
                line_items = inv_data.get("line_items", []) or []
                project_line_total = 0
                for item in line_items:
                    if isinstance(item, dict):
                        item_proj = str(item.get("project_number", "")).strip()
                        if item_proj == pnum:
                            project_line_total += _safe_float(item.get("amount", 0))

                # Get project's actual payments from payment_log (filtered by project_number)
                payment_log = inv_data.get("payment_log", []) or []
                project_payments = sum(_safe_float(p.get("amount", 0)) for p in payment_log if p.get("project_number", "") == pnum)

                # Add to P&L: invoiced = line items only (exclude tax), collected = actual payments only (exclude tax)
                p_invoiced += project_line_total
                p_collected += project_payments

        # Contract = stored contract_value (already reflects any CO approvals and stage edits)
        # Recalculate base from formula: Base = Total - Sum(Approved COs)
        p_cos = p.get("change_orders") or []
        if not isinstance(p_cos, list):
            p_cos = list(p_cos.values()) if isinstance(p_cos, dict) else []
        p_contract = _safe_float(p.get("contract_value", 0))
        # Calculate approved CO total (include Approved, Invoiced, Paid statuses)
        p_co_total = sum(_safe_float(co.get("amount", 0)) for co in p_cos
                        if co.get("status") in ("Approved", "Invoiced", "Paid"))
        # Base = Total - Sum(Approved COs)
        p_base_contract = max(0.0, p_contract - p_co_total)
        p_not_invoiced = p_contract - p_invoiced
        p_outstanding = max(0.0, p_contract - p_collected)
        p_expenses = sum(_safe_float(e.get("amount",0))                     for e in exp_list if e.get("project_number","") == pnum)
        p_gross_profit = p_collected - p_expenses
        p_labor_cost = labor_cost_by_project.get(pnum, 0.0)
        project_pnl.append({
            "project_number": pnum,
            "project_name":   p.get("project_name",""),
            "client_name":    p.get("client_name",""),
            "status":         p.get("status",""),
            "base_contract":  p_base_contract,
            "contract_value": p_contract,
            "co_total":       p_co_total,
            "invoiced":       p_invoiced,
            "not_invoiced":   p_not_invoiced,
            "paid":           p_collected,
            "outstanding":    p_outstanding,
            "expenses":       p_expenses,
            "gross_profit":   p_gross_profit,
            "labor_cost":     p_labor_cost,
            "net_profit":     p_gross_profit - p_labor_cost,
            "firebase_id":    p.get("firebase_id",""),
        })
    project_pnl.sort(key=lambda x: x["project_number"], reverse=True)

    # Show projects that have a contract value OR any financial activity this year
    project_pnl = [p for p in project_pnl if
                   _safe_float(p.get("contract_value", 0)) > 0
                   or _safe_float(p.get("invoiced", 0)) > 0
                   or _safe_float(p.get("paid", 0)) > 0
                   or _safe_float(p.get("expenses", 0)) > 0]

    # ── Monthly payment drill-down for Balance Sheet ──────────────────────────
    _proj_num_to_id = {p.get("project_number", ""): p.get("firebase_id", "") for p in projects_list}
    _proj_num_to_data = {p.get("project_number", ""): p for p in projects_list}
    monthly_payment_details = {str(i): [] for i in range(1, 13)}
    for _inv in inv_list:
        _inv_id   = _inv.get("firebase_id", "")
        _inv_meta = _inv.get("meta", {}) or {}
        _inv_num  = _inv_meta.get("invoice_number", "")
        _inv_total = _safe_float(_inv_meta.get("total", 0))
        for _pay in (_inv.get("payment_log", []) or []):
            _pay_ds = _pay.get("date", "") or ""
            try:
                _pay_dt = datetime.fromisoformat(_pay_ds[:10])
                if _pay_dt.year == current_year:
                    _mkey = str(_pay_dt.month)
                    _proj_num = _pay.get("project_number", "") or _inv_meta.get("project_number", "")
                    _stage = _pay.get("stage_name", "") or _inv_meta.get("payment_stage", "")

                    # If stage still empty, try to look it up from project payment_stages
                    if not _stage:
                        _stage_idx = _pay.get("stage_index")
                        if _stage_idx is None:
                            _stage_idx = _inv_meta.get("payment_stage_index")
                        if _stage_idx is not None:
                            try:
                                _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                                _proj_data = _proj_num_to_data.get(_proj_num, {})
                                _p_stages = _proj_data.get("payment_stages", [])
                                if isinstance(_p_stages, list) and 0 <= _stage_idx < len(_p_stages):
                                    _stage = _p_stages[_stage_idx].get("name", f"Stage {_stage_idx + 1}")
                            except (ValueError, TypeError, IndexError):
                                if _stage_idx is not None:
                                    try:
                                        _stage_idx = int(_stage_idx)
                                        _stage = f"Stage {_stage_idx + 1}"
                                    except (ValueError, TypeError):
                                        pass

                    # Final fallback: extract stage from invoice line items description
                    if not _stage:
                        for _li in (_inv.get("line_items", []) or []):
                            if not isinstance(_li, dict):
                                continue
                            _li_proj = _li.get("project_number", "")
                            if _li_proj and _li_proj != _proj_num:
                                continue
                            _li_desc = _li.get("description", "") or ""
                            if " — " in _li_desc:
                                _stage = _li_desc.split(" — ", 1)[1].strip()
                            elif _li_desc:
                                _stage = _li_desc
                            if _stage:
                                break

                    # If stage is generic "Stage N", upgrade to "Installment N of Total"
                    _stage_label = _stage or "—"
                    if _stage_label and re.match(r'^Stage\s+\d+$', _stage_label, re.IGNORECASE):
                        _proj_data_s = _proj_num_to_data.get(_proj_num, {})
                        _all_stages  = _proj_data_s.get("payment_stages", [])
                        if isinstance(_all_stages, list) and len(_all_stages) > 0:
                            try:
                                _snum = int(_stage_label.split()[-1])
                                if 1 <= _snum <= len(_all_stages):
                                    _sname = (_all_stages[_snum - 1].get("name", "") or "").strip()
                                    _stage_label = _sname if _sname else f"Installment {_snum} of {len(_all_stages)}"
                                else:
                                    _stage_label = f"Installment {_snum} of {len(_all_stages)}"
                            except (ValueError, TypeError):
                                pass

                    monthly_payment_details[_mkey].append({
                        "project_number": _proj_num,
                        "project_id":     _proj_num_to_id.get(_proj_num, ""),
                        "invoice_id":     _inv_id,
                        "invoice_number": _inv_num,
                        "stage":          _stage_label,
                        "total_amount":   _inv_total,
                        "paid_amount":    _safe_float(_pay.get("amount", 0)),
                        "paid_date":      _pay_ds,
                    })
            except Exception:
                pass
        for _tpay in (_inv.get("tax_payments", []) or []):
            _tpay_ds = _tpay.get("date", "") or ""
            try:
                _tpay_dt = datetime.fromisoformat(_tpay_ds[:10])
                if _tpay_dt.year == current_year:
                    monthly_payment_details[str(_tpay_dt.month)].append({
                        "project_number": "TAX",
                        "project_id":     "",
                        "invoice_id":     _inv_id,
                        "invoice_number": _inv_num,
                        "stage":          "Tax Payment",
                        "total_amount":   _safe_float(_inv_meta.get("tax_amount", 0)),
                        "paid_amount":    _safe_float(_tpay.get("amount", 0)),
                        "paid_date":      _tpay_ds,
                    })
            except Exception:
                pass
    # Merge multiple partial payments for the same invoice+project into one row
    for _mk in monthly_payment_details:
        _merged: dict = {}
        for _row in monthly_payment_details[_mk]:
            _key = (_row["invoice_id"], _row["project_number"])
            if _key in _merged:
                _merged[_key]["paid_amount"] += _row["paid_amount"]
                # keep latest payment date
                if _row["paid_date"] > _merged[_key]["paid_date"]:
                    _merged[_key]["paid_date"] = _row["paid_date"]
            else:
                _merged[_key] = dict(_row)
        monthly_payment_details[_mk] = sorted(_merged.values(), key=lambda x: x.get("paid_date", ""))

    # ── Chart data for overview pie charts ────────────────────────────────────
    inv_status_counts = {}
    for i in inv_list:
        st = i.get("meta", {}).get("status") or "Draft"
        inv_status_counts[st] = inv_status_counts.get(st, 0) + 1

    exp_cats = {}
    for e in exp_list:
        cat = e.get("category", "Other") or "Other"
        exp_cats[cat] = exp_cats.get(cat, 0) + _safe_float(e.get("amount", 0))

    # Load salaries data for Balance Sheet
    salaries = fb_get("/balance_sheet_salary") or {}
    salaries_domestic_raw = []
    salaries_international_raw = []
    total_salaries = 0.0

    if isinstance(salaries, dict):
        for sid, sdata in salaries.items():
            if isinstance(sdata, dict):
                sdata["firebase_id"] = sid
                sal_amount = _safe_float(sdata.get("amount", 0))
                total_salaries += sal_amount
                region = sdata.get("region", "").lower()
                if "international" in region or "outside" in region:
                    salaries_international_raw.append(sdata)
                else:
                    salaries_domestic_raw.append(sdata)

    # Balance Sheet: salary totals filtered by selected year (current_year)
    bs_sal_dom_raw = filter_by_year(list(salaries_domestic_raw), current_year)
    bs_sal_int_raw = filter_by_year(list(salaries_international_raw), current_year)
    bs_total_salaries = sum(_safe_float(s.get("amount", 0)) for s in bs_sal_dom_raw + bs_sal_int_raw)

    # Filter salaries by present running year for KPI cards (stat_card_year is always current year)
    salaries_domestic_raw = filter_by_year(salaries_domestic_raw, stat_card_year)
    salaries_international_raw = filter_by_year(salaries_international_raw, stat_card_year)
    total_salaries = sum(_safe_float(s.get("amount", 0)) for s in salaries_domestic_raw + salaries_international_raw)

    # ── Commission paid — treated as a cost on the balance sheet ────────────
    _cp_bs_raw = fb_get("/commission_payments") or {}
    bs_total_commission = 0.0
    total_commission_paid = 0.0
    monthly_commission_details: Dict[str, list] = {str(m): [] for m in range(1, 13)}
    if isinstance(_cp_bs_raw, dict):
        for _cpid, _cp in _cp_bs_raw.items():
            if not _cp or not isinstance(_cp, dict):
                continue
            _amt     = _safe_float(_cp.get("amount", 0))
            _paid_at = (_cp.get("paid_at") or "")
            try:
                _dt  = datetime.fromisoformat(_paid_at[:19])
                _yr  = _dt.year
                _mon = str(_dt.month)
            except Exception:
                _yr, _mon = 0, "0"
            if _yr == current_year:
                bs_total_commission += _amt
                monthly_commission_details[_mon].append({
                    "salesperson": _cp.get("salesperson", ""),
                    "period":      _cp.get("period", ""),
                    "amount":      _amt,
                    "paid_at":     _paid_at[:10],
                })
            if _yr == stat_card_year:
                total_commission_paid += _amt

    # Recalculate net profit now that total_salaries and commissions are known
    # Using total_collected (based on payment date) instead of total_paid (invoice date)
    net_profit             = total_collected - total_expenses - total_salaries - total_commission_paid
    net_profit_after_labor = net_profit - total_labor_cost

    # Count unique employees with salary entries in present running year
    employees_with_salaries = set()
    for s in salaries_domestic_raw + salaries_international_raw:
        emp_name = s.get("employee_name") or s.get("name")
        if emp_name:
            employees_with_salaries.add(emp_name)
    employee_count = len(employees_with_salaries)

    # Group salaries by name, sum amounts, track latest date; sort by date desc
    def group_by_name(items, entries_tracking=None):
        grouped = {}
        if entries_tracking is None:
            entries_tracking = {}
        for item in items:
            name = item.get("employee_name") or item.get("name", "—")
            amount = _safe_float(item.get("amount", 0))
            date = item.get("date", "")
            if name not in grouped:
                grouped[name] = {"name": name, "amount": 0, "date": date}
                entries_tracking[name] = []
            else:
                if date > grouped[name].get("date", ""):
                    grouped[name]["date"] = date
            grouped[name]["amount"] += amount
            entries_tracking[name].append({
                "date": date,
                "amount": amount,
                "notes": item.get("notes", "")
            })
        return sorted(grouped.values(), key=lambda x: x.get("date", ""), reverse=True), entries_tracking

    salary_entries_domestic = {}
    salary_entries_international = {}
    salaries_domestic, salary_entries_domestic = group_by_name(salaries_domestic_raw, salary_entries_domestic)
    salaries_international, salary_entries_international = group_by_name(salaries_international_raw, salary_entries_international)

    # ── Monthly salary details for drill-down (use BS-filtered salaries for selected year) ──
    # Build role lookup by username
    _role_lookup: Dict[str, str] = {}
    for _u in _load_all_users():
        _uname = (_u.get("username") or "").strip()
        if _uname:
            _role_lookup[_uname.lower()] = normalize_role(_u.get("role", "")).capitalize()

    monthly_salary_details = {str(i): [] for i in range(1, 13)}
    for _sal in bs_sal_dom_raw + bs_sal_int_raw:
        _ds = (_sal.get("date") or "")[:10]
        try:
            _d = datetime.fromisoformat(_ds)
            _emp_name = _sal.get("employee_name") or _sal.get("name") or "—"
            monthly_salary_details[str(_d.month)].append({
                "name":   _emp_name,
                "role":   _role_lookup.get(_emp_name.lower(), "—"),
                "region": "Inside America" if _sal in bs_sal_dom_raw else "Outside America",
                "amount": _safe_float(_sal.get("amount", 0)),
                "date":   _ds,
            })
        except Exception:
            pass

    # Calculate totals for Balance Sheet (use selected year data, not stat_card_year)
    bs_total_revenue = sum(annual_revenue.values())
    bs_total_expenses = sum(annual_expenses.values())
    total_revenue = total_collected  # kept for legacy/overview usage - uses payment date

    # Load custom expense categories from Firebase
    custom_categories = fb_get("/custom_categories") or {}
    expense_types = custom_categories.get("expense_type", []) if isinstance(custom_categories.get("expense_type"), list) else []
    categories_by_type = custom_categories.get("Categories", {}) if isinstance(custom_categories.get("Categories"), dict) else {}
    expense_names_by_category = custom_categories.get("expense_names", {}) if isinstance(custom_categories.get("expense_names"), dict) else {}

    # Provide default values if empty (matching reference structure)
    if not expense_types:
        expense_types = [
            "O & M (Operations & Maintenance)",
            "Capital Expenses",
            "Other Expenses"
        ]

    if not categories_by_type:
        categories_by_type = {
            "O & M (Operations & Maintenance)": [
                "Facilities & Utilities",
                "Office & Admin Overhead",
                "Engineering Software & IT",
                "Salaries, Labor & Related Costs",
                "Professional Services",
                "Insurance & Compliance",
                "Travel, Site Visits & Vehicles",
                "Marketing & Business Development",
                "Training, Licensure & Development",
                "Safety & Field Supplies",
                "Miscellaneous O & M"
            ],
            "Capital Expenses": [
                "Computer & Office Equipment",
                "Field & Inspection Equipment",
                "Furniture & Fixtures",
                "Vehicles",
                "Software (Capitalized)",
                "Leasehold Improvements",
                "Accumulated Depreciation"
            ],
            "Other Expenses": [
                "Salary/Bonuses",
                "Tax Expenses/Tax Deductions",
                "Medical/Benefits",
                "Meals & Entertainment",
                "Donations",
                "Bank Charges",
                "Contingency Funds",
                "Unexpected Costs"
            ]
        }

    if not expense_names_by_category:
        expense_names_by_category = {
            "Facilities & Utilities": ["Office rent or co-working space fees","Utilities (electricity, water, gas)","Internet service","Trash & cleaning services","Property taxes (for office, if applicable)","Office repairs & maintenance (HVAC, lights, minor repairs)"],
            "Office & Admin Overhead": ["Office supplies (paper, pens, notebooks, printer ink)","Printer/plotter maintenance & paper","Postage & shipping (documents, contracts, samples)","Bank fees & merchant processing fees","Software: Microsoft 365 / Google Workspace","Software: PDF tools (Bluebeam, Adobe, etc.)","Software: Password manager","Software: Others","Cloud storage (Dropbox, Google Drive, OneDrive)"],
            "Engineering Software & IT": ["Engineering software: SAP2000 / ETABS / STAAD / RAM / RISA","Engineering software: Others","CAD/BIM tools: AutoCAD, Civil 3D, Revit","License/maintenance fees for all software","IT support services","Computer maintenance & small repairs","Antivirus, backup services, security tools"],
            "Salaries, Labor & Related Costs": ["Owner draw/salary","Employee salaries & wages","Overtime or temporary staff","Payroll taxes paid by the company","Employee benefits: Health insurance","Employee benefits: Retirement plan contributions","Employee benefits: Paid time off costs","Payments to subcontract engineers, drafters"],
            "Professional Services": ["Accounting & bookkeeping fees","Tax preparation and consulting","Legal services (contracts, company setup)","Business consulting or coaching services","Registered agent fees (if applicable)"],
            "Insurance & Compliance": ["Professional liability / Errors & Omissions (E&O) insurance","General liability insurance","Business owner's policy (BOP)","Workers' comp insurance","Commercial auto insurance","License renewals (PE license, SE license)","Business license renewals","Memberships"],
            "Travel, Site Visits & Vehicles": ["Mileage (personal vehicle for business)","Fuel costs (company vehicles)","Parking fees & tolls","Vehicle maintenance","Airfare, hotels for out-of-town site visits","Rental cars or rideshare for business trips","Meals while traveling for business"],
            "Marketing & Business Development": ["Website hosting and domain expenses","Website maintenance & updates","Graphic design (logo, templates, brochures)","Online ads (Google, LinkedIn, Facebook)","Printing of business cards, brochures, banners","Sponsorships of events","Client entertainment (dinners, coffee meetings)"],
            "Training, Licensure & Development": ["Continuing education (PDH hours, webinars)","Training courses (technical or business)","Books, codes, and standards","Exam fees for additional licenses"],
            "Safety & Field Supplies": ["PPE: hard hats, safety vests, glasses, gloves, boots","Field tools for inspections","Calibration of field instruments","First-aid kits and safety equipment"],
            "Miscellaneous O & M": ["Subscriptions: LinkedIn Premium","Subscriptions: Industry journals","Project management tools","Document management tools or e-signature services"],
            "Computer & Office Equipment": ["Laptops","Desktops","Monitors","Printers/Scanners","Servers","Networking Equipment"],
            "Field & Inspection Equipment": ["Survey Equipment","Testing Equipment","Measurement Tools","Safety Equipment","Inspection Devices"],
            "Furniture & Fixtures": ["Office Desks","Chairs","Filing Cabinets","Shelving Units","Conference Room Furniture"],
            "Vehicles": ["Company Cars","Trucks","Vans","Heavy Equipment","Vehicle Accessories"],
            "Software (Capitalized)": ["Engineering Software License","ERP System","CRM System","Database Software","Custom Software Development"],
            "Leasehold Improvements": ["Office Renovations","Electrical Work","Plumbing Improvements","HVAC Installation","Security Systems"],
            "Accumulated Depreciation": ["Depreciation Expense - Computers","Depreciation Expense - Office Equipment","Depreciation Expense - Vehicles","Accumulated Depreciation"],
            "Salary/Bonuses": ["Employee Salary","Manager Salary","Executive Salary","Performance Bonus","Year-end Bonus","Commission Payments","Incentive Payments"],
            "Tax Expenses/Tax Deductions": ["Federal Income Tax","Tax Deduction","Payroll Tax","Sales Tax","Property Tax","Business Tax"],
            "Medical/Benefits": ["Health Insurance Premiums","Dental Insurance","Vision Insurance","Retirement Contributions","Life Insurance","Disability Insurance","Wellness Programs"],
            "Meals & Entertainment": ["Client Meals","Business Lunches","Team Dinners","Conference Meals","Entertainment Expenses","Team Building Events"],
            "Donations": ["Charitable Donations","Community Sponsorships","Educational Donations","Non-profit Contributions","Event Sponsorships"],
            "Bank Charges": ["Monthly Account Fees","Transaction Fees","Wire Transfer Fees","Credit Card Processing Fees","Check Printing Fees","Overdraft Fees"],
            "Contingency Funds": ["Emergency Funds","Reserve Funds","Project Contingency","Operational Reserve","Risk Management Fund"],
            "Unexpected Costs": ["Emergency Repairs","Unplanned Maintenance","Price Increases","Regulatory Changes","Market Fluctuations"]
        }

    # Flat list of all categories for expense filter dropdown
    all_categories = []
    for cats in categories_by_type.values():
        for cat in cats:
            if cat not in all_categories:
                all_categories.append(cat)

    # exp_list is already grouped by group_expenses_by_name(), use it directly
    # Build expense_entries_by_name for drill-down from raw filtered data
    expense_entries_by_name = {}
    for e in exp_list_raw_filtered:
        name = e.get("expense_name", "") or e.get("description", "—") or "—"
        if name not in expense_entries_by_name:
            expense_entries_by_name[name] = []
        expense_entries_by_name[name].append({
            "date": e.get("date", ""),
            "amount": _safe_float(e.get("amount", 0)),
            "category": e.get("category", ""),
            "vendor": e.get("vendor", ""),
            "notes": e.get("notes", "")
        })
    # Use the already-grouped exp_list as expenses_grouped, sorted by date
    expenses_grouped = sorted(exp_list, key=lambda x: x.get("date", ""), reverse=True)

    # Get list of available years from invoices and expenses
    available_years = set()
    for inv in inv_list:
        ds = inv.get("meta", {}).get("invoice_date", "") or ""
        try:
            year = datetime.fromisoformat(ds[:10]).year
            available_years.add(year)
        except Exception:
            pass
    for exp in exp_list:
        ds = exp.get("date", "") or ""
        try:
            year = datetime.fromisoformat(ds[:10]).year
            available_years.add(year)
        except Exception:
            pass

    # Add a broader range of years (past 15 years + future 10 years)
    _real_now_year = _now.year
    for year in range(_real_now_year - 15, _real_now_year + 11):
        available_years.add(year)

    available_years = sorted(list(available_years), reverse=True)

    # ── Accounts Receivable Aging ─────────────────────────────────────────────
    today_d = datetime.now().date()
    aging_buckets = {"current": [], "1_30": [], "31_60": [], "61_90": [], "90plus": []}
    for inv in inv_list:
        m = inv.get("meta", {}) or {}
        status = m.get("status", "Draft")
        if status not in ("Sent", "Viewed", "Partial", "Overdue"):
            continue
        total   = _safe_float(m.get("total", 0))
        paid    = _safe_float(m.get("amount_paid", 0))
        balance = total - paid
        if balance <= 0.01:
            continue
        due_str = m.get("due_date", "")
        try:
            due_d = datetime.fromisoformat(due_str[:10]).date()
            days_overdue = (today_d - due_d).days
        except Exception:
            days_overdue = 0
        entry = {
            "invoice_number": m.get("invoice_number", ""),
            "client_name":    m.get("client_name", ""),
            "invoice_date":   m.get("invoice_date", ""),
            "due_date":       due_str,
            "net_terms":      m.get("net_terms", ""),
            "days_overdue":   days_overdue,
            "balance":        balance,
            "status":         status,
            "firebase_id":    inv.get("firebase_id", ""),
        }
        if days_overdue <= 0:
            aging_buckets["current"].append(entry)
        elif days_overdue <= 30:
            aging_buckets["1_30"].append(entry)
        elif days_overdue <= 60:
            aging_buckets["31_60"].append(entry)
        elif days_overdue <= 90:
            aging_buckets["61_90"].append(entry)
        else:
            aging_buckets["90plus"].append(entry)

    # ── ALL YEARS totals (for Outstanding A/R calculation) ──
    # Total invoiced from ALL years (excluding Draft invoices)
    total_invoiced_all_years = sum(
        _safe_float(i.get("meta", {}).get("total", 0))
        for i in inv_list
        if i.get("meta", {}).get("status", "Draft") != "Draft"
    )

    # Total collected from ALL years (all payments, excluding payments on Draft invoices)
    total_collected_all_years = 0.0
    for inv_id, inv_data_r in invoices.items():
        if not isinstance(inv_data_r, dict):
            continue
        # Skip Draft invoices
        inv_status = inv_data_r.get("meta", {}).get("status", "Draft")
        if inv_status == "Draft":
            continue
        # Count ALL payments from ALL years (only non-Draft invoices)
        for pay in (inv_data_r.get("payment_log", []) or []):
            total_collected_all_years += _safe_float(pay.get("amount", 0))
        for tp in (inv_data_r.get("tax_payments", []) or []):
            total_collected_all_years += _safe_float(tp.get("amount", 0))

    # Outstanding A/R = All Year Invoiced (non-Draft) - All Year Collected (from non-Draft)
    overview_outstanding = total_invoiced_all_years - total_collected_all_years

    aging_totals = {k: sum(e["balance"] for e in v) for k, v in aging_buckets.items()}
    aging_total_outstanding = sum(aging_totals.values())
    outstanding_count = sum(len(v) for v in aging_buckets.values())  # Count of all open invoices

    # ── Year-filtered A/R for Balance Sheet (only invoices from selected_year) ──
    bs_aging_buckets = {"current": [], "1_30": [], "31_60": [], "61_90": [], "90plus": []}
    for _inv in inv_list:
        _m = _inv.get("meta", {}) or {}
        _inv_date = _m.get("invoice_date", "") or ""
        try:
            if int(_inv_date[:4]) != current_year:
                continue
        except (ValueError, TypeError):
            continue
        _status = _m.get("status", "Draft")
        if _status not in ("Sent", "Viewed", "Partial", "Overdue"):
            continue
        _balance = _safe_float(_m.get("total", 0)) - _safe_float(_m.get("amount_paid", 0))
        if _balance <= 0.01:
            continue
        _due_str = _m.get("due_date", "")
        try:
            _days_ov = (today_d - datetime.fromisoformat(_due_str[:10]).date()).days
        except Exception:
            _days_ov = 0
        _entry = {
            "invoice_number": _m.get("invoice_number", ""),
            "client_name":    _m.get("client_name", ""),
            "invoice_date":   _inv_date,
            "due_date":       _due_str,
            "net_terms":      _m.get("net_terms", ""),
            "days_overdue":   _days_ov,
            "balance":        _balance,
            "status":         _status,
            "firebase_id":    _inv.get("firebase_id", ""),
        }
        if _days_ov <= 0:       bs_aging_buckets["current"].append(_entry)
        elif _days_ov <= 30:    bs_aging_buckets["1_30"].append(_entry)
        elif _days_ov <= 60:    bs_aging_buckets["31_60"].append(_entry)
        elif _days_ov <= 90:    bs_aging_buckets["61_90"].append(_entry)
        else:                   bs_aging_buckets["90plus"].append(_entry)
    bs_aging_totals = {k: sum(e["balance"] for e in v) for k, v in bs_aging_buckets.items()}
    bs_aging_total_outstanding = sum(bs_aging_totals.values())

    # ── Monthly drill-down detail blocks (needs aging_buckets + salaries_domestic) ──
    monthly_expense_details = {str(i): [] for i in range(1, 13)}
    for _exp in exp_list_all:
        _ds = (_exp.get("date") or "")[:10]
        try:
            _d = datetime.fromisoformat(_ds)
            if _d.year == current_year:
                monthly_expense_details[str(_d.month)].append({
                    "name":     _exp.get("expense_name") or _exp.get("description") or "",
                    "category": _exp.get("category") or _exp.get("expense_type") or "",
                    "amount":   _safe_float(_exp.get("amount", 0)),
                    "date":     _ds,
                })
        except Exception:
            pass

    # Group outstanding by invoice_date month (invoices issued that month)
    monthly_outstanding_details = {str(i): [] for i in range(1, 13)}
    # Group outstanding by due_date month (invoices DUE that month)
    monthly_due_details = {str(i): [] for i in range(1, 13)}
    for _bucket in aging_buckets.values():
        for _entry in _bucket:
            # By invoice date
            _ds = (_entry.get("invoice_date") or "")[:10]
            try:
                _d = datetime.fromisoformat(_ds)
                if _d.year == current_year:
                    monthly_outstanding_details[str(_d.month)].append(_entry)
            except Exception:
                pass
            # By due date
            _ds2 = (_entry.get("due_date") or "")[:10]
            try:
                _d2 = datetime.fromisoformat(_ds2)
                if _d2.year == current_year:
                    monthly_due_details[str(_d2.month)].append(_entry)
            except Exception:
                pass

    today_date = datetime.now().strftime("%Y-%m-%d")
    active_tab = request.args.get("tab", "overview")
    _valid_fin_tabs = {'overview', 'income', 'expenses', 'by-project', 'balance-sheet', 'aging', 'commission'}
    if active_tab not in _valid_fin_tabs:
        active_tab = 'overview'

    # ── Commission Payable summary for Finance tab ────────────────────────────
    _all_quotes_fin = fb_get("/job_forms") or {}
    _all_proj_fin   = fb_get("/projects") or {}
    _proj_st_fin: Dict[str, str] = {}
    if isinstance(_all_proj_fin, dict):
        for _pid, _pd in _all_proj_fin.items():
            if _pd and isinstance(_pd, dict):
                _proj_st_fin[_pid] = _pd.get("status", "")

    _sales_users_fin: Dict[str, dict] = {}
    for _u in _load_all_users():
        if normalize_role(_u.get("role", "")) == "sales":
            _uname = (_u.get("username") or "").strip()
            if _uname:
                _sales_users_fin[_uname] = {
                    "commission_rate": _safe_float(_u.get("commission_rate", 0)),
                    "employee_type":   _u.get("employee_type", ""),
                    "email":           _u.get("email", ""),
                }

    _CONV_FIN = {"Converted", "Invoiced"}
    _fin_sp_totals: Dict[str, dict] = {}
    if isinstance(_all_quotes_fin, dict):
        for _fid, _fdata in _all_quotes_fin.items():
            if not _fdata or not isinstance(_fdata, dict):
                continue
            _sp = (_fdata.get("salesperson") or "").strip()
            if not _sp or _sp not in _sales_users_fin:
                continue
            _rate = _sales_users_fin[_sp]["commission_rate"]
            if not _rate:
                continue
            _linked = _fdata.get("linked_project_id", "")
            _is_conv = _fdata.get("status", "") in _CONV_FIN or bool(_linked)
            if not _is_conv:
                continue
            if _linked and _proj_st_fin.get(_linked) == "Cancelled":
                continue
            _qval = _safe_float(_fdata.get("total", 0))
            _period = (_fdata.get("date") or "")[:7]
            if _sp not in _fin_sp_totals:
                _fin_sp_totals[_sp] = {
                    "name":            _sp,
                    "email":           _sales_users_fin[_sp]["email"],
                    "employee_type":   _sales_users_fin[_sp]["employee_type"],
                    "commission_rate": _rate,
                    "total_revenue":   0.0,
                    "total_earned":    0.0,
                    "total_paid":      0.0,
                    "periods":         {},
                }
            _fin_sp_totals[_sp]["total_revenue"] += _qval
            _fin_sp_totals[_sp]["total_earned"]  += _qval * _rate / 100
            if _period:
                if _period not in _fin_sp_totals[_sp]["periods"]:
                    _fin_sp_totals[_sp]["periods"][_period] = {"earned": 0.0, "paid": False}
                _fin_sp_totals[_sp]["periods"][_period]["earned"] += _qval * _rate / 100

    # Apply payment records
    _comm_pay_fin = fb_get("/commission_payments") or {}
    if isinstance(_comm_pay_fin, dict):
        for _cpid, _cp in _comm_pay_fin.items():
            if not _cp or not isinstance(_cp, dict):
                continue
            _sp  = _cp.get("salesperson", "")
            _per = _cp.get("period", "")
            _amt = _safe_float(_cp.get("amount", 0))
            if _sp in _fin_sp_totals:
                _fin_sp_totals[_sp]["total_paid"] += _amt
                if _per in _fin_sp_totals[_sp]["periods"]:
                    _fin_sp_totals[_sp]["periods"][_per]["paid"] = True

    for _s in _fin_sp_totals.values():
        _s["outstanding"] = max(_s["total_earned"] - _s["total_paid"], 0.0)

    commission_summary = sorted(_fin_sp_totals.values(), key=lambda x: x["outstanding"], reverse=True)
    commission_total_earned    = sum(s["total_earned"]  for s in commission_summary)
    commission_total_paid      = sum(s["total_paid"]    for s in commission_summary)
    commission_total_outstanding = sum(s["outstanding"] for s in commission_summary)

    invoiced_years = f"{prev_year} & {stat_card_year}"

    return render_template("financial.html",
        total_invoiced=total_invoiced,
        invoiced_count=invoiced_count,
        invoiced_years=invoiced_years,
        total_outstanding=total_outstanding,
        outstanding_count=outstanding_count,
        total_expenses=total_expenses,
        prev_year_total_invoiced=prev_year_total_invoiced,
        prev_year_total_paid=prev_year_total_paid,
        prev_year_total_outstanding=prev_year_total_outstanding,
        prev_year_total_expenses=prev_year_total_expenses,
        exp_list_year_filtered_count=exp_list_year_filtered_count,
        total_revenue=total_revenue,
        bs_total_revenue=bs_total_revenue,
        bs_total_expenses=bs_total_expenses,
        bs_total_salaries=bs_total_salaries,
        total_salaries=total_salaries,
        employee_count=employee_count,
        net_profit=net_profit,
        total_labor_cost=total_labor_cost,
        net_profit_after_labor=net_profit_after_labor,
        stat_card_year=stat_card_year,
        prev_year=prev_year,
        chart_labels=json.dumps(all_months),
        chart_revenue=json.dumps(rev_data),
        chart_expenses=json.dumps(exp_data),
        annual_revenue=annual_revenue,
        annual_expenses=annual_expenses,
        annual_netpl=annual_netpl,
        expenses=exp_list_all,
        expenses_filtered=exp_list,
        filter_expense=filter_expense,
        selected_year=selected_year,
        rev_list=rev_list,
        prior_outstanding_invs=prior_outstanding_invs,
        total_collected=total_collected,
        overview_outstanding=overview_outstanding,
        projects=projects_list,
        project_pnl=project_pnl,
        salaries_domestic=salaries_domestic,
        salaries_international=salaries_international,
        salary_entries_domestic=json.dumps(salary_entries_domestic),
        salary_entries_international=json.dumps(salary_entries_international),
        available_years=available_years,
        expense_types=expense_types,
        all_categories=all_categories,
        expenses_grouped=expenses_grouped,
        expense_entries=json.dumps(expense_entries_by_name),
        categories_by_type=json.dumps(categories_by_type),
        expense_names_by_category=json.dumps(expense_names_by_category),
        all_vendors=all_vendors,
        today_date=today_date,
        active_tab=active_tab,
        inv_status_labels=json.dumps(list(inv_status_counts.keys())),
        inv_status_data=json.dumps(list(inv_status_counts.values())),
        exp_cat_labels=json.dumps(list(exp_cats.keys())),
        exp_cat_data=json.dumps(list(exp_cats.values())),
        ai_enabled=bool(_get_ai_client()),
        monthly_payment_details=json.dumps(monthly_payment_details),
        monthly_expense_details=json.dumps(monthly_expense_details),
        monthly_salary_details=json.dumps(monthly_salary_details),
        monthly_outstanding_details=json.dumps(monthly_outstanding_details),
        monthly_due_details=json.dumps(monthly_due_details),
        aging_buckets=aging_buckets,
        aging_totals=aging_totals,
        aging_total_outstanding=aging_total_outstanding,
        bs_aging_buckets=bs_aging_buckets,
        bs_aging_totals=bs_aging_totals,
        bs_aging_total_outstanding=bs_aging_total_outstanding,
        commission_summary=commission_summary,
        commission_total_earned=commission_total_earned,
        commission_total_paid=commission_total_paid,
        commission_total_outstanding=commission_total_outstanding,
        bs_total_commission=bs_total_commission,
        total_commission_paid=total_commission_paid,
        monthly_commission_details=json.dumps(monthly_commission_details),
    )

@app.route("/financial/expense/new", methods=["POST"])
@role_required("financial")
def expense_new():
    data = {
        "expense_type":      request.form.get("expense_type", ""),
        "expense_name":      request.form.get("expense_name", ""),
        "description":       request.form.get("description", "") or request.form.get("expense_name", ""),
        "amount":            _safe_float(request.form.get("amount", 0)),
        "category":          request.form.get("category", ""),
        "date":              request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
        "vendor":            request.form.get("vendor", ""),
        "project_number":    request.form.get("project_number", ""),
        "notes":             request.form.get("notes", ""),
        "created_by":        session.get("user_email", ""),
        "submitted_by_name": session.get("user_name", ""),
        "submitted_by_uid":  session.get("user_uid", ""),
        "status":            "Approved",
        "created_at":        datetime.now(timezone.utc).isoformat(),
        "updated_at":        datetime.now(timezone.utc).isoformat(),
    }
    # Handle receipt upload (base64 encoding)
    receipt_base64 = None
    receipt_filename = None
    receipt_type = None
    if 'receipt' in request.files:
        file = request.files['receipt']
        if file and file.filename:
            try:
                file_content = file.read()
                receipt_base64 = base64.b64encode(file_content).decode('utf-8')
                receipt_filename = file.filename
                receipt_type = file.content_type
                data['receipt_filename'] = receipt_filename
            except Exception as e:
                app.logger.error(f"Receipt upload error: {e}")
    # Finance expenses save ONLY to /balance_sheet_expenses (not to /expenses)
    # /expenses is for employee submissions only
    exp_id = fb_push("/balance_sheet_expenses", data)
    # Store receipt separately in /expense_receipts for fast loading
    if receipt_base64:
        fb_update(f"/expense_receipts/{exp_id}", {
            "receipt_base64":   receipt_base64,
            "receipt_filename": receipt_filename,
            "receipt_type":     receipt_type,
        })
    return jsonify({"success": True, "expense_id": exp_id})

@app.route("/financial/expense/<exp_id>/delete", methods=["POST"])
@role_required("financial")
def expense_delete(exp_id):
    # Check if this is an employee expense (in /expenses) or finance-only (just in /balance_sheet_expenses)
    is_emp_expense = fb_get(f"/expenses/{exp_id}") is not None

    # Always delete from balance_sheet_expenses
    fb_delete(f"/balance_sheet_expenses/{exp_id}")

    # Only delete from /expenses if it's an employee expense
    if is_emp_expense:
        fb_delete(f"/expenses/{exp_id}")

    flash("Expense deleted.", "success")
    return redirect(url_for("financial", tab="expenses"))

@app.route("/financial/expense/<exp_id>/remove-receipt", methods=["POST"])
@role_required("financial")
def remove_expense_receipt(exp_id):
    try:
        # Remove from /balance_sheet_expenses
        fb_update(f"/balance_sheet_expenses/{exp_id}", {
            "receipt_base64": "",
            "receipt_filename": "",
            "receipt_type": ""
        })

        # Remove from /expense_receipts (primary storage location)
        fb_delete(f"/expense_receipts/{exp_id}")

        # Sync removal to /expenses if it's an employee expense
        if fb_get(f"/expenses/{exp_id}"):
            fb_update(f"/expenses/{exp_id}", {
                "receipt_base64": "",
                "receipt_filename": "",
                "receipt_type": ""
            })

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/financial/expense/<exp_id>/edit", methods=["POST"])
@role_required("financial")
def expense_edit(exp_id):
    try:
        data = {
            "expense_type":   request.form.get("expense_type", ""),
            "expense_name":   request.form.get("expense_name", ""),
            "description":    request.form.get("description", "") or request.form.get("expense_name", ""),
            "amount":         request.form.get("amount", "0"),
            "category":       request.form.get("category", ""),
            "date":           request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
            "vendor":         request.form.get("vendor", ""),
            "project_number": request.form.get("project_number", ""),
            "notes":          request.form.get("notes", ""),
            "updated_at":     datetime.now(timezone.utc).isoformat(),
        }
        if 'receipt' in request.files:
            file = request.files['receipt']
            if file and file.filename:
                try:
                    file_content = file.read()
                    data['receipt_base64'] = base64.b64encode(file_content).decode('utf-8')
                    data['receipt_filename'] = file.filename
                    data['receipt_type'] = file.content_type
                except Exception as e:
                    app.logger.error(f"Receipt upload error: {e}")
        fb_update(f"/balance_sheet_expenses/{exp_id}", data)
        # If this expense originated as an employee submission, keep it in sync
        emp_rec = fb_get(f"/expenses/{exp_id}")
        if isinstance(emp_rec, dict):
            sync_fields = {k: v for k, v in data.items()
                           if k in ("expense_type","expense_name","description","amount",
                                    "category","date","vendor","project_number","notes","receipt_base64",
                                    "receipt_filename","receipt_type","updated_at")}
            fb_update(f"/expenses/{exp_id}", sync_fields)
            # Also sync receipt to separate receipt store if provided
            if any(k in data for k in ("receipt_base64","receipt_filename","receipt_type")):
                receipt_sync = {k: v for k, v in data.items()
                               if k in ("receipt_base64","receipt_filename","receipt_type")}
                if receipt_sync.get("receipt_base64"):
                    fb_update(f"/expense_receipts/{exp_id}", receipt_sync)
        return jsonify({"success": True, "expense_id": exp_id})
    except Exception as e:
        app.logger.error(f"Expense edit error: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/expense/<exp_id>/receipt", methods=["GET"])
@role_required("financial")
def get_expense_receipt(exp_id):
    """Retrieve receipt from Firebase - stored separately in /expense_receipts"""
    # First try /expense_receipts (primary storage location for all receipts)
    receipt_data = fb_get(f"/expense_receipts/{exp_id}") or {}
    if isinstance(receipt_data, dict) and receipt_data.get('receipt_base64'):
        return jsonify({
            "success": True,
            "receipt": receipt_data.get('receipt_base64'),
            "fileType": receipt_data.get('receipt_type', 'image/jpeg'),
            "filename": receipt_data.get('receipt_filename', 'receipt')
        })

    # Fallback: check if receipt is stored in expense data (legacy entries)
    expenses = fb_get("/balance_sheet_expenses") or {}
    if isinstance(expenses, dict) and exp_id in expenses:
        exp = expenses[exp_id]
        if isinstance(exp, dict) and 'receipt_base64' in exp:
            return jsonify({
                "success": True,
                "receipt": exp.get('receipt_base64'),
                "fileType": exp.get('receipt_type', 'image/jpeg'),
                "filename": exp.get('receipt_filename', 'receipt')
            })
    return jsonify({"success": False, "error": "Receipt not found"})

@app.route("/export/balance-sheet", methods=["GET"])
@role_required("financial")
def export_balance_sheet():
    """Export Balance Sheet to Excel - PIMS Format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.pagebreak import Break
        from io import BytesIO

        year_param = request.args.get("year", str(datetime.now().year))
        try:
            year = int(year_param)
        except (ValueError, TypeError):
            year = datetime.now().year

        # Get financial data
        invoices_data = fb_get("/invoices") or {}
        expenses_data = fb_get("/balance_sheet_expenses") or {}
        salaries_data = fb_get("/balance_sheet_salary") or {}

        # Calculate monthly data
        monthly_revenue = [0] * 12
        monthly_expenses = [0] * 12
        salary_inside = {}
        salary_outside = {}
        expense_breakdown = {}

        # Build revenue data from invoices based on PAYMENT DATES (not invoice dates)
        if isinstance(invoices_data, dict):
            for iid, inv_data in invoices_data.items():
                if isinstance(inv_data, dict):
                    for pay in (inv_data.get("payment_log", []) or []):
                        pay_ds = pay.get("date", "") or ""
                        try:
                            pay_date = datetime.fromisoformat(pay_ds[:10])
                            if pay_date.year == year:
                                monthly_revenue[pay_date.month - 1] += _safe_float(pay.get("amount", 0))
                        except Exception:
                            pass
                    for tax_pay in (inv_data.get("tax_payments", []) or []):
                        tax_ds = tax_pay.get("date", "") or ""
                        try:
                            tax_date = datetime.fromisoformat(tax_ds[:10])
                            if tax_date.year == year:
                                monthly_revenue[tax_date.month - 1] += _safe_float(tax_pay.get("amount", 0))
                        except Exception:
                            pass

        # Build expense data
        if isinstance(expenses_data, dict):
            for eid, edata in expenses_data.items():
                if isinstance(edata, dict):
                    date_str = edata.get("date", "")
                    try:
                        e_year = int(date_str[:4])
                        e_month = int(date_str[5:7])
                        if e_year == year and 1 <= e_month <= 12:
                            amt = _safe_float(edata.get("amount", 0))
                            monthly_expenses[e_month-1] += amt
                            # Use vendor name for breakdown, fallback to expense_name if no vendor
                            vendor_name = edata.get("vendor", "").strip() or edata.get("expense_name", "") or edata.get("description", "—")
                            expense_breakdown[vendor_name] = expense_breakdown.get(vendor_name, 0) + amt
                    except (ValueError, IndexError):
                        pass

        # Build salary data
        if isinstance(salaries_data, dict):
            for salary_id, sal_data in salaries_data.items():
                if isinstance(sal_data, dict):
                    date_str = sal_data.get("date", "")
                    try:
                        s_year = int(date_str[:4])
                        if s_year == year:
                            name = sal_data.get("name") or sal_data.get("employee_name", "—")
                            amt = _safe_float(sal_data.get("amount", 0))
                            region = sal_data.get("region", "").lower()
                            target_dict = salary_inside if "inside" in region or "usa" in region else salary_outside
                            target_dict[name] = target_dict.get(name, 0) + amt
                    except (ValueError, IndexError, TypeError):
                        pass

        total_revenue = sum(monthly_revenue)
        total_expenses = sum(monthly_expenses)
        total_salaries = sum(salary_inside.values()) + sum(salary_outside.values())
        net_profit = total_revenue - total_expenses

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Annual Summary"

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Set page break and print area
        ws.row_breaks.append(Break(id=35))
        ws.print_area = 'A1:O68'
        ws.print_title_rows = '1:5'

        # Professional colors matching PIMS
        dark_blue_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
        light_blue_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        light_gray_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

        # Fonts
        title_font = Font(name='Calibri', size=24, bold=True, color="FF1F4E79")
        subtitle_font = Font(name='Calibri', size=16, bold=True, color="FF1F4E79")
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFFFF")
        bold_font = Font(name='Calibri', size=10, bold=True)
        normal_font = Font(name='Calibri', size=10)

        # Alignments
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Borders
        thin_border = Border(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000')
        )

        # Column widths
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 15
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions['O'].width = 18

        # PAGE 1 - ANNUAL FINANCIAL SUMMARY
        # Add top spacing
        ws.row_dimensions[1].height = 15
        ws.row_dimensions[2].height = 30

        ws.merge_cells('B2:O2')
        title = ws['B2']
        title.value = "MABS ENGINEERING LLC"
        title.font = title_font
        title.alignment = center_align
        title.fill = white_fill

        ws.row_dimensions[3].height = 10
        ws.row_dimensions[4].height = 10

        ws.merge_cells('B5:O5')
        subtitle = ws['B5']
        subtitle.value = f"ANNUAL FINANCIAL SUMMARY - {year}"
        subtitle.font = subtitle_font
        subtitle.alignment = center_align
        subtitle.fill = white_fill

        # Month headers
        ws['B12'] = "Months"
        ws['B12'].font = header_font
        ws['B12'].fill = dark_blue_fill
        ws['B12'].alignment = center_align
        ws['B12'].border = thin_border

        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        for i, month in enumerate(months, 3):
            cell = ws.cell(row=12, column=i, value=month)
            cell.font = header_font
            cell.fill = dark_blue_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Revenue row
        ws['B13'] = "Revenue"
        ws['B13'].font = bold_font
        ws['B13'].fill = light_gray_fill
        ws['B13'].alignment = center_align
        ws['B13'].border = thin_border

        for i, val in enumerate(monthly_revenue, 3):
            cell = ws.cell(row=13, column=i, value=val)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

        # Expense row
        ws['B14'] = "Expense"
        ws['B14'].font = bold_font
        ws['B14'].fill = light_gray_fill
        ws['B14'].alignment = center_align
        ws['B14'].border = thin_border

        for i, val in enumerate(monthly_expenses, 3):
            cell = ws.cell(row=14, column=i, value=val)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = normal_font

        # Summary section
        ws['G20'] = "Revenue"
        ws['G20'].font = bold_font
        ws['H20'] = "="
        ws['H20'].font = bold_font
        ws['H20'].alignment = center_align
        ws.merge_cells('I20:J20')
        ws['I20'].value = total_revenue
        ws['I20'].number_format = '"$"#,##0.00'
        ws['I20'].font = bold_font
        ws['I20'].alignment = left_align

        ws['G21'] = "Total Expense"
        ws['G21'].font = bold_font
        ws['H21'] = "="
        ws['H21'].font = bold_font
        ws['H21'].alignment = center_align
        ws.merge_cells('I21:J21')
        ws['I21'].value = total_expenses
        ws['I21'].number_format = '"$"#,##0.00'
        ws['I21'].font = bold_font
        ws['I21'].alignment = left_align

        ws['G22'] = "Net Profit"
        ws['G22'].font = bold_font
        ws['H22'] = "="
        ws['H22'].font = bold_font
        ws['H22'].alignment = center_align
        ws.merge_cells('I22:J22')
        ws['I22'].value = net_profit
        ws['I22'].number_format = '"$"#,##0.00'
        ws['I22'].font = bold_font
        ws['I22'].fill = green_fill if net_profit > 0 else light_gray_fill
        ws['I22'].alignment = left_align

        ws['H34'] = "Page : 01"
        ws['H34'].font = Font(size=10, italic=True)
        ws['H34'].alignment = right_align

        for row in range(1, 35):
            for col in range(1, 16):
                if ws.cell(row=row, column=col).value is None:
                    ws.cell(row=row, column=col).fill = white_fill

        # PAGE 2 - SALARY & EXPENSE BREAKDOWN
        start_row = 35

        # Salary title
        ws.merge_cells(f"B{start_row}:E{start_row}")
        ws[f"B{start_row}"].value = "SALARY"
        ws[f"B{start_row}"].font = Font(name='Calibri', size=14, bold=True)
        ws[f"B{start_row}"].alignment = center_align

        # Inside America
        row = start_row + 2
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].value = "Inside America"
        ws[f"B{row}"].fill = light_blue_fill
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].alignment = center_align
        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"].value = "Emp. Nm"
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].fill = light_gray_fill
        ws[f"B{row}"].alignment = center_align

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"].value = "Amount"
        ws[f"D{row}"].font = bold_font
        ws[f"D{row}"].fill = light_gray_fill
        ws[f"D{row}"].alignment = center_align

        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        row += 1
        total_inside = 0
        for name in sorted(salary_inside.keys()):
            amt = salary_inside[name]
            ws.merge_cells(f"B{row}:C{row}")
            ws[f"B{row}"].value = name
            ws[f"B{row}"].alignment = left_align

            ws.merge_cells(f"D{row}:E{row}")
            ws[f"D{row}"].value = amt
            ws[f"D{row}"].number_format = '"$"#,##0.00'
            ws[f"D{row}"].alignment = center_align

            for col in range(2, 6):
                ws.cell(row, col).border = thin_border

            total_inside += amt
            row += 1

        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"].value = "Total"
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].alignment = center_align

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"].value = total_inside
        ws[f"D{row}"].number_format = '"$"#,##0.00'
        ws[f"D{row}"].fill = green_fill
        ws[f"D{row}"].font = bold_font
        ws[f"D{row}"].alignment = center_align

        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        # Outside America
        row += 2
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].value = "Outside America"
        ws[f"B{row}"].fill = light_blue_fill
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].alignment = center_align

        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        row += 1
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"].value = "Emp. Nm"
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].fill = light_gray_fill
        ws[f"B{row}"].alignment = center_align

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"].value = "Amount"
        ws[f"D{row}"].font = bold_font
        ws[f"D{row}"].fill = light_gray_fill
        ws[f"D{row}"].alignment = center_align

        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        row += 1
        total_outside = 0
        for name in sorted(salary_outside.keys()):
            amt = salary_outside[name]
            ws.merge_cells(f"B{row}:C{row}")
            ws[f"B{row}"].value = name
            ws[f"B{row}"].alignment = left_align

            ws.merge_cells(f"D{row}:E{row}")
            ws[f"D{row}"].value = amt
            ws[f"D{row}"].number_format = '"$"#,##0.00'
            ws[f"D{row}"].alignment = center_align

            for col in range(2, 6):
                ws.cell(row, col).border = thin_border

            total_outside += amt
            row += 1

        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"].value = "Total"
        ws[f"B{row}"].font = bold_font
        ws[f"B{row}"].alignment = center_align

        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"].value = total_outside
        ws[f"D{row}"].number_format = '"$"#,##0.00'
        ws[f"D{row}"].fill = green_fill
        ws[f"D{row}"].font = bold_font
        ws[f"D{row}"].alignment = center_align

        for col in range(2, 6):
            ws.cell(row, col).border = thin_border

        # Expense Breakdown
        ws.merge_cells("G35:K35")
        ws["G35"].value = "EXPENSE BREAK-DOWN"
        ws["G35"].font = Font(name='Calibri', size=14, bold=True)
        ws["G35"].alignment = center_align

        ws.merge_cells("G37:I37")
        ws["G37"].value = "Expense Item"
        ws["G37"].font = bold_font
        ws["G37"].fill = light_blue_fill
        ws["G37"].alignment = center_align

        ws.merge_cells("J37:K37")
        ws["J37"].value = "Amount"
        ws["J37"].font = bold_font
        ws["J37"].fill = light_blue_fill
        ws["J37"].alignment = center_align

        for col in range(7, 12):
            ws.cell(37, col).border = thin_border

        row = 38
        for exp_name in sorted(expense_breakdown.keys(), key=lambda x: expense_breakdown[x], reverse=True):
            amt = expense_breakdown[exp_name]

            ws.merge_cells(f"G{row}:I{row}")
            ws[f"G{row}"].value = exp_name
            ws[f"G{row}"].alignment = left_align

            ws.merge_cells(f"J{row}:K{row}")
            ws[f"J{row}"].value = amt
            ws[f"J{row}"].number_format = '"$"#,##0.00'
            ws[f"J{row}"].alignment = center_align

            for col in range(7, 12):
                ws.cell(row, col).border = thin_border

            row += 1

        ws.merge_cells(f"G{row}:I{row}")
        ws[f"G{row}"].value = "Total"
        ws[f"G{row}"].font = bold_font
        ws[f"G{row}"].alignment = center_align

        ws.merge_cells(f"J{row}:K{row}")
        ws[f"J{row}"].value = total_expenses
        ws[f"J{row}"].number_format = '"$"#,##0.00'
        ws[f"J{row}"].fill = green_fill
        ws[f"J{row}"].font = bold_font
        ws[f"J{row}"].alignment = center_align

        for col in range(7, 12):
            ws.cell(row, col).border = thin_border

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Balance_Sheet_Annual_{year}.xlsx'
        )

    except Exception as e:
        log.error(f"Export error: {e}")
        flash(f"Export failed: {str(e)}", "danger")
        return redirect(url_for("financial", tab="balance-sheet"))

# ── Routes: Employees ─────────────────────────────────────────────────────────
@app.route("/employees")
@role_required("employees")
def employees():
    uid        = session.get("user_uid", "")
    _role      = normalize_role(session.get("user_role", ""))
    is_admin   = _role == "admin"
    is_finance = _role == "finance"

    all_entries = _load_time_entries()
    all_time_off = _load_time_off_requests()
    active_projects = [p for p in _load_projects_list()
                       if p.get("status", "") not in ("Completed", "invoiced_Fully paid", "Cancelled")]

    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    month_start = now.strftime("%Y-%m-01")
    year_start = now.strftime("%Y-01-01")

    my_entries = [e for e in all_entries if e.get("employee_uid") == uid]
    my_open_entry = next((e for e in my_entries if e.get("status") == "open"), None)
    # Most recently used project — pre-selected as the default for the next clock-in
    last_project_number = my_entries[0].get("project_number", "") if my_entries else ""
    my_today_entries = [e for e in my_entries if e.get("status") == "closed" and e.get("date") == today_str]
    my_week_entries = [e for e in my_entries if e.get("status") == "closed" and e.get("date", "") >= week_start]
    my_month_entries = [e for e in my_entries if e.get("status") == "closed" and e.get("date", "") >= month_start]
    my_year_entries = [e for e in my_entries if e.get("status") == "closed" and e.get("date", "") >= year_start]
    my_today_minutes = _sum_minutes(my_today_entries)
    my_week_minutes = _sum_minutes(my_week_entries)
    my_month_minutes = _sum_minutes(my_month_entries)
    my_year_minutes = _sum_minutes(my_year_entries)
    my_time_off = [r for r in all_time_off if r.get("employee_uid") == uid]
    my_time_off_balance = _time_off_balance(all_time_off, uid, now.year)

    # "Hours by Project" breakdown for My Time Clock tab, with its own period selector
    my_period = request.args.get("myperiod", "week")
    my_period_start, my_period_end, my_period_label = _period_range(
        my_period, request.args.get("mystart", ""), request.args.get("myend", ""))
    my_period_entries = [e for e in my_entries
                          if e.get("status") == "closed" and my_period_start <= e.get("date", "") <= my_period_end]
    my_period_by_project = _sum_minutes_by_project(my_period_entries)
    my_period_minutes = _sum_minutes(my_period_entries)

    context = {
        "active_projects":  active_projects,
        "today_str":        today_str,
        "profile_user":     fb_get(f"/users/{uid}") or {},
        "last_project_number": last_project_number,
        "my_open_entry":    my_open_entry,
        "my_today_entries": my_today_entries,
        "my_today_minutes": my_today_minutes,
        "my_week_minutes":  my_week_minutes,
        "my_month_minutes": my_month_minutes,
        "my_year_minutes":  my_year_minutes,
        "my_period":            my_period,
        "my_period_start":      my_period_start,
        "my_period_end":        my_period_end,
        "my_period_label":      my_period_label,
        "my_period_by_project": my_period_by_project,
        "my_period_minutes":    my_period_minutes,
        "my_time_off":         my_time_off,
        "my_time_off_balance": my_time_off_balance,
        "current_year":        now.year,
    }

    if is_admin:
        period = request.args.get("period", "week")
        custom_start = request.args.get("start", "")
        custom_end = request.args.get("end", "")
        period_start, period_end, period_label = _period_range(period, custom_start, custom_end)
        period_entries = [e for e in all_entries if period_start <= e.get("date", "") <= period_end]

        stale_open_entries = [e for e in all_entries
                               if e.get("status") == "open" and e.get("date", "") < today_str]
        for e in stale_open_entries:
            e["_suggested_close"] = f"{e.get('date', today_str)}T17:00"

        context.update({
            "all_users":           _load_all_users(),
            "open_entries_by_uid": {e["employee_uid"]: e for e in all_entries if e.get("status") == "open"},
            "pending_time_off":    [r for r in all_time_off if r.get("status") == "Pending"],
            "all_time_off_balances": {
                u["firebase_uid"]: _time_off_balance(all_time_off, u["firebase_uid"], now.year)
                for u in _load_all_users() if u.get("firebase_uid")
            },
            "hours_by_project":    _aggregate_hours_by_project(period_entries),
            "stale_open_entries":  stale_open_entries,
            "period":              period,
            "period_start":        period_start,
            "period_end":          period_end,
            "period_label":        period_label,
        })
    # Medical allowance claims
    all_medical_raw = fb_get("/medical_claims") or {}
    all_medical_list = []
    for cid, cdata in (all_medical_raw.items() if isinstance(all_medical_raw, dict) else []):
        if cdata and isinstance(cdata, dict):
            cdata["firebase_id"] = cid
            all_medical_list.append(cdata)
    all_medical_list.sort(key=lambda x: x.get("submitted_at", ""), reverse=True)
    context["my_medical_claims"] = [c for c in all_medical_list if c.get("employee_uid") == uid]
    if is_admin:
        context["all_medical_claims"] = all_medical_list

    # Expense types / categories / names for employee submission form
    # "Other Expenses" (salary, tax, medical, bank charges) is excluded — those are admin/payroll entries
    # "Salaries, Labor & Related Costs" is excluded from O&M for the same reason
    _custom_cats = fb_get("/custom_categories") or {}
    _names_by_cat = _custom_cats.get("expense_names", {}) if isinstance(_custom_cats.get("expense_names"), dict) else {}

    EMP_EXPENSE_TYPES = ["O & M (Operations & Maintenance)", "Capital Expenses", "Other Expenses"]
    EMP_CATS_BY_TYPE = {
        "O & M (Operations & Maintenance)": [
            "Facilities & Utilities",
            "Office & Admin Overhead",
            "Engineering Software & IT",
            "Professional Services",
            "Insurance & Compliance",
            "Travel, Site Visits & Vehicles",
            "Marketing & Business Development",
            "Training, Licensure & Development",
            "Safety & Field Supplies",
            "Miscellaneous O & M",
        ],
        "Capital Expenses": [
            "Computer & Office Equipment",
            "Field & Inspection Equipment",
            "Furniture & Fixtures",
            "Vehicles",
            "Software (Capitalized)",
            "Leasehold Improvements",
        ],
        "Other Expenses": [],
    }
    context["emp_expense_types"]      = EMP_EXPENSE_TYPES
    context["emp_categories_by_type"] = json.dumps(EMP_CATS_BY_TYPE)
    context["emp_names_by_cat"]       = json.dumps(_names_by_cat)

    # Employee expense submissions — strip receipt_base64 (large binary) to keep page fast
    _LARGE_FIELDS = {"receipt_base64", "receipt_content"}
    raw_emp_exp = fb_get("/expenses") or {}
    all_emp_expenses = []
    if isinstance(raw_emp_exp, dict):
        for eid, edata in raw_emp_exp.items():
            if isinstance(edata, dict):
                slim = {k: v for k, v in edata.items() if k not in _LARGE_FIELDS}
                slim["firebase_id"] = eid
                all_emp_expenses.append(slim)
    all_emp_expenses.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    # For display, merge pending edit data into the main record
    # This shows the edited data while awaiting approval
    for e in all_emp_expenses:
        if e.get("edit_status") == "pending" and "pending_edit" in e:
            # Merge edited data for display
            pending_edit_data = e.get("pending_edit", {})
            for key, value in pending_edit_data.items():
                e[key] = value
            # Set status to Pending while awaiting edit approval
            e["status"] = "Pending"

    # "My Submissions" always shows only current user's submitted expenses
    context["my_expenses"] = [e for e in all_emp_expenses if e.get("submitted_by_uid") == uid]

    if is_admin:
        # Admins see all pending expenses for approval
        context["pending_expenses"] = [e for e in all_emp_expenses if e.get("status") == "Pending"]
    else:
        context["pending_expenses"] = []

    return render_template("employees.html", **context)

@app.route("/employees/medical-claims/form")
@role_required("employees")
def medical_claim_form_download():
    return render_template("medical_claim_form.html",
                           user_name=session.get("user_name", ""),
                           today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/employees/medical-claims/new", methods=["POST"])
@role_required("employees")
def medical_claim_new():
    uid  = session.get("user_uid", "")
    name = session.get("user_name", "")
    amount = request.form.get("amount_claimed", "0") or "0"
    try:
        amount = float(amount)
    except ValueError:
        amount = 0.0
    claim = {
        "employee_uid":    uid,
        "employee_name":   name,
        "claim_date":      request.form.get("claim_date", ""),
        "expense_type":    request.form.get("expense_type", "Medical"),
        "amount_claimed":  amount,
        "description":     request.form.get("description", "").strip(),
        "provider":        request.form.get("provider", "").strip(),
        "receipt_ref":     request.form.get("receipt_ref", "").strip(),
        "status":          "Pending",
        "amount_approved": None,
        "admin_notes":     None,
        "reviewed_by":     None,
        "reviewed_at":     None,
        "submitted_at":    datetime.now(timezone.utc).isoformat(),
    }
    fb_push("/medical_claims", claim)
    flash("Medical allowance claim submitted successfully.", "success")
    return redirect(url_for("employees") + "#medical")

@app.route("/employees/medical-claims/<claim_id>/review", methods=["POST"])
@role_required("employees")
def medical_claim_review(claim_id):
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("Admin access required.", "danger")
        return redirect(url_for("employees") + "#medical")
    action = request.form.get("action", "")
    status = "Approved" if action == "approve" else "Rejected"
    try:
        amt_approved = float(request.form.get("amount_approved", 0) or 0)
    except ValueError:
        amt_approved = 0.0
    fb_update(f"/medical_claims/{claim_id}", {
        "status":          status,
        "amount_approved": amt_approved if status == "Approved" else None,
        "admin_notes":     request.form.get("admin_notes", "").strip(),
        "reviewed_by":     session.get("user_name", ""),
        "reviewed_at":     datetime.now(timezone.utc).isoformat(),
    })
    flash(f"Claim {status.lower()} successfully.", "success")
    return redirect(url_for("employees") + "#medical")

@app.route("/employees/expenses/submit", methods=["POST"])
@role_required("employees")
def employee_expense_submit():
    app.logger.info(f"Employee expense submit - request.form keys: {list(request.form.keys())}")
    app.logger.info(f"Employee expense submit - request.form: {dict(request.form)}")

    editing_expense_id = request.form.get("editing_expense_id", "").strip()

    # If editing, receipt is optional (can keep existing receipt)
    # If creating new, receipt is mandatory
    has_receipt = 'receipt' in request.files and request.files['receipt'].filename
    if not editing_expense_id and not has_receipt:
        flash("A receipt is required to submit an expense.", "danger")
        return redirect(url_for("employees") + "#expenses")

    receipt_base64 = None
    receipt_filename = None
    receipt_type = None

    if has_receipt:
        receipt_file = request.files['receipt']
        try:
            file_content = receipt_file.read()
            receipt_base64 = base64.b64encode(file_content).decode('utf-8')
            receipt_filename = receipt_file.filename
            receipt_type = receipt_file.content_type
        except Exception as e:
            app.logger.error(f"Expense receipt upload error: {e}")
            flash("Receipt upload failed. Please try again.", "danger")
            return redirect(url_for("employees") + "#expenses")

    data = {
        "expense_type":      request.form.get("expense_type", ""),
        "expense_name":      request.form.get("expense_name", ""),
        "description":       request.form.get("expense_name", ""),
        "amount":            _safe_float(request.form.get("amount", 0)),
        "category":          request.form.get("category", ""),
        "date":              request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
        "vendor":            request.form.get("vendor", "").strip(),
        "project_number":    request.form.get("project_number", ""),
        "notes":             request.form.get("notes", "").strip(),
        "updated_at":        datetime.now(timezone.utc).isoformat(),
    }
    app.logger.info(f"Employee expense submit - expense_name={repr(request.form.get('expense_name'))}, category={repr(request.form.get('category'))}, project_number={repr(request.form.get('project_number'))}, notes={repr(request.form.get('notes'))}")

    if editing_expense_id:
        # Editing existing expense
        data.pop("expense_type", None)  # Can't change type after submission

        # Handle receipt updates
        if receipt_base64:
            # New receipt provided - update both expense records and receipt storage
            data["receipt_filename"] = receipt_filename
            fb_update(f"/expense_receipts/{editing_expense_id}", {
                "receipt_base64":   receipt_base64,
                "receipt_filename": receipt_filename,
                "receipt_type":     receipt_type,
            })
        else:
            # No new receipt - keep existing receipt_filename in expense data
            # (don't pop it from data - preserve it)
            pass

        # Update employee expense record directly (no approval needed for edits)
        fb_update(f"/expenses/{editing_expense_id}", data)

        # Also update /balance_sheet_expenses (Finance tab) if it exists (for approved expenses)
        if fb_get(f"/balance_sheet_expenses/{editing_expense_id}"):
            fb_update(f"/balance_sheet_expenses/{editing_expense_id}", data)

        flash("Expense updated successfully.", "success")
    else:
        # Creating new expense
        data.update({
            "submitted_by_name": session.get("user_name", ""),
            "submitted_by_uid":  session.get("user_uid", ""),
            "submitted_by_email":session.get("user_email", ""),
            "status":            "Pending",
            "created_at":        datetime.now(timezone.utc).isoformat(),
            "receipt_filename":  receipt_filename,
        })
        exp_id = fb_push("/expenses", data)
        # Store receipt binary in a separate path so /expenses stays lightweight
        if exp_id and receipt_base64:
            fb_update(f"/expense_receipts/{exp_id}", {
                "receipt_base64":   receipt_base64,
                "receipt_filename": receipt_filename,
                "receipt_type":     receipt_type,
            })
        flash("Expense submitted and pending admin approval.", "success")

    return redirect(url_for("employees") + "#expenses")


@app.route("/employees/expenses/<exp_id>/edit/review", methods=["POST"])
@role_required("employees")
def employee_expense_edit_review(exp_id):
    """Admin approve/reject pending edits to an expense"""
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("Admin access required.", "danger")
        return redirect(url_for("employees") + "#expenses")

    action = request.form.get("action", "")
    review_note = request.form.get("review_note", "").strip()
    if action not in ("approve", "reject"):
        flash("Invalid action.", "danger")
        return redirect(url_for("employees") + "#expenses")

    now_str = datetime.now(timezone.utc).isoformat()
    exp_data = fb_get(f"/expenses/{exp_id}") or {}

    if action == "approve":
        # Apply the pending edit
        if isinstance(exp_data, dict) and "pending_edit" in exp_data:
            edited_data = exp_data.get("pending_edit", {})
            # Merge edited data back into main record
            for key, value in edited_data.items():
                exp_data[key] = value
            # Clear pending edit status and set to approved
            exp_data.pop("pending_edit", None)
            exp_data["status"] = "Approved"  # Set status to Approved
            exp_data["edit_status"] = "approved"
            exp_data["edit_approved_by"] = session.get("user_name", "")
            exp_data["edit_approved_at"] = now_str
            exp_data["edit_review_note"] = review_note
            exp_data["updated_at"] = now_str

            # Update /expenses with merged data
            fb_update(f"/expenses/{exp_id}", exp_data)

            # Also update /balance_sheet_expenses (Finance tab) if it exists
            if fb_get(f"/balance_sheet_expenses/{exp_id}"):
                fb_update(f"/balance_sheet_expenses/{exp_id}", exp_data)

            flash("Edit approved and expense updated.", "success")
        else:
            flash("No pending edits found.", "warning")
    else:
        # Reject the edit - remove pending data
        exp_data.pop("pending_edit", None)
        exp_data["edit_status"] = "rejected"
        exp_data["edit_rejected_by"] = session.get("user_name", "")
        exp_data["edit_rejected_at"] = now_str
        exp_data["edit_review_note"] = review_note
        exp_data["updated_at"] = now_str

        fb_update(f"/expenses/{exp_id}", exp_data)
        flash("Edit rejected. Original expense data kept.", "success")

    return redirect(url_for("employees") + "#expenses")


@app.route("/employees/expenses/<exp_id>/review", methods=["POST"])
@role_required("employees")
def employee_expense_review(exp_id):
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("Admin access required.", "danger")
        return redirect(url_for("employees") + "#expenses")

    action = request.form.get("action", "")
    review_note = request.form.get("review_note", "").strip()
    if action not in ("approve", "reject"):
        flash("Invalid action.", "danger")
        return redirect(url_for("employees") + "#expenses")

    new_status = "Approved" if action == "approve" else "Rejected"
    now_str = datetime.now(timezone.utc).isoformat()
    fb_update(f"/expenses/{exp_id}", {
        "status":      new_status,
        "reviewed_by": session.get("user_name", ""),
        "reviewed_at": now_str,
        "review_note": review_note,
        "updated_at":  now_str,
    })

    if new_status == "Approved":
        # Mirror into balance_sheet_expenses so it appears in Financial tab
        exp_data = fb_get(f"/expenses/{exp_id}") or {}
        if isinstance(exp_data, dict):
            exp_data["firebase_id"] = exp_id
            exp_data["created_by"] = exp_data.get("submitted_by_email", "")
            fb_update(f"/balance_sheet_expenses/{exp_id}", exp_data)

    flash(f"Expense {new_status.lower()}.", "success")
    return redirect(url_for("employees") + "#expenses")


@app.route("/employees/expenses/<exp_id>/receipt", methods=["GET"])
@role_required("employees")
def employee_expense_receipt(exp_id):
    exp_data = fb_get(f"/expenses/{exp_id}") or {}
    uid = session.get("user_uid", "")
    role = normalize_role(session.get("user_role", ""))
    # Only admin or the submitter can view
    if role != "admin" and exp_data.get("submitted_by_uid") != uid:
        return jsonify({"error": "Not authorized"}), 403
    # Check separate receipt store first (new), fall back to inline field (legacy)
    receipt_rec = fb_get(f"/expense_receipts/{exp_id}") or {}
    receipt_b64      = receipt_rec.get("receipt_base64")      or exp_data.get("receipt_base64", "")
    receipt_filename = receipt_rec.get("receipt_filename")     or exp_data.get("receipt_filename", "receipt")
    receipt_type     = receipt_rec.get("receipt_type")         or exp_data.get("receipt_type", "application/octet-stream")
    if not receipt_b64:
        return jsonify({"error": "No receipt found"}), 404
    data_url = f"data:{receipt_type};base64,{receipt_b64}"
    return jsonify({"url": data_url, "filename": receipt_filename})


@app.route("/employees/expenses/<exp_id>/edit", methods=["POST"])
@role_required("employees")
def employee_expense_edit(exp_id):
    uid = session.get("user_uid", "")
    role = normalize_role(session.get("user_role", ""))
    exp_data = fb_get(f"/expenses/{exp_id}") or {}
    # Only admins can edit expenses
    if role != "admin":
        flash("Only admins can edit expenses.", "danger")
        return redirect(url_for("employees") + "#expenses")

    updates = {
        "expense_type":   request.form.get("expense_type", ""),
        "expense_name":   request.form.get("expense_name", "").strip(),
        "description":    request.form.get("description", "").strip() or request.form.get("expense_name", "").strip(),
        "amount":         _safe_float(request.form.get("amount", 0)),
        "date":           request.form.get("date", ""),
        "vendor":         request.form.get("vendor", "").strip(),
        "project_number": request.form.get("project_number", ""),
        "notes":          request.form.get("notes", "").strip(),
        "updated_at":     datetime.now(timezone.utc).isoformat(),
    }
    # Replace receipt only if a new file was provided
    if 'receipt' in request.files and request.files['receipt'].filename:
        receipt_file = request.files['receipt']
        try:
            new_b64 = base64.b64encode(receipt_file.read()).decode('utf-8')
            updates["receipt_filename"] = receipt_file.filename
            updates["receipt_base64"] = new_b64
            updates["receipt_type"] = receipt_file.content_type
            fb_update(f"/expense_receipts/{exp_id}", {
                "receipt_base64":   new_b64,
                "receipt_filename": receipt_file.filename,
                "receipt_type":     receipt_file.content_type,
            })
        except Exception as e:
            app.logger.error(f"Expense receipt update error: {e}")
            flash("Receipt upload failed. Other changes were not saved.", "danger")
            return redirect(url_for("employees") + "#expenses")

    # Update in employee expenses
    fb_update(f"/expenses/{exp_id}", updates)
    # Also sync to balance_sheet_expenses if it exists there
    if fb_get(f"/balance_sheet_expenses/{exp_id}"):
        fb_update(f"/balance_sheet_expenses/{exp_id}", updates)
    flash("Expense updated successfully.", "success")
    return redirect(url_for("employees") + "#expenses")


@app.route("/employees/expenses/<exp_id>/delete", methods=["POST"])
@role_required("employees")
def employee_expense_delete(exp_id):
    uid = session.get("user_uid", "")
    role = normalize_role(session.get("user_role", ""))
    exp_data = fb_get(f"/expenses/{exp_id}") or {}
    # Submitter can delete their own Pending; admin can delete anything
    if role != "admin" and exp_data.get("submitted_by_uid") != uid:
        flash("You can only delete your own expenses.", "danger")
        return redirect(url_for("employees") + "#expenses")
    if role != "admin" and exp_data.get("status", "Pending") != "Pending":
        flash("Only pending expenses can be deleted.", "danger")
        return redirect(url_for("employees") + "#expenses")

    fb_delete(f"/expenses/{exp_id}")
    flash("Expense deleted.", "success")
    return redirect(url_for("employees") + "#expenses")


@app.route("/employees/clock-in", methods=["POST"])
@role_required("employees")
def employee_clock_in():
    uid  = session.get("user_uid", "")
    name = session.get("user_name", "")
    project_number = request.form.get("project_number", "").strip()

    if any(e.get("employee_uid") == uid and e.get("status") == "open" for e in _load_time_entries()):
        flash("You're already clocked in. Clock out first.", "warning")
        return redirect(url_for("employees"))

    project_name = ""
    if project_number:
        proj = next((p for p in _load_projects_list() if p.get("project_number") == project_number), None)
        if proj:
            project_name = proj.get("project_name", "")

    now = datetime.now()
    fb_push("/time_entries", {
        "employee_uid":   uid,
        "employee_name":  name,
        "project_number": project_number,
        "project_name":   project_name,
        "clock_in":       now.isoformat(),
        "clock_out":      None,
        "duration_minutes": 0,
        "date":           now.strftime("%Y-%m-%d"),
        "status":         "open",
    })
    flash(f"Clocked in{' to ' + project_name if project_name else ''}.", "success")
    return redirect(url_for("employees"))

@app.route("/employees/switch-project", methods=["POST"])
@role_required("employees")
def employee_switch_project():
    uid  = session.get("user_uid", "")
    name = session.get("user_name", "")
    new_project_number = request.form.get("project_number", "").strip()

    open_entry = next((e for e in _load_time_entries()
                        if e.get("employee_uid") == uid and e.get("status") == "open"), None)
    if not open_entry:
        flash("You're not currently clocked in.", "warning")
        return redirect(url_for("employees"))

    if new_project_number == open_entry.get("project_number", ""):
        flash("You're already clocked into that project.", "warning")
        return redirect(url_for("employees"))

    now = datetime.now()
    clock_in = datetime.fromisoformat(open_entry["clock_in"])
    duration = round((now - clock_in).total_seconds() / 60.0, 1)
    fb_update(f"/time_entries/{open_entry['firebase_id']}", {
        "clock_out":        now.isoformat(),
        "duration_minutes": duration,
        "status":           "closed",
    })

    project_name = ""
    if new_project_number:
        proj = next((p for p in _load_projects_list() if p.get("project_number") == new_project_number), None)
        if proj:
            project_name = proj.get("project_name", "")

    fb_push("/time_entries", {
        "employee_uid":   uid,
        "employee_name":  name,
        "project_number": new_project_number,
        "project_name":   project_name,
        "clock_in":       now.isoformat(),
        "clock_out":      None,
        "duration_minutes": 0,
        "date":           now.strftime("%Y-%m-%d"),
        "status":         "open",
    })
    flash(f"Switched to {project_name or 'General / Admin'}.", "success")
    return redirect(url_for("employees"))

@app.route("/employees/clock-out", methods=["POST"])
@role_required("employees")
def employee_clock_out():
    uid = session.get("user_uid", "")
    open_entry = next((e for e in _load_time_entries()
                        if e.get("employee_uid") == uid and e.get("status") == "open"), None)
    if not open_entry:
        flash("You're not currently clocked in.", "warning")
        return redirect(url_for("employees"))

    clock_in = datetime.fromisoformat(open_entry["clock_in"])
    now = datetime.now()
    duration = round((now - clock_in).total_seconds() / 60.0, 1)
    fb_update(f"/time_entries/{open_entry['firebase_id']}", {
        "clock_out":        now.isoformat(),
        "duration_minutes": duration,
        "status":           "closed",
        "notes":            request.form.get("notes", "").strip(),
    })
    flash("Clocked out.", "success")
    return redirect(url_for("employees"))

@app.route("/employees/log-time", methods=["POST"])
@role_required("employees")
def employee_log_time():
    uid  = session.get("user_uid", "")
    name = session.get("user_name", "")
    date_str = request.form.get("date", "").strip()
    project_number = request.form.get("project_number", "").strip()
    hours = _safe_float(request.form.get("hours", 0))
    notes = request.form.get("notes", "").strip()

    today_str = datetime.now().strftime("%Y-%m-%d")
    if not date_str or date_str > today_str or hours <= 0:
        flash("Enter a valid date (not in the future) and number of hours.", "danger")
        return redirect(url_for("employees"))

    project_name = ""
    if project_number:
        proj = next((p for p in _load_projects_list() if p.get("project_number") == project_number), None)
        if proj:
            project_name = proj.get("project_name", "")

    fb_push("/time_entries", {
        "employee_uid":     uid,
        "employee_name":    name,
        "project_number":   project_number,
        "project_name":     project_name,
        "clock_in":         "",
        "clock_out":        "",
        "duration_minutes": round(hours * 60, 1),
        "date":             date_str,
        "status":           "closed",
        "notes":            notes,
        "source":           "manual",
    })
    flash(f"Logged {hours:g}h to {project_name or 'General / Admin'}.", "success")
    return redirect(url_for("employees"))

@app.route("/employees/time-off/new", methods=["POST"])
@role_required("employees")
def employee_time_off_new():
    start_date = request.form.get("start_date", "")
    end_date   = request.form.get("end_date", "")

    if not start_date or not end_date:
        flash("Start and end dates are required.", "danger")
        return redirect(url_for("employees"))

    req_type = request.form.get("type", "Vacation")
    hours_per_day = max(0.5, min(8.0, _safe_float(request.form.get("hours_per_day", 8))))
    if req_type == "Half Day":
        working_days = _count_working_days(start_date, end_date)
        hours_per_day = 4.0
    elif req_type == "Unpaid":
        working_days = _count_working_days(start_date, end_date)
        hours_per_day = hours_per_day
    else:
        working_days = _count_working_days(start_date, end_date)
    total_hours = round(working_days * hours_per_day, 2)
    fb_push("/time_off_requests", {
        "employee_uid":    session.get("user_uid", ""),
        "employee_name":   session.get("user_name", ""),
        "type":            req_type,
        "half_day_period": request.form.get("half_day_period", "") if req_type == "Half Day" else "",
        "start_date":      start_date,
        "end_date":        end_date,
        "working_days":    working_days,
        "hours_per_day":   hours_per_day,
        "total_hours":     total_hours,
        "reason":          request.form.get("reason", "").strip(),
        "status":          "Pending",
        "requested_at":    datetime.now(timezone.utc).isoformat(),
        "reviewed_by":     "",
        "reviewed_at":     "",
        "review_note":     "",
    })
    flash("Time off request submitted.", "success")
    return redirect(url_for("employees") + "#time-off")

@app.route("/employees/time-off/<request_id>/<action>", methods=["POST"])
@role_required("employees")
def employee_time_off_action(request_id, action):
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("You don't have permission to do that.", "danger")
        return redirect(url_for("employees"))

    if action not in ("approve", "reject"):
        flash("Invalid action.", "danger")
        return redirect(url_for("employees"))

    new_status = "Approved" if action == "approve" else "Rejected"
    fb_update(f"/time_off_requests/{request_id}", {
        "status":      new_status,
        "reviewed_by": session.get("user_name", ""),
        "reviewed_at": datetime.now(timezone.utc).isoformat(),
    })
    flash(f"Time off request {new_status.lower()}.", "success")
    return redirect(url_for("employees") + "#time-off")

@app.route("/employees/<uid>/update", methods=["POST"])
@role_required("employees")
def employee_update(uid):
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("You don't have permission to do that.", "danger")
        return redirect(url_for("employees"))

    updates = {
        "hourly_rate":     _safe_float(request.form.get("hourly_rate", 0)),
        "department":      request.form.get("department", "").strip(),
        "hire_date":       request.form.get("hire_date", "").strip(),
        "employee_status": request.form.get("employee_status", "Active"),
        "updated_at":      datetime.now(timezone.utc).isoformat(),
    }
    if request.form.get("title") is not None:
        updates["title"] = request.form.get("title", "").strip()
    if request.form.get("region") is not None:
        updates["region"] = request.form.get("region", "").strip()
    if request.form.get("monthly_salary") is not None:
        updates["monthly_salary"] = _safe_float(request.form.get("monthly_salary", 0))
    fb_update(f"/users/{uid}", updates)
    flash("Employee details updated.", "success")
    return redirect(url_for("employees") + "#team")

@app.route("/api/users/<uid>/pages", methods=["PATCH"])
@role_required("settings")
def user_pages_update(uid):
    """Set custom page list for a user. Empty list = revert to role default."""
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data  = request.get_json() or {}
    pages = data.get("custom_pages", [])
    # Validate — only known page keys accepted
    pages = [p for p in pages if p in ALL_PAGES]
    fb_update(f"/users/{uid}", {
        "custom_pages": pages if pages else None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    return jsonify({"ok": True, "custom_pages": pages})

@app.route("/api/users/<uid>/details", methods=["PATCH"])
@role_required("settings")
def user_details_update(uid):
    """Update employee profile fields (title, region, rates) from Settings or Directory."""
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data = request.get_json() or {}
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ("username", "title", "region", "employee_type"):
        if field in data:
            updates[field] = str(data[field]).strip()
    for field in ("hourly_rate", "monthly_salary", "commission_rate"):
        if field in data:
            updates[field] = _safe_float(data[field])
    if "role" in data:
        allowed_roles = {"sales", "projects", "finance", "engineer", "admin"}
        r = str(data["role"]).strip().lower()
        if r in allowed_roles:
            updates["role"] = r
            # Auto-reset custom_pages to role default when role changes
            updates["custom_pages"] = None
    fb_update(f"/users/{uid}", updates)
    return jsonify({"ok": True})

@app.route("/api/commission/mark-paid", methods=["POST"])
@role_required("quotes")
def commission_mark_paid():
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data = request.get_json() or {}
    period     = str(data.get("period", "")).strip()
    sp_name    = str(data.get("salesperson", "")).strip()
    amount     = _safe_float(data.get("amount", 0))
    action     = str(data.get("action", "pay")).strip()   # "pay" or "unpay"
    if not period or not sp_name:
        return jsonify({"error": "period and salesperson are required"}), 400
    if action == "unpay":
        # Remove the payment record
        raw = fb_get("/commission_payments") or {}
        if isinstance(raw, dict):
            for cpid, cp in raw.items():
                if cp and cp.get("period") == period and cp.get("salesperson") == sp_name:
                    fb_delete(f"/commission_payments/{cpid}")
                    break
        return jsonify({"ok": True, "action": "unpaid"})
    # Mark as paid — store record
    _new_id = f"{period}_{sp_name.replace(' ', '_')}_{int(datetime.now().timestamp())}"
    fb_update(f"/commission_payments/{_new_id}", {
        "period":      period,
        "salesperson": sp_name,
        "amount":      amount,
        "paid_at":     datetime.now(timezone.utc).isoformat(),
        "paid_by":     session.get("user_name", ""),
    })
    return jsonify({"ok": True, "action": "paid"})

@app.route("/employees/export-hours")
@role_required("employees")
def employee_export_hours():
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("You don't have permission to do that.", "danger")
        return redirect(url_for("employees") + "#team")

    import csv
    import io as _io

    period = request.args.get("period", "week")
    custom_start = request.args.get("start", "")
    custom_end = request.args.get("end", "")
    period_start, period_end, period_label = _period_range(period, custom_start, custom_end)

    all_entries = _load_time_entries()
    period_entries = [e for e in all_entries if period_start <= e.get("date", "") <= period_end]
    hours_by_project = _aggregate_hours_by_project(period_entries)

    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow([f"Hours by Project — {period_label}"])
    writer.writerow(["Project", "Employee", "Hours"])
    grand_total = 0.0
    for proj, data in hours_by_project.items():
        for emp, minutes in data.items():
            if emp == "_total":
                continue
            writer.writerow([proj, emp, round(minutes / 60, 2)])
        writer.writerow([proj, "TOTAL", round(data["_total"] / 60, 2)])
        grand_total += data["_total"]
    writer.writerow([])
    writer.writerow(["Grand Total", "", round(grand_total / 60, 2)])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    return send_file(
        _io.BytesIO(csv_bytes),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"hours_by_project_{period_start}_to_{period_end}.csv"
    )

@app.route("/employees/time-entries/<entry_id>/close", methods=["POST"])
@role_required("employees")
def employee_close_entry(entry_id):
    entry = fb_get(f"/time_entries/{entry_id}")
    if not entry or entry.get("status") != "open":
        flash("Time entry not found or already closed.", "danger")
        return redirect(url_for("employees") + "#team")

    is_owner = entry.get("employee_uid") == session.get("user_uid", "")
    is_admin = normalize_role(session.get("user_role", "")) == "admin"
    if not (is_owner or is_admin):
        flash("You don't have permission to do that.", "danger")
        return redirect(url_for("employees"))

    redirect_tab = "#clock" if is_owner else "#team"

    try:
        clock_out = datetime.fromisoformat(request.form.get("clock_out", ""))
        clock_in = datetime.fromisoformat(entry["clock_in"])
        duration = round((clock_out - clock_in).total_seconds() / 60.0, 1)
    except ValueError:
        duration = -1

    if duration < 0:
        flash("Invalid clock-out time — it must be after the clock-in time.", "danger")
        return redirect(url_for("employees") + redirect_tab)

    fb_update(f"/time_entries/{entry_id}", {
        "clock_out":        clock_out.isoformat(),
        "duration_minutes": duration,
        "status":           "closed",
        "closed_by":        session.get("user_name", ""),
    })
    flash(f"Closed {entry.get('employee_name','')}'s entry — {round(duration/60, 1)}h logged.", "success")
    return redirect(url_for("employees") + redirect_tab)

# ── Routes: Settings ──────────────────────────────────────────────────────────
@app.route("/settings")
@role_required("settings")
def settings():
    all_users = _load_all_users()
    # Normalize role to lowercase so template selectattr filters match reliably
    for _u in all_users:
        if isinstance(_u, dict) and _u.get("role"):
            _u["role"] = normalize_role(_u["role"])
    settings_data = load_settings()
    return render_template("settings.html", users=all_users, settings=settings_data,
                           role_pages=ROLE_PAGES, all_pages=ALL_PAGES, page_labels=PAGE_LABELS)

@app.route("/settings/company", methods=["POST"])
@role_required("settings")
def settings_company():
    existing = load_settings()
    co = existing.get("company", {})
    co.update({
        "name":             request.form.get("name", ""),
        "address":          request.form.get("address", ""),
        "email":            request.form.get("email", ""),
        "phone":            request.form.get("phone", ""),
        "website":          request.form.get("website", ""),
        "default_tax_rate": _safe_float(request.form.get("default_tax_rate", "0")),
        "default_terms":    request.form.get("default_terms", ""),
    })
    # Save to Firebase and local settings.json
    fb_update("/settings", {"company": co})
    _save_local_settings_key("company", co)
    flash("Company settings saved.", "success")
    return redirect(url_for("settings"))

@app.route("/settings/logo", methods=["POST"])
@role_required("settings")
def settings_logo():
    from werkzeug.utils import secure_filename
    logo_file = request.files.get("logo")
    if logo_file and logo_file.filename:
        ext = Path(logo_file.filename).suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".bmp"):
            flash("Unsupported file type. Use PNG or JPG.", "danger")
            return redirect(url_for("settings"))
        # Normalize extension: save .jpeg as .jpg
        if ext == ".jpeg":
            ext = ".jpg"
        ASSETS_DIR.mkdir(parents=True, exist_ok=True)
        # Delete old logo files to avoid conflicts
        for old_file in ASSETS_DIR.glob("company_logo.*"):
            try:
                old_file.unlink()
            except Exception:
                pass
        save_path = ASSETS_DIR / f"company_logo{ext}"
        logo_file.save(str(save_path))
        existing = load_settings()
        co = existing.get("company", {})
        co["logo_path"] = str(save_path.resolve())
        fb_update("/settings", {"company": co})
        _save_local_settings_key("company", co)
        flash("Logo uploaded successfully.", "success")
    else:
        flash("No file selected.", "warning")
    return redirect(url_for("settings"))

@app.route("/settings/email", methods=["POST"])
@role_required("settings")
def settings_email():
    email_cfg = {
        "enabled":              request.form.get("enabled") == "on",
        "smtp_host":            request.form.get("smtp_host", "smtp.gmail.com"),
        "smtp_port":            int(request.form.get("smtp_port", 587) or 587),
        "smtp_user":            request.form.get("smtp_user", ""),
        "smtp_password":        request.form.get("smtp_password", ""),
        "from_name":            request.form.get("from_name", ""),
        "reminder_days_before": int(request.form.get("reminder_days_before", 3) or 3),
    }
    fb_update("/settings", {"email": email_cfg})
    _save_local_settings_key("email", email_cfg)
    flash("Email settings saved.", "success")
    return redirect(url_for("settings"))

@app.route("/settings/app", methods=["POST"])
@role_required("settings")
def settings_app():
    app_cfg = {
        "theme":               request.form.get("theme", "light"),
        "log_level":           request.form.get("log_level", "INFO"),
        "auto_check_updates":  request.form.get("auto_check_updates") == "on",
    }
    fb_update("/settings", {"app": app_cfg})
    _save_local_settings_key("app", app_cfg)
    flash("App preferences saved.", "success")
    return redirect(url_for("settings"))

@app.route("/settings/ai", methods=["POST"])
@role_required("settings")
def settings_ai():
    ai_cfg = {"anthropic_key": request.form.get("anthropic_key", "").strip()}
    fb_update("/settings", {"ai": ai_cfg})
    _save_local_settings_key("ai", ai_cfg)
    flash("AI settings saved.", "success")
    return redirect(url_for("settings") + "?tab=ai")


# ── AI Routes ─────────────────────────────────────────────────────────────────

@app.route("/ai/extract-pdf", methods=["POST"])
@role_required("financial")
def ai_extract_pdf():
    """Extract expense fields from an uploaded receipt (PDF or image) using Claude."""
    f = request.files.get("pdf")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    custom_categories = fb_get("/custom_categories") or {}
    expense_types = custom_categories.get("expense_type", []) if isinstance(custom_categories.get("expense_type"), list) else []
    categories_by_type = custom_categories.get("Categories", {}) if isinstance(custom_categories.get("Categories"), dict) else {}

    # Provide default values if empty (matching reference structure)
    if not expense_types:
        expense_types = [
            "O & M (Operations & Maintenance)",
            "Capital Expenses",
            "Other Expenses"
        ]

    if not categories_by_type:
        categories_by_type = {
            "O & M (Operations & Maintenance)": [
                "Facilities & Utilities",
                "Office & Admin Overhead",
                "Engineering Software & IT",
                "Salaries, Labor & Related Costs",
                "Professional Services",
                "Insurance & Compliance",
                "Travel, Site Visits & Vehicles",
                "Marketing & Business Development",
                "Training, Licensure & Development",
                "Safety & Field Supplies",
                "Miscellaneous O & M"
            ],
            "Capital Expenses": [
                "Computer & Office Equipment",
                "Field & Inspection Equipment",
                "Furniture & Fixtures",
                "Vehicles",
                "Software (Capitalized)",
                "Leasehold Improvements",
                "Accumulated Depreciation"
            ],
            "Other Expenses": [
                "Salary/Bonuses",
                "Tax Expenses/Tax Deductions",
                "Medical/Benefits",
                "Meals & Entertainment",
                "Donations",
                "Bank Charges",
                "Contingency Funds",
                "Unexpected Costs"
            ]
        }

    type_hint = f"Pick from these expense types if one fits: {expense_types}. " if expense_types else ""
    cat_hint = f"Pick from these categories if one fits: {sorted(set(c for cats in categories_by_type.values() for c in cats))}. " if categories_by_type else ""

    fields_prompt = f"""Return ONLY valid JSON with these fields:
expense_name (short description of the expense), expense_type (a general type/department for this expense; {type_hint}otherwise make a reasonable guess), category (a specific category for this expense; {cat_hint}otherwise pick one of: Labor, Materials, Equipment, Subcontractor, Overhead, Travel, Other), amount (number, no currency symbol), date (YYYY-MM-DD or blank), vendor.
If a field is not found leave it blank. Return only the JSON object, nothing else."""

    content_type = (f.content_type or "").lower()

    if content_type.startswith("image/"):
        try:
            image_b64 = base64.b64encode(f.read()).decode("utf-8")
        except Exception as e:
            return jsonify({"error": f"Could not read image: {e}"}), 400
        prompt = f"This image is a receipt or invoice. {fields_prompt}"
        try:
            result = _ai_call_with_image(prompt, image_b64, content_type)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        if not PYPDF_AVAILABLE:
            return jsonify({"error": "pypdf not installed on server"}), 500
        try:
            reader = _PdfReader(f)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)[:4000]
        except Exception as e:
            return jsonify({"error": f"Could not read PDF: {e}"}), 400
        prompt = f"""Extract expense information from this document text. {fields_prompt}

Document text:
{text}"""
        try:
            result = _ai_call(prompt)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 503
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    try:
        return jsonify(json.loads(result))
    except json.JSONDecodeError:
        import re
        m = re.search(r'\{.*\}', result, re.DOTALL)
        if m:
            return jsonify(json.loads(m.group(0)))
        return jsonify({"error": "AI returned unexpected format", "raw": result}), 500


@app.route("/ai/draft-reminder/<invoice_id>", methods=["POST"])
@role_required("invoicing")
def ai_draft_reminder(invoice_id):
    """Draft a professional overdue reminder email for an invoice."""
    inv_data = fb_get(f"/invoices/{invoice_id}")
    if not inv_data:
        return jsonify({"error": "Invoice not found"}), 404
    meta = inv_data.get("meta", {})
    today = datetime.now().strftime("%Y-%m-%d")
    due   = meta.get("due_date", "")
    try:
        days_overdue = (datetime.fromisoformat(today) - datetime.fromisoformat(due)).days if due else 0
    except Exception:
        days_overdue = 0
    prompt = f"""Write a professional but friendly overdue payment reminder email.
Invoice: {meta.get('invoice_number')}
Client: {meta.get('client_name')}
Amount outstanding: ${_safe_float(meta.get('total',0)) - _safe_float(meta.get('amount_paid',0)):,.2f}
Due date: {due}
Days overdue: {days_overdue}
Company sending this: {company_info().get('name','MABS Engineering')}

Keep it concise (3-4 short paragraphs), professional, and include a clear call to action.
Use placeholders like [Your Name] for signature. Do not use asterisks for formatting."""
    try:
        email_text = _ai_call(prompt, max_tokens=512)
        subject = f"Payment Reminder — {meta.get('invoice_number')} (${_safe_float(meta.get('total',0)):,.0f})"
        return jsonify({"subject": subject, "body": email_text})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/ai/draft-quote-email/<quote_id>", methods=["POST"])
@role_required("quotes")
def ai_draft_quote_email(quote_id):
    """Generate a professional covering email for sending a quote PDF to the client."""
    ai_client = _get_ai_client()
    if not ai_client:
        return jsonify({"error": "AI not configured. Add your Anthropic API key in Settings → AI."})
    quote = fb_get(f"/job_forms/{quote_id}")
    if not quote:
        return jsonify({"error": "Quote not found."}), 404
    co       = company_info()
    services = ", ".join(quote.get("service_types") or []) or "engineering services"
    total    = _safe_float(quote.get("total", 0))
    prompt   = f"""Write a concise professional email to accompany a quote PDF sent to a client.

Sender company : {co.get('name', 'Our Company')}
Client         : {quote.get('client_name', 'Client')}
Quote number   : {quote.get('job_number', '')}
Project / Scope: {quote.get('project_name', '')}
Services       : {services}
Quote total    : ${total:,.2f}
Valid until    : {quote.get('valid_until', 'N/A')}
Salesperson    : {quote.get('salesperson', '')}

Write 3-4 short paragraphs:
1. Greeting + purpose of the email
2. Brief summary of the scope and quote total
3. Validity period + invitation to discuss or ask questions
4. Professional closing with [Your Name] placeholder

Return JSON only: {{"subject": "...", "body": "..."}}
No markdown, no asterisks in the body."""
    try:
        raw    = _ai_call(prompt, max_tokens=600)
        import re as _re
        match  = _re.search(r'\{.*\}', raw, _re.DOTALL)
        parsed = json.loads(match.group()) if match else {}
        subject = parsed.get("subject") or f"Quote {quote.get('job_number','')} — {quote.get('project_name','')}"
        body    = parsed.get("body") or raw
        return jsonify({"subject": subject, "body": body})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/ai/cash-flow-summary", methods=["POST"])
@role_required("financial")
def ai_cash_flow_summary():
    """Generate a plain-English cash flow narrative from upcoming payment data."""
    data = request.get_json(force=True) or {}
    upcoming = data.get("upcoming", [])
    total_expected = sum(float(u.get("amount", 0)) for u in upcoming)
    items_text = "\n".join(
        f"- {u.get('date','?')}: {u.get('label','?')} — ${float(u.get('amount',0)):,.0f}"
        for u in upcoming[:20]
    )
    prompt = f"""Based on these upcoming expected payments for the next 90 days, write a brief 2-3 sentence cash flow summary. Be specific about amounts and timing. No bullet points.

Total expected: ${total_expected:,.0f}
Upcoming items:
{items_text or 'No upcoming payments found.'}"""
    try:
        summary = _ai_call(prompt, max_tokens=200)
        return jsonify({"summary": summary})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/ai/project-health", methods=["POST"])
@role_required("projects")
def ai_project_health():
    """Return one-line health summaries for all projects."""
    projects_list = _load_projects_list()
    invoices_raw  = fb_get("/invoices") or {}
    inv_list = [v for v in invoices_raw.values() if isinstance(v, dict)] if isinstance(invoices_raw, dict) else []

    summaries = {}
    for p in projects_list:
        pnum   = p.get("project_number", "")
        paid   = _safe_float(p.get("amount_paid", 0))
        cv     = _safe_float(p.get("contract_value", 0))
        status = p.get("status", "Not Started")
        pct    = int(paid / cv * 100) if cv > 0 else 0
        stages = p.get("payment_stages", [])
        pending_stages = [s for s in stages if isinstance(s, dict) and s.get("status") in ("Pending", "Invoiced")]
        next_stage_amt = _safe_float(pending_stages[0].get("amount", 0)) if pending_stages else 0
        summary_input = (
            f"Project: {p.get('project_name','')} ({pnum}), Status: {status}, "
            f"Contract: ${cv:,.0f}, Paid: ${paid:,.0f} ({pct}%), "
            f"Pending installments: {len(pending_stages)}, Next due: ${next_stage_amt:,.0f}"
        )
        summaries[pnum] = summary_input

    if not summaries:
        return jsonify({"summaries": {}})

    bulk = "\n".join(f"{k}: {v}" for k, v in summaries.items())
    prompt = f"""For each project below, write exactly ONE short sentence (max 15 words) summarizing its health and next action. Return ONLY valid JSON like {{"PROJ-001": "sentence", "PROJ-002": "sentence"}}.

{bulk}"""
    try:
        result = _ai_call(prompt, max_tokens=600)
        import re
        m = re.search(r'\{.*\}', result, re.DOTALL)
        parsed = json.loads(m.group(0)) if m else {}
        return jsonify({"summaries": parsed})
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 503
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/settings/user/new", methods=["POST"])
@role_required("settings")
def user_new():
    username = request.form.get("username", "").strip()
    email    = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")
    role     = normalize_role(request.form.get("role", "sales"))

    if not all([username, email, password]):
        flash("All fields are required.", "danger")
        return redirect(url_for("settings") + "?tab=users")

    if not FIREBASE_AVAILABLE:
        flash("Firebase not available.", "danger")
        return redirect(url_for("settings") + "?tab=users")

    try:
        from firebase_admin import auth as fb_auth
        user = fb_auth.create_user(email=email, password=password,
                                   display_name=username, email_verified=False)
        user_data = {
            "username":       username,
            "email":          email,
            "role":           role,
            "active":         True,
            "firebase_uid":   user.uid,
            "title":           request.form.get("title", "").strip(),
            "region":          request.form.get("region", "").strip(),
            "hourly_rate":     _safe_float(request.form.get("hourly_rate", 0)),
            "monthly_salary":  _safe_float(request.form.get("monthly_salary", 0)),
            "employee_type":   request.form.get("employee_type", "").strip(),
            "commission_rate": _safe_float(request.form.get("commission_rate", 0)),
            "created_at":     datetime.now(timezone.utc).isoformat(),
            "updated_at":     datetime.now(timezone.utc).isoformat(),
        }
        fb_update(f"/users/{user.uid}", user_data)
        flash(f"User {username} created.", "success")
    except Exception as exc:
        flash(f"Error creating user: {exc}", "danger")

    return redirect(url_for("settings") + "?tab=users")

@app.route("/settings/user/<uid>/toggle", methods=["POST"])
@role_required("settings")
def user_toggle(uid):
    profile = fb_get(f"/users/{uid}") or {}
    current = profile.get("active", True)
    fb_update(f"/users/{uid}", {"active": not current,
                                "updated_at": datetime.now(timezone.utc).isoformat()})
    flash("User status updated.", "success")
    return redirect(url_for("settings") + "?tab=users")

@app.route("/settings/user/<uid>/role", methods=["POST"])
@role_required("settings")
def user_role_update(uid):
    new_role = normalize_role(request.form.get("role", "sales"))
    fb_update(f"/users/{uid}", {"role": new_role,
                                "updated_at": datetime.now(timezone.utc).isoformat()})
    flash("User role updated.", "success")
    return redirect(url_for("settings") + "?tab=users")

@app.route("/settings/user/<uid>/delete", methods=["POST"])
@role_required("settings")
def user_delete(uid):
    if uid == session.get("user_uid"):
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("settings") + "?tab=users")
    try:
        if FIREBASE_AVAILABLE:
            from firebase_admin import auth as fb_auth
            try:
                fb_auth.delete_user(uid)
            except Exception:
                pass
        fb_delete(f"/users/{uid}")
        flash("User deleted.", "success")
    except Exception as exc:
        flash(f"Error: {exc}", "danger")
    return redirect(url_for("settings") + "?tab=users")

# ── Helper: get company logo for PDF generation ────────────────────────────────
def _get_company_logo_path():
    """Get the company logo path from settings or fallback candidates."""
    try:
        settings = load_settings()
        logo_path = settings.get("company", {}).get("logo_path", "")
        candidates = [
            Path(logo_path) if logo_path else None,
            ASSETS_DIR / "company_logo.jpg",
            ASSETS_DIR / "company_logo.png",
            ASSETS_DIR / "company_logo.jpeg",
            DATA_DIR / "company_logo.jpg",
            DATA_DIR / "company_logo.png",
            DATA_DIR / "company_logo.jpeg",
            ASSETS_DIR / "logo.jpg",
            ASSETS_DIR / "logo.png",
            DATA_DIR / "logo.png",
        ]
        for p in candidates:
            if p:
                try:
                    p_obj = Path(p) if isinstance(p, str) else p
                    if p_obj.exists() and p_obj.is_file():
                        return str(p_obj.resolve())
                except Exception:
                    continue
    except Exception:
        pass
    return None

# ── Route: serve company logo ─────────────────────────────────────────────────
@app.route("/logo")
def company_logo():
    """Serve the company logo from its configured path."""
    settings = load_settings()
    logo_path = settings.get("company", {}).get("logo_path", "")
    candidates = [
        Path(logo_path) if logo_path else None,
        DATA_DIR / "company_logo.png",
        DATA_DIR / "company_logo.jpg",
        DATA_DIR / "logo.png",
        ASSETS_DIR / "logo.png",
        ASSETS_DIR / "logo.jpg",
    ]
    for p in candidates:
        if p and p.exists():
            mime = "image/png" if p.suffix.lower() == ".png" else "image/jpeg"
            return send_file(str(p), mimetype=mime)
    abort(404)

# ── API: invoice number ───────────────────────────────────────────────────────
@app.route("/api/next-invoice-number")
@login_required
def api_next_invoice():
    return jsonify({"number": _next_invoice_number()})

@app.route("/api/client/<client_name>")
@login_required
def api_client(client_name):
    data = fb_get(f"/clients/{client_name}") or {}
    return jsonify(data)

def _find_quote_by_number(job_number: str):
    """Look up a /job_forms quote by its job_number. Returns (firebase_id, data) or (None, None)."""
    job_number = (job_number or "").strip()
    if not job_number:
        return None, None
    raw = fb_get("/job_forms") or {}
    if isinstance(raw, dict):
        for qid, qdata in raw.items():
            if isinstance(qdata, dict) and qdata.get("job_number", "").strip() == job_number:
                return qid, qdata
    return None, None

def _quote_to_project_fields(quote: dict) -> dict:
    """Map a /job_forms quote record onto the New Project form's field names."""
    service_costs = {}
    for item in quote.get("line_items", []) or []:
        if isinstance(item, dict):
            desc = (item.get("description") or "").strip()
            if desc:
                service_costs[desc] = item.get("total") or item.get("unit_price") or "0"
    return {
        "project_name":   quote.get("project_name", "") or quote.get("description", ""),
        "client_name":    quote.get("client_name", ""),
        "sales":          quote.get("salesperson", ""),
        "site_address":   "",
        "mail_address":   "",
        "plant":          "",
        "service_types":  quote.get("service_types") or [],
        "service_costs":  service_costs,
        "scope_of_work":  quote.get("description", ""),
        "expedite":       "Yes" if quote.get("is_expedited") else "No",
        "rush_rate":      quote.get("rush_rate", "50"),
        "contract_value": f"{_safe_float(quote.get('total', 0)):.2f}",
        "start_date":     "",
        "end_date":       quote.get("expected_completion", ""),
    }

@app.route("/api/quote-by-number/<quote_number>")
@login_required
def api_quote_by_number(quote_number):
    """Return quote data for a given job_number so the project form can auto-fill."""
    _, qdata = _find_quote_by_number(quote_number)
    if qdata:
        fields = _quote_to_project_fields(qdata)
        fields["found"] = True
        return jsonify(fields)
    return jsonify({"found": False})

# ── Helpers ───────────────────────────────────────────────────────────────────
def _save_local_settings_key(key: str, value) -> None:
    """Persist a single top-level key in data/settings.json."""
    try:
        sf = DATA_DIR / "settings.json"
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        data = {}
        if sf.exists():
            with open(sf, encoding="utf-8") as f:
                data = json.load(f)
        data[key] = value
        with open(sf, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as exc:
        log.warning("Could not save settings.json: %s", exc)

def _safe_float(val) -> float:
    try:
        return float(str(val or 0).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0

def _format_date_display(date_str: str) -> str:
    """Convert date from YYYY-MM-DD format to MM-DD-YYYY for display."""
    if not date_str or date_str == "—":
        return "—"
    try:
        # Handle both YYYY-MM-DD and other formats
        if len(str(date_str)) >= 10:
            date_str = str(date_str)[:10]  # Take only the date part if it's a datetime
            parts = date_str.split("-")
            if len(parts) == 3:
                year, month, day = parts
                return f"{month}-{day}-{year}"
    except (ValueError, TypeError, IndexError):
        pass
    return str(date_str) if date_str else "—"

def _format_date_invoice(date_str: str) -> str:
    """Convert date from YYYY-MM-DD format to MM-DD-YY (2-digit year) for invoice display."""
    if not date_str or date_str == "—":
        return "—"
    try:
        date_str = str(date_str).strip()[:10]  # Take only the date part if it's a datetime
        parts = date_str.split("-")
        if len(parts) == 3:
            if len(parts[0]) == 4:  # Year is 4 digits (YYYY-MM-DD)
                year, month, day = parts
                year_2digit = year[-2:]  # Take last 2 digits (2026 → 26)
                return f"{month}-{day}-{year_2digit}"
    except (ValueError, TypeError, IndexError):
        pass
    return str(date_str) if date_str else "—"

def _format_date_input(date_str: str) -> str:
    """Convert date from display format MM-DD-YYYY to input format YYYY-MM-DD."""
    if not date_str:
        return ""
    try:
        # If already in YYYY-MM-DD format, return as-is
        if date_str and len(date_str) == 10 and date_str[4] == "-":
            return date_str
        # Convert from MM-DD-YYYY to YYYY-MM-DD
        parts = date_str.replace("/", "-").split("-")
        if len(parts) == 3:
            # Detect format based on part sizes
            if len(parts[0]) == 4:  # YYYY-MM-DD
                return date_str
            else:  # MM-DD-YYYY
                month, day, year = parts
                return f"{year}-{month}-{day}"
    except (ValueError, TypeError):
        pass
    return date_str

def _get_next_payment_stage(project: dict, all_invoices: dict = None) -> dict:
    """Get next uninvoiced payment stage.
    Returns: {stage_idx, stage_name, amount, blocked, reason}
    """
    if all_invoices is None:
        all_invoices = fb_get("/invoices") or {}

    proj_num = project.get("project_number", "")
    payment_stages = project.get("payment_stages", [])

    if not isinstance(payment_stages, list) or not payment_stages:
        return {"stage_idx": None, "stage_name": None, "amount": 0, "blocked": True, "reason": "No payment stages defined"}

    # Find which stages have been invoiced
    invoiced_stages = set()
    if isinstance(all_invoices, dict):
        for inv_id, inv_data in all_invoices.items():
            if isinstance(inv_data, dict):
                # Check single-project invoices
                inv_proj = inv_data.get("meta", {}).get("project_number", "")
                if inv_proj == proj_num:
                    stage_idx = inv_data.get("meta", {}).get("payment_stage_index")
                    if stage_idx is not None:
                        invoiced_stages.add(int(stage_idx))
                        print(f"[DETECT] Single-project invoice {inv_id}: proj={proj_num}, stage={stage_idx}", flush=True)

                # Check multi-project invoices (linked_projects field)
                linked_projects = inv_data.get("meta", {}).get("linked_projects", [])
                if isinstance(linked_projects, list) and len(linked_projects) > 0:
                    for linked in linked_projects:
                        if isinstance(linked, dict) and linked.get("project_number") == proj_num:
                            stage_idx = linked.get("payment_stage_index")
                            if stage_idx is not None:
                                invoiced_stages.add(int(stage_idx))
                                print(f"[DETECT] Multi-project invoice {inv_id}: proj={proj_num}, stage={stage_idx}, linked_projects={linked_projects}", flush=True)

    # Find first stage NOT invoiced
    print(f"[DETECT] Project {proj_num}: total_stages={len(payment_stages)}, invoiced_stages={invoiced_stages}", flush=True)
    for idx, stage in enumerate(payment_stages):
        if isinstance(stage, dict) and idx not in invoiced_stages:
            stage_name = stage.get("name", f"Stage {idx + 1}")
            amount = _safe_float(stage.get("amount", 0))
            blocked = False
            reason = "Ready to invoice"

            # Check if previous stages are invoiced
            if idx > 0:
                # All previous stages must be invoiced
                for prev_idx in range(idx):
                    if prev_idx not in invoiced_stages:
                        blocked = True
                        reason = f"Previous stage(s) not yet invoiced"
                        break

            print(f"[DETECT] Returning stage {idx} for {proj_num}: {stage_name}", flush=True)
            return {
                "stage_idx": idx,
                "stage_name": stage_name,
                "amount": amount,
                "blocked": blocked,
                "reason": reason
            }

    print(f"[DETECT] Project {proj_num} is fully invoiced", flush=True)
    return {"stage_idx": None, "stage_name": None, "amount": 0, "blocked": True, "reason": "All payment stages already invoiced"}

def _enrich_projects_with_next_stage(projects: list, all_invoices: dict = None) -> list:
    """Annotate each project dict with its next-pending-payment-stage info
    (used by the invoice form's "Active projects for this client" picker so
    it bills the next installment instead of the full outstanding balance)."""
    if all_invoices is None:
        all_invoices = fb_get("/invoices") or {}
    for p in projects:
        if isinstance(p, dict):
            detection = _get_next_payment_stage(p, all_invoices)
            p["next_stage_idx"]    = detection.get("stage_idx")
            p["next_stage_name"]   = detection.get("stage_name") or ""
            p["next_stage_amount"] = detection.get("amount", 0)
    return projects

def _find_project_by_number(project_number: str):
    """Return (firebase_id, project_dict) for the project with this number, or (None, None)."""
    raw_proj = fb_get("/projects") or {}
    if isinstance(raw_proj, dict):
        for pid, pdata in raw_proj.items():
            if isinstance(pdata, dict) and pdata.get("project_number", "") == project_number:
                return pid, pdata
    return None, None

def _mark_project_stage(project_number: str, stage_index: int, status: str, invoice_id: str = None, invoice_number: str = None, amount: float = None, amount_paid: float = None) -> None:
    """Update one stage's status (and optionally its linked invoice id/number/amount/amount_paid) within a project's payment plan."""
    pid, pdata = _find_project_by_number(project_number)
    print(f"[MARK_STAGE] project_number={project_number}, stage_idx={stage_index}, status={status}, amount={amount}, amount_paid={amount_paid}, pid={pid}", flush=True)
    if not pid:
        print(f"[MARK_STAGE] Project not found!", flush=True)
        return
    stages = pdata.get("payment_stages") or []
    if not (0 <= stage_index < len(stages)) or not isinstance(stages[stage_index], dict):
        print(f"[MARK_STAGE] Stage index out of range! stages_count={len(stages)}, idx={stage_index}", flush=True)
        return
    stages[stage_index]["status"] = status
    if invoice_id is not None:
        stages[stage_index]["invoice_id"] = invoice_id
    if invoice_number is not None:
        stages[stage_index]["invoice_number"] = invoice_number
    if amount is not None:
        stages[stage_index]["amount"] = amount
    if amount_paid is not None:
        stages[stage_index]["amount_paid"] = str(amount_paid)
    # When reverting to "Pending Invoice", clear the invoice tracking fields
    if status == "Pending Invoice":
        print(f"[MARK_STAGE] Clearing invoice_id and invoice_number", flush=True)
        stages[stage_index].pop("invoice_id", None)
        stages[stage_index].pop("invoice_number", None)
    print(f"[MARK_STAGE] Updated stage: {stages[stage_index]}", flush=True)

    proj_updates = {"payment_stages": stages, "updated_at": datetime.now(timezone.utc).isoformat()}

    # Auto-update project status when a stage first gets invoiced and $0 collected
    if status == "Invoiced":
        proj_status = pdata.get("status", "Not Started")
        proj_paid   = _safe_float(pdata.get("amount_paid", 0))
        if proj_paid == 0 and proj_status in ("In Progress", "Active", "Not Started"):
            proj_updates["status"] = "invoiced_Not paid yet"

    fb_update(f"/projects/{pid}", proj_updates)

def _calculate_invoice_status(inv_data: dict) -> str:
    """Calculate invoice status based on payments vs total (including tax).

    Returns: "Paid", "Partial", or "Overdue" based on actual payments.
    If no payments, returns the stored status (user-selected).
    """
    meta = inv_data.get("meta", {}) or {}

    # Always calculate from actual payments, not from stored status
    invoice_total = _safe_float(meta.get("total", 0))
    tax_amount = _safe_float(meta.get("tax_amount", 0))
    invoice_subtotal = invoice_total - tax_amount  # Subtract tax from total to get line items amount

    # Get invoice payments (line items)
    payment_log = inv_data.get("payment_log", [])
    if not isinstance(payment_log, list):
        payment_log = []
    invoice_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log)

    # Get tax payments
    tax_log = inv_data.get("tax_payments", [])
    if not isinstance(tax_log, list):
        tax_log = []
    tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)

    # If invoice_paid covers both line items and tax (overpayment towards tax), credit it
    if invoice_paid > invoice_subtotal:
        tax_paid += (invoice_paid - invoice_subtotal)
        invoice_paid = invoice_subtotal

    # Check due date for Overdue
    due_date = meta.get("due_date", "")
    today = datetime.now().strftime("%Y-%m-%d")
    is_overdue = due_date and due_date < today

    # Determine status based on actual amounts
    invoice_paid_enough = invoice_paid >= (invoice_subtotal - 0.01)
    tax_paid_enough = tax_amount <= 0.01 or tax_paid >= (tax_amount - 0.01)

    if invoice_paid_enough and tax_paid_enough:
        return "Paid"
    elif invoice_paid > 0 or tax_paid > 0:
        return "Partial"
    elif is_overdue:
        return "Overdue"
    else:
        return meta.get("status", "Draft")

def _derive_stage_index_from_line_items(project_number: str, line_items: list) -> int:
    """Try to find stage_index from line items by looking for paid stage references."""
    if not isinstance(line_items, list):
        return -1

    # Look for stage name patterns in line item descriptions
    # e.g., "Testing1 — Installment 2 of 3" -> find stage named "Installment 2 of 3"
    pid, pdata = _find_project_by_number(project_number)
    if not pid:
        return -1

    stages = pdata.get("payment_stages", []) or []
    if not isinstance(stages, list):
        return -1

    # Build a map of generated stage names to indices
    # Stage names are generated the same way as in project_form.html:
    # "Installment 1 of N", "Installment 2 of N", etc.
    num_stages = len(stages)
    stage_name_map = {}
    for idx in range(num_stages):
        if num_stages == 1:
            stage_name = "Balance Payment"
        else:
            stage_name = f"Installment {idx + 1} of {num_stages}"
        stage_name_map[stage_name.lower()] = idx

    # Search line items for stage references
    for item in line_items:
        if not isinstance(item, dict):
            continue
        desc = item.get("description", "").lower()
        # Look for stage name in description (e.g., "installment 2 of 3")
        for stage_name, stage_idx in stage_name_map.items():
            if stage_name in desc:
                print(f"[DERIVE] Found stage {stage_idx} from line item: {desc}", flush=True)
                return stage_idx

    print(f"[DERIVE] Could not derive stage from line items for project {project_number}", flush=True)
    return -1

def _update_single_project_stage_payment_status(invoice_id: str, project_number: str, stage_index: int) -> None:
    """Update payment status for a SINGLE project/stage linked to an invoice.

    Unlike _update_project_stage_payment_status which updates all linked projects,
    this only updates the specific project/stage being edited.
    Used when editing a stage amount from project details to avoid affecting other projects.
    """
    # Find the specific project
    pid, pdata = _find_project_by_number(project_number)
    if not pid:
        return

    stages = pdata.get("payment_stages") or []
    if not (0 <= stage_index < len(stages)):
        return

    stage = stages[stage_index]
    stage_amount = _safe_float(stage.get("amount", 0))

    if stage_amount <= 0:
        return

    # Get all invoices for payment calculation
    all_invoices = fb_get("/invoices") or {}

    # Calculate payments for THIS project/stage from all invoices
    linked_invoice_id = None
    linked_invoice_number = None
    project_paid = 0

    print(f"[UPDATE_SINGLE_STAGE] Looking for invoices for project {project_number} stage {stage_index}", flush=True)

    if isinstance(all_invoices, dict):
        for inv_id, inv in all_invoices.items():
            if not isinstance(inv, dict):
                continue
            inv_meta = inv.get("meta", {}) or {}

            # Check if invoice covers this project and stage
            is_for_this_project = False

            # Check if this is the current invoice being updated
            if inv_id == invoice_id:
                if inv_meta.get("project_number") == project_number:
                    is_for_this_project = True
                else:
                    # For multi-project invoices, check linked_projects
                    linked_projs = inv_meta.get("linked_projects", [])
                    if isinstance(linked_projs, list):
                        for lp in linked_projs:
                            if isinstance(lp, dict):
                                if lp.get("project_number") == project_number and lp.get("payment_stage_index") == stage_index:
                                    is_for_this_project = True
                                    break
            else:
                # For other invoices, check if they're linked to this project/stage
                if (inv_meta.get("project_number") == project_number and
                    inv_meta.get("payment_stage_index") == stage_index):
                    is_for_this_project = True
                else:
                    linked_projs = inv_meta.get("linked_projects", [])
                    if isinstance(linked_projs, list):
                        for lp in linked_projs:
                            if isinstance(lp, dict):
                                if lp.get("project_number") == project_number and lp.get("payment_stage_index") == stage_index:
                                    is_for_this_project = True
                                    break

            if is_for_this_project:
                print(f"[UPDATE_SINGLE_STAGE] Found matching invoice {inv_id} for project {project_number} stage {stage_index}", flush=True)
                # Sum payments for this project
                inv_payment_log = inv.get("payment_log", [])
                if isinstance(inv_payment_log, list):
                    inv_payments = sum(
                        _safe_float(p.get("amount", 0))
                        for p in inv_payment_log
                        if p.get("project_number") == project_number or not p.get("project_number")
                    )
                    project_paid += inv_payments
                    # Track the invoice if it has payments or is the current one
                    if inv_payments > 0 or inv_id == invoice_id:
                        linked_invoice_id = inv_id
                        linked_invoice_number = inv_meta.get("invoice_number", "")
                        print(f"[UPDATE_SINGLE_STAGE] Set invoice_number={linked_invoice_number} for invoice {inv_id}", flush=True)

    # Determine status based on actual payments
    if project_paid >= (stage_amount - 0.01):
        new_status = "Paid"
    elif project_paid > 0:
        new_status = "Partially Paid"
    else:
        new_status = "Invoiced"

    # Update ONLY this stage
    stage["status"] = new_status
    stage["amount_paid"] = str(project_paid) if project_paid > 0 else "0"

    print(f"[UPDATE_SINGLE_STAGE] Before update: stage has invoice_number='{stage.get('invoice_number', 'MISSING')}'", flush=True)

    if linked_invoice_id:
        stage["invoice_id"] = linked_invoice_id
    if linked_invoice_number:
        stage["invoice_number"] = linked_invoice_number
        print(f"[UPDATE_SINGLE_STAGE] Setting invoice_number={linked_invoice_number}", flush=True)
    else:
        print(f"[UPDATE_SINGLE_STAGE] No invoice_number found! linked_invoice_id={linked_invoice_id}, linked_invoice_number={linked_invoice_number}", flush=True)

    # Save back to Firebase
    print(f"[UPDATE_SINGLE_STAGE] After update: stage has invoice_number='{stage.get('invoice_number', 'MISSING')}'", flush=True)
    fb_update(f"/projects/{pid}", {
        "payment_stages": stages,
        "updated_at": datetime.now(timezone.utc).isoformat()
    })
    print(f"[UPDATE_SINGLE_STAGE] Saved stage with invoice_number='{stage.get('invoice_number', 'MISSING')}'", flush=True)

def _update_project_stage_payment_status(invoice_id: str) -> None:
    """Update project stage statuses based on invoice payments.

    For each project linked to the invoice, sum payments made FOR THAT PROJECT
    and update the stage status to Paid/Partially Paid/Invoiced.
    """
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    meta = inv_data.get("meta", {}) or {}

    # Get linked projects from invoice
    linked_projects = meta.get("linked_projects", [])

    # Convert string format to dict format if needed
    # Handles both ['MABS-202606-003'] and [{"project_number": "...", "payment_stage_index": ...}]
    normalized_projects = []
    if linked_projects:
        for item in linked_projects:
            if isinstance(item, dict):
                normalized_projects.append(item)
            elif isinstance(item, str):
                # Legacy string format: use main invoice's stage index
                stage_idx = meta.get("payment_stage_index", -1)
                if stage_idx >= 0:
                    normalized_projects.append({"project_number": item, "payment_stage_index": stage_idx})

    # Fallback: if no linked_projects, build from main project
    if not normalized_projects:
        main_project = meta.get("project_number", "")
        stage_index = meta.get("payment_stage_index", -1)

        # If no stage_index in meta, try to derive from line items
        if stage_index < 0:
            line_items = inv_data.get("line_items", [])
            stage_index = _derive_stage_index_from_line_items(main_project, line_items)

        if main_project and stage_index >= 0:
            normalized_projects = [{"project_number": main_project, "payment_stage_index": stage_index}]
        elif main_project and stage_index < 0:
            # Invoice doesn't have payment_stage_index - we'll search for it by invoice_id later
            normalized_projects = [{"project_number": main_project, "payment_stage_index": -1, "invoice_id": invoice_id}]
        else:
            return

    # Get ALL invoices to sum payments for each stage
    all_invoices = fb_get("/invoices") or {}

    # Update each project's stage status based on payments FOR THAT PROJECT
    for proj_info in normalized_projects:
        if not isinstance(proj_info, dict):
            continue

        project_number = proj_info.get("project_number", "")
        stage_index = proj_info.get("payment_stage_index", -1)
        inv_id_for_lookup = proj_info.get("invoice_id")

        if not project_number:
            continue

        # Get project and stage info
        pid, pdata = _find_project_by_number(project_number)
        if not pid:
            continue

        stages = pdata.get("payment_stages") or []

        # If stage_index not available, search for stage by invoice_id
        if stage_index < 0:
            if inv_id_for_lookup:
                # Find the stage that has this invoice_id
                for idx, s in enumerate(stages):
                    if isinstance(s, dict) and s.get("invoice_id") == inv_id_for_lookup:
                        stage_index = idx
                        break
            if stage_index < 0:
                # Still not found
                continue

        if not (0 <= stage_index < len(stages)):
            continue

        stage = stages[stage_index]
        stage_amount = _safe_float(stage.get("amount", 0))

        if stage_amount <= 0:
            continue

        # For multi-project invoices, use the amount_paid that was set by sequential allocation
        # For single-project invoices, sum payments from the invoice
        is_multi_project = len(linked_projects) > 1

        project_paid = 0
        linked_invoice_id = None
        linked_invoice_number = None

        if is_multi_project:
            # For multi-project invoices, use the stage's amount_paid (not the project total)
            project_paid = _safe_float(stage.get("amount_paid", 0))

            # Also find and set the invoice_number for this stage
            if isinstance(all_invoices, dict):
                for inv_id, inv in all_invoices.items():
                    if not isinstance(inv, dict):
                        continue
                    inv_meta = inv.get("meta", {}) or {}

                    # Check if this invoice is linked to this project and stage
                    linked_projs = inv_meta.get("linked_projects", [])
                    is_for_this_stage = False

                    if isinstance(linked_projs, list):
                        for lp in linked_projs:
                            if isinstance(lp, dict):
                                if lp.get("project_number") == project_number and lp.get("payment_stage_index") == stage_index:
                                    is_for_this_stage = True
                                    break
                            elif isinstance(lp, str) and lp == project_number:
                                if inv_meta.get("payment_stage_index") == stage_index:
                                    is_for_this_stage = True
                                    break

                    if is_for_this_stage:
                        if not linked_invoice_number:
                            linked_invoice_id = inv_id
                            linked_invoice_number = inv_meta.get("invoice_number", "")
                            print(f"[MULTI_INV] Found invoice {linked_invoice_number} for project {project_number} stage {stage_index}", flush=True)
                            # For current invoice, prioritize it
                            if inv_id == invoice_id:
                                break
        else:
            # Sum payments from ALL invoices linked to this stage for this project
            if isinstance(all_invoices, dict):
                for inv_id, inv in all_invoices.items():
                    if not isinstance(inv, dict):
                        continue
                    inv_meta = inv.get("meta", {}) or {}

                    # Check if invoice covers this project and stage
                    # Works for both single-project (project_number) and multi-project (linked_projects)
                    is_for_this_project = False

                    # First, check if this is the current invoice being updated (most direct match)
                    if inv_id == invoice_id:
                        # For the current invoice, if it's linked to this project, include it
                        if inv_meta.get("project_number") == project_number:
                            # Always preserve the current invoice's number, even if no payments found
                            if not linked_invoice_number:
                                linked_invoice_id = inv_id
                                # Invoice number might be in meta or at top level
                                linked_invoice_number = inv_meta.get("invoice_number") or inv.get("invoice_number", "")
                                print(f"[CURRENT_INV] Captured invoice_number for {inv_id}: {linked_invoice_number} (from meta: {inv_meta.get('invoice_number')}, from top: {inv.get('invoice_number')})", flush=True)
                            is_for_this_project = True
                    else:
                        # For other invoices, use the standard matching logic
                        # Single-project invoices use project_number
                        if (inv_meta.get("project_number") == project_number and
                            inv_meta.get("payment_stage_index") == stage_index):
                            is_for_this_project = True
                        else:
                            # Multi-project invoices use linked_projects array
                            # Handles both dict format [{"project_number": "...", "payment_stage_index": ...}]
                            # and legacy string format ['MABS-202606-003']
                            linked_projs = inv_meta.get("linked_projects", [])
                            if isinstance(linked_projs, list):
                                for lp in linked_projs:
                                    if isinstance(lp, dict):
                                        # Dict format: full metadata
                                        if lp.get("project_number") == project_number and lp.get("payment_stage_index") == stage_index:
                                            is_for_this_project = True
                                            break
                                    elif isinstance(lp, str) and lp == project_number:
                                        # Legacy string format: only has project number
                                        # For this case, use the invoice's payment_stage_index
                                        if inv_meta.get("payment_stage_index") == stage_index:
                                            is_for_this_project = True
                                            break

                    if is_for_this_project:
                        # Sum payments for this invoice (filter by project_number if available)
                        inv_payment_log = inv.get("payment_log", [])
                        if isinstance(inv_payment_log, list):
                            # Sum only payments for this project
                            inv_payments = sum(
                                _safe_float(p.get("amount", 0))
                                for p in inv_payment_log
                                if p.get("project_number") == project_number or not p.get("project_number")
                            )
                            project_paid += inv_payments
                            # Track invoice_id and invoice_number if this invoice has actual payments
                            # Prioritize the current invoice being updated (inv_id == invoice_id)
                            if inv_payments > 0:
                                if inv_id == invoice_id or not linked_invoice_id:
                                    linked_invoice_id = inv_id
                                    linked_invoice_number = inv_meta.get("invoice_number", "")
                            print(f"[PAYMENT_CALC] Invoice {inv_id}: project={project_number}, payments={inv_payments}, total_paid={project_paid}, current={inv_id == invoice_id}", flush=True)

        # Determine stage status based on actual payments for this project
        if project_paid >= (stage_amount - 0.01):
            new_status = "Paid"
        elif project_paid > 0:
            new_status = "Partially Paid"
        else:
            new_status = "Invoiced"

        log.info(f"[STATUS] Project {project_number} stage {stage_index}: amount={stage_amount}, paid={project_paid}, threshold={stage_amount - 0.01}, status={new_status}")

        # Update stage status with actual paid amount for this project, and track invoice_id and invoice_number
        stage["status"] = new_status
        stage["amount_paid"] = str(project_paid) if project_paid > 0 else "0"

        # If we have an invoice_id but amount_paid is still 0, try to use the invoice's recorded amount_paid
        # BUT NOT FOR MULTI-PROJECT INVOICES (sequential allocation may have correctly set this to 0)
        if project_paid == 0 and linked_invoice_id and not is_multi_project:
            current_invoice = fb_get(f"/invoices/{linked_invoice_id}") or {}
            inv_amount_paid = _safe_float(current_invoice.get("meta", {}).get("amount_paid", 0))
            if inv_amount_paid > 0:
                # Use the invoice's amount_paid if we calculated 0
                stage["amount_paid"] = str(inv_amount_paid)
                project_paid = inv_amount_paid
                log.info(f"[FALLBACK] Using invoice amount_paid: {inv_amount_paid}")

        if linked_invoice_id:
            stage["invoice_id"] = linked_invoice_id
            print(f"[SETTING_ID] Set invoice_id to {linked_invoice_id}", flush=True)
        if linked_invoice_number:
            stage["invoice_number"] = linked_invoice_number
            print(f"[SETTING_NUM] Set invoice_number to {linked_invoice_number}", flush=True)
        else:
            print("[SKIPPING_NUM] Skipped setting invoice_number (linked_invoice_number is empty/None)", flush=True)

        log.info(f"[SAVE_STATUS] Saving stage {stage_index} status={new_status} to project {pid}")
        log.info(f"[SAVE_STAGE] Full stage data: {stage}, amount_paid={stage.get('amount_paid')}, invoice_number={stage.get('invoice_number')}")

        fb_update(f"/projects/{pid}", {
            "payment_stages": stages,
            "updated_at": datetime.now(timezone.utc).isoformat()
        })

def _allocate_invoice_payment_sequential(invoice_id: str) -> None:
    """Allocate invoice payment SEQUENTIALLY across linked projects (sorted by project number).

    When invoice with multiple projects receives payment, allocate in strict order:
    - Sort projects by project_number alphabetically
    - Project 005 gets filled first up to its stage amount
    - Then Project 006 gets remaining amount
    - Continue until all projects allocated or payment exhausted

    Each project updated with ONLY its allocated amount, not invoice total.
    """
    invoice = fb_get(f"/invoices/{invoice_id}") or {}
    if not isinstance(invoice, dict):
        return

    meta = invoice.get("meta", {}) or {}
    total_paid = _safe_float(meta.get("amount_paid", 0))

    log.info(f"[SEQ_ALLOC] Starting sequential allocation for invoice {invoice_id}, total_paid={total_paid}")

    # If total_paid is $0, we need to CLEAR all project amounts, not skip
    # This handles the case when all payments are deleted
    if total_paid <= 0.01 and total_paid > 0:
        # Very small payment (rounding), skip
        return

    # Get linked projects with their stage indices
    linked_projects_meta = meta.get("linked_projects", [])
    if not isinstance(linked_projects_meta, list):
        linked_projects_meta = []

    log.info(f"[SEQ_ALLOC] linked_projects_meta from metadata: {linked_projects_meta}")

    # If no linked_projects metadata, try to build from line_items
    # (Works even when payment_log is empty after deletion)
    if len(linked_projects_meta) < 2:
        line_items = invoice.get("line_items", [])
        if isinstance(line_items, list) and len(line_items) > 0:
            # Extract unique project numbers from line_items
            projects_in_items = set()
            for item in line_items:
                if isinstance(item, dict):
                    proj_num = item.get("project_number", "")
                    if proj_num:
                        projects_in_items.add(proj_num)

            # If line_items has 2+ projects, build linked_projects from them
            if len(projects_in_items) >= 2:
                log.info(f"[SEQ_ALLOC] Found {len(projects_in_items)} projects in line_items, building linked_projects")
                # For each project found, use the main invoice's stage index
                main_stage_idx = meta.get("payment_stage_index", 0)
                linked_projects_meta = [
                    {"project_number": proj_num, "payment_stage_index": main_stage_idx}
                    for proj_num in sorted(projects_in_items)
                ]

        if len(linked_projects_meta) < 1:
            log.info(f"[SEQ_ALLOC] NO projects found, skipping sequential allocation")
            return

    log.info(f"[SEQ_ALLOC] Found {len(linked_projects_meta)} linked projects, starting allocation")

    # Load all projects once
    all_projects = fb_get("/projects") or {}

    # Track which projects we'll update (to reset and recalculate)
    projects_to_update = {}  # pid -> (pdata, stage_idx)

    # Build list AND reset stages in-memory (don't wait for Firebase)
    projects_data = []
    for proj_info in linked_projects_meta:
        if not isinstance(proj_info, dict):
            continue
        proj_num = proj_info.get("project_number", "")
        stage_idx = proj_info.get("payment_stage_index", -1)

        if not proj_num or stage_idx < 0:
            log.warning(f"[SEQ_ALLOC] Skipping invalid proj_info: {proj_info}")
            continue

        # Find project in all_projects
        proj_id = None
        proj_data = None
        if isinstance(all_projects, dict):
            for pid, pdata in all_projects.items():
                if isinstance(pdata, dict) and pdata.get("project_number") == proj_num:
                    proj_id = pid
                    proj_data = pdata
                    break

        if not proj_id or not proj_data:
            log.warning(f"[SEQ_ALLOC] Project {proj_num} not found in database")
            continue

        projects_data.append((proj_num, stage_idx, proj_id, proj_data))

    # SORT BY PROJECT NUMBER (extract last digits and sort numerically)
    # e.g., "MABS-202606-005" -> extract "005" and sort as 5, "MABS-202606-006" -> 6
    projects_data.sort(key=lambda x: int(x[0][-3:]) if x[0][-3:].isdigit() else x[0])
    log.info(f"[SEQ_ALLOC] Sorted projects: {[p[0] for p in projects_data]}")

    # Determine allocation strategy: use per-project payment_log sums when entries
    # are tagged with project_number (modern data); fall back to sequential distribution
    # for legacy invoices whose payment_log entries have no project_number.
    payment_log_entries = invoice.get("payment_log", []) or []
    has_tagged_payments = any(
        isinstance(p, dict) and p.get("project_number")
        for p in payment_log_entries
    )

    allocations = {}  # proj_num -> amount

    if has_tagged_payments:
        # Modern path: each payment_log entry knows which project it belongs to
        log.info("[SEQ_ALLOC] Using per-project payment_log sums (tagged payments)")
        for proj_num, stage_idx, proj_id, proj_data in projects_data:
            proj_paid = sum(
                _safe_float(p.get("amount", 0))
                for p in payment_log_entries
                if isinstance(p, dict) and p.get("project_number") == proj_num
            )
            allocations[proj_num] = proj_paid
            log.info(f"[SEQ_ALLOC] {proj_num}: payment_log sum=${proj_paid}")
    else:
        # Legacy path: no project_number tags — distribute total sequentially
        log.info("[SEQ_ALLOC] No tagged payments — using sequential allocation")
        remaining = total_paid
        for proj_num, stage_idx, proj_id, proj_data in projects_data:
            stages = proj_data.get("payment_stages") or []
            if not (0 <= stage_idx < len(stages)):
                log.warning(f"[SEQ_ALLOC] Invalid stage_idx {stage_idx} for {proj_num}")
                continue

            stage = stages[stage_idx]
            stage_amount = _safe_float(stage.get("amount", 0))

            if remaining <= 0.01:
                allocations[proj_num] = 0
                log.info(f"[SEQ_ALLOC] No remaining amount for {proj_num}")
                continue

            if stage_amount <= 0:
                log.info(f"[SEQ_ALLOC] Skipping {proj_num} - stage amount is 0")
                continue

            allocated = min(remaining, stage_amount)
            allocations[proj_num] = allocated
            remaining -= allocated
            log.info(f"[SEQ_ALLOC] {proj_num}: stage_amount=${stage_amount}, allocated=${allocated}, remaining=${remaining}")

    # Update each project with its allocated amount
    for proj_num, stage_idx, proj_id, proj_data in projects_data:
        allocated = allocations.get(proj_num, 0)

        stages = proj_data.get("payment_stages") or []

        # Update the stage if valid
        if 0 <= stage_idx < len(stages):
            stage = stages[stage_idx]
            stage_amount = _safe_float(stage.get("amount", 0))
            stage["amount_paid"] = str(allocated)

            # Set stage status based on allocated amount
            if allocated >= (stage_amount - 0.01):
                stage["status"] = "Paid"
            elif allocated > 0:
                stage["status"] = "Partially Paid"
            else:
                stage["status"] = "Pending Invoice"

            log.info(f"[SEQ_ALLOC] {proj_num}: allocated=${allocated}, status={stage['status']}")
        else:
            log.warning(f"[SEQ_ALLOC] Invalid stage_idx {stage_idx} for {proj_num}, still updating amount_paid")

        # Update project amount_paid — but never zero-out a Completed project
        contract_val = _safe_float(proj_data.get("contract_value", 0))
        current_status = proj_data.get("status", "Not Started")
        existing_paid = _safe_float(proj_data.get("amount_paid", 0))

        # Skip writing amount_paid=0 for a done project — it would corrupt the record
        if current_status in ("Completed", "invoiced_Fully paid") and allocated <= 0.01 and existing_paid > 0:
            updates = {
                "payment_stages": stages,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        else:
            updates = {
                "amount_paid": allocated,
                "payment_stages": stages,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }

        # Update project status if needed (never downgrade invoiced_Fully paid/Cancelled)
        if current_status not in ("invoiced_Fully paid", "Cancelled"):
            if contract_val > 0 and allocated >= contract_val - 0.01:
                updates["status"] = "invoiced_Fully paid"
            elif contract_val > 0 and 0 < allocated < contract_val - 0.01 and \
                    current_status not in ("On Hold", "invoiced_Fully paid", "Cancelled",
                                          "Ready to Sent", "Sent out_Invoiced", "Sent out_Not Invoiced"):
                updates["status"] = "invoiced_Partially paid"
            elif allocated > 0 and current_status == "Not Started":
                updates["status"] = "In Progress"

        fb_update(f"/projects/{proj_id}", updates)
        log.info(f"[SEQ_ALLOC] Updated project {proj_num}: amount_paid={allocated}")

def _sync_project_payment(project_number: str) -> None:
    """Recalculate project.amount_paid from invoice payment_logs (ground truth)
    and set project status based on that total.

    Using invoice payment_logs avoids corruption from stage.amount_paid resets
    that can occur during sequential allocation.
    """
    if not project_number:
        return

    raw_proj = fb_get("/projects") or {}
    pid = None
    pdata = None
    if isinstance(raw_proj, dict):
        for k, v in raw_proj.items():
            if isinstance(v, dict) and v.get("project_number", "") == project_number:
                pid, pdata = k, v
                break
    if not pid:
        return

    # Ground-truth: sum all invoice payments that belong to this project
    all_invoices = fb_get("/invoices") or {}
    total_paid = 0.0
    if isinstance(all_invoices, dict):
        for iid, inv in all_invoices.items():
            if not isinstance(inv, dict):
                continue
            meta = inv.get("meta", {}) or {}
            # Check if this invoice covers this project (single or multi-project)
            covers = (meta.get("project_number") == project_number)
            if not covers:
                linked = meta.get("linked_projects", [])
                if isinstance(linked, list):
                    for lp in linked:
                        if isinstance(lp, dict) and lp.get("project_number") == project_number:
                            covers = True
                            break
                        elif isinstance(lp, str) and lp == project_number:
                            covers = True
                            break
            if not covers:
                continue
            # Sum invoice line payments for this project
            payment_log = inv.get("payment_log", [])
            if isinstance(payment_log, list):
                for p in payment_log:
                    if isinstance(p, dict):
                        if p.get("project_number", project_number) == project_number:
                            total_paid += _safe_float(p.get("amount", 0))
                        elif not p.get("project_number"):
                            total_paid += _safe_float(p.get("amount", 0))

    updates = {
        "amount_paid": total_paid,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    contract_val   = _safe_float(pdata.get("contract_value", 0))
    current_status = pdata.get("status", "Not Started")

    # Never downgrade a cancelled project; always correct everything else.
    if current_status != "Cancelled":
        if contract_val > 0 and total_paid >= contract_val - 0.01:
            updates["status"] = "invoiced_Fully paid"
        elif contract_val > 0 and 0 < total_paid < contract_val - 0.01:
            if current_status not in ("On Hold", "invoiced_Fully paid", "Cancelled",
                                      "Ready to Sent", "Sent out_Invoiced", "Sent out_Not Invoiced"):
                updates["status"] = "invoiced_Partially paid"
        # Do NOT downgrade invoiced_Fully paid/invoiced_Not paid yet when total_paid == 0 — payment_log may be incomplete

    fb_update(f"/projects/{pid}", updates)

def _auto_flag_overdue() -> int:
    """Flip any Sent/Viewed/Partial invoice whose due_date < today to Overdue.
    Returns the number of invoices updated.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    raw = fb_get("/invoices") or {}
    count = 0
    if not isinstance(raw, dict):
        return 0
    for iid, inv in raw.items():
        if not isinstance(inv, dict):
            continue
        m = inv.get("meta", {})
        status   = m.get("status", "")
        due_date = m.get("due_date", "")
        if status in ("Sent", "Viewed", "Partial") and due_date and due_date < today:
            fb_update(f"/invoices/{iid}", {
                "meta/status":     "Overdue",
                "meta/updated_at": datetime.now(timezone.utc).isoformat(),
            })
            count += 1
            # Auto-send overdue reminder — only once per invoice
            if not m.get("reminder_sent_at"):
                ok, _ = _send_overdue_reminder_email(iid, inv)
                if ok:
                    fb_update(f"/invoices/{iid}", {
                        "meta/reminder_sent_at": datetime.now(timezone.utc).isoformat()
                    })
    return count

def _upsert_revenue_entry(invoice_id: str, inv_meta: dict) -> None:
    """Create or update a balance-sheet revenue entry for a paid/partial invoice."""
    paid_val = _safe_float(inv_meta.get("amount_paid", 0))
    if paid_val <= 0:
        return
    proj_num = inv_meta.get("project_number", "")
    entry = {
        "description":    f"Invoice {inv_meta.get('invoice_number','')}",
        "invoice_number": inv_meta.get("invoice_number", ""),
        "client_name":    inv_meta.get("client_name", ""),
        "status":         inv_meta.get("status", "Paid"),
        "total":          _safe_float(inv_meta.get("total", 0)),
        "amount_paid":    paid_val,
        "amount":         paid_val,
        "client":         inv_meta.get("client_name", ""),
        "invoice_id":     invoice_id,
        "project_number": proj_num,
        "date":           inv_meta.get("invoice_date", datetime.now().strftime("%Y-%m-%d")),
        "created_at":     datetime.now(timezone.utc).isoformat(),
    }
    raw_rev = fb_get("/balance_sheet_revenue") or {}
    existing_key = None
    if isinstance(raw_rev, dict):
        for rk, rv in raw_rev.items():
            if isinstance(rv, dict) and rv.get("invoice_id") == invoice_id:
                existing_key = rk
                break
    if existing_key:
        fb_update(f"/balance_sheet_revenue/{existing_key}", entry)
    else:
        fb_push("/balance_sheet_revenue", entry)

def _advance_project_to_in_progress(project_number: str) -> None:
    """Promote a project from Not Started to In Progress when an invoice is sent."""
    if not project_number:
        return
    raw_proj = fb_get("/projects") or {}
    for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
        if isinstance(pdata, dict) and pdata.get("project_number", "") == project_number:
            if pdata.get("status", "Not Started") == "Not Started":
                fb_update(f"/projects/{pid}", {
                    "status": "In Progress",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
            break

def _auto_complete_project_if_paid(project_number: str) -> None:
    """Mark project Completed once its amount paid covers the full contract value.

    Paying off a single installment (out of several) must not flip the whole
    project to Completed - only do so when total paid >= contract value.
    Call _sync_project_payment(project_number) first so amount_paid is current.
    """
    if not project_number:
        return
    raw_proj = fb_get("/projects") or {}
    for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
        if isinstance(pdata, dict) and pdata.get("project_number", "") == project_number:
            if pdata.get("status", "") not in ("invoiced_Fully paid", "Completed", "Cancelled"):
                contract_val = _safe_float(pdata.get("contract_value", 0))
                total_paid   = _safe_float(pdata.get("amount_paid", 0))
                if contract_val > 0 and total_paid >= contract_val - 0.01:
                    fb_update(f"/projects/{pid}", {
                        "status": "invoiced_Fully paid",
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    })
                    if not pdata.get("completion_email_sent"):
                        _send_project_completion_email(project_number, pdata)
                        fb_update(f"/projects/{pid}", {"completion_email_sent": True})
            break

def _send_project_completion_email(project_number: str, project_data: dict) -> None:
    """Email client when project is marked Completed / fully paid."""
    try:
        settings = load_settings()
        em = settings.get("email", {})
        if not em.get("enabled"):
            return
        co          = settings.get("company", {})
        client_name = (project_data.get("client_name") or project_data.get("client") or "").strip()
        if not client_name:
            return
        raw_clients  = fb_get("/clients") or {}
        client_email = ""
        if isinstance(raw_clients, dict):
            for ckey, cd in raw_clients.items():
                if isinstance(cd, dict) and (ckey.strip() == client_name or cd.get("company", "").strip() == client_name):
                    client_email = cd.get("email", "")
                    break
        if not client_email:
            return
        proj_name    = (project_data.get("project_name") or project_data.get("name") or project_number)
        contract_val = _safe_float(project_data.get("contract_value", 0))
        html_body = f"""<html><body style="font-family:Arial,sans-serif;color:#1a1a1a;">
<div style="max-width:600px;margin:0 auto;padding:24px;">
  <h2 style="color:#0D9488;">Project Complete — {proj_name}</h2>
  <p>Dear {client_name},</p>
  <p>We are pleased to confirm that project <strong>{proj_name}</strong>
     (#{project_number}) has been marked
     <strong style="color:#10B981;">Completed</strong>.
     All payments have been received in full — thank you!</p>
  <table style="width:100%;border-collapse:collapse;margin:16px 0;">
    <tr><td style="padding:8px;border-bottom:1px solid #eee;">Project</td>
        <td style="padding:8px;border-bottom:1px solid #eee;font-weight:bold;">{proj_name}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #eee;">Project #</td>
        <td style="padding:8px;border-bottom:1px solid #eee;">{project_number}</td></tr>
    <tr><td style="padding:8px;">Contract Value</td>
        <td style="padding:8px;font-weight:bold;">${contract_val:,.2f}</td></tr>
  </table>
  <p>We appreciate your business and look forward to working with you again.</p>
  <p style="margin-top:24px;">Best regards,<br><strong>{co.get('name','')}</strong></p>
</div></body></html>"""
        msg            = MIMEMultipart("alternative")
        msg["Subject"] = f"Project Completed — {proj_name} (#{project_number})"
        msg["From"]    = f"{em.get('from_name', co.get('name',''))} <{em.get('smtp_user','')}>"
        msg["To"]      = client_email
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP(em.get("smtp_host", "smtp.gmail.com"), int(em.get("smtp_port", 587)), timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(em.get("smtp_user", ""), em.get("smtp_password", ""))
            srv.sendmail(em.get("smtp_user", ""), [client_email], msg.as_string())
    except Exception as exc:
        log.error("Completion email error (project %s): %s", project_number, exc)

def _project_has_overdue_stage(payment_stages, raw_inv: dict) -> bool:
    """True if any payment stage is Invoiced/Partially Paid and its linked
    invoice's due date has passed without being fully paid."""
    if not isinstance(payment_stages, list):
        return False
    today = datetime.now().strftime("%Y-%m-%d")
    for stage in payment_stages:
        if not isinstance(stage, dict):
            continue
        if stage.get("status") not in ("Invoiced", "Partially Paid"):
            continue
        inv = raw_inv.get(stage.get("invoice_id", ""))
        if not isinstance(inv, dict):
            continue
        due_date = inv.get("meta", {}).get("due_date", "")
        if due_date and due_date < today:
            return True
    return False

def _load_clients() -> List[str]:
    raw = fb_get("/clients") or {}
    if isinstance(raw, dict):
        return sorted(raw.keys())
    return []

def _load_sales_people() -> List[dict]:
    """Return salesperson list sourced primarily from Settings /users (sales role only).
    Legacy /sales_persons entries are merged in as fallback so existing quote data
    that references old names is not orphaned.
    """
    seen_names: set = set()
    people: List[dict] = []

    # Primary source: Settings users with sales role only
    for u in _load_all_users():
        role = normalize_role(u.get("role", ""))
        if role != "sales":
            continue
        name = (u.get("username") or "").strip()
        if not name:
            continue
        people.append({
            "name":        name,
            "email":       u.get("email", ""),
            "phone":       u.get("phone", ""),
            "title":       u.get("title", ""),
            "firebase_id": u.get("firebase_uid", ""),
            "from_users":  True,
        })
        seen_names.add(name.lower())

    # Legacy fallback: /sales_persons (keeps old quote references valid)
    raw = fb_get("/sales_persons") or {}
    if isinstance(raw, dict):
        for pid, pdata in raw.items():
            if pdata and isinstance(pdata, dict):
                name = str(pdata.get("name", "")).strip()
                if name and name.lower() not in seen_names:
                    pdata["firebase_id"] = pid
                    people.append(pdata)
                    seen_names.add(name.lower())

    return sorted(people, key=lambda x: str(x.get("name", "")).lower())

def _load_projects_list() -> List[dict]:
    raw = fb_get("/projects") or {}
    if isinstance(raw, dict):
        items = []
        for pid, pdata in raw.items():
            if pdata and isinstance(pdata, dict):
                pdata["firebase_id"] = pid
                items.append(pdata)
        return sorted(items, key=lambda x: x.get("project_number", ""), reverse=True)
    return []

def _load_all_users() -> List[dict]:
    raw = fb_get("/users") or {}
    if isinstance(raw, dict):
        users = []
        for uid, udata in raw.items():
            if udata and isinstance(udata, dict):
                udata["firebase_uid"] = uid
                cp = udata.get("custom_pages")
                if isinstance(cp, dict):
                    udata["custom_pages"] = [cp[k] for k in sorted(cp.keys(), key=lambda x: int(x) if str(x).isdigit() else 0)]
                elif not isinstance(cp, list):
                    udata["custom_pages"] = None
                users.append(udata)
        return sorted(users, key=lambda x: x.get("username", "").lower())
    return []

def _load_time_entries() -> List[dict]:
    raw = fb_get("/time_entries") or {}
    if isinstance(raw, dict):
        items = []
        for tid, tdata in raw.items():
            if tdata and isinstance(tdata, dict):
                tdata["firebase_id"] = tid
                items.append(tdata)
        return sorted(items, key=lambda x: x.get("clock_in", ""), reverse=True)
    return []

def _load_time_off_requests() -> List[dict]:
    raw = fb_get("/time_off_requests") or {}
    if isinstance(raw, dict):
        items = []
        for rid, rdata in raw.items():
            if rdata and isinstance(rdata, dict):
                rdata["firebase_id"] = rid
                items.append(rdata)
        return sorted(items, key=lambda x: x.get("requested_at", ""), reverse=True)
    return []

DEFAULT_TIME_OFF_DAYS = 15
HOURS_PER_DAY = 8

def _count_working_days(start_str: str, end_str: str) -> int:
    """Count Mon–Fri working days between start and end dates (inclusive)."""
    from datetime import date as _date
    try:
        s = _date.fromisoformat(start_str[:10])
        e = _date.fromisoformat(end_str[:10])
        if e < s:
            return 0
        count = 0
        cur = s
        while cur <= e:
            if cur.weekday() < 5:
                count += 1
            cur += timedelta(days=1)
        return count
    except Exception:
        return 0

def _time_off_balance(all_requests: list, uid: str, year: int) -> dict:
    """Return {allotment_hours, used_hours, remaining_hours, allotment, used, remaining} for a user in a given year."""
    year_str = str(year)
    used_hours = 0.0
    for r in all_requests:
        if r.get("employee_uid") != uid:
            continue
        if r.get("status") != "Approved":
            continue
        if r.get("type") == "Unpaid":
            continue
        start = r.get("start_date", "")
        if not start or not start.startswith(year_str):
            continue
        # Use stored total_hours if available, otherwise fall back to working_days × 8
        if r.get("total_hours") is not None:
            used_hours += _safe_float(r.get("total_hours", 0))
        elif r.get("type") == "Half Day":
            used_hours += 4.0
        else:
            used_hours += _count_working_days(start, r.get("end_date", start)) * HOURS_PER_DAY
    allotment_hours = DEFAULT_TIME_OFF_DAYS * HOURS_PER_DAY
    remaining_hours = max(0.0, allotment_hours - used_hours)
    # Also keep day-based values for backwards compat
    used_days = used_hours / HOURS_PER_DAY
    remaining_days = remaining_hours / HOURS_PER_DAY
    return {
        "allotment_hours": allotment_hours,
        "used_hours": used_hours,
        "remaining_hours": remaining_hours,
        "allotment": DEFAULT_TIME_OFF_DAYS,
        "used": used_days,
        "remaining": remaining_days,
    }

def _sum_minutes_by_project(entries: List[dict]) -> dict:
    """Sum closed time-entry minutes grouped by project label."""
    totals: dict = {}
    for e in entries:
        if e.get("status") != "closed":
            continue
        minutes = _safe_float(e.get("duration_minutes", 0))
        proj = e.get("project_name") or e.get("project_number") or "General / Admin"
        totals[proj] = totals.get(proj, 0.0) + minutes
    return totals

def _sum_minutes(entries: List[dict]) -> float:
    """Total closed time-entry minutes."""
    return sum(_safe_float(e.get("duration_minutes", 0)) for e in entries if e.get("status") == "closed")

def _period_range(period: str, custom_start: str = "", custom_end: str = "") -> tuple:
    """Return (start_date, end_date, label) as YYYY-MM-DD strings for a named reporting period."""
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    if period == "month":
        return now.strftime("%Y-%m-01"), today_str, f"Month of {now.strftime('%B %Y')}"
    if period == "year":
        return now.strftime("%Y-01-01"), today_str, f"Year {now.year}"
    if period == "custom" and custom_start and custom_end:
        return custom_start, custom_end, f"{custom_start} → {custom_end}"
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    return week_start, today_str, f"Week of {week_start}"

def _aggregate_hours_by_project(entries: List[dict]) -> dict:
    """Group closed time-entry minutes into {project_label: {"_total": minutes, employee_name: minutes}}."""
    agg: dict = {}
    for e in entries:
        if e.get("status") != "closed":
            continue
        minutes = _safe_float(e.get("duration_minutes", 0))
        if minutes <= 0:
            continue
        proj = e.get("project_name") or e.get("project_number") or "General / Admin"
        emp = e.get("employee_name", "Unknown")
        bucket = agg.setdefault(proj, {"_total": 0.0})
        bucket["_total"] += minutes
        bucket[emp] = bucket.get(emp, 0.0) + minutes
    return agg

def _next_quote_number() -> str:
    now = datetime.now()
    prefix = f"QT-{now.strftime('%Y%m')}-"
    raw = fb_get("/job_forms") or {}
    nums = []
    for q in (raw.values() if isinstance(raw, dict) else []):
        if isinstance(q, dict):
            num = q.get("job_number", "") or ""
            if num.startswith(prefix):
                try:
                    nums.append(int(num[len(prefix):]))
                except ValueError:
                    pass
    next_n = (max(nums) + 1) if nums else 1
    return f"{prefix}{next_n:03d}"

def _next_invoice_number() -> str:
    now = datetime.now()
    prefix = f"INV-{now.strftime('%Y%m')}-"
    raw = fb_get("/invoices") or {}
    nums = []
    for inv in (raw.values() if isinstance(raw, dict) else []):
        if isinstance(inv, dict):
            num = inv.get("meta", {}).get("invoice_number", "") or ""
            if num.startswith(prefix):
                try:
                    nums.append(int(num[len(prefix):]))
                except ValueError:
                    pass
    next_n = (max(nums) + 1) if nums else 1
    return f"{prefix}{next_n:03d}"

def _next_project_number() -> str:
    now = datetime.now()
    prefix = f"MABS-{now.strftime('%Y%m')}-"
    raw = fb_get("/projects") or {}
    nums = []
    for p in (raw.values() if isinstance(raw, dict) else []):
        if isinstance(p, dict):
            num = p.get("project_number", "") or ""
            if num.startswith(prefix):
                try:
                    nums.append(int(num[len(prefix):]))
                except ValueError:
                    pass
    next_n = (max(nums) + 1) if nums else 1
    return f"{prefix}{next_n:03d}"

def _parse_service_types(form) -> list:
    """Return service_types list, substituting 'Other' with 'Other: {specify}' when filled."""
    types = form.getlist("service_types[]")
    if not types:
        return []  # Allow projects without services (return empty list, not None)
    specify = form.get("other_specify", "").strip()
    if "Other" in types and specify:
        idx = types.index("Other")
        types[idx] = f"Other: {specify}"
    return types

def _parse_quote_form(form) -> dict:
    line_items = []
    descriptions = form.getlist("item_description[]")
    quantities   = form.getlist("item_quantity[]")
    unit_prices  = form.getlist("item_unit_price[]")
    for desc, qty, price in zip(descriptions, quantities, unit_prices):
        if desc.strip():
            line_items.append({
                "description": desc,
                "quantity":    qty,
                "unit_price":  price,
                "total":       str(_safe_float(qty) * _safe_float(price)),
            })

    # Get client_id from selected company_name
    company_name = form.get("client_name", "")  # Form passes company_name as client_name value
    client_id = ""
    client_name_from_db = ""
    if company_name:
        client_data = fb_get(f"/clients/{company_name}") or {}
        client_id = client_data.get("client_id", "")
        client_name_from_db = client_data.get("client_name", "")
        print(f"[QUOTE_FORM] Quote created for client '{company_name}': client_id='{client_id}'", flush=True)

    return {
        "job_number":           form.get("job_number", ""),
        "client_id":            client_id,
        "company_name":         company_name,
        "client_name":          client_name_from_db,
        "project_name":         form.get("project_name", ""),
        "description":          form.get("description", ""),
        "status":               form.get("status", "Not Started"),
        "salesperson":          form.get("salesperson", ""),
        "date":                 form.get("date", datetime.now().strftime("%Y-%m-%d")),
        "valid_until":          form.get("valid_until", ""),
        "expected_completion":  form.get("expected_completion", ""),
        "service_types":        _parse_service_types(form),
        "priority":             form.get("priority", "Normal"),
        "is_expedited":         form.get("is_expedited") == "on",
        "rush_rate":            form.get("rush_rate", "0"),
        "rush_fee":             form.get("rush_fee", "0"),
        "line_items":           line_items,
        "subtotal":             form.get("subtotal", "0"),
        "tax_rate":             form.get("tax_rate", "0"),
        "tax_amount":           form.get("tax_amount", "0"),
        "total":                form.get("total", "0"),
        "notes":                form.get("notes", ""),
        "terms":                form.get("terms", ""),
        "follow_up_date":       form.get("follow_up_date", ""),
    }

def _resolve_installment_plan(data: dict) -> tuple:
    """Interpret the submitted installment selection into (mode, count, custom_amounts).

    Returns ('custom', n, [amounts]) when the user typed in their own irregular
    installment amounts (for clients who pay negotiated/random amounts rather than
    even splits), otherwise ('equal', n, None) for the standard equal-split plan.
    """
    raw = str(data.get("installment_count", "1")).strip().lower()
    if raw == "custom":
        amounts = [a for a in (_safe_float(x) for x in data.get("custom_installment_amounts", [])) if a > 0]
        if amounts:
            return "custom", len(amounts), amounts
        return "equal", 1, None
    return "equal", max(1, min(int(_safe_float(raw or 1)), 6)), None

def _compute_payment_stages(contract_value: float, down_pct: float, installments: int,
                            custom_amounts: list = None) -> list:
    """Build an ordered payment-stage plan from a contract value, down-payment %, and installment count.

    Mirrors the desktop app's down-payment + installment model: an optional down-payment
    stage up front, then the remaining balance either as one final payment, split evenly
    across 2-6 installments (the last absorbs any rounding remainder), or — when
    `custom_amounts` is given — billed out exactly as the user typed those amounts
    (for clients who pay irregular, negotiated amounts rather than equal splits).
    """
    contract_value = max(0.0, contract_value)
    down_pct = max(0.0, min(100.0, down_pct))

    stages = []
    remaining = contract_value
    if down_pct > 0:
        down_amt = round(contract_value * down_pct / 100.0, 2)
        stages.append({"name": f"Down Payment ({down_pct:.0f}%)", "amount": down_amt,
                       "status": "Pending Invoice", "invoice_id": ""})
        remaining = round(contract_value - down_amt, 2)

    custom_amounts = [round(max(0.0, a), 2) for a in (custom_amounts or []) if a > 0]
    if custom_amounts:
        for i, amt in enumerate(custom_amounts):
            stages.append({"name": f"Installment {i+1} of {len(custom_amounts)}", "amount": amt,
                           "status": "Pending Invoice", "invoice_id": ""})
        return stages

    installments = max(1, min(int(installments or 1), 6))
    if installments <= 1:
        if down_pct > 0:
            # Has down payment + 1 final payment
            label = "Final Payment"
        else:
            # Only 1 payment, no down payment - display as "Full Payment"
            label = "Full Payment"
        stages.append({"name": label, "amount": remaining, "status": "Pending Invoice", "invoice_id": ""})
    else:
        per_installment = round(remaining / installments, 2)
        running = 0.0
        for i in range(installments):
            amt = per_installment if i < installments - 1 else round(remaining - running, 2)
            running += amt
            stages.append({"name": f"Installment {i+1} of {installments}", "amount": amt,
                           "status": "Pending Invoice", "invoice_id": ""})
    return stages

def _parse_project_form(form) -> dict:
    # Extract service costs from form (service_cost_<service_name>)
    service_costs = {}
    service_types = _parse_service_types(form)
    for svc in service_types:
        # Remove "Other: " prefix if present to get the key
        svc_key = svc.replace("Other: ", "Other") if svc.startswith("Other:") else svc
        cost_val = form.get(f"service_cost_{svc_key}", "0")
        try:
            service_costs[svc] = float(cost_val) if cost_val else 0.0
        except (ValueError, TypeError):
            service_costs[svc] = 0.0

    # Get client_id from selected company_name
    company_name = form.get("client_name", "")  # Form passes company_name as client_name value
    client_id = ""
    client_data = {}
    if company_name:
        client_data = fb_get(f"/clients/{company_name}") or {}
        client_id = client_data.get("client_id", "")

    return {
        # ── identifiers (match desktop field names exactly) ──────────────────
        "project_number":  form.get("project_number", ""),
        "quote_number":    form.get("quote_number", ""),
        "po_wo_number":    form.get("po_wo_number", ""),
        # ── project info ─────────────────────────────────────────────────────
        "project_name":    form.get("project_name", ""),
        "client_id":       client_id,
        "company_name":    company_name,
        "company":         company_name,   # desktop key = company
        "client_name":     client_data.get("client_name", ""),   # keep for web queries
        "site_address":    form.get("site_address", ""),
        "mail_address":    form.get("mail_address", ""),
        "date_received":   form.get("date_received", ""),
        "plant":           form.get("plant", ""),          # 2-letter state code
        "sales":           form.get("sales", ""),
        "service_types":   service_types,
        "service_costs":   service_costs,
        "scope_of_work":   form.get("scope_of_work", ""),
        "expedite":        form.get("expedite", "No"),
        "rush_rate":       form.get("rush_rate", "0"),
        "rush_fee":        form.get("rush_fee", "0"),
        "description":     form.get("description", ""),
        "notes":           form.get("notes", ""),
        # ── dates ────────────────────────────────────────────────────────────
        "status":          form.get("status", "Not Started"),
        "start_date":      form.get("start_date", ""),
        "end_date":        form.get("end_date", ""),
        # ── financials ───────────────────────────────────────────────────────
        "contract_value":       form.get("contract_value", "0"),
        "project_amount":       form.get("contract_value", "0"),  # desktop key
        "payment_category":     form.get("payment_category", "Down Payment"),
        "amount_paid":          form.get("amount_paid", "0"),
        "down_payment_percent": form.get("down_payment_percent", "0"),
        "installment_count":    form.get("installment_count", "1"),
        "custom_installment_amounts": [a for a in form.getlist("custom_installment_amount[]") if str(a).strip()],
        # ── budgets ──────────────────────────────────────────────────────────
        "budget_labor":         _safe_float(form.get("budget_labor", 0)),
        "budget_expenses":      _safe_float(form.get("budget_expenses", 0)),
        "budget_subcontractor": _safe_float(form.get("budget_subcontractor", 0)),
        "actual_labor_cost":    _safe_float(form.get("actual_labor_cost", 0)),
    }

def _parse_invoice_form(form) -> dict:
    line_items = []
    descriptions    = form.getlist("item_description[]")
    quantities      = form.getlist("item_quantity[]")
    unit_prices     = form.getlist("item_unit_price[]")
    item_projects   = form.getlist("item_project[]")  # Form field name is "item_project[]" not "item_project_number[]"
    main_project    = form.get("project_number", "").strip()
    for i, (desc, qty, price) in enumerate(zip(descriptions, quantities, unit_prices)):
        if desc.strip():
            item_proj = (item_projects[i].strip() if i < len(item_projects) else "")
            line_items.append({
                "description":    desc,
                "quantity":       qty,
                "unit_price":     price,
                "amount":         str(_safe_float(qty) * _safe_float(price)),
                # Empty = "bill under the invoice's main project" (the common case).
                # Set explicitly when an item belongs to a *different* project than
                # the one selected above, so one invoice can span multiple projects.
                "project_number": item_proj,
            })

    # Every distinct project referenced anywhere on this invoice (main selection +
    # any per-item overrides) — used to link this invoice on each project's detail
    # page and to prorate payments across projects when the invoice spans more than one.
    linked_projects = sorted({p for p in
                              ([main_project] + [li["project_number"] for li in line_items])
                              if p})

    # Get client_id from selected company_name
    company_name = form.get("company_name", "")  # This comes from the form's company_name field
    client_id = ""
    client_data = {}
    if company_name:
        client_data = fb_get(f"/clients/{company_name}") or {}
        client_id = client_data.get("client_id", "")

    return {
        "meta": {
            "invoice_number": form.get("invoice_number", ""),
            "invoice_date":   form.get("invoice_date", datetime.now().strftime("%Y-%m-%d")),
            "due_date":       form.get("due_date", ""),
            "net_terms":      form.get("net_terms", ""),
            "client_id":      client_id,
            "company_name":   company_name,
            "client_name":    client_data.get("client_name", ""),
            "project_number": main_project,
            "linked_projects": linked_projects,
            "status":         form.get("status", "Draft"),
            "subtotal":       form.get("subtotal", "0"),
            "tax_rate":       form.get("tax_rate", "0"),
            "tax_amount":     form.get("tax_amount", "0"),
            "total":          form.get("total", "0"),
            "notes":          form.get("notes", ""),
            "terms":          form.get("terms", ""),
            "payment_method": form.get("payment_method", ""),
        },
        "line_items": line_items,
        "_payment_amount": form.get("payment_amount", ""),
        "_payment_date": form.get("payment_date", ""),
        "_payment_method": form.get("payment_method", ""),
        "_payment_reference": form.get("payment_reference", ""),
    }

def _invoice_project_share(invoice_data: dict, project_number: str) -> float:
    """Fraction (0-1) of an invoice's billed total attributable to one project.

    Single-project invoices (the common case) return 1.0. Multi-project invoices
    prorate by each project's share of the line-item subtotal, so payments and
    P&L roll up fairly across every project an invoice spans.
    """
    items = invoice_data.get("line_items", []) or []
    meta = invoice_data.get("meta", {}) or {}
    main_project = meta.get("project_number", "")

    item_amounts = [(str(it.get("project_number", "")).strip() or main_project,
                     _safe_float(it.get("amount", 0))) for it in items]
    total = sum(a for _, a in item_amounts)

    # Debug: print what we're calculating
    inv_num = meta.get("invoice_number", "?")
    print(f"[SHARE] Invoice {inv_num}, project={project_number}: item_amounts={item_amounts}, total={total}", flush=True)

    if total <= 0:
        # No usable line-item amounts — check if this project is in linked_projects
        # (for multi-project invoices where project_number isn't set in line items)
        linked_projects = meta.get("linked_projects", [])
        print(f"[SHARE] Total<=0, checking linked_projects={linked_projects}", flush=True)
        if isinstance(linked_projects, list):
            if any(isinstance(lp, dict) and lp.get("project_number") == project_number for lp in linked_projects):
                # This project is part of the multi-project invoice - equal share
                share = 1.0 / len(linked_projects) if linked_projects else 1.0
                print(f"[SHARE] Found in linked_projects, share={share}", flush=True)
                return share
        # Fall back to whole-invoice attribution for the (single) project this invoice names
        fallback = 1.0 if main_project == project_number else 0.0
        print(f"[SHARE] Fallback share={fallback}", flush=True)
        return fallback

    project_amount = sum(a for pn, a in item_amounts if pn == project_number)
    share = project_amount / total
    print(f"[SHARE] Calculated share={share} (project_amount={project_amount})", flush=True)
    return share

def _invoice_linked_projects(invoice_data: dict) -> set:
    """All project numbers an invoice is linked to (main selection + per-item overrides)."""
    meta = invoice_data.get("meta", {}) or {}
    linked = set()

    # Extract project numbers from linked_projects (list of dicts with {project_number, payment_stage_index})
    linked_projects = meta.get("linked_projects") or []
    if isinstance(linked_projects, list):
        for item in linked_projects:
            if isinstance(item, dict):
                pn = item.get("project_number", "")
                if pn:
                    linked.add(pn)

    if not linked:
        main_project = meta.get("project_number", "")
        if main_project:
            linked.add(main_project)
        for it in (invoice_data.get("line_items", []) or []):
            pn = str(it.get("project_number", "")).strip()
            if pn:
                linked.add(pn)
    return linked

# ── Routes: Workflow conversions ─────────────────────────────────────────────
@app.route("/quotes/<quote_id>/to-invoice", methods=["POST"])
@role_required("invoicing")
def quote_to_invoice(quote_id):
    quote = fb_get(f"/job_forms/{quote_id}")
    if not quote:
        abort(404)
    inv_num = _next_invoice_number()
    # Find the project linked to this quote (if any)
    linked_proj_num = ""
    raw_proj = fb_get("/projects") or {}
    if isinstance(raw_proj, dict):
        for p in raw_proj.values():
            if isinstance(p, dict) and p.get("source_quote") == quote_id:
                linked_proj_num = p.get("project_number", "")
                break
    # Map quote line items → invoice line items
    inv_items = []
    for item in quote.get("line_items", []):
        inv_items.append({
            "description": item.get("description", ""),
            "quantity":    item.get("quantity", "1"),
            "unit_price":  item.get("unit_price", "0"),
            "amount":      item.get("total", item.get("amount", "0")),
        })
    invoice_data = {
        "meta": {
            "invoice_number": inv_num,
            "invoice_date":   datetime.now().strftime("%Y-%m-%d"),
            "due_date":       (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
            "client_name":    quote.get("client_name", ""),
            "project_number": linked_proj_num,
            "status":         "Draft",
            "subtotal":       str(quote.get("subtotal", "0")),
            "tax_rate":       str(quote.get("tax_rate", "0")),
            "tax_amount":     str(quote.get("tax_amount", "0")),
            "total":          str(quote.get("total", "0")),
            "amount_paid":    "0",
            "notes":          quote.get("notes", ""),
            "terms":          quote.get("terms", ""),
            "payment_method": "",
            "source_quote":   quote_id,
            "created_at":     datetime.now(timezone.utc).isoformat(),
            "updated_at":     datetime.now(timezone.utc).isoformat(),
            "created_by":     session.get("user_email", ""),
        },
        "line_items": inv_items,
    }
    iid = fb_push("/invoices", invoice_data)
    fb_update(f"/job_forms/{quote_id}", {
        "status":     "Invoiced",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    flash(f"Invoice {inv_num} created from quote.", "success")
    return redirect(url_for("invoice_detail", invoice_id=iid))

@app.route("/quotes/<quote_id>/pdf")
@role_required("quotes")
def quote_pdf(quote_id):
    try:
        from reportlab.lib.pagesizes import letter
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("quote_detail", quote_id=quote_id))

    pdf_bytes = _generate_quote_pdf_bytes(quote_id)
    if not pdf_bytes:
        abort(404)

    quote = fb_get(f"/job_forms/{quote_id}")
    if not quote:
        abort(404)

    from flask import Response
    fname = f"Quote_{quote.get('job_number','')}.pdf"
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={"Content-Disposition": f"inline;filename={fname}"})

@app.route("/projects/<project_id>/pdf")
@role_required("projects")
def project_pdf(project_id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
    except ImportError:
        flash("reportlab not installed.", "danger")
        return redirect(url_for("project_detail", project_id=project_id))

    import io as _io
    project = fb_get(f"/projects/{project_id}")
    if not project:
        abort(404)

    co = company_info()
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    elems = []

    teal   = colors.HexColor("#0F766E")
    dark   = colors.HexColor("#0F172A")
    muted  = colors.HexColor("#64748B")
    light  = colors.HexColor("#F8FAFC")
    border = colors.HexColor("#E2E8F0")

    lbl = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=8,  fontName="Helvetica-Bold", textColor=muted, spaceAfter=1)
    val = ParagraphStyle("val", parent=styles["Normal"], fontSize=10, fontName="Helvetica",       textColor=dark, spaceAfter=8)
    sm  = ParagraphStyle("sm",  parent=styles["Normal"], fontSize=9,  fontName="Helvetica",       textColor=muted)
    h2  = ParagraphStyle("h2",  parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold",  textColor=dark, spaceBefore=12, spaceAfter=6)

    # ── Header ──
    addr = (co.get("address","") or "").replace("\n", " | ")
    hdr_data = [[
        Paragraph(f"<b>{co.get('name','')}</b>", ParagraphStyle("cn", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold", textColor=dark)),
        Paragraph("PROJECT DOCUMENT", ParagraphStyle("pd", parent=styles["Normal"], fontSize=18, fontName="Helvetica-Bold", textColor=teal, alignment=2)),
    ],[
        Paragraph(f"{addr}<br/>{co.get('phone','')}  |  {co.get('email','')}", sm),
        Paragraph(f"<b>{project.get('project_number','')}</b>", ParagraphStyle("pn", parent=styles["Normal"], fontSize=12, fontName="Helvetica-Bold", textColor=dark, alignment=2)),
    ]]
    hdr = Table(hdr_data, colWidths=[3.5*inch, 3.5*inch])
    hdr.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    elems.append(hdr)
    elems.append(HRFlowable(width="100%", thickness=2, color=teal, spaceAfter=12))

    # ── Project meta ──
    meta = [
        [Paragraph("CLIENT", lbl),       Paragraph("STATUS", lbl),          Paragraph("START DATE", lbl),       Paragraph("END DATE", lbl)],
        [Paragraph(project.get("client_name","—"), val), Paragraph(project.get("status","—"), val), Paragraph(project.get("start_date","—"), val), Paragraph(project.get("end_date","—") or "—", val)],
        [Paragraph("PROJECT NAME", lbl), Paragraph("", lbl), Paragraph("ASSIGNED TO", lbl), Paragraph("PAYMENT STAGE", lbl)],
        [Paragraph(project.get("project_name","—"), val), Paragraph("", val), Paragraph(project.get("assigned_to","—") or "—", val), Paragraph(project.get("payment_category","—") or "—", val)],
    ]
    mt = Table(meta, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    mt.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("LEFTPADDING",(0,0),(-1,-1),0), ("SPAN",(0,2),(1,2)), ("SPAN",(0,3),(1,3))]))
    elems.append(mt)

    # ── Scope ──
    if project.get("description"):
        elems.append(Paragraph("SCOPE OF WORK", h2))
        elems.append(HRFlowable(width="100%", thickness=0.5, color=border, spaceAfter=6))
        elems.append(Paragraph(project.get("description",""), sm))

    # ── Financial summary ──
    cv   = _safe_float(project.get("contract_value", 0))
    paid = _safe_float(project.get("amount_paid",    0))
    outstanding = cv - paid
    pct  = int(paid / cv * 100) if cv > 0 else 0

    elems.append(Paragraph("FINANCIAL SUMMARY", h2))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=border, spaceAfter=6))
    fin_data = [
        ["Contract Value",  f"${cv:,.2f}"],
        ["Amount Paid",     f"${paid:,.2f}"],
        ["Outstanding",     f"${outstanding:,.2f}"],
        ["Collection Rate", f"{pct}%"],
        ["Payment Stage",   project.get("payment_category","—") or "—"],
    ]
    fin_tbl = Table(fin_data, colWidths=[2.5*inch, 2*inch])
    fin_tbl.setStyle(TableStyle([
        ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,0), (-1,-1), 10),
        ("TEXTCOLOR",     (0,0), (0,-1),  muted),
        ("FONTNAME",      (1,0), (1,-1),  "Helvetica-Bold"),
        ("TEXTCOLOR",     (1,2), (1,2),   colors.HexColor("#DC2626") if outstanding > 0 else teal),
        ("TEXTCOLOR",     (1,0), (1,0),   teal),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [colors.white, light]),
        ("GRID",          (0,0), (-1,-1), 0.4, border),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
    ]))
    elems.append(fin_tbl)

    # ── Notes ──
    if project.get("notes"):
        elems.append(Paragraph("NOTES", h2))
        elems.append(HRFlowable(width="100%", thickness=0.5, color=border, spaceAfter=6))
        elems.append(Paragraph(project.get("notes",""), sm))

    # ── Signature block ──
    elems.append(Spacer(1, 30))
    sig_data = [[
        Paragraph("_" * 35, ParagraphStyle("sig", parent=styles["Normal"], fontSize=10)),
        Paragraph("_" * 35, ParagraphStyle("sig2", parent=styles["Normal"], fontSize=10)),
    ],[
        Paragraph(f"Client Signature — {project.get('client_name','')}", sm),
        Paragraph(f"Authorized — {co.get('name','')}", sm),
    ],[
        Paragraph("Date: _______________", sm),
        Paragraph("Date: _______________", sm),
    ]]
    sig = Table(sig_data, colWidths=[3.5*inch, 3.5*inch])
    sig.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("TOPPADDING",(0,0),(-1,-1),4)]))
    elems.append(sig)

    # ── Footer ──
    elems.append(Spacer(1, 20))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=border, spaceAfter=4))
    gen_date = datetime.now().strftime("%B %d, %Y")
    elems.append(Paragraph(f"Document generated {gen_date}  ·  {co.get('name','')}  ·  {co.get('phone','')}  ·  {co.get('email','')}", sm))

    doc.build(elems)
    buf.seek(0)
    from flask import Response
    fname = f"Project_{project.get('project_number','doc')}.pdf"
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"inline;filename={fname}"})

# ── Email helper ─────────────────────────────────────────────────────────────
def _generate_invoice_pdf_bytes(invoice_id: str):
    """Generate invoice PDF and return as bytes. Returns None on error."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import mm, inch
    except ImportError:
        return None

    import io as _io
    from pathlib import Path
    invoice = fb_get(f"/invoices/{invoice_id}")
    if not invoice:
        return None

    meta = invoice.get("meta", {})
    co = company_info()
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=5*mm, bottomMargin=5*mm)
    styles = getSampleStyleSheet()
    story = []

    styles.add(ParagraphStyle(name='CenteredBold20', alignment=1, fontName='Helvetica-Bold', fontSize=20, leading=24))
    styles.add(ParagraphStyle(name='Centered10', alignment=1, fontName='Helvetica', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='LeftBold16', alignment=0, fontName='Helvetica-Bold', fontSize=16, leading=18))
    styles.add(ParagraphStyle(name='LeftBold12', alignment=0, fontName='Helvetica-Bold', fontSize=12, leading=14))
    styles.add(ParagraphStyle(name='Left10', alignment=0, fontName='Helvetica', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='LeftBold10', alignment=0, fontName='Helvetica-Bold', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='Right10', alignment=2, fontName='Helvetica', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='RightBold10', alignment=2, fontName='Helvetica-Bold', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='Left9', alignment=0, fontName='Helvetica', fontSize=9, leading=11))
    styles.add(ParagraphStyle(name='center8', alignment=1, fontName='Helvetica', fontSize=8, leading=10))

    logo_path = _get_company_logo_path()
    logo_img = None
    if logo_path:
        try:
            logo_file = Path(logo_path)
            if logo_file.exists():
                logo_img = Image(str(logo_file.resolve()), width=0.95*inch, height=0.95*inch)
        except Exception:
            pass

    company_name = co.get('name', 'MABS Engineering LLC')
    address_text = ""
    for line in co.get('address', '').split('\n'):
        if line.strip():
            address_text += f"{line.strip()}<br/>"
    contact_text = f"Phone: {co.get('phone','')} • Email: {co.get('email','')} • {co.get('website','')}"
    header_html = f"<b><font size=16>{company_name}</font></b><br/><font size=9>{address_text}{contact_text}</font>"

    if logo_img:
        hdr_table_data = [[logo_img, Paragraph(header_html, ParagraphStyle("cn", parent=styles["Normal"], fontName="Helvetica", textColor=colors.black, alignment=1))]]
        hdr_table = Table(hdr_table_data, colWidths=[0.95*inch, doc.width - 0.95*inch])
        hdr_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (1,0), (1,0), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 0), ("TOPPADDING", (0,0), (-1,-1), 0), ("LINEBELOW", (0,0), (-1,-1), 1, colors.black)]))
        story.append(hdr_table)
    else:
        story.append(Paragraph(header_html, ParagraphStyle("cn", parent=styles["Normal"], fontName="Helvetica", textColor=colors.black, alignment=1)))
        story.append(Table([['']], colWidths=[doc.width], style=[('LINEBELOW', (0,0), (-1,-1), 1, colors.black)]))

    story.append(Spacer(1, 2*mm))

    # Get company name (primary identifier), fallback to client_name for backward compatibility
    company_identifier = meta.get('company_name', '') or meta.get('client_name', '')
    client_email = ""
    client_address = ""
    if company_identifier:
        try:
            client_data = fb_get(f"/clients/{company_identifier}") or {}
            client_email = client_data.get("email", "")
            client_address = client_data.get("address", "")
        except Exception:
            pass

    bill_to_lines = []
    if company_identifier:
        bill_to_lines.append(company_identifier)
    if client_email:
        bill_to_lines.append(client_email)
    if client_address:
        for line in client_address.split('\n'):
            if line.strip():
                bill_to_lines.append(line.strip())
    bill_to_text = "<br/>".join(bill_to_lines) if bill_to_lines else ""

    invoice_info = f"Invoice Number: {meta.get('invoice_number','')}<br/>Date: {meta.get('invoice_date','')}<br/>Due Date: {meta.get('due_date','')}"
    header_data = [
        [Paragraph("<b>Invoice:</b>", styles['Left10']), Paragraph("<b>Bill To:</b>", styles['Left10'])],
        [Paragraph(invoice_info, styles['Left9']), Paragraph(bill_to_text, styles['Left9']) if bill_to_text else Paragraph("", styles['Left9'])],
    ]
    header_table = Table(header_data, colWidths=[doc.width * 0.5, doc.width * 0.5])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0), ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2)]))
    story.append(header_table)
    story.append(Spacer(1, 8*mm))

    story.append(Paragraph("ITEMS", ParagraphStyle("items_title", parent=styles["Normal"], fontSize=12, fontName="Helvetica-Bold", textColor=colors.black, alignment=0)))
    story.append(Spacer(1, 3*mm))

    center_bold_style = ParagraphStyle("center_bold10", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", textColor=colors.black, alignment=1)
    center_style = ParagraphStyle("center10", parent=styles["Normal"], fontSize=10, fontName="Helvetica", textColor=colors.black, alignment=1)
    left_style = ParagraphStyle("left10", parent=styles["Normal"], fontSize=10, fontName="Helvetica", textColor=colors.black, alignment=0)
    item_data = [[Paragraph(h, center_bold_style) for h in ["Project Number", "Description (PO & Road Name)", "Plant", "Qty", "Unit Price", "Total Due"]]]

    linked_projects = meta.get("linked_projects", [])
    if isinstance(linked_projects, str):
        linked_projects = [{"project_number": linked_projects}]
    elif not isinstance(linked_projects, list):
        linked_projects = []

    for idx, item in enumerate(invoice.get("line_items", [])):
        qty_val = _safe_float(item.get("quantity", 1))
        qty = str(int(qty_val)) if qty_val == int(qty_val) else str(qty_val)
        unit_price_val = _safe_float(item.get('unit_price', 0))
        unit_price = f"${unit_price_val:,.2f}"
        project_number = ""
        project_name = ""
        plant = ""
        payment_stage = ""
        if idx < len(linked_projects) and isinstance(linked_projects[idx], dict):
            project_number = linked_projects[idx].get("project_number", "")
            project_name = linked_projects[idx].get("project_name", "")
        if not project_number:
            project_number = meta.get("project_number", "")

        po_wo = ""
        site_address = ""
        if project_number:
            try:
                raw_proj = fb_get("/projects") or {}
                for pid, pdata in (raw_proj.items() if isinstance(raw_proj, dict) else []):
                    if isinstance(pdata, dict) and pdata.get("project_number") == project_number:
                        if not project_name:
                            project_name = pdata.get("project_name", "")
                        plant = pdata.get("plant", "")
                        po_wo = pdata.get("po_wo_number", "")
                        site_address = pdata.get("site_address", "")
                        payment_stages = pdata.get("payment_stages", [])
                        payment_stage_index = None
                        if idx < len(linked_projects) and isinstance(linked_projects[idx], dict):
                            payment_stage_index = linked_projects[idx].get("payment_stage_index")
                        if payment_stage_index is None:
                            payment_stage_index = meta.get("payment_stage_index")
                        if payment_stage_index is not None and isinstance(payment_stages, list) and int(payment_stage_index) < len(payment_stages):
                            stage_data = payment_stages[int(payment_stage_index)]
                            if isinstance(stage_data, dict):
                                payment_stage = stage_data.get("name", "")
                        break
            except Exception:
                pass

        if not plant:
            plant = meta.get("plant", "")
        if not project_name:
            description = item.get("description", "")
            project_name = description.split("—")[0].strip() if "—" in description else description

        description = item.get("description", "")

        project_cell = Paragraph(project_number, center_style)
        if description and "co-" in description.lower():
            # Format: "Project Name — CO-002 – Title" or "CO-001 – Title"
            if "—" in description:
                # Split by em-dash first to get the CO part
                parts = description.split("—")
                co_section = parts[1].strip() if len(parts) > 1 else ""
                # Now extract CO number and title from co_section
                co_parts = co_section.split("–")
                co_part = co_parts[0].strip()
                title_part = co_parts[1].strip() if len(co_parts) > 1 else ""
            else:
                # Fallback: split by en-dash
                co_parts = description.split("–")
                co_part = co_parts[0].strip()
                title_part = co_parts[1].strip() if len(co_parts) > 1 else ""

            if co_part.upper().startswith("CO"):
                co_display = f"{co_part}_{title_part}" if title_part else co_part
                project_number_display = f"{project_number}<br/><font size=8>{co_display}</font>"
                project_cell = Paragraph(project_number_display, left_style)
        elif payment_stage and "change order" in payment_stage.lower():
            project_number_display = f"{project_number}<br/><font size=8>{payment_stage}</font>"
            project_cell = Paragraph(project_number_display, left_style)

        description_display = ""
        if po_wo or site_address:
            if site_address:
                import re
                us_states_abbr = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY"]

                state_idx = -1
                closest_state_idx = float('inf')
                site_upper = site_address.upper()

                for state in us_states_abbr:
                    idx = site_upper.find(f" {state} ")
                    if idx != -1:
                        next_char_idx = idx + len(state) + 2
                        if next_char_idx < len(site_upper):
                            next_char = site_upper[next_char_idx]
                            if next_char.isdigit() and idx < closest_state_idx:
                                state_idx = idx
                                closest_state_idx = idx

                    if state_idx == -1:
                        idx = site_upper.find(f" {state}")
                        if idx != -1:
                            next_char_idx = idx + len(state) + 1
                            if next_char_idx < len(site_upper):
                                next_char = site_upper[next_char_idx]
                                if next_char.isdigit() and idx < closest_state_idx:
                                    state_idx = idx
                                    closest_state_idx = idx

                if state_idx != -1:
                    site_address = site_address[:state_idx].strip()
                    if site_address.endswith(" -"):
                        site_address = site_address[:-2].strip()
                    if site_address.endswith("–"):
                        site_address = site_address[:-1].strip()
                elif plant and plant.strip():
                    plant_upper = plant.strip().upper()
                    site_upper = site_address.upper()
                    plant_idx = site_upper.find(f" {plant_upper} ")
                    if plant_idx == -1:
                        plant_idx = site_upper.find(f" {plant_upper}")
                    if plant_idx == -1:
                        plant_idx = site_upper.find(plant_upper)

                    if plant_idx != -1:
                        site_address = site_address[:plant_idx].strip()
                        if site_address.endswith(" -"):
                            site_address = site_address[:-2].strip()
                        if site_address.endswith("–"):
                            site_address = site_address[:-1].strip()

                if po_wo and site_address:
                    description_display = f"{po_wo} - {site_address}"
                elif site_address:
                    description_display = site_address
                else:
                    description_display = po_wo
            else:
                description_display = po_wo
        else:
            description_display = ""

        total_due_val = qty_val * unit_price_val
        total_due = f"${total_due_val:,.2f}"
        item_data.append([
            project_cell,
            Paragraph(description_display, center_style),
            Paragraph(plant or "", center_style),
            Paragraph(qty, center_style),
            Paragraph(unit_price, center_style),
            Paragraph(total_due, center_style)
        ])

    if len(item_data) == 1:
        item_data.append([Paragraph("", center_style) for _ in range(6)])

    item_table = Table(item_data, colWidths=[doc.width * 0.17, doc.width * 0.37, doc.width * 0.09, doc.width * 0.07, doc.width * 0.12, doc.width * 0.18])
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#CCCCCC")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(item_table)
    story.append(Spacer(1, 5*mm))

    subtotal = _safe_float(meta.get('subtotal', 0))
    tax_amount = _safe_float(meta.get('tax_amount', 0))
    tax_rate = _safe_float(meta.get('tax_rate', 0))
    total_amount = _safe_float(meta.get('total', 0))
    totals_data = [
        [Paragraph("Total", styles['Right10']), Paragraph(f"${subtotal:,.2f}", styles['Right10'])],
        [Paragraph(f"Tax ({tax_rate}%)", styles['Right10']), Paragraph(f"${tax_amount:,.2f}", styles['Right10'])],
        [Paragraph("TOTAL AMOUNT DUE:", styles['RightBold10']), Paragraph(f"${total_amount:,.2f}", styles['RightBold10'])],
    ]
    totals_table = Table(totals_data, colWidths=[doc.width * 0.5, doc.width * 0.5])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#CCCCCC")),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BACKGROUND', (0, len(totals_data)-1), (-1, len(totals_data)-1), colors.lightgrey),
        ('FONTNAME', (0, len(totals_data)-1), (-1, len(totals_data)-1), 'Helvetica-Bold'),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 5*mm))

    story.append(Paragraph("PAYMENT OPTIONS", styles['LeftBold12']))
    story.append(Spacer(1, 3*mm))

    InvLabel = ParagraphStyle("InvLabel", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold", textColor=colors.black, leading=12)
    InvValue = ParagraphStyle("InvValue", parent=styles["Normal"], fontSize=9, fontName="Helvetica", textColor=colors.black, leading=11, leftIndent=3, spaceAfter=3)
    TableCenter = ParagraphStyle("TableCenter", parent=styles["Normal"], fontSize=8, fontName="Helvetica", textColor=colors.black, alignment=1)

    qr_path = Path(__file__).parent / "static" / "venmo.png"
    qr_img = None
    if qr_path.exists():
        try:
            qr_img = Image(str(qr_path), width=1.0*inch, height=1.0*inch)
        except (IOError, OSError):
            qr_img = None

    right_section = [
        Table([[Paragraph("<b>Option 2: Zelle QR code</b>", InvLabel)]], colWidths=[doc.width * 0.40], rowHeights=[8*mm],
              style=TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#B6D7A8")), ("BOX", (0,0), (-1,-1), 0.8, colors.black), ("ALIGN", (0,0), (-1,-1), "LEFT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 2), ("BOTTOMPADDING", (0,0), (-1,-1), 2), ("LEFTPADDING", (0,0), (-1,-1), 3)])),
        Spacer(1, 1*mm),
    ]
    if qr_img:
        right_section.append(Table([[qr_img]], colWidths=[doc.width * 0.40],
                                   style=TableStyle([("ALIGN", (0,0), (-1,-1), "CENTER"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 1*mm), ("BOTTOMPADDING", (0,0), (-1,-1), 1*mm)])))
    right_section.append(Paragraph("Scan to pay with Zelle", TableCenter))

    left_section = [
        Table([[Paragraph("<b>Option 1: Check</b>", InvLabel)]], colWidths=[doc.width * 0.55], rowHeights=[7*mm],
              style=TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#B6DDE8")), ("BOX", (0,0), (-1,-1), 0.8, colors.black), ("ALIGN", (0,0), (-1,-1), "LEFT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 2), ("BOTTOMPADDING", (0,0), (-1,-1), 2), ("LEFTPADDING", (0,0), (-1,-1), 3)])),
        Spacer(1, 1*mm),
        Paragraph("<b>Payable to:</b> MABS Engineering LLC<br/><b>Mailing Address:</b> 15455 Manchester Rd, PO Box 1144 Manchester, MO 63011", InvValue),
        Spacer(1, 4*mm),
        Table([[Paragraph("<b>Option 3: Bank ACH Transfer</b>", InvLabel)]], colWidths=[doc.width * 0.55], rowHeights=[7*mm],
              style=TableStyle([("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#EA9999")), ("BOX", (0,0), (-1,-1), 0.8, colors.black), ("ALIGN", (0,0), (-1,-1), "LEFT"), ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 2), ("BOTTOMPADDING", (0,0), (-1,-1), 2), ("LEFTPADDING", (0,0), (-1,-1), 3)])),
        Spacer(1, 1*mm),
        Paragraph("Please contact MABS Admin to get our bank information for ACH transfers", InvValue),
    ]

    payment_table = Table([[left_section, right_section]], colWidths=[doc.width * 0.55, doc.width * 0.40])
    payment_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('LEFTPADDING', (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0), ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 0), ('BOX', (0,0), (-1,-1), 1, colors.black), ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black)]))
    story.append(payment_table)

    story.append(Spacer(1, 3*mm))
    default_terms = co.get('default_terms', 'Thank you for your business! Best regards, MABS Engineering LLC')
    notes_text = default_terms if default_terms else meta.get('notes', 'Thank you for your business!')
    story.append(Paragraph(f"<b>Note:</b> {notes_text}", styles['Left9']))

    calculated_status = _calculate_invoice_status(invoice)
    if (calculated_status or "").lower() == 'paid':
        def add_paid_watermark(canvas_obj, doc_obj):
            canvas_obj.saveState()
            center_x = A4[0] / 2
            center_y = A4[1] / 2
            canvas_obj.setFont("Helvetica-Bold", 130)
            canvas_obj.setFillColor(colors.HexColor("#00B050"))
            canvas_obj.setFillAlpha(0.25)
            canvas_obj.translate(center_x, center_y)
            canvas_obj.rotate(45)
            canvas_obj.drawCentredString(0, 0, "PAID")
            canvas_obj.restoreState()
        doc.build(story, onFirstPage=add_paid_watermark, onLaterPages=add_paid_watermark)
    else:
        doc.build(story)

    buf.seek(0)
    return buf.getvalue()

def _send_invoice_email(invoice_id: str):
    """Send invoice professional text email + PDF attachment to the client. Returns (ok: bool, message: str)."""
    settings = load_settings()
    em = settings.get("email", {})

    if not em.get("enabled"):
        return False, "Email sending is disabled. Enable it in Settings → Email/SMTP."

    invoice = fb_get(f"/invoices/{invoice_id}")
    if not invoice:
        return False, "Invoice not found."

    meta        = invoice.get("meta", {})
    client_name = meta.get("client_name", "")
    client_data = fb_get(f"/clients/{client_name}") or {}
    client_email = client_data.get("email", "")
    if not client_email:
        return False, f"No email on file for '{client_name}'. Add it in Clients."

    co = company_info()

    # Company info for signature
    company_name = co.get('name', 'MABS Engineering LLC')
    company_address = co.get('address', '15455 Manchester Rd, PO Box 1144, Manchester, MO 63011')
    company_email = co.get('email', 'info@mabs-engineering.com')
    company_phone = co.get('phone', '(314) 585-2003')

    signature = f"""{company_name}
{company_address}
{company_email}
{company_phone}"""

    # Professional text email
    invoice_number = meta.get('invoice_number', '')
    invoice_status = str(meta.get('status', 'Draft')).strip()  # Status values: Draft, Sent, Viewed, Paid, Partial, Overdue, Cancelled
    total_due = _safe_float(meta.get('total', 0))
    due_date = meta.get('due_date', '')
    paid_date = meta.get('paid_date', '')

    company_name = co.get('name', 'MABS Engineering LLC')
    company_address = co.get('address', '15455 Manchester Rd, PO Box 1144, Manchester, MO 63011')
    company_email = co.get('email', 'info@mabs-engineering.com')
    company_phone = co.get('phone', '(314) 585-2003')

    # Check if invoice is paid (handle capitalization and spacing)
    is_paid = invoice_status.lower() == 'paid'
    log.info(f"Invoice {invoice_number} status: '{invoice_status}' → is_paid: {is_paid}")

    if is_paid:
        # PAID INVOICE EMAIL
        text_body = f"""Hi {client_name},

Please find the attached paid invoice for your records.

Invoice Details:
Invoice Number: {invoice_number}
Amount Paid: ${total_due:,.2f}
Payment Date: {paid_date or 'Received'}

Thank you for your payment. If you require any additional information, please let us know.

Best regards,
{company_name}
{company_address}
{company_email}
{company_phone}
"""
        subject = f"{invoice_number} – Payment Received"
    else:
        # UNPAID INVOICE EMAIL
        text_body = f"""Hi {client_name},

Please find the attached invoice for your review.

Invoice Details:
Invoice Number: {invoice_number}
Amount Due: ${total_due:,.2f}
Due Date: {due_date}

Please review the attached invoice at your convenience. If you have any questions or require additional information, please feel free to contact us.

Thank you for your business. We appreciate your continued trust and support.

Best regards,
{company_name}
{company_address}
{company_email}
{company_phone}
"""
        subject = f"You have a new invoice from {company_name} ({invoice_number})"

    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"]    = f"{em.get('from_name', co.get('name',''))} <{em.get('smtp_user','')}>"
        msg["To"]      = client_email

        msg.attach(MIMEText(text_body, "plain"))

        try:
            pdf_bytes = _generate_invoice_pdf_bytes(invoice_id)
            if pdf_bytes:
                from email.mime.base import MIMEBase
                from email import encoders
                pdf_part = MIMEBase("application", "octet-stream")
                pdf_part.set_payload(pdf_bytes)
                encoders.encode_base64(pdf_part)
                pdf_part.add_header("Content-Disposition", "attachment",
                                   filename=f"Invoice_{meta.get('invoice_number','')}.pdf")
                msg.attach(pdf_part)
        except Exception as pdf_err:
            log.warning("Could not attach PDF to invoice email: %s", pdf_err)

        with smtplib.SMTP(em.get("smtp_host", "smtp.gmail.com"), int(em.get("smtp_port", 587)), timeout=15) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(em.get("smtp_user", ""), em.get("smtp_password", ""))
            srv.sendmail(em.get("smtp_user", ""), [client_email], msg.as_string())

        return True, f"Invoice emailed to {client_email}."
    except Exception as exc:
        log.error("Email send error: %s", exc)
        return False, f"Failed to send email: {exc}"

def _generate_quote_pdf_bytes(quote_id: str):
    """Generate quote PDF and return as bytes. Returns None on error."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch, mm
        from pathlib import Path
    except ImportError:
        return None

    import io as _io
    quote = fb_get(f"/job_forms/{quote_id}")
    if not quote:
        return None

    co = company_info()
    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=0.35*inch, rightMargin=0.35*inch,
                            topMargin=0.08*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elems = []

    dark_gray = colors.HexColor("#333333")
    light_blue = colors.HexColor("#B6DDE8")
    light_green = colors.HexColor("#B6D7A8")
    light_red = colors.HexColor("#EA9999")
    border_col = colors.HexColor("#000000")
    teal_line = colors.HexColor("#0D9488")

    form_label = ParagraphStyle("fl", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", textColor=dark_gray)
    form_value = ParagraphStyle("fv", parent=styles["Normal"], fontSize=9, fontName="Helvetica", textColor=dark_gray)
    section_title = ParagraphStyle("st", parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold", textColor=dark_gray)
    checkbox_style = ParagraphStyle("cs", parent=styles["Normal"], fontSize=8, fontName="Helvetica", textColor=dark_gray)

    logo_path = _get_company_logo_path()
    if logo_path:
        try:
            logo_file = Path(logo_path)
            if logo_file.exists():
                logo = Image(str(logo_file.resolve()), width=1.0*inch, height=0.85*inch)
                hdr_data = [[logo, Paragraph(f"<b>{co.get('name','MABS Engineering LLC')}</b>", ParagraphStyle("cn", parent=styles["Normal"], fontSize=22, fontName="Helvetica-Bold", textColor=dark_gray, alignment=1))]]
                hdr = Table(hdr_data, colWidths=[1.0*inch, doc.width - 1.0*inch])
                hdr.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("ALIGN", (1,0), (1,0), "CENTER"), ("LINEBELOW", (0,0), (-1,-1), 2, teal_line), ("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (1,0), (1,0), 0), ("BOTTOMPADDING", (0,0), (-1,-1), 0)]))
                elems.append(hdr)
            else:
                elems.append(Paragraph(co.get('name','MABS Engineering LLC'), ParagraphStyle("cn", parent=styles["Normal"], fontSize=16, fontName="Helvetica-Bold", textColor=dark_gray)))
        except Exception:
            elems.append(Paragraph(co.get('name','MABS Engineering LLC'), ParagraphStyle("cn", parent=styles["Normal"], fontSize=16, fontName="Helvetica-Bold", textColor=dark_gray)))
    else:
        elems.append(Paragraph(co.get('name','MABS Engineering LLC'), ParagraphStyle("cn", parent=styles["Normal"], fontSize=16, fontName="Helvetica-Bold", textColor=dark_gray)))

    elems.append(Spacer(1, 0.08*inch))

    sales_person = quote.get('salesperson', '')
    fixed_sales_width = 2.0*inch
    center_width = (doc.width - fixed_sales_width) / 2
    right_width = (doc.width - fixed_sales_width) / 2

    title_data = [
        [
            Paragraph(f"<b>Sales: {sales_person}</b>" if sales_person else "", ParagraphStyle("sp", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", textColor=dark_gray, alignment=1)),
            Paragraph("<u>New Job Request Form</u>", ParagraphStyle("title", parent=styles["Normal"], fontSize=12, fontName="Helvetica-Bold", textColor=dark_gray, alignment=1)),
            Paragraph("", ParagraphStyle("title", parent=styles["Normal"]))
        ]
    ]

    title_table = Table(title_data, colWidths=[fixed_sales_width, center_width, right_width])

    table_style = [
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]

    if sales_person:
        table_style.extend([
            ('BACKGROUND', (0,0), (0,0), colors.HexColor("#D9EEF7")),
            ('BORDER', (0,0), (0,0), 1, colors.HexColor("#5B9DBE")),
            ('ALIGN', (0,0), (0,0), 'CENTER'),
            ('LEFTPADDING', (0,0), (0,0), 3),
            ('RIGHTPADDING', (0,0), (0,0), 3),
            ('TOPPADDING', (0,0), (0,0), 2),
            ('BOTTOMPADDING', (0,0), (0,0), 2),
        ])

    title_table.setStyle(TableStyle(table_style))
    elems.append(title_table)
    elems.append(Spacer(1, 0.05*inch))
    elems.append(Spacer(1, 0.15*inch))

    def add_form_field(label, value):
        field_table = Table(
            [[Paragraph(f"<b>{label}</b>", form_label), Paragraph(":", form_label), Paragraph(value, form_value)]],
            colWidths=[55*mm, 5*mm, doc.width - 60*mm]
        )
        field_table.setStyle(TableStyle([
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 2),
        ]))
        elems.append(field_table)

    add_form_field("Quote Number", quote.get('job_number',''))
    add_form_field("Client / Company Name", quote.get('client_name',''))
    add_form_field("Project Name", quote.get('project_name',''))
    add_form_field("Scope of Work", quote.get('description',''))

    elems.append(Spacer(1, 1*mm))

    eng_rows = [
        [Paragraph("<b>Engineering Costs:</b>", section_title)]
    ]
    eng_cost_style = ParagraphStyle("ec", parent=styles["Normal"], fontSize=9, fontName="Helvetica", textColor=dark_gray, leftIndent=8*mm)

    agreed_cost = quote.get('subtotal', '')
    expedite = quote.get('is_expedited', False)
    expedite_amount = quote.get('rush_fee', '') or quote.get('rushFee', '')
    total_amount = quote.get('total', '')

    if agreed_cost:
        agreed_cost_val = str(agreed_cost).replace('$', '').strip()
        expedite_checkbox = "[✓]" if expedite else "[ ]"
        if expedite and expedite_amount and str(expedite_amount).strip():
            expedite_val = str(expedite_amount).replace('$', '').strip()
        else:
            expedite_val = "________"
        cost_line = f"Agreed Cost: {agreed_cost_val}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<b>Expedite? {expedite_checkbox} Yes, 50% Extra ( ) No:</b> {expedite_val}"
    else:
        cost_line = "Agreed Cost: ________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<b>Expedite? ( ) Yes, 50% Extra ( ) No:</b> ________"

    eng_rows.append([Paragraph(cost_line, eng_cost_style)])

    if total_amount:
        try:
            total_val = float(str(total_amount).replace('$', '').strip())
            total_line = f"<b>TOTAL: ${total_val:,.2f}</b>"
        except (ValueError, TypeError):
            total_line = "TOTAL: ________"
    else:
        total_line = "TOTAL: ________"

    eng_rows.append([Paragraph(total_line, eng_cost_style)])

    eng_table = Table(eng_rows, colWidths=[doc.width])
    eng_table.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("TOPPADDING", (0,0), (-1,-1), 1.5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    elems.append(eng_table)
    elems.append(Spacer(1, 1.5*mm))

    elems.append(Paragraph("<b>In the case of Court Appearance or Disposition:</b> N/A", form_value))
    elems.append(Paragraph("Rate: $250/hour (portal-to-portal)", form_value))
    elems.append(Spacer(1, 0.12*inch))

    def add_section_title(label):
        title_para = Paragraph(f"<b>{label}</b>", section_title)
        title_table = Table([[title_para]], colWidths=[doc.width])
        title_table.setStyle(TableStyle([
            ("TEXTCOLOR", (0,0), (-1,-1), dark_gray),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("TOPPADDING", (0,0), (-1,-1), 1),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ]))
        elems.append(title_table)
        elems.append(Spacer(1, 0.5*mm))

    add_section_title("Services Required :")

    all_services = ["Mechanical", "Electrical", "Civil", "Structural", "Plumbing", "HVAC", "Fire Protection", "Geotechnical", "Environmental", "Other"]

    selected_services = quote.get('service_types', [])

    custom_services = [svc for svc in selected_services if svc not in all_services]

    svc_items = []
    for svc in all_services:
        if svc == "Other" and custom_services:
            continue

        if svc in selected_services:
            svc_items.append(f"[✓] {svc}")
        else:
            svc_items.append(f"[ ] {svc}")

    for svc in custom_services:
        svc_items.append(f"[✓] {svc}")

    svc_grid = []
    for i in range(0, len(svc_items), 3):
        row = []
        for j in range(3):
            if i+j < len(svc_items):
                row.append(Paragraph(svc_items[i+j], checkbox_style))
            else:
                row.append(Paragraph("", checkbox_style))
        svc_grid.append(row)

    if not svc_grid:
        svc_grid = [[Paragraph("[ ] Mechanical", checkbox_style), Paragraph("[ ] Electrical", checkbox_style), Paragraph("[ ] Civil", checkbox_style)],
                    [Paragraph("[ ] Structural", checkbox_style), Paragraph("[ ] Plumbing", checkbox_style), Paragraph("[ ] HVAC", checkbox_style)],
                    [Paragraph("[ ] Fire Protection", checkbox_style), Paragraph("[ ] Geotechnical", checkbox_style), Paragraph("[ ] Environmental", checkbox_style)],
                    [Paragraph("[ ] Other", checkbox_style), Paragraph("", checkbox_style), Paragraph("", checkbox_style)]]

    svc_table = Table(svc_grid, colWidths=[doc.width/3, doc.width/3, doc.width/3])
    svc_table.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ("TOPPADDING", (0,0), (-1,-1), 1),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    elems.append(svc_table)
    elems.append(Spacer(1, 1.5*mm))

    elems.append(Spacer(1, 1.5*mm))
    add_section_title("Payment Information")

    pay_warning_table = Table([
        [Paragraph("<b>A 50% DOWN PAYMENT IS REQUIRED TO INITIATE</b>", ParagraphStyle("warning", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", textColor=colors.red, alignment=1))]
    ], colWidths=[doc.width])
    pay_warning_table.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 1, border_col),
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 1*mm),
        ("BOTTOMPADDING", (0,0), (-1,-1), 1*mm),
    ]))
    elems.append(pay_warning_table)

    qr_path = Path(__file__).parent / "static" / "venmo.png"
    qr_image = None
    if qr_path.exists():
        try:
            qr_image = Image(str(qr_path), width=35*mm, height=35*mm)
        except:
            pass

    available_width = doc.width

    left_section = [
        Table(
            [[Paragraph("<b>Option 1: Check</b>", form_label)]],
            colWidths=[available_width * 0.60],
            rowHeights=[7*mm],
            style=TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), light_blue),
                ("BOX", (0,0), (-1,-1), 0.7, colors.black),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ])
        ),
        Table(
            [[Paragraph(f"<b>Payable to:</b> {co.get('name','MABS Engineering LLC')}<br/><b>Mailing Address:</b> 15455 Manchester Rd, PO Box 1144 Manchester, MO 63011", form_value)]],
            colWidths=[available_width * 0.60],
            style=TableStyle([
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("TOPPADDING", (0,0), (-1,-1), 2*mm),
                ("BOTTOMPADDING", (0,0), (-1,-1), 2*mm),
                ("LEFTPADDING", (0,0), (-1,-1), 4*mm),
            ])
        ),
        Table(
            [[Paragraph("<b>Option 3: Bank ACH Transfer</b>", form_label)]],
            colWidths=[available_width * 0.60],
            rowHeights=[7*mm],
            style=TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), light_red),
                ("BOX", (0,0), (-1,-1), 0.7, colors.black),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ])
        ),
        Table(
            [[Paragraph("<b>Account Type:</b> Checking<br/><b>Bank Name:</b> BMO Harris Bank<br/><b>Routing Number:</b> 071025661<br/><b>Acct. Number:</b> 4834994317", form_value)]],
            colWidths=[available_width * 0.60],
            style=TableStyle([
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("TOPPADDING", (0,0), (-1,-1), 2*mm),
                ("BOTTOMPADDING", (0,0), (-1,-1), 2*mm),
                ("LEFTPADDING", (0,0), (-1,-1), 4*mm),
            ])
        ),
    ]

    right_section = [
        Table(
            [[Paragraph("<b>Option 2: Zelle QR code</b>", form_label)]],
            colWidths=[available_width * 0.40],
            rowHeights=[7*mm],
            style=TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), light_green),
                ("BOX", (0,0), (-1,-1), 0.8, colors.black),
                ("ALIGN", (0,0), (-1,-1), "LEFT"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ])
        ),
        Spacer(1, 1*mm),
    ]

    if qr_image:
        right_section.append(
            Table(
                [[qr_image]],
                style=TableStyle([
                    ("ALIGN", (0,0), (-1,-1), "CENTER"),
                    ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                    ("TOPPADDING", (0,0), (-1,-1), 1*mm),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 1*mm),
                ])
            )
        )

    right_section.append(
        Paragraph(
            "Scan to pay with Zelle",
            ParagraphStyle("qr_text", parent=styles["Normal"], fontSize=8, alignment=1)
        )
    )

    payment_data = [[left_section, right_section]]
    pay_table = Table(
        payment_data,
        colWidths=[available_width * 0.60, available_width * 0.40]
    )

    pay_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("BOX", (0,0), (-1,-1), 1, colors.black),
        ("INNERGRID", (0,0), (-1,-1), 0.5, colors.black),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))

    elems.append(pay_table)
    elems.append(Spacer(1, 1*mm))

    elems.append(Spacer(1, 1.5*mm))

    agreement_title = Paragraph("<b>Client Agreement :</b>", section_title)
    agreement_title_table = Table([[agreement_title]], colWidths=[doc.width])
    agreement_title_table.setStyle(TableStyle([
        ("TEXTCOLOR", (0,0), (-1,-1), dark_gray),
        ("BOTTOMPADDING", (0,0), (-1,-1), 1),
        ("TOPPADDING", (0,0), (-1,-1), 1),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
    ]))
    elems.append(agreement_title_table)
    elems.append(Spacer(1, 0.6*mm))

    agreement_style = ParagraphStyle(
        "agreement",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=dark_gray,
    )

    elems.append(Spacer(1, 0.5*mm))
    elems.append(Paragraph(
        "By signing below, the client agrees to provide necessary documents, respond to RFIs within 3 business days, "
        "and acknowledges that deliverables will be considered final if no response is received within 10 business days.",
        agreement_style
    ))

    elems.append(Spacer(1, 1.5*mm))

    sig_table = Table([
        [Paragraph("Client Signature :", form_value), Paragraph("Date :", form_value)]
    ], colWidths=[doc.width * 0.75, doc.width * 0.25])
    sig_table.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
    ]))
    elems.append(sig_table)
    elems.append(Spacer(1, -2*mm))

    def quote_footer(canvas, doc_obj):
        canvas.saveState()

        footer_blue = colors.HexColor("#003D82")

        canvas.setLineWidth(0.5)
        canvas.setStrokeColor(footer_blue)
        canvas.line(doc_obj.leftMargin, 0.72*inch, doc_obj.width + doc_obj.leftMargin, 0.72*inch)

        footer_style = ParagraphStyle(
            name="FooterStyle",
            alignment=1,
            fontName="Helvetica",
            fontSize=7,
            textColor=footer_blue,
            leading=9
        )

        footer_lines = [
            "Note: As the CEO of MABS Engineering LLC, Dr. Ashiq reserves the right to change or cancel this policy at any time, at his discretion.",
            f"Address: {co.get('address','15455 Manchester Rd, PO Box 1144, Ballwin, MO 63011')}",
            f"Telephone: {co.get('phone','(314) 585-2003')} • {co.get('email','info@mabs-engineering.com')}",
            co.get('website','www.mabs-engineering.com')
        ]

        y_position = 0.55*inch

        for line in footer_lines:
            p = Paragraph(line, footer_style)
            w, h = p.wrap(doc_obj.width - 1*inch, 12*mm)
            p.drawOn(canvas, doc_obj.leftMargin + 0.5*inch, y_position)
            y_position -= 3*mm

        canvas.restoreState()

    doc.build(elems, onFirstPage=quote_footer, onLaterPages=quote_footer)
    buf.seek(0)
    return buf.getvalue()

def _send_quote_email(quote_id: str):
    """Send quote professional text email + PDF attachment to the client. Returns (ok: bool, message: str)."""
    settings = load_settings()
    em = settings.get("email", {})

    if not em.get("enabled"):
        return False, "Email sending is disabled. Enable it in Settings → Email/SMTP."

    quote = fb_get(f"/job_forms/{quote_id}")
    if not quote:
        return False, "Quote not found."

    client_name = quote.get("client_name", "")
    client_data = fb_get(f"/clients/{client_name}") or {}
    client_email = client_data.get("email", "")
    if not client_email:
        return False, f"No email on file for '{client_name}'. Add it in Clients."

    co = company_info()

    # Get scope/description
    scope = quote.get('description', '').strip() or quote.get('scope', '').strip()
    project_name = quote.get('project_name', '').strip()

    # Company info for signature
    company_name = co.get('name', 'MABS Engineering LLC')
    company_address = co.get('address', '15455 Manchester Rd, PO Box 1144, Manchester, MO 63011')
    company_email = co.get('email', 'info@mabs-engineering.com')
    company_phone = co.get('phone', '(314) 585-2003')

    signature = f"""{company_name}
{company_address}
{company_email}
{company_phone}"""

    # Determine conditions
    has_project = bool(project_name)
    has_scope = bool(scope)

    # Build email based on conditions
    if has_project and has_scope:
        # Case 1: Project + Services
        text_body = f"""Hi {client_name},

Thank you for considering us for your {project_name} project. We've prepared a detailed quote for the {scope}.

The attached PDF shows the complete breakdown. Please let me know if you have any questions.

Best regards,

{signature}
"""
    elif has_project and not has_scope:
        # Case 2: Project + NO Services
        text_body = f"""Hi {client_name},

Thank you for considering us for your {project_name} project. We've prepared a detailed quote.

The attached PDF shows the complete breakdown. Please let me know if you have any questions.

Best regards,

{signature}
"""
    elif not has_project and has_scope:
        # Case 3: NO Project + Services
        text_body = f"""Hi {client_name},

Thank you for considering us for your project. We've prepared a detailed quote for the {scope}.

The attached PDF shows the complete breakdown. Please let me know if you have any questions.

Best regards,

{signature}
"""
    else:
        # Case 4: NO Project + NO Services
        text_body = f"""Hi {client_name},

Thank you for considering us for your project. We've prepared a detailed quote.

The attached PDF shows the complete breakdown. Please let me know if you have any questions.

Best regards,

{signature}
"""

    try:
        msg = MIMEMultipart("mixed")
        msg["Subject"] = f"Quote from {co.get('name','')} - {quote.get('job_number','')}"
        msg["From"]    = f"{em.get('from_name', co.get('name',''))} <{em.get('smtp_user','')}>"
        msg["To"]      = client_email

        msg.attach(MIMEText(text_body, "plain"))

        try:
            pdf_bytes = _generate_quote_pdf_bytes(quote_id)
            if pdf_bytes:
                from email.mime.base import MIMEBase
                from email import encoders
                pdf_part = MIMEBase("application", "octet-stream")
                pdf_part.set_payload(pdf_bytes)
                encoders.encode_base64(pdf_part)
                pdf_part.add_header("Content-Disposition", "attachment",
                                   filename=f"Quote_{quote.get('job_number','')}.pdf")
                msg.attach(pdf_part)
        except Exception as pdf_err:
            log.warning("Could not attach PDF to quote email: %s", pdf_err)

        with smtplib.SMTP(em.get("smtp_host", "smtp.gmail.com"), int(em.get("smtp_port", 587)), timeout=15) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(em.get("smtp_user", ""), em.get("smtp_password", ""))
            srv.sendmail(em.get("smtp_user", ""), [client_email], msg.as_string())

        return True, f"Quote emailed to {client_email}."
    except Exception as exc:
        log.error("Email send error: %s", exc)
        return False, f"Failed to send email: {exc}"

@app.route("/invoicing/<invoice_id>/send", methods=["POST"])
@role_required("invoicing")
def invoice_send(invoice_id):
    try:
        ok, msg = _send_invoice_email(invoice_id)
        if ok:
            inv_data = fb_get(f"/invoices/{invoice_id}") or {}
            current_status = (inv_data.get("meta", {}) or {}).get("status", "")
            # Only move to Sent if not already in a terminal state (Paid, Cancelled)
            if current_status not in ("Paid", "Cancelled"):
                fb_update(f"/invoices/{invoice_id}", {
                    "meta/status": "Sent",
                    "meta/updated_at": datetime.now(timezone.utc).isoformat(),
                })
            for proj_num in _invoice_linked_projects(inv_data):
                try:
                    _advance_project_to_in_progress(proj_num)
                except Exception as e:
                    log.warning("_advance_project_to_in_progress error for %s: %s", proj_num, e)
        flash(msg, "success" if ok else "danger")
    except Exception as exc:
        log.error("invoice_send error for %s: %s", invoice_id, exc)
        flash(f"Failed to send: {exc}", "danger")
    return redirect(url_for("invoice_detail", invoice_id=invoice_id))

@app.route("/quotes/<quote_id>/send", methods=["POST"])
@role_required("quotes")
def quote_send(quote_id):
    ok, msg = _send_quote_email(quote_id)
    flash(msg, "success" if ok else "danger")
    return redirect(url_for("quote_detail", quote_id=quote_id))

@app.route("/invoicing/<invoice_id>/payment/add", methods=["POST"])
@role_required("invoicing")
def payment_add(invoice_id):
    """Record a payment received against an invoice (for line items, not tax)."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    log = inv_data.get("payment_log", [])
    if not isinstance(log, list):
        log = []

    amount = _safe_float(request.form.get("amount", 0))
    if amount <= 0:
        return jsonify({"error": "Invalid payment amount"}), 400

    # For multi-project invoices, store which project this payment applies to
    # For single-project invoices, auto-fill from invoice metadata
    project_number = request.form.get("project_number", "") or inv_data.get("meta", {}).get("project_number", "")

    # Get invoice metadata for additional context
    inv_meta = inv_data.get("meta", {}) or {}

    payment_entry = {
        "amount":     str(amount),
        "date":       request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
        "method":     request.form.get("method", ""),
        "reference":  request.form.get("reference", ""),
        "notes":      request.form.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        # Metadata for tracking
        "invoice_number": inv_meta.get("invoice_number", ""),
        "stage_name": inv_meta.get("payment_stage", ""),
        "stage_index": inv_meta.get("payment_stage_index", ""),
    }
    if project_number:
        payment_entry["project_number"] = project_number

    log.append(payment_entry)

    amount_paid = sum(_safe_float(p.get("amount", 0)) for p in log)
    fresh_inv = dict(inv_data)
    fresh_inv["payment_log"] = log
    new_status = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "payment_log":      log,
        "meta/amount_paid": str(amount_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })

    # Use sequential allocation for ALL invoices - it handles both single and multi-project
    # This ensures stage.amount_paid is always updated the same way
    _allocate_invoice_payment_sequential(invoice_id)

    # Update project stage payment statuses based on payments (after allocation)
    # Must do this BEFORE _sync_project_payment since it updates payment_stages
    _update_project_stage_payment_status(invoice_id)

    # Sync project-level amount_paid (sum of all stages for each project)
    # This MUST run AFTER _update_project_stage_payment_status to override its data
    # Get fresh invoice data to ensure we have correct metadata after Firebase updates
    fresh_inv = fb_get(f"/invoices/{invoice_id}") or inv_data
    linked_projects = _invoice_linked_projects(fresh_inv)
    for proj_num in linked_projects:
        _sync_project_payment(proj_num)
        # Complete the project once its own amount_paid covers its contract value
        # (checked per-project, so this also fires for multi-project invoices)
        _auto_complete_project_if_paid(proj_num)

    fresh_meta = (fb_get(f"/invoices/{invoice_id}") or {}).get("meta", {})
    _upsert_revenue_entry(invoice_id, fresh_meta)

    return jsonify({"success": True, "amount": amount}), 200

@app.route("/invoicing/<invoice_id>/tax/add", methods=["POST"])
@role_required("invoicing")
def tax_payment_add(invoice_id):
    """Record a tax payment against an invoice."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    tax_log = inv_data.get("tax_payments", [])
    if not isinstance(tax_log, list):
        tax_log = []

    amount = _safe_float(request.form.get("amount") or request.form.get("tax_amount", 0))
    if amount <= 0:
        return jsonify({"error": "Invalid tax payment amount"}), 400

    tax_log.append({
        "amount":     str(amount),
        "date":       request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
        "method":     request.form.get("method", ""),
        "reference":  request.form.get("reference", ""),
        "notes":      request.form.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
    fresh_inv = dict(inv_data)
    fresh_inv["tax_payments"] = tax_log
    new_status = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "tax_payments":     tax_log,
        "meta/tax_paid":    str(tax_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })
    # Update project stage payment statuses based on payments
    _update_project_stage_payment_status(invoice_id)

    fresh_meta = (fb_get(f"/invoices/{invoice_id}") or {}).get("meta", {})
    _upsert_revenue_entry(invoice_id, fresh_meta)

    return jsonify({"success": True, "amount": amount}), 200

@app.route("/invoicing/<invoice_id>/payment/full", methods=["POST"])
@role_required("invoicing")
def payment_full(invoice_id):
    """Record a full invoice payment distributed proportionally across projects."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    log = inv_data.get("payment_log", [])
    if not isinstance(log, list):
        log = []

    amount = _safe_float(request.form.get("amount", 0))
    if amount <= 0:
        return jsonify({"error": "Invalid payment amount"}), 400

    # Get all linked projects
    linked_projects = inv_data.get("meta", {}).get("linked_projects", [])
    invoice_total = _safe_float(inv_data.get("meta", {}).get("total", 0))
    tax_amount = _safe_float(inv_data.get("meta", {}).get("tax_amount", 0))
    project_subtotal = invoice_total - tax_amount

    # Distribute payment proportionally across projects (based on REMAINING amounts)
    payment_date = request.form.get("date", datetime.now().strftime("%Y-%m-%d"))
    payment_method = request.form.get("method", "")
    payment_reference = request.form.get("reference", "")
    payment_notes = request.form.get("notes", "")

    if project_subtotal > 0 and linked_projects:
        # Calculate each project's remaining amount from invoice items and payments
        meta_items = inv_data.get("meta", {}).get("items", [])
        project_remainings = {}
        total_remaining = 0

        if isinstance(meta_items, list):
            for proj_info in linked_projects:
                if not isinstance(proj_info, dict):
                    continue

                project_number = proj_info.get("project_number", "")
                if not project_number:
                    continue

                # Find this project's original amount from invoice items
                proj_amount = 0
                for item in meta_items:
                    if isinstance(item, dict) and item.get("project_number") == project_number:
                        proj_amount = _safe_float(item.get("amount", 0))
                        break

                if proj_amount <= 0:
                    continue

                # Calculate how much this project has already been paid
                proj_paid = sum(
                    _safe_float(p.get("amount", 0))
                    for p in log
                    if p.get("project_number") == project_number
                )

                # Calculate remaining amount for this project
                proj_remaining = max(0, proj_amount - proj_paid)
                if proj_remaining > 0:
                    project_remainings[project_number] = proj_remaining
                    total_remaining += proj_remaining

        # Distribute payment proportionally across remaining amounts
        if total_remaining > 0:
            for project_number, remaining in project_remainings.items():
                # Calculate this project's share of the payment based on remaining amount
                project_share = (remaining / total_remaining) * amount

                # Create payment entry for this project
                payment_entry = {
                    "amount":     str(project_share),
                    "date":       payment_date,
                    "method":     payment_method,
                    "reference":  payment_reference,
                    "notes":      payment_notes,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "project_number": project_number,
                }
                log.append(payment_entry)
    else:
        # Single project or no projects - just add the amount
        payment_entry = {
            "amount":     str(amount),
            "date":       payment_date,
            "method":     payment_method,
            "reference":  payment_reference,
            "notes":      payment_notes,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        proj_num = inv_data.get("meta", {}).get("project_number", "")
        if proj_num:
            payment_entry["project_number"] = proj_num
        log.append(payment_entry)

    amount_paid = sum(_safe_float(p.get("amount", 0)) for p in log)
    fresh_inv = dict(inv_data)
    fresh_inv["payment_log"] = log
    new_status = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "payment_log":      log,
        "meta/amount_paid": str(amount_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })

    # Use sequential allocation for ALL invoices - it handles both single and multi-project
    # This ensures stage.amount_paid is always updated the same way
    _allocate_invoice_payment_sequential(invoice_id)

    # Update project stage payment statuses based on payments (after allocation)
    # Must do this BEFORE _sync_project_payment since it updates payment_stages
    _update_project_stage_payment_status(invoice_id)

    # Sync project-level amount_paid (sum of all stages for each project)
    # This MUST run AFTER _update_project_stage_payment_status to override its data
    # Get fresh invoice data to ensure we have correct metadata after Firebase updates
    fresh_inv = fb_get(f"/invoices/{invoice_id}") or inv_data
    linked_projects = _invoice_linked_projects(fresh_inv)
    for proj_num in linked_projects:
        _sync_project_payment(proj_num)
        # Complete the project once its own amount_paid covers its contract value
        # (checked per-project, so this also fires for multi-project invoices)
        _auto_complete_project_if_paid(proj_num)

    fresh_meta = (fb_get(f"/invoices/{invoice_id}") or {}).get("meta", {})
    _upsert_revenue_entry(invoice_id, fresh_meta)

    return jsonify({"success": True, "amount": amount}), 200

@app.route("/invoicing/<invoice_id>/payment/sequential", methods=["POST"])
@role_required("invoicing")
def payment_sequential(invoice_id):
    """Record payment distributed sequentially across projects then tax."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    meta = inv_data.get("meta", {}) or {}

    amount = _safe_float(request.form.get("amount", 0))
    status = request.form.get("status", "Partial")

    if amount <= 0:
        return jsonify({"error": "Invalid payment amount"}), 400

    payment_log = inv_data.get("payment_log", [])
    if not isinstance(payment_log, list):
        payment_log = []

    # Get invoice details
    linked_projects = meta.get("linked_projects", [])
    # Items are in line_items, not meta.items
    line_items = inv_data.get("line_items", []) or []
    tax_amount = _safe_float(meta.get("tax_amount", 0))
    main_project = meta.get("project_number", "")

    remaining_to_distribute = amount

    # Initialize tax_log - will be updated in Step 2 if needed
    tax_log = inv_data.get("tax_payments", [])
    if not isinstance(tax_log, list):
        tax_log = []

    # If no linked_projects, build from main_project
    if not linked_projects and main_project:
        linked_projects = [{"project_number": main_project, "payment_stage_index": meta.get("payment_stage_index", 0)}]

    # Step 1: For multi-project invoices, skip single primary_project allocation
    # Go straight to sequential distribution across all linked_projects (sorted)
    # For single-project invoices, handle main_project normally
    is_multi = len(linked_projects) > 1

    if not is_multi and remaining_to_distribute > 0 and main_project:
        # Single-project invoice: allocate to main_project
        # Calculate total project invoice amount from line items (excluding tax)
        proj_amount = 0
        for item in line_items:
            if isinstance(item, dict):
                proj_amount += _safe_float(item.get("amount", 0))

        if proj_amount > 0:
            # Calculate how much this project has already received
            proj_received = sum(
                _safe_float(p.get("amount", 0))
                for p in payment_log
                if p.get("project_number") == main_project
            )

            # How much more does this project need?
            proj_needs = max(0, proj_amount - proj_received)

            if proj_needs > 0:
                # Distribute amount to this project
                distribute_to_proj = min(proj_needs, remaining_to_distribute)

                # Determine stage name - try multiple sources
                _stage_name = meta.get("payment_stage", "")
                _stage_idx = meta.get("payment_stage_index")

                # Convert stage_index to int if available
                if _stage_idx is not None:
                    try:
                        _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                    except (ValueError, TypeError):
                        _stage_idx = None

                # If stage_name is empty but we have stage_index, generate a fallback
                if not _stage_name and _stage_idx is not None:
                    _stage_name = f"Stage {_stage_idx + 1}"

                payment_entry = {
                    "amount":     str(distribute_to_proj),
                    "date":       request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
                    "method":     request.form.get("method", ""),
                    "reference":  request.form.get("reference", ""),
                    "notes":      request.form.get("notes", ""),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "project_number": main_project,
                    # Metadata for tracking
                    "invoice_number": meta.get("invoice_number", ""),
                    "stage_name": _stage_name,
                    "stage_index": _stage_idx or "",
                }
                payment_log.append(payment_entry)
                remaining_to_distribute -= distribute_to_proj
                log.info(f"[PAYMENT_DIST] Recorded ${distribute_to_proj} payment for project {main_project}")

    # Fallback: if still have remaining_to_distribute and linked_projects, distribute there too
    elif remaining_to_distribute > 0 and linked_projects:
        # SORT linked_projects by project number (001, 002, 003...) before distributing
        # Handle both dict and string formats
        def get_sort_key(x):
            proj_num = x.get("project_number", "") if isinstance(x, dict) else x
            if proj_num and proj_num[-3:].isdigit():
                return int(proj_num[-3:])
            return proj_num
        sorted_projects = sorted(linked_projects, key=get_sort_key)
        for proj_info in sorted_projects:
            if remaining_to_distribute <= 0:
                continue

            # Handle both dict and string formats for project_info
            if isinstance(proj_info, dict):
                project_number = proj_info.get("project_number", "")
            else:
                project_number = proj_info
            if not project_number:
                continue

            # Find project amount from line items
            proj_amount = 0
            for item in line_items:
                if isinstance(item, dict):
                    item_proj = item.get("project_number", "").strip() or main_project
                    if item_proj == project_number:
                        proj_amount += _safe_float(item.get("amount", 0))

            if proj_amount <= 0:
                continue

            # Calculate how much this project has already received
            proj_received = sum(
                _safe_float(p.get("amount", 0))
                for p in payment_log
                if p.get("project_number") == project_number
            )

            # How much more does this project need?
            proj_needs = max(0, proj_amount - proj_received)

            if proj_needs > 0:
                # Distribute amount to this project
                distribute_to_proj = min(proj_needs, remaining_to_distribute)

                # Determine stage name - try multiple sources
                _stage_name = meta.get("payment_stage", "")
                _stage_idx = meta.get("payment_stage_index")

                # Convert stage_index to int if available
                if _stage_idx is not None:
                    try:
                        _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                    except (ValueError, TypeError):
                        _stage_idx = None

                # If stage_name is empty but we have stage_index, generate a fallback
                if not _stage_name and _stage_idx is not None:
                    _stage_name = f"Stage {_stage_idx + 1}"

                payment_entry = {
                    "amount":     str(distribute_to_proj),
                    "date":       request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
                    "method":     request.form.get("method", ""),
                    "reference":  request.form.get("reference", ""),
                    "notes":      request.form.get("notes", ""),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "project_number": project_number,
                    # Metadata for tracking
                    "invoice_number": meta.get("invoice_number", ""),
                    "stage_name": _stage_name,
                    "stage_index": _stage_idx or "",
                }
                payment_log.append(payment_entry)
                remaining_to_distribute -= distribute_to_proj

    # Step 2: Distribute remaining to tax
    if remaining_to_distribute > 0 and tax_amount > 0:
        tax_received = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
        tax_needs = max(0, tax_amount - tax_received)

        if tax_needs > 0:
            distribute_to_tax = min(tax_needs, remaining_to_distribute)
            tax_log.append({
                "amount":     str(distribute_to_tax),
                "date":       request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
                "method":     request.form.get("method", ""),
                "reference":  request.form.get("reference", ""),
                "notes":      request.form.get("notes", ""),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            remaining_to_distribute -= distribute_to_tax

        fb_update(f"/invoices/{invoice_id}", {"tax_payments": tax_log})

    # Calculate totals and update invoice
    amount_paid = sum(_safe_float(p.get("amount", 0)) for p in payment_log)
    tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)

    fresh_inv = dict(inv_data)
    fresh_inv["payment_log"] = payment_log
    new_status = _calculate_invoice_status(fresh_inv)

    log.info(f"[FIREBASE_SAVE] Saving payment_log with {len(payment_log)} entries for invoice {invoice_id}")
    log.info(f"[STATUS_UPDATE] Setting invoice status to: {new_status}")
    fb_update(f"/invoices/{invoice_id}", {
        "payment_log":      payment_log,
        "meta/amount_paid": str(amount_paid),
        "meta/tax_paid":    str(tax_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })
    log.info(f"[FIREBASE_SAVED] Invoice updated successfully with status={new_status}")

    # Use sequential allocation for multi-project invoices FIRST
    # Then update stage statuses using the allocated amounts
    log.info(f"[PAYMENT] Recording ${amount} for invoice {invoice_id}, updating stage statuses")
    fresh_inv = fb_get(f"/invoices/{invoice_id}") or inv_data
    linked_projects = _invoice_linked_projects(fresh_inv)
    if len(linked_projects) > 1:
        _allocate_invoice_payment_sequential(invoice_id)

    # Update project stage payment statuses (after allocation)
    _update_project_stage_payment_status(invoice_id)

    # Sync project-level amount_paid (sum of all stages for each project)
    # This MUST run AFTER both allocation and status update
    # This ensures Financial Summary shows TOTAL of all stages, not just this invoice
    for proj_num in linked_projects:
        _sync_project_payment(proj_num)
        # Complete the project once its own amount_paid covers its contract value
        # (checked per-project, so this also fires for multi-project invoices)
        _auto_complete_project_if_paid(proj_num)

    fresh_meta = (fb_get(f"/invoices/{invoice_id}") or {}).get("meta", {})
    _upsert_revenue_entry(invoice_id, fresh_meta)

    return jsonify({"success": True, "amount": amount}), 200

@app.route("/invoicing/<invoice_id>/payment/delete/<int:idx>", methods=["POST"])
@role_required("invoicing")
def payment_delete(invoice_id, idx):
    """Remove a payment entry from the log."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    log = inv_data.get("payment_log", [])
    if not isinstance(log, list) or idx >= len(log):
        return jsonify({"error": "Payment not found"}), 404

    log.pop(idx)
    total       = _safe_float(inv_data.get("meta", {}).get("total", 0))
    amount_paid = sum(_safe_float(p["amount"]) for p in log)
    any_paid    = amount_paid > 0
    fresh_inv = dict(inv_data)
    fresh_inv["payment_log"] = log
    new_status  = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "payment_log":      log,
        "meta/amount_paid": str(amount_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })
    # Use sequential allocation for ALL invoices - ensures both project.amount_paid and stage.amount_paid are updated
    _allocate_invoice_payment_sequential(invoice_id)

    # Sync project-level amount_paid (sum of all invoices for each project)
    # Use fresh invoice data after sequential allocation
    fresh_inv = fb_get(f"/invoices/{invoice_id}") or inv_data
    linked_projects = _invoice_linked_projects(fresh_inv)
    for proj_num in linked_projects:
        _sync_project_payment(proj_num)

    # Update project stage payment statuses based on payments (after allocation)
    _update_project_stage_payment_status(invoice_id)

    # Update project status based on remaining payments
    linked_projects = _invoice_linked_projects(fresh_inv)
    for proj_num in linked_projects:
        proj_id, pdata = _find_project_by_number(proj_num)
        if proj_id and pdata:
            amount_paid = _safe_float(pdata.get("amount_paid", 0))
            current_status = pdata.get("status", "Not Started")
            # If still has payments, change to In Progress
            if amount_paid > 0 and current_status in ("Completed", "invoiced_Fully paid"):
                fb_update(f"/projects/{proj_id}", {
                    "status": "In Progress",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
            # If no payments left, change to Not Started
            elif amount_paid == 0 and current_status != "Not Started":
                fb_update(f"/projects/{proj_id}", {
                    "status": "Not Started",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })

    return jsonify({"success": True}), 200

@app.route("/invoicing/<invoice_id>/tax/payment/delete/<int:idx>", methods=["POST"])
@role_required("invoicing")
def tax_payment_delete(invoice_id, idx):
    """Remove a tax payment entry from the log."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    tax_log = inv_data.get("tax_payments", [])
    if not isinstance(tax_log, list) or idx >= len(tax_log):
        return jsonify({"error": "Tax payment not found"}), 404

    tax_log.pop(idx)
    tax_paid = sum(_safe_float(p.get("amount", 0)) for p in tax_log)
    fresh_inv = dict(inv_data)
    fresh_inv["tax_payments"] = tax_log
    new_status = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "tax_payments":     tax_log,
        "meta/tax_paid":    str(tax_paid),
        "meta/status":      new_status,
        "meta/updated_at":  datetime.now(timezone.utc).isoformat(),
    })
    # Update project stage payment statuses based on payments
    _update_project_stage_payment_status(invoice_id)

    # Update project status based on remaining payments
    linked_projects = _invoice_linked_projects(fresh_inv)
    for proj_num in linked_projects:
        proj_id, pdata = _find_project_by_number(proj_num)
        if proj_id and pdata:
            amount_paid = _safe_float(pdata.get("amount_paid", 0))
            current_status = pdata.get("status", "Not Started")
            # If still has payments, change to In Progress
            if amount_paid > 0 and current_status in ("Completed", "invoiced_Fully paid"):
                fb_update(f"/projects/{proj_id}", {
                    "status": "In Progress",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
            # If no payments left, change to Not Started
            elif amount_paid == 0 and current_status != "Not Started":
                fb_update(f"/projects/{proj_id}", {
                    "status": "Not Started",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })

    return jsonify({"success": True}), 200

@app.route("/invoicing/<invoice_id>/payment/<int:idx>/edit", methods=["POST"])
@role_required("invoicing")
def payment_edit(invoice_id, idx):
    """Update a payment entry in the log."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    meta = inv_data.get("meta", {})
    log = inv_data.get("payment_log", [])

    if not isinstance(log, list) or idx >= len(log):
        return jsonify({"error": "Payment not found"}), 404

    # New amount being set
    new_amount = _safe_float(request.form.get("amount", 0))
    if new_amount <= 0:
        return jsonify({"error": "Invalid amount"}), 400

    # Old amount at this index
    old_amount = _safe_float(log[idx].get("amount", 0))
    amount_changed = abs(new_amount - old_amount) > 0.01

    # If only updating fields (date, method, etc.) without amount change, do simple update
    if not amount_changed:
        # Just update the single payment entry's fields
        log[idx].update({
            "amount": str(new_amount),
            "date": request.form.get("date", log[idx].get("date", "")),
            "method": request.form.get("method", ""),
            "reference": request.form.get("reference", ""),
            "notes": request.form.get("notes", ""),
        })

        invoice_paid = sum(_safe_float(p.get("amount", 0)) for p in log)
        fresh_inv = dict(inv_data)
        fresh_inv["payment_log"] = log
        fresh_inv["meta"] = meta
        new_status = _calculate_invoice_status(fresh_inv)

        meta["amount_paid"] = str(invoice_paid)
        meta["updated_at"] = datetime.now(timezone.utc).isoformat()

        fb_update(f"/invoices/{invoice_id}", {
            "payment_log": log,
            "meta/amount_paid": str(invoice_paid),
            "meta/status": new_status,
            "meta/updated_at": meta.get("updated_at"),
        })

        # Update project stages
        _update_project_stage_payment_status(invoice_id)
        return jsonify({"success": True}), 200

    # Amount changed - need to rebuild payment_log with sequential distribution
    # New total paid = (old total - old_amount) + new_amount
    old_total = sum(_safe_float(p.get("amount", 0)) for p in log)
    new_total_paid = old_total - old_amount + new_amount

    # Update meta.amount_paid
    meta["amount_paid"] = str(new_total_paid)
    meta["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Get invoice details for redistribution
    tax_amount = _safe_float(meta.get("tax_amount", 0))
    line_items = inv_data.get("line_items", []) or []
    main_project = meta.get("project_number", "")

    # Rebuild payment_log sequentially (same logic as invoice_update_amount)
    new_payment_log = []
    new_tax_log = []
    remaining_to_distribute = new_total_paid

    # Extract ALL projects from both line_items AND linked_projects metadata
    projects_from_items = set()
    for item in line_items:
        if isinstance(item, dict):
            proj_num = item.get("project_number", "")
            if proj_num:
                projects_from_items.add(proj_num)

    projects_from_meta = set()
    for proj_info in (meta.get("linked_projects") or []):
        if isinstance(proj_info, dict):
            proj_num = proj_info.get("project_number", "")
            if proj_num:
                projects_from_meta.add(proj_num)

    # Merge both sets
    all_projects = projects_from_items | projects_from_meta

    # Build linked_projects from merged list
    linked_projects = []
    if all_projects:
        linked_projects = [
            {"project_number": proj_num, "payment_stage_index": meta.get("payment_stage_index", 0)}
            for proj_num in sorted(all_projects)
        ]
    elif not linked_projects and main_project:
        linked_projects = [{"project_number": main_project, "payment_stage_index": meta.get("payment_stage_index", 0)}]

    # Get the project of the payment being edited (to apply form fields only to that project)
    old_project = log[idx].get("project_number", "") if idx < len(log) else ""

    # Step 1: Distribute to projects sequentially (same as invoice_update_amount)
    if linked_projects:
        def get_sort_key(x):
            proj_num = x.get("project_number", "") if isinstance(x, dict) else x
            if proj_num and proj_num[-3:].isdigit():
                return int(proj_num[-3:])
            return proj_num
        sorted_projects = sorted(linked_projects, key=get_sort_key)

        for proj_info in sorted_projects:
            if remaining_to_distribute <= 0:
                break

            proj_num = proj_info.get("project_number", "") if isinstance(proj_info, dict) else proj_info
            if not proj_num:
                continue

            # Get this project's line item amount
            proj_amount = sum(_safe_float(item.get("amount", 0)) for item in line_items
                            if isinstance(item, dict) and item.get("project_number", "").strip() == proj_num)

            if proj_amount > 0:
                distribute_to_proj = min(proj_amount, remaining_to_distribute)

                # Get stage info
                _stage_name = meta.get("payment_stage", "")
                _stage_idx = meta.get("payment_stage_index")
                if _stage_idx is not None:
                    try:
                        _stage_idx = int(_stage_idx) if not isinstance(_stage_idx, int) else _stage_idx
                    except (ValueError, TypeError):
                        _stage_idx = None

                if not _stage_name and _stage_idx is not None:
                    _stage_name = f"Stage {_stage_idx + 1}"

                # Only apply form fields (date, method, etc.) to the project being edited
                # Other projects get default values
                is_edited_project = (proj_num == old_project)

                new_payment_log.append({
                    "amount": str(distribute_to_proj),
                    "date": request.form.get("date", datetime.now().strftime("%Y-%m-%d")) if is_edited_project else datetime.now().strftime("%Y-%m-%d"),
                    "method": request.form.get("method", "") if is_edited_project else "",
                    "reference": request.form.get("reference", "") if is_edited_project else "",
                    "notes": request.form.get("notes", "") if is_edited_project else "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "project_number": proj_num,
                    "invoice_number": meta.get("invoice_number", ""),
                    "stage_name": _stage_name,
                    "stage_index": _stage_idx or "",
                })
                remaining_to_distribute -= distribute_to_proj

    # Step 2: Preserve existing tax payments and only add new ones if remainder
    old_tax_log = inv_data.get("tax_payments", []) or []
    if isinstance(old_tax_log, list):
        new_tax_log = list(old_tax_log)  # Preserve existing tax payments

    # Only add new tax allocation if there's remainder to distribute
    if remaining_to_distribute > 0 and tax_amount > 0:
        tax_needs = min(tax_amount, remaining_to_distribute)
        new_tax_log.append({
            "amount": str(tax_needs),
            "date": request.form.get("date", datetime.now().strftime("%Y-%m-%d")),
            "method": request.form.get("method", ""),
            "reference": request.form.get("reference", ""),
            "notes": request.form.get("notes", ""),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        remaining_to_distribute -= tax_needs

    # Calculate new status
    fresh_inv = dict(inv_data)
    fresh_inv["payment_log"] = new_payment_log
    fresh_inv["meta"] = meta
    new_status = _calculate_invoice_status(fresh_inv)

    # Update Firebase with rebuilt payment logs
    fb_update(f"/invoices/{invoice_id}", {
        "payment_log": new_payment_log,
        "tax_payments": new_tax_log if new_tax_log else [],
        "meta/amount_paid": str(new_total_paid),
        "meta/status": new_status,
        "meta/updated_at": meta.get("updated_at"),
    })

    # Update project stage payment statuses
    _allocate_invoice_payment_sequential(invoice_id)
    _update_project_stage_payment_status(invoice_id)

    return jsonify({"success": True}), 200

@app.route("/invoicing/<invoice_id>/tax/payment/edit/<int:idx>", methods=["POST"])
@role_required("invoicing")
def tax_payment_edit(invoice_id, idx):
    """Update a tax payment entry in the log."""
    inv_data = fb_get(f"/invoices/{invoice_id}") or {}
    log = inv_data.get("tax_payments", [])
    if not isinstance(log, list) or idx >= len(log):
        return jsonify({"error": "Tax payment not found"}), 404

    amount = _safe_float(request.form.get("amount", 0))
    if amount <= 0:
        return jsonify({"error": "Invalid amount"}), 400

    log[idx].update({
        "amount": str(amount),
        "date": request.form.get("date", ""),
        "method": request.form.get("method", ""),
        "reference": request.form.get("reference", ""),
        "notes": request.form.get("notes", ""),
    })

    tax_paid = sum(_safe_float(p.get("amount", 0)) for p in log)
    fresh_inv = dict(inv_data)
    fresh_inv["tax_payments"] = log
    new_status = _calculate_invoice_status(fresh_inv)

    fb_update(f"/invoices/{invoice_id}", {
        "tax_payments":         log,
        "meta/tax_paid":        str(tax_paid),
        "meta/status":          new_status,
        "meta/updated_at":      datetime.now(timezone.utc).isoformat(),
    })

    # For tax payments, just update project stage statuses (tax doesn't use sequential allocation)
    _update_project_stage_payment_status(invoice_id)

    return jsonify({"success": True}), 200

    return jsonify({"success": True}), 200

@app.route("/invoicing/send-reminders", methods=["POST"])
@role_required("invoicing")
def send_overdue_reminders():
    raw = fb_get("/invoices") or {}
    sent = 0
    errors = []
    for iid, idata in (raw.items() if isinstance(raw, dict) else []):
        if isinstance(idata, dict) and idata.get("meta", {}).get("status", "") == "Overdue":
            ok, msg = _send_overdue_reminder_email(iid, idata)
            if ok:
                sent += 1
            else:
                errors.append(msg)
    if sent:
        flash(f"Sent {sent} overdue reminder email{'s' if sent != 1 else ''}.", "success")
    if errors:
        flash(f"Some failed: {'; '.join(set(errors[:3]))}", "warning")
    if not sent and not errors:
        flash("No overdue invoices to send reminders for.", "info")
    return redirect(url_for("invoicing"))

def _send_overdue_reminder_email(invoice_id: str, invoice: dict):
    """Send overdue payment reminder to client. Returns (ok, message)."""
    settings = load_settings()
    em = settings.get("email", {})
    if not em.get("enabled"):
        return False, "Email sending disabled."
    meta = invoice.get("meta", {})
    client_name = meta.get("client_name", "")
    if not client_name:
        return False, f"No client on invoice {invoice_id}."
    raw_clients = fb_get("/clients") or {}
    client_email = ""
    if isinstance(raw_clients, dict):
        for ckey, cd in raw_clients.items():
            if isinstance(cd, dict) and (ckey.strip() == client_name or cd.get("company", "").strip() == client_name):
                client_email = cd.get("email", "")
                break
    if not client_email:
        return False, f"No email for {client_name}."
    co = settings.get("company", {})
    inv_num = meta.get("invoice_number", "")
    total   = _safe_float(meta.get("total", 0))
    paid    = _safe_float(meta.get("amount_paid", 0))
    balance = total - paid
    html_body = f"""<html><body style="font-family:Arial,sans-serif;color:#1a1a1a;">
<div style="max-width:600px;margin:0 auto;padding:24px;">
  <h2 style="color:#DC2626;">Payment Reminder — Invoice #{inv_num}</h2>
  <p>Dear {client_name},</p>
  <p>This is a reminder that invoice <strong>#{inv_num}</strong> is now
     <strong style="color:#DC2626;">overdue</strong>.</p>
  <table style="width:100%;border-collapse:collapse;margin:16px 0;">
    <tr><td style="padding:8px;border-bottom:1px solid #eee;">Invoice #</td>
        <td style="padding:8px;border-bottom:1px solid #eee;font-weight:bold;">{inv_num}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #eee;">Due Date</td>
        <td style="padding:8px;border-bottom:1px solid #eee;color:#DC2626;">{meta.get('due_date','—')}</td></tr>
    <tr><td style="padding:8px;border-bottom:1px solid #eee;">Invoice Total</td>
        <td style="padding:8px;border-bottom:1px solid #eee;">${total:,.2f}</td></tr>
    <tr><td style="padding:8px;">Balance Due</td>
        <td style="padding:8px;font-weight:bold;font-size:18px;color:#DC2626;">${balance:,.2f}</td></tr>
  </table>
  <p>Please arrange payment at your earliest convenience. Contact us at
     <a href="mailto:{em.get('smtp_user','')}">{em.get('smtp_user','')}</a> with any questions.</p>
  <p style="margin-top:24px;">Best regards,<br><strong>{co.get('name','')}</strong></p>
</div></body></html>"""
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Payment Reminder — Invoice #{inv_num} OVERDUE"
        msg["From"]    = f"{em.get('from_name', co.get('name',''))} <{em.get('smtp_user','')}>"
        msg["To"]      = client_email
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP(em.get("smtp_host", "smtp.gmail.com"), int(em.get("smtp_port", 587)), timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(em.get("smtp_user", ""), em.get("smtp_password", ""))
            srv.sendmail(em.get("smtp_user", ""), [client_email], msg.as_string())
        return True, f"Reminder sent to {client_email}."
    except Exception as exc:
        return False, str(exc)

# ── Payment Plan Management API ────────────────────────────────────────────────
@app.route("/api/projects/<project_id>/payment-plan", methods=["POST"])
@role_required("projects")
def update_project_payment_plan(project_id):
    """Update payment plan amounts with smart auto-adjustment and permission controls"""
    try:
        data = request.get_json()
        amounts = data.get("amounts", [])
        invoice_id = data.get("invoiceId", "")
        original_amount = _safe_float(data.get("originalAmount", 0))
        new_amount = _safe_float(data.get("newAmount", 0))
        permission_granted = data.get("permissionGranted", False)
        skip_auto_adjust = data.get("skipAutoAdjust", False)

        project = fb_get(f"/projects/{project_id}") or {}
        stages = project.get("payment_stages", [])
        old_contract_value = _safe_float(project.get("contract_value", 0))

        # Find which stage was edited (by comparing sent amounts with current)
        edited_idx = -1
        old_amount = 0
        new_amount = 0
        for idx, amount_data in enumerate(amounts):
            if idx < len(stages):
                current_stage_amount = _safe_float(stages[idx].get("amount", 0))
                new_amt = _safe_float(amount_data.get("amount", 0))
                if abs(current_stage_amount - new_amt) > 0.01:
                    edited_idx = idx
                    old_amount = current_stage_amount
                    new_amount = new_amt
                    break

        if edited_idx < 0:
            return {"success": False, "error": "No stage changed detected"}, 400

        # Count invoiced and non-invoiced stages
        invoiced_count = sum(1 for s in stages if s.get("status") in ["Invoiced", "Paid", "Partially Paid", "Overdue"])
        non_invoiced_count = sum(1 for s in stages if s.get("status") == "Pending Invoice")

        # Check if edited stage is invoiced
        edited_stage_invoiced = stages[edited_idx].get("status") in ["Invoiced", "Paid", "Partially Paid", "Overdue"]

        # Find non-invoiced stages (excluding the one being edited)
        uninvoiced_indices = [i for i in range(len(stages))
                             if i != edited_idx and stages[i].get("status") == "Pending Invoice"]

        # Determine if permission is needed
        # Never ask if user is manually editing all amounts (they balance themselves)
        needs_permission = False
        if not skip_auto_adjust:
            # Case 1: All stages are invoiced and user is editing one
            if non_invoiced_count == 0 and edited_stage_invoiced:
                needs_permission = True
            # Case 2: Editing a non-invoiced stage but all other stages are invoiced (no other stages to auto-adjust to)
            elif not edited_stage_invoiced and len(uninvoiced_indices) == 0:
                needs_permission = True

        # If permission needed and not granted, return request for permission
        if needs_permission and not permission_granted:
            new_contract_value = old_contract_value + (new_amount - old_amount)
            return {
                "success": False,
                "needs_permission": True,
                "message": f"This will change the contract value from ${old_contract_value:,.2f} to ${new_contract_value:,.2f}. Continue?",
                "new_contract_value": new_contract_value
            }, 400

        # Auto-adjust logic: ONLY adjust non-invoiced stages (never adjust invoiced)
        # Skip auto-adjust if user manually edited all amounts
        if not skip_auto_adjust:
            difference = new_amount - old_amount

            # If there are non-invoiced stages, distribute the difference to them
            if uninvoiced_indices:
                per_stage = difference / len(uninvoiced_indices)
                for idx in uninvoiced_indices:
                    current_amt = _safe_float(amounts[idx].get("amount", 0)) if idx < len(amounts) else _safe_float(stages[idx].get("amount", 0))
                    amounts[idx] = {"index": idx, "amount": max(0, current_amt - per_stage)}

        # Calculate total from (possibly adjusted) amounts
        total = sum(_safe_float(a.get("amount", 0)) for a in amounts)

        # Update contract value only if:
        # - NOT manually editing all amounts (skipAutoAdjust = user keeps CV stable)
        # - AND permission was granted OR no permission was needed
        if not skip_auto_adjust and abs(total - old_contract_value) > 0.01:
            if needs_permission or non_invoiced_count > 0 or invoiced_count > 0:
                project["contract_value"] = str(total)

        # Update all stages with new amounts; sync back to linked change orders
        change_orders = project.get("change_orders") or []
        if not isinstance(change_orders, list):
            change_orders = list(change_orders.values()) if isinstance(change_orders, dict) else []

        for amount_data in amounts:
            idx = amount_data.get("index", 0)
            if idx < len(stages):
                stage = stages[idx]
                new_stage_amt = _safe_float(amount_data.get("amount", 0))
                stage["amount"] = new_stage_amt
                # Sync CO amount — try co_index first, then fall back to stage name match
                co_idx = stage.get("co_index")
                if co_idx is not None and isinstance(co_idx, int) and co_idx < len(change_orders):
                    change_orders[co_idx]["amount"] = new_stage_amt
                else:
                    stage_name = stage.get("name", "")
                    if "CO-" in stage_name:
                        co_num = stage_name.split(" ")[0] if " " in stage_name else stage_name
                        for co in change_orders:
                            if isinstance(co, dict) and co.get("co_number", "") == co_num:
                                co["amount"] = str(new_stage_amt)
                                break

        project["payment_stages"] = stages
        project["change_orders"] = change_orders
        project["updated_at"] = datetime.now(timezone.utc).isoformat()
        fb_update(f"/projects/{project_id}", project)

        # Update invoice if this was an invoiced stage
        if invoice_id:
            invoice = fb_get(f"/invoices/{invoice_id}") or {}
            meta = invoice.get("meta", {})

            # Update invoice amount
            amount_diff = new_amount - original_amount
            old_invoice_total = _safe_float(meta.get("total", 0))
            new_invoice_total = old_invoice_total + amount_diff

            meta["total"] = str(new_invoice_total)
            meta["subtotal"] = str(new_invoice_total - _safe_float(meta.get("tax_amount", 0)))
            invoice["meta"] = meta

            # Update line items if they exist
            line_items = invoice.get("line_items", [])
            if line_items:
                line_items[0]["amount"] = str(new_amount)
                line_items[0]["unit_price"] = str(new_amount)
                invoice["line_items"] = line_items

            invoice["updated_at"] = datetime.now(timezone.utc).isoformat()
            fb_update(f"/invoices/{invoice_id}", invoice)

        return {"success": True, "message": "Payment plan updated"}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.route("/api/projects/<project_id>/stage/<int:stage_idx>/update", methods=["POST"])
@role_required("projects")
def update_project_stage(project_id, stage_idx):
    """Update a single stage payment amount"""
    try:
        data = request.get_json()
        new_amount = _safe_float(data.get("newAmount", 0))
        original_amount = _safe_float(data.get("originalAmount", 0))
        paid_amount = _safe_float(data.get("paidAmount", 0))

        project = fb_get(f"/projects/{project_id}") or {}
        stages = project.get("payment_stages", [])

        if stage_idx >= len(stages):
            return {"success": False, "error": "Invalid stage index"}, 400

        # Update the stage amount
        old_amount = stages[stage_idx].get("amount", 0)
        stages[stage_idx]["amount"] = new_amount

        # If there's an invoice for this stage, update it
        invoice_id = stages[stage_idx].get("invoice_id", "")
        if invoice_id:
            invoice = fb_get(f"/invoices/{invoice_id}") or {}
            meta = invoice.get("meta", {})

            # Update the invoice total
            old_invoice_total = _safe_float(meta.get("total", 0))
            amount_diff = new_amount - original_amount
            new_invoice_total = old_invoice_total + amount_diff

            meta["total"] = str(new_invoice_total)
            meta["subtotal"] = str(new_invoice_total - _safe_float(meta.get("tax_amount", 0)))

            # For multi-project invoices, also update line items and linked_projects to stay in sync
            line_items = invoice.get("line_items", [])
            if line_items and amount_diff != 0:
                # Find and update the line item for this project's stage
                project_number = project.get("project_number", "")

                # For each line item, check if it's for this stage/project
                for item in line_items:
                    if isinstance(item, dict):
                        # Check if this item is for the current project (or default project)
                        item_proj = item.get("project_number", "").strip() or meta.get("project_number", "")
                        stage_name = stages[stage_idx].get("name", "")
                        item_desc = item.get("description", "")

                        # Match by project number and stage name in description
                        if item_proj == project_number and stage_name and stage_name in item_desc:
                            item["amount"] = str(new_amount)
                            item["unit_price"] = str(new_amount)
                            break

                # Recalculate linked_projects from line items
                # For multi-project invoices on the same payment plan, all projects use the same stage_idx
                projects_in_items = {}
                for item in line_items:
                    if isinstance(item, dict):
                        proj_num = item.get("project_number", "")
                        if proj_num:
                            # Use the stage_idx being edited - all projects in same invoice use same stage
                            projects_in_items[proj_num] = stage_idx

                # Update linked_projects if multiple projects
                if len(projects_in_items) > 1:
                    meta["linked_projects"] = [
                        {"project_number": proj_num, "payment_stage_index": projects_in_items[proj_num]}
                        for proj_num in sorted(projects_in_items.keys())
                    ]

                invoice["line_items"] = line_items

            invoice["meta"] = meta
            fb_update(f"/invoices/{invoice_id}", invoice)

        # Redistribute remaining amounts to uninvoiced stages (all except the one being edited)
        amount_diff = new_amount - original_amount
        remaining_uninvoiced = [i for i in range(len(stages))
                               if i != stage_idx and stages[i].get("status") == "Pending Invoice"]

        if remaining_uninvoiced and abs(amount_diff) > 0.01:
            per_stage = amount_diff / len(remaining_uninvoiced)
            for idx in remaining_uninvoiced:
                stages[idx]["amount"] = _safe_float(stages[idx].get("amount", 0)) + per_stage

        project["payment_stages"] = stages
        project["updated_at"] = datetime.now(timezone.utc).isoformat()
        fb_update(f"/projects/{project_id}", project)

        # If this stage has an invoice, resync ONLY this project's stage (not all linked projects)
        invoice_id = stages[stage_idx].get("invoice_id", "")
        if invoice_id:
            # Only update this specific project/stage, not all projects in the invoice
            proj_num = project.get("project_number", "")
            print(f"[UPDATE_STAGE] Resyncing invoice {invoice_id} for project {proj_num} stage {stage_idx}", flush=True)
            _update_single_project_stage_payment_status(invoice_id, proj_num, stage_idx)

        return {"success": True, "message": "Stage updated"}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.route("/api/projects/<project_id>/stage/<int:stage_idx>/update-invoiced", methods=["POST"])
@role_required("projects")
def update_invoiced_stage(project_id, stage_idx):
    """Update an invoiced stage amount and redistribute remaining to non-invoiced stages"""
    try:
        data = request.get_json()
        invoice_id = data.get("invoiceId", "")
        new_amount = _safe_float(data.get("newAmount", 0))
        original_amount = _safe_float(data.get("originalAmount", 0))
        updated_stages = data.get("stages", [])

        project = fb_get(f"/projects/{project_id}") or {}
        stages = project.get("payment_stages", [])

        if stage_idx >= len(stages):
            return {"success": False, "error": "Invalid stage index"}, 400

        # Update all stages with redistributed amounts
        for stage_data in updated_stages:
            idx = stage_data.get("index", 0)
            if idx < len(stages):
                stages[idx]["amount"] = _safe_float(stage_data.get("amount", 0))

        project["payment_stages"] = stages
        project["updated_at"] = datetime.now(timezone.utc).isoformat()
        fb_update(f"/projects/{project_id}", project)

        # Update the invoice if it exists
        if invoice_id:
            invoice = fb_get(f"/invoices/{invoice_id}") or {}
            meta = invoice.get("meta", {})

            # Update invoice amount
            old_invoice_total = _safe_float(meta.get("total", 0))
            amount_diff = original_amount - new_amount
            new_invoice_total = old_invoice_total - amount_diff

            meta["total"] = str(new_invoice_total)
            meta["subtotal"] = str(new_invoice_total - _safe_float(meta.get("tax_amount", 0)))
            invoice["meta"] = meta

            # Update line items if they exist
            line_items = invoice.get("line_items", [])
            if line_items:
                line_items[0]["amount"] = str(new_amount)
                line_items[0]["unit_price"] = str(new_amount)
                invoice["line_items"] = line_items

            invoice["updated_at"] = datetime.now(timezone.utc).isoformat()
            fb_update(f"/invoices/{invoice_id}", invoice)

            # Resync project stage payment status to ensure invoice_number is properly linked
            _update_project_stage_payment_status(invoice_id)

        return {"success": True, "message": "Invoiced amount updated and redistributed"}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.route("/api/projects/<project_id>/stage/<int:stage_idx>/invoice", methods=["POST"])
@role_required("projects")
def quick_invoice_stage(project_id, stage_idx):
    """Quickly create an invoice for a specific payment stage - using invoice_new logic."""
    try:
        project = fb_get(f"/projects/{project_id}") or {}
        stages = project.get("payment_stages", [])

        # Validate stage
        if stage_idx >= len(stages) or not isinstance(stages[stage_idx], dict):
            return {"success": False, "error": "Invalid stage index"}, 400

        stage = stages[stage_idx]
        if stage.get("status") != "Pending Invoice":
            return {"success": False, "error": "Stage is not pending"}, 400

        # Get project details
        proj_num = project.get("project_number", "")
        client_name = project.get("client_name", "")
        stage_name = stage.get("name", f"Stage {stage_idx + 1}")
        stage_amount = _safe_float(stage.get("amount", 0))

        # Create invoice using invoice_new logic (with ALL proper fields)
        invoice_data = {
            "meta": {
                "invoice_number": _next_invoice_number(),
                "project_number": proj_num,
                "client_name": client_name,
                "invoice_date": datetime.now().strftime("%Y-%m-%d"),
                "due_date": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "created_by": session.get("user_email", ""),
                "status": "Draft",
                "subtotal": str(stage_amount),
                "tax_rate": "0",
                "tax_amount": "0",
                "total": str(stage_amount),
                "amount_paid": "0",
                "notes": "",
                "terms": "",
                "payment_method": "",
                "payment_stage_index": stage_idx,
                "payment_stage": stage_name,
                "linked_projects": [{"project_number": proj_num, "payment_stage_index": stage_idx}],
            },
            "line_items": [{
                "description": stage_name,
                "project_number": proj_num,
                "quantity": "1",
                "unit_price": str(stage_amount),
                "amount": str(stage_amount),
            }],
        }

        # Create invoice
        invoice_id = fb_push("/invoices", invoice_data)
        invoice_number = invoice_data["meta"].get("invoice_number", "")

        # Mark stage as invoiced with the stage amount
        _mark_project_stage(proj_num, stage_idx, "Invoiced", invoice_id=invoice_id, invoice_number=invoice_number, amount=stage_amount)

        return {"success": True, "invoice_id": invoice_id, "invoice_number": invoice_number}, 200
    except Exception as e:
        import traceback
        log.error(f"Quick invoice error: {str(e)}\n{traceback.format_exc()}")
        return {"success": False, "error": str(e)}, 500

# ── Timesheet helpers ─────────────────────────────────────────────────────────
def _fmt_time_12(t: str) -> str:
    """Convert HH:MM to '9:00 AM' format (cross-platform)."""
    if not t:
        return ""
    try:
        h, m = t.split(":")
        h = int(h)
        ampm = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m} {ampm}"
    except Exception:
        return t

def _week_monday(date_str: str = "") -> str:
    """Return Monday ISO date for given week (or current week if blank)."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        d = datetime.now().date()
    return (d - timedelta(days=d.weekday())).isoformat()

def _load_timesheets(week_of: str = "", employee_uid: str = "") -> list:
    raw = fb_get("/timesheets") or {}
    sheets = []
    if not isinstance(raw, dict):
        return sheets
    for tsid, tsdata in raw.items():
        if not isinstance(tsdata, dict):
            continue
        if week_of and tsdata.get("week_of") != week_of:
            continue
        if employee_uid and tsdata.get("employee_uid") != employee_uid:
            continue
        tsdata = dict(tsdata)
        tsdata["firebase_id"] = tsid
        entries = tsdata.get("entries")
        if isinstance(entries, dict):
            tsdata["entries"] = list(entries.values())
        elif not isinstance(entries, list):
            tsdata["entries"] = []
        sheets.append(tsdata)
    return sheets

# ── Routes: Timesheets ────────────────────────────────────────────────────────
@app.route("/timesheets")
@role_required("timesheets")
def timesheets():
    is_admin = normalize_role(session.get("user_role", "")) == "admin"
    uid = session.get("user_uid", "")

    week_of    = _week_monday(request.args.get("week", ""))
    week_start = datetime.strptime(week_of, "%Y-%m-%d").date()
    week_end   = week_start + timedelta(days=6)
    week_label = f"{week_start.strftime('%b %d')} — {week_end.strftime('%b %d, %Y')}"
    week_dates = [(week_start + timedelta(days=i)).isoformat() for i in range(7)]
    DAY_ABBRS  = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    day_labels = [
        {"abbr": DAY_ABBRS[i],
         "date": week_dates[i],
         "display": (week_start + timedelta(days=i)).strftime("%m/%d")}
        for i in range(7)
    ]
    prev_week = (week_start - timedelta(days=7)).isoformat()
    next_week = (week_start + timedelta(days=7)).isoformat()

    if is_admin:
        week_sheets    = _load_timesheets(week_of=week_of)
        all_users      = _load_all_users()
        employees      = [u for u in all_users if u.get("active", True)]
        sheets_by_uid  = {s["employee_uid"]: s for s in week_sheets if s.get("employee_uid")}

        grid_rows = []
        for emp in employees:
            emp_uid  = emp.get("firebase_uid", "")
            emp_name = emp.get("username") or emp.get("email", "Unknown")
            sheet    = sheets_by_uid.get(emp_uid)
            cells       = {}
            total_hours = 0.0
            if sheet:
                for entry in (sheet.get("entries") or []):
                    d   = entry.get("date", "")
                    hrs = _safe_float(entry.get("total_hours", 0))
                    if d in week_dates:
                        total_hours += hrs
                        if d in cells:
                            cells[d]["hours"]    += hrs
                            pn = entry.get("project_name", "")
                            if pn:
                                cells[d]["projects"].append(pn)
                        else:
                            st = entry.get("start_time", "")
                            et = entry.get("end_time", "")
                            cells[d] = {
                                "time_range": f"{_fmt_time_12(st)} – {_fmt_time_12(et)}" if st and et else "",
                                "hours":      hrs,
                                "projects":   [entry.get("project_name", "")] if entry.get("project_name") else [],
                                "status":     sheet.get("status", "Draft"),
                                "sheet_id":   sheet.get("firebase_id", ""),
                            }
            grid_rows.append({"uid": emp_uid, "name": emp_name,
                               "cells": cells, "total_hours": total_hours, "sheet": sheet})

        kpi_submitted = sum(1 for s in week_sheets if s.get("status") in ("Submitted", "Approved", "Rejected"))
        kpi_pending   = sum(1 for s in week_sheets if s.get("status") == "Submitted")
        kpi_approved  = sum(1 for s in week_sheets if s.get("status") == "Approved")
        kpi_hours     = sum(_safe_float(s.get("total_hours", 0)) for s in week_sheets)

        return render_template("timesheets.html",
            view="admin", week_of=week_of, week_label=week_label,
            week_dates=week_dates, day_labels=day_labels,
            prev_week=prev_week, next_week=next_week,
            grid_rows=grid_rows,
            kpi_employees=len(employees), kpi_submitted=kpi_submitted,
            kpi_pending=kpi_pending, kpi_approved=kpi_approved, kpi_hours=kpi_hours)
    else:
        # Load ALL sheets for this employee (history + current week)
        my_sheets = _load_timesheets(employee_uid=uid)
        my_sheets.sort(key=lambda x: x.get("week_of", ""), reverse=True)

        # Build personal grid cells for the selected week
        week_sheet = next((s for s in my_sheets if s.get("week_of") == week_of), None)
        emp_cells  = {}
        week_total = 0.0
        if week_sheet:
            for entry in (week_sheet.get("entries") or []):
                d   = entry.get("date", "")
                hrs = _safe_float(entry.get("total_hours", 0))
                if d in week_dates:
                    week_total += hrs
                    if d in emp_cells:
                        emp_cells[d]["hours"] += hrs
                        pn = entry.get("project_name", "")
                        if pn:
                            emp_cells[d]["projects"].append(pn)
                    else:
                        st = entry.get("start_time", "")
                        et = entry.get("end_time", "")
                        emp_cells[d] = {
                            "time_range": f"{_fmt_time_12(st)} – {_fmt_time_12(et)}" if st and et else "",
                            "hours":      hrs,
                            "projects":   [entry.get("project_name", "")] if entry.get("project_name") else [],
                            "status":     week_sheet.get("status", "Draft"),
                            "sheet_id":   week_sheet.get("firebase_id", ""),
                        }

        # Monthly stats (current calendar month)
        current_month = datetime.now().strftime("%Y-%m")
        month_sheets  = [s for s in my_sheets if s.get("week_of", "")[:7] == current_month]
        stat_month_hours   = sum(_safe_float(s.get("total_hours", 0)) for s in month_sheets)
        stat_month_ot      = sum(_safe_float(s.get("total_overtime_hours", 0)) for s in month_sheets)
        stat_approved      = sum(1 for s in my_sheets if s.get("status") == "Approved")
        stat_pending       = sum(1 for s in my_sheets if s.get("status") == "Submitted")

        current_week = _week_monday()
        return render_template("timesheets.html",
            view="my", my_sheets=my_sheets,
            week_of=week_of, week_label=week_label,
            week_dates=week_dates, day_labels=day_labels,
            prev_week=prev_week, next_week=next_week,
            week_sheet=week_sheet, emp_cells=emp_cells, week_total=week_total,
            current_week=current_week,
            stat_month_hours=stat_month_hours, stat_month_ot=stat_month_ot,
            stat_approved=stat_approved, stat_pending=stat_pending)


@app.route("/timesheets/submit")
@login_required
def timesheets_submit():
    week_of    = _week_monday(request.args.get("week", ""))
    uid        = session.get("user_uid", "")
    all_projs  = _load_projects_list()
    active_projects = [p for p in all_projs
                       if p.get("status", "") not in ("Completed", "Cancelled")]
    existing   = _load_timesheets(week_of=week_of, employee_uid=uid)
    existing_sheet = existing[0] if existing else None

    week_start = datetime.strptime(week_of, "%Y-%m-%d").date()
    week_end   = week_start + timedelta(days=6)
    week_label = f"{week_start.strftime('%b %d')} — {week_end.strftime('%b %d, %Y')}"
    DAY_NAMES  = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    week_days  = [
        {"name": DAY_NAMES[i],
         "abbr": DAY_NAMES[i][:3],
         "date": (week_start + timedelta(days=i)).isoformat()}
        for i in range(7)
    ]
    return render_template("timesheet_submit.html",
        week_of=week_of, week_label=week_label,
        week_days=week_days, active_projects=active_projects,
        existing_sheet=existing_sheet)


@app.route("/timesheets/<sheet_id>")
@login_required
def timesheet_detail(sheet_id):
    sheet = fb_get(f"/timesheets/{sheet_id}")
    if not sheet or not isinstance(sheet, dict):
        abort(404)
    sheet = dict(sheet)
    sheet["firebase_id"] = sheet_id
    entries = sheet.get("entries")
    if isinstance(entries, dict):
        entries = list(entries.values())
    elif not isinstance(entries, list):
        entries = []
    sheet["entries"] = sorted(entries, key=lambda e: (e.get("date", ""), e.get("start_time", "")))

    uid      = session.get("user_uid", "")
    is_admin = normalize_role(session.get("user_role", "")) == "admin"
    if not is_admin and sheet.get("employee_uid") != uid:
        flash("You don't have permission to view this timesheet.", "danger")
        return redirect(url_for("timesheets"))

    by_date = {}
    for entry in sheet["entries"]:
        d = entry.get("date", "")
        if d not in by_date:
            by_date[d] = []
        by_date[d].append(entry)

    return render_template("timesheet_detail.html",
        sheet=sheet, by_date=by_date, is_admin=is_admin)


@app.route("/api/timesheets", methods=["POST"])
@login_required
def api_timesheets_save():
    data    = request.get_json(force=True) or {}
    uid     = session.get("user_uid", "")
    name    = session.get("user_name", "")
    week_of = (data.get("week_of") or "").strip()
    action  = data.get("action", "draft")
    entries = data.get("entries") or []

    if not week_of:
        return jsonify({"error": "week_of is required"}), 400

    total_reg = sum(_safe_float(e.get("regular_hours", 0)) for e in entries)
    total_ot  = sum(_safe_float(e.get("overtime_hours", 0)) for e in entries)
    now_iso   = datetime.now(timezone.utc).isoformat()

    sheet_data = {
        "employee_uid":         uid,
        "employee_name":        name,
        "week_of":              week_of,
        "status":               "Submitted" if action == "submit" else "Draft",
        "entries":              entries,
        "total_regular_hours":  total_reg,
        "total_overtime_hours": total_ot,
        "total_hours":          total_reg + total_ot,
        "updated_at":           now_iso,
    }
    if action == "submit":
        sheet_data["submitted_at"] = now_iso

    existing = _load_timesheets(week_of=week_of, employee_uid=uid)
    if existing:
        sheet_id = existing[0]["firebase_id"]
        fb_update(f"/timesheets/{sheet_id}", sheet_data)
    else:
        sheet_data["created_at"] = now_iso
        sheet_id = fb_push("/timesheets", sheet_data)

    return jsonify({"success": True, "sheet_id": sheet_id, "status": sheet_data["status"]})


@app.route("/api/timesheets/<sheet_id>/approve", methods=["POST"])
@role_required("timesheets")
def api_timesheets_approve(sheet_id):
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    data   = request.get_json(force=True) or {}
    action = data.get("action", "approve")
    notes  = (data.get("notes") or "").strip()
    status = "Approved" if action == "approve" else "Rejected"
    fb_update(f"/timesheets/{sheet_id}", {
        "status":          status,
        "approved_by":     session.get("user_name", ""),
        "approved_at":     datetime.now(timezone.utc).isoformat(),
        "rejection_notes": notes,
    })
    return jsonify({"success": True, "status": status})


@app.route("/api/admin/reset-timesheets", methods=["POST"])
@login_required
def api_admin_reset_timesheets():
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    fb_delete("/timesheets")
    return jsonify({"success": True, "message": "All timesheets deleted."})


@app.route("/api/timesheets/<sheet_id>/delete", methods=["POST"])
@login_required
def api_timesheet_delete(sheet_id):
    if normalize_role(session.get("user_role", "")) != "admin":
        return jsonify({"error": "Admin access required"}), 403
    sheet = fb_get(f"/timesheets/{sheet_id}")
    if not sheet:
        return jsonify({"error": "Timesheet not found"}), 404
    fb_delete(f"/timesheets/{sheet_id}")
    return jsonify({"success": True})


@app.route("/api/timesheets/previous")
@login_required
def api_timesheet_previous():
    """Return project entries from the most recent timesheet before the given week."""
    uid = session.get("user_uid", "")
    week_of = request.args.get("week", "")
    all_sheets = fb_get("/timesheets") or {}
    my_sheets = []
    for k, v in all_sheets.items():
        if isinstance(v, dict) and v.get("employee_uid") == uid:
            v["firebase_id"] = k
            my_sheets.append(v)
    prev_sheets = [s for s in my_sheets if s.get("week_of", "") < week_of] if week_of else my_sheets
    if not prev_sheets:
        return jsonify({"entries": []})
    prev_sheets.sort(key=lambda x: x.get("week_of", ""), reverse=True)
    last = prev_sheets[0]
    carry = [{"project_number": e.get("project_number",""), "project_name": e.get("project_name",""),
              "project_id": e.get("project_id",""), "date": e.get("date","")}
             for e in (last.get("entries") or [])]
    return jsonify({"entries": carry, "week_of": last.get("week_of","")})

@app.route("/api/timesheets/export")
@role_required("timesheets")
def api_timesheets_export():
    import csv
    import io as _io
    if normalize_role(session.get("user_role", "")) != "admin":
        flash("Admin access required.", "danger")
        return redirect(url_for("timesheets"))

    week_of = request.args.get("week", _week_monday())
    fmt = request.args.get("fmt", "csv")
    sheets = _load_timesheets(week_of=week_of)
    co = company_info()

    headers = ["Week Of", "Employee", "Status", "Date", "Day", "Project",
               "Project #", "Start", "Lunch Mins", "End",
               "Regular Hrs", "OT Hrs", "Total Hrs", "Notes"]
    rows = []

    for sheet in sheets:
        for entry in (sheet.get("entries") or []):
            dt = entry.get("date", "")
            try:
                day_name = datetime.strptime(dt, "%Y-%m-%d").strftime("%A")
            except Exception:
                day_name = ""
            rows.append([
                sheet.get("week_of", ""), sheet.get("employee_name", ""),
                sheet.get("status", ""), dt, day_name,
                entry.get("project_name", ""), entry.get("project_number", ""),
                entry.get("start_time", ""), entry.get("lunch_break_mins", ""),
                entry.get("end_time", ""),
                entry.get("regular_hours", 0), entry.get("overtime_hours", 0),
                entry.get("total_hours", 0), entry.get("notes", ""),
            ])

    if fmt == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timesheets"

        hdr_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
        hdr_font = Font(color="FFFFFFFF", bold=True, size=11)
        title_font = Font(bold=True, size=13, color="FF0F766E")
        alt_fill = PatternFill(start_color="FFF8FAFC", end_color="FFF8FAFC", fill_type="solid")
        ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

        co = company_info()
        ws.merge_cells('A1:N1')
        title_cell = ws.cell(row=1, column=1, value=f"{co.get('name','')} - Timesheets Report (Week of {week_of})")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        header_row = 2
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr

        for ri, row in enumerate(rows, header_row + 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = ctr

        col_widths = [12, 18, 12, 12, 12, 20, 14, 10, 12, 10, 12, 10, 12, 20]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = _io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True, download_name=f"timesheets_{week_of}.xlsx")

    elif fmt == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=0.5*inch, rightMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        co = company_info()
        elems = []

        title_s = ParagraphStyle("T", parent=styles["Normal"], fontSize=14,
                                  fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#0F766E"), spaceAfter=3,
                                  alignment=1)
        elems.append(Paragraph(f"{co.get('name','')} — Timesheets Report", title_s))
        elems.append(Spacer(1, 0.15*inch))
        sub_s = ParagraphStyle("S", parent=styles["Normal"], fontSize=10,
                               textColor=colors.HexColor("#64748B"), spaceAfter=10,
                               alignment=0)
        elems.append(Paragraph(f"Week of {week_of}", sub_s))
        elems.append(Spacer(1, 0.15*inch))

        data = [headers]
        cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=1, leading=9, wordWrap='CJK')

        for row in rows:
            data.append([Paragraph(str(cell), cell_style) for cell in row])

        cw = [1.0*inch, 1.3*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.4*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.7*inch, 0.9*inch, 1.2*inch]
        tbl = Table(data, colWidths=cw, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 8),
            ("ALIGN",         (0,0), (-1,0), "CENTER"),
            ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,0), 6),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 7),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#E2E8F0")),
            ("TOPPADDING",    (0,1), (-1,-1), 3),
            ("BOTTOMPADDING", (0,1), (-1,-1), 3),
            ("ALIGN",         (0,1), (-1,-1), "CENTER"),
            ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
        ]))
        elems.append(tbl)
        doc.build(elems)
        buf.seek(0)

        return send_file(buf, mimetype="application/pdf",
                        as_attachment=True, download_name=f"timesheets_{week_of}.pdf")

    else:  # CSV
        import csv
        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        csv_bytes = output.getvalue().encode("utf-8-sig")
        return send_file(
            _io.BytesIO(csv_bytes),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"timesheets_{week_of}.csv",
        )

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(
        debug=os.environ.get("FLASK_DEBUG", "0") == "1",
        host=os.environ.get("FLASK_HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5000")),
    )
