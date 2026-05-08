from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtChart import QChart, QChartView, QBarSet, QBarSeries, QBarCategoryAxis, QValueAxis
from PyQt5.QtCore import Qt
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

class ExpensesViewerWindow(QtWidgets.QMainWindow):
    """Full window expenses viewer with category-wise display"""
    
    def __init__(self, expenses_data, parent=None):
        super().__init__(parent)
        self.expenses_data = expenses_data
        self.current_category = None
        self.category_colors = {
            "Office Supplies": "#3498db",
            "Equipment & Hardware": "#e74c3c", 
            "Travel & Transportation": "#2ecc71",
            "Meals & Entertainment": "#f39c12",
            "Software & Subscriptions": "#9b59b6",
            "Professional Services": "#1abc9c",
            "Utilities & Bills": "#34495e",
            "Rent & Lease": "#e67e22",
            "Marketing & Advertising": "#d35400",
            "Training & Development": "#16a085",
            "Vehicle & Fuel": "#27ae60",
            "Maintenance & Repairs": "#8e44ad",
            "Insurance Premiums": "#2c3e50",
            "Legal & Professional Fees": "#f1c40f",
            "Shipping & Delivery": "#7f8c8d",
            "Client Gifts & Relations": "#c0392b",
            "Communication": "#2980b9",
            "Health & Safety": "#e74c3c",
            "Bank & Finance Charges": "#95a5a6",
            "Miscellaneous": "#bdc3c7"
        }
        
        self.init_ui()
        self.load_categories()
        
        # Select first category automatically
        self.select_first_category()
        
    def select_first_category(self):
        """Select the first category in the list automatically"""
        if self.categories_list.count() > 0:
            first_item = self.categories_list.item(0)
            self.categories_list.setCurrentItem(first_item)
            self.on_category_selected(first_item)
        
    def init_ui(self):
        """Initialize the UI"""
        self.setWindowTitle("📊 Expenses Viewer - Category Analysis")
        self.setMinimumSize(1600, 1000)
        
        # Show maximized
        self.showMaximized()
        
        # Central widget
        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QtWidgets.QHBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Left panel - Categories
        left_panel = self.create_categories_panel()
        main_layout.addWidget(left_panel, 1)
        
        # Right panel - Details
        self.right_panel = self.create_details_panel()
        main_layout.addWidget(self.right_panel, 3)
        
        # Apply styles
        self.apply_styles()
        
    def create_categories_panel(self):
        """Create the left categories panel"""
        panel = QtWidgets.QFrame()
        panel.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e0e6ed;
                border-radius: 16px;
                box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(panel)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Header
        header = QtWidgets.QLabel("📂 Expense Categories")
        header.setStyleSheet("""
            QLabel {
                font-size: 22px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px 0px;
                border-bottom: 3px solid #3498db;
                background: transparent;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Search box
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("🔍 Search categories...")
        self.search_edit.setStyleSheet("""
            QLineEdit {
                padding: 14px;
                border: 2px solid #e0e6ed;
                border-radius: 12px;
                font-size: 14px;
                background: #f8fafc;
                font-weight: 500;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: white;
                box-shadow: 0 0 0 3px rgba(52, 152, 219, 0.1);
            }
        """)
        self.search_edit.textChanged.connect(self.filter_categories)
        layout.addWidget(self.search_edit)
        
        # Categories list
        self.categories_list = QtWidgets.QListWidget()
        self.categories_list.setStyleSheet("""
            QListWidget {
                background: white;
                border: 2px solid #f1f5f9;
                border-radius: 12px;
                font-size: 14px;
                outline: none;
                padding: 5px;
            }
            QListWidget::item {
                padding: 18px 15px;
                border-bottom: 1px solid #f8fafc;
                color: #2c3e50;
                border-radius: 10px;
                margin: 4px;
                font-weight: 500;
                transition: all 0.2s ease;
            }
            QListWidget::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                border-radius: 10px;
                border: 2px solid #2980b9;
                transform: scale(1.02);
            }
            QListWidget::item:hover:!selected {
                background: #f8fafc;
                border-radius: 10px;
                border: 2px solid #e0e6ed;
                transform: translateX(5px);
            }
        """)
        self.categories_list.itemClicked.connect(self.on_category_selected)
        layout.addWidget(self.categories_list, 1)
        
        return panel
        
    def create_details_panel(self):
        """Create the right details panel"""
        panel = QtWidgets.QFrame()
        panel.setStyleSheet("""
            QFrame {
                background: #ffffff;
                border: 1px solid #e0e6ed;
                border-radius: 16px;
                box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(panel)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(25)
        
        # Header with category name
        self.details_header = QtWidgets.QLabel("Select a category to view expenses")
        self.details_header.setStyleSheet("""
            QLabel {
                font-size: 26px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px 0px;
                border-bottom: 3px solid #3498db;
                background: transparent;
            }
        """)
        self.details_header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(self.details_header)
        
        # Statistics cards with action buttons
        stats_container = QtWidgets.QWidget()
        stats_layout = QtWidgets.QVBoxLayout(stats_container)
        stats_layout.setSpacing(15)
        
        # Statistics cards row - CENTERED with only 2 cards
        stats_cards_layout = QtWidgets.QHBoxLayout()
        stats_cards_layout.setSpacing(30)
        
        # Create container for centered cards
        cards_container = QtWidgets.QWidget()
        cards_container_layout = QtWidgets.QHBoxLayout(cards_container)
        cards_container_layout.setContentsMargins(0, 0, 0, 0)
        cards_container_layout.setSpacing(30)
        
        self.total_expenses_card = self.create_stat_card("Expense Items", "0", "background: #DFF0FA;", "📊")
        self.total_amount_card = self.create_stat_card("Total Expense", "$0.00", "background: #F7DDE2;", "💰")
        
        cards_container_layout.addWidget(self.total_expenses_card)
        cards_container_layout.addWidget(self.total_amount_card)
        
        # Center the cards container
        stats_cards_layout.addStretch()
        stats_cards_layout.addWidget(cards_container)
        stats_cards_layout.addStretch()
        
        stats_layout.addLayout(stats_cards_layout)
        
        layout.addWidget(stats_container)
        
        # Search and Filter Section (like ExpensesTab)
        search_filter_frame = QtWidgets.QFrame()
        search_filter_frame.setStyleSheet("QFrame { background: transparent; border: none; }")
        search_filter_layout = QtWidgets.QHBoxLayout(search_filter_frame)
        search_filter_layout.setSpacing(10)
        search_filter_layout.setContentsMargins(0, 0, 0, 0)
        
        # LEFT SIDE: Search and Date Filter
        left_section = QtWidgets.QHBoxLayout()
        left_section.setSpacing(10)
        
        # Search Bar - UPDATED placeholder text
        self.table_search_edit = QtWidgets.QLineEdit()
        self.table_search_edit.setPlaceholderText("🔍 Search by Expense Type, Expense Name, Category, Vendor...")
        self.table_search_edit.setMinimumHeight(36)
        self.table_search_edit.setStyleSheet("""
            QLineEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 10px;
                font-size: 13px;
                background: white;
                min-width: 300px;
            }
            QLineEdit:focus {
                border-color: #3498db;
                background: #fafbfc;
            }
        """)
        self.table_search_edit.textChanged.connect(self.filter_table_expenses)
        left_section.addWidget(self.table_search_edit)
        
        # Date Range Filter Button
        self.date_range_button = QtWidgets.QPushButton("📅")
        self.date_range_button.setMinimumHeight(36)
        self.date_range_button.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.date_range_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 16px;
                font-weight: bold;
                min-width: 40px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        self.date_range_button.clicked.connect(self.show_date_range_dialog)
        left_section.addWidget(self.date_range_button)
        
        search_filter_layout.addLayout(left_section)
        
        # RIGHT SIDE: Export Button
        search_filter_layout.addStretch(1)
        
        self.export_btn = QtWidgets.QPushButton("📤 Export")
        self.export_btn.setFixedHeight(400)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 10px;
                font-weight: bold;
                font-size: 18px;
                padding: 1px 6px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                transform: translateY(-1px);
            }
            QPushButton:pressed {
                background: #229954;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.export_btn.clicked.connect(self.open_pdf_export_dialog)
        self.export_btn.setEnabled(False)
        
        search_filter_layout.addWidget(self.export_btn)
        
        layout.addWidget(search_filter_frame)
        
        # Expenses table - WITHOUT the "Expense Details" label
        table_container = QtWidgets.QWidget()
        table_container.setStyleSheet("""
            background: white; 
            border-radius: 12px; 
            border: 1px solid #e0e6ed;
        """)
        table_layout = QtWidgets.QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)  # Removed margins since we don't have the label
        
        self.expenses_table = self.create_expenses_table()
        table_layout.addWidget(self.expenses_table)
        
        layout.addWidget(table_container, 1)
        
        return panel

    def update_statistics_cards(self, data, expenses):
        """Update statistics cards with category data - only Total Expense and Total Amount"""
        # Total Expense
        total_expenses_label = self.total_expenses_card.layout().itemAt(1).widget()
        total_expenses_label.setText(str(data['count']))
        
        # Total Amount
        total_amount_label = self.total_amount_card.layout().itemAt(1).widget()
        total_amount_label.setText(f"${data['total_amount']:,.2f}")
                
    def create_summary_card(self):
        """Create summary card for categories"""
        card = QtWidgets.QFrame()
        card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #34495e);
                border-radius: 12px;
                padding: 20px;
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(card)
        layout.setSpacing(12)
        
        total_categories = QtWidgets.QLabel("Total Categories: 0")
        total_categories.setStyleSheet("""
            color: white; 
            font-size: 16px; 
            font-weight: bold;
            padding: 5px;
        """)
        
        total_amount = QtWidgets.QLabel("Total Amount: $0.00")
        total_amount.setStyleSheet("""
            color: white; 
            font-size: 16px; 
            font-weight: bold;
            padding: 5px;
        """)
        
        layout.addWidget(total_categories)
        layout.addWidget(total_amount)
        
        return card
        
    def create_stat_card(self, title, value, color, icon):
        """Create a statistics card matching ExpensesTab style"""
        card = QtWidgets.QFrame()
        card.setFixedSize(220, 80)  # Fixed size for consistent layout

        card.setStyleSheet(f"""
            QFrame {{
                {color}
                border-radius: 18px;
                padding: 0px; 
                border: 1px solid rgba(0,0,0,0.08);
            }}
        """)

        # TRUE vertical + horizontal centering
        layout = QtWidgets.QVBoxLayout(card)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.setAlignment(QtCore.Qt.AlignCenter)

        # ---- Title ----
        title_label = QtWidgets.QLabel(title)
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        title_label.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                font-size: 14px;
                font-weight: bold;
            }
        """)

        # ---- Value ----
        value_label = QtWidgets.QLabel(value)
        value_label.setAlignment(QtCore.Qt.AlignCenter)
        value_label.setStyleSheet("""
            QLabel {
                background: transparent;
                border: none;
                color: #2c3e50;
                font-size: 20px;
                font-weight: 600;
            }
        """)

        layout.addWidget(title_label)
        layout.addWidget(value_label)

        return card

    def adjust_color_brightness(self, color, factor):
        """Adjust color brightness for gradients"""
        color = QtGui.QColor(color)
        h, s, v, a = color.getHsv()
        v = max(0, min(255, int(v * factor)))
        return QtGui.QColor.fromHsv(h, s, v, a).name()
        
    def create_expenses_table(self):
        """Create expenses table with UPDATED columns: S.No, Date, Expense Type, Category, Expense Name, Vendor, Description, Amount, Project, Payment Methods, Actions"""
        table = QtWidgets.QTableWidget()
        table.setColumnCount(10)  # Changed from 8 to 11
        table.setHorizontalHeaderLabels([
            "S.No", "📅 Date", "📂 Expense Type", "📝 Expense Name",
            "🏢 Vendor", "📝 Description", "💰 Amount", "🎯 Project", "💳 Payment Method", "⚡ Actions"
        ])
        
        table.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 1.5px solid #e1e8ed;
                border-radius: 12px;
                gridline-color: #e1e8ed;
                font-size: 11px;
                outline: none;
                selection-background-color: #e3f2fd;
            }
            QTableWidget::item {
                padding: 8px 4px;
                border-bottom: 1px solid #f8f9fa;
                border-right: 1px solid #e1e8ed;
                color: #2c3e50;
                font-size: 11px;
            }
            QTableWidget::item:selected {
                background: #e3f2fd;
                color: #2c3e50;
            }
            QTableWidget::item:hover {
                background: #f8f9fa;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2c3e50, stop:1 #34495e);
                color: white;
                font-weight: bold;
                font-size: 11px;
                padding: 8px 6px;
                border: none;
                border-right: 1px solid #3a506b;
            }
            QHeaderView::section:last {
                border-right: none;
            }
            QTableWidget QTableCornerButton::section {
                background: #2c3e50;
                border: none;
                border-bottom: 1px solid #3a506b;
                border-right: 1px solid #3a506b;
            }
        """)
        
        # Enable grid to show vertical lines
        table.setShowGrid(True)
        table.setGridStyle(QtCore.Qt.SolidLine)
        
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.setSortingEnabled(True)
        table.setWordWrap(True)
        
        # Set column widths
        header = table.horizontalHeader()
        header.setDefaultAlignment(QtCore.Qt.AlignCenter)
        header.setHighlightSections(False)
        header.setFixedHeight(35)
        
        # Make all columns manually sizeable
        for col in range(table.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Interactive)

        # Default manual column widths - UPDATED for new columns
        table.setColumnWidth(0, 60)    # S.No
        table.setColumnWidth(1, 120)   # Date
        table.setColumnWidth(2, 200)   # Expense Type
        table.setColumnWidth(3, 228)   # Expense Name
        table.setColumnWidth(4, 120)   # Vendor
        table.setColumnWidth(5, 120)   # Description
        table.setColumnWidth(6, 120)   # Amount
        table.setColumnWidth(7, 120)   # Project
        table.setColumnWidth(8, 170)   # Payment Method
        table.setColumnWidth(9, 80)  # Actions
        
        return table
        
    def apply_styles(self):
        """Apply window styles"""
        self.setStyleSheet("""
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
                font-family: 'Segoe UI', Arial, sans-serif;
            }
        """)
        
    def load_categories(self):
        """Load categories from expenses data"""
        categories = {}
        total_amount = 0
        
        for expense in self.expenses_data:
            # Use 'Category' field if available, otherwise 'type'
            category = expense.get('Category', expense.get('type', 'Unknown'))
            amount = expense.get('amount', 0)
            
            if category not in categories:
                categories[category] = {
                    'count': 0,
                    'total_amount': 0,
                    'expenses': []
                }
            
            categories[category]['count'] += 1
            categories[category]['total_amount'] += amount
            categories[category]['expenses'].append(expense)
            total_amount += amount
        
        # Sort categories by total amount (descending)
        sorted_categories = sorted(categories.items(), 
                                 key=lambda x: x[1]['total_amount'], 
                                 reverse=True)
        
        # Populate categories list
        self.categories_list.clear()
        self.categories_data = categories
        
        for category, data in sorted_categories:
            item = QtWidgets.QListWidgetItem(
                f"{category}\n"
                f"📊 {data['count']} expenses | 💰 ${data['total_amount']:,.2f}"
            )
            item.setData(QtCore.Qt.UserRole, category)
            
            # Set background color based on category
            color = self.category_colors.get(category, "#bdc3c7")
            item.setBackground(QtGui.QColor(color))
            
            self.categories_list.addItem(item)
        
    def filter_categories(self):
        """Filter categories based on search text"""
        search_text = self.search_edit.text().lower()
        
        for i in range(self.categories_list.count()):
            item = self.categories_list.item(i)
            category = item.data(QtCore.Qt.UserRole)
            item.setHidden(search_text not in category.lower())
            
    def on_category_selected(self, item):
        """Handle category selection"""
        category = item.data(QtCore.Qt.UserRole)
        self.current_category = category
        self.show_category_details(category)
        
    def show_category_details(self, category):
        """Show detailed view for selected category"""
        if category not in self.categories_data:
            return
            
        data = self.categories_data[category]
        expenses = data['expenses']
        
        # Update header
        self.details_header.setText(f"📊 {category} - Expense Details")
        
        # Enable export button
        self.export_btn.setEnabled(True)
        
        # Clear search filter when switching categories
        self.table_search_edit.clear()
        
        # Update statistics cards
        self.update_statistics_cards(data, expenses)
        
        # Update table
        self.update_expenses_table(expenses)
        
    def update_expenses_table(self, expenses):
        """Update expenses table with category expenses - UPDATED with new columns"""
        self.expenses_table.setRowCount(len(expenses))
        
        for row, expense in enumerate(reversed(expenses)):
            # S.No
            sno_item = QtWidgets.QTableWidgetItem(str(row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 0, sno_item)
            
            # Date - Convert to MM-dd-yy format
            iso_date = expense.get('date', '')
            try:
                # Try to parse the date and convert to MM-dd-yy format
                qd = QtCore.QDate.fromString(iso_date, "MM-dd-yyyy")
                if qd.isValid():
                    formatted_date = qd.toString("MM-dd-yy")
                else:
                    # Try ISO format
                    qd = QtCore.QDate.fromString(iso_date, "yyyy-MM-dd")
                    if qd.isValid():
                        formatted_date = qd.toString("MM-dd-yy")
                    else:
                        formatted_date = iso_date
            except:
                formatted_date = iso_date
            
            date_item = QtWidgets.QTableWidgetItem(formatted_date)
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 1, date_item)
            
            # Expense Type - AUTO-SET TO "Other Expenses" IF EMPTY
            expense_type = expense.get('expense_type', '')
            if not expense_type:
                expense_type = "Other Expenses"
            exp_type_item = QtWidgets.QTableWidgetItem(expense_type)
            exp_type_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 2, exp_type_item)

            # Expense Name
            expense_name = expense.get('expense_name', '')
            exp_name_item = QtWidgets.QTableWidgetItem(expense_name)
            exp_name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 3, exp_name_item)
            
            # Vendor
            vendor_item = QtWidgets.QTableWidgetItem(expense.get('vendor', ''))
            vendor_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 4, vendor_item)
            
            # Description
            desc_item = QtWidgets.QTableWidgetItem(expense.get('description', ''))
            desc_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 5, desc_item)
            
            # Amount
            amount = expense.get('amount', 0)
            # Format with US-style commas (e.g., 1,234.56 instead of 12,34,56)
            amount_item = QtWidgets.QTableWidgetItem(f"${amount:,.2f}")
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)

            # Color code amount
            if amount > 1000:
                amount_item.setForeground(QtGui.QColor('#e74c3c'))
            elif amount > 500:
                amount_item.setForeground(QtGui.QColor('#f39c12'))
            else:
                amount_item.setForeground(QtGui.QColor('#27ae60'))
                
            self.expenses_table.setItem(row, 6, amount_item)
            # Project
            project_item = QtWidgets.QTableWidgetItem(expense.get('project', ''))
            project_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 7, project_item)
            
            # Payment Method
            method_item = QtWidgets.QTableWidgetItem(expense.get('payment_method', ''))
            method_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 8, method_item)
            
            # Actions (only View button)
            actions_widget = self.create_action_buttons(expense)
            self.expenses_table.setCellWidget(row, 9, actions_widget)
        
        # Set row height
        for row in range(self.expenses_table.rowCount()):
            self.expenses_table.setRowHeight(row, 48)
            
    def create_action_buttons(self, expense):
        """Create action buttons for table - only View button"""
        actions_widget = QtWidgets.QWidget()
        actions_layout = QtWidgets.QHBoxLayout(actions_widget)
        actions_layout.setContentsMargins(3, 3, 3, 3)
        actions_layout.setSpacing(3)
        
        # View button only
        view_btn = QtWidgets.QPushButton("👁 View")
        view_btn.setToolTip("View Details")
        view_btn.setFixedSize(55, 28)
        view_btn.setStyleSheet("""
            QPushButton {
                background-color: #f8f9fa;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                font-size: 10px;
                font-weight: bold;
                padding: 2px;
            }
            QPushButton:hover {
                background-color: #e9ecef;
                border-color: #3498db;
            }
            QPushButton:pressed {
                background-color: #dee2e6;
            }
        """)
        view_btn.clicked.connect(lambda: self.view_expense_details(expense))
        
        actions_layout.addWidget(view_btn)
        actions_layout.addStretch()
        
        return actions_widget

    def filter_table_expenses(self):
        """Filter table expenses based on search text and date range - UPDATED search fields"""
        if not self.current_category:
            return
            
        search_text = self.table_search_edit.text().lower()
        category_expenses = self.categories_data[self.current_category]['expenses']
        
        # Check if date range filter is active
        date_range_active = "to" in self.date_range_button.text()
        from_date = None
        to_date = None
        
        if date_range_active:
            # Use stored QDate objects if available
            if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
                from_date = self.current_from_date
                to_date = self.current_to_date
            else:
                try:
                    date_text = self.date_range_button.text().replace("📅 ", "")
                    from_str, to_str = date_text.split(" to ")
                    # Parse from MMMM d, yyyy format
                    from_date = QtCore.QDate.fromString(from_str, "MMMM d, yyyy")
                    to_date = QtCore.QDate.fromString(to_str, "MMMM d, yyyy")
                except:
                    date_range_active = False
        
        filtered_expenses = []
        total_amount = 0
        count = 0
        
        for expense in category_expenses:
            # Get expense type and name for search
            expense_type = expense.get('expense_type', '').lower()
            if not expense_type:
                expense_type = "other expenses"
            expense_name = expense.get('expense_name', '').lower()
            category = expense.get('Category', expense.get('type', '')).lower()
            
            matches_search = (
                search_text in expense_type or
                search_text in expense_name or
                search_text in category or
                search_text in expense.get('vendor', '').lower() or
                search_text in expense.get('description', '').lower() or
                search_text in expense.get('project', '').lower() or
                search_text in expense.get('payment_method', '').lower() or
                str(expense.get('amount', 0)).lower().startswith(search_text)
            )
            
            # Date range filter - parse expense date as MM-dd-yyyy
            matches_date = True
            if date_range_active and from_date and to_date:
                try:
                    # Parse expense date as MM-dd-yyyy
                    expense_date = QtCore.QDate.fromString(expense.get('date', ''), "MM-dd-yyyy")
                    matches_date = (from_date <= expense_date <= to_date)
                except:
                    matches_date = False
            
            if matches_search and matches_date:
                filtered_expenses.append(expense)
                # Calculate statistics for filtered expenses
                total_amount += expense.get('amount', 0)
                count += 1
        
        # Update statistics cards with filtered data
        filtered_data = {
            'count': count,
            'total_amount': total_amount,
            'expenses': filtered_expenses
        }
        self.update_statistics_cards(filtered_data, filtered_expenses)
        
        # Update table with filtered expenses
        self.expenses_table.setRowCount(len(filtered_expenses))
        
        for row, expense in enumerate(reversed(filtered_expenses)):
            # S.No - Column 0
            sno_item = QtWidgets.QTableWidgetItem(str(row + 1))
            sno_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 0, sno_item)
            
            # Date - Convert to MM-dd-yy format - Column 1
            iso_date = expense.get('date', '')
            try:
                # Try to parse the date and convert to MM-dd-yy format
                qd = QtCore.QDate.fromString(iso_date, "MM-dd-yyyy")
                if qd.isValid():
                    formatted_date = qd.toString("MM-dd-yy")
                else:
                    # Try ISO format
                    qd = QtCore.QDate.fromString(iso_date, "yyyy-MM-dd")
                    if qd.isValid():
                        formatted_date = qd.toString("MM-dd-yy")
                    else:
                        formatted_date = iso_date
            except:
                formatted_date = iso_date
            
            date_item = QtWidgets.QTableWidgetItem(formatted_date)
            date_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 1, date_item)
            
            # Expense Type - AUTO-SET TO "Other Expenses" IF EMPTY - Column 2
            expense_type = expense.get('expense_type', '')
            if not expense_type:
                expense_type = "Other Expenses"
            exp_type_item = QtWidgets.QTableWidgetItem(expense_type)
            exp_type_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 2, exp_type_item)

            # Expense Name - Column 3
            expense_name = expense.get('expense_name', '')
            exp_name_item = QtWidgets.QTableWidgetItem(expense_name)
            exp_name_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 3, exp_name_item)
            
            # Vendor - Column 4
            vendor_item = QtWidgets.QTableWidgetItem(expense.get('vendor', ''))
            vendor_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 4, vendor_item)
            
            # Description - Column 5
            desc_item = QtWidgets.QTableWidgetItem(expense.get('description', ''))
            desc_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 5, desc_item)
            
            # Amount
            amount = expense.get('amount', 0)
            amount_item = QtWidgets.QTableWidgetItem(f"${amount:,.2f}")
            amount_item.setTextAlignment(QtCore.Qt.AlignCenter)

            # Color code amount
            if amount > 1000:
                amount_item.setForeground(QtGui.QColor('#e74c3c'))
            elif amount > 500:
                amount_item.setForeground(QtGui.QColor('#f39c12'))
            else:
                amount_item.setForeground(QtGui.QColor('#27ae60'))
                
            self.expenses_table.setItem(row, 6, amount_item)
            # Project - Column 7
            project_item = QtWidgets.QTableWidgetItem(expense.get('project', ''))
            project_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 7, project_item)
            
            # Payment Method - Column 8
            method_item = QtWidgets.QTableWidgetItem(expense.get('payment_method', ''))
            method_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.expenses_table.setItem(row, 8, method_item)
            
            # Actions (only View button) - Column 9
            actions_widget = self.create_action_buttons(expense)
            self.expenses_table.setCellWidget(row, 9, actions_widget)
        
        # Set row height
        for row in range(self.expenses_table.rowCount()):
            self.expenses_table.setRowHeight(row, 48)
            
    def show_date_range_dialog(self):
        """Show date range selection dialog"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("📅 Select Date Range")
        dialog.setModal(True)
        dialog.resize(400, 200)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        # Title
        title = QtWidgets.QLabel("Select Date Range")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(title)
        
        # Date inputs
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(20, 10, 20, 10)
        
        # Check if there's an active date range filter
        current_from_date = QtCore.QDate.currentDate().addMonths(-1)
        current_to_date = QtCore.QDate.currentDate()
        
        # Parse current filter if active - use stored QDate objects if available
        button_text = self.date_range_button.text()
        if hasattr(self, 'current_from_date') and hasattr(self, 'current_to_date'):
            # Use the stored QDate objects directly
            current_from_date = self.current_from_date
            current_to_date = self.current_to_date
        elif "to" in button_text and button_text != "📅":
            try:
                date_text = button_text.replace("📅 ", "")
                from_str, to_str = date_text.split(" to ")
                # Parse from MMMM d, yyyy format
                current_from_date = QtCore.QDate.fromString(from_str, "MMMM d, yyyy")
                current_to_date = QtCore.QDate.fromString(to_str, "MMMM d, yyyy")
            except:
                date_range_active = False
        
        # From Date - initialize with current filter or default
        self.from_date_edit = QtWidgets.QDateEdit()
        self.from_date_edit.setDate(current_from_date)
        self.from_date_edit.setCalendarPopup(True)
        self.from_date_edit.setDisplayFormat("MMMM d, yyyy")  # US format: Month Day, Year
        self.from_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """)
        
        # To Date - initialize with current filter or default
        self.to_date_edit = QtWidgets.QDateEdit()
        self.to_date_edit.setDate(current_to_date)
        self.to_date_edit.setCalendarPopup(True)
        self.to_date_edit.setDisplayFormat("MMMM d, yyyy")  # US format: Month Day, Year
        self.to_date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 1.5px solid #e1e8ed;
                border-radius: 8px;
                background: white;
                font-size: 13px;
            }
            QDateEdit:focus { border-color: #3498db; }
        """)
        
        form_layout.addRow("From Date:", self.from_date_edit)
        form_layout.addRow("To Date:", self.to_date_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons - Clear on left, Apply on right (Cancel button removed)
        button_layout = QtWidgets.QHBoxLayout()
        
        clear_btn = QtWidgets.QPushButton("Clear Filter")
        clear_btn.setMinimumHeight(40)
        clear_btn.setStyleSheet("""
            QPushButton {
                background: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #7f8c8d;
            }
        """)
        
        apply_btn = QtWidgets.QPushButton("Apply Filter")
        apply_btn.setMinimumHeight(40)
        apply_btn.setStyleSheet("""
            QPushButton {
                background: #27ae60;
                color: white;
                font-weight: bold;
                border-radius: 8px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #2ecc71;
            }
        """)
        
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()  # This pushes the Apply button to the right
        button_layout.addWidget(apply_btn)
        
        layout.addLayout(button_layout)
        
        # Connect signals
        def apply_filter():
            self.apply_date_range_filter()
            dialog.accept()
        
        def clear_filter():
            self.clear_date_range_filter()
            dialog.accept()
        
        apply_btn.clicked.connect(apply_filter)
        clear_btn.clicked.connect(clear_filter)
        
        dialog.exec_()

    def apply_date_range_filter(self):
        """Apply date range filter"""
        from_date_qdate = self.from_date_edit.date()
        to_date_qdate = self.to_date_edit.date()
        
        # Store the dates in both formats for different uses
        from_date_us = from_date_qdate.toString("MMMM d, yyyy")  # US format for display
        to_date_us = to_date_qdate.toString("MMMM d, yyyy")  # US format for display
        
        # Store the actual QDate objects for later use
        self.current_from_date = from_date_qdate
        self.current_to_date = to_date_qdate
        
        # Update button text to show active filter in US format
        self.date_range_button.setText(f"📅 {from_date_us} to {to_date_us}")
        
        # Clear search filter when applying date filter
        self.table_search_edit.clear()
        
        # Apply filter to table and update statistics cards
        self.filter_table_expenses()

    def clear_date_range_filter(self):
        """Clear date range filter"""
        self.date_range_button.setText("📅")
        
        # Clear the stored date objects
        if hasattr(self, 'current_from_date'):
            del self.current_from_date
        if hasattr(self, 'current_to_date'):
            del self.current_to_date
            
        # Clear search filter when clearing date filter
        self.table_search_edit.clear()
        
        # Refresh table without date filter
        self.filter_table_expenses()
        
        # Also refresh the statistics cards with the full category data
        if self.current_category and self.current_category in self.categories_data:
            data = self.categories_data[self.current_category]
            expenses = data['expenses']
            self.update_statistics_cards(data, expenses)

    def view_expense_details(self, expense):
        """View expense details - UPDATED with new fields"""
        # Format date as MM-dd-yy (matching table format)
        iso_date = expense.get('date', '')
        try:
            qd = QtCore.QDate.fromString(iso_date, "MM-dd-yyyy")
            if qd.isValid():
                formatted_date = qd.toString("MM-dd-yy")
            else:
                qd = QtCore.QDate.fromString(iso_date, "yyyy-MM-dd")
                if qd.isValid():
                    formatted_date = qd.toString("MM-dd-yy")
                else:
                    formatted_date = iso_date
        except:
            formatted_date = iso_date
        
        # Get expense type - auto-set to "Other Expenses" if empty
        expense_type = expense.get('expense_type', '')
        if not expense_type:
            expense_type = "Other Expenses"
        
        details_text = f"""
        <h3>Expense Details</h3>
        <table style="border-collapse: collapse; width: 100%;">
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Date:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{formatted_date}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Expense Type:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense_type}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Expense Name:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense.get('expense_name', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Vendor:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense.get('vendor', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Description:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense.get('description', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Amount:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">${expense.get('amount', 0):,.2f}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Project:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense.get('project', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee; font-weight: bold;">Payment Method:</td><td style="padding: 8px; border-bottom: 1px solid #eee;">{expense.get('payment_method', 'N/A')}</td></tr>
        </table>
        """
        
        msg = QtWidgets.QMessageBox(self)
        msg.setWindowTitle("Expense Details")
        msg.setTextFormat(QtCore.Qt.RichText)
        msg.setText(details_text)
        msg.exec_()
    
    def open_pdf_export_dialog(self):
        """Open the professional PDF export dialog for category expenses"""
        if not self.current_category:
            return
            
        try:
            # Collect expenses for the current category
            category_expenses = self.categories_data[self.current_category]['expenses']
            
            dialog = CategoryPDFExportDialog(self, self.current_category, category_expenses)
            result = dialog.exec_()
            
            # Only perform export if dialog was accepted AND has export parameters
            if result == QtWidgets.QDialog.Accepted and hasattr(dialog, '_export_params'):
                # Get export parameters and perform actual export
                export_params = dialog._export_params
                self.perform_pdf_export(export_params)
                    
        except Exception as e:
            print(f"Error opening PDF export dialog: {e}")
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error opening export dialog: {str(e)}")

    def perform_pdf_export(self, export_params):
        """Perform the actual export based on parameters"""
        try:
            # Get category expenses
            category_expenses = self.categories_data[self.current_category]['expenses']
            
            # Filter expenses based on export parameters
            expenses_to_export = []
            
            for expense in category_expenses:
                try:
                    # Parse expense date
                    expense_datetime = None
                    # CORRECTED: Use proper format strings with percent signs
                    date_formats = ["%Y-%m-%d", "%m-%d-%Y"]  # Changed from "MM-dd-yyyy" to "%m-%d-%Y"
                    
                    for date_format in date_formats:
                        try:
                            expense_datetime = datetime.strptime(expense.get('date', ''), date_format)
                            break
                        except ValueError:
                            continue
                    
                    # If we still couldn't parse the date, EXCLUDE the expense
                    if expense_datetime is None:
                        print(f"Warning: Could not parse date '{expense.get('date', '')}' for expense - EXCLUDING from export")
                        continue
                    
                    include_expense = False
                    
                    if export_params["range"] == "all":
                        include_expense = True
                    
                    elif export_params["range"] == "date_range":
                        from_date = export_params["from_date"]
                        to_date = export_params["to_date"]
                        
                        # Convert to date objects for comparison
                        expense_date_only = expense_datetime.date()
                        
                        # Ensure both from_date and to_date are date objects (not datetime)
                        if isinstance(from_date, datetime):
                            from_date_only = from_date.date()
                        else:
                            from_date_only = from_date
                        
                        if isinstance(to_date, datetime):
                            to_date_only = to_date.date()
                        else:
                            to_date_only = to_date
                        
                        # Check if expense date is within the range (inclusive)
                        if from_date_only <= expense_date_only <= to_date_only:
                            include_expense = True
                        else:
                            print(f"Excluding expense with date {expense_date_only} - not in export range {from_date_only} to {to_date_only}")
                    
                    elif export_params["range"] == "month":
                        month = export_params["month"]
                        year = export_params["year"]
                        if expense_datetime.month == month and expense_datetime.year == year:
                            include_expense = True
                        else:
                            print(f"Excluding expense with date {expense_datetime} - not in month {month}/{year}")
                    
                    elif export_params["range"] == "year":
                        year = export_params["year"]
                        if expense_datetime.year == year:
                            include_expense = True
                        else:
                            print(f"Excluding expense with date {expense_datetime} - not in year {year}")
                    
                    if include_expense:
                        expenses_to_export.append(expense)
                        
                except Exception as e:
                    print(f"Error processing expense: {e}")
                    continue
            
            if not expenses_to_export:
                QtWidgets.QMessageBox.warning(self, "Export Warning", 
                                            "No expenses found matching the selected criteria.")
                return
            
            print(f"Export: Found {len(expenses_to_export)} expenses to export")
            
            # Generate the export based on type
            if export_params["type"] == "pdf":
                self.generate_combined_pdf(expenses_to_export, export_params)
            elif export_params["type"] == "excel":
                self.generate_combined_excel(expenses_to_export, export_params)
                
        except Exception as e:
            print(f"Error performing export: {e}")
            QtWidgets.QMessageBox.critical(self, "Export Error", 
                                        f"Error during export: {str(e)}")
    
    def generate_combined_excel(self, expenses, export_params):
        """Generate a professional combined Excel report for category expenses"""
        try:
            # Create export directory if it doesn't exist
            export_dir = Path.home() / "Downloads" / "Expense_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters and category
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            category_name = self.remove_emojis(self.current_category)
            
            if export_params["range"] == "all":
                filename = f"{category_name}_All_Expenses_{timestamp}.xlsx"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"{category_name}_Expenses_{from_date}_to_{to_date}.xlsx"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"{category_name}_Expenses_{year}_{month:02d}.xlsx"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"{category_name}_Expenses_{year}.xlsx"

            excel_path = export_dir / filename

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Header information (ONLY MAIN TITLE - NO DATE)
            ws.merge_cells('A1:G1')
            ws['A1'] = f"MABS ENGINEERING LLC - {category_name.upper()} EXPENSES REPORT"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            # REMOVED DATE INFO SECTION
            # ws['A2'] = f"Generated: {generated_date}" - REMOVED
            
            # Export range info
            if export_params["range"] == "all":
                export_range_text = "All Expenses"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m-%d-%Y")
                to_date = export_params["to_date"].strftime("%m-%d-%Y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            ws['A2'] = f"Period: {export_range_text}"  # Moved to row 2

            # REMOVED STATISTICS SECTION
            # ws['A3'] = f"Total Expense: {total_expenses}" - REMOVED
            # ws['A4'] = f"Total Amount: ${total_amount:,.2f}" - REMOVED

            # Table headers - UPDATED columns (starting at row 4 instead of row 7)
            headers = ["S.No.", "Date", "Expense Type", "Expense Name", "Vendor", "Amount", "Payment Method"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)  # Changed from row 7 to row 4
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Expense data (starting at row 5 instead of row 8)
            for row_idx, expense in enumerate(expenses, 5):  # Changed from row 8 to row 5
                # Convert dates to MM-dd-YYYY format
                date_str = expense.get('date', '')
                try:
                    # Try to parse MM-dd-yyyy format
                    expense_date = datetime.strptime(date_str, "%m-%d-%Y")
                    display_date = expense_date.strftime("%m-%d-%Y")
                except:
                    try:
                        expense_date = datetime.strptime(date_str, "%m/%d/%Y")
                        display_date = expense_date.strftime("%m-%d-%Y")
                    except:
                        display_date = date_str
                
                # Get Expense Type
                expense_type = expense.get('expense_type', '')
                if not expense_type:
                    expense_type = "Other Expenses"
                
                # Get expense name
                expense_name = expense.get('expense_name', '')
                
                data = [
                    row_idx - 4,  # Changed from row_idx - 7
                    display_date,
                    self.remove_emojis(expense_type),
                    self.remove_emojis(
                        expense_name[:25] + "..." if len(expense_name) > 25 else expense_name
                    ),
                    self.remove_emojis(expense.get('vendor', '')),
                    expense.get('amount', 0),
                    self.remove_emojis(expense.get('payment_method', ''))
                ]

                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Style for sequential number column
                    if col == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                    
                    # Style for amount column
                    if col == 6:  # Amount column
                        if isinstance(value, (int, float)):
                            cell.number_format = '"$"#,##0.00'
                            if value > 1000:
                                cell.font = Font(color="FF0000", bold=True)
                            elif value > 500:
                                cell.font = Font(color="FFA500")

            # Auto-adjust column widths
            column_widths = {
                1: 8,   # S.No.
                2: 16,  # Date
                3: 30,  # Expense Type
                4: 30,  # Expense Name
                5: 20,  # Vendor
                6: 18,  # Amount
                7: 30   # Payment Method
            }
            
            for col_idx in range(1, len(headers) + 1):
                if col_idx in column_widths:
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[column_letter].width = column_widths[col_idx]
                else:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in ws[column_letter]:
                        if cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Add alternating row colors
            for row in range(5, ws.max_row + 1):  # Changed from row 8 to row 5
                if row % 2 == 0:  # Even rows
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.fill.start_color.index == '00000000':  # Default fill
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

            # Save the workbook
            wb.save(str(excel_path))

            # Open the Excel file
            if self.open_file(excel_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"The Excel file has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ Excel exported successfully!\n\n"
                                                f"File saved to: {excel_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            print(f"Error generating combined Excel: {e}")
            import traceback
            traceback.print_exc()
            QtWidgets.QMessageBox.critical(self, "Excel Generation Error", 
                                        f"Error generating Excel: {str(e)}")
            
    def generate_combined_pdf(self, expenses, export_params):
        """Generate a professional combined PDF report for category expenses - UPDATED with new columns"""
        try:
            # Create export directory if it doesn't exist
            export_dir = Path.home() / "Downloads" / "Expense_Exports"
            export_dir.mkdir(parents=True, exist_ok=True)

            # Generate filename based on export parameters and category
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            category_name = self.remove_emojis(self.current_category)
            
            if export_params["range"] == "all":
                filename = f"{category_name}_All_Expenses_{timestamp}.pdf"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%Y%m%d")
                to_date = export_params["to_date"].strftime("%Y%m%d")
                filename = f"{category_name}_Expenses_{from_date}_to_{to_date}.pdf"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                filename = f"{category_name}_Expenses_{year}_{month:02d}.pdf"
            elif export_params["range"] == "year":
                year = export_params["year"]
                filename = f"{category_name}_Expenses_{year}.pdf"

            pdf_path = export_dir / filename

            # Create PDF document with reduced margins
            doc = SimpleDocTemplate(str(pdf_path), pagesize=A4, 
                                topMargin=0.1*inch, bottomMargin=0.2*inch,
                                leftMargin=0.1*inch, rightMargin=0.1*inch)
            elements = []

            # Get styles
            styles = getSampleStyleSheet()
            
            # MABS Engineering LLC Header Style - Large and centered
            mabs_header_style = ParagraphStyle(
                'MABSHeader',
                parent=styles['Normal'],
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,  # Center aligned
                fontName='Helvetica-Bold',
                spaceAfter=5,
                spaceBefore=5
            )
            
            # Date Style - Small and gray
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#7f8c8d'),
                fontName='Helvetica',
                alignment=2,  # Right aligned
                spaceAfter=5
            )
            
            # Expense Report Title Style
            report_title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=12,
                spaceAfter=8,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,  # Center aligned
                fontName='Helvetica-Bold'
            )
            
            # Statistics Style
            stats_style = ParagraphStyle(
                'StatsStyle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=8,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1,
                fontName='Helvetica-Bold'
            )

            # 1. Header table with MABS Engineering LLC (centered) and Date (right)
            generated_date = datetime.now().strftime("%m/%d/%Y")
            header_data = [
                [Paragraph("MABS Engineering LLC", mabs_header_style), 
                Paragraph(f"{generated_date}", date_style)]
            ]

            header_table = Table(header_data, colWidths=[6.5*inch, 1.5*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                # Shift the first cell (MABS header) to the right
                ('LEFTPADDING', (0,0), (0,0), 115),
            ]))

            elements.append(header_table)
            
            # Add gap between Row 1 and Row 2
            elements.append(Spacer(1, 0.28*inch))
            
            # 2. Expense Report Title (centered below MABS)
            report_title = Paragraph(f"{category_name} - EXPENSE REPORT", report_title_style)
            elements.append(report_title)
            
            # Calculate statistics
            total_expenses = len(expenses)
            total_amount = sum(expense.get('amount', 0) for expense in expenses)
            
            # 3. Statistics (with vertical gap from row 2)
            elements.append(Spacer(1, 0.1*inch))
            
            stats_text = f"Total Entries: {total_expenses}&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Total Expense: ${total_amount:,.2f}"
            stats_paragraph = Paragraph(stats_text, stats_style)
            elements.append(stats_paragraph)
            
            # 4. Export range info
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#2c3e50'),
                fontName='Helvetica-Bold'
            )
            
            # Generate export range text
            if export_params["range"] == "all":
                export_range_text = "All Expenses"
            elif export_params["range"] == "date_range":
                from_date = export_params["from_date"].strftime("%m/%d/%y")
                to_date = export_params["to_date"].strftime("%m/%d/%y")
                export_range_text = f"{from_date} to {to_date}"
            elif export_params["range"] == "month":
                month = export_params["month"]
                year = export_params["year"]
                month_name = datetime(2000, month, 1).strftime("%B")
                export_range_text = f"{month_name} {year}"
            elif export_params["range"] == "year":
                year = export_params["year"]
                export_range_text = f"Year {year}"
            
            info_data = [[Paragraph(f"{export_range_text}", info_style)]]

            info_table = Table(info_data, colWidths=[7.5 * inch])
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (0, 0), 0),
                ('TOPPADDING', (0, 0), (0, 0), 4),
                ('BOTTOMPADDING', (0, 0), (0, 0), 4),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 0.2*inch))

            # 5. Expenses Table with UPDATED columns
            if expenses:
                # Prepare table data with required columns: S.No, Date, Expense Type, Expense Name, Vendor, Amount, Payment Method
                table_data = [["S.No", "Date", "Expense Type", "Expense Name", "Vendor", "Amount", "Payment Method"]]
                
                for idx, expense in enumerate(expenses, 1):
                    # Convert date to MM/dd/yyyy format
                    date_str = expense.get('date', '')
                    try:
                        # Try MM-dd-yyyy format first, then ISO
                        try:
                            expense_date = datetime.strptime(date_str, "%m-%d-%Y")
                        except ValueError:
                            # If MM-dd-yyyy fails, try yyyy-MM-dd format
                            expense_date = datetime.strptime(date_str, "%Y-%m-%d")
                        us_date = expense_date.strftime("%m/%d/%Y")
                    except:
                        us_date = date_str  # Fallback to original if parsing fails
                    
                    # Get expense type - auto-set to "Other Expenses" if empty
                    expense_type = expense.get('expense_type', '')
                    if not expense_type:
                        expense_type = "Other Expenses"
                    
                    # Remove emojis from fields
                    expense_type = self.remove_emojis(expense_type)
                    expense_name = self.remove_emojis(expense.get('expense_name', ''))
                    vendor = self.remove_emojis(expense.get('vendor', ''))
                    payment_method = self.remove_emojis(expense.get('payment_method', ''))
                    
                    table_data.append([
                        str(idx),
                        us_date,
                        expense_type[:25] + "..." if len(expense_type) > 25 else expense_type,
                        expense_name[:25] + "..." if len(expense_name) > 25 else expense_name,
                        vendor[:25] + "..." if len(vendor) > 25 else vendor,
                        f"${expense.get('amount', 0):,.2f}",
                        payment_method[:20] + "..." if len(payment_method) > 20 else payment_method
                    ])
                
                # Create table with adjusted column widths for required columns
                col_widths = [
                    0.5*inch,   # S.No
                    0.8*inch,   # Date
                    1.7*inch,   # Expense Type
                    1.8*inch,   # Expense Name
                    1.3*inch,   # Vendor
                    0.7*inch,   # Amount
                    1.0*inch    # Payment Method
                ]
                
                expense_table = Table(table_data, colWidths=col_widths)
                expense_table.setStyle(TableStyle([
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    
                    # Data row styling
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffffff')),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2c3e50')),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8f9fa'), colors.white]),
                    
                    # Row padding
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                
                elements.append(expense_table)
            else:
                no_data_style = ParagraphStyle(
                    'NoData',
                    parent=styles['Normal'],
                    fontSize=12,
                    textColor=colors.HexColor('#7f8c8d'),
                    alignment=1
                )
                elements.append(Paragraph("No expenses found for the selected criteria.", no_data_style))

            # Build PDF
            doc.build(elements)

            # Open the PDF
            if self.open_file(pdf_path):
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"The PDF has been opened automatically.")
            else:
                QtWidgets.QMessageBox.information(self, "Export Success", 
                                                f"✅ PDF exported successfully!\n\n"
                                                f"File saved to: {pdf_path}\n"
                                                f"Could not open automatically. Please open manually.")
                    
        except Exception as e:
            print(f"Error generating combined PDF: {e}")
            QtWidgets.QMessageBox.critical(self, "PDF Generation Error", 
                                        f"Error generating PDF: {str(e)}")
            
    def remove_emojis(self, text):
        """Remove emojis and special characters from text"""
        if not text:
            return text
        
        # Common emoji replacements
        emoji_replacements = {
            '🏢': '', '🎯': '', '💳': '', '💻': '', '✈️': '', 
            '🖥️': '', '📄': '', '🏦': '', '💵': '', '📱': '',
            '🔗': '', '✅': '', '💰': '', '📋': '', '📊': '',
            '⚡': '', '🗑️': '', '✏️': '', '👁️': '', '📅': '',
            '📂': '', '📝': '', '🥧': '', '🔍': '', '❌': '',
            '➕': '', '📤': '', '🚀': ''
        }
        
        # Replace emojis
        cleaned_text = text
        for emoji, replacement in emoji_replacements.items():
            cleaned_text = cleaned_text.replace(emoji, replacement)
        
        # Remove any remaining special characters and extra spaces
        cleaned_text = ' '.join(cleaned_text.split())
        
        return cleaned_text.strip()

    def open_file(self, file_path):
        """Open file with default application"""
        try:
            import os
            import platform
            import subprocess
            
            if platform.system() == "Darwin":  # macOS
                subprocess.call(("open", file_path))
            elif platform.system() == "Windows":  # Windows
                os.startfile(file_path)
            else:  # linux variants
                subprocess.call(("xdg-open", file_path))
            return True
        except Exception as e:
            print(f"Error opening file: {e}")
            return False
class YearCalendarGrid(QtWidgets.QWidget):
    """Professional 3x3 grid for year selection with unlimited past/future years"""
    
    def __init__(self, parent=None, start_year=1, end_year=9999):
        super().__init__(parent)
        self.selected_year = datetime.now().year
        self.start_year = start_year  # Minimum year (1 AD)
        self.end_year = end_year      # Maximum year (9999 AD)
        self.year_buttons = []
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(5)
        
        # Navigation buttons
        nav_layout = QtWidgets.QHBoxLayout()
        nav_layout.setSpacing(10)
        
        self.prev_block_btn = QtWidgets.QPushButton("◀◀")
        self.prev_block_btn.setFixedSize(40, 30)
        self.prev_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.prev_block_btn.clicked.connect(self.prev_nine_year_block)
        
        self.block_label = QtWidgets.QLabel("")
        self.block_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 14px;")
        self.block_label.setAlignment(QtCore.Qt.AlignCenter)
        
        self.next_block_btn = QtWidgets.QPushButton("▶▶")
        self.next_block_btn.setFixedSize(40, 30)
        self.next_block_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.next_block_btn.clicked.connect(self.next_nine_year_block)
        
        nav_layout.addWidget(self.prev_block_btn)
        nav_layout.addWidget(self.block_label)
        nav_layout.addWidget(self.next_block_btn)
        
        layout.addLayout(nav_layout)
        
        # Year grid container
        grid_container = QtWidgets.QWidget()
        grid_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 1px solid #dfe6e9;
                border-radius: 8px;
                padding: 10px;
            }
        """)
        grid_layout = QtWidgets.QGridLayout(grid_container)
        grid_layout.setSpacing(8)
        grid_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create 3x3 grid of year buttons
        self.year_buttons = []
        
        # Calculate current 9-year block start
        self.current_block_start = self.calculate_block_start(self.selected_year)
        
        for row in range(3):
            for col in range(3):
                year_btn = QtWidgets.QPushButton()
                year_btn.setFixedSize(70, 45)
                year_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
                self.year_buttons.append(year_btn)
                grid_layout.addWidget(year_btn, row, col)
        
        layout.addWidget(grid_container)
        
        # Current year display
        current_layout = QtWidgets.QHBoxLayout()
        current_layout.addStretch()
        
        self.current_year_label = QtWidgets.QLabel(f"Selected: {self.selected_year}")
        self.current_year_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 13px;
                background: #e8f6f3;
                padding: 6px 12px;
                border-radius: 6px;
                border: 1px solid #a3e4d7;
            }
        """)
        current_layout.addWidget(self.current_year_label)
        current_layout.addStretch()
        
        layout.addLayout(current_layout)
        
        # Update the grid
        self.update_nine_year_block_grid()
    
    def calculate_block_start(self, year):
        """Calculate which 9-year block a year belongs to"""
        # Formula: ((year - 1) // 9) * 9 + 1
        return ((year - 1) // 9) * 9 + 1
    
    def update_nine_year_block_grid(self):
        """Update the 3x3 grid with years from current 9-year block"""
        # Generate 9 consecutive years starting from current_block_start
        years = []
        
        for i in range(9):
            year = self.current_block_start + i
            years.append(year)
        
        # Update block label
        first_year = years[0]
        last_year = years[-1]
        self.block_label.setText(f"{first_year} - {last_year}")
        
        # Update button texts and styles
        current_year = datetime.now().year
        for i, year_btn in enumerate(self.year_buttons):
            year = years[i]
            
            # Check if year is within valid range (1-9999)
            if year < 1 or year > 9999:
                year_btn.setText("")
                year_btn.setEnabled(False)
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #f8f9fa;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        color: #bdc3c7;
                    }
                """)
                continue
            
            year_btn.setText(str(year))
            year_btn.setEnabled(True)
            
            # Style based on selection and current year
            if year == self.selected_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #27ae60, stop:1 #2ecc71);
                        color: white;
                        border: 2px solid #229954;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #229954, stop:1 #27ae60);
                    }
                """)
            elif year == current_year:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: #3498db;
                        color: white;
                        border: 2px solid #2980b9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #2980b9;
                    }
                """)
            else:
                year_btn.setStyleSheet("""
                    QPushButton {
                        background: white;
                        color: #2c3e50;
                        border: 1px solid #dfe6e9;
                        border-radius: 5px;
                        font-weight: bold;
                        font-size: 12px;
                    }
                    QPushButton:hover {
                        background: #f8f9fa;
                        border-color: #3498db;
                        color: #3498db;
                    }
                """)
            
            # Connect button click
            try:
                year_btn.clicked.disconnect()
            except TypeError:
                pass
            year_btn.clicked.connect(lambda checked, y=year: self.select_year(y))
    
    def select_year(self, year):
        """Select a year"""
        self.selected_year = year
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        self.year_selected.emit(year)
    
    def prev_nine_year_block(self):
        """Go to previous 9-year block (unlimited past)"""
        self.current_block_start -= 9
        
        # Unlimited past - no lower bound check
        # If we go below year 1, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def next_nine_year_block(self):
        """Go to next 9-year block (unlimited future)"""
        self.current_block_start += 9
        
        # Unlimited future - no upper bound check
        # If we go above year 9999, we'll show empty/disabled buttons
        self.update_nine_year_block_grid()
    
    def set_selected_year(self, year):
        """Set the selected year"""
        # Ensure year is within valid range
        if year < 1:
            year = 1
        elif year > 9999:
            year = 9999
        
        self.selected_year = year
        self.current_block_start = self.calculate_block_start(year)
        
        self.current_year_label.setText(f"Selected: {year}")
        self.update_nine_year_block_grid()
        
    def get_selected_year(self):
        """Get the selected year"""
        return self.selected_year
    
    # Add signal for year selection
    year_selected = QtCore.pyqtSignal(int)

class YearCalendarPopup(QtWidgets.QDialog):
    """Professional popup window for year selection with unlimited years"""
    
    year_selected = QtCore.pyqtSignal(int)
    
    def __init__(self, parent=None, current_year=None):
        super().__init__(parent)
        self.current_year = current_year or datetime.now().year
        self.setWindowTitle("Select Year")
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.WindowCloseButtonHint)
        self.setFixedSize(380, 450)
        self.setStyleSheet("""
            YearCalendarPopup {
                background: #ffffff;
                border: 1px solid #d1d8e0;
                border-radius: 12px;
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel("📅 Select Year")
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0;
                text-align: center;
                border-bottom: 2px solid #3498db;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(header)
        
        # Create YearCalendarGrid with unlimited years
        self.year_calendar = YearCalendarGrid(start_year=1, end_year=9999)
        self.year_calendar.set_selected_year(self.current_year)
        self.year_calendar.setStyleSheet("""
            YearCalendarGrid {
                background: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 10px;
            }
        """)
        layout.addWidget(self.year_calendar)
        
        # Selected year display
        selected_layout = QtWidgets.QHBoxLayout()
        selected_layout.addStretch()
        
        self.selected_label = QtWidgets.QLabel(f"")
        self.selected_label.setStyleSheet("""
            QLabel {
                font-weight: bold;
                color: #27ae60;
                font-size: 14px;
            }
        """)
        selected_layout.addWidget(self.selected_label)
        selected_layout.addStretch()
        
        layout.addLayout(selected_layout)
        
        # Action buttons
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.cancel_btn = QtWidgets.QPushButton("Cancel")
        self.cancel_btn.setFixedSize(120, 45)
        self.cancel_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: #c0392b;
                border: 2px solid #e74c3c;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        self.select_btn = QtWidgets.QPushButton("Select Year")
        self.select_btn.setFixedSize(120, 45)
        self.select_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
                border: 2px solid #27ae60;
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.select_btn.clicked.connect(self.on_select_clicked)
        
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.select_btn)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Connect signals
        self.year_calendar.year_selected.connect(self.on_year_changed)
    
    def on_year_changed(self, year):
        """Update selected year display when year is changed in calendar"""
        self.current_year = year
    
    def on_select_clicked(self):
        """Emit signal with selected year and close popup"""
        self.year_selected.emit(self.current_year)
        self.accept()
    
    def get_selected_year(self):
        """Get the selected year"""
        return self.current_year
    


class CategoryPDFExportDialog(QtWidgets.QDialog):
    """Professional PDF/Excel Export Dialog for Category Expenses with Tabs"""
    
    def __init__(self, parent=None, category_name=None, category_expenses=None):
        super().__init__(parent)
        self.category_name = category_name
        self.category_expenses = category_expenses or []
        self.export_range = "all"  # Default export range
        self.selected_dates = []
        self.export_type = "pdf"  # Default export type
        self.year_calendar_popup = None
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle(f"📊 Export {self.category_name} Expenses")
        self.setFixedSize(700, 750)
        self.setStyleSheet("""
            CategoryPDFExportDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
            }
        """)
        
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Header
        header = QtWidgets.QLabel(f"📤 Export {self.category_name} Expenses")
        header.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                color: white;
                border-radius: 10px;
                text-align: center;
            }
        """)
        header.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(header)
        
        # Export Type Tabs
        self.tab_widget = QtWidgets.QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #d5dbdb;
            }
        """)
        
        # PDF Export Tab
        self.pdf_tab = QtWidgets.QWidget()
        self.setup_pdf_tab()
        self.tab_widget.addTab(self.pdf_tab, "📄 PDF Export")
        
        # Excel Export Tab
        self.excel_tab = QtWidgets.QWidget()
        self.setup_excel_tab()
        self.tab_widget.addTab(self.excel_tab, "📊 Excel Export")
        
        layout.addWidget(self.tab_widget)
        
        # Connect tab change signal
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        
        # Progress Bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #ecf0f1;
            }
            QProgressBar::chunk {
                background-color: #27ae60;
                border-radius: 3px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Action Buttons
        button_layout = QtWidgets.QHBoxLayout()
        
        self.export_btn = QtWidgets.QPushButton("🚀 Export PDF")
        self.export_btn.setFixedHeight(45)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 150px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.cancel_btn = QtWidgets.QPushButton("❌ Cancel")
        self.cancel_btn.setFixedHeight(45)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #e74c3c;
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: #c0392b;
            }
        """)
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.cancel_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        
        layout.addLayout(button_layout)
    
    def setup_pdf_tab(self):
        """Setup the PDF export tab"""
        layout = QtWidgets.QVBoxLayout(self.pdf_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 PDF Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.all_radio = QtWidgets.QRadioButton("📋 Export All Category Expenses")
        self.all_radio.setChecked(True)
        self.all_radio.toggled.connect(lambda: self.on_range_changed("all"))
        
        self.date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.date_range_radio.toggled.connect(lambda: self.on_range_changed("date_range"))
        
        self.month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.month_radio.toggled.connect(lambda: self.on_range_changed("month"))
        
        self.year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.year_radio.toggled.connect(lambda: self.on_range_changed("year"))
        
        options_layout.addWidget(self.all_radio)
        options_layout.addWidget(self.date_range_radio)
        options_layout.addWidget(self.month_radio)
        options_layout.addWidget(self.year_radio)
        
        range_group.addButton(self.all_radio)
        range_group.addButton(self.date_range_radio)
        range_group.addButton(self.month_radio)
        range_group.addButton(self.year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container
        self.date_selection_container = QtWidgets.QWidget()
        self.date_selection_layout = QtWidgets.QVBoxLayout(self.date_selection_container)
        self.date_selection_layout.setSpacing(15)
        self.date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector
        self.date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.date_range_group.setMinimumHeight(120)
        self.date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        date_range_layout = QtWidgets.QHBoxLayout(self.date_range_group)
        date_range_layout.setSpacing(20)

        # From date section
        from_layout = QtWidgets.QVBoxLayout()
        from_label = QtWidgets.QLabel("From Date:")
        from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        from_layout.addWidget(from_label)
        self.from_date = QtWidgets.QDateEdit()
        self.from_date.setDisplayFormat("MM-dd-yyyy")
        self.from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setFixedSize(160, 45)
        self.from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        from_layout.addWidget(self.from_date)
        date_range_layout.addLayout(from_layout)

        # To date section
        to_layout = QtWidgets.QVBoxLayout()
        to_label = QtWidgets.QLabel("To Date:")
        to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        to_layout.addWidget(to_label)
        self.to_date = QtWidgets.QDateEdit()
        self.to_date.setDisplayFormat("MM-dd-yyyy")
        self.to_date.setDate(QtCore.QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setFixedSize(160, 45)
        self.to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        to_layout.addWidget(self.to_date)
        date_range_layout.addLayout(to_layout)

        date_range_layout.addStretch()
        self.date_selection_layout.addWidget(self.date_range_group)

        # Month Selector
        self.month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.month_group.setMinimumHeight(150)
        self.month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        month_layout = QtWidgets.QVBoxLayout(self.month_group)
        month_layout.setSpacing(15)

        # Month and Year selection in one row
        month_year_row_layout = QtWidgets.QHBoxLayout()
        month_year_row_layout.setSpacing(15)

        # Month selection
        month_container = QtWidgets.QHBoxLayout()
        month_label = QtWidgets.QLabel("Select Month:")
        month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        month_container.addWidget(month_label)
        self.month_combo = QtWidgets.QComboBox()
        self.month_combo.setFixedSize(200, 45)
        self.month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months()
        month_container.addWidget(self.month_combo)
        month_year_row_layout.addLayout(month_container)

        # Year selection for month export
        year_container = QtWidgets.QHBoxLayout()
        year_label_month = QtWidgets.QLabel("Select Year:")
        year_label_month.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_container.addWidget(year_label_month)

        # Year field
        self.year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit_month.setFixedSize(150, 45)
        self.year_edit_month.setReadOnly(True)
        self.year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.year_calendar_btn_month.setFixedSize(50, 45)
        self.year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month)

        year_container.addWidget(self.year_edit_month)
        year_container.addWidget(self.year_calendar_btn_month)
        month_year_row_layout.addLayout(year_container)

        month_year_row_layout.addStretch()
        month_layout.addLayout(month_year_row_layout)
        self.date_selection_layout.addWidget(self.month_group)

        # Year Selector
        self.year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.year_group.setMinimumHeight(120)
        self.year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        year_layout = QtWidgets.QVBoxLayout(self.year_group)
        year_layout.setSpacing(15)

        # Year selection row
        year_row_layout = QtWidgets.QHBoxLayout()
        year_label = QtWidgets.QLabel("Select Year:")
        year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        year_row_layout.addWidget(year_label)
        
        # Year field
        self.year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.year_edit.setFixedSize(150, 45)
        self.year_edit.setReadOnly(True)
        self.year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        
        # Calendar button
        self.year_calendar_btn = QtWidgets.QPushButton("📅")
        self.year_calendar_btn.setFixedSize(50, 45)
        self.year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.year_calendar_btn.clicked.connect(self.show_year_popup)
        
        year_row_layout.addWidget(self.year_edit)
        year_row_layout.addWidget(self.year_calendar_btn)
        year_row_layout.addStretch()
        year_layout.addLayout(year_row_layout)

        self.date_selection_layout.addWidget(self.year_group)

        layout.addWidget(self.date_selection_container)

        # Initially hide all date selection components
        self.date_selection_container.setVisible(False)
        self.date_range_group.setVisible(False)
        self.month_group.setVisible(False)
        self.year_group.setVisible(False)
        
        # Preview Section
        preview_card = QtWidgets.QGroupBox("👁️ PDF Export Preview")
        preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        preview_layout = QtWidgets.QVBoxLayout(preview_card)
        
        self.preview_label = QtWidgets.QLabel(f"Ready to export all {self.category_name} expenses as PDF")
        self.preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.preview_label.setWordWrap(True)
        preview_layout.addWidget(self.preview_label)
        
        layout.addWidget(preview_card)
        
        # Connect signals for live preview updates
        self.from_date.dateChanged.connect(self.update_preview)
        self.to_date.dateChanged.connect(self.update_preview)
        self.month_combo.currentTextChanged.connect(self.update_preview)
    
    def setup_excel_tab(self):
        """Setup the Excel export tab"""
        layout = QtWidgets.QVBoxLayout(self.excel_tab)
        layout.setSpacing(15)
        
        # Export Options Card
        options_card = QtWidgets.QGroupBox("🎯 Excel Export Options")
        options_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
            }
        """)
        options_layout = QtWidgets.QVBoxLayout(options_card)
        
        # Export Range Selection
        range_group = QtWidgets.QButtonGroup(self)
        
        self.excel_all_radio = QtWidgets.QRadioButton("📋 Export All Category Expenses")
        self.excel_all_radio.setChecked(True)
        self.excel_all_radio.toggled.connect(lambda: self.on_excel_range_changed("all"))
        
        self.excel_date_range_radio = QtWidgets.QRadioButton("📅 Export by Date Range")
        self.excel_date_range_radio.toggled.connect(lambda: self.on_excel_range_changed("date_range"))
        
        self.excel_month_radio = QtWidgets.QRadioButton("🗓️ Export by Month")
        self.excel_month_radio.toggled.connect(lambda: self.on_excel_range_changed("month"))
        
        self.excel_year_radio = QtWidgets.QRadioButton("📊 Export by Year")
        self.excel_year_radio.toggled.connect(lambda: self.on_excel_range_changed("year"))
        
        options_layout.addWidget(self.excel_all_radio)
        options_layout.addWidget(self.excel_date_range_radio)
        options_layout.addWidget(self.excel_month_radio)
        options_layout.addWidget(self.excel_year_radio)
        
        range_group.addButton(self.excel_all_radio)
        range_group.addButton(self.excel_date_range_radio)
        range_group.addButton(self.excel_month_radio)
        range_group.addButton(self.excel_year_radio)
        
        layout.addWidget(options_card)
        
        # Date Selection Container for Excel
        self.excel_date_selection_container = QtWidgets.QWidget()
        self.excel_date_selection_layout = QtWidgets.QVBoxLayout(self.excel_date_selection_container)
        self.excel_date_selection_layout.setSpacing(15)
        self.excel_date_selection_layout.setContentsMargins(10, 10, 10, 10)
        
        # Date Range Selector for Excel
        self.excel_date_range_group = QtWidgets.QGroupBox("📅 Select Date Range")
        self.excel_date_range_group.setMinimumHeight(120)
        self.excel_date_range_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_date_range_layout = QtWidgets.QHBoxLayout(self.excel_date_range_group)
        excel_date_range_layout.setSpacing(20)

        # From date section
        excel_from_layout = QtWidgets.QVBoxLayout()
        excel_from_label = QtWidgets.QLabel("From Date:")
        excel_from_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_from_layout.addWidget(excel_from_label)
        self.excel_from_date = QtWidgets.QDateEdit()
        self.excel_from_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_from_date.setDate(QtCore.QDate.currentDate().addMonths(-1))
        self.excel_from_date.setCalendarPopup(True)
        self.excel_from_date.setFixedSize(160, 45)
        self.excel_from_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_from_layout.addWidget(self.excel_from_date)
        excel_date_range_layout.addLayout(excel_from_layout)

        # To date section
        excel_to_layout = QtWidgets.QVBoxLayout()
        excel_to_label = QtWidgets.QLabel("To Date:")
        excel_to_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_to_layout.addWidget(excel_to_label)
        self.excel_to_date = QtWidgets.QDateEdit()
        self.excel_to_date.setDisplayFormat("MM-dd-yyyy")
        self.excel_to_date.setDate(QtCore.QDate.currentDate())
        self.excel_to_date.setCalendarPopup(True)
        self.excel_to_date.setFixedSize(160, 45)
        self.excel_to_date.setStyleSheet("""
            QDateEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QDateEdit:hover {
                border-color: #3498db;
            }
        """)
        excel_to_layout.addWidget(self.excel_to_date)
        excel_date_range_layout.addLayout(excel_to_layout)

        excel_date_range_layout.addStretch()
        self.excel_date_selection_layout.addWidget(self.excel_date_range_group)

        # Month Selector for Excel
        self.excel_month_group = QtWidgets.QGroupBox("🗓️ Select Month and Year")
        self.excel_month_group.setMinimumHeight(150)
        self.excel_month_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_month_layout = QtWidgets.QVBoxLayout(self.excel_month_group)
        excel_month_layout.setSpacing(15)

        # Month and Year selection in one row
        excel_month_year_row_layout = QtWidgets.QHBoxLayout()
        excel_month_year_row_layout.setSpacing(15)

        # Month selection
        excel_month_container = QtWidgets.QHBoxLayout()
        excel_month_label = QtWidgets.QLabel("Select Month:")
        excel_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_month_container.addWidget(excel_month_label)
        self.excel_month_combo = QtWidgets.QComboBox()
        self.excel_month_combo.setFixedSize(200, 45)
        self.excel_month_combo.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
            }
            QComboBox:hover {
                border-color: #3498db;
            }
        """)
        self.populate_months_excel()
        excel_month_container.addWidget(self.excel_month_combo)
        excel_month_year_row_layout.addLayout(excel_month_container)

        # Year selection for month
        excel_year_container = QtWidgets.QHBoxLayout()
        excel_year_month_label = QtWidgets.QLabel("Select Year:")
        excel_year_month_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_container.addWidget(excel_year_month_label)

        # Year field
        self.excel_year_edit_month = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit_month.setFixedSize(150, 45)
        self.excel_year_edit_month.setReadOnly(True)
        self.excel_year_edit_month.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)

        # Calendar button
        self.excel_year_calendar_btn_month = QtWidgets.QPushButton("📅")
        self.excel_year_calendar_btn_month.setFixedSize(50, 45)
        self.excel_year_calendar_btn_month.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn_month.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn_month.clicked.connect(self.show_year_popup_for_month_excel)

        excel_year_container.addWidget(self.excel_year_edit_month)
        excel_year_container.addWidget(self.excel_year_calendar_btn_month)
        excel_month_year_row_layout.addLayout(excel_year_container)

        excel_month_year_row_layout.addStretch()
        excel_month_layout.addLayout(excel_month_year_row_layout)
        self.excel_date_selection_layout.addWidget(self.excel_month_group)

        # Year Selector for Excel
        self.excel_year_group = QtWidgets.QGroupBox("📊 Select Year")
        self.excel_year_group.setMinimumHeight(120)
        self.excel_year_group.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 12px;
                background-color: #f8f9fa;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
            }
        """)
        excel_year_layout = QtWidgets.QVBoxLayout(self.excel_year_group)
        excel_year_layout.setSpacing(15)

        # Year selection row
        excel_year_row_layout = QtWidgets.QHBoxLayout()
        excel_year_label = QtWidgets.QLabel("Select Year:")
        excel_year_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 13px;")
        excel_year_row_layout.addWidget(excel_year_label)
        
        # Year field
        self.excel_year_edit = QtWidgets.QLineEdit(str(datetime.now().year))
        self.excel_year_edit.setFixedSize(150, 45)
        self.excel_year_edit.setReadOnly(True)
        self.excel_year_edit.setStyleSheet("""
            QLineEdit {
                padding: 12px;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
                font-weight: bold;
            }
        """)
        
        # Calendar button
        self.excel_year_calendar_btn = QtWidgets.QPushButton("📅")
        self.excel_year_calendar_btn.setFixedSize(50, 45)
        self.excel_year_calendar_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excel_year_calendar_btn.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                border: 2px solid #2980b9;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #2980b9;
                border-color: #21618c;
            }
            QPushButton:pressed {
                background: #21618c;
            }
        """)
        self.excel_year_calendar_btn.clicked.connect(self.show_year_popup_excel)
        
        excel_year_row_layout.addWidget(self.excel_year_edit)
        excel_year_row_layout.addWidget(self.excel_year_calendar_btn)
        excel_year_row_layout.addStretch()
        excel_year_layout.addLayout(excel_year_row_layout)

        self.excel_date_selection_layout.addWidget(self.excel_year_group)

        layout.addWidget(self.excel_date_selection_container)

        # Initially hide all date selection components for Excel
        self.excel_date_selection_container.setVisible(False)
        self.excel_date_range_group.setVisible(False)
        self.excel_month_group.setVisible(False)
        self.excel_year_group.setVisible(False)
        
        # Preview Section for Excel
        excel_preview_card = QtWidgets.QGroupBox("👁️ Excel Export Preview")
        excel_preview_card.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        excel_preview_layout = QtWidgets.QVBoxLayout(excel_preview_card)
        
        self.excel_preview_label = QtWidgets.QLabel(f"Ready to export all {self.category_name} expenses as Excel")
        self.excel_preview_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #7f8c8d;
                padding: 10px;
                background-color: #ecf0f1;
                border-radius: 5px;
            }
        """)
        self.excel_preview_label.setWordWrap(True)
        excel_preview_layout.addWidget(self.excel_preview_label)
        
        layout.addWidget(excel_preview_card)
        
        # Connect signals for live preview updates for Excel
        self.excel_from_date.dateChanged.connect(self.update_excel_preview)
        self.excel_to_date.dateChanged.connect(self.update_excel_preview)
        self.excel_month_combo.currentTextChanged.connect(self.update_excel_preview)
    
    def show_year_popup(self):
        """Show year calendar popup for PDF year selection"""
        try:
            current_year = int(self.year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected)
        
        # Center the popup
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month(self):
        """Show year calendar popup for PDF month+year selection"""
        try:
            current_year = int(self.year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_excel(self):
        """Show year calendar popup for Excel year selection"""
        try:
            current_year = int(self.excel_year_edit.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def show_year_popup_for_month_excel(self):
        """Show year calendar popup for Excel month+year selection"""
        try:
            current_year = int(self.excel_year_edit_month.text())
        except:
            current_year = datetime.now().year
        
        self.year_calendar_popup = YearCalendarPopup(self, current_year)
        self.year_calendar_popup.year_selected.connect(self.on_year_selected_for_month_excel)
        
        popup_rect = self.year_calendar_popup.geometry()
        main_rect = self.geometry()
        center_x = main_rect.x() + (main_rect.width() - popup_rect.width()) // 2
        center_y = main_rect.y() + (main_rect.height() - popup_rect.height()) // 2
        self.year_calendar_popup.move(center_x, center_y)
        
        self.year_calendar_popup.exec_()
    
    def on_year_selected(self, year):
        """Handle year selection from popup for PDF year export"""
        self.year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_for_month(self, year):
        """Handle year selection from popup for PDF month+year export"""
        self.year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_preview()
    
    def on_year_selected_excel(self, year):
        """Handle year selection from popup for Excel year export"""
        self.excel_year_edit.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def on_year_selected_for_month_excel(self, year):
        """Handle year selection from popup for Excel month+year export"""
        self.excel_year_edit_month.setText(str(year))
        self.year_calendar_popup = None
        self.update_excel_preview()
    
    def populate_months(self):
        """Populate months combo box for PDF"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)

    def populate_months_excel(self):
        """Populate months combo box for Excel"""
        months = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        self.excel_month_combo.addItems(months)
        self.excel_month_combo.setCurrentIndex(datetime.now().month - 1)
    
    def on_tab_changed(self, index):
        """Handle tab changes"""
        if index == 0:  # PDF tab
            self.export_type = "pdf"
            self.export_btn.setText("🚀 Export PDF")
            self.update_preview()
        elif index == 1:  # Excel tab
            self.export_type = "excel"
            self.export_btn.setText("🚀 Export Excel")
            self.update_excel_preview()
    
    def on_range_changed(self, range_type):
        """Handle export range changes for PDF"""
        self.export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.date_range_group.setVisible(date_range_visible)
        self.month_group.setVisible(month_visible)
        self.year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_preview()

    def on_excel_range_changed(self, range_type):
        """Handle export range changes for Excel"""
        self.excel_export_range = range_type
        
        # Show/hide specific date selection components based on the selected range
        date_range_visible = (range_type == "date_range")
        month_visible = (range_type == "month")
        year_visible = (range_type == "year")
        
        # Show/hide the specific group boxes
        self.excel_date_range_group.setVisible(date_range_visible)
        self.excel_month_group.setVisible(month_visible)
        self.excel_year_group.setVisible(year_visible)
        
        # Show the container if any date selection is needed
        self.excel_date_selection_container.setVisible(range_type != "all")
        
        # Update preview to show what will be exported
        self.update_excel_preview()
    
    def update_preview(self):
        """Update the PDF preview text"""
        if self.export_range == "all":
            self.preview_label.setText(f"📋 Will export ALL {self.category_name} expenses as PDF")
        
        elif self.export_range == "date_range":
            from_date = self.from_date.date().toString("MM/dd/yyyy")
            to_date = self.to_date.date().toString("MM/dd/yyyy")
            self.preview_label.setText(f"📅 Will export {self.category_name} expenses from {from_date} to {to_date} as PDF")
        
        elif self.export_range == "month":
            month = self.month_combo.currentText()
            year = self.year_edit_month.text()
            self.preview_label.setText(f"🗓️ Will export {self.category_name} expenses for {month} {year} as PDF")
        
        elif self.export_range == "year":
            year = self.year_edit.text()
            self.preview_label.setText(f"📊 Will export {self.category_name} expenses for the year {year} as PDF")

    def update_excel_preview(self):
        """Update the Excel preview text"""
        if hasattr(self, 'excel_export_range'):
            range_type = self.excel_export_range
        else:
            range_type = "all"
        
        if range_type == "all":
            self.excel_preview_label.setText(f"📋 Will export ALL {self.category_name} expenses as Excel")
        
        elif range_type == "date_range":
            from_date = self.excel_from_date.date().toString("MM/dd/yyyy")
            to_date = self.excel_to_date.date().toString("MM/dd/yyyy")
            self.excel_preview_label.setText(f"📅 Will export {self.category_name} expenses from {from_date} to {to_date} as Excel")
        
        elif range_type == "month":
            month = self.excel_month_combo.currentText()
            year = self.excel_year_edit_month.text()
            self.excel_preview_label.setText(f"🗓️ Will export {self.category_name} expenses for {month} {year} as Excel")
        
        elif range_type == "year":
            year = self.excel_year_edit.text()
            self.excel_preview_label.setText(f"📊 Will export {self.category_name} expenses for the year {year} as Excel")
    
    def get_export_parameters(self):
        """Get export parameters based on current selection"""
        if self.export_type == "pdf":
            if self.export_range == "all":
                return {"range": "all", "type": "pdf", "category": self.category_name}
            
            elif self.export_range == "date_range":
                from_date = self.from_date.date().toPyDate()
                to_date = self.to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "pdf", "category": self.category_name}
            
            elif self.export_range == "month":
                month = self.month_combo.currentIndex() + 1
                year = int(self.year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "pdf", "category": self.category_name}
            
            elif self.export_range == "year":
                year = int(self.year_edit.text())
                return {"range": "year", "year": year, "type": "pdf", "category": self.category_name}
        
        elif self.export_type == "excel":
            if hasattr(self, 'excel_export_range'):
                range_type = self.excel_export_range
            else:
                range_type = "all"
            
            if range_type == "all":
                return {"range": "all", "type": "excel", "category": self.category_name}
            
            elif range_type == "date_range":
                from_date = self.excel_from_date.date().toPyDate()
                to_date = self.excel_to_date.date().toPyDate()
                return {"range": "date_range", "from_date": from_date, "to_date": to_date, "type": "excel", "category": self.category_name}
            
            elif range_type == "month":
                month = self.excel_month_combo.currentIndex() + 1
                year = int(self.excel_year_edit_month.text())
                return {"range": "month", "month": month, "year": year, "type": "excel", "category": self.category_name}
            
            elif range_type == "year":
                year = int(self.excel_year_edit.text())
                return {"range": "year", "year": year, "type": "excel", "category": self.category_name}
    
    def start_export(self):
        """Start the export process based on selected type"""
        # Prevent multiple executions
        if hasattr(self, '_export_in_progress') and self._export_in_progress:
            return
            
        self._export_in_progress = True
        
        try:
            self.export_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            export_params = self.get_export_parameters()
            
            # Simulate export process
            for i in range(101):
                if not hasattr(self, '_export_in_progress'):  # Check if still valid
                    return
                QtWidgets.QApplication.processEvents()
                self.progress_bar.setValue(i)
                QtCore.QThread.msleep(10)
            
            # Store export parameters for parent to use after dialog closes
            self._export_params = export_params
            
            # Simply accept the dialog - let parent handle the actual export
            self.accept()
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.export_btn.setEnabled(True)
            self._export_in_progress = False